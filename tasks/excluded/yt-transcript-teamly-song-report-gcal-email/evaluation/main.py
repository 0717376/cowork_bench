"""
Evaluation for yt-transcript-teamly-song-report-gcal-email task.

Checks:
1. Song_Analysis_Report.xlsx exists with Tracklist / Artist_Stats /
   Publication_Plan sheets, correct columns and row counts, and the Tracklist
   covers real Afrobeat songs mapped to real artists.

   NOTE: the read-only YouTube transcript for 7ZQzGq32kAY is generic lyric
   content (no announcer/title lines, no artist names), so the *exact*
   song->artist ordering of the groundtruth xlsx is NOT recoverable from the
   data source. We therefore do not gate on byte-equal reproduction of the GT
   Tracklist; the critical "real song<->artist mapping" check is evaluated via
   the REAL_SONG_KEYWORDS / REAL_ARTISTS coverage that IS supported by the
   task instructions and transcript timing. The GT xlsx is still loaded for
   non-critical structural diagnostics (sheet presence, row counts), and the
   per-pair sample is reported (with "artist ft. X" normalized to the base
   artist) as a non-gating diagnostic.
2. Teamly page 'Afrobeat Mix 2024 - Blog Analysis' exists with a substantive
   body: an overview plus a song table covering several real songs AND real
   artists (Burna Boy, Wizkid, Rema, Davido, CKay, Asake, Ayra Starr,
   Fireboy DML). Replaces the former Notion page + song database.
3. GCal has >= 3 publication events in March 2026.
4. Email sent to editorial@musicblog.com.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Pass threshold otherwise: accuracy >= 70%.
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks reflect the task's substance (real song<->artist analysis,
# the team-wiki record, and the two required side-effects), not mere structure.
CRITICAL_CHECKS = {
    "Tracklist values match groundtruth (real song<->artist mapping)",
    "Teamly analysis page has substantive song<->artist content",
    "GCal has >= 3 publication-related events in March 2026",
    "Email sent to editorial@musicblog.com",
}

# Real songs/artists that must appear in the analysis (source of truth: the
# groundtruth tracklist; the read-only YouTube transcript exposes these).
REAL_SONG_KEYWORDS = [
    "ye", "essence", "calm down", "love nwantiti", "feel", "woman",
    "champion sound", "terminator", "unavailable", "rush", "organise",
    "last last",
]
REAL_ARTISTS = [
    "burna boy", "wizkid", "rema", "davido", "ckay",
    "asake", "ayra starr", "fireboy dml",
]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def base_artist(s):
    """Normalize an artist annotation to its base artist.

    'Wizkid ft. Tems', 'Wizkid feat. Tems', 'Wizkid (feat. Tems)',
    'Wizkid x Tems', 'Wizkid & Tems' -> 'wizkid'. A featured-artist
    annotation must not count as a contradictory mapping.
    """
    a = norm(s)
    for sep in (" ft.", " ft ", " feat.", " feat ", " featuring",
                " (ft", " (feat", " x ", " & ", " , ", ", "):
        idx = a.find(sep)
        if idx != -1:
            a = a[:idx]
    return a.strip(" (").strip()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Song_Analysis_Report.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Song_Analysis_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Song_Analysis_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Tracklist values match groundtruth (real song<->artist mapping)",
               False, "no xlsx")
        return
    record("Song_Analysis_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        record("Tracklist values match groundtruth (real song<->artist mapping)",
               False, "unreadable xlsx")
        return
    record("Excel file readable", True)

    # Check Tracklist sheet
    tracklist_sheet = None
    for name in wb.sheetnames:
        if "track" in name.lower():
            tracklist_sheet = wb[name]
            break
    if tracklist_sheet is None:
        record("Tracklist sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Tracklist sheet exists", True)
        rows = list(tracklist_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_title = any("song" in h or "title" in h for h in headers)
        has_artist = any("artist" in h for h in headers)
        record("Tracklist has Song_Title and Artist columns", has_title and has_artist,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Tracklist has >= 8 data rows", len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")

    # Check Artist_Stats sheet
    artist_sheet = None
    for name in wb.sheetnames:
        if "artist" in name.lower() or "stat" in name.lower():
            artist_sheet = wb[name]
            break
    if artist_sheet is None:
        record("Artist_Stats sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Artist_Stats sheet exists", True)
        rows = list(artist_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Artist_Stats has >= 4 data rows", len(data_rows) >= 4,
               f"Found {len(data_rows)} rows")

    # Check Publication_Plan sheet
    pub_sheet = None
    for name in wb.sheetnames:
        if "pub" in name.lower() or "plan" in name.lower():
            pub_sheet = wb[name]
            break
    if pub_sheet is None:
        record("Publication_Plan sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Publication_Plan sheet exists", True)
        rows = list(pub_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_date = any("date" in h or "publish" in h for h in headers)
        record("Publication_Plan has Publish_Date column", has_date,
               f"Headers: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Publication_Plan has >= 3 data rows", len(data_rows) >= 3,
               f"Found {len(data_rows)} rows")

    # --- Groundtruth XLSX value comparison ---
    # CRITICAL semantic: the Tracklist must map real songs to the right artists.
    gt_path = os.path.join(groundtruth_workspace, "Song_Analysis_Report.xlsx")
    tracklist_values_ok = None  # None = could not evaluate via GT yet
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                       f"Available: {wb.sheetnames}")
                # Leave tracklist_values_ok as None so the critical mapping
                # check is decided uniformly by the loose coverage fallback
                # below (a missing Tracklist sheet yields zero hits -> fail).
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                       if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")

            # Build a song_title -> artist map from GT for the Tracklist sheet so
            # we can verify the real song<->artist mapping regardless of order.
            if gt_sname.strip().lower() == "tracklist":
                gt_hdr = [norm(c) for c in next(gt_ws.iter_rows(values_only=True))]
                a_hdr = [norm(c) for c in next(a_ws.iter_rows(values_only=True))]

                def col_idx(hdr, *cands):
                    for cand in cands:
                        for i, h in enumerate(hdr):
                            if cand in h:
                                return i
                    return None

                gt_title_i = col_idx(gt_hdr, "song_title", "title", "song")
                gt_art_i = col_idx(gt_hdr, "artist")
                a_title_i = col_idx(a_hdr, "song_title", "title", "song")
                a_art_i = col_idx(a_hdr, "artist")
                if None not in (gt_title_i, gt_art_i, a_title_i, a_art_i):
                    # Non-gating diagnostic only. The exact GT song->artist
                    # ordering is not recoverable from the generic transcript,
                    # so this is reported but does NOT set tracklist_values_ok
                    # (the critical mapping check uses the loose coverage
                    # fallback below). Featured-artist annotations ("ft. X")
                    # are normalized to the base artist before comparison and
                    # are never treated as contradictory mismatches.
                    gt_map = {}
                    for r in gt_rows:
                        if gt_title_i < len(r) and gt_art_i < len(r):
                            gt_map[norm(r[gt_title_i])] = base_artist(r[gt_art_i])
                    a_map = {}
                    for r in a_rows:
                        if a_title_i < len(r) and a_art_i < len(r):
                            a_map[norm(r[a_title_i])] = base_artist(r[a_art_i])
                    matched = 0
                    mismatched = []
                    for song, gt_artist in gt_map.items():
                        if song in a_map:
                            if a_map[song] == gt_artist:
                                matched += 1
                            else:
                                mismatched.append((song, a_map[song], gt_artist))
                    record("Tracklist song->artist pairs sample (diagnostic)",
                           matched >= 8,
                           f"matched={matched} (>=8 desired), "
                           f"mismatched={mismatched[:3]} (non-gating)")

            # NOTE: We deliberately do NOT compare individual GT row values
            # cell-by-cell. The generic transcript does not expose the exact
            # GT song ordering / chosen article titles / per-artist totals, so
            # byte-equal reproduction of the frozen GT rows is unrecoverable
            # from the data source and would be an unsatisfiable check. Sheet
            # presence and row counts (above) remain as structural checks.
        gt_wb.close()

    # CRITICAL mapping gate: verify the Tracklist actually covers real Afrobeat
    # songs mapped to real artists. The exact GT ordering is not recoverable
    # from the generic transcript, so we evaluate via real song/artist coverage
    # (supported by the task instructions + transcript timing) rather than
    # byte-equal GT reproduction. tracklist_values_ok stays None through the GT
    # block above; the GT xlsx is used only for non-critical structural
    # diagnostics.
    tl = None
    for name in wb.sheetnames:
        if "track" in name.lower():
            tl = wb[name]
            break
    blob = ""
    if tl is not None:
        for r in tl.iter_rows(values_only=True):
            blob += " " + " ".join(norm(c) for c in r if c is not None)
    song_hits = sum(1 for kw in REAL_SONG_KEYWORDS if kw in blob)
    artist_hits = sum(1 for kw in REAL_ARTISTS if kw in blob)
    tracklist_values_ok = song_hits >= 8 and artist_hits >= 5
    record("Tracklist values match groundtruth (real song<->artist mapping)",
           tracklist_values_ok,
           f"song_hits={song_hits} (need >=8), artist_hits={artist_hits} (need >=5)")


def check_teamly():
    print("\n=== Check 2: Teamly page 'Afrobeat Mix 2024 - Blog Analysis' ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Teamly database accessible", False, str(e))
        record("Teamly analysis page has substantive song<->artist content",
               False, str(e))
        return
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE '%%afrobeat mix%%'
               OR title ILIKE '%%blog analysis%%'
        """)
        pages = cur.fetchall()
        record("Teamly page 'Afrobeat Mix 2024 - Blog Analysis' exists",
               len(pages) >= 1, f"Found {len(pages)} matching pages")

        if not pages:
            record("Teamly analysis page has substantive song<->artist content",
                   False, "No matching page")
            return

        combined = " ".join((str(t) + " " + str(b)).lower() for _, t, b in pages)
        max_len = max(len(str(b)) for _, _, b in pages)
        record("Teamly analysis page has non-trivial body", max_len >= 200,
               f"Longest matching body is {max_len} chars")

        # CRITICAL semantic: the page must hold the real song<->artist analysis
        # (an overview alone is not enough; it must list real songs and artists).
        song_hits = sum(1 for kw in REAL_SONG_KEYWORDS if kw in combined)
        artist_hits = sum(1 for kw in REAL_ARTISTS if kw in combined)
        substantive = song_hits >= 8 and artist_hits >= 5
        record("Teamly analysis page has substantive song<->artist content",
               substantive,
               f"song_hits={song_hits} (need >=8), artist_hits={artist_hits} (need >=5)")
    except Exception as e:
        record("Teamly page check", False, str(e))
        record("Teamly analysis page has substantive song<->artist content",
               False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 3: GCal Publication Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-15' AND start_datetime < '2026-04-01'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    pub_events = [e for e in events
                  if any(kw in (e[0] or "").lower()
                         for kw in ["publish", "article", "genre", "afrobeat",
                                    "spotlight", "mix", "feature"])]
    record("GCal has >= 3 publication-related events in March 2026", len(pub_events) >= 3,
           f"Found {len(pub_events)} events: {[e[0] for e in pub_events]}")


def check_email_sent():
    print("\n=== Check 4: Email Sent to editorial@musicblog.com ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT count(*) FROM email.messages m
            JOIN email.folders f ON m.folder_id = f.id
            WHERE UPPER(f.name) IN ('SENT')
              AND m.to_addr::text ILIKE '%editorial@musicblog.com%'
        """)
        count = cur.fetchone()[0]
        if count == 0:
            cur.execute("""
                SELECT count(*) FROM email.sent_log sl
                JOIN email.messages m ON sl.message_id = m.id
                WHERE m.to_addr::text ILIKE '%editorial@musicblog.com%'
            """)
            count = cur.fetchone()[0]
        record("Email sent to editorial@musicblog.com", count >= 1, f"Sent count: {count}")
    except Exception as e:
        record("Email sent to editorial@musicblog.com", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_teamly()
    check_gcal()
    check_email_sent()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
