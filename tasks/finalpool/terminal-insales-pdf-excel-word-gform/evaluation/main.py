"""Evaluation for terminal-insales-pdf-excel-word-gform.
Checks:
1. Product_Quality_Audit.xlsx with 4 sheets and correct data
2. Quality_Audit_Report.docx with required sections
3. Supplier Quality Feedback form (gform schema) with 5+ questions
4. quality_audit.py script exists

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Pass threshold otherwise: accuracy >= 70%.

Notes on russification:
- wc.product_categories names are russified centrally (Аудио/Электроника/Камеры/
  Бытовая техника/ТВ и домашний кинотеатр/Часы). Category greps accept RU OR EN.
- order status slugs stay English (completed/refunded/...). Those greps stay EN.
- the agent writes Russian prose in Word, so docx keyword greps accept RU OR EN.
- the form title may be Russian ("Обратная связь по качеству поставщиков"), so the
  title match accepts RU (поставщ+качеств) OR EN (supplier+quality).
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# RU OR EN category labels for the 6 in-scope categories (mirror wc_relabel_map).
CATEGORY_ALIASES = [
    ("аудио", "audio"),
    ("камер", "camera"),
    ("электроник", "electronic"),
    ("бытов", "home appliance"),
    ("тв и домашн", "tv & home theater"),
    ("час", "watch"),
]

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "CRITICAL: Review_Summary lists the 6 in-scope categories with valid metrics",
    "CRITICAL: Quality_Scorecard scores match the formula and risk banding",
    "CRITICAL: Refund_Analysis aggregates real orders (refunded row, pct ~100%)",
    "CRITICAL: Supplier-quality form has 5+ questions and 2+ types",
    "CRITICAL: quality_audit.py references the score formula",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _to_float(c):
    if c is None:
        return None
    try:
        return float(str(c).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def _find_sheet(wb, *keywords):
    sheets_lower = [s.lower() for s in wb.sheetnames]
    for i, s in enumerate(sheets_lower):
        if any(kw in s for kw in keywords):
            return wb.sheetnames[i]
    return None


def check_excel(workspace):
    print("\n=== Check 1: Product_Quality_Audit.xlsx ===")
    path = os.path.join(workspace, "Product_Quality_Audit.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    # ---- Review_Summary sheet ----
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s or "summary" in s), 0)
    ws = wb[sheets[rs_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    check("Review_Summary has 6 category rows", len(data_rows) >= 6, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    # RU OR EN (categories are russified centrally; agent may read live RU names)
    check("Contains Audio category", "аудио" in all_text or "audio" in all_text, f"Text: {all_text[:120]}")
    check("Contains Electronics category", "электроник" in all_text or "electronic" in all_text,
          f"Text: {all_text[:120]}")

    # CRITICAL: the 6 in-scope categories present, each with numeric avg_rating
    # in [0,5] and pct_below_3stars in [0,100]. Parse columns positionally by
    # locating the two plausible numeric columns per row.
    cats_found = []
    for ru, en in CATEGORY_ALIASES:
        if ru in all_text or en in all_text:
            cats_found.append(ru)
    valid_metric_rows = 0
    for r in data_rows:
        nums = [_to_float(c) for c in r]
        nums = [n for n in nums if n is not None]
        has_rating = any(0.0 <= n <= 5.0 for n in nums)
        has_pct = any(0.0 <= n <= 100.0 for n in nums)
        if has_rating and has_pct:
            valid_metric_rows += 1
    check("CRITICAL: Review_Summary lists the 6 in-scope categories with valid metrics",
          len(cats_found) >= 6 and valid_metric_rows >= 6,
          f"categories matched={len(cats_found)} ({cats_found}), rows with valid metrics={valid_metric_rows}")

    # Build a per-category map of (avg_rating, pct_below_3stars) for the scorecard
    # cross-check below. Locate columns BY HEADER (the formula tokens avg_rating /
    # pct_below_3stars stay English per policy; allow RU header fallbacks) so we
    # don't confuse product_count with the percentage column.
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]] if rows else []

    def _col(*tokens):
        for i, h in enumerate(header):
            if any(t in h for t in tokens):
                return i
        return None

    cat_col = _col("category", "категор")
    rating_col = _col("avg_rating", "avg rating", "рейтинг")
    pct_col = _col("pct_below_3stars", "pct_below", "below", "ниже", "процент")

    review_metrics = {}
    for r in data_rows:
        if cat_col is not None and cat_col < len(r) and r[cat_col]:
            label = str(r[cat_col])
        else:
            label = next((str(c) for c in r if c and _to_float(c) is None), None)
        if not label:
            continue
        rating = _to_float(r[rating_col]) if (rating_col is not None and rating_col < len(r)) else None
        pct = _to_float(r[pct_col]) if (pct_col is not None and pct_col < len(r)) else None
        if rating is None or pct is None:
            # header lookup failed; fall back to value heuristics (rating in [0,5],
            # pct as the LAST in-range numeric to avoid grabbing product_count).
            nums = [_to_float(c) for c in r if _to_float(c) is not None]
            if rating is None:
                rating = next((n for n in nums if 0.0 <= n <= 5.0), None)
            if pct is None:
                pcts = [n for n in nums if n != rating and 0.0 <= n <= 100.0]
                pct = pcts[-1] if pcts else None
        if rating is not None and pct is not None:
            review_metrics[label.strip().lower()] = (rating, pct)

    # ---- Refund_Analysis sheet ----
    ra_idx = next((i for i, s in enumerate(sheets_lower) if "refund" in s or "order" in s), 1)
    if ra_idx < len(sheets):
        ws2 = wb[sheets[ra_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Refund_Analysis has at least 5 status rows", len(data_rows2) >= 5, f"Found {len(data_rows2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        # order status slugs stay English per policy
        check("Contains refunded status", "refunded" in all_text2, f"Text: {all_text2[:120]}")
        check("Contains completed status", "completed" in all_text2, f"Text: {all_text2[:120]}")

        # CRITICAL: real aggregation -> refunded row present, positive counts/amounts,
        # pct_of_total column sums to ~100%. Locate the pct column by header
        # (pct_of_total / процент); fall back to the rightmost in-range numeric.
        rh = [str(c).strip().lower() if c is not None else "" for c in rows2[0]] if rows2 else []
        pct_col2 = next((i for i, h in enumerate(rh)
                         if "pct" in h or "percent" in h or "процент" in h or "доля" in h), None)
        has_refunded_row = any("refunded" in str(r).lower() for r in data_rows2)
        positive_numeric = False
        pct_sum = 0.0
        pct_count = 0
        for r in data_rows2:
            nums = [_to_float(c) for c in r if _to_float(c) is not None]
            if any(n > 0 for n in nums):
                positive_numeric = True
            pv = None
            if pct_col2 is not None and pct_col2 < len(r):
                pv = _to_float(r[pct_col2])
            if pv is None:
                pct_candidates = [n for n in nums if 0.0 < n <= 100.0]
                pv = pct_candidates[-1] if pct_candidates else None
            if pv is not None:
                pct_sum += pv
                pct_count += 1
        pct_ok = pct_count >= 5 and abs(pct_sum - 100.0) <= 1.5
        check("CRITICAL: Refund_Analysis aggregates real orders (refunded row, pct ~100%)",
              has_refunded_row and positive_numeric and pct_ok,
              f"refunded_row={has_refunded_row} positive={positive_numeric} pct_sum={pct_sum:.2f} (n={pct_count})")

    # ---- Quality_Scorecard sheet ----
    qs_idx = next((i for i, s in enumerate(sheets_lower) if "quality" in s or "scorecard" in s), 2)
    if qs_idx < len(sheets):
        ws3 = wb[sheets[qs_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Quality_Scorecard has 6 rows", len(data_rows3) >= 6, f"Found {len(data_rows3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        # risk levels: agent may write RU (низкий/средний/высокий) OR EN (low/medium/high)
        check("Contains risk levels",
              any(k in all_text3 for k in ("low", "medium", "high", "низк", "средн", "высок")),
              f"Text: {all_text3[:120]}")

        # CRITICAL: recompute quality_score from the values the agent itself wrote
        # in Review_Summary and verify the scorecard cell + risk banding.
        # (Uses agent-written values, not live wc, to avoid volatility.)
        verified = 0
        attempted = 0
        for r in data_rows3:
            label = next((str(c) for c in r if c and _to_float(c) is None and
                          str(c).strip().lower() not in ("low", "medium", "high",
                                                         "низкий", "средний", "высокий",
                                                         "risk", "риск")), None)
            if not label:
                continue
            key = label.strip().lower()
            # find a matching review-summary row (RU or EN substring either way)
            match = None
            for rk, (rating, pct) in review_metrics.items():
                if rk in key or key in rk:
                    match = (rating, pct)
                    break
            if match is None:
                continue
            rating, pct = match
            expected_score = (rating / 5.0) * 100.0 - 2.0 * pct
            nums = [_to_float(c) for c in r if _to_float(c) is not None]
            score_cell = next((n for n in nums if abs(n - expected_score) <= 1.0), None)
            # risk banding
            if expected_score > 85:
                exp_risk = ("low", "низк")
            elif expected_score >= 70:
                exp_risk = ("medium", "средн")
            else:
                exp_risk = ("high", "высок")
            row_text = " ".join(str(c) for c in r if c).lower()
            risk_ok = any(rk in row_text for rk in exp_risk)
            attempted += 1
            if score_cell is not None and risk_ok:
                verified += 1
        check("CRITICAL: Quality_Scorecard scores match the formula and risk banding",
              attempted >= 6 and verified >= 6,
              f"verified {verified}/{attempted} category rows against formula+banding")

    # ---- Survey_Questions sheet ----
    sq_idx = next((i for i, s in enumerate(sheets_lower) if "survey" in s or "question" in s), 3)
    if sq_idx < len(sheets):
        ws4 = wb[sheets[sq_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Survey_Questions has at least 5 rows", len(data_rows4) >= 5, f"Found {len(data_rows4)}")


def check_word(workspace):
    print("\n=== Check 2: Quality_Audit_Report.docx ===")
    path = os.path.join(workspace, "Quality_Audit_Report.docx")
    if not os.path.exists(path):
        check("Word file exists", False, f"Not found at {path}")
        return
    check("Word file exists", True)

    try:
        from docx import Document
        doc = Document(path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        # The agent writes Russian prose -> accept RU OR EN keywords.
        check("Contains executive summary",
              any(k in all_text for k in ("executive", "summary", "резюме", "сводка", "итог")),
              f"Text: {all_text[:150]}")
        check("Contains review analysis",
              ("review" in all_text or "отзыв" in all_text)
              and ("category" in all_text or "категор" in all_text),
              f"Text: {all_text[:150]}")
        check("Contains refund discussion",
              "refund" in all_text or "возврат" in all_text,
              f"Text: {all_text[:150]}")
        check("Contains recommendations",
              "recommend" in all_text or "рекоменд" in all_text,
              f"Text: {all_text[:150]}")
    except ImportError:
        check("python-docx available", False, "python-docx not installed")


def check_gform():
    print("\n=== Check 3: Supplier Quality Feedback Form ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Title may be Russian ("Обратная связь по качеству поставщиков") OR English
    # ("Supplier Quality Feedback"). Match either, and exclude the noise form.
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE (lower(title) LIKE '%%supplier%%quality%%'
               OR (lower(title) LIKE '%%поставщ%%' AND lower(title) LIKE '%%качеств%%'))
          AND lower(title) NOT LIKE '%%сотрудник%%'
          AND lower(title) NOT LIKE '%%employee%%'
    """)
    forms = cur.fetchall()
    check("Supplier Quality Feedback form exists", len(forms) >= 1,
          f"Found forms: {[f[1] for f in forms]}")

    qcount = 0
    types = set()
    if forms:
        form_id = forms[0][0]
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
        qcount = cur.fetchone()[0]
        check("Form has at least 5 questions", qcount >= 5, f"Found {qcount} questions")

        cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s", (form_id,))
        questions = cur.fetchall()
        types = set(q[1] for q in questions)
        check("Form has multiple question types", len(types) >= 2, f"Types: {types}")

    # CRITICAL: the real supplier-quality form (not the noise form) exists with
    # >=5 questions AND >=2 distinct question types.
    check("CRITICAL: Supplier-quality form has 5+ questions and 2+ types",
          len(forms) >= 1 and qcount >= 5 and len(types) >= 2,
          f"forms={len(forms)} questions={qcount} types={types}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: quality_audit.py ===")
    path = os.path.join(workspace, "quality_audit.py")
    exists = os.path.exists(path)
    check("quality_audit.py exists", exists)

    # CRITICAL: the script actually computes the audit (references the formula:
    # avg_rating and the 2x low-review penalty), not an empty stub.
    formula_ok = False
    if exists:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                src = f.read().lower()
            has_rating = "avg_rating" in src or "avg rating" in src
            has_penalty = "pct_below_3stars" in src or ("2 *" in src or "2*" in src)
            formula_ok = has_rating and has_penalty
        except Exception:
            formula_ok = False
    check("CRITICAL: quality_audit.py references the score formula", formula_ok,
          "script must reference avg_rating and the 2x low-review penalty")


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Product_Quality_Audit.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        # No unexpected sheets beyond the 4 required
        expected_keywords = {"review", "summary", "refund", "order", "quality", "scorecard", "survey", "question"}
        for s in wb.sheetnames:
            s_lower = s.lower()
            matched = any(kw in s_lower for kw in expected_keywords)
            if not matched:
                check(f"No unexpected sheet '{s}'", False, f"Sheet '{s}' not expected")
                break
        else:
            check("No unexpected sheets in Excel", True)

        # Review ratings should not be negative
        sheets_lower = [s.lower() for s in wb.sheetnames]
        rs_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s or "summary" in s), 0)
        ws = wb[wb.sheetnames[rs_idx]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        has_negative = False
        for r in rows:
            for c in r:
                v = _to_float(c)
                if v is not None and v < 0:
                    has_negative = True
                    break
            if has_negative:
                break
        check("No negative values in Review_Summary", not has_negative,
              "Found negative value in summary data")

    # Google Form: no duplicate supplier-quality forms (RU OR EN title)
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT title, COUNT(*) FROM gform.forms
            WHERE lower(title) LIKE '%%supplier%%quality%%'
               OR (lower(title) LIKE '%%поставщ%%' AND lower(title) LIKE '%%качеств%%')
            GROUP BY title HAVING COUNT(*) > 1
        """)
        dupes = cur.fetchall()
        check("No duplicate supplier quality forms", len(dupes) == 0,
              f"Duplicates: {dupes}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_gform()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
