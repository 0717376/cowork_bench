"""Evaluation for terminal-canvas-forms-gcal-excel-email (RU).

Gate: accuracy >= 70% AND no CRITICAL check failed => PASS.
Any CRITICAL failure => immediate FAIL (sys.exit(1)) regardless of accuracy.

Below-75 quiz set is computed LIVE from canvas.quiz_submissions, never hardcoded.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "[CRIT]" if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        d = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{d}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


# ---------------------------------------------------------------------------
# Live Canvas-derived expectations (never hardcoded values).
# ---------------------------------------------------------------------------
def live_canvas():
    """Return (quiz_count, {quiz_id: avg}, set(below75 quiz_ids), {quiz_id: title})."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM canvas.quizzes WHERE course_id IN (7, 11)")
    quiz_count = cur.fetchone()[0]
    cur.execute("""
        SELECT q.id, AVG(qs.score), MAX(q.title)
        FROM canvas.quizzes q
        JOIN canvas.quiz_submissions qs ON q.id = qs.quiz_id
        WHERE q.course_id IN (7, 11)
        GROUP BY q.id
    """)
    avgs, titles = {}, {}
    for qid, avg, title in cur.fetchall():
        avgs[str(qid)] = float(avg)
        titles[str(qid)] = title
    below75 = {qid for qid, a in avgs.items() if a < 75.0}
    cur.close(); conn.close()
    return quiz_count, avgs, below75, titles


def check_excel(ws_path, quiz_count, avgs, below75, titles):
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "Quiz_Performance_Report.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}", critical=True)
        return
    check("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e), critical=True)
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Sheet 1: Quiz_Performance
    qp_name = None
    for candidate in ["quiz_performance", "quiz performance"]:
        if candidate in sheet_names_lower:
            qp_name = sheet_names_lower[candidate]
            break
    if qp_name is None:
        check("Quiz_Performance sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Quiz_Performance sheet exists", True)
        ws = wb[qp_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        check(f"Quiz_Performance has {quiz_count} rows",
              len(data_rows) == quiz_count,
              f"Found {len(data_rows)} data rows")

        # Build a per-quiz map from the sheet: quiz_id -> (avg, needs_review_str)
        sheet_map = {}
        for r in data_rows:
            qid = None
            if len(r) > 2 and r[2] is not None:
                # Quiz_ID column
                qid = str(r[2]).strip()
            avg = safe_float(r[4]) if len(r) > 4 else None
            needs = (str(r[6]).strip().lower() if len(r) > 6 and r[6] is not None else "")
            if qid:
                sheet_map[qid] = (avg, needs, r)

        # CRITICAL: the set flagged Needs_Review=Yes equals the live below-75 set.
        flagged_yes = set()
        for qid, (_avg, needs, _r) in sheet_map.items():
            if needs.startswith("yes") or needs in ("да", "y", "true"):
                flagged_yes.add(qid)
        # Match on quiz-id substring to be robust to "CMA 25341" style Quiz_IDs.
        def norm_ids(idset):
            out = set()
            for qid in idset:
                out.add(qid)
                tail = qid.split()[-1] if qid else qid
                out.add(tail)
            return out
        # Map live below75 ids to whatever the sheet used: compare by numeric tail.
        live_below_tails = {qid for qid in below75}
        sheet_yes_tails = set()
        for qid in flagged_yes:
            tail = "".join(ch for ch in qid if ch.isdigit())
            sheet_yes_tails.add(tail or qid)
        live_tails = set()
        for qid in below75:
            tail = "".join(ch for ch in qid if ch.isdigit())
            live_tails.add(tail or qid)
        check("Flagged Needs_Review=Yes set equals live below-75 set",
              sheet_yes_tails == live_tails,
              f"sheet={sorted(sheet_yes_tails)} live={sorted(live_tails)}",
              critical=True)

        # CRITICAL: specific quiz avg + flag, values re-queried live.
        def find_row(numeric_id):
            for qid, (avg, needs, r) in sheet_map.items():
                if numeric_id in qid:
                    return avg, needs, r
            # fallback: search Quiz_Title column
            for r in data_rows:
                if len(r) > 3 and r[3] and numeric_id in str(r[3]):
                    avg = safe_float(r[4]) if len(r) > 4 else None
                    needs = (str(r[6]).strip().lower() if len(r) > 6 and r[6] else "")
                    return avg, needs, r
            return None, None, None

        if "25341" in avgs:
            avg341, needs341, _ = find_row("25341")
            exp = avgs["25341"]
            check(f"CMA 25341 Avg_Score within 1.0 of live ({exp:.2f}) and Needs_Review=Yes",
                  (avg341 is not None and abs(avg341 - exp) <= 1.0
                   and (needs341 or "").startswith("yes")),
                  f"avg={avg341} needs={needs341} expected~{exp:.2f}",
                  critical=True)
        if "24296" in avgs:
            avg296, needs296, _ = find_row("24296")
            exp = avgs["24296"]
            check(f"CMA 24296 Avg_Score within 1.0 of live ({exp:.2f}) and Needs_Review=No",
                  (avg296 is not None and abs(avg296 - exp) <= 1.0
                   and (needs296 or "").startswith("no")),
                  f"avg={avg296} needs={needs296} expected~{exp:.2f}",
                  critical=True)

    # Sheet 2: Feedback_Summary
    fb_name = None
    for candidate in ["feedback_summary", "feedback summary"]:
        if candidate in sheet_names_lower:
            fb_name = sheet_names_lower[candidate]
            break
    if fb_name is None:
        check("Feedback_Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Feedback_Summary sheet exists", True)
        ws2 = wb[fb_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Feedback_Summary has 5 rows", len(data_rows2) == 5,
              f"Found {len(data_rows2)} rows")

    # Sheet 3: Remediation_Schedule
    rs_name = None
    for candidate in ["remediation_schedule", "remediation schedule"]:
        if candidate in sheet_names_lower:
            rs_name = sheet_names_lower[candidate]
            break
    if rs_name is None:
        check("Remediation_Schedule sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Remediation_Schedule sheet exists", True)
        ws3 = wb[rs_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
        check(f"Remediation_Schedule has {len(below75)} rows",
              len(data_rows3) == len(below75),
              f"Found {len(data_rows3)} rows, expected {len(below75)}")

    wb.close()


def check_gform():
    """Check the assessment feedback form against the RU forms (gform.*) schema."""
    print("\n=== Checking Form ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    form_id = None
    for fid, title in forms:
        t = (title or "").lower()
        if "academic performance self-assessment" in t or (
                ("assessment" in t or "performance" in t or "academic" in t
                 or "самооцен" in t or "успеваем" in t) and "привыч" not in t):
            form_id = fid
            break

    check("Assessment feedback form created", form_id is not None,
          f"Forms: {[f[1] for f in forms]}")

    if not form_id:
        conn.close()
        return

    cur.execute(
        "SELECT id, title, question_type, position FROM gform.questions "
        "WHERE form_id = %s ORDER BY position", (form_id,))
    questions = cur.fetchall()
    q_count = len(questions)
    check("Form has exactly 5 questions", q_count == 5, f"Found {q_count}")

    def is_text(t):
        return (t or "") in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH")

    def is_choice(t):
        return (t or "") in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE",
                             "CHOICE", "CHECKBOX", "SCALE", "scaleQuestion")

    types = [q[2] for q in questions]
    text_count = sum(1 for t in types if is_text(t))
    choice_count = sum(1 for t in types if is_choice(t))

    # Collect all option labels for this form. The Forms MCP stores choice
    # options in gform.questions.config JSONB as {type, options:[{value/label}]}.
    all_opts = ""
    try:
        cur.execute(
            "SELECT config FROM gform.questions WHERE form_id = %s", (form_id,))
        opt_vals = []
        for (cfg,) in cur.fetchall():
            if not cfg:
                continue
            if isinstance(cfg, str):
                try:
                    cfg = json.loads(cfg)
                except Exception:
                    continue
            for opt in (cfg.get("options") or []):
                val = opt.get("value") or opt.get("label") or ""
                if val:
                    opt_vals.append(val.lower())
        all_opts = " ".join(opt_vals)
    except Exception:
        conn.rollback()

    # Study-hours option set (EN required; RU accepted if agent localized).
    study_ok = (
        ("less than 3" in all_opts and "more than 8" in all_opts)
        or ("менее 3" in all_opts and "более 8" in all_opts)
    )
    # Learning-format option set.
    format_ok = (
        ("lectures" in all_opts and "hands-on labs" in all_opts
         and "self-paced online" in all_opts)
        or ("лекции" in all_opts and "лаборатор" in all_opts)
    )

    check("Study-hours choice options present", study_ok, f"opts='{all_opts[:160]}'")
    check("Learning-format choice options present", format_ok, f"opts='{all_opts[:160]}'")

    # Q1 free text + Q5 free text => at least 2 text questions.
    # Q2 study-hours + Q3 format + Q4 rating(1-5 as choice) => at least 3 choice questions.
    # CRITICAL: structure + option sets must all hold.
    check("Assessment form: 5 questions with 2 text + 3 choice and correct option sets",
          (q_count == 5 and text_count >= 2 and choice_count >= 3
           and study_ok and format_ok),
          f"q={q_count} text={text_count} choice={choice_count} "
          f"study={study_ok} format={format_ok}",
          critical=True)

    conn.close()


def check_gcal(below75, titles):
    """Check calendar review events: count, first date/time, 1-hour duration, consecutive days."""
    print("\n=== Checking Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Pull start and (optionally) end datetimes.
    end_available = True
    try:
        cur.execute(
            "SELECT summary, start_datetime, end_datetime "
            "FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
    except Exception:
        conn.rollback()
        end_available = False
        cur.execute("SELECT summary, start_datetime FROM gcal.events ORDER BY start_datetime")
        events = [(r[0], r[1], None) for r in cur.fetchall()]

    review_events = [e for e in events
                     if "review" in (e[0] or "").lower()
                     or "quiz" in (e[0] or "").lower()
                     or "разбор" in (e[0] or "").lower()]
    n_expected = len(below75)
    check(f"At least {n_expected} quiz review events created",
          len(review_events) >= n_expected,
          f"Found {len(review_events)} review events")

    # Reference quiz titles in events (non-critical sanity).
    found_count = 0
    sample = list(titles.items())
    for qid, title in sample:
        tail = "".join(ch for ch in qid if ch.isdigit())
        if any(tail and tail in (e[0] or "") for e in review_events):
            found_count += 1
    check("Review events reference quiz titles", found_count >= 2,
          f"Found {found_count} quiz references in events")

    # --- CRITICAL: exactly n_expected events, first on 2026-03-16 14:00,
    # 1-hour duration, consecutive days at 14:00. ---
    import datetime as _dt

    def as_dt(v):
        if v is None:
            return None
        if isinstance(v, _dt.datetime):
            return v
        s = str(v)
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S%z", "%Y-%m-%dT%H:%M:%S%z",
                    "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%d %H:%M"):
            try:
                return _dt.datetime.strptime(s.replace("Z", ""), fmt)
            except ValueError:
                continue
        try:
            return _dt.datetime.fromisoformat(s.replace("Z", ""))
        except Exception:
            return None

    starts = sorted([as_dt(e[1]) for e in review_events if as_dt(e[1]) is not None])
    ok_count = (len(review_events) == n_expected)

    ok_first = bool(starts) and starts[0].date() == _dt.date(2026, 3, 16) \
        and starts[0].hour == 14 and starts[0].minute == 0

    ok_consecutive = True
    for i in range(1, len(starts)):
        prev, cur_s = starts[i - 1], starts[i]
        if cur_s.hour != 14 or cur_s.minute != 0:
            ok_consecutive = False
            break
        if (cur_s.date() - prev.date()).days != 1:
            ok_consecutive = False
            break

    ok_duration = True
    if end_available:
        durations = []
        for e in review_events:
            s, en = as_dt(e[1]), as_dt(e[2])
            if s and en:
                durations.append((en - s).total_seconds() / 3600.0)
        ok_duration = bool(durations) and all(abs(d - 1.0) <= 0.1 for d in durations)

    check("Review events: correct count, first on 2026-03-16 14:00, "
          "1-hour duration, consecutive days at 14:00",
          ok_count and ok_first and ok_consecutive and ok_duration,
          f"count={len(review_events)}/{n_expected} first_ok={ok_first} "
          f"consec={ok_consecutive} dur_ok={ok_duration} "
          f"first={starts[0] if starts else None}",
          critical=True)

    conn.close()


def check_email(quiz_count, below75):
    """Check the cross-course review email (RU/EN body facts)."""
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()

    target_email = None
    for subj, from_addr, to_addr, body in all_emails:
        recipients = []
        if isinstance(to_addr, list):
            recipients = [str(r).strip().lower() for r in to_addr]
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                recipients = ([str(r).strip().lower() for r in parsed]
                              if isinstance(parsed, list)
                              else [str(to_addr).strip().lower()])
            except (json.JSONDecodeError, TypeError):
                recipients = [str(to_addr).strip().lower()]
        if "faculty@assessment.example.com" in recipients:
            target_email = (subj, from_addr, to_addr, body)
            break

    check("Email sent to faculty@assessment.example.com", target_email is not None,
          f"Total emails: {len(all_emails)}")

    if not target_email:
        # CRITICAL gate still must register a failure for the email facts.
        check("Email from coordinator@ to faculty@ with correct subject and body facts",
              False, "No email to faculty@assessment.example.com", critical=True)
        conn.close()
        return

    subj, from_addr, to_addr, body = target_email
    subj_l = (subj or "").lower()
    from_l = (from_addr or "").lower()
    body_l = (body or "").lower()

    check("Email subject mentions quiz performance",
          "quiz" in subj_l or "performance" in subj_l,
          f"Subject: {subj}")
    check("Email from coordinator@assessment.example.com",
          "coordinator@assessment.example.com" in from_l, f"From: {from_addr}")
    check("Email body mentions remediation/review (RU/EN)",
          any(k in body_l for k in ("remediation", "review", "flagged",
                                    "коррекц", "разбор", "обзор", "помечен")),
          "Expected remediation/review content")
    check("Email body mentions survey/feedback (RU/EN)",
          any(k in body_l for k in ("survey", "feedback", "опрос", "обратн")),
          "Expected survey/feedback mention")

    # --- CRITICAL: correct subject, from, and body facts ---
    subject_ok = subj_l.strip() == "quiz performance analysis - cross-course review"
    from_ok = "coordinator@assessment.example.com" in from_l
    total_ok = str(quiz_count) in body_l
    flagged_ok = str(len(below75)) in body_l
    survey_ok = any(k in body_l for k in ("survey", "feedback", "опрос", "обратн"))
    sessions_ok = any(k in body_l for k in ("session", "review", "scheduled",
                                            "сесси", "разбор", "заплан"))

    check("Email: correct subject/sender and body has total quizzes, flagged count, "
          "survey + sessions mentioned",
          subject_ok and from_ok and total_ok and flagged_ok
          and survey_ok and sessions_ok,
          f"subject_ok={subject_ok} from_ok={from_ok} total({quiz_count})={total_ok} "
          f"flagged({len(below75)})={flagged_ok} survey={survey_ok} sessions={sessions_ok}",
          critical=True)

    conn.close()


def check_reverse_validation():
    print("\n=== Reverse Validation ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr FROM email.messages
            WHERE from_addr = 'coordinator@assessment.example.com'
        """)
        sent_emails = cur.fetchall()
        noise_recipients = ["all-staff@university.edu", "faculty@university.edu",
                            "all@university.edu", "researchers@university.edu"]
        for email_row in sent_emails:
            to_str = str(email_row[0]).lower()
            for noise in noise_recipients:
                if noise in to_str:
                    check("No email sent to noise recipients", False,
                          f"Sent to noise recipient: {noise}")
                    cur.close(); conn.close()
                    return
        check("No email sent to noise recipients", True)
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-CANVAS-FORMS-GCAL-EXCEL-EMAIL - EVALUATION")
    print("=" * 70)

    try:
        quiz_count, avgs, below75, titles = live_canvas()
    except Exception as e:
        print(f"[FATAL] Could not query live Canvas data: {e}")
        sys.exit(1)
    print(f"[live] quiz_count={quiz_count} below75={sorted(below75)}")

    check_excel(args.agent_workspace, quiz_count, avgs, below75, titles)
    check_gform()
    check_gcal(below75, titles)
    check_email(quiz_count, below75)
    check_reverse_validation()

    print("\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    if CRITICAL_FAILED:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILED}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0
    print(f"  Accuracy: {accuracy:.1f}%")
    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
