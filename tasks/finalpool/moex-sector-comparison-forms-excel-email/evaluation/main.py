"""Evaluation for moex-sector-comparison-forms-excel-email.

Checks:
1. Sector_Comparison.xlsx with Metrics sheet (5 rows) and Sector_Summary sheet (5 rows)
2. Forms (RU forms-mcp, schema gform.*) "Investment Preference Survey" with 4 questions
3. Email to investors@fund.example.com with "Sector Comparison" in subject

CRITICAL_CHECKS reflect the task's substance: a single critical failure => overall
FAIL (sys.exit(1)) regardless of accuracy. The accuracy>=70 gate applies afterward.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

SYMBOLS = ["SBER.ME", "GAZP.ME", "MGNT.ME", "MTSS.ME", "TCSG.ME"]
SECTORS = ["Financial Services", "Energy", "Consumer Defensive", "Communication Services", "Fintech"]

# Expected market caps in billions of RUB (marketCap / 1e9) from moex.stock_info.
EXPECTED_MKTCAP_B = {
    "SBER.ME": 2877.54, "GAZP.ME": 4681.71, "MGNT.ME": 435.44,
    "MTSS.ME": 465.16, "TCSG.ME": 401.20,
}
# Live current prices from moex.stock_info (RUB).
EXPECTED_PRICES = {
    "SBER.ME": 133.3, "GAZP.ME": 198.0, "MGNT.ME": 4439.0,
    "MTSS.ME": 275.05, "TCSG.ME": 2013.0,
}
# Average of the five current prices => 1411.67.
AVG_PRICE = sum(EXPECTED_PRICES.values()) / len(EXPECTED_PRICES)
# Correct Above/Below classification vs the 5-stock average.
EXPECTED_ASSESS = {
    sym: ("Above_Avg" if p > AVG_PRICE else "Below_Avg")
    for sym, p in EXPECTED_PRICES.items()
}
# Largest market cap => GAZP.ME (Газпром).  Highest P/E => MGNT (~9.08).
LARGEST_MKTCAP_SYM = "GAZP.ME"


def record(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        msg = f": {d}" if d else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


# Backwards-compatible alias for non-critical checks.
def check(name, condition, detail=""):
    record(name, condition, detail, critical=False)


def num_close(a, b, tol_pct=5.0):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(b) < 1e-6:
        return abs(a) < 0.01
    return abs(a - b) / abs(b) * 100 <= tol_pct


def check_excel(agent_ws):
    print("\n=== Проверка 1: Sector_Comparison.xlsx ===")
    path = os.path.join(agent_ws, "Sector_Comparison.xlsx")
    check("Файл Sector_Comparison.xlsx существует", os.path.isfile(path))
    if not os.path.isfile(path):
        record("Metrics: рыночная капитализация каждого тикера совпадает с moex (в млрд)",
               False, "no file", critical=True)
        record("Metrics: текущая цена каждого тикера совпадает с moex (live data)",
               False, "no file", critical=True)
        record("Sector_Summary: Price_Assessment вычислен корректно для MGNT/SBER/MTSS",
               False, "no file", critical=True)
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel читается", False, str(e))
        record("Metrics: рыночная капитализация каждого тикера совпадает с moex (в млрд)",
               False, "unreadable", critical=True)
        record("Metrics: текущая цена каждого тикера совпадает с moex (live data)",
               False, "unreadable", critical=True)
        record("Sector_Summary: Price_Assessment вычислен корректно для MGNT/SBER/MTSS",
               False, "unreadable", critical=True)
        return

    # --- Metrics sheet ---
    metrics_ws = None
    for sname in wb.sheetnames:
        if "metric" in sname.lower():
            metrics_ws = wb[sname]
            break
    check("Лист 'Metrics' существует", metrics_ws is not None, f"Sheets: {wb.sheetnames}")

    metrics_rows = []
    if metrics_ws is not None:
        rows = list(metrics_ws.iter_rows(min_row=2, values_only=True))
        metrics_rows = [r for r in rows if any(c is not None for c in r)]
        check("Лист Metrics содержит 5 строк", len(metrics_rows) == 5, f"Got {len(metrics_rows)}")

        all_text = " ".join(str(c) for row in metrics_rows for c in row if c is not None)
        for sym in SYMBOLS:
            check(f"Metrics содержит тикер {sym}", sym in all_text)

    # Per-symbol value lookup (find the row containing the symbol).
    def row_for(sym):
        for row in metrics_rows:
            row_text = " ".join(str(c) for c in row if c is not None)
            if sym in row_text:
                return row
        return None

    def row_has_value(row, expected, tol_pct=5.0):
        if row is None:
            return False
        for c in row:
            if num_close(c, expected, tol_pct=tol_pct):
                return True
        return False

    # --- CRITICAL: market caps (in billions) match moex within 5% ---
    mktcap_ok = True
    mktcap_detail = []
    for sym in SYMBOLS:
        r = row_for(sym)
        ok = row_has_value(r, EXPECTED_MKTCAP_B[sym], tol_pct=5.0)
        if not ok:
            mktcap_ok = False
            mktcap_detail.append(f"{sym}!={EXPECTED_MKTCAP_B[sym]}B (row={[str(x)[:14] for x in (r or [])][:8]})")
    record("Metrics: рыночная капитализация каждого тикера совпадает с moex (в млрд)",
           mktcap_ok, "; ".join(mktcap_detail), critical=True)

    # --- CRITICAL: current prices match live moex data within 5% ---
    price_ok = True
    price_detail = []
    for sym in SYMBOLS:
        r = row_for(sym)
        ok = row_has_value(r, EXPECTED_PRICES[sym], tol_pct=5.0)
        if not ok:
            price_ok = False
            price_detail.append(f"{sym}!={EXPECTED_PRICES[sym]} (row={[str(x)[:14] for x in (r or [])][:8]})")
    record("Metrics: текущая цена каждого тикера совпадает с moex (live data)",
           price_ok, "; ".join(price_detail), critical=True)

    # --- Sector_Summary sheet ---
    summary_ws = None
    for sname in wb.sheetnames:
        if "sector" in sname.lower() and "summary" in sname.lower():
            summary_ws = wb[sname]
            break
        elif "summary" in sname.lower():
            summary_ws = wb[sname]
    check("Лист 'Sector_Summary' существует", summary_ws is not None, f"Sheets: {wb.sheetnames}")

    summary_rows = []
    if summary_ws is not None:
        rows = list(summary_ws.iter_rows(min_row=2, values_only=True))
        summary_rows = [r for r in rows if any(c is not None for c in r)]
        check("Лист Sector_Summary содержит 5 строк", len(summary_rows) == 5, f"Got {len(summary_rows)}")

        all_text = " ".join(str(c) for row in summary_rows for c in row if c is not None)
        check("Sector_Summary содержит оценку цены (Above_Avg или Below_Avg)",
              "above_avg" in all_text.lower() or "below_avg" in all_text.lower()
              or "above" in all_text.lower() or "below" in all_text.lower(),
              f"Content: {all_text[:200]}")

    # --- CRITICAL: Price_Assessment is computed CORRECTLY vs the 5-stock average ---
    def assess_for(sym):
        for row in summary_rows:
            row_text = " ".join(str(c) for c in row if c is not None)
            if sym in row_text:
                low = row_text.lower()
                if "above_avg" in low or ("above" in low and "below" not in low):
                    return "Above_Avg"
                if "below_avg" in low or ("below" in low and "above" not in low):
                    return "Below_Avg"
                return None
        return None

    # Validate the three most-discriminating symbols: MGNT (highest, Above),
    # SBER & MTSS (low, Below).
    crit_syms = ["MGNT.ME", "SBER.ME", "MTSS.ME"]
    assess_ok = True
    assess_detail = []
    for sym in crit_syms:
        got = assess_for(sym)
        exp = EXPECTED_ASSESS[sym]
        if got != exp:
            assess_ok = False
            assess_detail.append(f"{sym}: got={got} exp={exp}")
    record("Sector_Summary: Price_Assessment вычислен корректно для MGNT/SBER/MTSS",
           assess_ok, "; ".join(assess_detail), critical=True)


def _option_values(config):
    """Extract option text values from a question config (RU forms-mcp shape)."""
    vals = []
    if not config:
        return vals
    opts = config.get("options") if isinstance(config, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    vals.append(str(v))
            else:
                vals.append(str(o))
    return vals


def check_gform():
    print("\n=== Проверка 2: Форма 'Investment Preference Survey' (forms / gform.*) ===")
    crit_sectors = ("Форма: вопрос о предпочтении сектора содержит все 5 секторов "
                    "(RU или EN)")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        check("Хотя бы одна форма существует", False, str(e))
        record(crit_sectors, False, "no db", critical=True)
        return
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    check("Хотя бы одна форма существует", len(forms) > 0,
          "No forms found in gform.forms")

    found_form_id = None
    for form_id, title in forms:
        t = (title or "").lower()
        if "investment" in t or "preference" in t or "sector" in t \
           or "инвест" in t or "предпочт" in t or "опрос" in t:
            found_form_id = form_id
            break
    if found_form_id is None and forms:
        found_form_id = forms[0][0]

    check("Форма 'Investment Preference Survey' существует",
          found_form_id is not None,
          f"Forms: {[(str(r[0])[:20], r[1]) for r in forms]}")

    if found_form_id is None:
        record(crit_sectors, False, "no form", critical=True)
        cur.close()
        conn.close()
        return

    cur.execute("SELECT title, question_type, config FROM gform.questions "
                "WHERE form_id = %s ORDER BY position", (found_form_id,))
    questions = cur.fetchall()
    cur.close()
    conn.close()

    check("Форма содержит ровно 4 вопроса", len(questions) == 4,
          f"Got {len(questions)} questions")

    parsed = []
    for q_title, q_type, q_config in questions:
        cfg = q_config if isinstance(q_config, dict) else (
            json.loads(q_config) if q_config else {})
        opts = _option_values(cfg)
        parsed.append({
            "title": (q_title or "").lower(),
            "type": q_type,
            "options_lower": [v.lower() for v in opts],
        })

    q_text = " ".join(p["title"] for p in parsed)
    check("Есть вопрос о предпочтении сектора",
          "sector" in q_text or "сектор" in q_text, f"Questions: {[p['title'] for p in parsed]}")
    check("Есть вопрос о терпимости к риску",
          "risk" in q_text or "риск" in q_text, f"Questions: {[p['title'] for p in parsed]}")
    check("Есть вопрос об инвестиционном горизонте",
          "horizon" in q_text or "term" in q_text or "горизонт" in q_text or "срок" in q_text,
          f"Questions: {[p['title'] for p in parsed]}")

    # --- CRITICAL: sector-preference question offers all 5 sectors (RU or EN) ---
    # Map each required sector to substrings accepted in either language.
    sector_keys = {
        "financial": ["financial services", "финанс"],
        "energy": ["energy", "энерг"],
        "consumer": ["consumer defensive", "потребит"],
        "communication": ["communication services", "телеком", "связ", "коммуникац"],
        "fintech": ["fintech", "финтех"],
    }
    sector_q = None
    best_hits = -1
    for p in parsed:
        joined = " ".join(p["options_lower"])
        hits = sum(1 for keys in sector_keys.values() if any(k in joined for k in keys))
        if hits > best_hits:
            best_hits = hits
            sector_q = p
    joined = " ".join(sector_q["options_lower"]) if sector_q else ""
    matched = {name: any(k in joined for k in keys) for name, keys in sector_keys.items()}
    all_sectors = all(matched.values())
    record(crit_sectors, all_sectors,
           f"options='{joined[:200]}' matched={matched}", critical=True)


def check_email():
    print("\n=== Проверка 3: Письмо ===")
    crit_largest = ("Письмо investors@fund: тема содержит 'Sector Comparison' И тело "
                    "верно называет крупнейшую капитализацию (GAZP / Газпром)")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        check("Письмо найдено", False, str(e))
        record(crit_largest, False, "no db", critical=True)
        return
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%sector%comparison%'
           OR subject ILIKE '%sector comparison%'
           OR to_addr::text ILIKE '%investors@fund%'
        LIMIT 10
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    check("Письмо с 'Sector Comparison' в теме найдено",
          len(rows) > 0, "No matching email found")

    to_addrs = [str(r[1]) for r in rows]
    check("Письмо отправлено на investors@fund.example.com",
          any("investors" in addr for addr in to_addrs),
          f"To addresses: {to_addrs}")

    bodies = [str(r[2] or "").lower() for r in rows]
    check("Тело письма упоминает рыночную капитализацию или данные по секторам",
          any(("market cap" in b or "капитализац" in b or "sector" in b or "сектор" in b
               or "gazp" in b or "sber" in b) for b in bodies),
          f"Body: {bodies[0][:200] if bodies else ''}")

    # --- CRITICAL: subject has Sector Comparison AND body names the largest-mktcap stock ---
    subj_ok = any(("sector" in str(r[0] or "").lower() and "comparison" in str(r[0] or "").lower())
                  for r in rows)
    # Largest market cap = GAZP / Газпром (4681.71B). Accept ticker or RU/EN name.
    largest_named = any(("gazp" in b or "газпром" in b or "gazprom" in b) for b in bodies)
    record(crit_largest, bool(rows) and subj_ok and largest_named,
           f"subj_ok={subj_ok} largest_named={largest_named} "
           f"body='{bodies[0][:160] if bodies else ''}'", critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Evaluation: moex-sector-comparison-forms-excel-email ===")

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")

    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"Overall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILS,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILS}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
