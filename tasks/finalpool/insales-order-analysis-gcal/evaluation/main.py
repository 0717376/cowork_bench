"""
Скрипт оценки для задачи insales-order-analysis-gcal (магазин InSales, схема wc.*).

Проверки:
1. Excel Order_Analysis.xlsx с листами Status_Summary и Low_Stock_Products
2. Событие Google Calendar "Restock Planning Meeting" в БД (дата, время, описание)

Все реалии магазина (имена клиентов, города и т.п.) русифицируются ЦЕНТРАЛЬНО
через db/zzz_wc_after_init.sql; groundtruth и ~11 литералов eval уже пропатчены
scripts/wc_patch_groundtruth.py. НЕ править значения данных wc.* вручную.
Идентификаторы (SKU, названия товаров, слаги статусов, заголовки колонок,
название события) остаются на английском — eval ищет их как подстроки.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []  # имена провалившихся обязательных проверок

# Содержательные проверки. Провал любой из них => итог FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "CRITICAL: Status_Summary статусы = распределение wc.orders",
    "CRITICAL: Low_Stock_Products SKU-набор = товары с stock_quantity<5",
    "CRITICAL: событие 'Restock Planning Meeting' 2026-03-10 10:00-11:00 UTC",
    "CRITICAL: описание события перечисляет товары с низким остатком",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300]) if len(detail) > 300 else detail
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}: {d}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def parse_hhmm(s):
    """Извлечь HH:MM из строки/времени как минуты от полуночи."""
    if s is None:
        return None
    s = str(s).strip()
    m = re.search(r"(\d{1,2}):(\d{2})", s)
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def compute_expected():
    """Вычислить ожидаемые значения из PostgreSQL (живые данные wc.*)."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    # Сводка по статусам
    cur.execute("""
        SELECT status, COUNT(*) as cnt,
               ROUND(SUM((total)::numeric), 2) as rev,
               ROUND(AVG((total)::numeric), 2) as avg_val
        FROM wc.orders GROUP BY status ORDER BY status
    """)
    status_rows = cur.fetchall()

    # Товары с низким остатком
    cur.execute("""
        SELECT name, sku, stock_quantity, total_sales
        FROM wc.products
        WHERE stock_quantity IS NOT NULL AND stock_quantity < 5
        ORDER BY stock_quantity, name
    """)
    low_stock_rows = cur.fetchall()

    conn.close()
    return {"status": status_rows, "low_stock": low_stock_rows}


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Order_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        # без файла содержательные проверки невозможны
        check("CRITICAL: Status_Summary статусы = распределение wc.orders", False, "нет файла")
        check("CRITICAL: Low_Stock_Products SKU-набор = товары с stock_quantity<5", False, "нет файла")
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    # Наличие листов (структурные, не критичные)
    check("Sheet 'Status_Summary' exists", get_sheet(wb, "Status_Summary") is not None,
          f"Found: {wb.sheetnames}")
    check("Sheet 'Low_Stock_Products' exists", get_sheet(wb, "Low_Stock_Products") is not None,
          f"Found: {wb.sheetnames}")

    # --- Status_Summary ---
    ws = get_sheet(wb, "Status_Summary")
    if ws and expected:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["status"]
        check("Status_Summary row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        agent_by_status = {}
        for row in agent_rows:
            if row and row[0]:
                agent_by_status[str(row[0]).strip().lower()] = row

        # CRITICAL: множество статусов совпадает точно
        exp_statuses = {str(r[0]).strip().lower() for r in exp}
        got_statuses = set(agent_by_status.keys())
        check("CRITICAL: Status_Summary статусы = распределение wc.orders",
              exp_statuses == got_statuses,
              f"Expected {sorted(exp_statuses)}, got {sorted(got_statuses)}")

        for exp_row in exp:
            status = exp_row[0]
            agent_row = agent_by_status.get(status.lower())
            if agent_row:
                check(f"Status '{status}' Order_Count",
                      num_close(agent_row[1], exp_row[1], 0),
                      f"Expected {exp_row[1]}, got {agent_row[1]}")
                check(f"Status '{status}' Total_Revenue",
                      num_close(agent_row[2], float(exp_row[2]), 5.0),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
                check(f"Status '{status}' Avg_Order_Value",
                      num_close(agent_row[3], float(exp_row[3]), 2.0),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
            else:
                check(f"Status '{status}' found", False, "Not in agent output")

        # Сортировка по алфавиту
        if len(agent_rows) >= 2:
            statuses = [str(r[0]).strip().lower() for r in agent_rows if r and r[0]]
            check("Status_Summary sorted alphabetically",
                  statuses == sorted(statuses),
                  f"Got: {statuses[:5]}")

    # --- Low_Stock_Products ---
    ws2 = get_sheet(wb, "Low_Stock_Products")
    if ws2 and expected:
        agent_rows = list(ws2.iter_rows(min_row=2, values_only=True))
        exp = expected["low_stock"]
        check("Low_Stock_Products row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        agent_by_sku = {}
        for row in agent_rows:
            if row and len(row) >= 4 and row[1]:
                agent_by_sku[str(row[1]).strip()] = row

        # CRITICAL: набор SKU совпадает ТОЧНО (без пропущенных/лишних)
        exp_skus = {str(r[1]).strip() for r in exp if r[1]}
        got_skus = set(agent_by_sku.keys())
        check("CRITICAL: Low_Stock_Products SKU-набор = товары с stock_quantity<5",
              exp_skus == got_skus,
              f"Missing {sorted(exp_skus - got_skus)[:5]}, extra {sorted(got_skus - exp_skus)[:5]}")

        # Сортировка по остатку (по возрастанию)
        stocks = [r[2] for r in agent_rows if r and r[2] is not None]
        if len(stocks) >= 2:
            check("Low_Stock sorted by Stock_Quantity ascending",
                  all(stocks[i] <= stocks[i + 1] for i in range(len(stocks) - 1)),
                  f"Stock values: {stocks[:5]}")

        # Точечная проверка значений по SKU + усечение Product_Name до 80
        checked = 0
        for exp_row in exp:
            if checked >= 5:
                break
            sku = str(exp_row[1]).strip() if exp_row[1] else None
            agent_row = agent_by_sku.get(sku)
            if agent_row:
                check(f"SKU '{sku}' Stock_Quantity",
                      num_close(agent_row[2], exp_row[2], 0),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
                check(f"SKU '{sku}' Total_Sales",
                      num_close(agent_row[3], exp_row[3], 1),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
                name_val = str(agent_row[0]) if agent_row[0] is not None else ""
                check(f"SKU '{sku}' Product_Name truncated <=80",
                      len(name_val) <= 80,
                      f"len={len(name_val)}")
                checked += 1


def check_calendar(expected):
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for gcal check", False, str(e))
        check("CRITICAL: событие 'Restock Planning Meeting' 2026-03-10 10:00-11:00 UTC", False, str(e))
        check("CRITICAL: описание события перечисляет товары с низким остатком", False, str(e))
        return

    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
    events = cur.fetchall()
    conn.close()

    check("At least one calendar event exists", len(events) > 0,
          f"Found {len(events)} events")

    meeting = None
    for ev in events:
        summary = str(ev[0] or "").strip().lower()
        if "restock" in summary and "planning" in summary:
            meeting = ev
            break

    check("Event title contains 'Restock Planning Meeting'", meeting is not None,
          f"Events found: {[e[0] for e in events]}")

    if meeting is None:
        check("CRITICAL: событие 'Restock Planning Meeting' 2026-03-10 10:00-11:00 UTC", False,
              "событие не найдено")
        check("CRITICAL: описание события перечисляет товары с низким остатком", False,
              "событие не найдено")
        return

    summary, description, start_dt, end_dt = meeting
    start_str = str(start_dt or "")
    end_str = str(end_dt or "")
    start_min = parse_hhmm(start_str)
    end_min = parse_hhmm(end_str)

    date_ok = "2026-03-10" in start_str
    time_ok = (start_min == 10 * 60) and (end_min == 11 * 60)
    check("Event is on 2026-03-10", date_ok, f"Got start: {start_str}")
    # CRITICAL: дата 2026-03-10 и время 10:00-11:00 UTC
    check("CRITICAL: событие 'Restock Planning Meeting' 2026-03-10 10:00-11:00 UTC",
          date_ok and time_ok,
          f"start={start_str}, end={end_str}")

    desc = str(description or "")
    check("Event description is not empty", len(desc) > 10,
          f"Description length: {len(desc)}")

    # CRITICAL: описание реально перечисляет товары с низким остатком.
    # Агент пишет описание по-русски, но имена товаров/SKU остаются английскими
    # идентификаторами — проверяем по подстрокам (имя ИЛИ SKU нескольких товаров).
    enumerated = False
    if expected and expected.get("low_stock"):
        low = expected["low_stock"]
        desc_l = desc.lower()
        hits = 0
        # каждый товар считается перечисленным, если в описании есть его SKU
        # или содержательная часть названия (>=4 символов первого токена)
        for name, sku, _stock, _sales in low:
            found = False
            if sku and str(sku).strip().lower() in desc_l:
                found = True
            elif name:
                token = re.split(r"[\s/_-]+", str(name).strip())
                token = token[0] if token else ""
                if len(token) >= 4 and token.lower() in desc_l:
                    found = True
            if found:
                hits += 1
        # требуем покрытия большинства товаров (минимум 3 или все, если их <3)
        need = min(3, len(low)) if low else 0
        enumerated = hits >= need and need > 0
        detail = f"hits={hits}/{len(low)}, need>={need}"
    else:
        detail = "нет ожидаемых данных (PG недоступен)"
    check("CRITICAL: описание события перечисляет товары с низким остатком",
          enumerated, detail)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()
    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
    else:
        # PG обязателен для содержательных проверок: без него критичные чеки падают.
        print("WARNING: PostgreSQL недоступен — содержательные проверки невозможны")
    check_excel(agent_workspace, expected)
    check_calendar(expected)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    all_ok = (not CRITICAL_FAILS) and accuracy >= 70

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"  Critical fails: {CRITICAL_FAILS}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "total_checks": total,
            "accuracy": accuracy,
            "critical_fails": CRITICAL_FAILS,
            "success": all_ok,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return all_ok, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, accuracy={accuracy:.1f}%"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
