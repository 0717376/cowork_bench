"""Evaluation for insales-coupon-campaign-gcal-gform.

Structure:
  - CRITICAL_CHECKS verify SEMANTIC correctness of the four core deliverables
    (Excel coupon analysis + priority rule, the 4 GCal launch events with
    correct dates/times, the GForm question structure/options, and the
    summary email subject+body substance). Any critical failure => sys.exit(1)
    BEFORE the accuracy gate, so a non-doing agent that only creates empty
    stubs cannot pass.
  - After criticals pass, a set of finer-grained scored checks must reach
    accuracy >= 70 to PASS.
"""
import os
import argparse
import sys
import re
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
          user="eigent", password="camel")

# Expected coupon semantics (Discount_Type, Usage_Count, expected Campaign_Priority).
# Source of truth = groundtruth_workspace/Coupon_Campaign_Plan.xlsx + the priority rule
# in task.md (>30 -> High, 10..30 -> Medium, <10 -> Low). Coupon CODES stay English.
EXPECTED_COUPONS = {
    "HOLIDAY30":     ("percent",     50, "high"),
    "VIP20":         ("percent",     41, "high"),
    "SAVE20":        ("fixed_cart",  39, "high"),
    "WELCOME10":     ("percent",     23, "medium"),
    "BULK10":        ("percent",     17, "medium"),
    "ELECTRONICS15": ("percent",     16, "medium"),
    "FREESHIP":      ("fixed_cart",  10, "medium"),
    "SUMMER25":      ("percent",      9, "low"),
    "FLASH50":       ("fixed_cart",   2, "low"),
    "NEWUSER5":      ("fixed_cart",   2, "low"),
}

# Expected GCal launch events: summary -> (date, start_hour, end_hour)
EXPECTED_EVENTS = {
    "HOLIDAY30 Campaign":  ("2026-04-01", 9, 18),
    "VIP20 Campaign":      ("2026-05-01", 9, 18),
    "SUMMER25 Campaign":   ("2026-06-01", 9, 18),
    "FLASH50 Flash Sale":  ("2026-06-15", 9, 18),
}


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def _extract_number(val):
    """Pull a numeric amount out of strings like '30%', '$20', '20.0'."""
    if val is None:
        return None
    m = re.search(r"-?\d+(?:[.,]\d+)?", str(val).replace(",", "."))
    return float(m.group(0)) if m else None


def connect():
    return psycopg2.connect(**DB)


# ----------------------------------------------------------------------------
# Excel analysis
# ----------------------------------------------------------------------------
def read_campaign_analysis(agent_workspace):
    """Return (rows_by_code, error_str). rows_by_code maps CODE -> dict."""
    import openpyxl
    path = os.path.join(agent_workspace, "Coupon_Campaign_Plan.xlsx")
    if not os.path.exists(path):
        return None, "Coupon_Campaign_Plan.xlsx not found"
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return None, f"Error opening Excel: {e}"
    rows = load_sheet_rows(wb, "Campaign Analysis")
    if rows is None:
        return None, "Sheet 'Campaign Analysis' not found"
    header = [str(c).strip() if c is not None else "" for c in rows[0]] if rows else []
    # Map header -> index, fall back to positional layout from task.md.
    def col(name, default):
        for i, h in enumerate(header):
            if h.strip().lower() == name.lower():
                return i
        return default
    ci_code = col("Code", 0)
    ci_type = col("Discount_Type", 1)
    ci_amt = col("Amount", 2)
    ci_usage = col("Usage_Count", 3)
    ci_prio = col("Campaign_Priority", 4)
    out = {}
    for r in rows[1:]:
        if not r or r[ci_code] is None:
            continue
        code = str(r[ci_code]).strip().upper()
        out[code] = {
            "discount_type": (str(r[ci_type]).strip().lower() if len(r) > ci_type and r[ci_type] is not None else ""),
            "amount": (r[ci_amt] if len(r) > ci_amt else None),
            "usage": (r[ci_usage] if len(r) > ci_usage else None),
            "priority": (str(r[ci_prio]).strip().lower() if len(r) > ci_prio and r[ci_prio] is not None else ""),
        }
    return out, None


# ----------------------------------------------------------------------------
# CRITICAL checks
# ----------------------------------------------------------------------------
def critical_excel(agent_workspace):
    """Core deliverable: every coupon present with correct discount_type,
    usage_count, and the priority RULE applied correctly."""
    rows, err = read_campaign_analysis(agent_workspace)
    if err:
        return [f"[Excel] {err}"]
    errors = []
    missing = [c for c in EXPECTED_COUPONS if c not in rows]
    if missing:
        errors.append(f"[Excel] missing coupon rows: {sorted(missing)}")
    # Require the full set so a stub with one row cannot pass.
    for code, (exp_type, exp_usage, exp_prio) in EXPECTED_COUPONS.items():
        row = rows.get(code)
        if row is None:
            continue
        if row["discount_type"] != exp_type:
            errors.append(f"[Excel] {code} Discount_Type={row['discount_type']!r}, expected {exp_type!r}")
        if not num_close(row["usage"], exp_usage, tol=0):
            errors.append(f"[Excel] {code} Usage_Count={row['usage']!r}, expected {exp_usage}")
        if row["priority"] != exp_prio:
            errors.append(f"[Excel] {code} Campaign_Priority={row['priority']!r}, expected {exp_prio}")
    return errors


def critical_gcal():
    """All 4 launch events present with correct date and 09:00-18:00 window."""
    errors = []
    try:
        conn = connect()
        cur = conn.cursor()
        # start_datetime/end_datetime are timestamptz; the calendar MCP stores a
        # naive 09:00 + Europe/Moscow as UTC (06:00). Normalize back to Moscow
        # wall-clock so date/hour match the task's Moscow-timezone schedule.
        cur.execute("""
            SELECT summary,
                   to_char(start_datetime AT TIME ZONE 'Europe/Moscow', 'YYYY-MM-DD') AS sdate,
                   EXTRACT(HOUR FROM start_datetime AT TIME ZONE 'Europe/Moscow')::int AS shour,
                   EXTRACT(HOUR FROM end_datetime   AT TIME ZONE 'Europe/Moscow')::int AS ehour
            FROM gcal.events
            WHERE start_datetime >= '2026-03-31' AND start_datetime < '2026-07-01'
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        return [f"[GCal] error querying events: {e}"]

    def find(summary):
        target = summary.strip().lower()
        for s, sdate, sh, eh in rows:
            if s and target in str(s).strip().lower():
                return (sdate, sh, eh)
        return None

    for summary, (exp_date, exp_sh, exp_eh) in EXPECTED_EVENTS.items():
        got = find(summary)
        if got is None:
            errors.append(f"[GCal] missing event '{summary}'")
            continue
        sdate, sh, eh = got
        if sdate != exp_date:
            errors.append(f"[GCal] '{summary}' date={sdate}, expected {exp_date}")
        # Time window check tolerant to +/-1h to absorb any timezone storage quirk.
        if sh is not None and abs(int(sh) - exp_sh) > 1:
            errors.append(f"[GCal] '{summary}' start hour={sh}, expected ~{exp_sh}")
        if eh is not None and abs(int(eh) - exp_eh) > 1:
            errors.append(f"[GCal] '{summary}' end hour={eh}, expected ~{exp_eh}")
    return errors


def _form_questions():
    """Return (form_title, [questions]) for the most recent feedback form, or (None, [])."""
    conn = connect()
    cur = conn.cursor()
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE title ILIKE '%coupon%' OR title ILIKE '%campaign%' OR title ILIKE '%feedback%'
        ORDER BY created_at DESC LIMIT 1
    """)
    frow = cur.fetchone()
    if not frow:
        cur.close(); conn.close()
        return None, []
    form_id, title = frow
    cur.execute("""
        SELECT title, question_type, required, config
        FROM gform.questions WHERE form_id = %s ORDER BY position ASC
    """, (form_id,))
    qrows = cur.fetchall()
    cur.close(); conn.close()
    return title, qrows


def _option_values(config):
    """Extract option value strings from a choiceQuestion config JSONB (dict already)."""
    if not isinstance(config, dict):
        return []
    opts = config.get("options") or []
    vals = []
    for o in opts:
        if isinstance(o, dict) and o.get("value") is not None:
            vals.append(str(o["value"]).strip().lower())
        elif isinstance(o, str):
            vals.append(o.strip().lower())
    return vals


def critical_gform():
    """Form must have: a choice question offering Percentage + Fixed Amount,
    a choice question with >=3 options (campaign timing), and a text question."""
    try:
        title, qrows = _form_questions()
    except Exception as e:
        return [f"[GForm] error querying form: {e}"]
    if title is None:
        return ["[GForm] no 'Coupon Campaign Feedback' form found"]
    errors = []
    choice_qs = [(t, cfg) for (t, qt, req, cfg) in qrows if qt == "choiceQuestion"]
    text_qs = [t for (t, qt, req, cfg) in qrows if qt == "textQuestion"]

    if len(text_qs) < 1:
        errors.append("[GForm] missing short-answer (textQuestion) for additional feedback")

    # Discount-type choice question with Percentage + Fixed Amount options.
    has_discount_q = False
    has_timing_q = False
    for t, cfg in choice_qs:
        vals = _option_values(cfg)
        blob = " ".join(vals)
        if ("percent" in blob or "процент" in blob) and ("fixed" in blob or "amount" in blob or "fixed_cart" in blob or "фикс" in blob or "сумм" in blob):
            has_discount_q = True
        if len(vals) >= 3:
            has_timing_q = True
    if not has_discount_q:
        errors.append("[GForm] missing single-choice discount-type question with Percentage + Fixed Amount options")
    if not has_timing_q:
        errors.append("[GForm] missing single-choice campaign-timing question with >=3 options")
    return errors


def critical_email():
    """Summary email to marketing@store.com with the exact subject and a body
    that actually summarizes coupons + the planned campaigns (not an empty stub)."""
    try:
        conn = connect()
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, COALESCE(body_text,'') || ' ' || COALESCE(body_html,'') AS body
            FROM email.messages
            WHERE to_addr::text ILIKE '%marketing@store.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        return [f"[Email] error querying email: {e}"]
    if not rows:
        return ["[Email] no email found to marketing@store.com"]
    errors = []
    # Subject must match (case-insensitive) the required English subject literal.
    subj_ok = any("q2 2026 coupon campaign plan" in str(s or "").strip().lower() for s, _, _ in rows)
    if not subj_ok:
        errors.append("[Email] no email with subject 'Q2 2026 Coupon Campaign Plan'")
    # Body substance: must reference coupon/campaign performance + at least one real coupon code.
    body_ok = False
    for _, _, body in rows:
        b = str(body or "").lower()
        mentions_coupon = ("coupon" in b or "купон" in b or "campaign" in b or "кампан" in b)
        mentions_code = any(code.lower() in b for code in EXPECTED_COUPONS)
        if mentions_coupon and mentions_code and len(b.strip()) >= 40:
            body_ok = True
            break
    if not body_ok:
        errors.append("[Email] email body does not substantively summarize coupons/campaigns")
    return errors


# ----------------------------------------------------------------------------
# Scored (non-critical) checks
# ----------------------------------------------------------------------------
def scored_checks(agent_workspace):
    """Return list of (name, passed_bool) for finer-grained correctness."""
    results = []

    # Amount tolerant numeric check per coupon (formatting like '$20'/'20%' allowed).
    import openpyxl
    gt = {
        "HOLIDAY30": 30, "VIP20": 20, "SAVE20": 20, "WELCOME10": 10, "BULK10": 10,
        "ELECTRONICS15": 15, "FREESHIP": 0, "SUMMER25": 25, "FLASH50": 50, "NEWUSER5": 5,
    }
    rows, err = read_campaign_analysis(agent_workspace)
    if err or rows is None:
        for code in gt:
            results.append((f"amount:{code}", False))
    else:
        for code, exp in gt.items():
            row = rows.get(code)
            val = _extract_number(row["amount"]) if row else None
            results.append((f"amount:{code}", val is not None and num_close(val, exp, tol=0.01)))

    # Next Quarter Plan sheet present with the 4 planned campaigns.
    try:
        wb = openpyxl.load_workbook(os.path.join(agent_workspace, "Coupon_Campaign_Plan.xlsx"), data_only=True)
        plan = load_sheet_rows(wb, "Next Quarter Plan")
        names = set()
        if plan:
            for r in plan[1:]:
                if r and r[0]:
                    names.add(str(r[0]).strip().lower())
        for cn in ["holiday30 campaign", "vip20 campaign", "summer25 campaign", "flash50 flash sale"]:
            results.append((f"plan:{cn}", any(cn in n for n in names)))
    except Exception:
        for cn in range(4):
            results.append((f"plan:{cn}", False))

    return results


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    # ---- CRITICAL checks: any failure => immediate FAIL before accuracy gate ----
    critical_errors = []
    print("== CRITICAL CHECKS ==")
    for label, fn in [
        ("Excel coupon analysis + priority rule", lambda: critical_excel(agent_ws)),
        ("GCal launch events (dates/times)", critical_gcal),
        ("GForm question structure/options", critical_gform),
        ("Summary email subject + body", critical_email),
    ]:
        errs = fn()
        if errs:
            critical_errors.extend(errs)
            print(f"  CRITICAL FAIL: {label}")
            for e in errs[:5]:
                print(f"    {e}")
        else:
            print(f"  OK: {label}")

    if critical_errors:
        print(f"\n=== RESULT: FAIL (critical: {len(critical_errors)} errors) ===")
        sys.exit(1)

    # ---- Scored checks: require accuracy >= 70 ----
    print("\n== SCORED CHECKS ==")
    results = scored_checks(agent_ws)
    passed = sum(1 for _, ok in results if ok)
    total = len(results) if results else 1
    accuracy = 100.0 * passed / total
    for name, ok in results:
        if not ok:
            print(f"  miss: {name}")
    print(f"\nAccuracy: {passed}/{total} = {accuracy:.1f}%")

    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("=== RESULT: FAIL (accuracy < 70) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
