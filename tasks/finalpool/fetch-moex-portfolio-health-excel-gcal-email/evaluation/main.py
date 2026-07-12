"""Evaluation script for fetch-moex-portfolio-health-excel-gcal-email.

Структурные проверки (наличие листов, колонок, минимального числа строк,
письма и события календаря) — НЕкритические, идут в общий процент accuracy.

CRITICAL-проверки (любой провал => немедленный FAIL, sys.exit(1)):
  C1: Data_Analysis покрывает реальные тикеры портфеля MOEX
      (SBER.ME, GAZP.ME, LKOH.ME присутствуют в колонке Symbol)
      и строки отсортированы по Symbol по возрастанию.
  C2: Колонка Upside численно согласована с Target_Price - Current_Price
      для каждой строки (в пределах допуска) — читается честно, не захардкожено.
  C3: Письмо с ТОЧНОЙ темой 'Analysis Report Complete' И получателем
      team-lead@company.com.
  C4: Событие календаря с названием 'Analysis Review' 14.03.2026,
      начало 14:00, конец 15:00.
  C5: Лист Metrics: Total_Stocks равен реальному числу проанализированных
      тикеров, а Avg_Upside совпадает со средним по строкам Upside.

PASS: нет ни одного CRITICAL-фейла И accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []
CRITICAL_CHECKS = {
    "CRITICAL: Data_Analysis покрывает тикеры MOEX и отсортирован по Symbol",
    "CRITICAL: Upside = Target_Price - Current_Price по каждой строке",
    "CRITICAL: письмо с темой 'Analysis Report Complete' получателю team-lead@company.com",
    "CRITICAL: событие 'Analysis Review' 14.03.2026 14:00-15:00",
    "CRITICAL: Metrics Total_Stocks и Avg_Upside согласованы со строками",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def _norm_label(s):
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


def find_metric(metrics, aliases):
    """Tolerant lookup: the task contract never fixes literal label strings,
    so match metric rows by normalized (alnum-only) substring against aliases."""
    norm = {_norm_label(k): v for k, v in metrics.items()}
    for a in aliases:
        na = _norm_label(a)
        if na in norm:
            return norm[na]
    for a in aliases:
        na = _norm_label(a)
        for k, v in norm.items():
            if na and na in k:
                return v
    return None


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    excel_path = os.path.join(agent_workspace, "Portfolio_Health_Report.xlsx")
    check("Portfolio_Health_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # ---------------- Data_Analysis ----------------
        da_rows = []
        da_headers = []
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                       if any(c is not None for c in r)]
            check("Data_Analysis has >= 5 rows", len(da_rows) >= 5, f"got {len(da_rows)}")

            da_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Symbol', 'Name', 'Sector', 'Current_Price', 'Target_Price', 'Upside']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in da_headers, f"headers: {da_headers[:8]}")

        def col_idx(name):
            try:
                return da_headers.index(name.lower())
            except ValueError:
                return None

        i_sym = col_idx("Symbol")
        i_cur = col_idx("Current_Price")
        i_tgt = col_idx("Target_Price")
        i_up = col_idx("Upside")

        # ---- CRITICAL C1: тикеры MOEX присутствуют + сортировка по Symbol ----
        symbols = []
        if i_sym is not None:
            symbols = [str(r[i_sym]).strip().upper() for r in da_rows
                       if r[i_sym] not in (None, "")]
        required_tickers = {"SBER.ME", "GAZP.ME", "LKOH.ME"}
        have_required = required_tickers.issubset(set(symbols))
        sorted_ok = symbols == sorted(symbols)
        check("CRITICAL: Data_Analysis покрывает тикеры MOEX и отсортирован по Symbol",
              have_required and sorted_ok and len(symbols) >= 5,
              f"symbols={symbols}, have_required={have_required}, sorted={sorted_ok}")

        # ---- CRITICAL C2: Upside == Target - Current по строкам ----
        upside_consistent = (i_cur is not None and i_tgt is not None and i_up is not None
                             and len(da_rows) >= 5)
        row_upsides = {}
        if upside_consistent:
            for r in da_rows:
                cur_p = safe_float(r[i_cur]); tgt_p = safe_float(r[i_tgt]); up = safe_float(r[i_up])
                sym = str(r[i_sym]).strip().upper() if i_sym is not None else ""
                if cur_p is None or tgt_p is None or up is None:
                    upside_consistent = False
                    upside_consistent_detail = f"missing numeric in row {r}"
                    break
                expected_abs = tgt_p - cur_p
                expected_pct = (tgt_p - cur_p) / cur_p * 100 if cur_p else None
                tol = max(0.5, abs(expected_abs) * 0.02)
                tol_pct = 0.5
                ok = (abs(up - expected_abs) <= tol) or \
                     (expected_pct is not None and abs(up - expected_pct) <= tol_pct)
                if not ok:
                    upside_consistent = False
                    upside_consistent_detail = (f"{sym}: upside={up} but tgt-cur={expected_abs:.2f} "
                                                f"({expected_pct:.2f}%)")
                    break
                row_upsides[sym] = up
            else:
                upside_consistent_detail = "all rows consistent"
        else:
            upside_consistent_detail = "missing Current/Target/Upside columns or <5 rows"
        check("CRITICAL: Upside = Target_Price - Current_Price по каждой строке",
              upside_consistent, upside_consistent_detail)

        # ---------------- Metrics ----------------
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            m_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None for c in r)]
            check("Metrics has >= 3 rows", len(m_rows) >= 3, f"got {len(m_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in m_rows:
                if r[0] is not None:
                    metrics[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        # ---- CRITICAL C5: Metrics Total_Stocks и Avg_Upside согласованы ----
        n_rows = len([s for s in symbols if s])
        total_ok = False
        avg_ok = False
        total_val = safe_float(find_metric(metrics,
            ["total_stocks", "total stocks", "total stocks count",
             "number of stocks", "stocks count", "total"]))
        if total_val is not None:
            total_ok = abs(total_val - n_rows) < 0.5
        avg_val = safe_float(find_metric(metrics,
            ["avg_upside", "average upside", "avg upside", "mean upside",
             "average upside (%)", "upside avg"]))
        if avg_val is not None and row_upsides:
            mean_up = sum(row_upsides.values()) / len(row_upsides)
            avg_ok = abs(avg_val - mean_up) <= max(0.5, abs(mean_up) * 0.03)
        check("CRITICAL: Metrics Total_Stocks и Avg_Upside согласованы со строками",
              total_ok and avg_ok,
              f"Total_Stocks={total_val} (rows={n_rows}, ok={total_ok}); "
              f"Avg_Upside={avg_val} (ok={avg_ok})")

        # ---------------- Recommendations ----------------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                         if any(c is not None for c in r)]
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Symbol']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ---------------- Email (CRITICAL C3) ----------------
        def to_addresses(to_addr):
            if isinstance(to_addr, list):
                return " ".join(str(r).lower() for r in to_addr)
            if to_addr:
                try:
                    parsed = json.loads(str(to_addr))
                    if isinstance(parsed, list):
                        return " ".join(str(r).lower() for r in parsed)
                except Exception:
                    pass
                return str(to_addr).lower()
            return ""

        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT subject, to_addr FROM email.messages")
            emails = cur.fetchall()
            conn.close()
            exact = [m for m in emails
                     if m[0] and str(m[0]).strip() == "Analysis Report Complete"
                     and "team-lead@company.com" in to_addresses(m[1])]
            check("CRITICAL: письмо с темой 'Analysis Report Complete' получателю team-lead@company.com",
                  len(exact) >= 1,
                  f"subjects={[m[0] for m in emails]}")
        except Exception as e:
            check("CRITICAL: письмо с темой 'Analysis Report Complete' получателю team-lead@company.com",
                  False, str(e))

        # ---------------- Calendar (CRITICAL C4) ----------------
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("""SELECT summary, start_datetime, end_datetime FROM gcal.events
                           WHERE summary ILIKE %s""", ('%Analysis Review%',))
            events = cur.fetchall()
            conn.close()
            ev_ok = False
            for summ, start, end in events:
                if not (summ and str(summ).strip() == "Analysis Review"):
                    continue
                if start is None or end is None:
                    continue
                if (start.year == 2026 and start.month == 3 and start.day == 14
                        and start.hour == 14 and start.minute == 0
                        and end.hour == 15 and end.minute == 0):
                    ev_ok = True
                    break
            check("CRITICAL: событие 'Analysis Review' 14.03.2026 14:00-15:00",
                  ev_ok, f"events={[(e[0], str(e[1]), str(e[2])) for e in events]}")
        except Exception as e:
            check("CRITICAL: событие 'Analysis Review' 14.03.2026 14:00-15:00", False, str(e))

        check("yf_portfolio_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "yf_portfolio_processor.py")))

    # ---- gate ----
    if CRITICAL_FAILS:
        total = PASS_COUNT + FAIL_COUNT
        print(f"\nCRITICAL FAILURES: {CRITICAL_FAILS}")
        print(f"Passed {PASS_COUNT}/{total} checks (CRITICAL fail => FAIL)")
        sys.exit(1)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    success = accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
