"""Evaluation for sf-support-resolution-analysis (ClickHouse swap, russified data values)."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Russified ISSUE_TYPE keys (from db/zzz_clickhouse_after_init.sql central map).
EXPECTED_ISSUE_KEYS = {
    "проблема производительности",
    "ошибка",
    "запрос функции",
    "инцидент",
    "запрос обслуживания",
    "обслуживание",
    "техническая проблема",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}{detail_str}")
        if critical:
            CRITICAL_FAILURES.append(name)


def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=1.0):
    fa, fb = to_float(a), to_float(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= tol


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_to_lookup(rows):
    """First column (lowercased, stripped) -> row."""
    data = rows[1:] if rows and len(rows) > 1 else []
    out = {}
    for row in data:
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row
    return out


def summary_lookup(rows):
    """Orientation-agnostic metric lookup for the Summary sheet.

    task.md describes the Summary fields (Total_Issue_Types, Total_Tickets,
    Most_Common_Issue, Best_CSAT_Issue) the same way as the Issue Analysis
    columns, so a correct agent may write them either:
      (a) vertically  -> column A = metric label, column B = value, or
      (b) horizontally -> row 0 = metric labels (header), row 1 = values.

    Build the metric->[label, value] lookup from BOTH orientations and pick
    whichever yields the expected metric keys. Each value is exposed as
    row[1] so downstream checks (which read [1]) stay unchanged.
    """
    expected = {"total_issue_types", "total_tickets",
                "most_common_issue", "best_csat_issue"}

    # (a) Vertical: column A key / column B value. Only skip row 0 when it is a
    # header (its first cell is not itself an expected metric key); a valid
    # headerless layout keeps all rows so the first metric is not dropped.
    start = 0
    if rows and rows[0] and rows[0][0] is not None \
            and str(rows[0][0]).strip().lower() not in expected:
        start = 1
    vertical = {}
    for row in (rows[start:] if rows else []):
        if row and row[0] is not None and str(row[0]).strip():
            key = str(row[0]).strip().lower()
            value = row[1] if len(row) > 1 else None
            vertical[key] = [row[0], value]

    # (b) Horizontal: row 0 labels / row 1 values.
    horizontal = {}
    if rows and len(rows) >= 2:
        header, values = rows[0], rows[1]
        for i, label in enumerate(header):
            if label is not None and str(label).strip():
                key = str(label).strip().lower()
                value = values[i] if values and i < len(values) else None
                horizontal[key] = [label, value]

    v_hits = len(expected & set(vertical.keys()))
    h_hits = len(expected & set(horizontal.keys()))
    base = horizontal if h_hits > v_hits else vertical
    # Merge so any metric resolvable in either orientation is found.
    merged = dict(vertical)
    for k, val in horizontal.items():
        if k not in merged or k not in base:
            merged[k] = val
    # Prefer the winning orientation for the expected metrics.
    for k in expected:
        if k in base:
            merged[k] = base[k]
    return merged


# --------------------------------------------------------------------------- #
# Google Sheet helpers
# --------------------------------------------------------------------------- #
def get_gsheet_dashboard():
    """Return dict with the 'Support Issue Dashboard' spreadsheet's 'Issues' sheet cells,
    or None on absence. Fails gracefully if DB/schema unavailable."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:  # noqa: BLE001
        print(f"  [WARN] gsheet DB unavailable: {e}")
        return None
    try:
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets ORDER BY created_at DESC")
        spreadsheets = cur.fetchall()
        target = None
        for ss_id, ss_title in spreadsheets:
            if "support issue dashboard" in (ss_title or "").lower():
                target = (ss_id, ss_title)
                break
        if target is None:
            return {"spreadsheet": None, "issues_cells": None,
                    "all_titles": [t for _, t in spreadsheets]}
        ss_id = target[0]
        cur.execute(
            "SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s ORDER BY index",
            (ss_id,),
        )
        sheets = cur.fetchall()
        issues_sheet_id = None
        for sheet_id, sheet_title in sheets:
            if "issues" in (sheet_title or "").lower():
                issues_sheet_id = sheet_id
                break
        issues_cells = None
        if issues_sheet_id is not None:
            cur.execute(
                "SELECT row_index, col_index, value FROM gsheet.cells "
                "WHERE spreadsheet_id = %s AND sheet_id = %s ORDER BY row_index, col_index",
                (ss_id, issues_sheet_id),
            )
            issues_cells = {}
            for r, c, v in cur.fetchall():
                issues_cells.setdefault(r, {})[c] = v
        return {"spreadsheet": target, "issues_cells": issues_cells,
                "all_titles": [t for _, t in spreadsheets]}
    finally:
        conn.close()


def check_gsheet(gt_issue_keys):
    print("\n=== Checking Google Sheet ===")
    data = get_gsheet_dashboard()
    if data is None:
        # DB unreachable: cannot verify. Treat as failure (non-fatal here, CRITICAL gate below).
        check("Google Sheet 'Support Issue Dashboard' verifiable", False,
              "gsheet DB unavailable", critical=True)
        return
    ss_present = data.get("spreadsheet") is not None
    check("Google Sheet 'Support Issue Dashboard' exists", ss_present,
          f"titles found: {data.get('all_titles')}", critical=True)
    if not ss_present:
        return
    issues_cells = data.get("issues_cells")
    check("'Issues' sheet exists with data", bool(issues_cells),
          "Issues sheet missing or empty", critical=True)
    if not issues_cells:
        return
    # Collect first-column values from data rows (skip header row 0).
    sheet_keys = set()
    for r, cols in issues_cells.items():
        if r == 0:
            continue
        v = cols.get(0)
        if v is not None and str(v).strip():
            sheet_keys.add(str(v).strip().lower())
    missing = gt_issue_keys - sheet_keys
    check("'Issues' sheet covers all issue types from Issue Analysis", not missing,
          f"missing in gsheet: {sorted(missing)}", critical=True)


# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Support_Resolution_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Support_Resolution_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ----------------------------------------------------------------- #
    # Issue Analysis
    # ----------------------------------------------------------------- #
    print("=== Checking Issue Analysis ===")
    a_rows = load_sheet_rows(agent_wb, "Issue Analysis")
    g_rows = load_sheet_rows(gt_wb, "Issue Analysis")
    check("Issue Analysis sheet present in agent output", a_rows is not None,
          "sheet missing", critical=True)
    check("Issue Analysis sheet present in groundtruth", g_rows is not None, "sheet missing")

    gt_keys = set()
    g_lookup = {}
    a_lookup = {}
    if a_rows is not None and g_rows is not None:
        a_lookup = rows_to_lookup(a_rows)
        g_lookup = rows_to_lookup(g_rows)
        gt_keys = set(g_lookup.keys())

        # CRITICAL: all 7 russified ISSUE_TYPE rows present and matched exactly.
        check("Groundtruth has exactly 7 expected russified issue types",
              gt_keys == EXPECTED_ISSUE_KEYS,
              f"gt keys: {sorted(gt_keys)}")
        missing_in_agent = gt_keys - set(a_lookup.keys())
        check("Issue Analysis contains all 7 russified ISSUE_TYPE rows", not missing_in_agent,
              f"missing: {sorted(missing_in_agent)}", critical=True)

        # Per-type metrics.
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                continue
            # Tickets: deterministic aggregate -> tight tolerance.
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"{key}.Tickets matches (tol=2)",
                      num_close(a_row[1], g_row[1], 2),
                      f"{a_row[1]} vs {g_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Avg_Response_Hrs matches (tol=1.0)",
                      num_close(a_row[2], g_row[2], 1.0),
                      f"{a_row[2]} vs {g_row[2]}")
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Avg_CSAT matches (tol=0.1)",
                      num_close(a_row[3], g_row[3], 0.1),
                      f"{a_row[3]} vs {g_row[3]}")

        # CRITICAL: top-3 by Tickets are deterministic.
        for k, expected in (
            ("проблема производительности", 8128),
            ("ошибка", 6804),
            ("запрос функции", 6118),
        ):
            a_row = a_lookup.get(k)
            check(f"{k}.Tickets ~= {expected}",
                  a_row is not None and len(a_row) > 1 and num_close(a_row[1], expected, 3),
                  f"got {a_row[1] if a_row and len(a_row) > 1 else None}",
                  critical=True)

    # ----------------------------------------------------------------- #
    # Summary
    # ----------------------------------------------------------------- #
    print("\n=== Checking Summary ===")
    a_srows = load_sheet_rows(agent_wb, "Summary")
    g_srows = load_sheet_rows(gt_wb, "Summary")
    check("Summary sheet present in agent output", a_srows is not None, "sheet missing", critical=True)
    check("Summary sheet present in groundtruth", g_srows is not None, "sheet missing")

    if a_srows is not None and g_srows is not None:
        a_s = summary_lookup(a_srows)
        g_s = summary_lookup(g_srows)

        # Numeric metrics (tolerant).
        for metric in ("total_issue_types", "total_tickets"):
            if metric in g_s and metric in a_s:
                tol = 0 if metric == "total_issue_types" else 10.0
                check(f"Summary.{metric} matches (tol={tol})",
                      num_close(a_s[metric][1], g_s[metric][1], tol),
                      f"{a_s[metric][1]} vs {g_s[metric][1]}")

        # CRITICAL: Total_Issue_Types == 7.
        check("Summary Total_Issue_Types == 7",
              "total_issue_types" in a_s and num_close(a_s["total_issue_types"][1], 7, 0),
              f"got {a_s.get('total_issue_types', [None, None])[1]}",
              critical=True)

        # CRITICAL: Total_Tickets internally consistent with per-type sum AND ~31588.
        per_type_sum = None
        if a_lookup:
            s = 0
            ok = True
            for k in EXPECTED_ISSUE_KEYS:
                row = a_lookup.get(k)
                v = to_float(row[1]) if row and len(row) > 1 else None
                if v is None:
                    ok = False
                    break
                s += v
            per_type_sum = s if ok else None
        a_total = to_float(a_s["total_tickets"][1]) if "total_tickets" in a_s else None
        consistent = (
            a_total is not None and per_type_sum is not None
            and abs(a_total - per_type_sum) <= 1
            and abs(a_total - 31588) <= 10
        )
        check("Summary Total_Tickets == sum(per-type Tickets) and ~31588", consistent,
              f"total={a_total} sum={per_type_sum}", critical=True)

        # CRITICAL: Most_Common_Issue / Best_CSAT_Issue match groundtruth russified values.
        if "most_common_issue" in g_s and "most_common_issue" in a_s:
            check("Summary Most_Common_Issue matches groundtruth (russified value)",
                  str_match(a_s["most_common_issue"][1], g_s["most_common_issue"][1]),
                  f"{a_s['most_common_issue'][1]} vs {g_s['most_common_issue'][1]}",
                  critical=True)
        else:
            check("Summary Most_Common_Issue present", False, "row missing", critical=True)

        if "best_csat_issue" in g_s and "best_csat_issue" in a_s:
            check("Summary Best_CSAT_Issue matches groundtruth (russified value)",
                  str_match(a_s["best_csat_issue"][1], g_s["best_csat_issue"][1]),
                  f"{a_s['best_csat_issue'][1]} vs {g_s['best_csat_issue'][1]}",
                  critical=True)
        else:
            check("Summary Best_CSAT_Issue present", False, "row missing", critical=True)

    # ----------------------------------------------------------------- #
    # Google Sheet deliverable
    # ----------------------------------------------------------------- #
    check_gsheet(gt_keys if gt_keys else EXPECTED_ISSUE_KEYS)

    # ----------------------------------------------------------------- #
    # Verdict
    # ----------------------------------------------------------------- #
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== Score: {PASS_COUNT}/{total} ({accuracy:.1f}%) ===")

    if CRITICAL_FAILURES:
        print("\n=== RESULT: FAIL (critical) ===")
        for c in CRITICAL_FAILURES:
            print(f"  CRITICAL FAIL: {c}")
        sys.exit(1)

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
