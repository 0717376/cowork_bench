"""Evaluation for sf-hr-manager-report (ClickHouse + Teamly russified)."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_lookup(rows):
    """rows: list including header. Returns {lower(key0): row} for data rows."""
    out = {}
    for row in rows[1:] if rows and len(rows) > 1 else []:
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Manager_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Manager_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # ---------------- Structural / accuracy checks (NON-critical) ----------------
    a_mr = load_sheet_rows(agent_wb, "Manager Report")
    g_mr = load_sheet_rows(gt_wb, "Manager Report")
    if a_mr is None:
        all_errors.append("Sheet 'Manager Report' not found in agent output")
    elif g_mr is None:
        all_errors.append("Sheet 'Manager Report' not found in groundtruth")
    else:
        print("  Checking Manager Report...")
        errors = []
        a_look = rows_lookup(a_mr)
        for key, g_row in rows_lookup(g_mr).items():
            a_row = a_look.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 5):
                errors.append(f"{key}.Total: {a_row[1]} vs {g_row[1]} (tol=5)")
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 5):
                errors.append(f"{key}.High_Performers: {a_row[2]} vs {g_row[2]} (tol=5)")
            if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 5):
                errors.append(f"{key}.Low_Performers: {a_row[3]} vs {g_row[3]} (tol=5)")
            if len(a_row) > 4 and len(g_row) > 4 and not num_close(a_row[4], g_row[4], 100.0):
                errors.append(f"{key}.Avg_Salary: {a_row[4]} vs {g_row[4]} (tol=100.0)")
            if len(a_row) > 5 and len(g_row) > 5 and not num_close(a_row[5], g_row[5], 0.5):
                errors.append(f"{key}.Avg_Experience: {a_row[5]} vs {g_row[5]} (tol=0.5)")
            if len(a_row) > 6 and len(g_row) > 6 and not num_close(a_row[6], g_row[6], 1.0):
                errors.append(f"{key}.High_Perf_Pct: {a_row[6]} vs {g_row[6]} (tol=1.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    a_sum = load_sheet_rows(agent_wb, "Summary")
    g_sum = load_sheet_rows(gt_wb, "Summary")
    if a_sum is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_sum is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        print("  Checking Summary...")
        errors = []
        a_look = rows_lookup(a_sum)
        for key, g_row in rows_lookup(g_sum).items():
            a_row = a_look.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 10.0):
                errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=10.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # ---------------- CRITICAL checks (substance; any fail => sys.exit(1)) -------
    print("\n  Checking CRITICAL...")
    critical_errors = []
    a_look = rows_lookup(a_mr) if a_mr else {}
    g_look = rows_lookup(g_mr) if g_mr else {}

    # 1. All 7 russified departments present.
    EXPECTED_DEPTS = ["инженерия", "финансы", "кадры", "операции", "ниокр", "продажи", "поддержка"]
    for d in EXPECTED_DEPTS:
        if d not in a_look:
            critical_errors.append(f"CRITICAL: department row missing: {d}")

    # 2. Per-department core counts within TIGHT tolerance + High_Perf_Pct correctness.
    for key, g_row in g_look.items():
        a_row = a_look.get(key)
        if a_row is None:
            continue  # already flagged in (1)
        total = to_float(a_row[1]) if len(a_row) > 1 else None
        high = to_float(a_row[2]) if len(a_row) > 2 else None
        pct = to_float(a_row[6]) if len(a_row) > 6 else None
        if not num_close(total, g_row[1], 1):
            critical_errors.append(f"CRITICAL: {key}.Total {total} vs {g_row[1]} (tol=1)")
        if not num_close(high, g_row[2], 2):
            critical_errors.append(f"CRITICAL: {key}.High_Performers {high} vs {g_row[2]} (tol=2)")
        if not num_close(pct, g_row[6], 0.2):
            critical_errors.append(f"CRITICAL: {key}.High_Perf_Pct {pct} vs {g_row[6]} (tol=0.2)")
        # Internal consistency: pct == high/total*100 rounded to 1 decimal.
        if total and high is not None and total != 0:
            expected_pct = round(high / total * 100, 1)
            if pct is None or abs(pct - expected_pct) > 0.2:
                critical_errors.append(
                    f"CRITICAL: {key}.High_Perf_Pct inconsistent: {pct} != {expected_pct} (high/total)"
                )

    # 3. Summary Total_Employees and Total_High_Performers.
    a_slook = rows_lookup(a_sum) if a_sum else {}
    g_slook = rows_lookup(g_sum) if g_sum else {}
    te = a_slook.get("total_employees")
    thp = a_slook.get("total_high_performers")
    g_te = g_slook.get("total_employees")
    g_thp = g_slook.get("total_high_performers")
    if g_te and (te is None or not num_close(to_float(te[1]), g_te[1], 2)):
        critical_errors.append(
            f"CRITICAL: Summary.Total_Employees {te[1] if te else None} vs {g_te[1]} (tol=2)"
        )
    if g_thp and (thp is None or not num_close(to_float(thp[1]), g_thp[1], 2)):
        critical_errors.append(
            f"CRITICAL: Summary.Total_High_Performers {thp[1] if thp else None} vs {g_thp[1]} (tol=2)"
        )

    # 4. Best_Dept == department with highest High_Perf_Pct (russified name).
    bd = a_slook.get("best_dept")
    g_bd = g_slook.get("best_dept")
    if g_bd:
        gt_best = str(g_bd[1]).strip().lower() if len(g_bd) > 1 else None
        ag_best = str(bd[1]).strip().lower() if bd and len(bd) > 1 else None
        # Recompute from agent's own Manager Report rows.
        computed_best = None
        best_pct = -1.0
        for key, a_row in a_look.items():
            p = to_float(a_row[6]) if len(a_row) > 6 else None
            if p is not None and p > best_pct:
                best_pct = p
                computed_best = key
        if ag_best != gt_best:
            critical_errors.append(f"CRITICAL: Summary.Best_Dept {bd[1] if bd else None} vs {g_bd[1]}")
        elif computed_best is not None and computed_best != gt_best:
            critical_errors.append(
                f"CRITICAL: Best_Dept '{ag_best}' is not the dept with highest High_Perf_Pct (got '{computed_best}')"
            )

    if critical_errors:
        print(f"    CRITICAL ERRORS: {len(critical_errors)}")
        for e in critical_errors:
            print(f"      {e}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    print("    PASS")

    # ---------------- Accuracy gate ----------------
    # Total checkable structural points (7 dept rows + 3 summary rows = 10 logical units).
    total_checks = (len(g_look) if g_look else 0) + (len(g_slook) if g_slook else 0)
    # Each "Missing row" / value mismatch counts as one failure unit; approximate by error count.
    failed = len(all_errors)
    denom = max(total_checks, 1)
    accuracy = max(0.0, (denom - failed) / denom * 100.0)
    print(f"\n  Accuracy: {accuracy:.1f}% ({denom - failed}/{denom})")

    if all_errors:
        print(f"=== Non-critical errors: {len(all_errors)} ===")
        for e in all_errors[:10]:
            print(f"  {e}")

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL (accuracy below 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
