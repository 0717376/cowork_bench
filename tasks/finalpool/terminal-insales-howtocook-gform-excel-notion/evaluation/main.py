"""Evaluation для terminal-insales-kulinar-forms-excel-teamly (RU-стек).

Проверяет:
1. Meal_Kit_Analysis.xlsx с 4 листами (Appliance_Catalog, Recipe_Matches,
   Survey_Results, Product_Roadmap).
2. Forms (gform.*): опрос «Meal Kit Interest Survey» (источник данных) с 6 вопросами.
3. Teamly (teamly.*): пространство/трекер «Meal Kit Development Tracker» с 5
   страницами наборов.
4. Скрипты appliance_recipe_matcher.py + appliance_recipe_matches.json.

КРИТИЧЕСКИЕ чеки (CRITICAL_CHECKS): любой их провал => задача FAIL до проверки
порога accuracy. Они отражают СУТЬ: верные числа, ключевые правила, реальность
рецептов. Структурные чеки (лист есть, колонка есть) — мягкие.
Порог: accuracy >= 70% И нет критических провалов.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Канонический набор блюд kulinar (источник: local_servers/kulinar-mcp/src/data/all_recipes.json).
KULINAR_RECIPES = {
    "Бефстроганов", "Блины тонкие", "Борщ", "Ватрушки с творогом", "Винегрет",
    "Голубцы", "Греческий салат", "Гречка с тушёнкой", "Гречневая каша",
    "Грибной суп", "Грибы маринованные", "Жаркое в горшочках", "Икра кабачковая",
    "Картофель отварной с укропом", "Картофельное пюре", "Квас домашний",
    "Кисель ягодный", "Компот из сухофруктов", "Котлеты домашние", "Крабовый салат",
    "Кулебяка с капустой и яйцом", "Куриный бульон с лапшой", "Курица в сметане",
    "Медовик", "Морс клюквенный", "Наполеон", "Окрошка", "Пасха творожная",
    "Пельмени домашние", "Перловая каша", "Пирожки с капустой жареные",
    "Пирожки с мясом печёные", "Плов узбекский", "Рассольник", "Расстегаи с рыбой",
    "Рис отварной", "Рыба запечённая по-русски", "Салат Мимоза", "Салат Оливье",
    "Салат с курицей и грибами", "Сало солёное", "Сбитень", "Сельдь под шубой",
    "Селёдка с луком", "Солянка мясная", "Сырники", "Уха", "Холодец",
    "Цыплёнок табака", "Щи из квашеной капусты",
}

# 5 реальных типов приборов для линейки наборов (EN названия товаров + RU токены,
# т.к. в RU-прогоне агент может писать «Блендер/Вакууматор/Кастрюля/Весы/Вытяжка»).
APPLIANCE_TYPES = {
    "blender", "vacuum", "seal", "cooking pot", "cooker", "steamer", "pot",
    "kitchen scale", "scale", "exhaust", "fan",
    "блендер", "вакуум", "вакууматор", "кастрюл", "пароварк", "вес", "вытяжк",
}

CRITICAL_CHECKS = {
    "Survey_Results: top-answer частоты готовки = 'Several times a week' с count==14",
    "Recipe_Matches: >=10 пар, приборы валидны, рецепты реальны (kulinar)",
    "Product_Roadmap: 5 наборов, estimated_price=price+15, priority по правилу >3",
    "Teamly: трекер с 5 страницами, Status=Planning, Revenue=estimated_price*100",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_num(v):
    """Извлечь число из ячейки (допускаем '229', '229.00', '229 руб')."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", ".")
    num = ""
    seen_dot = False
    for ch in s:
        if ch.isdigit():
            num += ch
        elif ch == "." and not seen_dot and num:
            num += ch
            seen_dot = True
        elif num:
            break
    try:
        return float(num) if num else None
    except ValueError:
        return None


def norm_recipe(s):
    s = (s or "").strip().lower().replace("ё", "е")
    return " ".join(s.split())


CANON_NORM = {norm_recipe(r) for r in KULINAR_RECIPES}


def recipe_is_real(name):
    n = norm_recipe(name)
    if not n:
        return False
    return n in CANON_NORM or any(c in n or n in c for c in CANON_NORM)


def check_excel(workspace):
    print("\n=== Check 1: Meal_Kit_Analysis.xlsx ===")
    path = os.path.join(workspace, "Meal_Kit_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        # критические, зависящие от файла, тоже падают
        check("Survey_Results: top-answer частоты готовки = 'Several times a week' с count==14", False, "нет файла")
        check("Recipe_Matches: >=10 пар, приборы валидны, рецепты реальны (kulinar)", False, "нет файла")
        check("Product_Roadmap: 5 наборов, estimated_price=price+15, priority по правилу >3", False, "нет файла")
        return None
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # ---------- Appliance_Catalog ----------
    ac_idx = next((i for i, s in enumerate(sheets_lower) if "appliance" in s or "catalog" in s), 0)
    ws1 = wb[sheets[ac_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Appliance_Catalog has 8 products", len(data1) >= 8, f"Found {len(data1)}")

    headers = []
    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has price column", any("price" in h for h in headers), f"Headers: {rows1[0]}")
        check("Has avg_rating column", any("rating" in h for h in headers), f"Headers: {rows1[0]}")
        check("Has compatible_recipes_count column",
              any("recipe" in h and "count" in h for h in headers) or any("compatible" in h for h in headers),
              f"Headers: {rows1[0]}")

    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Contains Blender product", "blender" in all_text1, f"Text snippet: {all_text1[:150]}")
    check("Contains Vacuum Sealer product", "vacuum" in all_text1 or "sealing" in all_text1,
          f"Text snippet: {all_text1[:150]}")
    check("Contains Cooking Pot product", "cooking pot" in all_text1 or "cooker" in all_text1 or "steamer" in all_text1,
          f"Text snippet: {all_text1[:150]}")

    # цена Blender ~214 (мягкий якорь — читается агентом честно из InSales)
    price_col = next((i for i, h in enumerate(headers) if "price" in h and "regular" not in h), -1)
    if price_col >= 0:
        prices = [to_num(r[price_col]) for r in data1 if price_col < len(r)]
        prices = [p for p in prices if p is not None]
        check("Blender price ~214.00", any(num_close(p, 214.0, 1.0) for p in prices), f"Prices: {prices}")

    # avg_rating присутствует и в диапазоне [1,5] (мягкий, но содержательный)
    rating_col = next((i for i, h in enumerate(headers) if "rating" in h), -1)
    if rating_col >= 0:
        ratings = [to_num(r[rating_col]) for r in data1 if rating_col < len(r)]
        ratings = [x for x in ratings if x is not None]
        check("avg_rating присутствует и в [1,5]",
              len(ratings) >= 6 and all(1.0 <= x <= 5.0 for x in ratings),
              f"ratings: {ratings}")

    # ---------- Recipe_Matches ----------
    rm_idx = next((i for i, s in enumerate(sheets_lower) if "recipe" in s and "match" in s), 1)
    recipe_critical = False
    if rm_idx < len(sheets):
        ws2 = wb[sheets[rm_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Recipe_Matches has 10+ pairings", len(data2) >= 10, f"Found {len(data2)}")
        headers2 = [str(c).lower() if c else "" for c in rows2[0]] if rows2 else []
        check("Has matched_appliance column",
              any("appliance" in h or "matched" in h for h in headers2), f"Headers: {headers2}")
        check("Has difficulty column", any("difficult" in h for h in headers2), f"Headers: {headers2}")

        recipe_col = next((i for i, h in enumerate(headers2) if "recipe" in h), 0)
        appl_col = next((i for i, h in enumerate(headers2) if "appliance" in h or "matched" in h), 1)
        recipe_names = [r[recipe_col] for r in data2 if recipe_col < len(r) and r[recipe_col]]
        appls = [str(r[appl_col]).lower() for r in data2 if appl_col < len(r) and r[appl_col]]
        real_count = sum(1 for n in recipe_names if recipe_is_real(n))
        appl_ok = all(any(t in a for t in APPLIANCE_TYPES) for a in appls) and len(appls) > 0
        unknown = [n for n in recipe_names if not recipe_is_real(n)][:5]
        recipe_critical = (len(data2) >= 10 and appl_ok and recipe_names and
                           real_count == len(recipe_names))
        check("Recipe_Matches: >=10 пар, приборы валидны, рецепты реальны (kulinar)",
              recipe_critical,
              f"pairs={len(data2)}, real={real_count}/{len(recipe_names)}, appl_ok={appl_ok}, unknown={unknown}")

    # ---------- Survey_Results ----------
    sr_idx = next((i for i, s in enumerate(sheets_lower) if "survey" in s and "result" in s), 2)
    if sr_idx < len(sheets):
        ws3 = wb[sheets[sr_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Survey_Results has 6 rows", len(data3) >= 6, f"Found {len(data3)}")
        headers3 = [str(c).lower() if c else "" for c in rows3[0]] if rows3 else []
        check("Has top_answer column", any("top" in h or "answer" in h for h in headers3), f"Headers: {headers3}")
        check("Has response_count column", any("count" in h or "response" in h for h in headers3), f"Headers: {headers3}")

        # CRITICAL: строка частоты готовки -> top='Several times a week', count==12
        q_col = next((i for i, h in enumerate(headers3) if "question" in h), 0)
        ans_col = next((i for i, h in enumerate(headers3) if ("top" in h or "answer" in h)), 1)
        cnt_col = next((i for i, h in enumerate(headers3) if ("count" in h or "response" in h)), 2)
        freq_row = None
        for r in data3:
            q = str(r[q_col]).lower() if q_col < len(r) and r[q_col] else ""
            if "cook" in q or "often" in q:
                freq_row = r
                break
        top_ok = cnt_ok = False
        if freq_row is not None:
            top_val = str(freq_row[ans_col]).lower() if ans_col < len(freq_row) and freq_row[ans_col] else ""
            cnt_val = to_num(freq_row[cnt_col]) if cnt_col < len(freq_row) else None
            top_ok = "several times a week" in top_val
            cnt_ok = cnt_val is not None and num_close(cnt_val, 14, 0)
        check("Survey_Results: top-answer частоты готовки = 'Several times a week' с count==14",
              top_ok and cnt_ok,
              f"freq_row={freq_row}")
    else:
        check("Survey_Results: top-answer частоты готовки = 'Several times a week' с count==14",
              False, "нет листа Survey_Results")

    # ---------- Product_Roadmap ----------
    pr_idx = next((i for i, s in enumerate(sheets_lower) if "roadmap" in s or "product" in s), 3)
    roadmap = {}  # kit_name(lower) -> dict(price, count, priority)
    if pr_idx < len(sheets):
        ws4 = wb[sheets[pr_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Product_Roadmap has 5 kits", len(data4) >= 5, f"Found {len(data4)}")
        headers4 = [str(c).lower() if c else "" for c in rows4[0]] if rows4 else []
        check("Has priority column", any("priority" in h for h in headers4), f"Headers: {headers4}")
        check("Has estimated_price column", any("price" in h or "estimated" in h for h in headers4),
              f"Headers: {headers4}")

        name_col = next((i for i, h in enumerate(headers4) if "kit" in h and "name" in h),
                        next((i for i, h in enumerate(headers4) if "kit" in h), 0))
        cnt_col4 = next((i for i, h in enumerate(headers4) if "recipe" in h and "count" in h),
                        next((i for i, h in enumerate(headers4) if "count" in h), -1))
        ep_col = next((i for i, h in enumerate(headers4) if "estimated" in h or "price" in h), -1)
        prio_col = next((i for i, h in enumerate(headers4) if "priority" in h), -1)

        # CRITICAL: priority по правилу (High iff recipe_count>3); estimated_price>15
        rule_ok = len(data4) >= 5
        for r in data4:
            cnt = to_num(r[cnt_col4]) if 0 <= cnt_col4 < len(r) else None
            ep = to_num(r[ep_col]) if 0 <= ep_col < len(r) else None
            prio = str(r[prio_col]).strip().lower() if 0 <= prio_col < len(r) and r[prio_col] else ""
            kname = str(r[name_col]).strip().lower() if 0 <= name_col < len(r) and r[name_col] else ""
            if ep is not None and kname:
                roadmap[kname] = {"price": ep, "count": cnt, "priority": prio}
            # estimated_price = appliance_price + 15  =>  ep должна быть > 15
            if ep is None or ep <= 15:
                rule_ok = False
            # правило приоритета
            if cnt is not None:
                expected = "high" if cnt > 3 else "medium"
                if prio not in (expected,):
                    rule_ok = False
        check("Product_Roadmap: 5 наборов, estimated_price=price+15, priority по правилу >3",
              rule_ok,
              f"rows={len(data4)}, roadmap={roadmap}")

        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has kit names", "kit" in all_text4, f"Text: {all_text4[:200]}")
    else:
        check("Product_Roadmap: 5 наборов, estimated_price=price+15, priority по правилу >3",
              False, "нет листа Product_Roadmap")

    return roadmap


def check_gform():
    print("\n=== Check 2: Forms (источник опроса) ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        meal_form = None
        for form_id, title in forms:
            if title and ("meal kit" in title.lower() or "meal_kit" in title.lower()):
                meal_form = (form_id, title)
                break
        check("Meal Kit Interest Survey form exists", meal_form is not None,
              f"Forms: {[f[1] for f in forms]}")

        if meal_form:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (meal_form[0],))
            q_count = cur.fetchone()[0]
            check("Survey has 6 questions", q_count >= 6, f"Found {q_count}")

            cur.execute("SELECT COUNT(*) FROM gform.responses WHERE form_id = %s", (meal_form[0],))
            r_count = cur.fetchone()[0]
            check("Survey has responses (>=20)", r_count >= 20, f"Found {r_count}")
    except Exception as e:
        check("Gform check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_teamly(roadmap):
    print("\n=== Check 3: Teamly Meal Kit Development Tracker ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # Трекер может быть смоделирован как пространство (space) c 5 страницами,
        # либо как родительская страница с 5 дочерними. Берём пространство по имени.
        cur.execute("SELECT id, key, name FROM teamly.spaces")
        spaces = cur.fetchall()
        tracker_space = None
        for sid, key, name in spaces:
            nm = (name or "").lower()
            if "meal kit" in nm and ("tracker" in nm or "development" in nm or "трекер" in nm):
                tracker_space = (sid, name)
                break
        check("Meal Kit Development Tracker создан (teamly space)",
              tracker_space is not None, f"Spaces: {[s[2] for s in spaces]}")

        pages = []
        if tracker_space:
            cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages WHERE space_id = %s",
                        (tracker_space[0],))
            pages = cur.fetchall()
        else:
            # запасной вариант: страницы трекера в любом пространстве по заголовку
            cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
            allp = cur.fetchall()
            pages = [(t, b) for t, b in allp
                     if "meal kit" in (t or "").lower() or "набор" in (t or "").lower()]

        check("Трекер содержит 5 страниц наборов", len(pages) >= 5, f"Found {len(pages)}")

        # Содержимое страниц: Status=Planning, Priority, Estimated_Revenue
        bodies = [(t or "") + "\n" + (b or "") for t, b in pages]
        joined = "\n".join(bodies).lower()
        all_planning = len(pages) >= 5 and all("planning" in s.lower() for s in bodies)
        check("Все страницы Status=Planning", all_planning,
              f"pages={len(pages)}")
        check("Страницы упоминают Priority", "priority" in joined or "приоритет" in joined,
              f"text: {joined[:150]}")

        # CRITICAL: 5 страниц, Status=Planning, Estimated_Revenue == estimated_price*100.
        # Сверяем каждую страницу с roadmap по числам: на странице есть число revenue,
        # равное какому-либо estimated_price*100 из roadmap.
        expected_revenues = {round(v["price"] * 100) for v in (roadmap or {}).values()
                             if v.get("price") is not None}
        revenue_ok = False
        if expected_revenues:
            matched = 0
            for s in bodies:
                digits = _all_numbers(s)
                if any(any(num_close(d, er, 0) for d in digits) for er in expected_revenues):
                    matched += 1
            revenue_ok = matched >= 5
        else:
            # нет roadmap-цен (Excel сломан) — критический чек считаем проваленным
            revenue_ok = False
        check("Teamly: трекер с 5 страницами, Status=Planning, Revenue=estimated_price*100",
              len(pages) >= 5 and all_planning and revenue_ok,
              f"pages={len(pages)}, planning={all_planning}, revenue_ok={revenue_ok}, "
              f"expected_revenues={sorted(expected_revenues)[:5]}")
    except Exception as e:
        check("Teamly check", False, str(e))
        check("Teamly: трекер с 5 страницами, Status=Planning, Revenue=estimated_price*100",
              False, str(e))
    finally:
        cur.close()
        conn.close()


def _all_numbers(s):
    # Убираем разделители групп разрядов (пробел/тонкий пробел/запятая) между
    # группами цифр, чтобы "22 900"/"22,900"/"22 900" парсились как 22900.
    # Натурально для RU-текста ("Estimated_Revenue: 22 900").
    s = re.sub(r"(?<=\d)[\s,   ](?=\d)", "", s)
    out = []
    cur = ""
    seen_dot = False
    for ch in s:
        if ch.isdigit():
            cur += ch
        elif ch == "." and cur and not seen_dot:
            cur += ch
            seen_dot = True
        else:
            if cur:
                try:
                    out.append(float(cur))
                except ValueError:
                    pass
            cur = ""
            seen_dot = False
    if cur:
        try:
            out.append(float(cur))
        except ValueError:
            pass
    return out


def check_scripts(workspace):
    print("\n=== Check 4: Scripts ===")
    matcher = os.path.join(workspace, "appliance_recipe_matcher.py")
    check("appliance_recipe_matcher.py exists", os.path.exists(matcher))

    matches_json = os.path.join(workspace, "appliance_recipe_matches.json")
    check("appliance_recipe_matches.json exists", os.path.exists(matches_json))
    if os.path.exists(matches_json):
        with open(matches_json) as f:
            data = json.load(f)
        check("JSON has appliance keys", len(data) >= 3, f"Keys: {list(data.keys())[:5]}")
        total_matches = sum(len(v) if isinstance(v, list) else 0 for v in data.values())
        check("JSON has 10+ total matches", total_matches >= 10, f"Total: {total_matches}")


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Meal_Kit_Analysis.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        check("No negative values in Excel", False,
                              f"Found {cell} in sheet {sheet_name}")
                        return
        check("No negative values in Excel", True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    roadmap = check_excel(args.agent_workspace)
    check_gform()
    check_teamly(roadmap)
    check_scripts(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_failed": CRITICAL_FAILED}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAIL: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
