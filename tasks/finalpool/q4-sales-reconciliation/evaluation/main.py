"""
Evaluation for q4-sales-reconciliation task (ClickHouse fork).
Compares agent's Q4_2025_Sales_Report.xlsx and Executive_Summary.docx against
groundtruth. After the snowflake->clickhouse swap the warehouse returns Russian
REGION/SEGMENT values, so the groundtruth cells and the agent output are both in
Russian; matching is done by exact (case-insensitive) region/segment strings.

Critical semantic checks (CRITICAL_CHECKS) run first; ANY critical failure =>
sys.exit(1) before the all-or-nothing accuracy gate.
"""
import argparse
import os
import re
import sys

import openpyxl

# Central English->Russian relabel map: region labels used for RU+EN matching
# in the free-text executive summary. Single source of truth.
_SCRIPTS = os.path.abspath(os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "..", "scripts"))
if _SCRIPTS not in sys.path:
    sys.path.insert(0, _SCRIPTS)
from clickhouse_relabel_map import REGIONS  # noqa: E402

# Frozen groundtruth actuals (total delivered-order revenue) per English region
# key — used by the CRITICAL semantic gate independent of the xlsx file.
GT_ACTUALS = {
    "Asia Pacific": 70510.11,
    "Europe": 54490.62,
    "Latin America": 57100.57,
    "Middle East": 57505.34,
    "North America": 51818.56,
}
GT_TARGETS = {
    "Asia Pacific": 65000, "Europe": 60000, "Latin America": 55000,
    "Middle East": 50000, "North America": 55000,
}
GT_TOTAL_ACTUAL = round(sum(GT_ACTUALS.values()), 2)  # ~291425.20
# Regions that fell short of target (Actual <= Target).
MISSED_REGIONS = [r for r in GT_ACTUALS if GT_ACTUALS[r] <= GT_TARGETS[r]]  # Europe, North America


def load_sheet_data(wb, sheet_name):
    """Load all rows from a sheet as list of lists. Case-insensitive sheet lookup."""
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    ws = wb[matched]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_regional_performance(agent_rows, gt_rows):
    """Check Sheet 1: Regional Performance."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Regional Performance' not found"]

    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    # Build lookup by region name (case-insensitive)
    agent_by_region = {}
    for row in agent_data:
        if row[0]:
            agent_by_region[str(row[0]).strip().lower()] = row

    for g_row in gt_data:
        region = str(g_row[0]).strip()
        region_key = region.lower()

        if region_key not in agent_by_region:
            errors.append(f"Missing region: {region}")
            continue

        a_row = agent_by_region[region_key]

        # Column mapping: Region, Target, Actual, Variance, Variance_Pct, Order_Count, Customer_Count
        field_names = ["Target", "Actual", "Variance", "Variance_Pct", "Order_Count", "Customer_Count"]
        tolerances = [1.0, 1.0, 1.0, 0.5, 0.5, 0.5]

        for j, (fname, tol) in enumerate(zip(field_names, tolerances), start=1):
            try:
                a_val = float(a_row[j]) if a_row[j] is not None else None
                g_val = float(g_row[j])
                if a_val is None:
                    errors.append(f"{region}.{fname}: missing value")
                elif abs(a_val - g_val) > tol:
                    errors.append(f"{region}.{fname}: {a_val} vs expected {g_val} (tol={tol})")
            except (ValueError, TypeError):
                errors.append(f"{region}.{fname}: invalid value '{a_row[j]}'")

    return len(errors) == 0, errors


def check_segment_breakdown(agent_rows, gt_rows):
    """Check Sheet 2: Segment Breakdown."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Segment Breakdown' not found"]

    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    # Build lookup by (region, segment) - case insensitive
    agent_lookup = {}
    for row in agent_data:
        if row[0] and row[1]:
            key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
            agent_lookup[key] = row

    for g_row in gt_data:
        region = str(g_row[0]).strip()
        segment = str(g_row[1]).strip()
        key = (region.lower(), segment.lower())

        if key not in agent_lookup:
            errors.append(f"Missing region-segment: {region}/{segment}")
            continue

        a_row = agent_lookup[key]

        # Check Revenue (col 2) with tolerance 5.0
        try:
            a_rev = float(a_row[2]) if a_row[2] is not None else None
            g_rev = float(g_row[2])
            if a_rev is None:
                errors.append(f"{region}/{segment}.Revenue: missing")
            elif abs(a_rev - g_rev) > 5.0:
                errors.append(f"{region}/{segment}.Revenue: {a_rev} vs expected {g_rev}")
        except (ValueError, TypeError):
            errors.append(f"{region}/{segment}.Revenue: invalid '{a_row[2]}'")

        # Check Orders (col 3) with tolerance 2
        try:
            a_ord = float(a_row[3]) if a_row[3] is not None else None
            g_ord = float(g_row[3])
            if a_ord is None:
                errors.append(f"{region}/{segment}.Orders: missing")
            elif abs(a_ord - g_ord) > 2:
                errors.append(f"{region}/{segment}.Orders: {a_ord} vs expected {g_ord}")
        except (ValueError, TypeError):
            errors.append(f"{region}/{segment}.Orders: invalid '{a_row[3]}'")

    return len(errors) == 0, errors


def _region_present(text_lower, en_region):
    """True if either the English or the Russian label of a region appears."""
    ru = REGIONS.get(en_region, "")
    return en_region.lower() in text_lower or (ru and ru.lower() in text_lower)


def _mentions_total(text):
    """True if the digit-normalized text contains the total actual revenue.

    Accepts '291,425' / '291 425' / '291425' (with or without the .20 cents).
    """
    digits = re.sub(r"[\s, ]", "", text)
    return ("291425" in digits)


def check_executive_summary(agent_workspace):
    """Check Executive_Summary.docx exists and contains the key semantic facts.

    Hardened: asserts the total actual revenue (~291425.20) is mentioned and the
    two under-target regions (Europe AND North America) are named (RU or EN).
    """
    errors = []
    docx_path = os.path.join(agent_workspace, "Executive_Summary.docx")

    if not os.path.exists(docx_path):
        return False, ["Executive_Summary.docx not found"]

    try:
        from docx import Document
        doc = Document(docx_path)
        raw_text = " ".join([p.text for p in doc.paragraphs])
        full_text = raw_text.lower()

        if len(full_text.strip()) < 50:
            errors.append("Executive summary is too short (less than 50 chars)")

        # At least one region named (RU or EN) — sanity that it discusses regions.
        if not any(_region_present(full_text, r) for r in GT_ACTUALS):
            errors.append("Executive summary mentions no recognizable region")

        # Total actual revenue must be present (~291,425).
        if not _mentions_total(raw_text):
            errors.append("Executive summary does not mention total actual revenue (~291,425)")

        # Both under-target regions must be named (RU or EN).
        for r in MISSED_REGIONS:
            if not _region_present(full_text, r):
                errors.append(
                    f"Executive summary does not identify under-target region "
                    f"'{r}' / '{REGIONS.get(r, '')}'")

    except ImportError:
        # If python-docx not available, just check file exists and has size.
        file_size = os.path.getsize(docx_path)
        if file_size < 100:
            errors.append("Executive_Summary.docx is suspiciously small")
    except Exception as e:
        errors.append(f"Error reading Executive_Summary.docx: {e}")

    return len(errors) == 0, errors


# ── CRITICAL semantic checks ────────────────────────────────────────────────
# Any failure here => sys.exit(1) before the all-or-nothing accuracy gate.
# These assert the SUBSTANCE: that the agent applied the delivered-only filter,
# joined PDF targets to warehouse actuals, and derived variances correctly.

def _rp_lookup(agent_rp_rows):
    """Map Russian region label (lower) -> row from the agent Regional Performance sheet."""
    out = {}
    for row in (agent_rp_rows or [])[1:]:
        if row and row[0]:
            out[str(row[0]).strip().lower()] = row
    return out


def critical_regional_actuals(agent_rp_rows):
    """CRITICAL: each region's Actual matches frozen groundtruth (tol 1.0).

    Region keyed by the central Russian label. Verifies delivered-only revenue.
    """
    errs = []
    lut = _rp_lookup(agent_rp_rows)
    for en, expected in GT_ACTUALS.items():
        ru = REGIONS[en].lower()
        row = lut.get(ru)
        if row is None:
            errs.append(f"Actual: missing region '{REGIONS[en]}'")
            continue
        try:
            actual = float(row[2])
        except (TypeError, ValueError, IndexError):
            errs.append(f"Actual[{REGIONS[en]}]: not numeric")
            continue
        if abs(actual - expected) > 1.0:
            errs.append(f"Actual[{REGIONS[en]}]: {actual} vs expected {expected}")
    return len(errs) == 0, errs


def critical_variance_derivation(agent_rp_rows):
    """CRITICAL: Variance==Actual-Target and Variance_Pct==Variance/Target*100.

    Confirms PDF targets were joined onto warehouse actuals. Also checks the two
    diagnostic signs: Europe negative, Middle East positive.
    """
    errs = []
    lut = _rp_lookup(agent_rp_rows)
    for en in GT_ACTUALS:
        ru = REGIONS[en].lower()
        row = lut.get(ru)
        if row is None:
            errs.append(f"Variance: missing region '{REGIONS[en]}'")
            continue
        try:
            target = float(row[1]); actual = float(row[2])
            variance = float(row[3]); vpct = float(row[4])
        except (TypeError, ValueError, IndexError):
            errs.append(f"Variance[{REGIONS[en]}]: non-numeric cells")
            continue
        if abs(variance - (actual - target)) > 1.0:
            errs.append(f"Variance[{REGIONS[en]}]: {variance} != Actual-Target ({actual - target:.2f})")
        if target:
            exp_pct = (actual - target) / target * 100
            if abs(vpct - exp_pct) > 0.5:
                errs.append(f"Variance_Pct[{REGIONS[en]}]: {vpct} != {exp_pct:.1f}")
    # Diagnostic signs.
    eu = lut.get(REGIONS["Europe"].lower())
    me = lut.get(REGIONS["Middle East"].lower())
    try:
        if eu is not None and float(eu[3]) >= 0:
            errs.append("Variance[Европа] should be negative (missed target)")
    except (TypeError, ValueError, IndexError):
        pass
    try:
        if me is not None and float(me[3]) <= 0:
            errs.append("Variance[Ближний Восток] should be positive (beat target)")
    except (TypeError, ValueError, IndexError):
        pass
    return len(errs) == 0, errs


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

    agent_excel = os.path.join(args.agent_workspace, "Q4_2025_Sales_Report.xlsx") if args.agent_workspace else None
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_excel = os.path.join(gt_dir, "Q4_2025_Sales_Report.xlsx")

    if not agent_excel or not os.path.exists(agent_excel):
        print(f"FAIL: Agent output file not found: {agent_excel}")
        sys.exit(1)

    if not os.path.exists(gt_excel):
        print(f"FAIL: Groundtruth file not found: {gt_excel}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_excel, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_excel, data_only=True)

    # Check sheet count
    print(f"Agent sheets: {agent_wb.sheetnames}")
    print(f"Groundtruth sheets: {gt_wb.sheetnames}")

    # Check 1: Regional Performance
    print("\n[1/3] Checking Regional Performance ...")
    a_rp = load_sheet_data(agent_wb, "Regional Performance")
    g_rp = load_sheet_data(gt_wb, "Regional Performance")
    rp_ok, rp_errors = check_regional_performance(a_rp, g_rp)
    if rp_ok:
        print("  PASS")
    else:
        for e in rp_errors:
            print(f"  ERROR: {e}")

    # Check 2: Segment Breakdown
    print("[2/3] Checking Segment Breakdown ...")
    a_sb = load_sheet_data(agent_wb, "Segment Breakdown")
    g_sb = load_sheet_data(gt_wb, "Segment Breakdown")
    sb_ok, sb_errors = check_segment_breakdown(a_sb, g_sb)
    if sb_ok:
        print("  PASS")
    else:
        for e in sb_errors:
            print(f"  ERROR: {e}")

    # Check 3: Executive Summary
    print("[3/3] Checking Executive Summary ...")
    es_ok, es_errors = check_executive_summary(args.agent_workspace)
    if es_ok:
        print("  PASS")
    else:
        for e in es_errors:
            print(f"  ERROR: {e}")

    # ── CRITICAL semantic gate (runs BEFORE the accuracy gate) ──────────────
    # Any critical failure => immediate FAIL (sys.exit(1)), regardless of the
    # all-or-nothing structural result below.
    print("\n[CRITICAL] Semantic checks ...")
    critical = {}
    critical["regional_actuals"] = critical_regional_actuals(a_rp)
    critical["variance_derivation"] = critical_variance_derivation(a_rp)
    # Segment breakdown correctness (delivered-only filter + per-cell values +
    # exact row count) is itself a critical semantic requirement.
    critical["segment_breakdown"] = (sb_ok, sb_errors)
    # Executive summary must state the total and name both missed regions.
    critical["executive_summary"] = (es_ok, es_errors)

    critical_failed = []
    for name, (ok, errs) in critical.items():
        if ok:
            print(f"  CRITICAL PASS: {name}")
        else:
            critical_failed.append(name)
            print(f"  CRITICAL FAIL: {name}")
            for e in errs:
                print(f"    - {e}")

    if critical_failed:
        print(f"\n=== RESULT: FAIL (critical) ===")
        print(f"CRITICAL FAILURES: {', '.join(critical_failed)}")
        sys.exit(1)

    overall = rp_ok and sb_ok and es_ok
    print(f"\n=== RESULT: {'PASS' if overall else 'FAIL'} ===")
    print(f"Regional Performance: {'PASS' if rp_ok else 'FAIL'}")
    print(f"Segment Breakdown: {'PASS' if sb_ok else 'FAIL'}")
    print(f"Executive Summary: {'PASS' if es_ok else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
