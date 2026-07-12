"""Evaluation for moex-stock-performance-email (RU swap of yf-stock-performance-email)."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

# Derived from the seeded moex.stock_prices boundary closes:
#   prior quarter = 2026-02-25 close, latest = 2026-05-26 close.
# YoY_Return_Pct = (latest - prior) / prior * 100, rounded 2dp.
# Verdict: >10 Outperform, [-10,10] Neutral, <-10 Underperform.
EXPECTED_PERF = {
    "GAZP.ME": {"company": "Газпром",  "price_1y": 190.9656, "latest": 208.6306, "yoy": 9.25,  "verdict": "Neutral"},
    "LKOH.ME": {"company": "ЛУКОЙЛ",   "price_1y": 3789.7453, "latest": 3727.0965, "yoy": -1.65, "verdict": "Neutral"},
    "MGNT.ME": {"company": "Магнит",   "price_1y": 4257.8235, "latest": 4182.1923, "yoy": -1.78, "verdict": "Neutral"},
    "MTSS.ME": {"company": "МТС",      "price_1y": 264.5709, "latest": 255.0623, "yoy": -3.59, "verdict": "Neutral"},
    "SBER.ME": {"company": "Сбербанк", "price_1y": 127.3413, "latest": 120.8171, "yoy": -5.12, "verdict": "Neutral"},
}

EXPECTED_SUMMARY = {
    "Best_Performer": "GAZP.ME",
    "Worst_Performer": "SBER.ME",
    "Avg_YoY_Return": -0.58,
    "Outperform_Count": 0,
    "Underperform_Count": 0,
}


def verdict_for(yoy):
    if yoy > 10:
        return "Outperform"
    if yoy < -10:
        return "Underperform"
    return "Neutral"


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.lower() in str(haystack).lower()


def str_match(a, b):
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_dir = args.agent_workspace or os.path.join(task_root, "initial_workspace")

    agent_file = os.path.join(agent_dir, "YoY_Stock_Performance.xlsx")
    gt_file = os.path.join(gt_dir, "YoY_Stock_Performance.xlsx")

    file_errors = []
    db_errors = []
    critical_failures = []

    if not os.path.exists(agent_file):
        file_errors.append(f"Agent output not found: {agent_file}")
        critical_failures.append("Agent workbook YoY_Stock_Performance.xlsx not produced")
    if not os.path.exists(gt_file):
        file_errors.append(f"Groundtruth not found: {gt_file}")

    # Holds the agent Performance rows keyed by symbol for semantic checks.
    agent_perf = {}
    agent_summary = {}

    if os.path.exists(agent_file) and os.path.exists(gt_file):
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Performance sheet
        print("  Checking Performance...")
        a_rows = load_sheet_rows(agent_wb, "Performance")
        g_rows = load_sheet_rows(gt_wb, "Performance")
        if a_rows is None:
            file_errors.append("Sheet 'Performance' not found in agent output")
            critical_failures.append("Performance sheet missing")
        elif g_rows is None:
            file_errors.append("Sheet 'Performance' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            if len(a_data) != len(g_data):
                file_errors.append(f"Performance row count: agent {len(a_data)} vs gt {len(g_data)}")

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    key = str(row[0]).strip().upper()
                    a_lookup[key] = row
                    agent_perf[key] = row

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().upper()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing stock: {key}")
                    continue

                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 2.0):
                        file_errors.append(f"{key} Price_1Y_Ago: {a_row[2]} vs {g_row[2]}")
                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 2.0):
                        file_errors.append(f"{key} Latest_Price: {a_row[3]} vs {g_row[3]}")
                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 2.0):
                        file_errors.append(f"{key} YoY_Return: {a_row[4]} vs {g_row[4]}")
                if len(a_row) > 5 and len(g_row) > 5:
                    if not str_match(a_row[5], g_row[5]):
                        file_errors.append(f"{key} Verdict: {a_row[5]} vs {g_row[5]}")

        # Check Summary sheet
        print("  Checking Summary...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            file_errors.append("Sheet 'Summary' not found in agent output")
            critical_failures.append("Summary sheet missing")
        elif g_rows is None:
            file_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row
                    agent_summary[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing summary metric: {g_row[0]}")
                    continue
                if len(a_row) > 1 and len(g_row) > 1:
                    g_val = g_row[1]
                    a_val = a_row[1]
                    try:
                        fa, fb = float(a_val), float(g_val)
                        if abs(fa - fb) > 2.0:
                            file_errors.append(f"Summary {key}: {a_val} vs {g_val}")
                    except (TypeError, ValueError):
                        if not str_match(a_val, g_val):
                            file_errors.append(f"Summary {key}: '{a_val}' vs '{g_val}'")

    # ===== CRITICAL semantic checks (substance, derived from seeded source) =====
    # 1. Every MOEX symbol present with correct YoY_Return_Pct (recomputed from seeded prices).
    if agent_perf:
        for sym, exp in EXPECTED_PERF.items():
            row = agent_perf.get(sym.upper())
            if row is None:
                critical_failures.append(f"Performance row for {sym} missing")
                continue
            if len(row) <= 4 or not num_close(row[4], exp["yoy"], 2.0):
                got = row[4] if len(row) > 4 else None
                critical_failures.append(f"{sym} YoY_Return_Pct wrong: got {got}, expected ~{exp['yoy']}")
            # 2. Verdict must match the threshold rule applied to the agent's own YoY value.
            got_yoy = to_float(row[4]) if len(row) > 4 else None
            got_verdict = str(row[5]).strip() if len(row) > 5 and row[5] is not None else None
            if got_yoy is None or got_verdict is None:
                critical_failures.append(f"{sym} verdict/return cell missing")
            else:
                rule = verdict_for(got_yoy)
                if got_verdict.lower() != rule.lower():
                    critical_failures.append(
                        f"{sym} Verdict '{got_verdict}' inconsistent with YoY {got_yoy} (rule -> {rule})")
    else:
        critical_failures.append("No Performance rows parsed from agent output")

    # 3 & 4. Summary Best/Worst/Avg/Counts must match the agent's own Performance sheet.
    if agent_perf:
        derived = []
        for sym, row in agent_perf.items():
            y = to_float(row[4]) if len(row) > 4 else None
            if y is not None:
                derived.append((sym, y))
        if len(derived) >= 2:
            best_sym = max(derived, key=lambda x: x[1])[0]
            worst_sym = min(derived, key=lambda x: x[1])[0]
            avg = round(sum(y for _, y in derived) / len(derived), 2)
            oc = sum(1 for _, y in derived if y > 10)
            uc = sum(1 for _, y in derived if y < -10)

            a_best = str(agent_summary.get("best_performer", "")).strip().upper()
            a_worst = str(agent_summary.get("worst_performer", "")).strip().upper()
            if a_best != best_sym.upper():
                critical_failures.append(f"Best_Performer '{a_best}' != actual max {best_sym}")
            if a_worst != worst_sym.upper():
                critical_failures.append(f"Worst_Performer '{a_worst}' != actual min {worst_sym}")

            a_avg = to_float(agent_summary.get("avg_yoy_return"))
            if a_avg is None or abs(a_avg - avg) > 0.1:
                critical_failures.append(f"Avg_YoY_Return {agent_summary.get('avg_yoy_return')} != {avg}")
            a_oc = to_float(agent_summary.get("outperform_count"))
            a_uc = to_float(agent_summary.get("underperform_count"))
            if a_oc is None or int(a_oc) != oc:
                critical_failures.append(f"Outperform_Count {agent_summary.get('outperform_count')} != {oc}")
            if a_uc is None or int(a_uc) != uc:
                critical_failures.append(f"Underperform_Count {agent_summary.get('underperform_count')} != {uc}")
        else:
            critical_failures.append("Could not derive summary from Performance rows")

    # 5. Email to investment-committee@fund.com with correct subject AND body naming best/worst (BLOCKING).
    print("  Checking email...")
    email_ok = False
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%investment-committee@fund.com%'
        """)
        email_rows = cur.fetchall()
        if not email_rows:
            cur.execute("SELECT COUNT(*) FROM email.messages")
            total = cur.fetchone()[0]
            db_errors.append(f"No email to investment-committee@fund.com found (total: {total})")
        else:
            best = EXPECTED_SUMMARY["Best_Performer"]
            worst = EXPECTED_SUMMARY["Worst_Performer"]
            for subj, to_addr, body in email_rows:
                subj_ok = str_contains(subj, "YoY Stock Performance Report")
                body = body or ""
                # Body must name actual best and worst performer symbols (with or without .ME suffix).
                best_ok = str_contains(body, best) or str_contains(body, best.split(".")[0])
                worst_ok = str_contains(body, worst) or str_contains(body, worst.split(".")[0])
                if subj_ok and best_ok and worst_ok:
                    email_ok = True
                    break
            if not email_ok:
                db_errors.append(
                    "Email found but subject/body invalid: need subject 'YoY Stock Performance Report' "
                    f"and body naming best {best} and worst {worst}")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Email check error: {e}")

    if not email_ok:
        critical_failures.append("Required report email (subject + best/worst symbols in body) missing")

    # ===== Final result =====
    print(f"\n=== SUMMARY ===")
    print(f"  File errors:        {len(file_errors)}")
    print(f"  DB errors:          {len(db_errors)}")
    print(f"  Critical failures:  {len(critical_failures)}")
    if db_errors:
        for e in db_errors[:15]:
            print(f"    [DB] {e}")
    if file_errors:
        for e in file_errors[:15]:
            print(f"    [FILE] {e}")
    if critical_failures:
        for e in critical_failures[:15]:
            print(f"    [CRITICAL] {e}")
        print(f"  Overall: FAIL (critical check failed)")
        sys.exit(1)

    if file_errors:
        print(f"  Overall: FAIL")
        sys.exit(1)

    print(f"  Overall: PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
