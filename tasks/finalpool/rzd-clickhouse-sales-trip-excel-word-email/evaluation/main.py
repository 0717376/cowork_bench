"""
Evaluation for train-sf-sales-trip-excel-word-email task (RU / rzd + ClickHouse).

Checks:
1. Sales_Trip_Plan.xlsx exists with Travel_Details, Customer_Priority, Summary sheets
2. Travel_Details has rzd train 752А with correct depart/arrive times and duration
3. Customer_Priority has 5 "Северная Америка" customers sorted by Total_Amount desc
4. Top customer is Артём Борисов (Ethan Brown) with ~7088.92 total amount
5. Sales_Trip_Brief.docx exists with required sections
6. Email sent to sales-manager@company.com (subject "Business Trip Confirmed")

Stable anchors (NOT russified): @example.com customer emails + numeric amounts.
Volatile (russified centrally): customer names + region — matched RU+EN, never
keyed alone for identity. Train realia read honestly from the rzd seed.
"""
import json
import os
import sys
import unicodedata
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []

# Pinned rzd seed values for train 752А (Москва-Ленинградская -> Санкт-Петербург-Главный, 2026-03-10)
TRAIN_CODE = "752А"          # cyrillic А
DEPART_TIME = "06:50"
ARRIVE_TIME = "10:50"
DURATION_MIN = 240

# True top-5 NA customers (ranked by summed ORDERS.TOTAL_AMOUNT) — emails are stable anchors.
TOP5_EMAILS = [
    "ethan.brown.1056@example.com",
    "noah.garcia.109@example.com",
    "isabella.smith.1373@example.com",
    "harper.jones.658@example.com",
    "mason.garcia.1213@example.com",
]
# Russified names (central clickhouse map) — matched RU+EN, used only as a secondary signal.
RANK1_NAME_RU = "артём борисов"
RANK1_NAME_EN = "ethan brown"
TOP3_NAME_TOKENS = [
    ("борисов", "brown"),    # Артём Борисов / Ethan Brown
    ("григорьев", "garcia"), # Ной Григорьев / Noah Garcia
    ("смирнов", "smith"),    # Изабелла Смирнов / Isabella Smith
]
RANK1_TOTAL = 7088.92
RANK2_TOTAL = 6133.40


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILURES.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def normalize(s):
    """NFKD + cyrillic->latin translit. Use ONLY for mixed cyr/lat ID matching (752А ~ 752A)."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    table = {
        "а": "a", "в": "b", "е": "e", "к": "k", "м": "m", "н": "h", "о": "o",
        "р": "p", "с": "c", "т": "t", "у": "y", "х": "x",
    }
    out = []
    for ch in s.lower():
        out.append(table.get(ch, ch))
    return "".join(out)


def check_excel(agent_workspace):
    print("\n=== Check 1: Sales_Trip_Plan.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Sales_Trip_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Sales_Trip_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Sales_Trip_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Travel_Details sheet",
           any("travel" in s for s in sheet_names_lower), f"Sheets: {wb.sheetnames}")
    record("Has Customer_Priority sheet",
           any("customer" in s or "priority" in s for s in sheet_names_lower), f"Sheets: {wb.sheetnames}")
    record("Has Summary sheet", "summary" in sheet_names_lower, f"Sheets: {wb.sheetnames}")

    # --- Travel_Details ---
    travel_sheet = next((wb[n] for n in wb.sheetnames if "travel" in n.lower()), None)
    if travel_sheet:
        rows = list(travel_sheet.iter_rows(values_only=True))
        all_text = " ".join(str(c) for row in rows for c in row if c is not None)
        norm_text = normalize(all_text)
        record("Travel_Details has rzd train code 752А", normalize(TRAIN_CODE) in norm_text,
               f"No {TRAIN_CODE} found", critical=True)
        record("Travel_Details has correct departure 06:50", DEPART_TIME in all_text,
               f"No {DEPART_TIME} departure found", critical=True)
        record("Travel_Details has correct arrival 10:50", ARRIVE_TIME in all_text,
               f"No {ARRIVE_TIME} arrival found", critical=True)
        record("Travel_Details has duration 240 min", "240" in all_text.replace(".0", ""),
               "No 240 min duration found")

    # --- Customer_Priority ---
    customer_sheet = next((wb[n] for n in wb.sheetnames
                           if "customer" in n.lower() or "priority" in n.lower()), None)
    cp_emails_in_order = []
    cp_amounts_in_order = []
    if customer_sheet:
        rows = list(customer_sheet.iter_rows(values_only=True))
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]] if rows else []
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
        record("Customer_Priority has 5 customers", len(data_rows) >= 5, f"Found {len(data_rows)} rows")

        # Locate columns
        def col_idx(*names):
            for nm in names:
                if nm in header:
                    return header.index(nm)
            return None
        i_email = col_idx("customer_email", "email")
        i_total = col_idx("total_amount")
        i_rank = col_idx("priority_rank", "rank")

        for r in data_rows[:5]:
            if i_email is not None and i_email < len(r) and r[i_email]:
                cp_emails_in_order.append(str(r[i_email]).strip().lower())
            if i_total is not None and i_total < len(r) and r[i_total] is not None:
                try:
                    cp_amounts_in_order.append(float(r[i_total]))
                except (TypeError, ValueError):
                    pass

        # Stable identity anchor: the 5 emails must be exactly the true top-5 (any order).
        record("Customer_Priority has the 5 correct @example.com emails",
               set(cp_emails_in_order) == set(TOP5_EMAILS),
               f"Got {cp_emails_in_order}", critical=True)

        # Sorting + ranking by amount (descending), rank-1 / rank-2 numeric anchors.
        sorted_desc = cp_amounts_in_order == sorted(cp_amounts_in_order, reverse=True) and len(cp_amounts_in_order) == 5
        rank_ok = (len(cp_amounts_in_order) >= 2
                   and num_close(cp_amounts_in_order[0], RANK1_TOTAL, 1.0)
                   and num_close(cp_amounts_in_order[1], RANK2_TOTAL, 1.0))
        record("Customer_Priority sorted by Total_Amount desc with correct rank-1/rank-2 amounts",
               sorted_desc and rank_ok,
               f"amounts={cp_amounts_in_order}", critical=True)

        # Rank-1 row must be the right customer by email.
        rank1_email_ok = bool(cp_emails_in_order) and cp_emails_in_order[0] == TOP5_EMAILS[0]
        record("Rank-1 customer is ethan.brown.1056 (Артём Борисов)", rank1_email_ok,
               f"rank1 email={cp_emails_in_order[0] if cp_emails_in_order else None}")

        # Region russified value present.
        all_cp_text = " ".join(str(c) for row in rows for c in row if c is not None).lower()
        record("Region uses russified 'Северная Америка'",
               "северная америка" in all_cp_text or "north america" in all_cp_text,
               "No region value found")

    # --- Summary ---
    summary_sheet = next((wb[n] for n in wb.sheetnames if n.lower() == "summary"), None)
    if summary_sheet:
        srows = list(summary_sheet.iter_rows(values_only=True))
        smap = {}
        for r in srows:
            if r and r[0] is not None and len(r) >= 2:
                smap[str(r[0]).strip().lower()] = ("" if r[1] is None else str(r[1]).strip())
        top_cust = smap.get("top_customer", "")
        record("Summary Top_Customer is rank-1 (Артём Борисов / Ethan Brown)",
               RANK1_NAME_RU in top_cust.lower() or RANK1_NAME_EN in top_cust.lower()
               or "борисов" in top_cust.lower(),
               f"Top_Customer={top_cust!r}", critical=True)
        record("Summary Train_Code is 752А",
               normalize(TRAIN_CODE) in normalize(smap.get("train_code", "")),
               f"Train_Code={smap.get('train_code')!r}")
        record("Summary Travel_Date 2026-03-10", "2026-03-10" in smap.get("travel_date", ""),
               f"Travel_Date={smap.get('travel_date')!r}")
        purpose = smap.get("trip_purpose", "").lower()
        record("Summary Trip_Purpose is Business Development Visit / RU equivalent",
               "business development" in purpose or "развит" in purpose,
               f"Trip_Purpose={smap.get('trip_purpose')!r}")


def check_word(agent_workspace):
    print("\n=== Check 2: Sales_Trip_Brief.docx ===")
    docx_path = os.path.join(agent_workspace, "Sales_Trip_Brief.docx")
    if not os.path.exists(docx_path):
        record("Sales_Trip_Brief.docx exists", False, f"Not found at {docx_path}")
        return
    record("Sales_Trip_Brief.docx exists", True)

    try:
        import docx
        doc = docx.Document(docx_path)
        full_text = " ".join(p.text for p in doc.paragraphs)
    except Exception as e:
        record("Word document readable", False, str(e))
        return
    record("Word document readable", True)

    low = full_text.lower()
    record("Contains Travel Logistics section",
           "travel logistics" in low or "logistics" in low or "логист" in low,
           "No Travel Logistics section found")
    record("Contains Customer Visit section",
           ("customer" in low and ("visit" in low or "priority" in low)) or "клиент" in low,
           "No Customer Visit section found")
    record("Contains Meeting Objectives section",
           "meeting objectives" in low or "objectives" in low or "цел" in low,
           "No Meeting Objectives section found")
    record("Contains Follow-up Actions section",
           ("follow" in low and ("action" in low or "up" in low)) or "действ" in low,
           "No Follow-up Actions section found")
    record("Mentions rzd train 752А", normalize(TRAIN_CODE) in normalize(full_text),
           "No 752А train mention")
    record("Mentions top customer (Артём Борисов / Ethan Brown)",
           RANK1_NAME_RU in low or RANK1_NAME_EN in low or "борисов" in low,
           "No top customer mention")


def check_email():
    print("\n=== Check 3: Email to sales-manager@company.com ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%sales-manager%'
               OR subject ILIKE '%trip%'
               OR subject ILIKE '%business%'
               OR subject ILIKE '%confirmed%'
        """)
        all_emails = cur.fetchall()
        record("Email sent to sales-manager@company.com", len(all_emails) >= 1,
               f"Found {len(all_emails)} matching emails", critical=True)

        if all_emails:
            # Prefer the one addressed to sales-manager.
            target = next((e for e in all_emails if e[1] and "sales-manager" in str(e[1]).lower()),
                          all_emails[0])
            subject, to_addr, body = target
            subject_lower = (subject or "").lower()
            record("Email subject mentions trip/business/confirmed",
                   any(k in subject_lower for k in ("trip", "business", "confirmed", "поезд")),
                   f"Subject: {subject}")

            body_lower = (body or "")
            body_norm = normalize(body_lower)
            body_lc = body_lower.lower()
            record("Email body mentions rzd train 752А", normalize(TRAIN_CODE) in body_norm,
                   "No 752А train in email body", critical=True)

            # Top-3 identity: count how many of the top-3 appear via email OR RU/EN name token.
            hits = 0
            for (ru_tok, en_tok), email in zip(TOP3_NAME_TOKENS, TOP5_EMAILS[:3]):
                if email.lower() in body_lc or ru_tok in body_lc or en_tok in body_lc:
                    hits += 1
            record("Email body lists the top-3 customers (RU name / EN name / @example.com email)",
                   hits >= 2, f"Matched {hits}/3 top customers", critical=True)
    except Exception as e:
        record("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILURES,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILURES:
        print(f"\nFAIL: {len(CRITICAL_FAILURES)} critical check(s) failed: {CRITICAL_FAILURES}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
