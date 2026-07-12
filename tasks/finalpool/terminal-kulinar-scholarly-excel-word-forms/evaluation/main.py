"""Evaluation for terminal-kulinar-scholarly-excel-word-forms.
Checks:
1. Nutrition_Program_Analysis.xlsx with 4 sheets and correct content
2. Nutrition_Program_Proposal.docx with required sections
3. Survey form with 4 multiple-choice questions
4. Python scripts exist (categorize_recipes.py, analyze_research.py, build_menus.py, validate_menus.py)
5. JSON output files exist
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}: {str(detail)[:200]}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except:
        return False


def check_excel(workspace):
    print("\n=== Check 1: Nutrition_Program_Analysis.xlsx ===")
    path = os.path.join(workspace, "Nutrition_Program_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Recipe_Database
    rd_idx = next((i for i, s in enumerate(sheets_lower) if "recipe" in s and "database" in s), 0)
    ws1 = wb[sheets[rd_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Recipe_Database has 10+ recipes", len(data1) >= 10, f"Found {len(data1)}")
    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has dietary_tags column", any("dietary" in h or "tag" in h for h in headers), f"Headers: {headers}")
        check("Has evidence_score column", any("evidence" in h or "score" in h for h in headers), f"Headers: {headers}")
        check("Has prep_time column", any("prep" in h or "time" in h for h in headers), f"Headers: {headers}")

    # Research_Summary
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s or "summary" in s), 1)
    if rs_idx < len(sheets):
        ws2 = wb[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Research_Summary has 4+ papers", len(data2) >= 4, f"Found {len(data2)}")
        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has confidence_level column", any("confidence" in h for h in headers2), f"Headers: {headers2}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Research mentions workplace nutrition",
              any(k in all_text2 for k in (
                  "workplace", "employee", "corporate", "wellness",
                  "рабоч", "сотрудник", "корпоратив", "питани", "благополучи")),
              all_text2[:150])

    # Weekly_Menu
    wm_idx = next((i for i, s in enumerate(sheets_lower) if "weekly" in s or "menu" in s), 2)
    if wm_idx < len(sheets):
        ws3 = wb[sheets[wm_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Weekly_Menu has 5 days", len(data3) >= 5, f"Found {len(data3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Menu includes Monday", "monday" in all_text3)
        check("Menu includes Friday", "friday" in all_text3)
        if rows3:
            headers3 = [str(c).lower() if c else "" for c in rows3[0]]
            check("Has dietary_compliance_pct column",
                  any("compliance" in h or "dietary" in h for h in headers3), f"Headers: {headers3}")
            check("Has est_calories column",
                  any("calori" in h for h in headers3), f"Headers: {headers3}")

    # Program_Budget
    pb_idx = next((i for i, s in enumerate(sheets_lower) if "budget" in s), 3)
    if pb_idx < len(sheets):
        ws4 = wb[sheets[pb_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Program_Budget has 5+ rows", len(data4) >= 5, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Budget includes ingredient costs",
              any(k in all_text4 for k in ("ingredient", "recipe", "ингредиент", "рецепт")))
        check("Budget includes workshop",
              any(k in all_text4 for k in ("workshop", "мастер")))
        # Check total_cost calculation for ALL rows (CRITICAL)
        if rows4:
            headers4 = [str(c).lower() if c else "" for c in rows4[0]]
            cost_idx = next((i for i, h in enumerate(headers4) if "total_cost" in h or h == "total_cost"), -1)
            cpp_idx = next((i for i, h in enumerate(headers4) if "cost_per" in h), -1)
            part_idx = next((i for i, h in enumerate(headers4) if "participant" in h), -1)
            if cost_idx >= 0 and cpp_idx >= 0 and part_idx >= 0 and data4:
                all_ok = True
                bad = None
                checked_any = False
                for row in data4:
                    if row[cost_idx] is not None and row[cpp_idx] is not None and row[part_idx] is not None:
                        checked_any = True
                        expected = float(row[cpp_idx]) * float(row[part_idx])
                        if not num_close(row[cost_idx], expected, 1.0):
                            all_ok = False
                            bad = f"{row[cpp_idx]}*{row[part_idx]}={expected} != {row[cost_idx]}"
                            break
                check("Budget total_cost = cost_per_person * participants (ALL rows)",
                      checked_any and all_ok, bad or "no numeric rows", critical=True)
                # Verify the 5 prescribed (cost_per_person, participants) pairs are present
                pairs = set()
                for row in data4:
                    try:
                        pairs.add((round(float(row[cpp_idx])), round(float(row[part_idx]))))
                    except (TypeError, ValueError):
                        pass
                required_pairs = {(35, 200), (15, 50), (25, 200), (2, 200), (10, 200)}
                check("Budget includes the 5 prescribed cost/participant pairs",
                      required_pairs.issubset(pairs),
                      f"missing {required_pairs - pairs}", critical=True)
            else:
                check("Budget total_cost = cost_per_person * participants (ALL rows)",
                      False, "missing budget columns", critical=True)


def check_word(workspace):
    print("\n=== Check 2: Nutrition_Program_Proposal.docx ===")
    path = os.path.join(workspace, "Nutrition_Program_Proposal.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()

    check("Has Executive Summary",
          "executive" in full_text or "резюме" in full_text or "исполнит" in full_text)
    check("Has Research Foundation",
          ("research" in full_text and "foundation" in full_text)
          or "научн" in full_text and ("основ" in full_text or "база" in full_text))
    check("Has Recipe Selection",
          ("recipe" in full_text and ("selection" in full_text or "methodology" in full_text))
          or ("рецепт" in full_text and ("отбор" in full_text or "методолог" in full_text)))
    check("Has Weekly Menus section",
          ("weekly" in full_text and "menu" in full_text)
          or ("меню" in full_text and ("недел" in full_text or "weekly" in full_text)))
    check("Has Survey Plan",
          ("survey" in full_text and "plan" in full_text)
          or ("опрос" in full_text and "план" in full_text))
    check("Has Budget section", "budget" in full_text or "бюджет" in full_text)
    check("Has Timeline section",
          "timeline" in full_text or "phase" in full_text
          or "график" in full_text or "этап" in full_text)
    check("Mentions 200 employees", "200" in full_text)
    check("Mentions evidence-based", "evidence" in full_text or "обоснован" in full_text)
    check("Mentions Phase 1",
          "phase 1" in full_text or "phase one" in full_text or "month 1" in full_text
          or "этап 1" in full_text or "первый месяц" in full_text)


def check_gform():
    print("\n=== Check 3: Survey form (forms RU fork, gform.* schema) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("Survey form: >=4 multiple-choice questions incl dietary-restrictions question",
              False, f"db connect failed: {e}", critical=True)
        return
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()
        diet_form = None
        for form_id, title in forms:
            tl = (title or "").lower()
            if any(k in tl for k in ("dietary", "nutrition", "preference",
                                     "диет", "питан", "предпочт", "опрос")):
                # skip the seeded IT-equipment noise form
                if "ит-оборуд" in tl or "it equipment" in tl:
                    continue
                diet_form = (form_id, title)
                break
        check("Dietary preferences survey form exists", diet_form is not None,
              f"Forms: {[f[1] for f in forms]}")

        # config may carry the real {'type': 'CHECKBOX'/'RADIO'} for the forms RU fork
        def _types_of(rows):
            out = []
            for r in rows:
                qt = (r[1] or "")
                cfg = r[2] if len(r) > 2 else None
                cfg_type = ""
                if isinstance(cfg, dict):
                    cfg_type = str(cfg.get("type", ""))
                elif isinstance(cfg, str):
                    cfg_type = cfg
                out.append((qt, cfg_type))
            return out

        def _is_choice(qt, cfg_type):
            # forms-mcp produces question_type='choiceQuestion' with config.type='RADIO'.
            # (There is no add_checkbox tool; CHECKBOX is accepted only if a future
            #  schema produces it, but is not required.)
            blob = (qt + " " + cfg_type).upper()
            return any(k in blob for k in ("CHOICE", "RADIO", "MULTIPLE_CHOICE", "CHECKBOX"))

        if diet_form:
            try:
                cur.execute("SELECT title, question_type, config FROM gform.questions "
                            "WHERE form_id = %s ORDER BY position", (diet_form[0],))
                qrows = cur.fetchall()
            except Exception:
                conn.rollback()
                cur.execute("SELECT title, question_type FROM gform.questions "
                            "WHERE form_id = %s ORDER BY position", (diet_form[0],))
                qrows = [(r[0], r[1], None) for r in cur.fetchall()]

            q_count = len(qrows)
            check("Survey has 4 questions", q_count >= 4, f"Found {q_count}")

            q_text = " ".join(str(q[0]) for q in qrows).lower()
            tpairs = _types_of(qrows)
            has_diet_q = any(k in q_text for k in ("dietary", "restriction", "ограничен",
                                                   "вегетариан", "веган", "глютен"))
            check("Has dietary restrictions question", has_diet_q, q_text[:150])
            check("Has meal focus question",
                  any(k in q_text for k in ("meal", "focus", "breakfast",
                                            "приём пищи", "прием пищи", "завтрак", "обед")),
                  q_text[:150])

            choice_count = sum(1 for qt, ct in tpairs if _is_choice(qt, ct))
            check("Has >=4 multiple-choice questions", choice_count >= 4,
                  f"choice_count={choice_count}, pairs={tpairs}")

            # CRITICAL: real survey shape — RU/EN title, >=4 choice questions,
            # incl. the substantive dietary-restrictions question.
            # NB: forms-mcp only offers add_text_question / add_multiple_choice_question
            # (question_type='choiceQuestion', config.type='RADIO'); there is no
            # checkbox tool, so we require the achievable choice-question shape.
            check("Survey form: >=4 multiple-choice questions incl dietary-restrictions question",
                  q_count >= 4 and has_diet_q and choice_count >= 4,
                  f"count={q_count}, diet_q={has_diet_q}, choice={choice_count}",
                  critical=True)
        else:
            check("Survey form: >=4 multiple-choice questions incl dietary-restrictions question",
                  False, "no dietary form", critical=True)
    except Exception as e:
        check("Gform check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python scripts ===")
    for script in ["categorize_recipes.py", "analyze_research.py", "build_menus.py", "validate_menus.py"]:
        path = os.path.join(workspace, script)
        check(f"{script} exists", os.path.exists(path))


def check_json_outputs(workspace):
    print("\n=== Check 5: JSON output files ===")
    for fname in ["categorized_recipes.json", "research_findings.json", "evidence_based_menus.json"]:
        path = os.path.join(workspace, fname)
        if not os.path.exists(path):
            check(f"{fname} exists", False)
            continue
        check(f"{fname} exists", True)
        try:
            with open(path) as f:
                data = json.load(f)
            check(f"{fname} is valid JSON", True)
            if fname == "categorized_recipes.json":
                if isinstance(data, list):
                    check("categorized_recipes has 10+ entries", len(data) >= 10, f"Found {len(data)}")
                elif isinstance(data, dict):
                    total = sum(len(v) if isinstance(v, list) else 1 for v in data.values())
                    check("categorized_recipes has 10+ entries", total >= 10, f"Found {total}")
            elif fname == "evidence_based_menus.json":
                if isinstance(data, list):
                    check("evidence_based_menus has 5 days", len(data) >= 5, f"Found {len(data)}")
                elif isinstance(data, dict):
                    check("evidence_based_menus has 5 days", len(data) >= 5, f"Found {len(data)} keys")
        except json.JSONDecodeError:
            check(f"{fname} is valid JSON", False, "Invalid JSON")


def check_reverse_validation(workspace):
    """Check that noise scholarly papers are NOT present in Excel output."""
    print("\n=== Reverse Validation ===")

    # Noise paper titles that should NOT appear in the Research_Summary sheet
    noise_titles = ["marathon runners", "carbohydrate loading", "elite marathon",
                    "agricultural policy", "food supply chain", "developing nations"]

    path = os.path.join(workspace, "Nutrition_Program_Analysis.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        sheets_lower = [s.lower().replace(" ", "_") for s in wb.sheetnames]

        # Check Research_Summary sheet for noise
        rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s or "summary" in s), None)
        if rs_idx is not None:
            ws = wb[wb.sheetnames[rs_idx]]
            rows = list(ws.iter_rows(values_only=True))
            all_text = " ".join(str(c) for r in rows for c in r if c).lower()

            no_noise = not any(nt in all_text for nt in noise_titles)
            rs_rows = [r for r in rows[1:] if any(c for c in r)]
            check("No noise scholarly papers in Research_Summary (marathon, agricultural policy)",
                  no_noise and len(rs_rows) >= 4,
                  f"no_noise={no_noise}, relevant_rows={len(rs_rows)}", critical=True)
        else:
            check("No noise scholarly papers in Research_Summary", True, "No Research_Summary sheet to check")

        # Also check all sheets for noise paper content
        all_wb_text = ""
        for sn in wb.sheetnames:
            ws = wb[sn]
            for row in ws.iter_rows(values_only=True):
                all_wb_text += " ".join(str(c) for c in row if c).lower() + " "

        no_noise_wb = not any(nt in all_wb_text for nt in noise_titles)
        check("No noise scholarly papers anywhere in Excel workbook",
              no_noise_wb,
              "Found noise paper content in workbook", critical=True)
        wb.close()
    else:
        check("No noise scholarly papers in Excel", True, "Excel file not found to check")


def _load_json(workspace, fname):
    path = os.path.join(workspace, fname)
    if not os.path.isfile(path):
        return None
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return None


def _recipe_category_map(workspace):
    """name(lower) -> category(lower) from categorized_recipes.json."""
    data = _load_json(workspace, "categorized_recipes.json")
    out = {}
    items = []
    if isinstance(data, list):
        items = data
    elif isinstance(data, dict):
        for v in data.values():
            if isinstance(v, list):
                items.extend(v)
    for it in items:
        if isinstance(it, dict):
            name = str(it.get("recipe_name") or it.get("name") or "").strip().lower()
            cat = str(it.get("category") or "").strip().lower()
            if name:
                out[name] = cat
    return out


def check_menu_rule(workspace):
    """CRITICAL: 5 days Mon-Fri and no two consecutive days share lunch category."""
    print("\n=== Semantic Check: Weekly menu rule ===")
    menus = _load_json(workspace, "evidence_based_menus.json")
    days = []
    if isinstance(menus, list):
        days = [d for d in menus if isinstance(d, dict)]
    elif isinstance(menus, dict):
        # dict keyed by day name
        order = ["monday", "tuesday", "wednesday", "thursday", "friday"]
        items = list(menus.items())
        items.sort(key=lambda kv: order.index(kv[0].lower())
                   if kv[0].lower() in order else 99)
        for k, v in items:
            if isinstance(v, dict):
                vv = dict(v)
                vv.setdefault("day", k)
                days.append(vv)

    check("evidence_based_menus has 5 day entries", len(days) >= 5,
          f"found {len(days)}", critical=True)
    if len(days) < 2:
        check("No two consecutive days share the same lunch category", False,
              "not enough days", critical=True)
        return

    cat_map = _recipe_category_map(workspace)

    def lunch_cat(day):
        # explicit category fields first
        for key in ("lunch_category", "lunch_cat"):
            if isinstance(day.get(key), str) and day[key].strip():
                return day[key].strip().lower()
        lunch = day.get("lunch")
        if isinstance(lunch, dict):
            for key in ("category", "lunch_category"):
                if isinstance(lunch.get(key), str) and lunch[key].strip():
                    return lunch[key].strip().lower()
            name = str(lunch.get("name") or lunch.get("recipe_name") or "").strip().lower()
            return cat_map.get(name, name)
        if isinstance(lunch, str):
            return cat_map.get(lunch.strip().lower(), lunch.strip().lower())
        return ""

    cats = [lunch_cat(d) for d in days[:5]]
    # only enforce when we could resolve categories for the whole week
    resolved = all(c for c in cats)
    no_consec = all(cats[i] != cats[i + 1] for i in range(len(cats) - 1))
    if resolved:
        check("No two consecutive days share the same lunch category",
              no_consec, f"lunch categories: {cats}", critical=True)
    else:
        # fall back to lunch recipe names if categories unresolved
        names = []
        for d in days[:5]:
            lunch = d.get("lunch")
            if isinstance(lunch, dict):
                names.append(str(lunch.get("name") or lunch.get("recipe_name") or "").lower())
            else:
                names.append(str(lunch or "").lower())
        no_consec_name = all(names[i] != names[i + 1] for i in range(len(names) - 1) if names[i])
        check("No two consecutive days share the same lunch (category unresolved -> by name)",
              no_consec_name, f"lunches: {names}", critical=True)


def check_confidence_mapping(workspace):
    """CRITICAL: confidence_level in Research_Summary maps to citation counts.
    Known relevant papers and citation counts:
      'workplace nutrition programs' -> 156 (High)
      'corporate settings' / 'employee health outcomes' -> 89 (Medium)
      'low-sodium' / 'longitudinal' -> 203 (High)
      'engagement' / 'survey design' -> 47 (Medium)
    """
    print("\n=== Semantic Check: confidence_level mapping ===")
    path = os.path.join(workspace, "Nutrition_Program_Analysis.xlsx")
    if not os.path.isfile(path):
        check("confidence_level maps to citation counts", False,
              "no workbook", critical=True)
        return
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets_lower = [s.lower().replace(" ", "_") for s in wb.sheetnames]
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s or "summary" in s), None)
    if rs_idx is None:
        check("confidence_level maps to citation counts", False,
              "no Research_Summary", critical=True)
        wb.close()
        return
    ws = wb[wb.sheetnames[rs_idx]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        check("confidence_level maps to citation counts", False, "empty sheet", critical=True)
        return
    headers = [str(c).lower() if c else "" for c in rows[0]]
    title_idx = next((i for i, h in enumerate(headers) if "title" in h or "paper" in h), 0)
    conf_idx = next((i for i, h in enumerate(headers) if "confidence" in h), -1)
    if conf_idx < 0:
        check("confidence_level maps to citation counts", False,
              "no confidence_level column", critical=True)
        return

    # title substring -> expected confidence
    expected = [
        (("workplace nutrition programs",), "high"),   # 156
        (("longitudinal", "low-sodium"), "high"),      # 203
        (("employee health outcomes", "corporate settings"), "medium"),  # 89
        (("engagement", "survey design", "feedback mechanisms"), "medium"),  # 47
    ]
    ok = True
    detail = []
    for row in rows[1:]:
        if not any(c for c in row):
            continue
        title = str(row[title_idx] or "").lower()
        conf = str(row[conf_idx] or "").lower()
        for subs, exp in expected:
            if any(s in title for s in subs):
                if exp not in conf:
                    ok = False
                    detail.append(f"'{title[:40]}' -> '{conf}' (exp {exp})")
                break
    check("confidence_level maps to citation counts (High>100 / Medium 20-100)",
          ok, "; ".join(detail) or "ok", critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_gform()
    check_scripts(args.agent_workspace)
    check_json_outputs(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)
    check_menu_rule(args.agent_workspace)
    check_confidence_mapping(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_fails": CRITICAL_FAILS}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Critical gate: any critical failure => FAIL regardless of accuracy
    if CRITICAL_FAILS:
        print(f"\nFAIL: {len(CRITICAL_FAILS)} critical check(s) failed: {CRITICAL_FAILS}")
        sys.exit(1)

    print(f"\nCritical checks: all passed.")
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()
