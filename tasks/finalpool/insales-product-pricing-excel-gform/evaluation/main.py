"""Evaluation for insales-product-pricing-excel-gform task (InSales + forms).

Checks:
1. Excel file Pricing_Audit.xlsx with correct data (both sheets)
2. Online feedback form titled "Pricing Review Feedback" with 4 questions (forms / gform.*)

Scoring model:
- CRITICAL_CHECKS reflect the task's substance. A single critical failure =>
  overall FAIL (sys.exit(1)) regardless of accuracy.
- Otherwise PASS requires accuracy >= 70.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# On-sale categories (russified centrally by db/zzz_wc_after_init.sql). These are
# the six categories that have products on sale in the groundtruth audit; Q2 of the
# form must offer exactly this set. DO NOT hand-edit these literals — they are
# map-synced with the wc seed + groundtruth_workspace.
EXPECTED_CATEGORIES = ["Электроника", "Аудио", "ТВ и домашний кинотеатр", "Камеры", "Бытовая техника", "Часы"]

ASSESSMENT_OPTIONS = ["too aggressive", "appropriate", "too conservative"]


def record(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300]) if len(detail) > 300 else detail
        msg = f": {d}" if d else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def check(name, condition, detail=""):
    record(name, condition, detail, critical=False)


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Проверка 1: Pricing_Audit.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Pricing_Audit.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Pricing_Audit.xlsx")

    crit_disc = "Excel 'Discounted Products': число строк совпадает и >=80% строк совпадают по ценам/скидке"
    crit_cat = "Excel 'Category Summary': ровно 6 категорий со скидками с верными Products_On_Sale"

    if not os.path.exists(agent_file):
        check("Файл Excel существует", False, f"Not found: {agent_file}")
        record(crit_disc, False, "no file", critical=True)
        record(crit_cat, False, "no file", critical=True)
        return
    check("Файл Excel существует", True)

    if not os.path.exists(gt_file):
        check("Файл groundtruth существует", False, f"Not found: {gt_file}")
        record(crit_disc, False, "no groundtruth", critical=True)
        record(crit_cat, False, "no groundtruth", critical=True)
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---------------- Discounted Products sheet ----------------
    print("  Проверка листа Discounted Products...")
    a_rows = load_sheet_rows(agent_wb, "Discounted Products")
    g_rows = load_sheet_rows(gt_wb, "Discounted Products")

    disc_count_ok = False
    disc_accuracy_ok = False
    if a_rows is None:
        check("Лист 'Discounted Products' существует", False, "Not found")
    elif g_rows is None:
        check("Лист groundtruth 'Discounted Products' существует", False, "Not found")
    else:
        check("Лист 'Discounted Products' существует", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        disc_count_ok = len(a_data) == len(g_data)
        check("Discounted Products: число строк совпадает",
              disc_count_ok, f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                key = str(row[0]).strip().lower()[:50]
                a_lookup[key] = row

        match_count = 0
        mismatch_count = 0
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()[:50]
            a_row = a_lookup.get(key)
            if a_row is None:
                mismatch_count += 1
                continue

            all_ok = True
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 0.5):
                all_ok = False
            if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 0.5):
                all_ok = False
            if len(a_row) > 4 and len(g_row) > 4 and not num_close(a_row[4], g_row[4], 1.0):
                all_ok = False

            if all_ok:
                match_count += 1
            else:
                mismatch_count += 1

        total_expected = len([r for r in g_data if r and r[0] is not None])
        disc_accuracy_ok = match_count >= total_expected * 0.8
        check(f"Discounted Products: точность данных ({match_count}/{total_expected})",
              disc_accuracy_ok, f"Matched {match_count}, mismatched {mismatch_count}")

        # Sort order (Discount_Pct descending) — structural, non-critical.
        if len(a_data) >= 2:
            sorted_ok = True
            for i in range(min(len(a_data) - 1, 10)):
                if a_data[i] and a_data[i+1] and len(a_data[i]) > 4 and len(a_data[i+1]) > 4:
                    try:
                        if float(a_data[i][4]) < float(a_data[i+1][4]) - 0.5:
                            sorted_ok = False
                            break
                    except (TypeError, ValueError):
                        pass
            check("Discounted Products отсортирован по Discount_Pct убыв.", sorted_ok)

    record(crit_disc, disc_count_ok and disc_accuracy_ok,
           f"count_ok={disc_count_ok}, accuracy_ok={disc_accuracy_ok}", critical=True)

    # ---------------- Category Summary sheet ----------------
    print("  Проверка листа Category Summary...")
    a_rows = load_sheet_rows(agent_wb, "Category Summary")
    g_rows = load_sheet_rows(gt_wb, "Category Summary")

    cat_ok = False
    if a_rows is None:
        check("Лист 'Category Summary' существует", False, "Not found")
    elif g_rows is None:
        check("Лист groundtruth 'Category Summary' существует", False, "Not found")
    else:
        check("Лист 'Category Summary' существует", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        count_ok = len(a_data) == len(g_data)
        check("Category Summary: число строк совпадает",
              count_ok, f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        cat_match = 0
        cat_total = 0
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            cat_total += 1
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Category '{g_row[0]}'", False, "Missing")
                continue

            pos_ok = True
            if len(a_row) > 1 and len(g_row) > 1:
                pos_match = num_close(a_row[1], g_row[1], 2)
                pos_ok = pos_ok and pos_match
                check(f"{key}.Products_On_Sale", pos_match, f"{a_row[1]} vs {g_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Avg_Discount_Pct",
                      num_close(a_row[2], g_row[2], 2.0), f"{a_row[2]} vs {g_row[2]}")
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Max_Discount_Pct",
                      num_close(a_row[3], g_row[3], 1.0), f"{a_row[3]} vs {g_row[3]}")
            if pos_ok:
                cat_match += 1

        cat_ok = count_ok and cat_total == len(EXPECTED_CATEGORIES) and cat_match == cat_total

    record(crit_cat, cat_ok, "категории/Products_On_Sale не совпадают с эталоном",
           critical=True)


def _is_choice(qtype):
    t = (qtype or "")
    return t in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE", "CHOICE", "CHECKBOX")


def _is_text(qtype):
    t = (qtype or "")
    return t in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH")


def _option_values(config):
    """Extract option text values from a question config (RU forms-mcp shape)."""
    vals = []
    if not config:
        return vals
    cfg = config if isinstance(config, dict) else (json.loads(config) if config else {})
    opts = cfg.get("options") if isinstance(cfg, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    vals.append(str(v))
            else:
                vals.append(str(o))
    return vals


def check_gform():
    print("\n=== Проверка 2: Форма 'Pricing Review Feedback' (forms / gform.*) ===")
    crit_form = ("Форма 'Pricing Review Feedback' существует с заголовком "
                 "(pricing/review/feedback) и ровно 4 вопросами в нужном порядке")
    crit_q1q4 = "Q1 имя рецензента = текст + обязательный; Q4 комментарии = текст + НЕ обязательный"
    crit_q2 = ("Q2 = обязательный множественный выбор; варианты = ровно 6 категорий "
               "со скидками (RU)")
    crit_q3 = ("Q3 = обязательный множественный выбор; присутствуют 3 варианта оценки "
               "(Too Aggressive / Appropriate / Too Conservative)")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record(crit_form, False, str(e), critical=True)
        record(crit_q1q4, False, "no db", critical=True)
        record(crit_q2, False, "no db", critical=True)
        record(crit_q3, False, "no db", critical=True)
        return
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    form_id = None
    form_title = None
    for fid, ftitle in forms:
        t = (ftitle or "").lower()
        if "pricing" in t and ("review" in t or "feedback" in t):
            form_id, form_title = fid, ftitle
            break
    if form_id is None:
        for fid, ftitle in forms:
            t = (ftitle or "").lower()
            if "pricing" in t or "review" in t or "feedback" in t:
                form_id, form_title = fid, ftitle
                break

    title_ok = bool(form_title) and "pricing" in (form_title or "").lower() \
        and ("review" in (form_title or "").lower() or "feedback" in (form_title or "").lower())
    check("Заголовок формы содержит pricing + review/feedback",
          title_ok, f"Got: {form_title}; all forms: {[r[1] for r in forms]}")

    questions = []
    if form_id is not None:
        cur.execute(
            "SELECT title, question_type, required, config FROM gform.questions "
            "WHERE form_id=%s ORDER BY position", (form_id,))
        questions = cur.fetchall()
    cur.close()
    conn.close()

    count_ok = len(questions) == 4
    check("Форма содержит ровно 4 вопроса", count_ok, f"Found {len(questions)} questions")
    record(crit_form, title_ok and count_ok,
           f"title_ok={title_ok}, count={len(questions)}", critical=True)

    parsed = []
    for q_title, q_type, q_req, q_cfg in questions:
        parsed.append({
            "title": (q_title or ""),
            "type": q_type,
            "required": q_req,
            "options_lower": [v.lower() for v in _option_values(q_cfg)],
        })

    # --- Q1 + Q4 (the two ends define the form shape) ---
    q1_ok = False
    q4_ok = False
    if len(parsed) >= 1:
        q = parsed[0]
        t1 = q["title"].lower()
        title1_ok = "name" in t1 or "имя" in t1
        q1_ok = title1_ok and _is_text(q["type"]) and q["required"] is True
        check("Q1: имя рецензента (текст, обязательный)", q1_ok,
              f"title='{q['title']}', type='{q['type']}', required={q['required']}")
    if len(parsed) >= 4:
        q = parsed[3]
        t4 = q["title"].lower()
        title4_ok = "comment" in t4 or "коммент" in t4
        q4_ok = title4_ok and _is_text(q["type"]) and q["required"] is not True
        check("Q4: комментарии (текст, не обязательный)", q4_ok,
              f"title='{q['title']}', type='{q['type']}', required={q['required']}")
    record(crit_q1q4, q1_ok and q4_ok, critical=True)

    # --- Q2: required multiple-choice, options == 6 on-sale categories ---
    q2_ok = False
    if len(parsed) >= 2:
        q = parsed[1]
        type_req_ok = _is_choice(q["type"]) and q["required"] is True
        joined = " ".join(q["options_lower"])
        cats_present = all(c.lower() in joined for c in EXPECTED_CATEGORIES)
        # no extra option beyond the 6 categories
        no_extra = len(q["options_lower"]) == len(EXPECTED_CATEGORIES)
        q2_ok = type_req_ok and cats_present and no_extra
        check("Q2: категория товара (множ. выбор, обязательный)", type_req_ok,
              f"type='{q['type']}', required={q['required']}")
        check("Q2: варианты = 6 категорий со скидками", cats_present and no_extra,
              f"options={q['options_lower']}")
    record(crit_q2, q2_ok, critical=True)

    # --- Q3: required multiple-choice with the three assessment options ---
    q3_ok = False
    if len(parsed) >= 3:
        q = parsed[2]
        type_req_ok = _is_choice(q["type"]) and q["required"] is True
        joined = " ".join(q["options_lower"])
        opts_present = all(o in joined for o in ASSESSMENT_OPTIONS)
        q3_ok = type_req_ok and opts_present
        check("Q3: оценка ценообразования (множ. выбор, обязательный)", type_req_ok,
              f"type='{q['type']}', required={q['required']}")
        check("Q3: варианты Too Aggressive/Appropriate/Too Conservative", opts_present,
              f"options={q['options_lower']}")
    record(crit_q3, q3_ok, critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    print(f"  Critical failures: {CRITICAL_FAILS}")

    if CRITICAL_FAILS:
        all_passed = False
    else:
        all_passed = accuracy >= 70
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "passed": PASS_COUNT,
                "failed": FAIL_COUNT,
                "accuracy": accuracy,
                "critical_failures": CRITICAL_FAILS,
                "success": all_passed,
            }, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILS}")
        sys.exit(1)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
