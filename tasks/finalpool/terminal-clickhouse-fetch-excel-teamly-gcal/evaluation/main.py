"""Evaluation for terminal-sf-fetch-excel-clickhouse-teamly-gcal.

Source: ClickHouse (sf_data.SUPPORT_CENTER, read-only, russified centrally).
Knowledge base: Teamly (schema teamly.*) — replaces the old Notion check.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.
"""
import argparse
import json
import os
import sys
from datetime import datetime

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Industry benchmark values from the static dashboard page (http://localhost:30185).
# These are read off a fixed page, so hardcoding them is safe.
BENCH_RESPONSE = 8.00
BENCH_SATISFACTION = 3.50
BENCH_COMPLIANCE = 85.00

# Any failure here => overall FAIL regardless of accuracy gate.
CRITICAL_CHECKS = {
    "Benchmark_Comparison contains the 3 industry benchmark values (8.00, 3.50, 85.00)",
    "Benchmark_Comparison status matches internal-vs-benchmark comparison (direction-aware)",
    "Ticket_Summary compliance rates are in [0,100] with >=3 priority rows",
    "Calendar has the 3 monthly SLA review events on first Mondays (Apr/May/Jun 2026), 10:00, 90 min",
    "Teamly 'SLA Compliance Dashboard' space has >=3 metric pages",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", ".").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def _numbers_in_rows(rows):
    out = []
    for r in rows:
        for c in r:
            f = safe_float(c) if isinstance(c, str) else (c if isinstance(c, (int, float)) else None)
            if isinstance(c, (int, float)):
                out.append(float(c))
            elif f is not None:
                out.append(f)
    return out


def check_excel(workspace):
    print("\n=== Check 1: SLA_Compliance_Report.xlsx ===")
    path = os.path.join(workspace, "SLA_Compliance_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        # Mark dependent critical checks failed.
        check("Benchmark_Comparison contains the 3 industry benchmark values (8.00, 3.50, 85.00)", False, "no excel")
        check("Benchmark_Comparison status matches internal-vs-benchmark comparison (direction-aware)", False, "no excel")
        check("Ticket_Summary compliance rates are in [0,100] with >=3 priority rows", False, "no excel")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")
    sheets_lower = [s.lower() for s in sheets]

    # ---- Ticket_Summary -----------------------------------------------------
    ts_idx = next((i for i, s in enumerate(sheets_lower) if "ticket" in s or "summary" in s), 0)
    ws = wb[sheets[ts_idx]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip().lower() if c is not None else "" for c in (rows[0] if rows else [])]
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    check("Ticket_Summary has 3 priority rows", len(data_rows) >= 3, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    # Priority "High" stays English in sf_data (not russified). Accept EN or RU.
    check("Contains High priority", "high" in all_text or "высок" in all_text, f"Text: {all_text[:120]}")

    # CRITICAL: compliance rate column present and in [0,100] for the priority rows.
    comp_i = next((i for i, h in enumerate(headers)
                   if "compliance" in h or "соответств" in h), None)
    comp_ok = False
    detail = ""
    if comp_i is not None:
        vals = []
        for r in data_rows:
            if comp_i < len(r):
                f = safe_float(r[comp_i])
                if f is not None:
                    vals.append(f)
        comp_ok = len(vals) >= 3 and all(0 <= v <= 100 for v in vals)
        detail = f"compliance values: {vals}"
    else:
        detail = f"no compliance column in headers: {headers}"
    check("Ticket_Summary compliance rates are in [0,100] with >=3 priority rows", comp_ok, detail)

    # ---- Benchmark_Comparison ----------------------------------------------
    bc_idx = next((i for i, s in enumerate(sheets_lower) if "benchmark" in s or "comparison" in s), 1)
    bench_vals_ok = False
    status_ok = False
    if bc_idx < len(sheets):
        ws2 = wb[sheets[bc_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        h2 = [str(c).strip().lower() if c is not None else "" for c in (rows2[0] if rows2 else [])]
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Benchmark_Comparison has at least 3 rows", len(data_rows2) >= 3, f"Found {len(data_rows2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        # RU+EN metric substrings.
        check("Contains response time metric",
              "response" in all_text2 or "реакц" in all_text2 or "отклик" in all_text2,
              f"Text: {all_text2[:120]}")
        check("Contains satisfaction metric",
              "satisfaction" in all_text2 or "удовлетвор" in all_text2,
              f"Text: {all_text2[:120]}")
        check("Contains status values",
              "meets" in all_text2 or "below" in all_text2
              or "соответ" in all_text2 or "ниже" in all_text2,
              f"Text: {all_text2[:120]}")

        # Column indices.
        def col(*subs):
            for i, h in enumerate(h2):
                if any(s in h for s in subs):
                    return i
            return None
        metric_i = col("metric", "метрик", "назван")
        internal_i = col("internal", "внутрен")
        bench_i = col("benchmark", "industry", "бенчмар", "отрасл")
        status_i = col("status", "статус")

        # CRITICAL: the 3 known industry benchmark values must be present.
        all_nums = _numbers_in_rows(data_rows2)
        def has_num(target, tol=0.01):
            return any(abs(n - target) <= tol for n in all_nums)
        bench_vals_ok = has_num(BENCH_RESPONSE) and has_num(BENCH_SATISFACTION) and has_num(BENCH_COMPLIANCE)
        check("Benchmark_Comparison contains the 3 industry benchmark values (8.00, 3.50, 85.00)",
              bench_vals_ok, f"numbers present: {sorted(set(round(n, 2) for n in all_nums))}")

        # CRITICAL: status must be direction-consistent with internal vs benchmark.
        # response/time -> lower is better; satisfaction/compliance -> higher is better.
        if metric_i is not None and internal_i is not None and bench_i is not None and status_i is not None:
            checked = 0
            consistent = 0
            for r in data_rows2:
                if max(metric_i, internal_i, bench_i, status_i) >= len(r):
                    continue
                name = str(r[metric_i]).lower() if r[metric_i] is not None else ""
                iv = safe_float(r[internal_i])
                bv = safe_float(r[bench_i])
                st = str(r[status_i]).strip().lower() if r[status_i] is not None else ""
                if iv is None or bv is None or not st:
                    continue
                lower_better = ("response" in name or "реакц" in name
                                or "отклик" in name or "time" in name)
                if lower_better:
                    meets = iv <= bv
                else:
                    meets = iv >= bv
                says_meets = ("meets" in st or "соответ" in st) and "ниже" not in st
                says_below = "below" in st or "ниже" in st
                if not (says_meets or says_below):
                    continue
                checked += 1
                if (meets and says_meets) or ((not meets) and says_below):
                    consistent += 1
            status_ok = checked >= 3 and consistent == checked
            check("Benchmark_Comparison status matches internal-vs-benchmark comparison (direction-aware)",
                  status_ok, f"consistent {consistent}/{checked}")
        else:
            check("Benchmark_Comparison status matches internal-vs-benchmark comparison (direction-aware)",
                  False, f"missing columns; headers={h2}")
    else:
        check("Benchmark_Comparison contains the 3 industry benchmark values (8.00, 3.50, 85.00)", False, "no sheet")
        check("Benchmark_Comparison status matches internal-vs-benchmark comparison (direction-aware)", False, "no sheet")

    # ---- Agent_Performance --------------------------------------------------
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "agent" in s or "performance" in s), 2)
    if ap_idx < len(sheets):
        ws3 = wb[sheets[ap_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Agent_Performance has at least 5 rows", len(data_rows3) >= 5, f"Found {len(data_rows3)}")

    # ---- Monthly_Trend ------------------------------------------------------
    mt_idx = next((i for i, s in enumerate(sheets_lower) if "monthly" in s or "trend" in s), 3)
    if mt_idx < len(sheets):
        ws4 = wb[sheets[mt_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Monthly_Trend has at least 3 rows", len(data_rows4) >= 3, f"Found {len(data_rows4)}")


def check_teamly():
    print("\n=== Check 2: Teamly SLA Compliance Dashboard ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
    except Exception as e:
        check("Teamly 'SLA Compliance Dashboard' space has >=3 metric pages", False, str(e))
        return

    try:
        # Find the dashboard space: agent-created (id > 2; seed spaces are TEAM/TRIPS).
        cur.execute("SELECT id, key, name, COALESCE(description, '') FROM teamly.spaces")
        spaces = cur.fetchall()
        dash = None
        for sid, key, name, desc in spaces:
            blob = f"{key} {name} {desc}".lower()
            if "sla" in blob and ("compliance" in blob or "dashboard" in blob
                                  or "соответ" in blob or "дашборд" in blob or "панель" in blob):
                dash = (sid, name)
                break

        # Fallback: dashboard may have been modeled as a single page whose title
        # carries the marker (within any space).
        page_count = 0
        if dash is not None:
            cur.execute("SELECT COUNT(*) FROM teamly.pages WHERE space_id = %s", (dash[0],))
            page_count = cur.fetchone()[0]
        check("SLA Compliance Dashboard space exists", dash is not None,
              f"spaces: {[(s[1], s[2]) for s in spaces]}")
        check("Teamly 'SLA Compliance Dashboard' space has >=3 metric pages",
              dash is not None and page_count >= 3,
              f"space={dash}, pages={page_count}")
    except Exception as e:
        check("Teamly 'SLA Compliance Dashboard' space has >=3 metric pages", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 3: Monthly SLA Review Events ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
    except Exception as e:
        check("At least 3 SLA review events", False, str(e))
        check("Calendar has the 3 monthly SLA review events on first Mondays (Apr/May/Jun 2026), 10:00, 90 min",
              False, str(e))
        return

    cur.execute("""
        SELECT summary, start_datetime, end_datetime FROM gcal.events
        WHERE lower(summary) LIKE '%%sla%%review%%'
           OR lower(summary) LIKE '%%monthly%%sla%%'
           OR (lower(summary) LIKE '%%sla%%' AND (lower(summary) LIKE '%%review%%' OR lower(summary) LIKE '%%обзор%%'))
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("At least 3 SLA review events", len(events) >= 3, f"Found {len(events)} events")

    # CRITICAL: first Monday of Apr/May/Jun 2026, 10:00 start, 90 min duration.
    # 2026-04-06 (Mon), 2026-05-04 (Mon), 2026-06-01 (Mon).
    expected_dates = {"2026-04-06", "2026-05-04", "2026-06-01"}
    matched = set()
    for summary, start, end in events:
        if start is None:
            continue
        sdt = start if isinstance(start, datetime) else None
        if sdt is None:
            try:
                sdt = datetime.fromisoformat(str(start))
            except Exception:
                continue
        dkey = sdt.strftime("%Y-%m-%d")
        if dkey not in expected_dates:
            continue
        if sdt.hour != 10 or sdt.minute != 0:
            continue
        dur_ok = False
        if end is not None:
            edt = end if isinstance(end, datetime) else None
            if edt is None:
                try:
                    edt = datetime.fromisoformat(str(end))
                except Exception:
                    edt = None
            if edt is not None:
                mins = (edt - sdt).total_seconds() / 60.0
                dur_ok = abs(mins - 90) <= 1
        if dur_ok:
            matched.add(dkey)
    check("Calendar has the 3 monthly SLA review events on first Mondays (Apr/May/Jun 2026), 10:00, 90 min",
          matched == expected_dates, f"matched dates: {sorted(matched)}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: sla_analyzer.py ===")
    path = os.path.join(workspace, "sla_analyzer.py")
    check("sla_analyzer.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no negative compliance rate values in Ticket_Summary compliance col.
    path = os.path.join(workspace, "SLA_Compliance_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        has_negative = False
        for sheet_name in wb.sheetnames:
            sl = sheet_name.lower()
            # Allow negative gap values in Benchmark_Comparison (gap can be < 0).
            if "benchmark" in sl or "comparison" in sl:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative values in SLA report (excl. gap column)", not has_negative,
              "Found negative compliance/metric value")

    # Teamly: no duplicate SLA dashboard spaces.
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM teamly.spaces
            WHERE LOWER(name) LIKE '%%sla%%' OR LOWER(key) LIKE '%%sla%%'
        """)
        dup = cur.fetchone()[0]
        check("No duplicate SLA dashboard spaces in Teamly", dup <= 1, f"Found {dup} SLA spaces")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()
    check_gcal()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {"total_passed": PASS_COUNT, "total_checks": total,
              "accuracy": accuracy, "critical_failed": critical_failed}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Critical gate before the accuracy gate.
    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
