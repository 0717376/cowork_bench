"""Evaluation for sf-department-budget-analysis (ClickHouse HR_ANALYTICS).

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Pass threshold otherwise: accuracy >= 70%.

The agent reads Budget_Targets.pdf (Russian department names, RUB budgets) and
pulls employee data from ClickHouse HR_ANALYTICS. Department data values are
russified centrally (db/zzz_clickhouse_after_init.sql) and the groundtruth xlsx
Department cells are substituted by the same map, so this eval iterates the
groundtruth rows by key and never hardcodes department-name literals.
"""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    """Numeric closeness. Unlike the original, non-numeric cells where a numeric
    groundtruth is expected are treated as a MISMATCH (no silent str-compare pass)."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


# --- check accounting ---
PASS_COUNT = 0
TOTAL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
# These are the core derived/business metrics. Structural presence checks
# (sheet exists, row present) are intentionally NON-critical.
CRITICAL_CHECKS = {
    "Budget Analysis: all 7 departments present",
    "Budget Analysis: Budget_Utilization_Pct correct (all depts)",
    "Budget Analysis: Total_Salary_Cost correct (all depts)",
    "Budget Analysis: Actual_Headcount correct (all depts)",
    "Summary: Over_Budget_Depts exact",
    "Summary: Avg_Budget_Utilization correct",
}


def record(name, ok):
    global PASS_COUNT, TOTAL_COUNT
    TOTAL_COUNT += 1
    if ok:
        PASS_COUNT += 1
        print(f"  PASS: {name}")
    else:
        FAILED_NAMES.append(name)
        print(f"  FAIL: {name}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Department_Budget_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Department_Budget_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ===================== Sheet: Budget Analysis =====================
    print("Checking Budget Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Budget Analysis")
    g_rows = load_sheet_rows(gt_wb, "Budget Analysis")

    if a_rows is None or g_rows is None:
        record("Budget Analysis: sheet present", False)
        record("Budget Analysis: all 7 departments present", False)
        record("Budget Analysis: Budget_Utilization_Pct correct (all depts)", False)
        record("Budget Analysis: Total_Salary_Cost correct (all depts)", False)
        record("Budget Analysis: Actual_Headcount correct (all depts)", False)
    else:
        record("Budget Analysis: sheet present", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        g_keys = [str(r[0]).strip().lower() for r in g_data if r and r[0] is not None]

        # critical: all departments present
        missing = [k for k in g_keys if k not in a_lookup]
        record("Budget Analysis: all 7 departments present", len(missing) == 0)
        if missing:
            print(f"    Missing rows: {missing}")

        # per-metric aggregates across all matched departments
        budget_ok = True
        planned_ok = True
        actual_ok = True
        salary_ok = True
        total_cost_ok = True
        util_ok = True

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                # already counted as missing; force the value checks to fail
                budget_ok = planned_ok = actual_ok = False
                salary_ok = total_cost_ok = util_ok = False
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 100.0):
                    budget_ok = False
                    print(f"    {key}.Budget: {a_row[1]} vs {g_row[1]} (tol=100.0)")
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1):
                    planned_ok = False
                    print(f"    {key}.Planned_Headcount: {a_row[2]} vs {g_row[2]} (tol=1)")
            if len(a_row) > 3 and len(g_row) > 3:
                # tightened from tol=5 to tol=1 so the GROUP BY count is verified
                if not num_close(a_row[3], g_row[3], 1):
                    actual_ok = False
                    print(f"    {key}.Actual_Headcount: {a_row[3]} vs {g_row[3]} (tol=1)")
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 10.0):
                    salary_ok = False
                    print(f"    {key}.Avg_Salary: {a_row[4]} vs {g_row[4]} (tol=10.0)")
            if len(a_row) > 5 and len(g_row) > 5:
                # tightened from tol=1000 to tol=100 to confirm SUM(SALARY) join
                if not num_close(a_row[5], g_row[5], 100.0):
                    total_cost_ok = False
                    print(f"    {key}.Total_Salary_Cost: {a_row[5]} vs {g_row[5]} (tol=100.0)")
            if len(a_row) > 6 and len(g_row) > 6:
                # tightened from tol=1.0 to tol=0.5 for the core derived metric
                if not num_close(a_row[6], g_row[6], 0.5):
                    util_ok = False
                    print(f"    {key}.Budget_Utilization_Pct: {a_row[6]} vs {g_row[6]} (tol=0.5)")

        record("Budget Analysis: Budget correct (all depts)", budget_ok)
        record("Budget Analysis: Planned_Headcount correct (all depts)", planned_ok)
        record("Budget Analysis: Actual_Headcount correct (all depts)", actual_ok)
        record("Budget Analysis: Avg_Salary correct (all depts)", salary_ok)
        record("Budget Analysis: Total_Salary_Cost correct (all depts)", total_cost_ok)
        record("Budget Analysis: Budget_Utilization_Pct correct (all depts)", util_ok)

    # ===================== Sheet: Summary =====================
    print("Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    if a_rows is None or g_rows is None:
        record("Summary: sheet present", False)
        record("Summary: Over_Budget_Depts exact", False)
        record("Summary: Avg_Budget_Utilization correct", False)
    else:
        record("Summary: sheet present", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        # Per-metric tolerances: split the old single tol=1000 so the
        # semantically meaningful metrics are not swallowed.
        # Big absolute sums keep a loose tol; ratio/count get tight tol.
        metric_tol = {
            "total_budget": 1000.0,
            "total_salary_cost": 1000.0,
            "avg_budget_utilization": 0.5,
            "over_budget_depts": 0.0,
        }

        results = {}
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            tol = metric_tol.get(key, 1.0)
            a_row = a_lookup.get(key)
            if a_row is None or len(a_row) < 2 or len(g_row) < 2:
                results[key] = False
                print(f"    Missing/short Summary row: {g_row[0]}")
                continue
            ok = num_close(a_row[1], g_row[1], tol)
            results[key] = ok
            if not ok:
                print(f"    {key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")

        record("Summary: Total_Budget correct", results.get("total_budget", False))
        record("Summary: Total_Salary_Cost correct", results.get("total_salary_cost", False))
        record("Summary: Avg_Budget_Utilization correct", results.get("avg_budget_utilization", False))
        record("Summary: Over_Budget_Depts exact", results.get("over_budget_depts", False))

    # ===================== Verdict =====================
    total = TOTAL_COUNT if TOTAL_COUNT else 1
    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")
        print("\n=== RESULT: FAIL (critical check failed) ===")
        sys.exit(1)

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL (accuracy below 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
