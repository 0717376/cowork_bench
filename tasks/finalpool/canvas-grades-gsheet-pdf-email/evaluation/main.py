"""Evaluation script for canvas-grades-gsheet-pdf-email.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. These capture the semantic substance of the deliverable (correct
aggregated values, the C>=70 pass rule from the policy PDF, the targeted email).
Structural checks (sheet exists, column present) are non-critical.
"""
import os
import argparse, json, os, sys
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Department_Summary Total_Students == 25820",
    "Department_Summary Highest_Avg_Course == Экологическая экономика и этика (Осень 2014)",
    "Department_Summary Lowest_Avg_Course == Креативные вычисления и культура (Весна 2014)",
    "Grade_Distribution Pass_Rate_Pct matches groundtruth for all courses",
    "Grade_Distribution Course_Avg matches groundtruth for all rows",
    "Email to dept-heads@university.edu with Grade Distribution subject",
}

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def _rows_as_dicts(ws):
    """Return list of {header: value} dicts for data rows of a worksheet."""
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                d[h] = row[i]
        out.append(d)
    return headers, out

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    summary_metrics = {}        # Metric -> Value from agent Department_Summary
    grade_dist_by_course = {}   # Course_Name -> dict from agent Grade_Distribution

    # Check Grade_Dashboard_Reference.xlsx
    excel_path = os.path.join(agent_workspace, "Grade_Dashboard_Reference.xlsx")
    check("Grade_Dashboard_Reference.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        gt_path = os.path.join(groundtruth_workspace, "Grade_Dashboard_Reference.xlsx")
        gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

        if gt_wb:
            for sheet_name in gt_wb.sheetnames:
                check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    gt_ws = gt_wb[sheet_name]
                    # Check headers
                    gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                    for h in gt_headers:
                        if h:
                            check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                    # Check row count
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                    min_rows = max(1, len(gt_rows) - 2)
                    check(f"{sheet_name} has >= {min_rows} data rows", len(data_rows) >= min_rows, f"got {len(data_rows)}")

                    # Cell value comparison against groundtruth for ALL rows.
                    header_map = {h: i for i, h in enumerate(headers)}
                    for ri in range(min(len(gt_rows), len(data_rows))):
                        gt_row = gt_rows[ri]
                        agent_row = data_rows[ri]
                        for ci, gt_h in enumerate(gt_headers):
                            if not gt_h or ci >= len(gt_row):
                                continue
                            gv = gt_row[ci]
                            agent_ci = header_map.get(gt_h)
                            if agent_ci is None or agent_ci >= len(agent_row):
                                continue
                            av = agent_row[agent_ci]
                            gf = safe_float(gv)
                            af = safe_float(av)
                            if gf is not None and af is not None:
                                check(f"{sheet_name} R{ri+2} {gt_h} ~{gf:.1f}",
                                      num_close(af, gf), f"got {af}")
                            elif gv is not None and av is not None:
                                gs = str(gv).strip().lower()
                                avs = str(av).strip().lower()
                                if gs:
                                    check(f"{sheet_name} R{ri+2} {gt_h} text",
                                          gs == avs or gs in avs or avs in gs,
                                          f"expected {gs[:50]}, got {avs[:50]}")

            # Build lookup dicts for the semantic CRITICAL checks.
            if "Department_Summary" in wb.sheetnames:
                _, rows = _rows_as_dicts(wb["Department_Summary"])
                for d in rows:
                    m = d.get("Metric")
                    if m is not None:
                        summary_metrics[str(m).strip()] = d.get("Value")
            if "Grade_Distribution" in wb.sheetnames:
                _, rows = _rows_as_dicts(wb["Grade_Distribution"])
                for d in rows:
                    cn = d.get("Course_Name")
                    if cn is not None:
                        grade_dist_by_course[str(cn).strip()] = d

            # Groundtruth per-course Pass_Rate_Pct / Course_Avg for critical comparison.
            gt_grade = {}
            if "Grade_Distribution" in gt_wb.sheetnames:
                _, gtrows = _rows_as_dicts(gt_wb["Grade_Distribution"])
                for d in gtrows:
                    cn = d.get("Course_Name")
                    if cn is not None:
                        gt_grade[str(cn).strip()] = d

            # --- CRITICAL: Total_Students == 25820 (real Canvas enrollment aggregated) ---
            ts = safe_float(summary_metrics.get("Total_Students"))
            check("Department_Summary Total_Students == 25820",
                  ts is not None and num_close(ts, 25820), f"got {ts}")

            # --- CRITICAL: Highest/Lowest avg course identity ---
            hi = str(summary_metrics.get("Highest_Avg_Course", "")).strip().lower()
            lo = str(summary_metrics.get("Lowest_Avg_Course", "")).strip().lower()
            check("Department_Summary Highest_Avg_Course == Экологическая экономика и этика (Осень 2014)",
                  "экологическая экономика и этика (осень 2014)" in hi, f"got {hi}")
            check("Department_Summary Lowest_Avg_Course == Креативные вычисления и культура (Весна 2014)",
                  "креативные вычисления и культура (весна 2014)" in lo, f"got {lo}")

            # --- CRITICAL: Pass_Rate_Pct matches groundtruth for ALL courses (C>=70 rule) ---
            pr_ok = bool(gt_grade) and len(grade_dist_by_course) > 0
            pr_detail = ""
            for cn, gtd in gt_grade.items():
                ad = grade_dist_by_course.get(cn)
                gpr = safe_float(gtd.get("Pass_Rate_Pct"))
                apr = safe_float(ad.get("Pass_Rate_Pct")) if ad else None
                if gpr is None:
                    continue
                if apr is None or not num_close(apr, gpr):
                    pr_ok = False
                    pr_detail = f"{cn}: expected {gpr}, got {apr}"
                    break
            check("Grade_Distribution Pass_Rate_Pct matches groundtruth for all courses",
                  pr_ok, pr_detail)

            # --- CRITICAL: Course_Avg matches groundtruth for ALL courses ---
            ca_ok = bool(gt_grade) and len(grade_dist_by_course) > 0
            ca_detail = ""
            for cn, gtd in gt_grade.items():
                ad = grade_dist_by_course.get(cn)
                gca = safe_float(gtd.get("Course_Avg"))
                aca = safe_float(ad.get("Course_Avg")) if ad else None
                if gca is None:
                    continue
                if aca is None or not num_close(aca, gca):
                    ca_ok = False
                    ca_detail = f"{cn}: expected {gca}, got {aca}"
                    break
            check("Grade_Distribution Course_Avg matches groundtruth for all rows",
                  ca_ok, ca_detail)

    # Check Python script exists (terminal usage)
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")

    # Database checks
    try:
        conn = get_conn()
        cur = conn.cursor()
        # CRITICAL: email to dept-heads@university.edu with "Grade Distribution" subject.
        cur.execute(
            "SELECT subject, to_addr, body_text FROM email.messages "
            "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
            "AND subject ILIKE '%grade distribution%'"
        )
        email_rows = cur.fetchall()
        matched = None
        for r in email_rows:
            to_addr = r[1]
            to_str = ""
            if to_addr is not None:
                try:
                    parsed = json.loads(to_addr) if isinstance(to_addr, str) else to_addr
                    to_str = " ".join(parsed) if isinstance(parsed, list) else str(parsed)
                except Exception:
                    to_str = str(to_addr)
            if "dept-heads@university.edu" in to_str.lower():
                matched = r
                break
        check("Email to dept-heads@university.edu with Grade Distribution subject",
              matched is not None,
              f"candidates: {[(r[0], r[1]) for r in email_rows]}")
        if matched:
            # Non-critical: body should mention the <70% review note.
            body = str(matched[2] or "").lower()
            check("Email body notes courses below 70% pass rate",
                  ("70" in body) and ("проверк" in body or "review" in body or "пропуск" in body
                                      or "ниже" in body or "below" in body),
                  f"body snippet: {body[:120]}")

        cur.execute("SELECT COUNT(*) FROM gsheet.spreadsheets")
        ss_count = cur.fetchone()[0]
        check("Google Sheet created", ss_count >= 1, f"spreadsheet count: {ss_count}")
        cur.execute("SELECT COUNT(*) FROM gsheet.cells")
        cell_count = cur.fetchone()[0]
        check("Sheet has data", cell_count >= 10, f"cell count: {cell_count}")
        # Reverse verification: noise emails should not be in Sent folder.
        cur.execute(
            "SELECT COUNT(*) FROM email.messages "
            "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
            "AND (subject ILIKE '%рассылка%' OR subject ILIKE '%техобслуживание%' OR subject ILIKE '%newsletter%')"
        )
        noise_sent = cur.fetchone()[0]
        check("No noise emails in Sent folder", noise_sent == 0, f"found {noise_sent} noise emails in Sent")
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
