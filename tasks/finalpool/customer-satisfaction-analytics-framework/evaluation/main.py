#!/usr/bin/env python3
"""Скрипт оценки задачи customer-satisfaction-analytics-framework.

Критические чеки (CRITICAL_CHECKS): любой их fail = задача FAIL, даже при высокой
общей accuracy. Это семантические проверки сути результата (правильные значения
консолидированных метрик, заполненный план мероприятий, ключевые числа в отчётах).
Структурные проверки (файл существует, лист есть, в листе есть строки) —
НЕкритические.
"""

from argparse import ArgumentParser
import json
import os
import sys

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Критические чеки (по строке name, как при record()).
CRITICAL_CHECKS = {
    "satisfaction_analysis.xlsx: Overall Satisfaction Score == 4.2",
    "satisfaction_analysis.xlsx: Net Promoter Score (NPS) == 42",
    "satisfaction_analysis.xlsx: First Contact Resolution Rate == 0.75",
    "satisfaction_analysis.xlsx: Average Resolution Time ~= 3.4 ч",
    "action_plans.xlsx: заголовки Action Plans + >=3 инициатив (Owner/метрики)",
    "satisfaction_report.docx: ключевые метрики (4.2, 42, 75%, 3.4) в тексте",
    "executive_summary.docx: ключевые метрики (4.2, 42, 75%) в тексте",
    "executive_summary.docx: вывод об отставании сегмента SMB",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _to_float(v):
    """Привести значение к float, учитывая проценты и '3.4 hrs/часа'."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().lower().replace(",", ".")
    pct = "%" in s or "процент" in s
    num = ""
    seen_dot = False
    for ch in s:
        if ch.isdigit():
            num += ch
        elif ch == "." and not seen_dot:
            num += ch
            seen_dot = True
        elif num:
            break
    if not num or num == ".":
        return None
    try:
        f = float(num)
    except ValueError:
        return None
    if pct and f > 1.5:
        f = f / 100.0
    return f


# ---------------------------------------------------------------------------
# Поиск метрики по метке в любом листе satisfaction_analysis.xlsx
# ---------------------------------------------------------------------------
def _find_metric(rows, label_substrings):
    """Вернуть числовое значение из строки, первая ячейка которой содержит
    одну из подстрок метки (без учёта регистра)."""
    for r in rows:
        if not r:
            continue
        first = str(r[0]).strip().lower() if r[0] is not None else ""
        if any(sub in first for sub in label_substrings):
            for cell in r[1:]:
                f = _to_float(cell)
                if f is not None:
                    return f
    return None


def check_satisfaction_analysis(workspace):
    print("\n=== Check: satisfaction_analysis.xlsx ===")
    import openpyxl

    path = os.path.join(workspace, "satisfaction_analysis.xlsx")
    if not os.path.isfile(path):
        record("xlsx satisfaction_analysis.xlsx exists", False, "Not found")
        # все зависимые критические чеки помечаем как fail
        record("satisfaction_analysis.xlsx: Overall Satisfaction Score == 4.2", False, "no file")
        record("satisfaction_analysis.xlsx: Net Promoter Score (NPS) == 42", False, "no file")
        record("satisfaction_analysis.xlsx: First Contact Resolution Rate == 0.75", False, "no file")
        record("satisfaction_analysis.xlsx: Average Resolution Time ~= 3.4 ч", False, "no file")
        return
    record("xlsx satisfaction_analysis.xlsx exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    all_rows = []
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        record(f"xlsx satisfaction_analysis.xlsx '{ws.title}' has data",
               len(rows) >= 2, f"{len(rows)} rows")
        all_rows.extend(rows)

    oss = _find_metric(all_rows, ["overall satisfaction"])
    record("satisfaction_analysis.xlsx: Overall Satisfaction Score == 4.2",
           oss is not None and num_close(oss, 4.2, 0.1), f"got {oss}")

    nps = _find_metric(all_rows, ["net promoter", "nps"])
    record("satisfaction_analysis.xlsx: Net Promoter Score (NPS) == 42",
           nps is not None and num_close(nps, 42, 1.0), f"got {nps}")

    fcr = _find_metric(all_rows, ["first contact resolution rate", "first contact"])
    record("satisfaction_analysis.xlsx: First Contact Resolution Rate == 0.75",
           fcr is not None and num_close(fcr, 0.75, 0.02), f"got {fcr}")

    art = _find_metric(all_rows, ["average resolution time", "resolution time"])
    record("satisfaction_analysis.xlsx: Average Resolution Time ~= 3.4 ч",
           art is not None and num_close(art, 3.4, 0.3), f"got {art}")

    wb.close()


def check_action_plans(workspace):
    print("\n=== Check: action_plans.xlsx ===")
    import openpyxl

    path = os.path.join(workspace, "action_plans.xlsx")
    if not os.path.isfile(path):
        record("xlsx action_plans.xlsx exists", False, "Not found")
        record("action_plans.xlsx: заголовки Action Plans + >=3 инициатив (Owner/метрики)",
               False, "no file")
        return
    record("xlsx action_plans.xlsx exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    for ws in wb.worksheets:
        rows.extend(list(ws.iter_rows(values_only=True)))

    expected_headers = ["initiative", "owner", "target metric", "baseline",
                        "goal", "timeline", "status"]
    header_idx = None
    for i, r in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in r]
        if all(h in cells for h in expected_headers):
            header_idx = i
            break

    record("xlsx action_plans.xlsx 'Action Plans' has data", len(rows) >= 2, f"{len(rows)} rows")

    if header_idx is None:
        record("action_plans.xlsx: заголовки Action Plans + >=3 инициатив (Owner/метрики)",
               False, "header row with all 7 columns not found")
        wb.close()
        return

    # Колонки Initiative (0) и Owner (1) по порядку заголовка.
    header = [str(c).strip().lower() if c is not None else "" for c in rows[header_idx]]
    init_col = header.index("initiative")
    owner_col = header.index("owner")

    init_rows = 0
    for r in rows[header_idx + 1:]:
        if not r or init_col >= len(r):
            continue
        init = r[init_col]
        owner = r[owner_col] if owner_col < len(r) else None
        # строка инициативы: непустое название и непустой ответственный
        if init and str(init).strip() and owner and str(owner).strip():
            init_rows += 1

    record("action_plans.xlsx: заголовки Action Plans + >=3 инициатив (Owner/метрики)",
           init_rows >= 3, f"{init_rows} populated initiative rows")
    wb.close()


def _docx_text(path):
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


def _has_num(text, target, tol):
    """Есть ли в тексте число, близкое к target (учёт '4.2', '42', '75%', '3.4')."""
    import re
    for m in re.findall(r"\d+(?:[.,]\d+)?", text):
        try:
            v = float(m.replace(",", "."))
        except ValueError:
            continue
        if abs(v - target) <= tol:
            return True
    return False


def check_docx_content(workspace):
    print("\n=== Check: DOCX files ===")

    # satisfaction_report.docx
    rp = os.path.join(workspace, "satisfaction_report.docx")
    if not os.path.isfile(rp):
        record("docx satisfaction_report.docx exists", False, "Not found")
        record("satisfaction_report.docx: ключевые метрики (4.2, 42, 75%, 3.4) в тексте",
               False, "no file")
    else:
        record("docx satisfaction_report.docx exists", True)
        try:
            txt = _docx_text(rp)
            record("docx satisfaction_report.docx has content",
                   len(txt.strip()) > 0, f"{len(txt)} chars")
            # 4.2, 42, 75%(=0.75 как доля или 75 как процент), 3.4
            ok = (_has_num(txt, 4.2, 0.05) and _has_num(txt, 42, 0.5)
                  and (_has_num(txt, 75, 0.5) or _has_num(txt, 0.75, 0.01))
                  and _has_num(txt, 3.4, 0.05))
            record("satisfaction_report.docx: ключевые метрики (4.2, 42, 75%, 3.4) в тексте",
                   ok, "missing one of 4.2/42/75/3.4")
        except Exception as e:
            record("docx satisfaction_report.docx readable", False, str(e))
            record("satisfaction_report.docx: ключевые метрики (4.2, 42, 75%, 3.4) в тексте",
                   False, str(e))

    # executive_summary.docx
    ep = os.path.join(workspace, "executive_summary.docx")
    if not os.path.isfile(ep):
        record("docx executive_summary.docx exists", False, "Not found")
        record("executive_summary.docx: ключевые метрики (4.2, 42, 75%) в тексте",
               False, "no file")
        record("executive_summary.docx: вывод об отставании сегмента SMB",
               False, "no file")
    else:
        record("docx executive_summary.docx exists", True)
        try:
            txt = _docx_text(ep)
            record("docx executive_summary.docx has content",
                   len(txt.strip()) > 0, f"{len(txt)} chars")
            ok = (_has_num(txt, 4.2, 0.05) and _has_num(txt, 42, 0.5)
                  and (_has_num(txt, 75, 0.5) or _has_num(txt, 0.75, 0.01)))
            record("executive_summary.docx: ключевые метрики (4.2, 42, 75%) в тексте",
                   ok, "missing one of 4.2/42/75")
            # Вывод об отставании сегмента SMB: упомянут SMB и слово об отставании
            # (underperform/lag/lowest/отстаёт/наименьш...). Конкретные числовые
            # значения сегментов (3.8/4.5) задачей не оговорены и не выводятся из
            # входных данных, поэтому литералы не требуем.
            low = txt.lower()
            lag_kw = ("underperform", "underperforms", "lag", "lags", "lagging",
                      "lowest", "weakest", "behind", "gap",
                      "отста", "наименьш", "ниже", "слаб", "хуже")
            smb_ok = ("smb" in low and any(k in low for k in lag_kw))
            record("executive_summary.docx: вывод об отставании сегмента SMB",
                   smb_ok, "need SMB + lag/underperform keyword")
        except Exception as e:
            record("docx executive_summary.docx readable", False, str(e))
            record("executive_summary.docx: ключевые метрики (4.2, 42, 75%) в тексте",
                   False, str(e))
            record("executive_summary.docx: вывод об отставании сегмента SMB",
                   False, str(e))


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    ws = args.agent_workspace
    if not os.path.isdir(ws):
        print(f"Agent workspace not found: {ws}")
        sys.exit(1)

    check_satisfaction_analysis(ws)
    check_action_plans(ws)
    check_docx_content(ws)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": CRITICAL_FAILED,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"CRITICAL FAIL: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL (accuracy < 70%)")
        sys.exit(1)


if __name__ == "__main__":
    main()
