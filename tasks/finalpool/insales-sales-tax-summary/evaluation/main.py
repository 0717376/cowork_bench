"""
Evaluation script for insales-sales-tax-summary task (InSales store, wc.* schema).

Checks:
1. Excel file Tax_Summary_Report.xlsx with By State and Overall sheets
2. Google Sheet "Tax Summary for Accounting" with summary data

All expected values are recomputed LIVE from wc.orders, so the eval stays in
sync with the centrally-russified seed (db/zzz_wc_after_init.sql). DO NOT
hardcode any wc realia data-value literals here.
"""
import argparse
import json
import os
import sys

import psycopg2


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT billing->>'state' as state,
               COUNT(*) as order_count,
               ROUND(SUM(total::numeric), 2) as total_sales,
               ROUND(SUM(total_tax::numeric), 2) as total_tax
        FROM wc.orders
        WHERE status IN ('completed', 'processing')
        AND billing->>'state' IS NOT NULL
        AND billing->>'state' != ''
        GROUP BY billing->>'state'
        ORDER BY SUM(total_tax::numeric) DESC
    """)
    state_rows = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*),
               ROUND(SUM(total::numeric), 2),
               ROUND(SUM(total_tax::numeric), 2)
        FROM wc.orders
        WHERE status IN ('completed', 'processing')
    """)
    ov = cur.fetchone()
    total_orders = int(ov[0])
    total_sales = float(ov[1])
    total_tax = float(ov[2])
    eff_rate = round(total_tax / total_sales * 100, 2) if total_sales > 0 else 0.0

    cur.close()
    conn.close()

    return {
        "states": [{"state": r[0], "count": int(r[1]), "sales": float(r[2]),
                     "tax": float(r[3])} for r in state_rows],
        "overall": {
            "total_orders": total_orders,
            "total_sales": total_sales,
            "total_tax": total_tax,
            "effective_rate": eff_rate,
        }
    }


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}{detail_str}")
        if critical:
            CRITICAL_FAILS.append(name)


def number_in_text(value, text, tolerance=0.01):
    text = str(text)
    val_str = str(value)
    if val_str in text:
        return True
    try:
        f2 = f"{float(value):.2f}"
        if f2 in text:
            return True
    except (ValueError, TypeError):
        pass
    try:
        formatted = f"{int(value):,}"
        if formatted in text:
            return True
    except (ValueError, TypeError):
        pass
    return False


def _to_float(v):
    """Parse a cell value into a float, stripping currency/percent/thousands."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    s = s.replace("%", "").replace("$", "").replace("₽", "")
    s = s.replace(",", "").replace(" ", "").replace(" ", "")
    try:
        return float(s)
    except ValueError:
        return None


def _rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Tax_Summary_Report.xlsx")
    file_ok = os.path.isfile(xlsx_path)
    check("Excel file Tax_Summary_Report.xlsx exists", file_ok,
          f"Expected {xlsx_path}", critical=True)
    if not file_ok:
        return

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        check("Excel file readable", False, str(e), critical=True)
        return

    def find_sheet(keywords):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(k in sl for k in keywords):
                return wb[s]
        return None

    def sheet_text(ws):
        txt = ""
        for row in ws.iter_rows(values_only=True):
            txt += " ".join(str(c) for c in row if c is not None) + " "
        return txt

    # ---------- By State sheet ----------
    ws_state = find_sheet(["state"])
    check("By State sheet exists", ws_state is not None, f"Sheets: {wb.sheetnames}")

    if ws_state is not None:
        txt = sheet_text(ws_state)
        # Loose (non-critical) presence checks for the top-5 states by tax.
        for s in expected["states"][:5]:
            check(f"State '{s['state']}' in sheet", s["state"] in txt)
            check(f"Tax {s['tax']} for {s['state']}",
                  number_in_text(s["tax"], txt))

        # Build a state -> per-row tax map keyed row-wise from the sheet.
        rows = _rows(ws_state)
        # Locate the State column and the Total Tax column from a header row.
        state_col = tax_col = None
        header_idx = None
        for ri, row in enumerate(rows):
            lc = [str(c).strip().lower() if c is not None else "" for c in row]
            if "state" in lc and any("tax" in c for c in lc):
                state_col = lc.index("state")
                tax_col = next(i for i, c in enumerate(lc) if "tax" in c)
                header_idx = ri
                break

        have_headers = state_col is not None and tax_col is not None
        check("By State sheet has State + Total Tax header columns", have_headers,
              f"rows[:3]={rows[:3]}", critical=True)

        if have_headers:
            data_rows = []
            for row in rows[header_idx + 1:]:
                if state_col >= len(row):
                    continue
                st = row[state_col]
                if st is None or str(st).strip() == "":
                    continue
                tax = _to_float(row[tax_col]) if tax_col < len(row) else None
                data_rows.append((str(st).strip(), tax))

            sheet_tax_by_state = {st: tx for st, tx in data_rows}

            # CRITICAL: every qualifying state present with row-wise correct tax.
            exp_states = expected["states"]
            all_states_ok = True
            mismatch = ""
            for s in exp_states:
                got = sheet_tax_by_state.get(s["state"])
                if got is None or not num_close(got, s["tax"]):
                    all_states_ok = False
                    mismatch = f"state={s['state']} expected_tax={s['tax']} got={got}"
                    break
            check("By State: every qualifying state present with row-wise correct Total Tax",
                  all_states_ok and len(exp_states) > 0, mismatch, critical=True)

            # CRITICAL: rows sorted by Total Tax descending (monotonic non-increasing).
            ordered_taxes = [tx for _, tx in data_rows if tx is not None]
            monotonic = all(
                ordered_taxes[i] >= ordered_taxes[i + 1] - 1e-6
                for i in range(len(ordered_taxes) - 1)
            )
            check("By State: rows ordered by Total Tax descending",
                  monotonic and len(ordered_taxes) >= 1,
                  f"taxes={ordered_taxes}", critical=True)

    # ---------- Overall sheet ----------
    ws_ov = find_sheet(["overall"])
    check("Overall sheet exists", ws_ov is not None, f"Sheets: {wb.sheetnames}")

    ov = expected["overall"]
    label_map = {
        "total orders": ("Total Orders", ov["total_orders"]),
        "total sales": ("Total Sales", ov["total_sales"]),
        "total tax": ("Total Tax", ov["total_tax"]),
        "effective tax rate": ("Effective Tax Rate", ov["effective_rate"]),
    }

    if ws_ov is not None:
        txt = sheet_text(ws_ov)
        # Loose presence (non-critical).
        check(f"Total Orders = {ov['total_orders']}",
              number_in_text(ov["total_orders"], txt))
        check(f"Total Sales = {ov['total_sales']}",
              number_in_text(ov["total_sales"], txt))
        check(f"Total Tax = {ov['total_tax']}",
              number_in_text(ov["total_tax"], txt))
        check(f"Effective Tax Rate = {ov['effective_rate']}",
              number_in_text(ov["effective_rate"], txt))

        # CRITICAL: Metric label associated row-wise to the correct Value.
        rows = _rows(ws_ov)
        found = {}
        for row in rows:
            cells = [c for c in row]
            for ci, c in enumerate(cells):
                if c is None:
                    continue
                key = str(c).strip().lower()
                if key in label_map:
                    # Value is the first numeric cell after the label in the row.
                    val = None
                    for c2 in cells[ci + 1:]:
                        f = _to_float(c2)
                        if f is not None:
                            val = f
                            break
                    found[key] = val

        all_ov_ok = True
        ov_detail = ""
        for key, (label, exp_val) in label_map.items():
            got = found.get(key)
            ok = got is not None and num_close(got, exp_val)
            if not ok:
                all_ov_ok = False
                ov_detail = f"{label}: expected {exp_val} got {got}"
                break
        check("Overall: all four metric labels associated row-wise to correct Value",
              all_ov_ok, ov_detail, critical=True)

        # CRITICAL: Effective Tax Rate is the derived metric, not arbitrary.
        got_rate = found.get("effective tax rate")
        rate_ok = got_rate is not None and num_close(got_rate, ov["effective_rate"], rel_tol=0.02, abs_tol=0.05)
        check("Overall: Effective Tax Rate == round(total_tax/total_sales*100, 2)",
              rate_ok, f"expected {ov['effective_rate']} got {got_rate}", critical=True)


def check_gsheet(expected):
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Broaden title match to RU + EN: accept either the English title or a
    # Russian rendering ("налог" + "бухгалтер") so a legitimately-russified
    # agent title still matches.
    cur.execute("""
        SELECT id, title FROM gsheet.spreadsheets
        WHERE (LOWER(title) LIKE '%%tax%%' AND LOWER(title) LIKE '%%accounting%%')
           OR (LOWER(title) LIKE '%%налог%%' AND LOWER(title) LIKE '%%бухгалт%%')
    """)
    sheets = cur.fetchall()
    check("Google Sheet 'Tax Summary for Accounting' exists",
          len(sheets) >= 1,
          f"Found {len(sheets)} matching spreadsheets", critical=True)

    if sheets:
        ss_id = sheets[0][0]
        cur.execute("""
            SELECT c.value FROM gsheet.cells c
            JOIN gsheet.sheets s ON c.spreadsheet_id = s.spreadsheet_id AND c.sheet_id = s.id
            WHERE c.spreadsheet_id = %s
        """, (ss_id,))
        cells = cur.fetchall()
        all_values = " ".join(str(c[0]) for c in cells if c[0])

        ov = expected["overall"]
        orders_ok = number_in_text(ov["total_orders"], all_values)
        sales_ok = number_in_text(ov["total_sales"], all_values)
        tax_ok = number_in_text(ov["total_tax"], all_values)
        check("GSheet contains Total Orders", orders_ok)
        check("GSheet contains Total Sales", sales_ok)
        check("GSheet contains Total Tax", tax_ok)
        # CRITICAL: the Summary sheet carries the live-DB Overall figures.
        check("GSheet Summary has live-DB Total Orders / Total Sales / Total Tax",
              orders_ok and sales_ok and tax_ok, critical=True)

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = load_expected()

    check_excel(args.agent_workspace, expected)
    check_gsheet(expected)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES ({len(CRITICAL_FAILS)}): {CRITICAL_FAILS}")
        print(f"  Overall: FAIL (critical)")
        sys.exit(1)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"  Accuracy: {accuracy:.1f}%")
    passed = accuracy >= 70.0
    print(f"  Overall: {'PASS' if passed else 'FAIL'}")
    sys.exit(0 if passed else 1)


if __name__ == "__main__":
    main()
