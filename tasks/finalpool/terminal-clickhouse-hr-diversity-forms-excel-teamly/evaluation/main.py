"""Evaluation for terminal-sf-hr-diversity-gform-excel-notion."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Семантические проверки (ядро задачи). Любой провал критической проверки =>
# overall FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Department_Breakdown Employee_Count all match groundtruth",
    "Department_Breakdown Diversity_Index all match groundtruth",
    "Summary headline results match groundtruth",
    "Teamly 'Diversity Metrics Dashboard' space has 7 department pages with content",
    "Diversity survey has exactly 5 questions",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def check_excel(agent_ws, gt_dir):
    print("\n=== Checking Diversity_Metrics_Report.xlsx ===")
    agent_file = os.path.join(agent_ws, "Diversity_Metrics_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Diversity_Metrics_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
        gwb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Department_Breakdown
    print("  Checking Department_Breakdown...")
    ws = get_sheet(awb, "Department_Breakdown")
    gws = get_sheet(gwb, "Department_Breakdown")
    check("Sheet Department_Breakdown exists", ws is not None, f"Sheets: {awb.sheetnames}")
    if ws and gws:
        a_rows = list(ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(gws.iter_rows(min_row=2, values_only=True))
        check("Department_Breakdown has 35 rows", len(a_rows) == 35, f"Got {len(a_rows)}")

        a_lookup = {}
        for row in a_rows:
            if row and row[0] and row[1]:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                a_lookup[key] = row

        count_mismatches = []
        index_mismatches = []
        for g_row in g_rows:
            if not g_row or not g_row[0]:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Row {g_row[0]}/{g_row[1]} exists", False, "Missing")
                count_mismatches.append(key)
                index_mismatches.append(key)
                continue
            ok_count = num_close(a_row[2], g_row[2], 1)
            check(f"{g_row[0]}/{g_row[1]} Employee_Count",
                  ok_count,
                  f"Expected {g_row[2]}, got {a_row[2]}")
            if not ok_count:
                count_mismatches.append(key)
            if len(a_row) > 4 and len(g_row) > 4:
                ok_idx = num_close(a_row[4], g_row[4], 0.01)
                check(f"{g_row[0]}/{g_row[1]} Diversity_Index",
                      ok_idx,
                      f"Expected {g_row[4]}, got {a_row[4]}")
                if not ok_idx:
                    index_mismatches.append(key)

        # CRITICAL: все 35 значений Employee_Count и Diversity_Index верны.
        check("Department_Breakdown Employee_Count all match groundtruth",
              len(count_mismatches) == 0,
              f"Mismatches: {count_mismatches[:5]}")
        check("Department_Breakdown Diversity_Index all match groundtruth",
              len(index_mismatches) == 0,
              f"Mismatches: {index_mismatches[:5]}")

    # Sheet 2: Education_Analysis
    print("  Checking Education_Analysis...")
    ws2 = get_sheet(awb, "Education_Analysis")
    gws2 = get_sheet(gwb, "Education_Analysis")
    check("Sheet Education_Analysis exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    if ws2 and gws2:
        a_rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        g_rows2 = list(gws2.iter_rows(min_row=2, values_only=True))
        check("Education_Analysis has 5 rows", len(a_rows2) == 5, f"Got {len(a_rows2)}")

        a_lookup2 = {str(r[0]).strip().lower(): r for r in a_rows2 if r and r[0]}
        for g_row in g_rows2:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup2.get(key)
            if a_row is None:
                check(f"Education '{g_row[0]}' exists", False, "Missing")
                continue
            check(f"'{g_row[0]}' Total_Count",
                  num_close(a_row[1], g_row[1], 1),
                  f"Expected {g_row[1]}, got {a_row[1]}")

    # Sheet 3: Survey_Config
    print("  Checking Survey_Config...")
    ws3 = get_sheet(awb, "Survey_Config")
    check("Sheet Survey_Config exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        a_rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        check("Survey_Config has 5 questions", len(a_rows3) == 5, f"Got {len(a_rows3)}")

    # Sheet 4: Summary
    print("  Checking Summary...")
    ws4 = get_sheet(awb, "Summary")
    gws4 = get_sheet(gwb, "Summary")
    check("Sheet Summary exists", ws4 is not None, f"Sheets: {awb.sheetnames}")
    if ws4 and gws4:
        a_summary = {}
        for row in ws4.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        g_summary = {}
        for row in gws4.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                g_summary[str(row[0]).strip().lower()] = row[1]

        # Dynamic DB queries for expected values
        expected_total = 50000
        expected_num_depts = 7
        try:
            conn2 = psycopg2.connect(**DB)
            cur2 = conn2.cursor()
            cur2.execute("SELECT COUNT(*) FROM sf_data.employees")
            result = cur2.fetchone()
            if result and result[0]:
                expected_total = result[0]
            cur2.execute("SELECT COUNT(DISTINCT department) FROM sf_data.employees")
            result = cur2.fetchone()
            if result and result[0]:
                expected_num_depts = result[0]
            cur2.close()
            conn2.close()
        except Exception:
            pass

        check("Total_Employees correct",
              num_close(a_summary.get("total_employees"), expected_total, 5),
              f"Got {a_summary.get('total_employees')}, expected {expected_total}")
        check("Num_Departments correct",
              num_close(a_summary.get("num_departments"), expected_num_depts, 0),
              f"Got {a_summary.get('num_departments')}, expected {expected_num_depts}")
        avg_ok = num_close(a_summary.get("avg_diversity_index"),
                           g_summary.get("avg_diversity_index", 1.3481), 0.02)
        check("Avg_Diversity_Index close to groundtruth", avg_ok,
              f"Got {a_summary.get('avg_diversity_index')}")
        meet_ok = num_close(a_summary.get("departments_meeting_target"),
                            g_summary.get("departments_meeting_target", 7), 0)
        check("Departments_Meeting_Target correct", meet_ok,
              f"Got {a_summary.get('departments_meeting_target')}")
        hdv = a_summary.get("highest_diversity_department")
        g_hdv = g_summary.get("highest_diversity_department", "Кадры")
        hdv_ok = hdv is not None and str(hdv).strip().lower() == str(g_hdv).strip().lower()
        check("Highest_Diversity_Department matches groundtruth", hdv_ok,
              f"Got {hdv}, expected {g_hdv}")

        # CRITICAL: ключевые итоговые результаты политики разнообразия.
        check("Summary headline results match groundtruth",
              avg_ok and meet_ok and hdv_ok,
              f"avg={avg_ok}, meeting={meet_ok}, highest={hdv_ok}")


def check_gform():
    print("\n=== Checking Survey Form ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Опрос разнообразия: заголовок может быть на EN ("diversity"/"workforce")
        # или RU ("разнообраз"/"персонал"). Шумовая форма удовлетворённости
        # ('удовлетвор'/'satisfaction') исключается.
        cur.execute("""
            SELECT id FROM gform.forms
            WHERE (
                LOWER(title) LIKE '%diversity%'
                OR LOWER(title) LIKE '%workforce%'
                OR LOWER(title) LIKE '%разнообраз%'
            )
            AND LOWER(title) NOT LIKE '%satisfaction%'
            AND LOWER(title) NOT LIKE '%удовлетвор%'
            ORDER BY id LIMIT 1
        """)
        row = cur.fetchone()
        check("Diversity survey form exists", row is not None, "No diversity form found")
        if row is not None:
            form_id = row[0]
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
            qcount = cur.fetchone()[0]
            # CRITICAL: ровно 5 вопросов в опросе разнообразия (а не в шумовой форме).
            check("Diversity survey has exactly 5 questions", qcount == 5, f"Got {qcount}")
        else:
            check("Diversity survey has exactly 5 questions", False, "No form")
        cur.close()
        conn.close()
    except Exception as e:
        check("Diversity survey has exactly 5 questions", False, str(e))


import re

# Названия отделов (realia, russified централизованно через
# db/zzz_clickhouse_after_init.sql -> live DB отдаёт RU-значения). Сравнение
# регистронезависимо; страницы отделов в Teamly содержат RU-названия.
GT_DEPARTMENTS = ["Инженерия", "Финансы", "Кадры", "Операции", "НИОКР", "Продажи", "Поддержка"]


def check_teamly():
    print("\n=== Checking Teamly 'Diversity Metrics Dashboard' space ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Найти пространство дашборда по названию (EN-идентификатор сохранён,
        # допускаем и RU-формулировку).
        cur.execute("SELECT id, key, name, description FROM teamly.spaces")
        spaces = cur.fetchall()
        found_space = None
        for sid, key, name, desc in spaces:
            blob = " ".join(str(x) for x in (key, name, desc) if x).lower()
            if "diversity" in blob and ("metric" in blob or "dashboard" in blob):
                found_space = sid
                break
        # Запасной вариант: пространство кадровой аналитики (seed-пространство).
        if found_space is None:
            for sid, key, name, desc in spaces:
                blob = " ".join(str(x) for x in (key, name, desc) if x).lower()
                if "hranalytics" in (str(key) or "").lower() or "разнообраз" in blob:
                    found_space = sid
                    break
        check("Teamly 'Diversity Metrics Dashboard' space exists",
              found_space is not None, f"Found {len(spaces)} spaces")

        if found_space is not None:
            cur.execute(
                "SELECT title, COALESCE(body, '') FROM teamly.pages WHERE space_id = %s",
                (found_space,))
            pages = cur.fetchall()

            # Шумовые страницы (создаются в preprocess) не относятся к дашборду.
            NOISE_MARKERS = ["протокол", "трекер проектов", "архив", "не относится"]
            dept_pages = []
            for title, body in pages:
                blob = f"{title} {body}".lower()
                if any(m in blob for m in NOISE_MARKERS):
                    continue
                dept_pages.append((title or "", body or ""))

            # Для каждой страницы отдела проверяем: упоминание названия отдела
            # из groundtruth + наличие числового значения индекса разнообразия.
            covered = set()
            pages_with_index = 0
            for title, body in dept_pages:
                blob = f"{title} {body}".lower()
                for dept in GT_DEPARTMENTS:
                    if dept.lower() in blob:
                        covered.add(dept.lower())
                # Индекс разнообразия — десятичное число (например 1.3481).
                if re.search(r"\d+[.,]\d+", body) or re.search(r"\d+[.,]\d+", title):
                    pages_with_index += 1

            # CRITICAL: 7 страниц отделов с содержательным контентом (название
            # отдела + значение индекса разнообразия), а не просто 7 страниц.
            ok_pages = (len(dept_pages) == 7
                        and len(covered) >= 7
                        and pages_with_index >= 7)
            check("Teamly 'Diversity Metrics Dashboard' space has 7 department pages with content",
                  ok_pages,
                  f"dept_pages={len(dept_pages)}, depts_covered={sorted(covered)}, "
                  f"pages_with_index={pages_with_index}")
        else:
            check("Teamly 'Diversity Metrics Dashboard' space has 7 department pages with content",
                  False, "No space")
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly 'Diversity Metrics Dashboard' space has 7 department pages with content",
              False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")

    # Excel: no unexpected sheets beyond the 4 required
    path = os.path.join(workspace, "Diversity_Metrics_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        check("Excel has no more than 6 sheets", len(wb.sheetnames) <= 6,
              f"Found {len(wb.sheetnames)} sheets: {wb.sheetnames}")

    # Teamly: нет дублирующихся пространств дашборда разнообразия
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM teamly.spaces
            WHERE LOWER(name) LIKE '%%diversity%%'
               OR LOWER(COALESCE(description, '')) LIKE '%%diversity metrics dashboard%%'
        """)
        space_count = cur.fetchone()[0]
        check("No duplicate Diversity Metrics dashboard spaces", space_count <= 1,
              f"Found {space_count} diversity spaces")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_teamly()
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    # CRITICAL-гейт: любой провал семантической проверки => FAIL независимо
    # от accuracy.
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILED: {critical_failed}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"  Accuracy: {accuracy:.1f}%")
    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
