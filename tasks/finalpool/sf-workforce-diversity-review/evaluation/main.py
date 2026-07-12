"""Evaluation for sf-workforce-diversity-review (ClickHouse warehouse, russified)."""
import argparse
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        msg = f" -- {detail}" if detail else ""
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def build_lookup(rows, ncols=2):
    """Map composite key of first `ncols` columns -> row."""
    data = rows[1:] if len(rows) > 1 else []
    lookup = {}
    for r in data:
        if not r or any(r[i] is None for i in range(min(ncols, len(r)))):
            continue
        k = "|".join(str(r[i]).strip().lower() for i in range(ncols))
        lookup[k] = r
    return lookup, data


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Diversity_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Diversity_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ===================== Education Distribution =====================
    print("  Checking Education Distribution...")
    a_rows = load_sheet_rows(agent_wb, "Education Distribution")
    g_rows = load_sheet_rows(gt_wb, "Education Distribution")
    if a_rows is None:
        record("Education Distribution sheet exists", False, "sheet missing", critical=True)
    elif g_rows is None:
        record("Education Distribution groundtruth exists", False, "gt missing")
    else:
        record("Education Distribution sheet exists", True)
        a_lookup, _ = build_lookup(a_rows, 2)
        _, g_data = build_lookup(g_rows, 2)
        # CRITICAL: per (Department, Education_Level) Internal_Pct within ~1.0pp AND
        # Count within a tight tolerance -- proves the agent queried the russified
        # warehouse and computed per-department education shares correctly.
        pct_ok, count_ok = True, True
        missing = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if a_row is None:
                missing.append(f"{g_row[0]}|{g_row[1]}")
                pct_ok = count_ok = False
                continue
            if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 1.0):
                pct_ok = False
                print(f"      {key}.Internal_Pct: {a_row[3]} vs {g_row[3]}")
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 10):
                count_ok = False
                print(f"      {key}.Count: {a_row[2]} vs {g_row[2]}")
        if missing:
            print(f"      missing keys: {missing[:5]}")
        record("Education Internal_Pct per (Dept,Edu) within 1.0pp", pct_ok,
               critical=True)
        record("Education Count per (Dept,Edu) within 10", count_ok, critical=True)
        # Non-critical: Benchmark_Pct transcribed from portal (Бакалавр 42, Диплом 22,
        # Среднее образование 18, Магистр 14, Кандидат наук 4).
        edu_bench = {"бакалавр": 42.0, "диплом": 22.0, "среднее образование": 18.0,
                     "магистр": 14.0, "кандидат наук": 4.0}
        bench_ok = True
        for g_row in g_data:
            if not g_row or g_row[1] is None or len(g_row) <= 4:
                continue
            exp = edu_bench.get(str(g_row[1]).strip().lower())
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if exp is not None and a_row is not None and len(a_row) > 4:
                if not num_close(a_row[4], exp, 0.5):
                    bench_ok = False
        record("Education Benchmark_Pct matches portal values", bench_ok)

    # ===================== Age Distribution =====================
    print("  Checking Age Distribution...")
    a_rows = load_sheet_rows(agent_wb, "Age Distribution")
    g_rows = load_sheet_rows(gt_wb, "Age Distribution")
    if a_rows is None:
        record("Age Distribution sheet exists", False, "sheet missing", critical=True)
    elif g_rows is None:
        record("Age Distribution groundtruth exists", False, "gt missing")
    else:
        record("Age Distribution sheet exists", True)
        a_lookup, _ = build_lookup(a_rows, 2)
        _, g_data = build_lookup(g_rows, 2)
        # CRITICAL: Internal_Pct per (Department, Age_Group) within ~1.0pp -- confirms
        # correct Under 30/30-39/40-49/50+ bucketing from AGE.
        pct_ok = True
        missing = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if a_row is None:
                missing.append(f"{g_row[0]}|{g_row[1]}")
                pct_ok = False
                continue
            if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 1.0):
                pct_ok = False
                print(f"      {key}.Internal_Pct: {a_row[3]} vs {g_row[3]}")
        if missing:
            print(f"      missing keys: {missing[:5]}")
        record("Age Internal_Pct per (Dept,Age_Group) within 1.0pp", pct_ok,
               critical=True)
        # Non-critical: Benchmark_Pct from portal (Under 30=35, 30-39=35, 40-49=20, 50+=10)
        age_bench = {"under 30": 35.0, "30-39": 35.0, "40-49": 20.0, "50+": 10.0}
        bench_ok = True
        for g_row in g_data:
            if not g_row or g_row[1] is None or len(g_row) <= 4:
                continue
            exp = age_bench.get(str(g_row[1]).strip().lower())
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if exp is not None and a_row is not None and len(a_row) > 4:
                if not num_close(a_row[4], exp, 0.5):
                    bench_ok = False
        record("Age Benchmark_Pct matches portal values", bench_ok)

    # ===================== Overall Scorecard =====================
    print("  Checking Overall Scorecard...")
    a_rows = load_sheet_rows(agent_wb, "Overall Scorecard")
    g_rows = load_sheet_rows(gt_wb, "Overall Scorecard")
    if a_rows is None:
        record("Overall Scorecard sheet exists", False, "sheet missing", critical=True)
    elif g_rows is None:
        record("Overall Scorecard groundtruth exists", False, "gt missing")
    else:
        record("Overall Scorecard sheet exists", True)
        a_lookup, _ = build_lookup(a_rows, 2)
        _, g_data = build_lookup(g_rows, 2)
        val_ok = True
        status_ok = True
        missing = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = f"{str(g_row[0]).strip().lower()}|{str(g_row[1]).strip().lower()}"
            a_row = a_lookup.get(key)
            if a_row is None:
                missing.append(f"{g_row[0]}|{g_row[1]}")
                val_ok = status_ok = False
                continue
            # Internal_Value (col idx 2)
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 1.0):
                val_ok = False
                print(f"      {key}.Internal: {a_row[2]} vs {g_row[2]}")
            # Status (col idx 5): Aligned / Above / Below at 3pp threshold
            if len(a_row) > 5 and len(g_row) > 5:
                if not str_match(a_row[5], g_row[5]):
                    status_ok = False
                    print(f"      {key}.Status: {a_row[5]} vs {g_row[5]}")
        if missing:
            print(f"      missing keys: {missing[:5]}")
        record("Scorecard Internal_Value within 1.0pp", val_ok)
        # CRITICAL: the Status classification is the core analytical deliverable.
        record("Scorecard Status (Aligned/Above/Below) matches groundtruth",
               status_ok, critical=True)

    # ===================== Summary =====================
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        record("Summary sheet exists", False, "sheet missing", critical=True)
    elif g_rows is None:
        record("Summary groundtruth exists", False, "gt missing")
    else:
        record("Summary sheet exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        avg_ok = True
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if key in ("total_employees", "total_departments", "education_levels"):
                continue  # handled as critical below
            if a_row is None or len(a_row) < 2 or len(g_row) < 2:
                avg_ok = False
                print(f"      Summary missing/short: {g_row[0]}")
                continue
            if not num_close(a_row[1], g_row[1], 1.0):
                avg_ok = False
                print(f"      {key}: {a_row[1]} vs {g_row[1]}")
        record("Summary averages within 1.0", avg_ok)
        # CRITICAL: full-warehouse aggregation -- Total_Employees (~50k, tol 100),
        # Total_Departments == 7 and Education_Levels == 5 must match exactly.
        gmap = {str(r[0]).strip().lower(): r for r in g_data if r and r[0] is not None}

        def cmp_exact(key, tol):
            a_row = a_lookup.get(key)
            g_row = gmap.get(key)
            if a_row is None or g_row is None or len(a_row) < 2 or len(g_row) < 2:
                return False
            return num_close(a_row[1], g_row[1], tol)

        record("Summary Total_Employees within 100", cmp_exact("total_employees", 100),
               critical=True)
        record("Summary Total_Departments == 7", cmp_exact("total_departments", 0),
               critical=True)
        record("Summary Education_Levels == 5 distinct", cmp_exact("education_levels", 0),
               critical=True)

    # ===================== PowerPoint =====================
    print("  Checking DEI_Board_Report.pptx...")
    pptx_file = os.path.join(args.agent_workspace, "DEI_Board_Report.pptx")
    if not os.path.exists(pptx_file):
        record("DEI_Board_Report.pptx exists", False, "not found", critical=True)
    else:
        try:
            from pptx import Presentation
            prs = Presentation(pptx_file)
            slide_count = len(prs.slides)
            record("DEI_Board_Report.pptx has >= 5 slides", slide_count >= 5,
                   f"{slide_count} slides", critical=True)
        except Exception as e:
            record("DEI_Board_Report.pptx readable", False, str(e), critical=True)

    # ===================== SUMMARY / GATE =====================
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (threshold accuracy>=70)")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
