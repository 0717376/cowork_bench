"""Evaluation for canvas-announcement-summary."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def summary_lookup(rows):
    """Build {metric_lower: value} from a Summary sheet (skip header row)."""
    out = {}
    for row in rows[1:] if len(rows) > 1 else []:
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
    return out


def find_sent_email(recipient, subject_substr):
    """Return matching sent email (subject, from_addr, to_addr, body) or None."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    # folder_id = 2 is the Sent folder; also accept sent_log linkage as a fallback.
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    def recipients_of(to_addr):
        if to_addr is None:
            return []
        if isinstance(to_addr, list):
            return [str(r).strip().lower() for r in to_addr]
        if isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                if isinstance(parsed, list):
                    return [str(r).strip().lower() for r in parsed]
            except (json.JSONDecodeError, TypeError):
                pass
            return [str(to_addr).strip().lower()]
        return [str(to_addr).strip().lower()]

    for subj, from_addr, to_addr, body in rows:
        if recipient.lower() in recipients_of(to_addr) and subject_substr.lower() in (subj or "").lower():
            return subj, from_addr, to_addr, body
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Announcements.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Announcements.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_failures = []
    # Courses tied at the maximum announcement count in the groundtruth.
    # task.md gives NO tiebreaker for Most_Active_Course, so any of these is a
    # legitimately correct answer; do not pin to a single backend-query-order pick.
    max_active_courses = set()

    # ---------------------------------------------------------------
    # Sheet: Announcement Stats (per-course counts)
    # ---------------------------------------------------------------
    print("  Checking Announcement Stats...")
    a_rows = load_sheet_rows(agent_wb, "Announcement Stats")
    g_rows = load_sheet_rows(gt_wb, "Announcement Stats")
    if a_rows is None:
        all_errors.append("Sheet 'Announcement Stats' not found in agent output")
        critical_failures.append("Announcement Stats sheet missing")
    elif g_rows is None:
        all_errors.append("Sheet 'Announcement Stats' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Compute tied-max courses from the groundtruth Announcement Stats.
        try:
            counts = []
            for g_row in g_data:
                if g_row and g_row[0] is not None and len(g_row) > 1 and g_row[1] is not None:
                    counts.append((str(g_row[0]).strip().lower(), float(g_row[1])))
            if counts:
                top = max(c for _, c in counts)
                max_active_courses = {code for code, c in counts if c == top}
        except (TypeError, ValueError):
            max_active_courses = set()

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        errors = []
        exact_mismatches = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                # Non-critical tolerant check (counts toward accuracy).
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Announcements: {a_row[1]} vs {g_row[1]} (tol=1)")
                # CRITICAL: exact per-course count for the top courses (>=5 in GT).
                try:
                    if float(g_row[1]) >= 5 and not num_close(a_row[1], g_row[1], 0):
                        exact_mismatches.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                except (TypeError, ValueError):
                    exact_mismatches.append(f"{key}: non-numeric {a_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")
        if exact_mismatches:
            critical_failures.append(
                "Per-course Announcements counts (top courses) not exact: " + "; ".join(exact_mismatches[:5])
            )

    # ---------------------------------------------------------------
    # Sheet: Summary (aggregate deliverables)
    # ---------------------------------------------------------------
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
        critical_failures.append("Summary sheet missing")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_sum = summary_lookup(a_rows)
        g_sum = summary_lookup(g_rows)

        # Tightened tolerances: tol=0 for the two numeric aggregates (was 5.0).
        errors = []
        for metric, tol in (("total_announcements", 0), ("courses_with_announcements", 0)):
            if metric in g_sum:
                if not num_close(a_sum.get(metric), g_sum.get(metric), tol):
                    errors.append(f"{metric}.Value: {a_sum.get(metric)} vs {g_sum.get(metric)} (tol={tol})")
        # Accept any course tied at the max count (task.md gives no tiebreaker).
        def most_active_ok(val):
            if val is None:
                return False
            v = str(val).strip().lower()
            if max_active_courses:
                return v in max_active_courses
            return str_match(val, g_sum.get("most_active_course"))

        if "most_active_course" in g_sum:
            if not most_active_ok(a_sum.get("most_active_course")):
                errors.append(f"most_active_course: {a_sum.get('most_active_course')} not among tied-max {sorted(max_active_courses) or g_sum.get('most_active_course')}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

        # CRITICAL: core aggregates must be exactly right.
        if not num_close(a_sum.get("total_announcements"), g_sum.get("total_announcements"), 0):
            critical_failures.append(
                f"Total_Announcements wrong: {a_sum.get('total_announcements')} vs {g_sum.get('total_announcements')} (must be exact)"
            )
        if not num_close(a_sum.get("courses_with_announcements"), g_sum.get("courses_with_announcements"), 0):
            critical_failures.append(
                f"Courses_With_Announcements wrong: {a_sum.get('courses_with_announcements')} vs {g_sum.get('courses_with_announcements')} (must be exact)"
            )
        # CRITICAL: must be a genuinely most-active course. With a 7-way tie at the
        # top count and no tiebreaker in task.md, accept any tied-max course.
        if not most_active_ok(a_sum.get("most_active_course")):
            critical_failures.append(
                f"Most_Active_Course wrong: {a_sum.get('most_active_course')} not among tied-max {sorted(max_active_courses) or g_sum.get('most_active_course')}"
            )

    # ---------------------------------------------------------------
    # CRITICAL: email actually sent to academic-affairs with required subject.
    # preprocess clears email.* so any matching row is the agent's deliverable.
    # ---------------------------------------------------------------
    print("  Checking sent email...")
    try:
        email = find_sent_email("academic-affairs@openuniversity.ac.uk", "Canvas Announcement Activity Report")
        if email is None:
            critical_failures.append(
                "No email sent to academic-affairs@openuniversity.ac.uk with subject 'Canvas Announcement Activity Report'"
            )
            all_errors.append("Required report email not found")
        else:
            print("    PASS (email found)")
    except Exception as e:
        # If the email backend is unreachable, fail critically rather than silently passing.
        critical_failures.append(f"Email verification failed: {e}")
        all_errors.append(f"Email check error: {e}")

    # ---------------------------------------------------------------
    # CRITICAL gate: any critical failure => FAIL regardless of accuracy.
    # ---------------------------------------------------------------
    if critical_failures:
        print(f"\n=== CRITICAL FAILURES ({len(critical_failures)}) ===")
        for c in critical_failures:
            print(f"  [CRITICAL] {c}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
