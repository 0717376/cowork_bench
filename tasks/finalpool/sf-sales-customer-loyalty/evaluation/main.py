"""Evaluation for sf-sales-customer-loyalty (ClickHouse / russified SALES_DW).

The live SALES_DW CUSTOMERS values (SEGMENT / REGION) are russified centrally by
db/zzz_clickhouse_after_init.sql, so the agent legitimately writes Russian
segment/region labels into the Segment_Region keys and Highest_LTV_Group. The
groundtruth Customer_Loyalty_Report.xlsx has been REGENERATED so its keys are the
russified strings (e.g. 'Частные клиенты - Латинская Америка'); all numeric
aggregates (Customers / Avg_LTV / Total_Orders / Overall_Avg_LTV) are unchanged by
russification (it only renames SEGMENT/REGION values), so they are checked directly
against the groundtruth.

Hardening over the original loose diff:
  - Avg_LTV / Overall_Avg_LTV tolerances tightened (1.0 instead of 50.0).
  - CRITICAL semantic checks: all 20 combos present, Highest_LTV_Group exact match,
    Total_Combinations==20, Overall_Avg_LTV correct.
  - Email-sent verification added (recipient + exact subject), previously unverified.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}: {d}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_loyalty_sheet(agent_wb, gt_wb):
    print("\n=== Checking 'Customer Loyalty' sheet ===")
    a_rows = load_sheet_rows(agent_wb, "Customer Loyalty")
    g_rows = load_sheet_rows(gt_wb, "Customer Loyalty")
    if a_rows is None:
        check("Sheet 'Customer Loyalty' present", False, "missing in agent output", critical=True)
        return
    if g_rows is None:
        check("Sheet 'Customer Loyalty' present in groundtruth", False, "missing in GT", critical=True)
        return

    a_data = a_rows[1:] if len(a_rows) > 1 else []
    g_data = g_rows[1:] if len(g_rows) > 1 else []

    a_lookup = {}
    for row in a_data:
        if row and row[0] is not None:
            a_lookup[str(row[0]).strip().lower()] = row

    missing = []
    val_errors = []
    for g_row in g_data:
        if not g_row or g_row[0] is None:
            continue
        key = str(g_row[0]).strip().lower()
        a_row = a_lookup.get(key)
        if a_row is None:
            missing.append(g_row[0])
            continue
        if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 5):
            val_errors.append(f"{key}.Customers: {a_row[1]} vs {g_row[1]} (tol=5)")
        # Tightened: Avg_LTV tol 1.0 (was 50.0)
        if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 1.0):
            val_errors.append(f"{key}.Avg_LTV: {a_row[2]} vs {g_row[2]} (tol=1.0)")
        if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 10):
            val_errors.append(f"{key}.Total_Orders: {a_row[3]} vs {g_row[3]} (tol=10)")

    # CRITICAL: all 20 russified segment-region combos present and matched by key.
    check(f"All {len(g_data)} segment-region combinations present (russified keys matched)",
          len(missing) == 0,
          f"missing={missing[:5]}", critical=True)
    # CRITICAL: per-group aggregates correct within tight tolerance.
    check("Per-group Customers / Avg_LTV (tol=1.0) / Total_Orders correct",
          len(val_errors) == 0,
          f"{len(val_errors)} mismatches: {val_errors[:5]}", critical=True)
    # Structural (non-critical): exactly 20 data rows.
    check("Exactly 20 data rows in 'Customer Loyalty'", len([r for r in a_data if r and r[0] is not None]) == 20,
          f"agent rows={len([r for r in a_data if r and r[0] is not None])}")


SUMMARY_KEYS = ("total_combinations", "highest_ltv_group", "overall_avg_ltv")


def summary_lookup(rows, keys):
    """Header/orientation-tolerant {metric_lower: value} map for a Summary sheet.

    Scans ALL cells; a metric's value is the cell to its right, or the cell
    below when the right neighbour is itself another metric key (horizontal
    layout). A 'Metric/Value' header never collides with metric keys.
    """
    keys = {str(k).strip().lower() for k in keys}
    out = {}
    rows = rows or []

    def _is_key(v):
        return v is not None and str(v).strip().lower() in keys

    for r, row in enumerate(rows):
        for c, cell in enumerate(row or []):
            if cell is None:
                continue
            k = str(cell).strip().lower()
            if k not in keys or k in out:
                continue
            right = row[c + 1] if c + 1 < len(row) else None
            below = None
            if r + 1 < len(rows) and rows[r + 1] and c < len(rows[r + 1]):
                below = rows[r + 1][c]
            if right is not None and not _is_key(right):
                out[k] = right
            elif below is not None and not _is_key(below):
                out[k] = below
    return out


def check_summary_sheet(agent_wb, gt_wb):
    print("\n=== Checking 'Summary' sheet ===")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        check("Sheet 'Summary' present", False, "missing in agent output", critical=True)
        return
    if g_rows is None:
        check("Sheet 'Summary' present in groundtruth", False, "missing in GT", critical=True)
        return

    a_map = summary_lookup(a_rows, SUMMARY_KEYS)
    g_map = summary_lookup(g_rows, SUMMARY_KEYS)

    # CRITICAL: Total_Combinations == 20.
    a_tc = a_map.get("total_combinations")
    check("Total_Combinations == 20", num_close(a_tc, g_map.get("total_combinations", 20), 0),
          f"agent={a_tc}", critical=True)

    # CRITICAL: Highest_LTV_Group exact match against regenerated (russified) GT.
    a_hi = a_map.get("highest_ltv_group")
    g_hi = g_map.get("highest_ltv_group")
    check(f"Highest_LTV_Group == '{g_hi}' (exact russified match)",
          a_hi is not None and str(a_hi).strip().lower() == str(g_hi).strip().lower(),
          f"agent={a_hi!r} gt={g_hi!r}", critical=True)

    # CRITICAL: Overall_Avg_LTV correct within tight tolerance (was 50.0).
    a_oa = a_map.get("overall_avg_ltv")
    check("Overall_Avg_LTV correct (tol=1.0)",
          num_close(a_oa, g_map.get("overall_avg_ltv"), 1.0),
          f"agent={a_oa} gt={g_map.get('overall_avg_ltv')}", critical=True)


def check_email():
    """Verify an email was sent to marketing@company.com with the required subject."""
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        check("Email DB reachable", False, str(e), critical=True)
        return
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        ORDER BY date DESC
    """)
    emails = cur.fetchall()
    cur.execute("""
        SELECT m.subject, m.from_addr, m.to_addr, m.body_text
        FROM email.sent_log s JOIN email.messages m ON s.message_id = m.id
        ORDER BY s.id DESC LIMIT 20
    """)
    sent = cur.fetchall()
    cur.close()
    conn.close()

    all_msgs = list(emails) + list(sent)
    check("At least 1 email present", len(all_msgs) >= 1,
          f"messages={len(emails)} sent_log={len(sent)}")

    mkt_msgs = [m for m in all_msgs if "marketing@company.com" in str(m[2] or "").lower()]
    subj_ok = any("customer loyalty analysis" in str(m[0] or "").lower() for m in mkt_msgs)

    # CRITICAL composite: correct recipient + exact subject.
    check("Email sent to marketing@company.com with subject 'Customer Loyalty Analysis'",
          len(mkt_msgs) >= 1 and subj_ok,
          f"recipients={[str(m[2]) for m in all_msgs[:5]]} subjects={[m[0] for m in mkt_msgs[:5]]}",
          critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("SF SALES CUSTOMER LOYALTY - EVALUATION (ClickHouse / RU)")
    print("=" * 70)

    agent_file = os.path.join(args.agent_workspace, "Customer_Loyalty_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Customer_Loyalty_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    check_loyalty_sheet(agent_wb, gt_wb)
    check_summary_sheet(agent_wb, gt_wb)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    critical_ok = len(CRITICAL_FAILS) == 0
    all_ok = critical_ok and accuracy >= 70.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy, "critical_fails": CRITICAL_FAILS,
                       "success": all_ok}, f, indent=2)

    if not critical_ok:
        print("FAIL: critical check(s) failed.")
        sys.exit(1)
    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
