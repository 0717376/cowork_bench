"""Evaluation for terminal-arxiv-moex-excel-word-teamly-email (RU / moex + teamly).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Tickers are bare-English MOEX symbols (SBER.ME / TCSG.ME / GAZP.ME). Latest
close prices are read HONESTLY at runtime from moex.stock_prices.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

RELEVANT_PAPER_IDS = {"2306.06031", "2304.07619", "2302.14040", "2311.10723"}
NOISE_PAPER_IDS = {"2305.18290", "2307.09288"}
STOCKS = ["SBER.ME", "TCSG.ME", "GAZP.ME"]

# Expected AI-exposure score + recommendation per chosen MOEX ticker.
# SBER (Сбербанк, GigaChat/AI leader) ~9 Overweight;
# TCSG (Т-Технологии, финтех/AI) ~8 Overweight;
# GAZP (Газпром, низкая доля ИИ) ~5 Hold.
EXPECTED_SCORES = {"SBER.ME": (9, "overweight"), "TCSG.ME": (8, "overweight"),
                   "GAZP.ME": (5, "hold")}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Research_Papers has exactly the 4 relevant papers",
    "No noise arxiv papers in Excel (DragGAN, Llama 2)",
    "No noise paper IDs in Excel (2305.18290, 2307.09288)",
    "AI scores+recommendations correct for all 3 tickers",
    "Current_Price matches moex.stock_prices within 10% for all 3 holdings",
    "Teamly Research Pipeline page exists (not the noise page)",
    "Teamly Research Pipeline has no noise papers (DragGAN, Llama 2)",
    "Both required emails sent to correct distinct recipients",
}


def get_expected_from_db():
    """Query moex schema for latest close price per chosen ticker (honest read)."""
    out = {"stock_prices": {}}
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        for symbol in STOCKS:
            cur.execute("""
                SELECT close FROM moex.stock_prices
                WHERE symbol = %s ORDER BY date DESC LIMIT 1
            """, (symbol,))
            row = cur.fetchone()
            if row and row[0]:
                out["stock_prices"][symbol] = float(row[0])
        cur.close()
        conn.close()
    except Exception as e:
        print(f"  [WARN] DB query for expected values failed: {e}")
    return out


EXPECTED = get_expected_from_db()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        d = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def check_excel(ws_path):
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "AI_Investment_Research.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        # Mark dependent critical checks failed.
        check("Research_Papers has exactly the 4 relevant papers", False, "no excel")
        check("AI scores+recommendations correct for all 3 tickers", False, "no excel")
        check("Current_Price matches moex.stock_prices within 10% for all 3 holdings", False, "no excel")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}

    # Portfolio_Holdings sheet
    ph_name = sn.get("portfolio_holdings")
    if ph_name is None:
        check("Portfolio_Holdings sheet exists", False, f"Sheets: {wb.sheetnames}")
        check("Current_Price matches moex.stock_prices within 10% for all 3 holdings", False, "no sheet")
    else:
        check("Portfolio_Holdings sheet exists", True)
        ws = wb[ph_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).lower() if h else "" for h in rows[0]] if rows else []
        data = [r for r in rows[1:] if r and r[0] is not None]
        check("Portfolio_Holdings has 3 rows", len(data) == 3, f"Found {len(data)}")

        symbols_found = {str(r[0]).strip().upper() for r in data}
        check("All 3 stocks present", symbols_found >= {s.upper() for s in STOCKS},
              f"Found: {symbols_found}")

        price_col = None
        for i, h in enumerate(headers):
            if "price" in h:
                price_col = i
                break
        if price_col is not None:
            prices = [r[price_col] for r in data if r[price_col] is not None]
            check("Prices are populated", len(prices) == 3, f"Prices: {prices}")
            # CRITICAL: validate against moex.stock_prices; hard-fail if DB
            # returned no rows (broken swap must not silently pass).
            if not EXPECTED["stock_prices"]:
                check("Current_Price matches moex.stock_prices within 10% for all 3 holdings",
                      False, "moex.stock_prices returned no rows for STOCKS")
            else:
                ok = True
                detail = []
                for row in data:
                    sym = str(row[0]).strip().upper()
                    exp = EXPECTED["stock_prices"].get(sym)
                    if exp is None or row[price_col] is None:
                        ok = False
                        detail.append(f"{sym}: no price/expected")
                        continue
                    if not num_close(row[price_col], exp, tol=exp * 0.1):
                        ok = False
                        detail.append(f"{sym}: got {row[price_col]} exp ~{exp:.2f}")
                check("Current_Price matches moex.stock_prices within 10% for all 3 holdings",
                      ok and len(data) == 3, "; ".join(detail))
        else:
            check("Price column exists", False, f"Headers: {headers}")
            check("Current_Price matches moex.stock_prices within 10% for all 3 holdings",
                  False, "no price column")

    # Research_Papers sheet
    rp_name = sn.get("research_papers")
    if rp_name is None:
        check("Research_Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
        check("Research_Papers has exactly the 4 relevant papers", False, "no sheet")
    else:
        check("Research_Papers sheet exists", True)
        ws2 = wb[rp_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Research_Papers has 4 rows (relevant only)", len(data2) == 4,
              f"Found {len(data2)}")

        all_titles = " ".join(str(r[0]) for r in data2).lower()
        check("FinGPT paper listed", "fingpt" in all_titles, f"Titles: {all_titles[:200]}")

        # CRITICAL: exactly the 4 relevant papers, no noise titles present.
        relevant_markers = ["fingpt", "predict stock price", "financial risk prediction",
                            "quantitative finance"]
        noise_markers = ["draggan", "drag your gan", "llama 2", "open foundation"]
        has_all_relevant = all(any(m in str(r[0]).lower() for r in data2) for m in relevant_markers)
        has_no_noise = not any(nm in all_titles for nm in noise_markers)
        check("Research_Papers has exactly the 4 relevant papers",
              len(data2) == 4 and has_all_relevant and has_no_noise,
              f"all_relevant={has_all_relevant} no_noise={has_no_noise}")

        all_stocks_text = " ".join(str(r[-1]) if r[-1] else "" for r in data2).upper()
        check("Applicable stocks mention SBER", "SBER" in all_stocks_text)
        check("Applicable stocks mention GAZP", "GAZP" in all_stocks_text)

    # AI_Impact_Assessment sheet
    ai_name = sn.get("ai_impact_assessment")
    if ai_name is None:
        check("AI_Impact_Assessment sheet exists", False, f"Sheets: {wb.sheetnames}")
        check("AI scores+recommendations correct for all 3 tickers", False, "no sheet")
    else:
        check("AI_Impact_Assessment sheet exists", True)
        ws3 = wb[ai_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("AI_Impact_Assessment has 3 rows", len(data3) == 3, f"Found {len(data3)}")

        # CRITICAL: per-ticker score + recommendation mapping.
        score_rec = {}
        for row in data3:
            stock = str(row[0]).strip().upper()
            score = row[1] if len(row) > 1 else None
            rec = str(row[-1]).lower() if row[-1] else ""
            score_rec[stock] = (score, rec)
        all_ok = True
        detail = []
        for sym, (exp_score, exp_rec) in EXPECTED_SCORES.items():
            got = score_rec.get(sym.upper())
            if got is None:
                all_ok = False
                detail.append(f"{sym}: missing")
                continue
            score, rec = got
            if not num_close(score, exp_score, tol=1):
                all_ok = False
                detail.append(f"{sym}: score {score} != ~{exp_score}")
            if exp_rec not in rec:
                all_ok = False
                detail.append(f"{sym}: rec '{rec}' != {exp_rec}")
        check("AI scores+recommendations correct for all 3 tickers", all_ok, "; ".join(detail))
        # Non-critical individual visibility
        for sym, (exp_score, exp_rec) in EXPECTED_SCORES.items():
            got = score_rec.get(sym.upper())
            if got:
                check(f"{sym} AI score ~{exp_score}", num_close(got[0], exp_score, tol=1),
                      f"Score: {got[0]}")
                check(f"{sym} recommendation {exp_rec}", exp_rec in got[1], f"Rec: {got[1]}")

    # Investment_Thesis sheet
    it_name = sn.get("investment_thesis")
    if it_name is None:
        check("Investment_Thesis sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Investment_Thesis sheet exists", True)
        ws4 = wb[it_name]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if r and r[0] is not None]
        check("Investment_Thesis has >= 2 rows", len(data4) >= 2, f"Found {len(data4)}")

        all_themes = " ".join(str(r[0]) for r in data4).lower()
        check("Theme mentions AI", "ai" in all_themes or "ии" in all_themes,
              f"Themes: {all_themes[:200]}")

    wb.close()


def check_word(ws_path):
    print("\n=== Checking Word Document ===")
    path = os.path.join(ws_path, "AI_Markets_Research_Report.docx")
    if not os.path.isfile(path):
        check("Word document exists", False, f"Not found: {path}")
        return
    check("Word document exists", True)

    from docx import Document
    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    check("Document title mentions AI and Markets",
          "ai" in full_text[:200] and ("market" in full_text[:200] or "рынк" in full_text[:200]))
    check("Document mentions SBER", "sber" in full_text)
    check("Document mentions TCSG", "tcsg" in full_text)
    check("Document mentions GAZP", "gazp" in full_text)
    check("Document has executive summary",
          "executive summary" in full_text or "executive" in full_text
          or "краткое резюме" in full_text or "резюме" in full_text)
    check("Document has risk assessment",
          ("risk" in full_text and ("assessment" in full_text or "factor" in full_text))
          or ("риск" in full_text and ("оцен" in full_text or "фактор" in full_text)))
    check("Document mentions research papers",
          "fingpt" in full_text or "language model" in full_text
          or "research" in full_text or "исследован" in full_text or "статьи" in full_text)
    check("Document mentions overweight or recommendation",
          "overweight" in full_text or "recommendation" in full_text
          or "увеличить" in full_text or "рекомендац" in full_text)
    check("Document length >= 800 chars", len(full_text) >= 800,
          f"Length: {len(full_text)}")


def _page_text(title, body):
    return ((title or "") + " " + (body or "")).lower()


def check_teamly():
    print("\n=== Checking Teamly Research Pipeline ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except Exception as e:
        check("Teamly Research Pipeline page exists (not the noise page)", False, str(e))
        check("Teamly Research Pipeline has no noise papers (DragGAN, Llama 2)", False, str(e))
        return

    cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
    pages = cur.fetchall()
    hub = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "архив протоколов" in tl:
            continue
        if "research pipeline" in tl or ("research" in tl and "pipeline" in tl) \
                or ("конвейер" in tl and ("статей" in tl or "исследован" in tl)):
            hub = (pid, title, body)
            break

    check("Teamly Research Pipeline page exists (not the noise page)", hub is not None,
          f"pages>3: {[(p[0], p[1]) for p in pages]}")

    if hub is not None:
        text = _page_text(hub[1], hub[2])
        # >= 4 paper entries: count relevant title markers present.
        relevant_markers = ["fingpt", "predict stock price", "financial risk prediction",
                            "quantitative finance"]
        covered = sum(1 for m in relevant_markers if m in text)
        check("Teamly page covers >= 4 relevant paper entries", covered >= 4,
              f"covered: {covered}/4")
        # Relevance + Status markers present (RU/EN).
        check("Teamly page has Relevance markers",
              "high" in text or "medium" in text or "релевант" in text,
              "no relevance markers")
        check("Teamly page has Applied status",
              "applied" in text or "примен" in text,
              "no Applied status")
        check("Teamly page links tickers (SBER/TCSG/GAZP)",
              ("sber" in text or "tcsg" in text or "gazp" in text), "no tickers")

        # CRITICAL: no noise papers on the pipeline page.
        noise_titles = ["draggan", "drag your gan", "interactive point-based",
                        "llama 2", "open foundation and fine-tuned"]
        no_noise = not any(nt in text for nt in noise_titles)
        check("Teamly Research Pipeline has no noise papers (DragGAN, Llama 2)", no_noise,
              "found noise paper on pipeline page")
    else:
        check("Teamly Research Pipeline has no noise papers (DragGAN, Llama 2)", False,
              "no pipeline page")

    cur.close()
    conn.close()


def check_emails():
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, body_text, to_addr FROM email.messages
        WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
    """)
    sent = cur.fetchall()

    portfolio_email = None
    risk_email = None
    portfolio_to_ok = False
    risk_to_ok = False
    for subj, body, to_addr in sent:
        to_str = json.dumps(to_addr).lower() if to_addr else ""
        subj_lower = (subj or "").lower()
        if "portfolio_team@firm.com" in to_str or "portfolio" in subj_lower:
            portfolio_email = (subj, body)
            if "portfolio_team@firm.com" in to_str:
                portfolio_to_ok = True
        if "risk_committee@firm.com" in to_str or ("risk" in subj_lower and "portfolio" in subj_lower):
            risk_email = (subj, body)
            if "risk_committee@firm.com" in to_str:
                risk_to_ok = True

    check("Portfolio team email sent", portfolio_email is not None,
          f"Sent: {[(s, t) for s, _, t in sent]}")
    if portfolio_email:
        body = (portfolio_email[1] or "").lower()
        check("Portfolio email mentions AI exposure",
              "ai" in body or "ии" in body or "exposure" in body or "экспозиц" in body,
              f"Body: {body[:200]}")

    check("Risk committee email sent", risk_email is not None,
          f"Sent: {[(s, t) for s, _, t in sent]}")
    if risk_email:
        body = (risk_email[1] or "").lower()
        check("Risk email mentions concentration",
              "concentrat" in body or "sector" in body or "risk" in body
              or "концентрац" in body or "сектор" in body or "риск" in body,
              f"Body: {body[:200]}")

    # CRITICAL: both emails to correct distinct recipients.
    check("Both required emails sent to correct distinct recipients",
          portfolio_to_ok and risk_to_ok,
          f"portfolio_to={portfolio_to_ok} risk_to={risk_to_ok}")

    cur.close()
    conn.close()


def check_terminal_outputs(ws_path):
    print("\n=== Checking Terminal Script Outputs ===")
    pa_path = os.path.join(ws_path, "portfolio_analysis.json")
    if os.path.isfile(pa_path):
        check("portfolio_analysis.json exists", True)
        with open(pa_path) as f:
            try:
                data = json.load(f)
                check("portfolio_analysis has content", len(data) > 0)
            except Exception:
                check("portfolio_analysis is valid JSON", False)
    else:
        check("portfolio_analysis.json exists", False)

    rm_path = os.path.join(ws_path, "research_stock_mapping.json")
    if os.path.isfile(rm_path):
        check("research_stock_mapping.json exists", True)
        with open(rm_path) as f:
            try:
                data = json.load(f)
                check("research_stock_mapping has content", len(data) > 0)
            except Exception:
                check("research_stock_mapping is valid JSON", False)
    else:
        check("research_stock_mapping.json exists", False)


def check_reverse_validation(ws_path):
    """No noise papers in Excel; no emails to wrong recipients."""
    print("\n=== Reverse Validation ===")
    noise_titles = ["draggan", "drag your gan", "interactive point-based",
                    "llama 2", "open foundation and fine-tuned"]
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # No emails to noise recipients. Match EXACT recipient tokens — note
        # 'team@firm.com' is a substring of the valid 'portfolio_team@firm.com',
        # so substring matching would false-positive; parse recipients instead.
        noise_recipients = {"team@firm.com", "office@firm.com", "social@firm.com", "admin@firm.com"}
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
        """)
        sent_emails = cur.fetchall()
        recipients = set()
        for _, to in sent_emails:
            if to is None:
                continue
            vals = to if isinstance(to, list) else None
            if vals is None:
                try:
                    parsed = json.loads(str(to))
                    vals = parsed if isinstance(parsed, list) else [str(to)]
                except Exception:
                    vals = [str(to)]
            for v in vals:
                recipients.add(str(v).strip().lower())
        no_noise_email = not (recipients & noise_recipients)
        check("No emails sent to noise recipients (team@/office@/social@/admin@firm)",
              no_noise_email, f"Recipients: {sorted(recipients)}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Reverse validation (email)", False, str(e))

    # No noise papers in Excel.
    path = os.path.join(ws_path, "AI_Investment_Research.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        all_text = ""
        for sname in wb.sheetnames:
            ws = wb[sname]
            for row in ws.iter_rows(values_only=True):
                all_text += " ".join(str(c) for c in row if c).lower() + " "
        wb.close()

        no_noise_excel = not any(nt in all_text for nt in noise_titles)
        check("No noise arxiv papers in Excel (DragGAN, Llama 2)", no_noise_excel,
              "Found noise paper content in Excel workbook")

        no_noise_ids = not any(nid in all_text for nid in NOISE_PAPER_IDS)
        check("No noise paper IDs in Excel (2305.18290, 2307.09288)", no_noise_ids,
              "Found noise paper ID in Excel")
    else:
        check("No noise arxiv papers in Excel (DragGAN, Llama 2)", False, "no excel")
        check("No noise paper IDs in Excel (2305.18290, 2307.09288)", False, "no excel")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-ARXIV-MOEX-EXCEL-WORD-TEAMLY-EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_teamly()
    check_emails()
    check_terminal_outputs(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {"total_passed": PASS_COUNT, "total_checks": total,
              "accuracy": accuracy, "critical_failed": critical_failed}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        sys.exit(1)
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()
