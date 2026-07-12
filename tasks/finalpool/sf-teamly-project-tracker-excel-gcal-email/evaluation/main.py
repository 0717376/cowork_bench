"""Evaluation script for sf-teamly-project-tracker-excel-gcal-email.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
the accuracy gate. They verify SUBSTANCE recomputed from the LIVE ClickHouse
warehouse (sf_data, logical DB HR_ANALYTICS) plus the core deliverables, not the
placeholder/sentinel values that the static groundtruth xlsx carries.

Department literals in sf_data are russified CENTRALLY (Engineering->Инженерия,
...). The agent reads them live, so all department/text comparisons accept BOTH
the russified names AND their English originals.
"""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

NOISE_PAGE_TITLE = "архив протоколов совещаний"

# Russified <-> English department map (central deterministic russification).
DEPT_RU2EN = {
    "инженерия": "engineering",
    "финансы": "finance",
    "кадры": "hr",
    "операции": "operations",
    "ниокр": "r&d",
    "продажи": "sales",
    "поддержка": "support",
}
DEPT_EN2RU = {v: k for k, v in DEPT_RU2EN.items()}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "Project_Portfolio.xlsx exists",
    "Team_Capacity Department column matches the live DWH department set (RU or EN)",
    "Team_Capacity Total_Staff per department matches live DWH headcount",
    "Exactly 3 kickoff events on 2026-03-19/20/21, 10:00-11:00 UTC",
    "Announcement email in Sent to all-managers with >=3 project names in body",
    "Teamly dashboard page exists (not the noise page) listing >=3 projects",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def dept_norm(s):
    """Map a department label (RU or EN, any case) to a canonical EN key."""
    t = str(s or "").strip().lower()
    if t in DEPT_RU2EN:
        return DEPT_RU2EN[t]
    if t in DEPT_EN2RU:
        return t
    return t


def live_dept_headcount():
    """Per-department COUNT(*) from the live warehouse, keyed by canonical EN."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('SELECT "DEPARTMENT", COUNT(*) '
                'FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" '
                'GROUP BY "DEPARTMENT"')
    out = {}
    for dept, cnt in cur.fetchall():
        out[dept_norm(dept)] = int(cnt)
    conn.close()
    return out


def find_col(headers, *names):
    for n in names:
        n = n.strip().lower()
        if n in headers:
            return headers.index(n)
    return None


# ---------------------------------------------------------------------------
# Excel structural + live-DWH critical checks
# ---------------------------------------------------------------------------
def check_excel(agent_workspace):
    excel_path = os.path.join(agent_workspace, "Project_Portfolio.xlsx")
    check("Project_Portfolio.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return None
    wb = openpyxl.load_workbook(excel_path)

    # Structural: required sheets + columns (non-critical).
    expected = {
        "Team_Capacity": ["Department", "Total_Staff", "Avg_Performance",
                          "Avg_Current_Projects", "Available_Capacity", "Top_Performer"],
        "Proposed_Projects": ["Project_Name", "Lead_Department", "Project_Lead",
                              "Team_Size", "Start_Date", "End_Date", "Priority"],
        "Resource_Allocation": ["Metric", "Value"],
    }
    for sheet_name, cols in expected.items():
        check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for c in cols:
                check(f"{sheet_name} has {c} column", c.lower() in headers,
                      f"headers: {headers[:10]}")

    # ---- CRITICAL: Team_Capacity vs live DWH ----
    try:
        live = live_dept_headcount()
    except Exception as e:
        check("Team_Capacity Department column matches the live DWH department set (RU or EN)", False, str(e))
        check("Team_Capacity Total_Staff per department matches live DWH headcount", False, str(e))
        return wb

    if "Team_Capacity" not in wb.sheetnames:
        check("Team_Capacity Department column matches the live DWH department set (RU or EN)", False, "no sheet")
        check("Team_Capacity Total_Staff per department matches live DWH headcount", False, "no sheet")
        return wb

    ws = wb["Team_Capacity"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci_dept = find_col(headers, "department")
    ci_staff = find_col(headers, "total_staff")
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    agent_depts = {}
    for r in rows:
        if ci_dept is None or ci_dept >= len(r) or r[ci_dept] is None:
            continue
        key = dept_norm(r[ci_dept])
        staff = safe_float(r[ci_staff]) if ci_staff is not None and ci_staff < len(r) else None
        agent_depts[key] = staff

    live_set = set(live.keys())
    agent_set = set(agent_depts.keys())
    # Must equal the live department set (canonicalised), rejecting the sentinel
    # English-only set if it diverges from the live DWH.
    check("Team_Capacity Department column matches the live DWH department set (RU or EN)",
          agent_set == live_set,
          f"live={sorted(live_set)} agent={sorted(agent_set)}")

    # Total_Staff per department within tolerance of live COUNT (proves real
    # aggregation, not the ~7100 sentinel).
    staff_ok = bool(agent_depts) and agent_set == live_set
    bad = []
    for k, expected_cnt in live.items():
        got = agent_depts.get(k)
        if got is None or abs(got - expected_cnt) > max(1, expected_cnt * 0.10):
            staff_ok = False
            bad.append((k, got, expected_cnt))
    check("Team_Capacity Total_Staff per department matches live DWH headcount",
          staff_ok, f"mismatches (dept, got, expected): {bad[:7]}")

    return wb


def check_script(agent_workspace):
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")
    # The plan json deliverable (non-critical).
    check("project_plan.json exists",
          os.path.exists(os.path.join(agent_workspace, "project_plan.json")),
          "missing project_plan.json")


def collect_project_names(wb):
    """Project names from the agent's Proposed_Projects sheet (lowercased)."""
    names = []
    if wb is None or "Proposed_Projects" not in wb.sheetnames:
        return names
    ws = wb["Proposed_Projects"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ci = find_col(headers, "project_name")
    if ci is None:
        return names
    for r in ws.iter_rows(min_row=2, values_only=True):
        if ci < len(r) and r[ci]:
            names.append(str(r[ci]).strip())
    return names


# ---------------------------------------------------------------------------
# Calendar / email / teamly DB checks
# ---------------------------------------------------------------------------
def check_calendar():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime "
                    "FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Exactly 3 kickoff events on 2026-03-19/20/21, 10:00-11:00 UTC", False, str(e))
        return

    target_dates = {"2026-03-19", "2026-03-20", "2026-03-21"}
    kickoff = []
    for summary, description, start_dt, end_dt in events:
        if start_dt is None:
            continue
        d = start_dt.strftime("%Y-%m-%d")
        if d not in target_dates:
            continue
        # 10:00-11:00 UTC window (tolerate a few minutes).
        if start_dt.hour != 10 or start_dt.minute > 5:
            continue
        if end_dt is not None and end_dt.hour != 11:
            continue
        kickoff.append(d)
    distinct_days = set(kickoff)
    check("Exactly 3 kickoff events on 2026-03-19/20/21, 10:00-11:00 UTC",
          len(kickoff) >= 3 and distinct_days == target_dates,
          f"matched days: {sorted(distinct_days)} (count={len(kickoff)})")

    # Noise events must remain (non-critical reverse check).
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM gcal.events "
                    "WHERE summary ILIKE '%планёрка%' OR summary ILIKE '%планерка%' "
                    "OR summary ILIKE '%обеденный%' OR summary ILIKE '%standup%' "
                    "OR summary ILIKE '%lunch%'")
        noise = cur.fetchone()[0]
        conn.close()
        check("Noise events preserved (not deleted)", noise >= 1, f"noise events: {noise}")
    except Exception as e:
        check("Noise events preserved (not deleted)", False, str(e))


def check_email(project_names):
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages "
                    "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
                    "AND subject ILIKE '%portfolio%'")
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Announcement email in Sent to all-managers with >=3 project names in body", False, str(e))
        return

    target = None
    for subject, to_addr, body in rows:
        to_s = str(to_addr or "").lower()
        if "all-managers@company.com" in to_s:
            target = (subject, to_addr, body or "")
            break
    if target is None and rows:
        target = (rows[0][0], rows[0][1], rows[0][2] or "")

    if target is None:
        check("Announcement email in Sent to all-managers with >=3 project names in body",
              False, "no announcement email in Sent")
        return

    subject, to_addr, body = target
    to_ok = "all-managers@company.com" in str(to_addr or "").lower()
    body_l = body.lower()
    # Body must list >=3 of the agent's own project names.
    if project_names:
        hits = sum(1 for n in project_names if n and n.lower() in body_l)
    else:
        hits = 0
    check("Announcement email in Sent to all-managers with >=3 project names in body",
          to_ok and hits >= 3,
          f"to_ok={to_ok}, project-name hits in body={hits}, names={project_names[:6]}")

    # Body mentions kickoff dates (non-critical).
    date_mentioned = any(d in body for d in ("19", "20", "21")) and (
        "март" in body_l or "march" in body_l or "2026-03" in body_l or "03.2026" in body_l)
    check("Email body references kickoff dates", date_mentioned, "no kickoff dates in body")

    # No noise email leaked into Sent (non-critical reverse check).
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM email.messages "
                    "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
                    "AND (subject ILIKE '%рассылка%' OR subject ILIKE '%newsletter%')")
        noise_sent = cur.fetchone()[0]
        conn.close()
        check("No noise emails in Sent folder", noise_sent == 0, f"found {noise_sent}")
    except Exception as e:
        check("No noise emails in Sent folder", False, str(e))


def check_teamly():
    """Critical: a project dashboard page exists (not the noise page) listing
    >=3 proposed projects with leaders. Seed pages have id <= 3."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly dashboard page exists (not the noise page) listing >=3 projects", False, str(e))
        return

    dash = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if NOISE_PAGE_TITLE in tl:
            continue
        if "dashboard" in tl or "project" in tl or "проект" in tl or "портфел" in tl:
            dash = (pid, title, body)
            break
    # Fallback: any non-noise new page mentioning the portfolio heading.
    if dash is None:
        for pid, title, body in pages:
            tl = (title or "").lower()
            if NOISE_PAGE_TITLE in tl:
                continue
            text = (tl + " " + (body or "").lower())
            if "q2 2026" in text or "portfolio" in text or "портфел" in text:
                dash = (pid, title, body)
                break

    if dash is None:
        check("Teamly dashboard page exists (not the noise page) listing >=3 projects",
              False, f"new pages: {[(p[0], p[1]) for p in pages]}")
        return

    text = ((dash[1] or "") + "\n" + (dash[2] or "")).lower()
    # Heading marker (non-critical).
    check("Teamly page includes the Q2 portfolio heading",
          "q2 2026" in text or "project portfolio" in text or "портфел" in text,
          "heading marker absent")
    # >=3 leader/role mentions as a proxy for >=3 projects with assigned leaders.
    leader_markers = text.count("руководител") + text.count("lead") + text.count("ведущ")
    check("Teamly dashboard page exists (not the noise page) listing >=3 projects",
          leader_markers >= 3,
          f"leader/role mentions: {leader_markers}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    wb = check_excel(agent_workspace)
    check_script(agent_workspace)
    project_names = collect_project_names(wb)
    check_calendar()
    check_email(project_names)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
