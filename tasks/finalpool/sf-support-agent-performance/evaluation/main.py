"""Evaluation for sf-support-agent-performance (ClickHouse fork).

The support data warehouse is queried LIVE via psycopg2 (schema sf_data, seeded
and russified centrally by db/zzz_clickhouse_after_init.sql). The static
groundtruth xlsx is intentionally NOT used as the source of truth: its
Tickets_Handled are all 0 / Avg_Satisfaction all None because the TICKETS table
has no agent foreign key (RESOLVER is empty), so per-agent ticket counts cannot
be derived from the data. We therefore validate:
  - the ACTIVE agent roster (names + TEAM + SKILL_LEVEL) against the live DB,
  - Summary counts against the live DB,
  - and the internal consistency of the agent's own deliverable (Top_Agent,
    descending sort, Total_Tickets sum, weighted Avg_Satisfaction_Overall).
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=int(os.environ.get("PGPORT", "5432")),
          dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {tag}{name}: {str(detail)[:300]}")
        if critical:
            CRITICAL_FAILS.append(name)


def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=1.0):
    fa, fb = num(a), num(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= tol


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_expected():
    """Live active-agent roster from the support data warehouse."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute('SELECT "AGENT_NAME","TEAM","SKILL_LEVEL" '
                'FROM sf_data."SUPPORT_CENTER__PUBLIC__AGENTS" '
                'WHERE "ACTIVE" = true ORDER BY "AGENT_NAME"')
    agents = [{"name": r[0], "team": r[1], "skill": r[2]} for r in cur.fetchall()]
    conn.close()
    return {"agents": agents, "total_active": len(agents)}


def sheet_rows(wb, name):
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            return [[c.value for c in row] for row in wb[sn].iter_rows()]
    return None


def sheet_dicts(rows):
    if not rows or len(rows) < 2:
        return []
    hdrs = [str(h).strip() if h is not None else "" for h in rows[0]]
    out = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        out.append({hdrs[i]: (row[i] if i < len(row) else None) for i in range(len(hdrs))})
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", default=".")
    parser.add_argument("--groundtruth_workspace", default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    exp = get_expected()
    exp_names = {a["name"].strip().lower(): a for a in exp["agents"]}

    agent_file = os.path.join(args.agent_workspace, "Support_Agent_Performance.xlsx")
    if not os.path.exists(agent_file):
        print(f"  [FAIL] CRITICAL Support_Agent_Performance.xlsx exists: {agent_file}")
        print("\n=== RESULT: FAIL (deliverable missing) ===")
        sys.exit(1)

    wb = openpyxl.load_workbook(agent_file, data_only=True)

    # ---- Agent Performance sheet ----
    print("=== Checking Agent Performance ===")
    ap_rows = sheet_rows(wb, "Agent Performance")
    perf = None
    if ap_rows is None:
        record("Sheet 'Agent Performance' present", False, str(wb.sheetnames), critical=True)
    else:
        record("Sheet 'Agent Performance' present", True)
        perf = sheet_dicts(ap_rows)
        produced = {}
        for r in perf:
            k = str(r.get("Agent_Name")).strip().lower() if r.get("Agent_Name") is not None else None
            if k:
                produced[k] = r

        # CRITICAL: roster set must exactly equal live active agents
        missing = [a["name"] for a in exp["agents"] if a["name"].strip().lower() not in produced]
        extra = [r.get("Agent_Name") for k, r in produced.items() if k not in exp_names]
        record("Active-agent roster exact match", not missing and not extra,
               f"missing={missing} extra={extra}", critical=True)

        # CRITICAL: TEAM + SKILL_LEVEL correct per agent (enums stay English)
        td_ok, td_bad = True, []
        for a in exp["agents"]:
            r = produced.get(a["name"].strip().lower())
            if r is None:
                td_ok = False
                continue
            if not str_match(r.get("Team"), a["team"]) or not str_match(r.get("Skill_Level"), a["skill"]):
                td_ok = False
                td_bad.append(f"{a['name']}: team={r.get('Team')}/{a['team']} skill={r.get('Skill_Level')}/{a['skill']}")
        record("Team/Skill_Level correct per agent", td_ok, td_bad[:5], critical=True)

        # Tickets_Handled present & numeric for every row
        th_ok = all(num(r.get("Tickets_Handled")) is not None for r in produced.values())
        record("Tickets_Handled numeric for all rows", th_ok)

        # CRITICAL: sheet sorted by Tickets_Handled descending
        seq = [num(r.get("Tickets_Handled")) for r in perf if r.get("Agent_Name") is not None]
        seq = [v for v in seq if v is not None]
        sorted_desc = all(seq[i] >= seq[i + 1] for i in range(len(seq) - 1))
        record("Sorted by Tickets_Handled descending", sorted_desc, seq, critical=True)

    # ---- Summary sheet ----
    print("=== Checking Summary ===")
    sm_rows = sheet_rows(wb, "Summary")
    if sm_rows is None:
        record("Sheet 'Summary' present", False, str(wb.sheetnames), critical=True)
    else:
        record("Sheet 'Summary' present", True)
        sm = sheet_dicts(sm_rows)
        metrics = {}
        for r in sm:
            mk = r.get("Metric")
            if mk is None:
                continue
            metrics[str(mk).strip()] = r.get("Value")

        # CRITICAL: Total_Active_Agents == live count of active agents (exact)
        record("Total_Active_Agents == live active count",
               num_close(metrics.get("Total_Active_Agents"), exp["total_active"], 0),
               f"{metrics.get('Total_Active_Agents')} vs {exp['total_active']}", critical=True)

        if perf is not None:
            rows = [r for r in perf if r.get("Agent_Name") is not None]
            tickets = [(r.get("Agent_Name"), num(r.get("Tickets_Handled")), num(r.get("Avg_Satisfaction")))
                       for r in rows]
            tickets = [(n, t, s) for (n, t, s) in tickets if t is not None]
            total_tickets = sum(t for _, t, _ in tickets)

            # CRITICAL: Top_Agent consistent with the agent's own max-tickets row
            if tickets:
                top_name = max(tickets, key=lambda x: x[1])[0]
                record("Top_Agent matches max Tickets_Handled in sheet",
                       str_match(metrics.get("Top_Agent"), top_name),
                       f"{metrics.get('Top_Agent')} vs {top_name}", critical=True)

            # Total_Tickets internally consistent with per-agent sum
            record("Total_Tickets == sum of per-agent Tickets_Handled",
                   num_close(metrics.get("Total_Tickets"), total_tickets, 1),
                   f"{metrics.get('Total_Tickets')} vs {total_tickets}")

            # Avg_Satisfaction_Overall: weighted average consistent with rows.
            # If all satisfactions are blank/zero (degenerate data), accept the
            # reported value as-is; otherwise require weighted-average consistency.
            wsum = sum(t * s for _, t, s in tickets if s is not None)
            wden = sum(t for _, t, s in tickets if s is not None)
            if wden > 0:
                expected_avg = round(wsum / wden, 2)
                record("Avg_Satisfaction_Overall == weighted average",
                       num_close(metrics.get("Avg_Satisfaction_Overall"), expected_avg, 0.05),
                       f"{metrics.get('Avg_Satisfaction_Overall')} vs {expected_avg}")
            else:
                record("Avg_Satisfaction_Overall present", "Avg_Satisfaction_Overall" in metrics,
                       str(metrics.keys()))

    wb.close()

    # ---- Word document ----
    print("=== Checking Agent_Review.docx ===")
    docx_path = os.path.join(args.agent_workspace, "Agent_Review.docx")
    if not os.path.exists(docx_path):
        record("Agent_Review.docx exists", False, docx_path, critical=True)
    else:
        record("Agent_Review.docx exists", True)
        try:
            from docx import Document
            doc = Document(docx_path)
            text = " ".join(p.text for p in doc.paragraphs).lower()
            record("Agent_Review.docx has substantive text (>=50 chars)",
                   len(text.strip()) >= 50, f"len={len(text.strip())}")
            kws = ["agent", "review", "агент", "обзор", "отчёт", "поддержк", "удовлетвор"]
            record("Agent_Review.docx mentions support/agent topic",
                   any(k in text for k in kws), text[:120])
        except ImportError:
            record("Agent_Review.docx non-trivial", os.path.getsize(docx_path) >= 100)
        except Exception as e:
            record("Agent_Review.docx readable", False, str(e))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY: {PASS_COUNT}/{total} passed ({accuracy:.1f}%), "
          f"{len(CRITICAL_FAILS)} critical fails ===")

    success = (not CRITICAL_FAILS) and accuracy >= 70
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy, "critical_fails": CRITICAL_FAILS,
                       "success": success}, f)

    if CRITICAL_FAILS:
        print(f"=== RESULT: FAIL (critical: {CRITICAL_FAILS}) ===")
        sys.exit(1)
    if accuracy < 70:
        print("=== RESULT: FAIL (accuracy < 70) ===")
        sys.exit(1)
    print("=== RESULT: PASS ===")
    sys.exit(0)


if __name__ == "__main__":
    main()
