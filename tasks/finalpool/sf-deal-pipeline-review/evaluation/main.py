"""Evaluation for sf-deal-pipeline-review (ClickHouse / sf_data, russified labels)."""
import argparse
import os
import sys

import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Pipeline_Review.xlsx")
    gt_file = os.path.join(gt_dir, "Pipeline_Review.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    # CRITICAL checks: each is (passed: bool, message: str). Any failure => hard FAIL.
    critical_results = []

    def crit(passed, msg):
        critical_results.append((bool(passed), msg))

    # ---- Build groundtruth lookups ----
    g_reg_rows = load_sheet_rows(gt_wb, "Regional Performance") or []
    g_reg_data = g_reg_rows[1:] if len(g_reg_rows) > 1 else []
    gt_regional = {}  # key -> row
    for r in g_reg_data:
        if r and r[0] is not None:
            gt_regional[str(r[0]).strip().lower()] = r

    g_seg_rows = load_sheet_rows(gt_wb, "Segment Breakdown") or []
    g_seg_data = g_seg_rows[1:] if len(g_seg_rows) > 1 else []
    gt_segment = {}
    for r in g_seg_data:
        if r and r[0] is not None and r[1] is not None:
            gt_segment[f"{str(r[0]).strip().lower()}|{str(r[1]).strip().lower()}"] = r

    g_sum_rows = load_sheet_rows(gt_wb, "Summary") or []
    g_sum_data = g_sum_rows[1:] if len(g_sum_rows) > 1 else []
    gt_summary = {}
    for r in g_sum_data:
        if r and r[0] is not None:
            gt_summary[str(r[0]).strip().lower()] = r

    # Check Regional Performance
    print("  Checking Regional Performance...")
    a_rows = load_sheet_rows(agent_wb, "Regional Performance")
    a_reg_lookup = {}
    if a_rows is None:
        all_errors.append("Sheet 'Regional Performance' not found")
        crit(False, "Regional Performance sheet missing")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        a_reg_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for key, g_row in gt_regional.items():
            a_row = a_reg_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing region: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 100):
                    errors.append(f"{key}.Target: {a_row[1]} vs {g_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 500):
                    errors.append(f"{key}.Actual: {a_row[2]} vs {g_row[2]}")
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 2.0):
                    errors.append(f"{key}.Gap_Pct: {a_row[4]} vs {g_row[4]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # CRITICAL 1: every groundtruth region present with correct Q1_Actual and On_Track='No'
    reg_actual_ok = bool(gt_regional)
    reg_ontrack_ok = bool(gt_regional)
    for key, g_row in gt_regional.items():
        a_row = a_reg_lookup.get(key)
        if a_row is None:
            reg_actual_ok = False
            reg_ontrack_ok = False
            continue
        if not (len(a_row) > 2 and num_close(a_row[2], g_row[2], 500)):
            reg_actual_ok = False
        # On_Track column index 5
        if not (len(a_row) > 5 and str_match(a_row[5], "No")):
            reg_ontrack_ok = False
    crit(reg_actual_ok,
         "Regional Performance: each russified region present with Q1_Actual within tolerance")
    crit(reg_ontrack_ok,
         "Regional Performance: On_Track='No' for all five regions")

    # Check Segment Breakdown
    print("  Checking Segment Breakdown...")
    a_rows = load_sheet_rows(agent_wb, "Segment Breakdown")
    a_seg_lookup = {}
    if a_rows is None:
        all_errors.append("Sheet 'Segment Breakdown' not found")
        crit(False, "Segment Breakdown sheet missing")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        for r in a_data:
            if r and r[0] is not None and r[1] is not None:
                k = f"{str(r[0]).strip().lower()}|{str(r[1]).strip().lower()}"
                a_seg_lookup[k] = r
        errors = []
        for key, g_row in gt_segment.items():
            a_row = a_seg_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}|{g_row[1]}")
                continue
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 500):
                    errors.append(f"{key}.Actual: {a_row[3]} vs {g_row[3]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # CRITICAL 2: all 20 region-segment rows present (keyed by Russian labels) with Actual within tolerance
    seg_ok = bool(gt_segment)
    missing_or_wrong = 0
    for key, g_row in gt_segment.items():
        a_row = a_seg_lookup.get(key)
        if a_row is None or not (len(a_row) > 3 and num_close(a_row[3], g_row[3], 500)):
            seg_ok = False
            missing_or_wrong += 1
    crit(seg_ok,
         f"Segment Breakdown: all {len(gt_segment)} russified region|segment rows present and Actual correct "
         f"(wrong/missing: {missing_or_wrong})")

    # Check Summary
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    a_sum_lookup = {}
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found")
        crit(False, "Summary sheet missing")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        a_sum_lookup = {str(r[0]).strip().lower(): r for r in a_data if r and r[0] is not None}
        errors = []
        for key, g_row in gt_summary.items():
            a_row = a_sum_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if key in ("worst_region", "best_region"):
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                elif key in ("regions_on_track", "regions_behind"):
                    if not num_close(a_row[1], g_row[1], 1):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
                else:
                    if not num_close(a_row[1], g_row[1], 1000):
                        errors.append(f"{key}: {a_row[1]} vs {g_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # CRITICAL 3: Summary counts and money totals match groundtruth
    def sum_val(key):
        r = a_sum_lookup.get(key)
        return r[1] if r and len(r) > 1 else None

    on_track_ok = num_close(sum_val("regions_on_track"), gt_summary.get("regions_on_track", [None, None])[1], 0)
    behind_ok = num_close(sum_val("regions_behind"), gt_summary.get("regions_behind", [None, None])[1], 0)
    crit(on_track_ok and behind_ok,
         "Summary: Regions_On_Track and Regions_Behind match groundtruth counts")

    tgt_ok = num_close(sum_val("total_q1_target"), gt_summary.get("total_q1_target", [None, None])[1], 100)
    act_ok = num_close(sum_val("total_q1_actual"), gt_summary.get("total_q1_actual", [None, None])[1], 1000)
    crit(tgt_ok and act_ok,
         "Summary: Total_Q1_Target and Total_Q1_Actual match groundtruth within tolerance")

    # CRITICAL 4: Worst_Region / Best_Region match russified groundtruth names
    worst_ok = str_match(sum_val("worst_region"), gt_summary.get("worst_region", [None, None])[1])
    best_ok = str_match(sum_val("best_region"), gt_summary.get("best_region", [None, None])[1])
    crit(worst_ok and best_ok,
         "Summary: Worst_Region/Best_Region equal russified groundtruth region names")

    # Check Google Calendar events
    print("  Checking Google Calendar events...")
    gcal_ok = False
    try:
        db_config = {
            "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
            "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
            "user": "eigent", "password": "camel",
        }
        conn = psycopg2.connect(**db_config)
        cur = conn.cursor()
        cur.execute("SELECT DISTINCT summary FROM gcal.events WHERE summary LIKE '%Pipeline Review%'")
        events = cur.fetchall()
        cur.close()
        conn.close()

        # All 5 regions are behind target (>15%), so expect 5 distinct review meetings.
        if len(events) < 5:
            all_errors.append(f"Expected 5 review meetings, found {len(events)}")
            print(f"    FAIL: expected 5 meetings, found {len(events)}")
        else:
            gcal_ok = True
            print("    PASS")
    except Exception as e:
        all_errors.append(f"GCal check error: {e}")
        print(f"    ERROR: {e}")

    # CRITICAL 5: exactly the 5 expected distinct 'Pipeline Review' events
    crit(gcal_ok,
         "Google Calendar: >=5 distinct events whose summary contains 'Pipeline Review'")

    # ---- CRITICAL gate (semantic substance). Any critical failure => hard FAIL. ----
    print("\n  CRITICAL checks:")
    failed_critical = [m for ok, m in critical_results if not ok]
    for ok, m in critical_results:
        print(f"    [{'PASS' if ok else 'FAIL'}] {m}")
    if failed_critical:
        print(f"\n=== RESULT: FAIL (critical check failed: {len(failed_critical)}) ===")
        sys.exit(1)

    # ---- Accuracy gate (non-critical structural/tolerance errors) ----
    # Total comparable cells across the three sheets:
    total_checks = (
        len(gt_regional) * 3        # Target, Actual, Gap_Pct
        + len(gt_segment) * 1       # Actual
        + len(gt_summary) * 1       # one value each
        + 1                         # gcal
    )
    accuracy = max(0.0, 100.0 * (total_checks - len(all_errors)) / total_checks) if total_checks else 0.0
    print(f"\n  Accuracy: {accuracy:.1f}% ({len(all_errors)} non-critical errors of {total_checks} checks)")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
