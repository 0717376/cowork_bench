"""Evaluation for terminal-sf-insales-order-audit-excel-word (RU swap: ClickHouse + InSales).
Checks:
1. Order_Audit_Report.xlsx with 4 sheets and correct data
2. Audit_Findings.docx with required sections
3. audit_analysis.py script exists

NOTE (RU swap): sf_data.* DATA VALUES are russified centrally
(db/zzz_clickhouse_after_init.sql): STATUS Delivered->Доставлен, Cancelled->Отменён,
Processing->В обработке, Shipped->Отправлен; SHIP_MODE Economy->Эконом, Express->Экспресс,
Next Day->На следующий день, Standard->Стандарт. The agent reads the DB honestly and
writes RU strings to Excel/Word, so all greps accept RU+EN alternatives. Live aggregates
are computed against the russified table sf_data."SALES_DW__PUBLIC__ORDERS".
CRITICAL_CHECKS run before the accuracy gate; any critical fail => sys.exit(1).
"""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

# RU+EN alternatives for honestly-read russified DB values / prose
STATUS_ALTS = {
    "delivered": ["delivered", "доставлен"],
    "cancelled": ["cancelled", "canceled", "отмен"],
    "processing": ["processing", "обработк"],
    "shipped": ["shipped", "отправлен"],
}
SHIPMODE_ALTS = {
    "economy": ["economy", "эконом"],
    "express": ["express", "экспресс"],
    "next day": ["next day", "следующ"],
    "standard": ["standard", "стандарт"],
}


def any_in(text, alts):
    """True if any alternative substring is present in text."""
    return any(a in text for a in alts)


PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}: {str(detail)[:200]}")


def fetch_dw_aggregates():
    """Live aggregates from the russified DW orders table.
    Returns dict or None on failure. Tries RU then EN status values, and
    SALES then TOTAL_AMOUNT revenue columns, for robustness across seeds."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Pick a revenue column that exists on the ORDERS table
        cur.execute("""
            SELECT column_name FROM information_schema.columns
            WHERE table_schema='sf_data' AND table_name='SALES_DW__PUBLIC__ORDERS'
        """)
        cols = {r[0].upper() for r in cur.fetchall()}
        rev_col = "SALES" if "SALES" in cols else ("TOTAL_AMOUNT" if "TOTAL_AMOUNT" in cols else None)
        if rev_col is None:
            cur.close(); conn.close(); return None
        # Delivered = russified 'Доставлен' (fallback English 'Delivered')
        cur.execute("""
            SELECT COUNT(*), COALESCE(SUM("%s"::float), 0)
            FROM sf_data."SALES_DW__PUBLIC__ORDERS"
            WHERE LOWER("STATUS") IN ('доставлен','delivered')
        """ % rev_col)
        dcount, drev = cur.fetchone()
        cur.execute('SELECT COUNT(*), COALESCE(SUM("%s"::float), 0) FROM sf_data."SALES_DW__PUBLIC__ORDERS"' % rev_col)
        total_count, total_rev = cur.fetchone()
        cur.close(); conn.close()
        return {
            "delivered_count": int(dcount),
            "delivered_revenue": float(drev),
            "total_count": int(total_count),
            "total_revenue": float(total_rev),
        }
    except Exception as e:
        print(f"  [warn] fetch_dw_aggregates failed: {e}")
        return None


# Groundtruth fallback if DB unreachable (Superstore DW snapshot)
GT_FALLBACK = {
    "delivered_count": 14033,
    "delivered_revenue": 2177149.66,
    "total_count": 20000,
    "total_revenue": 3048998.33,
}


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').replace('%', '').strip())
    except Exception:
        return default


def check_excel(workspace, dw):
    print("\n=== Check 1: Order_Audit_Report.xlsx ===")
    path = os.path.join(workspace, "Order_Audit_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}", critical=True)
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheets_lower = [s.lower() for s in sheets]

    # Check 4 sheets
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    # ---- DW_Summary sheet ----
    dw_idx = next((i for i, s in enumerate(sheets_lower) if "dw" in s or "warehouse" in s), 0)
    ws_dw = wb[sheets[dw_idx]]
    rows_dw = list(ws_dw.iter_rows(values_only=True))
    if len(rows_dw) > 1:
        all_text = " ".join(str(c) for r in rows_dw for c in r if c).lower()
        # RU+EN status presence
        check("DW sheet has Delivered/Доставлен status",
              any_in(all_text, STATUS_ALTS["delivered"]), f"Content: {all_text[:120]}")
        check("DW sheet has Cancelled/Отменён status",
              any_in(all_text, STATUS_ALTS["cancelled"]), f"Content: {all_text[:120]}")
        check("DW sheet has Processing/В обработке status",
              any_in(all_text, STATUS_ALTS["processing"]), f"Content: {all_text[:120]}")
        check("DW sheet has Shipped/Отправлен status",
              any_in(all_text, STATUS_ALTS["shipped"]), f"Content: {all_text[:120]}")

        exp_d = dw["delivered_count"]
        exp_rev = dw["delivered_revenue"]

        # Find Delivered row (RU or EN) and check count + revenue (CRITICAL)
        delivered_row_found = False
        for row in rows_dw[1:]:
            label = str(row[0]).lower() if row[0] else ""
            if any_in(label, STATUS_ALTS["delivered"]):
                delivered_row_found = True
                count = safe_float(row[1])
                check(f"Delivered order count ~{exp_d}",
                      count is not None and abs(count - exp_d) < max(100, exp_d * 0.01),
                      f"Got {count}, expected ~{exp_d}", critical=True)
                rev = safe_float(row[2])
                check(f"Delivered revenue ~{exp_rev:.0f}",
                      rev is not None and abs(rev - exp_rev) < max(5000, exp_rev * 0.01),
                      f"Got {rev}, expected ~{exp_rev:.0f}", critical=True)
                break
        if not delivered_row_found:
            check("DW Delivered row present", False,
                  "No Delivered/Доставлен row found", critical=True)

        # Totals row Order_Count ~ live total order count (CRITICAL: completeness)
        exp_total = dw["total_count"]
        total_count_match = False
        for row in rows_dw[1:]:
            label = str(row[0]).lower() if row[0] else ""
            if "total" in label or "итог" in label or "всего" in label:
                tc = safe_float(row[1])
                if tc is not None and abs(tc - exp_total) < max(100, exp_total * 0.02):
                    total_count_match = True
                    break
        # also accept: sum of status Order_Counts equals total
        if not total_count_match:
            status_sum = 0.0
            for row in rows_dw[1:]:
                label = str(row[0]).lower() if row[0] else ""
                if any(any_in(label, alts) for alts in STATUS_ALTS.values()):
                    v = safe_float(row[1])
                    if v is not None:
                        status_sum += v
            total_count_match = abs(status_sum - exp_total) < max(100, exp_total * 0.02)
        check(f"DW totals row / status sum ~{exp_total} orders",
              total_count_match,
              f"No totals row and status counts do not sum to ~{exp_total}", critical=True)
    else:
        check("DW sheet has data", False, "Sheet is empty", critical=True)

    # ---- Store_Summary sheet ----
    store_idx = next((i for i, s in enumerate(sheets_lower) if "store" in s), 1)
    if store_idx < len(sheets):
        ws_store = wb[sheets[store_idx]]
        rows_store = list(ws_store.iter_rows(values_only=True))
        all_text_store = " ".join(str(c) for r in rows_store for c in r if c).lower()
        # Real assertion: a Total Products metric row exists with a positive numeric value
        prod_count_ok = False
        for row in rows_store[1:]:
            label = str(row[0]).lower() if row[0] else ""
            if "product" in label or "товар" in label:
                v = safe_float(row[1])
                if v is not None and v > 0:
                    prod_count_ok = True
                    break
        check("Store sheet has positive Total Products metric", prod_count_ok,
              f"Content snippet: {all_text_store[:120]}")
        # Store summary should also carry orders & customers metrics
        check("Store sheet has orders metric",
              "order" in all_text_store or "заказ" in all_text_store,
              f"Content: {all_text_store[:120]}")
        check("Store sheet has customers metric",
              "customer" in all_text_store or "клиент" in all_text_store,
              f"Content: {all_text_store[:120]}")

    # ---- ShipMode sheet ----
    ship_idx = next((i for i, s in enumerate(sheets_lower) if "ship" in s or "mode" in s), 2)
    if ship_idx < len(sheets):
        ws_ship = wb[sheets[ship_idx]]
        rows_ship = list(ws_ship.iter_rows(values_only=True))
        all_text_ship = " ".join(str(c) for r in rows_ship for c in r if c).lower()
        check("ShipMode sheet has Economy/Эконом",
              any_in(all_text_ship, SHIPMODE_ALTS["economy"]), f"Content: {all_text_ship[:120]}")
        check("ShipMode sheet has Express/Экспресс",
              any_in(all_text_ship, SHIPMODE_ALTS["express"]), f"Content: {all_text_ship[:120]}")
        check("ShipMode sheet has Next Day/На следующий день",
              any_in(all_text_ship, SHIPMODE_ALTS["next day"]), f"Content: {all_text_ship[:120]}")
        check("ShipMode sheet has Standard/Стандарт",
              any_in(all_text_ship, SHIPMODE_ALTS["standard"]), f"Content: {all_text_ship[:120]}")

        # CRITICAL: all 4 ship modes present, each with positive Order_Count,
        # and Total_Revenue sums to ~ live DW total revenue.
        exp_total_rev = dw["total_revenue"]
        modes_with_positive = 0
        rev_sum = 0.0
        for row in rows_ship[1:]:
            label = str(row[0]).lower() if row[0] else ""
            if any(any_in(label, alts) for alts in SHIPMODE_ALTS.values()):
                cnt = safe_float(row[1])
                rev = safe_float(row[2])
                if cnt is not None and cnt > 0:
                    modes_with_positive += 1
                if rev is not None:
                    rev_sum += rev
        check("ShipMode: all 4 modes have positive Order_Count",
              modes_with_positive >= 4,
              f"Modes with positive count: {modes_with_positive}", critical=True)
        check(f"ShipMode Total_Revenue sums to DW total ~{exp_total_rev:.0f}",
              abs(rev_sum - exp_total_rev) < max(20000, exp_total_rev * 0.02),
              f"Got sum {rev_sum:.0f}, expected ~{exp_total_rev:.0f}", critical=True)

    # ---- Reconciliation sheet ----
    recon_idx = next((i for i, s in enumerate(sheets_lower) if "recon" in s), 3)
    if recon_idx < len(sheets):
        ws_recon = wb[sheets[recon_idx]]
        rows_recon = list(ws_recon.iter_rows(values_only=True))
        check("Reconciliation sheet has data", len(rows_recon) > 1,
              f"Found {len(rows_recon)} rows")
        if len(rows_recon) > 1:
            all_text_recon = " ".join(str(c) for r in rows_recon for c in r if c).lower()
            check("Reconciliation has order count comparison",
                  "order" in all_text_recon or "заказ" in all_text_recon
                  or str(dw["total_count"]) in all_text_recon,
                  f"Content: {all_text_recon[:120]}")
            # CRITICAL: Difference == DW_Value - Store_Value arithmetic consistency
            # on the order-count row (Metric, DW_Value, Store_Value, Difference).
            recon_arith_ok = False
            for row in rows_recon[1:]:
                label = str(row[0]).lower() if row[0] else ""
                if ("order" in label and "count" in label) or "количеств" in label or "заказ" in label:
                    dwv = safe_float(row[1]) if len(row) > 1 else None
                    stv = safe_float(row[2]) if len(row) > 2 else None
                    diff = safe_float(row[3]) if len(row) > 3 else None
                    if dwv is not None and stv is not None and diff is not None:
                        if abs(diff - (dwv - stv)) <= max(1.0, abs(dwv - stv) * 0.01):
                            recon_arith_ok = True
                            break
            check("Reconciliation order-count Difference = DW_Value - Store_Value",
                  recon_arith_ok,
                  "No order-count row with consistent Difference = DW - Store", critical=True)


def check_word(workspace):
    print("\n=== Check 2: Audit_Findings.docx ===")
    path = os.path.join(workspace, "Audit_Findings.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    try:
        from docx import Document
        doc = Document(path)
        full_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Doc mentions reconciliation or audit (RU+EN)",
              any(k in full_text for k in ["reconcil", "audit", "сверк", "аудит", "реконсил"]),
              f"Text snippet: {full_text[:120]}")
        check("Doc mentions data warehouse (RU+EN)",
              any(k in full_text for k in ["warehouse", "dw", "хранилищ"]),
              f"Text snippet: {full_text[:120]}")
        check("Doc mentions online store / InSales (RU+EN)",
              any(k in full_text for k in ["store", "insales", "ecommerce", "магазин"]),
              f"Text snippet: {full_text[:120]}")
        check("Doc mentions shipping or ship mode (RU+EN)",
              any(k in full_text for k in ["ship", "доставк", "способ доставки"]),
              f"Text snippet: {full_text[:120]}")
        check("Doc has recommendation section (RU+EN)",
              any(k in full_text for k in ["recommend", "рекоменд"]),
              f"Text snippet: {full_text[:200]}")
    except Exception as e:
        check("Word document readable", False, str(e))


def check_reverse_validation(workspace):
    """Reverse validation: check things that should NOT exist."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Order_Audit_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
        sheets_lower = [s.lower() for s in wb.sheetnames]
        # No unexpected/garbage sheets
        valid_keywords = ["dw", "warehouse", "store", "ship", "mode", "recon", "summary", "audit", "order"]
        unexpected = [s for s in wb.sheetnames
                      if not any(k in s.lower() for k in valid_keywords) and s.lower() != "sheet1"]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected sheets: {unexpected}")

        # Check no test/debug/placeholder data leaked (real assertion)
        all_text = ""
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                all_text += " ".join(str(c) for c in row if c) + " "
        all_lower = all_text.lower()
        leaked = [tok for tok in ["lorem", "todo", "fixme", "placeholder", "xxxxx", "dummy", "заглушк"]
                  if tok in all_lower]
        check("No debug/placeholder tokens in Excel", len(leaked) == 0,
              f"Found tokens: {leaked}")
        check("No negative order counts", all(
            safe_float(c, 0) >= 0 for ws in wb.worksheets
            for row in ws.iter_rows(min_row=2, values_only=True)
            for c in [row[1]] if c is not None and isinstance(safe_float(c), (int, float))
        ), "Found negative counts")


def check_script(workspace):
    print("\n=== Check 3: audit_analysis.py ===")
    path = os.path.join(workspace, "audit_analysis.py")
    check("audit_analysis.py exists", os.path.exists(path), f"Not found at {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # Live DW aggregates (russified table); fall back to groundtruth snapshot.
    dw = fetch_dw_aggregates()
    if dw is None:
        print("  [warn] Using groundtruth DW fallback (DB unreachable).")
        dw = dict(GT_FALLBACK)

    check_excel(args.agent_workspace, dw)
    check_word(args.agent_workspace)
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_failed": CRITICAL_FAILED}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # CRITICAL gate: any critical failure => FAIL regardless of accuracy.
    if CRITICAL_FAILED:
        print(f"\nCRITICAL CHECK(S) FAILED: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
