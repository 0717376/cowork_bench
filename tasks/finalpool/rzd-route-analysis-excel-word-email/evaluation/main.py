"""
Evaluation для rzd-route-analysis-excel-word-email (РЖД, «Сапсан» Москва → СПб).

Источник данных — rzd.* (только чтение). На 2026-03-10 маршрут MOW → SPB:
  4 поезда: 752А (06:50→10:50, 240 мин), 754А (09:30→13:25, 235 мин),
            756А (13:30→17:30, 240 мин), 758А (19:00→23:00, 240 мин).
  Цены (₽): Эконом 5500 (Second_Class), Эконом+ 7500, Бизнес 14000 (First_Class).
  Доступность мест — текст «много» для всех (число не задаётся).
  Самый быстрый — 754А (235 мин, единственный минимум). Это и есть рекомендация.

Что проверяем:
1. Route_Analysis.xlsx существует, листы Routes и Summary.
2. Routes: ровно 4 строки с поездами 752А/754А/756А/758А, назначение Санкт-Петербург,
   цена эконома 5500 ₽.
3. Summary: Fastest_Route = 754А, Total_Routes = 4.
4. Route_Analysis_Report.docx: 4 раздела, в Recommendation указан 754А (скорость).
5. Письмо на logistics@company.com: тело содержит код 754А и его время.

CRITICAL_CHECKS — содержательные проверки: любой провал => FAIL независимо от accuracy.
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Содержательные проверки. Провал любой => вся задача FAIL.
CRITICAL_CHECKS = {
    "Routes: ровно 4 поезда 752А/754А/756А/758А",
    "Routes: цена эконом-класса 5500 ₽ присутствует",
    "Summary: Fastest_Route = 754А",
    "Word: Recommendation указывает 754А как самый быстрый",
    "Email: тело содержит рекомендованный поезд 754А и его время",
}

# Ожидаемые данные
EXPECTED_TRAINS = ["752А", "754А", "756А", "758А"]
FASTEST = "754А"
ECONOM_PRICE = 5500
BIZ_PRICE = 14000
FASTEST_DEPART = "09:30"
FASTEST_ARRIVE = "13:25"


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def normalize_train(s):
    """Кириллица → латиница для номеров поездов (752А ≈ 752A)."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    cyr_to_lat = str.maketrans({
        "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M",
        "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T",
        "У": "Y", "Х": "X", "Г": "G",
    })
    return s.translate(cyr_to_lat)


def check_excel(agent_workspace):
    print("\n=== Проверка 1: Route_Analysis.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Route_Analysis.xlsx")
    if not os.path.exists(xlsx_path):
        record("Route_Analysis.xlsx существует", False, f"Не найден: {xlsx_path}")
        return
    record("Route_Analysis.xlsx существует", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel читается", False, str(e))
        return
    record("Excel читается", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # --- Routes ---
    if "routes" not in sheet_names_lower:
        record("Лист Routes существует", False, f"Листы: {wb.sheetnames}")
        record("Routes: ровно 4 поезда 752А/754А/756А/758А", False, "нет листа Routes")
        record("Routes: цена эконом-класса 5500 ₽ присутствует", False, "нет листа Routes")
    else:
        record("Лист Routes существует", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("routes")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
        record("Routes: 4 строки данных", len(data_rows) == 4, f"Найдено {len(data_rows)}")

        all_text_norm = normalize_train(
            " ".join(str(c) for r in rows for c in r if c is not None)
        )
        found = [t for t in EXPECTED_TRAINS if normalize_train(t) in all_text_norm]
        record("Routes: ровно 4 поезда 752А/754А/756А/758А",
               len(found) == 4 and len(data_rows) == 4,
               f"Найдены коды: {found}, строк данных: {len(data_rows)}")

        # назначение — Санкт-Петербург
        all_text_lower = " ".join(str(c) for r in rows for c in r if c is not None).lower()
        record("Routes: назначение Санкт-Петербург",
               "санкт" in all_text_lower or "spb" in all_text_lower or "петербург" in all_text_lower,
               all_text_lower[:150])

        # числовые значения
        numeric_vals = []
        for r in data_rows:
            for c in r:
                try:
                    numeric_vals.append(float(str(c).replace(",", ".").replace("₽", "").strip()))
                except (TypeError, ValueError):
                    pass
        has_econom = any(abs(v - ECONOM_PRICE) < 1 for v in numeric_vals)
        record("Routes: цена эконом-класса 5500 ₽ присутствует", has_econom,
               f"Числа: {numeric_vals[:20]}")
        has_biz = any(abs(v - BIZ_PRICE) < 1 for v in numeric_vals)
        record("Routes: цена бизнес-класса 14000 ₽ присутствует", has_biz,
               f"Числа: {numeric_vals[:20]}")
        # длительность 235 (самый быстрый) и 240 присутствуют
        has_235 = any(abs(v - 235) < 2 for v in numeric_vals)
        has_240 = any(abs(v - 240) < 2 for v in numeric_vals)
        record("Routes: длительности 235 и 240 мин присутствуют", has_235 and has_240,
               f"Числа: {numeric_vals[:20]}")

    # --- Summary ---
    if "summary" not in sheet_names_lower:
        record("Лист Summary существует", False, f"Листы: {wb.sheetnames}")
        record("Summary: Fastest_Route = 754А", False, "нет листа Summary")
    else:
        record("Лист Summary существует", True)
        ws_s = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows_s = list(ws_s.iter_rows(values_only=True))

        # Найдём значение в строке, где Label содержит "fastest"
        fastest_val = None
        total_val = None
        for r in rows_s:
            cells = [str(c).strip() if c is not None else "" for c in r]
            joined_low = " ".join(cells).lower()
            if "fastest" in joined_low and len(cells) >= 2:
                # значение — последняя непустая ячейка после метки
                for c in cells[1:]:
                    if c:
                        fastest_val = c
            if "total_routes" in joined_low and len(cells) >= 2:
                for c in cells[1:]:
                    if c:
                        total_val = c

        fastest_ok = fastest_val is not None and normalize_train(FASTEST) in normalize_train(fastest_val)
        record("Summary: Fastest_Route = 754А", fastest_ok,
               f"Fastest_Route = {fastest_val}")

        total_ok = total_val is not None and "4" in str(total_val)
        record("Summary: Total_Routes = 4", total_ok, f"Total_Routes = {total_val}")


def check_word(agent_workspace):
    print("\n=== Проверка 2: Route_Analysis_Report.docx ===")

    docx_path = os.path.join(agent_workspace, "Route_Analysis_Report.docx")
    if not os.path.exists(docx_path):
        record("Route_Analysis_Report.docx существует", False, f"Не найден: {docx_path}")
        record("Word: Recommendation указывает 754А как самый быстрый", False, "нет файла")
        return
    record("Route_Analysis_Report.docx существует", True)

    try:
        from docx import Document
        doc = Document(docx_path)
    except Exception as e:
        record("Word читается", False, str(e))
        record("Word: Recommendation указывает 754А как самый быстрый", False, str(e))
        return
    record("Word читается", True)

    full_text_orig = "\n".join(p.text for p in doc.paragraphs)
    full_text = full_text_orig.lower()

    record("Есть раздел Executive Summary", "executive summary" in full_text)
    record("Есть раздел Route Details", "route details" in full_text or "route detail" in full_text)
    record("Есть раздел Recommendation", "recommendation" in full_text)
    record("Есть раздел Cost Comparison", "cost comparison" in full_text)

    # Recommendation называет 754А и обосновывает скоростью
    norm_text = normalize_train(full_text_orig)
    names_fastest = normalize_train(FASTEST) in norm_text
    speed_words = any(w in full_text for w in ["быстр", "скорост", "минимальн", "fastest", "speed", "235"])
    record("Word: Recommendation указывает 754А как самый быстрый",
           names_fastest and speed_words,
           f"754А={names_fastest}, speed-слова={speed_words}")


def check_email():
    print("\n=== Проверка 3: Письмо на logistics@company.com ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%route analysis%'
           OR subject ILIKE '%москва%'
           OR subject ILIKE '%санкт%'
           OR subject ILIKE '%спб%'
    """)
    messages = cur.fetchall()
    cur.close()
    conn.close()

    def to_str(to_addr):
        if to_addr is None:
            return ""
        if isinstance(to_addr, list):
            return " ".join(str(r).lower() for r in to_addr)
        try:
            parsed = json.loads(str(to_addr))
            if isinstance(parsed, list):
                return " ".join(str(r).lower() for r in parsed)
        except Exception:
            pass
        return str(to_addr).lower()

    record("Письмо с анализом маршрутов отправлено", len(messages) >= 1,
           f"Найдено {len(messages)} писем")

    if not messages:
        record("Письмо отправлено на logistics@company.com", False, "нет писем")
        record("Email: тело содержит рекомендованный поезд 754А и его время", False, "нет писем")
        return

    # выбираем письмо адресованное logistics@company.com (или первое)
    target = None
    for m in messages:
        if "logistics@company.com" in to_str(m[1]):
            target = m
            break
    record("Письмо отправлено на logistics@company.com", target is not None,
           f"Адресаты: {[to_str(m[1])[:60] for m in messages]}")
    if target is None:
        target = messages[0]

    body = str(target[2] or "")
    body_norm = normalize_train(body)
    has_train = normalize_train(FASTEST) in body_norm
    has_time = FASTEST_DEPART in body or FASTEST_ARRIVE in body
    record("Email: тело содержит рекомендованный поезд 754А и его время",
           has_train and has_time,
           f"754А={has_train}, время({FASTEST_DEPART}/{FASTEST_ARRIVE})={has_time}")

    # эконом-цены в теле (некритично)
    has_price = str(ECONOM_PRICE) in body
    record("Email: тело содержит цену эконом-класса 5500", has_price,
           body[:150])


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"Critical fails: {CRITICAL_FAILS}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_fails": CRITICAL_FAILS,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"FAIL: критичные чеки провалены ({len(CRITICAL_FAILS)})")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
