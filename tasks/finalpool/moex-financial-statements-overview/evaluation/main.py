"""Evaluation для moex-financial-statements-overview (RU / moex-finance).

Проверяем YF_Financial_Statements.xlsx, который агент строит по данным MCP
`moex-finance` (схема moex.*). По реальному сиду доступна только годовая
income_stmt за 2 периода (2024, 2025) для 6 отслеживаемых тикеров; balance_sheet
и cashflow данных не возвращают, поэтому в итог попадают только income_stmt-строки.

CRITICAL_CHECKS (семантические): любой их провал => общий FAIL независимо от
accuracy. Структурные проверки (наличие листов, заголовки, сортировка) —
не критические.
"""
import argparse
import os
import sys

import openpyxl

# Отслеживаемые тикеры MOEX (источник данных moex-finance)
TRACKED_TICKERS = {"SBER.ME", "GAZP.ME", "LKOH.ME", "TCSG.ME", "MGNT.ME", "MTSS.ME"}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = False

# Имена критических проверок (семантика, не структура)
CRITICAL_CHECKS = {
    "Financial Overview: строки и Periods совпадают с эталоном",
    "Financial Overview: тикеры из отслеживаемого набора MOEX (.ME)",
    "Summary: Total_Records == число строк Financial Overview",
    "Summary: Unique_Symbols == число уникальных тикеров",
    "Summary: Statement_Types == число уникальных типов отчётности",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        tag = " (CRITICAL)" if name in CRITICAL_CHECKS else ""
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED = True


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def to_int(v):
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "YF_Financial_Statements.xlsx")
    gt_file = os.path.join(gt_dir, "YF_Financial_Statements.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: вывод агента не найден: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: эталон не найден: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---------------- Financial Overview ----------------
    a_rows = load_sheet_rows(agent_wb, "Financial Overview")
    g_rows = load_sheet_rows(gt_wb, "Financial Overview")

    # Структурная: лист существует
    record("Лист 'Financial Overview' присутствует", a_rows is not None)

    if a_rows is None or g_rows is None:
        # без листа дальнейшие проверки невозможны — критическая строковая проверка провалится
        record("Financial Overview: строки и Periods совпадают с эталоном", False,
               "лист отсутствует")
        record("Financial Overview: тикеры из отслеживаемого набора MOEX (.ME)", False,
               "лист отсутствует")
    else:
        a_header = [str(c).strip() if c is not None else "" for c in (a_rows[0] if a_rows else [])]
        # Структурная: заголовки
        record("Financial Overview: заголовки Symbol_Type/Periods",
               len(a_header) >= 2
               and a_header[0].lower() == "symbol_type"
               and a_header[1].lower() == "periods",
               f"header={a_header}")

        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}

        # CRITICAL: каждая эталонная строка присутствует и Periods совпадает (tol=1)
        row_errors = []
        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                row_errors.append(f"нет строки {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 1):
                row_errors.append(f"{g_row[0]}.Periods {a_row[1]} != {g_row[1]}")
        record("Financial Overview: строки и Periods совпадают с эталоном",
               not row_errors, "; ".join(row_errors))

        # CRITICAL: все Symbol_Type используют отслеживаемые тикеры MOEX (.ME)
        bad_syms = []
        for r in a_data:
            sym = str(r[0]).split(" - ")[0].strip()
            if sym not in TRACKED_TICKERS:
                bad_syms.append(sym)
        record("Financial Overview: тикеры из отслеживаемого набора MOEX (.ME)",
               len(a_data) > 0 and not bad_syms,
               f"посторонние тикеры: {sorted(set(bad_syms))}" if bad_syms else
               ("нет строк данных" if not a_data else ""))

        # Структурная: сортировка по Symbol_Type A-Z
        keys = [str(r[0]).strip().lower() for r in a_data]
        record("Financial Overview: строки отсортированы по Symbol_Type (A-Z)",
               keys == sorted(keys), f"первые: {keys[:4]}")

    # ---------------- Summary ----------------
    a_sum = load_sheet_rows(agent_wb, "Summary")
    record("Лист 'Summary' присутствует", a_sum is not None)

    if a_sum is None:
        record("Summary: Total_Records == число строк Financial Overview", False, "нет листа")
        record("Summary: Unique_Symbols == число уникальных тикеров", False, "нет листа")
        record("Summary: Statement_Types == число уникальных типов отчётности", False, "нет листа")
    else:
        s_header = [str(c).strip() if c is not None else "" for c in (a_sum[0] if a_sum else [])]
        record("Summary: заголовки Metric/Value",
               len(s_header) >= 2
               and s_header[0].lower() == "metric"
               and s_header[1].lower() == "value",
               f"header={s_header}")

        summary = {}
        for r in a_sum[1:]:
            if r and r[0] is not None:
                summary[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        # Истинные значения по фактическим строкам Financial Overview агента
        a_over = []
        if a_rows is not None:
            a_over = [r for r in a_rows[1:] if r and r[0] is not None]
        true_total = len(a_over)
        true_unique_syms = len({str(r[0]).split(" - ")[0].strip() for r in a_over})
        true_stmt_types = len({str(r[0]).split(" - ")[1].strip()
                               for r in a_over if " - " in str(r[0])})

        # CRITICAL: Total_Records == число строк Financial Overview
        tr = to_int(summary.get("total_records"))
        record("Summary: Total_Records == число строк Financial Overview",
               tr is not None and tr == true_total,
               f"{tr} != {true_total}")

        # CRITICAL: Unique_Symbols == число уникальных тикеров (исправленная семантика)
        us = to_int(summary.get("unique_symbols"))
        record("Summary: Unique_Symbols == число уникальных тикеров",
               us is not None and us == true_unique_syms,
               f"{us} != {true_unique_syms}")

        # CRITICAL: Statement_Types == число уникальных типов отчётности
        st = to_int(summary.get("statement_types"))
        record("Summary: Statement_Types == число уникальных типов отчётности",
               st is not None and st == true_stmt_types,
               f"{st} != {true_stmt_types}")

    # ---------------- Итог ----------------
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    if CRITICAL_FAILED:
        print("=== RESULT: FAIL (провалена критическая проверка) ===")
        sys.exit(1)

    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print("=== RESULT: FAIL (accuracy < 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
