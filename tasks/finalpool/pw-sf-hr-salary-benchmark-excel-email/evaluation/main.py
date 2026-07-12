"""Evaluation script for pw-sf-hr-salary-benchmark-excel-email (russified)."""
import os
import argparse, json, os, sys, re
import urllib.request
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

BENCHMARK_URL = "http://localhost:30301"

# Authoritative industry benchmark figures served on the mock page (RU dept labels,
# aligned with the central ClickHouse russification map). Used as fallback if the
# live page cannot be fetched during evaluation.
FALLBACK_BENCHMARK = {
    "Инженерия": 121500.0,
    "Финансы": 110000.0,
    "Кадры": 95000.0,
    "Операции": 98000.0,
    "НИОКР": 121500.0,
    "Продажи": 105000.0,
    "Поддержка": 78000.0,
}

PASS_COUNT = 0
FAIL_COUNT = 0

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def fetch_benchmark():
    """Fetch the served benchmark page and parse Department -> benchmark salary.
    Department labels on the page are the RU names that match the warehouse."""
    try:
        with urllib.request.urlopen(BENCHMARK_URL, timeout=10) as resp:
            html = resp.read().decode("utf-8", errors="replace")
    except Exception as e:
        print(f"  [WARN] could not fetch benchmark page ({e}); using fallback")
        return dict(FALLBACK_BENCHMARK)

    result = {}
    # rows look like: <td>Инженерия</td> <td>121500</td> <td>27.9</td> <td>P50-P75</td>
    rows = re.findall(r"<tr>(.*?)</tr>", html, re.DOTALL)
    for row in rows:
        cells = re.findall(r"<td>(.*?)</td>", row, re.DOTALL)
        if len(cells) >= 2:
            dept = cells[0].strip()
            sal = safe_float(cells[1])
            if dept and sal is not None:
                result[dept] = sal
    if not result:
        print("  [WARN] benchmark page parsed empty; using fallback")
        return dict(FALLBACK_BENCHMARK)
    return result


def get_warehouse_stats():
    """Per-department stats from ClickHouse warehouse (sf_data.HR_ANALYTICS).
    DEPARTMENT values are russified centrally so they align with the benchmark page."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               COUNT(*) AS emp_count,
               AVG("SALARY") AS avg_salary,
               AVG("YEARS_EXPERIENCE") AS avg_exp,
               AVG("PERFORMANCE_RATING") AS avg_perf
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    stats = {}
    for dept, cnt, sal, exp, perf in cur.fetchall():
        stats[dept] = {
            "count": int(cnt),
            "avg_salary": float(sal) if sal is not None else None,
            "avg_exp": float(exp) if exp is not None else None,
            "avg_perf": float(perf) if perf is not None else None,
        }
    cur.close()
    conn.close()
    return stats


def compute_expected():
    """Recompute the authoritative comparison from warehouse + served benchmark."""
    bench = fetch_benchmark()
    wh = get_warehouse_stats()
    # Only departments present in BOTH sources participate in the comparison.
    depts = sorted(set(bench) & set(wh))
    rows = {}
    for d in depts:
        our = round(wh[d]["avg_salary"], 2)
        ind = round(bench[d], 2)
        diff = round(our - ind, 2)
        diff_pct = round(diff / ind * 100, 1) if ind else 0.0
        status = "Above" if diff >= 0 else "Below"
        rows[d] = {
            "our_avg": our, "industry": ind, "diff": diff,
            "diff_pct": diff_pct, "status": status,
            "count": wh[d]["count"],
            "avg_exp": wh[d]["avg_exp"], "avg_perf": wh[d]["avg_perf"],
        }
    above = [d for d in depts if rows[d]["diff"] >= 0]
    below = [d for d in depts if rows[d]["diff"] < 0]
    highest = max(depts, key=lambda d: rows[d]["diff"]) if depts else None
    lowest = min(depts, key=lambda d: rows[d]["diff"]) if depts else None
    avg_diff = round(sum(rows[d]["diff"] for d in depts) / len(depts), 2) if depts else 0.0
    overall = "Competitive" if len(above) > len(below) else "Needs Attention"
    return {
        "rows": rows, "depts": depts,
        "above": above, "below": below,
        "highest": highest, "lowest": lowest,
        "avg_diff": avg_diff, "overall": overall,
        "total": len(depts),
    }


# ----- helpers to read the produced workbook -----

def read_sheet_rows(ws):
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    hmap = {h.lower(): i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(r)
    return headers, hmap, rows


def cell(row, hmap, col):
    idx = hmap.get(col.lower())
    if idx is None or idx >= len(row):
        return None
    return row[idx]


STATUS_ABOVE = {"above", "выше"}
STATUS_BELOW = {"below", "ниже"}


def run_critical_checks(agent_workspace):
    """SEMANTIC gate. Any failure here => hard FAIL (sys.exit(1)).
    Recomputes correct values from ClickHouse + served benchmark page and
    verifies the produced workbook/email against them."""
    print("=== CRITICAL CHECKS ===")
    crit_fail = 0

    def crit(name, cond, detail=""):
        nonlocal crit_fail
        if cond:
            print(f"  [CRIT-PASS] {name}")
        else:
            crit_fail += 1
            print(f"  [CRIT-FAIL] {name}: {str(detail)[:300]}")

    excel_path = os.path.join(agent_workspace, "Salary_Benchmark_Report.xlsx")
    if not os.path.exists(excel_path):
        crit("Salary_Benchmark_Report.xlsx exists", False, excel_path)
        return False

    try:
        exp = compute_expected()
    except Exception as e:
        crit("recompute from warehouse+benchmark", False, str(e))
        return False

    if not exp["depts"]:
        crit("warehouse/benchmark departments overlap", False,
             "no common departments between ClickHouse and served page (JOIN-KEY MISMATCH)")
        return False

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # --- Compensation_Comparison: per-department Difference / Difference_Pct / Status ---
    if "Compensation_Comparison" not in wb.sheetnames:
        crit("Compensation_Comparison sheet exists", False)
        return False
    ws = wb["Compensation_Comparison"]
    _, hmap, rows = read_sheet_rows(ws)
    produced = {}
    for r in rows:
        dn = cell(r, hmap, "Department")
        if dn is None:
            continue
        produced[str(dn).strip()] = r

    diff_ok = True
    pct_ok = True
    status_ok = True
    diff_detail = []
    for d, e in exp["rows"].items():
        r = produced.get(d)
        if r is None:
            diff_ok = False
            diff_detail.append(f"{d}: missing row")
            continue
        pdiff = safe_float(cell(r, hmap, "Difference"))
        ppct = safe_float(cell(r, hmap, "Difference_Pct"))
        pstat = str(cell(r, hmap, "Status") or "").strip().lower()
        if pdiff is None or abs(pdiff - e["diff"]) > 1.0:
            diff_ok = False
            diff_detail.append(f"{d}: Difference got {pdiff} exp {e['diff']}")
        if ppct is None or abs(ppct - e["diff_pct"]) > 0.5:
            pct_ok = False
            diff_detail.append(f"{d}: Difference_Pct got {ppct} exp {e['diff_pct']}")
        exp_set = STATUS_ABOVE if e["status"] == "Above" else STATUS_BELOW
        if pstat not in exp_set:
            status_ok = False
            diff_detail.append(f"{d}: Status got '{pstat}' exp {e['status']}")

    crit("Compensation_Comparison Difference matches warehouse-vs-benchmark for every dept",
         diff_ok, "; ".join(diff_detail[:6]))
    crit("Compensation_Comparison Difference_Pct correct for every dept",
         pct_ok, "; ".join(diff_detail[:6]))
    crit("Status column derived correctly (Above/Выше if Difference>=0 else Below/Ниже)",
         status_ok, "; ".join(diff_detail[:6]))

    # --- Executive_Summary: highest/lowest gap, counts, overall status ---
    if "Executive_Summary" not in wb.sheetnames:
        crit("Executive_Summary sheet exists", False)
    else:
        es = wb["Executive_Summary"]
        summary = {}
        for r in es.iter_rows(min_row=2, values_only=True):
            if r and r[0] is not None:
                key = str(r[0]).strip()
                val = r[1] if len(r) > 1 else None
                summary[key.lower()] = val

        def sval(name):
            return summary.get(name.lower())

        total = safe_float(sval("Total_Departments"))
        above_n = safe_float(sval("Departments_Above_Benchmark"))
        below_n = safe_float(sval("Departments_Below_Benchmark"))
        highest = str(sval("Highest_Gap_Department") or "").strip()
        lowest = str(sval("Lowest_Gap_Department") or "").strip()
        avg_diff = safe_float(sval("Average_Difference"))
        overall = str(sval("Overall_Status") or "").strip().lower()

        crit("Highest_Gap_Department = dept with max recomputed Difference",
             highest == exp["highest"], f"got '{highest}' exp '{exp['highest']}'")
        crit("Lowest_Gap_Department = dept with min recomputed Difference",
             lowest == exp["lowest"], f"got '{lowest}' exp '{exp['lowest']}'")
        crit("Above + Below counts sum to Total_Departments and match recompute",
             (above_n == len(exp["above"]) and below_n == len(exp["below"])
              and (total is not None and int(total) == exp["total"])
              and len(exp["above"]) + len(exp["below"]) == exp["total"]),
             f"above {above_n}/{len(exp['above'])} below {below_n}/{len(exp['below'])} total {total}/{exp['total']}")
        crit("Average_Difference matches recomputed mean",
             avg_diff is not None and abs(avg_diff - exp["avg_diff"]) <= 1.0,
             f"got {avg_diff} exp {exp['avg_diff']}")
        if exp["overall"] == "Competitive":
            overall_ok = overall in {"competitive", "конкурентоспособно"}
        else:
            overall_ok = overall in {"needs attention", "требует внимания"}
        crit("Overall_Status = Competitive/Конкурентоспособно iff more depts Above than Below",
             overall_ok, f"got '{overall}' exp '{exp['overall']}'")

    # --- Email: correct above/below counts + biggest-gap dept, RU+EN keywords ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages WHERE subject ILIKE %s",
                    ('%benchmark%',))
        emails = cur.fetchall()
        conn.close()
    except Exception as e:
        emails = []
        crit("email lookup", False, str(e))

    email_ok = False
    edetail = ""
    for subj, to_addr, body in emails:
        if "hr-director@company.com" not in str(to_addr).lower():
            continue
        b = (body or "").lower()
        # dept keyword RU+EN
        has_dept_kw = ("отдел" in b) or ("department" in b)
        # above/below counts present
        has_above = str(len(exp["above"])) in b
        has_below = str(len(exp["below"])) in b
        # biggest-gap department named
        big = (exp["highest"] or "").lower()
        has_big = big and big in b
        if has_dept_kw and has_above and has_below and has_big:
            email_ok = True
            break
        edetail = (f"dept_kw={has_dept_kw} above({len(exp['above'])})={has_above} "
                   f"below({len(exp['below'])})={has_below} biggest('{exp['highest']}')={has_big}")
    crit("Benchmark email to hr-director states correct above/below counts and biggest-gap dept "
         "(RU+EN keywords accepted)", email_ok, edetail or "no matching email")

    # --- salary_comparison.json consistent with recomputed merge ---
    cmp_path = os.path.join(agent_workspace, "salary_comparison.json")
    json_ok = False
    jdetail = ""
    if os.path.exists(cmp_path):
        try:
            with open(cmp_path, encoding="utf-8") as f:
                raw = f.read()
            data = json.loads(raw)
            # find numeric salary values for each department somewhere in the structure
            flat = json.dumps(data, ensure_ascii=False)
            hits = 0
            for d, e in exp["rows"].items():
                # require the department name AND its benchmark or our-avg value to appear
                if d in flat and (str(int(e["industry"])) in flat or
                                  any(abs(safe_float(x, 0) - e["our_avg"]) <= 1.0
                                      for x in re.findall(r"-?\d+\.?\d*", flat))):
                    hits += 1
            json_ok = hits >= max(1, len(exp["rows"]) - 1)
            jdetail = f"matched {hits}/{len(exp['rows'])} departments"
        except Exception as e:
            jdetail = str(e)
    else:
        jdetail = "salary_comparison.json missing"
    crit("salary_comparison.json contains values consistent with recomputed merge",
         json_ok, jdetail)

    print(f"=== CRITICAL: {'PASS' if crit_fail == 0 else f'{crit_fail} FAILED'} ===")
    return crit_fail == 0


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    excel_path = os.path.join(agent_workspace, "Salary_Benchmark_Report.xlsx")
    check("Salary_Benchmark_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Compensation_Comparison sheet exists", "Compensation_Comparison" in wb.sheetnames)
        if "Compensation_Comparison" in wb.sheetnames:
            ws = wb["Compensation_Comparison"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Compensation_Comparison has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Employee_Count', 'Our_Avg_Salary', 'Industry_Benchmark', 'Difference']:
                check(f"Compensation_Comparison has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Department_Details sheet exists", "Department_Details" in wb.sheetnames)
        if "Department_Details" in wb.sheetnames:
            ws = wb["Department_Details"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Department_Details has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Department', 'Avg_Experience', 'Avg_Performance']:
                check(f"Department_Details has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Executive_Summary sheet exists", "Executive_Summary" in wb.sheetnames)
        if "Executive_Summary" in wb.sheetnames:
            ws = wb["Executive_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Executive_Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Executive_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # Check email was sent (structural; RU+EN body keyword)
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT subject, to_addr, body_text FROM email.messages WHERE subject ILIKE %s", ('%benchmark%',))
            emails = cur.fetchall()
            check("Benchmark email sent", len(emails) >= 1, f"found {len(emails)} matching emails")
            if emails:
                check("Email to hr-director", "hr-director@company.com" in str(emails[0][1]).lower())
                body = str(emails[0][2]).lower() if emails[0][2] else ""
                check("Email mentions departments (RU/EN)",
                      ("department" in body) or ("отдел" in body), body[:120])
            conn.close()
        except Exception as e:
            check("Email verification", False, str(e))

        # Check terminal artifacts
        check("salary_processor.py exists", os.path.exists(os.path.join(agent_workspace, "salary_processor.py")))
        check("benchmark_raw.json exists", os.path.exists(os.path.join(agent_workspace, "benchmark_raw.json")))
        check("salary_comparison.json exists", os.path.exists(os.path.join(agent_workspace, "salary_comparison.json")))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    return accuracy >= 70, f"Passed {PASS_COUNT}/{total} checks (accuracy {accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # CRITICAL semantic gate first: any failure => hard FAIL regardless of accuracy.
    critical_ok = run_critical_checks(args.agent_workspace)
    if not critical_ok:
        print("CRITICAL checks failed -> FAIL")
        sys.exit(1)

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
