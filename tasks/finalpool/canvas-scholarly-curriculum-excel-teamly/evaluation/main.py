"""Evaluation script for canvas-scholarly-curriculum-excel-teamly.

All expected values are recomputed LIVE from the source systems:
  - canvas.courses / canvas.enrollments  (courses, enrollment, avg score)
  - scholarly.scholar_papers             (relevant papers, citations, years)
  - teamly.pages                         (the Curriculum Innovation Tracker page)
Nothing is compared against a hardcoded/fabricated groundtruth xlsx, so the
eval stays honest if the seeds change.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Otherwise PASS requires accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Relevant (on-topic) seeded papers the agent must surface in Research_Trends.
# Substrings of the titles/topics that the eval matches; kept in sync with
# preprocess RELEVANT_PAPERS.
RELEVANT_PAPER_TITLES = [
    "deep learning advances",
    "scalable data analytics",
    "computational thinking",
]

# Relevance enum, English + Russian equivalents accepted.
RELEVANCE_VALUES = {
    "high", "medium", "low",
    "высокая", "высокий", "средняя", "средний", "низкая", "низкий",
}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Teamly Curriculum Innovation Tracker page exists (not the noise page)",
    "Teamly page has the alignment-review heading + >=3 content paragraphs",
    "Current_Courses matches live canvas courses (name+code, correct count)",
    "Gap_Analysis Total_Courses/Papers_Reviewed correct and internally consistent",
    "curriculum_reviewer.py + curriculum_review.json + input JSONs present",
    "Research_Trends derives from seeded papers, sorted by Citations desc, valid Relevance",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:300] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_sheet(wb, name):
    target = name.strip().lower().replace(" ", "_")
    for n in wb.sheetnames:
        if n.strip().lower().replace(" ", "_") == target:
            return wb[n]
    return None


def header_map(ws):
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def data_rows(ws):
    return [r for r in ws.iter_rows(min_row=2, values_only=True)
            if any(c is not None for c in r)]


# ---------------------------------------------------------------------------
# Live expectations from source systems
# ---------------------------------------------------------------------------

def get_expected_courses():
    """Live canvas courses with name, code and enrollment count."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, name, course_code FROM canvas.courses ORDER BY name")
    courses = []
    for cid, name, code in cur.fetchall():
        cur.execute(
            """SELECT COUNT(*) FROM canvas.enrollments
               WHERE course_id = %s AND type = 'StudentEnrollment'""",
            (cid,))
        enroll = cur.fetchone()[0]
        courses.append({"id": cid, "name": name, "code": code, "enroll": enroll})
    conn.close()
    return courses


def get_relevant_paper_count():
    """Count of on-topic seeded scholar papers (ML / data analytics / comp thinking)."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT title FROM scholarly.scholar_papers")
    titles = [str(t[0]).lower() for t in cur.fetchall()]
    conn.close()
    count = 0
    for sub in RELEVANT_PAPER_TITLES:
        if any(sub in t for t in titles):
            count += 1
    return count


# ---------------------------------------------------------------------------
# Excel checks
# ---------------------------------------------------------------------------

def check_current_courses(wb, courses):
    """Critical: every live canvas course present by name+code, correct count."""
    ws = get_sheet(wb, "Current_Courses")
    if ws is None:
        check("Current_Courses matches live canvas courses (name+code, correct count)",
              False, "no sheet")
        return
    headers, hmap = header_map(ws)
    # Structural (non-critical) header presence.
    for col in ["course_name", "course_code", "enrollment_count", "avg_score"]:
        check(f"Current_Courses has {col} column", col in hmap, f"headers: {headers}")
    rows = data_rows(ws)

    name_i = hmap.get("course_name")
    code_i = hmap.get("course_code")
    avg_i = hmap.get("avg_score")

    # Each live course present by name and code.
    present = 0
    if name_i is not None and code_i is not None:
        joined = []
        for r in rows:
            nm = str(r[name_i]).strip().lower() if name_i < len(r) and r[name_i] is not None else ""
            cd = str(r[code_i]).strip().lower() if code_i < len(r) and r[code_i] is not None else ""
            joined.append((nm, cd))
        for c in courses:
            cn = c["name"].strip().lower()
            cc = (c["code"] or "").strip().lower()
            if any((cn in nm or nm in cn) and (not cc or cc in cd or cd in cc) for nm, cd in joined):
                present += 1
    n = len(courses)
    check("Current_Courses matches live canvas courses (name+code, correct count)",
          n > 0 and present >= n and len(rows) == n,
          f"matched {present}/{n}, rows={len(rows)}")

    # Sorted by Course_Name ascending (non-critical).
    if name_i is not None:
        names = [str(r[name_i]).strip().lower() for r in rows
                 if name_i < len(r) and r[name_i] is not None]
        sorted_ok = names == sorted(names)
        check("Current_Courses sorted by Course_Name ascending", sorted_ok, f"{names[:5]}")

    # Avg_Score rounded to 1 decimal and within range (non-critical, lenient).
    if avg_i is not None:
        scores = [safe_float(r[avg_i]) for r in rows if avg_i < len(r)]
        scores = [s for s in scores if s is not None]
        range_ok = all(0 <= s <= 100 for s in scores) and len(scores) == n
        check("Current_Courses Avg_Score values are 0-100 for every course", range_ok,
              f"scores: {scores[:8]}")


def check_research_trends(wb):
    """Critical: rows derive from seeded papers, sorted by Citations desc, valid Relevance."""
    ws = get_sheet(wb, "Research_Trends")
    if ws is None:
        check("Research_Trends derives from seeded papers, sorted by Citations desc, valid Relevance",
              False, "no sheet")
        return
    headers, hmap = header_map(ws)
    for col in ["paper_title", "topic_area", "year", "citations", "relevance_to_curriculum"]:
        check(f"Research_Trends has {col} column", col in hmap, f"headers: {headers}")
    rows = data_rows(ws)

    title_i = hmap.get("paper_title")
    cit_i = hmap.get("citations")
    rel_i = hmap.get("relevance_to_curriculum")

    # Titles derive from the seeded on-topic papers.
    derived = 0
    if title_i is not None:
        titles = [str(r[title_i]).strip().lower() for r in rows
                  if title_i < len(r) and r[title_i] is not None]
        for sub in RELEVANT_PAPER_TITLES:
            if any(sub in t for t in titles):
                derived += 1

    # Sorted by Citations descending.
    cits = []
    if cit_i is not None:
        cits = [safe_float(r[cit_i]) for r in rows if cit_i < len(r)]
        cits = [c for c in cits if c is not None]
    sorted_ok = all(cits[i] >= cits[i + 1] for i in range(len(cits) - 1)) and len(cits) >= 3

    # Relevance values valid (RU/EN).
    rel_ok = False
    if rel_i is not None:
        vals = [str(r[rel_i]).strip().lower() for r in rows
                if rel_i < len(r) and r[rel_i] is not None]
        rel_ok = len(vals) >= 3 and all(v in RELEVANCE_VALUES for v in vals)

    check("Research_Trends derives from seeded papers, sorted by Citations desc, valid Relevance",
          derived >= 3 and sorted_ok and rel_ok,
          f"derived={derived}/3, cits={cits}, sorted={sorted_ok}, relevance_ok={rel_ok}")


def check_gap_analysis(wb, n_courses, n_relevant):
    """Critical: Total_Courses/Papers_Reviewed correct, metrics internally consistent."""
    ws = get_sheet(wb, "Gap_Analysis")
    if ws is None:
        check("Gap_Analysis Total_Courses/Papers_Reviewed correct and internally consistent",
              False, "no sheet")
        return
    headers, hmap = header_map(ws)
    for col in ["metric", "value"]:
        check(f"Gap_Analysis has {col} column", col in hmap, f"headers: {headers}")
    rows = data_rows(ws)
    metric_i = hmap.get("metric")
    value_i = hmap.get("value")

    metrics = {}
    if metric_i is not None and value_i is not None:
        for r in rows:
            if metric_i < len(r) and r[metric_i] is not None:
                key = str(r[metric_i]).strip().lower()
                val = r[value_i] if value_i < len(r) else None
                metrics[key] = val

    total_courses = safe_float(metrics.get("total_courses"))
    papers_reviewed = safe_float(metrics.get("papers_reviewed"))
    high_rel = safe_float(metrics.get("high_relevance_papers"))
    coverage = safe_float(metrics.get("curriculum_coverage_pct"))
    top_gap = metrics.get("top_gap_area")

    # Total_Courses == live canvas course count.
    tc_ok = total_courses is not None and int(total_courses) == n_courses
    # Papers_Reviewed >= number of seeded relevant papers (agent may include noise).
    pr_ok = papers_reviewed is not None and papers_reviewed >= n_relevant
    # Coverage numeric in 0-100.
    cov_ok = coverage is not None and 0 <= coverage <= 100
    # High_Relevance_Papers <= Papers_Reviewed.
    hr_ok = (high_rel is not None and papers_reviewed is not None
             and high_rel <= papers_reviewed)
    # Top_Gap_Area present.
    gap_ok = top_gap is not None and str(top_gap).strip() != ""

    check("Gap_Analysis Total_Courses/Papers_Reviewed correct and internally consistent",
          tc_ok and pr_ok and cov_ok and hr_ok,
          f"Total_Courses={total_courses} (exp {n_courses}), Papers_Reviewed={papers_reviewed} "
          f"(>= {n_relevant}), Coverage={coverage}, High<=Reviewed={hr_ok}")
    check("Gap_Analysis includes a Top_Gap_Area value", gap_ok, f"top_gap={top_gap}")


def check_excel(agent_workspace, courses, n_relevant):
    excel_path = os.path.join(agent_workspace, "Curriculum_Review_Report.xlsx")
    exists = os.path.exists(excel_path)
    check("Curriculum_Review_Report.xlsx exists", exists)
    if not exists:
        check("Current_Courses matches live canvas courses (name+code, correct count)", False, "no excel")
        check("Research_Trends derives from seeded papers, sorted by Citations desc, valid Relevance", False, "no excel")
        check("Gap_Analysis Total_Courses/Papers_Reviewed correct and internally consistent", False, "no excel")
        return
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for sn in ["Current_Courses", "Research_Trends", "Gap_Analysis"]:
        check(f"{sn} sheet exists", get_sheet(wb, sn) is not None)
    check_current_courses(wb, courses)
    check_research_trends(wb)
    check_gap_analysis(wb, len(courses), n_relevant)


# ---------------------------------------------------------------------------
# Terminal-step deliverables
# ---------------------------------------------------------------------------

def check_terminal(agent_workspace):
    """Critical: the reviewer script + its output + input JSONs are present."""
    files = set(os.listdir(agent_workspace)) if os.path.isdir(agent_workspace) else set()
    reviewer = "curriculum_reviewer.py" in files
    review_out = "curriculum_review.json" in files
    course_json = "course_data.json" in files
    papers_json = "research_papers.json" in files
    # curriculum_review.json must be valid JSON if present.
    review_valid = False
    if review_out:
        try:
            with open(os.path.join(agent_workspace, "curriculum_review.json"), "r",
                      encoding="utf-8") as f:
                json.load(f)
            review_valid = True
        except Exception:
            review_valid = False
    check("curriculum_reviewer.py + curriculum_review.json + input JSONs present",
          reviewer and review_out and review_valid and course_json and papers_json,
          f"reviewer={reviewer}, review_json={review_out}(valid={review_valid}), "
          f"course_data={course_json}, research_papers={papers_json}")


# ---------------------------------------------------------------------------
# Teamly checks
# ---------------------------------------------------------------------------

def check_teamly():
    """Critical: Curriculum Innovation Tracker page (not the noise page) with the
    alignment-review heading and >=3 content paragraphs/recommendations.

    Seed pages have id <= 3; the noise page ('Архив протоколов совещаний...') is
    a leftover and must NOT satisfy the deliverable check.
    """
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly Curriculum Innovation Tracker page exists (not the noise page)", False, str(e))
        check("Teamly page has the alignment-review heading + >=3 content paragraphs", False, str(e))
        return

    tracker = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "архив протоколов" in tl:
            continue
        if ("curriculum innovation tracker" in tl
                or ("curriculum" in tl and ("tracker" in tl or "innovation" in tl))
                or ("учебн" in tl and ("трекер" in tl or "иннова" in tl))):
            tracker = (pid, title, body)
            break
    check("Teamly Curriculum Innovation Tracker page exists (not the noise page)",
          tracker is not None, f"new pages: {[(p[0], p[1]) for p in pages]}")

    if tracker is None:
        check("Teamly page has the alignment-review heading + >=3 content paragraphs", False, "no page")
        return

    text = ((tracker[1] or "") + "\n" + (tracker[2] or ""))
    text_l = text.lower()
    # Heading marker, RU or EN (preserved English literal from task.md).
    heading_ok = ("research-curriculum alignment review" in text_l
                  or ("alignment" in text_l and "review" in text_l)
                  or ("соответств" in text_l and "учебн" in text_l)
                  or "обзор соответствия" in text_l)
    # >=3 content paragraphs/lines (findings, top-3 trends, recommendations).
    paras = [ln for ln in tracker[2].splitlines() if len(ln.strip()) >= 20]
    body_ok = len(tracker[2]) >= 200 and len(paras) >= 3
    check("Teamly page has the alignment-review heading + >=3 content paragraphs",
          heading_ok and body_ok,
          f"heading={heading_ok}, body_len={len(tracker[2])}, paragraphs={len(paras)}")
    # Recommendation / trends keywords (non-critical).
    check("Teamly page mentions trends and recommendations",
          any(k in text_l for k in ["recommend", "рекоменд", "trend", "тренд", "направлен"]),
          "no trend/recommendation keywords")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    try:
        courses = get_expected_courses()
        n_relevant = get_relevant_paper_count()
    except Exception as e:
        check("Live source systems reachable", False, str(e))
        courses, n_relevant = [], 0

    check_excel(agent_workspace, courses, n_relevant)
    check_terminal(agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
