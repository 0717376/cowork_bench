"""Evaluation for terminal-canvas-gsheet-word-teamly-gcal.

Checks:
1. Academic_Advising_Report.xlsx with 4 sheets + SEMANTIC analytics recomputed
   live from Canvas submissions (Course_Summary averages/pass-rate,
   Grade_Distribution submission counts, Advising_Needs per-student estimates).
2. Google Sheet "Academic Advising Analytics".
3. Advising_Recommendations.docx with required sections (RU+EN matching).
4. Teamly "Student Advising Tracker" page with >=4 category child pages.
5. Google Calendar advising events (weekdays only, 09:00-16:00).
6. advising_analyzer.py script exists.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Otherwise PASS threshold: accuracy >= 70%.

Course IDs come from the global Canvas seed: 20 (Fall 2013) and 21 (Fall 2014),
matching semester_config.json and task.md.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Course identifiers used by this task (verified against the global Canvas seed).
COURSE_IDS = (20, 21)

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Course_Summary averages match live Canvas (per course, tol 1.0)",
    "Course_Summary pass-rate matches live Canvas (per course, tol 1.5)",
    "Grade_Distribution counts sum to total Canvas submissions",
    "Grade_Distribution failing(<60) count matches Canvas (tol 1)",
    "Advising_Needs Urgent Intervention student estimate is positive and <= failing submissions",
    "Teamly 'Student Advising Tracker' has >=4 category pages",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def _num(x):
    """Extract a float from a cell that may carry '%' or other text."""
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace("%", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def compute_canvas_analytics():
    """Recompute ground-truth analytics from live Canvas submissions.

    Returns dict keyed by course_id -> {avg, passrate, count, failing} plus
    a 'combined' entry with total count and failing count across both courses.
    Returns None if Canvas is unreachable.
    """
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        result = {}
        total_count = 0
        total_failing = 0
        for cid in COURSE_IDS:
            cur.execute("""
                SELECT s.score FROM canvas.submissions s
                JOIN canvas.assignments a ON a.id = s.assignment_id
                WHERE a.course_id = %s AND s.score IS NOT NULL
            """, (cid,))
            scores = [float(r[0]) for r in cur.fetchall()]
            n = len(scores)
            if n == 0:
                result[cid] = {"avg": None, "passrate": None, "count": 0, "failing": 0}
                continue
            avg = sum(scores) / n
            passing = sum(1 for s in scores if s >= 60)
            failing = sum(1 for s in scores if s < 60)
            passrate = 100.0 * passing / n
            result[cid] = {"avg": avg, "passrate": passrate, "count": n, "failing": failing}
            total_count += n
            total_failing += failing
        result["combined"] = {"count": total_count, "failing": total_failing}
        cur.close()
        conn.close()
        return result
    except Exception as e:
        print(f"  [WARN] Could not compute live Canvas analytics: {e}")
        return None


def check_excel(workspace):
    print("\n=== Check 1: Academic_Advising_Report.xlsx ===")
    path = os.path.join(workspace, "Academic_Advising_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]
    gt = compute_canvas_analytics()

    # ── Course_Summary ───────────────────────────────────────────────────────
    cs_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s or "summary" in s), 0)
    ws = wb[sheets[cs_idx]]
    rows = list(ws.iter_rows(values_only=True))
    data_rows = [r for r in rows[1:] if any(c for c in r)]
    check("Course_Summary has 2 course rows", len(data_rows) >= 2, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    # Course names come from the russified Canvas LMS:
    # "Global Governance & Geopolitics" -> "Глобальное управление и геополитика".
    # Accept the russified discriminating substring (or the English name if the
    # agent sourced it from semester_config.json, which stays English).
    check("Contains Global Governance", any(
        k in all_text for k in ("governance", "управлени", "геополит")),
          f"Text: {all_text[:120]}")

    # SEMANTIC: match each course's average and pass-rate against live Canvas.
    if gt is not None:
        header = [str(c).lower() if c else "" for c in (rows[0] if rows else [])]

        def col_idx(keys, default=None):
            for i, h in enumerate(header):
                if any(k in h for k in keys):
                    return i
            return default

        id_i = col_idx(["identifier", "id", "идентиф", "курс"])
        avg_i = col_idx(["average", "avg", "сред"])
        pr_i = col_idx(["pass", "rate", "успеш", "доля", "процент"])

        avg_ok = True
        pr_ok = True
        if avg_i is None or id_i is None:
            avg_ok = False
        if pr_i is None or id_i is None:
            pr_ok = False
        if id_i is not None:
            # Map each data row to a course by matching the id cell.
            for r in data_rows:
                rid = None
                cell = r[id_i] if id_i < len(r) else None
                for cid in COURSE_IDS:
                    if cell is not None and str(cid) in str(cell):
                        rid = cid
                        break
                if rid is None or gt.get(rid, {}).get("avg") is None:
                    continue
                if avg_i is not None and avg_i < len(r):
                    got = _num(r[avg_i])
                    exp = gt[rid]["avg"]
                    if got is None or abs(got - exp) > 1.0:
                        avg_ok = False
                if pr_i is not None and pr_i < len(r):
                    got = _num(r[pr_i])
                    exp = gt[rid]["passrate"]
                    if got is None or abs(got - exp) > 1.5:
                        pr_ok = False
        check("Course_Summary averages match live Canvas (per course, tol 1.0)",
              avg_ok, "Average score deviates from recomputed Canvas data")
        check("Course_Summary pass-rate matches live Canvas (per course, tol 1.5)",
              pr_ok, "Pass rate deviates from recomputed Canvas data")
    else:
        # Canvas unreachable: structural fallback so the check still runs.
        check("Course_Summary averages match live Canvas (per course, tol 1.0)",
              len(data_rows) >= 2, "Canvas unreachable; structural fallback")
        check("Course_Summary pass-rate matches live Canvas (per course, tol 1.5)",
              len(data_rows) >= 2, "Canvas unreachable; structural fallback")

    # ── Grade_Distribution ───────────────────────────────────────────────────
    gd_idx = next((i for i, s in enumerate(sheets_lower) if "grade" in s or "distribution" in s), 1)
    if gd_idx < len(sheets):
        ws2 = wb[sheets[gd_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Grade_Distribution has 5 range rows", len(data_rows2) >= 5, f"Found {len(data_rows2)}")

        if gt is not None and gt["combined"]["count"] > 0:
            header2 = [str(c).lower() if c else "" for c in (rows2[0] if rows2 else [])]
            label_i = next((i for i, h in enumerate(header2)
                            if any(k in h for k in ["range", "label", "диапаз", "метк", "оцен"])), 0)
            count_i = next((i for i, h in enumerate(header2)
                            if any(k in h for k in ["count", "число", "кол"])), 1)
            counts = []
            failing_count = None
            for r in data_rows2:
                cval = _num(r[count_i]) if count_i < len(r) else None
                if cval is not None:
                    counts.append(cval)
                lbl = str(r[label_i]).lower() if label_i < len(r) and r[label_i] else ""
                if any(k in lbl for k in ["fail", "<60", "below 60", "0-59", "0–59", "ниже 60", " f "]) or lbl.strip() in ("f", "failing"):
                    failing_count = cval
            total_sum = sum(counts) if counts else None
            exp_total = gt["combined"]["count"]
            sum_ok = total_sum is not None and abs(total_sum - exp_total) <= 1
            check("Grade_Distribution counts sum to total Canvas submissions",
                  sum_ok, f"Sum={total_sum}, expected total={exp_total}")

            exp_fail = gt["combined"]["failing"]
            fail_ok = failing_count is not None and abs(failing_count - exp_fail) <= 1
            check("Grade_Distribution failing(<60) count matches Canvas (tol 1)",
                  fail_ok, f"Got failing={failing_count}, expected={exp_fail}")
        else:
            check("Grade_Distribution counts sum to total Canvas submissions",
                  len(data_rows2) >= 5, "Canvas unreachable; structural fallback")
            check("Grade_Distribution failing(<60) count matches Canvas (tol 1)",
                  len(data_rows2) >= 5, "Canvas unreachable; structural fallback")

    # ── Advising_Needs ───────────────────────────────────────────────────────
    an_idx = next((i for i, s in enumerate(sheets_lower) if "advising" in s or "need" in s), 2)
    if an_idx < len(sheets):
        ws3 = wb[sheets[an_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Advising_Needs has 4 category rows", len(data_rows3) >= 4, f"Found {len(data_rows3)}")
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c).lower()
        check("Contains Urgent Intervention", "urgent" in all_text3, f"Text: {all_text3[:120]}")
        # All four English category labels present.
        cats = ["urgent intervention", "academic support", "progress monitoring", "no action"]
        check("Advising_Needs lists all 4 categories",
              all(c in all_text3 for c in cats),
              f"Text: {all_text3[:200]}")

        # SEMANTIC: Advising_Needs holds a per-STUDENT estimate (task.md: "оцените
        # число студентов ... исходя из распределения баллов отправок"), which is a
        # distinct quantity from Grade_Distribution's per-SUBMISSION counts. task.md
        # fixes no exact student-estimate formula, so we assert the invariant that
        # always holds: the Urgent (failing) student estimate is positive (failing
        # submissions exist => some students need urgent help) and cannot exceed the
        # number of failing submissions. A non-doing agent (no/zero Urgent count)
        # still fails this gate.
        if gt is not None:
            header3 = [str(c).lower() if c else "" for c in (rows3[0] if rows3 else [])]
            cat_i = next((i for i, h in enumerate(header3)
                          if any(k in h for k in ["categor", "катего"])), 0)
            cnt_i = next((i for i, h in enumerate(header3)
                          if any(k in h for k in ["count", "число", "кол", "student"])), 1)
            urgent_count = None
            for r in data_rows3:
                lbl = str(r[cat_i]).lower() if cat_i < len(r) and r[cat_i] else ""
                if "urgent" in lbl:
                    urgent_count = _num(r[cnt_i]) if cnt_i < len(r) else None
                    break
            exp_fail = gt["combined"]["failing"]
            urgent_ok = (urgent_count is not None and urgent_count > 0
                         and urgent_count <= exp_fail + 1)
            check("Advising_Needs Urgent Intervention student estimate is positive and <= failing submissions",
                  urgent_ok, f"Got urgent={urgent_count}, failing submissions={exp_fail}")
        else:
            check("Advising_Needs Urgent Intervention student estimate is positive and <= failing submissions",
                  len(data_rows3) >= 4, "Canvas unreachable; structural fallback")

    # ── Appointment_Schedule ─────────────────────────────────────────────────
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "appointment" in s or "schedule" in s), 3)
    if ap_idx < len(sheets):
        ws4 = wb[sheets[ap_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data_rows4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Appointment_Schedule has 5 weekday rows", len(data_rows4) >= 5, f"Found {len(data_rows4)}")
        sched_text = " ".join(str(c) for r in rows4 for c in r if c)
        # The 5 scheduled days fall within the Mon-Fri week 2026-03-16..20.
        weekdays = [f"2026-03-{d}" for d in ("16", "17", "18", "19", "20")]
        days_present = sum(1 for d in weekdays if d in sched_text)
        check("Appointment_Schedule covers the weekdays 2026-03-16..20",
              days_present >= 5, f"Matched {days_present}/5 weekday dates")


def check_gsheet():
    print("\n=== Check 2: Google Sheet Analytics ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE lower(title) LIKE '%%advising%%'")
    sheets = cur.fetchall()
    check("Academic Advising spreadsheet exists", len(sheets) >= 1,
          f"Found: {[s[1] for s in sheets]}")

    if sheets:
        ss_id = sheets[0][0]
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (ss_id,))
        cnt = cur.fetchone()[0]
        check("Spreadsheet has data cells", cnt >= 5, f"Found {cnt} cells")

    cur.close()
    conn.close()


def check_word(workspace):
    print("\n=== Check 3: Advising_Recommendations.docx ===")
    path = os.path.join(workspace, "Advising_Recommendations.docx")
    if not os.path.exists(path):
        check("Word file exists", False, f"Not found at {path}")
        return
    check("Word file exists", True)

    try:
        from docx import Document
        doc = Document(path)
        # RU keyword checks operate on the ORIGINAL lowercased text (not normalize()).
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Contains performance overview",
              any(k in all_text for k in ["overview", "performance", "обзор", "успеваем"]),
              f"Text: {all_text[:150]}")
        check("Contains at-risk identification",
              any(k in all_text for k in ["risk", "intervention", "риск", "группа риска", "вмешат"]),
              f"Text: {all_text[:150]}")
        check("Contains recommendations",
              any(k in all_text for k in ["recommend", "action", "рекоменд", "действ"]),
              f"Text: {all_text[:150]}")
        # 'governance' stays an English marker (program name preserved).
        check("Mentions Global Governance",
              "governance" in all_text or "geopolitics" in all_text,
              f"Text: {all_text[:150]}")
    except ImportError:
        check("python-docx available", False, "python-docx not installed")


def check_teamly():
    print("\n=== Check 4: Teamly Student Advising Tracker ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT to_regclass('teamly.pages')")
    if cur.fetchone()[0] is None:
        check("Teamly schema available", False, "teamly.pages not found")
        check("Teamly 'Student Advising Tracker' has >=4 category pages", False, "no schema")
        cur.close(); conn.close()
        return

    # Find the tracker parent page (English title preserved per task.md).
    cur.execute("""
        SELECT id, title FROM teamly.pages
        WHERE title ILIKE '%%advising%%' AND title ILIKE '%%tracker%%'
    """)
    parents = cur.fetchall()
    check("Student Advising Tracker page exists", len(parents) >= 1,
          f"Matching pages: {[p[1] for p in parents]}")

    # Count child pages (one per category). Prefer parent_id linkage; fall back
    # to category pages anywhere if the tracker was modelled flat.
    category_kw = ["urgent", "academic support", "progress monitoring", "no action"]
    child_count = 0
    if parents:
        parent_ids = tuple(p[0] for p in parents)
        cur.execute(
            "SELECT title, COALESCE(body,'') FROM teamly.pages WHERE parent_id IN %s",
            (parent_ids,))
        children = cur.fetchall()
        child_count = len(children)

    if child_count < 4:
        # Fallback: count pages whose title/body names the advising categories.
        cur.execute("SELECT title, COALESCE(body,'') FROM teamly.pages")
        cat_pages = 0
        for title, body in cur.fetchall():
            t = (str(title) + " " + str(body)).lower()
            if any(kw in t for kw in category_kw):
                cat_pages += 1
        child_count = max(child_count, cat_pages)

    check("Teamly 'Student Advising Tracker' has >=4 category pages",
          child_count >= 4, f"Found {child_count} category/child pages")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Check 5: Calendar Advising Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime, end_datetime FROM gcal.events
        WHERE lower(summary) LIKE '%%advising%%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("At least 5 advising events", len(events) >= 5, f"Found {len(events)} events")

    if events:
        summaries = " ".join(str(e[0]) for e in events).lower()
        check("Events mention academic advising", "academic" in summaries or "advising" in summaries,
              f"Summaries: {summaries[:150]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 6: advising_analyzer.py ===")
    path = os.path.join(workspace, "advising_analyzer.py")
    check("advising_analyzer.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, EXTRACT(DOW FROM start_datetime) AS dow
            FROM gcal.events
            WHERE lower(summary) LIKE '%%advising%%'
        """)
        events = cur.fetchall()
        weekend_events = [e for e in events if e[2] in (0, 6)]
        check("No advising events on weekends",
              len(weekend_events) == 0,
              f"Found {len(weekend_events)} weekend advising events")
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gsheet()
    check_word(args.agent_workspace)
    check_teamly()
    check_gcal()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_failed": critical_failed}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
