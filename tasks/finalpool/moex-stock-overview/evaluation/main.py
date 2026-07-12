"""Evaluation for yf-stock-overview (moex-finance, deterministic PG seed)."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Deterministic groundtruth from moex.stock_info seed.
# 5 tickers WITH a sector, sorted by Market_Cap descending. TCSG.ME excluded (empty sector).
EXPECTED_ROWS = [
    # symbol,    longName,        sector,                  market_cap,    pe_ratio
    ("GAZP.ME", "ПАО Газпром",   "Energy",                4681709912064, 2.2370353),
    ("SBER.ME", "ПАО Сбербанк",  "Financial Services",    2877540270080, 2.6448412),
    ("LKOH.ME", "ПАО ЛУКОЙЛ",    "Energy",                2543495413760, 3.4640048),
    ("MTSS.ME", "ПАО МТС",       "Communication Services", 465162960896, 8.8383665),
    ("MGNT.ME", "ПАО Магнит",    "Consumer Defensive",     435439271936, 9.077895),
]
EXPECTED_SYMBOLS = [r[0] for r in EXPECTED_ROWS]
EXPECTED_CAP = {r[0]: r[3] for r in EXPECTED_ROWS}
EXCLUDED_SYMBOL = "TCSG.ME"  # empty sector -> must NOT appear
TOTAL_CAP = sum(r[3] for r in EXPECTED_ROWS)  # 11003347828736
LARGEST_SYMBOL = "GAZP.ME"
LARGEST_LONG = "ПАО Газпром"
LARGEST_SHORT = "Газпром"
LARGEST_CAP = 4681709912064

CAP_TOL = 1.0  # moex is deterministic; tight tolerance

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}{detail_str}")
        if critical:
            CRITICAL_FAILURES.append(name)


def num_close(a, b, tol=CAP_TOL):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_gsheet_data():
    """Read the 'Stock Watch Dashboard' Google Sheet from PG gsheet.* schema."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gsheet.spreadsheets ORDER BY created_at DESC")
    spreadsheets = cur.fetchall()

    result = {"spreadsheet": None, "sheets": {}, "cells": {}}
    target = None
    for ss_id, ss_title in spreadsheets:
        if "stock watch dashboard" in (ss_title or "").lower():
            target = (ss_id, ss_title)
            break
    if target is None and spreadsheets:
        # fall back to most recent so we can still report partial detail
        target = spreadsheets[0]
    result["spreadsheet"] = target

    if target:
        ss_id = target[0]
        cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s", (ss_id,))
        for sheet_id, sheet_title in cur.fetchall():
            result["sheets"][(sheet_title or "").lower()] = sheet_id
            cur.execute("""
                SELECT row_index, col_index, value FROM gsheet.cells
                WHERE spreadsheet_id = %s AND sheet_id = %s
                ORDER BY row_index, col_index
            """, (ss_id, sheet_id))
            cells = {}
            for row_idx, col_idx, value in cur.fetchall():
                cells.setdefault(row_idx, {})[col_idx] = value
            result["cells"][(sheet_title or "").lower()] = cells

    cur.close()
    conn.close()
    return result


def check_excel(agent_wb):
    print("\n=== Checking Excel: YF_Stock_Overview.xlsx ===")

    # --- Stock Overview ---
    a_rows = load_sheet_rows(agent_wb, "Stock Overview")
    check("Sheet 'Stock Overview' exists", a_rows is not None)
    a_overview = a_rows[1:] if a_rows and len(a_rows) > 1 else []
    a_overview = [r for r in a_overview if r and r[0] is not None]

    # Header column positions (default to canonical order if header missing/odd).
    header = [norm(c) for c in (a_rows[0] if a_rows else [])]
    def col(name, default):
        return header.index(name) if name in header else default
    c_sym, c_name, c_sec, c_cap = col("symbol", 0), col("name", 1), col("sector", 2), col("market_cap", 3)

    present = [norm(r[c_sym]) for r in a_overview]
    present_set = set(present)
    expected_set = {s.lower() for s in EXPECTED_SYMBOLS}

    # CRITICAL: exact 5-symbol set with sector, TCSG.ME excluded
    check("Stock Overview has exactly the 5 sectored tickers (TCSG.ME excluded)",
          present_set == expected_set,
          f"got {sorted(present_set)}, expected {sorted(expected_set)}", critical=True)
    check("TCSG.ME (empty sector) is excluded",
          EXCLUDED_SYMBOL.lower() not in present_set,
          f"present={sorted(present_set)}", critical=True)

    # CRITICAL: per-row Market_Cap exact
    cap_ok = True
    cap_detail = []
    a_lookup = {norm(r[c_sym]): r for r in a_overview}
    for sym in EXPECTED_SYMBOLS:
        r = a_lookup.get(sym.lower())
        v = r[c_cap] if (r and len(r) > c_cap) else None
        if not num_close(v, EXPECTED_CAP[sym]):
            cap_ok = False
            cap_detail.append(f"{sym}: {v} vs {EXPECTED_CAP[sym]}")
    check("Each Market_Cap matches deterministic seed (tol<=1)", cap_ok,
          "; ".join(cap_detail), critical=True)

    # CRITICAL: sorted by Market_Cap descending
    caps = [r[c_cap] for r in a_overview if len(r) > c_cap and isinstance(r[c_cap], (int, float))]
    sorted_desc = all(caps[i] >= caps[i + 1] for i in range(len(caps) - 1)) and len(caps) == len(a_overview)
    check("Stock Overview sorted by Market_Cap descending", sorted_desc,
          f"caps order={caps}", critical=True)

    # NON-critical: every row has a non-empty sector
    sectors_ok = all(norm(r[c_sec]).strip() != "" for r in a_overview if len(r) > c_sec)
    check("Every row has a non-empty Sector", sectors_ok)

    # --- Summary ---
    s_rows = load_sheet_rows(agent_wb, "Summary")
    check("Sheet 'Summary' exists", s_rows is not None)
    summ = {}
    if s_rows:
        for r in s_rows[1:]:
            if r and r[0] is not None and len(r) > 1:
                summ[norm(r[0])] = r[1]

    check("Summary.Total_Stocks == 5",
          num_close(summ.get("total_stocks"), 5, 0.5))
    check("Summary.Total_Market_Cap matches sum",
          num_close(summ.get("total_market_cap"), TOTAL_CAP, 1.0), critical=True)
    check("Summary.Largest_Market_Cap == GAZP cap",
          num_close(summ.get("largest_market_cap"), LARGEST_CAP, 1.0))
    lc = norm(summ.get("largest_company"))
    check("Summary.Largest_Company is Газпром (long or short name)",
          lc in (norm(LARGEST_LONG), norm(LARGEST_SHORT)) or "газпром" in lc,
          f"got {summ.get('largest_company')!r}")


def check_gsheet():
    print("\n=== Checking Google Sheet: Stock Watch Dashboard ===")
    try:
        data = get_gsheet_data()
    except Exception as e:
        check("Google Sheet 'Stock Watch Dashboard' exists", False, f"DB error: {e}")
        return

    ss = data["spreadsheet"]
    has_title = ss is not None and "stock watch dashboard" in (ss[1] or "").lower()
    check("Google Sheet titled 'Stock Watch Dashboard' exists", has_title,
          f"spreadsheet={ss[1] if ss else None}")
    if not has_title:
        return

    stocks_key = next((k for k in data["sheets"] if k == "stocks"), None)
    check("'Stocks' sheet exists", stocks_key is not None,
          f"sheets={list(data['sheets'].keys())}")
    if not stocks_key:
        return

    cells = data["cells"].get(stocks_key, {})
    # Header map
    header = cells.get(0, {})
    hmap = {norm(v): c for c, v in header.items() if v is not None}
    sym_col = hmap.get("symbol")
    cap_col = hmap.get("market_cap")

    data_rows = [cells[r] for r in sorted(cells) if r > 0]
    syms = set()
    if sym_col is not None:
        syms = {norm(row.get(sym_col)) for row in data_rows if row.get(sym_col)}
    check("'Stocks' sheet contains the 5 sectored tickers",
          syms == {s.lower() for s in EXPECTED_SYMBOLS},
          f"got {sorted(syms)}")

    if cap_col is not None and sym_col is not None:
        cap_ok = True
        detail = []
        for row in data_rows:
            sym = norm(row.get(sym_col))
            if sym in EXPECTED_CAP:
                if not num_close(row.get(cap_col), EXPECTED_CAP[sym], CAP_TOL):
                    cap_ok = False
                    detail.append(f"{sym}: {row.get(cap_col)}")
        check("'Stocks' sheet Market_Cap values match seed", cap_ok, "; ".join(detail))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    agent_file = os.path.join(args.agent_workspace, "YF_Stock_Overview.xlsx")
    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

    check_excel(agent_wb)
    check_gsheet()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    print(f"\n=== {PASS_COUNT}/{total} checks passed (accuracy {accuracy:.1f}%) ===")

    if CRITICAL_FAILURES:
        print(f"=== RESULT: FAIL (critical checks failed: {CRITICAL_FAILURES}) ===")
        sys.exit(1)

    if accuracy >= 70.0:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
