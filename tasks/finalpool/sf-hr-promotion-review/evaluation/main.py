"""Evaluation for sf-hr-promotion-review (ClickHouse / RU)."""
import argparse
import os
import re
import sys
import json
import openpyxl


def normalize_ru_numbers(text):
    """RU number normalization for substring/regex checks: collapse digit-group
    separators (space/NBSP/NNBSP/dot/comma before a 3-digit group) and turn
    decimal commas into dots ("31 588" -> "31588", "4 586,91" -> "4586.91")."""
    t = str(text or "")
    t = re.sub(r"(?<=\d)[ \xa0\u202f\u2009.,](?=\d{3}\b)", "", t)
    return re.sub(r"(?<=\d),(?=\d)", ".", t)

try:
    import psycopg2
except Exception:
    psycopg2 = None

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432,
      "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

# Department names are russified CENTRALLY in the ClickHouse seed
# (db/zzz_clickhouse_after_init.sql). The agent reads RU dept names from the DB
# and writes them into the xlsx, while the groundtruth xlsx keeps the English
# literals. We canonicalize both sides through this deterministic map so the
# per-department key matching stays in sync without hand-editing groundtruth.
DEPT_EN2RU = {
    "engineering": "инженерия",
    "finance": "финансы",
    "hr": "кадры",
    "operations": "операции",
    "r&d": "ниокр",
    "sales": "продажи",
    "support": "поддержка",
}
DEPT_RU2EN = {v: k for k, v in DEPT_EN2RU.items()}

# Highest Eligible_Count department in groundtruth: Finance (1444) -> Финансы.
TOP_DEPT_RU = "финансы"
TOP_DEPT_EN = "finance"

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "Summary.Total_Eligible matches groundtruth total of eligible employees",
    "By Department per-department Eligible_Count matches groundtruth for all departments",
    "Summary weighted Avg_Salary_Eligible / Avg_Experience_Eligible match (weighted, not mean-of-means)",
    "Email sent to hr-director@company.com with exact subject",
    "Email body mentions total eligible count and names the top department",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        d = (str(detail)[:300] + "...") if len(str(detail)) > 300 else str(detail)
        print(f"  [FAIL] {name}: {d}")


def dept_key(raw):
    """Canonicalize a department label to a language-neutral key (English lower)."""
    s = str(raw).strip().lower()
    if s in DEPT_EN2RU:
        return s
    if s in DEPT_RU2EN:
        return DEPT_RU2EN[s]
    return s


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_file, gt_file):
    if not os.path.exists(agent_file):
        check("Promotion_Candidates.xlsx exists", False, agent_file)
        return
    check("Promotion_Candidates.xlsx exists", True)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---------- By Department ----------
    a_rows = load_sheet_rows(agent_wb, "By Department")
    g_rows = load_sheet_rows(gt_wb, "By Department")
    check("Sheet 'By Department' exists", a_rows is not None,
          f"Sheets: {agent_wb.sheetnames}")

    if a_rows is not None and g_rows is not None:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[dept_key(row[0])] = row
        g_lookup = {}
        for row in g_data:
            if row and row[0] is not None:
                g_lookup[dept_key(row[0])] = row

        # NON-critical: row count present
        check("By Department has all department rows",
              len(a_lookup) >= len(g_lookup),
              f"agent depts={sorted(a_lookup)} gt depts={sorted(g_lookup)}")

        # CRITICAL: per-department Eligible_Count for ALL departments.
        count_ok = True
        count_detail = []
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                count_ok = False
                count_detail.append(f"missing dept {key}")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 1):
                count_ok = False
                count_detail.append(f"{key}.Eligible_Count {a_row[1]} vs {g_row[1]}")
        check("By Department per-department Eligible_Count matches groundtruth for all departments",
              count_ok, "; ".join(count_detail))  # CRITICAL

        # NON-critical: averages per department (loose tolerances).
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                continue
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"By Department {key}.Avg_Salary",
                      num_close(a_row[2], g_row[2], 5.0),
                      f"{a_row[2]} vs {g_row[2]}")
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"By Department {key}.Avg_Experience",
                      num_close(a_row[3], g_row[3], 0.5),
                      f"{a_row[3]} vs {g_row[3]}")
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"By Department {key}.Avg_Rating",
                      num_close(a_row[4], g_row[4], 0.1),
                      f"{a_row[4]} vs {g_row[4]}")

    # ---------- Summary ----------
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    check("Sheet 'Summary' exists", a_rows is not None,
          f"Sheets: {agent_wb.sheetnames}")

    if a_rows is not None and g_rows is not None:
        a_sum = {}
        for row in a_rows[1:]:
            if row and row[0] is not None:
                a_sum[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
        g_sum = {}
        for row in g_rows[1:]:
            if row and row[0] is not None:
                g_sum[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        # CRITICAL: Total_Eligible (headline number), tol=1.
        check("Summary.Total_Eligible matches groundtruth total of eligible employees",
              num_close(a_sum.get("total_eligible"), g_sum.get("total_eligible"), 1),
              f"{a_sum.get('total_eligible')} vs {g_sum.get('total_eligible')}")  # CRITICAL

        # CRITICAL: weighted averages with tight tolerances (distinguishes
        # weighted-over-all-eligible from mean-of-department-means).
        sal_ok = num_close(a_sum.get("avg_salary_eligible"),
                           g_sum.get("avg_salary_eligible"), 1.0)
        exp_ok = num_close(a_sum.get("avg_experience_eligible"),
                           g_sum.get("avg_experience_eligible"), 0.2)
        check("Summary weighted Avg_Salary_Eligible / Avg_Experience_Eligible match (weighted, not mean-of-means)",
              sal_ok and exp_ok,
              f"salary {a_sum.get('avg_salary_eligible')} vs {g_sum.get('avg_salary_eligible')}; "
              f"exp {a_sum.get('avg_experience_eligible')} vs {g_sum.get('avg_experience_eligible')}")  # CRITICAL

        # Store total for email body check.
        return g_sum.get("total_eligible")
    return None


def check_email(gt_total):
    if psycopg2 is None:
        check("Email sent to hr-director@company.com with exact subject", False,
              "psycopg2 unavailable")
        check("Email body mentions total eligible count and names the top department",
              False, "psycopg2 unavailable")
        return
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE %s
        """, ("%hr-director@company.com%",))
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Email sent to hr-director@company.com with exact subject", False, str(e))
        check("Email body mentions total eligible count and names the top department",
              False, str(e))
        return

    EXACT_SUBJECT = "promotion cycle: eligible candidates summary"
    subj_rows = [r for r in rows if (r[0] or "").strip().lower() == EXACT_SUBJECT]
    # CRITICAL: an email to the director with the exact subject exists.
    check("Email sent to hr-director@company.com with exact subject",
          len(subj_rows) > 0,
          f"matching rows={len(rows)} to dir; subjects={[r[0] for r in rows]}")  # CRITICAL

    # CRITICAL: body mentions correct total eligible count AND the top department.
    body_ok = False
    target_subj_rows = subj_rows if subj_rows else rows
    total_str = str(int(gt_total)) if gt_total is not None else None
    for _, _, body in target_subj_rows:
        b = (body or "").lower()  # ORIGINAL lower (RU keywords), not normalize()
        bn = normalize_ru_numbers(b)
        total_hit = (total_str is not None and total_str in bn)
        dept_hit = (TOP_DEPT_RU in b) or (TOP_DEPT_EN in b)
        if total_hit and dept_hit:
            body_ok = True
            break
    check("Email body mentions total eligible count and names the top department",
          body_ok,
          f"need total '{total_str}' and dept '{TOP_DEPT_RU}'/'{TOP_DEPT_EN}' in body")  # CRITICAL


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_file = os.path.join(args.agent_workspace, "Promotion_Candidates.xlsx")
    gt_file = os.path.join(gt_dir, "Promotion_Candidates.xlsx")

    print("=== Evaluation: sf-hr-promotion-review (ClickHouse/RU) ===")
    gt_total = check_excel(agent_file, gt_file)
    check_email(gt_total)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "pass": PASS_COUNT, "fail": FAIL_COUNT,
                "total_passed": PASS_COUNT, "total_checks": total,
                "accuracy": accuracy, "critical_failed": critical_failed,
            }, f)

    success = (not critical_failed) and accuracy >= 70
    if success:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
