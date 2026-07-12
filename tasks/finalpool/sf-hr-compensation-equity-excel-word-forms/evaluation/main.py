"""Evaluation for sf-hr-compensation-equity-excel-word-gform (ClickHouse + forms fork).

The HR source data in sf_data (HR_ANALYTICS) is russified centrally by
db/zzz_clickhouse_after_init.sql, so the agent reads RUSSIAN department/education
values and writes them into Excel. The frozen groundtruth Compensation_Equity.xlsx
has been relabeled through scripts/clickhouse_relabel_map.FLAT_VALUE_MAP, so its
Department/Education_Level/Highest_Paid_Group/Lowest_Paid_Group/Highest_Gap_Dept
cells already hold the same Russian values. Experience bands (Junior/Mid/Senior/
Expert) and equity statuses (Concerning/Monitor/Acceptable) stay English tokens.

Scoring: CRITICAL_CHECKS gate (any critical failure => FAIL) is applied BEFORE an
accuracy>=70 gate over the non-critical structural checks.
"""
import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        d = (str(detail)[:300] + "...") if len(str(detail)) > 300 else detail
        msg = f": {d}" if d else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def check(name, condition, detail=""):
    record(name, condition, detail, critical=False)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_ws, gt_dir):
    import openpyxl

    agent_file = os.path.join(agent_ws, "Compensation_Equity.xlsx")
    gt_file = os.path.join(gt_dir, "Compensation_Equity.xlsx")

    print("\n=== Проверка 1: Compensation_Equity.xlsx ===")
    if not os.path.exists(agent_file):
        check("Файл Compensation_Equity.xlsx существует", False, "not found")
        # critical deliverables cannot be verified -> fail them
        record("Salary Analysis: строки (dept,edu,band) совпадают по count/avg/median",
               False, "no file", critical=True)
        record("Equity Metrics: Pay_Gap_Pct/Equity_Ratio/Equity_Status корректны",
               False, "no file", critical=True)
        record("Summary: счётчики разрывов и Overall_Equity_Score корректны",
               False, "no file", critical=True)
        record("Summary: Highest_Gap_Dept и Highest_Gap_Band корректны",
               False, "no file", critical=True)
        return

    check("Файл Compensation_Equity.xlsx существует", True)
    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # --- Sheet: Salary Analysis (CRITICAL deliverable) ---
    print("  -- Salary Analysis --")
    a_rows = load_sheet_rows(agent_wb, "Salary Analysis")
    g_rows = load_sheet_rows(gt_wb, "Salary Analysis")
    sa_ok = True
    sa_detail = ""
    if a_rows is None:
        sa_ok = False
        sa_detail = "sheet 'Salary Analysis' missing"
    elif g_rows is None:
        sa_ok = False
        sa_detail = "groundtruth sheet missing"
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] and row[1] and row[2]:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower(), str(row[2]).strip().lower())
                a_lookup[key] = row
        errors = []
        for g_row in g_data:
            if not g_row or not g_row[0]:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower(), str(g_row[2]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}/{g_row[1]}/{g_row[2]}")
                continue
            if not num_close(a_row[3], g_row[3], 1):
                errors.append(f"{key} Employee_Count: {a_row[3]} vs {g_row[3]}")
            if not num_close(a_row[4], g_row[4], 5.0):
                errors.append(f"{key} Avg_Salary: {a_row[4]} vs {g_row[4]}")
            if not num_close(a_row[5], g_row[5], 5.0):
                errors.append(f"{key} Median_Salary: {a_row[5]} vs {g_row[5]}")
            if not num_close(a_row[6], g_row[6], 5.0):
                errors.append(f"{key} Min_Salary: {a_row[6]} vs {g_row[6]}")
            if not num_close(a_row[7], g_row[7], 5.0):
                errors.append(f"{key} Max_Salary: {a_row[7]} vs {g_row[7]}")
            if not num_close(a_row[8], g_row[8], 5.0):
                errors.append(f"{key} Salary_Std_Dev: {a_row[8]} vs {g_row[8]}")
        sa_ok = not errors
        sa_detail = "; ".join(errors[:5])
    record("Salary Analysis: строки (dept,edu,band) совпадают по count/avg/median/min/max/std",
           sa_ok, sa_detail, critical=True)

    # --- Sheet: Equity Metrics (CRITICAL deliverable) ---
    print("  -- Equity Metrics --")
    a_rows = load_sheet_rows(agent_wb, "Equity Metrics")
    g_rows = load_sheet_rows(gt_wb, "Equity Metrics")
    em_ok = True
    em_detail = ""
    em_groups_ok = True
    em_groups_detail = ""
    if a_rows is None:
        em_ok = em_groups_ok = False
        em_detail = em_groups_detail = "sheet 'Equity Metrics' missing"
    elif g_rows is None:
        em_ok = em_groups_ok = False
        em_detail = em_groups_detail = "groundtruth sheet missing"
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] and row[1]:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                a_lookup[key] = row
        num_errors = []
        grp_errors = []
        for g_row in g_data:
            if not g_row or not g_row[0]:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                num_errors.append(f"Missing equity row: {g_row[0]}/{g_row[1]}")
                grp_errors.append(f"Missing equity row: {g_row[0]}/{g_row[1]}")
                continue
            # Pay_Gap_Pct / Equity_Ratio / Equity_Status -> core metric correctness
            if not num_close(a_row[4], g_row[4], 0.5):
                num_errors.append(f"{key} Pay_Gap_Pct: {a_row[4]} vs {g_row[4]}")
            if not num_close(a_row[5], g_row[5], 0.01):
                num_errors.append(f"{key} Equity_Ratio: {a_row[5]} vs {g_row[5]}")
            if not str_match(a_row[6], g_row[6]):
                num_errors.append(f"{key} Equity_Status: '{a_row[6]}' vs '{g_row[6]}'")
            # Highest/Lowest paid education-level group (russified labels)
            if not str_match(a_row[2], g_row[2]):
                grp_errors.append(f"{key} Highest_Paid_Group: '{a_row[2]}' vs '{g_row[2]}'")
            if not str_match(a_row[3], g_row[3]):
                grp_errors.append(f"{key} Lowest_Paid_Group: '{a_row[3]}' vs '{g_row[3]}'")
        em_ok = not num_errors
        em_detail = "; ".join(num_errors[:5])
        em_groups_ok = not grp_errors
        em_groups_detail = "; ".join(grp_errors[:5])
    record("Equity Metrics: Pay_Gap_Pct/Equity_Ratio/Equity_Status корректны для каждой dept+band",
           em_ok, em_detail, critical=True)
    # Highest/Lowest paid group is structural correctness (non-critical).
    check("Equity Metrics: Highest_Paid_Group/Lowest_Paid_Group совпадают (RU метки)",
          em_groups_ok, em_groups_detail)

    # --- Sheet: Summary (CRITICAL deliverable) ---
    print("  -- Summary --")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    counts_ok = True
    counts_detail = ""
    gap_ok = True
    gap_detail = ""
    if a_rows is None or g_rows is None:
        counts_ok = gap_ok = False
        counts_detail = gap_detail = "Summary sheet missing"
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        count_errors = []
        gap_errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                (gap_errors if key in ("highest_gap_dept", "highest_gap_band")
                 else count_errors).append(f"Missing summary row: {g_row[0]}")
                continue
            if key == "overall_equity_score":
                if not num_close(a_row[1], g_row[1], 0.02):
                    count_errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol=0.02)")
            elif key in ("total_employees_analyzed", "departments_analyzed",
                         "concerning_gaps_count", "monitor_gaps_count", "acceptable_gaps_count"):
                if not num_close(a_row[1], g_row[1], 1):
                    count_errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol=1)")
            elif key in ("highest_gap_dept", "highest_gap_band"):
                if not str_match(a_row[1], g_row[1]):
                    gap_errors.append(f"{key}: '{a_row[1]}' vs '{g_row[1]}'")
            else:
                if not str_match(a_row[1], g_row[1]):
                    count_errors.append(f"{key}: '{a_row[1]}' vs '{g_row[1]}'")
        counts_ok = not count_errors
        counts_detail = "; ".join(count_errors[:5])
        gap_ok = not gap_errors
        gap_detail = "; ".join(gap_errors[:5])
    record("Summary: Concerning/Monitor/Acceptable_Gaps_Count и Overall_Equity_Score корректны",
           counts_ok, counts_detail, critical=True)
    record("Summary: Highest_Gap_Dept и Highest_Gap_Band корректны (RU метка отдела)",
           gap_ok, gap_detail, critical=True)


def check_word(agent_ws):
    print("\n=== Проверка 2: Equity_Report.docx ===")
    docx_path = os.path.join(agent_ws, "Equity_Report.docx")
    if not os.path.exists(docx_path):
        check("Файл Equity_Report.docx существует", False, "not found")
        return
    check("Файл Equity_Report.docx существует", True)
    try:
        from docx import Document as _DocCheck
        _doc = _DocCheck(docx_path)
        _text = " ".join(p.text for p in _doc.paragraphs).lower()
        _headings = " ".join(
            p.text for p in _doc.paragraphs if p.style.name.startswith("Heading")
        ).lower()
        joined = _text + " " + _headings

        check("Документ содержит достаточно текста (>=100 символов)", len(_text.strip()) >= 100)

        # Required sections: accept RU or EN substrings.
        required_sections = {
            "executive summary": ["executive summary", "резюме", "краткое резюме"],
            "methodology": ["methodology", "методолог", "методик"],
            "findings": ["finding", "вывод", "наход"],
            "recommendation": ["recommendation", "рекомендац"],
            "compliance": ["compliance", "соответств", "комплаенс"],
        }
        missing = [name for name, alts in required_sections.items()
                   if not any(a in joined for a in alts)]
        check("Документ содержит обязательные разделы (резюме/методология/выводы/рекомендации/соответствие)",
              len(missing) <= 2, f"missing: {missing}")

        # Content keywords: accept RU or EN.
        content_kws = {
            "equity": ["equity", "справедлив"],
            "salary": ["salary", "зарплат", "оклад"],
            "department": ["department", "отдел", "подразделен"],
            "gap": ["gap", "разрыв"],
        }
        missing_kw = [name for name, alts in content_kws.items()
                      if not any(a in _text for a in alts)]
        check("Документ содержит ключевые термины (equity/salary/department/gap, RU или EN)",
              len(missing_kw) < len(content_kws), f"missing: {missing_kw}")
    except ImportError:
        check("Equity_Report.docx непустой", os.path.getsize(docx_path) >= 100)
    except Exception as _e:
        check("Equity_Report.docx читается", False, str(_e))


def _option_values(cfg):
    if not isinstance(cfg, dict):
        return []
    opts = cfg.get("options", [])
    out = []
    for o in opts:
        if isinstance(o, dict):
            out.append(str(o.get("value", "")))
        else:
            out.append(str(o))
    return out


def check_form():
    print("\n=== Проверка 3: Google Form 'Compensation Fairness Survey' ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("Подключение к БД для проверки формы", False, str(e))
        return
    try:
        cur.execute("SELECT id FROM gform.forms WHERE LOWER(title) LIKE '%compensation fairness%'")
        forms = cur.fetchall()
        check("Форма 'Compensation Fairness Survey' существует", bool(forms))
        if not forms:
            record("Форма содержит 5 требуемых вопросов нужных типов", False, "no form", critical=True)
            return
        form_id = forms[0][0]
        cur.execute(
            "SELECT title, question_type, config FROM gform.questions "
            "WHERE form_id = %s ORDER BY position", (form_id,))
        rows = cur.fetchall()
    finally:
        cur.close()
        conn.close()

    check("Форма содержит не менее 5 вопросов", len(rows) >= 5, f"got {len(rows)}")

    parsed = []
    for q_title, q_type, q_config in rows:
        cfg = q_config if isinstance(q_config, dict) else (
            json.loads(q_config) if q_config else {})
        parsed.append({
            "title": (q_title or "").lower(),
            "type": q_type,
            "options_lower": [v.lower() for v in _option_values(cfg)],
        })

    # The forms MCP only emits 'textQuestion' (text) and 'choiceQuestion' (RADIO).
    choice_qs = [p for p in parsed if p["type"] == "choiceQuestion"]
    text_qs = [p for p in parsed if p["type"] == "textQuestion"]

    # Department question: a choice question whose options include >=3 of the
    # russified department names (Инженерия/Финансы/Кадры/Операции/НИОКР/Продажи/Поддержка).
    dept_names = ["инженер", "финанс", "кадр", "операц", "ниокр", "продаж", "поддержк",
                  "engineering", "finance", "hr", "operations", "r&d", "sales", "support"]
    dept_q_ok = False
    for p in choice_qs:
        joined = " ".join(p["options_lower"])
        hits = sum(1 for d in dept_names if d in joined)
        if hits >= 3:
            dept_q_ok = True
            break

    # Two rating questions: choice questions offering a 1..5 scale.
    rating_count = 0
    for p in choice_qs:
        opts = set(p["options_lower"])
        if all(str(n) in opts for n in (1, 2, 3, 4, 5)):
            rating_count += 1

    # Free-text questions: experience (years) + comments => at least 2 text questions.
    has_two_text = len(text_qs) >= 2

    record(
        "Форма содержит 5 требуемых вопросов нужных типов "
        "(отдел=выбор с названиями отделов, 2 рейтинга 1-5, 2 текстовых)",
        dept_q_ok and rating_count >= 2 and has_two_text,
        f"dept_q_ok={dept_q_ok}, rating_count={rating_count}, text_qs={len(text_qs)}",
        critical=True,
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    check_excel(agent_ws, gt_dir)
    check_word(agent_ws)
    check_form()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    print(f"\n=== ИТОГ: passed {PASS_COUNT}/{total} ({accuracy:.1f}%) ===")
    if CRITICAL_FAILS:
        print(f"FAIL: провалены критические проверки: {CRITICAL_FAILS}")
        sys.exit(1)
    if accuracy < 70.0:
        print(f"FAIL: точность {accuracy:.1f}% < 70%")
        sys.exit(1)
    print("PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
