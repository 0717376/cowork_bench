"""Evaluation script for pw-insales-review-analysis-excel-notion (russified, teamly).

Критические проверки (CRITICAL_CHECKS): любой провал => общий FAIL независимо
от точности. Структурные проверки (лист существует, колонка присутствует) —
некритические.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Источник истины: значения, опубликованные на mock-сайте http://localhost:30307.
# Product-имена и числа на сайте оставлены EXACTLY как есть (ключи join + числа).
EXPECTED_PRODUCTS = {
    "Wireless Headphones": 4.2,
    "Smart Watch": 3.8,
    "Bluetooth Speaker": 4.5,
    "USB-C Hub": 3.2,
    "Webcam HD": 4.0,
    "Mechanical Keyboard": 4.7,
    "Gaming Mouse": 4.3,
    "Laptop Stand": 4.1,
}
# Товары, требующие внимания (тональность Negative/Mixed): USB-C Hub + Smart Watch.
NEEDS_ATTENTION_PRODUCTS = ["USB-C Hub", "Smart Watch"]
EXPECTED_TOTAL = 8
EXPECTED_AVG = round(sum(EXPECTED_PRODUCTS.values()) / len(EXPECTED_PRODUCTS), 2)  # 4.1
EXPECTED_NEEDS_ATTENTION = 2
EXPECTED_POSITIVE = 6

# Тональности с негативным/смешанным оттенком (RU + EN), для подсчёта Needs_Attention.
NEG_MIXED_TOKENS = ["negative", "mixed", "негатив", "смешан"]

CRITICAL_CHECKS = {
    "External_Reviews contains all 8 products with correct External_Rating",
    "Review_Summary Total_Products_Reviewed == 8 and Avg_External_Rating ~= 4.1",
    "Review_Summary Needs_Attention == 2 and Positive_Products == 6",
    "Teamly review-tracker page exists and mentions a needs-attention product with complaint",
    "review_analyzer.py + external/internal/insights JSON artifacts exist",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _header_map(ws):
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def check_external_reviews_values(wb):
    """Критическая: на листе External_Reviews есть все 8 товаров с верным
    External_Rating (как опубликовано на mock-сайте)."""
    name = "External_Reviews contains all 8 products with correct External_Rating"
    if "External_Reviews" not in wb.sheetnames:
        check(name, False, "no External_Reviews sheet")
        return
    ws = wb["External_Reviews"]
    headers, hmap = _header_map(ws)
    p_i = hmap.get("product")
    r_i = hmap.get("external_rating")
    if p_i is None or r_i is None:
        check(name, False, f"missing Product/External_Rating columns: {headers}")
        return
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    ok = True
    bad = []
    for prod, exp_rating in EXPECTED_PRODUCTS.items():
        found = False
        for r in rows:
            rp = str(r[p_i]).strip() if p_i < len(r) and r[p_i] is not None else ""
            if rp.lower() == prod.lower():
                rr = safe_float(r[r_i] if r_i < len(r) else None)
                if rr is not None and abs(rr - exp_rating) < 0.05:
                    found = True
                break
        if not found:
            ok = False
            bad.append(prod)
    check(name, ok, f"missing/wrong: {bad}")


def _summary_metrics(wb):
    """Возвращает dict нормализованных метрик -> значение из листа Review_Summary."""
    if "Review_Summary" not in wb.sheetnames:
        return {}
    ws = wb["Review_Summary"]
    headers, hmap = _header_map(ws)
    m_i = hmap.get("metric", 0)
    v_i = hmap.get("value", 1)
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[m_i] is None:
            continue
        key = str(r[m_i]).strip().lower().replace(" ", "_")
        out[key] = r[v_i] if v_i < len(r) else None
    return out


def check_summary_totals(wb):
    """Критическая: Total_Products_Reviewed==8 и Avg_External_Rating~=4.1."""
    name = "Review_Summary Total_Products_Reviewed == 8 and Avg_External_Rating ~= 4.1"
    m = _summary_metrics(wb)
    total = safe_float(m.get("total_products_reviewed"))
    avg = safe_float(m.get("avg_external_rating"))
    ok = (total is not None and abs(total - EXPECTED_TOTAL) < 0.5
          and avg is not None and abs(avg - EXPECTED_AVG) < 0.05)
    check(name, ok, f"total={total}, avg={avg} (expected {EXPECTED_TOTAL}, {EXPECTED_AVG})")


def check_summary_attention(wb):
    """Критическая: Needs_Attention==2 и Positive_Products==6."""
    name = "Review_Summary Needs_Attention == 2 and Positive_Products == 6"
    m = _summary_metrics(wb)
    na = safe_float(m.get("needs_attention"))
    pos = safe_float(m.get("positive_products"))
    ok = (na is not None and abs(na - EXPECTED_NEEDS_ATTENTION) < 0.5
          and pos is not None and abs(pos - EXPECTED_POSITIVE) < 0.5)
    check(name, ok, f"needs_attention={na}, positive={pos} "
                    f"(expected {EXPECTED_NEEDS_ATTENTION}, {EXPECTED_POSITIVE})")


def check_artifacts(agent_workspace):
    """Критическая: review_analyzer.py + три JSON-артефакта существуют."""
    name = "review_analyzer.py + external/internal/insights JSON artifacts exist"
    needed = [
        "review_analyzer.py",
        "external_reviews.json",
        "internal_reviews.json",
        "review_insights.json",
    ]
    missing = [f for f in needed if not os.path.exists(os.path.join(agent_workspace, f))]
    check(name, not missing, f"missing: {missing}")


def check_teamly():
    """Критическая: существует не-архивная страница трекера отзывов (RU+EN
    заголовок), в теле которой упомянут хотя бы один товар, требующий внимания
    (USB-C Hub или Smart Watch), вместе с жалобой.

    Seed-страницы имеют id <= 3; шумовая страница ('Заметки команды магазина')
    не должна удовлетворять проверке.
    """
    name = "Teamly review-tracker page exists and mentions a needs-attention product with complaint"
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check(name, False, str(e))
        return

    tracker = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "заметки команды" in tl:
            continue
        if ("review" in tl and ("tracker" in tl or "track" in tl)) \
                or "product review" in tl \
                or "отзыв" in tl or "трекер" in tl:
            tracker = (pid, title, body)
            break
    if tracker is None:
        check(name, False, f"new pages: {[(p[0], p[1]) for p in pages]}")
        return

    text = ((tracker[1] or "") + " " + (tracker[2] or "")).lower()
    # упомянут хотя бы один товар, требующий внимания
    mentions_product = any(p.lower() in text for p in NEEDS_ATTENTION_PRODUCTS)
    # и есть упоминание жалобы (RU complaint-проза из mock-сайта или общее слово)
    complaint_tokens = [
        "complaint", "жалоб", "проблем", "совместимост", "прочность экрана",
        "экран", "compatibility", "durability",
    ]
    mentions_complaint = any(t in text for t in complaint_tokens)
    check(name, mentions_product and mentions_complaint,
          f"product={mentions_product}, complaint={mentions_complaint}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Review_Analysis_Report.xlsx")
    check("Review_Analysis_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # --- структурные (некритические) проверки ---
        check("External_Reviews sheet exists", "External_Reviews" in wb.sheetnames)
        if "External_Reviews" in wb.sheetnames:
            ws = wb["External_Reviews"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("External_Reviews has >= 8 rows", len(data_rows) >= 8, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Product', 'External_Rating', 'Review_Count', 'Sentiment', 'Common_Complaint']:
                check(f"External_Reviews has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Review_Summary sheet exists", "Review_Summary" in wb.sheetnames)
        if "Review_Summary" in wb.sheetnames:
            ws = wb["Review_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Review_Summary has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Review_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # --- КРИТИЧЕСКИЕ семантические проверки значений ---
        check_external_reviews_values(wb)
        check_summary_totals(wb)
        check_summary_attention(wb)
    else:
        # без файла критические Excel-проверки тоже фейлятся
        check_external_reviews_values(openpyxl.Workbook())
        check_summary_totals(openpyxl.Workbook())
        check_summary_attention(openpyxl.Workbook())

    # review_analyzer.py + JSON-артефакты (критическая)
    check_artifacts(agent_workspace)
    # дублируем структурную проверку наличия скрипта (некритическая)
    check("review_analyzer.py exists",
          os.path.exists(os.path.join(agent_workspace, "review_analyzer.py")))

    # Teamly страница трекера отзывов (критическая)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
