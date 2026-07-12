"""Evaluation for sf-support-coaching-plan-excel-clickhouse-teamly-gcal.

Checks:
1. Excel Agent_Scorecard.xlsx with Performance Metrics, Coaching Plan, and Summary sheets
2. Teamly space "Agent Coaching Tracker" with one page per agent
3. Google Calendar with 5 coaching session events in March 16-20, 2026

CRITICAL_CHECKS gate the run: any critical failure => sys.exit(1) before the
accuracy gate. The semantic deliverables (per-agent tier / SLA / focus area /
coaching frequency, and the summary aggregates) are critical.
"""
import argparse
import json
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_CRITICAL = []

# ClickHouse fork russifies REPORTER values; the agent reads RUSSIAN agent names.
AGENT_NAMES = ["Алиса", "Борис", "Иван", "Карл", "Эмилия"]

# CRITICAL check names (semantic, value-bearing). Any failure here -> hard FAIL.
CRITICAL_CHECKS = set()


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if critical:
        CRITICAL_CHECKS.add(name)
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            FAILED_CRITICAL.append(name)
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}: {str(detail)[:200]}")


def num_close(a, b, tol=1.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def norm_str(x):
    return str(x).strip().lower() if x is not None else None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return norm_str(a) == norm_str(b)


def str_match_any(a, options):
    """True if a equals (case-insensitive) any option in the list."""
    na = norm_str(a)
    return na is not None and na in {norm_str(o) for o in options}


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    print("\n=== Checking Excel ===")
    try:
        import openpyxl
    except ImportError:
        check("openpyxl installed", False, "openpyxl not available")
        return

    agent_file = os.path.join(agent_workspace, "Agent_Scorecard.xlsx")
    gt_file = os.path.join(gt_dir, "Agent_Scorecard.xlsx")

    check("Agent_Scorecard.xlsx exists", os.path.exists(agent_file),
          f"Not found at {agent_file}", critical=True)
    if not os.path.exists(agent_file) or not os.path.exists(gt_file):
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # --- Performance Metrics sheet ---
    a_rows = load_sheet_rows(agent_wb, "Performance Metrics")
    g_rows = load_sheet_rows(gt_wb, "Performance Metrics")
    check("Performance Metrics sheet exists", a_rows is not None,
          f"Sheets: {agent_wb.sheetnames}")

    if a_rows and g_rows:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        check("Performance Metrics has 5 data rows", len(a_data) >= 5,
              f"Got {len(a_data)}")

        a_lookup = {norm_str(r[0]): r for r in a_data}
        for g_row in g_data:
            key = norm_str(g_row[0])
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Agent {g_row[0]} in Performance Metrics", False, "Missing",
                      critical=True)
                continue
            # Total_Tickets (col 1)
            check(f"{g_row[0]} Total_Tickets",
                  num_close(a_row[1], g_row[1], 10),
                  f"got {a_row[1]}, expected {g_row[1]}")
            # Resolved_Tickets (col 2) -- previously unchecked
            check(f"{g_row[0]} Resolved_Tickets",
                  num_close(a_row[2], g_row[2], 10),
                  f"got {a_row[2]}, expected {g_row[2]}")
            # SLA_Compliance_Pct (col 3) -- central computed metric (CRITICAL, tol 1.0)
            check(f"{g_row[0]} SLA_Compliance_Pct",
                  num_close(a_row[3], g_row[3], 1.0),
                  f"got {a_row[3]}, expected {g_row[3]}", critical=True)
            # Avg_CSAT (col 4)
            check(f"{g_row[0]} Avg_CSAT",
                  num_close(a_row[4], g_row[4], 0.1),
                  f"got {a_row[4]}, expected {g_row[4]}")

    # --- Coaching Plan sheet ---
    a_cp = load_sheet_rows(agent_wb, "Coaching Plan")
    g_cp = load_sheet_rows(gt_wb, "Coaching Plan")
    check("Coaching Plan sheet exists", a_cp is not None,
          f"Sheets: {agent_wb.sheetnames}")

    # Coaching_Frequency is fully determined by tier per the framework.
    TIER_FREQ = {
        "elite": "quarterly", "strong": "monthly",
        "developing": "bi-weekly", "needs improvement": "weekly",
    }

    if a_cp and g_cp:
        a_cp_data = [r for r in a_cp[1:] if r and r[0] is not None]
        g_cp_data = [r for r in g_cp[1:] if r and r[0] is not None]
        check("Coaching Plan has 5 data rows", len(a_cp_data) >= 5,
              f"Got {len(a_cp_data)}")

        a_cp_lookup = {norm_str(r[0]): r for r in a_cp_data}
        for g_row in g_cp_data:
            key = norm_str(g_row[0])
            a_row = a_cp_lookup.get(key)
            if a_row is None:
                check(f"Agent {g_row[0]} in Coaching Plan", False, "Missing",
                      critical=True)
                continue
            # Performance_Tier (col 1) -- core analytical deliverable (CRITICAL)
            check(f"{g_row[0]} Performance_Tier",
                  str_match(a_row[1], g_row[1]),
                  f"got '{a_row[1]}', expected '{g_row[1]}'", critical=True)
            # Coaching_Frequency (col 2) -- tier-derived cadence (CRITICAL)
            expected_freq = TIER_FREQ.get(norm_str(g_row[1]), g_row[2])
            check(f"{g_row[0]} Coaching_Frequency",
                  str_match(a_row[2], g_row[2]) or str_match(a_row[2], expected_freq),
                  f"got '{a_row[2]}', expected '{g_row[2]}'", critical=True)
            # Primary_Focus_Area (col 3) -- priority-ordered focus rule (CRITICAL)
            check(f"{g_row[0]} Primary_Focus_Area",
                  str_match(a_row[3], g_row[3]),
                  f"got '{a_row[3]}', expected '{g_row[3]}'", critical=True)
            # Secondary_Focus_Area (col 4) -- non-critical
            check(f"{g_row[0]} Secondary_Focus_Area",
                  str_match(a_row[4], g_row[4]),
                  f"got '{a_row[4]}', expected '{g_row[4]}'")

    # --- Summary sheet ---
    a_sum = load_sheet_rows(agent_wb, "Summary")
    check("Summary sheet exists", a_sum is not None,
          f"Sheets: {agent_wb.sheetnames}")

    if a_sum:
        a_sum_data = {norm_str(r[0]): r[1] for r in a_sum[1:] if r and r[0]}
        tt = a_sum_data.get("total_tickets")
        # Summary Total_Tickets aggregate (CRITICAL, tol 50)
        check("Summary Total_Tickets", num_close(tt, 31588, 50),
              f"got {tt}, expected 31588", critical=True)
        ta = a_sum_data.get("total_agents")
        # Summary Total_Agents == 5 (CRITICAL)
        check("Summary Total_Agents", num_close(ta, 5, 0),
              f"got {ta}, expected 5", critical=True)
        asla = a_sum_data.get("avg_sla_compliance")
        check("Summary Avg_SLA_Compliance", num_close(asla, 20.9, 2.0),
              f"got {asla}, expected ~20.9")
        acsat = a_sum_data.get("avg_csat")
        check("Summary Avg_CSAT", num_close(acsat, 3.26, 0.1),
              f"got {acsat}, expected ~3.26")


def check_teamly():
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except Exception as e:
        check("Teamly Agent Coaching Tracker space exists", False, str(e))
        return

    # The 'Agent Coaching Tracker' deliverable: a space holding one page per agent.
    # Accept either a dedicated space named like the tracker, OR (fallback) pages
    # whose title carries the tracker marker. Seed pages have id <= 3.
    space_id = None
    try:
        cur.execute("""
            SELECT id FROM teamly.spaces
            WHERE LOWER(name) LIKE '%agent%coaching%'
               OR LOWER(name) LIKE '%coaching%tracker%'
               OR LOWER(key) LIKE '%coaching%'
        """)
        row = cur.fetchone()
        if row:
            space_id = row[0]
    except Exception as e:
        print(f"  [warn] space lookup: {e}")

    check("Teamly Agent Coaching Tracker space exists", space_id is not None,
          "No space named like 'Agent Coaching Tracker'")

    # Gather candidate agent pages: user-created (id > 3), in the tracker space if
    # found, excluding the noise pages by title.
    try:
        if space_id is not None:
            cur.execute(
                "SELECT id, title, COALESCE(body,'') FROM teamly.pages WHERE id > 3 AND space_id = %s",
                (space_id,))
        else:
            cur.execute("SELECT id, title, COALESCE(body,'') FROM teamly.pages WHERE id > 3")
        rows = cur.fetchall()
    except Exception as e:
        check("At least 5 agent pages in tracker", False, str(e))
        conn.close()
        return
    finally:
        pass

    NOISE_MARKERS = ["планёрк", "okr", "праздник"]
    agent_pages = [
        (t, b) for (_id, t, b) in rows
        if not any(m in (t or "").lower() for m in NOISE_MARKERS)
    ]
    check("At least 5 agent pages in tracker", len(agent_pages) >= 5,
          f"Found {len(agent_pages)} candidate pages")

    # Each agent name appears as a page title (CRITICAL: >=4 of 5).
    titles_text = " ".join((t or "") for t, _ in agent_pages).lower()
    found_titles = sum(1 for a in AGENT_NAMES if a.lower() in titles_text)
    check("Agent names appear as Teamly page titles", found_titles >= 4,
          f"Found {found_titles}/5 agent names in titles", critical=True)

    # Page bodies carry the tier (all agents are 'Developing' per groundtruth) and
    # the 'General Support' team marker -- verifies real content, not empty pages.
    all_body = " ".join((b or "") for _, b in agent_pages).lower()
    check("Teamly pages mention performance tier (Developing)",
          "developing" in all_body,
          "tier value not found in any page body", critical=True)
    check("Teamly pages mention team 'General Support'",
          "general support" in all_body,
          "team marker not found in page bodies")

    cur.close()
    conn.close()


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Coaching events: title carries the English marker 'Coaching Session'.
    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%coaching%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("Coaching session events exist", len(events) >= 5,
          f"Found {len(events)} coaching events", critical=True)

    if events:
        # Check they are in March 16-20, 2026
        march_events = [e for e in events
                        if e[2] and e[2].year == 2026 and e[2].month == 3
                        and 16 <= e[2].day <= 20]
        check("Coaching events in March 16-20 2026", len(march_events) >= 5,
              f"Found {len(march_events)} events in target week, total={len(events)}")

        # Check agent names appear in event summaries (RU names from ClickHouse).
        all_summaries = " ".join((e[0] or "") for e in events).lower()
        found = sum(1 for a in AGENT_NAMES if a.lower() in all_summaries)
        check("Agent names in coaching event titles", found >= 4,
              f"Found {found}/5 agent names in summaries", critical=True)

        # Check 30-minute duration
        valid_duration = 0
        for e in events:
            if e[2] and e[3]:
                delta = (e[3] - e[2]).total_seconds()
                if 25 * 60 <= delta <= 35 * 60:
                    valid_duration += 1
        check("Coaching sessions are ~30 minutes", valid_duration >= 4,
              f"{valid_duration} events have 30-min duration")

        # Descriptions should carry tier+focus context (all 'Developing').
        all_desc = " ".join((e[1] or "") for e in events).lower()
        check("Coaching event descriptions mention tier/focus",
              "developing" in all_desc or "communication skills" in all_desc,
              "no tier/focus context in event descriptions")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_teamly()
    check_gcal()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    critical_ok = len(FAILED_CRITICAL) == 0
    if not critical_ok:
        print(f"  CRITICAL FAILURES: {FAILED_CRITICAL}")

    # PASS requires: no critical failure AND accuracy >= 70.
    overall = critical_ok and accuracy >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "passed": PASS_COUNT, "failed": FAIL_COUNT,
                "accuracy": accuracy, "critical_failures": FAILED_CRITICAL,
                "success": overall,
            }, f, indent=2)

    # Critical failures hard-fail before the accuracy gate.
    if not critical_ok:
        sys.exit(1)
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
