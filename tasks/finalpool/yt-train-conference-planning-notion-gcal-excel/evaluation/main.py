"""
Evaluation for yt-train-conference-planning-notion-gcal-excel task.

Checks:
1. Conference_Planning.xlsx exists with Pre_Conference_Videos, Speaker_Travel, Summary sheets
2. Pre_Conference_Videos has 6 AI-related Fireship videos (DeepSeek top video present)
3. Speaker_Travel has 3 speakers with 818А (Moscow) and 822А (St. Petersburg) trains
4. Summary has correct conference name and total travel budget (6500 RUB)
5. Teamly page 'Technology Innovation Forum 2026' exists in the EVENTS space with content
6. 3 GCal events for conference days (Mar 12-14)
7. 3 emails sent to the three speakers with train details

A subset of semantic checks is marked CRITICAL: if any of them fails the task
FAILS immediately (sys.exit(1)) regardless of overall accuracy. This prevents
an agent that merely produces an empty/placeholder artifact from passing on the
strength of weak existence checks.
"""
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = " [CRITICAL]" if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if critical:
            CRITICAL_FAILURES.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Conference_Planning.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Conference_Planning.xlsx")
    if not os.path.exists(xlsx_path):
        record("Conference_Planning.xlsx exists", False, f"Not found at {xlsx_path}", critical=True)
        return
    record("Conference_Planning.xlsx exists", True, critical=True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e), critical=True)
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Pre_Conference_Videos sheet",
           any("video" in s or "pre" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has Speaker_Travel sheet",
           any("speaker" in s or "travel" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has Summary sheet", "summary" in sheet_names_lower, f"Sheets: {wb.sheetnames}")

    # Check Pre_Conference_Videos
    video_sheet = None
    for name in wb.sheetnames:
        if "video" in name.lower() or "pre" in name.lower():
            video_sheet = wb[name]
            break

    if video_sheet:
        rows = list(video_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Pre_Conference_Videos has 6 videos", len(data_rows) >= 6,
               f"Found {len(data_rows)} rows", critical=True)
        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        # The most-viewed AI Fireship video in the GT is the DeepSeek R1 video.
        record("Videos include DeepSeek AI content", "deepseek" in all_text,
               "No DeepSeek videos found", critical=True)
        # The real top-viewed AI video id must be present (anti-placeholder).
        record("Top video id Nl7aCUsWykg present", "nl7acuswykg" in all_text,
               "Expected real Fireship DeepSeek R1 video id")

    # Check Speaker_Travel
    speaker_sheet = None
    for name in wb.sheetnames:
        if "speaker" in name.lower() or "travel" in name.lower():
            speaker_sheet = wb[name]
            break

    if speaker_sheet:
        rows = list(speaker_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Speaker_Travel has 3 speakers", len(data_rows) >= 3,
               f"Found {len(data_rows)} rows", critical=True)
        all_text = " ".join(str(c) for row in rows for c in row if c).lower()
        record("Speaker_Travel has 818А for Moscow speakers", ("818а" in all_text or "818a" in all_text),
               "No 818А found", critical=True)
        record("Speaker_Travel has 822А for St. Petersburg speaker", ("822а" in all_text or "822a" in all_text),
               "No 822А found", critical=True)
        record("Speaker Alex Kim listed", "alex" in all_text or "kim" in all_text,
               "No Alex Kim found")
        record("Speaker James Wu listed", "james" in all_text or "wu" in all_text,
               "No James Wu found")

    # Check Summary
    if "summary" in sheet_names_lower:
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        all_text = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c).lower()
        record("Summary has conference name", "technology innovation forum" in all_text,
               "No conference name found")
        # Validate the budget numerically (sum of 2500 + 2500 + 1500 = 6500),
        # not a loose substring match.
        budget_ok = False
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if num_close(c, 6500, 0.5):
                    budget_ok = True
        record("Summary has correct travel budget (6500)", budget_ok,
               "Total_Travel_Budget_RUB != 6500", critical=True)

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Conference_Planning.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_teamly():
    print("\n=== Check 2: Teamly page 'Technology Innovation Forum 2026' (EVENTS) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("Teamly page exists", False, f"DB error: {e}", critical=True)
        return

    try:
        cur.execute("""
            SELECT p.id, p.title, p.body
              FROM teamly.pages p
              JOIN teamly.spaces s ON s.id = p.space_id
             WHERE s.key = 'EVENTS'
               AND (p.title ILIKE '%technology innovation forum%'
                    OR p.title ILIKE '%innovation forum 2026%')
             ORDER BY p.id DESC
        """)
        pages = cur.fetchall()
        record("Teamly page 'Technology Innovation Forum 2026' exists in EVENTS",
               len(pages) >= 1, f"Found {len(pages)} matching pages", critical=True)

        if pages:
            body = (pages[0][2] or "").lower()
            record("Teamly page body is non-trivial", len(body) >= 80,
                   f"Body length {len(body)}")
            # Both required content domains must appear in the page body.
            record("Teamly page mentions speaker travel",
                   "818а" in body or "822а" in body or "alex" in body or "speaker" in body,
                   "No speaker travel info in page", critical=True)
            record("Teamly page mentions curated videos",
                   "deepseek" in body or "video" in body or "fireship" in body,
                   "No video resources in page", critical=True)
    except Exception as e:
        record("Teamly page check", False, str(e), critical=True)
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 3: Google Calendar Conference Events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE start_datetime >= '2026-03-12' AND start_datetime < '2026-03-15'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()

    conf_events = [e for e in events if
                   any(kw in (e[0] or "").lower() for kw in
                       ["conference", "opening", "session", "closing", "forum"])]
    record("At least 3 conference day events scheduled", len(conf_events) >= 3,
           f"Found {len(conf_events)} conference events. All events: {[e[0] for e in events]}",
           critical=True)

    # Check Day 1 event on March 12
    day1_events = [e for e in events if e[1] and e[1].month == 3 and e[1].day == 12]
    record("Conference Day 1 event on 2026-03-12", len(day1_events) >= 1,
           f"Events on Mar 12: {[e[0] for e in day1_events]}")

    # Check Day 3 event on March 14
    day3_events = [e for e in events if e[1] and e[1].month == 3 and e[1].day == 14]
    record("Conference Day 3 event on 2026-03-14", len(day3_events) >= 1,
           f"Events on Mar 14: {[e[0] for e in day3_events]}")

    cur.close()
    conn.close()


def check_emails():
    print("\n=== Check 4: Speaker Confirmation Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE subject ILIKE '%conference%' OR subject ILIKE '%travel%' OR subject ILIKE '%confirmation%'
        """)
        emails = cur.fetchall()
        record("At least 3 confirmation emails sent", len(emails) >= 3,
               f"Found {len(emails)} conference emails", critical=True)

        # Check emails for each speaker
        all_to = " ".join(json.dumps(e[1]) if e[1] else "" for e in emails).lower()
        record("Email sent to Alex Kim (alex.kim@tech.edu)",
               "alex.kim" in all_to, f"Recipients: {all_to[:200]}", critical=True)
        record("Email sent to Maria Chen (maria.chen@research.org)",
               "maria.chen" in all_to, f"Recipients: {all_to[:200]}")
        record("Email sent to James Wu (james.wu@university.edu)",
               "james.wu" in all_to, f"Recipients: {all_to[:200]}", critical=True)

        all_bodies = " ".join((e[2] or "").lower() for e in emails)
        record("Email bodies mention train details",
               "818а" in all_bodies or "822а" in all_bodies or "13:50" in all_bodies,
               "No train details in email bodies")
    except Exception as e:
        record("Email check", False, str(e), critical=True)
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_teamly()
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILURES,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # CRITICAL gate: any failed critical check => hard fail before accuracy gate.
    if CRITICAL_FAILURES:
        print(f"\nFAIL: {len(CRITICAL_FAILURES)} critical check(s) failed: {CRITICAL_FAILURES}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
