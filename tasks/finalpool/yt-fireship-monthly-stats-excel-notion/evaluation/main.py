"""
Evaluation для yt-fireship-monthly-stats-excel-teamly (RU: notion -> teamly).

Семантический эталон пересчитывается из схемы youtube.videos (канал Fireship),
поэтому проверки не зависят от захардкоженных значений и от groundtruth-файла.

Проверки:
1. Fireship_Monthly_Stats.xlsx существует с листами Monthly_Stats и Summary.
2. Monthly_Stats: по строке на каждый активный месяц с корректными
   Month/Video_Count/Avg_Views/Avg_Likes/Total_Views (сверка с пересчётом из БД).
3. Summary: корректные пары Label/Value (Total_Videos, Peak_Month и т.д.).
4. Teamly: страница «Fireship Channel Analysis 2024-2025» существует с осмысленным
   содержимым (ключевые выводы), а не пустая.
5. Письмо на analytics@company.com с корректной темой.

CRITICAL_CHECKS: любой их провал => общий FAIL независимо от accuracy.
Иначе PASS требует accuracy >= 70%.
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

TEAMLY_PAGE_TITLE = "Fireship Channel Analysis 2024-2025"

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Семантические критические проверки: их провал = общий FAIL.
CRITICAL_CHECKS = {
    "Fireship_Monthly_Stats.xlsx exists with Monthly_Stats and Summary sheets",
    "Monthly_Stats rows match per-month stats recomputed from youtube.videos",
    "Summary Total_Videos and Peak_Month match recomputed values",
    "Teamly page 'Fireship Channel Analysis 2024-2025' exists with key findings",
    "Email sent to analytics@company.com with correct subject",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


# --------------------------------------------------------------------------- #
# Семантический эталон из youtube.videos
# --------------------------------------------------------------------------- #
def compute_groundtruth():
    """Пересчитывает помесячную статистику канала Fireship напрямую из БД."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT to_char(published_at, 'YYYY-MM') AS m,
               count(*)::int,
               round(avg(view_count::numeric))::bigint,
               round(avg(like_count::numeric))::bigint,
               sum(view_count::bigint)::bigint
        FROM youtube.videos
        WHERE channel_title ILIKE 'Fireship'
        GROUP BY m
        ORDER BY m
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    monthly = {}
    for m, cnt, avg_v, avg_l, tot in rows:
        monthly[m] = {
            "count": int(cnt),
            "avg_views": int(avg_v),
            "avg_likes": int(avg_l),
            "total_views": int(tot),
        }
    total_videos = sum(d["count"] for d in monthly.values())
    months_active = len(monthly)
    # Пик по числу видео; при равенстве берём наименьший месяц хронологически.
    peak_month = sorted(monthly.items(),
                        key=lambda kv: (-kv[1]["count"], kv[0]))[0][0]
    best_avg_month = sorted(monthly.items(),
                            key=lambda kv: (-kv[1]["avg_views"], kv[0]))[0][0]
    return {
        "monthly": monthly,
        "total_videos": total_videos,
        "months_active": months_active,
        "peak_month": peak_month,
        "peak_month_videos": monthly[peak_month]["count"],
        "best_avg_month": best_avg_month,
        "best_avg_views": monthly[best_avg_month]["avg_views"],
    }


def _num(v):
    try:
        return int(round(float(str(v).replace(",", "").strip())))
    except Exception:
        return None


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def check_excel(agent_workspace, gt):
    print("\n=== Check 1: Fireship_Monthly_Stats.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Fireship_Monthly_Stats.xlsx")
    if not os.path.exists(xlsx_path):
        record("Fireship_Monthly_Stats.xlsx exists with Monthly_Stats and Summary sheets",
               False, f"Not found at {xlsx_path}")
        return
    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Fireship_Monthly_Stats.xlsx exists with Monthly_Stats and Summary sheets",
               False, str(e))
        return

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    has_both = "monthly_stats" in sheet_names_lower and "summary" in sheet_names_lower
    record("Fireship_Monthly_Stats.xlsx exists with Monthly_Stats and Summary sheets",
           has_both, f"Sheets: {wb.sheetnames}")
    if not has_both:
        return

    # ---- Monthly_Stats: сверка построчно с пересчётом из БД ----
    ws = wb[wb.sheetnames[sheet_names_lower.index("monthly_stats")]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        record("Monthly_Stats rows match per-month stats recomputed from youtube.videos",
               False, "empty sheet")
        return

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]

    def col_idx(*keys):
        for i, h in enumerate(headers):
            if all(k in h for k in keys):
                return i
        return None

    i_month = col_idx("month")
    i_count = col_idx("video", "count") or col_idx("count")
    i_avgv = col_idx("avg", "view")
    i_avgl = col_idx("avg", "like")
    i_totv = col_idx("total", "view")

    data_rows = [r for r in rows[1:] if any(c is not None and str(c) != "" for c in r)]
    record("Monthly_Stats row count equals active months",
           len(data_rows) == gt["months_active"],
           f"Found {len(data_rows)}, expected {gt['months_active']}")

    # Построим словарь агента month -> кортеж значений.
    agent = {}
    if i_month is not None:
        for r in data_rows:
            m = str(r[i_month]).strip() if r[i_month] is not None else ""
            agent[m] = r

    mismatches = []
    matched_months = 0
    for m, d in gt["monthly"].items():
        r = agent.get(m)
        if r is None:
            mismatches.append(f"{m}: missing")
            continue
        ok = True
        bad = []
        if i_count is not None and _num(r[i_count]) != d["count"]:
            ok = False
            bad.append(f"cnt={_num(r[i_count])} want {d['count']}")
        # ±1: PG round() is half-away-from-zero, python round() is banker's
        if i_avgv is not None and abs(_num(r[i_avgv]) - d["avg_views"]) > 1:
            ok = False
            bad.append(f"avgv={_num(r[i_avgv])} want {d['avg_views']}")
        if i_avgl is not None and abs(_num(r[i_avgl]) - d["avg_likes"]) > 1:
            ok = False
            bad.append(f"avgl={_num(r[i_avgl])} want {d['avg_likes']}")
        if i_totv is not None and _num(r[i_totv]) != d["total_views"]:
            ok = False
            bad.append(f"totv={_num(r[i_totv])} want {d['total_views']}")
        if ok:
            matched_months += 1
        else:
            mismatches.append(f"{m}: " + ", ".join(bad))

    cols_present = all(x is not None for x in (i_month, i_count, i_avgv, i_avgl, i_totv))
    record("Monthly_Stats has all required columns",
           cols_present, f"Headers: {rows[0]}")
    record("Monthly_Stats rows match per-month stats recomputed from youtube.videos",
           cols_present and matched_months == gt["months_active"],
           "; ".join(mismatches[:6]) or f"matched {matched_months}/{gt['months_active']}")

    # Ascending chronological order.
    months_seq = [str(r[i_month]).strip() for r in data_rows] if i_month is not None else []
    record("Monthly_Stats sorted ascending by Month",
           months_seq == sorted(months_seq), f"{months_seq}")

    # ---- Summary ----
    ws2 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
    rows2 = list(ws2.iter_rows(values_only=True))
    kv = {}
    for r in rows2:
        if r and r[0] is not None:
            key = str(r[0]).strip().lower()
            val = r[1] if len(r) > 1 else None
            kv[key] = val

    total_ok = _num(kv.get("total_videos")) == gt["total_videos"]
    peak_ok = (str(kv.get("peak_month", "")).strip() == gt["peak_month"])
    record("Summary Total_Videos and Peak_Month match recomputed values",
           total_ok and peak_ok,
           f"Total_Videos got={kv.get('total_videos')} want={gt['total_videos']}; "
           f"Peak_Month got={kv.get('peak_month')} want={gt['peak_month']}")

    months_active_ok = _num(kv.get("total_months_active")) == gt["months_active"]
    best_month_ok = (str(kv.get("best_avg_views_month", "")).strip() == gt["best_avg_month"])
    record("Summary Total_Months_Active and Best_Avg_Views_Month correct",
           months_active_ok and best_month_ok,
           f"months_active got={kv.get('total_months_active')} want={gt['months_active']}; "
           f"best_month got={kv.get('best_avg_views_month')} want={gt['best_avg_month']}")


# --------------------------------------------------------------------------- #
# Teamly
# --------------------------------------------------------------------------- #
def check_teamly(gt):
    print("\n=== Check 2: Teamly page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE %s
        """, (f"%{TEAMLY_PAGE_TITLE}%",))
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Teamly page 'Fireship Channel Analysis 2024-2025' exists with key findings",
               False, str(e))
        return

    if not pages:
        record("Teamly page 'Fireship Channel Analysis 2024-2025' exists with key findings",
               False, "no matching page")
        return

    # Содержимое должно отражать ключевые выводы: общее число видео и пиковый месяц.
    best = None
    for title, body in pages:
        text = f"{title} {body}".lower()
        score = 0
        if str(gt["total_videos"]) in text:
            score += 1
        if gt["peak_month"].lower() in text:
            score += 1
        if gt["best_avg_month"].lower() in text:
            score += 1
        if str(gt["months_active"]) in text:
            score += 1
        if best is None or score > best[0]:
            best = (score, title, len(body))

    # Требуем хотя бы 2 из 4 ключевых фактов и непустое тело.
    ok = best is not None and best[0] >= 2 and best[2] > 0
    record("Teamly page 'Fireship Channel Analysis 2024-2025' exists with key findings",
           ok, f"best score={best[0] if best else 0}, body_len={best[2] if best else 0}")


# --------------------------------------------------------------------------- #
# Email
# --------------------------------------------------------------------------- #
def check_email():
    print("\n=== Check 3: Email sent ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT m.to_addr, m.subject FROM email.messages m
        JOIN email.sent_log sl ON sl.message_id = m.id
        WHERE m.to_addr::text ILIKE %s
        ORDER BY sl.sent_at DESC LIMIT 5
    """, ("%analytics@company.com%",))
    emails = cur.fetchall()
    if not emails:
        cur.execute("""
            SELECT to_addr, subject FROM email.messages
            WHERE to_addr::text ILIKE %s
            ORDER BY date DESC LIMIT 5
        """, ("%analytics@company.com%",))
        emails = cur.fetchall()
    cur.close()
    conn.close()

    subj_ok = False
    if emails:
        for _, subject in emails:
            s = str(subject or "").lower()
            if "fireship" in s and "analysis" in s:
                subj_ok = True
                break
            if "channel analysis complete" in s:
                subj_ok = True
                break

    record("Email sent to analytics@company.com with correct subject",
           len(emails) >= 1 and subj_ok,
           f"Found: {emails}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    gt = compute_groundtruth()
    print(f"[eval] Эталон из youtube.videos: total={gt['total_videos']}, "
          f"months={gt['months_active']}, peak={gt['peak_month']}, "
          f"best_avg_month={gt['best_avg_month']}")

    check_excel(args.agent_workspace, gt)
    check_teamly(gt)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Критический гейт ДО порога точности.
    if critical_failed:
        print("FAIL (critical)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
