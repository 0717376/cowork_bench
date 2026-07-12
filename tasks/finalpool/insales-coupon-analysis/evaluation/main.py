"""Evaluation for insales-coupon-analysis (InSales store).

Структурные проверки (наличие листов, наличие строк) — НЕ критичные.
Содержательные проверки — КРИТИЧНЫЕ: провал любой => FAIL независимо от accuracy:
  - Summary.Most_Used_Code совпадает с эталоном (ключевой результат, точное сравнение строки);
  - Summary.Total_Coupons и Total_Usage совпадают с эталоном (tol<=1);
  - для каждого купона Usage_Count (tol=1) И вычисленный Utilization_Pct (tol=0.5)
    совпадают с эталоном — проверяет правило деления на лимит.
Эталон gt уже синхронизирован с централизованно русифицированной схемой wc.*.
Коды купонов, имена колонок и листов — английские (eval их не переводит).
"""
import argparse
import os
import sys
import openpyxl


# Содержательные (обязательные) проверки. Провал любой => итог FAIL.
CRITICAL_CHECKS = {
    "CRITICAL: Summary.Most_Used_Code совпадает с эталоном",
    "CRITICAL: Summary.Total_Coupons совпадает с эталоном",
    "CRITICAL: Summary.Total_Usage совпадает с эталоном",
    "CRITICAL: Coupon Analysis Usage_Count+Utilization_Pct совпадают с эталоном",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def summary_lookup(rows, keys):
    """Header/orientation-tolerant {metric_lower: value} map for a Summary sheet.

    Scans ALL cells; a metric's value is the cell to its right, or the cell
    below when the right neighbour is itself another metric key (horizontal
    layout). A 'Metric/Value' header never collides with metric keys.
    """
    keys = {str(k).strip().lower() for k in keys}
    out = {}
    rows = rows or []

    def _is_key(v):
        return v is not None and str(v).strip().lower() in keys

    for r, row in enumerate(rows):
        for c, cell in enumerate(row or []):
            if cell is None:
                continue
            k = str(cell).strip().lower()
            if k not in keys or k in out:
                continue
            right = row[c + 1] if c + 1 < len(row) else None
            below = None
            if r + 1 < len(rows) and rows[r + 1] and c < len(rows[r + 1]):
                below = rows[r + 1][c]
            if right is not None and not _is_key(right):
                out[k] = right
            elif below is not None and not _is_key(below):
                out[k] = below
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Coupon_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Coupon_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- Coupon Analysis ----
    print("  Checking Coupon Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Coupon Analysis")
    g_rows = load_sheet_rows(gt_wb, "Coupon Analysis")
    check("Лист 'Coupon Analysis' присутствует", a_rows is not None,
          "sheet missing in agent output")
    if a_rows is not None and g_rows is not None:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Coupon Analysis содержит строки", len(a_data) > 0, f"rows={len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        # Структурная (не критичная): Amount соответствует эталону.
        amount_errors = []
        # Содержательная (критичная): Usage_Count + Utilization_Pct.
        usage_util_errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                usage_util_errors.append(f"Missing row: {g_row[0]}")
                continue
            # Amount (col 2) — структурная.
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1.0):
                    amount_errors.append(f"{key}.Amount: {a_row[2]} vs {g_row[2]}")
            # Usage_Count (col 3) — критичная, tol=1.
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1):
                    usage_util_errors.append(
                        f"{key}.Usage_Count: {a_row[3]} vs {g_row[3]} (tol=1)")
            # Utilization_Pct (col 5) — критичная, tol=0.5, правило деления на лимит.
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0.5):
                    usage_util_errors.append(
                        f"{key}.Utilization_Pct: {a_row[5]} vs {g_row[5]} (tol=0.5)")

        check("Coupon Analysis Amount совпадает с эталоном",
              not amount_errors, amount_errors[:5])
        check("CRITICAL: Coupon Analysis Usage_Count+Utilization_Pct совпадают с эталоном",
              not usage_util_errors, usage_util_errors[:5])

    # ---- Summary ----
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    check("Лист 'Summary' присутствует", a_rows is not None,
          "sheet missing in agent output")
    if a_rows is not None and g_rows is not None:
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        g_lookup = {}
        for row in g_data:
            if row and row[0] is not None:
                g_lookup[str(row[0]).strip().lower()] = (row[1] if len(row) > 1 else None)
        a_lookup = summary_lookup(a_rows, g_lookup.keys())

        # Most_Used_Code — критичная, точное сравнение строки.
        check("CRITICAL: Summary.Most_Used_Code совпадает с эталоном",
              str_match(a_lookup.get("most_used_code"), g_lookup.get("most_used_code")),
              f"{a_lookup.get('most_used_code')} vs {g_lookup.get('most_used_code')}")

        # Total_Coupons / Total_Usage — критичные, tol=1.
        check("CRITICAL: Summary.Total_Coupons совпадает с эталоном",
              num_close(a_lookup.get("total_coupons"), g_lookup.get("total_coupons"), 1.0),
              f"{a_lookup.get('total_coupons')} vs {g_lookup.get('total_coupons')}")
        check("CRITICAL: Summary.Total_Usage совпадает с эталоном",
              num_close(a_lookup.get("total_usage"), g_lookup.get("total_usage"), 1.0),
              f"{a_lookup.get('total_usage')} vs {g_lookup.get('total_usage')}")

        # Avg_Utilization — структурная (мягкий показатель), tol=5.0.
        if "avg_utilization" in g_lookup:
            check("Summary.Avg_Utilization близко к эталону",
                  num_close(a_lookup.get("avg_utilization"), g_lookup.get("avg_utilization"), 5.0),
                  f"{a_lookup.get('avg_utilization')} vs {g_lookup.get('avg_utilization')}")

    # ---- Coupon_Strategy.docx ----
    docx_path = os.path.join(args.agent_workspace, "Coupon_Strategy.docx")
    if not os.path.exists(docx_path):
        check("Coupon_Strategy.docx существует", False, docx_path)
    else:
        check("Coupon_Strategy.docx существует", True)
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs).lower()
            _headings = " ".join(
                p.text for p in _doc.paragraphs if p.style.name.startswith("Heading")).lower()
            check("Coupon_Strategy.docx содержит текст (>=50 символов)",
                  len(_text.strip()) >= 50, f"len={len(_text.strip())}")
            # Группы ключевых слов RU+EN: достаточно одного слова из каждой группы.
            _kw_groups = [["coupon", "купон"], ["strategy", "стратеги", "рекомендаци"]]
            _missing = []
            for grp in _kw_groups:
                if not any(k in _text or k in _headings for k in grp):
                    _missing.append(grp)
            # Слабая проверка: провал только если ВСЕ группы отсутствуют.
            check("Coupon_Strategy.docx содержит ожидаемые ключевые слова (RU+EN)",
                  len(_missing) < len(_kw_groups), f"missing={_missing}")
        except ImportError:
            check("Coupon_Strategy.docx достаточного размера",
                  os.path.getsize(docx_path) >= 100, f"size={os.path.getsize(docx_path)}")
        except Exception as _e:
            check("Coupon_Strategy.docx читается", False, str(_e))

    # ---- Gate ----
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"Critical fails: {CRITICAL_FAILS}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
