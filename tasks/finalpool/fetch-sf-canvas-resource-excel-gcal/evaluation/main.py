"""Evaluation script for fetch-sf-canvas-resource-excel-gcal (ClickHouse / RU)."""
import os
import argparse, json, os, sys
from datetime import datetime, timezone, timedelta
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# data.json industry_avg benchmarks keyed by English department name.
BENCHMARKS = {
    "Engineering": 66816, "Finance": 79324, "HR": 67899, "Operations": 72812,
    "R&D": 74771, "Sales": 77904, "Support": 61709,
}
# Central russification map (db/zzz_clickhouse_after_init.sql).
RU2EN = {
    "инженерия": "Engineering", "финансы": "Finance", "кадры": "HR",
    "операции": "Operations", "ниокр": "R&D", "продажи": "Sales",
    "поддержка": "Support",
}
EN_SET = {k.lower() for k in BENCHMARKS}
# Expected primary-dimension set (RU canonical) sorted alphabetically.
RU_DEPTS_SORTED = ["Инженерия", "Кадры", "НИОКР", "Операции", "Поддержка", "Продажи", "Финансы"]


def dept_to_en(name):
    """Map a department label (RU or EN) to its canonical English key."""
    if name is None:
        return None
    s = str(name).strip()
    low = s.lower()
    if low in RU2EN:
        return RU2EN[low]
    if low in EN_SET:
        for k in BENCHMARKS:
            if k.lower() == low:
                return k
    return None


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [{'CRIT-' if critical else ''}PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [{'CRIT-' if critical else ''}FAIL] {name}: {detail_str}")
        if critical:
            CRITICAL_FAILS.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def find_col(headers, *aliases):
    """Return index of first header matching any alias (case-insensitive)."""
    low = [h.lower() for h in headers]
    for a in aliases:
        if a.lower() in low:
            return low.index(a.lower())
    return -1


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    excel_path = os.path.join(agent_workspace, "Canvas_Resource_Report.xlsx")
    check("Canvas_Resource_Report.xlsx exists", os.path.exists(excel_path))

    da_rows = []          # parsed (en_dept, internal, benchmark, gap)
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---------- Data_Analysis ----------
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            # Primary dimension header (Department / Подразделение), plus Internal/Benchmark/Gap.
            i_dept = find_col(headers, "Department", "Подразделение")
            i_int = find_col(headers, "Internal", "Внутреннее", "Внутр")
            i_bench = find_col(headers, "Benchmark", "Эталон")
            i_gap = find_col(headers, "Gap", "Разрыв", "Разница")
            check("Data_Analysis has Department column", i_dept >= 0, f"headers: {headers}")
            check("Data_Analysis has Internal column", i_int >= 0, f"headers: {headers}")
            check("Data_Analysis has Benchmark column", i_bench >= 0, f"headers: {headers}")
            check("Data_Analysis has Gap column", i_gap >= 0, f"headers: {headers}")

            if min(i_dept, i_int, i_bench, i_gap) >= 0:
                seen = []
                for r in data_rows:
                    if r is None or all(v is None for v in r):
                        continue
                    en = dept_to_en(r[i_dept])
                    internal = safe_float(r[i_int])
                    bench = safe_float(r[i_bench])
                    gap = safe_float(r[i_gap])
                    seen.append((str(r[i_dept]).strip() if r[i_dept] is not None else "", en, internal, bench, gap))
                    if en is not None:
                        da_rows.append((en, internal, bench, gap))

                # CRITICAL 1: 7 departments matching the russified sf_data set, sorted alphabetically.
                labels = [s[0] for s in seen if s[0]]
                en_keys = sorted([s[1] for s in seen if s[1] is not None])
                unique_en = sorted(set(en_keys))
                check("CRITICAL: 7 departments match sf_data set",
                      unique_en == sorted(BENCHMARKS.keys()),
                      f"got: {unique_en}", critical=True)
                check("CRITICAL: Department column sorted alphabetically",
                      labels == sorted(labels),
                      f"labels: {labels}", critical=True)

                # CRITICAL 2: Benchmark == data.json industry_avg for that department.
                bench_ok = True
                bench_detail = ""
                for en, internal, bench, gap in da_rows:
                    exp = BENCHMARKS.get(en)
                    if exp is None or bench is None or abs(bench - exp) > 1.0:
                        bench_ok = False
                        bench_detail = f"{en}: got {bench}, expected {exp}"
                        break
                check("CRITICAL: Benchmark equals data.json industry_avg per row",
                      bench_ok and len(da_rows) == 7, bench_detail, critical=True)

                # CRITICAL 3: Gap == Internal - Benchmark per row.
                gap_ok = True
                gap_detail = ""
                for en, internal, bench, gap in da_rows:
                    if internal is None or bench is None or gap is None:
                        gap_ok = False
                        gap_detail = f"{en}: missing value"
                        break
                    if abs(gap - (internal - bench)) > 1.0:
                        gap_ok = False
                        gap_detail = f"{en}: gap {gap} != {internal}-{bench}={internal-bench}"
                        break
                check("CRITICAL: Gap == Internal - Benchmark per row",
                      gap_ok and len(da_rows) == 7, gap_detail, critical=True)

        # ---------- Metrics ----------
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            mrows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(mrows) >= 4, f"got {len(mrows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Metric', 'Value']:
                check(f"Metrics has {col} column", col.lower() in headers, f"headers: {headers[:8]}")

        # ---------- Recommendations ----------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        rec_rows = []
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            rec_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")
            headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
            i_pri = find_col(headers, "Priority", "Приоритет")
            i_act = find_col(headers, "Action", "Действие")
            i_dep = find_col(headers, "Department", "Подразделение")
            check("Recommendations has Priority column", i_pri >= 0, f"headers: {headers}")
            check("Recommendations has Action column", i_act >= 0, f"headers: {headers}")
            check("Recommendations has Department column", i_dep >= 0, f"headers: {headers}")

            # CRITICAL 5: top-priority Action targets the largest-negative-gap department.
            if da_rows and i_pri >= 0 and i_dep >= 0:
                neg = sorted([r for r in da_rows if r[3] is not None], key=lambda r: r[3])
                worst_en = neg[0][0] if neg else None
                # find priority==1 row (or first data row)
                top_dep_en = None
                for r in rec_rows:
                    if r is None:
                        continue
                    p = safe_float(r[i_pri])
                    if p == 1:
                        top_dep_en = dept_to_en(r[i_dep])
                        break
                if top_dep_en is None and rec_rows:
                    top_dep_en = dept_to_en(rec_rows[0][i_dep])
                check("CRITICAL: top-priority action targets largest-negative-gap department",
                      top_dep_en is not None and top_dep_en == worst_en,
                      f"top={top_dep_en}, expected={worst_en}", critical=True)

        # ---------- processor ----------
        check("sf_canvas_resource_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "sf_canvas_resource_processor.py")))

    # ---------- Calendar (CRITICAL date/time) ----------
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT summary, start_datetime, end_datetime FROM gcal.events "
            "WHERE summary ILIKE %s OR summary ILIKE %s",
            ('%analysis review%', '%review%'))
        events = cur.fetchall()
        conn.close()
        check("Review event created", len(events) >= 1, f"found {len(events)} events")

        target_start = datetime(2026, 3, 14, 14, 0, tzinfo=timezone.utc)
        target_end = datetime(2026, 3, 14, 15, 0, tzinfo=timezone.utc)
        matched = False
        detail = ""
        for summ, sdt, edt in events:
            if sdt is None or edt is None:
                continue
            s = sdt.astimezone(timezone.utc) if sdt.tzinfo else sdt.replace(tzinfo=timezone.utc)
            e = edt.astimezone(timezone.utc) if edt.tzinfo else edt.replace(tzinfo=timezone.utc)
            if abs((s - target_start).total_seconds()) <= 60 and abs((e - target_end).total_seconds()) <= 60:
                if "review" in str(summ).lower():
                    matched = True
                    break
            detail = f"{summ}: {s.isoformat()} - {e.isoformat()}"
        check("CRITICAL: 'Analysis Review' on 2026-03-14 14:00-15:00 UTC",
              matched, detail, critical=True)
    except Exception as e:
        check("CRITICAL: calendar check", False, str(e), critical=True)

    # ---------- gates ----------
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"Accuracy: {accuracy:.1f}% ({PASS_COUNT}/{total})")
    if CRITICAL_FAILS:
        print(f"CRITICAL FAILURES: {CRITICAL_FAILS}")
        sys.exit(1)
    success = accuracy >= 70.0
    return success, f"Passed {PASS_COUNT}/{total} checks (accuracy {accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
