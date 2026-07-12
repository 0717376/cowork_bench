"""Evaluation для yf-macro-investment-report (RU / moex-finance).

Агент строит Macro_Investment.xlsx и Investment_Report.docx по данным MCP
`moex-finance` (схема moex.*) и макроиндикаторам из mock-API. Портфель из пяти
тикеров MOEX: SBER.ME, GAZP.ME, LKOH.ME, MGNT.ME, MTSS.ME.

Эталонные значения пересчитаны из сида moex.stock_prices (окно 2026-02-25 ..
2026-05-26) и moex.stock_info (trailingPE). Доходности за 90 дней:
  GAZP.ME +9.25, LKOH.ME -1.65, MGNT.ME -1.78, MTSS.ME -3.59, SBER.ME -5.12.
  Average_90d_Return = -0.58, Best = GAZP.ME, Worst = SBER.ME, Avg_Trailing_PE = 5.25.

CRITICAL_CHECKS (семантика): любой их провал => общий FAIL независимо от
accuracy. Структурные проверки (наличие листов, заголовков) — не критические.
"""
import argparse
import os
import sys

# Эталон, пересчитанный из сида moex (НЕ переносить старые US-числа).
EXPECTED_SYMBOLS = ["GAZP.ME", "LKOH.ME", "MGNT.ME", "MTSS.ME", "SBER.ME"]
# Доходности за 90 дней (2dp) по сиду moex.stock_prices.
EXPECTED_RETURNS = {
    "GAZP.ME": 9.25,
    "LKOH.ME": -1.65,
    "MGNT.ME": -1.78,
    "MTSS.ME": -3.59,
    "SBER.ME": -5.12,
}
EXPECTED_BEST = "GAZP.ME"
EXPECTED_WORST = "SBER.ME"
EXPECTED_AVG_RETURN = -0.58
EXPECTED_AVG_PE = 5.25
# Макроиндикаторы (English keys — eval их грепает) и ожидаемые Impact_Assessment.
EXPECTED_MACRO = {
    "us_10y_yield": (4.25, "headwind"),
    "fed_funds_rate": (5.25, "restrictive"),
    "cpi_yoy": (3.1, "inflationary"),
    "unemployment": (3.7, "tight labor"),
    "gdp_growth_q4": (2.8, "supportive"),
    "sp500_pe_ratio": (22.5, "above historical"),
    "vix": (18.3, "low volatility"),
}
EXPECTED_RISK_SCORE = 3  # headwind + restrictive + inflationary

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = False

CRITICAL_CHECKS = {
    "Stock Performance: ровно 5 тикеров MOEX (отсортированы), Current_Price и Trailing_PE заданы",
    "Stock Performance: Return_90d_Pct совпадает с эталоном (>=2 тикеров)",
    "Portfolio Summary: Best_Performer и Worst_Performer соответствуют эталону",
    "Macro Context: 7 индикаторов со значениями и корректным Impact_Assessment",
    "Investment_Report.docx: содержательный отчёт (macro/portfolio/sector, RU+EN)",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        tag = " (CRITICAL)" if name in CRITICAL_CHECKS else ""
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED = True


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    import openpyxl
    path = os.path.join(agent_workspace, "Macro_Investment.xlsx")
    if not os.path.exists(path):
        record("Macro_Investment.xlsx существует", False, "файл не найден")
        # критические проверки, зависящие от файла, тоже фейлят
        for n in [
            "Stock Performance: ровно 5 тикеров MOEX (отсортированы), Current_Price и Trailing_PE заданы",
            "Stock Performance: Return_90d_Pct совпадает с эталоном (>=2 тикеров)",
            "Portfolio Summary: Best_Performer и Worst_Performer соответствуют эталону",
            "Macro Context: 7 индикаторов со значениями и корректным Impact_Assessment",
        ]:
            record(n, False, "нет xlsx")
        return
    record("Macro_Investment.xlsx существует", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        record("Macro_Investment.xlsx читается", False, str(e))
        return
    record("Macro_Investment.xlsx читается", True)

    # ---- Stock Performance ----
    rows = load_sheet_rows(wb, "Stock Performance")
    if rows is None:
        record("Лист 'Stock Performance' существует", False)
        record("Stock Performance: ровно 5 тикеров MOEX (отсортированы), Current_Price и Trailing_PE заданы", False, "нет листа")
        record("Stock Performance: Return_90d_Pct совпадает с эталоном (>=2 тикеров)", False, "нет листа")
    else:
        record("Лист 'Stock Performance' существует", True)
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        by_sym = {}
        for r in data_rows:
            by_sym[str(r[0]).strip().upper()] = r
        col = {h.lower(): i for i, h in enumerate(header)}

        # CRITICAL: ровно 5 ожидаемых тикеров, по порядку (алфавит), цена и PE заданы
        present = [str(r[0]).strip().upper() for r in data_rows]
        sorted_ok = present == sorted(set(present))
        symbols_ok = set(present) == set(EXPECTED_SYMBOLS) and len(data_rows) == 5
        cp_i = col.get("current_price")
        pe_i = col.get("trailing_pe")
        prices_ok = True
        if cp_i is not None and pe_i is not None:
            for sym in EXPECTED_SYMBOLS:
                r = by_sym.get(sym)
                if not r or cp_i >= len(r) or pe_i >= len(r):
                    prices_ok = False
                    break
                if r[cp_i] is None or r[pe_i] is None:
                    prices_ok = False
                    break
        else:
            prices_ok = False
        record(
            "Stock Performance: ровно 5 тикеров MOEX (отсортированы), Current_Price и Trailing_PE заданы",
            symbols_ok and sorted_ok and prices_ok,
            f"present={present} sorted={sorted_ok} prices_pe_ok={prices_ok}",
        )

        # CRITICAL: Return_90d_Pct совпадает с эталоном минимум для 2 тикеров
        ret_i = col.get("return_90d_pct")
        matched = 0
        details = []
        if ret_i is not None:
            for sym, exp in EXPECTED_RETURNS.items():
                r = by_sym.get(sym)
                if r and ret_i < len(r) and r[ret_i] is not None:
                    if num_close(r[ret_i], exp, abs_tol=0.15, rel_tol=0.0):
                        matched += 1
                    else:
                        details.append(f"{sym}={r[ret_i]} (ожид {exp})")
        record(
            "Stock Performance: Return_90d_Pct совпадает с эталоном (>=2 тикеров)",
            matched >= 2,
            f"совпало {matched}/5; расхождения: {details}",
        )

    # ---- Macro Context ----
    rows2 = load_sheet_rows(wb, "Macro Context")
    if rows2 is None:
        record("Лист 'Macro Context' существует", False)
        record("Macro Context: 7 индикаторов со значениями и корректным Impact_Assessment", False, "нет листа")
    else:
        record("Лист 'Macro Context' существует", True)
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        record("Macro Context: 7 строк индикаторов", len(data2) >= 7, f"{len(data2)} строк")
        lookup = {str(r[0]).strip().lower(): r for r in data2 if r[0]}
        good = 0
        bad = []
        for key, (val, impact_kw) in EXPECTED_MACRO.items():
            r = lookup.get(key)
            if not r:
                bad.append(f"{key}: нет строки")
                continue
            v_ok = len(r) > 1 and num_close(r[1], val, abs_tol=0.1, rel_tol=0.0)
            impact_text = " ".join(str(c).lower() for c in r[2:] if c is not None)
            i_ok = impact_kw in impact_text
            if v_ok and i_ok:
                good += 1
            else:
                bad.append(f"{key}: value_ok={v_ok} impact_ok={i_ok}")
        record(
            "Macro Context: 7 индикаторов со значениями и корректным Impact_Assessment",
            good == 7,
            f"корректно {good}/7; проблемы: {bad}",
        )

    # ---- Portfolio Summary ----
    rows3 = load_sheet_rows(wb, "Portfolio Summary")
    if rows3 is None:
        record("Лист 'Portfolio Summary' существует", False)
        record("Portfolio Summary: Best_Performer и Worst_Performer соответствуют эталону", False, "нет листа")
    else:
        record("Лист 'Portfolio Summary' существует", True)
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        lk = {str(r[0]).strip().lower(): (r[1] if len(r) > 1 else None) for r in data3 if r[0]}

        record(
            "Portfolio Summary: Average_90d_Return ~ эталон",
            "average_90d_return" in lk and num_close(lk["average_90d_return"], EXPECTED_AVG_RETURN, abs_tol=0.2, rel_tol=0.0),
            f"={lk.get('average_90d_return')} (ожид {EXPECTED_AVG_RETURN})",
        )
        record(
            "Portfolio Summary: Avg_Trailing_PE ~ эталон",
            "avg_trailing_pe" in lk and num_close(lk["avg_trailing_pe"], EXPECTED_AVG_PE, abs_tol=0.3, rel_tol=0.0),
            f"={lk.get('avg_trailing_pe')} (ожид {EXPECTED_AVG_PE})",
        )
        record(
            "Portfolio Summary: Macro_Risk_Score == 3",
            "macro_risk_score" in lk and num_close(lk["macro_risk_score"], EXPECTED_RISK_SCORE, abs_tol=0.0, rel_tol=0.0),
            f"={lk.get('macro_risk_score')} (ожид {EXPECTED_RISK_SCORE})",
        )
        best = str(lk.get("best_performer", "")).strip().upper()
        worst = str(lk.get("worst_performer", "")).strip().upper()
        record(
            "Portfolio Summary: Best_Performer и Worst_Performer соответствуют эталону",
            best == EXPECTED_BEST and worst == EXPECTED_WORST,
            f"best={best} (ожид {EXPECTED_BEST}), worst={worst} (ожид {EXPECTED_WORST})",
        )


def check_word(agent_workspace):
    path = os.path.join(agent_workspace, "Investment_Report.docx")
    if not os.path.exists(path):
        record("Investment_Report.docx: содержательный отчёт (macro/portfolio/sector, RU+EN)", False, "файл не найден")
        return
    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join([p.text for p in doc.paragraphs]).lower()
    except Exception as e:
        record("Investment_Report.docx: содержательный отчёт (macro/portfolio/sector, RU+EN)", False, str(e))
        return

    long_ok = len(text) >= 200
    # RU+EN ключевые понятия: макро / портфель / сектор
    kw_groups = [
        ["macro", "макро"],
        ["portfolio", "портфел"],
        ["sector", "сектор", "отрасл"],
    ]
    missing = []
    for grp in kw_groups:
        if not any(k in text for k in grp):
            missing.append("/".join(grp))
    record(
        "Investment_Report.docx: содержательный отчёт (macro/portfolio/sector, RU+EN)",
        long_ok and not missing,
        f"len={len(text)} missing={missing}",
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace")

    print("  Проверка Excel...")
    check_excel(agent_ws)
    print("  Проверка Word...")
    check_word(agent_ws)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    if CRITICAL_FAILED:
        print("=== RESULT: FAIL (провалена критическая проверка) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print("=== RESULT: FAIL (accuracy < 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
