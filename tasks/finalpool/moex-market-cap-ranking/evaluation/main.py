"""Evaluation for moex-market-cap-ranking.

The agent pulls stock info for our tracked MOEX tickers (via moex-finance),
ranks companies by market capitalization in billions of RUB, and writes
YF_Market_Cap.xlsx with two sheets:
  - "Market Cap Ranking": Symbol, Name, Sector, Market_Cap_B, Recommendation,
    sorted by Market_Cap_B descending.
  - "Summary": Metric / Value rows Total_Companies, Largest, Smallest,
    Total_Market_Cap_B.
The agent also creates a teamly page titled "Market Cap Rankings".

Critical checks (see CRITICAL_CHECKS): any failure => overall FAIL regardless
of accuracy. Otherwise pass threshold: accuracy >= 70%.

Groundtruth (groundtruth_workspace/YF_Market_Cap.xlsx) is regenerated from the
deterministic moex.* seed (6 .ME tickers). marketCap is raw RUB; /1e9 ->
billions of RUB. Tolerances on Market_Cap_B are RELATIVE (~2%), not absolute.
"""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=5432,
    dbname=os.environ.get("PGDATABASE", "cowork_gym"),
    user="eigent",
    password="camel",
)

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "All tracked tickers present with correct Market_Cap_B",
    "Market Cap Ranking sorted by Market_Cap_B descending",
    "Sector matches moex stock_info for each ticker",
    "Summary Largest / Smallest / Total match agent data",
    "Teamly 'Market Cap Rankings' page exists and names the largest company",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        d = (str(detail)[:300]) if detail else ""
        print(f"  [FAIL] {name}: {d}")


def parse_float(cell):
    try:
        return float(str(cell).replace(",", "").replace(" ", "").strip())
    except (ValueError, AttributeError, TypeError):
        return None


def rel_close(a, b, rel=0.02, abs_floor=1.0):
    fa, fb = parse_float(a), parse_float(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= max(abs_floor, abs(fb) * rel)


def str_eq(a, b):
    return str(a or "").strip().lower() == str(b or "").strip().lower()


LEGAL_FORMS = {"пао", "мкпао", "оао", "ао", "пк", "pjsc", "ojsc", "jsc"}


def name_eq(a, b):
    """Variant-tolerant company-name match (shortName vs longName, ± legal form)."""
    if str_eq(a, b):
        return True
    ta = [t for t in str(a or "").lower().replace('"', " ").split() if t not in LEGAL_FORMS]
    tb = [t for t in str(b or "").lower().replace('"', " ").split() if t not in LEGAL_FORMS]
    if not ta or not tb:
        return False
    if ta == tb or set(tb) <= set(ta) or set(ta) <= set(tb):
        return True
    # distinctive last token containment ("т-технологии" in "мкт-технологии")
    return tb[-1] in " ".join(ta) or ta[-1] in " ".join(tb)


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_col(header, *aliases):
    norm = {str(h).strip().lower(): i for i, h in enumerate(header) if h is not None}
    for a in aliases:
        if a.lower() in norm:
            return norm[a.lower()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "YF_Market_Cap.xlsx")
    gt_file = os.path.join(gt_dir, "YF_Market_Cap.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- Build groundtruth ranking lookup ----
    g_rows = load_sheet_rows(gt_wb, "Market Cap Ranking")
    g_header = g_rows[0]
    gi_sym = get_col(g_header, "Symbol")
    gi_name = get_col(g_header, "Name")
    gi_sec = get_col(g_header, "Sector")
    gi_mc = get_col(g_header, "Market_Cap_B")
    gt = {}
    for r in g_rows[1:]:
        if not r or r[gi_sym] is None:
            continue
        gt[str(r[gi_sym]).strip().lower()] = {
            "symbol": str(r[gi_sym]).strip(),
            "name": r[gi_name],
            "sector": r[gi_sec],
            "mc": parse_float(r[gi_mc]),
        }
    gt_total = round(sum(v["mc"] for v in gt.values() if v["mc"] is not None), 2)
    gt_sorted = sorted(gt.values(), key=lambda v: v["mc"], reverse=True)
    gt_largest = gt_sorted[0]["name"]
    gt_smallest = gt_sorted[-1]["name"]

    # ---- Load agent ranking sheet ----
    print("\n=== Checking Market Cap Ranking ===")
    a_rows = load_sheet_rows(agent_wb, "Market Cap Ranking")
    if a_rows is None or len(a_rows) < 2:
        for c in ("All tracked tickers present with correct Market_Cap_B",
                  "Market Cap Ranking sorted by Market_Cap_B descending",
                  "Sector matches moex stock_info for each ticker"):
            check(c, False, "Sheet 'Market Cap Ranking' missing or empty")
    else:
        a_header = a_rows[0]
        ai_sym = get_col(a_header, "Symbol")
        ai_sec = get_col(a_header, "Sector")
        ai_mc = get_col(a_header, "Market_Cap_B")
        a_data = [r for r in a_rows[1:] if r and ai_sym is not None and r[ai_sym] is not None]
        a_lookup = {str(r[ai_sym]).strip().lower(): r for r in a_data}

        # CRITICAL: each gt ticker present with correct (relative) Market_Cap_B.
        mc_ok = True
        mc_detail = []
        for key, gv in gt.items():
            ar = a_lookup.get(key)
            if ar is None:
                mc_ok = False
                mc_detail.append(f"missing {gv['symbol']}")
                continue
            av = parse_float(ar[ai_mc]) if ai_mc is not None else None
            if not rel_close(av, gv["mc"], rel=0.02):
                mc_ok = False
                mc_detail.append(f"{gv['symbol']}: {av} vs {gv['mc']}")
        check("All tracked tickers present with correct Market_Cap_B", mc_ok, "; ".join(mc_detail))

        # CRITICAL: sorted descending by Market_Cap_B (actual row order).
        agent_mcs = [parse_float(r[ai_mc]) for r in a_data if ai_mc is not None]
        agent_mcs = [m for m in agent_mcs if m is not None]
        sorted_ok = all(agent_mcs[i] >= agent_mcs[i + 1] - 1e-6 for i in range(len(agent_mcs) - 1))
        check("Market Cap Ranking sorted by Market_Cap_B descending", sorted_ok,
              f"order={agent_mcs}")

        # CRITICAL: Sector matches moex stock_info (skip empty-sector tickers).
        sec_ok = True
        sec_detail = []
        for key, gv in gt.items():
            gsec = str(gv["sector"] or "").strip()
            if not gsec:
                continue  # ticker with no sector in seed (e.g. TCSG.ME)
            ar = a_lookup.get(key)
            if ar is None or ai_sec is None:
                sec_ok = False
                sec_detail.append(f"{gv['symbol']}: missing")
                continue
            if not str_eq(ar[ai_sec], gsec):
                sec_ok = False
                sec_detail.append(f"{gv['symbol']}: {ar[ai_sec]} vs {gsec}")
        check("Sector matches moex stock_info for each ticker", sec_ok, "; ".join(sec_detail))

    # ---- Summary sheet ----
    print("\n=== Checking Summary ===")
    s_rows = load_sheet_rows(agent_wb, "Summary")
    if s_rows is None or len(s_rows) < 2:
        check("Summary Largest / Smallest / Total match agent data", False, "Summary sheet missing")
        check("Summary Total_Companies present", False, "Summary sheet missing")
    else:
        smap = {}
        for r in s_rows[1:]:
            if r and r[0] is not None and len(r) > 1:
                smap[str(r[0]).strip().lower()] = r[1]
        total_companies = smap.get("total_companies")
        largest = smap.get("largest")
        smallest = smap.get("smallest")
        total_mc = parse_float(smap.get("total_market_cap_b"))

        # NON-critical structural: Total_Companies equals number of gt tickers.
        try:
            tc_ok = int(parse_float(total_companies)) == len(gt)
        except (TypeError, ValueError):
            tc_ok = False
        check("Summary Total_Companies present", tc_ok,
              f"{total_companies} vs {len(gt)}")

        # CRITICAL: Largest/Smallest names match gt extremes, Total ~ sum.
        largest_ok = name_eq(largest, gt_largest)
        smallest_ok = name_eq(smallest, gt_smallest)
        total_ok = total_mc is not None and rel_close(total_mc, gt_total, rel=0.02)
        check("Summary Largest / Smallest / Total match agent data",
              largest_ok and smallest_ok and total_ok,
              f"largest={largest} (exp {gt_largest}); smallest={smallest} "
              f"(exp {gt_smallest}); total={total_mc} (exp {gt_total})")

    # ---- Teamly page ----
    print("\n=== Checking Teamly Knowledge Base ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly 'Market Cap Rankings' page exists and names the largest company",
              False, str(e))
        rows = []

    pages = [
        (t, b) for t, b in rows
        if t and ("market cap" in t.lower() or "rankings" in t.lower())
    ]
    page_text = " ".join((str(t) + " " + str(b)) for t, b in pages).lower()
    largest_name = str(gt_largest or "").strip().lower()
    # Match either full longName or a distinctive token of it (e.g. "газпром").
    largest_token = largest_name.split()[-1] if largest_name else ""
    names_largest = bool(pages) and largest_name and (
        largest_name in page_text or (largest_token and largest_token in page_text)
    )
    check("Teamly 'Market Cap Rankings' page exists and names the largest company",
          bool(pages) and names_largest,
          f"pages={len(pages)}, largest={gt_largest}")

    # ---- Verdict ----
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    print(f"\n=== Score: {PASS_COUNT}/{total} = {accuracy:.1f}% ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print("CRITICAL FAILURE(S): " + "; ".join(critical_failed))
        print("=== RESULT: FAIL ===")
        sys.exit(1)

    if accuracy >= 70.0:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print("=== RESULT: FAIL (accuracy below 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
