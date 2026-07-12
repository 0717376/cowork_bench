"""Evaluation для kulinar-diet-forms-teamly-excel (RU-стек: kulinar/forms/teamly).

Проверяет четыре части задачи:
  - Forms (gform.*): опрос «Dietary Preference Survey» с >=4 вопросами. RU forms-mcp
    (local_servers/forms-mcp) умеет ровно два инструмента: add_text_question
    (question_type='textQuestion') и add_multiple_choice_question
    (question_type='choiceQuestion'). Поэтому проверяем достижимую схему: >=4
    вопроса, есть текстовый вопрос (textQuestion) и есть вопрос(ы) с выбором
    (choiceQuestion), а 4 содержательных вопроса (приёмы пищи / ограничения /
    шкала готовки / избегаемые ингредиенты) опознаём по заголовкам.
  - Teamly (teamly.pages): страница «Healthy Recipe Knowledge Base» с содержимым по
    рецептам (название + категория + описание), >=6 реальных блюд kulinar.
  - Excel Recipe_Overview.xlsx: лист Recipes (колонки Recipe_Name/Category/Description,
    >=6 строк) и лист Category Summary (>=3 строки). ВНУТРЕННЯЯ согласованность:
    суммы по категориям в Category Summary совпадают с числом строк по категориям в
    Recipes; рецепты реальны (из базы kulinar), охватывают >=3 категории.
  - Email на wellness.team@company.com с темой про здоровое питание/рецепты.

Критические чеки (CRITICAL_CHECKS): любой их fail => задача FAIL, даже если общая
accuracy >= 70%. Остальные (структурные) — мягкие. Порог: accuracy >= 70% И нет
критических провалов.
"""
import argparse
import os
import sys

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# Канонический набор блюд kulinar (источник: local_servers/kulinar-mcp -> all_recipes.json).
KULINAR_RECIPES = {
    "Салат Оливье", "Винегрет", "Сельдь под шубой", "Салат Мимоза", "Крабовый салат",
    "Греческий салат", "Салат с курицей и грибами", "Холодец", "Икра кабачковая",
    "Грибы маринованные", "Сало солёное", "Селёдка с луком", "Борщ",
    "Щи из квашеной капусты", "Солянка мясная", "Уха", "Окрошка", "Грибной суп",
    "Рассольник", "Куриный бульон с лапшой", "Бефстроганов", "Пельмени домашние",
    "Голубцы", "Котлеты домашние", "Жаркое в горшочках", "Курица в сметане",
    "Рыба запечённая по-русски", "Цыплёнок табака", "Гречка с тушёнкой",
    "Плов узбекский", "Картофельное пюре", "Гречневая каша", "Перловая каша",
    "Картофель отварной с укропом", "Рис отварной", "Пирожки с капустой жареные",
    "Пирожки с мясом печёные", "Блины тонкие", "Кулебяка с капустой и яйцом",
    "Расстегаи с рыбой", "Медовик", "Наполеон", "Сырники", "Пасха творожная",
    "Ватрушки с творогом", "Кисель ягодный", "Морс клюквенный",
    "Компот из сухофруктов", "Сбитень", "Квас домашний",
}


def norm_recipe(s):
    s = (s or "").strip().lower().replace("ё", "е")
    return " ".join(s.split())


CANON_NORM = {norm_recipe(r) for r in KULINAR_RECIPES}

# Критические чеки по имени record()
CRITICAL_CHECKS = {
    "Forms: опрос >=4 вопросов (текст+выбор), охватывает приёмы пищи/ограничения/шкалу/текст",
    "Excel: лист Recipes согласован, >=6 реальных блюд kulinar из >=3 категорий",
    "Excel: Category Summary согласован с Recipes",
    "Teamly: страница базы знаний с реальными рецептами (>=6)",
    "Email на wellness.team@company.com отправлено",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


# ---------------------------------------------------------------------------
# Forms (gform)
# ---------------------------------------------------------------------------
def check_gform():
    print("\n=== Проверка Google-формы (gform) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()

        def title_matches(title):
            t = (title or "").lower()
            ru = any(k in t for k in ("предпочт", "питани", "опрос", "анкет", "диет"))
            en = any(k in t for k in ("dietary", "preference", "survey"))
            return ru or en

        target_form = None
        for fid, title in forms:
            if title_matches(title):
                target_form = fid
                break
        record("GForm 'Dietary Preference Survey' существует",
               target_form is not None,
               f"Найденные формы: {[t for _, t in forms]}")

        questions = []
        if target_form is not None:
            cur.execute(
                "SELECT title, question_type, required FROM gform.questions "
                "WHERE form_id = %s ORDER BY position", (target_form,))
            questions = cur.fetchall()

        record("GForm имеет минимум 4 вопроса", len(questions) >= 4,
               f"Найдено {len(questions)} вопросов")

        # RU forms-mcp пишет ТОЛЬКО два значения question_type:
        #   'textQuestion'  (add_text_question)
        #   'choiceQuestion' (add_multiple_choice_question, config.type='RADIO')
        # CHECKBOX/SCALE/LINEAR_SCALE инструментами недостижимы — опознаём 4
        # содержательных вопроса по заголовкам, а структуру — по этим двум типам.
        q_types = [q[1] for q in questions]
        has_text = any(
            (qt or "") in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH")
            for qt in q_types)
        has_choice = any(
            (qt or "") in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE", "CHECKBOX")
            for qt in q_types)
        record("GForm: есть текстовый вопрос (textQuestion)", has_text, f"Типы: {q_types}")
        record("GForm: есть вопрос(ы) с выбором (choiceQuestion)", has_choice, f"Типы: {q_types}")

        # Содержательные вопросы по заголовкам (RU+EN ключи в ОРИГИНАЛЬНОМ .lower())
        q_titles_lower = [(q[0] or "").lower() for q in questions]

        def any_title(*keys):
            return any(any(k in qt for k in keys) for qt in q_titles_lower)

        has_meals_q = any_title("приём", "прием", "завтрак", "обед", "ужин",
                                "meal", "breakfast", "lunch", "dinner")
        has_dietary_q = any_title("ограничен", "вегетариан", "веган", "глютен", "молочн",
                                  "dietary", "restriction", "vegetarian", "vegan", "gluten")
        has_scale_q = any_title("сколько", "недел", "готов", "many", "week", "cook", "scale")
        has_avoid_q = any_title("избег", "ингредиент", "avoid", "ingredient")

        record("GForm: есть вопрос про приёмы пищи", has_meals_q, f"Заголовки: {q_titles_lower}")
        record("GForm: есть вопрос про пищевые ограничения", has_dietary_q,
               f"Заголовки: {q_titles_lower}")
        record("GForm: есть вопрос-шкала про готовку блюд в неделю", has_scale_q,
               f"Заголовки: {q_titles_lower}")
        record("GForm: есть текстовый вопрос про избегаемые ингредиенты", has_avoid_q,
               f"Заголовки: {q_titles_lower}")

        # CRITICAL: реальная достижимая схема — >=4 вопроса, есть text и choice,
        # и покрыты 4 содержательных вопроса (приёмы пищи, ограничения, шкала, текст).
        record("Forms: опрос >=4 вопросов (текст+выбор), охватывает приёмы пищи/ограничения/шкалу/текст",
               target_form is not None and len(questions) >= 4
               and has_text and has_choice
               and has_meals_q and has_dietary_q and has_scale_q and has_avoid_q,
               f"types={q_types}, titles={q_titles_lower}")

        conn.close()
    except Exception as e:
        record("Forms: опрос >=4 вопросов (текст+выбор), охватывает приёмы пищи/ограничения/шкалу/текст", False, str(e))


# ---------------------------------------------------------------------------
# Teamly
# ---------------------------------------------------------------------------
def check_teamly():
    print("\n=== Проверка Teamly (страница базы знаний) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Teamly: страница базы знаний с реальными рецептами (>=6)", False, str(e))
        return

    def title_matches(title):
        t = (title or "").lower()
        ru = ("рецепт" in t or "блюд" in t) and ("баз" in t or "знани" in t or "полезн" in t or "здоров" in t)
        en = "recipe" in t and ("knowledge" in t or "healthy" in t or "base" in t)
        return ru or en

    candidates = [(t, b) for t, b in pages if title_matches(t)]
    record("Teamly: страница 'Healthy Recipe Knowledge Base' создана",
           len(candidates) >= 1, f"Заголовки: {[t for t, _ in pages]}")

    body = "\n\n".join(b for _t, b in candidates)
    nb = norm_recipe(body)
    dishes_found = sum(1 for c in CANON_NORM if c and c in nb)

    record("Teamly: на странице перечислены блюда kulinar (>=6)",
           dishes_found >= 6, f"найдено блюд={dishes_found}")

    # CRITICAL: страница есть и содержит >=6 реальных блюд
    record("Teamly: страница базы знаний с реальными рецептами (>=6)",
           bool(candidates) and dishes_found >= 6,
           f"candidates={len(candidates)}, dishes={dishes_found}")


# ---------------------------------------------------------------------------
# Excel (внутренняя согласованность, без хардкод-GT)
# ---------------------------------------------------------------------------
def check_excel(agent_workspace):
    print("\n=== Проверка Excel Recipe_Overview.xlsx ===")
    xl_path = os.path.join(agent_workspace, "Recipe_Overview.xlsx")
    if not os.path.isfile(xl_path):
        record("Excel Recipe_Overview.xlsx существует", False, f"Не найден: {xl_path}")
        record("Excel: лист Recipes согласован, >=6 реальных блюд kulinar из >=3 категорий", False, "нет файла")
        record("Excel: Category Summary согласован с Recipes", False, "нет файла")
        return
    record("Excel Recipe_Overview.xlsx существует", True)

    try:
        wb = openpyxl.load_workbook(xl_path, data_only=True)
    except Exception as e:
        record("Excel читается", False, str(e))
        record("Excel: лист Recipes согласован, >=6 реальных блюд kulinar из >=3 категорий", False, "не читается")
        record("Excel: Category Summary согласован с Recipes", False, "не читается")
        return
    record("Excel читается", True)

    sheet_names = [s.lower() for s in wb.sheetnames]
    record("Excel имеет лист 'Recipes'", any("recipe" in s for s in sheet_names),
           f"Листы: {wb.sheetnames}")
    record("Excel имеет лист 'Category Summary'",
           any("category" in s or "summary" in s for s in sheet_names),
           f"Листы: {wb.sheetnames}")

    # --- Лист Recipes ---
    recipes_ws = None
    for sname in wb.sheetnames:
        if "recipe" in sname.lower():
            recipes_ws = wb[sname]
            break

    recipes_rows = []        # (name, category)
    name_idx, cat_idx = 0, 1
    if recipes_ws is not None:
        headers = [c.value for c in list(recipes_ws.rows)[0]] if recipes_ws.max_row > 0 else []
        hl = [str(h).strip().lower() if h is not None else "" for h in headers]
        record("Recipes: колонка Recipe_Name", any("recipe" in h or "name" in h for h in hl), f"Headers: {headers}")
        record("Recipes: колонка Category", any("category" in h for h in hl), f"Headers: {headers}")
        record("Recipes: колонка Description", any("descr" in h for h in hl), f"Headers: {headers}")

        for i, h in enumerate(hl):
            if "name" in h or "recipe" in h:
                name_idx = i
            if "category" in h:
                cat_idx = i

        for row in recipes_ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(v is not None and str(v).strip() for v in row):
                continue
            nm = row[name_idx] if name_idx < len(row) else None
            ct = row[cat_idx] if cat_idx < len(row) else None
            recipes_rows.append((nm, ct))

    data_rows = len(recipes_rows)
    record("Excel Recipes: >=6 строк блюд", data_rows >= 6, f"Найдено {data_rows} строк")

    # Реальность блюд против kulinar
    matched = 0
    unknown = []
    for nm, _ct in recipes_rows:
        n = norm_recipe(nm)
        if n and (n in CANON_NORM or any(c in n or n in c for c in CANON_NORM)):
            matched += 1
        elif nm is not None and str(nm).strip():
            unknown.append(nm)
    record("Excel Recipes: блюда реальны (из базы kulinar)",
           data_rows > 0 and matched == data_rows,
           f"совпало {matched}/{data_rows}; неизвестные: {unknown[:5]}")

    # Категории по строкам Recipes
    cat_counts = {}
    for _nm, ct in recipes_rows:
        key = norm_recipe(ct)
        if key:
            cat_counts[key] = cat_counts.get(key, 0) + 1
    distinct_cats = len(cat_counts)
    record("Excel Recipes: >=3 различных категорий", distinct_cats >= 3,
           f"Категории: {list(cat_counts.keys())}")

    # CRITICAL: Recipes согласован — >=6 реальных блюд из >=3 категорий
    record("Excel: лист Recipes согласован, >=6 реальных блюд kulinar из >=3 категорий",
           data_rows >= 6 and matched == data_rows and distinct_cats >= 3,
           f"rows={data_rows}, matched={matched}, cats={distinct_cats}")

    # --- Лист Category Summary ---
    cat_ws = None
    for sname in wb.sheetnames:
        if "category" in sname.lower() or "summary" in sname.lower():
            cat_ws = wb[sname]
            break

    summary_pairs = []   # (category_norm, count)
    if cat_ws is not None:
        for row in cat_ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(v is not None and str(v).strip() for v in row):
                continue
            cname = norm_recipe(row[0]) if len(row) > 0 else ""
            cval = None
            for v in row[1:]:
                if isinstance(v, (int, float)):
                    cval = v
                    break
                if v is not None and str(v).strip().isdigit():
                    cval = int(str(v).strip())
                    break
            if cname:
                summary_pairs.append((cname, cval))

    record("Excel Category Summary: >=3 категорий", len(summary_pairs) >= 3,
           f"Найдено {len(summary_pairs)} категорий")

    # Внутренняя согласованность: счётчики Summary == счётчикам по строкам Recipes
    consistent = len(summary_pairs) >= 3 and distinct_cats >= 3
    mismatches = []
    if consistent:
        for cname, cval in summary_pairs:
            expected = cat_counts.get(cname)
            if expected is None or cval is None or not num_close(cval, expected, 0):
                consistent = False
                mismatches.append((cname, cval, expected))
    record("Excel: Category Summary согласован с Recipes",
           consistent and not mismatches,
           f"summary={summary_pairs}, recipes_counts={cat_counts}, mismatches={mismatches[:5]}")


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------
def check_email():
    print("\n=== Проверка письма ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE LOWER(subject) LIKE '%healthy%' OR LOWER(subject) LIKE '%recipe%' OR LOWER(subject) LIKE '%eating%'
        """)
        emails = cur.fetchall()
        record("Письмо о здоровом питании/рецептах отправлено", len(emails) > 0,
               f"Найдено {len(emails)} подходящих писем")

        target_found = False
        for subject, to_addr in emails:
            to_str = str(to_addr).lower() if to_addr else ""
            if "wellness.team@company.com" in to_str:
                target_found = True
                break
        record("Email на wellness.team@company.com отправлено", target_found,
               f"Получатели: {[e[1] for e in emails]}")
        conn.close()
    except Exception as e:
        record("Email на wellness.team@company.com отправлено", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("KULINAR-DIET (RU: kulinar/forms/teamly) - EVALUATION")
    print("=" * 70)

    check_gform()
    check_teamly()
    check_excel(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    pct = 100.0 * PASS_COUNT / total if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {pct:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAIL: {CRITICAL_FAILED}")
        print("  Overall: FAIL")
        sys.exit(1)

    overall = pct >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
