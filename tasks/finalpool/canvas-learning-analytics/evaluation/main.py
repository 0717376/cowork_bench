"""Evaluation for canvas-learning-analytics.

Source of truth is the LIVE Canvas MCP instance; all expected KPIs are derived
from canvas.* at eval time (get_expected_kpis), so course names / grades are
read honestly and never hardcoded.

Critical checks (CRITICAL_CHECKS): any failure there => overall FAIL regardless
of accuracy. Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Learning_Analytics.xlsx exists",
    "Course KPIs: all course avg_grade values correct",
    "Course KPIs: all course completion_rate values correct",
    "Course KPIs: all course late_rate values correct",
    "Course KPIs: Meets_Benchmark column correct for every course",
    "Summary: Total_Courses, Meets_All_Benchmarks, Below_Benchmark, Benchmark_Rate correct",
    "Trends: 66 rows with Performance consistent with KPIs",
    "Teamly: 'Learning Analytics Dashboard' page with multiple content blocks",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _to_float(x):
    # Tolerate percent/locale formatting, e.g. '45.5%', '45,5'.
    if isinstance(x, str):
        x = x.replace("%", "").replace("\xa0", "").replace(",", ".").strip()
    return float(x)


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(_to_float(a) - _to_float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_kpis():
    """Get expected KPIs from the live Canvas DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    results = {}
    # Only the 22 OU courses (task.md), excluding courses leaked from other tasks
    cur.execute("SELECT id, name FROM canvas.courses WHERE id BETWEEN 1 AND 22 ORDER BY name")
    courses = cur.fetchall()

    for cid, cname in courses:
        # Completion rate
        cur.execute("""
            SELECT COUNT(*) FILTER (WHERE s.workflow_state = 'graded'),
                   COUNT(*)
            FROM canvas.submissions s
            JOIN canvas.assignments a ON s.assignment_id = a.id
            WHERE a.course_id = %s
        """, (cid,))
        graded, total = cur.fetchone()
        comp_rate = round(graded / total * 100, 1) if total > 0 else 0

        # Avg grade
        cur.execute("""
            SELECT ROUND(AVG((grades->>'current_score')::numeric), 1)
            FROM canvas.enrollments
            WHERE course_id = %s AND type = 'StudentEnrollment'
              AND grades->>'current_score' IS NOT NULL
        """, (cid,))
        avg_grade = float(cur.fetchone()[0] or 0)

        # Late rate
        cur.execute("""
            SELECT SUM(CASE WHEN s.late THEN 1 ELSE 0 END)::numeric / NULLIF(COUNT(s.id), 0) * 100
            FROM canvas.submissions s
            JOIN canvas.assignments a ON s.assignment_id = a.id
            WHERE a.course_id = %s
        """, (cid,))
        late_rate = round(float(cur.fetchone()[0] or 0), 1)

        meets = comp_rate >= 75 and avg_grade >= 70 and late_rate <= 15
        results[cname.lower()] = {
            "completion_rate": comp_rate,
            "avg_grade": avg_grade,
            "late_rate": late_rate,
            "meets": meets,
        }

    cur.close()
    conn.close()
    return results


def find_col(header_lower, key):
    for i, h in enumerate(header_lower):
        if key in h:
            return i
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Learning_Analytics.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Learning_Analytics.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Learning_Analytics.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    expected = get_expected_kpis()
    n_courses = len(expected)
    meets_count = sum(1 for v in expected.values() if v["meets"])
    below_count = n_courses - meets_count

    # ── Course KPIs sheet ────────────────────────────────────────────────────
    kpi_rows = load_sheet_rows(wb, "Course KPIs")
    if kpi_rows is None:
        check("Sheet 'Course KPIs' exists", False, f"Available: {wb.sheetnames}")
        # Cannot evaluate per-course critical checks without the sheet.
        check("Course KPIs: all course avg_grade values correct", False, "no sheet")
        check("Course KPIs: all course completion_rate values correct", False, "no sheet")
        check("Course KPIs: all course late_rate values correct", False, "no sheet")
        check("Course KPIs: Meets_Benchmark column correct for every course", False, "no sheet")
    else:
        check("Sheet 'Course KPIs' exists", True)
        data_rows = kpi_rows[1:] if len(kpi_rows) > 1 else []
        check(f"Course KPIs has {n_courses} rows", len(data_rows) == n_courses,
              f"Found {len(data_rows)}")

        header = kpi_rows[0] if kpi_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "completion_rate", "avg_grade", "late_rate",
                    "discussion_count", "meets_benchmark"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

        ci_course = find_col(header_lower, "course")
        ci_comp = find_col(header_lower, "completion_rate")
        ci_grade = find_col(header_lower, "avg_grade")
        ci_late = find_col(header_lower, "late_rate")
        ci_meets = find_col(header_lower, "meets_benchmark")
        # Fallback to positional columns if headers not found.
        if ci_course is None:
            ci_course = 0
        if ci_comp is None:
            ci_comp = 1
        if ci_grade is None:
            ci_grade = 2
        if ci_late is None:
            ci_late = 3
        if ci_meets is None:
            ci_meets = 5

        def cell(row, idx):
            return row[idx] if row and idx is not None and idx < len(row) else None

        # Build a lookup of agent rows by lowercased course name.
        agent_by_name = {}
        for row in data_rows:
            cname = cell(row, ci_course)
            if cname:
                agent_by_name[str(cname).strip().lower()] = row

        grade_all_ok = True
        comp_all_ok = True
        late_all_ok = True
        meets_all_ok = True
        for cname, exp in expected.items():
            row = agent_by_name.get(cname)
            if row is None:
                grade_all_ok = comp_all_ok = late_all_ok = meets_all_ok = False
                print(f"    -> course '{cname}' missing from Course KPIs")
                continue
            if not num_close(cell(row, ci_grade), exp["avg_grade"], 1.0):
                grade_all_ok = False
                print(f"    -> {cname} avg_grade got {cell(row, ci_grade)}, expected {exp['avg_grade']}")
            if not num_close(cell(row, ci_comp), exp["completion_rate"], 1.0):
                comp_all_ok = False
                print(f"    -> {cname} completion_rate got {cell(row, ci_comp)}, expected {exp['completion_rate']}")
            if not num_close(cell(row, ci_late), exp["late_rate"], 1.0):
                late_all_ok = False
                print(f"    -> {cname} late_rate got {cell(row, ci_late)}, expected {exp['late_rate']}")
            mval = str(cell(row, ci_meets)).strip().lower()
            agent_meets = mval in ("yes", "да", "true")
            if agent_meets != exp["meets"]:
                meets_all_ok = False
                print(f"    -> {cname} Meets_Benchmark got '{cell(row, ci_meets)}', expected {exp['meets']}")

        check("Course KPIs: all course avg_grade values correct", grade_all_ok,
              "one or more avg_grade wrong")
        check("Course KPIs: all course completion_rate values correct", comp_all_ok,
              "one or more completion_rate wrong")
        check("Course KPIs: all course late_rate values correct", late_all_ok,
              "one or more late_rate wrong")
        check("Course KPIs: Meets_Benchmark column correct for every course", meets_all_ok,
              "one or more Meets_Benchmark wrong")

    # ── Summary sheet ────────────────────────────────────────────────────────
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
        check("Summary: Total_Courses, Meets_All_Benchmarks, Below_Benchmark, Benchmark_Rate correct",
              False, "no sheet")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        bench_rate = round(meets_count / n_courses * 100, 1) if n_courses else 0
        tot_ok = num_close(lookup.get("total_courses"), n_courses)
        meets_ok = num_close(lookup.get("meets_all_benchmarks"), meets_count)
        below_ok = num_close(lookup.get("below_benchmark"), below_count)
        rate_ok = num_close(lookup.get("benchmark_rate"), bench_rate, 1.0)

        check("Summary: Total_Courses correct", tot_ok,
              f"Got {lookup.get('total_courses')}, expected {n_courses}")
        check("Summary: Meets_All_Benchmarks correct", meets_ok,
              f"Got {lookup.get('meets_all_benchmarks')}, expected {meets_count}")
        check("Summary: Below_Benchmark correct", below_ok,
              f"Got {lookup.get('below_benchmark')}, expected {below_count}")
        check("Summary: Benchmark_Rate correct", rate_ok,
              f"Got {lookup.get('benchmark_rate')}, expected {bench_rate}")
        check("Summary: Total_Courses, Meets_All_Benchmarks, Below_Benchmark, Benchmark_Rate correct",
              tot_ok and meets_ok and below_ok and rate_ok,
              "one or more summary values wrong")

    # ── Trends sheet ─────────────────────────────────────────────────────────
    tr_rows = load_sheet_rows(wb, "Trends")
    if tr_rows is None:
        check("Sheet 'Trends' exists", False, f"Available: {wb.sheetnames}")
        check("Trends: 66 rows with Performance consistent with KPIs", False, "no sheet")
    else:
        check("Sheet 'Trends' exists", True)
        data_rows = [r for r in tr_rows[1:] if r and any(c is not None for c in r)]
        # 22 courses * 3 categories = 66 rows
        rows_ok = len(data_rows) == 3 * n_courses
        check(f"Trends has {3 * n_courses} rows", rows_ok, f"Found {len(data_rows)}")

        header = tr_rows[0] if tr_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        ci_course = find_col(header_lower, "course")
        ci_cat = find_col(header_lower, "category")
        ci_perf = find_col(header_lower, "performance")
        if ci_course is None:
            ci_course = 0
        if ci_cat is None:
            ci_cat = 1
        if ci_perf is None:
            ci_perf = 2

        def tcell(row, idx):
            return row[idx] if row and idx < len(row) else None

        consistent = rows_ok
        for row in data_rows:
            cname = tcell(row, ci_course)
            cat = tcell(row, ci_cat)
            perf = tcell(row, ci_perf)
            if not cname or not cat or perf is None:
                consistent = False
                continue
            exp = expected.get(str(cname).strip().lower())
            if exp is None:
                consistent = False
                continue
            cat_l = str(cat).strip().lower()
            perf_l = str(perf).strip().lower()
            if "completion" in cat_l or "завершен" in cat_l:
                want = "above" if exp["completion_rate"] >= 75 else "below"
                ok = (want in perf_l) or ("выше" in perf_l if want == "above" else "ниже" in perf_l)
            elif "grade" in cat_l or "оценк" in cat_l or "балл" in cat_l:
                want = "above" if exp["avg_grade"] >= 70 else "below"
                ok = (want in perf_l) or ("выше" in perf_l if want == "above" else "ниже" in perf_l)
            elif "late" in cat_l or "просроч" in cat_l:
                want = "acceptable" if exp["late_rate"] <= 15 else "high"
                ok = (want in perf_l) or ("приемлем" in perf_l if want == "acceptable" else "высок" in perf_l)
            else:
                ok = False
            if not ok:
                consistent = False
                print(f"    -> Trends inconsistent: {cname} / {cat} / {perf}")

        check("Trends: 66 rows with Performance consistent with KPIs", consistent,
              "row count or Above/Below/Acceptable/High logic wrong")


def check_teamly():
    print("\n=== Checking Teamly Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("Teamly: 'Learning Analytics Dashboard' page with multiple content blocks",
              False, f"DB error: {e}")
        return

    try:
        # Only user-created pages (seed pages have id <= 3).
        cur.execute("SELECT id, title, body FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
    except Exception as e:
        check("Teamly: 'Learning Analytics Dashboard' page with multiple content blocks",
              False, f"DB error: {e}")
        cur.close()
        conn.close()
        return

    found = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if ("learning analytics" in tl or "analytics dashboard" in tl
                or "панель учебной аналитики" in tl or "учебная аналитика" in tl):
            found = (pid, title, body)
            break

    check("Teamly: analytics dashboard page exists", found is not None,
          f"Found {len(pages)} user pages")

    if found is None:
        check("Teamly: 'Learning Analytics Dashboard' page with multiple content blocks",
              False, "page not found")
        cur.close()
        conn.close()
        return

    body = (found[2] or "")
    bl = body.lower()
    # "Multiple content blocks" ~ a body that covers the four required sections.
    # Match each section by RU or EN keywords.
    sec_overview = any(k in bl for k in ["overview", "обзор", "сводк", "benchmark", "бенчмарк"])
    sec_meeting = any(k in bl for k in ["meet", "удовлетвор", "достиг", "соответств"])
    sec_attention = any(k in bl for k in ["attention", "вниман", "needing", "требующ", "ниже"])
    sec_reco = any(k in bl for k in ["recommend", "рекоменд", "улучшен", "action"])
    sections_ok = sum([sec_overview, sec_meeting, sec_attention, sec_reco]) >= 3
    nonempty = len(body.strip()) >= 50

    check("Teamly: 'Learning Analytics Dashboard' page with multiple content blocks",
          sections_ok and nonempty,
          f"sections={[sec_overview, sec_meeting, sec_attention, sec_reco]}, body_len={len(body)}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {critical_failed}")

    all_passed = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "critical_failed": critical_failed,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        sys.exit(0)
    sys.exit(1)


if __name__ == "__main__":
    main()
