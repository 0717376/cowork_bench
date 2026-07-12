"""Оценка задачи мониторинга экспирации опционов (MOEX swap).

Источник данных детерминирован (см. preprocess SEED_PLAN), поэтому все
ожидаемые агрегаты вычисляются точно. CRITICAL_CHECKS проверяют существо
аналитики (значения Avg_IV / Risk_Level, состав near-expiry, сводные счётчики,
события календаря). Любой провал критической проверки => немедленный FAIL
(sys.exit(1)) ещё до порога точности.
"""
import argparse
import os
import statistics
import sys
from datetime import date

import psycopg2

CURRENT_DATE = date(2026, 3, 7)
CUTOFF_DATE = date(2026, 3, 14)

# Тот же план, что и в preprocess/main.py (источник истины для ожиданий).
SEED_PLAN = {
    "SBER.ME": {
        "2026-03-10": {"calls": [0.55, 0.58, 0.61], "puts": [0.50, 0.52, 0.54]},
        "2026-03-13": {"calls": [0.30, 0.32], "puts": [0.28, 0.31]},
        "2026-04-17": {"calls": [0.22, 0.24, 0.26], "puts": [0.20, 0.23]},
    },
    "GAZP.ME": {
        "2026-03-12": {"calls": [0.45, 0.47], "puts": [0.41, 0.43]},
        "2026-05-15": {"calls": [0.33, 0.35, 0.30], "puts": [0.29, 0.31]},
    },
    "LKOH.ME": {
        "2026-03-14": {"calls": [0.38, 0.36], "puts": [0.35, 0.37, 0.39]},
        "2026-06-19": {"calls": [0.28, 0.30], "puts": [0.26, 0.27]},
    },
    "MGNT.ME": {
        "2026-03-11": {"calls": [0.62, 0.65, 0.60], "puts": [0.58, 0.61]},
        "2026-04-17": {"calls": [0.42, 0.44], "puts": [0.40, 0.45, 0.43]},
    },
    "MTSS.ME": {
        "2026-03-09": {"calls": [0.33, 0.35], "puts": [0.31, 0.34, 0.30]},
        "2026-07-17": {"calls": [0.25, 0.27, 0.29], "puts": [0.24, 0.26]},
    },
}

SYMBOLS = sorted(SEED_PLAN.keys())


def _risk_level(avg_iv_pct, near):
    high = avg_iv_pct > 40
    if high and near:
        return "High IV + Near Expiry"
    if high:
        return "High IV"
    if near:
        return "Near Expiry"
    return "Normal"


def _build_expected():
    """Вычислить эталонные группы и сводные значения из SEED_PLAN."""
    groups = {}  # (sym, exp, type) -> dict
    for sym, exps in SEED_PLAN.items():
        for exp, types in exps.items():
            ed = date.fromisoformat(exp)
            near = ed <= CUTOFF_DATE
            for typ in ("calls", "puts"):
                ivs = types[typ]
                avg = round(statistics.mean(ivs) * 100, 1)
                groups[(sym, exp, typ)] = {
                    "n": len(ivs),
                    "avg_iv": avg,
                    "near": near,
                    "high": avg > 40,
                    "risk": _risk_level(avg, near),
                    "days": (ed - CURRENT_DATE).days,
                }
    near_groups = {k: v for k, v in groups.items() if v["near"]}
    summary = {
        "total_positions": len(groups),
        "near_expiry_count": len(near_groups),
        "high_iv_count": sum(1 for v in groups.values() if v["high"]),
        "total_near_expiry_contracts": sum(v["n"] for v in near_groups.values()),
        "stocks_with_near_expiry": sorted({k[0] for k in near_groups}),
    }
    return groups, near_groups, summary


EXP_GROUPS, EXP_NEAR_GROUPS, EXP_SUMMARY = _build_expected()


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        if isinstance(a, str):
            a = a.replace("%", "").replace("\xa0", "").replace(",", ".").strip()
        if isinstance(b, str):
            b = b.replace("%", "").replace("\xa0", "").replace(",", ".").strip()
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _norm_date(v):
    """Привести значение ячейки экспирации к 'YYYY-MM-DD'."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v).strip()
    return s[:10]


def _header_index(header, names):
    low = [str(h).strip().lower() if h is not None else "" for h in header]
    out = {}
    for n in names:
        out[n] = low.index(n.lower()) if n.lower() in low else None
    return out


# ----------------------------- Excel checks --------------------------------

def parse_excel(agent_workspace):
    """Вернуть (wb или None, ошибка-строка или None)."""
    import openpyxl

    path = os.path.join(agent_workspace, "Options_Monitor.xlsx")
    if not os.path.exists(path):
        return None, "Options_Monitor.xlsx not found"
    try:
        return openpyxl.load_workbook(path, data_only=True), None
    except Exception as e:  # noqa: BLE001
        return None, f"Error reading Excel: {e}"


def check_excel_structural(wb):
    """Нестрогие структурные проверки (наличие листов/столбцов/строк)."""
    errors = []
    rows = load_sheet_rows(wb, "Position Analysis")
    if rows is None:
        errors.append("Sheet 'Position Analysis' not found")
    else:
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        if len(data_rows) < 20:
            errors.append(f"Position Analysis has {len(data_rows)} rows, expected at least 20")
        symbols = {str(r[0]).strip().upper() for r in data_rows if r[0]}
        for sym in SYMBOLS:
            if sym.upper() not in symbols:
                errors.append(f"Symbol {sym} missing from Position Analysis")

    rows2 = load_sheet_rows(wb, "Expiry Alerts")
    if rows2 is None:
        errors.append("Sheet 'Expiry Alerts' not found")
    else:
        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        if len(data_rows2) < 5:
            errors.append(f"Expiry Alerts has {len(data_rows2)} rows, expected at least 5")

    rows3 = load_sheet_rows(wb, "Summary")
    if rows3 is None:
        errors.append("Sheet 'Summary' not found")
    return errors


def critical_position_analysis(wb):
    """CRITICAL: проверить Avg_IV и Risk_Level для каждой засеянной группы."""
    errors = []
    rows = load_sheet_rows(wb, "Position Analysis")
    if not rows:
        return ["CRITICAL: Position Analysis sheet missing/empty"]
    header = rows[0]
    idx = _header_index(header, ["Symbol", "Expiration", "Type", "Avg_IV", "Risk_Level"])
    if any(idx[c] is None for c in ["Symbol", "Expiration", "Type", "Avg_IV", "Risk_Level"]):
        return [f"CRITICAL: Position Analysis missing required columns (got {header})"]

    found = {}
    for r in rows[1:]:
        if not r or r[idx["Symbol"]] is None:
            continue
        sym = str(r[idx["Symbol"]]).strip().upper()
        exp = _norm_date(r[idx["Expiration"]])
        typ = str(r[idx["Type"]]).strip().lower()
        if typ in ("call", "calls"):
            typ = "calls"
        elif typ in ("put", "puts"):
            typ = "puts"
        found[(sym, exp, typ)] = (r[idx["Avg_IV"]], str(r[idx["Risk_Level"]]).strip())

    for (sym, exp, typ), exp_vals in EXP_GROUPS.items():
        key = (sym.upper(), exp, typ)
        if key not in found:
            errors.append(f"CRITICAL: Position Analysis missing group {sym} {exp} {typ}")
            continue
        avg_iv, risk = found[key]
        if not num_close(avg_iv, exp_vals["avg_iv"], abs_tol=0.15, rel_tol=0.0):
            errors.append(
                f"CRITICAL: {sym} {exp} {typ} Avg_IV={avg_iv}, expected {exp_vals['avg_iv']}"
            )
        if risk.lower() != exp_vals["risk"].lower():
            errors.append(
                f"CRITICAL: {sym} {exp} {typ} Risk_Level='{risk}', expected '{exp_vals['risk']}'"
            )
    return errors


def critical_expiry_alerts(wb):
    """CRITICAL: множество near-expiry строк точно совпадает; Days_Until_Expiry верны."""
    errors = []
    rows = load_sheet_rows(wb, "Expiry Alerts")
    if not rows:
        return ["CRITICAL: Expiry Alerts sheet missing/empty"]
    header = rows[0]
    idx = _header_index(header, ["Symbol", "Expiration", "Type", "Days_Until_Expiry"])
    if any(idx[c] is None for c in ["Symbol", "Expiration", "Type", "Days_Until_Expiry"]):
        return [f"CRITICAL: Expiry Alerts missing required columns (got {header})"]

    seen = set()
    for r in rows[1:]:
        if not r or r[idx["Symbol"]] is None:
            continue
        sym = str(r[idx["Symbol"]]).strip().upper()
        exp = _norm_date(r[idx["Expiration"]])
        typ = str(r[idx["Type"]]).strip().lower()
        if typ in ("call", "calls"):
            typ = "calls"
        elif typ in ("put", "puts"):
            typ = "puts"
        days = r[idx["Days_Until_Expiry"]]
        try:
            ed = date.fromisoformat(exp)
        except Exception:  # noqa: BLE001
            errors.append(f"CRITICAL: Expiry Alerts bad date '{exp}' for {sym} {typ}")
            continue
        if ed > CUTOFF_DATE:
            errors.append(f"CRITICAL: Expiry Alerts row {sym} {exp} {typ} is past cutoff {CUTOFF_DATE}")
        exp_days = (ed - CURRENT_DATE).days
        if not num_close(days, exp_days, abs_tol=0, rel_tol=0.0):
            errors.append(
                f"CRITICAL: {sym} {exp} {typ} Days_Until_Expiry={days}, expected {exp_days}"
            )
        seen.add((sym.upper(), exp, typ))

    expected = {(s.upper(), e, t) for (s, e, t) in EXP_NEAR_GROUPS}
    missing = expected - seen
    extra = seen - expected
    if missing:
        errors.append(f"CRITICAL: Expiry Alerts missing near-expiry rows: {sorted(missing)}")
    if extra:
        errors.append(f"CRITICAL: Expiry Alerts has unexpected rows: {sorted(extra)}")
    return errors


def critical_summary(wb):
    """CRITICAL: точные сводные счётчики и состав near-expiry тикеров."""
    errors = []
    rows = load_sheet_rows(wb, "Summary")
    if not rows:
        return ["CRITICAL: Summary sheet missing/empty"]
    data_rows = [r for r in rows[1:] if r and r[0] is not None]
    lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows if r[0]}

    def need(metric, expected, abs_tol=0):
        if metric not in lookup:
            errors.append(f"CRITICAL: Summary missing metric '{metric}'")
        elif not num_close(lookup[metric], expected, abs_tol=abs_tol, rel_tol=0.0):
            errors.append(f"CRITICAL: Summary {metric}={lookup[metric]}, expected {expected}")

    need("total_positions", EXP_SUMMARY["total_positions"])
    need("near_expiry_count", EXP_SUMMARY["near_expiry_count"])
    need("high_iv_count", EXP_SUMMARY["high_iv_count"])
    need("total_near_expiry_contracts", EXP_SUMMARY["total_near_expiry_contracts"])

    if "stocks_with_near_expiry" not in lookup:
        errors.append("CRITICAL: Summary missing metric 'stocks_with_near_expiry'")
    else:
        val = str(lookup["stocks_with_near_expiry"]).upper()
        for sym in EXP_SUMMARY["stocks_with_near_expiry"]:
            if sym.upper() not in val:
                errors.append(f"CRITICAL: {sym} missing from Stocks_With_Near_Expiry")
    return errors


# ----------------------------- GCal checks ---------------------------------

def _gcal_events():
    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"),
        port=int(os.environ.get("PGPORT", "5432")),
        dbname="cowork_gym",
        user="eigent",
        password="camel",
    )
    cur = conn.cursor()
    cur.execute(
        """
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE summary ILIKE '%options expiry%'
        ORDER BY start_datetime
        """
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_gcal_structural(rows):
    errors = []
    if len(rows) < len(EXP_SUMMARY["stocks_with_near_expiry"]):
        errors.append(
            f"Only {len(rows)} options expiry calendar events, "
            f"expected at least {len(EXP_SUMMARY['stocks_with_near_expiry'])}"
        )
    return errors


def critical_gcal(rows):
    """CRITICAL: по одному корректному событию на каждый near-expiry тикер,
    окно 09:00-09:30 на одной из валидных дат экспирации этого тикера."""
    errors = []
    # допустимые даты экспирации near-expiry на тикер
    valid_dates = {}
    for (sym, exp, _typ) in EXP_NEAR_GROUPS:
        valid_dates.setdefault(sym.upper(), set()).add(exp)

    for sym in EXP_SUMMARY["stocks_with_near_expiry"]:
        ticker = sym.upper()
        match = None
        for summary, start_dt, end_dt in rows:
            s = (summary or "").upper()
            if ticker in s and "OPTIONS EXPIRY ALERT" in s:
                match = (summary, start_dt, end_dt)
                break
        if match is None:
            errors.append(f"CRITICAL: no calendar event 'Options Expiry Alert: {sym}'")
            continue
        summary, start_dt, end_dt = match
        if start_dt is None:
            errors.append(f"CRITICAL: event for {sym} has no start datetime")
            continue
        ev_date = start_dt.date().isoformat()
        if ev_date not in valid_dates.get(ticker, set()):
            errors.append(
                f"CRITICAL: event for {sym} on {ev_date}, expected one of "
                f"{sorted(valid_dates.get(ticker, set()))}"
            )
        if (start_dt.hour, start_dt.minute) != (9, 0):
            errors.append(
                f"CRITICAL: event for {sym} starts at {start_dt.hour:02d}:{start_dt.minute:02d}, expected 09:00"
            )
        if end_dt is not None and (end_dt.hour, end_dt.minute) != (9, 30):
            errors.append(
                f"CRITICAL: event for {sym} ends at {end_dt.hour:02d}:{end_dt.minute:02d}, expected 09:30"
            )
    return errors


# ------------------------------- main --------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    critical_errors = []
    soft_errors = []

    # --- Excel ---
    wb, perr = parse_excel(agent_ws)
    if perr:
        critical_errors.append(f"CRITICAL: {perr}")
    else:
        soft_errors += check_excel_structural(wb)
        critical_errors += critical_position_analysis(wb)
        critical_errors += critical_expiry_alerts(wb)
        critical_errors += critical_summary(wb)

    # --- GCal ---
    try:
        gcal_rows = _gcal_events()
        soft_errors += check_gcal_structural(gcal_rows)
        critical_errors += critical_gcal(gcal_rows)
    except Exception as e:  # noqa: BLE001
        critical_errors.append(f"CRITICAL: Error checking GCal: {e}")

    # --- Critical gate ---
    if critical_errors:
        print("=== CRITICAL CHECK FAILURES ===")
        for e in critical_errors[:20]:
            print(f"  {e}")
        print(f"\n=== RESULT: FAIL ({len(critical_errors)} critical errors) ===")
        sys.exit(1)

    # --- Accuracy gate (>=70) over soft/structural checks ---
    total_soft = 8  # число структурных пунктов как знаменатель
    passed = max(0, total_soft - len(soft_errors))
    accuracy = passed / total_soft * 100
    print(f"  Structural soft errors: {len(soft_errors)}")
    for e in soft_errors[:10]:
        print(f"    NOTE: {e}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
