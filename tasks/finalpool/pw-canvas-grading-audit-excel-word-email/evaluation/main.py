"""Evaluation script for pw-canvas-grading-audit-excel-word-email."""
import os
import argparse, json, os, sys, glob as globmod
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Benchmark values published on the mock dashboard at http://localhost:30336.
# An honest playwright extraction must surface at least one of these (numbers
# or level labels) somewhere in the deliverables.
BENCHMARK_NUMBERS = ["78.5", "71.2", "65.8", "82.1"]
BENCHMARK_LEVELS = ["introductory", "intermediate", "advanced", "graduate"]
# RU localizations the agent may legitimately use instead of the EN labels.
BENCHMARK_LEVELS_RU = ["начальн", "средн", "продвинут", "магистр", "выпускн", "базов"]


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}: {detail_str}")
        if critical:
            CRITICAL_FAILS.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    # ---- shared accumulators for cross-deliverable benchmark search ----
    benchmark_corpus = []  # lower-cased strings from json/word for benchmark check

    excel_path = os.path.join(agent_workspace, "Grading_Audit_Report.xlsx")
    check("Grading_Audit_Report.xlsx exists", os.path.exists(excel_path))

    data_rows = []          # Data_Analysis data rows (tuples)
    da_headers = []
    metrics_map = {}        # lower metric name -> Value
    rec_rows = []
    rec_headers = []

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---------- Data_Analysis ----------
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")
            da_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Course', 'Code', 'Enrollment', 'Avg_Score', 'Pass_Rate']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in da_headers, f"headers: {da_headers[:8]}")

        # ---------- Metrics ----------
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            m_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(m_rows) >= 4, f"got {len(m_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in m_rows:
                if r and r[0] is not None:
                    metrics_map[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        # ---------- Recommendations ----------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            rec_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")
            rec_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Course']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in rec_headers, f"headers: {rec_headers[:8]}")

        # ---------- Email (structural, broadened RU+EN) ----------
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "SELECT subject, to_addr, body_text FROM email.messages "
                "WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s "
                "OR subject ILIKE %s OR subject ILIKE %s",
                ('%report%', '%analysis%', '%отчёт%', '%отчет%', '%анализ%'))
            emails = cur.fetchall()
            check("Analysis email sent", len(emails) >= 1, f"found {len(emails)} matching emails")
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        # ---------- Word document ----------
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        word_text = ""
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            word_text = " ".join(p.text for p in doc.paragraphs)
            check("Word has content", len(word_text.strip()) > 50, f"text length: {len(word_text)}")
            benchmark_corpus.append(word_text.lower())

        check("course_grading_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "course_grading_processor.py")))

    # ---- gather course_grading_results.json into benchmark corpus ----
    results_path = os.path.join(agent_workspace, "course_grading_results.json")
    if os.path.exists(results_path):
        try:
            with open(results_path, "r", encoding="utf-8") as f:
                benchmark_corpus.append(json.dumps(json.load(f), ensure_ascii=False).lower())
        except Exception:
            try:
                with open(results_path, "r", encoding="utf-8", errors="ignore") as f:
                    benchmark_corpus.append(f.read().lower())
            except Exception:
                pass

    # ============================================================
    # CRITICAL CHECKS — semantic substance. Any failure => exit(1).
    # ============================================================

    # 1) Data_Analysis numeric integrity + sorted by Course.
    has_da_cols = all(c in da_headers for c in ['course', 'code', 'enrollment', 'avg_score', 'pass_rate'])
    if has_da_cols and len(data_rows) >= 6:
        idx = {h: da_headers.index(h) for h in ['course', 'enrollment', 'avg_score', 'pass_rate']}
        ok = True
        bad = ""
        courses = []
        for r in data_rows:
            if r is None or all(c is None for c in r):
                continue
            sc = safe_float(r[idx['avg_score']])
            pr = safe_float(r[idx['pass_rate']])
            en = safe_float(r[idx['enrollment']])
            cv = r[idx['course']]
            courses.append(str(cv).strip().lower() if cv is not None else "")
            if sc is None or not (0 <= sc <= 100):
                ok = False; bad = f"Avg_Score out of range: {r[idx['avg_score']]}"; break
            if pr is None or not (0 <= pr <= 100):
                ok = False; bad = f"Pass_Rate out of range: {r[idx['pass_rate']]}"; break
            if en is None or en <= 0 or abs(en - round(en)) > 1e-6:
                ok = False; bad = f"Enrollment not positive int: {r[idx['enrollment']]}"; break
        sorted_ok = courses == sorted(courses)
        check("CRITICAL Data_Analysis numeric integrity (scores/rates 0-100, enrollment positive int)",
              ok, bad, critical=True)
        check("CRITICAL Data_Analysis sorted alphabetically by Course",
              sorted_ok, "rows not alphabetically sorted by Course", critical=True)
    else:
        check("CRITICAL Data_Analysis numeric integrity (scores/rates 0-100, enrollment positive int)",
              False, "required columns/rows missing", critical=True)
        check("CRITICAL Data_Analysis sorted alphabetically by Course",
              False, "required columns/rows missing", critical=True)

    # 2) Metrics consistency against Data_Analysis.
    if has_da_cols and len(data_rows) >= 6 and metrics_map:
        idx = {h: da_headers.index(h) for h in ['enrollment', 'avg_score', 'pass_rate']}
        valid = [r for r in data_rows if r is not None and not all(c is None for c in r)]
        n = len(valid)
        enr_sum = sum((safe_float(r[idx['enrollment']]) or 0) for r in valid)
        sc_vals = [safe_float(r[idx['avg_score']]) for r in valid if safe_float(r[idx['avg_score']]) is not None]
        pr_vals = [safe_float(r[idx['pass_rate']]) for r in valid if safe_float(r[idx['pass_rate']]) is not None]
        mean_sc = sum(sc_vals) / len(sc_vals) if sc_vals else None
        mean_pr = sum(pr_vals) / len(pr_vals) if pr_vals else None

        tc = safe_float(metrics_map.get('total_courses'))
        te = safe_float(metrics_map.get('total_enrollment'))
        m_sc = safe_float(metrics_map.get('avg_score'))
        m_pr = safe_float(metrics_map.get('avg_pass_rate'))

        problems = []
        if tc is None or int(round(tc)) != n:
            problems.append(f"Total_Courses={metrics_map.get('total_courses')} != rows {n}")
        if te is None or abs(te - enr_sum) > 0.5:
            problems.append(f"Total_Enrollment={metrics_map.get('total_enrollment')} != sum {enr_sum}")
        if m_sc is None or mean_sc is None or abs(m_sc - mean_sc) > 1.0:
            problems.append(f"Avg_Score={metrics_map.get('avg_score')} != mean {mean_sc}")
        if m_pr is None or mean_pr is None or abs(m_pr - mean_pr) > 1.0:
            problems.append(f"Avg_Pass_Rate={metrics_map.get('avg_pass_rate')} != mean {mean_pr}")
        check("CRITICAL Metrics consistent with Data_Analysis (counts/sum/means)",
              len(problems) == 0, "; ".join(problems), critical=True)
    else:
        check("CRITICAL Metrics consistent with Data_Analysis (counts/sum/means)",
              False, "Data_Analysis or Metrics missing required content", critical=True)

    # 3) Benchmark cross-reference: proof the localhost:30336 dashboard was consumed.
    corpus = " ".join(benchmark_corpus)
    found_num = [v for v in BENCHMARK_NUMBERS if v in corpus]
    found_lvl = [v for v in BENCHMARK_LEVELS if v in corpus]
    found_lvl_ru = [v for v in BENCHMARK_LEVELS_RU if v in corpus]
    benchmark_ok = len(found_num) >= 1 or len(found_lvl) >= 2 or len(found_lvl_ru) >= 2
    check("CRITICAL Benchmark cross-reference present (dashboard values consumed)",
          benchmark_ok,
          f"numbers={found_num} en_levels={found_lvl} ru_levels={found_lvl_ru}", critical=True)

    # 4) Recommendations are gap-driven: Course values are real, ordered by Priority,
    #    and the top-priority row targets a low-Pass_Rate course.
    if (has_da_cols and len(data_rows) >= 6 and rec_rows
            and all(c in rec_headers for c in ['priority', 'action', 'course'])):
        ridx = {h: rec_headers.index(h) for h in ['priority', 'course']}
        didx = {h: da_headers.index(h) for h in ['course', 'code', 'pass_rate']}
        real_courses = {}
        for r in data_rows:
            if r is None or all(c is None for c in r):
                continue
            pr = safe_float(r[didx['pass_rate']])
            for key in (r[didx['course']], r[didx['code']]):
                if key is not None:
                    real_courses[str(key).strip().lower()] = pr
        valid_recs = [r for r in rec_rows if r is not None and not all(c is None for c in r)]
        # every recommendation references a real course
        all_real = True
        for r in valid_recs:
            cv = r[ridx['course']]
            if cv is None or str(cv).strip().lower() not in real_courses:
                all_real = False
                break
        check("CRITICAL Recommendations reference real Data_Analysis courses",
              all_real and len(valid_recs) >= 2,
              f"recs={len(valid_recs)} all_real={all_real}", critical=True)

        # ordered by priority + top targets a below-median pass-rate course
        prio = [(safe_float(r[ridx['priority']]), r[ridx['course']]) for r in valid_recs]
        prio_clean = [(p, c) for (p, c) in prio if p is not None]
        ordered_ok = prio_clean == sorted(prio_clean, key=lambda x: x[0])
        top_ok = True
        if prio_clean:
            top_course = str(min(prio_clean, key=lambda x: x[0])[1]).strip().lower()
            top_pr = real_courses.get(top_course)
            pr_values = [v for v in real_courses.values() if v is not None]
            if top_pr is not None and pr_values:
                median = sorted(pr_values)[len(pr_values) // 2]
                top_ok = top_pr <= median  # highest-priority targets a weaker (<=median) course
        check("CRITICAL Recommendations ordered by Priority, top targets low Pass_Rate course",
              ordered_ok and top_ok,
              f"ordered={ordered_ok} top_targets_low={top_ok}", critical=True)
    else:
        check("CRITICAL Recommendations reference real Data_Analysis courses",
              False, "missing Recommendations/Data_Analysis content", critical=True)
        check("CRITICAL Recommendations ordered by Priority, top targets low Pass_Rate course",
              False, "missing Recommendations/Data_Analysis content", critical=True)

    # 5) Email deliverable: exact subject + body references a concrete finding.
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT to_addr, body_text FROM email.messages WHERE subject = %s",
            ('Analysis Report Complete',))
        rows = cur.fetchall()
        conn.close()
        email_ok = False
        detail = "no email with exact subject 'Analysis Report Complete'"
        if rows:
            # build set of concrete findings: course codes + numeric metric values
            findings = set()
            if has_da_cols:
                didx = {h: da_headers.index(h) for h in ['code', 'course']}
                for r in data_rows:
                    if r is None or all(c is None for c in r):
                        continue
                    for k in (r[didx['code']], r[didx['course']]):
                        if k:
                            findings.add(str(k).strip().lower())
            for v in metrics_map.values():
                if v is not None:
                    findings.add(str(v).strip().lower())
            for to_addr, body in rows:
                addr_blob = json.dumps(to_addr).lower() if to_addr is not None else ""
                if 'team-lead@company.com' not in addr_blob:
                    continue
                b = (body or "").lower()
                if any(f and f in b for f in findings):
                    email_ok = True
                    break
            if not email_ok:
                detail = "email exists but missing recipient team-lead@company.com or a concrete finding in body"
        check("CRITICAL Email to team-lead@company.com (exact subject) references a concrete finding",
              email_ok, detail, critical=True)
    except Exception as e:
        check("CRITICAL Email to team-lead@company.com (exact subject) references a concrete finding",
              False, str(e), critical=True)

    # ---- critical gate before accuracy ----
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    if CRITICAL_FAILS:
        print(f"CRITICAL FAILURES: {CRITICAL_FAILS}")
        print(f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%) but critical checks failed.")
        sys.exit(1)

    success = accuracy >= 70.0
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
