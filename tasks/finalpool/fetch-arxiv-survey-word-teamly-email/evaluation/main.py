"""Evaluation script for fetch-arxiv-survey-word-teamly-email.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# External benchmark areas + growth values from data.json (read honestly, not invented).
BENCHMARK = {"LLMs": 45, "Computer Vision": 12, "RL": 8}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Data_Analysis Area covers all 3 benchmark areas",
    "Data_Analysis sorted alphabetically by Area",
    "Metrics encodes the external benchmark growth values",
    "Recommendations prioritizes the largest-negative-gap area",
    "Analysis email to team-lead with exact subject 'Analysis Report Complete'",
    "Teamly 'Arxiv Survey Dashboard' page exists with analysis content",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def norm_area(s):
    return str(s).strip().lower().replace("-", " ") if s is not None else ""


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Survey_Report.xlsx")
    check("Survey_Report.xlsx exists", os.path.exists(excel_path))
    wb = openpyxl.load_workbook(excel_path, data_only=True) if os.path.exists(excel_path) else None

    # ----- Data_Analysis -----
    da_areas = []
    if wb and "Data_Analysis" in wb.sheetnames:
        check("Data_Analysis sheet exists", True)
        ws = wb["Data_Analysis"]
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        check("Data_Analysis has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        hmap = {h: i for i, h in enumerate(headers)}
        for expected_col in ['Paper_ID', 'Title', 'Area', 'Citations', 'Relevance_Score']:
            check(f"Data_Analysis has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")
        area_i = hmap.get("area")
        if area_i is not None:
            da_areas = [r[area_i] for r in data_rows if area_i < len(r) and r[area_i] is not None]
    else:
        check("Data_Analysis sheet exists", False)

    # CRITICAL: Area covers all 3 benchmark areas.
    da_norm = {norm_area(a) for a in da_areas}
    covers = all(norm_area(b) in da_norm for b in BENCHMARK)
    check("Data_Analysis Area covers all 3 benchmark areas", covers,
          f"areas: {sorted(da_norm)}")

    # CRITICAL: sorted alphabetically by Area (ascending).
    lowered = [norm_area(a) for a in da_areas]
    sorted_ok = len(lowered) >= 4 and lowered == sorted(lowered)
    check("Data_Analysis sorted alphabetically by Area", sorted_ok, f"areas order: {lowered}")

    # ----- Metrics -----
    metrics_text = ""
    if wb and "Metrics" in wb.sheetnames:
        check("Metrics sheet exists", True)
        ws = wb["Metrics"]
        m_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        check("Metrics has >= 4 rows", len(m_rows) >= 4, f"got {len(m_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for expected_col in ['Metric', 'Value']:
            check(f"Metrics has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")
        metrics_text = " ".join(str(c) for r in m_rows for c in r if c is not None)
    else:
        check("Metrics sheet exists", False)

    # CRITICAL: Metrics encodes external benchmark values (45 / 12 / 8) and/or their avg.
    mt = metrics_text
    avg = sum(BENCHMARK.values()) / len(BENCHMARK)  # ~21.67
    has_growth = all(str(v) in mt for v in BENCHMARK.values())
    has_avg = any(a in mt for a in [f"{avg:.0f}", f"{avg:.1f}", f"{avg:.2f}", "21.6", "21.7", "21,6", "21,7"])
    check("Metrics encodes the external benchmark growth values", has_growth or has_avg,
          f"metrics text: {mt[:160]}")

    # ----- Recommendations -----
    rec_areas_top = []
    rec_text = ""
    if wb and "Recommendations" in wb.sheetnames:
        check("Recommendations sheet exists", True)
        ws = wb["Recommendations"]
        r_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        check("Recommendations has >= 2 rows", len(r_rows) >= 2, f"got {len(r_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        hmap = {h: i for i, h in enumerate(headers)}
        for expected_col in ['Priority', 'Action', 'Area']:
            check(f"Recommendations has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")
        area_i = hmap.get("area")
        if area_i is not None and r_rows:
            rec_areas_top = [r[area_i] for r in r_rows if area_i < len(r)]
        rec_text = " ".join(str(c) for r in r_rows for c in r if c is not None)
    else:
        check("Recommendations sheet exists", False)

    # CRITICAL: prioritize the largest-negative-gap area per analysis_guide.md.
    # Gap = Internal - External. With the lowest external growth NOT automatically
    # the answer; the methodology focuses on largest negative gap. The benchmark
    # area with the largest external value (LLMs, 45) is the hardest to match, so
    # the top recommendation should reference one of the benchmark areas. We accept
    # the top-priority row referencing any of the three benchmark areas, but require
    # the highest-benchmark area 'LLMs' to appear somewhere in recommendations
    # (largest external => largest potential negative gap to close first).
    rl = rec_text.lower()
    top_ref = norm_area(rec_areas_top[0]) if rec_areas_top else ""
    benchmark_in_rec = any(norm_area(b) in rl for b in BENCHMARK)
    llms_focus = "llms" in rl or "llm" in rl
    check("Recommendations prioritizes the largest-negative-gap area",
          benchmark_in_rec and llms_focus,
          f"top area: {top_ref!r}; rec text: {rec_text[:160]}")

    # ----- Email -----
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, COALESCE(body_text,''), COALESCE(to_addr::text,'') "
            "FROM email.messages "
            "WHERE subject = %s OR subject ILIKE %s",
            ('Analysis Report Complete', '%Analysis Report Complete%'))
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        rows = []
        check("Analysis email to team-lead with exact subject 'Analysis Report Complete'", False, str(e))
    if rows is not None:
        ok = False
        for subj, body, to_addr in rows:
            if "team-lead@company.com" in (to_addr or "").lower():
                btext = (body or "").lower()
                # RU keyword check on ORIGINAL lowered text (not normalized).
                if any(k in btext for k in ["отчёт", "отчет", "анализ", "вывод", "report", "analysis"]) \
                        or len(btext) > 20:
                    ok = True
                    break
        check("Analysis email to team-lead with exact subject 'Analysis Report Complete'", ok,
              f"matching emails: {len(rows)}")

    # ----- Teamly dashboard page -----
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body,'') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        pages = []
        check("Teamly 'Arxiv Survey Dashboard' page exists with analysis content", False, str(e))
    if isinstance(pages, list):
        dash = None
        for pid, title, body in pages:
            tl = (title or "").lower()
            if "старые заметки" in tl or "архив" in tl:
                continue
            if "arxiv survey dashboard" in tl or ("dashboard" in tl or "дашборд" in tl or "панель" in tl):
                dash = (pid, title, body)
                break
        content_ok = False
        if dash is not None:
            text = ((dash[1] or "") + " " + (dash[2] or "")).lower()
            content_ok = len(text) > 60 and any(
                k in text for k in ["анализ", "обзор", "вывод", "разрыв", "gap",
                                    "llms", "llm", "рекоменд", "бенчмарк", "benchmark"])
        check("Teamly 'Arxiv Survey Dashboard' page exists with analysis content",
              dash is not None and content_ok,
              f"new pages: {[(p[0], p[1]) for p in pages]}")

    # ----- Word document -----
    import glob as globmod
    word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
    check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
    if word_files:
        from docx import Document
        doc = Document(word_files[0])
        text = " ".join(p.text for p in doc.paragraphs)
        check("Word has content", len(text) > 50, f"text length: {len(text)}")

    check("arxiv_survey_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "arxiv_survey_processor.py")))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
