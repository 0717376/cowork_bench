"""Оценка для задачи moex-sector-comparison-pdf.

CRITICAL_CHECKS (семантические): любой их провал => общий FAIL независимо от
accuracy. Структурные проверки (наличие листов, заголовки) — НЕ критические;
они учитываются только в проценте accuracy (порог >= 70%).

Критические проверки:
  - Latest_Close каждого тикера MOEX совпадает с эталонным close (tol=2.0)
  - YTD_Return_Pct каждого тикера совпадает с эталоном (tol=1.0)
  - Met_Target (Yes/No) корректен для каждого тикера
  - Target_Return по сектору совпадает с значением из Sector_Targets.pdf (tol=0.5)
"""
import argparse
import os
import sys
import openpyxl


# Множество имён критических проверок. Любой провал => sys.exit(1).
CRITICAL_NAMES = {"Latest_Close", "YTD_Return_Pct", "Met_Target", "Target_Return"}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = False


def check(name, condition, detail=""):
    """Регистрирует результат проверки. name входит в CRITICAL_NAMES => критическая."""
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    critical = name in CRITICAL_NAMES
    tag = " (CRITICAL)" if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  PASS{tag}: {name} {detail}")
    else:
        FAIL_COUNT += 1
        print(f"  FAIL{tag}: {name} {detail}")
        if critical:
            CRITICAL_FAILED = True


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sector_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Sector_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: вывод агента не найден: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: эталон не найден: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- Лист: Stock Prices ----
    print("Проверка листа 'Stock Prices'...")
    a_rows = load_sheet_rows(agent_wb, "Stock Prices")
    g_rows = load_sheet_rows(gt_wb, "Stock Prices")
    check("Sheet_StockPrices", a_rows is not None, "лист присутствует")
    if a_rows is not None and g_rows is not None:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().upper()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(key)
            check("Row_StockPrices", a_row is not None, f"строка тикера {key}")
            if a_row is None:
                continue

            # Sector (структурная строка-идентификатор)
            if len(a_row) > 1 and len(g_row) > 1:
                check("Sector_StockPrices", str_match(a_row[1], g_row[1]),
                      f"{key}: {a_row[1]} vs {g_row[1]}")
            # Latest_Close (CRITICAL, tol=2.0)
            if len(a_row) > 2 and len(g_row) > 2:
                check("Latest_Close", num_close(a_row[2], g_row[2], 2.0),
                      f"{key}: {a_row[2]} vs {g_row[2]} (tol=2.0)")
            # YTD_Return_Pct (CRITICAL, tol=1.0)
            if len(a_row) > 3 and len(g_row) > 3:
                check("YTD_Return_Pct", num_close(a_row[3], g_row[3], 1.0),
                      f"{key}: {a_row[3]} vs {g_row[3]} (tol=1.0)")

    # ---- Лист: Sector Comparison ----
    print("Проверка листа 'Sector Comparison'...")
    a_rows = load_sheet_rows(agent_wb, "Sector Comparison")
    g_rows = load_sheet_rows(gt_wb, "Sector Comparison")
    check("Sheet_SectorComparison", a_rows is not None, "лист присутствует")
    if a_rows is not None and g_rows is not None:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Сопоставление по колонке Stock (col 1)
        a_lookup = {}
        for row in a_data:
            if row and len(row) > 1 and row[1] is not None:
                a_lookup[str(row[1]).strip().upper()] = row
        for g_row in g_data:
            if not g_row or len(g_row) < 2 or g_row[1] is None:
                continue
            key = str(g_row[1]).strip().upper()
            a_row = a_lookup.get(key)
            check("Row_SectorComparison", a_row is not None, f"строка акции {key}")
            if a_row is None:
                continue

            # Sector (структурная)
            check("Sector_SectorComparison", str_match(a_row[0], g_row[0]),
                  f"{key}: {a_row[0]} vs {g_row[0]}")
            # Target_Return (CRITICAL, tol=0.5)
            if len(a_row) > 2 and len(g_row) > 2:
                check("Target_Return", num_close(a_row[2], g_row[2], 0.5),
                      f"{key}: {a_row[2]} vs {g_row[2]} (tol=0.5)")
            # Actual_Return (CRITICAL — это YTD, tol=1.0)
            if len(a_row) > 3 and len(g_row) > 3:
                check("YTD_Return_Pct", num_close(a_row[3], g_row[3], 1.0),
                      f"{key} Actual_Return: {a_row[3]} vs {g_row[3]} (tol=1.0)")
            # Met_Target (CRITICAL)
            if len(a_row) > 4 and len(g_row) > 4:
                check("Met_Target", str_match(a_row[4], g_row[4]),
                      f"{key}: {a_row[4]} vs {g_row[4]}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    if CRITICAL_FAILED:
        print("=== RESULT: FAIL (провалена критическая проверка) ===")
        sys.exit(1)

    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)

    print("=== RESULT: FAIL (accuracy < 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
