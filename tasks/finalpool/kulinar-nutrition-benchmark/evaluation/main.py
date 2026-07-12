"""
Evaluation for kulinar-nutrition-benchmark task.

The agent freely chooses >= 8 kulinar recipes from >= 3 categories, estimates
their nutrition from ingredient_nutrition.csv, classifies them against the
benchmark JSON, builds a 3-day meal plan and a Summary, and writes a Word report.

Because recipe selection is the agent's free choice, the evaluation does NOT
compare against any hardcoded ground-truth workbook. Instead it derives semantic
correctness directly from the agent's own data plus the live reference sources
(benchmark JSON, ingredient_nutrition.csv, kulinar recipe catalog).

Critical checks (see CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import sys
import urllib.request

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

MOCK_URL = "http://localhost:30232/api/nutrition_benchmarks.json"

# Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Est_Calories consistent with macros (Atwater) for analyzed recipes",
    "Meets_Guidelines classification matches per-meal benchmark",
    "Each meal-plan day totals within daily calorie benchmark",
    "Each meal-plan day has exactly 3 meals",
    "Shellfish/allergy excluded from analysis and meal plan",
    "Summary aggregates consistent with Recipe Analysis",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num(v):
    try:
        return float(str(v).replace(",", ".").strip())
    except (TypeError, ValueError, AttributeError):
        return None


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


# ---------------------------------------------------------------------------
# Reference data loaders
# ---------------------------------------------------------------------------
def task_root():
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def gym_root():
    return os.path.normpath(os.path.join(task_root(), "..", "..", ".."))


def load_benchmarks():
    """Live fetch from the mock server; fall back to the shipped JSON file."""
    try:
        with urllib.request.urlopen(MOCK_URL, timeout=5) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception:
        for p in (
            os.path.join(task_root(), "files", "mock_pages", "api", "nutrition_benchmarks.json"),
            os.path.join(task_root(), "tmp", "mock_pages", "api", "nutrition_benchmarks.json"),
        ):
            if os.path.isfile(p):
                with open(p, encoding="utf-8") as f:
                    return json.load(f)
    return None


def load_kulinar_shellfish_names():
    """Names of kulinar recipes whose ingredients include shellfish/seafood."""
    path = os.path.join(gym_root(), "local_servers", "kulinar-mcp",
                        "src", "data", "all_recipes.json")
    out = set()
    kw = ("креветк", "морепродукт", "краб", "мидии", "кальмар", "устриц", "лангуст", "омар")
    try:
        with open(path, encoding="utf-8") as f:
            recipes = json.load(f)
        for r in recipes:
            ings = " ".join(norm(i.get("name", "")) for i in r.get("ingredients", []))
            if any(k in ings for k in kw):
                out.add(norm(r.get("name", "")))
    except Exception:
        pass
    return out


# ---------------------------------------------------------------------------
# Excel checks
# ---------------------------------------------------------------------------
def parse_recipe_analysis(wb):
    rows = load_sheet_rows(wb, "Recipe Analysis") or load_sheet_rows(wb, "Recipe_Analysis")
    if rows is None:
        return None, None
    header = rows[0]
    data = [r for r in rows[1:] if r and r[0] is not None and str(r[0]).strip()]
    cols = {
        "name": find_col(header, ["Recipe_Name", "Recipe Name", "Name"]),
        "cat": find_col(header, ["Category"]),
        "cal": find_col(header, ["Est_Calories", "Est Calories", "Estimated_Calories", "Calories"]),
        "prot": find_col(header, ["Est_Protein", "Est Protein", "Protein"]),
        "carb": find_col(header, ["Est_Carbs", "Est Carbs", "Carbs"]),
        "fat": find_col(header, ["Est_Fat", "Est Fat", "Fat"]),
        "meets": find_col(header, ["Meets_Guidelines", "Meets Guidelines", "Guidelines"]),
    }
    parsed = []
    for r in data:
        def g(key):
            i = cols[key]
            return r[i] if (i is not None and i < len(r)) else None
        parsed.append({
            "name": g("name"), "cat": g("cat"),
            "cal": num(g("cal")), "prot": num(g("prot")),
            "carb": num(g("carb")), "fat": num(g("fat")),
            "meets": norm(g("meets")),
        })
    return parsed, cols


def meets_is_yes(v):
    return v in ("yes", "да", "y", "true", "+")


def meets_is_no(v):
    return v in ("no", "нет", "n", "false", "-")


def check_excel(agent_workspace, benchmarks):
    print("\n=== Checking Excel ===")
    path = os.path.join(agent_workspace, "Nutrition_Benchmark.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Structural (non-critical) ----
    ra, cols = parse_recipe_analysis(wb)
    if ra is None:
        record("Sheet 'Recipe Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Sheet 'Recipe Analysis' exists", True)
    record("Recipe Analysis has >= 8 rows", len(ra) >= 8, f"Found {len(ra)}")

    cats = {norm(r["cat"]) for r in ra if r["cat"]}
    record("Recipes span >= 3 categories", len(cats) >= 3, f"Categories: {sorted(cats)}")

    if cols["meets"] is not None:
        has = any(meets_is_yes(r["meets"]) or meets_is_no(r["meets"]) for r in ra)
        record("Meets_Guidelines has Yes/No values", has,
               f"Values: {sorted({r['meets'] for r in ra})}")
    else:
        record("Meets_Guidelines column exists", False)

    # ---- CRITICAL 1: per-recipe calories consistent with reported macros (Atwater) ----
    # calories ~= 4*protein + 4*carbs + 9*fat. Verifies the estimates are
    # internally derived rather than invented. Generous tolerance.
    checkable = [r for r in ra if None not in (r["cal"], r["prot"], r["carb"], r["fat"])]
    if len(checkable) >= 3:
        ok_count = 0
        sample = checkable[:8]
        for r in sample:
            atwater = 4 * r["prot"] + 4 * r["carb"] + 9 * r["fat"]
            tol = max(0.35 * max(atwater, r["cal"]), 250)
            if abs(r["cal"] - atwater) <= tol and 50 <= r["cal"] <= 2500:
                ok_count += 1
        # require a clear majority of sampled recipes to be macro-consistent
        passed = ok_count >= max(3, (len(sample) + 1) // 2)
        record("Est_Calories consistent with macros (Atwater) for analyzed recipes",
               passed, f"{ok_count}/{len(sample)} consistent")
    else:
        record("Est_Calories consistent with macros (Atwater) for analyzed recipes",
               False, f"Only {len(checkable)} recipes have full macro columns")

    # ---- CRITICAL 2: Meets_Guidelines vs per-meal benchmark ----
    # Per-meal calorie ceiling = daily max / 3. A recipe whose Est_Calories clearly
    # exceeds it must be classified No/Нет; one clearly within range must be Yes/Да.
    if benchmarks and "calories" in benchmarks and cols["meets"] is not None:
        cal_max = benchmarks["calories"].get("max")
        cal_min = benchmarks["calories"].get("min")
        meal_max = cal_max / 3.0 if cal_max else None
        meal_min = cal_min / 3.0 if cal_min else None
        violations = []
        classified = [r for r in ra if r["cal"] is not None and (meets_is_yes(r["meets"]) or meets_is_no(r["meets"]))]
        for r in classified:
            if meal_max is not None and r["cal"] > meal_max * 1.10 and not meets_is_no(r["meets"]):
                violations.append(f"{r['name']} ({r['cal']} kcal) should be No")
            if meal_min is not None and meal_max is not None and \
               (meal_min * 0.9) <= r["cal"] <= meal_max and not meets_is_yes(r["meets"]):
                # in-range recipes should generally be Yes (other nutrients may flip it,
                # so only flag if calories are comfortably inside the window)
                if (meal_min) <= r["cal"] <= (meal_max * 0.95) and meets_is_no(r["meets"]):
                    violations.append(f"{r['name']} ({r['cal']} kcal) in-range but No")
        record("Meets_Guidelines classification matches per-meal benchmark",
               len(violations) == 0 and len(classified) >= 4,
               f"violations={violations[:4]}; classified={len(classified)}")
    else:
        record("Meets_Guidelines classification matches per-meal benchmark",
               False, "Missing benchmarks or Meets column")

    # ---- Meal Plan ----
    mp_rows = load_sheet_rows(wb, "Meal Plan Suggestions") or load_sheet_rows(wb, "Meal_Plan_Suggestions")
    mp_recipe_names = set()
    if mp_rows is None:
        record("Sheet 'Meal Plan Suggestions' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Meal Plan Suggestions' exists", True)
        h = mp_rows[0]
        day_c = find_col(h, ["Day"])
        meal_c = find_col(h, ["Meal"])
        name_c = find_col(h, ["Recipe_Name", "Recipe Name", "Name"])
        cal_c = find_col(h, ["Est_Calories", "Est Calories", "Calories"])
        data2 = [r for r in mp_rows[1:] if r and any(c is not None for c in r)]
        record("Meal Plan has >= 9 rows (3 days x 3 meals)", len(data2) >= 9, f"Found {len(data2)}")

        # group by day
        days = {}
        for r in data2:
            day = norm(r[day_c]) if (day_c is not None and day_c < len(r)) else ""
            cal = num(r[cal_c]) if (cal_c is not None and cal_c < len(r)) else None
            nm = norm(r[name_c]) if (name_c is not None and name_c < len(r)) else ""
            if nm:
                mp_recipe_names.add(nm)
            days.setdefault(day, []).append(cal)

        # CRITICAL 3: each day total within daily calorie range (live from benchmark)
        if benchmarks and "calories" in benchmarks and days:
            lo = benchmarks["calories"].get("min", 0)
            hi = benchmarks["calories"].get("max", 10 ** 9)
            day_ok = []
            for day, cals in days.items():
                valid = [c for c in cals if c is not None]
                total = sum(valid)
                # allow 5% slack on each side for rounding
                day_ok.append(lo * 0.95 <= total <= hi * 1.05)
            passed = len(days) >= 3 and all(day_ok)
            record("Each meal-plan day totals within daily calorie benchmark", passed,
                   f"daily totals={[round(sum(c for c in v if c is not None)) for v in days.values()]}, range={lo}-{hi}")
        else:
            record("Each meal-plan day totals within daily calorie benchmark", False,
                   "Missing benchmarks or day grouping")

        # CRITICAL 4: each day has exactly 3 meals
        if days:
            counts = {d: len(v) for d, v in days.items()}
            passed = len(days) >= 3 and all(c == 3 for c in counts.values())
            record("Each meal-plan day has exactly 3 meals", passed, f"meals per day={counts}")
        else:
            record("Each meal-plan day has exactly 3 meals", False, "No days parsed")

    # ---- CRITICAL 5: shellfish/allergy exclusion ----
    shellfish = load_kulinar_shellfish_names()
    if shellfish:
        ra_names = {norm(r["name"]) for r in ra if r["name"]}
        offenders = (ra_names | mp_recipe_names) & shellfish
        record("Shellfish/allergy excluded from analysis and meal plan",
               len(offenders) == 0, f"offending recipes: {sorted(offenders)}")
    else:
        # no shellfish recipe in catalog => exclusion is vacuously satisfied
        record("Shellfish/allergy excluded from analysis and meal plan", True,
               "no shellfish recipe in catalog")

    # ---- CRITICAL 6: Summary aggregates consistent with Recipe Analysis ----
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        record("Summary aggregates consistent with Recipe Analysis", False, "No Summary sheet")
    else:
        record("Sheet 'Summary' exists", True)
        metrics = {}
        for row in sum_rows[1:]:
            if row and row[0]:
                metrics[norm(row[0]).replace(" ", "_")] = row[1] if len(row) > 1 else None

        def metric(*needles):
            for k, v in metrics.items():
                if all(n in k for n in needles):
                    return v
            return None

        problems = []
        total = num(metric("total", "recip"))
        if total is None or int(total) != len(ra):
            problems.append(f"Total_Recipes_Analyzed={metric('total','recip')} != {len(ra)}")

        meeting = num(metric("meeting") or metric("meet"))
        yes_count = sum(1 for r in ra if meets_is_yes(r["meets"]))
        if meeting is None or int(meeting) != yes_count:
            problems.append(f"Recipes_Meeting_Guidelines={metric('meet')} != {yes_count}")

        cals = [(r["name"], r["cal"]) for r in ra if r["cal"] is not None]
        if cals:
            avg_m = num(metric("average", "calor") or metric("avg", "calor"))
            real_avg = sum(c for _, c in cals) / len(cals)
            if avg_m is None or abs(avg_m - real_avg) > max(0.05 * real_avg, 5):
                problems.append(f"Average_Calories={metric('average','calor')} != {round(real_avg,1)}")

            hi_name = max(cals, key=lambda x: x[1])[0]
            lo_name = min(cals, key=lambda x: x[1])[0]
            hi_m = metric("highest")
            lo_m = metric("lowest")
            # containment: tolerate suffixes like " (554 ккал)"
            if hi_m is None or norm(hi_name) not in norm(hi_m):
                problems.append(f"Highest_Calorie_Recipe={hi_m} != {hi_name}")
            if lo_m is None or norm(lo_name) not in norm(lo_m):
                problems.append(f"Lowest_Calorie_Recipe={lo_m} != {lo_name}")

        record("Summary aggregates consistent with Recipe Analysis",
               len(problems) == 0, "; ".join(problems[:5]))


# ---------------------------------------------------------------------------
# Word checks
# ---------------------------------------------------------------------------
def check_word(workspace):
    print("\n=== Checking Word Document ===")
    path = os.path.join(workspace, "Nutrition_Report.docx")
    if not os.path.isfile(path):
        record("Word document exists", False, f"Not found: {path}")
        return
    record("Word document exists", True)
    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join(p.text for p in doc.paragraphs).lower()
        record("Report has substantial content", len(text) > 300, f"Only {len(text)} chars")
        record("Mentions nutrition/calories",
               any(k in text for k in ("calori", "nutrition", "питани", "калори", "пищев")))
        record("Mentions benchmark/guideline",
               any(k in text for k in ("benchmark", "guideline", "норматив", "рекоменд")))
        record("Mentions meal plan",
               ("meal" in text and "plan" in text) or ("план" in text and ("пита" in text or "меню" in text)) or "меню" in text)
    except Exception as e:
        record("Word readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    benchmarks = load_benchmarks()
    if benchmarks is None:
        print("[warn] Could not load nutrition benchmarks (mock server + files unavailable)")

    check_excel(args.agent_workspace, benchmarks)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
