"""
Evaluation for scholarly-arxiv-cross-reference task.
Checks Excel sheets and email.

Structure:
  - CRITICAL_CHECKS: semantic checks reflecting the core deliverable. Any failure
    => sys.exit(1) immediately, regardless of overall accuracy.
  - Non-critical structural checks (sheet exists, column present): gated by
    accuracy >= 70.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# scholarly.arxiv_papers has 5 papers, arxiv.papers has 6
# Overlap IDs: 1602.05629, 1812.06127, 1908.07873 (3 papers)
# Scholarly only: 2001.08361, 2005.14165
# Arxiv only: 1207.00580, 1502.03167, 1912.04977
OVERLAP_IDS = {"1602.05629", "1812.06127", "1908.07873"}
ARXIV_ONLY_IDS = {"1207.00580", "1502.03167", "1912.04977"}
SCHOLARLY_ONLY_IDS = {"2001.08361", "2005.14165"}


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILED.append(name)


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == c:
                return i
    return None


def check_excel(agent_workspace):
    """Check Excel file."""
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Citation_Cross_Reference.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}", critical=True)
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e), critical=True)
        return
    record("Excel readable", True)

    # --- Scholarly_Papers sheet ---
    scholarly_rows = load_sheet_rows(wb, "Scholarly_Papers")
    if scholarly_rows is None:
        scholarly_rows = load_sheet_rows(wb, "Scholarly Papers")
    if scholarly_rows is not None:
        record("Sheet 'Scholarly_Papers' exists", True)
        data_rows = [r for r in scholarly_rows[1:] if any(c is not None and str(c).strip() for c in r)]
        # scholarly.arxiv_papers has exactly 5 papers (CRITICAL exact count)
        record("Scholarly_Papers has exactly 5 rows", len(data_rows) == 5,
               f"Found {len(data_rows)} rows", critical=True)
    else:
        record("Sheet 'Scholarly_Papers' exists", False, f"Available: {wb.sheetnames}", critical=True)

    # --- Arxiv_Papers sheet ---
    arxiv_rows = load_sheet_rows(wb, "Arxiv_Papers")
    if arxiv_rows is None:
        arxiv_rows = load_sheet_rows(wb, "Arxiv Papers")
    if arxiv_rows is not None:
        record("Sheet 'Arxiv_Papers' exists", True)
        data_rows = [r for r in arxiv_rows[1:] if any(c is not None and str(c).strip() for c in r)]
        # arxiv.papers has exactly 6 papers (CRITICAL exact count)
        record("Arxiv_Papers has exactly 6 rows", len(data_rows) == 6,
               f"Found {len(data_rows)} rows", critical=True)
    else:
        record("Sheet 'Arxiv_Papers' exists", False, f"Available: {wb.sheetnames}", critical=True)

    # --- Overlap_Analysis sheet ---
    overlap_rows = load_sheet_rows(wb, "Overlap_Analysis")
    if overlap_rows is None:
        overlap_rows = load_sheet_rows(wb, "Overlap Analysis")
    if overlap_rows is not None:
        record("Sheet 'Overlap_Analysis' exists", True)
        header = overlap_rows[0] if overlap_rows else []
        data_rows = [r for r in overlap_rows[1:] if any(c is not None and str(c).strip() for c in r)]

        id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        in_scholarly_col = find_col(header, ["In_Scholarly", "In Scholarly", "in_scholarly"])
        in_arxiv_col = find_col(header, ["In_Arxiv", "In Arxiv", "in_arxiv"])

        record("Overlap has Paper_ID column", id_col is not None, f"Header: {header}")
        record("Overlap has In_Scholarly column", in_scholarly_col is not None, f"Header: {header}")
        record("Overlap has In_Arxiv column", in_arxiv_col is not None, f"Header: {header}")

        # Total rows should be the union of both databases = 8 (exact count)
        record("Overlap_Analysis has exactly 8 rows", len(data_rows) == 8,
               f"Found {len(data_rows)} rows")

        # Verify the Yes/No flags per category.
        if id_col is not None and in_scholarly_col is not None and in_arxiv_col is not None:
            def flags(pid):
                for row in data_rows:
                    if id_col < len(row) and row[id_col] and str(row[id_col]).strip() == pid:
                        s_val = str(row[in_scholarly_col]).strip().lower() if in_scholarly_col < len(row) and row[in_scholarly_col] is not None else ""
                        a_val = str(row[in_arxiv_col]).strip().lower() if in_arxiv_col < len(row) and row[in_arxiv_col] is not None else ""
                        return s_val, a_val
                return None, None

            # Overlap IDs: In_Scholarly=Yes AND In_Arxiv=Yes
            overlap_ok = 0
            for pid in OVERLAP_IDS:
                s_val, a_val = flags(pid)
                if s_val is not None and "yes" in s_val and "yes" in a_val:
                    overlap_ok += 1
            record("Overlap IDs marked In_Scholarly=Yes / In_Arxiv=Yes",
                   overlap_ok == len(OVERLAP_IDS),
                   f"{overlap_ok}/{len(OVERLAP_IDS)} overlap IDs correct", critical=True)

            # Arxiv-only IDs: In_Scholarly=No AND In_Arxiv=Yes
            arxiv_only_ok = 0
            for pid in ARXIV_ONLY_IDS:
                s_val, a_val = flags(pid)
                if s_val is not None and "no" in s_val and "yes" in a_val:
                    arxiv_only_ok += 1
            record("Arxiv-only IDs marked In_Scholarly=No / In_Arxiv=Yes",
                   arxiv_only_ok == len(ARXIV_ONLY_IDS),
                   f"{arxiv_only_ok}/{len(ARXIV_ONLY_IDS)} arxiv-only IDs correct", critical=True)

            # Scholarly-only IDs: In_Scholarly=Yes AND In_Arxiv=No
            scholarly_only_ok = 0
            for pid in SCHOLARLY_ONLY_IDS:
                s_val, a_val = flags(pid)
                if s_val is not None and "yes" in s_val and "no" in a_val:
                    scholarly_only_ok += 1
            record("Scholarly-only IDs marked In_Scholarly=Yes / In_Arxiv=No",
                   scholarly_only_ok == len(SCHOLARLY_ONLY_IDS),
                   f"{scholarly_only_ok}/{len(SCHOLARLY_ONLY_IDS)} scholarly-only IDs correct", critical=True)
    else:
        record("Sheet 'Overlap_Analysis' exists", False, f"Available: {wb.sheetnames}", critical=True)


def check_email():
    """Check email was sent."""
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find sent folder ID
        cur.execute("SELECT id FROM email.folders WHERE name='Sent'")
        sent_row = cur.fetchone()
        if not sent_row:
            cur.execute("SELECT id FROM email.folders WHERE name ILIKE '%%sent%%' LIMIT 1")
            sent_row = cur.fetchone()

        # Check for email with cross-reference subject (subject stays English)
        cur.execute("""
            SELECT id, subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE subject ILIKE '%%cross%%reference%%'
               OR subject ILIKE '%%cross-reference%%'
        """)
        emails = cur.fetchall()

        if not emails:
            # Broader search
            cur.execute("""
                SELECT id, subject, from_addr, to_addr, body_text
                FROM email.messages
                WHERE subject ILIKE '%%scholarly%%'
                   OR subject ILIKE '%%arxiv%%'
            """)
            emails = cur.fetchall()

        record("Email with cross-reference subject sent", len(emails) > 0,
               "No matching email found", critical=True)

        if emails:
            email = emails[0]
            from_addr = email[2]
            to_addr = email[3]
            if isinstance(to_addr, str):
                try:
                    to_addr = json.loads(to_addr)
                except Exception:
                    pass

            # from_addr validation (CRITICAL)
            from_str = str(from_addr).lower()
            record("Email from librarian@university.edu",
                   "librarian@university.edu" in from_str,
                   f"From: {from_addr}", critical=True)

            # to_addr validation (CRITICAL)
            to_str = str(to_addr).lower()
            record("Email to research-lead@university.edu",
                   "research-lead@university.edu" in to_str,
                   f"To: {to_addr}", critical=True)

            # Body keyword check on ORIGINAL .lower() text (NOT normalized).
            # Accept RU or EN summary wording.
            body = str(email[4]).lower() if email[4] else ""
            en_kw = ("overlap" in body or "common" in body or "shared" in body)
            ru_kw = ("пересеч" in body or "совпад" in body or "общи" in body or "уникальн" in body)
            record("Email body mentions overlap/unique analysis (RU or EN)",
                   en_kw or ru_kw,
                   f"Body preview: {body[:200]}", critical=True)

            # Body should state correct totals: 5 scholarly, 6 arxiv, 3 overlapping.
            has_5 = "5" in body
            has_6 = "6" in body
            has_3 = "3" in body
            record("Email body states totals 5 / 6 / 3",
                   has_5 and has_6 and has_3,
                   f"5={has_5} 6={has_6} 3={has_3} | preview: {body[:200]}", critical=True)

        conn.close()
    except Exception as e:
        record("Email check", False, str(e), critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed (accuracy {accuracy:.1f}%) ===")

    if CRITICAL_FAILED:
        print(f"CRITICAL checks failed ({len(CRITICAL_FAILED)}): {CRITICAL_FAILED}")
        sys.exit(1)

    if accuracy >= 70.0:
        print("All critical checks passed and accuracy >= 70%. PASS")
        sys.exit(0)
    else:
        print(f"Accuracy {accuracy:.1f}% < 70%. FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
