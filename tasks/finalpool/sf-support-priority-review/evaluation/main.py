"""Evaluation for sf-support-priority-review.

Swap: snowflake -> clickhouse. The TICKETS data lives in sf_data."SUPPORT_CENTER__PUBLIC__TICKETS".
Realia data values (REPORTER/ISSUE_TYPE/SHORT_DESCRIPTION/...) are russified CENTRALLY by the
ClickHouse relabel map; PRIORITY stays English (High/Medium/Low). All expectations here are read
LIVE from the DB (counts, avg response, avg CSAT) so the eval auto-syncs with the seed -- no English
realia literals and no volatile numbers are hardcoded.

Scoring: any CRITICAL_CHECKS failure => overall FAIL regardless of accuracy. Otherwise PASS requires
accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

# SLA target response time (hours) per priority.
SLA_TARGETS = {"high": 4, "medium": 8, "low": 24}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "CRITICAL: High SLA_Met == No",
    "CRITICAL: Medium SLA_Met == No",
    "CRITICAL: Low SLA_Met == No",
    "CRITICAL: Priorities_Meeting_SLA exact",
    "CRITICAL: Priorities_Missing_SLA exact",
    "CRITICAL: Total_Tickets exact",
    "CRITICAL: High Avg_Response_Hrs",
    "CRITICAL: Medium Avg_Response_Hrs",
    "CRITICAL: Low Avg_Response_Hrs",
    "CRITICAL: Email to support-manager flags SLA miss",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_expected():
    """Compute all expected values LIVE from the DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute(
        'SELECT "PRIORITY", COUNT(*), '
        'ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2), '
        'ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2) '
        'FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" '
        'GROUP BY "PRIORITY"'
    )
    pstats = {}
    for prio, cnt, avg_resp, avg_csat in cur.fetchall():
        key = str(prio).strip().lower()
        target = SLA_TARGETS.get(key)
        avg_resp_f = float(avg_resp)
        pstats[key] = {
            "priority": prio,
            "count": int(cnt),
            "avg_resp": avg_resp_f,
            "avg_csat": float(avg_csat),
            "sla_target": target,
            "sla_met": "Yes" if (target is not None and avg_resp_f <= target) else "No",
        }
    cur.execute(
        'SELECT COUNT(*), '
        'ROUND((SUM("CUSTOMER_SATISFACTION") / COUNT(*))::numeric, 2) '
        'FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"'
    )
    total, overall_csat = cur.fetchone()
    conn.close()

    meeting = sum(1 for p in pstats.values() if p["sla_met"] == "Yes")
    missing = sum(1 for p in pstats.values() if p["sla_met"] == "No")
    return {
        "priority_stats": pstats,
        "total_tickets": int(total),
        "overall_csat": float(overall_csat),
        "priorities_meeting_sla": meeting,
        "priorities_missing_sla": missing,
    }


def sheet_dicts(wb, name):
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return []
            hdrs = [str(h).strip() if h is not None else "" for h in rows[0]]
            return [{hdrs[i]: row[i] for i in range(len(hdrs))} for row in rows[1:]
                    if not all(v is None for v in row)]
    return None


def check_excel(ws_path, exp):
    print("\n=== Checking Excel ===")
    p = os.path.join(ws_path, "Support_Priority_Report.xlsx")
    if not os.path.isfile(p):
        record("Excel file exists", False, p)
        # Cascade-fail the criticals that depend on the file so the run cannot pass.
        for n in ("CRITICAL: High SLA_Met == No", "CRITICAL: Medium SLA_Met == No",
                  "CRITICAL: Low SLA_Met == No", "CRITICAL: Total_Tickets exact",
                  "CRITICAL: Priorities_Meeting_SLA exact", "CRITICAL: Priorities_Missing_SLA exact",
                  "CRITICAL: High Avg_Response_Hrs", "CRITICAL: Medium Avg_Response_Hrs",
                  "CRITICAL: Low Avg_Response_Hrs"):
            record(n, False, "Excel missing")
        return
    record("Excel file exists", True)
    wb = openpyxl.load_workbook(p, data_only=True)

    # --- Priority Analysis ---
    d = sheet_dicts(wb, "Priority Analysis")
    if d is None:
        record("Sheet Priority Analysis", False, str(wb.sheetnames))
        d = []
    else:
        record("Sheet Priority Analysis", True)
    lookup = {str(r.get("Priority")).strip().lower(): r for r in d if r.get("Priority") is not None}

    for key, e in exp["priority_stats"].items():
        label = e["priority"]
        m = lookup.get(key)
        if m is None:
            record(f"Priority {label} row present", False, "Missing")
            record(f"CRITICAL: {label} SLA_Met == No", False, "row missing")
            record(f"CRITICAL: {label} Avg_Response_Hrs", False, "row missing")
            continue
        record(f"Priority {label} row present", True)
        record(f"Priority {label} Ticket_Count", num_close(m.get("Ticket_Count"), e["count"], 1),
               f"{m.get('Ticket_Count')} vs {e['count']}")
        # Avg_Response_Hrs drives the SLA conclusion -> tight tol, critical.
        record(f"CRITICAL: {label} Avg_Response_Hrs", num_close(m.get("Avg_Response_Hrs"), e["avg_resp"], 0.1),
               f"{m.get('Avg_Response_Hrs')} vs {e['avg_resp']}")
        record(f"Priority {label} SLA_Target_Hrs", num_close(m.get("SLA_Target_Hrs"), e["sla_target"], 0.1),
               f"{m.get('SLA_Target_Hrs')} vs {e['sla_target']}")
        # Core analytical conclusion -> strict string match, critical.
        record(f"CRITICAL: {label} SLA_Met == No", str_match(m.get("SLA_Met"), e["sla_met"]),
               f"{m.get('SLA_Met')} vs {e['sla_met']}")
        record(f"Priority {label} Avg_CSAT", num_close(m.get("Avg_CSAT"), e["avg_csat"], 0.1),
               f"{m.get('Avg_CSAT')} vs {e['avg_csat']}")

    # Sort-order: Priority alphabetical.
    prios = [str(r.get("Priority")).strip() for r in d if r.get("Priority") is not None]
    record("Priority Analysis sorted alphabetically", prios == sorted(prios), f"{prios}")

    # --- Summary ---
    d = sheet_dicts(wb, "Summary")
    if d is None:
        record("Sheet Summary", False, str(wb.sheetnames))
        record("CRITICAL: Total_Tickets exact", False, "Summary missing")
        record("CRITICAL: Priorities_Meeting_SLA exact", False, "Summary missing")
        record("CRITICAL: Priorities_Missing_SLA exact", False, "Summary missing")
    else:
        record("Sheet Summary", True)
        ms = {str(r.get("Metric", "")).strip().lower(): r.get("Value") for r in d}
        # Deterministic COUNT(*) -> exact (tol<=1), critical.
        record("CRITICAL: Total_Tickets exact", num_close(ms.get("total_tickets"), exp["total_tickets"], 1),
               f"{ms.get('total_tickets')} vs {exp['total_tickets']}")
        record("Summary Overall_Avg_CSAT", num_close(ms.get("overall_avg_csat"), exp["overall_csat"], 0.1),
               f"{ms.get('overall_avg_csat')} vs {exp['overall_csat']}")
        # SLA pass/fail counts -> EXACT integer match, critical (was meaningless at tol=10).
        record("CRITICAL: Priorities_Meeting_SLA exact",
               num_close(ms.get("priorities_meeting_sla"), exp["priorities_meeting_sla"], 0),
               f"{ms.get('priorities_meeting_sla')} vs {exp['priorities_meeting_sla']}")
        record("CRITICAL: Priorities_Missing_SLA exact",
               num_close(ms.get("priorities_missing_sla"), exp["priorities_missing_sla"], 0),
               f"{ms.get('priorities_missing_sla')} vs {exp['priorities_missing_sla']}")
    wb.close()


def check_email(exp):
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
    rows = cur.fetchall()
    conn.close()

    found = None
    for subject, to_addr, body_text in rows:
        subj = (subject or "").lower()
        to_str = str(to_addr or "").lower()
        if "support priority analysis report" in subj and "support-manager@company.com" in to_str:
            found = (subject, to_addr, body_text)
            break
    if found is None:
        # Fallback: match by recipient alone (subject still checked separately).
        for subject, to_addr, body_text in rows:
            if "support-manager@company.com" in str(to_addr or "").lower():
                found = (subject, to_addr, body_text)
                break

    record("Email sent to support-manager@company.com",
           found is not None, f"emails={len(rows)}")
    if found is None:
        record("CRITICAL: Email to support-manager flags SLA miss", False, "No email found")
        return

    subject, to_addr, body_text = found
    record("Email subject is 'Support Priority Analysis Report'",
           "support priority analysis report" in (subject or "").lower(), f"{subject}")

    # Body must flag at least one priority missing SLA. RU + EN keywords on ORIGINAL .lower()
    # (NOT normalize() -- these are RU free-text keywords).
    body = (body_text or "").lower()
    kw = ["sla", "не уклад", "превыша", "просроч", "не соответств", "нарушен",
          "miss", "not met", "exceed", "breach"]
    record("CRITICAL: Email to support-manager flags SLA miss",
           any(k in body for k in kw),
           f"No SLA-miss keyword in body: {body[:150]}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    exp = get_expected()
    check_excel(args.agent_workspace, exp)
    check_email(exp)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy, "critical_failed": critical_failed,
                       "success": success}, f)
    print(f"\n=== RESULT: {'PASS' if success else 'FAIL'} ===")
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
