"""Evaluation for terminal-sf-hr-excel-word-gcal-email (ClickHouse, RU).

Data lives in PostgreSQL schema sf_data, table HR_ANALYTICS__PUBLIC__EMPLOYEES
(the ClickHouse-fronted HR_ANALYTICS warehouse). Department labels are russified
centrally by the deterministic relabel map, so the agent reads/writes Russian
department names. This eval derives ALL expected numbers and department labels
LIVE from the DB, so seed<->eval<->agent stay in sync regardless of language.
No realia literals are hardcoded here.

CRITICAL_FAILS gate any single semantic failure to a hard FAIL before the
accuracy threshold (accuracy >= 70). Structural checks (sheet/column/file
existence) are non-critical.

Checks:
1. Performance_Review_Report.xlsx with 4 sheets (+ CRITICAL per-dept avg_rating,
   summary totals, highest/lowest dept derived from DB)
2. Review_Policy_Memo.docx (RU+EN headings)
3. Calendar events (CRITICAL: 7 Performance Review events, Mar 9-13 2026,
   Friday 14:00 + 16:00, 1-hour duration, all 7 departments)
4. Email to HR leadership (CRITICAL: body contains overall avg + underperformer count)
5. rating_analysis.py script exists
"""
import argparse
import json
import os
import sys
from datetime import datetime

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

EMP_TABLE = 'sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"'

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        tag = "FAIL-CRITICAL" if critical else "FAIL"
        print(f"  [{tag}] {name}: {str(detail)[:200]}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=0.5):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected():
    """Compute expected per-department and org-wide stats live from the DB.

    Returns a dict, or None if the DB/table is unreachable (then numeric
    CRITICAL checks are skipped rather than firing on infra failure)."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(f'''
            SELECT "DEPARTMENT",
                   COUNT(*) AS headcount,
                   ROUND(AVG("PERFORMANCE_RATING")::numeric, 2) AS avg_rating,
                   SUM(CASE WHEN "PERFORMANCE_RATING" >= 4 THEN 1 ELSE 0 END) AS high,
                   SUM(CASE WHEN "PERFORMANCE_RATING" < 2 THEN 1 ELSE 0 END) AS under
            FROM {EMP_TABLE}
            GROUP BY "DEPARTMENT"
            ORDER BY "DEPARTMENT"
        ''')
        depts = {}
        for r in cur.fetchall():
            depts[str(r[0]).strip()] = {
                "headcount": int(r[1]), "avg": float(r[2]),
                "high": int(r[3]), "under": int(r[4]),
            }
        cur.execute(f'''
            SELECT COUNT(*),
                   ROUND(AVG("PERFORMANCE_RATING")::numeric, 2),
                   SUM(CASE WHEN "PERFORMANCE_RATING" >= 4 THEN 1 ELSE 0 END),
                   SUM(CASE WHEN "PERFORMANCE_RATING" < 2 THEN 1 ELSE 0 END)
            FROM {EMP_TABLE}
        ''')
        row = cur.fetchone()
        conn.close()
        total = int(row[0])
        overall_avg = float(row[1])
        total_high = int(row[2])
        total_under = int(row[3])
        highest = max(depts.items(), key=lambda kv: kv[1]["avg"])[0]
        lowest = min(depts.items(), key=lambda kv: kv[1]["avg"])[0]
        return {
            "depts": depts, "total": total, "overall_avg": overall_avg,
            "total_high": total_high, "total_under": total_under,
            "highest": highest, "lowest": lowest,
        }
    except Exception as e:
        print(f"  [WARN] Could not compute expected values from DB: {e}")
        return None


def _find_col(headers, *needles):
    """Return index of first header containing ALL needles (lowercased)."""
    for i, h in enumerate(headers):
        hl = str(h).lower() if h else ""
        if all(n in hl for n in needles):
            return i
    return None


def check_excel(workspace, exp):
    print("\n=== Check 1: Performance_Review_Report.xlsx ===")
    path = os.path.join(workspace, "Performance_Review_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}", critical=True)
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")
    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # ---- Department_Ratings ----
    dr_idx = next((i for i, s in enumerate(sheets_lower) if "department" in s and "rating" in s), 0)
    ws1 = wb[sheets[dr_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Department_Ratings has 7 rows", len(data1) >= 7, f"Found {len(data1)}")

    headers1 = list(rows1[0]) if rows1 else []
    avg_col = _find_col(headers1, "avg", "rating")
    dept_col = _find_col(headers1, "department")
    if dept_col is None:
        dept_col = 0
    check("Has avg_rating column", avg_col is not None, f"Headers: {headers1}")
    check("Has pct_above_4 column",
          _find_col(headers1, "pct", "4") is not None or _find_col(headers1, "above") is not None,
          f"Headers: {headers1}")

    # CRITICAL: per-department avg_rating matches DB (RU dept labels from DB)
    if exp and avg_col is not None:
        agent_avg = {}
        for r in data1:
            dname = str(r[dept_col]).strip() if dept_col < len(r) and r[dept_col] else ""
            try:
                agent_avg[dname] = float(r[avg_col])
            except (TypeError, ValueError, IndexError):
                pass
        mismatches = []
        for dname, info in exp["depts"].items():
            got = agent_avg.get(dname)
            if got is None or not num_close(got, info["avg"], 0.05):
                mismatches.append(f"{dname}: got {got}, expected {info['avg']}")
        check("Department avg_rating matches DB for all departments",
              len(mismatches) == 0,
              "; ".join(mismatches), critical=True)

    # ---- Rating_Distribution ----
    rd_idx = next((i for i, s in enumerate(sheets_lower) if "rating" in s and "dist" in s), 1)
    if rd_idx < len(sheets):
        ws2 = wb[sheets[rd_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Rating_Distribution has 5 rows", len(data2) >= 5, f"Found {len(data2)}")

    # ---- Review_Calendar ----
    rc_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s and "calendar" in s), 2)
    if rc_idx < len(sheets):
        ws3 = wb[sheets[rc_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Review_Calendar has 7 rows", len(data3) >= 7, f"Found {len(data3)}")

    # ---- Policy_Summary ----
    ps_idx = next((i for i, s in enumerate(sheets_lower) if "policy" in s or "summary" in s), 3)
    if ps_idx < len(sheets):
        ws4 = wb[sheets[ps_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c for c in r)]
        check("Policy_Summary has 5+ metrics", len(data4) >= 5, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Has total_employees metric", "total" in all_text4 and "employee" in all_text4)

        if exp:
            # Build a metric -> value map from the (metric, value) rows.
            mvals = {}
            for r in data4:
                if r and r[0] is not None:
                    key = str(r[0]).strip().lower()
                    mvals[key] = r[1] if len(r) > 1 else None

            def get_metric(*needles):
                for k, v in mvals.items():
                    if all(n in k for n in needles):
                        return v
                return None

            # CRITICAL: total_employees / high / under match DB
            te = get_metric("total_employee")
            th = get_metric("high_performer") or get_metric("high", "performer")
            tu = get_metric("underperformer") or get_metric("under")
            check("Policy_Summary total_employees correct",
                  num_close(te, exp["total"], 1),
                  f"Got {te}, expected {exp['total']}", critical=True)
            check("Policy_Summary total_high_performers correct",
                  num_close(th, exp["total_high"], 1),
                  f"Got {th}, expected {exp['total_high']}", critical=True)
            check("Policy_Summary total_underperformers correct",
                  num_close(tu, exp["total_under"], 1),
                  f"Got {tu}, expected {exp['total_under']}", critical=True)

            # CRITICAL: highest/lowest rated department match DB argmax/argmin
            hi = get_metric("highest")
            lo = get_metric("lowest")
            hi_s = str(hi).strip() if hi is not None else ""
            lo_s = str(lo).strip() if lo is not None else ""
            check("Policy_Summary highest_rated_department correct",
                  hi_s == exp["highest"],
                  f"Got '{hi_s}', expected '{exp['highest']}'", critical=True)
            check("Policy_Summary lowest_rated_department correct",
                  lo_s == exp["lowest"],
                  f"Got '{lo_s}', expected '{exp['lowest']}'", critical=True)


def check_word(workspace, exp):
    print("\n=== Check 2: Review_Policy_Memo.docx ===")
    path = os.path.join(workspace, "Review_Policy_Memo.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()

    # Required section headings, RU+EN alternatives.
    heading_groups = [
        ["overview", "обзор"],
        ["rating distribution", "distribution", "распределение оценок", "распределение"],
        ["calibration", "калибровк"],
    ]
    for grp in heading_groups:
        check(f"Mentions section ({'/'.join(grp[:2])})",
              any(g in full_text for g in grp), f"Looked for {grp}")

    # Mentions at least one (RU) department label from the DB.
    if exp:
        check("Mentions a department",
              any(d.lower() in full_text for d in exp["depts"].keys()),
              "No DB department label found in memo")
    check("Has substantial content", len(full_text) > 200, f"Length: {len(full_text)}")


def _parse_dt(v):
    if isinstance(v, datetime):
        return v
    try:
        return datetime.fromisoformat(str(v).replace("Z", "").strip()[:19])
    except Exception:
        return None


def check_gcal(exp):
    print("\n=== Check 3: Calendar Department Reviews ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
        review_events = [e for e in events
                         if "performance review" in str(e[0]).lower()
                         or ("performance" in str(e[0]).lower() and "review" in str(e[0]).lower())]
        check("Exactly 7 Performance Review events", len(review_events) == 7,
              f"Found {len(review_events)} review events out of {len(events)} total",
              critical=True)

        # Departments covered (RU labels from DB).
        if exp and review_events:
            summaries = " ".join(str(e[0]) for e in review_events).lower()
            covered = sum(1 for d in exp["depts"].keys() if d.lower() in summaries)
            check("Events cover all 7 departments", covered >= 7,
                  f"Found {covered} of {len(exp['depts'])} departments in summaries",
                  critical=True)

        # All events fall within Mar 9-13 2026.
        window_lo = datetime(2026, 3, 9)
        window_hi = datetime(2026, 3, 13, 23, 59, 59)
        starts = [_parse_dt(e[2]) for e in review_events]
        in_window = [d for d in starts if d and window_lo <= d <= window_hi]
        check("All review events within Mar 9-13 2026",
              len(review_events) == 7 and len(in_window) == 7,
              f"{len(in_window)}/{len(review_events)} in window; starts={[str(s) for s in starts]}",
              critical=True)

        # Two events on Friday (Mar 13): one ~14:00 and one ~16:00.
        fri = [d for d in starts if d and d.date() == datetime(2026, 3, 13).date()]
        fri_hours = sorted(d.hour for d in fri)
        check("Two events on Friday Mar 13 at 14:00 and 16:00",
              len(fri) == 2 and 14 in fri_hours and 16 in fri_hours,
              f"Friday starts hours={fri_hours}", critical=True)

        # All start times are at 14:00 or 16:00 (top of hour).
        valid_times = all(d and d.hour in (14, 16) and d.minute == 0 for d in starts)
        check("Review events start at 14:00 or 16:00", valid_times,
              f"start hours={[d.hour if d else None for d in starts]}")

        # 1-hour duration.
        durations_ok = True
        for e in review_events:
            s, en = _parse_dt(e[2]), _parse_dt(e[3])
            if not s or not en or abs((en - s).total_seconds() - 3600) > 60:
                durations_ok = False
                break
        check("Review events last one hour", durations_ok,
              "Some event is not 1 hour long", critical=True)
    except Exception as e:
        check("Gcal check", False, str(e), critical=True)
    finally:
        cur.close()
        conn.close()


def check_email(exp):
    print("\n=== Check 4: Email to HR Leadership ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%hr_leadership%%'
               OR subject ILIKE '%%Annual Performance Review Cycle%%'
               OR subject ILIKE '%%Department Summary%%'
        """)
        emails = cur.fetchall()
        if not emails:
            cur.execute("""
                SELECT id, subject, to_addr, body_text
                FROM email.drafts
                WHERE to_addr::text ILIKE '%%hr_leadership%%'
                   OR subject ILIKE '%%Performance Review%%'
            """)
            emails = cur.fetchall()
        check("Email about performance review sent", len(emails) >= 1,
              "No matching email found", critical=True)
        if not emails:
            return

        # Pick the best-matching email: addressed to hr_leadership AND right subject.
        target = None
        for e in emails:
            to_s = str(e[2]).lower() if e[2] else ""
            subj_s = str(e[1]).lower() if e[1] else ""
            if "hr_leadership" in to_s and "performance review" in subj_s:
                target = e
                break
        if target is None:
            target = emails[0]

        to_s = str(target[2]).lower() if target[2] else ""
        subj_s = str(target[1]).lower() if target[1] else ""
        check("Email addressed to hr_leadership@company.com", "hr_leadership" in to_s,
              f"to_addr: {target[2]}", critical=True)
        check("Email subject relevant",
              "performance" in subj_s or "review" in subj_s or "department" in subj_s,
              f"Subject: {target[1]}")

        # CRITICAL: body contains correct overall avg rating AND underperformer count.
        if exp:
            body = str(target[3] or "").lower()
            avg_str = f"{exp['overall_avg']:.2f}"
            avg_str1 = f"{round(exp['overall_avg'], 1)}"
            under_str = str(exp["total_under"])
            avg_ok = avg_str in body or avg_str1 in body
            under_ok = under_str in body
            check("Email body contains correct overall avg rating",
                  avg_ok, f"Looked for {avg_str} / {avg_str1} in body", critical=True)
            check("Email body contains correct underperformer count",
                  under_ok, f"Looked for '{under_str}' in body", critical=True)
            # RU/EN keyword presence (soft).
            check("Email body mentions a relevant keyword",
                  any(k in body for k in ["отчёт", "отчет", "средн", "недостаточн",
                                          "underperformer", "average", "rating", "оцен"]),
                  "No RU/EN summary keyword found")
    except Exception as e:
        check("Email check", False, str(e), critical=True)
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: rating_analysis.py ===")
    path = os.path.join(workspace, "rating_analysis.py")
    check("rating_analysis.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Performance_Review_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative values in Excel", not has_negative,
              "Found negative rating/count value")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE (subject ILIKE '%%performance review%%' OR subject ILIKE '%%review cycle%%')
              AND to_addr::text ILIKE '%%competitor%%'
        """)
        bad_count = cur.fetchone()[0]
        check("No performance review emails to competitor addresses", bad_count == 0,
              f"Found {bad_count}")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    exp = get_expected()

    check_excel(args.agent_workspace, exp)
    check_word(args.agent_workspace, exp)
    check_gcal(exp)
    check_email(exp)
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_fails": CRITICAL_FAILS}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: {len(CRITICAL_FAILS)} CRITICAL check(s) failed: {CRITICAL_FAILS}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
