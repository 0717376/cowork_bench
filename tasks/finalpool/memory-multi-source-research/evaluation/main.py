"""
Evaluation для memory-multi-source-research.

Что проверяем:
1. Research_Analysis.xlsx: листы Paper Summary и Research Progress с корректными данными
   - >=5 целевых статей, шумовые статьи (Deep Learning / Attention Is All You Need) НЕ попали
   - значения Citations совпадают с источником (3000/1200/800/...) — данные взяты из базы, не выдуманы
2. Research_Report.docx: разделы Introduction / Literature Review / Key Findings /
   Research Gaps / Conclusion (строгое сопоставление заголовков, EN или RU), >=3 статьи упомянуты
3. memory.json: сущность отслеживания исследования + >=5 сущностей-статей
   (каждая несёт title+year+citations), >=2 раунда поиска зафиксированы

CRITICAL_CHECKS — содержательные проверки. Провал любой => итог FAIL независимо от accuracy.
Структурные проверки (лист есть, колонка есть) — некритичные.

Usage:
    python -m evaluation.main --agent_workspace <path> --groundtruth_workspace <path>
"""
import argparse
import json
import os
import re
import sys

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []  # имена провалившихся обязательных проверок

# Содержательные проверки: провал => вся задача FAIL, независимо от accuracy.
CRITICAL_CHECKS = {
    "Excel: >=5 целевых статей по безопасности ИИ",
    "Excel: шумовые статьи (Deep Learning / Attention) НЕ засчитаны как safety-статьи",
    "Excel: для >=3 статей значение Citations совпадает с источником",
    "Word: все 5 обязательных разделов присутствуют (EN или RU заголовок)",
    "Memory: есть сущность отслеживания исследования + >=5 сущностей-статей с title/year/citations",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def normalize(text):
    return re.sub(r'\s+', ' ', text.lower().strip())


EXPECTED_PAPER_FRAGMENTS = [
    "concrete problems in ai safety",
    "ai safety via debate",
    "risks from learned optimization",
    "red teaming language models",
    "alignment of language agents",
    "scalable oversight",
]

NOISE_FRAGMENTS = [
    "deep learning",
    "attention is all you need",
]

# Эталонные значения цитирований из источника (preprocess TARGET_PAPERS).
# Ключ — фрагмент названия, значение — citation_count в базе.
EXPECTED_CITATIONS = {
    "concrete problems in ai safety": 3000,
    "ai safety via debate": 500,
    "risks from learned optimization": 400,
    "red teaming language models": 800,
    "alignment of language agents": 1200,
    "scalable oversight": 350,
}

# Заголовки разделов: каждый раздел задаётся набором допустимых форм (EN + RU).
REQUIRED_SECTIONS = {
    "Introduction": ["introduction", "введение"],
    "Literature Review": ["literature review", "обзор литературы"],
    "Key Findings": ["key findings", "ключевые выводы", "основные выводы"],
    "Research Gaps": ["research gaps", "пробелы в исследованиях", "направления исследований", "нерешённые"],
    "Conclusion": ["conclusion", "заключение"],
}


def check_excel(agent_workspace):
    """Check Research_Analysis.xlsx has correct data."""
    print("\n=== Checking Excel Output ===")

    import openpyxl

    excel_path = os.path.join(agent_workspace, "Research_Analysis.xlsx")
    if not os.path.isfile(excel_path):
        check("Research_Analysis.xlsx exists", False, f"Not found: {excel_path}")
        return

    check("Research_Analysis.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    # --- Paper Summary sheet ---
    paper_sheet = None
    for name in wb.sheetnames:
        if "paper" in name.lower() and "summary" in name.lower():
            paper_sheet = wb[name]
            break

    if paper_sheet is None:
        check("Paper Summary sheet exists", False,
              f"Sheets found: {wb.sheetnames}")
    else:
        check("Paper Summary sheet exists", True)

        rows = list(paper_sheet.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        check("Paper Summary has at least 5 data rows",
              len(data_rows) >= 5,
              f"Found {len(data_rows)} data rows")

        # Текст всех data-строк
        all_text = " ".join(
            str(cell).lower() for row in data_rows for cell in row if cell is not None
        )

        # CRITICAL: >=5 целевых статей
        found_count = sum(1 for f in EXPECTED_PAPER_FRAGMENTS if f in all_text)
        check("Excel: >=5 целевых статей по безопасности ИИ",
              found_count >= 5,
              f"Found {found_count} of {len(EXPECTED_PAPER_FRAGMENTS)} target papers")

        # CRITICAL: шумовые статьи не засчитаны. Допускается, что одно из шумовых
        # названий случайно встретится в прозе, но оба сразу — почти наверняка
        # означает, что агент не отфильтровал нерелевантные статьи.
        noise_hits = sum(1 for f in NOISE_FRAGMENTS if f in all_text)
        check("Excel: шумовые статьи (Deep Learning / Attention) НЕ засчитаны как safety-статьи",
              noise_hits < 2,
              f"Noise fragments present: {noise_hits}/2 ({[f for f in NOISE_FRAGMENTS if f in all_text]})")

        # CRITICAL: значения цитирований совпадают с источником (>=3 статьи).
        # Для каждой целевой статьи находим строку, где встречается её фрагмент,
        # и проверяем, что эталонное число цитирований присутствует в этой же строке.
        citation_matches = 0
        for frag, expected_cit in EXPECTED_CITATIONS.items():
            for row in data_rows:
                row_text = " ".join(str(c).lower() for c in row if c is not None)
                if frag in row_text:
                    # Принимаем эталонное число цитирований в обычной форме
                    # либо с разделителем тысяч (3000 / 3,000 / 3 000).
                    plain = str(expected_cit)
                    grouped = f"{expected_cit:,}"          # 3,000
                    spaced = grouped.replace(",", " ")     # 3 000
                    forms = {plain, grouped, spaced}
                    if any(re.search(rf"(?<!\d){re.escape(form)}(?!\d)", row_text)
                           for form in forms):
                        citation_matches += 1
                    break
        check("Excel: для >=3 статей значение Citations совпадает с источником",
              citation_matches >= 3,
              f"Matched citation values: {citation_matches}")

        # Заголовки колонок
        if rows:
            header = " ".join(str(h).lower() for h in rows[0] if h is not None)
            check("Paper Summary has Title column", "title" in header, f"Header: {header}")
            check("Paper Summary has Year column", "year" in header, f"Header: {header}")
            check("Paper Summary has Citations column", "citation" in header, f"Header: {header}")

    # --- Research Progress sheet ---
    progress_sheet = None
    for name in wb.sheetnames:
        if "progress" in name.lower():
            progress_sheet = wb[name]
            break

    if progress_sheet is None:
        check("Research Progress sheet exists", False,
              f"Sheets found: {wb.sheetnames}")
    else:
        check("Research Progress sheet exists", True)

        rows = list(progress_sheet.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        check("Research Progress has at least 2 data rows",
              len(data_rows) >= 2,
              f"Found {len(data_rows)} data rows")

        if rows:
            header = " ".join(str(h).lower() for h in rows[0] if h is not None)
            check("Research Progress has Phase column", "phase" in header, f"Header: {header}")
            check("Research Progress has Status column", "status" in header, f"Header: {header}")

        # >=2 фазы, отражающие раунды поиска плюс этап анализа/отчёта.
        progress_text = " ".join(
            str(cell).lower() for row in data_rows for cell in row if cell is not None
        )
        has_search_phase = ("search" in progress_text or "поиск" in progress_text
                            or "round" in progress_text or "раунд" in progress_text)
        has_report_phase = ("report" in progress_text or "отчёт" in progress_text
                            or "отчет" in progress_text or "analysis" in progress_text
                            or "анализ" in progress_text)
        check("Research Progress содержит фазы поиска и анализа/отчёта",
              len(data_rows) >= 2 and has_search_phase and has_report_phase,
              f"search_phase={has_search_phase}, report_phase={has_report_phase}")


def check_word(agent_workspace):
    """Check Research_Report.docx has required content."""
    print("\n=== Checking Word Report ===")

    from docx import Document

    docx_path = os.path.join(agent_workspace, "Research_Report.docx")
    if not os.path.isfile(docx_path):
        check("Research_Report.docx exists", False, f"Not found: {docx_path}")
        return

    check("Research_Report.docx exists", True)

    try:
        doc = Document(docx_path)
        paragraphs = [para.text for para in doc.paragraphs]
        full_text = "\n".join(paragraphs)
    except Exception as e:
        check("Word document readable", False, str(e))
        return

    normalized = normalize(full_text)

    check("Report has at least 500 characters",
          len(full_text.strip()) >= 500,
          f"Document has {len(full_text.strip())} characters")

    # CRITICAL: строгое сопоставление заголовков разделов (EN или RU).
    # Заголовок засчитывается, если какая-либо его форма встречается как
    # отдельная строка-параграф (а не просто слово где-то в прозе).
    para_norms = [normalize(p) for p in paragraphs if p.strip()]

    def section_present(forms):
        for form in forms:
            for pn in para_norms:
                # заголовок: строка, начинающаяся с формы (допускаем нумерацию/двоеточие),
                # и достаточно короткая, чтобы это был заголовок, а не абзац прозы.
                stripped = pn.lstrip("0123456789.) -")
                if (stripped == form or stripped.startswith(form)) and len(pn) <= len(form) + 30:
                    return True
        return False

    missing = []
    for canonical, forms in REQUIRED_SECTIONS.items():
        if not section_present(forms):
            missing.append(canonical)

    check("Word: все 5 обязательных разделов присутствуют (EN или RU заголовок)",
          len(missing) == 0,
          f"Missing sections: {missing}")

    # CRITICAL входит выше; ниже — отдельные некритичные индикаторы для прозрачности
    for canonical, forms in REQUIRED_SECTIONS.items():
        check(f"Report has {canonical} heading",
              section_present(forms),
              f"Section '{canonical}' heading not found")

    # Упоминание статей
    paper_mention_count = sum(1 for f in EXPECTED_PAPER_FRAGMENTS if f in normalized)
    check("Report mentions at least 3 papers",
          paper_mention_count >= 3,
          f"Found {paper_mention_count} paper mentions")


def check_memory(agent_workspace):
    """Check that memory.json has research tracking entities."""
    print("\n=== Checking Memory ===")

    memory_path = os.path.join(agent_workspace, "memory", "memory.json")
    if not os.path.isfile(memory_path):
        check("memory.json exists", False, f"Not found: {memory_path}")
        return

    check("memory.json exists", True)

    with open(memory_path, "r") as f:
        content = f.read().strip()

    if not content or content in ("{}", '{"entities": [], "relations": []}'):
        check("Memory has content", False, "memory.json is empty or unchanged")
        return

    check("Memory has content", True)

    try:
        memory_data = json.loads(content)
    except json.JSONDecodeError:
        check("Memory is valid JSON", False, "Cannot parse memory.json")
        return

    check("Memory is valid JSON", True)

    entities = memory_data.get("entities", []) if isinstance(memory_data, dict) else []
    if isinstance(memory_data, list):
        entities = memory_data

    check("Memory has at least 3 entities",
          len(entities) >= 3,
          f"Found {len(entities)} entities")

    def ent_text(ent):
        return json.dumps(ent, ensure_ascii=False).lower() if isinstance(ent, dict) else ""

    # Сущности-статьи: каждая несёт название статьи + год + цитирования.
    paper_entities = 0
    for ent in entities:
        t = ent_text(ent)
        has_title = any(f in t for f in EXPECTED_PAPER_FRAGMENTS)
        has_year = bool(re.search(r"\b(20[0-2]\d)\b", t))
        has_cit = ("citation" in t or "цитир" in t)
        if has_title and has_year and has_cit:
            paper_entities += 1

    # Сущность отслеживания исследования: фиксирует выполненные поиски/раунды И число статей.
    tracking_entity = None
    for ent in entities:
        t = ent_text(ent)
        mentions_search = ("search" in t or "round" in t or "поиск" in t or "раунд" in t)
        mentions_count = ("papers found" in t or "total papers" in t
                          or re.search(r"(найден|статей|papers)\D*\d", t) is not None)
        if mentions_search and mentions_count:
            tracking_entity = ent
            break

    # CRITICAL: tracking entity + >=5 paper entities с полными полями
    check("Memory: есть сущность отслеживания исследования + >=5 сущностей-статей с title/year/citations",
          tracking_entity is not None and paper_entities >= 5,
          f"tracking_entity={'yes' if tracking_entity else 'no'}, paper_entities={paper_entities}")

    # Некритично: >=2 раунда поиска зафиксированы в tracking-сущности
    rounds_recorded = 0
    if tracking_entity is not None:
        tt = ent_text(tracking_entity)
        rounds_recorded = len(re.findall(r"(search round|раунд поиска|round \d)", tt))
    check("Memory: tracking-сущность фиксирует >=2 раунда поиска",
          rounds_recorded >= 2,
          f"Rounds recorded in tracking entity: {rounds_recorded}")

    # Некритично: ключевые слова статей в памяти
    entity_text = " ".join(ent_text(e) for e in entities)
    paper_keywords = ["concrete problems", "debate", "mesa-optimization",
                      "learned optimization", "red teaming", "alignment",
                      "reward model"]
    kw_count = sum(1 for kw in paper_keywords if kw in entity_text)
    check("Memory has paper-related keywords (at least 2)",
          kw_count >= 2,
          f"Found {kw_count} paper keywords in memory")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_memory(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== SUMMARY ===")
    print(f"Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%), {FAIL_COUNT} failed")
    if CRITICAL_FAILS:
        print(f"Critical fails: {CRITICAL_FAILS}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_fails": CRITICAL_FAILS,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"FAIL: критичные чеки провалены ({len(CRITICAL_FAILS)})")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
