"""Evaluation for terminal-sf-insales-support-gsheet-word-email.

Checks:
1. Support_Quality_Audit.xlsx with 3 sheets
2. Support_Audit_Report.docx
3. Google Sheet "Support Quality Audit"
4. Email sent to cs_leadership@company.com
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_CRITICAL = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            FAILED_CRITICAL.append(name)
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_dynamic_ticket_data():
    """Dynamically query ticket priority data from the DB."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT "PRIORITY", COUNT(*) as cnt,
                   ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp,
                   ROUND(AVG("SATISFACTION_SCORE")::numeric, 2) as avg_sat
            FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
            GROUP BY "PRIORITY" ORDER BY "PRIORITY"
        """)
        rows = cur.fetchall()
        cur.execute('SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
        total = cur.fetchone()[0]
        cur.execute("""
            SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
            WHERE "PRIORITY" = 'High'
        """)
        high_count = cur.fetchone()[0]
        cur.close()
        conn.close()
        return {r[0].strip().lower(): {"count": r[1], "avg_resp": float(r[2]), "avg_sat": float(r[3])} for r in rows}, total, high_count
    except Exception:
        return None, None, None


def get_dynamic_order_data():
    """Dynamically query WC order data from the DB."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM wc.orders")
        total_orders = cur.fetchone()[0]
        cur.execute("""
            SELECT c.name as category, COUNT(*) as total,
                   SUM(CASE WHEN o.status IN ('refunded','cancelled','failed') THEN 1 ELSE 0 END) as problems
            FROM wc.orders o
            JOIN wc.order_items oi ON o.id = oi.order_id
            JOIN wc.products p ON oi.product_id = p.id
            JOIN wc.categories c ON p.category_id = c.id
            GROUP BY c.name ORDER BY c.name
        """)
        cat_rows = cur.fetchall()
        cur.close()
        conn.close()
        return total_orders, {r[0].strip().lower(): {"total": r[1], "problems": r[2]} for r in cat_rows}
    except Exception:
        return None, None


def get_highest_problem_category():
    """Live winner: category name (original case) with the max problem rate."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT c.name,
                   SUM(CASE WHEN o.status IN ('refunded','cancelled','failed') THEN 1 ELSE 0 END)::float
                   / NULLIF(COUNT(*), 0) AS rate
            FROM wc.orders o
            JOIN wc.order_items oi ON o.id = oi.order_id
            JOIN wc.products p ON oi.product_id = p.id
            JOIN wc.categories c ON p.category_id = c.id
            GROUP BY c.name
            ORDER BY rate DESC, c.name
            LIMIT 1
        """)
        row = cur.fetchone()
        cur.close()
        conn.close()
        return row[0].strip() if row else None
    except Exception:
        return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Support_Quality_Audit.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Support_Quality_Audit.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Support_Quality_Audit.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Get dynamic expected data
    ticket_data, total_tickets, high_count = get_dynamic_ticket_data()
    total_orders, cat_data = get_dynamic_order_data()

    # Fallback: use groundtruth file if DB queries fail
    gt_wb = None
    try:
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception:
        pass

    # Sheet 1: Ticket_Priority_Summary
    print("  Checking Ticket_Priority_Summary...")
    a_sheet = get_sheet(agent_wb, "Ticket_Priority_Summary")
    check("Sheet 'Ticket_Priority_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        check("Ticket_Priority_Summary has 3 rows", len(a_rows) == 3, f"Got {len(a_rows)}")

        a_lookup = {str(r[0]).strip().lower(): r for r in a_rows if r and r[0]}

        if ticket_data:
            # Use dynamic DB data
            for priority, vals in ticket_data.items():
                # The 'High' priority row carries the core support-quality signal.
                is_high = priority == "high"
                a_row = a_lookup.get(priority)
                if a_row is None:
                    check(f"Priority '{priority}' present", False, "Missing", critical=is_high)
                    continue
                if len(a_row) > 1:
                    check(f"'{priority}' Ticket_Count",
                          num_close(a_row[1], vals["count"], 50),
                          f"Expected {vals['count']}, got {a_row[1]}", critical=is_high)
                if len(a_row) > 2:
                    check(f"'{priority}' Avg_Response_Hours",
                          num_close(a_row[2], vals["avg_resp"], 1.0),
                          f"Expected {vals['avg_resp']}, got {a_row[2]}", critical=is_high)
                if len(a_row) > 3:
                    check(f"'{priority}' Avg_Satisfaction",
                          num_close(a_row[3], vals["avg_sat"], 0.1),
                          f"Expected {vals['avg_sat']}, got {a_row[3]}", critical=is_high)
        elif gt_wb:
            # Fallback to groundtruth
            g_sheet = get_sheet(gt_wb, "Ticket_Priority_Summary")
            if g_sheet:
                g_rows = list(g_sheet.iter_rows(min_row=2, values_only=True))
                for g_row in g_rows:
                    if not g_row or not g_row[0]:
                        continue
                    key = str(g_row[0]).strip().lower()
                    a_row = a_lookup.get(key)
                    if a_row is None:
                        check(f"Priority '{g_row[0]}' present", False, "Missing")
                        continue
                    if len(a_row) > 1 and len(g_row) > 1:
                        check(f"'{key}' Ticket_Count",
                              num_close(a_row[1], g_row[1], 50),
                              f"Expected {g_row[1]}, got {a_row[1]}")

    # Sheet 2: Product_Problem_Rates
    print("  Checking Product_Problem_Rates...")
    a_sheet = get_sheet(agent_wb, "Product_Problem_Rates")
    check("Sheet 'Product_Problem_Rates' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        a_lookup = {}
        for r in a_rows:
            if r and r[0]:
                a_lookup[str(r[0]).strip().lower()] = r

        if cat_data:
            expected_cats = len(cat_data)
            check(f"Product_Problem_Rates has {expected_cats} rows",
                  len(a_rows) == expected_cats, f"Got {len(a_rows)}")
            for cat_name, vals in cat_data.items():
                a_row = a_lookup.get(cat_name)
                if a_row is None:
                    for ak, av in a_lookup.items():
                        if cat_name.split()[0] in ak:
                            a_row = av
                            break
                if a_row is None:
                    check(f"Category '{cat_name}' present", False,
                          f"Missing from {list(a_lookup.keys())}")
                    continue
                if len(a_row) > 1:
                    check(f"'{cat_name}' Total_Orders",
                          num_close(a_row[1], vals["total"], 10),
                          f"Expected {vals['total']}, got {a_row[1]}")
        else:
            check("Product_Problem_Rates has 8 rows", len(a_rows) == 8, f"Got {len(a_rows)}")

    # Sheet 3: Audit_Summary
    print("  Checking Audit_Summary...")
    a_sheet = get_sheet(agent_wb, "Audit_Summary")
    check("Sheet 'Audit_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_data = {}
        for row in a_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_data[str(row[0]).strip().lower()] = row[1]

        if total_tickets is not None and total_orders is not None:
            check("Total_Tickets",
                  num_close(a_data.get("total_tickets"), total_tickets, 100),
                  f"Expected {total_tickets}, got {a_data.get('total_tickets')}")
            check("Total_WC_Orders",
                  num_close(a_data.get("total_wc_orders"), total_orders, 5),
                  f"Expected {total_orders}, got {a_data.get('total_wc_orders')}")
            if total_orders > 0 and cat_data:
                total_problems = sum(v["problems"] for v in cat_data.values())
                expected_rate = round(total_problems / sum(v["total"] for v in cat_data.values()) * 100, 1)
                check("Overall_Problem_Rate_Pct",
                      num_close(a_data.get("overall_problem_rate_pct"), expected_rate, 3.0),
                      f"Expected {expected_rate}, got {a_data.get('overall_problem_rate_pct')}",
                      critical=True)
            if high_count is not None and total_tickets > 0:
                expected_hp = round(high_count / total_tickets * 100, 1)
                check("High_Priority_Pct",
                      num_close(a_data.get("high_priority_pct"), expected_hp, 2.0),
                      f"Expected {expected_hp}, got {a_data.get('high_priority_pct')}",
                      critical=True)

            # Highest_Problem_Category must equal the live DB winner (russified,
            # e.g. 'Камеры'). Asserted explicitly, case-insensitive substring-safe.
            winner = get_highest_problem_category()
            if winner:
                got = a_data.get("highest_problem_category")
                got_s = str(got).strip().lower() if got is not None else ""
                win_s = winner.strip().lower()
                match = bool(got_s) and (win_s in got_s or got_s in win_s)
                check("Highest_Problem_Category == live DB winner",
                      match, f"Expected '{winner}', got '{got}'", critical=True)
        elif gt_wb:
            g_sheet = get_sheet(gt_wb, "Audit_Summary")
            if g_sheet:
                g_data = {}
                for row in g_sheet.iter_rows(min_row=2, values_only=True):
                    if row and row[0]:
                        g_data[str(row[0]).strip().lower()] = row[1]
                check("Total_Tickets",
                      num_close(a_data.get("total_tickets"), g_data.get("total_tickets"), 100),
                      f"Expected {g_data.get('total_tickets')}, got {a_data.get('total_tickets')}")
                check("Total_WC_Orders",
                      num_close(a_data.get("total_wc_orders"), g_data.get("total_wc_orders"), 5),
                      f"Expected {g_data.get('total_wc_orders')}, got {a_data.get('total_wc_orders')}")


def check_word(agent_workspace):
    print("\n=== Checking Support_Audit_Report.docx ===")
    docx_path = os.path.join(agent_workspace, "Support_Audit_Report.docx")
    check("Support_Audit_Report.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 200, f"Length: {len(text)}")
        check("Contains support/ticket reference",
              ("поддерж" in text or "support" in text)
              and ("тикет" in text or "заяв" in text or "ticket" in text))
        check("Contains problem/complaint reference",
              "проблем" in text or "жалоб" in text or "возврат" in text
              or "problem" in text or "complaint" in text or "refund" in text)
        check("Contains recommendation",
              "рекоменд" in text or "предлож" in text or "меропр" in text
              or "recommend" in text or "suggestion" in text or "action" in text,
              critical=True)
        check("Contains highest-problem category reference (RU/EN)",
              "камер" in text or "camera" in text,
              "Missing highest-problem category discussion", critical=True)
    except ImportError:
        check("python-docx available", False, "Cannot verify Word content")
    except Exception as e:
        check("Word document readable", False, str(e))


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title FROM gsheet.spreadsheets
            WHERE lower(title) LIKE '%%support%%audit%%'
        """)
        rows = cur.fetchall()
        check("Google Sheet 'Support Quality Audit' exists", len(rows) > 0,
              f"Found {len(rows)} matching sheets")

        if rows:
            ss_id = rows[0][0]
            cur.execute("""
                SELECT title FROM gsheet.sheets WHERE spreadsheet_id = %s
            """, (ss_id,))
            sheet_names = [r[0].lower() for r in cur.fetchall()]
            check("GSheet has Ticket_Priority_Summary",
                  any("ticket" in s and "priority" in s for s in sheet_names),
                  f"Sheets: {sheet_names}")
            check("GSheet has Product_Problem_Rates",
                  any("product" in s and "problem" in s for s in sheet_names),
                  f"Sheets: {sheet_names}")
            check("GSheet has Audit_Summary",
                  any("audit" in s and "summary" in s for s in sheet_names),
                  f"Sheets: {sheet_names}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Google Sheet check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Check sent_log joined with messages
        cur.execute("""
            SELECT m.subject, m.to_addr, m.body_text
            FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            WHERE lower(m.subject) LIKE '%%support%%quality%%audit%%'
               OR lower(m.subject) LIKE '%%q1%%support%%'
        """)
        rows = cur.fetchall()
        if not rows:
            # Also check messages directly for sent emails
            cur.execute("""
                SELECT subject, to_addr, body_text FROM email.messages
                WHERE lower(subject) LIKE '%%support%%quality%%audit%%'
                   OR lower(subject) LIKE '%%q1%%support%%'
            """)
            rows = cur.fetchall()
        check("Audit email sent", len(rows) > 0, f"Found {len(rows)} matching emails")
        if rows:
            subj, to_addr, body = rows[0]
            to_str = str(to_addr).lower() if to_addr else ""
            check("Email to cs_leadership",
                  "cs_leadership" in to_str,
                  f"To: {to_addr}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    agent_file = os.path.join(workspace, "Support_Quality_Audit.xlsx")
    if os.path.exists(agent_file):
        wb = openpyxl.load_workbook(agent_file, data_only=True)
        # No unexpected sheets
        expected_keywords = {"ticket", "priority", "product", "problem", "audit", "summary"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        # Problem_Rate_Pct should not exceed 100
        ppr = get_sheet(wb, "Product_Problem_Rates")
        if ppr:
            for row in ppr.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 3 and row[3] is not None:
                    try:
                        rate = float(row[3])
                        if rate > 100:
                            check("No problem rate > 100%", False, f"Found {rate} for {row[0]}")
                            break
                    except (ValueError, TypeError):
                        pass
            else:
                check("No problem rate > 100%", True)

    # Email: no emails to wrong recipients
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE lower(subject) LIKE '%%support%%quality%%audit%%'
              AND to_addr::text NOT ILIKE '%%cs_leadership%%'
        """)
        wrong_emails = cur.fetchone()[0]
        check("No audit emails to wrong recipients", wrong_emails == 0,
              f"Found {wrong_emails} misrouted emails")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gsheet()
    check_email()
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    critical_ok = len(FAILED_CRITICAL) == 0
    if not critical_ok:
        print(f"  CRITICAL FAILURES: {FAILED_CRITICAL}")

    # PASS requires: no critical failure AND accuracy >= 70.
    overall = critical_ok and accuracy >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "passed": PASS_COUNT, "failed": FAIL_COUNT,
                "accuracy": accuracy, "critical_failures": FAILED_CRITICAL,
                "success": overall,
            }, f, indent=2)

    # Critical failures hard-fail before the accuracy gate.
    if not critical_ok:
        sys.exit(1)
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
