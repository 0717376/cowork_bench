"""Evaluation for sf-hr-education-teamly (clickhouse + teamly swap).

Critical checks (CRITICAL_CHECKS): любой провал => overall FAIL независимо от
accuracy. Иначе PASS требует отсутствия критических провалов И accuracy >= 70%.

Данные в ClickHouse русифицированы централизованно (db/zzz_clickhouse_after_init.sql):
DEPARTMENT (Инженерия/Финансы/Кадры/Операции/НИОКР/Продажи/Поддержка) и
EDUCATION_LEVEL (Бакалавр/Магистр/Кандидат наук/...). Числовые агрегаты не меняются.
Groundtruth Education_Analysis.xlsx содержит русифицированные ключи. Строковые
проверки принимают RU+EN. Страница в Teamly — блокирующая критическая проверка.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# RU<->EN алиасы названий подразделений (для приёма обеих форм).
DEPT_ALIASES = {
    "инженерия": "eng", "engineering": "eng",
    "финансы": "fin", "finance": "fin",
    "кадры": "hr", "hr": "hr",
    "операции": "ops", "operations": "ops",
    "ниокр": "rnd", "r&d": "rnd",
    "продажи": "sal", "sales": "sal",
    "поддержка": "sup", "support": "sup",
}

# Самый распространённый уровень образования (бакалавр), RU+EN.
BACHELOR_FORMS = {"бакалавр", "bachelor's", "bachelors", "bachelor"}

CRITICAL_CHECKS = {
    "By Department: PhD_Count и PhD_Pct совпадают с эталоном по всем подразделениям",
    "Summary Total_Employees == 50000",
    "Summary Most_Common_Education = бакалавр (RU+EN)",
    "Summary Department_Highest_PhD_Pct = подразделение с макс. долей PhD (RU+EN, с учётом ничьей)",
    "Teamly 'Workforce Education Analysis' page exists with non-empty content",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def dept_key(label):
    return DEPT_ALIASES.get(str(label or "").strip().lower())


def get_sheet(wb, target):
    for name in wb.sheetnames:
        if name.strip().lower() == target.strip().lower():
            return wb[name]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Education_Analysis.xlsx against groundtruth (русифицированные ключи)."""
    print("\n=== Checking Education_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Education_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Education_Analysis.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return

    # --- Sheet: By Department ---
    agent_ws = get_sheet(agent_wb, "By Department")
    gt_ws = get_sheet(gt_wb, "By Department")

    # Соберём данные для критической проверки PhD_Count / PhD_Pct.
    phd_critical_ok = True
    phd_detail = []

    if agent_ws is None:
        record("Sheet 'By Department' exists", False, f"Sheets: {agent_wb.sheetnames}")
        phd_critical_ok = False
        phd_detail.append("missing sheet")
    else:
        record("Sheet 'By Department' exists", True)
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("By Department row count", len(agent_rows) == len(gt_rows),
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        # Ключуем по канонизированному (RU/EN-устойчивому) названию подразделения.
        agent_lookup = {}
        for r in agent_rows:
            if r and r[0] is not None:
                k = dept_key(r[0]) or str(r[0]).strip().lower()
                agent_lookup[k] = r

        for gt_row in gt_rows:
            if not gt_row or gt_row[0] is None:
                continue
            key = dept_key(gt_row[0]) or str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Department '{gt_row[0]}' present", False, "Missing")
                phd_critical_ok = False
                phd_detail.append(f"{gt_row[0]}: missing")
                continue

            ok_total = num_close(a_row[1], gt_row[1], 10)
            record(f"'{gt_row[0]}' Total_Employees", ok_total,
                   f"Expected {gt_row[1]}, got {a_row[1]}")

            ok_bach = num_close(a_row[2], gt_row[2], 10)
            record(f"'{gt_row[0]}' Bachelors_Count", ok_bach,
                   f"Expected {gt_row[2]}, got {a_row[2]}")

            ok_mast = num_close(a_row[3], gt_row[3], 10)
            record(f"'{gt_row[0]}' Masters_Count", ok_mast,
                   f"Expected {gt_row[3]}, got {a_row[3]}")

            ok_phd = num_close(a_row[4], gt_row[4], 10)
            record(f"'{gt_row[0]}' PhD_Count", ok_phd,
                   f"Expected {gt_row[4]}, got {a_row[4]}")

            ok_bpct = num_close(a_row[5], gt_row[5], 1.0)
            record(f"'{gt_row[0]}' Bachelors_Pct", ok_bpct,
                   f"Expected {gt_row[5]}, got {a_row[5]}")

            ok_ppct = num_close(a_row[7], gt_row[7], 0.5)
            record(f"'{gt_row[0]}' PhD_Pct", ok_ppct,
                   f"Expected {gt_row[7]}, got {a_row[7]}")

            if not (ok_phd and ok_ppct):
                phd_critical_ok = False
                phd_detail.append(
                    f"{gt_row[0]}: PhD_Count {a_row[4]}!={gt_row[4]} or "
                    f"PhD_Pct {a_row[7]}!={gt_row[7]}")

    # CRITICAL: PhD_Count + PhD_Pct по всем подразделениям.
    record("By Department: PhD_Count и PhD_Pct совпадают с эталоном по всем подразделениям",
           phd_critical_ok, "; ".join(phd_detail))

    # --- Sheet: Summary ---
    agent_ws2 = get_sheet(agent_wb, "Summary")
    gt_ws2 = get_sheet(gt_wb, "Summary")

    agent_summary = {}
    if agent_ws2 is not None:
        record("Sheet 'Summary' exists", True)
        for row in agent_ws2.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None:
                agent_summary[str(row[0]).strip().lower()] = row[1]
    else:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")

    gt_summary = {}
    for row in gt_ws2.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            gt_summary[str(row[0]).strip().lower()] = row[1]

    # Структурные (числовые/строковые) проверки Summary, non-critical кроме явных critical ниже.
    for metric, expected in gt_summary.items():
        actual = agent_summary.get(metric)
        if metric in ("total_employees", "most_common_education",
                      "department_highest_phd_pct"):
            continue  # обрабатываются отдельными критическими проверками
        if actual is None:
            record(f"Summary '{metric}' present", False, "Missing")
        else:
            if isinstance(expected, (int, float)):
                ok = num_close(actual, expected, max(abs(expected) * 0.02, 50))
            else:
                ok = str(actual).strip().lower() == str(expected).strip().lower()
            record(f"Summary '{metric}'", ok, f"Expected {expected}, got {actual}")

    # CRITICAL: Total_Employees == 50000
    total_emp = agent_summary.get("total_employees")
    record("Summary Total_Employees == 50000",
           num_close(total_emp, 50000, 1),
           f"got {total_emp}")

    # CRITICAL: Most_Common_Education = бакалавр (RU+EN)
    mce = agent_summary.get("most_common_education")
    mce_ok = mce is not None and str(mce).strip().lower() in BACHELOR_FORMS
    record("Summary Most_Common_Education = бакалавр (RU+EN)",
           mce_ok, f"got {mce}")

    # CRITICAL: Department_Highest_PhD_Pct = подразделение с макс. долей PhD.
    # Кадры (2.1) и НИОКР (2.1) — ничья; принимаем любой из них (RU+EN).
    hpd = agent_summary.get("department_highest_phd_pct")
    hpd_key = dept_key(hpd)
    record("Summary Department_Highest_PhD_Pct = подразделение с макс. долей PhD (RU+EN, с учётом ничьей)",
           hpd_key in ("hr", "rnd"),
           f"got {hpd} (ожидается Кадры/HR или НИОКР/R&D)")


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def check_teamly():
    """CRITICAL: страница 'Workforce Education Analysis' (RU+EN) в teamly.pages
    с непустым содержимым, упоминающим хотя бы одно подразделение и PhD/кандидат наук.
    Независимо от Excel."""
    print("\n=== Checking Teamly Page ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, COALESCE(title,''), COALESCE(body,'') FROM teamly.pages")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Teamly 'Workforce Education Analysis' page exists with non-empty content",
               False, str(e))
        return

    page = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if ("workforce" in tl and "education" in tl) or \
           ("образовани" in tl and ("персонал" in tl or "сотрудник" in tl or "кадр" in tl)) or \
           ("анализ" in tl and "образовани" in tl):
            page = (pid, title, body)
            break

    if page is None:
        record("Teamly 'Workforce Education Analysis' page exists with non-empty content",
               False, f"pages: {[(p[0], p[1]) for p in pages]}")
        return

    text = ((page[1] or "") + " " + (page[2] or "")).lower()
    non_empty = len((page[2] or "").strip()) >= 20
    dept_ref = any(d in text for d in [
        "инженерия", "финансы", "кадры", "операции", "ниокр", "продажи", "поддержка",
        "engineering", "finance", "hr", "operations", "r&d", "sales", "support",
        "подразделен", "department", "отдел",
    ])
    phd_ref = "phd" in text or "кандидат наук" in text or "кандидат" in text
    record("Teamly 'Workforce Education Analysis' page exists with non-empty content",
           non_empty and dept_ref and phd_ref,
           f"body_len={len((page[2] or '').strip())} dept={dept_ref} phd={phd_ref}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
