"""
Evaluation for sales-target-reconciliation task.
Compares agent's Q1_Sales_Review.xlsx against groundtruth.
"""
import argparse
import os
import sys

import openpyxl


def load_sheet_data(wb, sheet_name):
    """Load all rows from a sheet as list of lists."""
    # Case-insensitive sheet name lookup
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    sheet_name = matched
    ws = wb[sheet_name]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_regional_performance(agent_rows, gt_rows):
    """Check Sheet 1: Regional Performance."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Regional Performance' not found"]

    # Skip header row
    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    for i, (a_row, g_row) in enumerate(zip(agent_data, gt_data)):
        region_name = g_row[0]
        # Check region name (case-insensitive)
        if str(a_row[0]).strip().lower() != str(g_row[0]).strip().lower():
            errors.append(f"Row {i+1}: Region mismatch: '{a_row[0]}' vs '{g_row[0]}'")
            continue

        # Check numeric fields with tolerance
        field_names = ["Target", "Actual", "Variance", "Variance_Pct", "Industry_Benchmark"]
        for j, fname in enumerate(field_names, start=1):
            try:
                a_val = float(a_row[j]) if a_row[j] is not None else None
                g_val = float(g_row[j])
                if a_val is None:
                    errors.append(f"{region_name}.{fname}: missing value")
                elif abs(a_val - g_val) > 1.0:  # tolerance of $1 or 1%
                    errors.append(f"{region_name}.{fname}: {a_val} vs expected {g_val}")
            except (ValueError, TypeError):
                errors.append(f"{region_name}.{fname}: invalid value '{a_row[j]}'")

        # Check vs_Benchmark (case-insensitive)
        a_vs = str(a_row[6]).strip().lower() if a_row[6] else ""
        g_vs = str(g_row[6]).strip().lower()
        if a_vs != g_vs:
            errors.append(f"{region_name}.vs_Benchmark: '{a_row[6]}' vs expected '{g_row[6]}'")

    return len(errors) == 0, errors


def check_top_products(agent_rows, gt_rows):
    """Check Sheet 2: Top Products."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Top Products' not found"]

    agent_data = agent_rows[1:] if len(agent_rows) > 1 else []
    gt_data = gt_rows[1:] if len(gt_rows) > 1 else []

    if len(agent_data) != len(gt_data):
        errors.append(f"Row count mismatch: agent={len(agent_data)}, expected={len(gt_data)}")
        return False, errors

    # Group by region and check top 3 revenue values match
    from collections import defaultdict
    agent_by_region = defaultdict(list)
    gt_by_region = defaultdict(list)

    for row in agent_data:
        if row[0]:
            agent_by_region[str(row[0]).strip()].append(row)
    for row in gt_data:
        if row[0]:
            gt_by_region[str(row[0]).strip()].append(row)

    for region in gt_by_region:
        # Case-insensitive region lookup
        agent_region_key = None
        for k in agent_by_region:
            if k.lower() == region.lower():
                agent_region_key = k
                break
        if agent_region_key is None:
            errors.append(f"Missing region in Top Products: {region}")
            continue
        a_products = agent_by_region[agent_region_key]
        g_products = gt_by_region[region]
        if len(a_products) != len(g_products):
            errors.append(f"{region}: product count {len(a_products)} vs expected {len(g_products)}")
            continue
        for k, (a_p, g_p) in enumerate(zip(a_products, g_products)):
            try:
                a_rev = float(a_p[3]) if a_p[3] is not None else 0
                g_rev = float(g_p[3])
                if abs(a_rev - g_rev) > 1.0:
                    errors.append(f"{region} product #{k+1}: revenue {a_rev} vs expected {g_rev}")
            except (ValueError, TypeError):
                errors.append(f"{region} product #{k+1}: invalid revenue '{a_p[3]}'")

    return len(errors) == 0, errors


def check_summary(agent_rows, gt_rows):
    """Check Sheet 3: Summary."""
    errors = []
    if agent_rows is None:
        return False, ["Sheet 'Summary' not found"]

    if len(agent_rows) < 7:
        errors.append(f"Summary has only {len(agent_rows)} rows, expected 7")
        return False, errors

    gt_dict = {str(r[0]).strip(): r[1] for r in gt_rows if r[0]}
    agent_dict = {str(r[0]).strip().lower(): r[1] for r in agent_rows if r[0]}

    for key, expected in gt_dict.items():
        if key.lower() not in agent_dict:
            errors.append(f"Missing key: {key}")
            continue
        try:
            a_val = float(agent_dict[key.lower()])
            g_val = float(expected)
            if abs(a_val - g_val) > 1.0:
                errors.append(f"{key}: {a_val} vs expected {g_val}")
        except (ValueError, TypeError):
            errors.append(f"{key}: invalid value '{agent_dict[key]}'")

    return len(errors) == 0, errors


# ---------------------------------------------------------------------------
# CRITICAL CHECKS: load-bearing semantic values of the reconciliation.
# Any failure here => sys.exit(1) BEFORE the structural accuracy gate.
# Region names are the Russian strings from the central ClickHouse map
# (db/zzz_clickhouse_after_init.sql); benchmarks/targets/aggregates stay numeric.
# ---------------------------------------------------------------------------

# Russian region -> (Target from PDF, Actual revenue from DW, Industry_Benchmark from dashboard)
CRITICAL_REGIONS = {
    "азиатско-тихоокеанский регион": (70000, 68298.9, 65000),
    "европа": (85000, 87065.22, 82000),
    "латинская америка": (60000, 65042.33, 62000),
    "ближний восток": (75000, 72524.95, 70000),
    "северная америка": (78000, 80916.1, 76000),
}

CRITICAL_SUMMARY = {
    "total_target": 368000,
    "total_actual": 373847.5,
    "regions_met_target": 3,
    "regions_missed_target": 2,
    "regions_above_benchmark": 5,
}


def _num(v):
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def run_critical_checks(agent_wb):
    """Semantic checks on the core reconciliation. Returns list of failures."""
    failures = []

    # --- Regional Performance: Target/Actual + internal consistency + vs_Benchmark ---
    rp = load_sheet_data(agent_wb, "Regional Performance")
    if not rp or len(rp) < 6:
        failures.append("CRITICAL: Regional Performance sheet missing or incomplete")
    else:
        seen = {}
        for row in rp[1:]:
            key = str(row[0]).strip().lower() if row[0] else ""
            if key in CRITICAL_REGIONS:
                seen[key] = row
        for key, (tgt, act, bench) in CRITICAL_REGIONS.items():
            if key not in seen:
                failures.append(f"CRITICAL: region '{key}' missing from Regional Performance")
                continue
            row = seen[key]
            a_tgt, a_act = _num(row[1]), _num(row[2])
            a_var, a_pct = _num(row[3]), _num(row[4])
            a_bench = _num(row[5])
            a_vs = str(row[6]).strip().lower() if row[6] else ""
            if a_tgt is None or abs(a_tgt - tgt) > 1.0:
                failures.append(f"CRITICAL: {key} Target {row[1]} != {tgt} (from PDF)")
            if a_act is None or abs(a_act - act) > 1.0:
                failures.append(f"CRITICAL: {key} Actual {row[2]} != {act} (from DW)")
            if a_bench is None or abs(a_bench - bench) > 1.0:
                failures.append(f"CRITICAL: {key} Industry_Benchmark {row[5]} != {bench} (dashboard)")
            # internal consistency: Variance == Actual - Target
            if a_act is not None and a_tgt is not None:
                if a_var is None or abs(a_var - (a_act - a_tgt)) > 1.0:
                    failures.append(f"CRITICAL: {key} Variance {row[3]} != Actual-Target")
                exp_pct = round((a_act - a_tgt) / a_tgt * 100, 1)
                if a_pct is None or abs(a_pct - exp_pct) > 1.0:
                    failures.append(f"CRITICAL: {key} Variance_Pct {row[4]} != {exp_pct}")
            # vs_Benchmark boundary logic vs dashboard benchmark
            if a_act is not None:
                exp_vs = "above" if a_act >= bench else "below"
                if a_vs != exp_vs:
                    failures.append(f"CRITICAL: {key} vs_Benchmark '{row[6]}' != '{exp_vs}'")

    # --- Summary: core reconciliation aggregates ---
    sm = load_sheet_data(agent_wb, "Summary")
    if not sm:
        failures.append("CRITICAL: Summary sheet missing")
    else:
        sm_dict = {str(r[0]).strip().lower(): r[1] for r in sm if r and r[0]}
        for key, expected in CRITICAL_SUMMARY.items():
            if key not in sm_dict:
                failures.append(f"CRITICAL: Summary missing '{key}'")
                continue
            a_val = _num(sm_dict[key])
            if a_val is None or abs(a_val - expected) > 1.0:
                failures.append(f"CRITICAL: Summary {key} {sm_dict[key]} != {expected}")

    # --- Top Products: per-region top-3 revenue (descending) ---
    tp = load_sheet_data(agent_wb, "Top Products")
    gt_tp_rev = {
        "азиатско-тихоокеанский регион": [3101.08, 2560.2, 2380.56],
        "ближний восток": [3161.12, 3036.06, 2614.43],
        "европа": [4156.38, 3837.78, 3753.96],
        "латинская америка": [6614.35, 2484.9, 2200.14],
        "северная америка": [5639.7, 3373.3, 3315.58],
    }
    if not tp:
        failures.append("CRITICAL: Top Products sheet missing")
    else:
        from collections import defaultdict
        by_region = defaultdict(list)
        for row in tp[1:]:
            if row and row[0]:
                by_region[str(row[0]).strip().lower()].append(_num(row[3]))
        for key, gt_revs in gt_tp_rev.items():
            a_revs = by_region.get(key, [])
            if len(a_revs) < 3:
                failures.append(f"CRITICAL: Top Products '{key}' has <3 rows")
                continue
            for rank, g in enumerate(gt_revs):
                a = a_revs[rank]
                if a is None or abs(a - g) > 1.0:
                    failures.append(f"CRITICAL: Top Products {key} #{rank+1} revenue {a_revs[rank]} != {g}")

    return failures


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # Resolve paths
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))

    agent_file = os.path.join(args.agent_workspace, "Q1_Sales_Review.xlsx") if args.agent_workspace else None
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_file = os.path.join(gt_dir, "Q1_Sales_Review.xlsx")

    if not agent_file or not os.path.exists(agent_file):
        print(f"FAIL: Agent output file not found: {agent_file}")
        sys.exit(1)

    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth file not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # CRITICAL CHECKS (semantic): any failure => hard FAIL before accuracy gate.
    print("[CRITICAL] Checking load-bearing reconciliation values ...")
    critical_failures = run_critical_checks(agent_wb)
    if critical_failures:
        for f in critical_failures:
            print(f"  {f}")
        print("\n=== RESULT: FAIL (critical check failed) ===")
        sys.exit(1)
    print("  PASS")

    # Check 1: Regional Performance
    print("[1/3] Checking Regional Performance ...")
    a_rp = load_sheet_data(agent_wb, "Regional Performance")
    g_rp = load_sheet_data(gt_wb, "Regional Performance")
    rp_ok, rp_errors = check_regional_performance(a_rp, g_rp)
    if rp_ok:
        print("  PASS")
    else:
        for e in rp_errors:
            print(f"  ERROR: {e}")

    # Check 2: Top Products
    print("[2/3] Checking Top Products ...")
    a_tp = load_sheet_data(agent_wb, "Top Products")
    g_tp = load_sheet_data(gt_wb, "Top Products")
    tp_ok, tp_errors = check_top_products(a_tp, g_tp)
    if tp_ok:
        print("  PASS")
    else:
        for e in tp_errors:
            print(f"  ERROR: {e}")

    # Check 3: Summary
    print("[3/3] Checking Summary ...")
    a_sm = load_sheet_data(agent_wb, "Summary")
    g_sm = load_sheet_data(gt_wb, "Summary")
    sm_ok, sm_errors = check_summary(a_sm, g_sm)
    if sm_ok:
        print("  PASS")
    else:
        for e in sm_errors:
            print(f"  ERROR: {e}")

    overall = rp_ok and tp_ok and sm_ok
    print(f"\n=== RESULT: {'PASS' if overall else 'FAIL'} ===")
    print(f"Regional Performance: {'PASS' if rp_ok else 'FAIL'}")
    print(f"Top Products: {'PASS' if tp_ok else 'FAIL'}")
    print(f"Summary: {'PASS' if sm_ok else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
