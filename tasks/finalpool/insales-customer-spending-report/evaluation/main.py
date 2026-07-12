"""Evaluation for insales-customer-spending-report (InSales / wc.* schema)."""
import argparse
import os
import sys
import openpyxl
import psycopg2


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_gsheet_data():
    """Read Google Sheet data from the gsheet.* PG schema (written by google_sheet MCP)."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets ORDER BY created_at DESC")
    spreadsheets = cur.fetchall()

    result = {"spreadsheet": None, "sheets": {}, "cells": {}}

    # Prefer the spreadsheet titled "Customer Analysis"
    for ss_id, ss_title in spreadsheets:
        if "customer analysis" in (ss_title or "").lower():
            result["spreadsheet"] = (ss_id, ss_title)
            break
    if not result["spreadsheet"] and spreadsheets:
        result["spreadsheet"] = spreadsheets[0]

    if result["spreadsheet"]:
        ss_id = result["spreadsheet"][0]
        cur.execute(
            "SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s ORDER BY index",
            (ss_id,),
        )
        for sheet_id, sheet_title in cur.fetchall():
            key = (sheet_title or "").lower()
            result["sheets"][key] = sheet_id
            cur.execute(
                """SELECT row_index, col_index, value FROM gsheet.cells
                   WHERE spreadsheet_id = %s AND sheet_id = %s
                   ORDER BY row_index, col_index""",
                (ss_id, sheet_id),
            )
            cells = {}
            for row_idx, col_idx, value in cur.fetchall():
                cells.setdefault(row_idx, {})[col_idx] = value
            result["cells"][key] = cells

    cur.close()
    conn.close()
    return result


def _summary_map(rows):
    """Return {metric_lower: value} from a Summary sheet's data rows."""
    out = {}
    for r in rows:
        if r and r[0] is not None and len(r) > 1:
            out[str(r[0]).strip().lower()] = r[1]
    return out


def _summary_lookup(rows, keys):
    """Header/orientation-tolerant {metric_lower: value} map for a Summary sheet.

    Scans ALL cells; a metric's value is the cell to its right, or the cell
    below when the right neighbour is itself another metric key (horizontal
    layout). A 'Metric/Value' header never collides with metric keys.
    """
    keys = {str(k).strip().lower() for k in keys}
    out = {}
    rows = rows or []

    def _is_key(v):
        return v is not None and str(v).strip().lower() in keys

    for r, row in enumerate(rows):
        for c, cell in enumerate(row or []):
            if cell is None:
                continue
            k = str(cell).strip().lower()
            if k not in keys or k in out:
                continue
            right = row[c + 1] if c + 1 < len(row) else None
            below = None
            if r + 1 < len(rows) and rows[r + 1] and c < len(rows[r + 1]):
                below = rows[r + 1][c]
            if right is not None and not _is_key(right):
                out[k] = right
            elif below is not None and not _is_key(below):
                out[k] = below
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Customer_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Customer_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_failures = []

    # ---------- Customer Spending sheet ----------
    print("  Checking Customer Spending...")
    a_rows = load_sheet_rows(agent_wb, "Customer Spending")
    g_rows = load_sheet_rows(gt_wb, "Customer Spending")
    a_spending_data = []
    g_spending_data = []
    if a_rows is None:
        all_errors.append("Sheet 'Customer Spending' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Customer Spending' not found in groundtruth")
    else:
        errors = []
        a_spending_data = a_rows[1:] if len(a_rows) > 1 else []
        g_spending_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_spending_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_spending_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 0):
                    errors.append(f"{key}.Orders: {a_row[3]} vs {g_row[3]} (tol=0)")
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 2.0):
                    errors.append(f"{key}.Total_Spent: {a_row[4]} vs {g_row[4]} (tol=2.0)")
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1.0):
                    errors.append(f"{key}.Avg_Order_Value: {a_row[5]} vs {g_row[5]} (tol=1.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # ---------- Summary sheet ----------
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    a_summary = {}
    g_summary = {}
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        g_summary = _summary_map(g_rows[1:] if len(g_rows) > 1 else [])
        a_summary = _summary_lookup(a_rows, g_summary.keys())
        for key, g_val in g_summary.items():
            a_val = a_summary.get(key)
            if a_val is None:
                errors.append(f"Missing Summary metric: {key}")
                continue
            # tighten generic tolerance; string metrics fall back to str compare
            if not num_close(a_val, g_val, 5.0):
                errors.append(f"{key}.Value: {a_val} vs {g_val} (tol=5.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # ---------- Google Sheet "Customer Analysis" / "Spending" ----------
    print("  Checking Google Sheet 'Customer Analysis'...")
    gsheet_top_id = None
    gsheet_top_spent = None
    gsheet_spending_rows = 0
    try:
        gdata = get_gsheet_data()
    except Exception as e:
        gdata = {"spreadsheet": None, "sheets": {}, "cells": {}}
        print(f"    WARN: could not read gsheet schema: {e}")

    if not gdata["spreadsheet"]:
        all_errors.append("Google Sheet 'Customer Analysis' not found")
    else:
        ss_title = gdata["spreadsheet"][1]
        print(f"    Found spreadsheet: {ss_title}")
        if "customer analysis" not in (ss_title or "").lower():
            all_errors.append(f"Google Sheet title is not 'Customer Analysis': {ss_title}")
        spending_key = None
        for k in gdata["sheets"]:
            if "spending" in k:
                spending_key = k
                break
        if spending_key is None:
            all_errors.append(f"Google Sheet tab 'Spending' not found; tabs={list(gdata['sheets'].keys())}")
        else:
            cells = gdata["cells"].get(spending_key, {})
            header = {str(v).strip().lower(): c for c, v in cells.get(0, {}).items() if v is not None}
            data_rows = sorted(r for r in cells if r > 0)
            gsheet_spending_rows = len(data_rows)
            cid_col = header.get("customer_id")
            spent_col = header.get("total_spent")
            if data_rows and cid_col is not None and spent_col is not None:
                # top-ranked row = first data row (sorted descending in deliverable)
                first = cells[data_rows[0]]
                gsheet_top_id = first.get(cid_col)
                gsheet_top_spent = first.get(spent_col)
            print(f"    Spending tab: {gsheet_spending_rows} data rows, top_id={gsheet_top_id}, top_spent={gsheet_top_spent}")

    # ============================================================
    # CRITICAL CHECKS (semantic core; any fail => sys.exit(1))
    # ============================================================
    print("\n  === CRITICAL checks ===")

    # GT-derived reference values
    gt_spending_lookup = {}
    for g_row in g_spending_data:
        if g_row and g_row[0] is not None:
            gt_spending_lookup[str(g_row[0]).strip().lower()] = g_row
    gt_top_name = g_summary.get("top_customer")
    gt_active = g_summary.get("total_active_customers")
    gt_spent_all = g_summary.get("total_spent_all")
    # GT top-spending row (sorted desc => first data row)
    gt_top_row = g_spending_data[0] if g_spending_data else None

    # CRITICAL 1: Top_Customer name matches GT (RU, case-insensitive on ORIGINAL strings)
    a_top_name = a_summary.get("top_customer")
    if str_match(a_top_name, gt_top_name):
        # also verify that customer's Total_Spent in Customer Spending matches GT (tol<=2.0)
        ok_spent = False
        if gt_top_row is not None and len(gt_top_row) > 0:
            tkey = str(gt_top_row[0]).strip().lower()
            a_top = None
            for row in a_spending_data:
                if row and row[0] is not None and str(row[0]).strip().lower() == tkey:
                    a_top = row
                    break
            if a_top is not None and len(a_top) > 4 and len(gt_top_row) > 4:
                ok_spent = num_close(a_top[4], gt_top_row[4], 2.0)
        if ok_spent:
            print(f"    CRITICAL PASS: Top_Customer='{a_top_name}' & their Total_Spent matches GT")
        else:
            critical_failures.append(
                f"CRITICAL: Top_Customer name OK but their Total_Spent does not match GT (expected ~{gt_top_row[4] if gt_top_row and len(gt_top_row) > 4 else '?'})"
            )
    else:
        critical_failures.append(f"CRITICAL: Top_Customer '{a_top_name}' != GT '{gt_top_name}'")

    # CRITICAL 2: Total_Active_Customers == GT exact integer (no tolerance)
    a_active = a_summary.get("total_active_customers")
    if num_close(a_active, gt_active, 0):
        print(f"    CRITICAL PASS: Total_Active_Customers={a_active}")
    else:
        critical_failures.append(f"CRITICAL: Total_Active_Customers {a_active} != GT {gt_active}")

    # CRITICAL 3: Total_Spent_All within tol<=2.0 of GT AND equals sum of per-customer Total_Spent
    a_spent_all = a_summary.get("total_spent_all")
    ok3 = num_close(a_spent_all, gt_spent_all, 2.0)
    # integrity: sum of agent per-customer Total_Spent ~ reported Total_Spent_All
    per_sum = 0.0
    sum_ok = True
    for row in a_spending_data:
        if row and len(row) > 4 and row[4] is not None:
            try:
                per_sum += float(row[4])
            except (TypeError, ValueError):
                sum_ok = False
    try:
        integrity = abs(per_sum - float(a_spent_all)) <= 2.0 if a_spent_all is not None else False
    except (TypeError, ValueError):
        integrity = False
    if ok3 and sum_ok and integrity:
        print(f"    CRITICAL PASS: Total_Spent_All={a_spent_all} (sum of rows={round(per_sum, 2)})")
    else:
        critical_failures.append(
            f"CRITICAL: Total_Spent_All {a_spent_all} vs GT {gt_spent_all} (tol=2.0); rows sum={round(per_sum, 2)} integrity_ok={integrity}"
        )

    # CRITICAL 4: Google Sheet 'Customer Analysis'/'Spending' exists and its top row matches GT top customer
    if gt_top_row is not None and gsheet_top_id is not None and gsheet_top_spent is not None:
        id_ok = str(gsheet_top_id).strip().lower() == str(gt_top_row[0]).strip().lower()
        spent_ok = num_close(gsheet_top_spent, gt_top_row[4] if len(gt_top_row) > 4 else None, 2.0)
        if id_ok and spent_ok:
            print(f"    CRITICAL PASS: gsheet 'Spending' top row matches GT (id={gsheet_top_id})")
        else:
            critical_failures.append(
                f"CRITICAL: gsheet 'Spending' top row id={gsheet_top_id}/spent={gsheet_top_spent} != GT id={gt_top_row[0]}/spent={gt_top_row[4] if len(gt_top_row) > 4 else '?'}"
            )
    else:
        critical_failures.append(
            "CRITICAL: gsheet 'Customer Analysis'/'Spending' missing or lacks Customer_ID/Total_Spent header"
        )

    # CRITICAL 5: Customer Spending sorted by Total_Spent desc AND row count == active customers
    spent_vals = []
    for row in a_spending_data:
        if row and len(row) > 4 and row[4] is not None:
            try:
                spent_vals.append(float(row[4]))
            except (TypeError, ValueError):
                pass
    is_sorted = all(spent_vals[i] >= spent_vals[i + 1] - 0.01 for i in range(len(spent_vals) - 1))
    count_ok = num_close(len(a_spending_data), gt_active, 0)
    if is_sorted and count_ok:
        print(f"    CRITICAL PASS: rows sorted desc & count={len(a_spending_data)}")
    else:
        critical_failures.append(
            f"CRITICAL: sorted_desc={is_sorted}, row_count={len(a_spending_data)} (expected {gt_active})"
        )

    if critical_failures:
        print(f"\n=== RESULT: FAIL ({len(critical_failures)} CRITICAL failures) ===")
        for e in critical_failures:
            print(f"  {e}")
        sys.exit(1)

    # ============================================================
    # Accuracy gate (non-critical structural/tolerant checks)
    # ============================================================
    # Treat the per-row/summary tolerant checks as the accuracy basis.
    total_checks = len(g_spending_data) * 3 + len(g_summary) + 1  # +1 gsheet structural
    failed = len(all_errors)
    passed = max(total_checks - failed, 0)
    accuracy = (passed / total_checks * 100.0) if total_checks else 0.0
    print(f"\n  Accuracy: {accuracy:.1f}% ({passed}/{total_checks}, {failed} non-critical errors)")

    if all_errors:
        print(f"\n  Non-critical errors ({len(all_errors)}):")
        for e in all_errors[:10]:
            print(f"    {e}")

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
