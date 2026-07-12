"""Evaluation для kulinar-catering-forms-teamly-excel (RU-стек: kulinar/forms/teamly).

Проверяет четыре части задачи:
  - Excel Catering_Budget.xlsx: лист Menu (4 колонки в порядке + 6-8 строк),
    лист Summary (Total_Dishes / Total_Budget=500 / Budget_Per_Person=25).
  - Forms (gform.*): опрос пищевых предпочтений с >=3 вопросами.
  - Teamly (teamly.pages): страница «Меню корпоративного обеда» с описанием
    мероприятия и перечнем выбранных блюд.
  - Рецепты соответствуют реальным блюдам из базы kulinar (не выдуманы).

Критические чеки (CRITICAL_CHECKS): любой их fail => задача FAIL, даже если общая
accuracy >= 70%. Остальные (структурные) — мягкие. Порог: accuracy >= 70% И нет
критических провалов.
"""
import argparse
import json
import os
import sys
import unicodedata

import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=5432,
    dbname=os.environ.get("PGDATABASE", "cowork_gym"),
    user="eigent",
    password="camel",
)

# Канонический набор блюд kulinar (источник: local_servers/kulinar-mcp -> all_recipes.json).
# Критический чек "рецепты реальны" сверяет названия из Menu с этим множеством.
KULINAR_RECIPES = {
    "Салат Оливье", "Винегрет", "Сельдь под шубой", "Салат Мимоза", "Крабовый салат",
    "Греческий салат", "Салат с курицей и грибами", "Холодец", "Икра кабачковая",
    "Грибы маринованные", "Сало солёное", "Селёдка с луком", "Борщ",
    "Щи из квашеной капусты", "Солянка мясная", "Уха", "Окрошка", "Грибной суп",
    "Рассольник", "Куриный бульон с лапшой", "Бефстроганов", "Пельмени домашние",
    "Голубцы", "Котлеты домашние", "Жаркое в горшочках", "Курица в сметане",
    "Рыба запечённая по-русски", "Цыплёнок табака", "Гречка с тушёнкой",
    "Плов узбекский", "Картофельное пюре", "Гречневая каша", "Перловая каша",
    "Картофель отварной с укропом", "Рис отварной", "Пирожки с капустой жареные",
    "Пирожки с мясом печёные", "Блины тонкие", "Кулебяка с капустой и яйцом",
    "Расстегаи с рыбой", "Медовик", "Наполеон", "Сырники", "Пасха творожная",
    "Ватрушки с творогом", "Кисель ягодный", "Морс клюквенный",
    "Компот из сухофруктов", "Сбитень", "Квас домашний",
}

# Критические чеки по имени record()
CRITICAL_CHECKS = {
    "Excel Menu: 4 колонки в порядке + 6-8 строк данных",
    "Excel Summary: Total_Budget=500 и Budget_Per_Person=25",
    "Forms: опрос пищевых предпочтений с >=3 вопросами",
    "Teamly: страница «Меню корпоративного обеда» с мероприятием и блюдами",
    "Рецепты Menu — реальные блюда из базы kulinar",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:250]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def normalize(s: str) -> str:
    """Lowercase + схлопывание кириллических/латинских двойников (А/A, С/C...).
    Только для ID-сопоставлений (translit), НЕ для поиска русских ключевых слов."""
    s = (s or "").lower()
    s = unicodedata.normalize("NFKD", s)
    table = str.maketrans({"а": "a", "о": "o", "е": "e", "р": "p",
                           "с": "c", "у": "y", "к": "k", "х": "x"})
    return s.translate(table)


def norm_recipe(s: str) -> str:
    """Нормализация названия блюда для сравнения с каноном (регистр, ё/е, пробелы)."""
    s = (s or "").strip().lower().replace("ё", "е")
    return " ".join(s.split())


CANON_NORM = {norm_recipe(r) for r in KULINAR_RECIPES}


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def check_excel(agent_workspace):
    print("\n=== Проверка Excel Catering_Budget.xlsx ===")
    menu_recipe_cells = []
    try:
        import openpyxl
    except ImportError:
        record("openpyxl доступен", False, "openpyxl не установлен")
        return menu_recipe_cells

    agent_file = os.path.join(agent_workspace, "Catering_Budget.xlsx")
    record("Catering_Budget.xlsx существует", os.path.isfile(agent_file), f"Ожидался {agent_file}")
    if not os.path.isfile(agent_file):
        record("Excel Menu: 4 колонки в порядке + 6-8 строк данных", False, "нет файла")
        record("Excel Summary: Total_Budget=500 и Budget_Per_Person=25", False, "нет файла")
        return menu_recipe_cells

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel читается", False, str(e))
        record("Excel Menu: 4 колонки в порядке + 6-8 строк данных", False, "файл не читается")
        record("Excel Summary: Total_Budget=500 и Budget_Per_Person=25", False, "файл не читается")
        return menu_recipe_cells

    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    # --- Лист Menu ---
    print("\n--- Лист Menu ---")
    menu_ws = get_sheet(wb, "Menu")
    record("Лист 'Menu' существует", menu_ws is not None, f"Найдены: {wb.sheetnames}")

    if menu_ws is not None:
        headers = [c.value for c in list(menu_ws.rows)[0]] if menu_ws.max_row > 0 else []
        hl = [str(h).strip().lower() if h is not None else "" for h in headers]

        # Структурные (мягкие) чеки на наличие отдельных колонок
        record("Menu: колонка Recipe_Name", any("recipe" in h for h in hl), f"Headers: {headers}")
        record("Menu: колонка Category", any("category" in h for h in hl), f"Headers: {headers}")
        record("Menu: колонка Servings", any("serving" in h for h in hl), f"Headers: {headers}")
        record("Menu: колонка Estimated_Cost_Per_Person",
               any("cost" in h for h in hl), f"Headers: {headers}")

        data_rows = [row for row in menu_ws.iter_rows(min_row=2, values_only=True)
                     if any(v is not None and str(v).strip() for v in row)]

        # CRITICAL: все 4 колонки строго в порядке + 6-8 строк
        expected = ["recipe", "category", "serving", "cost"]
        order_ok = all(i < len(hl) and expected[i] in hl[i] for i in range(4))
        rows_ok = 6 <= len(data_rows) <= 8
        record("Excel Menu: 4 колонки в порядке + 6-8 строк данных",
               order_ok and rows_ok,
               f"order_ok={order_ok}, rows={len(data_rows)}, headers={headers}")

        # Собираем названия блюд из первой колонки для критического чека рецептов
        for row in data_rows:
            if row and row[0] is not None and str(row[0]).strip():
                menu_recipe_cells.append(str(row[0]).strip())

    # --- Лист Summary ---
    print("\n--- Лист Summary ---")
    sum_ws = get_sheet(wb, "Summary")
    record("Лист 'Summary' существует", sum_ws is not None, f"Найдены: {wb.sheetnames}")

    if sum_ws is not None:
        summary_data = {}
        for row in sum_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary_data[str(row[0]).strip().lower()] = row[1]

        record("Summary: есть Total_Dishes",
               any("total_dishes" in k or "total dishes" in k for k in summary_data),
               f"Keys: {list(summary_data.keys())}")

        budget_ok = any(("total_budget" in k or "total budget" in k) and num_close(v, 500, 1)
                        for k, v in summary_data.items())
        per_person_ok = any(("budget_per_person" in k or "budget per person" in k) and num_close(v, 25, 1)
                            for k, v in summary_data.items())
        record("Excel Summary: Total_Budget=500 и Budget_Per_Person=25",
               budget_ok and per_person_ok, f"Data: {summary_data}")

        # Мягкий чек: Total_Dishes == числу строк Menu
        td_val = None
        for k, v in summary_data.items():
            if "total_dishes" in k or "total dishes" in k:
                td_val = v
                break
        record("Summary: Total_Dishes равно числу строк Menu",
               td_val is not None and num_close(td_val, len(menu_recipe_cells), 0),
               f"Total_Dishes={td_val}, строк Menu={len(menu_recipe_cells)}")

    return menu_recipe_cells


# ---------------------------------------------------------------------------
# Recipes vs kulinar
# ---------------------------------------------------------------------------
def check_recipes(menu_recipe_cells):
    print("\n=== Проверка рецептов против базы kulinar ===")
    if not menu_recipe_cells:
        record("Рецепты Menu — реальные блюда из базы kulinar", False,
               "нет названий блюд в листе Menu")
        return
    matched = 0
    unknown = []
    for name in menu_recipe_cells:
        n = norm_recipe(name)
        # допускаем точное совпадение или вхождение каноничного названия в ячейку
        if n in CANON_NORM or any(c in n or n in c for c in CANON_NORM):
            matched += 1
        else:
            unknown.append(name)
    # Критично: подавляющее большинство блюд должны быть реальными (>= все, допускаем 0 ошибок)
    record("Рецепты Menu — реальные блюда из базы kulinar",
           matched == len(menu_recipe_cells) and matched >= 6,
           f"совпало {matched}/{len(menu_recipe_cells)}; неизвестные: {unknown[:5]}")


# ---------------------------------------------------------------------------
# Forms (gform)
# ---------------------------------------------------------------------------
def check_forms():
    print("\n=== Проверка Forms (опрос предпочтений) ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        record("Forms: опрос пищевых предпочтений с >=3 вопросами", False, str(e))
        return
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    record("Создана хотя бы одна форма", len(forms) >= 1, f"Найдено {len(forms)} форм")

    # Сопоставление заголовка: RU ключевые слова в .lower() ИЛИ EN-токены ИЛИ translit
    def title_matches(title):
        t = (title or "").lower()
        ru_hit = any(k in t for k in ("обед", "пищев", "предпочт", "питани", "аллерг", "диет"))
        en_hit = any(k in t for k in ("dietary", "lunch", "team", "menu", "preference"))
        return ru_hit or en_hit

    target = None
    for fid, title in forms:
        if title_matches(title):
            target = fid
            break

    q_count = 0
    if target is not None:
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (target,))
        q_count = cur.fetchone()[0]
    elif forms:
        # запасной вариант: берём форму с наибольшим числом вопросов
        best_fid, best_q = None, -1
        for fid, _t in forms:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (fid,))
            qc = cur.fetchone()[0]
            if qc > best_q:
                best_fid, best_q = fid, qc
        target, q_count = best_fid, best_q

    record("Forms: найден опрос с подходящим заголовком",
           any(title_matches(t) for _f, t in forms),
           f"Заголовки: {[t for _f, t in forms]}")

    # Тематическое покрытие вопросов (мягкий чек)
    topics_ok = False
    if target is not None:
        cur.execute(
            "SELECT lower(title) || ' ' || coalesce(lower(description),'') "
            "FROM gform.questions WHERE form_id = %s", (target,))
        qtext = " ".join(r[0] for r in cur.fetchall())
        has_diet = any(k in qtext for k in ("ограничен", "предпочт", "вегетариан", "веган",
                                            "глютен", "diet", "restrict", "preference"))
        has_allergy = any(k in qtext for k in ("аллерг", "allerg"))
        has_comment = any(k in qtext for k in ("комментар", "пожелан", "дополнит", "comment", "request", "additional"))
        topics_ok = has_diet and has_allergy and has_comment
        record("Forms: вопросы покрывают ограничения/аллергии/комментарии",
               topics_ok,
               f"diet={has_diet}, allergy={has_allergy}, comment={has_comment}")
    else:
        record("Forms: вопросы покрывают ограничения/аллергии/комментарии", False, "форма не найдена")

    # CRITICAL: подходящий опрос с >=3 вопросами
    record("Forms: опрос пищевых предпочтений с >=3 вопросами",
           target is not None and any(title_matches(t) for _f, t in forms) and q_count >= 3,
           f"вопросов={q_count}")

    conn.close()


# ---------------------------------------------------------------------------
# Teamly
# ---------------------------------------------------------------------------
def check_teamly():
    print("\n=== Проверка Teamly (страница меню) ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        record("Teamly: страница «Меню корпоративного обеда» с мероприятием и блюдами", False, str(e))
        return
    cur = conn.cursor()
    cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
    pages = cur.fetchall()
    conn.close()

    # Поиск страницы меню: RU ключевые слова или EN/translit
    def title_matches(title):
        t = (title or "").lower()
        ru = ("меню" in t and ("обед" in t or "корпоратив" in t or "команд" in t))
        en = ("menu" in t and ("lunch" in t or "team" in t))
        return ru or en

    candidates = [(t, b) for t, b in pages if title_matches(t)]
    record("Teamly: страница с заголовком меню создана",
           len(candidates) >= 1, f"Заголовки: {[t for t, _b in pages]}")

    # Содержимое: упоминание мероприятия (20 человек / 500) + список блюд из kulinar
    body = "\n\n".join(b for _t, b in candidates)
    bl = body.lower()
    has_people = "20" in body
    has_budget = ("500" in body) or ("25" in body)
    nb = norm_recipe(body)
    dishes_found = sum(1 for c in CANON_NORM if c and c in nb)

    record("Teamly: упомянуты 20 человек и бюджет (500/25)",
           bool(candidates) and has_people and has_budget,
           f"people={has_people}, budget={has_budget}")
    record("Teamly: на странице перечислены блюда из kulinar (>=6)",
           dishes_found >= 6, f"найдено блюд={dishes_found}")

    # CRITICAL: страница есть, документирует мероприятие и содержит >=4 реальных блюда
    record("Teamly: страница «Меню корпоративного обеда» с мероприятием и блюдами",
           bool(candidates) and has_people and has_budget and dishes_found >= 4,
           f"candidates={len(candidates)}, people={has_people}, budget={has_budget}, dishes={dishes_found}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("KULINAR CATERING FORMS TEAMLY EXCEL - EVALUATION")
    print("=" * 70)

    menu_recipes = check_excel(args.agent_workspace)
    check_recipes(menu_recipes or [])
    check_forms()
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    pct = 100.0 * PASS_COUNT / total if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {pct:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAIL: {CRITICAL_FAILED}")
        print("  Overall: FAIL")
        sys.exit(1)

    overall = pct >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
