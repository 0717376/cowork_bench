"""Evaluation script for terminal-sf-insales-support-email-gcal-excel.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

NOTE: all substantive values (ticket counts per priority, order counts per
status) are read LIVE from the DB (sf_data / wc schemas), never hardcoded, so
the central russification map stays in sync.
"""
import os
import argparse, json, os, re, sys
import openpyxl


def normalize_ru_numbers(text):
    """RU number normalization for substring/regex checks: collapse digit-group
    separators (space/NBSP/NNBSP/dot/comma before a 3-digit group) and turn
    decimal commas into dots ("31 588" -> "31588", "4 586,91" -> "4586.91")."""
    t = str(text or "")
    t = re.sub(r"(?<=\d)[ \xa0\u202f\u2009.,](?=\d{3}\b)", "", t)
    return re.sub(r"(?<=\d),(?=\d)", ".", t)

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Per-priority Ticket_Count matches live DB",
    "Pct_of_Total matches count/total within tolerance",
    "Order_Status_Breakdown counts match live DB",
    "Quality_Flag correct for both Review and OK statuses",
    "Quality analysis email has substantive body (counts + defect rate)",
    "Quality Review Meeting on 2026-03-19",
    "Quality Improvement Planning on 2026-03-21",
    "quality_analysis_results.json exists with per-priority defect rate",
}


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_expected_ticket_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT "PRIORITY", COUNT(*) as cnt, ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2) as avg_resp
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY" ORDER BY cnt DESC
    """)
    rows = cur.fetchall()
    cur.execute('SELECT COUNT(*) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return rows, total


def get_expected_order_data():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT status, COUNT(*) FROM wc.orders GROUP BY status ORDER BY COUNT(*) DESC")
    rows = cur.fetchall()
    cur.execute("SELECT COUNT(*) FROM wc.orders")
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return rows, total


# Statuses that must be flagged for review vs. flagged OK.
REVIEW_STATUSES = {"refunded", "cancelled", "failed"}
OK_STATUSES = {"completed", "processing", "on-hold", "pending"}


def check_ticket_by_priority(wb, ticket_data, total_tickets):
    """CRITICAL: per-priority Ticket_Count + Pct_of_Total match live DB."""
    if "Ticket_by_Priority" not in wb.sheetnames:
        check("Ticket_by_Priority sheet exists", False)
        check("Per-priority Ticket_Count matches live DB", False, "no sheet", critical=True)
        check("Pct_of_Total matches count/total within tolerance", False, "no sheet", critical=True)
        return
    check("Ticket_by_Priority sheet exists", True)
    ws = wb["Ticket_by_Priority"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    check("Ticket_by_Priority has 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    hmap = {h: i for i, h in enumerate(headers)}
    for col in ['Priority', 'Ticket_Count', 'Avg_Response_Hours', 'Pct_of_Total']:
        check(f"Ticket_by_Priority has {col}", col.lower() in headers, f"headers: {headers[:5]}")

    pri_i = hmap.get("priority")
    cnt_i = hmap.get("ticket_count")
    pct_i = hmap.get("pct_of_total")
    row_dict = {}
    if pri_i is not None:
        for r in data_rows:
            if pri_i < len(r) and r[pri_i]:
                row_dict[str(r[pri_i]).strip()] = r

    # CRITICAL: every DB priority count must be present and exact (tight tol).
    counts_ok = pri_i is not None and cnt_i is not None and len(row_dict) > 0
    pct_ok = counts_ok and pct_i is not None
    for priority, cnt, avg_resp in ticket_data:
        r = row_dict.get(priority)
        if r is None:
            counts_ok = False
            pct_ok = False
            continue
        agent_cnt = safe_float(r[cnt_i]) if cnt_i < len(r) else None
        if agent_cnt is None or abs(agent_cnt - cnt) >= 1:
            counts_ok = False
        if pct_i is not None and pct_i < len(r):
            agent_pct = safe_float(r[pct_i])
            expected_pct = round(cnt / total_tickets * 100, 1) if total_tickets else 0
            if agent_pct is None or abs(agent_pct - expected_pct) > 1.0:
                pct_ok = False
    check("Per-priority Ticket_Count matches live DB", counts_ok,
          f"expected {[(p, c) for p, c, _ in ticket_data]}", critical=True)
    check("Pct_of_Total matches count/total within tolerance", pct_ok,
          f"total={total_tickets}", critical=True)


def check_order_status(wb, order_data, total_orders):
    """CRITICAL: per-status counts match DB + Quality_Flag both directions."""
    if "Order_Status_Breakdown" not in wb.sheetnames:
        check("Order_Status_Breakdown sheet exists", False)
        check("Order_Status_Breakdown counts match live DB", False, "no sheet", critical=True)
        check("Quality_Flag correct for both Review and OK statuses", False, "no sheet", critical=True)
        return
    check("Order_Status_Breakdown sheet exists", True)
    ws = wb["Order_Status_Breakdown"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    check("Order_Status_Breakdown has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    hmap = {h: i for i, h in enumerate(headers)}
    for col in ['Status', 'Order_Count', 'Pct_of_Orders', 'Quality_Flag']:
        check(f"Order_Status_Breakdown has {col}", col.lower() in headers, f"headers: {headers[:5]}")

    st_i = hmap.get("status")
    cnt_i = hmap.get("order_count")
    flag_i = hmap.get("quality_flag")
    row_dict = {}
    if st_i is not None:
        for r in data_rows:
            if st_i < len(r) and r[st_i]:
                row_dict[str(r[st_i]).strip().lower()] = r

    counts_ok = st_i is not None and cnt_i is not None and len(row_dict) > 0
    for status, cnt in order_data:
        r = row_dict.get(str(status).lower())
        if r is None:
            counts_ok = False
            continue
        agent_cnt = safe_float(r[cnt_i]) if cnt_i < len(r) else None
        if agent_cnt is None or abs(agent_cnt - cnt) >= 1:
            counts_ok = False
    check("Order_Status_Breakdown counts match live DB", counts_ok,
          f"expected {order_data}", critical=True)

    # CRITICAL: Quality_Flag both directions. Review for refunded/cancelled/failed,
    # OK for completed/processing/on-hold/pending.
    flag_ok = flag_i is not None and len(row_dict) > 0
    for status, r in row_dict.items():
        flag = str(r[flag_i]).strip().lower() if flag_i is not None and flag_i < len(r) and r[flag_i] else ""
        if status in REVIEW_STATUSES:
            if "review" not in flag:
                flag_ok = False
        elif status in OK_STATUSES:
            if flag != "ok":
                flag_ok = False
    check("Quality_Flag correct for both Review and OK statuses", flag_ok,
          f"flags: {[(s, str(r[flag_i]) if flag_i is not None and flag_i < len(r) else '?') for s, r in row_dict.items()]}",
          critical=True)


def check_action_plan(wb):
    if "Quality_Action_Plan" not in wb.sheetnames:
        check("Quality_Action_Plan sheet exists", False)
        return
    check("Quality_Action_Plan sheet exists", True)
    ws = wb["Quality_Action_Plan"]
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    check("Quality_Action_Plan has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    for col in ['Issue', 'Severity', 'Owner', 'Deadline', 'Action_Item']:
        check(f"Quality_Action_Plan has {col}", col.lower() in headers, f"headers: {headers[:6]}")
    # Severity values within allowed set (soft).
    sev_i = headers.index("severity") if "severity" in headers else None
    if sev_i is not None:
        bad = [r[sev_i] for r in data_rows
               if r[sev_i] and str(r[sev_i]).strip().lower() not in ("high", "medium", "low")]
        check("Quality_Action_Plan Severity values valid", len(bad) == 0, f"bad: {bad[:3]}")


def check_results_json(agent_workspace, ticket_data, order_data):
    """CRITICAL: quality_analysis_results.json exists with per-priority defect rate."""
    path = os.path.join(agent_workspace, "quality_analysis_results.json")
    if not os.path.exists(path):
        check("quality_analysis_results.json exists with per-priority defect rate", False,
              "file missing", critical=True)
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        check("quality_analysis_results.json exists with per-priority defect rate", False,
              f"invalid JSON: {e}", critical=True)
        return
    blob = json.dumps(data, ensure_ascii=False).lower()
    # Must mention each priority and a defect-rate notion (defect / per 100 / дефект).
    priorities = [str(p).lower() for p, _, _ in ticket_data]
    has_priorities = all(p in blob for p in priorities) if priorities else False
    has_defect = any(k in blob for k in ("defect", "per_100", "per100", "per 100",
                                          "rate", "дефект", "на 100"))
    check("quality_analysis_results.json exists with per-priority defect rate",
          has_priorities and has_defect,
          f"priorities_found={has_priorities} defect_kw={has_defect}", critical=True)


def check_emails(total_tickets, ticket_data):
    """Quality analysis email must have substantive body (CRITICAL)."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, to_addr, body_text FROM email.messages WHERE subject ILIKE %s",
            ('%quality%analysis%',))
        quality_emails = cur.fetchall()
        check("Quality analysis email sent", len(quality_emails) >= 1, f"found {len(quality_emails)}")
        if quality_emails:
            check("Email to quality-team",
                  any("quality-team" in str(e[1]).lower() for e in quality_emails),
                  f"to: {quality_emails[0][1]}")
            # CRITICAL: body must reference the total ticket count, the highest
            # volume priority, and a defect-rate notion (RU+EN keywords).
            body = normalize_ru_numbers(" ".join(str(e[2] or "") for e in quality_emails).lower())
            top_priority = str(ticket_data[0][0]).lower() if ticket_data else "medium"
            has_total = str(total_tickets) in body
            has_top = top_priority in body
            has_defect = any(k in body for k in (
                "defect", "rate", "per 100", "per_100",
                "дефект", "процент", "на 100", "уровень"))
            check("Quality analysis email has substantive body (counts + defect rate)",
                  has_total and has_top and has_defect,
                  f"total={has_total} top({top_priority})={has_top} defect={has_defect}",
                  critical=True)
        else:
            check("Quality analysis email has substantive body (counts + defect rate)", False,
                  "no email", critical=True)

        cur.execute(
            "SELECT subject, to_addr FROM email.messages WHERE subject ILIKE %s",
            ('%quality%review%meeting%',))
        review_emails = cur.fetchall()
        check("Quality review meeting email sent", len(review_emails) >= 1, f"found {len(review_emails)}")
        if review_emails:
            check("Review meeting email to operations",
                  any("operations" in str(e[1]).lower() for e in review_emails),
                  f"to: {review_emails[0][1]}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email checks", False, str(e))
        check("Quality analysis email has substantive body (counts + defect rate)", False,
              str(e), critical=True)


def check_calendar():
    """CRITICAL: both events exist on exact dates Mar 19 / Mar 21 2026."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime FROM gcal.events
            WHERE summary ILIKE %s
              AND start_datetime >= '2026-03-19' AND start_datetime < '2026-03-20'
        """, ('%quality%review%',))
        review = cur.fetchall()
        check("Quality Review Meeting on 2026-03-19", len(review) >= 1,
              f"found {review}", critical=True)

        cur.execute("""
            SELECT summary, start_datetime FROM gcal.events
            WHERE summary ILIKE %s
              AND start_datetime >= '2026-03-21' AND start_datetime < '2026-03-22'
        """, ('%quality%improvement%',))
        plan = cur.fetchall()
        check("Quality Improvement Planning on 2026-03-21", len(plan) >= 1,
              f"found {plan}", critical=True)

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar checks", False, str(e))
        check("Quality Review Meeting on 2026-03-19", False, str(e), critical=True)
        check("Quality Improvement Planning on 2026-03-21", False, str(e), critical=True)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    ticket_data, total_tickets = get_expected_ticket_data()
    order_data, total_orders = get_expected_order_data()

    # Check Excel
    excel_path = os.path.join(agent_workspace, "Product_Quality_Report.xlsx")
    check("Product_Quality_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        check_ticket_by_priority(wb, ticket_data, total_tickets)
        check_order_status(wb, order_data, total_orders)
        check_action_plan(wb)
    else:
        # Excel-dependent critical checks fail.
        for n in ("Per-priority Ticket_Count matches live DB",
                  "Pct_of_Total matches count/total within tolerance",
                  "Order_Status_Breakdown counts match live DB",
                  "Quality_Flag correct for both Review and OK statuses"):
            check(n, False, "no Excel", critical=True)

    # Check terminal script
    check("defect_correlation.py exists",
          os.path.exists(os.path.join(agent_workspace, "defect_correlation.py")))

    check_results_json(agent_workspace, ticket_data, order_data)
    check_emails(total_tickets, ticket_data)
    check_calendar()

    check_reverse_validation(agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"   - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w", encoding="utf-8") as f:
                json.dump({"passed": PASS_COUNT, "total": total,
                           "accuracy": accuracy, "critical_failed": critical_failed}, f,
                          ensure_ascii=False, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Product_Quality_Report.xlsx")
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        expected_sheets = {"ticket_by_priority", "order_status_breakdown", "quality_action_plan"}
        unexpected = [s for s in wb.sheetnames if s.lower().replace(" ", "_") not in expected_sheets]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected sheets: {unexpected}")

        if "Ticket_by_Priority" in wb.sheetnames:
            ws = wb["Ticket_by_Priority"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 1 and row[1] is not None:
                    val = safe_float(row[1])
                    if val is not None and val < 0:
                        check("No negative ticket counts", False, f"Found negative: {val}")
                        break
            else:
                check("No negative ticket counts", True)

    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE summary ILIKE '%%quality%%'
              AND start_datetime < '2026-03-01'
        """)
        old_events = cur.fetchone()[0]
        check("No quality events before March 2026", old_events == 0,
              f"Found {old_events} old events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
