"""
Evaluation for arxiv-method-benchmark-tracker task.
Checks Excel (Method_Benchmark.xlsx) and the Teamly knowledge-base page.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Pass threshold otherwise: accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Leaderboard: all 5 methods present with correct Task and Score",
    "Leaderboard: Paper_ID linkage correct (GPT-4 blank)",
    "Summary: Total_Methods = 5",
    "Summary: Methods_With_Papers = 4",
    "Summary: Total_Tasks = 3",
    "Summary: Top_Score = 95.0",
    "Method Details: 4 paper rows with correct Paper_IDs and non-empty Key_Contribution + Dataset_Used",
    "Teamly tracker page has substantive method/benchmark content",
}

LEADERBOARD = [
    {"task": "Image Classification", "method": "ViT-Large", "score": 91.2, "paper_id": "2402.10001"},
    {"task": "Image Classification", "method": "ConvNeXt-XL", "score": 89.5, "paper_id": "2402.10002"},
    {"task": "Text Generation", "method": "GPT-4", "score": 95.0, "paper_id": ""},
    {"task": "Text Generation", "method": "LLaMA-3", "score": 92.3, "paper_id": "2402.10003"},
    {"task": "Image Generation", "method": "DiffusionXL", "score": 2.1, "paper_id": "2402.10004"},
]

PAPERS_WITH_ID = {"2402.10001", "2402.10002", "2402.10003", "2402.10004"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Method_Benchmark.xlsx")
    if not os.path.isfile(path):
        record("Excel file exists", False, f"Not found: {path}")
        return False
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Leaderboard sheet ───────────────────────────────────────────────────
    lb_rows = load_sheet_rows(wb, "Leaderboard")
    if lb_rows is None:
        record("Sheet 'Leaderboard' exists", False, f"Sheets: {wb.sheetnames}")
        record("Leaderboard: all 5 methods present with correct Task and Score", False, "no sheet")
        record("Leaderboard: Paper_ID linkage correct (GPT-4 blank)", False, "no sheet")
    else:
        record("Sheet 'Leaderboard' exists", True)
        header = lb_rows[0] if lb_rows else []
        data = lb_rows[1:]
        record("Leaderboard has 5 rows", len(data) == 5, f"Found {len(data)}")

        method_col = find_col(header, ["Method", "method"])
        score_col = find_col(header, ["Score", "score", "Accuracy"])
        task_col = find_col(header, ["Task", "task"])
        pid_col = find_col(header, ["Paper_ID", "Paper ID", "paper_id"])

        # Build method -> row map (lowercased method name)
        def cell(r, c):
            return r[c] if (c is not None and c < len(r)) else None

        rows_by_method = {}
        for r in data:
            m = cell(r, method_col)
            if m:
                rows_by_method[str(m).strip().lower()] = r

        # Per-method presence (non-critical informational)
        for entry in LEADERBOARD:
            present = entry["method"].lower() in rows_by_method
            record(f"Method '{entry['method']}' present", present,
                   f"Found: {list(rows_by_method.keys())}")

        # CRITICAL: all 5 methods with correct Task + Score
        all_ms_ok = True
        for entry in LEADERBOARD:
            r = rows_by_method.get(entry["method"].lower())
            if r is None:
                all_ms_ok = False
                continue
            sc = cell(r, score_col)
            tk = cell(r, task_col)
            sc_ok = num_close(sc, entry["score"], tol=0.05)
            tk_ok = (tk is not None and
                     str(tk).strip().lower() == entry["task"].strip().lower())
            if not (sc_ok and tk_ok):
                all_ms_ok = False
        record("Leaderboard: all 5 methods present with correct Task and Score",
               all_ms_ok, "method/task/score mismatch")

        # CRITICAL: Paper_ID linkage correct, GPT-4 blank
        link_ok = True
        if pid_col is None:
            link_ok = False
        else:
            for entry in LEADERBOARD:
                r = rows_by_method.get(entry["method"].lower())
                if r is None:
                    link_ok = False
                    continue
                pid = cell(r, pid_col)
                pid_s = "" if pid is None else str(pid).strip()
                if entry["paper_id"] == "":
                    if pid_s not in ("", "None"):
                        link_ok = False
                else:
                    if entry["paper_id"] not in pid_s:
                        link_ok = False
        record("Leaderboard: Paper_ID linkage correct (GPT-4 blank)",
               link_ok, "paper_id linkage mismatch")

    # ── Method Details sheet ────────────────────────────────────────────────
    md_rows = load_sheet_rows(wb, "Method Details") or load_sheet_rows(wb, "Method_Details")
    if md_rows is None:
        record("Sheet 'Method Details' exists", False, f"Sheets: {wb.sheetnames}")
        record("Method Details: 4 paper rows with correct Paper_IDs and non-empty Key_Contribution + Dataset_Used",
               False, "no sheet")
    else:
        record("Sheet 'Method Details' exists", True)
        header2 = md_rows[0] if md_rows else []
        data2 = md_rows[1:]
        record("Method Details has 4 rows", len(data2) == 4, f"Found {len(data2)}")

        id_col = find_col(header2, ["Paper_ID", "Paper ID", "paper_id"])
        kc_col = find_col(header2, ["Key_Contribution", "Key Contribution", "key_contribution"])
        ds_col = find_col(header2, ["Dataset_Used", "Dataset Used", "dataset_used", "Dataset"])

        def cell2(r, c):
            return r[c] if (c is not None and c < len(r)) else None

        found_ids = {}
        for r in data2:
            v = cell2(r, id_col)
            if v:
                found_ids[str(v).strip()] = r

        for pid in PAPERS_WITH_ID:
            record(f"Paper {pid} in Method Details", pid in found_ids,
                   f"Found: {set(found_ids)}")

        # CRITICAL: 4 correct IDs each with non-empty Key_Contribution + Dataset_Used
        details_ok = (id_col is not None and kc_col is not None and ds_col is not None
                      and set(found_ids) == PAPERS_WITH_ID)
        if details_ok:
            for pid, r in found_ids.items():
                kc = cell2(r, kc_col)
                ds = cell2(r, ds_col)
                if not (kc and str(kc).strip()) or not (ds and str(ds).strip()):
                    details_ok = False
        record("Method Details: 4 paper rows with correct Paper_IDs and non-empty Key_Contribution + Dataset_Used",
               details_ok, f"IDs={set(found_ids)}")

    # ── Summary sheet ───────────────────────────────────────────────────────
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        record("Summary: Total_Methods = 5", False, "no sheet")
        record("Summary: Methods_With_Papers = 4", False, "no sheet")
        record("Summary: Total_Tasks = 3", False, "no sheet")
        record("Summary: Top_Score = 95.0", False, "no sheet")
    else:
        record("Sheet 'Summary' exists", True)
        metrics = {}
        for row in sum_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        tm_key = next((k for k in metrics if "total" in k and "method" in k), None)
        record("Summary: Total_Methods = 5",
               tm_key is not None and num_close(metrics[tm_key], 5, tol=0),
               f"Got {metrics.get(tm_key)}")

        mwp_key = next((k for k in metrics if "with" in k and "paper" in k), None)
        record("Summary: Methods_With_Papers = 4",
               mwp_key is not None and num_close(metrics[mwp_key], 4, tol=0),
               f"Got {metrics.get(mwp_key)}")

        tt_key = next((k for k in metrics if "total" in k and "task" in k), None)
        record("Summary: Total_Tasks = 3",
               tt_key is not None and num_close(metrics[tt_key], 3, tol=0),
               f"Got {metrics.get(tt_key)}")

        ts_key = next((k for k in metrics if "top" in k and "score" in k), None)
        record("Summary: Top_Score = 95.0",
               ts_key is not None and num_close(metrics[ts_key], 95.0, tol=0.05),
               f"Got {metrics.get(ts_key)}")

    return True


def check_teamly():
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find the tracker page by title (English title preserved per task.md).
        cur.execute("""
            SELECT id, title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE '%%benchmark%%'
               OR title ILIKE '%%method%%tracker%%'
               OR title ILIKE '%%leaderboard%%'
               OR title ILIKE '%%трекер%%'
        """)
        pages = cur.fetchall()

        if not pages:
            cur.execute("SELECT COUNT(*) FROM teamly.pages")
            total = cur.fetchone()[0]
            record("Teamly tracker page exists", False,
                   f"Found {total} total pages but none matching")
            record("Teamly tracker page has substantive method/benchmark content", False,
                   "no matching page")
            cur.close()
            conn.close()
            return False

        record("Teamly tracker page exists", True)

        combined = " ".join((str(t) + " " + str(b)).lower() for _, t, b in pages)
        max_len = max(len(str(b)) for _, _, b in pages)
        record("Teamly tracker page has non-trivial body", max_len >= 100,
               f"Longest matching body is {max_len} chars")

        # CRITICAL: mention >=3 of the actual methods/papers + per-task scores.
        method_keywords = ["vit", "convnext", "llama", "diffusionxl", "diffusion"]
        method_hits = sum(1 for kw in ["vit", "convnext", "llama", "diffusion"]
                          if kw in combined)
        # accept RU + EN task framing
        task_keywords = [
            "image classification", "классификац",
            "text generation", "генерац текст", "генерация текст",
            "image generation", "генерац изображ", "генерация изображ",
            "imagenet", "mmlu", "fid", "benchmark", "бенчмарк", "лидерборд", "leaderboard",
        ]
        has_task = any(t in combined for t in task_keywords)
        score_keywords = ["91.2", "89.5", "95.0", "95", "92.3", "2.1"]
        has_score = any(s in combined for s in score_keywords)

        substantive = (method_hits >= 3) and has_task and has_score
        record("Teamly tracker page has substantive method/benchmark content",
               substantive,
               f"method hits={method_hits} (need>=3), task={has_task}, score={has_score}")

        cur.close()
        conn.close()
        return True
    except Exception as e:
        record("Teamly accessible", False, str(e))
        record("Teamly tracker page has substantive method/benchmark content", False, str(e))
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
