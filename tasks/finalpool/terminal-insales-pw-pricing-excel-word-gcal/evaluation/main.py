"""Evaluation for terminal-insales-pw-pricing-excel-word-gcal (InSales / russified)."""
import argparse
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def record_critical(name, passed, detail=""):
    record(name, passed, detail)
    if not passed:
        CRITICAL_FAILS.append(name)


def num_close(a, b, tol=5.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_float(c):
    try:
        return float(str(c).replace(',', '').replace('$', '').strip())
    except (TypeError, ValueError):
        return None


# Russified category -> expected deterministic gap. Derived from the LIVE wc.products
# catalog (82 products, flattened over the jsonb categories array) joined with the
# fixed 19-row competitor page. Gap_Pct = (our_avg - comp_avg) / comp_avg * 100, 1dp.
EXPECTED_GAPS = {
    "аудио": 35.9,
    "камеры": -22.9,
    "электроника": 206.7,
    "наушники": -16.8,
    "колонки": 43.4,
}

# Expected recommendation per gap rule (>10% above -> reduction; >10% below ->
# maintain advantage; else monitor). Each accepted in EN (frozen GT) or RU.
RECO_EXPECTED = {
    "аудио": "reduction",
    "камеры": "maintain",
    "электроника": "reduction",
    "наушники": "maintain",
    "колонки": "reduction",
}

RECO_PATTERNS = {
    # english (frozen groundtruth) OR russian equivalents
    "reduction": ["consider price reduction", "снижение цены", "снизить цен"],
    "maintain": ["maintain competitive advantage", "сохранять конкурентное",
                 "сохранение конкурентного", "конкурентное преимущество"],
    "monitor": ["monitor pricing", "мониторинг цен", "наблюдать за цен"],
}


def classify_reco(text):
    t = (text or "").lower()
    for kind, pats in RECO_PATTERNS.items():
        if any(p in t for p in pats):
            return kind
    return None


def expected_reco_kind(gap):
    if gap is None:
        return None
    if gap > 10:
        return "reduction"
    if gap < -10:
        return "maintain"
    return "monitor"


def wc_product_count():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM wc.products WHERE regular_price > 0")
        n = cur.fetchone()[0]
        cur.close()
        conn.close()
        return n
    except Exception:
        return None


def read_gap_rows(agent_workspace):
    """Return list of (cat_lower, our_avg, comp_avg, gap, reco) from Price_Gap_Analysis."""
    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Report.xlsx")
    if not os.path.isfile(fpath):
        return None
    wb = openpyxl.load_workbook(fpath, data_only=True)
    gap_sheet = None
    for name in wb.sheetnames:
        if "gap" in name.lower() or "analysis" in name.lower():
            gap_sheet = name
            break
    if not gap_sheet:
        wb.close()
        return []
    ws = wb[gap_sheet]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        cat = str(row[0]).strip().lower()
        our = to_float(row[1]) if len(row) > 1 else None
        comp = to_float(row[2]) if len(row) > 2 else None
        gap = to_float(row[3]) if len(row) > 3 else None
        reco = str(row[4]) if len(row) > 4 and row[4] is not None else ""
        out.append((cat, our, comp, gap, reco))
    wb.close()
    return out


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Report.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(fpath, data_only=True)
    expected_products = wc_product_count()

    # Sheet 1: Our_Products
    our_sheet = None
    for name in wb.sheetnames:
        if "our" in name.lower() and "product" in name.lower():
            our_sheet = name
            break
    if not our_sheet:
        record("Our_Products sheet exists", False, f"Sheets: {wb.sheetnames}")
        our_rows = 0
    else:
        record("Our_Products sheet exists", True)
        ws = wb[our_sheet]
        our_rows = len(list(ws.iter_rows(min_row=2, values_only=True)))
        if expected_products is not None:
            record(f"Our_Products has >= {expected_products} rows",
                   our_rows >= expected_products, f"Found {our_rows}")
        else:
            record("Our_Products has >= 30 rows", our_rows >= 30, f"Found {our_rows}")

    # Sheet 2: Competitor_Prices
    comp_sheet = None
    for name in wb.sheetnames:
        if "competitor" in name.lower():
            comp_sheet = name
            break
    if not comp_sheet:
        record("Competitor_Prices sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Competitor_Prices sheet exists", True)
        ws = wb[comp_sheet]
        comp_rows = len(list(ws.iter_rows(min_row=2, values_only=True)))
        record("Competitor_Prices has 19 rows", comp_rows == 19, f"Found {comp_rows}")

    # Sheet 3: Price_Gap_Analysis (non-critical structural)
    gap_sheet = None
    for name in wb.sheetnames:
        if "gap" in name.lower() or "analysis" in name.lower():
            gap_sheet = name
            break
    if not gap_sheet:
        record("Price_Gap_Analysis sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Price_Gap_Analysis sheet exists", True)
        ws = wb[gap_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Price_Gap_Analysis has 5 rows", len(rows) == 5, f"Found {len(rows)}")
        for row in rows:
            if row and row[0]:
                cat = str(row[0]).strip().lower()
                if cat in EXPECTED_GAPS:
                    gap_val = row[3] if len(row) > 3 else None
                    record(f"Gap for {cat} is correct",
                           num_close(gap_val, EXPECTED_GAPS[cat], tol=10.0),
                           f"Got {gap_val}, expected ~{EXPECTED_GAPS[cat]}")

    wb.close()
    return True


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    fpath = os.path.join(agent_workspace, "Pricing_Strategy_Report.docx")
    if not os.path.isfile(fpath):
        record("Word document exists", False, f"Not found: {fpath}")
        return False
    record("Word document exists", True)

    from docx import Document
    doc = Document(fpath)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    record("Document mentions competitive pricing",
           "competitive" in full_text or "pricing" in full_text
           or "конкурент" in full_text or "цен" in full_text)
    record("Document mentions Electronics (RU/EN)",
           "electronics" in full_text or "электроника" in full_text or "электрон" in full_text)
    record("Document mentions Headphones (RU/EN)",
           "headphones" in full_text or "наушник" in full_text)
    return True


def check_terminal_output(agent_workspace):
    print("\n=== Checking Terminal Output ===")
    fpath = os.path.join(agent_workspace, "price_analysis_output.txt")
    if not os.path.isfile(fpath):
        record("price_analysis_output.txt exists", False)
        return False
    record("price_analysis_output.txt exists", True)
    with open(fpath, encoding="utf-8", errors="replace") as f:
        content = f.read().lower()
    record("Output mentions price gap or comparison",
           "gap" in content or "comparison" in content or "%" in content
           or "разрыв" in content or "сравнен" in content)
    return True


def _gcal_events():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def is_price_review(summary):
    s = (summary or "").lower()
    return "price review" in s or "обзор цен" in s or "пересмотр цен" in s


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    try:
        events = _gcal_events()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        return False

    categories_found = set()
    for summary, description, _s, _e in events:
        if not is_price_review(summary):
            continue
        sl = (summary or "").lower()
        for cat in EXPECTED_GAPS:
            if cat in sl:
                categories_found.add(cat)

    record("Price review events for >= 5 categories", len(categories_found) >= 5,
           f"Found events for: {categories_found}")
    return len(categories_found) >= 5


# --------------------------------------------------------------------------
# CRITICAL semantic checks (any failure => sys.exit(1) before accuracy gate)
# --------------------------------------------------------------------------
def critical_checks(agent_workspace):
    print("\n=== CRITICAL CHECKS ===")
    rows = read_gap_rows(agent_workspace)

    if rows is None:
        record_critical("[CRIT] Price_Gap_Analysis present", False, "Excel missing")
        return
    by_cat = {c: (our, comp, gap, reco) for (c, our, comp, gap, reco) in rows}

    # C1: exactly the 5 russian-category gap rows, gaps match expected
    have = set(by_cat) & set(EXPECTED_GAPS)
    record_critical("[CRIT] 5 overlapping RU-category gap rows present",
                    len(have) == 5 and len(rows) == 5,
                    f"have={sorted(have)} total_rows={len(rows)}")
    for cat, exp in EXPECTED_GAPS.items():
        gap = by_cat.get(cat, (None, None, None, None))[2]
        record_critical(f"[CRIT] Gap_Pct for {cat} ~= {exp}",
                        num_close(gap, exp, tol=5.0), f"got {gap}")

    # C2: recommendation correct per rule
    for cat, kind in RECO_EXPECTED.items():
        reco = by_cat.get(cat, (None, None, None, ""))[3]
        got = classify_reco(reco)
        record_critical(f"[CRIT] Recommendation for {cat} == {kind}",
                        got == kind, f"got '{reco}' -> {got}")

    # C3: each gap internally consistent with own our_avg vs competitor_avg
    for cat in EXPECTED_GAPS:
        our, comp, gap, _ = by_cat.get(cat, (None, None, None, ""))
        if our is None or comp is None or comp == 0 or gap is None:
            record_critical(f"[CRIT] {cat} gap internally consistent", False,
                            f"our={our} comp={comp} gap={gap}")
            continue
        recomputed = round((our - comp) / comp * 100, 1)
        record_critical(f"[CRIT] {cat} gap matches own avgs",
                        num_close(recomputed, gap, tol=1.5),
                        f"row gap={gap}, (our-comp)/comp*100={recomputed}")

    # C4: catalogs actually pulled (not fabricated)
    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Report.xlsx")
    if os.path.isfile(fpath):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        comp_rows = our_rows = None
        for name in wb.sheetnames:
            ws = wb[name]
            n = len(list(ws.iter_rows(min_row=2, values_only=True)))
            if "competitor" in name.lower():
                comp_rows = n
            elif "our" in name.lower() and "product" in name.lower():
                our_rows = n
        wb.close()
        record_critical("[CRIT] Competitor_Prices has exactly 19 rows",
                        comp_rows == 19, f"got {comp_rows}")
        expected_products = wc_product_count()
        if expected_products is not None:
            record_critical(f"[CRIT] Our_Products >= {expected_products} (real catalog pull)",
                            our_rows is not None and our_rows >= expected_products,
                            f"got {our_rows}")

    # C5: calendar — one price-review event per RU category, 1h @ 14:00, distinct
    #     days from 2026-03-10 onward, description mentions the gap percentage.
    try:
        events = _gcal_events()
    except Exception as e:
        record_critical("[CRIT] Calendar accessible", False, str(e))
        return

    cat_events = {}
    for summary, description, start, end in events:
        if not is_price_review(summary):
            continue
        sl = (summary or "").lower()
        for cat in EXPECTED_GAPS:
            if cat in sl:
                cat_events.setdefault(cat, []).append((summary, description, start, end))

    record_critical("[CRIT] one price-review event per RU category (5)",
                    set(cat_events) == set(EXPECTED_GAPS),
                    f"covered={sorted(cat_events)}")

    days = set()
    for cat, evs in cat_events.items():
        summary, description, start, end = evs[0]
        # 1 hour duration
        dur_ok = False
        try:
            dur_ok = abs((end - start).total_seconds() - 3600) < 1
        except Exception:
            dur_ok = False
        # 14:00 start
        time_ok = False
        try:
            time_ok = start.hour == 14 and start.minute == 0
        except Exception:
            time_ok = False
        # on/after 2026-03-10
        date_ok = False
        try:
            date_ok = (start.year, start.month, start.day) >= (2026, 3, 10)
            days.add((start.year, start.month, start.day))
        except Exception:
            date_ok = False
        record_critical(f"[CRIT] {cat} meeting 1h @14:00 on/after 2026-03-10",
                        dur_ok and time_ok and date_ok,
                        f"start={start} end={end}")
        # description mentions the gap percentage for that category
        desc = (description or "")
        exp = EXPECTED_GAPS[cat]
        # accept the rounded gap with/without sign, 1-decimal
        gap_str_variants = {
            f"{exp}", f"{abs(exp)}", f"{exp:.1f}", f"{abs(exp):.1f}",
        }
        # also accept any number within 1.0 of expected found in the description
        nums = re.findall(r"-?\d+\.?\d*", desc)
        num_match = any(num_close(n, exp, tol=1.0) for n in nums)
        desc_ok = any(g in desc for g in gap_str_variants) or num_match
        record_critical(f"[CRIT] {cat} meeting description mentions its gap %",
                        desc_ok, f"desc={desc[:120]!r}")

    record_critical("[CRIT] price-review meetings on 5 distinct days",
                    len(days) >= 5, f"distinct days={sorted(days)}")


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    fpath = os.path.join(workspace, "Competitive_Pricing_Report.xlsx")
    if os.path.isfile(fpath):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        expected_keywords = {"our", "product", "competitor", "price", "gap", "analysis"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        record("No unexpected sheets in Excel", len(unexpected) == 0,
               f"Unexpected: {unexpected}")

        for sname in wb.sheetnames:
            ws = wb[sname]
            bad = False
            for row in ws.iter_rows(min_row=2, values_only=True):
                for c in row:
                    v = to_float(c)
                    if v is not None and v < 0 and "gap" not in sname.lower():
                        record(f"No negative prices in {sname}", False, f"Found {v}")
                        bad = True
                        break
                if bad:
                    break
            if bad:
                break
        else:
            record("No negative prices in product sheets", True)
        wb.close()

    # Calendar: no price review events before March 10, 2026 (EN or RU summary)
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE (summary ILIKE '%%price review%%'
                   OR summary ILIKE '%%обзор цен%%'
                   OR summary ILIKE '%%пересмотр цен%%')
              AND start_datetime < '2026-03-10'
        """)
        early_events = cur.fetchone()[0]
        record("No price review events before March 10", early_events == 0,
               f"Found {early_events} early events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_terminal_output(args.agent_workspace)
    check_calendar()
    critical_checks(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES ({len(CRITICAL_FAILS)}): {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical)")
        sys.exit(1)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"  Accuracy: {accuracy:.1f}%")
    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
