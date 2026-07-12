"""Evaluation for insales-category-revenue."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Category_Revenue.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Category_Revenue.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # ===== CRITICAL CHECKS (semantic core deliverables) =====
    # Any critical failure => immediate FAIL, before accumulating soft errors.
    critical_errors = []

    a_cat = load_sheet_rows(agent_wb, "Category Revenue")
    g_cat = load_sheet_rows(gt_wb, "Category Revenue")
    a_sum = load_sheet_rows(agent_wb, "Summary")
    g_sum = load_sheet_rows(gt_wb, "Summary")

    # CRITICAL (structural): both required sheets exist with header columns.
    if a_cat is None:
        critical_errors.append("CRITICAL: sheet 'Category Revenue' missing")
    elif not a_cat or not a_cat[0] or len(a_cat[0]) < 5:
        critical_errors.append("CRITICAL: 'Category Revenue' header missing required columns (Category..Est_Revenue)")
    if a_sum is None:
        critical_errors.append("CRITICAL: sheet 'Summary' missing")
    elif not a_sum or not a_sum[0] or len(a_sum[0]) < 2:
        critical_errors.append("CRITICAL: 'Summary' header missing required columns (Metric, Value)")

    if a_cat and g_cat and a_sum and g_sum:
        a_cat_lookup = {}
        for row in a_cat[1:]:
            if row and row[0] is not None:
                a_cat_lookup[str(row[0]).strip().lower()] = row
        a_sum_lookup = {}
        for row in a_sum[1:]:
            if row and row[0] is not None:
                a_sum_lookup[str(row[0]).strip().lower()] = row
        g_sum_lookup = {}
        for row in g_sum[1:]:
            if row and row[0] is not None:
                g_sum_lookup[str(row[0]).strip().lower()] = row

        # CRITICAL: every GT category present with Est_Revenue within tol (primary deliverable).
        for g_row in g_cat[1:]:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_cat_lookup.get(key)
            if a_row is None:
                critical_errors.append(f"CRITICAL: missing category row '{g_row[0]}'")
            elif len(a_row) > 4 and len(g_row) > 4 and not num_close(a_row[4], g_row[4], 100.0):
                critical_errors.append(
                    f"CRITICAL: category '{g_row[0]}' Est_Revenue {a_row[4]} vs {g_row[4]} (tol=100.0)")

        # CRITICAL: Top_Category matches GT (highest-revenue category, exact string).
        g_top = g_sum_lookup.get("top_category")
        a_top = a_sum_lookup.get("top_category")
        if g_top is not None and len(g_top) > 1:
            if a_top is None or len(a_top) < 2:
                critical_errors.append("CRITICAL: Summary 'Top_Category' missing")
            elif not str_match(a_top[1], g_top[1]):
                critical_errors.append(
                    f"CRITICAL: Top_Category '{a_top[1] if len(a_top)>1 else None}' vs '{g_top[1]}'")

        # CRITICAL: Total_Est_Revenue within tol (end-to-end revenue computation).
        g_tot = g_sum_lookup.get("total_est_revenue")
        a_tot = a_sum_lookup.get("total_est_revenue")
        if g_tot is not None and len(g_tot) > 1:
            if a_tot is None or len(a_tot) < 2:
                critical_errors.append("CRITICAL: Summary 'Total_Est_Revenue' missing")
            elif not num_close(a_tot[1], g_tot[1], 100.0):
                critical_errors.append(
                    f"CRITICAL: Total_Est_Revenue {a_tot[1] if len(a_tot)>1 else None} vs {g_tot[1]} (tol=100.0)")

    if critical_errors:
        print("=== CRITICAL CHECK FAILED ===")
        for e in critical_errors:
            print(f"  {e}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    print("  Critical checks PASSED")


    # Check sheet: Category Revenue
    print(f"  Checking Category Revenue...")
    a_rows = load_sheet_rows(agent_wb, "Category Revenue")
    g_rows = load_sheet_rows(gt_wb, "Category Revenue")
    if a_rows is None:
        all_errors.append("Sheet 'Category Revenue' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Category Revenue' not found in groundtruth")
    else:
        sheet_name = "Category Revenue"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 2):
                    errors.append(f"{key}.Products: {a_row[1]} vs {g_row[1]} (tol=2)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1.0):
                    errors.append(f"{key}.Avg_Price: {a_row[2]} vs {g_row[2]} (tol=1.0)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 10):
                    errors.append(f"{key}.Units_Sold: {a_row[3]} vs {g_row[3]} (tol=10)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 100.0):
                    errors.append(f"{key}.Est_Revenue: {a_row[4]} vs {g_row[4]} (tol=100.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 100.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=100.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
