"""Evaluation script for fetch-arxiv-citation-network-excel-notion (teamly swap).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Seeded relevant papers (English identifiers preserved). The noise paper
# 9999.99999 (quant-ph) must be EXCLUDED.
EXPECTED_PAPERS = ["2301.01234", "2302.05678", "2303.09012", "2304.03456"]
NOISE_PAPER = "9999.99999"

CRITICAL_CHECKS = {
    "Data_Analysis contains all 4 seeded papers and excludes the noise paper",
    "Citation values match groundtruth per paper",
    "Metrics: Total_Papers=4, Avg_Citations=300, Top_Area=LLMs",
    "External benchmark (LLMs highest growth) reflected as Top_Area / priority",
    "Teamly 'Arxiv Citation Dashboard' page exists with non-empty analysis content",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_sheet(wb, name):
    target = name.strip().lower().replace(" ", "_")
    for n in wb.sheetnames:
        if n.strip().lower().replace(" ", "_") == target:
            return wb[n]
    return None


def _header_map(ws):
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def check_data_analysis(wb, gt_wb):
    """CRITICAL: all 4 seeded papers present, noise excluded, citations match gt."""
    ws = get_sheet(wb, "Data_Analysis")
    if ws is None:
        check("Data_Analysis contains all 4 seeded papers and excludes the noise paper", False, "no sheet")
        check("Citation values match groundtruth per paper", False, "no sheet")
        return
    headers, hmap = _header_map(ws)
    # Structural (non-critical) column checks.
    for col in ['Paper_ID', 'Title', 'Area', 'Citations', 'Relevance_Score']:
        check(f"Data_Analysis has {col} column", col.lower() in hmap, f"headers: {headers[:8]}")

    id_i = hmap.get("paper_id")
    cit_i = hmap.get("citations")
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    check("Data_Analysis has >= 4 rows", len(rows) >= 4, f"got {len(rows)}")

    # Build pid -> citations map from agent sheet.
    agent = {}
    if id_i is not None:
        for r in rows:
            rid = str(r[id_i]).strip() if id_i < len(r) and r[id_i] is not None else ""
            cit = safe_float(r[cit_i]) if cit_i is not None and cit_i < len(r) else None
            for pid in EXPECTED_PAPERS + [NOISE_PAPER]:
                if pid in rid:
                    agent[pid] = cit

    all_present = all(p in agent for p in EXPECTED_PAPERS)
    noise_excluded = NOISE_PAPER not in agent
    check("Data_Analysis contains all 4 seeded papers and excludes the noise paper",
          all_present and noise_excluded,
          f"present={[p for p in EXPECTED_PAPERS if p in agent]} noise_in={NOISE_PAPER in agent}")

    # Citations match groundtruth (read from gt_wb, not ignored).
    gt_cit = {}
    if gt_wb is not None:
        gws = get_sheet(gt_wb, "Data_Analysis")
        if gws is not None:
            _, ghmap = _header_map(gws)
            gid_i, gcit_i = ghmap.get("paper_id"), ghmap.get("citations")
            for r in gws.iter_rows(min_row=2, values_only=True):
                if gid_i is None or gcit_i is None:
                    break
                rid = str(r[gid_i]).strip() if gid_i < len(r) and r[gid_i] is not None else ""
                for pid in EXPECTED_PAPERS:
                    if pid in rid:
                        gt_cit[pid] = safe_float(r[gcit_i])
    if gt_cit:
        match = all(agent.get(p) is not None and gt_cit.get(p) is not None
                    and abs(agent[p] - gt_cit[p]) < 0.5 for p in EXPECTED_PAPERS)
        check("Citation values match groundtruth per paper", match,
              f"agent={agent} gt={gt_cit}")
    else:
        check("Citation values match groundtruth per paper", False, "no gt citations")


def check_metrics(wb):
    """CRITICAL: Total_Papers=4, Avg_Citations=300, Top_Area=LLMs."""
    ws = get_sheet(wb, "Metrics")
    if ws is None:
        check("Metrics: Total_Papers=4, Avg_Citations=300, Top_Area=LLMs", False, "no sheet")
        check("External benchmark (LLMs highest growth) reflected as Top_Area / priority", False, "no sheet")
        return
    headers, hmap = _header_map(ws)
    for col in ['Metric', 'Value']:
        check(f"Metrics has {col} column", col.lower() in hmap, f"headers: {headers[:8]}")
    m_i, v_i = hmap.get("metric"), hmap.get("value")
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    check("Metrics has >= 4 rows", len(rows) >= 4, f"got {len(rows)}")

    metrics = {}
    if m_i is not None and v_i is not None:
        for r in rows:
            key = str(r[m_i]).strip().lower() if m_i < len(r) and r[m_i] is not None else ""
            val = r[v_i] if v_i < len(r) else None
            metrics[key] = val

    total_ok = safe_float(metrics.get("total_papers")) == 4
    avg_ok = safe_float(metrics.get("avg_citations")) is not None and abs(safe_float(metrics.get("avg_citations")) - 300) < 0.5
    top_ok = "llm" in str(metrics.get("top_area", "")).lower()
    check("Metrics: Total_Papers=4, Avg_Citations=300, Top_Area=LLMs",
          total_ok and avg_ok and top_ok,
          f"total={metrics.get('total_papers')} avg={metrics.get('avg_citations')} top={metrics.get('top_area')}")

    # External benchmark: LLMs has highest growth_pct (45) in data.json, so it
    # should be the Top_Area (and/or surface as priority-1 in Recommendations).
    bench_ok = top_ok
    return metrics, bench_ok


def check_recommendations(wb, bench_ok):
    """CRITICAL benchmark reflection check (also looks at Recommendations)."""
    ws = get_sheet(wb, "Recommendations")
    rec_top = False
    if ws is not None:
        headers, hmap = _header_map(ws)
        for col in ['Priority', 'Action', 'Area']:
            check(f"Recommendations has {col} column", col.lower() in hmap, f"headers: {headers[:8]}")
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        check("Recommendations has >= 2 rows", len(rows) >= 2, f"got {len(rows)}")
        p_i, a_i = hmap.get("priority"), hmap.get("area")
        # Priority-1 row references LLMs.
        for r in rows:
            prio = safe_float(r[p_i]) if p_i is not None and p_i < len(r) else None
            area = str(r[a_i]).lower() if a_i is not None and a_i < len(r) and r[a_i] is not None else ""
            if (prio == 1 or "1" in str(r[p_i] if p_i is not None and p_i < len(r) else "")) and "llm" in area:
                rec_top = True
                break
    else:
        check("Recommendations has >= 2 rows", False, "no sheet")

    check("External benchmark (LLMs highest growth) reflected as Top_Area / priority",
          bool(bench_ok) or rec_top,
          f"top_area_llms={bench_ok} rec_priority1_llms={rec_top}")


def check_excel(agent_workspace, groundtruth_workspace):
    excel_path = os.path.join(agent_workspace, "Citation_Network_Report.xlsx")
    check("Citation_Network_Report.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        check("Data_Analysis contains all 4 seeded papers and excludes the noise paper", False, "no excel")
        check("Citation values match groundtruth per paper", False, "no excel")
        check("Metrics: Total_Papers=4, Avg_Citations=300, Top_Area=LLMs", False, "no excel")
        check("External benchmark (LLMs highest growth) reflected as Top_Area / priority", False, "no excel")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    gt_path = os.path.join(groundtruth_workspace, "Citation_Network_Report.xlsx")
    gt_wb = openpyxl.load_workbook(gt_path, data_only=True) if os.path.exists(gt_path) else None

    check("Data_Analysis sheet exists", get_sheet(wb, "Data_Analysis") is not None)
    check("Metrics sheet exists", get_sheet(wb, "Metrics") is not None)
    check("Recommendations sheet exists", get_sheet(wb, "Recommendations") is not None)

    check_data_analysis(wb, gt_wb)
    metrics_res = check_metrics(wb)
    bench_ok = metrics_res[1] if isinstance(metrics_res, tuple) else False
    check_recommendations(wb, bench_ok)

    check("arxiv_citation_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "arxiv_citation_processor.py")))


def check_teamly():
    """CRITICAL: 'Arxiv Citation Dashboard' page exists (non-seed, archived=false)
    with non-empty analysis content. Runs independently of Excel."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Seed pages have id <= 3; noise page is 'старые заметки проекта'.
        cur.execute("SELECT id, COALESCE(title,''), COALESCE(body,'') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Arxiv Citation Dashboard' page exists with non-empty analysis content", False, str(e))
        return

    dash = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "старые заметки" in tl:
            continue
        if "dashboard" in tl or ("arxiv" in tl and "citation" in tl) or "дашборд" in tl:
            dash = (pid, title, body)
            break

    if dash is None:
        check("Teamly 'Arxiv Citation Dashboard' page exists with non-empty analysis content",
              False, f"new pages: {[(p[0], p[1]) for p in pages]}")
        return

    text = ((dash[1] or "") + " " + (dash[2] or "")).lower()
    # Non-empty body that references the analysis: an area code or a metric value.
    refs = any(k in text for k in ["llm", "cv", "rl", "fl", "цитир", "citation",
                                   "300", "450", "area", "област", "метрик", "metric"])
    non_empty = len((dash[2] or "").strip()) >= 20
    check("Teamly 'Arxiv Citation Dashboard' page exists with non-empty analysis content",
          non_empty and refs, f"body_len={len((dash[2] or '').strip())} refs={refs}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    check_excel(agent_workspace, groundtruth_workspace)
    # Teamly check runs independently of Excel existence.
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
