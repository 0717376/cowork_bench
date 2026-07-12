"""
Evaluation script for gsheet-insales-inventory task.

Check 1: Inventory_Restock_Report.xlsx
  - "Category Inventory" sheet: 8 rows, correct values per category
  - "Restock Alerts" sheet: products with stock < 10
  - "Cost Trend" sheet: 5 rows of LKOH.ME closing prices
Check 2: Google Sheet "Inventory Dashboard" exists with category data
Check 3: Email to warehouse@company.com with low stock alert

CRITICAL semantic checks (any fail => sys.exit(1) before the accuracy gate):
  C1: Category Inventory has all 8 (RU) categories with correct
      Product_Count / Total_Stock / Avg_Price (tight tolerance).
  C2: Restock Alerts lists the EXACT set of SKUs with stock < 10,
      with correct Current_Stock per SKU.
  C3: gsheet "Inventory Dashboard" contains the same 8 category rows.
  C4: Email body lists most of the low-stock SKUs/names.
  C5: Cost Trend has the 5 most-recent LKOH.ME rows (dates + closes).
"""

import argparse
import os
import sys
import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# MOEX ticker used as the raw-material / cost-trend proxy (Лукойл).
COST_TREND_SYMBOL = "LKOH.ME"

# Category names as russified centrally by db/zzz_wc_after_init.sql.
# The agent reads these live from wc.products and writes them verbatim,
# so eval compares against the RU names (accepting EN as a fallback).
RU_TO_EN_CATEGORY = {
    "Аудио": "Audio",
    "Камеры": "Cameras",
    "Электроника": "Electronics",
    "Наушники": "Headphones",
    "Бытовая техника": "Home Appliances",
    "Колонки": "Speakers",
    "ТВ и домашний кинотеатр": "TV & Home Theater",
    "Часы": "Watches",
}
# Accept either the RU name or its EN original for any given category cell.
CATEGORY_ALIASES = {ru.lower() for ru in RU_TO_EN_CATEGORY}
CATEGORY_ALIASES |= {en.lower() for en in RU_TO_EN_CATEGORY.values()}


# The store has exactly 8 canonical product categories. They live in
# wc.product_categories (russified centrally by db/zzz_wc_after_init.sql), and
# the per-product wc.products.categories jsonb was ALSO russified — it carries
# the RU names (Аудио, Камеры, ..., Часы), NOT the English originals. The
# distinct names actually present in wc.products.categories are exactly the 8
# RU names. Restricting the expected grouping to the known aliases (RU names
# plus their EN originals, all lowercased) keeps the expected inventory at
# exactly the 8 canonical categories — so a correct 8-row agent sheet matches —
# even if a product's categories array ever picks up a stray, non-canonical
# member. CATEGORY_ALIASES (defined above) holds both the RU and EN spellings.
CANONICAL_PRODUCT_CATEGORIES_EN = (
    "Audio",
    "Cameras",
    "Electronics",
    "Headphones",
    "Home Appliances",
    "Speakers",
    "TV & Home Theater",
    "Watches",
)


def get_expected_category_inventory():
    """Query InSales (wc.*) data for expected category inventory.

    Groups wc.products by the category names carried in the categories jsonb,
    restricted to the known canonical categories. The live jsonb carries the
    russified RU names, so we match case-insensitively against CATEGORY_ALIASES
    (which holds both the RU names and their EN originals) and report whatever
    name is actually stored (the RU name). _category_matches() accepts the RU
    name or its EN original for any cell, so a correct agent sheet still
    matches regardless of which spelling it used.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT
                cat,
                COUNT(*) AS product_count,
                SUM(stock_quantity) AS total_stock,
                ROUND(AVG(regular_price::float)::numeric, 2) AS avg_price
            FROM (
                SELECT
                    jsonb_array_elements(categories)->>'name' AS cat,
                    stock_quantity,
                    regular_price
                FROM wc.products
            ) AS exploded
            WHERE LOWER(cat) = ANY(%s)
            GROUP BY cat
            ORDER BY cat
            """,
            (sorted(CATEGORY_ALIASES),),
        )
        rows = cur.fetchall()
        return rows
    finally:
        cur.close()
        conn.close()


def get_expected_low_stock():
    """Query InSales (wc.*) for products with stock < 10."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT
                name,
                sku,
                stock_quantity,
                categories->0->>'name' AS category,
                regular_price
            FROM wc.products
            WHERE stock_quantity < 10
            ORDER BY stock_quantity, name
        """)
        rows = cur.fetchall()
        return rows
    finally:
        cur.close()
        conn.close()


def get_expected_cost_trend():
    """Query MOEX Finance for the latest 5 LKOH.ME closing prices."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute(
            """
            SELECT date, close
            FROM moex.stock_prices
            WHERE symbol = %s
            ORDER BY date DESC
            LIMIT 5
        """,
            (COST_TREND_SYMBOL,),
        )
        rows = cur.fetchall()
        return rows
    finally:
        cur.close()
        conn.close()


# EN canonical name -> RU name (inverse of RU_TO_EN_CATEGORY), so a cell may be
# matched against the expected category regardless of which name the agent used.
EN_TO_RU_CATEGORY = {en.lower(): ru for ru, en in RU_TO_EN_CATEGORY.items()}


def _category_matches(row_cat, exp_cat):
    """True if the cell value names the expected category (RU or EN alias).

    exp_cat may be supplied as either the RU name (RU_TO_EN_CATEGORY key) or its
    EN original; the cell may likewise be RU or EN. We accept any pairing.
    """
    rc = str(row_cat).strip().lower()
    ec = str(exp_cat).strip().lower()
    if rc == ec:
        return True
    # exp_cat is RU -> accept its EN original.
    en = RU_TO_EN_CATEGORY.get(exp_cat)
    if en and rc == en.lower():
        return True
    # exp_cat is EN -> accept its RU equivalent.
    ru = EN_TO_RU_CATEGORY.get(ec)
    if ru and rc == ru.lower():
        return True
    return False


# RU+EN header tokens accepted as a match for each expected (English) column.
RU_HEADER_TOKENS = {
    "Category": ["category", "категория", "категории"],
    "Product_Count": ["product_count", "product count", "productcount",
                      "количество товаров", "кол-во товаров", "количество"],
    "Total_Stock": ["total_stock", "total stock", "totalstock",
                    "суммарный остаток", "общий остаток", "остаток"],
    "Avg_Price": ["avg_price", "avg price", "avgprice",
                  "средняя цена", "ср. цена", "средняя розничная цена"],
    "Product_Name": ["product_name", "product name", "productname",
                     "название товара", "наименование", "товар"],
    "SKU": ["sku", "артикул"],
    "Current_Stock": ["current_stock", "current stock", "currentstock",
                      "текущий остаток", "остаток"],
    "Regular_Price": ["regular_price", "regular price", "regularprice",
                      "розничная цена", "цена"],
    "Date": ["date", "дата"],
    "Close_Price": ["close_price", "close price", "closeprice",
                    "цена закрытия", "закрытие"],
}


def build_col_map(headers, expected_cols, errors, sheet_label):
    """Map each expected English column to a header index, RU+EN tolerant."""
    col_map = {}
    norm_headers = [str(h).strip().lower() if h is not None else "" for h in headers]
    for ec in expected_cols:
        # 1) exact normalized match against EN variant
        for idx, h in enumerate(headers):
            if h and ec.lower().replace("_", "") == str(h).lower().replace("_", "").replace(" ", ""):
                col_map[ec] = idx
                break
        # 2) substring match against EN variant
        if ec not in col_map:
            for idx, h in enumerate(headers):
                if h and ec.lower().replace("_", " ") in str(h).lower().replace("_", " "):
                    col_map[ec] = idx
                    break
        # 3) RU+EN token match
        if ec not in col_map:
            for token in RU_HEADER_TOKENS.get(ec, []):
                for idx, nh in enumerate(norm_headers):
                    if nh and (nh == token or token in nh):
                        col_map[ec] = idx
                        break
                if ec in col_map:
                    break
        if ec not in col_map:
            errors.append(f"Column '{ec}' not found in {sheet_label} headers: {headers}")
    return col_map


def check_excel(agent_workspace):
    """Check the Excel report against expected data."""
    print("[eval] Checking Excel file...")
    errors = []

    excel_path = os.path.join(agent_workspace, "Inventory_Restock_Report.xlsx")
    if not os.path.exists(excel_path):
        errors.append(f"Excel file not found: {excel_path}")
        return False, errors

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        errors.append(f"Cannot open Excel file: {e}")
        return False, errors

    # --- Check Category Inventory sheet ---
    if "Category Inventory" not in wb.sheetnames:
        errors.append(f"Sheet 'Category Inventory' not found. Available: {wb.sheetnames}")
    else:
        ws = wb["Category Inventory"]
        headers = [cell.value for cell in ws[1]]

        expected_cols = ["Category", "Product_Count", "Total_Stock", "Avg_Price"]
        col_map = build_col_map(headers, expected_cols, errors, "Category Inventory")

        if not errors:
            data_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_map["Category"]] is not None:
                    data_rows.append(row)

            if len(data_rows) != 8:
                errors.append(f"Category Inventory: expected 8 rows, found {len(data_rows)}")

            expected_cats = get_expected_category_inventory()
            for exp_cat, exp_count, exp_stock, exp_price in expected_cats:
                found = False
                for row in data_rows:
                    row_cat = str(row[col_map["Category"]]).strip()
                    if _category_matches(row_cat, exp_cat):
                        found = True
                        # Check Product_Count (tolerance 1)
                        pc = row[col_map["Product_Count"]]
                        if pc is not None and abs(int(pc) - exp_count) > 1:
                            errors.append(
                                f"Category '{exp_cat}': Product_Count {pc} != expected {exp_count}"
                            )
                        # Check Total_Stock (tolerance 5)
                        ts = row[col_map["Total_Stock"]]
                        if ts is not None and abs(int(ts) - exp_stock) > 5:
                            errors.append(
                                f"Category '{exp_cat}': Total_Stock {ts} != expected {exp_stock}"
                            )
                        # Check Avg_Price (tolerance 1.0)
                        ap = row[col_map["Avg_Price"]]
                        if ap is not None and abs(float(ap) - float(exp_price)) > 1.0:
                            errors.append(
                                f"Category '{exp_cat}': Avg_Price {ap} != expected {exp_price}"
                            )
                        break
                if not found:
                    errors.append(f"Category '{exp_cat}' not found in Category Inventory sheet")

    # --- Check Restock Alerts sheet ---
    if "Restock Alerts" not in wb.sheetnames:
        errors.append(f"Sheet 'Restock Alerts' not found. Available: {wb.sheetnames}")
    else:
        ws = wb["Restock Alerts"]
        headers = [cell.value for cell in ws[1]]

        expected_cols_ra = ["Product_Name", "SKU", "Current_Stock", "Category", "Regular_Price"]
        col_map_ra = build_col_map(headers, expected_cols_ra, errors, "Restock Alerts")

        if all(ec in col_map_ra for ec in expected_cols_ra):
            data_rows_ra = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_map_ra["Product_Name"]] is not None:
                    data_rows_ra.append(row)

            expected_low = get_expected_low_stock()
            expected_count = len(expected_low)

            if abs(len(data_rows_ra) - expected_count) > 2:
                errors.append(
                    f"Restock Alerts: expected ~{expected_count} rows, found {len(data_rows_ra)}"
                )

            # Verify a sample of products exist
            sample_skus = [row[1] for row in expected_low[:5]]
            found_skus = set()
            for row in data_rows_ra:
                sku_val = str(row[col_map_ra["SKU"]]).strip() if row[col_map_ra["SKU"]] else ""
                found_skus.add(sku_val)

            for sku in sample_skus:
                if sku not in found_skus:
                    errors.append(f"Restock Alerts: expected SKU '{sku}' not found")

    # --- Check Cost Trend sheet (LKOH.ME closing prices) ---
    if "Cost Trend" not in wb.sheetnames:
        errors.append(f"Sheet 'Cost Trend' not found. Available: {wb.sheetnames}")
    else:
        ws = wb["Cost Trend"]
        headers = [cell.value for cell in ws[1]]

        col_map_gt = build_col_map(headers, ["Date", "Close_Price"], errors, "Cost Trend")

        if all(ec in col_map_gt for ec in ["Date", "Close_Price"]):
            data_rows_gt = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[col_map_gt["Date"]] is not None:
                    data_rows_gt.append(row)

            if len(data_rows_gt) != 5:
                errors.append(f"Cost Trend: expected 5 rows, found {len(data_rows_gt)}")

            # Check first row close price matches expected
            expected_cost = get_expected_cost_trend()
            if data_rows_gt and expected_cost:
                actual_close = float(data_rows_gt[0][col_map_gt["Close_Price"]])
                expected_close = float(expected_cost[0][1])
                if abs(actual_close - expected_close) > 5.0:
                    errors.append(
                        f"Cost Trend: first row Close_Price {actual_close} != expected {expected_close}"
                    )

    if errors:
        return False, errors

    print("  Excel check passed.")
    return True, []


def check_gsheet():
    """Check that an Inventory Dashboard spreadsheet exists in gsheet schema."""
    print("[eval] Checking Google Sheet...")
    errors = []
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        # Find spreadsheet with "inventory" in title (case-insensitive)
        cur.execute(
            "SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%inventory%'"
        )
        spreadsheets = cur.fetchall()
        if not spreadsheets:
            errors.append("No spreadsheet found with 'inventory' in the title")
            return False, errors

        ss_id = spreadsheets[0][0]
        ss_title = spreadsheets[0][1]
        print(f"  Found spreadsheet: '{ss_title}' (id={ss_id})")

        # Check that cells exist
        cur.execute(
            "SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (ss_id,)
        )
        cell_count = cur.fetchone()[0]
        if cell_count < 9:  # At least header + 8 data rows
            errors.append(
                f"Spreadsheet has only {cell_count} cells, expected at least 9 rows of data"
            )
            return False, errors

        # Check that category data is present by looking for known category
        # names (RU names russified centrally by db/zzz_wc_after_init.sql, with
        # the EN originals accepted as a fallback).
        alias_list = sorted(CATEGORY_ALIASES)
        placeholders = ", ".join(["%s"] * len(alias_list))
        cur.execute(
            f"""SELECT value FROM gsheet.cells
               WHERE spreadsheet_id = %s AND LOWER(value) IN ({placeholders})""",
            (ss_id, *alias_list),
        )
        found_cats = cur.fetchall()
        if len(found_cats) < 5:
            errors.append(
                f"Spreadsheet has only {len(found_cats)} recognizable category names, expected at least 5"
            )
            return False, errors

        print(f"  Found {len(found_cats)} category entries in spreadsheet.")

    finally:
        cur.close()
        conn.close()

    if errors:
        return False, errors

    print("  Google Sheet check passed.")
    return True, []


def check_email():
    """Check that a low stock alert email was sent to warehouse@company.com."""
    print("[eval] Checking email...")
    errors = []
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        # Look for email with relevant subject
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE (LOWER(subject) LIKE '%low stock%' OR LOWER(subject) LIKE '%alert%' OR LOWER(subject) LIKE '%restock%')
        """)
        emails = cur.fetchall()

        if not emails:
            errors.append("No email found with 'low stock', 'alert', or 'restock' in subject")
            return False, errors

        # Check at least one email goes to warehouse@company.com
        found_target = False
        matched_email = None
        for subject, to_addr, body_text in emails:
            to_str = str(to_addr).lower() if to_addr else ""
            if "warehouse@company.com" in to_str:
                found_target = True
                matched_email = (subject, body_text)
                break

        if not found_target:
            errors.append(
                "No email to warehouse@company.com found among alert emails"
            )
            return False, errors

        subject, body = matched_email
        print(f"  Found email: subject='{subject}'")

        # Check body contains some product info
        if body:
            body_lower = body.lower()
            # Check that at least a few known low-stock SKUs or product names appear
            expected_low = get_expected_low_stock()
            matches = 0
            for name, sku, stock, cat, price in expected_low[:10]:
                if sku.lower() in body_lower or name[:30].lower() in body_lower:
                    matches += 1
            if matches < 3:
                errors.append(
                    f"Email body mentions only {matches} of the first 10 low-stock products (expected >= 3)"
                )
                return False, errors
        else:
            errors.append("Email body is empty")
            return False, errors

    finally:
        cur.close()
        conn.close()

    if errors:
        return False, errors

    print("  Email check passed.")
    return True, []


# ---------------------------------------------------------------------------
# CRITICAL semantic checks. Any failure => sys.exit(1) BEFORE the accuracy
# gate. These verify SUBSTANCE (correct values from the live source), not just
# structure. Tolerances are intentionally tighter than the soft checks above.
# ---------------------------------------------------------------------------

def _load_excel(agent_workspace):
    path = os.path.join(agent_workspace, "Inventory_Restock_Report.xlsx")
    if not os.path.exists(path):
        return None, None
    try:
        return openpyxl.load_workbook(path), path
    except Exception:
        return None, None


def critical_category_inventory(agent_workspace):
    """C1: all 8 (RU) categories present with correct Count/Stock/Avg (tight)."""
    errors = []
    wb, _ = _load_excel(agent_workspace)
    if wb is None or "Category Inventory" not in wb.sheetnames:
        return False, ["C1: Inventory_Restock_Report.xlsx / 'Category Inventory' missing"]
    ws = wb["Category Inventory"]
    headers = [c.value for c in ws[1]]
    cmap = build_col_map(headers, ["Category", "Product_Count", "Total_Stock", "Avg_Price"], errors, "Category Inventory")
    if errors:
        return False, errors
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[cmap["Category"]] is not None]
    expected = get_expected_category_inventory()
    if len(expected) != 8:
        errors.append(f"C1: source has {len(expected)} categories, expected 8")
    for exp_cat, exp_count, exp_stock, exp_price in expected:
        match = None
        for r in rows:
            if _category_matches(str(r[cmap["Category"]]).strip(), exp_cat):
                match = r
                break
        if match is None:
            errors.append(f"C1: category '{exp_cat}' missing from sheet")
            continue
        pc = match[cmap["Product_Count"]]
        if pc is None or int(pc) != int(exp_count):
            errors.append(f"C1: '{exp_cat}' Product_Count {pc} != {exp_count}")
        ts = match[cmap["Total_Stock"]]
        if ts is None or int(ts) != int(exp_stock):
            errors.append(f"C1: '{exp_cat}' Total_Stock {ts} != {exp_stock}")
        ap = match[cmap["Avg_Price"]]
        if ap is None or abs(float(ap) - float(exp_price)) > 0.1:
            errors.append(f"C1: '{exp_cat}' Avg_Price {ap} != {exp_price}")
    return (not errors), errors


def critical_restock_alerts(agent_workspace):
    """C2: EXACT set of SKUs with stock < 10, correct Current_Stock each."""
    errors = []
    wb, _ = _load_excel(agent_workspace)
    if wb is None or "Restock Alerts" not in wb.sheetnames:
        return False, ["C2: 'Restock Alerts' sheet missing"]
    ws = wb["Restock Alerts"]
    headers = [c.value for c in ws[1]]
    cmap = build_col_map(headers, ["Product_Name", "SKU", "Current_Stock"], errors, "Restock Alerts")
    if errors:
        return False, errors
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[cmap["Product_Name"]] is not None]
    sheet_stock = {}
    for r in rows:
        sku = str(r[cmap["SKU"]]).strip() if r[cmap["SKU"]] is not None else ""
        if sku:
            sheet_stock[sku] = r[cmap["Current_Stock"]]
    expected = get_expected_low_stock()
    expected_skus = {str(row[1]).strip() for row in expected}
    sheet_skus = set(sheet_stock)
    missing = expected_skus - sheet_skus
    extra = sheet_skus - expected_skus
    if missing:
        errors.append(f"C2: missing low-stock SKUs: {sorted(missing)}")
    if extra:
        errors.append(f"C2: unexpected SKUs in Restock Alerts: {sorted(extra)}")
    for name, sku, stock, cat, price in expected:
        s = str(sku).strip()
        if s in sheet_stock:
            v = sheet_stock[s]
            if v is None or int(v) != int(stock):
                errors.append(f"C2: SKU '{s}' Current_Stock {v} != {stock}")
    return (not errors), errors


def critical_cost_trend(agent_workspace):
    """C5: exactly the 5 most-recent LKOH.ME rows (all dates + closes)."""
    errors = []
    wb, _ = _load_excel(agent_workspace)
    if wb is None or "Cost Trend" not in wb.sheetnames:
        return False, ["C5: 'Cost Trend' sheet missing"]
    ws = wb["Cost Trend"]
    headers = [c.value for c in ws[1]]
    cmap = build_col_map(headers, ["Date", "Close_Price"], errors, "Cost Trend")
    if errors:
        return False, errors
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[cmap["Date"]] is not None]
    expected = get_expected_cost_trend()
    if len(rows) != 5:
        errors.append(f"C5: expected 5 rows, found {len(rows)}")
    if len(expected) < 5:
        errors.append(f"C5: source has {len(expected)} LKOH.ME rows, expected >=5")
        return False, errors
    # Compare closes as a multiset, order-independent on value but expecting the
    # same 5 most-recent closing prices.
    exp_closes = sorted(float(e[1]) for e in expected)
    act_closes = sorted(float(r[cmap["Close_Price"]]) for r in rows
                        if r[cmap["Close_Price"]] is not None)
    if len(act_closes) != len(exp_closes):
        errors.append(f"C5: {len(act_closes)} close values vs {len(exp_closes)} expected")
    else:
        for a, e in zip(act_closes, exp_closes):
            if abs(a - e) > 1.0:
                errors.append(f"C5: close {a} not among expected {exp_closes}")
                break
    return (not errors), errors


def critical_gsheet():
    """C3: Inventory Dashboard contains >=8 of the expected category names."""
    errors = []
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%inventory%'")
        rows = cur.fetchall()
        if not rows:
            return False, ["C3: no 'inventory' spreadsheet found"]
        ss_id = rows[0][0]
        alias_list = sorted(CATEGORY_ALIASES)
        placeholders = ", ".join(["%s"] * len(alias_list))
        cur.execute(
            f"SELECT DISTINCT LOWER(value) FROM gsheet.cells "
            f"WHERE spreadsheet_id = %s AND LOWER(value) IN ({placeholders})",
            (ss_id, *alias_list),
        )
        found = {r[0] for r in cur.fetchall()}
        # Each of the 8 source categories must appear (RU or EN alias).
        # get_expected_category_inventory() reports EN canonical names; accept
        # either the EN name or its russified equivalent in the sheet cell.
        expected = get_expected_category_inventory()
        for exp_cat, *_ in expected:
            ec = exp_cat.lower()
            ru = EN_TO_RU_CATEGORY.get(ec, "")
            en = RU_TO_EN_CATEGORY.get(exp_cat, "")
            if ec not in found and ru.lower() not in found and en.lower() not in found:
                errors.append(f"C3: category '{exp_cat}' missing from Inventory Dashboard")
    finally:
        cur.close()
        conn.close()
    return (not errors), errors


def critical_email():
    """C4: email to warehouse@company.com lists most low-stock SKUs/names."""
    errors = []
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute(
            """SELECT subject, to_addr, body_text FROM email.messages
               WHERE (LOWER(subject) LIKE '%low stock%' OR LOWER(subject) LIKE '%alert%'
                      OR LOWER(subject) LIKE '%restock%')"""
        )
        emails = cur.fetchall()
        matched = None
        for subject, to_addr, body in emails:
            if "warehouse@company.com" in (str(to_addr).lower() if to_addr else ""):
                matched = (subject, body)
                break
        if matched is None:
            return False, ["C4: no low-stock email to warehouse@company.com"]
        body = (matched[1] or "").lower()
        if not body:
            return False, ["C4: email body is empty"]
        expected = get_expected_low_stock()
        total = len(expected)
        matches = 0
        for name, sku, stock, cat, price in expected:
            if str(sku).lower() in body or str(name)[:30].lower() in body:
                matches += 1
        # Require the email to cover most of the low-stock set.
        need = max(3, (total * 3 + 3) // 4)  # ~75%
        if matches < need:
            errors.append(f"C4: email lists {matches}/{total} low-stock products (need >= {need})")
    finally:
        cur.close()
        conn.close()
    return (not errors), errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, help="Launch time")
    args = parser.parse_args()

    # ---- CRITICAL semantic gate (any fail => exit 1, before accuracy) ----
    critical = [
        ("C1 Category Inventory values", lambda: critical_category_inventory(args.agent_workspace)),
        ("C2 Restock Alerts exact SKU set", lambda: critical_restock_alerts(args.agent_workspace)),
        ("C3 Inventory Dashboard categories", critical_gsheet),
        ("C4 Low-stock email coverage", critical_email),
        ("C5 Cost Trend latest-5 closes", lambda: critical_cost_trend(args.agent_workspace)),
    ]
    critical_failed = False
    for label, fn in critical:
        try:
            ok, errs = fn()
        except Exception as e:
            ok, errs = False, [f"{label}: exception {e}"]
        if ok:
            print(f"[CRITICAL PASS] {label}")
        else:
            critical_failed = True
            print(f"[CRITICAL FAIL] {label}")
            for e in errs:
                print(f"  - {e}")
    if critical_failed:
        print("\nCritical semantic check(s) failed. => FAIL")
        sys.exit(1)

    # ---- Accuracy gate (structural / soft checks), threshold >= 70 ----
    checks = []

    excel_pass, excel_errors = check_excel(args.agent_workspace)
    checks.append(("Excel", excel_pass, excel_errors))

    gsheet_pass, gsheet_errors = check_gsheet()
    checks.append(("Google Sheet", gsheet_pass, gsheet_errors))

    email_pass, email_errors = check_email()
    checks.append(("Email", email_pass, email_errors))

    passed = 0
    for name, ok, errs in checks:
        if ok:
            passed += 1
            print(f"[PASS] {name} check passed.")
        else:
            print(f"[FAIL] {name} check failed:")
            for e in errs:
                print(f"  - {e}")

    accuracy = 100.0 * passed / len(checks)
    print(f"\nAccuracy: {accuracy:.1f}% ({passed}/{len(checks)})")

    if accuracy >= 70:
        print("All critical checks passed and accuracy >= 70. => PASS")
        sys.exit(0)
    else:
        print("Accuracy below threshold. => FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
