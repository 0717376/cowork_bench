"""Evaluation for insales-product-review-analysis-gform-gcal (InSales + RU Forms/Calendar/Email).

Russified task: prose is Russian, but ALL English literals the eval greps are
preserved (sheet/column/metric names, English product names, file/form/event/email
literals). The agent legitimately writes Russian form titles, event summaries and
email subjects/bodies, so all free-text matching accepts RU+EN keywords. English
product NAMES (never translated) drive the CRITICAL semantic checks.
"""
import argparse
import json
import os
import sys

import psycopg2
import openpyxl


DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Free-text keyword sets accepted in either Russian or English (agent may write RU).
QUALITY_KEYS = ["quality", "качеств", "product", "продук", "товар"]
REVIEW_KEYS = ["review", "провер", "обзор", "отзыв", "качеств"]
SURVEY_TITLE_KEYS = ["product quality", "quality improvement", "качеств", "опрос", "улучшен"]


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {tag}{name}: {str(detail)[:300]}")
        if critical:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_review_data():
    """Source of truth: average review rating per product from live wc.* seed.

    Tie-break: more reviews. English product names are preserved by the central
    wc seed, so they drive the semantic checks below.
    """
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id, p.name, p.categories,
            count(pr.id) as review_count,
            avg(pr.rating) as avg_rating
        FROM wc.products p
        LEFT JOIN wc.product_reviews pr ON p.id = pr.product_id
        WHERE pr.id IS NOT NULL
        GROUP BY p.id, p.name, p.categories
        HAVING count(pr.id) >= 1
        ORDER BY avg(pr.rating) ASC, count(pr.id) DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    total = len(rows)
    lowest = rows[0] if rows else None
    overall_avg = round(sum(round(float(r[4]), 2) for r in rows) / len(rows), 2) if rows else 0
    return {
        "all_rows": rows,
        "total": total,
        "lowest_name": lowest[1][:40] if lowest else "",
        "lowest_avg": round(float(lowest[4]), 2) if lowest else 0,
        "top5": rows[:5],
        "top5_names": [r[1] for r in rows[:5]],
        "overall_avg": overall_avg,
    }


def check_review_summary(agent_wb, review_data):
    print("\n=== Checking Review Summary sheet ===")
    a_rows = load_sheet_rows(agent_wb, "Review Summary")
    if a_rows is None:
        record("Sheet 'Review Summary' exists", False, "not found")
        record("Review Summary: top-5 lowest-rated products present in ascending order",
               False, "no sheet", critical=True)
        return
    record("Sheet 'Review Summary' exists", True)

    data_rows = [r for r in a_rows[1:] if r and any(c is not None for c in r)]
    record("Review Summary has >= 5 data rows", len(data_rows) >= 5,
           f"{len(data_rows)} rows")

    # Non-critical structural: ascending sort (first <= last).
    if len(data_rows) >= 2:
        first_rating, last_rating = data_rows[0][1], data_rows[-1][1]
        try:
            record("Review Summary sorted by Avg_Rating ASC",
                   float(first_rating) <= float(last_rating),
                   f"first={first_rating}, last={last_rating}")
        except (TypeError, ValueError):
            record("Review Summary sorted by Avg_Rating ASC", False,
                   f"non-numeric ratings first={first_rating}, last={last_rating}")

    # CRITICAL: the actual top-5 lowest-rated products (live SQL) appear in the
    # first 5 data rows, in ascending Avg_Rating order, with the lowest first.
    top5 = review_data["top5"]
    if not top5:
        record("Review Summary: top-5 lowest-rated products present in ascending order",
               False, "no SQL data", critical=True)
        return

    # Build a name->row map from the agent's first column (Product_Name, truncated).
    def norm(s):
        return str(s).strip().lower() if s is not None else ""

    agent_names = [norm(r[0]) for r in data_rows]

    # Each of the 5 expected English product names must appear in the sheet.
    matched = 0
    for _id, name, _cat, _cnt, _avg in top5:
        nm = name[:60].strip().lower()
        nm_short = name[:30].strip().lower()
        if any(nm in an or an in nm or nm_short in an for an in agent_names):
            matched += 1
    record("Review Summary contains all 5 lowest-rated product names (EN)",
           matched >= 5,
           f"matched {matched}/5", critical=True)

    # First data row must be the single lowest-rated product (top5[0]).
    expected_first = top5[0][1][:60].strip().lower()
    actual_first = agent_names[0] if agent_names else ""
    record("Review Summary first row == lowest-rated product",
           expected_first in actual_first or actual_first in expected_first
           or top5[0][1][:30].strip().lower() in actual_first,
           f"expected '{top5[0][1]}', got '{data_rows[0][0]}'", critical=True)


def check_stats(agent_wb, review_data):
    print("\n=== Checking Stats sheet ===")
    a_rows = load_sheet_rows(agent_wb, "Stats")
    if a_rows is None:
        record("Sheet 'Stats' exists", False, "not found")
        record("Stats: Total_Products_Reviewed & Lowest_Rated_Product match live SQL",
               False, "no sheet", critical=True)
        return
    record("Sheet 'Stats' exists", True)

    a_data = {str(r[0]).strip().lower(): r[1] for r in a_rows[1:] if r and r[0] is not None}

    total_val = a_data.get("total_products_reviewed")
    lowest_val = a_data.get("lowest_rated_product")
    avg_overall_val = a_data.get("avg_rating_overall")

    # Non-critical: presence + loose overall-avg.
    record("Stats has Avg_Rating_Overall (tol=0.2)",
           avg_overall_val is not None and num_close(avg_overall_val, review_data["overall_avg"], 0.2),
           f"{avg_overall_val} vs {review_data['overall_avg']}")

    # CRITICAL: total reviewed products exact AND lowest-rated product name matches.
    total_ok = total_val is not None and num_close(total_val, review_data["total"], 0)
    expected_lower = review_data["lowest_name"].lower().strip()
    agent_lower = str(lowest_val).lower().strip() if lowest_val is not None else ""
    name_ok = bool(expected_lower) and (expected_lower in agent_lower or agent_lower in expected_lower)
    record("Stats: Total_Products_Reviewed & Lowest_Rated_Product match live SQL",
           total_ok and name_ok,
           f"total={total_val} vs {review_data['total']}; "
           f"lowest='{lowest_val}' vs '{review_data['lowest_name']}'", critical=True)


def check_gform(review_data):
    print("\n=== Checking Google Form (forms / gform.*) ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except Exception as e:
        record("Form 'Product Quality Improvement Survey' exists", False, str(e))
        record("Form covers the correct 5 lowest-rated products with required questions",
               False, "no db", critical=True)
        return

    cur.execute("SELECT id, title, description FROM gform.forms")
    forms = cur.fetchall()

    def title_match(t):
        tl = (t or "").lower()
        return any(k in tl for k in SURVEY_TITLE_KEYS)

    target = None
    for fid, title, _desc in forms:
        if title_match(title):
            target = fid
            break
    if target is None and forms:
        target = forms[0][0]

    record("Survey form created (title matches product/quality, RU+EN)",
           target is not None,
           f"forms found: {[f[1] for f in forms]}")
    if target is None:
        record("Form covers the correct 5 lowest-rated products with required questions",
               False, "no form", critical=True)
        cur.close()
        conn.close()
        return

    cur.execute(
        "SELECT title, question_type, required FROM gform.questions "
        "WHERE form_id = %s ORDER BY position",
        (target,),
    )
    questions = cur.fetchall()
    cur.close()
    conn.close()

    record("Form has >= 6 questions (5 rating + 1 text)", len(questions) >= 6,
           f"found {len(questions)}")

    required_count = sum(1 for q in questions if q[2])
    # 5 required product questions + 1 optional text question.
    structure_ok = len(questions) >= 6 and required_count >= 5
    record("Form has >= 5 required questions and >= 1 optional question",
           structure_ok,
           f"required={required_count}, total={len(questions)}")

    # CRITICAL: question texts cover the correct 5 lowest-rated products (EN names).
    matched = 0
    for name in review_data["top5_names"]:
        nm = name[:30].lower()
        nm_short = name[:20].lower()
        for q_title, _qt, _req in questions:
            qt = (q_title or "").lower()
            if nm in qt or nm_short in qt:
                matched += 1
                break
    record("Form covers the correct 5 lowest-rated products with required questions",
           matched >= 5 and structure_ok,
           f"product names matched {matched}/5, required={required_count}, total={len(questions)}",
           critical=True)


def check_gcal_event():
    print("\n=== Checking Calendar event (gcal.*) ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except Exception as e:
        record("Calendar event on 2026-03-20 14:00-15:00 exists", False, str(e), critical=True)
        return

    # Keep the date filter (do NOT drop it). Accept RU+EN summaries.
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE DATE(start_datetime) = '2026-03-20'
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    def summ_match(s):
        sl = (s or "").lower()
        return any(k in sl for k in QUALITY_KEYS)

    on_date = [e for e in events if summ_match(e[0])]
    record("Quality/product event exists on 2026-03-20 (RU+EN)",
           len(on_date) >= 1,
           f"events on date: {[(e[0], str(e[1])) for e in events]}")

    # CRITICAL: an event on 2026-03-20 with start 14:00 / end 15:00 (1h duration).
    time_ok = False
    detail = ""
    for summ, sdt, edt in on_date:
        if sdt is None or edt is None:
            continue
        dur_h = (edt - sdt).total_seconds() / 3600.0
        # start hour 14 (allow +/-1h for tz storage), duration ~1h.
        if 13 <= sdt.hour <= 15 and abs(dur_h - 1.0) <= 0.05:
            time_ok = True
            detail = f"start={sdt}, end={edt}, dur={dur_h:.2f}h"
            break
        detail = f"start={sdt}, end={edt}, dur={dur_h:.2f}h"
    record("Calendar event on 2026-03-20 starts ~14:00 and lasts 1h",
           time_ok, detail or "no matching event", critical=True)


def check_email_sent(review_data):
    print("\n=== Checking email to product.team ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
    except Exception as e:
        record("Email to product.team@company.com mentions all 5 products", False, str(e), critical=True)
        return

    cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
    rows = cur.fetchall()
    cur.close()
    conn.close()

    def to_blob(to):
        if isinstance(to, list):
            return " ".join(str(r).lower() for r in to)
        s = str(to or "")
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return " ".join(str(r).lower() for r in parsed)
        except Exception:
            pass
        return s.lower()

    target = [(subj, to, body) for (subj, to, body) in rows if "product.team" in to_blob(to)]
    record("Email sent to product.team@company.com", len(target) >= 1,
           f"recipients: {[to_blob(r[1]) for r in rows]}")

    # Subject accepts RU+EN keywords (agent may write a Russian subject).
    subj_ok = any(
        any(k in (subj or "").lower() for k in QUALITY_KEYS + REVIEW_KEYS)
        for subj, _to, _body in target
    )
    record("Email subject matches quality/review (RU+EN)", subj_ok,
           f"subjects: {[r[0] for r in target]}")

    # CRITICAL: at least one email to product.team whose body mentions all 5
    # lowest-rated product names (English names, preserved by the wc seed).
    best = 0
    for _subj, _to, body in target:
        bl = (body or "").lower()
        cnt = sum(1 for name in review_data["top5_names"]
                  if name[:30].lower() in bl or name[:20].lower() in bl)
        best = max(best, cnt)
    record("Email body mentions all 5 lowest-rated product names",
           len(target) >= 1 and best >= 5,
           f"matched {best}/5 product names", critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Product_Review_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Product_Review_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    review_data = get_review_data()

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

    check_review_summary(agent_wb, review_data)
    check_stats(agent_wb, review_data)
    check_gform(review_data)
    check_gcal_event()
    check_email_sent(review_data)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed, accuracy={accuracy:.1f}% ===")

    success = (not CRITICAL_FAILED) and accuracy >= 70
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy, "critical_failed": CRITICAL_FAILED,
                       "success": success}, f)

    if CRITICAL_FAILED:
        print(f"\n=== RESULT: FAIL (critical checks failed: {CRITICAL_FAILED}) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    print("\n=== RESULT: FAIL (accuracy < 70) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
