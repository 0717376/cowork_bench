"""Evaluation script for pw-sf-hr-dept-efficiency-excel-gcal (ClickHouse, RU)."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Department labels may be written in Russian (DWH russified centrally) or English.
# Alias map -> canonical English key so checks are language-agnostic.
DEPT_ALIASES = {
    "engineering": "Engineering", "инженерия": "Engineering",
    "finance": "Finance", "финансы": "Finance",
    "hr": "HR", "кадры": "HR", "human resources": "HR",
    "operations": "Operations", "операции": "Operations",
    "r&d": "R&D", "rd": "R&D", "ниокр": "R&D",
    "sales": "Sales", "продажи": "Sales",
    "support": "Support", "поддержка": "Support",
}


def canon_dept(val):
    if val is None:
        return None
    return DEPT_ALIASES.get(str(val).strip().lower(), str(val).strip())


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def critical(name, condition, detail=""):
    """A semantic check whose failure forces overall FAIL."""
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS][CRITICAL] {name}")
    else:
        FAIL_COUNT += 1
        CRITICAL_FAILS.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL][CRITICAL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    excel_path = os.path.join(agent_workspace, "Hr_Dept_Efficiency_Report.xlsx")
    check("Hr_Dept_Efficiency_Report.xlsx exists", os.path.exists(excel_path))

    data_rows_parsed = []   # list of dicts with parsed Data_Analysis values
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---------- Data_Analysis ----------
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            headers_lc = [h.lower() for h in headers]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 7 rows", len(data_rows) >= 7, f"got {len(data_rows)}")

            for expected_col in ['Department', 'Employee_Count', 'Our_Avg_Salary', 'Avg_Experience',
                                 'Avg_Performance', 'Industry_Benchmark', 'Salary_Gap']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers_lc, f"headers: {headers[:8]}")

            # Build column index map
            def idx(col):
                return headers_lc.index(col.lower()) if col.lower() in headers_lc else None

            i_dep = idx('Department')
            i_emp = idx('Employee_Count')
            i_sal = idx('Our_Avg_Salary')
            i_bench = idx('Industry_Benchmark')
            i_gap = idx('Salary_Gap')

            if None not in (i_dep, i_sal, i_bench, i_gap):
                for r in data_rows:
                    if r is None or all(c is None for c in r):
                        continue
                    dep_raw = r[i_dep]
                    if dep_raw is None or str(dep_raw).strip() == "":
                        continue
                    data_rows_parsed.append({
                        "dept_raw": dep_raw,
                        "dept": canon_dept(dep_raw),
                        "emp": safe_float(r[i_emp]) if i_emp is not None else None,
                        "salary": safe_float(r[i_sal]),
                        "bench": safe_float(r[i_bench]),
                        "gap": safe_float(r[i_gap]),
                    })

            # CRITICAL: all 7 expected departments present (RU or EN)
            expected_depts = {"Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"}
            present = {row["dept"] for row in data_rows_parsed}
            critical("Data_Analysis covers all 7 departments",
                     expected_depts.issubset(present),
                     f"present={sorted(present)}")

            # CRITICAL: gap arithmetic self-consistent per row
            #   Salary_Gap == round(Our_Avg_Salary - Industry_Benchmark, 2)
            bad_gap = []
            for row in data_rows_parsed:
                if None in (row["salary"], row["bench"], row["gap"]):
                    bad_gap.append((row["dept"], "missing value"))
                    continue
                expected_gap = round(row["salary"] - row["bench"], 2)
                if abs(expected_gap - row["gap"]) > 0.5:
                    bad_gap.append((row["dept"], f"got {row['gap']} expected {expected_gap}"))
            critical("Salary_Gap = round(Our_Avg_Salary - Industry_Benchmark, 2) for every row",
                     len(bad_gap) == 0 and len(data_rows_parsed) >= 7, f"errors: {bad_gap[:5]}")

            # CRITICAL: rows sorted alphabetically by Department (as written)
            written_depts = [str(row["dept_raw"]).strip() for row in data_rows_parsed]
            critical("Data_Analysis sorted alphabetically by Department",
                     written_depts == sorted(written_depts),
                     f"order={written_depts}")

        # ---------- Metrics ----------
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            headers_lc = [h.lower() for h in headers]
            m_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 5 rows", len(m_rows) >= 5, f"got {len(m_rows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers_lc, f"headers: {headers[:8]}")
            for r in m_rows:
                if r and r[0] is not None:
                    metrics_map[str(r[0]).strip()] = r[1] if len(r) > 1 else None

            # CRITICAL: Total_Departments == 7 and Departments_Below == count of negative gaps
            total_dep = safe_float(metrics_map.get("Total_Departments"))
            critical("Metrics Total_Departments == 7", total_dep == 7,
                     f"got {metrics_map.get('Total_Departments')}")

            neg_count = sum(1 for row in data_rows_parsed
                            if row["gap"] is not None and row["gap"] < 0)
            below = safe_float(metrics_map.get("Departments_Below"))
            critical("Metrics Departments_Below == number of negative-gap departments",
                     below is not None and int(below) == neg_count,
                     f"reported {metrics_map.get('Departments_Below')}, actual {neg_count}")

        # ---------- Recommendations ----------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            headers_lc = [h.lower() for h in headers]
            r_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 3 rows", len(r_rows) >= 3, f"got {len(r_rows)}")
            for expected_col in ['Priority', 'Action', 'Department', 'Impact']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers_lc, f"headers: {headers[:8]}")

    # ---------- Calendar (CRITICAL: title + date + time window) ----------
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""SELECT summary, start_datetime, end_datetime
                       FROM gcal.events
                       WHERE summary ILIKE %s""", ('%Analysis Review%',))
        events = cur.fetchall()
        conn.close()

        match = None
        for summary, start_dt, end_dt in events:
            if start_dt is None:
                continue
            if start_dt.strftime("%Y-%m-%d") != "2026-03-14":
                continue
            # start at 14:00, ~1h duration
            start_ok = start_dt.hour == 14 and start_dt.minute == 0
            dur_ok = False
            if end_dt is not None:
                dur_min = (end_dt - start_dt).total_seconds() / 60.0
                dur_ok = 55 <= dur_min <= 65 and end_dt.hour == 15
            if start_ok and dur_ok:
                match = (summary, start_dt, end_dt)
                break
        critical("Calendar event 'Analysis Review' on 2026-03-14 14:00-15:00 UTC",
                 match is not None,
                 f"candidates={[(s, str(sd), str(ed)) for s, sd, ed in events][:5]}")
    except Exception as e:
        critical("Calendar event 'Analysis Review' on 2026-03-14 14:00-15:00 UTC", False, str(e))

    # ---------- Python deliverables (CRITICAL: script + valid results JSON) ----------
    proc_path = os.path.join(agent_workspace, "sf_efficiency_processor.py")
    res_path = os.path.join(agent_workspace, "sf_efficiency_results.json")
    check("sf_efficiency_processor.py exists", os.path.exists(proc_path))
    results_ok = False
    if os.path.exists(res_path):
        try:
            with open(res_path, "r", encoding="utf-8") as f:
                json.load(f)
            results_ok = True
        except Exception as e:
            results_ok = False
            print(f"  [WARN] sf_efficiency_results.json not valid JSON: {e}")
    critical("sf_efficiency_processor.py exists AND sf_efficiency_results.json is valid JSON",
             os.path.exists(proc_path) and results_ok,
             f"proc={os.path.exists(proc_path)} results_ok={results_ok}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    return accuracy


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    accuracy = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )

    print(f"\nAccuracy: {accuracy:.1f}% ({PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks)")

    # CRITICAL gate: any critical failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILS:
        print(f"CRITICAL FAILURES: {CRITICAL_FAILS}")
        sys.exit(1)

    # Accuracy gate.
    if accuracy >= 70.0:
        print("RESULT: PASS")
        sys.exit(0)
    else:
        print("RESULT: FAIL (accuracy below 70%)")
        sys.exit(1)


if __name__ == "__main__":
    main()
