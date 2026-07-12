"""Evaluation for canvas-late-submission-gcal."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": "cowork_gym", "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Top-3 courses by Late_Count read from the live Canvas source data.
# These are SEMANTIC ground-truth values proving real aggregation, not guessing.
TOP3_BY_COURSE = [
    ("Биохимия и биоинформатика (Осень 2013)", 7696, 1564),
    ("Креативные вычисления и культура (Осень 2014)", 6638, 1965),
    ("Биохимия и биоинформатика (Весна 2013)", 6350, 1286),
]


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "[CRITICAL] " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILS.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def int_close(a, b, tol=5):
    try:
        return abs(int(float(a)) - int(float(b))) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Late_Submissions.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Late_Submissions.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Late_Submissions.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    # Sheet 1: By Course
    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    agent_ws = get_sheet(agent_wb, "By Course")
    gt_ws = get_sheet(gt_wb, "By Course")

    if agent_ws is None:
        record("Sheet 'By Course' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'By Course' exists", True)
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("By Course row count", len(agent_rows) >= len(gt_rows) - 2,
               f"Expected ~{len(gt_rows)}, got {len(agent_rows)}")

        # Check top 5 courses by late count
        gt_lookup = {}
        for r in gt_rows:
            if r and r[0]:
                gt_lookup[str(r[0]).strip().lower()] = r

        agent_lookup = {}
        for r in agent_rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for gt_row in gt_rows[:5]:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Course '{gt_row[0]}' present", False, "Missing")
                all_ok = False
            else:
                ok = int_close(a_row[1], gt_row[1], 50) and int_close(a_row[2], gt_row[2], 20)
                record(f"Course '{gt_row[0]}' data", ok,
                       f"Late: {a_row[1]} vs {gt_row[1]}, Students: {a_row[2]} vs {gt_row[2]}")
                if not ok:
                    all_ok = False

        # CRITICAL: top-3 courses must match the real source aggregation tightly
        # (tol<=2), proving honest counting rather than guessing/presence-only.
        for cname, exp_late, exp_stud in TOP3_BY_COURSE:
            a_row = agent_lookup.get(cname.strip().lower())
            if a_row is None:
                record(f"Top-3 course '{cname}' exact aggregation", False,
                       "Missing from 'By Course'", critical=True)
                all_ok = False
            else:
                ok = int_close(a_row[1], exp_late, 2) and int_close(a_row[2], exp_stud, 2)
                record(f"Top-3 course '{cname}' exact aggregation", ok,
                       f"Late: {a_row[1]} vs {exp_late}, Students: {a_row[2]} vs {exp_stud}",
                       critical=True)
                if not ok:
                    all_ok = False

        # CRITICAL: 'By Course' must be sorted by Late_Count descending.
        counts = []
        for r in agent_rows:
            if r and r[0] is not None and r[1] is not None:
                try:
                    counts.append(int(float(r[1])))
                except (TypeError, ValueError):
                    pass
        sorted_desc = all(counts[i] >= counts[i + 1] for i in range(len(counts) - 1))
        record("'By Course' sorted by Late_Count descending", sorted_desc and len(counts) >= 2,
               f"First few: {counts[:5]}", critical=True)
        if not (sorted_desc and len(counts) >= 2):
            all_ok = False

    # Sheet 2: Top Offenders
    agent_ws2 = get_sheet(agent_wb, "Top Offenders")
    gt_ws2 = get_sheet(gt_wb, "Top Offenders")

    if agent_ws2 is None:
        record("Sheet 'Top Offenders' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Top Offenders' exists", True)
        agent_rows2 = list(agent_ws2.iter_rows(min_row=2, values_only=True))
        gt_rows2 = list(gt_ws2.iter_rows(min_row=2, values_only=True))

        record("Top Offenders has ~10 rows", 8 <= len(agent_rows2) <= 12,
               f"Got {len(agent_rows2)}")

        # Top Offenders must be sorted by Late_Count descending (structural).
        off_counts = []
        for r in agent_rows2:
            if r and r[2] is not None:
                try:
                    off_counts.append(int(float(r[2])))
                except (TypeError, ValueError):
                    pass
        off_sorted = all(off_counts[i] >= off_counts[i + 1] for i in range(len(off_counts) - 1))
        record("'Top Offenders' sorted by Late_Count descending", off_sorted,
               f"First few: {off_counts[:5]}")

        # Check top 3 offenders
        for i, gt_row in enumerate(gt_rows2[:3]):
            if not gt_row or not gt_row[0]:
                continue
            found = False
            for a_row in agent_rows2:
                if a_row and a_row[0] and str(a_row[0]).strip().lower() == str(gt_row[0]).strip().lower():
                    ok = int_close(a_row[2], gt_row[2], 3)
                    record(f"Top offender '{gt_row[0]}'", ok,
                           f"Late: {a_row[2]} vs {gt_row[2]}")
                    if not ok:
                        all_ok = False
                    found = True
                    break
            if not found:
                record(f"Top offender '{gt_row[0]}' present", False, "Missing")
                all_ok = False

    return all_ok


def check_gcal():
    """Check calendar events for review meetings."""
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description FROM gcal.events ORDER BY summary")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")

    # Should have review meetings (22 courses with late submissions)
    record("At least 10 calendar events created", len(events) >= 10,
           f"Found {len(events)}")

    review_events = [e for e in events if "review meeting" in (e[0] or "").lower()]
    # CRITICAL: the calendar deliverable must actually exist.
    record("Review Meeting events found (>=10)", len(review_events) >= 10,
           f"Found {len(review_events)} review meeting events", critical=True)

    # CRITICAL: top-3 courses must each have an event whose description carries
    # the correct late-count / unique-student numbers from the real source data.
    for cname, exp_late, exp_stud in TOP3_BY_COURSE:
        matched = [e for e in events if cname.lower() in (e[0] or "").lower()]
        if not matched:
            record(f"Calendar event for '{cname[:40]}...' present", False,
                   "No event with this course title", critical=True)
            continue
        desc = " ".join((e[1] or "") for e in matched)
        ok = (str(exp_late) in desc) and (str(exp_stud) in desc)
        record(f"Calendar event '{cname[:40]}...' description numbers", ok,
               f"Expected late={exp_late}, students={exp_stud} in: {desc[:120]}",
               critical=True)

    return len(review_events) >= 10


def check_emails():
    """Check that summary email was sent."""
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    found_report = False
    for subject, from_addr, to_addr, body in all_emails:
        subj_lower = (subject or "").lower()
        if "late" in subj_lower and "submission" in subj_lower:
            found_report = True
            to_str = str(to_addr or "").lower()
            # CRITICAL: report email must reach the academic-affairs recipient.
            record("Email sent to academic.affairs@university.example.com",
                   "academic.affairs@university.example.com" in to_str,
                   f"To: {to_addr}", critical=True)

            body_lower = (body or "").lower()
            # Total course count (22) should appear; tolerate RU/EN wording.
            record("Email mentions course count",
                   "22" in body_lower or "course" in body_lower or "курс" in body_lower)

            # CRITICAL (semantic): body must name the real top-3 courses,
            # not merely contain the generic word 'course'/'курс'.
            top3_names = [c[0].lower() for c in TOP3_BY_COURSE]
            named = sum(1 for n in top3_names if n in body_lower)
            record("Email body names the real top-3 courses", named >= 2,
                   f"Matched {named}/3 top course names in body", critical=True)
            break

    # CRITICAL: the summary-email deliverable must exist.
    record("Late Submission Report email found", found_report, critical=True)
    return found_report


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES ({len(CRITICAL_FAILS)}):")
        for c in CRITICAL_FAILS:
            print(f"    - {c}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
