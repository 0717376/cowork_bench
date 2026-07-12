"""Оценка для insales-coupon-marketing-roi (InSales + RU forms / gform.*).

Критические проверки (CRITICAL_CHECKS): любой провал => общий FAIL независимо от
accuracy. Иначе PASS требует accuracy >= 70%.

Ожидаемые значения берутся из files/groundtruth_data.json — это тот же источник,
из которого агент косвенно считывает данные (mock API + купоны магазина InSales).
Названия кампаний и каналов держим английскими (как в mock API и в магазине),
поэтому десинхронизации нет. Хардкодить волатильные значения здесь не нужно —
все агрегаты вычисляются из groundtruth_data.json.
"""
import argparse
import json
import os
import sys

import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# Критические проверки: любой провал => общий FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Campaign Performance: ровно 10 строк (только кампании с валидным купоном)",
    "Campaign Performance: Revenue и ROI_Pct верны для ключевых кампаний",
    "Campaign Performance: Meets_ROI_Target/Meets_Conversion_Target корректны",
    "Recommendations: счётчики ROI и Best/Worst Channel корректны",
    "Форма 'Campaign Effectiveness Feedback': >=3 вопроса нужных типов",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def yes_no(val):
    """Нормализует значение Yes/No (RU+EN) к булеву или None."""
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("yes", "да", "true", "1", "y"):
        return True
    if s in ("no", "нет", "false", "0", "n"):
        return False
    return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def expected_channel_summary(gt):
    ch = {}
    for c in gt["campaigns"]:
        a = ch.setdefault(c["channel"], {"budget": 0, "revenue": 0, "count": 0})
        a["budget"] += c["budget"]
        a["revenue"] += c["revenue"]
        a["count"] += 1
    for k, v in ch.items():
        v["roi"] = round((v["revenue"] - v["budget"]) / v["budget"] * 100, 2)
    return ch


def check_excel(agent_workspace, gt):
    try:
        import openpyxl
    except ImportError:
        check("Campaign_ROI.xlsx существует", False, "openpyxl не установлен")
        for c in CRITICAL_CHECKS:
            if c != "Форма 'Campaign Effectiveness Feedback': >=3 вопроса нужных типов":
                check(c, False, "нет openpyxl")
        return

    path = os.path.join(agent_workspace, "Campaign_ROI.xlsx")
    exists = os.path.exists(path)
    check("Campaign_ROI.xlsx существует", exists, path)
    if not exists:
        for c in CRITICAL_CHECKS:
            if c != "Форма 'Campaign Effectiveness Feedback': >=3 вопроса нужных типов":
                check(c, False, "файл не найден")
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Campaign_ROI.xlsx читается", False, str(e))
        for c in CRITICAL_CHECKS:
            if c != "Форма 'Campaign Effectiveness Feedback': >=3 вопроса нужных типов":
                check(c, False, "ошибка чтения")
        return

    # ===== Лист Campaign Performance =====
    rows = load_sheet_rows(wb, "Campaign Performance")
    if rows is None:
        check("Лист 'Campaign Performance' существует", False, "лист не найден")
        check("Campaign Performance: ровно 10 строк (только кампании с валидным купоном)",
              False, "нет листа")
        check("Campaign Performance: Revenue и ROI_Pct верны для ключевых кампаний",
              False, "нет листа")
        check("Campaign Performance: Meets_ROI_Target/Meets_Conversion_Target корректны",
              False, "нет листа")
        data_rows = []
        lookup = {}
    else:
        check("Лист 'Campaign Performance' существует", True)
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        expected_n = gt["total_campaigns"]
        # CRITICAL: ровно столько строк, сколько кампаний с валидным купоном (10);
        # Brand Awareness (пустой код) и Influencer Collab (INFLUENCER20 нет в магазине)
        # должны быть исключены. Допуск 0 — это ровно 10.
        check("Campaign Performance: ровно 10 строк (только кампании с валидным купоном)",
              len(data_rows) == expected_n,
              f"строк {len(data_rows)}, ожидалось {expected_n}")

        lookup = {}
        for r in data_rows:
            if r[0] is not None:
                lookup[str(r[0]).strip().lower()] = r

        # CRITICAL: Revenue + ROI_Pct для всех кампаний groundtruth (не молча
        # пропускать, если строка не найдена).
        # Колонки: 0=Name 1=Channel 2=Coupon 3=Budget 4=Usage 5=Revenue 6=ROI 7=Conv 8=MeetsROI 9=MeetsConv
        roi_errors = []
        for gc in gt["campaigns"]:
            row = lookup.get(gc["campaign_name"].strip().lower())
            if row is None:
                roi_errors.append(f"{gc['campaign_name']}: строка отсутствует")
                continue
            if len(row) < 10:
                roi_errors.append(f"{gc['campaign_name']}: мало столбцов ({len(row)})")
                continue
            if not num_close(row[5], gc["revenue"], 1):
                roi_errors.append(f"{gc['campaign_name']} Revenue={row[5]} ожид {gc['revenue']}")
            if not num_close(row[6], gc["roi_pct"], 1):
                roi_errors.append(f"{gc['campaign_name']} ROI={row[6]} ожид {gc['roi_pct']}")
        check("Campaign Performance: Revenue и ROI_Pct верны для ключевых кампаний",
              not roi_errors, "; ".join(roi_errors[:6]))

        # Non-critical: Usage_Count и Conversion_Rate.
        uc_errors = []
        for gc in gt["campaigns"]:
            row = lookup.get(gc["campaign_name"].strip().lower())
            if row is None or len(row) < 8:
                continue
            if not num_close(row[4], gc["usage_count"], 0):
                uc_errors.append(f"{gc['campaign_name']} Usage={row[4]} ожид {gc['usage_count']}")
            if not num_close(row[7], gc["conversion_rate"], 0.5):
                uc_errors.append(f"{gc['campaign_name']} Conv={row[7]} ожид {gc['conversion_rate']}")
        check("Campaign Performance: Usage_Count и Conversion_Rate верны",
              not uc_errors, "; ".join(uc_errors[:6]))

        # CRITICAL: Meets_ROI_Target / Meets_Conversion_Target.
        # ROI ни у кого не достигает 200% => все No. Conversion >=3% только у VIP Exclusive.
        meets_errors = []
        for gc in gt["campaigns"]:
            row = lookup.get(gc["campaign_name"].strip().lower())
            if row is None or len(row) < 10:
                meets_errors.append(f"{gc['campaign_name']}: нет столбцов Meets_*")
                continue
            got_roi = yes_no(row[8])
            got_conv = yes_no(row[9])
            if got_roi != gc["meets_roi_target"]:
                meets_errors.append(
                    f"{gc['campaign_name']} Meets_ROI={row[8]} ожид {gc['meets_roi_target']}")
            if got_conv != gc["meets_conversion_target"]:
                meets_errors.append(
                    f"{gc['campaign_name']} Meets_Conv={row[9]} ожид {gc['meets_conversion_target']}")
        check("Campaign Performance: Meets_ROI_Target/Meets_Conversion_Target корректны",
              not meets_errors, "; ".join(meets_errors[:6]))

    # ===== Лист Channel Summary =====
    exp_ch = expected_channel_summary(gt)
    rows2 = load_sheet_rows(wb, "Channel Summary")
    if rows2 is None:
        check("Лист 'Channel Summary' существует", False, "лист не найден")
        ch_lookup = {}
    else:
        check("Лист 'Channel Summary' существует", True)
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        ch_lookup = {str(r[0]).strip().lower(): r for r in data2}
        # Non-critical: агрегаты по каналам (Total_Budget/Total_Revenue/Channel_ROI_Pct/Count).
        # Колонки: 0=Channel 1=Total_Budget 2=Total_Revenue 3=Channel_ROI_Pct 4=Campaign_Count
        ch_errors = []
        for name, v in exp_ch.items():
            r = ch_lookup.get(name.strip().lower())
            if r is None:
                ch_errors.append(f"канал {name} отсутствует")
                continue
            if len(r) < 5:
                ch_errors.append(f"{name}: мало столбцов")
                continue
            if not num_close(r[1], v["budget"], 1):
                ch_errors.append(f"{name} Budget={r[1]} ожид {v['budget']}")
            if not num_close(r[2], v["revenue"], 1):
                ch_errors.append(f"{name} Revenue={r[2]} ожид {v['revenue']}")
            if not num_close(r[3], v["roi"], 1):
                ch_errors.append(f"{name} ROI={r[3]} ожид {v['roi']}")
            if not num_close(r[4], v["count"], 0):
                ch_errors.append(f"{name} Count={r[4]} ожид {v['count']}")
        check("Channel Summary: агрегаты по каналам верны",
              not ch_errors, "; ".join(ch_errors[:6]))

    # ===== Лист Recommendations =====
    best_channel = max(exp_ch, key=lambda k: exp_ch[k]["roi"])
    worst_channel = min(exp_ch, key=lambda k: exp_ch[k]["roi"])
    rows3 = load_sheet_rows(wb, "Recommendations")
    if rows3 is None:
        check("Лист 'Recommendations' существует", False, "лист не найден")
        check("Recommendations: счётчики ROI и Best/Worst Channel корректны",
              False, "нет листа")
    else:
        check("Лист 'Recommendations' существует", True)
        # Собираем пары (Category, Finding) -> текстовый блок.
        rec_map = {}
        blob_parts = []
        for r in rows3[1:]:
            if not r or r[0] is None:
                continue
            cat = str(r[0]).strip().lower()
            fnd = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
            rec_map[cat] = fnd
            blob_parts.append(cat + " " + fnd.lower())
        blob = " ".join(blob_parts)

        rec_errors = []
        # Campaigns Meeting ROI Target = 0
        meet_val = None
        for cat, fnd in rec_map.items():
            if "meeting" in cat or "достиг" in cat:
                meet_val = fnd
        if meet_val is None or not num_close(
                "".join(ch for ch in str(meet_val) if ch.isdigit() or ch == "-") or "x",
                gt["meets_roi_count"], 0):
            rec_errors.append(f"Meeting ROI Target ожид {gt['meets_roi_count']}, got '{meet_val}'")
        # Campaigns Below ROI Target = 10
        below_val = None
        for cat, fnd in rec_map.items():
            if "below" in cat or "ниже" in cat:
                below_val = fnd
        if below_val is None or not num_close(
                "".join(ch for ch in str(below_val) if ch.isdigit() or ch == "-") or "x",
                gt["below_roi_count"], 0):
            rec_errors.append(f"Below ROI Target ожид {gt['below_roi_count']}, got '{below_val}'")
        # Best / Worst Channel by ROI
        best_val = worst_val = None
        for cat, fnd in rec_map.items():
            if "best" in cat or "лучш" in cat:
                best_val = fnd.lower()
            if "worst" in cat or "худш" in cat or "наимень" in cat:
                worst_val = fnd.lower()
        if best_val is None or best_channel.lower() not in best_val:
            rec_errors.append(f"Best Channel ожид '{best_channel}', got '{best_val}'")
        if worst_val is None or worst_channel.lower() not in worst_val:
            rec_errors.append(f"Worst Channel ожид '{worst_channel}', got '{worst_val}'")
        check("Recommendations: счётчики ROI и Best/Worst Channel корректны",
              not rec_errors, "; ".join(rec_errors[:6]))


def _option_values(config):
    vals = []
    if not config:
        return vals
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except json.JSONDecodeError:
            return vals
    opts = config.get("options") if isinstance(config, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    vals.append(str(v))
            else:
                vals.append(str(o))
    return vals


def _is_text(t):
    return (t or "") in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH", "text")


def _is_choice(t):
    return (t or "") in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE", "CHOICE",
                         "CHECKBOX", "choice")


def check_gform():
    crit = "Форма 'Campaign Effectiveness Feedback': >=3 вопроса нужных типов"
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        check(crit, False, f"нет БД: {e}")
        return
    cur = conn.cursor()
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE title ILIKE '%%campaign%%' OR title ILIKE '%%effectiveness%%'
           OR title ILIKE '%%feedback%%' OR title ILIKE '%%кампан%%'
           OR title ILIKE '%%эффективн%%'
        ORDER BY created_at DESC LIMIT 5
    """)
    forms = cur.fetchall()
    check("Форма 'Campaign Effectiveness Feedback' существует", bool(forms),
          "форма не найдена")
    if not forms:
        check(crit, False, "форма не найдена")
        cur.close()
        conn.close()
        return

    form_id = forms[0][0]
    cur.execute(
        "SELECT title, question_type, config FROM gform.questions "
        "WHERE form_id = %s ORDER BY position", (form_id,))
    questions = cur.fetchall()
    cur.close()
    conn.close()

    parsed = []
    for q_title, q_type, q_config in questions:
        parsed.append({
            "title": (q_title or "").lower(),
            "type": q_type,
            "options_lower": [v.lower() for v in _option_values(q_config)],
        })

    text_count = sum(1 for p in parsed if _is_text(p["type"]))
    choice_count = sum(1 for p in parsed if _is_choice(p["type"]))

    # CRITICAL: >=3 вопроса, минимум 2 с множественным выбором (канал + частота)
    # и минимум 1 с коротким ответом (предложения).
    struct_ok = (len(parsed) >= 3 and choice_count >= 2 and text_count >= 1)
    check(crit, struct_ok,
          f"n={len(parsed)} text={text_count} choice={choice_count} "
          f"types={[p['type'] for p in parsed]}")

    # Non-critical: есть вопрос о канале и вопрос о частоте с нужными вариантами.
    freq_q = None
    best_hits = -1
    freq_opts = {
        "monthly": ["monthly", "ежемесяч", "ежемесячно"],
        "biweekly": ["bi-weekly", "biweekly", "раз в две недели", "двухнедель"],
        "weekly": ["weekly", "еженедель", "еженедельно"],
        "quarterly": ["quarterly", "ежекварт", "квартал"],
    }
    for p in parsed:
        joined = " ".join(p["options_lower"])
        hits = sum(1 for keys in freq_opts.values() if any(k in joined for k in keys))
        if hits > best_hits:
            best_hits = hits
            freq_q = p
    joined = " ".join(freq_q["options_lower"]) if freq_q else ""
    matched = {n: any(k in joined for k in keys) for n, keys in freq_opts.items()}
    check("Форма: вопрос о частоте содержит Monthly/Bi-weekly/Weekly/Quarterly",
          all(matched.values()), f"options='{joined[:160]}' matched={matched}")


def run_evaluation(agent_workspace, gt, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    print("\n=== Проверка Excel (Campaign_ROI.xlsx) ===")
    check_excel(agent_workspace, gt)
    print("\n=== Проверка формы (forms / gform.*) ===")
    check_gform()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Итог: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"КРИТИЧЕСКИЕ ПРОВАЛЫ: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                    "success": success,
                }, f, indent=2)
        except Exception:
            pass

    return success


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    success = run_evaluation(agent_ws, gt_data, args.res_log_file)

    if success:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
