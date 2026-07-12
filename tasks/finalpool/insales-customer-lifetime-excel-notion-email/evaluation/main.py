"""
Evaluation script for insales-customer-lifetime-excel-notion-email task.

Checks:
1. Excel file (Customer_CLV_Report.xlsx) - 3 sheets with correct data
2. Teamly space "Customer CRM" with >= N customer pages (hub+child pattern)
3. Emails sent to at-risk customers with retention content

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Pass threshold otherwise: accuracy >= 70%.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# Teamly space the agent must populate with one page per customer.
SPACE_KEY = "CUSTCRM"
SPACE_NAME = "Customer CRM"

# Minimum number of customer pages expected in the Customer CRM space.
MIN_CUSTOMER_PAGES = 45

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Excel file exists",
    "CLV Analysis data accuracy",
    "Tier Summary data accuracy",
    "Teamly 'Customer CRM' space exists",
    "Teamly CRM has >= 45 customer pages",
    "At-risk customers received emails (>= 50%)",
}


def check(name, condition, detail="", db=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_at_risk_emails():
    """Get expected at-risk customer emails from DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.email, c.first_name, c.last_name, c.orders_count, c.total_spent
        FROM wc.customers c
        WHERE c.orders_count <= 1
        ORDER BY c.email
    """)
    single_order = {row[0].lower() for row in cur.fetchall()}

    # Also get customers whose last order > 60 days ago
    cur.execute("""
        SELECT c.email
        FROM wc.customers c
        LEFT JOIN (
            SELECT customer_id, MAX(date_created) as last_order
            FROM wc.orders
            WHERE status IN ('completed', 'processing', 'on-hold', 'pending')
            GROUP BY customer_id
        ) o ON c.id = o.customer_id
        WHERE c.orders_count > 1
          AND (o.last_order IS NULL OR o.last_order < NOW() - INTERVAL '60 days')
    """)
    stale_order = {row[0].lower() for row in cur.fetchall()}

    cur.close()
    conn.close()
    return single_order | stale_order


def get_at_risk_customers():
    """Expected at-risk customers (name/email/orders/spent) from DB at eval time."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.first_name || ' ' || c.last_name, c.email,
               c.orders_count, c.total_spent
        FROM wc.customers c
        LEFT JOIN (
            SELECT customer_id, MAX(date_created) as last_order
            FROM wc.orders
            WHERE status IN ('completed', 'processing', 'on-hold', 'pending')
            GROUP BY customer_id
        ) o ON c.id = o.customer_id
        WHERE c.orders_count <= 1
           OR o.last_order IS NULL
           OR o.last_order < NOW() - INTERVAL '60 days'
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Customer_CLV_Report.xlsx")
    gt_path = os.path.join(gt_workspace, "Customer_CLV_Report.xlsx")

    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        return False

    if not os.path.isfile(gt_path):
        print(f"  WARNING: Groundtruth not found at {gt_path}, skipping comparison")
        return False

    try:
        agent_wb = openpyxl.load_workbook(excel_path, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # At-risk membership is time-dependent (60-day window) — derive it from
    # the DB at eval time instead of the frozen groundtruth columns.
    try:
        at_risk_set = get_at_risk_emails()
    except Exception as e:
        print(f"  WARNING: could not compute at-risk set from DB: {e}")
        at_risk_set = None

    # --- Sheet 1: CLV Analysis ---
    print("  Checking CLV Analysis sheet...")
    a_rows = load_sheet_rows(agent_wb, "CLV Analysis")
    g_rows = load_sheet_rows(gt_wb, "CLV Analysis")

    if a_rows is None:
        check("CLV Analysis sheet exists", False, f"Sheets: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("CLV Analysis sheet in groundtruth", False)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("CLV Analysis row count", abs(len(a_data) - len(g_data)) <= 2,
              f"Agent {len(a_data)} vs GT {len(g_data)}")

        # Build lookup by customer name (col 0)
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing customer: {g_row[0]}")
                continue

            # Col 2: Orders_Count
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1):
                    errors.append(f"{key}.Orders_Count: {a_row[2]} vs {g_row[2]}")

            # Col 3: Total_Spent
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5.0):
                    errors.append(f"{key}.Total_Spent: {a_row[3]} vs {g_row[3]}")

            # Col 4: Avg_Order_Value
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 5.0):
                    errors.append(f"{key}.Avg_Order_Value: {a_row[4]} vs {g_row[4]}")

            # Col 5: CLV_Tier
            if len(a_row) > 5 and len(g_row) > 5:
                if not str_match(a_row[5], g_row[5]):
                    errors.append(f"{key}.CLV_Tier: {a_row[5]} vs {g_row[5]}")

            # Col 6: Is_At_Risk — expected value derived from DB at eval time
            if len(a_row) > 6 and len(g_row) > 1:
                email = str(g_row[1] or "").strip().lower()
                if at_risk_set is not None and email:
                    expected_risk = "Yes" if email in at_risk_set else "No"
                else:
                    expected_risk = g_row[6] if len(g_row) > 6 else None
                if not str_match(a_row[6], expected_risk):
                    errors.append(f"{key}.Is_At_Risk: {a_row[6]} vs {expected_risk}")

        if errors:
            check("CLV Analysis data accuracy", False,
                  f"{len(errors)} errors: {'; '.join(errors[:5])}")
        else:
            check("CLV Analysis data accuracy", True)

    # Dynamic per-tier at-risk counts: tier mapping from GT (seed-stable),
    # at-risk membership from DB at eval time.
    expected_tier_risk = None
    if at_risk_set is not None and g_rows:
        expected_tier_risk = {}
        for g_row in g_rows[1:]:
            if not g_row or len(g_row) < 6 or g_row[1] is None or g_row[5] is None:
                continue
            tier = str(g_row[5]).strip().lower()
            if str(g_row[1]).strip().lower() in at_risk_set:
                expected_tier_risk[tier] = expected_tier_risk.get(tier, 0) + 1

    # --- Sheet 2: Tier Summary ---
    print("  Checking Tier Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Tier Summary")
    g_rows = load_sheet_rows(gt_wb, "Tier Summary")

    if a_rows is None:
        check("Tier Summary sheet exists", False, f"Sheets: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Tier Summary sheet in groundtruth", False)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing tier: {g_row[0]}")
                continue

            # Col 1: Customer_Count
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Customer_Count: {a_row[1]} vs {g_row[1]}")

            # Col 2: Total_Revenue
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 10.0):
                    errors.append(f"{key}.Total_Revenue: {a_row[2]} vs {g_row[2]}")

            # Col 3: Avg_CLV
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 10.0):
                    errors.append(f"{key}.Avg_CLV: {a_row[3]} vs {g_row[3]}")

            # Col 4: At_Risk_Count — expected value derived from DB at eval time
            if len(a_row) > 4 and len(g_row) > 4:
                if expected_tier_risk is not None:
                    expected_count = expected_tier_risk.get(key, 0)
                else:
                    expected_count = g_row[4]
                if not num_close(a_row[4], expected_count, 2):
                    errors.append(f"{key}.At_Risk_Count: {a_row[4]} vs {expected_count}")

        if errors:
            check("Tier Summary data accuracy", False,
                  f"{len(errors)} errors: {'; '.join(errors[:5])}")
        else:
            check("Tier Summary data accuracy", True)

    # --- Sheet 3: At Risk Customers ---
    print("  Checking At Risk Customers sheet...")
    a_rows = load_sheet_rows(agent_wb, "At Risk Customers")
    g_rows = load_sheet_rows(gt_wb, "At Risk Customers")

    if a_rows is None:
        check("At Risk Customers sheet exists", False, f"Sheets: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("At Risk Customers sheet in groundtruth", False)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        # Expected rows derived from DB at eval time (at-risk is time-dependent);
        # fallback to frozen GT if the DB is unavailable.
        try:
            e_data = get_at_risk_customers()
        except Exception as e:
            print(f"  WARNING: could not compute at-risk rows from DB: {e}")
            e_data = [(r[0], r[1], r[2], r[3]) for r in (g_rows[1:] if len(g_rows) > 1 else [])
                      if r and r[0] is not None]

        check("At Risk row count", abs(len(a_data) - len(e_data)) <= 3,
              f"Agent {len(a_data)} vs expected {len(e_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
            if row and len(row) > 1 and row[1] is not None:
                a_lookup[str(row[1]).strip().lower()] = row

        errors = []
        for e_row in e_data:
            if not e_row or e_row[0] is None:
                continue
            key = str(e_row[0]).strip().lower()
            email_key = str(e_row[1] or "").strip().lower()
            a_row = a_lookup.get(key) or a_lookup.get(email_key)
            if a_row is None:
                errors.append(f"Missing at-risk customer: {e_row[0]}")
                continue

            # Col 3: Total_Spent
            if len(a_row) > 3 and len(e_row) > 3:
                if not num_close(a_row[3], e_row[3], 5.0):
                    errors.append(f"{key}.Total_Spent: {a_row[3]} vs {e_row[3]}")

        if errors:
            check("At Risk data accuracy", False,
                  f"{len(errors)} errors: {'; '.join(errors[:5])}")
        else:
            check("At Risk data accuracy", True)

    return True


def _has_tier(text):
    t = (text or "").lower()
    return any(tier in t for tier in ["platinum", "gold", "silver", "bronze"])


def check_teamly():
    """Check Teamly 'Customer CRM' space holds one page per customer.

    The space is identified by key/name; customer pages live directly inside the
    space (flat one-page-per-customer layout). A hub + child-page layout is also
    accepted: if a hub page is detected, children are counted by parent_id, with
    a keyword fallback (tier/customer content) over all space pages.
    """
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("Teamly 'Customer CRM' space exists", False, f"DB error: {e}")
        check("Teamly CRM has >= 45 customer pages", False, f"DB error: {e}")
        return

    # Locate the Customer CRM space(s) by key or name. The agent may have
    # created its own space alongside the preprocess-seeded one, so collect
    # ALL matching spaces instead of an arbitrary LIMIT 1 pick.
    cur.execute(
        "SELECT id FROM teamly.spaces WHERE key = %s OR LOWER(name) = LOWER(%s) "
        "OR (LOWER(name) LIKE '%%customer%%' AND LOWER(name) LIKE '%%crm%%')",
        (SPACE_KEY, SPACE_NAME),
    )
    space_ids = [r[0] for r in cur.fetchall()]
    check("Teamly 'Customer CRM' space exists", bool(space_ids),
          "No space with key/name 'Customer CRM' found")

    if not space_ids:
        cur.close()
        conn.close()
        check("Teamly CRM has >= 45 customer pages", False, "space missing")
        return

    # User-created pages across all matching spaces (seeds have id <= 3).
    cur.execute(
        "SELECT id, title, body, parent_id FROM teamly.pages "
        "WHERE space_id = ANY(%s) AND id > 3",
        (space_ids,),
    )
    pages = cur.fetchall()
    cur.close()
    conn.close()

    # Detect an optional hub page (title says CRM/customers but is not a single
    # customer record). Count its direct children.
    hub_id = None
    for pid, title, body, parent_id in pages:
        tl = (title or "").lower()
        if ("crm" in tl or "customer crm" in tl) and "@" not in tl:
            hub_id = pid
            break

    child_pages = [p for p in pages if hub_id is not None and p[3] == hub_id and p[0] != hub_id]

    if len(child_pages) >= MIN_CUSTOMER_PAGES:
        customer_pages = child_pages
    else:
        # Flat layout / keyword fallback: pages (excluding hub) that look like a
        # customer record (mention a tier or an email address).
        customer_pages = [
            p for p in pages
            if p[0] != hub_id
            and (_has_tier((p[1] or "") + " " + (p[2] or "")) or "@" in (p[2] or ""))
        ]
        # Last resort: count all non-hub user pages in the space.
        if len(customer_pages) < MIN_CUSTOMER_PAGES:
            non_hub = [p for p in pages if p[0] != hub_id]
            if len(non_hub) >= MIN_CUSTOMER_PAGES:
                customer_pages = non_hub

    check("Teamly CRM has >= 45 customer pages", len(customer_pages) >= MIN_CUSTOMER_PAGES,
          f"Found {len(customer_pages)} customer pages (total {len(pages)}, children {len(child_pages)})")


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Get all sent emails
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()

    # Get expected at-risk emails
    at_risk_emails = get_at_risk_emails()

    # Filter for retention emails (not noise)
    retention_emails = []
    for subject, from_addr, to_addr, body_text in all_emails:
        subj_lower = (subject or "").lower()
        if "retention" in subj_lower or "customer" in subj_lower:
            # Parse to_addr
            recipients = set()
            if isinstance(to_addr, list):
                for r in to_addr:
                    recipients.add(str(r).strip().lower())
            elif isinstance(to_addr, str):
                try:
                    parsed = json.loads(to_addr)
                    if isinstance(parsed, list):
                        for r in parsed:
                            recipients.add(str(r).strip().lower())
                    else:
                        recipients.add(to_addr.strip().lower())
                except (json.JSONDecodeError, TypeError):
                    recipients.add(to_addr.strip().lower())
            retention_emails.append({
                "subject": subject,
                "from": from_addr,
                "to": recipients,
                "body": body_text or "",
            })

    print(f"  Found {len(retention_emails)} retention emails out of {len(all_emails)} total")
    print(f"  Expected at-risk customers: {len(at_risk_emails)}")

    # Check that we have a reasonable number of retention emails
    check("Retention emails sent (>= 50% of at-risk)",
          len(retention_emails) >= len(at_risk_emails) * 0.5,
          f"Got {len(retention_emails)}, expected >= {int(len(at_risk_emails) * 0.5)}")

    # Check that at-risk email addresses received emails
    matched = 0
    all_recipients = set()
    for em in retention_emails:
        all_recipients.update(em["to"])
    for ar_email in at_risk_emails:
        if ar_email in all_recipients:
            matched += 1

    check("At-risk customers received emails (>= 50%)",
          matched >= len(at_risk_emails) * 0.5,
          f"Matched {matched} out of {len(at_risk_emails)} at-risk customers")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_teamly()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "total_passed": PASS_COUNT,
            "total_checks": total,
            "accuracy": accuracy,
            "critical_failed": critical_failed,
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
