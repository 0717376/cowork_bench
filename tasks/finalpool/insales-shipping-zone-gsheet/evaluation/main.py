"""Evaluation for insales-shipping-zone-gsheet (InSales)."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

CRITICAL_FAILURES = []


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def zone_lookup(rows):
    """Return {zone_name_lower: row} from data rows (header skipped)."""
    data = rows[1:] if rows and len(rows) > 1 else []
    out = {}
    for row in data:
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row
    return out


def get_gsheet_summary():
    """Read the 'Shipping Zone Dashboard' spreadsheet 'Summary' sheet from gsheet.*.
    Returns {zone_name_lower: [zone, order_count, total_revenue, avg]} or None."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gsheet.spreadsheets ORDER BY created_at DESC")
    spreadsheets = cur.fetchall()
    ss = None
    for ss_id, ss_title in spreadsheets:
        if "shipping zone dashboard" in (ss_title or "").strip().lower():
            ss = (ss_id, ss_title)
            break
    if ss is None and spreadsheets:
        ss = spreadsheets[0]
    if ss is None:
        cur.close(); conn.close()
        return None, None
    ss_id = ss[0]
    cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id=%s ORDER BY index", (ss_id,))
    sheets = cur.fetchall()
    summary_sid = None
    for sid, stitle in sheets:
        if (stitle or "").strip().lower() == "summary":
            summary_sid = sid
            break
    if summary_sid is None:
        cur.close(); conn.close()
        return ss[1], None
    cur.execute(
        "SELECT row_index, col_index, value FROM gsheet.cells WHERE spreadsheet_id=%s AND sheet_id=%s ORDER BY row_index, col_index",
        (ss_id, summary_sid),
    )
    cells = {}
    for r, c, v in cur.fetchall():
        cells.setdefault(r, {})[c] = v
    cur.close(); conn.close()
    rows = []
    for r in sorted(cells.keys()):
        maxc = max(cells[r].keys())
        rows.append([cells[r].get(c) for c in range(maxc + 1)])
    return ss[1], rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Shipping_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Shipping_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # ---- Zone Summary (Excel) -------------------------------------------------
    print("  Checking Zone Summary (Excel)...")
    a_rows = load_sheet_rows(agent_wb, "Zone Summary")
    g_rows = load_sheet_rows(gt_wb, "Zone Summary")
    a_zone = {}
    if a_rows is None:
        all_errors.append("Sheet 'Zone Summary' not found in agent output")
        CRITICAL_FAILURES.append("zone_summary_missing")
    elif g_rows is None:
        all_errors.append("Sheet 'Zone Summary' not found in groundtruth")
    else:
        a_zone = zone_lookup(a_rows)
        g_zone = zone_lookup(g_rows)

        # Identify GT zones by their numbers (sync-safe: read RU names from GT).
        primary_key = None   # the zone with real orders (Domestic US -> Доставка по РФ)
        zero_keys = []       # zones with 0 orders (California/International)
        for k, g_row in g_zone.items():
            oc = to_num(g_row[1]) if len(g_row) > 1 else None
            if oc is not None and oc > 0:
                primary_key = k
            else:
                zero_keys.append(k)

        for key, g_row in g_zone.items():
            a_row = a_zone.get(key)
            if a_row is None:
                all_errors.append(f"Missing zone: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 5):
                all_errors.append(f"{key}.Order_Count: {a_row[1]} vs {g_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 500.0):
                all_errors.append(f"{key}.Total_Revenue: {a_row[2]} vs {g_row[2]} (tol=500)")
            if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 10.0):
                all_errors.append(f"{key}.Avg_Order_Value: {a_row[3]} vs {g_row[3]} (tol=10)")

        # CRITICAL 1: primary zone (Domestic US bucket) Order_Count + Total_Revenue.
        if primary_key is not None:
            g_row = g_zone[primary_key]
            a_row = a_zone.get(primary_key)
            ok = (
                a_row is not None
                and len(a_row) > 2
                and num_close(a_row[1], g_row[1], 5)
                and num_close(a_row[2], g_row[2], 100.0)
            )
            if not ok:
                CRITICAL_FAILURES.append(
                    f"primary_zone '{g_row[0]}' count/revenue "
                    f"(got {a_row[1:3] if a_row else None}, want {g_row[1:3]} tol count=5 rev=100)"
                )
        else:
            CRITICAL_FAILURES.append("no primary (non-zero) zone found in groundtruth")

        # CRITICAL 2: all three zones present; zero-order zones report 0/0.
        for key, g_row in g_zone.items():
            if key not in a_zone:
                CRITICAL_FAILURES.append(f"zone '{g_row[0]}' missing from Zone Summary")
        for key in zero_keys:
            a_row = a_zone.get(key)
            if a_row is not None:
                if not (num_close(a_row[1], 0, 0.001) and num_close(a_row[2], 0, 0.001)):
                    CRITICAL_FAILURES.append(
                        f"zero-order zone '{g_zone[key][0]}' must be 0/0, got {a_row[1:3]}"
                    )

    # ---- Methods (Excel) ------------------------------------------------------
    print("  Checking Methods (Excel)...")
    a_rows = load_sheet_rows(agent_wb, "Methods")
    g_rows = load_sheet_rows(gt_wb, "Methods")
    if a_rows is None:
        all_errors.append("Sheet 'Methods' not found in agent output")
        CRITICAL_FAILURES.append("methods_sheet_missing")
    elif g_rows is None:
        all_errors.append("Sheet 'Methods' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        if len(a_data) < len(g_data):
            all_errors.append(f"Methods: expected {len(g_data)} rows, got {len(a_data)}")

        # CRITICAL 3: method-identifier multiset (col index 1) must equal GT.
        def method_multiset(data):
            from collections import Counter
            return Counter(
                str(r[1]).strip().lower()
                for r in data
                if r and len(r) > 1 and r[1] is not None
            )
        a_ms = method_multiset(a_data)
        g_ms = method_multiset(g_data)
        if a_ms != g_ms:
            CRITICAL_FAILURES.append(
                f"Methods method-id multiset mismatch: got {dict(a_ms)}, want {dict(g_ms)}"
            )

    # ---- Google Sheet 'Shipping Zone Dashboard' / 'Summary' -------------------
    print("  Checking Google Sheet 'Shipping Zone Dashboard'...")
    try:
        ss_title, gs_rows = get_gsheet_summary()
    except Exception as e:
        ss_title, gs_rows = None, None
        all_errors.append(f"gsheet read error: {e}")

    if ss_title is None or "shipping zone dashboard" not in (ss_title or "").strip().lower():
        CRITICAL_FAILURES.append("Google Sheet 'Shipping Zone Dashboard' not found")
    elif gs_rows is None:
        CRITICAL_FAILURES.append("Google Sheet 'Summary' sheet not found")
    else:
        gs_zone = zone_lookup(gs_rows)
        # CRITICAL 4: gsheet Summary zone numbers match the Excel Zone Summary.
        if not gs_zone:
            CRITICAL_FAILURES.append("Google Sheet 'Summary' has no zone rows")
        else:
            for key, x_row in a_zone.items():
                g_row = gs_zone.get(key)
                if g_row is None:
                    CRITICAL_FAILURES.append(f"gsheet Summary missing zone '{x_row[0]}'")
                    continue
                if len(g_row) > 1 and len(x_row) > 1 and not num_close(g_row[1], x_row[1], 5):
                    CRITICAL_FAILURES.append(
                        f"gsheet '{key}'.Order_Count {g_row[1]} != Excel {x_row[1]}"
                    )
                if len(g_row) > 2 and len(x_row) > 2 and not num_close(g_row[2], x_row[2], 100.0):
                    CRITICAL_FAILURES.append(
                        f"gsheet '{key}'.Total_Revenue {g_row[2]} != Excel {x_row[2]}"
                    )

    # ---- Critical gate (before accuracy aggregation) --------------------------
    if CRITICAL_FAILURES:
        print("\n=== CRITICAL FAILURE(S) ===")
        for c in CRITICAL_FAILURES:
            print(f"  CRITICAL: {c}")
        sys.exit(1)

    # ---- Accuracy aggregation (non-critical structural/numeric details) -------
    total_checks = 6  # Domestic count, revenue, avg; two zero zones present; methods rows
    fail = len(all_errors)
    accuracy = max(0.0, 100.0 * (1 - fail / max(total_checks, 1)))
    if all_errors:
        print(f"\n=== {fail} non-critical error(s) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
    print(f"\nAccuracy: {accuracy:.1f}%")
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
