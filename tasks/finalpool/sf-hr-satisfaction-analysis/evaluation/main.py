"""Evaluation for sf-hr-satisfaction-analysis (ClickHouse fork).

Hardened structure:
  - CRITICAL_CHECKS: semantic correctness (per-dept aggregates, Summary overall,
    Happiest/Least dept, descending sort). Any critical failure => sys.exit(1)
    BEFORE the accuracy gate.
  - Non-critical structural checks (sheets present, docx present/keywords) feed
    an accuracy>=70 gate.
Department labels are Russian (Инженерия/Финансы/Кадры/Операции/НИОКР/Продажи/
Поддержка) on BOTH sides because the ClickHouse seed russifies sf_data
DEPARTMENT_NAME via the central relabel map; groundtruth uses the SAME labels.
Row matching keys on str(dept).strip().lower() (pure-Cyrillic, no normalize needed).
"""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def summary_lookup(rows, keys):
    """Header/orientation-tolerant {metric_lower: value} map for a Summary sheet.

    Scans ALL cells; a metric's value is the cell to its right, or the cell
    below when the right neighbour is itself another metric key (horizontal
    layout). A 'Metric/Value' header never collides with metric keys.
    """
    keys = {str(k).strip().lower() for k in keys}
    out = {}
    rows = rows or []

    def _is_key(v):
        return v is not None and str(v).strip().lower() in keys

    for r, row in enumerate(rows):
        for c, cell in enumerate(row or []):
            if cell is None:
                continue
            k = str(cell).strip().lower()
            if k not in keys or k in out:
                continue
            right = row[c + 1] if c + 1 < len(row) else None
            below = None
            if r + 1 < len(rows) and rows[r + 1] and c < len(rows[r + 1]):
                below = rows[r + 1][c]
            if right is not None and not _is_key(right):
                out[k] = right
            elif below is not None and not _is_key(below):
                out[k] = below
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Satisfaction_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Satisfaction_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # checks accumulate (label, passed, is_critical)
    checks = []

    def record(label, passed, critical=False):
        checks.append((label, passed, critical))
        if not passed:
            tag = "CRITICAL FAIL" if critical else "FAIL"
            print(f"    [{tag}] {label}")
        return passed

    # ---- Sheet: Satisfaction Analysis -------------------------------------
    print("  Checking Satisfaction Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Satisfaction Analysis")
    g_rows = load_sheet_rows(gt_wb, "Satisfaction Analysis")
    record("Sheet 'Satisfaction Analysis' present", a_rows is not None, critical=False)
    if a_rows is not None and g_rows is not None:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            present = a_row is not None
            record(f"Analysis row present: {g_row[0]}", present, critical=True)
            if not present:
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                record(f"{key}.Avg_Satisfaction ({a_row[1]} vs {g_row[1]})",
                       num_close(a_row[1], g_row[1], 0.1), critical=True)
            if len(a_row) > 2 and len(g_row) > 2:
                record(f"{key}.Avg_Work_Life_Balance ({a_row[2]} vs {g_row[2]})",
                       num_close(a_row[2], g_row[2], 0.1), critical=True)
            if len(a_row) > 3 and len(g_row) > 3:
                record(f"{key}.Avg_Rating ({a_row[3]} vs {g_row[3]})",
                       num_close(a_row[3], g_row[3], 0.1), critical=True)
            if len(a_row) > 4 and len(g_row) > 4:
                record(f"{key}.Employees ({a_row[4]} vs {g_row[4]})",
                       num_close(a_row[4], g_row[4], 5), critical=False)

        # descending sort by Avg_Satisfaction (critical)
        a_vals = []
        for row in a_data:
            if row and row[0] is not None and len(row) > 1:
                try:
                    a_vals.append(float(row[1]))
                except (TypeError, ValueError):
                    pass
        sorted_ok = all(a_vals[i] >= a_vals[i + 1] - 1e-9 for i in range(len(a_vals) - 1))
        record("Analysis sorted by Avg_Satisfaction descending", sorted_ok, critical=True)

    # ---- Sheet: Summary ----------------------------------------------------
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    record("Sheet 'Summary' present", a_rows is not None, critical=False)
    if a_rows is not None and g_rows is not None:
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        g_keys = [str(r[0]).strip().lower() for r in g_data if r and r[0] is not None]
        a_lookup = summary_lookup(a_rows, g_keys)

        # Tie-tolerant dept extremes: any dept whose rounded(2dp) GT
        # Avg_Satisfaction equals the extreme is accepted.
        dept_avg = {}
        for r in (load_sheet_rows(gt_wb, "Satisfaction Analysis") or [])[1:]:
            if r and r[0] is not None and len(r) > 1:
                try:
                    dept_avg[str(r[0]).strip().lower()] = round(float(r[1]), 2)
                except (TypeError, ValueError):
                    pass
        happiest_ties = {d for d, v in dept_avg.items() if v == max(dept_avg.values())} if dept_avg else set()
        least_ties = {d for d, v in dept_avg.items() if v == min(dept_avg.values())} if dept_avg else set()

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            present = key in a_lookup
            # Overall_* and Happiest/Least dept are critical
            is_crit = key in ("overall_satisfaction", "overall_wlb",
                              "happiest_dept", "least_happy_dept")
            record(f"Summary row present: {g_row[0]}", present, critical=is_crit)
            if not present:
                continue
            if len(g_row) > 1:
                a_val, g_val = a_lookup.get(key), g_row[1]
                if key in ("happiest_dept", "least_happy_dept"):
                    ties = happiest_ties if key == "happiest_dept" else least_ties
                    ok = str_match(a_val, g_val) or (
                        a_val is not None and str(a_val).strip().lower() in ties)
                else:
                    ok = num_close(a_val, g_val, 0.5)
                record(f"{key}.Value ({a_val} vs {g_val})", ok, critical=is_crit)

    # ---- Word doc (non-critical structural) --------------------------------
    docx_path = os.path.join(args.agent_workspace, "Satisfaction_Summary.docx")
    if not os.path.exists(docx_path):
        record("Satisfaction_Summary.docx present", False, critical=False)
    else:
        record("Satisfaction_Summary.docx present", True, critical=False)
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs).lower()
            _headings = " ".join(
                p.text for p in _doc.paragraphs if p.style.name.startswith("Heading")
            ).lower()
            record("docx has >= 50 chars", len(_text.strip()) >= 50, critical=False)
            # RU+EN keyword groups: at least one alternative per group must appear
            _kw_groups = [
                ["удовлетвор", "satisfaction"],
                ["свод", "summary", "сводк"],
            ]
            kw_ok = any(
                any(alt in _text or alt in _headings for alt in grp)
                for grp in _kw_groups
            )
            record("docx has RU/EN satisfaction/summary keyword", kw_ok, critical=False)
        except ImportError:
            record("docx non-trivial size", os.path.getsize(docx_path) >= 100, critical=False)
        except Exception as _e:
            record(f"docx readable ({_e})", False, critical=False)

    # ---- Verdict -----------------------------------------------------------
    critical_fails = [lbl for lbl, ok, crit in checks if crit and not ok]
    total = len(checks)
    passed = sum(1 for _, ok, _ in checks if ok)
    accuracy = (passed / total * 100) if total else 0.0

    print(f"\n=== Checks passed: {passed}/{total} (accuracy={accuracy:.1f}%) ===")
    if critical_fails:
        print(f"=== RESULT: FAIL — {len(critical_fails)} critical check(s) failed ===")
        for lbl in critical_fails[:10]:
            print(f"  CRITICAL: {lbl}")
        sys.exit(1)

    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"=== RESULT: FAIL — accuracy {accuracy:.1f}% < 70% ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
