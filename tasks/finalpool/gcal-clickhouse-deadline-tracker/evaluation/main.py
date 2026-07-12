"""
Evaluation script for gcal-clickhouse-deadline-tracker task.

Checks:
1. Excel SLA_Compliance_Audit.xlsx - "Breached Tickets" and "Summary" sheets
2. Google Calendar - SLA Review events for top-5 breaches on 2026-03-07
3. Teamly - knowledge-base page with SLA breach dashboard (March 2026)
4. Email - alert email to support-lead@company.com about SLA breaches

CRITICAL checks: any failure => overall FAIL regardless of accuracy.
Pass condition: no critical failure AND accuracy >= 70.
"""

import argparse
import json
import os
import sys
from decimal import Decimal

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Top 5 worst breach ticket IDs (for calendar/email checks).
# Derived from the current sf_data seed; numeric values in sf_data are NOT
# russified, so these English INC#### identifiers stay stable.
TOP5_TICKET_IDS = ["INC8275", "INC7833", "INC6422", "INC6854", "INC0925"]

# Critical checks: any failure here => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "CRITICAL: Breached Tickets row count matches DB",
    "CRITICAL: all TOP5 breach tickets present in Breached Tickets",
    "CRITICAL: Summary High Breached_Count matches DB",
    "CRITICAL: Summary High Breach_Rate_Pct matches DB",
    "CRITICAL: Summary Medium Breached_Count matches DB",
    "CRITICAL: Summary Medium Breach_Rate_Pct matches DB",
    "CRITICAL: all 5 TOP5 SLA Review events exist on 2026-03-07",
    "CRITICAL: SLA alert email to support-lead lists >=3 worst tickets",
    "CRITICAL: Teamly SLA breach page (March 2026) exists with content",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def int_close(a, b, tol=10):
    try:
        return abs(int(float(a)) - int(float(b))) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_summary():
    """Query PostgreSQL for expected summary data."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT t."PRIORITY",
          COUNT(*) as total_tickets,
          COUNT(*) FILTER (WHERE t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS") as breached_count,
          COUNT(*) FILTER (WHERE t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS" * 0.8
                           AND t."RESPONSE_TIME_HOURS" <= p."RESPONSE_TARGET_HOURS") as near_breach_count,
          COUNT(*) FILTER (WHERE t."RESPONSE_TIME_HOURS" <= p."RESPONSE_TARGET_HOURS" * 0.8) as compliant_count,
          ROUND(100.0 * COUNT(*) FILTER (WHERE t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS")
                / COUNT(*), 1) as breach_rate_pct,
          ROUND(AVG(t."RESPONSE_TIME_HOURS")::numeric, 2) as avg_response_hours,
          ROUND(AVG(t."CUSTOMER_SATISFACTION")::numeric, 2) as avg_csat
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
        JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" p ON t."PRIORITY" = p."PRIORITY"
        WHERE t."PRIORITY" IN ('High', 'Medium')
        GROUP BY t."PRIORITY"
        ORDER BY CASE t."PRIORITY" WHEN 'High' THEN 1 WHEN 'Medium' THEN 2 END
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def get_expected_breached_count():
    """Get expected number of breached tickets."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t
        JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" p ON t."PRIORITY" = p."PRIORITY"
        WHERE t."PRIORITY" IN ('High', 'Medium')
          AND t."RESPONSE_TIME_HOURS" > p."RESPONSE_TARGET_HOURS"
    """)
    count = cur.fetchone()[0]
    cur.close()
    conn.close()
    return count


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    """Check SLA_Compliance_Audit.xlsx content."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "SLA_Compliance_Audit.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "SLA_Compliance_Audit.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False
    record("Excel file readable", True)

    all_ok = True

    # --- Sheet 1: Breached Tickets ---
    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    ws1 = get_sheet(agent_wb, "Breached Tickets")
    if ws1 is None:
        record("Sheet 'Breached Tickets' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Breached Tickets' exists", True)

        # Check headers
        agent_headers = [str(c.value).strip() if c.value else "" for c in ws1[1]]
        expected_headers = [
            "Ticket_ID", "Priority", "Issue_Type", "Created_At",
            "Response_Time_Hours", "SLA_Target_Hours", "Hours_Over_SLA",
            "Customer_Satisfaction"
        ]
        headers_ok = all(
            a.lower().replace(" ", "_") == e.lower().replace(" ", "_")
            for a, e in zip(agent_headers[:len(expected_headers)], expected_headers)
        ) and len(agent_headers) >= len(expected_headers)
        record("Breached Tickets headers match", headers_ok,
               f"Expected: {expected_headers}, Got: {agent_headers}")

        # Check row count (tight: must equal DB-computed breached count)
        data_rows = [r for r in ws1.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in r)]
        expected_count = get_expected_breached_count()
        record("CRITICAL: Breached Tickets row count matches DB",
               int_close(len(data_rows), expected_count, 1),
               f"Expected {expected_count}, got {len(data_rows)}")

        # Check that top 5 tickets are present (single critical gate)
        ticket_ids = [str(r[0]).strip() if r[0] else "" for r in data_rows]
        missing_top = [tid for tid in TOP5_TICKET_IDS if tid not in ticket_ids]
        record("CRITICAL: all TOP5 breach tickets present in Breached Tickets",
               not missing_top,
               f"Missing {missing_top} among {len(ticket_ids)} rows")

        # Check sort order: High before Medium
        priorities = [str(r[1]).strip() if r[1] else "" for r in data_rows]
        high_indices = [i for i, p in enumerate(priorities) if p == "High"]
        medium_indices = [i for i, p in enumerate(priorities) if p == "Medium"]

        if high_indices and medium_indices:
            record("Sort: High before Medium",
                   max(high_indices) < min(medium_indices),
                   f"Last High at {max(high_indices)}, first Medium at {min(medium_indices)}")
        else:
            record("Sort: both priorities present",
                   bool(high_indices) and bool(medium_indices),
                   f"High: {len(high_indices)}, Medium: {len(medium_indices)}")

        # Check that Hours_Over_SLA is descending within each priority group
        high_overs = []
        medium_overs = []
        for r in data_rows:
            try:
                over = float(r[6])
                if str(r[1]).strip() == "High":
                    high_overs.append(over)
                elif str(r[1]).strip() == "Medium":
                    medium_overs.append(over)
            except (TypeError, ValueError, IndexError):
                pass

        if len(high_overs) > 1:
            is_desc = all(high_overs[i] >= high_overs[i+1] - 0.01
                         for i in range(min(20, len(high_overs) - 1)))
            record("High tickets sorted by Hours_Over_SLA desc", is_desc,
                   f"First 5: {high_overs[:5]}")

        # Spot-check top ticket data
        if data_rows and data_rows[0][0]:
            top_ticket = data_rows[0]
            record("Top breach ticket is INC8275",
                   str(top_ticket[0]).strip() == "INC8275",
                   f"Got {top_ticket[0]}")
            if str(top_ticket[0]).strip() == "INC8275":
                record("INC8275 response time ~13.48",
                       num_close(top_ticket[4], 13.48, 0.1),
                       f"Got {top_ticket[4]}")
                record("INC8275 hours over SLA ~9.48",
                       num_close(top_ticket[6], 9.48, 0.1),
                       f"Got {top_ticket[6]}")

    # --- Sheet 2: Summary ---
    ws2 = get_sheet(agent_wb, "Summary")
    if ws2 is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)

        summary_rows = list(ws2.iter_rows(min_row=2, values_only=True))
        record("Summary has 2 data rows", len(summary_rows) == 2,
               f"Got {len(summary_rows)}")

        expected_summary = get_expected_summary()

        for i, expected in enumerate(expected_summary):
            if i >= len(summary_rows):
                record(f"Summary row for {expected[0]}", False, "Missing")
                all_ok = False
                continue

            agent_row = summary_rows[i]
            priority = expected[0]

            record(f"Summary {priority}: Priority label",
                   str(agent_row[0]).strip() == priority,
                   f"Expected {priority}, got {agent_row[0]}")

            # Total_Tickets (tight)
            record(f"Summary {priority}: Total_Tickets",
                   int_close(agent_row[1], expected[1], 1),
                   f"Expected {expected[1]}, got {agent_row[1]}")

            # Breached_Count (CRITICAL, exact)
            record(f"CRITICAL: Summary {priority} Breached_Count matches DB",
                   int_close(agent_row[2], expected[2], 0),
                   f"Expected {expected[2]}, got {agent_row[2]}")

            # Breach_Rate_Pct (CRITICAL, tight +/-0.2)
            record(f"CRITICAL: Summary {priority} Breach_Rate_Pct matches DB",
                   num_close(agent_row[5], float(expected[5]), 0.2),
                   f"Expected {expected[5]}, got {agent_row[5]}")

            # Avg_Response_Hours
            record(f"Summary {priority}: Avg_Response_Hours",
                   num_close(agent_row[6], float(expected[6]), 1.0),
                   f"Expected {expected[6]}, got {agent_row[6]}")

    return all_ok


# ============================================================================
# Check 2: Google Calendar events
# ============================================================================

def check_gcal():
    """Verify SLA Review events created for top 5 breaches."""
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY summary
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  [INFO] Found {len(events)} calendar events.")

    record("At least 5 calendar events created", len(events) >= 5,
           f"Found {len(events)}")

    all_ok = True
    correct_top5 = 0  # events titled with the ticket ID AND dated 2026-03-07

    for tid in TOP5_TICKET_IDS:
        found = False
        date_ok = False
        for summary, description, start_dt, end_dt in events:
            summary_str = str(summary or "").lower()
            if tid.lower() in summary_str and "sla" in summary_str:
                found = True

                # Check date is 2026-03-07
                if start_dt is not None:
                    start_date_str = start_dt.strftime("%Y-%m-%d")
                    date_ok = start_date_str == "2026-03-07"
                    record(f"gcal {tid}: date is 2026-03-07",
                           date_ok,
                           f"Got {start_date_str}")
                else:
                    record(f"gcal {tid}: has start datetime", False,
                           "start_dt is None")

                # Check description mentions ticket ID
                desc_str = str(description or "").lower()
                record(f"gcal {tid}: description mentions ticket",
                       tid.lower() in desc_str,
                       f"Description: {description[:100] if description else 'None'}")

                break

        if not found:
            # Also check for events with just the ticket ID
            for summary, description, start_dt, end_dt in events:
                if tid.lower() in str(summary or "").lower():
                    found = True
                    if start_dt is not None:
                        date_ok = start_dt.strftime("%Y-%m-%d") == "2026-03-07"
                    break

            record(f"gcal {tid}: event exists", found,
                   f"No event containing '{tid}' found")
            if not found:
                all_ok = False

        if found and date_ok:
            correct_top5 += 1

    record("CRITICAL: all 5 TOP5 SLA Review events exist on 2026-03-07",
           correct_top5 == len(TOP5_TICKET_IDS),
           f"{correct_top5}/{len(TOP5_TICKET_IDS)} TOP5 events present on 2026-03-07")

    return all_ok


# ============================================================================
# Check 3: Teamly knowledge-base page
# ============================================================================

def check_teamly():
    """Verify the SLA breach dashboard page exists in the Teamly KB.

    Seed pages have id <= 3 and are pre-existing noise; only user/agent-created
    pages (id > 3) may satisfy this. Title may be written in Russian or English,
    so keyword matching covers both. Content must be non-trivial.
    """
    print("\n=== Checking Teamly ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
    pages = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  [INFO] Found {len(pages)} agent-created Teamly pages.")

    # Title keywords: SLA/breach/нарушен plus dashboard/audit/report/march/март.
    found_page = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        topic = ("sla" in tl or "breach" in tl or "нарушен" in tl)
        kind = ("dashboard" in tl or "report" in tl or "audit" in tl
                or "march" in tl or "март" in tl or "дашборд" in tl
                or "отчёт" in tl or "отчет" in tl or "аудит" in tl)
        if topic and kind:
            found_page = (pid, title, body)
            break

    record("CRITICAL: Teamly SLA breach page (March 2026) exists with content",
           found_page is not None and len((found_page[2] or "").strip()) > 0,
           f"pages: {[(p[0], p[1]) for p in pages]}")

    if found_page is None:
        return False

    # Content substance (non-critical): mentions priority breakdown / CSAT.
    text = ((found_page[1] or "") + " " + (found_page[2] or "")).lower()
    record("Teamly page mentions priority breakdown (High/Medium)",
           ("high" in text or "medium" in text or "приоритет" in text),
           "no priority mention")
    record("Teamly page mentions breach counts or rate",
           ("breach" in text or "нарушен" in text or "rate" in text or "доля" in text),
           "no breach stats mention")
    record("Teamly page mentions customer satisfaction (CSAT)",
           ("csat" in text or "satisfaction" in text or "удовлетвор" in text),
           "no CSAT mention")

    return True


# ============================================================================
# Check 4: Email
# ============================================================================

def check_email():
    """Verify alert email sent to support-lead@company.com."""
    print("\n=== Checking Email ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  [INFO] Found {len(all_emails)} total emails.")

    record("At least 1 email sent", len(all_emails) >= 1,
           f"Found {len(all_emails)}")

    all_ok = True
    found = False

    for subject, from_addr, to_addr, body_text in all_emails:
        subject_lower = (subject or "").lower()
        if "sla" in subject_lower and ("breach" in subject_lower or "compliance" in subject_lower or "alert" in subject_lower):
            found = True

            # Check recipient
            to_str = ""
            if isinstance(to_addr, list):
                to_str = " ".join(str(r).lower() for r in to_addr)
            elif isinstance(to_addr, str):
                try:
                    parsed = json.loads(to_addr)
                    if isinstance(parsed, list):
                        to_str = " ".join(str(r).lower() for r in parsed)
                    else:
                        to_str = str(to_addr).lower()
                except (json.JSONDecodeError, TypeError):
                    to_str = str(to_addr).lower()

            record("Email sent to support-lead@company.com",
                   "support-lead@company.com" in to_str,
                   f"To: {to_addr}")

            # Check body mentions priority breakdown (RU or EN)
            body_lower = (body_text or "").lower()
            record("Email body mentions High priority",
                   "high" in body_lower or "приоритет" in body_lower,
                   "No mention of 'High'/'приоритет' in body")

            record("Email body mentions breach count or rate",
                   "breach" in body_lower or "нарушен" in body_lower
                   or "rate" in body_lower or "доля" in body_lower,
                   "No breach statistics found")

            # Check body lists worst offending ticket IDs (CRITICAL)
            top_mentioned = sum(1 for tid in TOP5_TICKET_IDS
                               if tid in (body_text or ""))
            record("CRITICAL: SLA alert email to support-lead lists >=3 worst tickets",
                   top_mentioned >= 3,
                   f"Found {top_mentioned} of 5 top ticket IDs")

            break

    if not found:
        record("SLA alert email exists", False,
               "No email with 'SLA' + 'breach/compliance/alert' in subject")
        # No matching email => the worst-tickets critical check could not run; fail it.
        record("CRITICAL: SLA alert email to support-lead lists >=3 worst tickets",
               False, "No SLA alert email found")
        all_ok = False
    else:
        record("SLA alert email exists", True)

    return all_ok


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)
    gcal_ok = check_gcal()
    teamly_ok = check_teamly()
    email_ok = check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    all_passed = (not critical_failed) and accuracy >= 70

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:   {'PASS' if excel_ok else 'FAIL'}")
    print(f"  GCal:    {'PASS' if gcal_ok else 'FAIL'}")
    print(f"  Teamly:  {'PASS' if teamly_ok else 'FAIL'}")
    print(f"  Email:   {'PASS' if email_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"    - {n}")
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "success": all_passed,
            "critical_failed": critical_failed,
            "details": {
                "excel": excel_ok,
                "gcal": gcal_ok,
                "teamly": teamly_ok,
                "email": email_ok,
            },
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if all_passed else 1)


if __name__ == "__main__":
    main()
