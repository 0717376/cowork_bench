"""Evaluation script for fetch-kulinar-catering-excel-gcal-email (russified).

Critical checks (CRITICAL_CHECKS): any failure there => overall FAIL regardless
of accuracy. Otherwise pass threshold: accuracy >= 70% (AND no critical fail).
"""
import os
import argparse, json, os, sys
import openpyxl


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

# --- Kulinar recipe catalog (name -> (unique_ingredient_count, is_vegetarian)).
# Source of truth: local_servers/kulinar-mcp/src/data/all_recipes.json.
# is_vegetarian = no meat/poultry/fish tag AND no meat/fish ingredient (eggs &
# dairy count as vegetarian). Embedded statically so eval is self-contained; if
# the live JSON is found it overrides this map.
CATALOG = {
    "Салат Оливье": (10, False),
    "Винегрет": (8, True),
    "Сельдь под шубой": (7, False),
    "Салат Мимоза": (6, False),
    "Крабовый салат": (7, False),
    "Греческий салат": (10, True),
    "Салат с курицей и грибами": (8, False),
    "Холодец": (8, False),
    "Икра кабачковая": (8, True),
    "Грибы маринованные": (9, True),
    "Сало солёное": (5, False),
    "Селёдка с луком": (6, False),
    "Борщ": (12, False),
    "Щи из квашеной капусты": (9, False),
    "Солянка мясная": (11, False),
    "Уха": (8, False),
    "Окрошка": (11, False),
    "Грибной суп": (8, True),
    "Рассольник": (9, False),
    "Куриный бульон с лапшой": (7, False),
    "Бефстроганов": (8, False),
    "Пельмени домашние": (8, False),
    "Голубцы": (9, False),
    "Котлеты домашние": (9, False),
    "Жаркое в горшочках": (9, False),
    "Курица в сметане": (8, False),
    "Рыба запечённая по-русски": (9, False),
    "Цыплёнок табака": (6, False),
    "Гречка с тушёнкой": (7, False),
    "Плов узбекский": (9, False),
    "Картофельное пюре": (4, True),
    "Гречневая каша": (4, True),
    "Перловая каша": (5, True),
    "Картофель отварной с укропом": (5, True),
    "Рис отварной": (4, True),
    "Пирожки с капустой жареные": (9, True),
    "Пирожки с мясом печёные": (9, False),
    "Блины тонкие": (6, True),
    "Кулебяка с капустой и яйцом": (8, True),
    "Расстегаи с рыбой": (9, False),
    "Медовик": (8, True),
    "Наполеон": (8, True),
    "Сырники": (7, True),
    "Пасха творожная": (8, True),
    "Ватрушки с творогом": (8, True),
    "Кисель ягодный": (4, True),
    "Морс клюквенный": (3, True),
    "Компот из сухофруктов": (3, True),
    "Сбитень": (7, True),
    "Квас домашний": (5, True),
}

_MEAT_TAGS = {"мясное", "рыбное", "куриное", "говядина"}
_NONVEG = [
    "мясо", "мясн", "говяд", "свин", "баран", "курица", "куриц", "куриный",
    "куриная", "цыпл", "индейк", "птиц", "фарш", "колбас", "сосиск", "ветчин",
    "бекон", "грудинк", "окорок", "тушёнк", "тушенк", "холодец", "язык говяж",
    "рыб", "сельд", "селёдк", "селедк", "сайр", "горбуш", "лосос", "форел",
    "треск", "краб", "креветк", "кальмар", "моллюск", "икра красн", "шпрот",
    "тунец", "сало",
]


def _ingredient_is_meat(name):
    nm = name.lower()
    if "яйц" in nm or "желток" in nm or "белок" in nm or "икра кабачк" in nm:
        return False
    return any(kw in nm for kw in _NONVEG)


def _load_live_catalog():
    """Override CATALOG from the live kulinar JSON if reachable (best-effort)."""
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = []
    cur = here
    for _ in range(7):
        cur = os.path.dirname(cur)
        candidates.append(os.path.join(
            cur, "local_servers", "kulinar-mcp", "src", "data", "all_recipes.json"))
    for path in candidates:
        if os.path.exists(path):
            try:
                recs = json.load(open(path, encoding="utf-8"))
                cat = {}
                for r in recs:
                    ic = len({i["name"] for i in r["ingredients"]})
                    veg = not (set(r.get("tags", [])) & _MEAT_TAGS) and not any(
                        _ingredient_is_meat(i["name"]) for i in r["ingredients"])
                    cat[r["name"].strip()] = (ic, veg)
                if cat:
                    return cat
            except Exception:
                return None
    return None


_live = _load_live_catalog()
if _live:
    CATALOG = _live

# Normalize keys for tolerant lookup (strip, collapse spaces).
CATALOG_NORM = {" ".join(k.split()).lower(): v for k, v in CATALOG.items()}


def catalog_lookup(name):
    if name is None:
        return None
    key = " ".join(str(name).split()).lower()
    return CATALOG_NORM.get(key)


PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic (critical) checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Wellness_Menu_Plan.xlsx exists",
    "Daily Menu sheet exists",
    "Daily Menu has 10 rows",
    "Days sorted Monday-Friday",
    "At least 2 vegetarian options",
    "Recipe names exist in kulinar catalog",
    "Ingredient_Count matches kulinar catalog",
    "Vegetarian flags marked Yes are truly vegetarian",
    "Estimated_Cost = Ingredient_Count * 0.50 * 50",
    "Budget Overview: Total_Budget is 2000",
    "Budget Overview arithmetic consistent",
    "5 meal prep calendar events",
    "Calendar events on March 16-20",
    "All calendar events 07:00-08:00 with correct day suffix",
    "Vendor email body has ingredient list and total cost",
    "Committee email body has daily menu and budget summary",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:300] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


DAY_ORDER = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    # --- Check Excel file ---
    excel_path = os.path.join(agent_workspace, "Wellness_Menu_Plan.xlsx")
    check("Wellness_Menu_Plan.xlsx exists", os.path.exists(excel_path))

    daily_total_cost = None  # computed from Daily Menu for cross-sheet checks

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # Sheet 1: Daily Menu
        check("Daily Menu sheet exists", "Daily Menu" in wb.sheetnames)
        if "Daily Menu" in wb.sheetnames:
            ws = wb["Daily Menu"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [r for r in data_rows if any(c is not None for c in r)]

            check("Daily Menu has 10 rows", len(data_rows) == 10, f"got {len(data_rows)}")

            for col in ["Day", "Meal_Type", "Recipe_Name", "Servings", "Ingredient_Count",
                        "Estimated_Cost", "Is_Vegetarian"]:
                check(f"Daily Menu has {col} column",
                      col in headers, f"headers: {headers}")

            # Verify days are sorted Monday-Friday (2 per day)
            if "Day" in headers and len(data_rows) == 10:
                day_col = headers.index("Day")
                days = [str(r[day_col]) for r in data_rows]
                expected_days = [d for d in DAY_ORDER for _ in range(2)]
                check("Days sorted Monday-Friday", days == expected_days, f"got {days}")

            # Verify servings are all 50
            if "Servings" in headers:
                serv_col = headers.index("Servings")
                servings = [safe_float(r[serv_col]) for r in data_rows]
                check("All servings are 50",
                      all(s == 50 for s in servings if s is not None),
                      f"servings: {servings}")

            # Verify Is_Vegetarian has at least 2 Yes values
            veg_col = headers.index("Is_Vegetarian") if "Is_Vegetarian" in headers else None
            name_col = headers.index("Recipe_Name") if "Recipe_Name" in headers else None
            if veg_col is not None:
                veg_vals = [str(r[veg_col]).strip() for r in data_rows if r[veg_col] is not None]
                yes_count = sum(1 for v in veg_vals if v.lower() in ("yes", "да"))
                check("At least 2 vegetarian options", yes_count >= 2,
                      f"found {yes_count} vegetarian")

            # CRITICAL: recipe names must exist in the kulinar catalog
            if name_col is not None:
                names = [str(r[name_col]).strip() for r in data_rows if r[name_col]]
                unknown = [n for n in names if catalog_lookup(n) is None]
                check("Recipe names exist in kulinar catalog", len(unknown) == 0,
                      f"unknown: {unknown}")

            # CRITICAL: Ingredient_Count per row matches the real catalog count
            ic_col = headers.index("Ingredient_Count") if "Ingredient_Count" in headers else None
            if name_col is not None and ic_col is not None:
                ic_ok = True
                bad = []
                for r in data_rows:
                    nm = str(r[name_col]).strip() if r[name_col] else None
                    info = catalog_lookup(nm)
                    ic = safe_float(r[ic_col])
                    if info is None or ic is None or int(round(ic)) != info[0]:
                        ic_ok = False
                        bad.append((nm, ic, info[0] if info else None))
                check("Ingredient_Count matches kulinar catalog", ic_ok, f"mismatches: {bad}")

            # CRITICAL: rows marked vegetarian must actually be vegetarian per catalog
            if name_col is not None and veg_col is not None:
                veg_ok = True
                bad = []
                for r in data_rows:
                    nm = str(r[name_col]).strip() if r[name_col] else None
                    info = catalog_lookup(nm)
                    flag = str(r[veg_col]).strip().lower() if r[veg_col] is not None else ""
                    if flag in ("yes", "да") and info is not None and not info[1]:
                        veg_ok = False
                        bad.append(nm)
                check("Vegetarian flags marked Yes are truly vegetarian", veg_ok,
                      f"non-veg marked Yes: {bad}")

            # CRITICAL: cost formula Ingredient_Count * 0.50 * 50
            ec_col = headers.index("Estimated_Cost") if "Estimated_Cost" in headers else None
            if ic_col is not None and ec_col is not None:
                cost_ok = True
                running = 0.0
                for row in data_rows:
                    ic = safe_float(row[ic_col])
                    ec = safe_float(row[ec_col])
                    if ic is not None and ec is not None:
                        expected = round(ic * 0.50 * 50, 2)
                        running += ec
                        if abs(ec - expected) > 0.01:
                            cost_ok = False
                check("Estimated_Cost = Ingredient_Count * 0.50 * 50", cost_ok)
                daily_total_cost = round(running, 2)

        # Sheet 2: Ingredient Summary
        check("Ingredient Summary sheet exists", "Ingredient Summary" in wb.sheetnames)
        if "Ingredient Summary" in wb.sheetnames:
            ws = wb["Ingredient Summary"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [r for r in data_rows if any(c is not None for c in r)]

            check("Ingredient Summary has >= 15 rows", len(data_rows) >= 15,
                  f"got {len(data_rows)}")

            for col in ["Ingredient_Name", "Total_Quantity", "Unit", "Times_Used", "Total_Cost"]:
                check(f"Ingredient Summary has {col} column",
                      col in headers, f"headers: {headers}")

            # Verify sorted alphabetically
            if "Ingredient_Name" in headers:
                name_col = headers.index("Ingredient_Name")
                names = [str(r[name_col]) for r in data_rows if r[name_col]]
                check("Ingredients sorted alphabetically",
                      names == sorted(names), f"first few: {names[:5]}")

            # Verify cost formula: Times_Used * 0.50 * 50
            if "Times_Used" in headers and "Total_Cost" in headers:
                tu_col = headers.index("Times_Used")
                tc_col = headers.index("Total_Cost")
                cost_ok = True
                for row in data_rows:
                    tu = safe_float(row[tu_col])
                    tc = safe_float(row[tc_col])
                    if tu is not None and tc is not None:
                        expected = round(tu * 0.50 * 50, 2)
                        if abs(tc - expected) > 0.01:
                            cost_ok = False
                            break
                check("Ingredient Total_Cost = Times_Used * 0.50 * 50", cost_ok)

        # Sheet 3: Budget Overview
        check("Budget Overview sheet exists", "Budget Overview" in wb.sheetnames)
        if "Budget Overview" in wb.sheetnames:
            ws = wb["Budget Overview"]
            headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [r for r in data_rows if any(c is not None for c in r)]

            check("Budget Overview has 6 rows", len(data_rows) == 6, f"got {len(data_rows)}")

            for col in ["Label", "Value"]:
                check(f"Budget Overview has {col} column",
                      col in headers, f"headers: {headers}")

            label_col = headers.index("Label") if "Label" in headers else 0
            value_col = headers.index("Value") if "Value" in headers else 1
            budget_map = {}
            for row in data_rows:
                if row[label_col]:
                    budget_map[str(row[label_col]).strip()] = safe_float(row[value_col])

            check("Has Total_Budget label", "Total_Budget" in budget_map)
            check("Budget Overview: Total_Budget is 2000",
                  budget_map.get("Total_Budget") == 2000.0,
                  f"got {budget_map.get('Total_Budget')}")

            tc = budget_map.get("Total_Estimated_Cost")
            br = budget_map.get("Budget_Remaining")
            acd = budget_map.get("Avg_Cost_Per_Day")
            acm = budget_map.get("Avg_Cost_Per_Meal")

            # CRITICAL cross-sheet/internal arithmetic consistency
            arith_ok = True
            details = []
            if tc is None:
                arith_ok = False
                details.append("no Total_Estimated_Cost")
            else:
                if daily_total_cost is not None and abs(tc - daily_total_cost) > 0.05:
                    arith_ok = False
                    details.append(f"TEC {tc} != sum of Daily Menu {daily_total_cost}")
                if br is None or abs(br - (2000.0 - tc)) > 0.05:
                    arith_ok = False
                    details.append(f"Budget_Remaining {br} != {round(2000.0 - tc, 2)}")
                if acd is None or abs(acd - tc / 5) > 0.05:
                    arith_ok = False
                    details.append(f"Avg_Cost_Per_Day {acd} != {round(tc / 5, 2)}")
                if acm is None or abs(acm - tc / 10) > 0.05:
                    arith_ok = False
                    details.append(f"Avg_Cost_Per_Meal {acm} != {round(tc / 10, 2)}")
            check("Budget Overview arithmetic consistent", arith_ok, "; ".join(details))

    # --- Check Calendar Events ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE summary LIKE '%Wellness Week Meal Prep%'
            ORDER BY start_datetime
        """)
        cal_rows = cur.fetchall()
        cur.close()
        conn.close()

        check("5 meal prep calendar events", len(cal_rows) == 5,
              f"found {len(cal_rows)}")

        if len(cal_rows) == 5:
            dates = [str(r[1])[:10] for r in cal_rows]
            expected_dates = ["2026-03-16", "2026-03-17", "2026-03-18",
                              "2026-03-19", "2026-03-20"]
            check("Calendar events on March 16-20",
                  dates == expected_dates, f"dates: {dates}")

            # CRITICAL: ALL five events 07:00-08:00 with correct Monday..Friday suffix
            all_ok = True
            details = []
            for idx, row in enumerate(cal_rows):
                summary = str(row[0])
                start_h = str(row[1])[11:16]
                end_h = str(row[2])[11:16]
                day_name = DAY_ORDER[idx] if idx < 5 else "?"
                if start_h != "07:00":
                    all_ok = False
                    details.append(f"{summary}: start {start_h}")
                if end_h != "08:00":
                    all_ok = False
                    details.append(f"{summary}: end {end_h}")
                if not summary.strip().endswith(day_name):
                    all_ok = False
                    details.append(f"'{summary}' should end with {day_name}")
            check("All calendar events 07:00-08:00 with correct day suffix",
                  all_ok, "; ".join(details))
        else:
            check("Calendar events on March 16-20", False, "wrong count")
            check("All calendar events 07:00-08:00 with correct day suffix", False,
                  "wrong count")
    except Exception as e:
        check("5 meal prep calendar events", False, str(e))
        check("Calendar events on March 16-20", False, str(e))
        check("All calendar events 07:00-08:00 with correct day suffix", False, str(e))

    # --- Check Emails ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr, subject, body_text
            FROM email.messages
            WHERE subject LIKE '%Wellness Week%'
            ORDER BY subject
        """)
        email_rows = cur.fetchall()
        cur.close()
        conn.close()

        check("At least 2 wellness emails sent", len(email_rows) >= 2,
              f"found {len(email_rows)}")

        recipients = [str(r[0]) for r in email_rows]
        subjects = [str(r[1]) for r in email_rows]

        vendor_found = any("catering_vendor" in r for r in recipients)
        check("Email sent to catering_vendor@company.com", vendor_found,
              f"recipients: {recipients}")

        committee_found = any("wellness_committee" in r for r in recipients)
        check("Email sent to wellness_committee@company.com", committee_found,
              f"recipients: {recipients}")

        ingredient_subj = any("ingredient" in s.lower() for s in subjects)
        check("Vendor email has ingredient-related subject", ingredient_subj,
              f"subjects: {subjects}")

        menu_subj = any("menu" in s.lower() for s in subjects)
        check("Committee email has menu-related subject", menu_subj,
              f"subjects: {subjects}")

        def body_for(recipient_substr):
            for r in email_rows:
                if recipient_substr in str(r[0]):
                    return str(r[2] or "")
            return ""

        # CRITICAL: vendor email body must list ingredients + total estimated cost.
        vb = body_for("catering_vendor").lower()
        # count distinct kulinar ingredient names that appear in the body
        ing_hits = 0
        seen = set()
        if _live is None:
            ing_names = set()  # fall back to keyword heuristic below
        else:
            here = os.path.dirname(os.path.abspath(__file__))
            ing_names = set()
        # gather ingredient names from live JSON if available
        try:
            cur_path = os.path.dirname(os.path.abspath(__file__))
            for _ in range(7):
                cur_path = os.path.dirname(cur_path)
                p = os.path.join(cur_path, "local_servers", "kulinar-mcp",
                                 "src", "data", "all_recipes.json")
                if os.path.exists(p):
                    for r in json.load(open(p, encoding="utf-8")):
                        for i in r["ingredients"]:
                            ing_names.add(i["name"].lower())
                    break
        except Exception:
            ing_names = set()
        for nm in ing_names:
            base = nm.split()[0] if nm else ""
            if len(base) >= 4 and base in vb and base not in seen:
                seen.add(base)
                ing_hits += 1
        has_cost = any(tok in vb for tok in ["$", "стоимост", "итог", "cost", "total"])
        # require either several recognizable ingredients OR an ingredient-list cue
        list_cue = any(tok in vb for tok in ["ингредиент", "ingredient", "список", "- ", "\n-"])
        vendor_body_ok = has_cost and (ing_hits >= 5 or list_cue)
        check("Vendor email body has ingredient list and total cost", vendor_body_ok,
              f"ing_hits={ing_hits} has_cost={has_cost} list_cue={list_cue} len={len(vb)}")

        # CRITICAL: committee email body must give per-day menu + budget summary.
        cb = body_for("wellness_committee").lower()
        day_tokens_en = ["monday", "tuesday", "wednesday", "thursday", "friday"]
        day_tokens_ru = ["понедельник", "вторник", "среда", "четверг", "пятниц"]
        day_hits = sum(1 for d in day_tokens_en if d in cb) + \
            sum(1 for d in day_tokens_ru if d in cb)
        has_budget = any(tok in cb for tok in
                         ["бюджет", "budget", "$", "стоимост", "итог", "total"])
        committee_body_ok = day_hits >= 3 and has_budget
        check("Committee email body has daily menu and budget summary",
              committee_body_ok,
              f"day_hits={day_hits} has_budget={has_budget} len={len(cb)}")

    except Exception as e:
        check("At least 2 wellness emails sent", False, str(e))
        check("Vendor email body has ingredient list and total cost", False, str(e))
        check("Committee email body has daily menu and budget summary", False, str(e))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT,
                    "total_checks": total,
                    "accuracy": accuracy,
                    "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
