"""
Evaluation for kulinar-event-menu-planner task.
Checks the Event_Menu.xlsx workbook + the email.

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Otherwise pass threshold: accuracy >= 70%.

Note: dish names come from the kulinar RU recipe base, so we DO NOT compare
against any hardcoded dish names. Semantic anchors are taken from
event_details.json (guest counts, budget, cost arithmetic) instead.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Semantic anchors from files/mock_pages/api/event_details.json
GUEST_COUNT = 50
BUDGET_PER_PERSON = 30
DIETARY = {"vegetarian": 10, "gluten_free": 5, "nut_allergy": 3}

# Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Excel exists",
    "Menu Plan covers all 3 courses",
    "Menu Plan has >= 6 rows (2 per course)",
    "Cost_Per_Person <= 30",
    "Budget_Variance == budget - cost_per_person",
    "Service_Fee_Estimate ~= 15% of food cost",
    "Dietary Accommodations covers 3 restrictions with correct guest counts",
    "Menu/dinner email sent",
    "Email to catering@vendor.com",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def to_float(v):
    try:
        return float(str(v).replace(",", ".").replace("₽", "").replace("RUB", "").strip())
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Event_Menu.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Menu Plan ---
    mp_rows = load_sheet_rows(wb, "Menu Plan") or load_sheet_rows(wb, "Menu_Plan")
    if mp_rows is None:
        record("Sheet 'Menu Plan' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Menu Plan' exists", True)
        data = [r for r in mp_rows[1:] if r and r[0] is not None]
        record("Menu Plan has >= 6 rows (2 per course)", len(data) >= 6, f"Found {len(data)}")

        course_col = find_col(mp_rows[0], ["Course", "course"])
        all_courses_ok = False
        if course_col is not None:
            courses = {str(r[course_col]).strip().lower() for r in data if course_col < len(r) and r[course_col]}
            # accept main / main course
            has_main = any("main" in c for c in courses)
            all_courses_ok = ("appetizer" in courses) and has_main and ("dessert" in courses)
            for c in ["appetizer", "main", "dessert"]:
                present = (c in courses) if c != "main" else has_main
                record(f"Course '{c}' present", present, f"Found: {courses}")
        record("Menu Plan covers all 3 courses", all_courses_ok,
               "Need appetizer + main + dessert in Course column")

        diet_col = find_col(mp_rows[0], ["Dietary_Tags", "Dietary Tags", "Tags"])
        record("Dietary_Tags column exists", diet_col is not None, f"Header: {mp_rows[0]}")

    # --- Ingredient List ---
    il_rows = load_sheet_rows(wb, "Ingredient List") or load_sheet_rows(wb, "Ingredient_List")
    if il_rows is None:
        record("Sheet 'Ingredient List' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Ingredient List' exists", True)
        data = [r for r in il_rows[1:] if r and r[0] is not None]
        record("Ingredient List has >= 5 items", len(data) >= 5, f"Found {len(data)}")

    # --- Cost Summary ---
    cs_rows = load_sheet_rows(wb, "Cost Summary") or load_sheet_rows(wb, "Cost_Summary")
    if cs_rows is None:
        record("Sheet 'Cost Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Cost Summary' exists", True)
        metrics = {}
        for row in cs_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        total_key = next((k for k in metrics if "total" in k and "food" in k and "cost" in k), None)
        cpp_key = next((k for k in metrics if "cost" in k and "per" in k and "person" in k), None)
        bud_key = next((k for k in metrics if "budget" in k and "per" in k and "person" in k), None)
        bv_key = next((k for k in metrics if "budget" in k and "var" in k), None)
        fee_key = next((k for k in metrics if "service" in k and "fee" in k), None)

        total_food = to_float(metrics.get(total_key)) if total_key else None
        cpp = to_float(metrics.get(cpp_key)) if cpp_key else None
        budget = to_float(metrics.get(bud_key)) if bud_key else None
        bv = to_float(metrics.get(bv_key)) if bv_key else None
        fee = to_float(metrics.get(fee_key)) if fee_key else None

        # Cost per person within budget (semantic, critical)
        record("Cost_Per_Person <= 30", cpp is not None and cpp <= BUDGET_PER_PERSON,
               f"Got {cpp}")

        # Budget_Per_Person should equal the source value (non-critical structural)
        record("Budget_Per_Person == 30", budget is not None and num_close(budget, BUDGET_PER_PERSON, 0.01),
               f"Got {budget}")

        # Budget_Variance arithmetic: budget - cost_per_person (critical)
        b = budget if budget is not None else BUDGET_PER_PERSON
        record("Budget_Variance == budget - cost_per_person",
               bv is not None and cpp is not None and num_close(bv, b - cpp, 0.5),
               f"bv={bv}, expected {None if cpp is None else b - cpp}")

        # Service fee = 15% of total food cost (critical)
        record("Service_Fee_Estimate ~= 15% of food cost",
               fee is not None and total_food is not None and num_close(fee, 0.15 * total_food, max(abs(0.15 * total_food) * 0.05, 1.0)),
               f"fee={fee}, expected {None if total_food is None else round(0.15 * total_food, 2)}")

        # Cross-check: total_food / guests ~= cost_per_person (non-critical)
        if total_food is not None and cpp is not None:
            record("Total_Food_Cost / 50 ~= Cost_Per_Person",
                   num_close(total_food / GUEST_COUNT, cpp, max(abs(cpp) * 0.1, 1.0)),
                   f"total={total_food}, cpp={cpp}")

    # --- Dietary Accommodations ---
    da_rows = load_sheet_rows(wb, "Dietary Accommodations") or load_sheet_rows(wb, "Dietary_Accommodations")
    if da_rows is None:
        record("Sheet 'Dietary Accommodations' exists", False, f"Sheets: {wb.sheetnames}")
        record("Dietary Accommodations covers 3 restrictions with correct guest counts", False, "sheet missing")
    else:
        record("Sheet 'Dietary Accommodations' exists", True)
        data = [r for r in da_rows[1:] if r and r[0] is not None]
        record("Dietary Accommodations has >= 3 rows", len(data) >= 3, f"Found {len(data)}")

        rest_col = find_col(da_rows[0], ["Restriction", "restriction"])
        gc_col = find_col(da_rows[0], ["Guest_Count", "Guest Count", "Guests"])
        # Map each restriction to its guest count from the source
        found = {}
        if rest_col is not None and gc_col is not None:
            for r in data:
                rname = str(r[rest_col]).strip().lower() if rest_col < len(r) and r[rest_col] is not None else ""
                gval = to_float(r[gc_col]) if gc_col < len(r) else None
                found[rname] = gval

        def match_count(keys, expected):
            for rn, gv in found.items():
                if any(k in rn for k in keys) and gv is not None and num_close(gv, expected, 0.01):
                    return True
            return False

        veg_ok = match_count(["vegetarian", "вегетариан"], DIETARY["vegetarian"])
        gf_ok = match_count(["gluten", "глютен"], DIETARY["gluten_free"])
        nut_ok = match_count(["nut", "орех"], DIETARY["nut_allergy"])
        record("Vegetarian guest count == 10", veg_ok, f"found={found}")
        record("Gluten-free guest count == 5", gf_ok, f"found={found}")
        record("Nut allergy guest count == 3", nut_ok, f"found={found}")
        record("Dietary Accommodations covers 3 restrictions with correct guest counts",
               veg_ok and gf_ok and nut_ok, f"found={found}")

    return True


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%%menu%%' OR subject ILIKE '%%dinner%%' OR subject ILIKE '%%catering%%'
    """)
    emails = cur.fetchall()

    record("Menu/dinner email sent", len(emails) >= 1, f"Found {len(emails)}")

    if emails:
        e = emails[0]
        to = e[3]
        if isinstance(to, str):
            try:
                to = json.loads(to)
            except Exception:
                pass
        to_str = str(to).lower()
        record("Email to catering@vendor.com", "catering@vendor.com" in to_str, f"To: {to}")

        # body keyword check: original-text .lower() with RU + EN alternatives
        body = str(e[4]).lower() if e[4] else ""
        kw = ["guest", "menu", "dinner", "гост", "меню", "ужин"]
        record("Email body mentions guests/menu (RU/EN)",
               any(k in body for k in kw),
               f"Body preview: {body[:200]}")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")

    if total == 0:
        print("FAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"  Accuracy: {accuracy:.1f}%")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
