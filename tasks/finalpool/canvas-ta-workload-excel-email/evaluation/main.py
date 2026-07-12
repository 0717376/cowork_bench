"""Evaluation for canvas-ta-workload-excel-email.

The agent must:
  1. Build TA_Workload_Report.xlsx with a "Course Workload" sheet (one row per
     course: Course_Name, Course_Code, TA_Count, Assignment_Count,
     Submission_Count, Submissions_Per_TA) and a "Summary" sheet
     (Total_Courses, Total_TAs, Avg_TAs_Per_Course, Max_Assignment_Count,
     Most_Loaded_Course).
  2. Create a Teamly page "TA Staffing Overview" whose body actually contains
     the key findings (total courses, total TAs, most-loaded course).
  3. Email dept_chair@university.edu (subject "TA Workload Report") with a
     summary in the body.

The aggregate values (22 courses, Total_TAs=29, Max_Assignment_Count=14, and
Основы финансов (Осень 2013) Submission_Count=16240) are deterministic
from the seeded Canvas dataset. Submissions_Per_TA correctness is verified
INTERNALLY from the agent's own sheet (Submission_Count / TA_Count) so it stays
honest regardless of the exact seed.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Expected aggregates, deterministic from the seeded Canvas dataset.
EXP_TOTAL_COURSES = 22
EXP_TOTAL_TAS = 29
EXP_MAX_ASSIGNMENT_COUNT = 14
EXP_FFF_SUBMISSIONS = 16240
EXP_MOST_LOADED = "основы финансов (осень 2013)"

# Critical semantic checks — reflect the task's substance (correct values from
# Canvas, correct aggregates, correct most-loaded identification, correct
# per-TA arithmetic, and a Teamly page that actually contains the findings),
# not mere structure (sheet/column existence).
CRITICAL_CHECKS = {
    "Course Workload has 22 data rows",
    "FFF-2013J Submission_Count=16240",
    "Total_Courses = 22",
    "Total_TAs = 29",
    "Max_Assignment_Count = 14",
    "Most_Loaded_Course is Основы финансов (Осень 2013)",
    "Submissions_Per_TA computed correctly",
    "Teamly page 'TA Staffing Overview' contains the key findings",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def num_variants(n):
    """RU + EN renderings of an integer (thousands separators tolerated)."""
    n = int(n)
    s = str(n)
    grouped_space = f"{n:,}".replace(",", " ")   # 16 240
    grouped_comma = f"{n:,}"                      # 16,240
    grouped_dot = f"{n:,}".replace(",", ".")     # 16.240
    return {s, grouped_space, grouped_comma, grouped_dot}


def body_has_number(text, n):
    """True if any RU/EN rendering of n appears as a standalone token."""
    for v in num_variants(n):
        pat = r"(?<!\d)" + re.escape(v) + r"(?!\d)"
        if re.search(pat, text):
            return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "TA_Workload_Report.xlsx")
    if not os.path.isfile(xlsx_path):
        check("TA_Workload_Report.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("TA_Workload_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # --- Course Workload sheet ---
    cw_rows = load_sheet_rows(wb, "Course Workload")
    if cw_rows is None:
        check("Sheet 'Course Workload' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Course Workload' exists", True)
        data_rows = cw_rows[1:] if len(cw_rows) > 1 else []
        check("Course Workload has 22 data rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        # Header columns -> index map (structural).
        header = cw_rows[0] if cw_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        col_idx = {}
        for col_name in ["course_name", "course_code", "ta_count", "assignment_count",
                         "submission_count", "submissions_per_ta"]:
            idx = next((i for i, h in enumerate(header_lower) if col_name in h), None)
            col_idx[col_name] = idx
            check(f"Column '{col_name}' present", idx is not None, f"Header: {header}")

        # CRITICAL: Основы финансов (Осень 2013) — correct value from Canvas.
        found_fff = False
        for row in data_rows:
            if row and row[0] and "основы финансов (осень 2013)" in str(row[0]).lower():
                found_fff = True
                ta_i = col_idx.get("ta_count")
                sub_i = col_idx.get("submission_count")
                ta_count = row[ta_i] if ta_i is not None and ta_i < len(row) else None
                submission_count = row[sub_i] if sub_i is not None and sub_i < len(row) else None
                check("FFF-2013J has TA_Count=1", num_close(ta_count, 1, 0), f"Got {ta_count}")
                check("FFF-2013J Submission_Count=16240",
                      num_close(submission_count, EXP_FFF_SUBMISSIONS, 10), f"Got {submission_count}")
        check("Основы финансов (Осень 2013) row found", found_fff)

        # CRITICAL: Submissions_Per_TA arithmetic verified from the agent's own
        # numbers — Submission_Count / TA_Count rounded to 1 dp, 0 when no TAs.
        ta_i = col_idx.get("ta_count")
        sub_i = col_idx.get("submission_count")
        spt_i = col_idx.get("submissions_per_ta")
        spt_ok = (ta_i is not None and sub_i is not None and spt_i is not None)
        mismatches = []
        if spt_ok:
            for row in data_rows:
                try:
                    ta = float(row[ta_i]) if row[ta_i] is not None else 0.0
                    sub = float(row[sub_i]) if row[sub_i] is not None else 0.0
                    got = float(row[spt_i]) if row[spt_i] is not None else None
                except (TypeError, ValueError, IndexError):
                    mismatches.append((row[0] if row else "?", "unparsable"))
                    continue
                expected = 0.0 if ta == 0 else round(sub / ta, 1)
                if got is None or abs(got - expected) > 0.15:
                    mismatches.append((row[0] if row else "?", f"got {got}, exp {expected}"))
        check("Submissions_Per_TA computed correctly",
              spt_ok and not mismatches,
              f"missing cols={not spt_ok}; mismatches={mismatches[:5]}")

    # --- Summary sheet ---
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        check("Total_Courses = 22", num_close(lookup.get("total_courses"), EXP_TOTAL_COURSES, 0),
              f"Got {lookup.get('total_courses')}")
        check("Total_TAs = 29", num_close(lookup.get("total_tas"), EXP_TOTAL_TAS, 0),
              f"Got {lookup.get('total_tas')}")
        check("Max_Assignment_Count = 14",
              num_close(lookup.get("max_assignment_count"), EXP_MAX_ASSIGNMENT_COUNT, 0),
              f"Got {lookup.get('max_assignment_count')}")
        most_loaded = lookup.get("most_loaded_course")
        check("Most_Loaded_Course is Основы финансов (Осень 2013)",
              bool(most_loaded) and EXP_MOST_LOADED in str(most_loaded).lower(),
              f"Got: {most_loaded}")


def check_teamly():
    print("\n=== Checking Teamly Page ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # English title 'TA Staffing Overview' is preserved per task.md; accept a
        # couple of lenient fallbacks (incl. a russified title variant).
        cur.execute("""
            SELECT id, title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE '%ta staffing%'
               OR title ILIKE '%staffing overview%'
               OR title ILIKE '%кадров%ассистент%'
        """)
        pages = cur.fetchall()
        if not pages:
            cur.execute("SELECT COUNT(*) FROM teamly.pages")
            total = cur.fetchone()[0]
            check("Teamly page 'TA Staffing Overview' exists", False,
                  f"Found {total} pages, none matching 'TA Staffing Overview'")
            conn.close()
            return
        check("Teamly page 'TA Staffing Overview' exists", True)

        body = "\n".join(str(b) for _, _, b in pages)
        body_lower = body.lower()
        check("Teamly page has non-trivial body", len(body) >= 80,
              f"Combined body is {len(body)} chars")

        # CRITICAL: the page must actually contain the key findings — total
        # courses, total TAs, and the most-loaded course name — not just a title.
        has_courses = body_has_number(body, EXP_TOTAL_COURSES)
        has_tas = body_has_number(body, EXP_TOTAL_TAS)
        has_most_loaded = EXP_MOST_LOADED in body_lower or "основы финансов" in body_lower
        check("Teamly page 'TA Staffing Overview' contains the key findings",
              has_courses and has_tas and has_most_loaded,
              f"total_courses={has_courses}, total_tas={has_tas}, most_loaded={has_most_loaded}")

        conn.close()
    except Exception as e:
        check("Teamly check", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, subject, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()

        # Locate the report email by recipient + subject; fall back to any email.
        chosen = None
        for _id, subj, to_addr, body in emails:
            to_str = json.dumps(to_addr).lower() if isinstance(to_addr, list) else str(to_addr).lower()
            subj_l = (subj or "").lower()
            if "dept_chair@university.edu" in to_str and "ta workload report" in subj_l:
                chosen = (subj, to_addr, body)
                break
        if chosen is None:
            for _id, subj, to_addr, body in emails:
                to_str = json.dumps(to_addr).lower() if isinstance(to_addr, list) else str(to_addr).lower()
                if "dept_chair@university.edu" in to_str:
                    chosen = (subj, to_addr, body)
                    break
        if chosen is None and emails:
            chosen = (emails[0][1], emails[0][2], emails[0][3])

        check("Email sent to dept_chair@university.edu", chosen is not None,
              "No matching email found")
        if chosen:
            subj, to_addr, body = chosen
            to_str = json.dumps(to_addr).lower() if isinstance(to_addr, list) else str(to_addr).lower()
            check("Email recipient is dept_chair@university.edu",
                  "dept_chair@university.edu" in to_str, f"to={to_addr}")
            subject = (subj or "").lower()
            check("Email subject is 'TA Workload Report'",
                  "ta workload report" in subject, f"Subject: {subj}")
            body = body or ""
            check("Email body has content", len(body) > 20, f"Body length: {len(body)}")
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and (accuracy >= 70)
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy,
                       "critical_failed": critical_failed,
                       "success": success}, f)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
