"""Evaluation script for fetch-moex-market-sentiment-excel-teamly.

Structural checks (sheets/columns/row floors, teamly page exists) are
NON-critical. Semantic checks in CRITICAL_CHECKS verify that the agent
actually merged live MOEX finance data with the seeded benchmark JSON
(values, sort, derived metrics, teamly summary). Any critical failure =>
overall FAIL regardless of accuracy. Otherwise pass threshold: accuracy >= 70%.
"""
import os
import argparse, json, os, sys, tarfile
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Data_Analysis Symbol set matches benchmark MOEX tickers",
    "Target_Price per row matches benchmark",
    "Upside per row equals Target_Price - Current_Price",
    "Metrics Total_Stocks and Avg_Upside are consistent with Data_Analysis",
    "Data_Analysis sorted alphabetically ascending by Symbol",
    "Teamly dashboard page exists and references a seeded MOEX ticker",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def load_benchmark():
    """Read the seeded benchmark JSON (same data served at /api/data.json)."""
    # Prefer the live extracted copy, fall back to the packaged tarball.
    candidates = [
        os.path.join(TASK_ROOT, "tmp", "mock_pages", "api", "data.json"),
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                with open(p, encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
    tar_path = os.path.join(TASK_ROOT, "files", "mock_pages.tar.gz")
    if os.path.exists(tar_path):
        try:
            with tarfile.open(tar_path, "r:gz") as tar:
                m = tar.extractfile("mock_pages/api/data.json")
                if m is not None:
                    return json.load(m)
        except Exception:
            pass
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    bench = load_benchmark()
    bench_map = {}
    if bench and isinstance(bench.get("market_indicators"), list):
        for it in bench["market_indicators"]:
            sym = str(it.get("symbol", "")).strip()
            if sym:
                bench_map[sym] = safe_float(it.get("target_price"))
    bench_syms = set(bench_map.keys())

    excel_path = os.path.join(agent_workspace, "Market_Sentiment_Report.xlsx")
    check("Market_Sentiment_Report.xlsx exists", os.path.exists(excel_path))

    da_rows = []  # list of (Symbol, Current_Price, Target_Price, Upside)
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # --- Data_Analysis (structural) ---
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            for expected_col in ['Symbol', 'Name', 'Sector', 'Current_Price', 'Target_Price', 'Upside']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            def col(name):
                return headers.index(name.lower()) if name.lower() in headers else None
            ci_sym, ci_cur, ci_tgt, ci_up = (col('Symbol'), col('Current_Price'),
                                             col('Target_Price'), col('Upside'))
            for r in data_rows:
                if ci_sym is None or r[ci_sym] is None:
                    continue
                da_rows.append((
                    str(r[ci_sym]).strip(),
                    safe_float(r[ci_cur]) if ci_cur is not None else None,
                    safe_float(r[ci_tgt]) if ci_tgt is not None else None,
                    safe_float(r[ci_up]) if ci_up is not None else None,
                ))

        # --- Metrics (structural) ---
        metrics_map = {}
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            mrows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(mrows) >= 3, f"got {len(mrows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for row in mrows:
                if row and row[0] is not None:
                    metrics_map[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        # --- Recommendations (structural) ---
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            rrows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(rrows) >= 2, f"got {len(rrows)}")
            for expected_col in ['Priority', 'Action', 'Symbol']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ===== CRITICAL SEMANTIC CHECKS =====
        # 1. Symbol set must match the benchmark MOEX tickers exactly.
        agent_syms = set(s for (s, _c, _t, _u) in da_rows)
        check("Data_Analysis Symbol set matches benchmark MOEX tickers",
              bool(bench_syms) and agent_syms == bench_syms,
              f"agent={sorted(agent_syms)} bench={sorted(bench_syms)}")

        # 2. Target_Price per row == benchmark target_price (within 0.01).
        tgt_ok = bool(da_rows) and bool(bench_map)
        for (s, _c, t, _u) in da_rows:
            bt = bench_map.get(s)
            if bt is None or t is None or abs(t - bt) > 0.01:
                tgt_ok = False
                break
        check("Target_Price per row matches benchmark", tgt_ok,
              f"rows={da_rows}")

        # 3. Upside per row == round(Target_Price - Current_Price, 2) (within 0.05).
        up_ok = bool(da_rows)
        for (s, c, t, u) in da_rows:
            if c is None or t is None or u is None or abs(u - round(t - c, 2)) > 0.05:
                up_ok = False
                break
        check("Upside per row equals Target_Price - Current_Price", up_ok,
              f"rows={da_rows}")

        # 4. Metrics: Total_Stocks == #rows, Avg_Upside == rounded mean(Upside).
        total_ok = avg_ok = False
        if da_rows:
            total_v = safe_float(metrics_map.get('total_stocks'))
            total_ok = total_v is not None and int(total_v) == len(da_rows)
            ups = [u for (_s, _c, _t, u) in da_rows if u is not None]
            if ups:
                exp_avg = round(sum(ups) / len(ups), 2)
                avg_v = safe_float(metrics_map.get('avg_upside'))
                avg_ok = avg_v is not None and abs(avg_v - exp_avg) <= 0.05
        check("Metrics Total_Stocks and Avg_Upside are consistent with Data_Analysis",
              total_ok and avg_ok,
              f"metrics={metrics_map}")

        # 5. Data_Analysis sorted alphabetically ascending by Symbol.
        syms_in_order = [s for (s, _c, _t, _u) in da_rows]
        check("Data_Analysis sorted alphabetically ascending by Symbol",
              syms_in_order == sorted(syms_in_order) and len(syms_in_order) >= 5,
              f"order={syms_in_order}")

        # yf_sentiment_processor.py exists (structural)
        check("yf_sentiment_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "yf_sentiment_processor.py")))

    # --- Teamly dashboard (CRITICAL: exists + references a seeded ticker) ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        rows = []
        check("Teamly dashboard page exists and references a seeded MOEX ticker", False, str(e))
        rows = None

    if rows is not None:
        def is_dashboard(t):
            tl = (t or "").lower()
            return any(k in tl for k in ("дашборд", "панель", "сводка",
                                         "dashboard", "настроени"))
        dash = [(t, b) for (t, b) in rows if is_dashboard(t)]
        page_text = " ".join((str(t) + " " + str(b)) for (t, b) in dash).lower()
        ticker_ok = any(
            (s.lower() in page_text) or (s.split(".")[0].lower() in page_text)
            for s in bench_syms
        )
        check("Teamly dashboard page exists and references a seeded MOEX ticker",
              len(dash) >= 1 and ticker_ok,
              f"dashboard_pages={len(dash)}, total_pages={len(rows)}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks (accuracy {accuracy:.1f}%)", critical_failed

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message, critical_failed = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    if critical_failed:
        print("Overall: FAIL (critical check failed)")
        sys.exit(1)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
