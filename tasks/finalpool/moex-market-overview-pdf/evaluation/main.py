"""
Evaluation для moex-market-overview-pdf (RU / moex-finance).

Агент строит:
  - Excel Market_Overview_Report.xlsx (листы "Stock Data" и "Summary")
  - PDF Market_Report.pdf
по данным MCP `moex-finance` (схемы moex.stock_prices / moex.stock_info)
для пяти тикеров Московской биржи: SBER.ME, GAZP.ME, LKOH.ME, MGNT.ME, MTSS.ME.

Эталонные значения НЕ захардкожены: они читаются вживую из PostgreSQL на
момент проверки (последний торговый день в moex.stock_prices).

CRITICAL_CHECKS (семантика): любой их провал => общий FAIL независимо от
accuracy. Иначе порог: accuracy >= 70%. Структурные проверки
(наличие листа, столбца, файла) — не критические.
"""
from argparse import ArgumentParser
import sys
import os
from pathlib import Path


SYMBOLS = ['SBER.ME', 'GAZP.ME', 'LKOH.ME', 'MGNT.ME', 'MTSS.ME']

# 13 обязательных столбцов листа "Stock Data".
REQUIRED_HEADERS = [
    "Symbol", "Company_Name", "Date", "Open", "High", "Low", "Close",
    "Volume", "Market_Cap", "PE_Ratio", "52W_High", "52W_Low", "Sector",
]

# Заголовок PDF: допускаем английский и русский вариант.
PDF_TITLE_VARIANTS = ["market overview report", "обзор рынка"]

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Семантические критические проверки.
CRITICAL_CHECKS = {
    "Stock Data: Close/Volume всех 5 тикеров корректны",
    "Summary: Highest_Close_Symbol и Highest_Volume_Symbol корректны",
    "Summary: Report_Date равен последнему торговому дню",
    "PDF: заголовок и все 5 тикеров присутствуют в тексте",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        tag = " (CRITICAL)" if name in CRITICAL_CHECKS else ""
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")


def get_expected_data():
    """Query PostgreSQL to get expected stock data (moex schema)."""
    import psycopg2

    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
        user="eigent", password="camel"
    )
    cur = conn.cursor()

    syms = tuple(SYMBOLS)
    cur.execute("""
        SELECT sp.symbol, sp.date, sp.open, sp.high, sp.low, sp.close, sp.volume
        FROM moex.stock_prices sp
        WHERE sp.symbol IN %s
          AND sp.date = (SELECT MAX(date) FROM moex.stock_prices WHERE symbol IN %s)
        ORDER BY sp.symbol
    """, (syms, syms))
    prices = cur.fetchall()

    cur.execute("""
        SELECT symbol,
               data->>'shortName' as name,
               data->>'marketCap' as market_cap,
               data->>'trailingPE' as pe_ratio,
               data->>'fiftyTwoWeekHigh' as high_52w,
               data->>'fiftyTwoWeekLow' as low_52w,
               data->>'sector' as sector
        FROM moex.stock_info
        WHERE symbol IN %s
        ORDER BY symbol
    """, (syms,))
    info = {r[0]: r for r in cur.fetchall()}

    conn.close()
    return prices, info


def check_excel(workspace, prices, info):
    """Проверка Market_Overview_Report.xlsx."""
    import openpyxl

    xlsx_path = Path(workspace) / "Market_Overview_Report.xlsx"
    if not xlsx_path.exists():
        record("Excel: файл Market_Overview_Report.xlsx существует", False, f"не найден в {workspace}")
        # Файла нет — все критические семантические проверки Excel/Summary
        # фиксируем как провал, чтобы сработал critical-gate.
        record("Stock Data: Close/Volume всех 5 тикеров корректны", False,
               "файл Excel отсутствует")
        record("Summary: Highest_Close_Symbol и Highest_Volume_Symbol корректны", False,
               "файл Excel отсутствует")
        record("Summary: Report_Date равен последнему торговому дню", False,
               "файл Excel отсутствует")
        return
    record("Excel: файл Market_Overview_Report.xlsx существует", True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    has_stock = "Stock Data" in wb.sheetnames
    has_summary = "Summary" in wb.sheetnames
    record("Excel: лист 'Stock Data' присутствует", has_stock, wb.sheetnames)
    record("Excel: лист 'Summary' присутствует", has_summary, wb.sheetnames)
    if not has_stock or not has_summary:
        # Без обоих листов критические семантические проверки невыполнимы —
        # явно фиксируем их провал, чтобы сработал critical-gate.
        record("Stock Data: Close/Volume всех 5 тикеров корректны", False,
               "лист отсутствует")
        record("Summary: Highest_Close_Symbol и Highest_Volume_Symbol корректны", False,
               "лист отсутствует")
        record("Summary: Report_Date равен последнему торговому дню", False,
               "лист отсутствует")
        wb.close()
        return

    # ---- Лист Stock Data ----
    ws1 = wb["Stock Data"]
    rows1 = list(ws1.iter_rows(values_only=True))
    if len(rows1) < 2:
        record("Stock Data: есть строки данных", False, "нет строк")
        # Нет строк данных — критические семантические проверки невыполнимы.
        record("Stock Data: Close/Volume всех 5 тикеров корректны", False,
               "нет строк данных")
        record("Summary: Highest_Close_Symbol и Highest_Volume_Symbol корректны", False,
               "нет строк данных в Stock Data")
        record("Summary: Report_Date равен последнему торговому дню", False,
               "нет строк данных в Stock Data")
        wb.close()
        return

    header = [str(h).strip() if h else "" for h in rows1[0]]

    # Структурная: все 13 заголовков присутствуют.
    missing = [c for c in REQUIRED_HEADERS if c not in header]
    record("Stock Data: присутствуют все 13 столбцов", not missing,
           f"нет: {missing}; есть: {header}")

    # Минимально необходимые столбцы для дальнейших проверок.
    for col in ("Symbol", "Close", "Volume"):
        if col not in header:
            record("Stock Data: Close/Volume всех 5 тикеров корректны", False,
                   f"нет обязательного столбца '{col}'")
            record("Summary: Highest_Close_Symbol и Highest_Volume_Symbol корректны", False,
                   f"нет обязательного столбца '{col}' в Stock Data")
            record("Summary: Report_Date равен последнему торговому дню", False,
                   f"нет обязательного столбца '{col}' в Stock Data")
            wb.close()
            return

    sym_idx = header.index("Symbol")
    close_idx = header.index("Close")
    vol_idx = header.index("Volume")

    data_rows = rows1[1:]
    # Иногда последняя строка пустая — отфильтруем полностью пустые.
    data_rows = [r for r in data_rows if any(c is not None and str(c).strip() != "" for c in r)]

    record("Stock Data: ровно 5 строк данных", len(data_rows) == 5,
           f"получено {len(data_rows)}")

    syms_in_order = [str(r[sym_idx]).strip() if r[sym_idx] else "" for r in data_rows]
    record("Stock Data: строки отсортированы по Symbol по алфавиту",
           syms_in_order == sorted(syms_in_order),
           f"порядок: {syms_in_order}")

    price_map = {r[0]: r for r in prices}

    close_vol_ok = True
    close_vol_detail = ""
    seen = set()
    for row in data_rows:
        sym = str(row[sym_idx]).strip() if row[sym_idx] else ""
        seen.add(sym)
        if sym not in price_map:
            close_vol_ok = False
            close_vol_detail = f"неизвестный тикер '{sym}'"
            break
        exp = price_map[sym]
        exp_close = float(exp[5])
        exp_vol = int(exp[6])
        if row[close_idx] is None or abs(float(row[close_idx]) - exp_close) > 0.5:
            close_vol_ok = False
            close_vol_detail = f"'{sym}' Close: ожидалось {exp_close}, получено {row[close_idx]}"
            break
        if row[vol_idx] is None or abs(int(row[vol_idx]) - exp_vol) > exp_vol * 0.05:
            close_vol_ok = False
            close_vol_detail = f"'{sym}' Volume: ожидалось {exp_vol}, получено {row[vol_idx]}"
            break
    if close_vol_ok and seen != set(SYMBOLS):
        close_vol_ok = False
        close_vol_detail = f"набор тикеров {seen} != {set(SYMBOLS)}"
    record("Stock Data: Close/Volume всех 5 тикеров корректны", close_vol_ok, close_vol_detail)

    # ---- Лист Summary ----
    ws2 = wb["Summary"]
    rows2 = list(ws2.iter_rows(values_only=True))
    summary_map = {}
    for row in rows2[1:]:
        if row and row[0]:
            summary_map[str(row[0]).strip()] = row[1] if len(row) > 1 else None

    record("Summary: Stocks_Covered == 5",
           "Stocks_Covered" in summary_map and int(summary_map["Stocks_Covered"]) == 5,
           summary_map.get("Stocks_Covered"))

    best_close = max(prices, key=lambda x: float(x[5]))[0]
    best_vol = max(prices, key=lambda x: int(x[6]))[0]
    got_close_sym = str(summary_map.get("Highest_Close_Symbol", "")).strip()
    got_vol_sym = str(summary_map.get("Highest_Volume_Symbol", "")).strip()
    record("Summary: Highest_Close_Symbol и Highest_Volume_Symbol корректны",
           got_close_sym == best_close and got_vol_sym == best_vol,
           f"close: ожид {best_close}/получ {got_close_sym}; vol: ожид {best_vol}/получ {got_vol_sym}")

    exp_date = prices[0][1]  # date последнего торгового дня
    got_date = summary_map.get("Report_Date")
    record("Summary: Report_Date равен последнему торговому дню",
           got_date is not None and str(exp_date) in str(got_date),
           f"ожидалось {exp_date}, получено {got_date}")

    wb.close()


def _extract_pdf_text(pdf_path):
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                text += (page.extract_text() or "") + "\n"
        if text.strip():
            return text
    except Exception:
        pass
    try:
        from pypdf import PdfReader
        reader = PdfReader(str(pdf_path))
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
        if text.strip():
            return text
    except Exception:
        pass
    try:
        import PyPDF2
        reader = PyPDF2.PdfReader(str(pdf_path))
        for page in reader.pages:
            text += (page.extract_text() or "") + "\n"
    except Exception:
        pass
    return text


def check_pdf(workspace):
    """Проверка Market_Report.pdf: содержимое (заголовок + 5 тикеров)."""
    pdf_path = Path(workspace) / "Market_Report.pdf"
    if not pdf_path.exists():
        record("PDF: файл Market_Report.pdf существует", False, f"не найден в {workspace}")
        record("PDF: заголовок и все 5 тикеров присутствуют в тексте", False, "файл отсутствует")
        return
    record("PDF: файл Market_Report.pdf существует", True)

    text = _extract_pdf_text(pdf_path)
    tl = text.lower()

    has_title = any(v in tl for v in PDF_TITLE_VARIANTS)
    # Тикеры могут быть как с суффиксом .ME, так и базовым кодом (SBER).
    missing_syms = []
    for s in SYMBOLS:
        base = s.split(".")[0].lower()
        if s.lower() not in tl and base not in tl:
            missing_syms.append(s)
    record("PDF: заголовок и все 5 тикеров присутствуют в тексте",
           has_title and not missing_syms,
           f"title={has_title}; нет тикеров: {missing_syms}; извлечено {len(text)} симв.")


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, help="Launch time")
    args = parser.parse_args()

    workspace = args.agent_workspace
    if not workspace:
        print("Error: --agent_workspace is required")
        sys.exit(1)

    print("Fetching expected data from database...")
    try:
        prices, info = get_expected_data()
        print(f"  Stocks: {len(prices)}")
        if len(prices) != 5:
            print(f"FATAL: ожидалось 5 тикеров в moex.stock_prices, получено {len(prices)}")
            sys.exit(1)
    except Exception as e:
        print(f"Error querying database: {e}")
        sys.exit(1)

    print("\n--- Check 1: Excel File ---")
    try:
        check_excel(workspace, prices, info)
    except Exception as e:
        record("Excel: проверка без ошибок", False, f"исключение: {e}")

    print("\n--- Check 2: PDF File ---")
    try:
        check_pdf(workspace)
    except Exception as e:
        record("PDF: проверка без ошибок", False, f"исключение: {e}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\nPassed {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"CRITICAL FAILED: {critical_failed}")
        print("\nSome checks failed.")
        sys.exit(1)

    if accuracy >= 70:
        print("\nPass all tests!")
        sys.exit(0)
    else:
        print(f"\nAccuracy {accuracy:.1f}% < 70%.")
        sys.exit(1)
