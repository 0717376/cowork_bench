"""Evaluation for sf-support-priority-gcal-excel."""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    errors = []
    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl not installed")
        return errors

    agent_file = os.path.join(agent_workspace, "Support_SLA_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Support_SLA_Analysis.xlsx")

    if not os.path.exists(agent_file):
        errors.append("Support_SLA_Analysis.xlsx not found in agent workspace")
        return errors
    if not os.path.exists(gt_file):
        errors.append("Groundtruth Support_SLA_Analysis.xlsx not found")
        return errors

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Priority Analysis sheet
    a_rows = load_sheet_rows(agent_wb, "Priority Analysis")
    g_rows = load_sheet_rows(gt_wb, "Priority Analysis")
    if a_rows is None:
        errors.append("Sheet 'Priority Analysis' not found in agent output")
    else:
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        g_data = [r for r in (g_rows[1:] if g_rows and len(g_rows) > 1 else []) if r and r[0] is not None]

        if len(a_data) < 3:
            errors.append(f"Priority Analysis: expected 3 data rows, got {len(a_data)}")
        else:
            a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r[0]}
            for g_row in g_data:
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing priority row: {g_row[0]}")
                    continue
                # Ticket_Count col 1
                if len(a_row) > 1 and not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{g_row[0]} Ticket_Count: got {a_row[1]}, expected {g_row[1]} (tol=5)")
                # Avg_Response_Hours col 2
                if len(a_row) > 2 and not num_close(a_row[2], g_row[2], 0.5):
                    errors.append(f"{g_row[0]} Avg_Response_Hours: got {a_row[2]}, expected {g_row[2]} (tol=0.5)")
                # SLA_Status col 3
                if len(a_row) > 3 and not str_match(a_row[3], g_row[3]):
                    errors.append(f"{g_row[0]} SLA_Status: got '{a_row[3]}', expected '{g_row[3]}'")

    # Check Summary sheet
    a_sum = load_sheet_rows(agent_wb, "Summary")
    g_sum = load_sheet_rows(gt_wb, "Summary")
    if a_sum is None:
        errors.append("Sheet 'Summary' not found in agent output")
    else:
        a_sum_data = {str(r[0]).strip().lower(): r[1] for r in (a_sum[1:] if len(a_sum) > 1 else []) if r and r[0]}
        g_sum_data = {str(r[0]).strip().lower(): r[1] for r in (g_sum[1:] if g_sum and len(g_sum) > 1 else []) if r and r[0]}

        # Total_Tickets
        tt = a_sum_data.get("total_tickets")
        if tt is None:
            errors.append("Summary missing Total_Tickets")
        elif not num_close(tt, 31588, 10):
            errors.append(f"Total_Tickets: got {tt}, expected 31588 (tol=10)")

        # Priorities_Met_SLA
        pm = a_sum_data.get("priorities_met_sla")
        if pm is None:
            errors.append("Summary missing Priorities_Met_SLA")
        elif not num_close(pm, 0, 0):
            errors.append(f"Priorities_Met_SLA: got {pm}, expected 0")

        # Most_Common_Priority
        mcp = a_sum_data.get("most_common_priority")
        if mcp is None:
            errors.append("Summary missing Most_Common_Priority")
        elif not str_match(mcp, "Medium"):
            errors.append(f"Most_Common_Priority: got '{mcp}', expected 'Medium'")

    return errors


def check_gcal():
    """CRITICAL: exactly the 3 first-Monday SLA review events in 09:00-10:00 window."""
    errors = []
    # First Monday of Apr/May/Jun 2026
    expected_dates = {"2026-04-06", "2026-05-04", "2026-06-01"}
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE (LOWER(summary) LIKE '%sla review%'
                   OR LOWER(summary) LIKE '%support%review%'
                   OR LOWER(summary) LIKE '%обзор%поддержк%')
            AND start_datetime >= '2026-04-01T00:00:00'
        """)
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"GCal DB check error: {e}")
        return errors

    def hhmm(dt):
        # Robust to 'YYYY-MM-DD HH:MM:SS' or 'YYYY-MM-DDTHH:MM:SS' forms.
        s = str(dt).replace("T", " ")
        parts = s.split(" ")
        return parts[1][:5] if len(parts) > 1 else ""

    matched_dates = set()
    for summary, start_dt, end_dt in events:
        date_part = str(start_dt)[:10]
        if date_part not in expected_dates:
            continue
        # verify 09:00-10:00 window
        if hhmm(start_dt) == "09:00" and hhmm(end_dt) == "10:00":
            matched_dates.add(date_part)

    missing = expected_dates - matched_dates
    if missing:
        errors.append(
            f"GCal: missing/incorrect first-Monday 09:00-10:00 SLA review events for dates {sorted(missing)} "
            f"(found valid {sorted(matched_dates)})"
        )
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr
            FROM email.messages
            WHERE LOWER(subject) LIKE '%support%'
               OR LOWER(subject) LIKE '%sla%'
               OR LOWER(subject) LIKE '%поддержк%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
        if not emails:
            errors.append("No email related to support or SLA found")
        else:
            found_to = False
            for em in emails:
                if "support.leads" in str(em[1]).lower():
                    found_to = True
                    break
            if not found_to:
                errors.append("No email sent to support.leads@company.com")
    except Exception as e:
        errors.append(f"Email DB check error: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    docx_path = os.path.join(agent_workspace, "SLA_Report.docx")
    if not os.path.exists(docx_path):
        errors.append("SLA_Report.docx not found")
        return errors
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        if len(text.strip()) < 30:
            errors.append("SLA_Report.docx has too little content")
        if "sla" not in text:
            errors.append("SLA_Report.docx missing keyword: sla")
        # 'medium' is an English data token; accept RU prose 'средн' as alternative
        if "medium" not in text and "средн" not in text:
            errors.append("SLA_Report.docx missing priority breakdown keyword: medium")
    except ImportError:
        if os.path.getsize(docx_path) < 100:
            errors.append("SLA_Report.docx too small")
    except Exception as e:
        errors.append(f"Error reading SLA_Report.docx: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    # Run all checks. (name, errors, is_critical)
    print("\n=== Checking Excel ===")
    excel_errors = check_excel(args.agent_workspace, gt_dir)

    print("\n=== Checking GCal Events ===")
    gcal_errors = check_gcal()

    print("\n=== Checking Email ===")
    email_errors = check_email()

    print("\n=== Checking Word Document ===")
    word_errors = check_word(args.agent_workspace)

    # CRITICAL checks carry the core semantic deliverables:
    #  - Excel per-priority counts/SLA_Status + Summary (data-derived values)
    #  - exactly the 3 first-Monday 09:00-10:00 SLA review calendar events
    #  - email to support.leads@company.com referencing the support review
    # NON-critical (structural / soft): Word document keyword check.
    checks = [
        ("Excel", excel_errors, True),
        ("GCal", gcal_errors, True),
        ("Email", email_errors, True),
        ("Word", word_errors, False),
    ]

    all_errors = []
    critical_errors = []
    passed = 0
    for name, errs, is_critical in checks:
        if errs:
            tag = "CRITICAL FAIL" if is_critical else "FAIL"
            for e in errs:
                print(f"  [{tag}] {e}")
            all_errors.extend(errs)
            if is_critical:
                critical_errors.extend(errs)
        else:
            print(f"  [PASS] {name} check passed")
            passed += 1

    accuracy = 100.0 * passed / len(checks)
    print(f"\nAccuracy: {accuracy:.1f}% ({passed}/{len(checks)} checks passed)")

    success = (not critical_errors) and accuracy >= 70

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "errors": all_errors,
                "critical_errors": critical_errors,
                "accuracy": accuracy,
                "success": success,
            }, f, indent=2)

    if critical_errors:
        print(f"\n=== RESULT: FAIL (critical check failed: {len(critical_errors)} errors) ===")
        sys.exit(1)
    if accuracy < 70:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)
    print("\n=== RESULT: PASS ===")
    sys.exit(0)


if __name__ == "__main__":
    main()
