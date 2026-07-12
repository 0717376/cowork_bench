"""Evaluation for sf-hr-satisfaction-forms-excel-gcal.

Критические проверки (CRITICAL_CHECKS): любой их провал => общий FAIL независимо
от accuracy. Иначе PASS требует accuracy >= 70%.

Примечание: значения отделов (Инженерия/Финансы/...) русифицированы централизованно
в db/zzz_clickhouse_after_init.sql и читаются здесь ЖИВЫМ запросом к sf_data, поэтому
seed<->eval остаются синхронизированы автоматически. НЕ хардкодим литералы отделов
и числовые значения — читаем их из БД.
"""
import argparse
import json
import os
import sys

import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

# Критические проверки: любой провал => общий FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Department_Analysis: 7 russified departments match live warehouse",
    "Department_Analysis: Priority_Level derived from rule per department",
    "Calendar: 'Wellness Program Kickoff' on 2026-03-15 14:00-15:30 UTC",
    "Forms: feedback survey distinct from noise with >=5 questions incl. 7-option department choice",
    "Metrics: required metric rows with valid russified department names",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    a = safe_float(a)
    b = safe_float(b)
    if a is None or b is None:
        return False
    return abs(a - b) <= max(abs_tol, abs(b) * rel_tol)


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def priority_from_sat(sat):
    if sat < 3.0:
        return "high"
    if sat < 3.5:
        return "medium"
    return "low"


def get_warehouse_aggregates():
    """Живой запрос к складу: на отдел кол-во/средние + общие показатели."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               COUNT(*),
               AVG("JOB_SATISFACTION"),
               AVG("WORK_LIFE_BALANCE"),
               AVG("PERFORMANCE_RATING")
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
    """)
    rows = cur.fetchall()
    cur.execute("""
        SELECT AVG("JOB_SATISFACTION"), AVG("WORK_LIFE_BALANCE")
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
    """)
    overall = cur.fetchone()
    cur.close()
    conn.close()
    depts = {}
    for name, cnt, sat, wlb, perf in rows:
        depts[str(name).strip().lower()] = {
            "name": str(name).strip(),
            "count": int(cnt),
            "sat": round(float(sat), 2),
            "wlb": round(float(wlb), 2),
            "perf": round(float(perf), 2),
            "priority": priority_from_sat(float(sat)),
        }
    overall_sat = round(float(overall[0]), 2)
    overall_wlb = round(float(overall[1]), 2)
    return depts, overall_sat, overall_wlb


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[c.value for c in row] for row in wb[name].iter_rows()]
    return None


def header_index(rows):
    return {str(h).strip().lower(): i for i, h in enumerate(rows[0])} if rows else {}


def check_excel(agent_workspace, depts, overall_sat, overall_wlb):
    excel_path = os.path.join(agent_workspace, "Satisfaction_Program_Report.xlsx")
    exists = os.path.exists(excel_path)
    check("Satisfaction_Program_Report.xlsx exists", exists, excel_path)
    if not exists:
        check("Department_Analysis: 7 russified departments match live warehouse", False, "no file")
        check("Department_Analysis: Priority_Level derived from rule per department", False, "no file")
        check("Metrics: required metric rows with valid russified department names", False, "no file")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    valid_dept_names = {d["name"].lower() for d in depts.values()}

    # ---- Department_Analysis ----
    da = load_sheet_rows(wb, "Department_Analysis")
    if da is None:
        check("Department_Analysis: 7 russified departments match live warehouse", False, "sheet missing")
        check("Department_Analysis: Priority_Level derived from rule per department", False, "sheet missing")
    else:
        hidx = header_index(da)
        data = [r for r in da[1:] if r and r[0] is not None]
        # Header presence (non-critical)
        for h in ["department", "employee_count", "avg_satisfaction",
                  "avg_worklife_balance", "avg_performance", "priority_level"]:
            check(f"Department_Analysis has {h} column", h in hidx, f"headers: {list(hidx)}")

        d_col = hidx.get("department", 0)
        agent_depts = {str(r[d_col]).strip().lower(): r for r in data if r[d_col]}

        # CRITICAL: exactly 7 rows, names match the russified warehouse set.
        names_match = (len(agent_depts) == len(depts)
                       and set(agent_depts) == set(depts))
        check("Department_Analysis: 7 russified departments match live warehouse",
              names_match,
              f"agent={sorted(agent_depts)} expected={sorted(depts)}")

        # Non-critical: numeric averages per department.
        if "avg_satisfaction" in hidx:
            sat_errors = []
            for key, exp in depts.items():
                row = agent_depts.get(key)
                if row is None:
                    continue
                if not num_close(row[hidx["avg_satisfaction"]], exp["sat"]):
                    sat_errors.append(f"{key} sat got {row[hidx['avg_satisfaction']]} exp {exp['sat']}")
                if "avg_worklife_balance" in hidx and not num_close(row[hidx["avg_worklife_balance"]], exp["wlb"]):
                    sat_errors.append(f"{key} wlb got {row[hidx['avg_worklife_balance']]} exp {exp['wlb']}")
            check("Department_Analysis: per-department averages match warehouse",
                  not sat_errors, "; ".join(sat_errors[:6]))

        # CRITICAL: Priority_Level derived correctly from the rule.
        prio_errors = []
        if "priority_level" not in hidx:
            prio_errors.append("priority_level column missing")
        else:
            for key, exp in depts.items():
                row = agent_depts.get(key)
                if row is None:
                    prio_errors.append(f"missing dept {key}")
                    continue
                got = str(row[hidx["priority_level"]]).strip().lower()
                if got != exp["priority"]:
                    prio_errors.append(f"{key} priority got '{got}' exp '{exp['priority']}'")
        check("Department_Analysis: Priority_Level derived from rule per department",
              not prio_errors, "; ".join(prio_errors[:6]))

    # ---- Program_Plan ----
    pp = load_sheet_rows(wb, "Program_Plan")
    if pp is None:
        check("Program_Plan sheet exists with >=5 initiatives", False, "sheet missing")
    else:
        pp_data = [r for r in pp[1:] if r and r[0] is not None]
        hidx = header_index(pp)
        for h in ["initiative", "target_department", "expected_impact", "timeline"]:
            check(f"Program_Plan has {h} column", h in hidx, f"headers: {list(hidx)}")
        check("Program_Plan sheet exists with >=5 initiatives", len(pp_data) >= 5,
              f"got {len(pp_data)}")

    # ---- Metrics ----
    mt = load_sheet_rows(wb, "Metrics")
    if mt is None:
        check("Metrics: required metric rows with valid russified department names", False, "sheet missing")
    else:
        mdata = {str(r[0]).strip().lower(): r[1]
                 for r in mt[1:] if r and r[0] is not None}
        required = ["overall_avg_satisfaction", "overall_avg_worklife",
                    "lowest_satisfaction_dept", "highest_satisfaction_dept",
                    "departments_below_avg"]
        missing = [m for m in required if m not in mdata]
        # The two *_dept metrics must hold valid russified department names.
        dept_metric_ok = True
        bad = []
        for m in ("lowest_satisfaction_dept", "highest_satisfaction_dept"):
            v = mdata.get(m)
            if v is None or str(v).strip().lower() not in valid_dept_names:
                dept_metric_ok = False
                bad.append(f"{m}={v}")
        check("Metrics: required metric rows with valid russified department names",
              (not missing) and dept_metric_ok,
              f"missing={missing} bad_dept_metrics={bad}")

        # Non-critical: overall averages match warehouse.
        check("Metrics: overall averages match warehouse",
              num_close(mdata.get("overall_avg_satisfaction"), overall_sat)
              and num_close(mdata.get("overall_avg_worklife"), overall_wlb),
              f"got sat={mdata.get('overall_avg_satisfaction')} wlb={mdata.get('overall_avg_worklife')} "
              f"exp sat={overall_sat} wlb={overall_wlb}")


def check_script(agent_workspace):
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")


def check_calendar():
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Дата/время, не подстрока ключевого слова. Заголовок 'Wellness Program Kickoff'
        # допускаем EN или RU; дата 2026-03-15, начало 14:00, конец 15:30 UTC.
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE (LOWER(summary) LIKE '%%wellness%%'
                   OR LOWER(summary) LIKE '%%kickoff%%'
                   OR LOWER(summary) LIKE '%%благополуч%%'
                   OR LOWER(summary) LIKE '%%старт%%')
              AND start_datetime::date = DATE '2026-03-15'
        """)
        rows = cur.fetchall()
        ok = False
        detail = f"candidates: {rows}"
        for summary, sdt, edt in rows:
            if sdt is None:
                continue
            start_ok = sdt.hour == 14 and sdt.minute == 0
            end_ok = edt is not None and edt.hour == 15 and edt.minute == 30
            if start_ok and end_ok:
                ok = True
                break
        check("Calendar: 'Wellness Program Kickoff' on 2026-03-15 14:00-15:30 UTC", ok, detail)

        # Non-critical: шумовые события не удалены агентом.
        cur.execute("""SELECT COUNT(*) FROM gcal.events
            WHERE summary ILIKE '%%планёрк%%' OR summary ILIKE '%%планерк%%'
               OR summary ILIKE '%%обеденн%%' OR summary ILIKE '%%перерыв%%'""")
        noise = cur.fetchone()[0]
        check("Noise calendar events preserved (not deleted)", noise >= 1, f"noise events: {noise}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar: 'Wellness Program Kickoff' on 2026-03-15 14:00-15:30 UTC", False, str(e))


def check_forms():
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Опрос агента, отличный от шумовой формы noise-form-001.
        cur.execute("SELECT id, title FROM gform.forms WHERE id != 'noise-form-001'")
        forms = cur.fetchall()
        if not forms:
            check("Forms: feedback survey distinct from noise with >=5 questions incl. 7-option department choice",
                  False, "no non-noise form")
            cur.close()
            conn.close()
            return

        best = None
        best_detail = ""
        for fid, title in forms:
            cur.execute(
                "SELECT title, question_type, config FROM gform.questions WHERE form_id = %s ORDER BY position",
                (fid,))
            qs = cur.fetchall()
            n = len(qs)
            # Ищем вопрос-выбор с >=7 вариантами (отдел).
            dept_choice_ok = False
            for qtitle, qtype, cfg in qs:
                if qtype in ("choiceQuestion", "RADIO", "CHOICE", "MULTIPLE_CHOICE"):
                    opts = []
                    if isinstance(cfg, dict):
                        opts = cfg.get("options", []) or []
                    elif isinstance(cfg, str) and cfg:
                        try:
                            opts = (json.loads(cfg) or {}).get("options", []) or []
                        except json.JSONDecodeError:
                            opts = []
                    if len(opts) >= 7:
                        dept_choice_ok = True
                        break
            if n >= 5 and dept_choice_ok:
                best = (fid, title, n)
                break
            if best_detail == "":
                best_detail = f"form '{title}' n={n} dept_choice7={dept_choice_ok}"
        ok = best is not None
        check("Forms: feedback survey distinct from noise with >=5 questions incl. 7-option department choice",
              ok, best_detail if not ok else f"matched {best}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Forms: feedback survey distinct from noise with >=5 questions incl. 7-option department choice",
              False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    try:
        depts, overall_sat, overall_wlb = get_warehouse_aggregates()
    except Exception as e:
        check("Warehouse reachable", False, str(e))
        depts, overall_sat, overall_wlb = {}, 0.0, 0.0

    print("\n=== Excel ===")
    check_excel(agent_workspace, depts, overall_sat, overall_wlb)
    print("\n=== Script ===")
    check_script(agent_workspace)
    print("\n=== Calendar ===")
    check_calendar()
    print("\n=== Forms ===")
    check_forms()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                    "success": (not critical_failed) and accuracy >= 70,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
