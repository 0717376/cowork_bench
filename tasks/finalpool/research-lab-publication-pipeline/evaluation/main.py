#!/usr/bin/env python3
"""Evaluation script for research-lab-publication-pipeline task validation.

The agent reads research_data.xlsx (sheets Paper1_Results / Paper2_Analysis /
Paper3_Experiments) plus figure_templates.pptx, writes three manuscripts
(manuscript_paper1/2/3.docx) whose Results/Methods prose derives from the data,
and fills submission_checklist.xlsx mirroring the groundtruth structure.

English identifiers (sheet names, column headers, file names, Status literals
like 'Complete') are preserved; free-text prose (Notes, manuscript body) is
expected in Russian. Matching is therefore broadened to RU+EN where relevant.

CRITICAL_CHECKS capture the substance of the deliverable: any failure forces a
hard FAIL (sys.exit(1)) BEFORE the >=70% accuracy gate.
"""

from argparse import ArgumentParser
import json
import os
import sys

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_xlsx_content(workspace, groundtruth_workspace="."):
    print("\n=== Check: XLSX submission_checklist.xlsx ===")
    import openpyxl
    xlsx_path = os.path.join(workspace, "submission_checklist.xlsx")
    if not os.path.isfile(xlsx_path):
        record("submission_checklist.xlsx exists", False, "Not found", critical=True)
        return
    record("submission_checklist.xlsx exists", True, critical=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # --- Locate the 'Submission Checklist' sheet (English identifier preserved) ---
    target_ws = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == "submission checklist":
            target_ws = wb[sn]
            break
    record("sheet 'Submission Checklist' present", target_ws is not None,
           f"sheets={wb.sheetnames}", critical=True)

    # --- Header row [Item, Status, Notes, Reviewer] must stay English ---
    header_ok = False
    if target_ws is not None:
        for row in target_ws.iter_rows(min_row=1, max_row=6, values_only=True):
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if cells[:4] == ["item", "status", "notes", "reviewer"]:
                header_ok = True
                break
    record("header row == ['Item','Status','Notes','Reviewer'] (EN preserved)",
           header_ok, critical=True)

    # --- Substance: core checklist items (Item + Status) match groundtruth ---
    # Build a map Item(lower) -> Status(lower) from the agent sheet.
    agent_items = {}
    if target_ws is not None:
        for row in target_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] is not None and len(row) >= 2:
                agent_items[str(row[0]).strip().lower()] = (
                    str(row[1]).strip().lower() if row[1] is not None else "")

    EXPECTED = [
        ("Manuscript formatting", "Complete"),
        ("Title page included", "Complete"),
        ("Abstract (150-200 words)", "Complete"),
        ("References formatted", "Complete"),
        ("Copyright permissions", "In Progress"),
    ]
    matched = 0
    for item, status in EXPECTED:
        got = agent_items.get(item.lower())
        if got is not None and got == status.lower():
            matched += 1
    record(f"core checklist items+statuses match GT ({matched}/{len(EXPECTED)})",
           matched >= 4, f"agent items: {list(agent_items)[:8]}", critical=True)

    # --- Groundtruth row-count comparison (non-critical structural) ---
    gt_path = os.path.join(groundtruth_workspace, "submission_checklist.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                       f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                       if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", abs(len(a_rows) - len(gt_rows)) <= 2,
                   f"Expected ~{len(gt_rows)}, got {len(a_rows)}")
        gt_wb.close()
    wb.close()


# Provenance numbers expected to appear in each manuscript (from research_data.xlsx).
DOCX_PROVENANCE = {
    "manuscript_paper1.docx": ["156", "0.847"],
    "manuscript_paper2.docx": ["45.2", "52.8"],
    "manuscript_paper3.docx": ["87.5"],
}

# Section markers accepted in Russian OR English.
REQUIRED_SECTIONS = [
    ["аннотац", "abstract", "реферат"],
    ["введени", "introduction"],
    ["метод", "method"],
    ["результат", "result"],
]


def check_docx_content(workspace):
    print("\n=== Check: DOCX manuscripts ===")
    from docx import Document
    for fname in ["manuscript_paper1.docx", "manuscript_paper2.docx", "manuscript_paper3.docx"]:
        path = os.path.join(workspace, fname)
        if not os.path.isfile(path):
            record(f"{fname} exists", False, "Not found", critical=True)
            continue
        record(f"{fname} exists", True, critical=True)
        try:
            doc = Document(path)
        except Exception as e:
            record(f"{fname} readable", False, str(e), critical=True)
            continue

        full = "\n".join(p.text for p in doc.paragraphs)
        low = full.lower()

        # Non-trivial prose body (>= 200 chars), not an empty placeholder.
        record(f"{fname} has substantial prose (>=200 chars)", len(full.strip()) >= 200,
               f"{len(full.strip())} chars", critical=True)

        # Required sections (RU or EN).
        sect_hits = sum(1 for alts in REQUIRED_SECTIONS if any(a in low for a in alts))
        record(f"{fname} contains key sections (abstract/intro/methods/results)",
               sect_hits >= 3, f"{sect_hits}/4 section groups found")

        # Provenance: derived from seeded research_data.xlsx values.
        # Russian prose naturally uses the comma decimal separator (45,2), so
        # match locale-agnostically against both comma- and dot-form decimals.
        full_dot = full.replace(",", ".")
        prov = DOCX_PROVENANCE.get(fname, [])
        prov_hits = sum(1 for v in prov if v in full or v in full_dot)
        record(f"{fname} references data values from research_data.xlsx ({prov_hits}/{len(prov)})",
               prov_hits >= 1, f"looked for {prov}", critical=True)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    ws = args.agent_workspace
    if not os.path.isdir(ws):
        print(f"Agent workspace not found: {ws}")
        sys.exit(1)

    check_xlsx_content(ws, args.groundtruth_workspace)
    check_docx_content(ws)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": CRITICAL_FAILED,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # Critical gate first: any critical failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILED:
        print(f"\nFAIL: {len(CRITICAL_FAILED)} critical check(s) failed: {CRITICAL_FAILED}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
