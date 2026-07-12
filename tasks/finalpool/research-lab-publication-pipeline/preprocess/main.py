#!/usr/bin/env python3
"""Preprocess script for research-lab-publication-pipeline task setup.

Copies the real source files (research_data.xlsx, figure_templates.pptx) into the
agent workspace and injects light Russian lab context into the *Notes* column of
research_data.xlsx. The Notes column is NOT graded (the evaluator only compares
submission_checklist.xlsx and the manuscripts), so this adds RU specificity
without pre-seeding any answer and without touching English sheet names, column
headers or numeric values the agent must read.
"""

from argparse import ArgumentParser
from pathlib import Path
import shutil

TASK_ROOT = Path(__file__).resolve().parent.parent
INITIAL_WS = TASK_ROOT / "initial_workspace"

# Only the genuine source files are handed to the agent. The generic
# config.json / data.csv / README.md placeholders are intentionally skipped.
SOURCE_FILES = ["research_data.xlsx", "figure_templates.pptx"]

# RU lab context injected into the (ungraded) Notes column of research_data.xlsx.
# Keyed by sheet -> list of notes per data row (header row excluded).
RU_NOTES = {
    "Paper1_Results": [
        "Ответственный: Иванов И. И., НИИ прикладных исследований",
        "Проверила: Петрова А. С.",
        "Требуется повторный анализ подгрупп",
        "Согласовано с научным руководителем",
        "Доверительный интервал, нижняя граница",
        "Доверительный интервал, верхняя граница",
    ],
    "Paper2_Analysis": [
        "Контрольная группа, лаборатория молекулярной биологии",
        "Опытная группа, лаборатория молекулярной биологии",
        "Сравнение групп, ответственный Сидоров П. Н.",
        "Подгруппа когорты А",
        "Подгруппа когорты Б",
    ],
    "Paper3_Experiments": [
        "Серия экспериментов, отв. Кузнецова Е. В.",
        "Серия экспериментов, отв. Кузнецова Е. В.",
        "Повторный прогон на спектрометре А",
        "Сводные результаты по всем установкам",
        "Контроль воспроизводимости, независимая лаборатория",
    ],
}


def inject_ru_context(xlsx_path: Path):
    try:
        import openpyxl
    except ImportError:
        print("[preprocess] openpyxl unavailable, skipping RU note injection")
        return
    wb = openpyxl.load_workbook(xlsx_path)
    for sheet, notes in RU_NOTES.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        header = [c.value for c in ws[1]]
        # Find the Notes column index (only Paper1 has one originally); if absent,
        # append it so each sheet carries RU context.
        if "Notes" in header:
            ncol = header.index("Notes") + 1
        else:
            ncol = len(header) + 1
            ws.cell(row=1, column=ncol, value="Notes")
        for i, note in enumerate(notes):
            ws.cell(row=2 + i, column=ncol, value=note)
    wb.save(xlsx_path)
    print(f"[preprocess] Injected RU lab context into {xlsx_path.name}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    if args.agent_workspace:
        agent_ws = Path(args.agent_workspace)
        agent_ws.mkdir(parents=True, exist_ok=True)
        for fname in SOURCE_FILES:
            src = INITIAL_WS / fname
            if src.is_file():
                dst = agent_ws / fname
                shutil.copy2(src, dst)
                print(f"[preprocess] Copied {fname} -> {dst}")
            else:
                print(f"[preprocess] WARNING: source missing: {src}")
        ws_xlsx = agent_ws / "research_data.xlsx"
        if ws_xlsx.is_file():
            inject_ru_context(ws_xlsx)

    print("Preprocess completed successfully")


if __name__ == "__main__":
    main()
