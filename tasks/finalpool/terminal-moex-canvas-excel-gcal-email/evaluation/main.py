"""Evaluation for terminal-moex-canvas-excel-gcal-email.

Checks:
1. Financial_Literacy_Workshops.xlsx with 3 sheets (Student_Tiers, Market_Events, Workshop_Schedule)
2. Google Calendar events for 3 workshops (14:00, 90 min, distinct weekdays, no conflict-window overlap)
3. Emails sent to finance_students and department_head (exact subjects)
4. workshop_materials.txt exists
5. market_events.json exists
6. categorize_students.py and find_market_events.py scripts exist

CRITICAL checks (any failure => exit 1 regardless of accuracy):
- Per-course tier counts exactly match live Canvas groundtruth (courses 16, 17).
- Top-3 Market_Events symbols AND change_pct match live moex groundtruth in order.
- Exactly 3 workshops at 14:00 / 90 min on distinct weekdays, none overlapping injected 13-16h conflicts.
- Both emails sent to correct recipients with exact subjects.
- Workshop_Schedule Expected_Attendance per tier equals combined both-course count.
"""
import argparse
import json
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

# MOEX tickers whose synthetic series give distinct, non-tied top-3 daily moves.
MARKET_SYMBOLS = ['GAZP.ME', 'MGNT.ME', 'MTSS.ME']

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def check_critical(name, condition, detail=""):
    """A semantic check that gates PASS/FAIL. Also counted in accuracy."""
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS][CRITICAL] {name}")
    else:
        FAIL_COUNT += 1
        CRITICAL_FAILED.append(name)
        print(f"  [FAIL][CRITICAL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_expected_tiers():
    """Query Canvas DB for expected tier counts."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT sub.course_id,
              SUM(CASE WHEN avg_pct < 60 THEN 1 ELSE 0 END) as needs_support,
              SUM(CASE WHEN avg_pct >= 60 AND avg_pct < 75 THEN 1 ELSE 0 END) as developing,
              SUM(CASE WHEN avg_pct >= 75 THEN 1 ELSE 0 END) as proficient,
              COUNT(*) as total
            FROM (
              SELECT a.course_id, s.user_id,
                AVG(CASE WHEN a.points_possible > 0 THEN s.score / a.points_possible * 100 ELSE NULL END) as avg_pct
              FROM canvas.submissions s
              JOIN canvas.assignments a ON s.assignment_id = a.id
              WHERE a.course_id IN (16, 17) AND s.score IS NOT NULL AND a.points_possible > 0
              GROUP BY a.course_id, s.user_id
            ) sub
            GROUP BY sub.course_id
            ORDER BY sub.course_id
        """)
        result = {}
        for row in cur.fetchall():
            result[int(row[0])] = {
                'needs_support': int(row[1]),
                'developing': int(row[2]),
                'proficient': int(row[3]),
                'total': int(row[4])
            }
        cur.close()
        conn.close()
        return result
    except Exception as e:
        print(f"  [WARN] Could not query Canvas: {e}")
        return {}


def get_expected_market_events():
    """Query MOEX DB for expected top 3 market events.

    task.md asks for the largest daily moves over the last 30 *trading days*
    (за последние 30 дней торговых данных), i.e. the last 30 trading rows
    per symbol -- NOT a 30 calendar-day window. We take, for each symbol, its
    most recent 30 trading rows and pick the top-3 moves across that union.
    """
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            WITH ranked AS (
              SELECT symbol, date, open, close,
                ROW_NUMBER() OVER (PARTITION BY symbol ORDER BY date DESC) AS rn
              FROM moex.stock_prices
              WHERE symbol IN %s
            )
            SELECT symbol, date, open, close,
              ROUND(((close - open) / open * 100)::numeric, 2) as change_pct
            FROM ranked
            WHERE rn <= 30 AND open > 0
            ORDER BY ABS((close - open) / open) DESC
            LIMIT 3
        """, (tuple(MARKET_SYMBOLS),))
        events = []
        for row in cur.fetchall():
            events.append({
                'symbol': row[0],
                'date': str(row[1]),
                'change_pct': float(row[4])
            })
        cur.close()
        conn.close()
        return events
    except Exception as e:
        print(f"  [WARN] Could not query MOEX: {e}")
        return []


def get_injected_conflicts():
    """Return list of (start_datetime, end_datetime) for the injected 13-16h conflict events."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT start_datetime, end_datetime FROM gcal.events
            WHERE summary IN ('Заседание учёного совета', 'Бюджетный комитет кафедры')
        """)
        conflicts = cur.fetchall()
        cur.close()
        conn.close()
        return conflicts
    except Exception as e:
        print(f"  [WARN] Could not query conflicts: {e}")
        return []


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Financial_Literacy_Workshops.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Financial_Literacy_Workshops.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        # Cannot verify critical tier/attendance deliverables -> fail them.
        check_critical("Student_Tiers per-course counts (no Excel)", False, "Excel missing")
        check_critical("Market_Events top-3 (no Excel)", False, "Excel missing")
        check_critical("Workshop_Schedule attendance (no Excel)", False, "Excel missing")
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        check_critical("Student_Tiers per-course counts (unreadable)", False, str(e))
        check_critical("Market_Events top-3 (unreadable)", False, str(e))
        check_critical("Workshop_Schedule attendance (unreadable)", False, str(e))
        return

    check("Has 3 sheets", len(agent_wb.sheetnames) >= 3, f"Got {agent_wb.sheetnames}")

    expected_tiers = get_expected_tiers()
    expected_events = get_expected_market_events()

    # ---- Student_Tiers sheet (CRITICAL: exact per-course tier counts) ----
    print("  Checking Student_Tiers...")
    st_sheet = get_sheet(agent_wb, "Student_Tiers")
    check("Sheet 'Student_Tiers' exists", st_sheet is not None, f"Sheets: {agent_wb.sheetnames}")

    # Un-guarded: missing groundtruth is a CRITICAL failure, not a silent skip.
    check_critical("Canvas groundtruth tiers available",
                   bool(expected_tiers) and 16 in expected_tiers and 17 in expected_tiers,
                   f"expected_tiers keys: {list(expected_tiers.keys())}")

    rows_by_cid = {}
    if st_sheet:
        rows = list(st_sheet.iter_rows(min_row=2, values_only=True))
        check("Student_Tiers has 2 rows", len(rows) == 2, f"Got {len(rows)}")
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                rows_by_cid[int(row[0])] = row
            except (TypeError, ValueError):
                pass

    if expected_tiers and 16 in expected_tiers and 17 in expected_tiers:
        for cid in (16, 17):
            exp = expected_tiers[cid]
            row = rows_by_cid.get(cid)
            if row is None:
                check_critical(f"Course {cid} tier counts present", False, "row missing")
                continue
            check_critical(f"Course {cid} Needs_Support exact",
                           num_close(row[2], exp['needs_support'], 0),
                           f"Expected {exp['needs_support']}, got {row[2]}")
            check_critical(f"Course {cid} Developing exact",
                           num_close(row[3], exp['developing'], 0),
                           f"Expected {exp['developing']}, got {row[3]}")
            check_critical(f"Course {cid} Proficient exact",
                           num_close(row[4], exp['proficient'], 0),
                           f"Expected {exp['proficient']}, got {row[4]}")
            check_critical(f"Course {cid} Total exact",
                           num_close(row[5], exp['total'], 0),
                           f"Expected {exp['total']}, got {row[5]}")

    # ---- Market_Events sheet (CRITICAL: top-3 symbol + change_pct in order) ----
    print("  Checking Market_Events...")
    me_sheet = get_sheet(agent_wb, "Market_Events")
    check("Sheet 'Market_Events' exists", me_sheet is not None, f"Sheets: {agent_wb.sheetnames}")

    # Un-guarded: missing groundtruth is a CRITICAL failure.
    check_critical("MOEX groundtruth events available",
                   len(expected_events) == 3, f"Got {len(expected_events)} events")

    if me_sheet:
        rows = list(me_sheet.iter_rows(min_row=2, values_only=True))
        check("Market_Events has 3 rows", len(rows) == 3, f"Got {len(rows)}")
        if len(expected_events) == 3:
            for i, exp_event in enumerate(expected_events):
                row = rows[i] if i < len(rows) else None
                row_sym = str(row[1]).strip().upper() if row and row[1] else ""
                exp_sym = exp_event['symbol'].upper()
                # Accept ticker with or without .ME suffix.
                sym_ok = row_sym == exp_sym or row_sym == exp_sym.replace('.ME', '')
                check_critical(f"Event {i+1} symbol is {exp_event['symbol']}",
                               sym_ok, f"Got {row_sym}")
                pct_val = row[2] if row else None
                check_critical(f"Event {i+1} change_pct ~ {exp_event['change_pct']}",
                               pct_val is not None and num_close(pct_val, exp_event['change_pct'], 0.1),
                               f"Expected {exp_event['change_pct']}, got {pct_val}")
        else:
            check_critical("Market_Events matches groundtruth", False, "no groundtruth")
    else:
        check_critical("Market_Events sheet present", False, "missing sheet")

    # ---- Workshop_Schedule sheet (CRITICAL: per-tier combined attendance, tol=0) ----
    print("  Checking Workshop_Schedule...")
    ws_sheet = get_sheet(agent_wb, "Workshop_Schedule")
    check("Sheet 'Workshop_Schedule' exists", ws_sheet is not None, f"Sheets: {agent_wb.sheetnames}")
    if ws_sheet:
        rows = list(ws_sheet.iter_rows(min_row=2, values_only=True))
        check("Workshop_Schedule has 3 rows", len(rows) == 3, f"Got {len(rows)}")

        topics_found = set()
        for row in rows:
            if row and row[1]:
                topics_found.add(str(row[1]).strip().lower())

        check("Has 'Intro to Markets' workshop",
              any("intro" in t and "market" in t for t in topics_found),
              f"Topics: {topics_found}")
        check("Has 'Portfolio Basics' workshop",
              any("portfolio" in t and "basic" in t for t in topics_found),
              f"Topics: {topics_found}")
        check("Has 'Risk Management' workshop",
              any("risk" in t and "manage" in t for t in topics_found),
              f"Topics: {topics_found}")

        if expected_tiers and 16 in expected_tiers and 17 in expected_tiers:
            ns_total = sum(t.get('needs_support', 0) for t in expected_tiers.values())
            dev_total = sum(t.get('developing', 0) for t in expected_tiers.values())
            prof_total = sum(t.get('proficient', 0) for t in expected_tiers.values())
            seen = set()
            for row in rows:
                if not row or not row[2]:
                    continue
                tier = str(row[2]).strip().lower()
                if "needs" in tier or "support" in tier:
                    seen.add("ns")
                    check_critical("Needs Support attendance exact",
                                   num_close(row[3], ns_total, 0),
                                   f"Expected {ns_total}, got {row[3]}")
                elif "develop" in tier:
                    seen.add("dev")
                    check_critical("Developing attendance exact",
                                   num_close(row[3], dev_total, 0),
                                   f"Expected {dev_total}, got {row[3]}")
                elif "proficient" in tier:
                    seen.add("prof")
                    check_critical("Proficient attendance exact",
                                   num_close(row[3], prof_total, 0),
                                   f"Expected {prof_total}, got {row[3]}")
            check_critical("All three tier attendance rows present",
                           seen == {"ns", "dev", "prof"}, f"Saw {seen}")
        else:
            check_critical("Workshop_Schedule attendance verifiable", False, "no tier groundtruth")
    else:
        check_critical("Workshop_Schedule sheet present", False, "missing sheet")


def _overlaps(s1, e1, s2, e2):
    return s1 < e2 and s2 < e1


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE lower(summary) LIKE '%%intro%%market%%'
               OR lower(summary) LIKE '%%portfolio%%basic%%'
               OR lower(summary) LIKE '%%risk%%manage%%'
            ORDER BY start_datetime
        """)
        workshops = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))
        check_critical("3 workshops scheduled correctly", False, str(e))
        return

    check("3 workshop calendar events created", len(workshops) >= 3,
          f"Found {len(workshops)} workshop events")

    # CRITICAL: exactly three distinct workshops scheduled with correct timing/no conflicts.
    check_critical("Exactly 3 workshop events", len(workshops) == 3,
                   f"Found {len(workshops)}")

    if workshops:
        topics = [w[0].lower() for w in workshops]
        check("Calendar has Intro to Markets",
              any("intro" in t and "market" in t for t in topics))
        check("Calendar has Portfolio Basics",
              any("portfolio" in t and "basic" in t for t in topics))
        check("Calendar has Risk Management",
              any("risk" in t and "manage" in t for t in topics))

    conflicts = get_injected_conflicts()

    # CRITICAL: each workshop at 14:00, 90 min, distinct weekday, no overlap with injected conflicts.
    weekdays = set()
    dates = set()
    timing_ok = len(workshops) == 3
    distinct_ok = len(workshops) == 3
    conflict_ok = True
    for w in workshops:
        start, end = w[2], w[3]
        if not start or not end:
            timing_ok = False
            continue
        # 14:00 start
        if not (start.hour == 14 and start.minute == 0):
            timing_ok = False
        # 90 minutes duration
        if abs((end - start).total_seconds() - 90 * 60) > 60:
            timing_ok = False
        # weekday Mon-Fri
        if start.weekday() >= 5:
            timing_ok = False
        # distinct date / weekday
        if start.date() in dates:
            distinct_ok = False
        dates.add(start.date())
        weekdays.add(start.weekday())
        # no overlap with injected 13-16h conflicts
        for cs, ce in conflicts:
            if cs and ce and _overlaps(start, end, cs, ce):
                conflict_ok = False

    check_critical("Each workshop at 14:00 for 90 min on a weekday", timing_ok,
                   f"workshops: {[(str(w[2]), str(w[3])) for w in workshops]}")
    check_critical("Workshops on distinct days", distinct_ok and len(dates) == len(workshops),
                   f"dates: {sorted(str(d) for d in dates)}")
    check_critical("No workshop overlaps injected 13-16h conflicts", conflict_ok,
                   f"conflicts: {[(str(c[0]), str(c[1])) for c in conflicts]}")


def check_emails():
    print("\n=== Checking Emails ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()

        # Announcement email (exact subject).
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE subject = 'Financial Literacy Workshop Series Announcement'
        """)
        student_emails = cur.fetchall()
        check_critical("Announcement email with exact subject sent",
                       len(student_emails) > 0, f"Found {len(student_emails)}")
        ann_to_ok = False
        if student_emails:
            to_str = str(student_emails[0][1]).lower()
            ann_to_ok = "finance_students@university.edu" in to_str
            body = (student_emails[0][2] or "").lower()
            check("Announcement mentions workshops",
                  "intro" in body or "portfolio" in body or "risk" in body or "workshop" in body or "семинар" in body,
                  f"Body length: {len(body)}")
        check_critical("Announcement to finance_students@university.edu", ann_to_ok,
                       f"To: {student_emails[0][1] if student_emails else None}")

        # Department head summary email (exact subject).
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE subject = 'Workshop Planning Summary'
        """)
        head_emails = cur.fetchall()
        check_critical("Summary email with exact subject sent",
                       len(head_emails) > 0, f"Found {len(head_emails)}")
        head_to_ok = False
        if head_emails:
            to_str = str(head_emails[0][1]).lower()
            head_to_ok = "department_head@university.edu" in to_str
        check_critical("Summary to department_head@university.edu", head_to_ok,
                       f"To: {head_emails[0][1] if head_emails else None}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))
        check_critical("Both emails sent correctly", False, str(e))


def check_scripts_and_outputs(agent_workspace):
    print("\n=== Checking Scripts and Output Files ===")

    check("categorize_students.py exists",
          os.path.isfile(os.path.join(agent_workspace, "categorize_students.py")),
          agent_workspace)
    check("find_market_events.py exists",
          os.path.isfile(os.path.join(agent_workspace, "find_market_events.py")),
          agent_workspace)
    check("generate_outline.py exists",
          os.path.isfile(os.path.join(agent_workspace, "generate_outline.py")),
          agent_workspace)

    mej = os.path.join(agent_workspace, "market_events.json")
    check("market_events.json exists", os.path.isfile(mej), agent_workspace)
    if os.path.isfile(mej):
        try:
            with open(mej) as f:
                events = json.load(f)
            if isinstance(events, list):
                check("market_events.json has 3 events", len(events) >= 3,
                      f"Got {len(events)}")
            elif isinstance(events, dict) and "events" in events:
                check("market_events.json has 3 events", len(events["events"]) >= 3,
                      f"Got {len(events['events'])}")
            else:
                check("market_events.json is valid list/dict", False, f"Type: {type(events)}")
        except Exception as e:
            check("market_events.json parseable", False, str(e))

    wmt = os.path.join(agent_workspace, "workshop_materials.txt")
    check("workshop_materials.txt exists", os.path.isfile(wmt), agent_workspace)
    if os.path.isfile(wmt):
        with open(wmt) as f:
            content = f.read().lower()
        check("workshop_materials.txt has content", len(content) > 200,
              f"Length: {len(content)}")
        check("Materials mentions Intro to Markets",
              "intro" in content and "market" in content)
        check("Materials mentions Portfolio",
              "portfolio" in content)
        check("Materials mentions Risk",
              "risk" in content)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_calendar()
    check_emails()
    check_scripts_and_outputs(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    if CRITICAL_FAILED:
        print(f"CRITICAL FAILURES ({len(CRITICAL_FAILED)}):")
        for n in CRITICAL_FAILED:
            print(f"  - {n}")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_failed": CRITICAL_FAILED}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        sys.exit(1)
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()
