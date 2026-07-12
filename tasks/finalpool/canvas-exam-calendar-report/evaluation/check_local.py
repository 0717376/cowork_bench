"""
Check the agent's exam_review_plan.xlsx against the groundtruth.
"""

import os
import openpyxl


def str_match(a, b):
    """Case-insensitive, whitespace-normalized comparison."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=1.0):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_local(agent_workspace, groundtruth_workspace):
    """
    Compare exam_review_plan.xlsx from agent workspace against groundtruth.
    Returns (pass: bool, error_msg: str or None).
    """
    agent_file = os.path.join(agent_workspace, "exam_review_plan.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "exam_review_plan.xlsx")

    if not os.path.isfile(agent_file):
        return False, f"Agent workspace file does not exist: {agent_file}"
    if not os.path.isfile(gt_file):
        return False, f"Groundtruth file does not exist: {gt_file}"

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
        gt_wb = openpyxl.load_workbook(gt_file)
    except Exception as e:
        return False, f"Error reading Excel files: {e}"

    # Find the sheet (case-insensitive)
    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    agent_ws = get_sheet(agent_wb, "Exam Plan")
    gt_ws = get_sheet(gt_wb, "Exam Plan")

    if agent_ws is None:
        return False, "Agent Excel file is missing 'Exam Plan' sheet"
    if gt_ws is None:
        return False, "Groundtruth Excel file is missing 'Exam Plan' sheet"

    # Read rows (skip header)
    agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

    if len(agent_rows) != len(gt_rows):
        return False, (
            f"Row count mismatch: expected {len(gt_rows)} rows, "
            f"got {len(agent_rows)} rows"
        )

    # Build lookup by Course_Code (column index 1)
    differences = []
    matched = 0

    col_names = [
        "Course_Name", "Course_Code", "Instructor_Name", "Instructor_Email",
        "Exam_Name", "Points_Possible", "Exam_Date",
        "Study_Session_1", "Study_Session_2",
    ]

    for gt_row in gt_rows:
        gt_code = str(gt_row[1]).strip() if gt_row[1] else ""
        # Find matching agent row by Course_Code
        agent_match = None
        for ar in agent_rows:
            if ar and str_match(ar[1], gt_code):
                agent_match = ar
                break

        if agent_match is None:
            differences.append(f"Course {gt_code} missing from agent output")
            continue

        row_diffs = []
        for i, col in enumerate(col_names):
            gt_val = gt_row[i] if i < len(gt_row) else None
            ag_val = agent_match[i] if i < len(agent_match) else None

            if col == "Points_Possible":
                if not num_close(ag_val, gt_val, 0.5):
                    row_diffs.append(
                        f"{col}: expected '{gt_val}', got '{ag_val}'"
                    )
            elif (
                col in ("Instructor_Name", "Instructor_Email")
                and gt_code in EXPECTED_INSTRUCTORS
            ):
                # Курсы с несколькими преподавателями: любой допустимый вариант.
                vi = 0 if col == "Instructor_Name" else 1
                if not any(
                    str_match(ag_val, v[vi])
                    for v in EXPECTED_INSTRUCTORS[gt_code]
                ):
                    row_diffs.append(
                        f"{col}: expected '{gt_val}', got '{ag_val}'"
                    )
            else:
                if not str_match(ag_val, gt_val):
                    row_diffs.append(
                        f"{col}: expected '{gt_val}', got '{ag_val}'"
                    )

        if row_diffs:
            differences.append(f"{gt_code}: {'; '.join(row_diffs)}")
        else:
            matched += 1

    if differences:
        return False, (
            f"Matched {matched}/{len(gt_rows)} courses. "
            f"Differences: {'; '.join(differences[:5])}"
        )

    print(f"All {matched} courses matched perfectly.")
    return True, None


# ---------------------------------------------------------------------------
# Semantic sub-checks for CRITICAL gating (read agent xlsx directly).
# These verify the core date-math deliverable and TBD/N-A branching against
# the known-correct values derived from the shipped Canvas dump.
# ---------------------------------------------------------------------------

# code -> (Exam_Date, Study_Session_1, Study_Session_2)
EXPECTED_DATES = {
    "AAA-2013J": ("TBD", "N/A", "N/A"),
    "BBB-2013J": ("TBD", "N/A", "N/A"),
    "DDD-2013J": ("2014-06-19", "2014-06-17 15:00", "2014-06-18 10:00"),
    "EEE-2013J": ("2014-05-24", "2014-05-22 15:00", "2014-05-23 10:00"),
    "FFF-2013J": ("2014-05-25", "2014-05-23 15:00", "2014-05-24 10:00"),
    "GGG-2013J": ("2014-05-18", "2014-05-16 15:00", "2014-05-17 10:00"),
}

# code -> кортеж допустимых вариантов (Instructor_Name, Instructor_Email).
# У BBB-2013J несколько преподавателей: «первый по алфавиту по имени» в русском
# сиде даёт Скворцова (Давид < Кирилл), а по фамилии — Моргунова; принимаем оба.
EXPECTED_INSTRUCTORS = {
    "AAA-2013J": (("Д-р Давид Скворцов", "david.scott@openuniversity.ac.uk"),),
    "BBB-2013J": (
        ("Д-р Кирилл Моргунов", "caleb.morgan@openuniversity.ac.uk"),
        ("Д-р Давид Скворцов", "david.scott@openuniversity.ac.uk"),
    ),
    "DDD-2013J": (("Д-р Эмилия Королёва", "emily.king@openuniversity.ac.uk"),),
    "EEE-2013J": (("Д-р Григорий Власов", "harry.wilson@openuniversity.ac.uk"),),
    "FFF-2013J": (("Д-р Георгий Морозов", "george.allen@openuniversity.ac.uk"),),
    "GGG-2013J": (("Д-р Анна Скворцова", "hannah.scott@openuniversity.ac.uk"),),
}

# Column indices in the sheet.
COL = {
    "Course_Code": 1,
    "Instructor_Name": 2,
    "Instructor_Email": 3,
    "Exam_Date": 6,
    "Study_Session_1": 7,
    "Study_Session_2": 8,
}


def _agent_rows(agent_workspace):
    agent_file = os.path.join(agent_workspace, "exam_review_plan.xlsx")
    if not os.path.isfile(agent_file):
        return None, f"Agent workspace file does not exist: {agent_file}"
    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        return None, f"Error reading Excel file: {e}"
    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "exam plan":
            ws = wb[name]
            break
    if ws is None:
        return None, "Agent Excel file is missing 'Exam Plan' sheet"
    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and r[1]:
            rows[str(r[1]).strip().lower()] = r
    return rows, None


def check_exam_dates(agent_workspace):
    """CRITICAL: exam dates + derived study sessions correct (core date math)."""
    rows, err = _agent_rows(agent_workspace)
    if rows is None:
        return False, err
    diffs = []
    for code, (exam, s1, s2) in EXPECTED_DATES.items():
        r = rows.get(code.lower())
        if r is None:
            diffs.append(f"{code} missing")
            continue
        for label, exp, idx in (
            ("Exam_Date", exam, COL["Exam_Date"]),
            ("Study_Session_1", s1, COL["Study_Session_1"]),
            ("Study_Session_2", s2, COL["Study_Session_2"]),
        ):
            if not str_match(r[idx] if idx < len(r) else None, exp):
                diffs.append(
                    f"{code} {label}: expected '{exp}', got '{r[idx] if idx < len(r) else None}'"
                )
    if diffs:
        return False, "; ".join(diffs[:6])
    return True, None


def check_instructors(agent_workspace):
    """CRITICAL: instructor name/email cross-referenced from Canvas (1st alphabetically)."""
    rows, err = _agent_rows(agent_workspace)
    if rows is None:
        return False, err
    diffs = []
    for code, variants in EXPECTED_INSTRUCTORS.items():
        r = rows.get(code.lower())
        if r is None:
            diffs.append(f"{code} missing")
            continue
        ag_name = r[COL["Instructor_Name"]]
        ag_email = r[COL["Instructor_Email"]]
        # Пара (имя, email) должна целиком совпасть с одним из вариантов.
        if not any(
            str_match(ag_name, name) and str_match(ag_email, email)
            for name, email in variants
        ):
            expected = " | ".join(f"'{n}' <{e}>" for n, e in variants)
            diffs.append(
                f"{code} Instructor: expected {expected}, "
                f"got '{ag_name}' <{ag_email}>"
            )
    if diffs:
        return False, "; ".join(diffs[:6])
    return True, None
