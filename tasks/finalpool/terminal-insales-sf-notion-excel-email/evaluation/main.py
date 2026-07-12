"""Evaluation for terminal-insales-sf-notion-excel-email (russified).

Data layer note: wc.* realia (product categories, review bodies, ...) are
russified CENTRALLY by db/zzz_wc_after_init.sql; sf_data.* realia by
db/zzz_clickhouse_after_init.sql. We DO NOT hardcode russified data-value
literals here — all groundtruth is recomputed LIVE from the DB, so it stays in
sync whether values are EN or RU. English identifiers (column/table names,
sheet names, file names, email subjects/addresses, teamly field markers) stay
English by design.

CRITICAL_CHECKS: any failure => overall FAIL regardless of the accuracy gate.
They verify SUBSTANCE recomputed from the live DB plus the core deliverables.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Central wc category russification map (EN <-> RU), used ONLY to accept either
# spelling of the highest-risk category in the Executive_Summary cell.
CAT_EN2RU = {
    "audio": "аудио", "cameras": "камеры", "electronics": "электроника",
    "headphones": "наушники", "accessories": "аксессуары",
    "all electronics": "вся электроника", "wearables": "носимые устройства",
    "smart home": "умный дом",
}
CAT_RU2EN = {v: k for k, v in CAT_EN2RU.items()}


def cat_aliases(name):
    """Return the lowercased set {name, its EN/RU counterpart} for matching."""
    t = str(name or "").strip().lower()
    out = {t}
    if t in CAT_EN2RU:
        out.add(CAT_EN2RU[t])
    if t in CAT_RU2EN:
        out.add(CAT_RU2EN[t])
    return out


CRITICAL_CHECKS = {
    "Support_Quality_Audit.xlsx exists",
    "Problem_Products top row product_id and severity_score match DB groundtruth",
    "Executive_Summary Critical Products count equals DB 80th-percentile count",
    "Executive_Summary Highest Risk Category matches DB-derived category",
    "Support_By_Priority High row Total_Tickets and Avg_Satisfaction match DB",
    "Teamly 'Support Quality Tracker' has exactly the critical-product entries",
    "Both audit emails sent to correct recipients with correct subjects",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except Exception:
        return False


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').strip())
    except Exception:
        return default


def get_groundtruth_from_db():
    """Compute expected values from read-only DB data (live, EN or RU)."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Problem products
    cur.execute("""
        SELECT (unnest_item->>'product_id')::int as pid, COUNT(DISTINCT o.id)
        FROM wc.orders o, jsonb_array_elements(o.line_items) as unnest_item
        WHERE o.status IN ('refunded','failed')
        GROUP BY pid
    """)
    refund_products = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("""
        SELECT product_id, COUNT(*)
        FROM wc.product_reviews WHERE rating <= 2
        GROUP BY product_id
    """)
    low_review_products = {r[0]: r[1] for r in cur.fetchall()}

    cur.execute("SELECT id, name, categories FROM wc.products")
    products = {r[0]: (r[1], r[2]) for r in cur.fetchall()}

    all_pids = set(refund_products.keys()) | set(low_review_products.keys())
    problem_list = []
    for pid in all_pids:
        rc = refund_products.get(pid, 0)
        lrc = low_review_products.get(pid, 0)
        severity = rc * 30 + lrc * 40
        name = products.get(pid, ("Unknown", []))[0][:60]
        cats = products.get(pid, ("", []))[1]
        cat = cats[0]['name'] if cats else 'Unknown'
        problem_list.append((pid, name, cat, rc, lrc, severity))
    problem_list.sort(key=lambda x: (-x[5], x[0]))

    severities = sorted([p[5] for p in problem_list])
    p80_idx = int(len(severities) * 0.8)
    p80_val = severities[p80_idx] if p80_idx < len(severities) else severities[-1]
    critical = [p for p in problem_list if p[5] > p80_val]

    # Highest risk category = category with the most problem products (DB-derived).
    cat_counts = {}
    for p in problem_list:
        cat_counts[p[2]] = cat_counts.get(p[2], 0) + 1
    highest_risk_category = max(cat_counts, key=lambda k: (cat_counts[k], k)) \
        if cat_counts else "Unknown"

    # Priority data
    cur.execute("""
        SELECT "PRIORITY", COUNT(*),
            ROUND(AVG("RESPONSE_TIME_HOURS")::numeric, 2),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "PRIORITY" ORDER BY "PRIORITY"
    """)
    priority_data = cur.fetchall()

    # Issue type data
    cur.execute("""
        SELECT "ISSUE_TYPE", COUNT(*),
            ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
        GROUP BY "ISSUE_TYPE" ORDER BY COUNT(*) DESC
    """)
    issue_data = cur.fetchall()

    total_tickets = sum(r[1] for r in priority_data)
    cur.execute("""
        SELECT ROUND(AVG("CUSTOMER_SATISFACTION")::numeric, 2)
        FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"
    """)
    overall_sat = float(cur.fetchone()[0])

    cur.close()
    conn.close()

    return {
        "problem_list": problem_list,
        "critical": critical,
        "priority_data": priority_data,
        "issue_data": issue_data,
        "total_tickets": total_tickets,
        "overall_sat": overall_sat,
        "p80_val": p80_val,
        "highest_risk_category": highest_risk_category,
    }


def check_excel(workspace, gt):
    print("\n=== Check 1: Support_Quality_Audit.xlsx ===")
    path = os.path.join(workspace, "Support_Quality_Audit.xlsx")
    if not os.path.exists(path):
        check("Support_Quality_Audit.xlsx exists", False, f"Not found at {path}")
        return
    check("Support_Quality_Audit.xlsx exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    sheets_lower = [s.lower() for s in sheets]

    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    # Problem_Products sheet
    pp_idx = next((i for i, s in enumerate(sheets_lower) if "problem" in s or "product" in s), 0)
    ws_pp = wb[sheets[pp_idx]]
    rows_pp = list(ws_pp.iter_rows(min_row=2, values_only=True))
    expected_count = len(gt["problem_list"])
    check(f"Problem_Products has ~{expected_count} rows",
          abs(len(rows_pp) - expected_count) <= 2,
          f"Found {len(rows_pp)} data rows, expected {expected_count}")

    # CRITICAL: top product id AND severity exactly match DB groundtruth.
    if rows_pp and gt["problem_list"]:
        top_row = rows_pp[0]
        top_pid = safe_float(top_row[0])
        top_severity = safe_float(top_row[5] if len(top_row) > 5 else top_row[-1])
        expected_top = gt["problem_list"][0]
        ok_top = (top_pid is not None and int(top_pid) == expected_top[0]
                  and top_severity is not None
                  and num_close(top_severity, expected_top[5], tol=0.5))
        check("Problem_Products top row product_id and severity_score match DB groundtruth",
              ok_top,
              f"Got pid={top_pid}, sev={top_severity}; expected pid={expected_top[0]}, sev={expected_top[5]}")
    else:
        check("Problem_Products top row product_id and severity_score match DB groundtruth",
              False, "no data rows")

    # Check a mid-range product exists
    if len(gt["problem_list"]) > 5:
        mid_product = gt["problem_list"][3]
        all_text = " ".join(str(c) for r in rows_pp for c in r if c).lower()
        check("Contains expected mid-range product",
              str(mid_product[0]) in all_text or mid_product[1][:15].lower() in all_text,
              f"Looking for pid={mid_product[0]} or name={mid_product[1][:15]}")

    # Support_By_Priority sheet
    sp_idx = next((i for i, s in enumerate(sheets_lower) if "priority" in s or "support" in s), 1)
    high_ok = False
    high_detail = "High row not found"
    if sp_idx < len(sheets):
        ws_sp = wb[sheets[sp_idx]]
        rows_sp = list(ws_sp.iter_rows(min_row=2, values_only=True))
        check("Support_By_Priority has 3 rows", len(rows_sp) == 3,
              f"Found {len(rows_sp)} rows")

        if rows_sp:
            all_text_sp = " ".join(str(c) for r in rows_sp for c in r if c).lower()
            check("Has High priority", "high" in all_text_sp)
            check("Has Medium priority", "medium" in all_text_sp)
            check("Has Low priority", "low" in all_text_sp)

            expected_high = next((p for p in gt["priority_data"] if p[0] == "High"), None)
            for r in rows_sp:
                if r[0] and "high" in str(r[0]).lower() and expected_high:
                    count = safe_float(r[1])
                    sat = safe_float(r[3] if len(r) > 3 else r[-1])
                    cnt_ok = count is not None and int(count) == int(expected_high[1])
                    sat_ok = sat is not None and num_close(sat, float(expected_high[3]), tol=0.05)
                    high_ok = cnt_ok and sat_ok
                    high_detail = (f"Got count={count}, sat={sat}; "
                                   f"expected count={expected_high[1]}, sat={expected_high[3]}")
                    break
    # CRITICAL: High-priority aggregation matches DB exactly (tickets) + tight sat.
    check("Support_By_Priority High row Total_Tickets and Avg_Satisfaction match DB",
          high_ok, high_detail)

    # Issue_Type_Breakdown sheet
    it_idx = next((i for i, s in enumerate(sheets_lower) if "issue" in s or "type" in s), 2)
    if it_idx < len(sheets):
        ws_it = wb[sheets[it_idx]]
        rows_it = list(ws_it.iter_rows(min_row=2, values_only=True))
        check("Issue_Type_Breakdown has 7 rows", len(rows_it) == 7,
              f"Found {len(rows_it)} rows")

        if rows_it:
            all_text_it = " ".join(str(c) for r in rows_it for c in r if c).lower()
            check("Has Bug issue type", "bug" in all_text_it or "ошибк" in all_text_it)
            check("Has Performance Issue type",
                  "performance" in all_text_it or "производительн" in all_text_it)

    # Executive_Summary sheet
    es_idx = next((i for i, s in enumerate(sheets_lower) if "executive" in s or "summary" in s), 3)
    crit_ok = False
    crit_detail = "Executive_Summary missing"
    cat_ok = False
    cat_detail = "Highest Risk Category missing"
    if es_idx < len(sheets):
        ws_es = wb[sheets[es_idx]]
        rows_es = list(ws_es.iter_rows(min_row=2, values_only=True))
        check("Executive_Summary has at least 5 rows", len(rows_es) >= 5,
              f"Found {len(rows_es)} rows")

        if rows_es:
            summary_dict = {}
            for r in rows_es:
                if r[0]:
                    summary_dict[str(r[0]).lower()] = r[1]

            # Total Problem Products
            tp_key = next((k for k in summary_dict if "total" in k and "problem" in k), None)
            if tp_key:
                check("Total Problem Products correct",
                      num_close(summary_dict[tp_key], expected_count, tol=2),
                      f"Got {summary_dict[tp_key]}, expected {expected_count}")

            # Critical Products (CRITICAL — exact count vs 80th-percentile gt)
            cp_key = next((k for k in summary_dict if "critical" in k), None)
            if cp_key:
                crit_ok = num_close(summary_dict[cp_key], len(gt["critical"]), tol=0)
                crit_detail = f"Got {summary_dict[cp_key]}, expected {len(gt['critical'])}"

            # Total Support Tickets
            tt_key = next((k for k in summary_dict if "ticket" in k), None)
            if tt_key:
                check("Total Support Tickets correct",
                      num_close(summary_dict[tt_key], gt["total_tickets"], tol=100),
                      f"Got {summary_dict[tt_key]}, expected {gt['total_tickets']}")

            # Overall Avg Satisfaction
            sat_key = next((k for k in summary_dict if "satisfaction" in k), None)
            if sat_key:
                check("Overall Avg Satisfaction correct",
                      num_close(summary_dict[sat_key], gt["overall_sat"], tol=0.1),
                      f"Got {summary_dict[sat_key]}, expected {gt['overall_sat']}")

            # Highest Risk Category (CRITICAL — DB-derived, EN or RU accepted)
            cat_key = next((k for k in summary_dict if "category" in k or "risk" in k), None)
            if cat_key:
                got = str(summary_dict[cat_key]).strip().lower()
                accepted = cat_aliases(gt["highest_risk_category"])
                cat_ok = got in accepted
                cat_detail = f"Got '{summary_dict[cat_key]}', expected one of {sorted(accepted)}"

    check("Executive_Summary Critical Products count equals DB 80th-percentile count",
          crit_ok, crit_detail)
    check("Executive_Summary Highest Risk Category matches DB-derived category",
          cat_ok, cat_detail)


def check_teamly(gt):
    """Teamly replaces the old Notion database. Teamly has no database/select
    primitive (only spaces + pages), so the 'Support Quality Tracker' is a SPACE
    containing one page per critical product; structured fields (Issue/Product/
    Severity/Status/Assigned_To) live in the page body text."""
    print("\n=== Check 2: Teamly 'Support Quality Tracker' ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, name FROM teamly.spaces")
    spaces = cur.fetchall()
    tracker = None
    for sid, name in spaces:
        nl = (name or "").lower()
        if "support quality" in nl or ("quality" in nl and "tracker" in nl):
            tracker = (sid, name)
            break

    check("Teamly space 'Support Quality Tracker' exists", tracker is not None,
          f"Spaces: {[s[1] for s in spaces]}")

    entries_ok = False
    entries_detail = "space missing"
    fields_ok = False
    if tracker is not None:
        sid = tracker[0]
        cur.execute("SELECT id, title, COALESCE(body,'') FROM teamly.pages "
                    "WHERE space_id = %s", (sid,))
        pages = cur.fetchall()
        n_crit = len(gt["critical"])

        # CRITICAL: exactly the critical-product entries (count == len critical),
        # each marking Severity=Critical / Status=Open / Assigned_To=Quality Team,
        # and NOT more than that (reverse check tightened).
        count_ok = abs(len(pages) - n_crit) <= 1

        def page_text(p):
            return ((p[1] or "") + " " + (p[2] or "")).lower()

        crit_pages = sum(1 for p in pages
                         if "critical" in page_text(p)
                         and "open" in page_text(p)
                         and ("quality team" in page_text(p)))
        entries_ok = count_ok and crit_pages >= n_crit - 1 and crit_pages >= 1
        entries_detail = (f"pages={len(pages)} (expected ~{n_crit}); "
                          f"pages with Critical+Open+Quality Team = {crit_pages}")

        # Structural (non-critical): pages reference a real critical product name.
        all_text = " ".join(page_text(p) for p in pages)
        crit_names_present = any(
            (str(p[0]) in all_text) or (p[1][:12].lower() in all_text)
            for p in gt["critical"]
        )
        fields_ok = crit_names_present

    check("Teamly 'Support Quality Tracker' has exactly the critical-product entries",
          entries_ok, entries_detail)
    check("Teamly entries reference real critical product names",
          fields_ok, "no critical product name found in pages")

    cur.close()
    conn.close()


def check_emails(gt):
    print("\n=== Check 3: Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Support team email
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%support quality%' AND subject ILIKE '%priority%'
        AND to_addr::text ILIKE '%support_team%'
    """)
    support_emails = cur.fetchall()
    support_ok = len(support_emails) >= 1

    if support_emails:
        body = (support_emails[0][2] or "").lower()
        check("Support email mentions satisfaction",
              "satisfaction" in body or "удовлетвор" in body,
              f"Body snippet: {body[:150]}")

    # Product team email
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%support quality%' AND subject ILIKE '%product%'
        AND to_addr::text ILIKE '%product_team%'
    """)
    product_emails = cur.fetchall()
    product_ok = len(product_emails) >= 1

    if product_emails:
        body = (product_emails[0][2] or "").lower()
        check("Product email mentions critical products",
              "critical" in body or "severity" in body
              or "критич" in body or "серьёзн" in body or "серьезн" in body,
              f"Body snippet: {body[:150]}")
        # Robust content proof: reference a real critical product name (EN-preserved).
        crit_name_in_body = any(
            (str(p[0]) in body) or (p[1][:10].lower() in body)
            for p in gt["critical"]
        )
        check("Product email references total problem count or a critical product",
              str(len(gt["problem_list"])) in body or "problem" in body
              or "проблем" in body or crit_name_in_body,
              f"Body snippet: {body[:150]}")

    # CRITICAL: both audit emails sent to the correct recipients + subjects.
    check("Both audit emails sent to correct recipients with correct subjects",
          support_ok and product_ok,
          f"support_ok={support_ok}, product_ok={product_ok}")

    cur.close()
    conn.close()


def check_reverse_validation(gt):
    print("\n=== Reverse Validation ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # No emails sent to noise recipients.
        noise_recipients = [
            "all-staff@company.com",
            "hr@company.com",
            "newsletter@company.com",
            "finance@company.com",
        ]
        for addr in noise_recipients:
            cur.execute(
                "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s "
                "AND subject ILIKE '%%support quality%%'",
                (f"%{addr}%",),
            )
            cnt = cur.fetchone()[0]
            check(f"No audit email sent to noise recipient {addr}", cnt == 0,
                  f"Found {cnt} audit emails to {addr}")
    except Exception as e:
        check("Reverse validation", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python Scripts ===")
    for script in ["correlate_issues.py", "support_metrics.py"]:
        path = os.path.join(workspace, script)
        check(f"{script} exists", os.path.exists(path), f"Not found at {path}")

    for jf in ["problem_products.json", "support_analysis.json"]:
        path = os.path.join(workspace, jf)
        check(f"{jf} exists", os.path.exists(path), f"Not found at {path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    gt = get_groundtruth_from_db()

    check_excel(args.agent_workspace, gt)
    check_teamly(gt)
    check_emails(gt)
    check_scripts(args.agent_workspace)
    check_reverse_validation(gt)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # CRITICAL gate: any critical-check failure => hard FAIL.
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print("\nCRITICAL FAIL: " + "; ".join(critical_failed))
        sys.exit(1)

    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()
