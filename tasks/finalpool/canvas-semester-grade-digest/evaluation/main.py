"""Evaluation for canvas-semester-grade-digest."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def summary_lookup(rows):
    """Build {metric_lower: value} from a Summary sheet's data rows."""
    out = {}
    for row in rows or []:
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
    return out


def course_name_match(a, b, min_shared=2):
    """Permissive course-name match used for strong-name comparisons.

    Course names come from LIVE Canvas and stay English in both groundtruth
    and agent output, so a shared-word overlap is robust to minor formatting.
    """
    a_str = str(a or "").strip().lower()
    b_str = str(b or "").strip().lower()
    if not a_str or not b_str:
        return False
    if a_str == b_str:
        return True
    return len(set(a_str.split()) & set(b_str.split())) >= min_shared


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Grade_Digest.xlsx")
    gt_file = os.path.join(gt_dir, "Grade_Digest.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_failures = []

    # --- Check Course Grades sheet ---
    print("  Checking Course Grades sheet...")
    a_rows = load_sheet_rows(agent_wb, "Course Grades")
    g_rows = load_sheet_rows(gt_wb, "Course Grades")
    if a_rows is None:
        all_errors.append("Sheet 'Course Grades' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Course Grades' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Build lookup by partial course name match (case-insensitive)
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)

            # Try partial match if exact match fails
            if a_row is None:
                for akey, aval in a_lookup.items():
                    # Match on key words from the course name
                    g_words = set(key.split())
                    a_words = set(akey.split())
                    if len(g_words & a_words) >= 3:
                        a_row = aval
                        break

            if a_row is None:
                all_errors.append(f"Missing course: {g_row[0]}")
                continue

            # Enrolled_Students (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 20):
                    all_errors.append(f"{key[:40]}.Enrolled: {a_row[1]} vs {g_row[1]}")

            # Avg_Score (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 2.0):
                    all_errors.append(f"{key[:40]}.Avg_Score: {a_row[2]} vs {g_row[2]}")

            # Min_Score (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5.0):
                    all_errors.append(f"{key[:40]}.Min_Score: {a_row[3]} vs {g_row[3]}")

            # Max_Score (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 5.0):
                    all_errors.append(f"{key[:40]}.Max_Score: {a_row[4]} vs {g_row[4]}")

        if not [e for e in all_errors if "Course Grades" in e or "Missing course" in e]:
            print("    PASS")

    # --- Check Summary sheet ---
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    a_summary = {}
    g_summary = {}
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_summary = summary_lookup(a_data)
        g_summary = summary_lookup(g_data)

        for g_metric, g_val in g_summary.items():
            a_val = a_summary.get(g_metric)
            if a_val is None and g_metric not in a_summary:
                all_errors.append(f"Missing row in Summary: {g_metric}")
                continue

            try:
                float(a_val); float(g_val)
                if not num_close(a_val, g_val, 2.0):
                    all_errors.append(f"Summary.{g_metric}: {a_val} vs {g_val} (tol=2.0)")
            except (TypeError, ValueError):
                # For string comparisons (course names), check if key part matches
                if not course_name_match(a_val, g_val, min_shared=2):
                    all_errors.append(f"Summary.{g_metric}: '{a_val}' vs '{g_val}'")

        if not [e for e in all_errors if "Summary" in e]:
            print("    PASS")

    # --- Check Word document ---
    print("  Checking Word document...")
    word_text = ""
    word_file = os.path.join(args.agent_workspace, "Grade_Summary.docx")
    if not os.path.exists(word_file):
        all_errors.append("Grade_Summary.docx not found")
    else:
        try:
            from docx import Document
            doc = Document(word_file)
            word_text = " ".join(p.text for p in doc.paragraphs)
            if len(word_text.strip()) <= 50:
                all_errors.append("Grade_Summary.docx has too little content")
            else:
                print("    PASS")
        except ImportError:
            if os.path.getsize(word_file) > 100:
                print("    PASS (file exists, size OK)")
            else:
                all_errors.append("Grade_Summary.docx is too small")

    # ============================================================
    # CRITICAL CHECKS — semantic substance. Any failure => exit(1).
    # These verify the core deliverable numbers and rules rather than
    # mere structure. They compare the agent's Summary against the
    # honestly-read groundtruth (both derived from the same live Canvas).
    # ============================================================
    print("\n  Running CRITICAL checks...")

    gt_total_courses = g_summary.get("total_courses")
    gt_overall_avg = g_summary.get("overall_average")
    gt_below70 = g_summary.get("courses_below_70")
    gt_highest = g_summary.get("highest_avg_course")
    gt_lowest = g_summary.get("lowest_avg_course")

    a_total_courses = a_summary.get("total_courses")
    a_overall_avg = a_summary.get("overall_average")
    a_below70 = a_summary.get("courses_below_70")
    a_highest = a_summary.get("highest_avg_course")
    a_lowest = a_summary.get("lowest_avg_course")

    # 1) Total_Courses exactly matches groundtruth (core count of Fall 2014 courses)
    if not (gt_total_courses is not None and a_total_courses is not None
            and num_close(a_total_courses, gt_total_courses, 0)):
        critical_failures.append(
            f"CRITICAL: Total_Courses {a_total_courses} != groundtruth {gt_total_courses}")

    # 2) Overall_Average matches enrollment-WEIGHTED groundtruth within 0.5
    if not (gt_overall_avg is not None and a_overall_avg is not None
            and num_close(a_overall_avg, gt_overall_avg, 0.5)):
        critical_failures.append(
            f"CRITICAL: Overall_Average {a_overall_avg} != groundtruth {gt_overall_avg} (tol=0.5, weighted)")

    # 3) Highest/Lowest avg courses identify the correct courses
    if not course_name_match(a_highest, gt_highest, min_shared=2):
        critical_failures.append(
            f"CRITICAL: Highest_Avg_Course '{a_highest}' != groundtruth '{gt_highest}'")
    if not course_name_match(a_lowest, gt_lowest, min_shared=2):
        critical_failures.append(
            f"CRITICAL: Lowest_Avg_Course '{a_lowest}' != groundtruth '{gt_lowest}'")

    # 4) Courses_Below_70 exactly matches (strictly-below-70 rule)
    if not (gt_below70 is not None and a_below70 is not None
            and num_close(a_below70, gt_below70, 0)):
        critical_failures.append(
            f"CRITICAL: Courses_Below_70 {a_below70} != groundtruth {gt_below70}")

    # 5) Word narrative substantively reports key figures: total course count
    #    AND the best/worst course names. The narrative is written in Russian,
    #    but course names stay English (live Canvas), so match on the English
    #    course-name substrings plus the numeric course count.
    if word_text:
        wt_low = word_text.lower()

        def course_key_tokens(name):
            # Strongest non-trivial token(s) of a course name, e.g. the first
            # significant word, used as a substring probe in the narrative.
            toks = [t for t in str(name or "").lower().replace("&", " ").split()
                    if len(t) >= 4 and t not in ("fall", "2014")]
            return toks

        # total course count present as a number
        count_present = gt_total_courses is not None and (
            str(int(float(gt_total_courses))) in wt_low)

        def narrative_mentions(name):
            toks = course_key_tokens(name)
            # require at least one strong token of the course name to appear
            return any(t in wt_low for t in toks) if toks else False

        if not count_present:
            critical_failures.append(
                f"CRITICAL: Word narrative does not mention total course count ({gt_total_courses})")
        if not narrative_mentions(gt_highest):
            critical_failures.append(
                f"CRITICAL: Word narrative does not mention best course '{gt_highest}'")
        if not narrative_mentions(gt_lowest):
            critical_failures.append(
                f"CRITICAL: Word narrative does not mention worst course '{gt_lowest}'")
    else:
        # No readable text (e.g. python-docx missing) — cannot verify narrative
        # substance; treat as critical only if the file was missing entirely.
        if not os.path.exists(word_file):
            critical_failures.append("CRITICAL: Grade_Summary.docx narrative missing")

    if critical_failures:
        print(f"\n=== CRITICAL FAILURES ({len(critical_failures)}) ===")
        for c in critical_failures:
            print(f"  {c}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    print("    CRITICAL checks PASS")

    # --- Accuracy gate (>=70) over non-critical structural/value checks ---
    # Total structural/value comparison points (Course Grades cells + Summary
    # rows + Word existence). We approximate "checks performed" via error count
    # against a baseline; keep the original strict all_errors gate as well.
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
