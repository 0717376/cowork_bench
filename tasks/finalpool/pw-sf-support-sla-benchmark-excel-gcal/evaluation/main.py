"""Evaluation script for pw-sf-support-sla-benchmark-excel-gcal (ClickHouse variant)."""
import os
import argparse, json, os, sys
from datetime import datetime, timezone
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Industry benchmark response hours published on the source web page
# (http://localhost:30303). These are fixed and must be reproduced verbatim.
WEB_BENCHMARK_RESPONSE = {"Critical": 0.8, "High": 3.5, "Medium": 7.0, "Low": 20.0}

# Expected analysis values derived from the warehouse data (groundtruth).
# Only High / Medium / Low priorities have tickets (no Critical tickets present).
EXPECTED_COMPARISON = {
    "High":   {"Ticket_Count": 6466,  "Our": 6.23,  "Industry": 3.5,  "Gap": 2.73, "CSAT": 3.26, "Status": "Non-Compliant"},
    "Medium": {"Ticket_Count": 15774, "Our": 12.28, "Industry": 7.0,  "Gap": 5.28, "CSAT": 3.26, "Status": "Non-Compliant"},
    "Low":    {"Ticket_Count": 9348,  "Our": 25.76, "Industry": 20.0, "Gap": 5.76, "CSAT": 3.25, "Status": "Non-Compliant"},
}
EXPECTED_ACTION = {
    "High":   {"Gap": 2.73, "Pct": 78.0, "Action": "Process optimization required"},
    "Medium": {"Gap": 5.28, "Pct": 75.4, "Action": "Urgent review needed"},
    "Low":    {"Gap": 5.76, "Pct": 28.8, "Action": "Urgent review needed"},
}
EXPECTED_SUMMARY = {
    "Total_Tickets": 31588,
    "Compliant_Priorities": 0,
    "Non_Compliant_Priorities": 3,
    "Worst_Priority": "Low",
    "Best_Priority": "High",
    "Overall_CSAT": 3.26,
}


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        tag = "CRITICAL FAIL" if critical else "FAIL"
        print(f"  [{tag}] {name}: {detail_str}")
        if critical:
            CRITICAL_FAILS.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def approx(a, b, tol=0.05):
    fa, fb = safe_float(a), safe_float(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= tol


def norm_str(v):
    return str(v).strip().lower() if v is not None else ""


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def header_index(ws):
    """Return {lowercased_header: col_index} for row 1."""
    idx = {}
    for i, c in enumerate(ws[1]):
        if c.value is not None:
            idx[str(c.value).strip().lower()] = i
    return idx


def rows_by_priority(ws):
    """Map priority (capitalized) -> row tuple, keyed by the Priority column."""
    idx = header_index(ws)
    pcol = idx.get("priority")
    out = {}
    if pcol is None:
        return out, idx
    for row in ws.iter_rows(min_row=2, values_only=True):
        if pcol < len(row) and row[pcol] is not None:
            out[str(row[pcol]).strip().capitalize()] = row
    return out, idx


def summary_map(ws):
    """Map Metric -> Value from a two-column Metric/Value sheet."""
    idx = header_index(ws)
    mcol = idx.get("metric", 0)
    vcol = idx.get("value", 1)
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if mcol < len(row) and row[mcol] is not None:
            key = str(row[mcol]).strip()
            out[key] = row[vcol] if vcol < len(row) else None
    return out


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    excel_path = os.path.join(agent_workspace, "SLA_Benchmark_Report.xlsx")
    check("SLA_Benchmark_Report.xlsx exists", os.path.exists(excel_path), critical=True)

    wb = openpyxl.load_workbook(excel_path) if os.path.exists(excel_path) else None

    # ---- Structural checks (non-critical) ----
    if wb is not None:
        check("SLA_Comparison sheet exists", "SLA_Comparison" in wb.sheetnames)
        check("Action_Items sheet exists", "Action_Items" in wb.sheetnames)
        check("Summary sheet exists", "Summary" in wb.sheetnames)

        if "SLA_Comparison" in wb.sheetnames:
            ws = wb["SLA_Comparison"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("SLA_Comparison has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
            headers = [norm_str(c.value) for c in ws[1]]
            for col in ['Priority', 'Ticket_Count', 'Our_Avg_Response_Hrs',
                        'Industry_Avg_Response_Hrs', 'Response_Gap', 'Avg_CSAT', 'Compliance_Status']:
                check(f"SLA_Comparison has {col} column", col.lower() in headers, f"headers: {headers[:8]}")

        if "Action_Items" in wb.sheetnames:
            ws = wb["Action_Items"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Action_Items has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
            headers = [norm_str(c.value) for c in ws[1]]
            for col in ['Priority', 'Response_Gap', 'Improvement_Needed_Pct', 'Recommended_Action']:
                check(f"Action_Items has {col} column", col.lower() in headers, f"headers: {headers[:8]}")

        if "Summary" in wb.sheetnames:
            ws = wb["Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Summary has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            headers = [norm_str(c.value) for c in ws[1]]
            for col in ['Metric', 'Value']:
                check(f"Summary has {col} column", col.lower() in headers, f"headers: {headers[:8]}")

    check("sla_analyzer.py exists", os.path.exists(os.path.join(agent_workspace, "sla_analyzer.py")))

    # ============ CRITICAL SEMANTIC CHECKS ============

    # CRITICAL 1: Industry response benchmarks pulled from the web page match exactly.
    if wb is not None and "SLA_Comparison" in wb.sheetnames:
        by_pri, idx = rows_by_priority(wb["SLA_Comparison"])
        icol = idx.get("industry_avg_response_hrs")
        bench_ok = True
        details = []
        if icol is None:
            bench_ok = False
            details.append("Industry_Avg_Response_Hrs column missing")
        else:
            for pri, row in by_pri.items():
                if pri in WEB_BENCHMARK_RESPONSE:
                    val = row[icol] if icol < len(row) else None
                    if not approx(val, WEB_BENCHMARK_RESPONSE[pri], tol=0.01):
                        bench_ok = False
                        details.append(f"{pri}: got {val}, want {WEB_BENCHMARK_RESPONSE[pri]}")
        check("CRITICAL: Industry benchmarks match web page", bench_ok, "; ".join(details), critical=True)
    else:
        check("CRITICAL: Industry benchmarks match web page", False, "SLA_Comparison missing", critical=True)

    # CRITICAL 2: Response_Gap math + Compliance_Status logic per priority.
    if wb is not None and "SLA_Comparison" in wb.sheetnames:
        by_pri, idx = rows_by_priority(wb["SLA_Comparison"])
        gcol = idx.get("response_gap")
        scol = idx.get("compliance_status")
        ocol = idx.get("our_avg_response_hrs")
        tcol = idx.get("ticket_count")
        ok = True
        details = []
        for pri, exp in EXPECTED_COMPARISON.items():
            if pri not in by_pri:
                ok = False
                details.append(f"missing {pri} row")
                continue
            row = by_pri[pri]
            if ocol is None or not approx(row[ocol], exp["Our"], tol=0.05):
                ok = False; details.append(f"{pri} Our={row[ocol] if ocol is not None else '?'} want {exp['Our']}")
            if tcol is None or safe_float(row[tcol]) != float(exp["Ticket_Count"]):
                ok = False; details.append(f"{pri} Count={row[tcol] if tcol is not None else '?'} want {exp['Ticket_Count']}")
            if gcol is None or not approx(row[gcol], exp["Gap"], tol=0.02):
                ok = False; details.append(f"{pri} Gap={row[gcol] if gcol is not None else '?'} want {exp['Gap']}")
            if scol is None or norm_str(row[scol]) != exp["Status"].lower():
                ok = False; details.append(f"{pri} Status={row[scol] if scol is not None else '?'} want {exp['Status']}")
        check("CRITICAL: Response_Gap math and Compliance_Status correct", ok, "; ".join(details), critical=True)
    else:
        check("CRITICAL: Response_Gap math and Compliance_Status correct", False, "SLA_Comparison missing", critical=True)

    # CRITICAL 3: Summary totals, weighted CSAT, worst/best, compliant counts.
    if wb is not None and "Summary" in wb.sheetnames:
        sm = summary_map(wb["Summary"])
        # normalize keys lowercased for lookup tolerance
        sml = {k.lower(): v for k, v in sm.items()}
        ok = True
        details = []
        def sval(metric):
            return sml.get(metric.lower())
        if safe_float(sval("Total_Tickets")) != float(EXPECTED_SUMMARY["Total_Tickets"]):
            ok = False; details.append(f"Total_Tickets={sval('Total_Tickets')} want {EXPECTED_SUMMARY['Total_Tickets']}")
        if safe_float(sval("Compliant_Priorities")) != float(EXPECTED_SUMMARY["Compliant_Priorities"]):
            ok = False; details.append(f"Compliant_Priorities={sval('Compliant_Priorities')}")
        if safe_float(sval("Non_Compliant_Priorities")) != float(EXPECTED_SUMMARY["Non_Compliant_Priorities"]):
            ok = False; details.append(f"Non_Compliant_Priorities={sval('Non_Compliant_Priorities')}")
        if norm_str(sval("Worst_Priority")) != EXPECTED_SUMMARY["Worst_Priority"].lower():
            ok = False; details.append(f"Worst_Priority={sval('Worst_Priority')} want Low")
        if norm_str(sval("Best_Priority")) != EXPECTED_SUMMARY["Best_Priority"].lower():
            ok = False; details.append(f"Best_Priority={sval('Best_Priority')} want High")
        if not approx(sval("Overall_CSAT"), EXPECTED_SUMMARY["Overall_CSAT"], tol=0.02):
            ok = False; details.append(f"Overall_CSAT={sval('Overall_CSAT')} want 3.26")
        check("CRITICAL: Summary totals/CSAT/worst-best correct", ok, "; ".join(details), critical=True)
    else:
        check("CRITICAL: Summary totals/CSAT/worst-best correct", False, "Summary missing", critical=True)

    # CRITICAL 4: Action_Items recommended actions + improvement pct per rule thresholds.
    if wb is not None and "Action_Items" in wb.sheetnames:
        by_pri, idx = rows_by_priority(wb["Action_Items"])
        acol = idx.get("recommended_action")
        pcol = idx.get("improvement_needed_pct")
        ok = True
        details = []
        for pri, exp in EXPECTED_ACTION.items():
            if pri not in by_pri:
                ok = False; details.append(f"missing {pri}"); continue
            row = by_pri[pri]
            if acol is None or norm_str(row[acol]) != exp["Action"].lower():
                ok = False; details.append(f"{pri} action={row[acol] if acol is not None else '?'} want {exp['Action']}")
            if pcol is None or not approx(row[pcol], exp["Pct"], tol=0.2):
                ok = False; details.append(f"{pri} pct={row[pcol] if pcol is not None else '?'} want {exp['Pct']}")
        check("CRITICAL: Action_Items recommendations and improvement pct correct", ok, "; ".join(details), critical=True)
    else:
        check("CRITICAL: Action_Items recommendations and improvement pct correct", False, "Action_Items missing", critical=True)

    # CRITICAL 5: Both calendar events exist with correct UTC start/end datetimes.
    review_dt = (datetime(2026, 3, 14, 14, 0), datetime(2026, 3, 14, 15, 30))
    workshop_dt = (datetime(2026, 3, 21, 10, 0), datetime(2026, 3, 21, 12, 0))
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT summary, start_datetime, end_datetime, description FROM gcal.events")
        events = cur.fetchall()
        conn.close()

        def naive_utc(dt):
            if dt is None:
                return None
            if getattr(dt, "tzinfo", None) is not None:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            return dt

        def match_event(name_pred, start, end):
            for summ, sdt, edt, desc in events:
                if not name_pred(str(summ).lower()):
                    continue
                s, e = naive_utc(sdt), naive_utc(edt)
                if s == start and e == end:
                    return True, desc
            return False, None

        review_ok, review_desc = match_event(
            lambda s: "sla" in s and "review" in s, review_dt[0], review_dt[1])
        workshop_ok, _ = match_event(
            lambda s: "sla" in s and ("workshop" in s or "improvement" in s),
            workshop_dt[0], workshop_dt[1])

        cal_ok = review_ok and workshop_ok
        # Review description must reference at least one non-compliant priority.
        desc_ok = bool(review_desc) and any(
            p.lower() in str(review_desc).lower() for p in ("high", "medium", "low"))
        check("CRITICAL: Both SLA calendar events at correct UTC datetimes",
              cal_ok, f"review={review_ok} workshop={workshop_ok}", critical=True)
        check("Review event description references non-compliant priorities",
              desc_ok, f"desc={str(review_desc)[:120]}")
    except Exception as e:
        check("CRITICAL: Both SLA calendar events at correct UTC datetimes", False, str(e), critical=True)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    if CRITICAL_FAILS:
        print(f"\nCRITICAL checks failed: {CRITICAL_FAILS}")
        print(f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%) -- FAIL (critical)")
        sys.exit(1)

    success = accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
