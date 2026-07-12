"""Evaluation script for fetch-insales-shipping-audit-excel-word (InSales store).

Структурные проверки (наличие листов/колонок, пороги по строкам) — НЕ критичные.
Содержательные проверки (значения из реальных данных магазина InSales + бенчмарка,
правило расчёта разрыва, агрегаты, алфавитная сортировка, артефакт-результат
Python-скрипта) — КРИТИЧНЫЕ: провал любой => FAIL независимо от accuracy.
Эталон gt_wb уже синхронизирован с централизованно русифицированной схемой wc.*.
"""
import os
import argparse, json, os, sys
import glob as globmod
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Содержательные (обязательные) проверки. Провал любой => итог FAIL.
CRITICAL_CHECKS = {
    "CRITICAL: Category содержит русские категории магазина InSales",
    "CRITICAL: Price_Gap_Pct рассчитан как (Our-Market)/Market*100",
    "CRITICAL: Metrics Total_Products = сумме Product_Count (как в эталоне)",
    "CRITICAL: Data_Analysis отсортирован по Category по алфавиту",
    "CRITICAL: wc_shipping_results.json существует и содержит анализ",
}

# Ожидаемые русские категории магазина InSales (из центральной карты wc.*).
RU_CATEGORIES = {
    "аудио", "камеры", "электроника", "бытовая техника",
    "тв и домашний кинотеатр", "часы",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def col_index(headers, name):
    """Индекс колонки по имени (без учёта регистра)."""
    name = name.strip().lower()
    for i, h in enumerate(headers):
        if (str(h).strip().lower() if h is not None else "") == name:
            return i
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    excel_path = os.path.join(agent_workspace, "Shipping_Audit_Report.xlsx")
    check("Shipping_Audit_Report.xlsx exists", os.path.exists(excel_path))

    # ---- Эталон ----
    gt_path = os.path.join(groundtruth_workspace, "Shipping_Audit_Report.xlsx")
    gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None
    gt_market = {}     # category(lower) -> Market_Avg_Price
    gt_total_products = None
    gt_cat_order = []  # порядок категорий в эталоне (lower)
    if gt_wb is not None and "Data_Analysis" in gt_wb.sheetnames:
        gws = gt_wb["Data_Analysis"]
        gheaders = [c.value for c in gws[1]]
        gi_cat = col_index(gheaders, "Category")
        gi_mkt = col_index(gheaders, "Market_Avg_Price")
        gi_pc = col_index(gheaders, "Product_Count")
        gt_sum = 0
        for row in gws.iter_rows(min_row=2, values_only=True):
            if gi_cat is not None and row[gi_cat]:
                cat = str(row[gi_cat]).strip().lower()
                gt_cat_order.append(cat)
                if gi_mkt is not None:
                    gt_market[cat] = safe_float(row[gi_mkt])
                if gi_pc is not None:
                    pc = safe_float(row[gi_pc])
                    if pc is not None:
                        gt_sum += pc
        gt_total_products = int(gt_sum) if gt_sum else None

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ===== Data_Analysis (структура) =====
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        da_categories = []
        da_rows_dict = {}     # category(lower) -> dict(our, market, gap, pc)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            raw_headers = [c.value for c in ws[1]]
            headers = [str(c).strip().lower() if c else "" for c in raw_headers]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            for expected_col in ['Category', 'Product_Count', 'Our_Avg_Price', 'Total_Sales', 'Market_Avg_Price', 'Price_Gap_Pct']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            i_cat = col_index(raw_headers, "Category")
            i_our = col_index(raw_headers, "Our_Avg_Price")
            i_mkt = col_index(raw_headers, "Market_Avg_Price")
            i_gap = col_index(raw_headers, "Price_Gap_Pct")
            i_pc = col_index(raw_headers, "Product_Count")
            for row in data_rows:
                if i_cat is None or not row[i_cat]:
                    continue
                cat_raw = str(row[i_cat]).strip()
                cat = cat_raw.lower()
                da_categories.append(cat_raw)
                da_rows_dict[cat] = {
                    "our": safe_float(row[i_our]) if i_our is not None else None,
                    "market": safe_float(row[i_mkt]) if i_mkt is not None else None,
                    "gap": safe_float(row[i_gap]) if i_gap is not None else None,
                    "pc": safe_float(row[i_pc]) if i_pc is not None else None,
                }

            # ---- CRITICAL: русские категории InSales ----
            matched = sum(1 for c in da_rows_dict if c in RU_CATEGORIES)
            check("CRITICAL: Category содержит русские категории магазина InSales",
                  matched >= 5,
                  f"совпало {matched} из {len(RU_CATEGORIES)}; категории: {da_categories[:8]}")

            # ---- CRITICAL: Price_Gap_Pct = (Our-Market)/Market*100 ----
            gap_ok = 0
            gap_total = 0
            for cat, d in da_rows_dict.items():
                our, mkt, gap = d["our"], d["market"], d["gap"]
                if our is None or mkt is None or gap is None or mkt == 0:
                    continue
                gap_total += 1
                expected = (our - mkt) / mkt * 100
                if abs(gap - expected) <= 1.0:
                    gap_ok += 1
            check("CRITICAL: Price_Gap_Pct рассчитан как (Our-Market)/Market*100",
                  gap_total >= 5 and gap_ok >= gap_total,
                  f"корректно {gap_ok}/{gap_total} строк")

            # ---- CRITICAL: Market_Avg_Price совпадает с эталоном (реальный бенчмарк) ----
            if gt_market:
                mkt_ok = 0
                mkt_total = 0
                for cat, gm in gt_market.items():
                    if cat in da_rows_dict and da_rows_dict[cat]["market"] is not None and gm is not None:
                        mkt_total += 1
                        if abs(da_rows_dict[cat]["market"] - gm) <= max(1.0, gm * 0.05):
                            mkt_ok += 1
                check("Market_Avg_Price соответствует эталонному бенчмарку",
                      mkt_total >= 5 and mkt_ok >= mkt_total,
                      f"совпало {mkt_ok}/{mkt_total}")

            # ---- CRITICAL: алфавитная сортировка по Category ----
            # Принимаем русскую сортировку ИЛИ порядок эталона (эталон собран
            # по англо-алфавитному порядку исходных категорий-бенчмарка).
            ru_sorted = sorted(da_categories, key=lambda s: s.lower())
            agent_lower = [c.lower() for c in da_categories]
            is_ru_sorted = (da_categories == ru_sorted)
            is_gt_order = bool(gt_cat_order) and agent_lower == gt_cat_order
            check("CRITICAL: Data_Analysis отсортирован по Category по алфавиту",
                  (is_ru_sorted or is_gt_order) and len(da_categories) >= 6,
                  f"got order: {da_categories}")

        # ===== Metrics =====
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # собрать пары Metric -> Value
            metrics = {}
            for row in data_rows:
                if row and row[0]:
                    metrics[str(row[0]).strip().lower()] = safe_float(row[1]) if len(row) > 1 else None

            # ---- CRITICAL: Total_Products = сумме Product_Count ----
            agent_sum = sum(d["pc"] for d in da_rows_dict.values() if d.get("pc") is not None)
            tp = metrics.get("total_products")
            expected_tp = gt_total_products if gt_total_products is not None else agent_sum
            ok_tp = tp is not None and abs(tp - agent_sum) <= 0.5
            if gt_total_products is not None:
                ok_tp = ok_tp and abs(agent_sum - gt_total_products) <= 0.5
            check("CRITICAL: Metrics Total_Products = сумме Product_Count (как в эталоне)",
                  ok_tp,
                  f"Total_Products={tp}, сумма Product_Count={agent_sum}, эталон={expected_tp}")

            # Total_Categories = числу строк Data_Analysis (структурный, не критичный)
            tc = metrics.get("total_categories")
            check("Metrics Total_Categories = числу строк Data_Analysis",
                  tc is not None and abs(tc - len(da_categories)) <= 0.5,
                  f"Total_Categories={tc}, строк={len(da_categories)}")

        # ===== Recommendations =====
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Category']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ===== Word document =====
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        # точное имя (структурно желательно)
        exact_docx = os.path.join(agent_workspace, "Shipping_Audit_Analysis.docx")
        check("Shipping_Audit_Analysis.docx exists (точное имя)", os.path.exists(exact_docx))
        if word_files:
            from docx import Document
            target = exact_docx if os.path.exists(exact_docx) else word_files[0]
            doc = Document(target)
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content", len(text) > 50, f"text length: {len(text)}")

        # ===== Python-скрипт =====
        check("wc_shipping_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "wc_shipping_processor.py")))

        # ---- CRITICAL: wc_shipping_results.json существует и непустой ----
        results_path = os.path.join(agent_workspace, "wc_shipping_results.json")
        results_ok = False
        if os.path.exists(results_path):
            try:
                with open(results_path, "r", encoding="utf-8") as f:
                    payload = json.load(f)
                results_ok = bool(payload) and (
                    isinstance(payload, (list, dict)) and len(payload) > 0
                )
            except Exception as e:
                results_ok = False
        check("CRITICAL: wc_shipping_results.json существует и содержит анализ",
              results_ok,
              f"path={results_path}, exists={os.path.exists(results_path)}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"Critical fails: {CRITICAL_FAILS}")

    success = (not CRITICAL_FAILS) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%), critical_fails={len(CRITICAL_FAILS)}"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
