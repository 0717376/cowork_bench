"""Evaluation script for fetch-sf-hr-worklife-excel-word-gcal (ClickHouse fork).

Internal HR data is read LIVE from the ClickHouse DWH schema sf_data
(HR_ANALYTICS__PUBLIC__EMPLOYEES); the external benchmark is read LIVE from the
fetched mock API (http://localhost:30339/api/data.json, with the on-disk tmp copy
as a fallback). Department literals are russified CENTRALLY by
db/zzz_clickhouse_after_init.sql and the mock JSON is russified in lockstep, so the
agent's department JOIN/gap analysis lines up. The eval reads the department set
dynamically -- no russified literals are hardcoded here.

Structural checks (sheet/column/row-count presence, Word content length, processor
file) are NON-critical. CRITICAL semantic checks verify the agent actually pulled
live warehouse data and computed the Gap = Internal - External rule, and that the
exact 'Analysis Review' meeting (2026-03-14 14:00-15:00 UTC) exists. Any critical
fail => sys.exit(1) before the accuracy gate.
"""
import argparse
import json
import os
import sys
import urllib.request

import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

BENCHMARK_URL = "http://localhost:30339/api/data.json"
TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = " [CRIT]" if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def load_live_internal():
    """Per-department live aggregates from the DWH, keyed by lowercase dept name."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        'SELECT "DEPARTMENT", COUNT(*), '
        'ROUND(AVG("SALARY")::numeric, 2) '
        'FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" '
        'GROUP BY "DEPARTMENT" ORDER BY "DEPARTMENT"'
    )
    rows = cur.fetchall()
    cur.execute('SELECT COUNT(*) FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    total_emp = cur.fetchone()[0]
    cur.close()
    conn.close()
    internal = {}
    for dept, cnt, avg_sal in rows:
        internal[str(dept).strip().lower()] = {
            "department": str(dept).strip(),
            "count": int(cnt),
            "avg_salary": float(avg_sal),
        }
    return internal, int(total_emp)


def load_live_benchmark():
    """Per-department industry_avg from the fetched mock API (live), tmp file fallback."""
    data = None
    try:
        with urllib.request.urlopen(BENCHMARK_URL, timeout=5) as resp:
            data = json.loads(resp.read().decode("utf-8"))
    except Exception:
        path = os.path.join(TASK_ROOT, "tmp", "mock_pages", "api", "data.json")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as fh:
                data = json.load(fh)
    bench = {}
    if data and "benchmarks" in data:
        for b in data["benchmarks"]:
            bench[str(b["department"]).strip().lower()] = float(b["industry_avg"])
    return bench


def find_col(headers, name):
    name = name.lower()
    for i, h in enumerate(headers):
        if h == name:
            return i
    return -1


def run_evaluation(agent_workspace):
    # --- load live ground sources ---
    try:
        internal, total_emp = load_live_internal()
    except Exception as e:
        record("DWH (sf_data) accessible", False, str(e), critical=True)
        internal, total_emp = {}, 0
    bench = load_live_benchmark()
    record("Benchmark API data fetched (>=1 dept)", len(bench) >= 1,
           f"depts: {list(bench)}")

    dept_set = set(internal.keys())  # russified, dynamic
    n_dept = len(dept_set)

    # expected per-dept gaps (Our_Avg_Salary - Industry_Benchmark)
    expected_gap = {}
    for d, info in internal.items():
        if d in bench:
            expected_gap[d] = round(info["avg_salary"] - bench[d], 2)
    avg_gap_expected = (sum(expected_gap.values()) / len(expected_gap)) if expected_gap else None

    # ============ Excel ============
    excel_path = os.path.join(agent_workspace, "Hr_Worklife_Report.xlsx")
    record("Hr_Worklife_Report.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ---- Data_Analysis (structural + CRITICAL semantic) ----
    record("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
    da_by_dept = {}
    if "Data_Analysis" in wb.sheetnames:
        ws = wb["Data_Analysis"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Data_Analysis has >= 7 rows", len(rows) >= 7, f"got {len(rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Department', 'Employee_Count', 'Our_Avg_Salary', 'Avg_Experience',
                    'Avg_Performance', 'Industry_Benchmark', 'Salary_Gap']:
            record(f"Data_Analysis has {col} column", col.lower() in headers,
                   f"headers: {headers[:8]}")

        ci_dept = find_col(headers, 'department')
        ci_avg = find_col(headers, 'our_avg_salary')
        ci_bench = find_col(headers, 'industry_benchmark')
        ci_gap = find_col(headers, 'salary_gap')

        dept_col_values = []
        for row in rows:
            if ci_dept >= 0 and ci_dept < len(row) and row[ci_dept]:
                d = str(row[ci_dept]).strip().lower()
                dept_col_values.append(d)
                da_by_dept[d] = row

        # CRITICAL 1: department rows EXACTLY match the live DWH department set, sorted A-Z
        present = set(dept_col_values)
        sorted_ok = dept_col_values == sorted(dept_col_values)
        record("Data_Analysis departments exactly match live DWH set, sorted A-Z",
               n_dept >= 1 and present == dept_set and sorted_ok,
               f"expected={sorted(dept_set)} got={dept_col_values}", critical=True)

        # CRITICAL 2: Our_Avg_Salary per dept matches live AVG(salary)
        avg_ok = bool(internal)
        avg_bad = []
        for d, info in internal.items():
            row = da_by_dept.get(d)
            v = safe_float(row[ci_avg]) if (row and ci_avg >= 0 and ci_avg < len(row)) else None
            if v is None or not num_close(v, info["avg_salary"], tol=2.0):
                avg_ok = False
                avg_bad.append((info["department"], v, info["avg_salary"]))
        record("Our_Avg_Salary matches live DWH AVG(salary) per department", avg_ok,
               f"mismatches={avg_bad[:4]}", critical=True)

        # CRITICAL 3: Industry_Benchmark matches fetched data.json AND
        #             Salary_Gap == round(Our_Avg_Salary - Industry_Benchmark, 2)
        gap_ok = bool(expected_gap)
        gap_bad = []
        for d in expected_gap:
            row = da_by_dept.get(d)
            if not row:
                gap_ok = False
                gap_bad.append((d, "row missing"))
                continue
            vb = safe_float(row[ci_bench]) if (ci_bench >= 0 and ci_bench < len(row)) else None
            vg = safe_float(row[ci_gap]) if (ci_gap >= 0 and ci_gap < len(row)) else None
            va = safe_float(row[ci_avg]) if (ci_avg >= 0 and ci_avg < len(row)) else None
            if vb is None or not num_close(vb, bench[d], tol=1.0):
                gap_ok = False
                gap_bad.append((d, "bench", vb, bench[d]))
                continue
            # gap must equal internal_avg - benchmark within tolerance
            if vg is None or va is None or not num_close(vg, round(va - vb, 2), tol=2.0):
                gap_ok = False
                gap_bad.append((d, "gap", vg, (va, vb)))
        record("Industry_Benchmark matches API & Salary_Gap = Internal - External",
               gap_ok, f"mismatches={gap_bad[:4]}", critical=True)

    # ---- Metrics (structural + CRITICAL totals) ----
    record("Metrics sheet exists", "Metrics" in wb.sheetnames)
    if "Metrics" in wb.sheetnames:
        ws = wb["Metrics"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Metrics has >= 5 rows", len(rows) >= 5, f"got {len(rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Metric', 'Value']:
            record(f"Metrics has {col} column", col.lower() in headers,
                   f"headers: {headers[:8]}")

        # Build a metric->value map (first col = metric label, find first numeric in row)
        mmap = {}
        for row in rows:
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip().lower()
            val = None
            for c in row[1:]:
                f = safe_float(c)
                if f is not None:
                    val = f
                    break
            mmap[key] = val

        def metric_lookup(*needles):
            for k, v in mmap.items():
                if all(nd in k for nd in needles):
                    return v
            return None

        td = metric_lookup("total", "depart")
        te = metric_lookup("total", "employ")
        ag = metric_lookup("avg", "gap") or metric_lookup("average", "gap") or \
            metric_lookup("salary", "gap")

        ok_td = td is not None and num_close(td, n_dept, tol=0.5)
        ok_te = te is not None and num_close(te, total_emp, tol=0.5)
        ok_ag = (avg_gap_expected is not None and ag is not None
                 and num_close(ag, avg_gap_expected, tol=5.0))
        record("Metrics Total_Departments & Total_Employees & Avg_Salary_Gap correct",
               ok_td and ok_te and ok_ag,
               f"td={td}/{n_dept} te={te}/{total_emp} ag={ag}/{avg_gap_expected}",
               critical=True)

    # ---- Recommendations (structural) ----
    record("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
    if "Recommendations" in wb.sheetnames:
        ws = wb["Recommendations"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Recommendations has >= 3 rows", len(rows) >= 3, f"got {len(rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Priority', 'Action', 'Department', 'Impact']:
            record(f"Recommendations has {col} column", col.lower() in headers,
                   f"headers: {headers[:8]}")

    # ============ Calendar (CRITICAL exact event) ============
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Match summary 'Analysis Review' on 2026-03-14 with start 14:00 and end 15:00.
        # Time-of-day comparison is tz-tolerant: we check the HH:MM regardless of any
        # uniform tz offset by also accepting the literal 14:00->15:00 window on that date.
        cur.execute(
            "SELECT summary, start_datetime, end_datetime FROM gcal.events "
            "WHERE summary ILIKE %s", ('%analysis review%',)
        )
        ev_rows = cur.fetchall()
        conn.close()
    except Exception as e:
        ev_rows = []
        record("Calendar DB accessible", False, str(e), critical=True)

    def fmt(dt):
        try:
            return dt.strftime("%Y-%m-%d %H:%M")
        except Exception:
            return str(dt)

    exact_match = False
    for summary, sdt, edt in ev_rows:
        s = fmt(sdt)
        e = fmt(edt)
        if s.startswith("2026-03-14") and s.endswith("14:00") and \
           e.startswith("2026-03-14") and e.endswith("15:00"):
            exact_match = True
            break
    record("Exactly the 'Analysis Review' event on 2026-03-14 14:00-15:00 UTC exists",
           exact_match,
           f"matching-title events: {[(s, fmt(a), fmt(b)) for s, a, b in ev_rows][:4]}",
           critical=True)

    # ============ Word (structural) ============
    import glob as globmod
    word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
    record("Word document exists", len(word_files) >= 1, f"found {len(word_files)}")
    if word_files:
        from docx import Document
        doc = Document(word_files[0])
        text = " ".join(p.text for p in doc.paragraphs)
        record("Word has content", len(text) > 50, f"len {len(text)}")

    # ============ Processor script (structural) ============
    record("sf_worklife_processor.py exists",
           os.path.exists(os.path.join(agent_workspace, "sf_worklife_processor.py")))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    run_evaluation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (threshold accuracy>=70)")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
