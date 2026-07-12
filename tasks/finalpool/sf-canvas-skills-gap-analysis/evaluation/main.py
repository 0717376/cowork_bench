"""Evaluation for sf-canvas-skills-gap-analysis (ClickHouse / sf_data, russified).

Department-name DATA VALUES come from the ClickHouse sf_data schema, which is
russified CENTRALLY by db/zzz_clickhouse_after_init.sql:
  Engineering->Инженерия, Sales->Продажи, Finance->Финансы, HR->Кадры,
  Operations->Операции, R&D->НИОКР, Support->Поддержка.
The Excel "Department" column therefore holds Russian dept names, so the eval
greps on the Russian values. Skill-area / course names stay English-anchored on
the portal, so skill greps stay English. Column/sheet/file names stay English.

Critical checks (see CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS threshold: accuracy >= 70%.
"""
import argparse
import os
import sys

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Russian dept names (central map) and an English alias kept for robustness.
DEPT_ALIASES = {
    "engineering": ["инженер"],
    "sales": ["продаж"],
    "finance": ["финанс"],
    "hr": ["кадр"],
    "operations": ["операц"],
    "r&d": ["ниокр"],
    "support": ["поддержк"],
}

# Critical semantic checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Operations Gap_Score > 0",
    "Engineering Gap_Score ~0 (threshold 3.21)",
    ">=3 High and >=1 Low priority departments",
    "training emails to all 3 gap-dept heads (Sales/Operations/Support)",
    "all 7 departments present in Department Skills",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def dept_matches(cell, key):
    """True if Excel dept cell `cell` denotes department `key` (RU value or EN alias)."""
    low = str(cell).lower()
    if key in low:  # english alias fallback
        return True
    return any(tok in low for tok in DEPT_ALIASES.get(key, []))


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    import openpyxl

    path = os.path.join(agent_workspace, "Skills_Gap.xlsx")
    if not os.path.exists(path):
        check("Skills_Gap.xlsx exists", False, "file not found")
        return
    check("Skills_Gap.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Skills_Gap.xlsx readable", False, e)
        return
    check("Skills_Gap.xlsx readable", True)

    # ---- Sheet 1: Department Skills ----
    rows = load_sheet_rows(wb, "Department Skills")
    if rows is None:
        check("Sheet 'Department Skills' present", False)
        check("all 7 departments present in Department Skills", False, "sheet missing")
    else:
        check("Sheet 'Department Skills' present", True)
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        check(
            "Department Skills has 7 rows",
            len(data_rows) >= 7,
            f"got {len(data_rows)}",
        )
        dept_names = [str(r[0]).strip() for r in data_rows if r[0]]

        # CRITICAL: all 7 departments present with non-null core fields.
        all_present = True
        missing = []
        for key in DEPT_ALIASES:
            row = next((r for r in data_rows if dept_matches(r[0], key)), None)
            if row is None:
                all_present = False
                missing.append(key)
            else:
                # Required_Skills(1), Employee_Count(2), Avg_Performance(3) non-null
                if len(row) < 4 or row[1] is None or row[2] is None or row[3] is None:
                    all_present = False
                    missing.append(f"{key}(null-field)")
        check(
            "all 7 departments present in Department Skills",
            all_present,
            f"missing/incomplete: {missing}",
        )

        # CRITICAL: Operations gap > 0 (max(0, 3.21 - avg_perf) from warehouse).
        ops = next((r for r in data_rows if dept_matches(r[0], "operations")), None)
        ops_ok = False
        if ops and len(ops) > 5 and ops[5] is not None:
            try:
                ops_ok = float(ops[5]) > 0
            except (ValueError, TypeError):
                ops_ok = False
        check("Operations Gap_Score > 0", ops_ok, ops[5] if ops else "no Operations row")

        # CRITICAL: Engineering gap ~0 (avg perf 3.21 at threshold).
        eng = next((r for r in data_rows if dept_matches(r[0], "engineering")), None)
        eng_ok = False
        if eng and len(eng) > 5 and eng[5] is not None:
            try:
                eng_ok = float(eng[5]) <= 0.05
            except (ValueError, TypeError):
                eng_ok = False
        check(
            "Engineering Gap_Score ~0 (threshold 3.21)",
            eng_ok,
            eng[5] if eng else "no Engineering row",
        )

    # ---- Sheet 2: Training Mapping ----
    rows2 = load_sheet_rows(wb, "Training Mapping")
    if rows2 is None:
        check("Sheet 'Training Mapping' present", False)
    else:
        check("Sheet 'Training Mapping' present", True)
        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        check(
            "Training Mapping has >=10 rows",
            len(data_rows2) >= 10,
            f"got {len(data_rows2)}",
        )
        # Skill areas stay English-anchored on the portal.
        skill_names = [str(r[0]).strip().lower() for r in data_rows2 if r[0]]
        for skill in ["python", "sql", "negotiation"]:
            check(
                f"Training Mapping contains skill '{skill}'",
                any(skill in s for s in skill_names),
            )

    # ---- Sheet 3: Priority Actions ----
    rows3 = load_sheet_rows(wb, "Priority Actions")
    if rows3 is None:
        check("Sheet 'Priority Actions' present", False)
        check(">=3 High and >=1 Low priority departments", False, "sheet missing")
    else:
        check("Sheet 'Priority Actions' present", True)
        data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
        check(
            "Priority Actions has 7 rows",
            len(data_rows3) >= 7,
            f"got {len(data_rows3)}",
        )
        high = [
            r for r in data_rows3
            if len(r) > 2 and r[2] and str(r[2]).strip().lower() == "high"
        ]
        low = [
            r for r in data_rows3
            if len(r) > 2 and r[2] and str(r[2]).strip().lower() == "low"
        ]
        check(
            ">=3 High and >=1 Low priority departments",
            len(high) >= 3 and len(low) >= 1,
            f"high={len(high)} low={len(low)}",
        )


def check_emails():
    try:
        import psycopg2

        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=int(os.environ.get("PGPORT", "5432")),
            dbname=os.environ.get("PGDATABASE", "cowork_gym"),
            user=os.environ.get("PGUSER", "eigent"),
            password=os.environ.get("PGPASSWORD", "camel"),
        )
        cur = conn.cursor()
        # Subject may be EN "Training Recommendations" or RU "Рекомендации по обучению".
        cur.execute(
            """
            SELECT subject, to_addr FROM email.messages
            WHERE subject ILIKE '%training%'
               OR subject ILIKE '%recommendation%'
               OR subject ILIKE '%обучени%'
               OR subject ILIKE '%рекомендац%'
            ORDER BY id DESC LIMIT 20
            """
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("training emails sent (>=3)", False, e)
        check("training emails to all 3 gap-dept heads (Sales/Operations/Support)", False, e)
        return

    check(
        "training emails sent (>=3)",
        len(rows) >= 3,
        f"found {len(rows)}",
    )
    all_to = " ".join(str(r[1]) for r in rows).lower()
    # Sales=m.rodriguez, Operations=r.kim, Support=a.foster — emails kept English.
    gap_dept_emails = ["m.rodriguez", "r.kim", "a.foster"]
    found = [e for e in gap_dept_emails if e in all_to]
    check(
        "training emails to all 3 gap-dept heads (Sales/Operations/Support)",
        len(found) == 3,
        f"found {found}",
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    print("=== Checking Excel file ===")
    check_excel(agent_ws)

    print("\n=== Checking emails ===")
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")
        print("FAIL (critical check failed)")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
