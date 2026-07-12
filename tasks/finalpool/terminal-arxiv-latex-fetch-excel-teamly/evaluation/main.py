"""Evaluation script for terminal-arxiv-latex-fetch-excel-teamly.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Target papers (English identifiers preserved across RU localization).
TARGET_PAPER_IDS = ["2301.07041", "2203.11171", "2205.01068"]

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Conference_Prep_Tracker.xlsx exists",
    "Paper_Sections covers all 3 target papers with >=12 rows and positive word counts",
    "Conference_Schedule covers S1-S5 with correct dates and Related_Papers mapping",
    "Teamly Conference Prep page exists with Status, Paper_Count and conference name",
    "Teamly page summarizes all 3 papers",
    "Presentation_Notes has >=8 slides referencing real papers with intro and conclusion",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Conference_Prep_Tracker.xlsx")
    check("Conference_Prep_Tracker.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        check("Paper_Sections covers all 3 target papers with >=12 rows and positive word counts", False, "no excel")
        check("Conference_Schedule covers S1-S5 with correct dates and Related_Papers mapping", False, "no excel")
        check("Presentation_Notes has >=8 slides referencing real papers with intro and conclusion", False, "no excel")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ---- Paper_Sections (structural + CRITICAL) ----
    check("Paper_Sections sheet exists", "Paper_Sections" in wb.sheetnames)
    if "Paper_Sections" in wb.sheetnames:
        ws = wb["Paper_Sections"]
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in r)]
        check("Paper_Sections has >= 10 rows", len(data_rows) >= 10, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Paper_ID', 'Paper_Title', 'Section_Title', 'Section_Word_Count']:
            check(f"Paper_Sections has {col}", col.lower() in headers, f"headers: {headers[:5]}")
        hmap = {h: i for i, h in enumerate(headers)}
        id_i = hmap.get("paper_id")
        wc_i = hmap.get("section_word_count")
        paper_ids = set()
        if id_i is not None:
            paper_ids = set(str(r[id_i]) for r in data_rows if id_i < len(r) and r[id_i])
        check("Has scaling laws paper", any("2301.07041" in pid for pid in paper_ids), f"IDs: {paper_ids}")
        check("Has RLHF paper", any("2203.11171" in pid for pid in paper_ids), f"IDs: {paper_ids}")
        check("Has OPT paper", any("2205.01068" in pid for pid in paper_ids), f"IDs: {paper_ids}")

        # CRITICAL: all 3 IDs + >=12 rows + every word count positive int.
        all_ids = all(any(tid in pid for pid in paper_ids) for tid in TARGET_PAPER_IDS)
        wc_ok = wc_i is not None
        if wc_ok:
            for r in data_rows:
                if wc_i < len(r) and r[wc_i] is not None:
                    v = safe_float(r[wc_i])
                    if v is None or v <= 0:
                        wc_ok = False
                        break
        critical_ok = all_ids and len(data_rows) >= 12 and wc_ok
        check("Paper_Sections covers all 3 target papers with >=12 rows and positive word counts",
              critical_ok, f"ids={all_ids} rows={len(data_rows)} wc_ok={wc_ok}")
    else:
        check("Paper_Sections covers all 3 target papers with >=12 rows and positive word counts",
              False, "no sheet")

    # ---- Conference_Schedule (structural + CRITICAL) ----
    check("Conference_Schedule sheet exists", "Conference_Schedule" in wb.sheetnames)
    if "Conference_Schedule" in wb.sheetnames:
        ws = wb["Conference_Schedule"]
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in r)]
        check("Conference_Schedule has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Session_ID', 'Session_Title', 'Date', 'Related_Papers']:
            check(f"Conference_Schedule has {col}", col.lower() in headers, f"headers: {headers[:7]}")
        hmap = {h: i for i, h in enumerate(headers)}
        sid_i = hmap.get("session_id")
        date_i = hmap.get("date")
        rel_i = hmap.get("related_papers")

        # Build per-session row map.
        sess = {}
        for r in data_rows:
            if sid_i is not None and sid_i < len(r) and r[sid_i]:
                sess[str(r[sid_i]).strip().upper()] = r

        has_all_sessions = all(s in sess for s in ["S1", "S2", "S3", "S4", "S5"])
        expected_dates = {"S1": "2026-03-28", "S2": "2026-03-28", "S3": "2026-03-29",
                          "S4": "2026-03-29", "S5": "2026-03-30"}
        dates_ok = True
        if date_i is not None and has_all_sessions:
            for s, exp in expected_dates.items():
                cell = sess[s][date_i] if date_i < len(sess[s]) else None
                if cell is None or exp not in str(cell):
                    dates_ok = False
                    break
        else:
            dates_ok = False

        # Related_Papers mapping: scaling-laws paper -> S1, RLHF paper -> S2.
        mapping_ok = False
        if rel_i is not None and "S1" in sess and "S2" in sess:
            s1_rel = str(sess["S1"][rel_i]).lower() if rel_i < len(sess["S1"]) and sess["S1"][rel_i] else ""
            s2_rel = str(sess["S2"][rel_i]).lower() if rel_i < len(sess["S2"]) and sess["S2"][rel_i] else ""
            s1_ok = "scaling law" in s1_rel or "2301.07041" in s1_rel
            s2_ok = ("instruction" in s2_rel or "human feedback" in s2_rel
                     or "rlhf" in s2_rel or "2203.11171" in s2_rel)
            mapping_ok = s1_ok and s2_ok
        check("Conference_Schedule covers S1-S5 with correct dates and Related_Papers mapping",
              has_all_sessions and dates_ok and mapping_ok,
              f"sessions={has_all_sessions} dates={dates_ok} mapping={mapping_ok}")
    else:
        check("Conference_Schedule covers S1-S5 with correct dates and Related_Papers mapping",
              False, "no sheet")

    # ---- Presentation_Notes (structural + CRITICAL) ----
    check("Presentation_Notes sheet exists", "Presentation_Notes" in wb.sheetnames)
    if "Presentation_Notes" in wb.sheetnames:
        ws = wb["Presentation_Notes"]
        data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in r)]
        check("Presentation_Notes has >= 8 rows", len(data_rows) >= 8, f"got {len(data_rows)}")
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for col in ['Slide_Number', 'Topic', 'Key_Points', 'Source_Paper']:
            check(f"Presentation_Notes has {col}", col.lower() in headers, f"headers: {headers[:5]}")
        hmap = {h: i for i, h in enumerate(headers)}
        topic_i = hmap.get("topic")
        src_i = hmap.get("source_paper")

        all_text = " ".join(
            str(c).lower() for r in data_rows for c in r if c is not None
        )
        # At least 2 of the 3 real papers referenced somewhere in notes.
        refs = sum(1 for tid, ttl in [
            ("2301.07041", "scaling law"),
            ("2203.11171", "instruction"),
            ("2205.01068", "opt"),
        ] if tid in all_text or ttl in all_text)
        # Intro + conclusion present among topics.
        topic_text = ""
        if topic_i is not None:
            topic_text = " ".join(
                str(r[topic_i]).lower() for r in data_rows
                if topic_i < len(r) and r[topic_i] is not None
            )
        has_intro = "introduction" in topic_text or "введен" in topic_text or "вступлен" in topic_text
        has_concl = "conclusion" in topic_text or "заключен" in topic_text or "вывод" in topic_text
        crit_ok = len(data_rows) >= 8 and refs >= 2 and has_intro and has_concl
        check("Presentation_Notes has >=8 slides referencing real papers with intro and conclusion",
              crit_ok, f"rows={len(data_rows)} refs={refs} intro={has_intro} concl={has_concl}")
    else:
        check("Presentation_Notes has >=8 slides referencing real papers with intro and conclusion",
              False, "no sheet")


def check_teamly():
    """Teamly knowledge base: the 'Conference Prep Notes' page.

    Seed pages have id <= 3; the agent-created page is queried by id > 3.
    Teamly pages expose title + body (no notion-style properties JSON), so the
    Status / Paper_Count / conference-name values must appear in the body text.
    """
    print("\n=== Checking Teamly ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly accessible", False, str(e))
        check("Teamly Conference Prep page exists with Status, Paper_Count and conference name", False, str(e))
        check("Teamly page summarizes all 3 papers", False, str(e))
        return

    # Identify the page by title: English marker OR RU equivalent.
    page = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if ("conference" in tl and "prep" in tl) or ("подготовк" in tl and "конференц" in tl):
            page = (pid, title, body)
            break
    check("Conference Prep page exists", page is not None,
          f"pages: {[(p[0], p[1]) for p in pages]}")

    if page is None:
        check("Teamly Conference Prep page exists with Status, Paper_Count and conference name",
              False, "no page")
        check("Teamly page summarizes all 3 papers", False, "no page")
        check("Teamly page content blocks present", False, "no page")
        return

    # Use .lower() on ORIGINAL text (never normalize) for RU/EN keyword checks.
    text = ((page[1] or "") + " " + (page[2] or "")).lower()

    has_status = "status" in text and ("in progress" in text or "в работе" in text or "в процессе" in text)
    has_count = ("paper_count" in text or "paper count" in text or "количество стат" in text
                 or "статей: 3" in text) and "3" in text
    has_conf = ("machine learning methods" in text
                or "методам машинного обучения" in text
                or "conference_name" in text or "conference name" in text
                or "название конференции" in text)
    check("Page has Status property", has_status, f"status missing")
    check("Page has Paper_Count property", has_count, f"count missing")
    check("Page has Conference_Name property", has_conf, f"conf name missing")
    check("Teamly Conference Prep page exists with Status, Paper_Count and conference name",
          has_status and has_count and has_conf,
          f"status={has_status} count={has_count} conf={has_conf}")

    # CRITICAL: each of the 3 papers is summarized (id or distinctive title).
    paper_refs = sum(1 for tid, ttl in [
        ("2301.07041", "scaling law"),
        ("2203.11171", "instruction"),
        ("2205.01068", "opt"),
    ] if tid in text or ttl in text)
    check("Teamly page summarizes all 3 papers", paper_refs >= 3,
          f"papers referenced: {paper_refs}/3")

    # Body has substantive content.
    check("Teamly page content blocks present", len((page[2] or "").strip()) >= 100,
          f"body length: {len(page[2] or '')}")


def check_script(agent_workspace):
    print("\n=== Checking Terminal Script ===")
    check("conference_prep_builder.py exists",
          os.path.exists(os.path.join(agent_workspace, "conference_prep_builder.py")))


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output."""
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Conference_Prep_Tracker.xlsx")
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        expected_keywords = {"paper", "section", "conference", "schedule", "presentation", "note"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        if "Paper_Sections" in wb.sheetnames:
            ws = wb["Paper_Sections"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            wc_idx = next((i for i, h in enumerate(headers) if "word_count" in h), None)
            if wc_idx is not None:
                negative = False
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > wc_idx and row[wc_idx] is not None:
                        v = safe_float(row[wc_idx])
                        if v is not None and v < 0:
                            negative = True
                            break
                check("No negative word counts", not negative)

    # Teamly: no duplicate Conference Prep pages.
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
        conf_pages = 0
        for pid, title in pages:
            tl = (title or "").lower()
            if ("conference" in tl and "prep" in tl) or ("подготовк" in tl and "конференц" in tl):
                conf_pages += 1
        check("No duplicate Conference Prep pages", conf_pages <= 1,
              f"Found {conf_pages} matching pages")
    except Exception:
        pass


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    check_excel(agent_workspace)
    check_teamly()
    check_script(agent_workspace)
    check_reverse_validation(agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
