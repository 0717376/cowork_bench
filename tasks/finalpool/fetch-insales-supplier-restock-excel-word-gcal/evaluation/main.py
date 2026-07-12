"""Evaluation for fetch-insales-supplier-restock-excel-word-gcal."""
import argparse
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic critical checks: any failure here => overall FAIL regardless of
# accuracy. These verify the load-bearing deliverables demanded by task.md
# (the computed reorder qty/cost for Электроника, its computed margin, the
# zero-stock products that drive Critical urgency, the aggregate procurement
# cost in the Word report, and the three dated April supplier meetings) so a
# non-doing agent cannot slip through on the accuracy gate, while a correctly
# doing RU agent passes.
CRITICAL_CHECKS = {
    "CRITICAL: Электроника reorder qty=200 and cost=6000",
    "CRITICAL: Электроника computed margin ~50.9%",
    "CRITICAL: Current Inventory contains zero-stock products",
    "CRITICAL: Word report total procurement cost ~15,940",
    "CRITICAL: three April supplier meetings (Critical/High/Medium)",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    import openpyxl
    path = os.path.join(agent_workspace, "Restocking_Plan.xlsx")
    if not os.path.exists(path):
        check("Restocking_Plan.xlsx exists", False, "file not found")
        return
    # data_only=True returns cached formula results; if a cell holds an
    # uncomputed formula it reads as None. We compare against expected numbers
    # with tolerance, so a populated-but-uncached cell still fails honestly
    # rather than false-passing.
    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Current Inventory ----
    rows = load_sheet_rows(wb, "Current Inventory")
    if rows is None:
        check("Sheet 'Current Inventory' present", False, "missing")
        check("CRITICAL: Current Inventory contains zero-stock products", False,
              "sheet missing")
    else:
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        check("Current Inventory has ~30 rows", len(data_rows) >= 25,
              f"{len(data_rows)} rows")
        zero_stock = [r for r in data_rows if r[1] is not None and float(r[1]) == 0]
        # CRITICAL: at least one zero-stock product must be present; these are
        # what make Электроника (and other categories) Critical urgency.
        check("CRITICAL: Current Inventory contains zero-stock products",
              len(zero_stock) >= 1, f"{len(zero_stock)} zero-stock rows")

    # ---- Supplier Pricing ----
    rows2 = load_sheet_rows(wb, "Supplier Pricing")
    if rows2 is None:
        check("Sheet 'Supplier Pricing' present", False, "missing")
        check("CRITICAL: Электроника computed margin ~50.9%", False, "sheet missing")
    else:
        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Supplier Pricing has 6 rows", len(data_rows2) >= 6,
              f"{len(data_rows2)} rows")
        cats = {str(r[0]).strip().lower() for r in data_rows2 if r[0]}
        # Categories are russified centrally (scripts/wc_relabel_map.py CATEGORIES);
        # the supplier catalog, store inventory and groundtruth all serve Russian
        # category names, so the eval must expect Russian here.
        for c in ["электроника", "камеры", "часы"]:
            check(f"Supplier Pricing has category '{c}'", c in cats, sorted(cats))
        elec = [r for r in data_rows2 if r[0] and "электроника" == str(r[0]).strip().lower()]
        # CRITICAL: the computed margin for Электроника proves the agent applied
        # the margin formula against the right wholesale/retail figures.
        margin_ok = bool(elec) and len(elec[0]) >= 6 and num_close(elec[0][5], 50.9, 3.0)
        check("CRITICAL: Электроника computed margin ~50.9%", margin_ok,
              f"margin={elec[0][5] if elec else 'N/A'}")
        retail_ok = bool(elec) and len(elec[0]) >= 5 and num_close(elec[0][4], 61.13, 5.0)
        check("Электроника retail avg ~61.13", retail_ok,
              f"retail_avg={elec[0][4] if elec else 'N/A'}")

    # ---- Reorder Recommendations ----
    rows3 = load_sheet_rows(wb, "Reorder Recommendations")
    if rows3 is None:
        check("Sheet 'Reorder Recommendations' present", False, "missing")
        check("CRITICAL: Электроника reorder qty=200 and cost=6000", False,
              "sheet missing")
    else:
        data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("Reorder Recommendations has 6 rows", len(data_rows3) >= 6,
              f"{len(data_rows3)} rows")
        critical = [r for r in data_rows3 if r[1] and str(r[1]).strip().lower() == "critical"]
        check("5 Critical categories", len(critical) >= 5,
              f"{len(critical)} critical")
        watches = [r for r in data_rows3 if r[0] and "часы" in str(r[0]).lower()]
        check("Часы urgency=High",
              bool(watches) and str(watches[0][1]).strip().lower() == "high",
              f"urgency={watches[0][1] if watches else 'N/A'}")
        elec_r = [r for r in data_rows3 if r[0] and "электроника" == str(r[0]).strip().lower()]
        # CRITICAL: Электроника is Critical (a zero-stock category), so qty must
        # be 2x its min order (100 -> 200) and cost = 200 * wholesale 30 = 6000.
        # This proves the urgency classification and the cost computation.
        qcost_ok = (bool(elec_r) and len(elec_r[0]) >= 4
                    and num_close(elec_r[0][2], 200, 10)
                    and num_close(elec_r[0][3], 6000, 100))
        check("CRITICAL: Электроника reorder qty=200 and cost=6000", qcost_ok,
              f"qty={elec_r[0][2] if elec_r else 'N/A'}, cost={elec_r[0][3] if elec_r else 'N/A'}")


def check_word(agent_workspace):
    path = os.path.join(agent_workspace, "Procurement_Report.docx")
    if not os.path.exists(path):
        check("Procurement_Report.docx exists", False, "file not found")
        check("CRITICAL: Word report total procurement cost ~15,940", False,
              "file not found")
        return
    from docx import Document
    import re
    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()
    # Locale-agnostic digit normalization: a Russian report uses a space (or NBSP/
    # thin space) thousands separator (15 940), English uses a comma (15,940).
    # Strip those so "15 940", "15 940", "15,940" and "15940" all match.
    full_text_digits = re.sub(r"[\s,  ]", "", full_text)
    check("Word doc has expected title content",
          "q2 2026" in full_text or "procurement" in full_text,
          "title content missing")
    check("Word doc discusses inventory status",
          "zero stock" in full_text or "critical" in full_text or "0 stock" in full_text,
          "status discussion missing")
    # CRITICAL: the aggregate procurement cost (sum of all Estimated_Cost rows =
    # 4500+480+6000+1080+3600+280 = 15,940) is the headline computed figure of
    # the report; its presence proves the agent aggregated the reorder costs.
    check("CRITICAL: Word report total procurement cost ~15,940",
          "15940" in full_text_digits,
          "total cost ~15,940 not found")


def check_gcal():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date FROM gcal.events
            WHERE start_datetime >= '2026-04-01' AND start_datetime < '2026-05-01'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("CRITICAL: three April supplier meetings (Critical/High/Medium)",
              False, f"DB error: {e}")
        return

    summaries = [r[0].lower() if r[0] else "" for r in rows]
    has_critical = any("critical" in s for s in summaries)
    has_high = any("high" in s for s in summaries)
    has_medium = any("medium" in s for s in summaries)
    # CRITICAL: all three dated supplier meetings must exist with the correct
    # priority summaries; these are the calendar deliverable of the task.
    check("CRITICAL: three April supplier meetings (Critical/High/Medium)",
          len(rows) >= 3 and has_critical and has_high and has_medium,
          f"found {len(rows)} events; "
          f"critical={has_critical} high={has_high} medium={has_medium}")


def run_evaluation(agent_ws):
    print("  Checking Excel file...")
    check_excel(agent_ws)
    print("  Checking Word document...")
    check_word(agent_ws)
    print("  Checking GCal events...")
    check_gcal()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"  - {n}")

    # Convention: any critical failure => FAIL (sys.exit(1)) before the
    # accuracy gate; otherwise PASS requires accuracy >= 70%.
    if critical_failed:
        print("=== RESULT: FAIL (critical check failed) ===")
        return False
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        return True
    print(f"=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
    return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace")

    success = run_evaluation(agent_ws)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
