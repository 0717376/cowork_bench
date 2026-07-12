"""Evaluation for insales-customer-retention-email."""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2


def normalize_ru_numbers(text):
    """RU number normalization for substring/regex checks: collapse digit-group
    separators (space/NBSP/NNBSP/dot/comma before a 3-digit group) and turn
    decimal commas into dots ("31 588" -> "31588", "4 586,91" -> "4586.91")."""
    t = str(text or "")
    t = re.sub(r"(?<=\d)[ \xa0\u202f\u2009.,](?=\d{3}\b)", "", t)
    return re.sub(r"(?<=\d),(?=\d)", ".", t)

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}{detail_str}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check the Excel output against groundtruth."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "VIP_Customer_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "VIP_Customer_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    # Find VIP Customers sheet (case-insensitive)
    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    agent_ws = get_sheet(agent_wb, "VIP Customers")
    gt_ws = get_sheet(gt_wb, "VIP Customers")

    check("Sheet 'VIP Customers' exists", agent_ws is not None,
          f"Found sheets: {agent_wb.sheetnames}")
    if not agent_ws or not gt_ws:
        return False

    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
    agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True)
                  if r and any(c is not None for c in r)]

    check("Row count matches (10 VIP customers)", len(agent_rows) == 10,
          f"Expected 10, got {len(agent_rows)}")

    # Authoritative top-10 from DB (wc.customers ORDER BY total_spent DESC LIMIT 10)
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT email, total_spent::numeric
        FROM wc.customers
        ORDER BY total_spent::numeric DESC
        LIMIT 10
    """)
    db_top10 = cur.fetchall()
    conn.close()
    db_top10_emails = [str(e).strip().lower() for e, _ in db_top10]
    db_top10_set = set(db_top10_emails)
    db_total_by_email = {str(e).strip().lower(): float(t) for e, t in db_top10}

    agent_emails_ordered = [str(ar[2]).strip().lower() if len(ar) >= 3 and ar[2] is not None else None
                            for ar in agent_rows]
    agent_set = set(e for e in agent_emails_ordered if e)

    # CRITICAL: exact top-10 set match
    check("Excel top-10 customer set matches DB exactly", agent_set == db_top10_set,
          f"Missing: {db_top10_set - agent_set}; Extra: {agent_set - db_top10_set}",
          critical=True)

    # CRITICAL: rank order matches total_spent DESC (rank tol=0)
    rank_order_ok = True
    for ar in agent_rows:
        if not ar or len(ar) < 5 or ar[2] is None:
            rank_order_ok = False
            continue
        em = str(ar[2]).strip().lower()
        if em in db_top10_emails:
            expected_rank = db_top10_emails.index(em) + 1
            if not num_close(ar[0], expected_rank, 0):
                rank_order_ok = False
    check("Excel rank order matches total_spent DESC (tol=0)", rank_order_ok,
          "One or more ranks do not match DB ordering", critical=True)

    # CRITICAL: total_spent values match DB rounded to 2dp (tight tol<=0.01)
    totals_ok = True
    for ar in agent_rows:
        if not ar or len(ar) < 5 or ar[2] is None:
            totals_ok = False
            continue
        em = str(ar[2]).strip().lower()
        if em in db_total_by_email:
            exp = round(db_total_by_email[em], 2)
            if not num_close(ar[4], exp, 0.01):
                totals_ok = False
    check("Excel total_spent values match DB (tol<=0.01)", totals_ok,
          "One or more total_spent values mismatch DB", critical=True)

    all_ok = True
    for gt_row in gt_rows:
        rank, name, email, orders, total = gt_row
        matched = None
        for ar in agent_rows:
            if ar and len(ar) >= 5 and str_match(ar[2], email):
                matched = ar
                break
        if matched:
            check(f"Customer {email} rank", num_close(matched[0], rank, 0),
                  f"Expected {rank}, got {matched[0]}")
            check(f"Customer {email} orders_count",
                  num_close(matched[3], orders, 0),
                  f"Expected {orders}, got {matched[3]}")
            check(f"Customer {email} total_spent",
                  num_close(matched[4], total, 0.01),
                  f"Expected {total}, got {matched[4]}")
        else:
            check(f"Customer {email} found", False, "Not in agent output")
            all_ok = False

    return all_ok


def check_emails():
    """Check that VIP emails were sent to the correct addresses."""
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Get expected top 10 customer emails
    cur.execute("""
        SELECT email, first_name, total_spent
        FROM wc.customers
        ORDER BY total_spent::numeric DESC
        LIMIT 10
    """)
    expected_customers = cur.fetchall()

    # Get sent emails
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE folder_id = 2
    """)
    sent_emails = cur.fetchall()

    # Also check all messages
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    conn.close()

    def find_email_for_recipient(recipient):
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    EXPECTED_SUBJECT = "thank you, vip customer!"
    EXPECTED_FROM = "vip-program@store.example.com"

    expected_recipient_set = set(str(e).strip().lower() for e, _, _ in expected_customers)

    def recipients_of(to_addr):
        if not to_addr:
            return []
        if isinstance(to_addr, list):
            return [str(r).strip().lower() for r in to_addr]
        try:
            parsed = json.loads(to_addr)
            if isinstance(parsed, list):
                return [str(r).strip().lower() for r in parsed]
        except (json.JSONDecodeError, TypeError):
            pass
        return [str(to_addr).strip().lower()]

    def number_variants(amount):
        """RU/EN formatting variants of a money amount for substring matching."""
        f = float(amount)
        variants = set()
        for s in (f"{f:.2f}", f"{round(f, 2)}", str(f)):
            variants.add(s)
            variants.add(s.replace(".", ","))
        # integer part alone (in case body drops decimals)
        variants.add(str(int(round(f))))
        return variants

    # CRITICAL: each top-10 customer received >=1 matching email
    each_received = True
    subject_ok_all = True
    from_ok_all = True
    body_ok_all = True
    for email_addr, fname, total in expected_customers:
        result = find_email_for_recipient(email_addr)
        if result is None:
            each_received = False
            check(f"Email sent to {email_addr}", False, "No email found")
            continue
        check(f"Email sent to {email_addr}", True)
        subj, from_addr, to_addr, body = result
        if str(subj or "").strip().lower() != EXPECTED_SUBJECT:
            subject_ok_all = False
        if str(from_addr or "").strip().lower() != EXPECTED_FROM:
            from_ok_all = False
        body_l = normalize_ru_numbers(str(body or "")).lower()
        # russified first_name (assert against DB first_name, NOT normalized)
        has_name = bool(fname) and str(fname).strip().lower() in body_l
        has_total = any(v in body_l for v in number_variants(total))
        if not (has_name and has_total):
            body_ok_all = False
            check(f"Body to {email_addr} has name+total", False,
                  f"name={has_name} total={has_total}")

    check("All top-10 customers received an email", each_received,
          "One or more top-10 recipients missing", critical=True)
    check("All VIP emails subject == 'Thank You, VIP Customer!'", subject_ok_all,
          "Subject mismatch on one or more emails", critical=True)
    check(f"All VIP emails from == {EXPECTED_FROM}", from_ok_all,
          "from_addr mismatch on one or more emails", critical=True)
    check("All VIP email bodies contain first_name + total_spent", body_ok_all,
          "Body personalization missing on one or more emails", critical=True)

    # CRITICAL: recipient set is exactly the top-10 (no extra recipients)
    actual_recipients = set()
    for subj, from_addr, to_addr, body in all_emails:
        if str(subj or "").strip().lower() == EXPECTED_SUBJECT:
            for r in recipients_of(to_addr):
                actual_recipients.add(r)
    extras = actual_recipients - expected_recipient_set
    check("No VIP email sent to a non-top-10 recipient", len(extras) == 0,
          f"Unexpected recipients: {extras}", critical=True)

    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("WC CUSTOMER RETENTION EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, gt_dir)
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
