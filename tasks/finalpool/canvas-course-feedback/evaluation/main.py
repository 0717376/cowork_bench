"""
Evaluation для canvas-course-feedback (RU stack: canvas + forms + excel).

Canvas — read-only английские данные курсов (названия/коды на английском).
Форма обратной связи создаётся через RU forms MCP (schema gform.*).
Excel-отчёт Fall_2014_Course_Report.xlsx — лист 'Course Statistics'.

Критические чеки (CRITICAL_CHECKS): любой их fail => задача FAIL, независимо
от общей accuracy. Это семантические чеки сути задачи (форма-ядро присутствует,
правильные значения и вычисляемые колонки), а НЕ структура (лист/колонка есть).
Порог прохождения: нет критических fail И accuracy >= 70%.
"""

import argparse
import os
import sys
import json
import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# Ground truth data from Canvas (read-only). Названия курсов — английские
# (canvas хранит их по-английски, не русифицируем).
EXPECTED_COURSES = [
    {
        "Course_Code": "AAA-2014J",
        "Course_Name": "Прикладная аналитика и алгоритмы",
        "Enrollment": 365,
        "Avg_Score": 67.85,
        "Assignment_Count": 6,
        "Quiz_Count": 0,
        "Total_Assessments": 6,
    },
    {
        "Course_Code": "BBB-2014J",
        "Course_Name": "Биохимия и биоинформатика",
        "Enrollment": 2292,
        "Avg_Score": 64.31,
        "Assignment_Count": 6,
        "Quiz_Count": 0,
        "Total_Assessments": 6,
    },
    {
        "Course_Code": "CCC-2014J",
        "Course_Name": "Креативные вычисления и культура",
        "Enrollment": 2498,
        "Avg_Score": 70.22,
        "Assignment_Count": 10,
        "Quiz_Count": 4,
        "Total_Assessments": 14,
    },
    {
        "Course_Code": "DDD-2014J",
        "Course_Name": "Проектирование на основе данных",
        "Enrollment": 1803,
        "Avg_Score": 69.99,
        "Assignment_Count": 7,
        "Quiz_Count": 0,
        "Total_Assessments": 7,
    },
    {
        "Course_Code": "EEE-2014J",
        "Course_Name": "Экологическая экономика и этика",
        "Enrollment": 1188,
        "Avg_Score": 81.27,
        "Assignment_Count": 5,
        "Quiz_Count": 0,
        "Total_Assessments": 5,
    },
    {
        "Course_Code": "FFF-2014J",
        "Course_Name": "Основы финансов",
        "Enrollment": 2365,
        "Avg_Score": 76.51,
        "Assignment_Count": 13,
        "Quiz_Count": 7,
        "Total_Assessments": 20,
    },
    {
        "Course_Code": "GGG-2014J",
        "Course_Name": "Глобальное управление и геополитика",
        "Enrollment": 749,
        "Avg_Score": 76.60,
        "Assignment_Count": 10,
        "Quiz_Count": 6,
        "Total_Assessments": 16,
    },
]

EXPECTED_BY_CODE = {c["Course_Code"]: c for c in EXPECTED_COURSES}

# Критические чеки — по строке name, как передаётся в record()
CRITICAL_CHECKS = {
    "Форма Fall 2014 Feedback: 5 вопросов, >=2 choiceQuestion и >=2 textQuestion",
    "Форма: вопрос выбора курса со всеми 7 курсами (7 вариантов)",
    "Excel: лист 'Course Statistics' с 7 строками по всем кодам AAA..GGG-2014J",
    "Excel: Avg_Score (±2.0) И Total_Assessments == Assignment_Count + Quiz_Count",
    "Excel: Course_Name — короткие названия БЕЗ суффикса семестра/2014J",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def _config_options(cfg):
    cfg = cfg if isinstance(cfg, dict) else json.loads(cfg) if cfg else {}
    return cfg.get("options", []) or []


def check_gform():
    """Форма Fall 2014 Feedback создана в gform.* с правильной структурой."""
    print("\n=== Check 1: Форма обратной связи (forms / gform.*) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Форма обратной связи Fall 2014 создана", False, str(e))
        record("Форма Fall 2014 Feedback: 5 вопросов, >=2 choiceQuestion и >=2 textQuestion", False, "no db")
        record("Форма: вопрос выбора курса со всеми 7 курсами (7 вариантов)", False, "no db")
        return
    cur = conn.cursor()

    # Заголовок формы содержит маркеры 'fall 2014' + 'feedback' (eval-идентификатор,
    # остаётся на английском). Запасной вариант — RU 'отзыв'/'обратн'.
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE (LOWER(title) LIKE '%fall 2014%' AND LOWER(title) LIKE '%feedback%')
           OR (LOWER(title) LIKE '%2014%' AND (LOWER(title) LIKE '%отзыв%'
                                               OR LOWER(title) LIKE '%обратн%'))
        ORDER BY id
    """)
    forms = cur.fetchall()
    record("Форма обратной связи Fall 2014 создана", len(forms) >= 1,
           f"Found forms: {[f[1] for f in forms]}")

    if not forms:
        record("Форма Fall 2014 Feedback: 5 вопросов, >=2 choiceQuestion и >=2 textQuestion", False, "no form")
        record("Форма: вопрос выбора курса со всеми 7 курсами (7 вариантов)", False, "no form")
        cur.close()
        conn.close()
        return

    form_id, form_title = forms[0][0], forms[0][1]
    print(f"  Найдена форма: '{form_title}' (id={form_id})")

    cur.execute(
        "SELECT id, title, question_type, config FROM gform.questions WHERE form_id = %s ORDER BY position",
        (form_id,),
    )
    questions = cur.fetchall()
    cur.close()
    conn.close()

    n_q = len(questions)
    radio_count = sum(1 for q in questions if q[2] == "choiceQuestion")
    text_count = sum(1 for q in questions if q[2] == "textQuestion")

    record("Форма содержит ровно 5 вопросов", n_q == 5, f"found {n_q}")
    record("Не менее 2 вопросов с выбором (choiceQuestion)", radio_count >= 2, f"found {radio_count}")
    record("Не менее 2 свободных текстовых вопросов (textQuestion)", text_count >= 2, f"found {text_count}")

    # CRITICAL: ядро формы присутствует
    record("Форма Fall 2014 Feedback: 5 вопросов, >=2 choiceQuestion и >=2 textQuestion",
           (n_q == 5 and radio_count >= 2 and text_count >= 2),
           f"n={n_q} choice={radio_count} text={text_count}")

    # Вопрос выбора курса: единственный choice-вопрос ровно с 7 вариантами.
    # Сопоставляем по числу вариантов (язык заголовка не важен).
    seven_opt_questions = [
        q for q in questions
        if q[2] == "choiceQuestion" and len(_config_options(q[3])) == 7
    ]
    record("Форма: вопрос выбора курса со всеми 7 курсами (7 вариантов)",
           len(seven_opt_questions) >= 1,
           f"choice-вопросов с 7 вариантами: {len(seven_opt_questions)}")


def check_excel(agent_workspace):
    """Excel-отчёт против эталонных данных курсов."""
    print("\n=== Check 2: Excel-отчёт Fall_2014_Course_Report.xlsx ===")

    excel_path = os.path.join(agent_workspace, "Fall_2014_Course_Report.xlsx")
    if not os.path.exists(excel_path):
        record("Файл Fall_2014_Course_Report.xlsx существует", False, excel_path)
        record("Excel: лист 'Course Statistics' с 7 строками по всем кодам AAA..GGG-2014J", False, "no file")
        record("Excel: Avg_Score (±2.0) И Total_Assessments == Assignment_Count + Quiz_Count", False, "no file")
        record("Excel: Course_Name — короткие названия БЕЗ суффикса семестра/2014J", False, "no file")
        return
    record("Файл Fall_2014_Course_Report.xlsx существует", True)

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        record("Файл открывается openpyxl", False, str(e))
        record("Excel: лист 'Course Statistics' с 7 строками по всем кодам AAA..GGG-2014J", False, "open failed")
        record("Excel: Avg_Score (±2.0) И Total_Assessments == Assignment_Count + Quiz_Count", False, "open failed")
        record("Excel: Course_Name — короткие названия БЕЗ суффикса семестра/2014J", False, "open failed")
        return
    record("Файл открывается openpyxl", True)

    if "Course Statistics" not in wb.sheetnames:
        record("Лист 'Course Statistics' присутствует", False, f"sheets={wb.sheetnames}")
        record("Excel: лист 'Course Statistics' с 7 строками по всем кодам AAA..GGG-2014J", False, "no sheet")
        record("Excel: Avg_Score (±2.0) И Total_Assessments == Assignment_Count + Quiz_Count", False, "no sheet")
        record("Excel: Course_Name — короткие названия БЕЗ суффикса семестра/2014J", False, "no sheet")
        return
    record("Лист 'Course Statistics' присутствует", True)

    ws = wb["Course Statistics"]
    headers = [cell.value for cell in ws[1]]
    expected_headers = [
        "Course_Code", "Course_Name", "Enrollment", "Avg_Score",
        "Assignment_Count", "Quiz_Count", "Total_Assessments",
    ]

    col_map = {}
    missing_cols = []
    for eh in expected_headers:
        found = False
        for idx, h in enumerate(headers):
            if h and eh.lower().replace("_", "") == str(h).lower().replace("_", "").replace(" ", ""):
                col_map[eh] = idx
                found = True
                break
        if not found:
            for idx, h in enumerate(headers):
                if h and eh.lower().replace("_", " ") in str(h).lower().replace("_", " "):
                    col_map[eh] = idx
                    found = True
                    break
        if not found:
            missing_cols.append(eh)
    record("Все ожидаемые колонки присутствуют", not missing_cols,
           f"missing={missing_cols} headers={headers}")
    if missing_cols:
        record("Excel: лист 'Course Statistics' с 7 строками по всем кодам AAA..GGG-2014J", False, "missing cols")
        record("Excel: Avg_Score (±2.0) И Total_Assessments == Assignment_Count + Quiz_Count", False, "missing cols")
        record("Excel: Course_Name — короткие названия БЕЗ суффикса семестра/2014J", False, "missing cols")
        return

    data_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[col_map["Course_Code"]] is not None:
            data_rows.append(row)

    codes_present = {str(r[col_map["Course_Code"]]).strip() for r in data_rows}
    expected_codes = set(EXPECTED_BY_CODE.keys())
    # CRITICAL: ровно 7 строк по всем 7 кодам, без лишних/недостающих
    record("Excel: лист 'Course Statistics' с 7 строками по всем кодам AAA..GGG-2014J",
           len(data_rows) == 7 and codes_present == expected_codes,
           f"rows={len(data_rows)} codes={sorted(codes_present)}")

    # Сортировка по Course_Code (некритично)
    row_codes = [str(r[col_map["Course_Code"]]).strip() for r in data_rows]
    record("Строки отсортированы по алфавиту по Course_Code",
           row_codes == sorted(row_codes), f"order={row_codes}")

    # Значения: Avg_Score и вычисляемая колонка Total_Assessments
    values_ok = True
    name_ok = True
    enroll_ok = True
    for row in data_rows:
        code = str(row[col_map["Course_Code"]]).strip()
        exp = EXPECTED_BY_CODE.get(code)
        if not exp:
            values_ok = False
            name_ok = False
            continue

        enr = row[col_map["Enrollment"]]
        if enr is None or abs(int(enr) - exp["Enrollment"]) > 10:
            enroll_ok = False

        avg = row[col_map["Avg_Score"]]
        if avg is None or abs(float(avg) - exp["Avg_Score"]) > 2.0:
            values_ok = False

        ac = row[col_map["Assignment_Count"]]
        qc = row[col_map["Quiz_Count"]]
        ta = row[col_map["Total_Assessments"]]
        try:
            ac_i, qc_i, ta_i = int(ac), int(qc), int(ta)
        except (TypeError, ValueError):
            values_ok = False
            continue
        if abs(ac_i - exp["Assignment_Count"]) > 1 or abs(qc_i - exp["Quiz_Count"]) > 1:
            values_ok = False
        if ta_i != ac_i + qc_i:
            values_ok = False

        # Course_Name — короткое название без суффикса семестра/2014J.
        name = str(row[col_map["Course_Name"]] or "").strip()
        if "2014" in name or "2014j" in name.lower():
            name_ok = False
        # И совпадает с ожидаемым коротким именем (без учёта регистра/пробелов).
        if name.lower().replace(" ", "") != exp["Course_Name"].lower().replace(" ", ""):
            name_ok = False

    record("Enrollment в пределах допуска (±10)", enroll_ok)

    # CRITICAL: значения и вычисляемая колонка
    record("Excel: Avg_Score (±2.0) И Total_Assessments == Assignment_Count + Quiz_Count",
           values_ok, "см. строки выше")

    # CRITICAL: имена курсов без суффикса
    record("Excel: Course_Name — короткие названия БЕЗ суффикса семестра/2014J",
           name_ok, "см. колонку Course_Name")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False, help="Launch time")
    args = parser.parse_args()

    check_gform()
    check_excel(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "pass": PASS_COUNT,
                    "fail": FAIL_COUNT,
                    "accuracy": accuracy,
                    "critical_failed": CRITICAL_FAILED,
                }, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"[warn] cannot write res_log_file: {e}")

    if CRITICAL_FAILED:
        print(f"CRITICAL FAIL: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL (accuracy < 70%)")
        sys.exit(1)


if __name__ == "__main__":
    main()
