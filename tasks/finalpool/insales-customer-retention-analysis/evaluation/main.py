"""Evaluation for insales-customer-retention-analysis (InSales / wc.* schema)."""
import argparse
import os
import sys
from datetime import date

import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CHURN_THRESHOLD = 90
REFERENCE_DATE = date(2026, 3, 7)

# Collects human-readable reasons for any CRITICAL failure.
CRITICAL_FAILURES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def critical(name, passed, detail=""):
    """A semantic check that, if failed, forces an overall FAIL."""
    record(name, passed, detail)
    if not passed:
        CRITICAL_FAILURES.append(f"{name}{(': ' + detail[:200]) if detail else ''}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def safe_str(v):
    return str(v).strip() if v is not None else ""


def is_yes(v):
    """At_Risk affirmative in RU or EN."""
    return safe_str(v).lower() in ("yes", "да", "true", "1")


def get_expected_data():
    """Compute expected customer retention data live from the wc.* database.

    Data VALUES (names/cities) are russified centrally by the seed; we read them
    honestly here and never hardcode realia literals.
    """
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id, c.email, c.first_name, c.last_name,
               c.total_spent::numeric, c.orders_count,
               MAX(o.date_created) as last_order_date
        FROM wc.customers c
        LEFT JOIN wc.orders o ON c.id = o.customer_id
        WHERE c.total_spent::numeric > 0
        GROUP BY c.id, c.email, c.first_name, c.last_name, c.total_spent, c.orders_count
        ORDER BY c.total_spent::numeric DESC
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    customers = []
    for cid, email, first, last, spent, orders, last_order in rows:
        days_since = (REFERENCE_DATE - last_order.date()).days if last_order else 999
        if spent > 500:
            segment = "VIP"
        elif spent >= 100:
            segment = "Regular"
        else:
            segment = "New"
        at_risk = "Yes" if days_since > CHURN_THRESHOLD else "No"
        customers.append({
            "name": f"{first} {last}",
            "email": email,
            "first": first,
            "spent": float(spent),
            "orders": orders,
            "last_date": last_order.strftime("%Y-%m-%d") if last_order else "",
            "days": days_since,
            "segment": segment,
            "at_risk": at_risk,
        })
    return customers


def check_excel(agent_workspace):
    """Check Customer_Retention.xlsx. Returns (structural_ok, expected)."""
    print("\n=== Checking Customer_Retention.xlsx ===")

    excel_path = os.path.join(agent_workspace, "Customer_Retention.xlsx")
    expected = get_expected_data()

    if not os.path.isfile(excel_path):
        critical("Excel file exists", False, f"Not found: {excel_path}")
        return False, expected
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        critical("Excel readable", False, str(e))
        return False, expected

    all_ok = True

    # --- Customer Analysis sheet ---
    ca_sheet = None
    for name in wb.sheetnames:
        if "customer" in name.lower() and "analy" in name.lower():
            ca_sheet = wb[name]
            break
    if ca_sheet is None:
        record("Sheet 'Customer Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
        headers = []
        rows = []
    else:
        record("Sheet 'Customer Analysis' exists", True)
        headers = [safe_str(ca_sheet.cell(1, c).value).lower() for c in range(1, 10)]
        record("Has Customer_Name column", any("customer" in h or "name" in h for h in headers))
        record("Has Email column", any("email" in h for h in headers))
        record("Has Total_Spent column", any("spent" in h or "total" in h for h in headers))
        record("Has Segment column", any("segment" in h for h in headers))
        record("Has At_Risk column", any("risk" in h for h in headers))

        rows = list(ca_sheet.iter_rows(min_row=2, values_only=True))
        record("Customer Analysis row count matches",
               abs(len(rows) - len(expected)) <= 2,
               f"Expected ~{len(expected)}, got {len(rows)}")

        seg_col = risk_col = None
        for ci, h in enumerate(headers):
            if "segment" in h:
                seg_col = ci
            if "risk" in h:
                risk_col = ci

        def find_row(exp):
            last_part = exp["name"].split()[-1].lower()
            for r in rows:
                if r and r[0] and last_part in safe_str(r[0]).lower() and exp["first"].lower() in safe_str(r[0]).lower():
                    return r
            return None

        # CRITICAL: top-5 by Total_Spent present with matching spend + correct segment.
        for exp in expected[:5]:
            r = find_row(exp)
            if r is None:
                critical(f"Top customer '{exp['name']}' present in Customer Analysis", False)
                all_ok = False
                continue
            ok_spent = num_close(r[2], exp["spent"], 5.0)
            critical(f"{exp['name']} Total_Spent ~{exp['spent']}", ok_spent, f"Got {r[2]}")
            if not ok_spent:
                all_ok = False
            if seg_col is not None:
                ok_seg = safe_str(r[seg_col]).lower() == exp["segment"].lower()
                critical(f"{exp['name']} Segment={exp['segment']}", ok_seg, f"Got {r[seg_col]}")
                if not ok_seg:
                    all_ok = False

        # CRITICAL: every VIP-at-risk customer present with Segment=VIP and At_Risk=Yes/Да.
        vip_at_risk = [c for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes"]
        for exp in vip_at_risk:
            r = find_row(exp)
            if r is None:
                critical(f"VIP-at-risk '{exp['name']}' present in Customer Analysis", False)
                all_ok = False
                continue
            if seg_col is not None:
                ok = safe_str(r[seg_col]).lower() == "vip"
                critical(f"{exp['name']} segment=VIP", ok, f"Got {r[seg_col]}")
                if not ok:
                    all_ok = False
            if risk_col is not None:
                ok = is_yes(r[risk_col])
                critical(f"{exp['name']} at_risk=Yes/Да", ok, f"Got {r[risk_col]}")
                if not ok:
                    all_ok = False

    # --- Retention Metrics sheet ---
    rm_sheet = None
    for name in wb.sheetnames:
        if "retention" in name.lower() and "metric" in name.lower():
            rm_sheet = wb[name]
            break
    total_active = len(expected)
    vip_count = sum(1 for c in expected if c["segment"] == "VIP")
    at_risk_total = sum(1 for c in expected if c["at_risk"] == "Yes")
    at_risk_vip = sum(1 for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes")

    if rm_sheet is None:
        critical("Sheet 'Retention Metrics' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Retention Metrics' exists", True)
        metrics = {}
        for row in rm_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                metrics[safe_str(row[0]).lower().replace(" ", "_")] = row[1]

        got_total = got_vip = got_arisk_total = got_arisk_vip = None
        for key, val in metrics.items():
            if "at_risk_vip" in key or ("risk" in key and "vip" in key):
                got_arisk_vip = val
            elif "at_risk_total" in key or ("risk" in key and "total" in key):
                got_arisk_total = val
            elif "vip_count" in key and "risk" not in key:
                got_vip = val
            elif "total_active" in key or ("total" in key and "customer" in key):
                got_total = val

        # CRITICAL: the four key counts match the live-derived expected values.
        ok = num_close(got_total, total_active, 2)
        critical(f"Total_Active_Customers={total_active}", ok, f"Got {got_total}")
        if not ok:
            all_ok = False
        ok = num_close(got_vip, vip_count, 2)
        critical(f"VIP_Count={vip_count}", ok, f"Got {got_vip}")
        if not ok:
            all_ok = False
        ok = num_close(got_arisk_total, at_risk_total, 2)
        critical(f"At_Risk_Total={at_risk_total}", ok, f"Got {got_arisk_total}")
        if not ok:
            all_ok = False
        ok = num_close(got_arisk_vip, at_risk_vip, 2)
        critical(f"At_Risk_VIP_Count={at_risk_vip}", ok, f"Got {got_arisk_vip}")
        if not ok:
            all_ok = False

    # --- Action Plan sheet ---
    ap_sheet = None
    for name in wb.sheetnames:
        if "action" in name.lower() and "plan" in name.lower():
            ap_sheet = wb[name]
            break
    if ap_sheet is None:
        record("Sheet 'Action Plan' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Action Plan' exists", True)
        ap_rows = list(ap_sheet.iter_rows(min_row=2, values_only=True))
        vip_at_risk = [c for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes"]
        record("Action Plan row count matches",
               abs(len(ap_rows) - len(vip_at_risk)) <= 1,
               f"Expected {len(vip_at_risk)}, got {len(ap_rows)}")

    return all_ok, expected


def check_emails(expected):
    """CRITICAL: a retention email was sent to each VIP-at-risk customer.

    Match by recipient address (RU/EN subject agnostic), NOT by English subject.
    """
    print("\n=== Checking Emails Sent ===")
    vip_at_risk = [c for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes"]

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    if not vip_at_risk:
        # Avoid a vacuous critical check when the at-risk-VIP set is empty.
        record("No VIP-at-risk customers -> email check skipped", True)
        cur.close()
        conn.close()
        return True

    all_ok = True
    matched = 0
    for c in vip_at_risk:
        cur.execute("SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s",
                    (f"%{c['email']}%",))
        cnt = cur.fetchone()[0]
        ok = cnt >= 1
        if ok:
            matched += 1
        else:
            all_ok = False
    critical("Retention email sent to every VIP-at-risk customer",
             all_ok, f"Matched {matched}/{len(vip_at_risk)} recipient addresses")

    cur.close()
    conn.close()
    return all_ok


def check_gsheet(expected):
    """CRITICAL: 'Retention Outreach Log' sheet with one row per VIP-at-risk customer."""
    print("\n=== Checking Google Sheet ===")
    vip_at_risk = [c for c in expected if c["segment"] == "VIP" and c["at_risk"] == "Yes"]
    needed_rows = len(vip_at_risk)

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Title may be RU or EN: match retention/outreach (EN) or удержани/рассылк/аутрич (RU).
    cur.execute("""
        SELECT id, title FROM gsheet.spreadsheets
        WHERE title ILIKE '%retention%' OR title ILIKE '%outreach%'
           OR title ILIKE '%удержани%' OR title ILIKE '%рассылк%' OR title ILIKE '%аутрич%'
    """)
    rows = cur.fetchall()
    if not rows:
        critical("Outreach-log Google Sheet exists", False, "No matching spreadsheet found")
        cur.close()
        conn.close()
        return False
    record("Outreach-log Google Sheet exists", True)

    ss_id = rows[0][0]
    cur.execute("SELECT id FROM gsheet.sheets WHERE spreadsheet_id = %s", (ss_id,))
    sheets = cur.fetchall()
    if not sheets:
        critical("Outreach-log sheet has a tab", False, "No sheets")
        cur.close()
        conn.close()
        return False
    record("Outreach-log sheet has a tab", True)

    sheet_id = sheets[0][0]
    cur.execute(
        "SELECT COUNT(DISTINCT row_index) FROM gsheet.cells "
        "WHERE spreadsheet_id = %s AND sheet_id = %s AND row_index > 0",
        (ss_id, sheet_id))
    row_count = cur.fetchone()[0]
    # >= one data row per VIP-at-risk customer (>=1 floor so it isn't vacuous).
    ok = row_count >= max(needed_rows, 1)
    critical("Outreach log has one data row per VIP-at-risk customer",
             ok, f"Expected >= {max(needed_rows, 1)} data rows, got {row_count}")

    cur.close()
    conn.close()
    return ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok, expected = check_excel(args.agent_workspace)
    email_ok = check_emails(expected)
    gsheet_ok = check_gsheet(expected)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    # CRITICAL gate: any critical failure => FAIL regardless of accuracy.
    if CRITICAL_FAILURES:
        print("  CRITICAL FAILURES:")
        for cf in CRITICAL_FAILURES:
            print(f"    - {cf}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
