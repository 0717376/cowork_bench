#!/usr/bin/env python3
"""Evaluation for arxiv-research-workflow-pipeline (RU, teamly swap).

Validates four artifacts the agent produces in the workspace plus the Teamly
research hub:
  - paper_analysis.xlsx (Metadata / Statistics / Recommendations).  Values are
    checked for INTERNAL CONSISTENCY against the agent's own data (not against a
    fabricated answer key): Statistics 'Total Papers' must equal the Metadata
    data-row count, citation/year columns must be numeric, etc.
  - literature_review.docx (RU prose, >= ~1200 words / RU headings).
  - distributed_ml_papers.bib (>= 10 valid bib entries).
  - Teamly hub page + child pages (replacing the former notion deliverable).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS threshold: accuracy >= 70%.

RU keyword checks search the ORIGINAL lowercased text (never normalized/translit).
"""
from argparse import ArgumentParser
import json
import os
import re
import sys

try:
    import psycopg2
except Exception:  # pragma: no cover
    psycopg2 = None

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Any failure here => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "xlsx paper_analysis.xlsx exists with Metadata schema + >=10 rows",
    "xlsx internal consistency: Statistics 'Total Papers' == Metadata data rows",
    "docx literature_review.docx is non-trivial Russian prose",
    "distributed_ml_papers.bib has >=10 valid entries",
    "Teamly research hub page exists",
}

METADATA_COLS = ["Paper ID", "Title", "Authors", "Year", "Category"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _is_num(v):
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


def _norm(s):
    return str(s).strip().lower().replace("_", " ") if s is not None else ""


def find_col(header, names):
    for i, cell in enumerate(header):
        c = _norm(cell)
        for n in names:
            if _norm(n) == c:
                return i
    return None


def check_xlsx(workspace):
    print("\n=== Check: XLSX paper_analysis.xlsx ===")
    import openpyxl
    xlsx_path = os.path.join(workspace, "paper_analysis.xlsx")
    if not os.path.isfile(xlsx_path):
        record("xlsx paper_analysis.xlsx exists with Metadata schema + >=10 rows", False, "Not found")
        return
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("xlsx paper_analysis.xlsx exists with Metadata schema + >=10 rows", False, f"unreadable: {e}")
        return

    sheets = {n.strip().lower(): n for n in wb.sheetnames}

    # --- Metadata sheet ---
    meta_name = sheets.get("metadata")
    meta_rows = []
    meta_header = []
    if meta_name:
        ws = wb[meta_name]
        rows = list(ws.iter_rows(values_only=True))
        meta_header = list(rows[0]) if rows else []
        meta_rows = [r for r in rows[1:] if any(c is not None for c in r)]

    cols_present = meta_name is not None and all(
        find_col(meta_header, [c]) is not None for c in METADATA_COLS
    )
    record(
        "xlsx paper_analysis.xlsx exists with Metadata schema + >=10 rows",
        cols_present and len(meta_rows) >= 10,
        f"sheet={meta_name}, header={meta_header}, data_rows={len(meta_rows)}",
    )

    # --- Statistics sheet + internal consistency ---
    stats_name = sheets.get("statistics")
    stats_metrics = {}
    if stats_name:
        record("xlsx Statistics sheet exists", True)
        for r in wb[stats_name].iter_rows(values_only=True):
            if r and r[0] is not None:
                stats_metrics[_norm(r[0])] = r[1] if len(r) > 1 else None
    else:
        record("xlsx Statistics sheet exists", False, f"sheets={wb.sheetnames}")

    # Critical: Total Papers == Metadata data rows
    total_papers = None
    for k, v in stats_metrics.items():
        if "total" in k and "paper" in k:
            total_papers = v
            break
    consistent = (
        total_papers is not None
        and _is_num(total_papers)
        and len(meta_rows) >= 10
        and int(float(total_papers)) == len(meta_rows)
    )
    record(
        "xlsx internal consistency: Statistics 'Total Papers' == Metadata data rows",
        consistent,
        f"Total Papers={total_papers}, Metadata rows={len(meta_rows)}",
    )

    # Year/Citation columns numeric in Metadata (non-critical structural).
    if meta_name:
        year_i = find_col(meta_header, ["Year"])
        cit_i = find_col(meta_header, ["Citation Count"])
        if year_i is not None:
            ok = sum(1 for r in meta_rows if year_i < len(r) and _is_num(r[year_i]))
            record("xlsx Metadata Year column numeric", ok >= len(meta_rows) * 0.8,
                   f"{ok}/{len(meta_rows)} numeric")
        if cit_i is not None:
            ok = sum(1 for r in meta_rows if cit_i < len(r) and _is_num(r[cit_i]))
            record("xlsx Metadata Citation Count numeric", ok >= len(meta_rows) * 0.8,
                   f"{ok}/{len(meta_rows)} numeric")

    # Average Citations consistent with data (within rounding).
    avg_val = None
    for k, v in stats_metrics.items():
        if "average" in k and "citation" in k:
            avg_val = v
            break
    if avg_val is not None and meta_name:
        cit_i = find_col(meta_header, ["Citation Count"])
        if cit_i is not None:
            nums = [float(r[cit_i]) for r in meta_rows if cit_i < len(r) and _is_num(r[cit_i])]
            if nums:
                computed = sum(nums) / len(nums)
                ok = _is_num(avg_val) and abs(float(avg_val) - computed) <= max(computed * 0.1, 5)
                record("xlsx Average Citations consistent with Metadata", ok,
                       f"reported={avg_val}, computed={computed:.0f}")

    # --- Recommendations sheet ---
    rec_name = sheets.get("recommendations")
    if rec_name:
        rows = list(wb[rec_name].iter_rows(values_only=True))
        header = list(rows[0]) if rows else []
        data = [r for r in rows[1:] if any(c is not None for c in r)]
        has_cols = find_col(header, ["Paper ID"]) is not None and find_col(header, ["Priority"]) is not None
        record("xlsx Recommendations sheet has rows + schema", has_cols and len(data) >= 1,
               f"header={header}, rows={len(data)}")
    else:
        record("xlsx Recommendations sheet exists", False, f"sheets={wb.sheetnames}")

    wb.close()


def check_docx(workspace):
    print("\n=== Check: DOCX literature_review.docx ===")
    from docx import Document
    path = os.path.join(workspace, "literature_review.docx")
    if not os.path.isfile(path):
        record("docx literature_review.docx is non-trivial Russian prose", False, "Not found")
        return
    try:
        doc = Document(path)
    except Exception as e:
        record("docx literature_review.docx is non-trivial Russian prose", False, f"unreadable: {e}")
        return

    paras = [p.text for p in doc.paragraphs]
    full = "\n".join(paras)
    low = full.lower()  # ORIGINAL lowercased text — NOT normalized.
    words = re.findall(r"\S+", full)
    has_cyrillic = bool(re.search(r"[а-яё]", low))
    ru_terms = ["обзор", "литератур", "исследован", "рекоменд", "тенденц", "пробел", "введен"]
    ru_hits = sum(1 for t in ru_terms if t in low)

    non_trivial = (len(words) >= 1200 or len([p for p in paras if p.strip()]) >= 20) \
        and has_cyrillic and ru_hits >= 2
    record("docx literature_review.docx is non-trivial Russian prose", non_trivial,
           f"words={len(words)}, paras={len([p for p in paras if p.strip()])}, "
           f"cyrillic={has_cyrillic}, ru_terms={ru_hits}")

    # Section headings as list-of-alternatives (RU + EN), non-critical.
    required_headings = [
        ["введен", "introduction"],
        ["рекоменд", "recommendation"],
        ["тенденц", "trend", "пробел", "gap"],
    ]
    for alts in required_headings:
        found = any(a in low for a in alts)
        record(f"docx mentions section ({'/'.join(alts)})", found)


def check_bib(workspace):
    print("\n=== Check: distributed_ml_papers.bib ===")
    # Accept the named file, or any *.bib the agent produced (excluding the seed).
    path = os.path.join(workspace, "distributed_ml_papers.bib")
    if not os.path.isfile(path):
        cands = [f for f in os.listdir(workspace)
                 if f.endswith(".bib") and f != "bibliography_seed.bib"]
        path = os.path.join(workspace, cands[0]) if cands else None
    if not path or not os.path.isfile(path):
        record("distributed_ml_papers.bib has >=10 valid entries", False, "no .bib found")
        return
    try:
        text = open(path, encoding="utf-8", errors="ignore").read()
    except Exception as e:
        record("distributed_ml_papers.bib has >=10 valid entries", False, str(e))
        return
    entries = re.findall(r"@(article|inproceedings|misc|book)\s*\{", text, re.IGNORECASE)
    has_fields = text.lower().count("title") >= 10 and text.lower().count("year") >= 10
    record("distributed_ml_papers.bib has >=10 valid entries",
           len(entries) >= 10 and has_fields,
           f"entries={len(entries)}, titles={text.lower().count('title')}")


def check_teamly():
    print("\n=== Check: Teamly research hub ===")
    if psycopg2 is None:
        record("Teamly research hub page exists", False, "psycopg2 unavailable")
        return
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title, body, parent_id FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Teamly research hub page exists", False, f"DB error: {e}")
        return

    hub_id = None
    for pid, title, body, parent_id in pages:
        tl = _norm(title)
        # RU keywords in original lowercased text (not normalized).
        if any(k in tl for k in ["база знани", "распределённ", "распределенн",
                                 "машинн", "исследован", "обзор", "knowledge", "hub", "research"]):
            hub_id = pid
            break
    if hub_id is None and pages:
        hub_id = pages[0][0]
    record("Teamly research hub page exists", hub_id is not None,
           f"new pages found: {len(pages)}")

    children = [p for p in pages if p[3] == hub_id and p[0] != hub_id]
    entries = children if len(children) >= 5 else [p for p in pages if p[0] != hub_id]
    record("Teamly: >= 5 child/library pages", len(entries) >= 5,
           f"children={len(children)}, other={len(entries)}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    ws = args.agent_workspace
    if not os.path.isdir(ws):
        print(f"Agent workspace not found: {ws}")
        sys.exit(1)

    check_xlsx(ws)
    check_docx(ws)
    check_bib(ws)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
