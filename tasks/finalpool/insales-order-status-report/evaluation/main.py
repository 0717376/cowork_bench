"""Evaluation for insales-order-status-report (InSales)."""
import argparse
import os
import sys
import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# SEMANTIC checks whose failure forces FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "completed row Order_Count & Total_Revenue",
    "Summary Total_Orders & Total_Revenue",
    "Summary Completed_Revenue & Fulfillment_Rate_Pct",
    "All 7 status rows present with exact Order_Count",
    "Email delivered to operations@shop.com with required subject",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def lookup_by_key(rows):
    data = rows[1:] if rows and len(rows) > 1 else []
    out = {}
    for row in data:
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row
    return out


def check_order_status(agent_wb, gt_wb):
    print("\n=== Checking 'Order Status' sheet ===")
    a_rows = load_sheet_rows(agent_wb, "Order Status")
    g_rows = load_sheet_rows(gt_wb, "Order Status")
    if a_rows is None:
        check("'Order Status' sheet exists", False, "Sheet not found in agent output")
        return
    if g_rows is None:
        check("'Order Status' groundtruth sheet exists", False, "missing in groundtruth")
        return
    check("'Order Status' sheet exists", True)

    a_lookup = lookup_by_key(a_rows)
    g_lookup = lookup_by_key(g_rows)

    STATUSES = ["completed", "processing", "on-hold", "cancelled",
                "pending", "refunded", "failed"]

    # CRITICAL: completeness of breakdown -- all 7 status rows present, count exact.
    counts_ok = True
    detail = []
    for st in STATUSES:
        a_row = a_lookup.get(st)
        g_row = g_lookup.get(st)
        if a_row is None or g_row is None:
            counts_ok = False
            detail.append(f"{st}: missing (a={a_row is not None}, g={g_row is not None})")
            continue
        a_cnt = a_row[1] if len(a_row) > 1 else None
        g_cnt = g_row[1] if len(g_row) > 1 else None
        if not num_close(a_cnt, g_cnt, 0):
            counts_ok = False
            detail.append(f"{st}.Order_Count: {a_cnt} vs {g_cnt}")
    check("All 7 status rows present with exact Order_Count", counts_ok, "; ".join(detail))

    # CRITICAL: headline fulfillment metric -- completed count exact, revenue tol<=1.0
    a_c = a_lookup.get("completed")
    g_c = g_lookup.get("completed")
    if a_c is not None and g_c is not None:
        cnt_ok = num_close(a_c[1], g_c[1], 0) if len(a_c) > 1 and len(g_c) > 1 else False
        rev_ok = num_close(a_c[2], g_c[2], 1.0) if len(a_c) > 2 and len(g_c) > 2 else False
        check("completed row Order_Count & Total_Revenue", cnt_ok and rev_ok,
              f"count={a_c[1] if len(a_c)>1 else None}/{g_c[1] if len(g_c)>1 else None}, "
              f"rev={a_c[2] if len(a_c)>2 else None}/{g_c[2] if len(g_c)>2 else None}")
    else:
        check("completed row Order_Count & Total_Revenue", False, "completed row missing")

    # NON-critical granular per-status revenue / avg checks (tightened revenue tol).
    for st in STATUSES:
        a_row = a_lookup.get(st)
        g_row = g_lookup.get(st)
        if a_row is None or g_row is None:
            continue
        if len(a_row) > 2 and len(g_row) > 2:
            check(f"{st}.Total_Revenue", num_close(a_row[2], g_row[2], 1.0),
                  f"{a_row[2]} vs {g_row[2]} (tol=1.0)")
        if len(a_row) > 3 and len(g_row) > 3:
            check(f"{st}.Avg_Order_Value", num_close(a_row[3], g_row[3], 1.0),
                  f"{a_row[3]} vs {g_row[3]} (tol=1.0)")
        if len(a_row) > 4 and len(g_row) > 4:
            check(f"{st}.Revenue_Share_Pct", num_close(a_row[4], g_row[4], 1.0),
                  f"{a_row[4]} vs {g_row[4]} (tol=1.0)")

    # NON-critical: sort by Total_Revenue descending.
    a_data = a_rows[1:] if len(a_rows) > 1 else []
    revs = []
    for row in a_data:
        if row and row[0] is not None and len(row) > 2 and row[2] is not None:
            try:
                revs.append(float(row[2]))
            except (TypeError, ValueError):
                pass
    descending = all(revs[i] >= revs[i + 1] for i in range(len(revs) - 1))
    check("Order Status sorted by Total_Revenue descending", descending, f"revenues={revs}")


def check_summary(agent_wb, gt_wb):
    print("\n=== Checking 'Summary' sheet ===")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        check("'Summary' sheet exists", False, "Sheet not found in agent output")
        return
    if g_rows is None:
        check("'Summary' groundtruth sheet exists", False, "missing in groundtruth")
        return
    check("'Summary' sheet exists", True)

    a_lookup = lookup_by_key(a_rows)
    g_lookup = lookup_by_key(g_rows)

    def val(lookup, key):
        row = lookup.get(key.lower())
        return row[1] if row and len(row) > 1 else None

    # CRITICAL: core aggregate substance.
    a_to = val(a_lookup, "total_orders")
    g_to = val(g_lookup, "total_orders")
    a_tr = val(a_lookup, "total_revenue")
    g_tr = val(g_lookup, "total_revenue")
    check("Summary Total_Orders & Total_Revenue",
          num_close(a_to, g_to, 0) and num_close(a_tr, g_tr, 2.0),
          f"orders={a_to}/{g_to}, revenue={a_tr}/{g_tr}")

    # CRITICAL: headline KPI -- Completed_Revenue (tol<=2.0) + Fulfillment_Rate_Pct (tol<=1.0).
    a_cr = val(a_lookup, "completed_revenue")
    g_cr = val(g_lookup, "completed_revenue")
    a_fr = val(a_lookup, "fulfillment_rate_pct")
    g_fr = val(g_lookup, "fulfillment_rate_pct")
    check("Summary Completed_Revenue & Fulfillment_Rate_Pct",
          num_close(a_cr, g_cr, 2.0) and num_close(a_fr, g_fr, 1.0),
          f"completed_rev={a_cr}/{g_cr}, fulfillment={a_fr}/{g_fr}")


def check_email():
    print("\n=== Checking Email deliverable ===")
    import psycopg2

    target = "operations@shop.com"
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
            user="eigent", password="camel"
        )
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Email delivered to operations@shop.com with required subject", False,
              f"DB error: {e}")
        return

    found = None
    for subj, to_addr, body in rows:
        if target in str(to_addr or "").lower():
            found = (subj, to_addr, body)
            break

    if found is None:
        check("Email delivered to operations@shop.com with required subject", False,
              f"No email to {target} ({len(rows)} total emails)")
        return

    subj, _, body = found
    subj_lc = (subj or "").lower()
    # Subject is an English-preserved literal that the task requires verbatim.
    subj_ok = "order fulfillment status report" in subj_lc
    check("Email delivered to operations@shop.com with required subject", subj_ok,
          f"Subject: {(subj or '')[:120]}")

    # NON-critical: body mentions key metrics. RU/EN keywords on ORIGINAL lowered text.
    body_lc = (body or "").lower()
    BODY_TOKENS = ["статус", "заказ", "выручк", "отчёт", "отчет",
                   "status", "order", "revenue", "report", "fulfillment"]
    check("Email body references status/orders/revenue (RU+EN)",
          any(t in body_lc for t in BODY_TOKENS),
          f"Body[:120]: {(body or '')[:120]}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Order_Status_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Order_Status_Report.xlsx")

    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    if not os.path.exists(agent_file):
        check("WC_Order_Status_Report.xlsx exists", False, f"Not found: {agent_file}")
    else:
        check("WC_Order_Status_Report.xlsx exists", True)
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
        check_order_status(agent_wb, gt_wb)
        check_summary(agent_wb, gt_wb)

    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print("\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILS:
        print(f"  CRITICAL failures: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    all_ok = accuracy >= 70
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")
    if all_ok:
        print("\nPass all tests!")
        sys.exit(0)
    else:
        print("\nAccuracy below threshold.")
        sys.exit(1)


if __name__ == "__main__":
    main()
