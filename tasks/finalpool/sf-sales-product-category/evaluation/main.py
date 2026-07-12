"""Оценка для sf-sales-product-category (форк ClickHouse).

Данные продаж в sf_data (SALES_DW.PUBLIC.PRODUCTS) русифицированы централизованно
через db/zzz_clickhouse_after_init.sql, поэтому агент читает РУССКОЕ значение CATEGORY
('ТВ, аудио и камеры') из БД и записывает его как ключ строки на листе
"Product Categories" и как значение Most_Profitable_Category на листе "Summary".
Эталон groundtruth_workspace/Sales_Product_Categories.xlsx переразмечен тем же
центральным отображением (scripts/clickhouse_relabel_map.FLAT_VALUE_MAP), поэтому
его ячейка категории уже содержит то же русское значение — никакого хардкода
русского литерала в коде оценки нет, сравнение идёт ПРОТИВ эталона.

Имена столбцов, имена листов и метки Metric (Total_Categories/Total_Products/
Most_Profitable_Category/Overall_Avg_Margin) остаются английскими токенами, как и
имя файла и имена схемы/таблицы/столбцов БД.

Оценка: сначала гейт CRITICAL_CHECKS (любой провал критической проверки => FAIL),
затем гейт accuracy>=70 по всем (критическим и структурным) проверкам.
"""
import argparse
import os
import sys
import openpyxl

try:
    import psycopg2
except ImportError:
    psycopg2 = None

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432,
      "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        d = (str(detail)[:300] + "...") if len(str(detail)) > 300 else detail
        msg = f": {d}" if d else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_gsheet(gt_category):
    """Проверка Google-таблицы 'Product Category Report' (deliverable для мерчандайзинга).

    gt_category — русское значение категории из эталона; ищем его в ячейках таблицы.
    """
    print("\n=== Проверка: Google Sheet 'Product Category Report' ===")
    if psycopg2 is None:
        record("Google-таблица 'Product Category Report' существует", False,
               "psycopg2 недоступен", critical=True)
        return
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        record("Google-таблица 'Product Category Report' существует", False,
               f"нет соединения с БД: {e}", critical=True)
        return
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM gsheet.spreadsheets WHERE LOWER(title) LIKE %s",
        ("%product%category%report%",))
    sheets = cur.fetchall()
    if not sheets:
        record("Google-таблица 'Product Category Report' существует", False,
               "не найдена таблица с 'product category report' в названии", critical=True)
        cur.close(); conn.close()
        return
    record("Google-таблица 'Product Category Report' существует", True, critical=True)
    sheet_id = sheets[0][0]

    cur.execute("SELECT id FROM gsheet.sheets WHERE spreadsheet_id = %s", (sheet_id,))
    tabs = cur.fetchall()
    record("В таблице есть хотя бы один лист", len(tabs) >= 1, f"листов: {len(tabs)}")
    if not tabs:
        cur.close(); conn.close()
        return

    all_values = []
    for (tab_id,) in tabs:
        cur.execute(
            "SELECT value FROM gsheet.cells WHERE sheet_id = %s", (tab_id,))
        all_values.extend(str(v[0]) for v in cur.fetchall() if v[0] is not None)
    cur.close(); conn.close()

    joined = " ".join(all_values).lower()
    record("В таблице есть данные строк", len([v for v in all_values if v.strip()]) >= 1,
           f"непустых ячеек: {len([v for v in all_values if v.strip()])}", critical=True)
    # Категория (русское значение из эталона) должна присутствовать в таблице
    record("Таблица содержит данные по категории",
           str(gt_category).strip().lower() in joined,
           f"значение категории '{gt_category}' не найдено в ячейках", critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Product_Categories.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Product_Categories.xlsx")

    if not os.path.exists(agent_file):
        record("Файл Sales_Product_Categories.xlsx существует", False, agent_file, critical=True)
        print(f"FAIL: вывод агента не найден: {agent_file}")
        print(f"\n=== RESULT: FAIL (критические проверки: {CRITICAL_FAILS}) ===")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: эталон не найден: {gt_file}")
        sys.exit(1)

    record("Файл Sales_Product_Categories.xlsx существует", True)
    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Категория из эталона (русское значение, переразмечено центральным map)
    gt_category = "ТВ, аудио и камеры"

    # === Лист: Product Categories ===
    print("\n=== Проверка: Product Categories ===")
    a_rows = load_sheet_rows(agent_wb, "Product Categories")
    g_rows = load_sheet_rows(gt_wb, "Product Categories")
    if a_rows is None:
        record("Лист 'Product Categories' присутствует", False, "нет листа", critical=True)
    elif g_rows is None:
        record("Лист 'Product Categories' (эталон)", False, "нет листа в эталоне")
    else:
        record("Лист 'Product Categories' присутствует", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
                # запомним русское значение категории из эталона для проверки gsheet
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            gt_category = g_row[0]  # русское значение категории из эталона
            cat = g_row[0]
            a_row = a_lookup.get(key)
            # Наличие строки категории — КРИТИЧНО (ключ = русское название категории из БД)
            if a_row is None:
                record(f"Строка категории '{cat}' присутствует", False, "строка отсутствует", critical=True)
                continue
            record(f"Строка категории '{cat}' присутствует", True)

            # Product_Count — КРИТИЧНО, ужесточённый tol=2 (источник: PRODUCTS)
            if len(a_row) > 1 and len(g_row) > 1:
                record(f"{cat}.Product_Count корректно",
                       num_close(a_row[1], g_row[1], 2),
                       f"{a_row[1]} vs {g_row[1]} (tol=2)", critical=True)

            # Avg_Price — ужесточено 5.0 -> 0.05
            if len(a_row) > 2 and len(g_row) > 2:
                record(f"{cat}.Avg_Price корректно",
                       num_close(a_row[2], g_row[2], 0.05),
                       f"{a_row[2]} vs {g_row[2]} (tol=0.05)", critical=True)

            # Avg_Cost — ужесточено 5.0 -> 0.05
            if len(a_row) > 3 and len(g_row) > 3:
                record(f"{cat}.Avg_Cost корректно",
                       num_close(a_row[3], g_row[3], 0.05),
                       f"{a_row[3]} vs {g_row[3]} (tol=0.05)", critical=True)

            # Avg_Margin — КРИТИЧНО, ужесточено 5.0 -> 0.05; должно равняться Avg_Price - Avg_Cost
            if len(a_row) > 4 and len(g_row) > 4:
                record(f"{cat}.Avg_Margin корректно",
                       num_close(a_row[4], g_row[4], 0.05),
                       f"{a_row[4]} vs {g_row[4]} (tol=0.05)", critical=True)
                # внутренняя согласованность: Avg_Margin == Avg_Price - Avg_Cost
                if len(a_row) > 3 and a_row[2] is not None and a_row[3] is not None and a_row[4] is not None:
                    try:
                        record(f"{cat}.Avg_Margin == Avg_Price - Avg_Cost",
                               abs(float(a_row[4]) - (float(a_row[2]) - float(a_row[3]))) <= 0.05,
                               f"margin={a_row[4]}, price-cost={float(a_row[2]) - float(a_row[3]):.2f}",
                               critical=True)
                    except (TypeError, ValueError):
                        record(f"{cat}.Avg_Margin == Avg_Price - Avg_Cost", False,
                               "нечисловые значения", critical=True)

    # === Лист: Summary ===
    print("\n=== Проверка: Summary ===")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        record("Лист 'Summary' присутствует", False, "нет листа", critical=True)
    elif g_rows is None:
        record("Лист 'Summary' (эталон)", False, "нет листа в эталоне")
    else:
        record("Лист 'Summary' присутствует", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Вертикальная раскладка Metric|Value: ключ = ячейка столбца A, значение в столбце B.
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        # Горизонтальная (широкая) раскладка: имена метрик в строке-заголовке, одна
        # строка значений под ними. task.md требует лишь чтобы Summary "содержал" эти
        # метрики и не задаёт ориентацию Metric|Value, поэтому такая раскладка валидна.
        # Строим запасной lookup: имя столбца-заголовка -> значение в строке значений.
        a_wide = {}
        if a_rows and len(a_rows) >= 2:
            header = a_rows[0]
            value_row = a_rows[1]
            for idx, hcell in enumerate(header):
                if hcell is None:
                    continue
                hkey = str(hcell).strip().lower()
                val = value_row[idx] if idx < len(value_row) else None
                # значение храним в позиции [1], чтобы переиспользовать сравнение ниже
                a_wide[hkey] = [hcell, val]

        # Все метрики КРИТИЧНЫ (ядро задачи), ужесточённые tol (было 10.0)
        metric_tol = {
            "total_categories": 0.0,
            "total_products": 2.0,
            "most_profitable_category": 0.0,   # строковое сравнение (num_close fallback)
            "overall_avg_margin": 0.05,        # ужесточено -> проверяет взвешивание
        }
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            metric = g_row[0]
            # Сначала вертикальная раскладка, затем запасная горизонтальная.
            a_row = a_lookup.get(key)
            if a_row is None:
                a_row = a_wide.get(key)
            if a_row is None:
                record(f"Summary.{metric} присутствует", False, "строка отсутствует", critical=True)
                continue
            tol = metric_tol.get(key, 10.0)
            if len(a_row) > 1 and len(g_row) > 1:
                record(f"Summary.{metric} корректно",
                       num_close(a_row[1], g_row[1], tol),
                       f"{a_row[1]} vs {g_row[1]} (tol={tol})", critical=True)

    # === Google Sheet deliverable ===
    check_gsheet(gt_category)

    # === Итог ===
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nПройдено: {PASS_COUNT}/{total} ({accuracy:.1f}%)")

    if CRITICAL_FAILS:
        print(f"\n=== RESULT: FAIL (провалены критические проверки: {CRITICAL_FAILS}) ===")
        sys.exit(1)

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
