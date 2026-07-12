"""Evaluation для pw-moex-sector-analysis-excel-word (RU / moex-finance).

Агент строит отчёт по отраслевому анализу акций портфеля MOEX, получая данные
через MCP `moex-finance` (схема moex.*) и эталонные отраслевые мультипликаторы
с локальной mock-страницы http://localhost:30315.

Проверяется Excel Sector_Analysis_Report.xlsx (листы Data_Analysis / Metrics /
Recommendations), Word Sector_Analysis_Analysis.docx и наличие скрипта
yf_sector_processor.py.

CRITICAL_CHECKS — семантические проверки СУЩЕСТВА результата. Любой их провал
=> общий FAIL независимо от accuracy. Они НЕ хардкодят волатильные цены
(их агент честно читает из MCP), а проверяют внутреннюю согласованность:
  * тикеры — из отслеживаемого набора MOEX (.ME), а не выдуманные US-тикеры;
  * Upside согласован с Current_Price и Target_Price построчно;
  * строки отсортированы по алфавиту по Symbol;
  * Avg_Upside == среднее Upside, Best_Opportunity == Symbol с макс. Upside;
  * Word содержит три раздела (RU+EN), Recommendations ссылается на тикеры
    из Data_Analysis.

Структурные проверки (наличие листов/колонок/файлов, пороги строк) — НЕ
критические.
"""
import argparse
import glob as globmod
import os
import sys

import openpyxl

# Отслеживаемые тикеры MOEX — источник данных moex-finance
MOEX_TICKERS = {"SBER.ME", "GAZP.ME", "LKOH.ME", "TCSG.ME", "MGNT.ME", "MTSS.ME"}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = False

# Имена критических (семантических) проверок
CRITICAL_CHECKS = {
    "Data_Analysis: Symbol — тикеры из набора MOEX (.ME)",
    "Data_Analysis: Upside согласован с Current_Price и Target_Price",
    "Data_Analysis: строки отсортированы по Symbol",
    "Metrics: Avg_Upside и Best_Opportunity выведены из данных",
    "Word + Recommendations: три раздела и ссылки на тикеры Data_Analysis",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        tag = " (CRITICAL)" if name in CRITICAL_CHECKS else ""
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED = True


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def norm_header(ws):
    """Список заголовков (lower, strip)."""
    return [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]


def col_index(headers, name):
    name = name.strip().lower()
    return headers.index(name) if name in headers else None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = False

    excel_path = os.path.join(agent_workspace, "Sector_Analysis_Report.xlsx")
    check("Sector_Analysis_Report.xlsx exists", os.path.exists(excel_path))

    # значения, нужные критическим проверкам между листами
    da_symbols = []          # Symbol по строкам Data_Analysis
    da_upsides = {}          # Symbol -> Upside

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # ---------------- Data_Analysis ----------------
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = norm_header(ws)
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [r for r in data_rows if r and any(c is not None for c in r)]

            # структурные
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
            for expected_col in ['Symbol', 'Name', 'Sector', 'Current_Price', 'Target_Price', 'Upside']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            i_sym = col_index(headers, 'Symbol')
            i_cur = col_index(headers, 'Current_Price')
            i_tgt = col_index(headers, 'Target_Price')
            i_ups = col_index(headers, 'Upside')

            if i_sym is not None:
                da_symbols = [str(r[i_sym]).strip() for r in data_rows if r[i_sym] is not None]

            # CRITICAL: тикеры из набора MOEX
            symbols_ok = (len(da_symbols) >= 5
                          and all(s in MOEX_TICKERS for s in da_symbols))
            check("Data_Analysis: Symbol — тикеры из набора MOEX (.ME)",
                  symbols_ok, f"symbols={da_symbols}")

            # CRITICAL: Upside согласован с Current/Target построчно
            if None not in (i_sym, i_cur, i_tgt, i_ups) and data_rows:
                consistent = True
                bad = None
                for r in data_rows:
                    cur = safe_float(r[i_cur])
                    tgt = safe_float(r[i_tgt])
                    ups = safe_float(r[i_ups])
                    if cur is None or tgt is None or ups is None or cur == 0:
                        consistent = False
                        bad = (r[i_sym], cur, tgt, ups)
                        break
                    expected = (tgt - cur) / cur * 100.0
                    if abs(expected - ups) > 1.0:  # допуск 1 п.п.
                        consistent = False
                        bad = (r[i_sym], cur, tgt, ups, round(expected, 3))
                        break
                    da_upsides[str(r[i_sym]).strip()] = ups
                check("Data_Analysis: Upside согласован с Current_Price и Target_Price",
                      consistent, f"строка не сходится: {bad}")
            else:
                check("Data_Analysis: Upside согласован с Current_Price и Target_Price",
                      False, "не найдены нужные колонки")

            # CRITICAL: сортировка по Symbol (по алфавиту)
            sorted_ok = da_symbols == sorted(da_symbols)
            check("Data_Analysis: строки отсортированы по Symbol",
                  sorted_ok and len(da_symbols) >= 5,
                  f"order={da_symbols}")
        else:
            check("Data_Analysis: Symbol — тикеры из набора MOEX (.ME)", False, "нет листа")
            check("Data_Analysis: Upside согласован с Current_Price и Target_Price", False, "нет листа")
            check("Data_Analysis: строки отсортированы по Symbol", False, "нет листа")

        # ---------------- Metrics ----------------
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        avg_ok = best_ok = False
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = norm_header(ws)
            mrows = list(ws.iter_rows(min_row=2, values_only=True))
            mrows = [r for r in mrows if r and any(c is not None for c in r)]
            check("Metrics has >= 3 rows", len(mrows) >= 3, f"got {len(mrows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            i_m = col_index(headers, 'Metric')
            i_v = col_index(headers, 'Value')
            metrics = {}
            if None not in (i_m, i_v):
                for r in mrows:
                    if r[i_m] is None:
                        continue
                    metrics[str(r[i_m]).strip().lower()] = r[i_v]

            # CRITICAL: Avg_Upside == среднее Upside, Best_Opportunity == max Upside
            if da_upsides:
                ref_avg = sum(da_upsides.values()) / len(da_upsides)
                ref_best = max(da_upsides, key=da_upsides.get)
                avg_val = safe_float(metrics.get('avg_upside'))
                best_val = metrics.get('best_opportunity')
                avg_ok = avg_val is not None and abs(avg_val - ref_avg) <= 1.0
                best_ok = (best_val is not None
                           and str(best_val).strip() == ref_best)
            check("Metrics: Avg_Upside и Best_Opportunity выведены из данных",
                  avg_ok and best_ok,
                  f"avg_ok={avg_ok}, best_ok={best_ok}")
        else:
            check("Metrics: Avg_Upside и Best_Opportunity выведены из данных", False, "нет листа")

        # ---------------- Recommendations ----------------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        rec_symbols = []
        rec_rowcount = 0
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = norm_header(ws)
            rrows = list(ws.iter_rows(min_row=2, values_only=True))
            rrows = [r for r in rrows if r and any(c is not None for c in r)]
            rec_rowcount = len(rrows)
            check("Recommendations has >= 2 rows", len(rrows) >= 2, f"got {len(rrows)}")
            for expected_col in ['Priority', 'Action', 'Symbol']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            i_rs = col_index(headers, 'Symbol')
            if i_rs is not None:
                rec_symbols = [str(r[i_rs]).strip() for r in rrows if r[i_rs] is not None]

        # ---------------- Word ----------------
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        word_text = ""
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            word_text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content", len(word_text) > 50, f"text length: {len(word_text)}")

        # CRITICAL: Word три раздела (RU+EN) + Recommendations ссылается на тикеры Data_Analysis
        required_sections = [
            ["executive summary", "краткое резюме", "резюме"],
            ["key findings", "ключевые выводы", "выводы"],
            ["recommendations", "рекомендации"],
        ]
        sections_ok = all(any(kw in word_text for kw in grp) for grp in required_sections)
        rec_refs_ok = (rec_rowcount >= 2
                       and len(rec_symbols) >= 2
                       and all(s in da_symbols for s in rec_symbols)
                       and len(da_symbols) >= 5)
        check("Word + Recommendations: три раздела и ссылки на тикеры Data_Analysis",
              sections_ok and rec_refs_ok,
              f"sections_ok={sections_ok}, rec_refs_ok={rec_refs_ok}, rec_symbols={rec_symbols}")

        check("yf_sector_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "yf_sector_processor.py")))
    else:
        # без Excel остальные критические проверки тоже падают
        for nm in CRITICAL_CHECKS:
            check(nm, False, "нет Excel-файла")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    message = f"Passed {PASS_COUNT}/{total} checks (accuracy={accuracy:.1f}%)"

    if CRITICAL_FAILED:
        message += " | КРИТИЧЕСКАЯ проверка провалена => FAIL"
        return False, message

    success = accuracy >= 70.0
    return success, message


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
