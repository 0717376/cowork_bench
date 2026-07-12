"""
Evaluation script for yf-portfolio-risk-excel task (MOEX, RU).

Проверяет Excel-файл Portfolio_Risk_Analysis.xlsx (3 листа), построенный
по живым данным moex.stock_prices (рублёвые тикеры Московской биржи).

CRITICAL чеки (любой провал => немедленный FAIL вне зависимости от accuracy):
  C1: Risk Metrics — avg, std, min И max каждого тикера совпадают с расчётными
      (по строке тикера, через number_close_match).
  C2: Risk Assessment — коэффициент вариации каждого тикера совпадает с round(cv,4)
      И категория риска корректно выведена из CV (Low/<0.10, Medium/0.10-0.20, High/>0.20)
      по строке самого тикера.
  C3: Price History — выборочная цена закрытия по каждому тикеру совпадает с moex.stock_prices.
  C4: Книга содержит ровно 3 требуемых листа и все 5 тикеров MOEX в каждом релевантном листе.

PASS, если нет провалов критичных чеков И accuracy >= 70%.
"""
import argparse
import math
import os
import sys
from collections import defaultdict

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

TICKERS = ["SBER.ME", "GAZP.ME", "LKOH.ME", "TCSG.ME", "MGNT.ME"]

CRITICAL_CHECKS = set()


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    prices = defaultdict(dict)
    for ticker in TICKERS:
        cur.execute("SELECT date, close FROM moex.stock_prices WHERE symbol = %s ORDER BY date",
                    (ticker,))
        for date, close in cur.fetchall():
            prices[date][ticker] = float(close)

    valid_dates = sorted([d for d, p in prices.items() if len(p) == len(TICKERS)])

    stats = {}
    for ticker in TICKERS:
        vals = [prices[d][ticker] for d in valid_dates]
        avg_val = sum(vals) / len(vals)
        std_val = math.sqrt(sum((v - avg_val)**2 for v in vals) / (len(vals) - 1))
        cv = std_val / avg_val
        if cv < 0.10:
            category = "Low Risk"
        elif cv <= 0.20:
            category = "Medium Risk"
        else:
            category = "High Risk"
        stats[ticker] = {
            "avg": round(avg_val, 2),
            "std": round(std_val, 2),
            "min": round(min(vals), 2),
            "max": round(max(vals), 2),
            "cv": round(cv, 4),
            "category": category,
        }

    cur.close()
    conn.close()
    # вернём и сырые цены для выборочной проверки Price History
    return stats, valid_dates, prices


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if critical:
        CRITICAL_CHECKS.add(name)
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:200]}" if detail else ""
        marker = " [CRITICAL]" if critical else ""
        print(f"  [FAIL]{marker} {name}{detail_str}")
        if critical:
            CRITICAL_FAILS.append(name)


def number_close_match(expected, actual, tolerance=0.05):
    """Check if actual is within tolerance of expected."""
    try:
        return abs(float(actual) - float(expected)) <= abs(float(expected) * tolerance) + 0.01
    except (ValueError, TypeError):
        return False


def _norm_cell(v):
    """Нормализация значения ячейки к строке для сопоставления тикеров."""
    return str(v).strip().upper() if v is not None else ""


def _row_for_ticker(ws, ticker):
    """Возвращает список значений строки, в которой встречается тикер (любая ячейка)."""
    tu = ticker.upper()
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if c is not None and _norm_cell(c) == tu:
                return list(row)
    return None


def _find_col_in_row(row, ticker):
    """Индекс ячейки тикера в строке (для row-aligned чтения)."""
    tu = ticker.upper()
    for i, c in enumerate(row):
        if c is not None and _norm_cell(c) == tu:
            return i
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Portfolio_Risk_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(xlsx_path), f"Expected {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        return

    try:
        wb = load_workbook(xlsx_path)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    expected_stats, valid_dates, prices = load_expected()
    expected_rows = len(valid_dates)

    def find_sheet(keywords):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(k in sl for k in keywords):
                return wb[s]
        return None

    # C4: ровно 3 листа
    check("Workbook has exactly 3 sheets",
          len(wb.sheetnames) == 3,
          f"Sheets: {wb.sheetnames}",
          critical=True)

    # ---- Price History ----
    ws_ph = find_sheet(["price"]) or find_sheet(["history"])
    check("Price History sheet exists", ws_ph is not None, f"Sheets: {wb.sheetnames}")
    if ws_ph:
        header = [_norm_cell(c.value) for c in ws_ph[1]]
        ph_all_tickers = all(t.upper() in header for t in TICKERS)
        check("C4: Price History header contains all 5 MOEX tickers",
              ph_all_tickers, f"Header: {header}", critical=True)

        data_rows = ws_ph.max_row - 1
        check(f"Price History has ~{expected_rows} data rows",
              abs(data_rows - expected_rows) <= 5,
              f"Found {data_rows} rows, expected ~{expected_rows}")

        # дату-столбец найдём как первый столбец, не являющийся тикером
        # построим карту тикер -> индекс столбца по заголовку
        col_idx = {}
        for i, h in enumerate(header):
            for t in TICKERS:
                if h == t.upper():
                    col_idx[t] = i

        # выборочная сверка цены закрытия (середина диапазона) по каждому тикеру
        if valid_dates and col_idx:
            # соберём прочитанные значения по дате из листа: ключ = строковая дата
            sheet_rows = {}
            for row in ws_ph.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                date_key = str(row[0]).strip()[:10] if row[0] is not None else None
                if date_key:
                    sheet_rows[date_key] = row

            sample_date = valid_dates[len(valid_dates) // 2]
            sample_key = sample_date.strftime("%Y-%m-%d") if hasattr(sample_date, "strftime") else str(sample_date)
            srow = sheet_rows.get(sample_key)
            ph_sample_ok = srow is not None
            detail = f"sample date {sample_key} not found in Price History"
            if srow is not None:
                for t in TICKERS:
                    if t not in col_idx or col_idx[t] >= len(srow):
                        ph_sample_ok = False
                        detail = f"column for {t} missing"
                        break
                    exp_close = round(prices[sample_date][t], 2)
                    if not number_close_match(exp_close, srow[col_idx[t]]):
                        ph_sample_ok = False
                        detail = f"{t} on {sample_key}: expected ~{exp_close}, got {srow[col_idx[t]]}"
                        break
            check("C3: Price History sampled closing prices match moex.stock_prices",
                  ph_sample_ok, detail, critical=True)
        else:
            check("C3: Price History sampled closing prices match moex.stock_prices",
                  False, "no valid dates or ticker columns", critical=True)

    # ---- Risk Metrics ----
    ws_rm = find_sheet(["risk", "metric"]) or find_sheet(["metric"])
    check("Risk Metrics sheet exists", ws_rm is not None, f"Sheets: {wb.sheetnames}")
    if ws_rm:
        all_text = ""
        for row in ws_rm.iter_rows(values_only=True):
            all_text += " ".join(str(c) for c in row if c is not None) + " "

        for ticker in TICKERS:
            s = expected_stats[ticker]
            check(f"Risk Metrics: {ticker} present", ticker.upper() in all_text.upper())

            row = _row_for_ticker(ws_rm, ticker)
            ok = False
            detail = f"row for {ticker} not found"
            if row is not None:
                # числовые ячейки в строке тикера (без самого тикера)
                ti = _find_col_in_row(row, ticker)
                nums = []
                for i, c in enumerate(row):
                    if i == ti:
                        continue
                    try:
                        nums.append(float(c))
                    except (ValueError, TypeError):
                        pass
                # требуем, чтобы все 4 ожидаемых метрики нашли соответствие в строке
                exp_metrics = {"avg": s["avg"], "std": s["std"], "min": s["min"], "max": s["max"]}
                missing = []
                for key, ev in exp_metrics.items():
                    if not any(number_close_match(ev, n) for n in nums):
                        missing.append(f"{key}={ev}")
                ok = len(missing) == 0
                detail = f"{ticker}: missing/mismatched {missing} in row nums {nums}"
            check(f"C1: Risk Metrics {ticker} avg/std/min/max match expected",
                  ok, detail, critical=True)

    # ---- Risk Assessment ----
    ws_ra = find_sheet(["risk", "assess"]) or find_sheet(["assess"])
    check("Risk Assessment sheet exists", ws_ra is not None, f"Sheets: {wb.sheetnames}")
    if ws_ra:
        all_text = ""
        for row in ws_ra.iter_rows(values_only=True):
            all_text += " ".join(str(c) for c in row if c is not None) + " "

        for ticker in TICKERS:
            s = expected_stats[ticker]
            check(f"Risk Assessment: {ticker} present", ticker.upper() in all_text.upper())

            row = _row_for_ticker(ws_ra, ticker)
            ok = False
            detail = f"row for {ticker} not found"
            if row is not None:
                ti = _find_col_in_row(row, ticker)
                row_text = " ".join(str(c) for c in row if c is not None).lower()
                # CV: проверяем, что round(cv,4) встречается в числах строки
                nums = []
                for i, c in enumerate(row):
                    if i == ti:
                        continue
                    try:
                        nums.append(float(c))
                    except (ValueError, TypeError):
                        pass
                cv_ok = any(number_close_match(s["cv"], n, tolerance=0.02) for n in nums)
                # категория: корректная строка категории присутствует в строке тикера
                cat_ok = s["category"].lower() in row_text
                ok = cv_ok and cat_ok
                detail = (f"{ticker}: cv_expected={s['cv']} (cv_ok={cv_ok}), "
                          f"category_expected='{s['category']}' (cat_ok={cat_ok}), row_nums={nums}")
            check(f"C2: Risk Assessment {ticker} CV value and derived Risk Category correct",
                  ok, detail, critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0
    no_critical = len(CRITICAL_FAILS) == 0
    success = no_critical and accuracy >= 70

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    print(f"  Critical fails: {CRITICAL_FAILS}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if not no_critical:
        print(f"FAIL: провалены критичные чеки: {CRITICAL_FAILS}")
        sys.exit(1)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
