"""Evaluation for arxiv-reading-plan-excel-gcal-email.

Checks:
1. Reading_Plan.xlsx with Papers sheet (8 rows) and Schedule sheet (8 rows)
2. 8 Google Calendar events for reading sessions
3. Email to reading-group@lab.example.com with "LLM Agent Research Reading Plan" in subject

CRITICAL_CHECKS: a curated subset of semantic checks. Any failure in this set
=> overall FAIL regardless of accuracy. Structural checks (sheet exists, column
present) are non-critical. Accuracy threshold stays >= 70.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# paper_ids.txt order (== Assigned_Session order)
ARXIV_IDS = ["2301.13379", "2302.01560", "2303.12528", "2305.10403",
             "2308.12950", "2309.17453", "2201.11903", "2310.06825"]

# Seeded (arxiv.papers) ID -> canonical English title. Used to verify the
# agent mapped each ID to the right paper. Match is substring-based on a few
# distinctive keywords so abbreviated titles still pass.
ID_TITLE_KEYWORDS = {
    "2301.13379": ["zero-shot communicat"],
    "2302.01560": ["toolformer"],
    "2303.12528": ["hugginggpt"],
    "2305.10403": ["tree of thoughts"],
    "2308.12950": ["agentbench"],
    "2309.17453": ["self-rag"],
    "2201.11903": ["chain-of-thought", "chain of thought"],
    "2310.06825": ["mistral 7b", "mistral-7b"],
}

# The 8 consecutive Mondays starting 2026-03-09 (one per session).
EXPECTED_DATES = ["2026-03-09", "2026-03-16", "2026-03-23", "2026-03-30",
                  "2026-04-06", "2026-04-13", "2026-04-20", "2026-04-27"]

# Critical checks: any failure here => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Papers sheet has 8 rows",
    "Papers maps >=6/8 arXiv IDs to correct titles",
    "Assigned_Session numbering 1..8 matches paper_ids.txt order",
    "Schedule dates are the 8 consecutive Mondays from 2026-03-09",
    "8 'Reading Session' events on the expected Mondays",
    "Email subject contains 'LLM Agent Research Reading Plan'",
    "Email body mentions paper count and >=3 topics",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_ws, groundtruth_ws="."):
    print("\n=== Check 1: Reading_Plan.xlsx ===")
    path = os.path.join(agent_ws, "Reading_Plan.xlsx")
    check("File Reading_Plan.xlsx exists", os.path.isfile(path))
    if not os.path.isfile(path):
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Excel is readable", False, str(e))
        return

    # ---- Papers sheet ----
    papers_ws = None
    for sname in wb.sheetnames:
        if "paper" in sname.lower():
            papers_ws = wb[sname]
            break
    check("Sheet 'Papers' exists", papers_ws is not None, f"Sheets: {wb.sheetnames}")

    if papers_ws is not None:
        header_row = list(papers_ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        header_cells = [str(c).strip().lower() if c is not None else "" for c in header_row]
        header_text = " ".join(header_cells)

        rows = list(papers_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Papers sheet has 8 rows", len(non_empty) == 8, f"Got {len(non_empty)}")  # CRITICAL

        # column indices
        def col_idx(*keys):
            for i, h in enumerate(header_cells):
                if any(k in h for k in keys):
                    return i
            return None

        id_i = col_idx("arxiv")
        title_i = col_idx("title")
        sess_i = col_idx("assigned", "session")

        # arXiv IDs present anywhere (structural, non-critical)
        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        found_ids = sum(1 for arxiv_id in ARXIV_IDS if arxiv_id in all_text)
        check("Papers sheet contains at least 6 of 8 arXiv IDs",
              found_ids >= 6, f"Found {found_ids}/8 IDs in: {all_text[:200]}")

        # CRITICAL: each ID mapped to its correct title (>=6/8 pairings)
        pair_ok = 0
        if id_i is not None and title_i is not None:
            for r in non_empty:
                if id_i >= len(r) or title_i >= len(r):
                    continue
                rid = str(r[id_i]).strip() if r[id_i] is not None else ""
                rtitle = str(r[title_i]).strip().lower() if r[title_i] is not None else ""
                kws = ID_TITLE_KEYWORDS.get(rid)
                if kws and any(k in rtitle for k in kws):
                    pair_ok += 1
        check("Papers maps >=6/8 arXiv IDs to correct titles",
              pair_ok >= 6, f"Correct ID->title pairings: {pair_ok}/8")  # CRITICAL

        # CRITICAL: Assigned_Session 1..8 in paper_ids.txt order
        session_ok = False
        if id_i is not None and sess_i is not None:
            id_to_sess = {}
            for r in non_empty:
                if id_i >= len(r) or sess_i >= len(r):
                    continue
                rid = str(r[id_i]).strip() if r[id_i] is not None else ""
                try:
                    rsess = int(float(r[sess_i])) if r[sess_i] is not None else None
                except (TypeError, ValueError):
                    rsess = None
                if rid:
                    id_to_sess[rid] = rsess
            session_ok = all(id_to_sess.get(aid) == n + 1 for n, aid in enumerate(ARXIV_IDS))
        check("Assigned_Session numbering 1..8 matches paper_ids.txt order",
              session_ok, f"Mapping: {id_to_sess if id_i is not None and sess_i is not None else 'cols missing'}")  # CRITICAL

        # structural header checks (non-critical)
        check("Papers header has ArXiv_ID or arxiv column", "arxiv" in header_text, f"Header: {header_row}")
        check("Papers header has Title column", "title" in header_text, f"Header: {header_row}")
        check("Papers header has Session column",
              "session" in header_text or "assigned" in header_text, f"Header: {header_row}")

    # ---- Schedule sheet ----
    schedule_ws = None
    for sname in wb.sheetnames:
        if "schedule" in sname.lower():
            schedule_ws = wb[sname]
            break
    check("Sheet 'Schedule' exists", schedule_ws is not None, f"Sheets: {wb.sheetnames}")

    if schedule_ws is not None:
        rows = list(schedule_ws.iter_rows(min_row=2, values_only=True))
        non_empty = [r for r in rows if any(c is not None for c in r)]
        check("Schedule sheet has 8 rows", len(non_empty) == 8, f"Got {len(non_empty)}")

        all_text = " ".join(str(c) for row in non_empty for c in row if c is not None)
        # CRITICAL: the 8 consecutive Mondays appear (parsed as dates, not just 'march')
        found_dates = sum(1 for d in EXPECTED_DATES if d in all_text)
        check("Schedule dates are the 8 consecutive Mondays from 2026-03-09",
              found_dates == 8, f"Found {found_dates}/8 expected dates in: {all_text[:200]}")  # CRITICAL

        check("Schedule has March 2026 dates",
              "2026" in all_text and ("march" in all_text.lower() or "2026-03" in all_text or "март" in all_text.lower()),
              f"Date content: {all_text[:200]}")

    # ---- Groundtruth value comparison (non-critical, soft) ----
    gt_path = os.path.join(groundtruth_ws, "Reading_Plan.xlsx")
    if not os.path.isfile(gt_path):
        check("Groundtruth xlsx exists", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws_sheet = gt_wb[gt_sheet_name]
        agent_ws_sheet = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws_sheet = wb[asn]
                break
        if agent_ws_sheet is None:
            check(f"GT sheet '{gt_sheet_name}' exists in agent", False, f"Available: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws_sheet.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws_sheet.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        check(f"GT '{gt_sheet_name}' row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        check_indices_list = list(range(min(3, len(gt_rows))))
        if len(gt_rows) > 3:
            check_indices_list.append(len(gt_rows) - 1)
        for idx in check_indices_list:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                row_ok = True
                for col_idx2 in range(min(len(gt_row), len(a_row) if a_row else 0)):
                    gt_val = gt_row[col_idx2]
                    a_val = a_row[col_idx2]
                    if gt_val is None:
                        continue
                    if isinstance(gt_val, (int, float)):
                        ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                    else:
                        ok = str_match(a_val, gt_val)
                    if not ok:
                        check(f"GT '{gt_sheet_name}' row {idx+1} col {col_idx2+1}",
                              False, f"Expected {gt_val}, got {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    check(f"GT '{gt_sheet_name}' row {idx+1} values match", True)
            else:
                check(f"GT '{gt_sheet_name}' row {idx+1} exists", False, "Row missing in agent")
    gt_wb.close()


def check_gcal():
    print("\n=== Check 2: Google Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, summary, start_datetime, description FROM gcal.events
        WHERE summary ILIKE '%reading session%'
           OR summary ILIKE '%reading%session%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    check("At least 8 'Reading Session' calendar events created",
          len(events) >= 8, f"Found {len(events)} events")

    summaries = [e[1] or "" for e in events]
    start_dates = [str(e[2]) for e in events]
    descriptions = [str(e[3] or "") for e in events]

    if events:
        check("Events start in March 2026",
              any("2026-03" in d for d in start_dates),
              f"Dates: {start_dates[:4]}")
        check("Events cover 8 weeks (April 2026 included)",
              any("2026-04" in d for d in start_dates),
              f"Dates: {start_dates}")
        check("Events have 'Reading Session' in title",
              all("reading" in s.lower() for s in summaries[:8]),
              f"Titles: {summaries[:4]}")

    # CRITICAL: 8 events whose start falls on the 8 expected Mondays, and each
    # of those events carries an arXiv ID in its description.
    matched_dates = 0
    desc_with_id = 0
    for exp in EXPECTED_DATES:
        for s, d, desc in zip(summaries, start_dates, descriptions):
            if exp in d:
                matched_dates += 1
                if any(aid in desc for aid in ARXIV_IDS):
                    desc_with_id += 1
                break
    check("8 'Reading Session' events on the expected Mondays",
          matched_dates == 8 and desc_with_id >= 6,
          f"date-matched={matched_dates}/8, with-arxiv-id-in-desc={desc_with_id}/8")  # CRITICAL

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE subject ILIKE '%reading plan%'
           OR subject ILIKE '%LLM%reading%'
           OR subject ILIKE '%LLM Agent%'
           OR to_addr::text ILIKE '%reading-group%'
        LIMIT 10
    """)
    rows = cur.fetchall()
    check("Email with reading plan subject found",
          len(rows) > 0, "No matching email found")

    if rows:
        to_addrs = [str(r[1]) for r in rows]
        check("Email sent to reading-group@lab.example.com",
              any("reading-group" in addr for addr in to_addrs),
              f"To addresses: {to_addrs}")
        subjects = [r[0] or "" for r in rows]

        # CRITICAL: required subject substring (literal identifier, English)
        check("Email subject contains 'LLM Agent Research Reading Plan'",
              any("llm agent research reading plan" in s.lower() for s in subjects),
              f"Subjects: {subjects}")  # CRITICAL

        bodies = [str(r[2] or "").lower() for r in rows]

        # CRITICAL: body mentions the paper count (8 / восемь) AND at least 3
        # distinct topics/titles (RU or EN equivalents accepted).
        topic_keywords = [
            "toolformer", "agentbench", "self-rag", "hugginggpt",
            "tree of thoughts", "chain-of-thought", "chain of thought",
            "mistral", "zero-shot", "rag", "цепочк", "дерев", "агент",
            "рассужден", "инструмент",
        ]
        body_ok = False
        for b in bodies:
            has_count = ("8" in b or "восемь" in b or "eight" in b)
            topic_hits = sum(1 for k in topic_keywords if k in b)
            if has_count and topic_hits >= 3:
                body_ok = True
                break
        check("Email body mentions paper count and >=3 topics",
              body_ok,
              f"Body sample: {bodies[0][:200] if bodies else ''}")  # CRITICAL

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Evaluation: arxiv-reading-plan-excel-gcal-email ===")

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "pass": PASS_COUNT, "fail": FAIL_COUNT,
                "total_passed": PASS_COUNT, "total_checks": total,
                "accuracy": accuracy, "critical_failed": critical_failed,
            }, f)

    success = (not critical_failed) and accuracy >= 70
    if success:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
