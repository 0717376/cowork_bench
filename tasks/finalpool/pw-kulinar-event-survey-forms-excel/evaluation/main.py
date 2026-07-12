"""Evaluation для pw-kulinar-event-survey-forms-excel (RU-стек: kulinar/forms).

Проверяет отчёт по опросу поваров:
  - Excel Event_Survey_Report.xlsx: листы Data_Analysis (Recipe/Category/Calories/
    Protein_g/Meets_Guidelines), Metrics (Metric/Value), Recommendations (Priority/Action).
  - Форма forms (gform.*) с заголовком ровно «Cook Survey Feedback» и >=1 вопросом.
  - cook_survey_processor.py создан в рабочей директории.
  - cook_survey_results.json содержит нормативы, считанные с mock-страницы.

Критические чеки (CRITICAL_CHECKS): любой их fail => задача FAIL, даже если общая
accuracy >= 70%. Остальные (структурные) — мягкие. Порог: accuracy >= 70% И нет
критических провалов.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

# Критические чеки по имени check() — провал любого => FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Data_Analysis: >=5 реальных блюд с числовыми Calories/Protein_g и валидным Meets_Guidelines",
    "Metrics согласован с Data_Analysis (Total_Recipes/Avg_Calories/Recipes_Meeting_Guidelines)",
    "Data_Analysis отсортирован по алфавиту по колонке Recipe",
    "Форма «Cook Survey Feedback» существует и содержит >=1 вопрос",
    "cook_survey_results.json содержит нормативы со страницы (Protein=50000)",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '.').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

# Значения Meets_Guidelines — RU+EN положительные/отрицательные.
YES_VALUES = {"yes", "да", "true", "соответствует"}
NO_VALUES = {"no", "нет", "false", "не соответствует"}


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    excel_path = os.path.join(agent_workspace, "Event_Survey_Report.xlsx")
    check("Event_Survey_Report.xlsx exists", os.path.exists(excel_path))

    da_rows = []          # содержимое Data_Analysis (для критических перерасчётов)
    da_headers = []
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---------------- Data_Analysis ----------------
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_rows = list(ws.iter_rows(min_row=2, values_only=True))
            da_rows = [r for r in da_rows if r and any(c is not None and str(c).strip() for c in r)]
            check("Data_Analysis has >= 5 rows", len(da_rows) >= 5, f"got {len(da_rows)}")

            da_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Recipe', 'Category', 'Calories', 'Protein_g', 'Meets_Guidelines']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in da_headers, f"headers: {da_headers[:8]}")

        # ---------------- Metrics ----------------
        metrics_map = {}
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            mrows = list(ws.iter_rows(min_row=2, values_only=True))
            mrows = [r for r in mrows if r and r[0]]
            check("Metrics has >= 4 rows", len(mrows) >= 4, f"got {len(mrows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in mrows:
                key = str(r[0]).strip().lower() if r[0] is not None else ""
                val = r[1] if len(r) > 1 else None
                metrics_map[key] = val

        # ---------------- Recommendations ----------------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            rrows = list(ws.iter_rows(min_row=2, values_only=True))
            rrows = [r for r in rrows if r and any(c is not None and str(c).strip() for c in r)]
            check("Recommendations has >= 2 rows", len(rrows) >= 2, f"got {len(rrows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ============ КРИТИЧЕСКИЕ ЧЕКИ по данным Data_Analysis ============
        # Индексы колонок (по заголовкам, с дефолтами 0..4).
        def col_idx(name, default):
            return da_headers.index(name) if name in da_headers else default
        i_recipe = col_idx('recipe', 0)
        i_cal = col_idx('calories', 2)
        i_prot = col_idx('protein_g', 3)
        i_meet = col_idx('meets_guidelines', 4)

        recipes = []
        cals = []
        prots = []
        meet_yes = 0
        all_meet_valid = True
        for r in da_rows:
            if i_recipe < len(r) and r[i_recipe]:
                recipes.append(str(r[i_recipe]).strip())
            c = safe_float(r[i_cal]) if i_cal < len(r) else None
            p = safe_float(r[i_prot]) if i_prot < len(r) else None
            cals.append(c)
            prots.append(p)
            mv = str(r[i_meet]).strip().lower() if i_meet < len(r) and r[i_meet] is not None else ""
            if mv in YES_VALUES:
                meet_yes += 1
            elif mv in NO_VALUES:
                pass
            else:
                all_meet_valid = False

        distinct_real = len([x for x in set(recipes) if x and "placeholder" not in x.lower()
                             and "recipe" != x.lower() and "пример" not in x.lower()])
        nums_ok = all(c is not None for c in cals) and all(p is not None for p in prots)
        check("Data_Analysis: >=5 реальных блюд с числовыми Calories/Protein_g и валидным Meets_Guidelines",
              len(da_rows) >= 5 and distinct_real >= 5 and nums_ok and all_meet_valid,
              f"rows={len(da_rows)} distinct={distinct_real} nums_ok={nums_ok} meet_valid={all_meet_valid}")

        # Metrics согласован с Data_Analysis (перерасчёт, не хардкод).
        consistent = True
        detail = []
        if cals and all(c is not None for c in cals):
            exp_total = len(da_rows)
            exp_avg = round(sum(cals) / len(cals))
            got_total = safe_float(metrics_map.get("total_recipes"))
            got_avg = safe_float(metrics_map.get("avg_calories"))
            got_meet = safe_float(metrics_map.get("recipes_meeting_guidelines"))
            if got_total is None or int(got_total) != exp_total:
                consistent = False; detail.append(f"Total_Recipes exp={exp_total} got={got_total}")
            if got_avg is None or abs(got_avg - exp_avg) > 1:
                consistent = False; detail.append(f"Avg_Calories exp={exp_avg} got={got_avg}")
            if got_meet is None or int(got_meet) != meet_yes:
                consistent = False; detail.append(f"Meeting exp={meet_yes} got={got_meet}")
        else:
            consistent = False; detail.append("Calories не числовые")
        check("Metrics согласован с Data_Analysis (Total_Recipes/Avg_Calories/Recipes_Meeting_Guidelines)",
              consistent, "; ".join(detail))

        # Сортировка по алфавиту по Recipe (RU collation: locale-независимый lower-codepoint порядок).
        sorted_ok = recipes == sorted(recipes, key=lambda s: s.lower())
        check("Data_Analysis отсортирован по алфавиту по колонке Recipe",
              sorted_ok and len(recipes) >= 5, f"recipes={recipes}")
    else:
        # Файла нет — критические Excel-чеки проваливаем явно.
        check("Data_Analysis: >=5 реальных блюд с числовыми Calories/Protein_g и валидным Meets_Guidelines",
              False, "нет файла")
        check("Metrics согласован с Data_Analysis (Total_Recipes/Avg_Calories/Recipes_Meeting_Guidelines)",
              False, "нет файла")
        check("Data_Analysis отсортирован по алфавиту по колонке Recipe", False, "нет файла")

    # ---------------- Форма (gform.*) ----------------
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gform.forms WHERE lower(trim(title)) = %s",
                    ('cook survey feedback',))
        forms = cur.fetchall()
        has_form = len(forms) >= 1
        n_questions = 0
        if has_form:
            fid = forms[0][0]
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (fid,))
            n_questions = cur.fetchone()[0]
        check("Форма «Cook Survey Feedback» существует и содержит >=1 вопрос",
              has_form and n_questions >= 1, f"forms={len(forms)} questions={n_questions}")
        conn.close()
    except Exception as e:
        check("Форма «Cook Survey Feedback» существует и содержит >=1 вопрос", False, str(e))

    # ---------------- cook_survey_processor.py ----------------
    check("cook_survey_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "cook_survey_processor.py")))

    # ---------------- cook_survey_results.json: нормативы со страницы ----------------
    results_path = os.path.join(agent_workspace, "cook_survey_results.json")
    bench_ok = False
    bench_detail = ""
    if os.path.exists(results_path):
        try:
            raw = open(results_path, encoding="utf-8").read()
            # Грубая, но устойчивая проверка: присутствуют токены страницы и значение Protein=50000.
            low = raw.lower()
            has_protein_val = "50000" in raw
            has_tokens = "daily_recommended_mg" in low or "nutrient" in low or "protein" in low
            bench_ok = has_protein_val and has_tokens
            bench_detail = f"protein_val={has_protein_val} tokens={has_tokens}"
        except Exception as e:
            bench_detail = str(e)
    else:
        bench_detail = "нет cook_survey_results.json"
    check("cook_survey_results.json содержит нормативы со страницы (Protein=50000)",
          bench_ok, bench_detail)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    no_critical = len(CRITICAL_FAILED) == 0
    success = accuracy >= 70 and no_critical
    msg = (f"Passed {PASS_COUNT}/{total} checks (accuracy={accuracy:.1f}%); "
           f"critical_failed={CRITICAL_FAILED}")
    return success, msg

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
