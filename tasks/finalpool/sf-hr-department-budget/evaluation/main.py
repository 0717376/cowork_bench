"""Evaluation for sf-hr-department-budget (ClickHouse fork).

The HR_ANALYTICS departments in sf_data are russified centrally
(db/zzz_clickhouse_after_init.sql): Engineering->Инженерия, Finance->Финансы,
HR->Кадры, Operations->Операции, R&D->НИОКР, Sales->Продажи, Support->Поддержка.
The groundtruth (_data.json + Budget_Analysis.xlsx Department column) carries those
Cyrillic names, so every value-based department lookup matches the agent's output
only if it queried the live ClickHouse-backed DWH. Numeric budget/salary values are
NOT translated, so the groundtruth figures are identical to the English original.
"""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def check_excel(agent_workspace, gt_dir):
    """Check Budget_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Budget_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Budget_Analysis.xlsx")

    if not os.path.exists(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_ok = True

    # Sheet 1: Department Budget
    a_rows = load_sheet_rows(agent_wb, "Department Budget")
    g_rows = load_sheet_rows(gt_wb, "Department Budget")

    if a_rows is None:
        record("Sheet 'Department Budget' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Department Budget' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        record("Department Budget row count", len(a_data) == len(g_data),
               f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                # CRITICAL: the russified department must be present (proves the agent
                # queried the live ClickHouse sf_data, not the English original).
                record(f"Department '{g_row[0]}' present", False, "Missing", critical=True)
                all_ok = False
                continue

            errors = []
            critical_err = False
            # Budget (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 100000):
                    errors.append(f"Budget: {a_row[1]} vs {g_row[1]}")

            # Actual_Spend (col 2) - CRITICAL: must match the sum-of-salaries from sf_data
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 100000):
                    errors.append(f"Actual_Spend: {a_row[2]} vs {g_row[2]}")
                    critical_err = True

            # Variance (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 200000):
                    errors.append(f"Variance: {a_row[3]} vs {g_row[3]}")

            # Variance_Pct (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1.0):
                    errors.append(f"Variance_Pct: {a_row[4]} vs {g_row[4]}")

            # Employee_Count (col 5)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 10):
                    errors.append(f"Employee_Count: {a_row[5]} vs {g_row[5]}")

            # Avg_Salary (col 6)
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 500):
                    errors.append(f"Avg_Salary: {a_row[6]} vs {g_row[6]}")

            if errors:
                record(f"Department '{g_row[0]}' data", False, "; ".join(errors),
                       critical=critical_err)
                all_ok = False
            else:
                record(f"Department '{g_row[0]}' data", True)

    # Sheet 2: Summary
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    if a_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary: {g_row[0]} present", False, "Missing")
                all_ok = False
                continue

            # Large values need larger tolerance
            tol = 1000000 if "budget" in key or "spend" in key or "variance" in key.replace("_pct", "").replace("pct", "") else 1.0
            if "depts" in key or "over" in key or "under" in key:
                tol = 1.0

            # CRITICAL: the headline figures that prove the analysis is correct.
            crit = ("total_spend" in key or "overall_variance" in key
                    or "depts_over" in key or "depts_under" in key)

            if len(a_row) > 1 and len(g_row) > 1:
                ok = num_close(a_row[1], g_row[1], tol)
                record(f"Summary: {g_row[0]}", ok,
                       f"Expected {g_row[1]}, got {a_row[1]} (tol={tol})",
                       critical=crit)
                if not ok:
                    all_ok = False

    return all_ok


def check_gsheet():
    """Check Google Sheet was created."""
    print("\n=== Checking Google Sheet ===")

    try:
        conn = psycopg2.connect(**{"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym",
                                   "user": "eigent", "password": "camel"})
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%budget%'")
        sheets = cur.fetchall()
        if not sheets:
            # CRITICAL: the 'Department Budget Overview' Google Sheet deliverable must exist.
            record("Google Sheet with 'Budget' in title", False, "Not found", critical=True)
            cur.close()
            conn.close()
            return False
        record("Google Sheet with 'Budget' in title", True)

        sid = sheets[0][0]
        cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
        cell_count = cur.fetchone()[0]
        # CRITICAL: the sheet must be populated with the Department Budget data.
        record("Google Sheet has data (>=20 cells)", cell_count >= 20,
               f"Cell count: {cell_count}", critical=True)

        cur.close()
        conn.close()
        return cell_count >= 20
    except Exception as e:
        record("Google Sheet check", False, str(e), critical=True)
        return False


def check_emails():
    """Check email was sent to CFO."""
    print("\n=== Checking Emails ===")

    try:
        conn = psycopg2.connect(**{"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym",
                                   "user": "eigent", "password": "camel"})
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        all_emails = cur.fetchall()
        cur.close()
        conn.close()

        record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

        found = False
        for subject, from_addr, to_addr, body in all_emails:
            subj_lower = (subject or "").lower()
            # EN subject literal is preserved ('Department Budget Analysis').
            if "budget" in subj_lower or "department" in subj_lower:
                found = True
                to_str = str(to_addr or "").lower()
                # CRITICAL: addressed to the CFO.
                record("Email sent to cfo@company.example.com",
                       "cfo@company.example.com" in to_str, f"To: {to_addr}", critical=True)

                body_lower = (body or "").lower()
                # CRITICAL: body mentions budget figures. Agent body is Russian, so accept
                # RU (бюджет / отклонен / расход) OR preserved EN (budget / variance).
                body_ok = any(k in body_lower for k in (
                    "budget", "variance", "бюджет", "отклонен", "расход"))
                record("Email body mentions budget figures (RU/EN)", body_ok,
                       critical=True)
                break

        # CRITICAL: the CFO email deliverable must exist.
        record("Budget email found", found, critical=True)
        return found
    except Exception as e:
        record("Email check", False, str(e), critical=True)
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gsheet()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")

    # CRITICAL gate: any critical failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILS:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (threshold accuracy>=70)")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
