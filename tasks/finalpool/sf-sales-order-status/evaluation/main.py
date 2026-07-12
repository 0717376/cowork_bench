"""Evaluation for sf-sales-order-status (ClickHouse fork).

STATUS values are russified centrally by db/zzz_clickhouse_after_init.sql
(Delivered->Доставлен, Shipped->Отправлен, Processing->В обработке,
Cancelled->Отменён). The groundtruth Excel 'Order Status' sheet is keyed by
those Russian values; the agent reads them honestly from the DB.
"""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def build_lookup(rows):
    data = rows[1:] if rows and len(rows) > 1 else []
    lookup = {}
    for row in data:
        if row and row[0] is not None:
            lookup[str(row[0]).strip().lower()] = row
    return data, lookup


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Order_Status.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Order_Status.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Russified STATUS keys produced by the ClickHouse central map.
    STATUS_KEYS = ["доставлен", "отправлен", "в обработке", "отменён"]

    all_errors = []          # non-critical: feed accuracy ratio
    total_checks = 0
    passed_checks = 0
    critical_fail = []       # any non-empty => sys.exit(1) before accuracy gate

    def soft(ok, label):
        nonlocal total_checks, passed_checks
        total_checks += 1
        if ok:
            passed_checks += 1
        else:
            all_errors.append(label)

    # ------------------------------------------------------------------
    # Sheet: Order Status
    # ------------------------------------------------------------------
    print("  Checking Order Status...")
    a_rows = load_sheet_rows(agent_wb, "Order Status")
    g_rows = load_sheet_rows(gt_wb, "Order Status")

    a_os_lookup = {}
    a_os_data = []
    if a_rows is None:
        critical_fail.append("Sheet 'Order Status' not found in agent output")
    elif g_rows is None:
        critical_fail.append("Sheet 'Order Status' not found in groundtruth")
    else:
        a_os_data, a_os_lookup = build_lookup(a_rows)
        _, g_os_lookup = build_lookup(g_rows)

        for g_row in (g_rows[1:] if len(g_rows) > 1 else []):
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_os_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row: {g_row[0]}")
                total_checks += 1
                continue

            # Orders (tightened tol 5 -> 2)
            if len(a_row) > 1 and len(g_row) > 1:
                soft(num_close(a_row[1], g_row[1], 2),
                     f"{key}.Orders: {a_row[1]} vs {g_row[1]} (tol=2)")
            # Revenue (tightened tol 100 -> 1.0)
            if len(a_row) > 2 and len(g_row) > 2:
                soft(num_close(a_row[2], g_row[2], 1.0),
                     f"{key}.Revenue: {a_row[2]} vs {g_row[2]} (tol=1.0)")
            # Avg_Quantity (tightened tol 0.5 -> 0.15)
            if len(a_row) > 3 and len(g_row) > 3:
                soft(num_close(a_row[3], g_row[3], 0.15),
                     f"{key}.Avg_Quantity: {a_row[3]} vs {g_row[3]} (tol=0.15)")
            # Avg_Discount (tightened tol 0.5 -> 0.02; values ~0.07)
            if len(a_row) > 4 and len(g_row) > 4:
                soft(num_close(a_row[4], g_row[4], 0.02),
                     f"{key}.Avg_Discount: {a_row[4]} vs {g_row[4]} (tol=0.02)")
            # Revenue_Pct (NEW: previously never checked, tol 0.2)
            if len(a_row) > 5 and len(g_row) > 5:
                soft(num_close(a_row[5], g_row[5], 0.2),
                     f"{key}.Revenue_Pct: {a_row[5]} vs {g_row[5]} (tol=0.2)")

    # ------------------------------------------------------------------
    # Sheet: Summary
    # ------------------------------------------------------------------
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    a_sum_lookup = {}
    # tightened, metric-aware tolerances (was a blanket tol=100.0)
    SUMMARY_TOL = {
        "total_orders": 1.0,
        "total_revenue": 1.0,
        "delivered_revenue": 1.0,
        "cancel_rate_pct": 0.2,
    }
    if a_rows is None:
        critical_fail.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        critical_fail.append("Sheet 'Summary' not found in groundtruth")
    else:
        _, a_sum_lookup = build_lookup(a_rows)
        for g_row in (g_rows[1:] if len(g_rows) > 1 else []):
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_sum_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing Summary row: {g_row[0]}")
                total_checks += 1
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                tol = SUMMARY_TOL.get(key, 1.0)
                soft(num_close(a_row[1], g_row[1], tol),
                     f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")

    # ==================================================================
    # CRITICAL CHECKS (semantic). Any failure => FAIL regardless of accuracy.
    # ==================================================================
    print("  Checking CRITICAL semantic constraints...")

    # GT reference values (regenerated, Russian status keys).
    _, g_os_lookup = build_lookup(g_rows := load_sheet_rows(gt_wb, "Order Status"))
    _, g_sum_lookup = build_lookup(load_sheet_rows(gt_wb, "Summary"))

    # C1: all four russified STATUS rows present with exact-ish Orders (tol<=2)
    #     and Revenue within tol<=1.0.
    for k in STATUS_KEYS:
        a_row = a_os_lookup.get(k)
        g_row = g_os_lookup.get(k)
        if a_row is None:
            critical_fail.append(f"[C1] Missing russified status row: {k}")
            continue
        if g_row is None:
            continue
        if not num_close(a_row[1] if len(a_row) > 1 else None,
                         g_row[1] if len(g_row) > 1 else None, 2):
            critical_fail.append(
                f"[C1] {k}.Orders {a_row[1] if len(a_row)>1 else None} vs {g_row[1]} (tol=2)")
        if not num_close(a_row[2] if len(a_row) > 2 else None,
                         g_row[2] if len(g_row) > 2 else None, 1.0):
            critical_fail.append(
                f"[C1] {k}.Revenue {a_row[2] if len(a_row)>2 else None} vs {g_row[2]} (tol=1.0)")

    # C2: Revenue_Pct per status correct (tol<=0.2) AND sum ~= 100.
    pct_sum = 0.0
    pct_ok = True
    for k in STATUS_KEYS:
        a_row = a_os_lookup.get(k)
        g_row = g_os_lookup.get(k)
        if a_row is None or g_row is None or len(a_row) < 6 or len(g_row) < 6:
            pct_ok = False
            critical_fail.append(f"[C2] Revenue_Pct missing for {k}")
            continue
        v = to_float(a_row[5])
        if v is None:
            pct_ok = False
            critical_fail.append(f"[C2] Revenue_Pct non-numeric for {k}: {a_row[5]}")
            continue
        pct_sum += v
        if not num_close(a_row[5], g_row[5], 0.2):
            critical_fail.append(
                f"[C2] {k}.Revenue_Pct {a_row[5]} vs {g_row[5]} (tol=0.2)")
    if pct_ok and not num_close(pct_sum, 100.0, 0.5):
        critical_fail.append(f"[C2] sum(Revenue_Pct)={pct_sum:.2f} != ~100 (tol=0.5)")

    # C3: Order Status sorted by Revenue DESCENDING (Доставлен first).
    revs = []
    for row in a_os_data:
        if row and len(row) > 2 and to_float(row[2]) is not None:
            revs.append(to_float(row[2]))
    if len(revs) < 2:
        critical_fail.append("[C3] Order Status has fewer than 2 numeric Revenue rows")
    else:
        if any(revs[i] < revs[i + 1] - 1e-6 for i in range(len(revs) - 1)):
            critical_fail.append(f"[C3] Order Status not sorted by Revenue desc: {revs}")
        first_key = str(a_os_data[0][0]).strip().lower() if a_os_data and a_os_data[0] else None
        if first_key != "доставлен":
            critical_fail.append(f"[C3] First row is '{first_key}', expected 'доставлен'")

    # C4: Summary.Total_Orders == 20000 exact; Total_Revenue within tol<=1.0.
    a_to = a_sum_lookup.get("total_orders")
    if a_to is None or len(a_to) < 2 or not num_close(a_to[1], 20000, 0.5):
        critical_fail.append(
            f"[C4] Total_Orders {a_to[1] if a_to and len(a_to)>1 else None} != 20000")
    a_tr = a_sum_lookup.get("total_revenue")
    g_tr = g_sum_lookup.get("total_revenue")
    if a_tr is None or g_tr is None or not num_close(a_tr[1], g_tr[1], 1.0):
        critical_fail.append(
            f"[C4] Total_Revenue {a_tr[1] if a_tr and len(a_tr)>1 else None} "
            f"vs {g_tr[1] if g_tr and len(g_tr)>1 else None} (tol=1.0)")

    # C5: Delivered_Revenue (Доставлен only) and Cancel_Rate_Pct (gt 4.6) tight.
    a_dr = a_sum_lookup.get("delivered_revenue")
    g_dr = g_sum_lookup.get("delivered_revenue")
    if a_dr is None or g_dr is None or not num_close(a_dr[1], g_dr[1], 1.0):
        critical_fail.append(
            f"[C5] Delivered_Revenue {a_dr[1] if a_dr and len(a_dr)>1 else None} "
            f"vs {g_dr[1] if g_dr and len(g_dr)>1 else None} (tol=1.0)")
    else:
        # cross-check it equals the Доставлен row Revenue
        dv = a_os_lookup.get("доставлен")
        if dv is not None and len(dv) > 2 and not num_close(a_dr[1], dv[2], 1.0):
            critical_fail.append(
                f"[C5] Delivered_Revenue {a_dr[1]} != Доставлен.Revenue {dv[2]}")
    a_cr = a_sum_lookup.get("cancel_rate_pct")
    g_cr = g_sum_lookup.get("cancel_rate_pct")
    if a_cr is None or g_cr is None or not num_close(a_cr[1], g_cr[1], 0.2):
        critical_fail.append(
            f"[C5] Cancel_Rate_Pct {a_cr[1] if a_cr and len(a_cr)>1 else None} "
            f"vs {g_cr[1] if g_cr and len(g_cr)>1 else None} (tol=0.2)")

    # ------------------------------------------------------------------
    # Verdict
    # ------------------------------------------------------------------
    if critical_fail:
        print(f"\n=== CRITICAL FAIL ({len(critical_fail)}) ===")
        for e in critical_fail[:15]:
            print(f"  {e}")
        sys.exit(1)

    accuracy = (passed_checks / total_checks * 100.0) if total_checks else 0.0
    print(f"\n  Soft accuracy: {passed_checks}/{total_checks} = {accuracy:.1f}%")
    if all_errors:
        print(f"  Soft errors ({len(all_errors)}):")
        for e in all_errors[:10]:
            print(f"    {e}")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
