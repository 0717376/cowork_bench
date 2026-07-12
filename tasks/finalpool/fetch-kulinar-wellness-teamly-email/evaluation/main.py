"""Evaluation script for fetch-kulinar-wellness-teamly-email (russified).

Uses Teamly and kulinar. Critical checks (CRITICAL_CHECKS):
any failure => overall FAIL regardless of accuracy. Otherwise PASS requires
accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "cook_wellness_results.json exists and parsed nutrition guidelines",
    "Data_Analysis Meets_Guidelines logic consistent with Protein_g >= 50",
    "Data_Analysis sorted alphabetically by Recipe with >= 5 real recipes",
    "Analysis email to team-lead@company.com with exact subject",
    "Teamly 'Cook Wellness Dashboard' page exists with non-empty body",
}

# Положительные / отрицательные обозначения соответствия норме (RU + EN).
YES_VALUES = {"yes", "y", "true", "1", "да", "соответствует", "pass", "meets"}
NO_VALUES = {"no", "n", "false", "0", "нет", "не соответствует", "fail"}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def header_map(ws):
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def check_results_json(agent_workspace):
    """Critical: cook_wellness_results.json exists, valid JSON, references the
    parsed nutrition guidelines (Protein=50). cook_wellness_processor.py exists.
    """
    script_path = os.path.join(agent_workspace, "cook_wellness_processor.py")
    check("cook_wellness_processor.py exists", os.path.exists(script_path))

    res_path = os.path.join(agent_workspace, "cook_wellness_results.json")
    if not os.path.exists(res_path):
        check("cook_wellness_results.json exists and parsed nutrition guidelines",
              False, "missing file")
        return
    try:
        with open(res_path, "r", encoding="utf-8") as f:
            txt = f.read()
        json.loads(txt)  # must be valid JSON
    except Exception as e:
        check("cook_wellness_results.json exists and parsed nutrition guidelines",
              False, str(e))
        return
    low = txt.lower()
    # Protein guideline value (50) consumed from data.json must be present.
    has_protein = "protein" in low and "50" in txt
    check("cook_wellness_results.json exists and parsed nutrition guidelines",
          has_protein, f"protein/50 present: {has_protein}")


def check_excel(agent_workspace):
    excel_path = os.path.join(agent_workspace, "Wellness_Report.xlsx")
    exists = os.path.exists(excel_path)
    check("Wellness_Report.xlsx exists", exists)
    if not exists:
        # Mark dependent critical checks failed.
        check("Data_Analysis Meets_Guidelines logic consistent with Protein_g >= 50",
              False, "no excel")
        check("Data_Analysis sorted alphabetically by Recipe with >= 5 real recipes",
              False, "no excel")
        return

    wb = openpyxl.load_workbook(excel_path)

    # --- Data_Analysis sheet ---
    da_ok = "Data_Analysis" in wb.sheetnames
    check("Data_Analysis sheet exists", da_ok)
    if da_ok:
        ws = wb["Data_Analysis"]
        headers, hmap = header_map(ws)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")
        for expected_col in ['Recipe', 'Category', 'Calories', 'Protein_g', 'Meets_Guidelines']:
            check(f"Data_Analysis has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")

        recipe_i = hmap.get("recipe")
        protein_i = hmap.get("protein_g")
        meets_i = hmap.get("meets_guidelines")

        # Critical: alphabetical sort by Recipe + >= 5 non-empty real names.
        recipes = []
        for r in data_rows:
            if recipe_i is not None and recipe_i < len(r) and r[recipe_i] is not None:
                recipes.append(str(r[recipe_i]).strip())
        non_empty = [x for x in recipes if x]
        is_sorted = non_empty == sorted(non_empty, key=lambda s: s.lower())
        check("Data_Analysis sorted alphabetically by Recipe with >= 5 real recipes",
              len(non_empty) >= 5 and is_sorted,
              f"count={len(non_empty)} sorted={is_sorted}")

        # Critical: Meets_Guidelines derived from Protein_g >= 50.
        if protein_i is not None and meets_i is not None:
            consistent = True
            checked = 0
            for r in data_rows:
                if protein_i >= len(r) or meets_i >= len(r):
                    continue
                p = safe_float(r[protein_i])
                m_raw = r[meets_i]
                if p is None or m_raw is None:
                    continue
                m = str(m_raw).strip().lower()
                if m in YES_VALUES:
                    m_bool = True
                elif m in NO_VALUES:
                    m_bool = False
                else:
                    # Unknown token; treat as inconsistent only if clearly mismatched.
                    continue
                checked += 1
                if m_bool != (p >= 50):
                    consistent = False
                    break
            check("Data_Analysis Meets_Guidelines logic consistent with Protein_g >= 50",
                  consistent and checked >= 3,
                  f"checked={checked} consistent={consistent}")
        else:
            check("Data_Analysis Meets_Guidelines logic consistent with Protein_g >= 50",
                  False, "Protein_g or Meets_Guidelines column missing")
    else:
        check("Data_Analysis sorted alphabetically by Recipe with >= 5 real recipes",
              False, "no sheet")
        check("Data_Analysis Meets_Guidelines logic consistent with Protein_g >= 50",
              False, "no sheet")

    # --- Metrics sheet ---
    if "Metrics" in wb.sheetnames:
        check("Metrics sheet exists", True)
        ws = wb["Metrics"]
        headers, _ = header_map(ws)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")
        for expected_col in ['Metric', 'Value']:
            check(f"Metrics has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")
    else:
        check("Metrics sheet exists", False)

    # --- Recommendations sheet ---
    if "Recommendations" in wb.sheetnames:
        check("Recommendations sheet exists", True)
        ws = wb["Recommendations"]
        headers, _ = header_map(ws)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
        for expected_col in ['Priority', 'Action']:
            check(f"Recommendations has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")
    else:
        check("Recommendations sheet exists", False)


def check_email():
    """Critical: email to team-lead@company.com with exact subject
    'Analysis Report Complete' and non-empty body."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages "
                    "WHERE subject ILIKE %s", ('%analysis report complete%',))
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Analysis email to team-lead@company.com with exact subject", False, str(e))
        return

    matched = None
    for subject, to_addr, body in rows:
        to_str = str(to_addr).lower() if to_addr is not None else ""
        if "team-lead@company.com" in to_str:
            matched = (subject, to_addr, body)
            break
    check("Analysis email to team-lead@company.com with exact subject",
          matched is not None,
          f"candidates={[(s, t) for s, t, _ in rows]}")
    if matched is not None:
        body = str(matched[2]) if matched[2] else ""
        check("Analysis email has non-empty body", len(body.strip()) > 20,
              f"body len {len(body)}")


def check_teamly():
    """Critical: a 'Cook Wellness Dashboard' page exists with non-empty body.
    Seed pages have id <= 3; the noise page ('Старые заметки проекта') is a
    user-leftover and must NOT satisfy the dashboard check."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Cook Wellness Dashboard' page exists with non-empty body", False, str(e))
        return

    dash = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "старые заметки" in tl:
            continue
        if "dashboard" in tl or "дашборд" in tl or "панель" in tl or \
                ("cook" in tl and "wellness" in tl):
            dash = (pid, title, body)
            break
    body_ok = dash is not None and len(str(dash[2]).strip()) > 20
    check("Teamly 'Cook Wellness Dashboard' page exists with non-empty body",
          body_ok, f"new pages: {[(p[0], p[1]) for p in pages]}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    check_excel(agent_workspace)
    check_results_json(agent_workspace)
    check_email()
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
