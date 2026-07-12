"""
Evaluation script for fetch-sf-support-survey-gform-gcal-excel task.

Checks:
1. Support_Satisfaction_Analysis.xlsx with 4 sheets and correct data.
2. Customer feedback Form (RU forms-mcp, schema gform.*).
3. Calendar events for 4 quarterly reviews (correct dates/times).

CRITICAL_CHECKS reflect the task's substance: a single critical failure =>
overall FAIL (sys.exit(1)) regardless of accuracy. The accuracy>=70 gate
applies afterward.

Ticket-system numbers (ClickHouse / sf_data SUPPORT_CENTER) are recomputed
LIVE from the DB -- never hardcoded -- so they stay in sync with the russified
seed. Survey truth is recomputed LIVE from the served mock JSON.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [{tag}PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [{tag}FAIL] {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


# ---------------------------------------------------------------------------
# Source-of-truth helpers (computed live, never hardcoded)
# ---------------------------------------------------------------------------
def load_survey_truth():
    """Load the served mock survey JSON and compute the ground-truth aggregates.

    Returns dict or None if unavailable.
    """
    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    candidates = [
        os.path.join(task_root, "tmp", "mock_pages", "api", "survey_results.json"),
    ]
    data = None
    for p in candidates:
        if os.path.isfile(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                break
            except Exception:
                continue
    if data is None:
        # Fall back to the served endpoint.
        try:
            import urllib.request
            with urllib.request.urlopen(
                "http://localhost:30205/api/survey_results.json", timeout=5
            ) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except Exception:
            return None

    resp = data.get("responses", [])
    n = len(resp)
    if n == 0:
        return None

    def avg(key):
        return sum(r[key] for r in resp) / n

    by_priority = {}
    for r in resp:
        by_priority.setdefault(r["priority_experienced"], []).append(
            r["overall_satisfaction"]
        )
    prio_avg = {p: sum(v) / len(v) for p, v in by_priority.items()}
    lowest = min(prio_avg, key=prio_avg.get)
    highest = max(prio_avg, key=prio_avg.get)

    return {
        "n": n,
        "avg_overall": avg("overall_satisfaction"),
        "avg_response_time": avg("response_time_rating"),
        "avg_resolution": avg("resolution_quality_rating"),
        "avg_professionalism": avg("agent_professionalism"),
        "prio_avg": prio_avg,
        "lowest_priority": lowest,   # expected: Low
        "highest_priority": highest,  # expected: High
    }


def load_ticket_truth():
    """Recompute SUPPORT_CENTER ticket aggregates live from ClickHouse/sf_data.

    Returns {priority_lower: {avg_resp, avg_csat, count}} or None.
    """
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            'SELECT "PRIORITY", '
            'AVG("RESPONSE_TIME_HOURS")::float, '
            'AVG("CUSTOMER_SATISFACTION")::float, '
            'COUNT(*) '
            'FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" '
            'GROUP BY "PRIORITY"'
        )
        out = {}
        for prio, avg_resp, avg_csat, cnt in cur.fetchall():
            out[(prio or "").strip().lower()] = {
                "avg_resp": float(avg_resp) if avg_resp is not None else None,
                "avg_csat": float(avg_csat) if avg_csat is not None else None,
                "count": int(cnt),
            }
        cur.close()
        conn.close()
        return out or None
    except Exception as e:
        print(f"  [warn] could not load ticket truth: {e}")
        return None


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def check_excel(agent_workspace, survey_truth, ticket_truth):
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Support_Satisfaction_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}", critical=True)
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e), critical=True)
        return

    # --- Sheet 1: Survey Results ---
    survey_sheet = None
    for name in wb.sheetnames:
        if "survey" in name.lower() and "summary" not in name.lower():
            survey_sheet = name
            break
    if not survey_sheet:
        record("Survey Results sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Survey Results sheet exists", True)
        ws = wb[survey_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        record("Survey Results has 20 rows", len(data_rows) == 20, f"Found {len(data_rows)}")

    # --- Sheet 2: Survey Summary ---
    summary_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            summary_sheet = name
            break
    if not summary_sheet:
        record("Survey Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Survey Summary: Avg_Overall correct", False, "no Summary sheet", critical=True)
        record("Survey Summary: Lowest=Low / Highest=High", False, "no Summary sheet", critical=True)
    else:
        record("Survey Summary sheet exists", True)
        ws = wb[summary_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        # Index metric rows by lowercased first-column label.
        metric_map = {}
        for row in data_rows:
            if row and row[0] is not None and str(row[0]).strip():
                metric_map[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        def find_metric(*substrs):
            """Return value of first row whose label contains any substr."""
            for label, val in metric_map.items():
                if any(s in label for s in substrs):
                    return label, val
            return None, None

        # Total respondents
        _, total_val = find_metric("total_respondents", "total")
        record("Total respondents = 20", num_close(total_val, 20, tol=0), f"Got {total_val}")

        # CRITICAL: Avg_Overall_Satisfaction within 0.15 of true mean.
        true_overall = survey_truth["avg_overall"] if survey_truth else 3.25
        _, overall_val = find_metric("avg_overall", "overall")
        record(
            "Survey Summary: Avg_Overall_Satisfaction within 0.15 of true mean",
            overall_val is not None and num_close(overall_val, true_overall, tol=0.15),
            f"Got {overall_val}, expected ~{true_overall:.3f}",
            critical=True,
        )

        # Avg response/resolution/professionalism (non-critical, structural-ish).
        _, art = find_metric("avg_response_time", "response_time")
        if art is not None and survey_truth:
            record("Avg_Response_Time_Rating correct",
                   num_close(art, survey_truth["avg_response_time"], tol=0.15),
                   f"Got {art}")
        _, arq = find_metric("avg_resolution", "resolution")
        if arq is not None and survey_truth:
            record("Avg_Resolution_Quality correct",
                   num_close(arq, survey_truth["avg_resolution"], tol=0.15),
                   f"Got {arq}")

        # CRITICAL: Lowest=Low and Highest=High. Rows MUST exist.
        low_label, low_val = find_metric("lowest")
        high_label, high_val = find_metric("highest")
        # Accept either an English priority word or its RU translation written
        # by the agent. Truth from survey (Low lowest, High highest).
        def is_priority(val, en, ru_words):
            if val is None:
                return False
            s = str(val).strip().lower()
            return en in s or any(w in s for w in ru_words)

        lowest_ok = (low_label is not None and
                     is_priority(low_val, "low", ["низк"]))
        highest_ok = (high_label is not None and
                      is_priority(high_val, "high", ["высок"]))
        record(
            "Survey Summary: Lowest_Rated_Priority=Low AND Highest_Rated_Priority=High (rows present)",
            lowest_ok and highest_ok,
            f"lowest_label={low_label}, lowest={low_val}; highest_label={high_label}, highest={high_val}",
            critical=True,
        )

    # --- Sheet 3: Ticket System Comparison ---
    comp_sheet = None
    for name in wb.sheetnames:
        if "ticket" in name.lower() or "comparison" in name.lower():
            comp_sheet = name
            break
    if not comp_sheet:
        record("Ticket System Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Ticket Comparison: High avg response matches DB", False,
               "no Comparison sheet", critical=True)
    else:
        record("Ticket System Comparison sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        # CRITICAL: exactly 3 priority rows.
        record("Ticket Comparison has exactly 3 priority rows",
               len(data_rows) == 3, f"Found {len(data_rows)}", critical=True)

        # Build per-priority row map by first column.
        prio_rows = {}
        for row in data_rows:
            if not row or row[0] is None:
                continue
            key = str(row[0]).strip().lower()
            for canon, words in (("high", ["высок"]),
                                 ("medium", ["средн"]),
                                 ("low", ["низк"])):
                if canon in key or any(w in key for w in words):
                    prio_rows[canon] = row
                    break

        # CRITICAL: High-priority ticket avg response hours matches live DB value.
        if ticket_truth and "high" in ticket_truth and ticket_truth["high"]["avg_resp"] is not None:
            true_high_resp = ticket_truth["high"]["avg_resp"]
            high_row = prio_rows.get("high")
            found = False
            if high_row:
                for cell in high_row[1:]:
                    if num_close(cell, true_high_resp, tol=0.5):
                        found = True
                        break
            record(
                "Ticket Comparison: High Ticket_Avg_Response_Hours matches DB (live, tol 0.5)",
                found,
                f"true~{true_high_resp:.3f}; row={str(high_row)[:200]}",
                critical=True,
            )
        else:
            record("Ticket Comparison: High Ticket_Avg_Response_Hours matches DB (live, tol 0.5)",
                   False, "could not derive truth from DB", critical=True)

    # --- Sheet 4: Improvement Areas ---
    imp_sheet = None
    for name in wb.sheetnames:
        if "improvement" in name.lower():
            imp_sheet = name
            break
    if not imp_sheet:
        record("Improvement Areas sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Improvement Areas: correct metrics + gaps", False,
               "no Improvement sheet", critical=True)
    else:
        record("Improvement Areas sheet exists", True)
        ws = wb[imp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        record("Improvement Areas has >= 2 rows", len(data_rows) >= 2, f"Found {len(data_rows)}")

        # Build area -> (current, gap). Columns: Area, Current_Score, Target_Score, Gap.
        def area_match(label, en_words, ru_words):
            s = (label or "").lower()
            return any(w in s for w in en_words) or any(w in s for w in ru_words)

        listed = {"response": None, "resolution": None, "professionalism": None}
        for row in data_rows:
            if not row or row[0] is None:
                continue
            label = str(row[0])
            current = row[1] if len(row) > 1 else None
            gap = row[3] if len(row) > 3 else None
            if area_match(label, ["response"], ["время", "отклик"]):
                listed["response"] = (current, gap)
            elif area_match(label, ["resolution"], ["качеств", "решени"]):
                listed["resolution"] = (current, gap)
            elif area_match(label, ["professionalism", "agent"], ["профессионал", "агент"]):
                listed["professionalism"] = (current, gap)

        # CRITICAL: Response Time and Resolution Quality present (both < 4.0),
        # Professionalism (avg 4.0) NOT present, and Gap = 4.5 - Current to 0.1.
        true_rt = survey_truth["avg_response_time"] if survey_truth else 3.1
        true_rq = survey_truth["avg_resolution"] if survey_truth else 3.55

        def gap_ok(entry, true_current):
            if entry is None:
                return False
            current, gap = entry
            cur_f = to_float(current)
            gap_f = to_float(gap)
            if cur_f is None or gap_f is None:
                return False
            # Current should equal the survey average; gap = 4.5 - current.
            return (num_close(cur_f, true_current, tol=0.15) and
                    num_close(gap_f, 4.5 - cur_f, tol=0.1))

        rt_present_correct = gap_ok(listed["response"], true_rt)
        rq_present_correct = gap_ok(listed["resolution"], true_rq)
        prof_absent = listed["professionalism"] is None

        record(
            "Improvement Areas: Response Time + Resolution Quality listed with correct Gap, "
            "Professionalism (>=4.0) excluded",
            rt_present_correct and rq_present_correct and prof_absent,
            f"response={listed['response']}, resolution={listed['resolution']}, "
            f"professionalism={listed['professionalism']}",
            critical=True,
        )

    wb.close()


# ---------------------------------------------------------------------------
# Forms (RU forms-mcp, schema gform.*)
# ---------------------------------------------------------------------------
def _config_options(config):
    """Extract option value strings from a question config (RU forms-mcp)."""
    if not config:
        return []
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except Exception:
            return []
    opts = config.get("options") if isinstance(config, dict) else None
    out = []
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                out.append(str(o.get("value", o.get("label", ""))))
            else:
                out.append(str(o))
    return out


def check_gform():
    print("\n=== Checking Feedback Form (gform.*) ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title, description FROM gform.forms")
        forms = cur.fetchall()
    except Exception as e:
        record("Feedback form exists", False, str(e), critical=True)
        return

    found_form = False
    form_id = None
    for fid, title, desc in forms:
        title_lower = (title or "").lower()
        # Match support/feedback/customer/satisfaction (EN) or RU equivalents.
        if any(w in title_lower for w in
               ["support", "feedback", "customer", "satisfaction",
                "поддерж", "обратн", "клиент", "удовлетвор"]):
            if "employee" not in title_lower and "сотрудник" not in title_lower:
                found_form = True
                form_id = fid
                break

    record("Customer feedback form exists (not the noise 'employee' form)",
           found_form,
           f"Found forms: {[(t, (d or '')[:50]) for _, t, d in forms]}",
           critical=True)

    if not form_id:
        cur.close()
        conn.close()
        return

    cur.execute(
        "SELECT title, question_type, config FROM gform.questions WHERE form_id = %s",
        (form_id,),
    )
    questions = cur.fetchall()
    cur.close()
    conn.close()

    q_count = len(questions)
    record("Form has >= 4 questions", q_count >= 4, f"Found {q_count}")

    q_titles = " ".join((t or "").lower() for t, _, _ in questions)
    has_satisfaction = any(w in q_titles for w in
                           ["satisfaction", "overall", "rating",
                            "удовлетвор", "оценк", "общ"])
    record("Has satisfaction/rating question", has_satisfaction, f"Q titles: {q_titles[:200]}")

    # CRITICAL: a priority multiple-choice question with High/Medium/Low options
    # and an open-ended comments/paragraph question.
    has_priority_choice = False
    priority_detail = ""
    for title, qtype, config in questions:
        tl = (title or "").lower()
        qt = (qtype or "").lower()
        if ("priorit" in tl or "приоритет" in tl):
            opts = " ".join(_config_options(config)).lower()
            has_hml = (("high" in opts and "medium" in opts and "low" in opts) or
                       ("высок" in opts and "средн" in opts and "низк" in opts))
            is_choice = ("choice" in qt or "radio" in qt or "select" in qt or
                         "high" in opts)
            if has_hml and is_choice:
                has_priority_choice = True
                priority_detail = f"type={qtype}, opts={opts[:120]}"
                break
            priority_detail = f"type={qtype}, opts={opts[:120]}"

    has_comment = any(
        ("text" in (qt or "").lower() or "paragraph" in (qt or "").lower())
        for _, qt, _ in questions
    )
    has_comment = has_comment or any(
        ("comment" in (t or "").lower() or "feedback" in (t or "").lower() or
         "коммент" in (t or "").lower() or "отзыв" in (t or "").lower())
        for t, _, _ in questions
    )

    record(
        "Form has a priority (High/Medium/Low) multiple-choice question AND an open comments question",
        has_priority_choice and has_comment,
        f"priority_choice={has_priority_choice} ({priority_detail}); comment={has_comment}",
        critical=True,
    )


# ---------------------------------------------------------------------------
# Calendar
# ---------------------------------------------------------------------------
def check_calendar():
    print("\n=== Checking Google Calendar ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e), critical=True)
        return

    # Collect the quarterly review events.
    review_events = []
    for summary, description, start_dt, end_dt in events:
        sl = (summary or "").lower()
        if ("review" in sl or "обзор" in sl or "support satisfaction" in sl or
                "поддерж" in sl) and any(q in sl for q in ["q1", "q2", "q3", "q4"]):
            review_events.append((summary, start_dt, end_dt))

    record("4 quarterly review events found",
           len(review_events) == 4,
           f"Found {len(review_events)}: {[e[0] for e in review_events]}")

    # CRITICAL: correct dates (Mar/Jun/Sep/Dec 15, 2026) and 10:00-11:30 window.
    expected_dates = {"2026-03-15", "2026-06-15", "2026-09-15", "2026-12-15"}
    found_dates = set()
    times_ok = True
    time_detail = []
    for summary, start_dt, end_dt in review_events:
        if start_dt is None:
            times_ok = False
            time_detail.append(f"{summary}: no start")
            continue
        found_dates.add(start_dt.strftime("%Y-%m-%d"))
        # 10:00 local in America/New_York. Stored value may be local or UTC.
        # Accept the start hour as 10 (local) or 14/15 (UTC, EST/EDT).
        h = start_dt.hour
        start_hour_ok = h in (10, 14, 15)
        dur_ok = True
        if end_dt is not None:
            dur = (end_dt - start_dt).total_seconds() / 3600.0
            dur_ok = abs(dur - 1.5) <= 0.1
        if not (start_hour_ok and dur_ok):
            times_ok = False
            time_detail.append(f"{summary}: hour={h}, end={end_dt}")

    dates_ok = expected_dates.issubset(found_dates)
    record(
        "Calendar: 4 events on 2026-03-15/06-15/09-15/12-15 with 10:00-11:30 window",
        len(review_events) == 4 and dates_ok and times_ok,
        f"dates={sorted(found_dates)}; time_issues={time_detail}",
        critical=True,
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    survey_truth = load_survey_truth()
    if survey_truth is None:
        print("  [warn] survey truth unavailable; using fallback constants")
    ticket_truth = load_ticket_truth()

    check_excel(args.agent_workspace, survey_truth, ticket_truth)
    check_gform()
    check_calendar()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ===")

    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"Overall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILS,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILS}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
