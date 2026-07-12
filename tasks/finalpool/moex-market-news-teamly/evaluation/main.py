"""Evaluation for moex-market-news-teamly.

Critical checks (CRITICAL_CHECKS): any of them failing => task FAIL even if the
overall accuracy is high. They verify the SUBSTANCE of the digest:
  - Excel News Articles row count == COUNT(*) FROM moex.news exactly.
  - Excel Summary Unique_Tickers == COUNT(DISTINCT symbol) FROM moex.news.
  - Excel Summary Top_Publisher == the actual top publisher from moex.news.
  - Teamly page exists (RU/EN title) AND body mentions every seeded ticker.
  - Email to portfolio-team@investco.com, subject 'Weekly Market News Digest',
    body mentions the correct total article count and the top publisher.

Expected values are read live from moex.news (read-only), never hardcoded.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

CRITICAL_CHECKS = {
    "News Articles row count == moex.news (exact)",
    "Summary Unique_Tickers == moex.news (exact)",
    "Summary Top_Publisher == actual top publisher",
    "Teamly digest page mentions every ticker",
    "Email body mentions total count and top publisher",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

TICKERS = ["SBER.ME", "GAZP.ME", "LKOH.ME", "TCSG.ME", "MGNT.ME", "MTSS.ME"]


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=2):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected values from read-only moex.news."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM moex.news")
    total_articles = cur.fetchone()[0]
    cur.execute("SELECT COUNT(DISTINCT symbol) FROM moex.news")
    unique_tickers = cur.fetchone()[0]
    cur.execute("""
        SELECT data->'content'->'provider'->>'displayName' as publisher, COUNT(*) as cnt
        FROM moex.news GROUP BY 1 ORDER BY 2 DESC, 1 ASC LIMIT 1
    """)
    row = cur.fetchone()
    top_publisher = row[0] if row else ""
    cur.close()
    conn.close()
    return total_articles, unique_tickers, top_publisher


def check_excel(agent_workspace, expected):
    """Check Market_News_Digest.xlsx."""
    print("\n=== Checking Market_News_Digest.xlsx ===")
    total_articles, unique_tickers, top_publisher = expected

    agent_file = os.path.join(agent_workspace, "Market_News_Digest.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        # cascade-fail criticals tied to excel
        record("News Articles row count == moex.news (exact)", False, "no file")
        record("Summary Unique_Tickers == moex.news (exact)", False, "no file")
        record("Summary Top_Publisher == actual top publisher", False, "no file")
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        record("News Articles row count == moex.news (exact)", False, "unreadable")
        record("Summary Unique_Tickers == moex.news (exact)", False, "unreadable")
        record("Summary Top_Publisher == actual top publisher", False, "unreadable")
        return
    record("Excel readable", True)

    # News Articles sheet
    news_sheet = None
    for name in wb.sheetnames:
        if "news" in name.lower() and "article" in name.lower():
            news_sheet = wb[name]
            break
    if news_sheet is None:
        record("Sheet 'News Articles' exists", False, f"Sheets: {wb.sheetnames}")
        record("News Articles row count == moex.news (exact)", False, "no sheet")
    else:
        record("Sheet 'News Articles' exists", True)
        rows = [r for r in news_sheet.iter_rows(min_row=2, values_only=True)
                if r and any(c is not None and str(c).strip() for c in r)]
        # CRITICAL: exact row count proves full RU news pull.
        record("News Articles row count == moex.news (exact)",
               len(rows) == total_articles,
               f"Expected {total_articles}, got {len(rows)}")

    # Summary sheet
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        record("Summary Unique_Tickers == moex.news (exact)", False, "no sheet")
        record("Summary Top_Publisher == actual top publisher", False, "no sheet")
        return
    record("Sheet 'Summary' exists", True)

    summary = {}
    for row in sum_sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            summary[str(row[0]).strip().lower()] = row[1]

    total_val = unique_val = top_val = None
    for key, val in summary.items():
        if "total" in key and "article" in key:
            total_val = val
        elif "unique" in key and "ticker" in key:
            unique_val = val
        elif "top" in key and "publisher" in key:
            top_val = val

    record("Summary Total_Articles", num_close(total_val, total_articles, 2),
           f"Expected {total_articles}, got {total_val}")
    # CRITICAL: exact distinct ticker count.
    record("Summary Unique_Tickers == moex.news (exact)",
           str(unique_val).strip() == str(unique_tickers),
           f"Expected {unique_tickers}, got {unique_val}")
    # CRITICAL: top publisher exact (case-insensitive, substring tolerant).
    tp_ok = bool(top_publisher) and top_publisher.lower() in str(top_val or "").lower()
    record("Summary Top_Publisher == actual top publisher", tp_ok,
           f"Expected '{top_publisher}', got '{top_val}'")


def check_teamly():
    """Teamly digest page exists with every seeded ticker in the body."""
    print("\n=== Checking Teamly Page ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, title, body FROM teamly.pages
            WHERE title ILIKE '%market%' OR title ILIKE '%рыночн%'
               OR title ILIKE '%рынк%' OR title ILIKE '%аналитик%'
               OR title ILIKE '%дайджест%' OR title ILIKE '%news%'
            ORDER BY id DESC
        """)
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Teamly digest page found", False, f"DB error: {e}")
        record("Teamly digest page mentions every ticker", False, str(e))
        return

    record("Teamly digest page found", len(pages) >= 1,
           f"Found {len(pages)} pages: {[p[1] for p in pages]}")

    if not pages:
        record("Teamly digest page mentions every ticker", False, "no page")
        return

    # Pick the page covering the most tickers.
    best_body = ""
    best_cov = -1
    for _id, _title, body in pages:
        bl = (body or "").lower()
        cov = sum(1 for t in TICKERS if t.lower() in bl)
        if cov > best_cov:
            best_cov = cov
            best_body = bl

    missing = [t for t in TICKERS if t.lower() not in best_body]
    # CRITICAL: per-ticker digest built.
    record("Teamly digest page mentions every ticker", not missing,
           f"Missing tickers: {missing}; body head: {best_body[:200]}")


def check_email(expected):
    """Email digest with correct subject, recipient, and body content."""
    print("\n=== Checking Email ===")
    total_articles, _unique, top_publisher = expected
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE subject ILIKE '%news%' OR subject ILIKE '%digest%'
               OR subject ILIKE '%дайджест%' OR subject ILIKE '%новост%'
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email digest exists", False, f"DB error: {e}")
        record("Email body mentions total count and top publisher", False, str(e))
        return

    record("Email digest exists", len(rows) > 0, "No matching email found")

    target = None
    for subj, to_addr, body in rows:
        if "portfolio-team@investco.com" in str(to_addr).lower():
            target = (subj, to_addr, body)
            break
    record("Email to portfolio-team@investco.com", target is not None,
           f"To addresses: {[str(r[1]) for r in rows]}")

    if target is None:
        record("Email body mentions total count and top publisher", False,
               "no email to portfolio-team@investco.com")
        return

    subj, _to, body = target
    record("Email subject is 'Weekly Market News Digest'",
           "weekly market news digest" in str(subj).lower(),
           f"subject: {subj}")

    bl = str(body or "").lower()
    total_ok = str(total_articles) in bl
    top_ok = bool(top_publisher) and top_publisher.lower() in bl
    # CRITICAL: body summary content verified.
    record("Email body mentions total count and top publisher",
           total_ok and top_ok,
           f"total({total_articles})={total_ok}, top('{top_publisher}')={top_ok}; "
           f"body head: {bl[:200]}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = get_expected_data()
    print(f"Expected: total={expected[0]}, unique_tickers={expected[1]}, "
          f"top_publisher='{expected[2]}'")

    check_excel(args.agent_workspace, expected)
    check_teamly()
    check_email(expected)

    total = PASS_COUNT + FAIL_COUNT
    pct = 100.0 * PASS_COUNT / total if total else 0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({pct:.1f}%)")

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAIL: {CRITICAL_FAILED}")
        print("  Overall: FAIL")
        sys.exit(1)
    if pct < 70.0:
        print("  Overall: FAIL (accuracy < 70%)")
        sys.exit(1)
    print("  Overall: PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
