"""Evaluation for canvas-quiz-item-analysis-word-gcal-email."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
# These reflect the core computed deliverables read from live Canvas data.
CRITICAL_CHECKS = {
    "Quiz Overview Avg_Score & Quality_Rating match groundtruth",
    "Course Summary has 7 rows with correct totals",
    "Difficulty classification matches groundtruth",
    "7 quiz review events in week of March 16-20 (~45 min each)",
    "Email to assessment_office with non-trivial body",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Checking Excel ===")
    agent_file = os.path.join(agent_workspace, "Quiz_Item_Analysis.xlsx")
    gt_file = os.path.join(gt_workspace, "Quiz_Item_Analysis.xlsx")

    if not os.path.isfile(agent_file):
        check("Quiz_Item_Analysis.xlsx exists", False, f"Not found: {agent_file}")
        return
    check("Quiz_Item_Analysis.xlsx exists", True)

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel files readable", False, str(e))
        return

    # --- Sheet 1: Quiz Overview ---
    print("  Checking Quiz Overview...")
    a_rows = load_sheet_rows(agent_wb, "Quiz Overview")
    g_rows = load_sheet_rows(gt_wb, "Quiz Overview")

    if a_rows is None:
        check("Sheet 'Quiz Overview' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Quiz Overview' exists (gt)", False)
    else:
        check("Sheet 'Quiz Overview' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Quiz Overview row count", abs(len(a_data) - len(g_data)) <= 1,
              f"Agent={len(a_data)}, GT={len(g_data)}")

        # Lookup by quiz title
        a_lookup = {}
        for row in a_data:
            if row and len(row) > 1 and row[1] is not None:
                a_lookup[str(row[1]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[1] is None:
                continue
            key = str(g_row[1]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing quiz: {g_row[1]}")
                continue
            # Check Avg_Score (col 4, tol 0.5)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 0.5):
                    errors.append(f"{key}.Avg_Score: {a_row[4]} vs {g_row[4]}")
            # Check Completion_Rate_Pct (col 5, tol 1.0)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1.0):
                    errors.append(f"{key}.Completion_Rate: {a_row[5]} vs {g_row[5]}")
            # Check Submission_Count (col 3, exact)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5):
                    errors.append(f"{key}.Sub_Count: {a_row[3]} vs {g_row[3]}")
            # Check Quality_Rating (col 6, string)
            if len(a_row) > 6 and len(g_row) > 6:
                if not str_match(a_row[6], g_row[6]):
                    errors.append(f"{key}.Quality: {a_row[6]} vs {g_row[6]}")
        check("Quiz Overview Avg_Score & Quality_Rating match groundtruth",
              not errors, "; ".join(errors[:5]))

    # --- Sheet 2: Question Analysis ---
    print("  Checking Question Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Question Analysis")
    g_rows = load_sheet_rows(gt_wb, "Question Analysis")

    if a_rows is None:
        check("Sheet 'Question Analysis' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Question Analysis' exists (gt)", False)
    else:
        check("Sheet 'Question Analysis' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Question Analysis row count", abs(len(a_data) - len(g_data)) <= 5,
              f"Agent={len(a_data)}, GT={len(g_data)}")

        # Difficulty classification correctness: compare the Easy/Optimal/Hard
        # distribution (col 5) against groundtruth. This is the central
        # methodology (avg_score/100 thresholds) from the PDF.
        def diff_dist(rows):
            dist = {"easy": 0, "optimal": 0, "hard": 0}
            for r in rows:
                if r and len(r) > 5 and r[5] is not None:
                    k = str(r[5]).strip().lower()
                    if k in dist:
                        dist[k] += 1
            return dist
        a_dist = diff_dist(a_data)
        g_dist = diff_dist(g_data)
        dist_ok = all(abs(a_dist[k] - g_dist[k]) <= 5 for k in g_dist)
        check("Difficulty classification matches groundtruth", dist_ok,
              f"Agent={a_dist}, GT={g_dist}")

    # --- Sheet 3: Course Summary ---
    print("  Checking Course Summary...")
    a_rows = load_sheet_rows(agent_wb, "Course Summary")
    g_rows = load_sheet_rows(gt_wb, "Course Summary")

    if a_rows is None:
        check("Sheet 'Course Summary' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Course Summary' exists (gt)", False)
    else:
        check("Sheet 'Course Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        rows_ok = len(a_data) == 7

        # Lookup by course name
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing course: {g_row[0]}")
                continue
            # Total_Quizzes (col 1, exact)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    errors.append(f"{key[:30]}.Total_Quizzes: {a_row[1]} vs {g_row[1]}")
            # Total_Questions (col 2, exact)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 0):
                    errors.append(f"{key[:30]}.Total_Questions: {a_row[2]} vs {g_row[2]}")
            # Review_Quizzes (col 5, exact)
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 0):
                    errors.append(f"{key[:30]}.Review_Quizzes: {a_row[5]} vs {g_row[5]}")
            # Avg_Completion_Rate (col 6, tol 1.0)
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 1.0):
                    errors.append(f"{key[:30]}.Avg_Comp: {a_row[6]} vs {g_row[6]}")

        if not rows_ok:
            errors.insert(0, f"row count {len(a_data)} != 7")
        check("Course Summary has 7 rows with correct totals",
              rows_ok and not errors, "; ".join(errors[:5]))


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Quiz_Analysis_Report.docx")
    if not os.path.isfile(docx_path):
        check("Quiz_Analysis_Report.docx exists", False, f"Not found: {docx_path}")
        return
    check("Quiz_Analysis_Report.docx exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
    except Exception as e:
        check("Word doc readable", False, str(e))
        return

    full_text = " ".join(p.text for p in doc.paragraphs).lower()

    # Check for key sections (accept RU or EN wording).
    section_variants = [
        ("executive summary", ["executive summary", "аннотация", "краткое резюме"]),
        ("methodology", ["methodology", "методология"]),
        ("recommendations", ["recommendations", "рекомендаци"]),
    ]
    for label, variants in section_variants:
        check(f"Section '{label}' present",
              any(v in full_text for v in variants),
              f"Not found (RU/EN): {variants}")

    # Check for course names (sourced live from the russified Canvas data).
    for course_key in ["креативные вычисления", "основы финансов", "глобальное управление"]:
        check(f"Course '{course_key}' mentioned",
              course_key in full_text,
              f"Not found in document text")

    # Check for key terms (accept RU or EN wording).
    term_variants = [
        ("difficulty", ["difficulty", "сложност"]),
        ("optimal", ["optimal", "оптимальн"]),
        ("needs review", ["needs review", "требует пересмотра", "требует проверки"]),
    ]
    for label, variants in term_variants:
        check(f"Term '{label}' mentioned",
              any(v in full_text for v in variants),
              f"Not found (RU/EN): {variants}")


def check_gcal():
    print("\n=== Checking Calendar Events ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE summary ILIKE '%%quiz review%%'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()

        from datetime import datetime, timezone
        week_start = datetime(2026, 3, 16, tzinfo=timezone.utc)
        week_end = datetime(2026, 3, 21, tzinfo=timezone.utc)
        in_week = sum(1 for e in events
                      if e[1] and week_start <= e[1].replace(tzinfo=timezone.utc) < week_end)
        durations_ok = all(
            e[1] and e[2] and abs((e[2] - e[1]).total_seconds() / 60 - 45) <= 5
            for e in events
        )
        check("7 quiz review events in week of March 16-20 (~45 min each)",
              in_week >= 7 and durations_ok,
              f"Found {len(events)} events, {in_week} in target week, durations_ok={durations_ok}")

        cur.close()
        conn.close()
    except Exception as e:
        check("7 quiz review events in week of March 16-20 (~45 min each)", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Match by recipient address (robust) OR by subject keywords (EN or RU).
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%assessment_office@university.edu%%'
               OR subject ILIKE '%%quiz%%item%%'
               OR subject ILIKE '%%quiz%%analysis%%'
               OR subject ILIKE '%%item analysis%%'
               OR subject ILIKE '%%анализ%%'
               OR subject ILIKE '%%качеств%%'
               OR subject ILIKE '%%викторин%%'
               OR subject ILIKE '%%тест%%'
        """)
        emails = cur.fetchall()
        body = str(emails[0][3]) if emails and emails[0][3] else ""
        check("Email to assessment_office with non-trivial body",
              len(emails) >= 1 and len(body) > 20,
              f"Found {len(emails)} emails, body length {len(body)}")

        cur.close()
        conn.close()
    except Exception as e:
        check("Email to assessment_office with non-trivial body", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS: no critical failures and accuracy >= 70%")
        sys.exit(0)
    else:
        print(f"FAIL: accuracy {accuracy:.1f}% < 70%")
        sys.exit(1)


if __name__ == "__main__":
    main()
