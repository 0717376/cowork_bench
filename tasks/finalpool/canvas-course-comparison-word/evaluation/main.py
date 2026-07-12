"""Evaluation for canvas-course-comparison-word."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        tag = " (CRITICAL)" if critical else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if critical:
            CRITICAL_FAILURES.append(name)


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected course comparison data from DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        WITH course_data AS (
            SELECT
                SPLIT_PART(c.course_code, '-', 1) as prefix,
                c.course_code,
                c.name,
                c.total_students,
                COUNT(DISTINCT a.id) as assignment_count,
                ROUND(AVG(s.score)::numeric, 2) as avg_grade
            FROM canvas.courses c
            LEFT JOIN canvas.assignments a ON a.course_id = c.id
            LEFT JOIN canvas.submissions s ON s.assignment_id = a.id AND s.score IS NOT NULL
            WHERE c.course_code LIKE '%%2013J' OR c.course_code LIKE '%%2014J'
            GROUP BY c.id, c.course_code, c.name, c.total_students
        )
        SELECT
            f13.prefix,
            f13.name as name_2013,
            f13.total_students as enroll_2013,
            f14.total_students as enroll_2014,
            f13.assignment_count as assign_2013,
            f14.assignment_count as assign_2014,
            f13.avg_grade as grade_2013,
            f14.avg_grade as grade_2014
        FROM course_data f13
        JOIN course_data f14 ON f13.prefix = f14.prefix
        WHERE f13.course_code LIKE '%%2013J' AND f14.course_code LIKE '%%2014J'
        ORDER BY f13.prefix
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def match_agent_row(prefix, course_name_2013, agent_lookup):
    """Find the agent row corresponding to an expected course pair."""
    for key, r in agent_lookup.items():
        if prefix.lower() in key or course_name_2013.lower().split("(")[0].strip().lower() in key:
            return r
    return None


def check_excel(agent_workspace, groundtruth_workspace, expected):
    """Check Year_Over_Year_Comparison.xlsx. Returns (all_ok, agent_lookup)."""
    print("\n=== Checking Year_Over_Year_Comparison.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Year_Over_Year_Comparison.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}", critical=True)
        return False, {}
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e), critical=True)
        return False, {}

    all_ok = True

    # Check Course Comparison sheet
    comp_sheet = None
    for name in wb.sheetnames:
        if "comparison" in name.lower() or "course" in name.lower():
            comp_sheet = wb[name]
            break
    agent_lookup = {}
    if comp_sheet is None:
        record("Sheet 'Course Comparison' exists", False, f"Sheets: {wb.sheetnames}", critical=True)
        all_ok = False
    else:
        record("Sheet 'Course Comparison' exists", True)
        rows = list(comp_sheet.iter_rows(min_row=2, values_only=True))

        # CRITICAL: exact row count == number of expected 2013J/2014J pairs
        record("Course Comparison row count matches expected pairs",
               len(rows) == len(expected),
               f"Expected {len(expected)}, got {len(rows)}", critical=True)
        if len(rows) != len(expected):
            all_ok = False

        for r in rows:
            if r and r[0]:
                key = str(r[0]).strip().lower()
                agent_lookup[key] = r

        for exp_row in expected:
            prefix = exp_row[0]
            course_name_2013 = exp_row[1]
            matched = match_agent_row(prefix, course_name_2013, agent_lookup)

            if matched is None:
                record(f"Course '{prefix}' found", False, "Missing", critical=True)
                all_ok = False
                continue

            # Check enrollment values
            ok_e13 = num_close(matched[1], exp_row[2], 10)
            record(f"'{prefix}' Fall_2013_Enrollment", ok_e13,
                   f"Expected {exp_row[2]}, got {matched[1]}")
            if not ok_e13:
                all_ok = False

            ok_e14 = num_close(matched[2], exp_row[3], 10)
            record(f"'{prefix}' Fall_2014_Enrollment", ok_e14,
                   f"Expected {exp_row[3]}, got {matched[2]}")
            if not ok_e14:
                all_ok = False

            # CRITICAL: avg grades within tightened tolerance (proves real join+avg)
            ok_g13 = num_close(matched[6], exp_row[6], 0.5)
            record(f"'{prefix}' Fall_2013_Avg_Grade", ok_g13,
                   f"Expected {exp_row[6]}, got {matched[6]}", critical=True)
            if not ok_g13:
                all_ok = False

            ok_g14 = num_close(matched[7], exp_row[7], 0.5)
            record(f"'{prefix}' Fall_2014_Avg_Grade", ok_g14,
                   f"Expected {exp_row[7]}, got {matched[7]}", critical=True)
            if not ok_g14:
                all_ok = False

    # Check Summary sheet (values, not just existence)
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}", critical=True)
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        # Build metric->value lookup
        summary = {}
        for row in sum_sheet.iter_rows(min_row=1, values_only=True):
            if row and row[0] is not None:
                summary[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        exp_count = len(expected)
        exp_enroll_changes = [float(e[3]) - float(e[2]) for e in expected]
        exp_grade_changes = [float(e[7]) - float(e[6]) for e in expected
                             if e[6] is not None and e[7] is not None]
        exp_avg_enroll = round(sum(exp_enroll_changes) / len(exp_enroll_changes), 1) if exp_enroll_changes else 0.0
        exp_avg_grade = round(sum(exp_grade_changes) / len(exp_grade_changes), 2) if exp_grade_changes else 0.0

        def find_metric(*subs):
            for k, v in summary.items():
                if any(sub in k for sub in subs):
                    return v
            return None

        v_count = find_metric("courses_compared", "courses compared")
        ok_count = num_close(v_count, exp_count, 0)
        record("Summary Courses_Compared matches pair count", ok_count,
               f"Expected {exp_count}, got {v_count}", critical=True)
        if not ok_count:
            all_ok = False

        v_avg_e = find_metric("avg_enrollment_change", "avg enrollment")
        ok_avg_e = num_close(v_avg_e, exp_avg_enroll, 1.0)
        record("Summary Avg_Enrollment_Change matches recomputed", ok_avg_e,
               f"Expected {exp_avg_enroll}, got {v_avg_e}", critical=True)
        if not ok_avg_e:
            all_ok = False

        v_avg_g = find_metric("avg_grade_change", "avg grade")
        ok_avg_g = num_close(v_avg_g, exp_avg_grade, 0.5)
        record("Summary Avg_Grade_Change matches recomputed", ok_avg_g,
               f"Expected {exp_avg_grade}, got {v_avg_g}", critical=True)
        if not ok_avg_g:
            all_ok = False

    return all_ok, agent_lookup


def check_word(agent_workspace, expected):
    """Check Academic_Year_Comparison.docx. Returns real pass state."""
    print("\n=== Checking Academic_Year_Comparison.docx ===")
    from docx import Document

    docx_file = os.path.join(agent_workspace, "Academic_Year_Comparison.docx")
    if not os.path.isfile(docx_file):
        record("Word file exists", False, f"Not found: {docx_file}", critical=True)
        return False
    record("Word file exists", True)

    try:
        doc = Document(docx_file)
    except Exception as e:
        record("Word readable", False, str(e), critical=True)
        return False

    ok = True
    # ORIGINAL (lowercased) text — RU keywords go here, NOT normalized.
    all_text = " ".join(p.text for p in doc.paragraphs)
    low = all_text.lower()

    has_2013 = "2013" in low
    has_2014 = "2014" in low
    record("Word mentions '2013'", has_2013, "No mention of '2013'", critical=True)
    record("Word mentions '2014'", has_2014, "No mention of '2014'", critical=True)
    if not (has_2013 and has_2014):
        ok = False

    # Accept English OR Russian narrative keywords.
    kw_en = ["performance", "comparison", "review", "enrollment", "grade"]
    kw_ru = ["год", "сравнен", "успеваемост", "набор", "оценк", "обзор"]
    has_kw = any(k in low for k in kw_en + kw_ru)
    record("Word has narrative keywords (EN or RU)", has_kw,
           "No relevant keywords found", critical=True)
    if not has_kw:
        ok = False

    # Title present (EN identifier preserved).
    has_title = "academic performance review" in low or (
        "fall 2013" in low and "fall 2014" in low)
    record("Word title present", has_title,
           "Title 'Fall 2013 vs Fall 2014 Academic Performance Review' missing", critical=True)
    if not has_title:
        ok = False

    # At least one narrative paragraph per course pair naming the course.
    n_pairs = len(expected)
    named = 0
    for exp_row in expected:
        prefix = exp_row[0].lower()
        cname = exp_row[1].lower().split("(")[0].strip()
        if prefix in low or (cname and cname in low):
            named += 1
    record(f"Word names course pairs ({named}/{n_pairs})",
           named >= n_pairs,
           f"Only {named} of {n_pairs} course pairs named in document", critical=True)
    if named < n_pairs:
        ok = False

    return ok


def check_gsheet(expected, agent_lookup):
    """Check Google Sheet 'Academic Year Comparison' + 'Comparison Data' rows. Gated."""
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    ok = True

    cur.execute("SELECT id FROM gsheet.spreadsheets WHERE title ILIKE '%academic%' OR title ILIKE '%year%comparison%'")
    rows = cur.fetchall()
    found = len(rows) > 0
    record("GSheet 'Academic Year Comparison' exists", found,
           "No matching spreadsheet found", critical=True)
    if not found:
        cur.close()
        conn.close()
        return False
    ss_id = rows[0][0]

    # Find 'Comparison Data' sheet
    cur.execute("""SELECT id FROM gsheet.sheets
                   WHERE spreadsheet_id = %s
                   AND (title ILIKE '%%comparison%%data%%' OR title ILIKE '%%comparison%%')""", (ss_id,))
    srows = cur.fetchall()
    has_sheet = len(srows) > 0
    record("GSheet 'Comparison Data' sheet exists", has_sheet,
           "No 'Comparison Data' sheet found", critical=True)
    if not has_sheet:
        ok = False
        cur.close()
        conn.close()
        return ok
    sheet_id = srows[0][0]

    # Gather all cell values for this sheet, grouped by row.
    cur.execute("""SELECT row_index, col_index, COALESCE(value, formatted_value, formula)
                   FROM gsheet.cells
                   WHERE spreadsheet_id = %s AND sheet_id = %s""", (ss_id, sheet_id))
    cells = cur.fetchall()
    cur.close()
    conn.close()

    grid = {}
    for ri, ci, val in cells:
        grid.setdefault(ri, {})[ci] = val

    # Build per-row text blobs (skip empty rows).
    row_blobs = []
    for ri in sorted(grid):
        vals = [str(v) for v in grid[ri].values() if v is not None and str(v).strip()]
        if vals:
            row_blobs.append(" | ".join(vals).lower())

    # CRITICAL: data rows present (header + one per course pair).
    # Match each expected course pair to at least one GSheet row.
    matched_pairs = 0
    for exp_row in expected:
        prefix = exp_row[0].lower()
        cname = exp_row[1].lower().split("(")[0].strip()
        for blob in row_blobs:
            if prefix in blob or (cname and cname in blob):
                matched_pairs += 1
                break
    record(f"GSheet 'Comparison Data' has rows per course pair ({matched_pairs}/{len(expected)})",
           matched_pairs >= len(expected),
           f"Only {matched_pairs} of {len(expected)} course pairs present in GSheet", critical=True)
    if matched_pairs < len(expected):
        ok = False

    return ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = get_expected_data()
    print(f"Expected course pairs from canvas DB: {len(expected)}")

    excel_ok, agent_lookup = check_excel(args.agent_workspace, args.groundtruth_workspace, expected)
    word_ok = check_word(args.agent_workspace, expected)
    gsheet_ok = check_gsheet(expected, agent_lookup)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILURES:
        print(f"  CRITICAL FAILURES ({len(CRITICAL_FAILURES)}):")
        for c in CRITICAL_FAILURES:
            print(f"    - {c}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70 and excel_ok and word_ok and gsheet_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
