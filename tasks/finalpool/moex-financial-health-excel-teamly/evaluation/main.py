"""Evaluation for yf-financial-health-excel-notion (russified: moex + teamly).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Expected numeric values are the seeded ground-truth from
db/zzz_moex_after_init.sql (moex.financial_statements, 2025 annual). They are
deterministic seed constants, not volatile market data, so it is correct for the
eval to assert against them.

Ticker set: OZON/PHOR/ROSN/VTBR/YNDX. These are the five symbols that have BOTH
income_stmt AND balance_sheet annual rows seeded, so the balance figures are
reachable via get_financial_statement(ticker, "balance_sheet"). (The earlier set
SBER/GAZP/LKOH/MGNT/MTSS had no balance_sheet rows, making the assets/liabilities/
equity values impossible to obtain.)
"""
import os
import argparse
import sys
import psycopg2


# --- Seeded ground truth (2025 annual, moex.financial_statements) -------------
# Income: Symbol -> (Total_Revenue, Net_Income)
INCOME = {
    "OZON.ME": (812400000000.0, 41200000000.0),
    "PHOR.ME": (101500000000.0, 29400000000.0),
    "ROSN.ME": (320800000000.0, 47300000000.0),
    "VTBR.ME": (351200000000.0, 55900000000.0),
    "YNDX.ME": (471900000000.0, 158600000000.0),
}
# Balance: Symbol -> (Total_Assets, Total_Liabilities, Equity)
BALANCE = {
    "OZON.ME": (905300000000.0, 497915000000.0, 407385000000.0),
    "PHOR.ME": (214600000000.0, 118030000000.0, 96570000000.0),
    "ROSN.ME": (548900000000.0, 301895000000.0, 247005000000.0),
    "VTBR.ME": (1044000000000.0, 574200000000.0, 469800000000.0),
    "YNDX.ME": (651100000000.0, 358105000000.0, 292995000000.0),
}
SYMBOLS = ["OZON.ME", "PHOR.ME", "ROSN.ME", "VTBR.ME", "YNDX.ME"]

# Derived superlatives (semantic, from the seed):
HIGHEST_REVENUE = max(INCOME, key=lambda s: INCOME[s][0])          # OZON.ME
BEST_MARGIN = max(INCOME, key=lambda s: INCOME[s][1] / INCOME[s][0])  # YNDX.ME

# RU/EN names so the teamly dashboard can name the company in either form.
# Aliases cover the ticker, the source longName/shortName (moex.stock_info), and
# the common Russian transliteration.
COMPANY_ALIASES = {
    "OZON.ME": ["ozon", "озон"],
    "PHOR.ME": ["phor", "phosagro", "фосагро"],
    "ROSN.ME": ["rosn", "rosneft", "роснефть"],
    "VTBR.ME": ["vtbr", "vtb", "втб"],
    "YNDX.ME": ["yndx", "yandex", "яндекс"],
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "Income Statement values match seeded MOEX income data (revenue/net income/margin)",
    "Balance Sheet values match seeded MOEX balance data (assets/liabilities/equity)",
    "Teamly 'Financial Health Dashboard' names highest-revenue and best-margin companies",
    "Email to finance_team has subject 'Financial Health Report' and references the report + dashboard",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def rel_close(a, b, rel=0.02):
    """Relative tolerance comparison for large currency figures."""
    try:
        a = float(a)
        b = float(b)
    except (TypeError, ValueError):
        return False
    if b == 0:
        return abs(a) <= 1e-6
    return abs(a - b) / abs(b) <= rel


def get_conn():
    return psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432,
                            dbname=os.environ.get("PGDATABASE", "cowork_gym"),
                            user="eigent", password="camel")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def col_index(header_row, *aliases):
    norm = [str(c).strip().lower() if c is not None else "" for c in header_row]
    for a in aliases:
        a = a.lower()
        for i, h in enumerate(norm):
            if h == a:
                return i
    return None


def check_excel(agent_workspace):
    import openpyxl
    path = os.path.join(agent_workspace, "Financial_Health_Report.xlsx")
    check("Financial_Health_Report.xlsx exists", os.path.exists(path))
    if not os.path.exists(path):
        check("Income Statement values match seeded MOEX income data (revenue/net income/margin)", False, "no excel")
        check("Balance Sheet values match seeded MOEX balance data (assets/liabilities/equity)", False, "no excel")
        return

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Income Statement values match seeded MOEX income data (revenue/net income/margin)", False, str(e))
        check("Balance Sheet values match seeded MOEX balance data (assets/liabilities/equity)", False, str(e))
        return

    # ---- Income Statement ----
    rows = load_sheet_rows(wb, "Income Statement")
    check("Income Statement sheet exists", rows is not None)
    income_ok = True
    income_detail = ""
    if rows is None:
        income_ok = False
        income_detail = "sheet missing"
    else:
        header = rows[0] if rows else []
        sym_i = col_index(header, "symbol")
        rev_i = col_index(header, "total_revenue")
        ni_i = col_index(header, "net_income")
        margin_i = col_index(header, "profit_margin_pct")
        check("Income Statement has required columns",
              None not in (sym_i, rev_i, ni_i, margin_i), f"header: {header}")
        data_rows = [r for r in rows[1:] if r and sym_i is not None and r[sym_i] is not None]
        by_sym = {}
        for r in data_rows:
            by_sym[str(r[sym_i]).strip().upper()] = r
        for sym in SYMBOLS:
            check(f"Income Statement contains {sym}", sym in by_sym)
        if None in (sym_i, rev_i, ni_i, margin_i):
            income_ok = False
            income_detail = "missing columns"
        else:
            for sym in SYMBOLS:
                r = by_sym.get(sym)
                exp_rev, exp_ni = INCOME[sym]
                exp_margin = round(exp_ni / exp_rev * 100, 1)
                if r is None:
                    income_ok = False
                    income_detail = f"{sym} row missing"
                    break
                if not rel_close(r[rev_i], exp_rev):
                    income_ok = False
                    income_detail = f"{sym} revenue {r[rev_i]} != {exp_rev}"
                    break
                if not rel_close(r[ni_i], exp_ni):
                    income_ok = False
                    income_detail = f"{sym} net_income {r[ni_i]} != {exp_ni}"
                    break
                try:
                    if abs(float(r[margin_i]) - exp_margin) > 0.2:
                        income_ok = False
                        income_detail = f"{sym} margin {r[margin_i]} != {exp_margin}"
                        break
                except (TypeError, ValueError):
                    income_ok = False
                    income_detail = f"{sym} margin not numeric: {r[margin_i]}"
                    break
    check("Income Statement values match seeded MOEX income data (revenue/net income/margin)",
          income_ok, income_detail)

    # ---- Balance Sheet ----
    rows2 = load_sheet_rows(wb, "Balance Sheet")
    check("Balance Sheet sheet exists", rows2 is not None)
    bal_ok = True
    bal_detail = ""
    if rows2 is None:
        bal_ok = False
        bal_detail = "sheet missing"
    else:
        header = rows2[0] if rows2 else []
        sym_i = col_index(header, "symbol")
        ta_i = col_index(header, "total_assets")
        tl_i = col_index(header, "total_liabilities")
        eq_i = col_index(header, "equity")
        check("Balance Sheet has required columns",
              None not in (sym_i, ta_i, tl_i, eq_i), f"header: {header}")
        data_rows = [r for r in rows2[1:] if r and sym_i is not None and r[sym_i] is not None]
        by_sym = {}
        for r in data_rows:
            by_sym[str(r[sym_i]).strip().upper()] = r
        for sym in SYMBOLS:
            check(f"Balance Sheet contains {sym}", sym in by_sym)
        if None in (sym_i, ta_i, tl_i, eq_i):
            bal_ok = False
            bal_detail = "missing columns"
        else:
            for sym in SYMBOLS:
                r = by_sym.get(sym)
                exp_a, exp_l, exp_e = BALANCE[sym]
                if r is None:
                    bal_ok = False
                    bal_detail = f"{sym} row missing"
                    break
                if not rel_close(r[ta_i], exp_a):
                    bal_ok = False
                    bal_detail = f"{sym} assets {r[ta_i]} != {exp_a}"
                    break
                if not rel_close(r[tl_i], exp_l):
                    bal_ok = False
                    bal_detail = f"{sym} liabilities {r[tl_i]} != {exp_l}"
                    break
                if not rel_close(r[eq_i], exp_e):
                    bal_ok = False
                    bal_detail = f"{sym} equity {r[eq_i]} != {exp_e}"
                    break
    check("Balance Sheet values match seeded MOEX balance data (assets/liabilities/equity)",
          bal_ok, bal_detail)


def check_teamly():
    """Critical: a NEW Teamly page (id>3) named like the dashboard that names the
    correct highest-revenue and best-margin companies (RU or EN forms)."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Financial Health Dashboard' page exists", False, str(e))
        check("Teamly 'Financial Health Dashboard' names highest-revenue and best-margin companies", False, str(e))
        return

    dash = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "financial health dashboard" in tl or ("financial" in tl and "dashboard" in tl) \
                or ("финанс" in tl and ("дашборд" in tl or "панель" in tl or "здоров" in tl)):
            dash = (pid, title, body)
            break
    check("Teamly 'Financial Health Dashboard' page exists", dash is not None,
          f"new pages: {[(p[0], p[1]) for p in pages]}")

    if dash is None:
        check("Teamly 'Financial Health Dashboard' names highest-revenue and best-margin companies",
              False, "no dashboard page")
        return

    text = ((dash[1] or "") + " " + (dash[2] or "")).lower()
    names_revenue = any(a in text for a in COMPANY_ALIASES[HIGHEST_REVENUE])
    names_margin = any(a in text for a in COMPANY_ALIASES[BEST_MARGIN])
    check("Teamly 'Financial Health Dashboard' names highest-revenue and best-margin companies",
          names_revenue and names_margin,
          f"highest_rev({HIGHEST_REVENUE})={names_revenue} best_margin({BEST_MARGIN})={names_margin}")


def check_email():
    """Critical: email to finance_team with the right subject and body refs."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, COALESCE(body_text, '') FROM email.messages
            WHERE to_addr::text ILIKE '%finance_team@company.com%'
            ORDER BY id DESC LIMIT 10
        """)
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Email to finance_team exists", False, str(e))
        check("Email to finance_team has subject 'Financial Health Report' and references the report + dashboard",
              False, str(e))
        return

    check("Email to finance_team exists", bool(rows), "no email to finance_team@company.com")
    ok = False
    detail = "no matching email"
    for subject, body in rows:
        subj = (subject or "").lower()
        text = ((subject or "") + " " + (body or "")).lower()
        subj_ok = "financial health report" in subj
        file_ok = "financial_health_report.xlsx" in text or "financial_health_report" in text
        dash_ok = "dashboard" in text or "teamly" in text or "дашборд" in text or "панель" in text
        kw_ok = "отчёт" in text or "отчет" in text or "анализ" in text or "report" in text
        if subj_ok and file_ok and dash_ok and kw_ok:
            ok = True
            detail = ""
            break
        detail = f"subj={subj_ok} file={file_ok} dash={dash_ok} kw={kw_ok}"
    check("Email to finance_team has subject 'Financial Health Report' and references the report + dashboard",
          ok, detail)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    print("  Checking Excel file...")
    check_excel(agent_ws)
    print("  Checking Teamly dashboard...")
    check_teamly()
    print("  Checking email...")
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if args.res_log_file:
        try:
            import json
            with open(args.res_log_file, "w") as f:
                json.dump({"total_passed": PASS_COUNT, "total_checks": total,
                           "accuracy": accuracy, "critical_failed": critical_failed}, f, indent=2)
        except Exception:
            pass

    if critical_failed or accuracy < 70:
        print("\n=== RESULT: FAIL ===")
        sys.exit(1)
    print("\n=== RESULT: PASS ===")
    sys.exit(0)


if __name__ == "__main__":
    main()
