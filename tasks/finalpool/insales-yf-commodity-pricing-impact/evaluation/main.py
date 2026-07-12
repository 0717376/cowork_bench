"""Evaluation for insales-yf-commodity-pricing-impact."""
import argparse
import os
import sys

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



def nums_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a, b = float(a), float(b)
    except (TypeError, ValueError):
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Commodity_Impact.xlsx")
    if not os.path.exists(path):
        return ["Commodity_Impact.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Gold Price Trend sheet
        rows = load_sheet_rows(wb, "Gold Price Trend")
        if rows is None:
            errors.append("Sheet 'Gold Price Trend' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 25:
                errors.append(f"Gold Price Trend has {len(data_rows)} rows, expected ~30")
            else:
                # Check that gold prices are reasonable (between 4000 and 6000)
                prices_ok = 0
                for r in data_rows:
                    if r[1] and 4000 < float(r[1]) < 6000:
                        prices_ok += 1
                if prices_ok < 20:
                    errors.append(f"Only {prices_ok} gold prices in expected range 4000-6000")
                # Check last row has recent date
                last_row = data_rows[-1]
                if last_row[0] and "2026" in str(last_row[0]):
                    pass
                else:
                    errors.append(f"Last date does not contain 2026: {last_row[0]}")
                # Check Trend_Direction column exists
                has_direction = any(r[3] and str(r[3]).strip().lower() in ("up", "down") for r in data_rows if len(r) > 3)
                if not has_direction:
                    errors.append("Trend_Direction column missing or has no Up/Down values")

        # Check Category Analysis sheet
        rows2 = load_sheet_rows(wb, "Category Analysis")
        if rows2 is None:
            errors.append("Sheet 'Category Analysis' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 5:
                errors.append(f"Category Analysis has {len(data_rows2)} rows, expected ~7")
            # Check Watches row exists
            watches_rows = [r for r in data_rows2 if r[0] and "час" in str(r[0]).lower()]
            if not watches_rows:
                errors.append("Watches category not found in Category Analysis")
            else:
                # Watches avg price should be ~60.69
                if len(watches_rows[0]) > 1 and watches_rows[0][1]:
                    if not nums_close(watches_rows[0][1], 60.69, abs_tol=5.0):
                        errors.append(f"Watches avg price {watches_rows[0][1]}, expected ~60.69")
            # Check Electronics row
            electronics_rows = [r for r in data_rows2 if r[0] and "электроник" in str(r[0]).lower()]
            if not electronics_rows:
                errors.append("Electronics category not found")
            else:
                if len(electronics_rows[0]) > 2 and electronics_rows[0][2]:
                    if not nums_close(electronics_rows[0][2], 30, abs_tol=3.0):
                        errors.append(f"Electronics product count {electronics_rows[0][2]}, expected ~30")

        # Check Correlation Summary sheet
        rows3 = load_sheet_rows(wb, "Correlation Summary")
        if rows3 is None:
            errors.append("Sheet 'Correlation Summary' not found")
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 3:
                errors.append(f"Correlation Summary has {len(data_rows3)} rows, expected >= 3")
            # Check gold price is present
            gold_row = [r for r in data_rows3 if r[0] and "gold" in str(r[0]).lower() and "price" in str(r[0]).lower()]
            gold_vals = [r[1] for r in gold_row if len(r) > 1 and r[1] is not None]
            if gold_vals and not any(nums_close(v, 5093.30, abs_tol=100.0) for v in gold_vals):
                errors.append(f"Current gold price {gold_vals}, expected ~5093")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word(agent_workspace):
    errors = []
    from docx import Document
    path = os.path.join(agent_workspace, "Commodity_Report.docx")
    if not os.path.exists(path):
        return ["Commodity_Report.docx not found"]
    try:
        doc = Document(path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
        if len(full_text) < 200:
            errors.append(f"Word doc too short ({len(full_text)} chars)")
        # Russian-aware substring checks (accept RU or EN keywords)
        if not any(k in full_text for k in ("золот", "gold")):
            errors.append("Word doc does not mention gold ('золот')")
        if not any(k in full_text for k in ("час", "watch")):
            errors.append("Word doc does not mention watches ('час')")
        if not any(k in full_text for k in ("маржа", "маржи", "марж", "margin")):
            errors.append("Word doc does not mention margin ('маржа')")
        if not any(k in full_text for k in ("тренд", "скользящ", "trend", "moving")):
            errors.append("Word doc does not mention price trend or moving average")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def run_critical_checks(agent_ws):
    """Semantic gating checks. Any failure => hard FAIL (sys.exit(1)) before accuracy.

    A correct agent that actually queried the moex-finance gold series (GLDRUB_TOM),
    the InSales catalog and produced both deliverables passes these. A non-doer fails.
    """
    import openpyxl
    crit = []

    xlsx = os.path.join(agent_ws, "Commodity_Impact.xlsx")
    docx = os.path.join(agent_ws, "Commodity_Report.docx")

    # 1) Both deliverables must exist
    if not os.path.exists(xlsx):
        crit.append("CRITICAL: Commodity_Impact.xlsx missing")
        return crit  # nothing else checkable
    if not os.path.exists(docx):
        crit.append("CRITICAL: Commodity_Report.docx missing")

    try:
        wb = openpyxl.load_workbook(xlsx, data_only=True)
    except Exception as e:
        crit.append(f"CRITICAL: cannot open Commodity_Impact.xlsx: {e}")
        return crit

    # 2) Gold series present with ~30 rows and last close in moex RUB/gram scale
    gold = load_sheet_rows(wb, "Gold Price Trend")
    if gold is None:
        crit.append("CRITICAL: 'Gold Price Trend' sheet missing")
    else:
        data = [r for r in gold[1:] if r and r[0] is not None]
        if len(data) < 25:
            crit.append(f"CRITICAL: Gold Price Trend has {len(data)} rows (expected ~30)")
        last = data[-1] if data else None
        if not (last and last[1] and nums_close(last[1], 5093.30, abs_tol=100.0)):
            crit.append("CRITICAL: latest gold close not ~5093.30 (moex GLDRUB_TOM last close)")

    # 3) Category Analysis: Часы avg ~60.69 and Электроника count ~30 (real WC catalog)
    cat = load_sheet_rows(wb, "Category Analysis")
    if cat is None:
        crit.append("CRITICAL: 'Category Analysis' sheet missing")
    else:
        rows = [r for r in cat[1:] if r and r[0] is not None]
        watches = [r for r in rows if r[0] and "час" in str(r[0]).lower()]
        if not (watches and len(watches[0]) > 1 and watches[0][1]
                and nums_close(watches[0][1], 60.69, abs_tol=5.0)):
            crit.append("CRITICAL: Часы avg price not ~60.69 (real catalog value)")
        elec = [r for r in rows if r[0] and "электроник" in str(r[0]).lower()]
        if not (elec and len(elec[0]) > 2 and elec[0][2]
                and nums_close(elec[0][2], 30, abs_tol=3.0)):
            crit.append("CRITICAL: Электроника product count not ~30 (real catalog value)")

    # 4) Correlation Summary gold price ~5093.30
    corr = load_sheet_rows(wb, "Correlation Summary")
    if corr is None:
        crit.append("CRITICAL: 'Correlation Summary' sheet missing")
    else:
        rows = [r for r in corr[1:] if r and r[0] is not None]
        grow = [r for r in rows if r[0] and "gold" in str(r[0]).lower() and "price" in str(r[0]).lower()]
        if not any(len(r) > 1 and r[1] is not None and nums_close(r[1], 5093.30, abs_tol=100.0) for r in grow):
            crit.append("CRITICAL: Correlation Summary current gold price not ~5093.30")

    return crit


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    # --- CRITICAL CHECKS (gate before accuracy) ---
    print("  Running CRITICAL checks...")
    crit = run_critical_checks(agent_ws)
    if crit:
        print(f"\n=== RESULT: FAIL (critical: {len(crit)}) ===")
        for e in crit:
            print(f"  {e}")
        sys.exit(1)
    print("    PASS (critical)")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Word document...")
    errs = check_word(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
