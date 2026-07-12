"""
Evaluation для yf-sector-analysis-word (RU / moex-finance).

Агент собирает данные через MCP `moex-finance` (get_stock_info, схема moex.*)
по пяти тикерам MOEX и формирует:
  - sector_analysis_report.docx (Word): заголовок 'Cross-Sector Stock Analysis
    Report', подзаголовок с датой 2026-03-06, раздел на каждый тикер
    '[Ticker] - [Company Name]', итоговый раздел 'Summary';
  - sector_analysis_data.xlsx (Excel): лист 'Stock Data' (7 столбцов) и лист
    'Sector Summary' (метка/значение).

Данные читаются живьём из сервиса, поэтому конкретные цены/капитализацию НЕ
хардкодим. Проверяем структуру, присутствие тикеров и ВНУТРЕННЮЮ согласованность.

CRITICAL_CHECKS (семантические): любой их провал => общий FAIL независимо от
accuracy. Структурные проверки (наличие листов/файла, заголовки, ISO-дата,
сортировка) — не критические. Порог: accuracy >= 70 И отсутствие критических
провалов => PASS.
"""

import argparse
import json
import os
import sys

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    from docx import Document
except ImportError:
    Document = None

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = False

# Пять выбранных тикеров MOEX (алфавитный порядок)
TICKERS = ["GAZP.ME", "LKOH.ME", "MGNT.ME", "MTSS.ME", "SBER.ME"]

# Имена критических проверок (семантика, не структура)
CRITICAL_CHECKS = {
    "Stock Data: ровно 5 строк с ожидаемыми тикерами MOEX",
    "Stock Data: все Market_Cap > 1e9 и все Current_Price > 0",
    "Stock Data: строки отсортированы по Ticker",
    "Summary: Highest_Market_Cap_Ticker = тикер с реальным макс. Market_Cap",
    "Summary: Lowest_Price_Ticker = тикер с реальной мин. Current_Price",
    "Summary: Average_Current_Price согласовано со средним Current_Price",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        tag = " (CRITICAL)" if name in CRITICAL_CHECKS else ""
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED = True


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def contains_any(haystack, options):
    """haystack — уже .lower() оригинальный текст; options — RU/EN варианты."""
    return any(opt in haystack for opt in options)


# ============================================================================
# Check 1: Word document
# ============================================================================

def check_word(agent_workspace):
    print("\n=== Проверка документа Word ===")

    doc_path = os.path.join(agent_workspace, "sector_analysis_report.docx")

    if not os.path.isfile(doc_path):
        record("Word: файл существует", False, f"Не найден: {doc_path}")
        return False
    record("Word: файл существует", True)

    if Document is None:
        record("Word: python-docx установлен", False, "Не удалось импортировать docx")
        return False

    doc = Document(doc_path)
    full_text = "\n".join(p.text for p in doc.paragraphs).lower()

    # Заголовок: EN-литерал сохранён, но допускаем RU-вариант
    record("Word: есть заголовок 'cross-sector'",
           contains_any(full_text, ["cross-sector", "межсектор"]),
           "Заголовок не найден")
    record("Word: упоминается дата 2026-03-06",
           "2026-03-06" in full_text,
           "Дата не найдена")

    for ticker in TICKERS:
        record(f"Word: упоминается {ticker}",
               ticker.lower() in full_text,
               f"{ticker} не найден в документе")

    record("Word: есть раздел Summary",
           contains_any(full_text, ["summary", "итоги", "итог"]),
           "Раздел Summary не найден")

    record("Word: упоминается рыночная капитализация",
           contains_any(full_text, ["market cap", "рыночная капитализация",
                                     "рыночной капитализации", "капитализаци"]),
           "Рыночная капитализация не упомянута")

    return True


# ============================================================================
# Check 2: Excel file
# ============================================================================

def check_excel(agent_workspace):
    print("\n=== Проверка файла Excel ===")

    xlsx_path = os.path.join(agent_workspace, "sector_analysis_data.xlsx")

    if not os.path.isfile(xlsx_path):
        record("Excel: файл существует", False, f"Не найден: {xlsx_path}")
        return False
    record("Excel: файл существует", True)

    if openpyxl is None:
        record("Excel: openpyxl установлен", False, "Не удалось импортировать openpyxl")
        return False

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    # -------- Лист 1: Stock Data --------
    ws1 = get_sheet(wb, "Stock Data")
    if ws1 is None:
        record("Excel: лист 'Stock Data' существует", False, f"Листы: {wb.sheetnames}")
        return False
    record("Excel: лист 'Stock Data' существует", True)

    headers = [str(c.value).strip() if c.value else "" for c in ws1[1]]
    expected_headers = ["Ticker", "Company_Name", "Current_Price", "Market_Cap",
                        "Week52_High", "Week52_Low", "Sector"]
    headers_ok = all(str_match(h, e) for h, e in zip(headers, expected_headers))
    record("Stock Data: заголовки столбцов совпадают", headers_ok,
           f"Ожидалось: {expected_headers}, получено: {headers}")

    rows = list(ws1.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if r and r[0]]

    agent_tickers = [str(r[0]).strip().upper() for r in rows]

    # CRITICAL: ровно 5 строк с ожидаемыми тикерами
    expected_set = {t.upper() for t in TICKERS}
    record("Stock Data: ровно 5 строк с ожидаемыми тикерами MOEX",
           len(rows) == 5 and set(agent_tickers) == expected_set,
           f"Тикеры: {agent_tickers}")

    # CRITICAL: сортировка по тикеру
    record("Stock Data: строки отсортированы по Ticker",
           agent_tickers == sorted(agent_tickers),
           f"Получено: {agent_tickers}")

    # CRITICAL: все цены > 0 и капитализация > 1e9
    sane = True
    sane_detail = ""
    for r in rows:
        t = str(r[0]).strip().upper()
        try:
            if not (float(r[2]) > 0):
                sane = False; sane_detail = f"{t} цена={r[2]}"; break
            if not (float(r[3]) > 1e9):
                sane = False; sane_detail = f"{t} mcap={r[3]}"; break
        except (TypeError, ValueError):
            sane = False; sane_detail = f"{t} нечисловое значение"; break
    record("Stock Data: все Market_Cap > 1e9 и все Current_Price > 0",
           sane, sane_detail)

    # -------- Лист 2: Sector Summary --------
    ws2 = get_sheet(wb, "Sector Summary")
    if ws2 is None:
        record("Excel: лист 'Sector Summary' существует", False, f"Листы: {wb.sheetnames}")
        return False
    record("Excel: лист 'Sector Summary' существует", True)

    summary = {}
    for row in ws2.iter_rows(min_row=1, values_only=True):
        if row and row[0]:
            summary[str(row[0]).strip().lower()] = row[1]

    record("Sector Summary: Total_Stocks = 5",
           str(summary.get("total_stocks", "")).strip() == "5",
           f"Получено {summary.get('total_stocks')}")

    hmc_ticker = str(summary.get("highest_market_cap_ticker", "")).strip().upper()
    lp_ticker = str(summary.get("lowest_price_ticker", "")).strip().upper()

    # Реальные max mcap / min price из данных листа Stock Data
    real_max_mcap_ticker = None
    real_min_price_ticker = None
    try:
        all_mcaps = [(str(r[0]).strip().upper(), float(r[3])) for r in rows]
        real_max_mcap_ticker = max(all_mcaps, key=lambda x: x[1])[0]
    except (TypeError, ValueError):
        pass
    try:
        all_prices = [(str(r[0]).strip().upper(), float(r[2])) for r in rows]
        real_min_price_ticker = min(all_prices, key=lambda x: x[1])[0]
    except (TypeError, ValueError):
        pass

    # CRITICAL: Highest_Market_Cap_Ticker = реальный макс. из живых данных
    record("Summary: Highest_Market_Cap_Ticker = тикер с реальным макс. Market_Cap",
           real_max_mcap_ticker is not None and hmc_ticker == real_max_mcap_ticker,
           f"Summary: {hmc_ticker}, реальный максимум: {real_max_mcap_ticker}")

    # CRITICAL: Lowest_Price_Ticker = реальный мин. из живых данных
    record("Summary: Lowest_Price_Ticker = тикер с реальной мин. Current_Price",
           real_min_price_ticker is not None and lp_ticker == real_min_price_ticker,
           f"Summary: {lp_ticker}, реальный минимум: {real_min_price_ticker}")

    # CRITICAL: среднее согласовано
    avg_price = summary.get("average_current_price")
    computed_avg = None
    try:
        computed_avg = sum(float(r[2]) for r in rows) / len(rows) if rows else None
    except (TypeError, ValueError):
        computed_avg = None
    record("Summary: Average_Current_Price согласовано со средним Current_Price",
           avg_price is not None and computed_avg is not None
           and num_close(avg_price, computed_avg, 1.0),
           f"Summary: {avg_price}, вычислено: {computed_avg}")

    # Доп. структурные проверки значений капитализации/цены в Summary
    hmc_val = summary.get("highest_market_cap_value")
    if hmc_val is not None and real_max_mcap_ticker is not None:
        try:
            real_mcap = [float(r[3]) for r in rows
                         if str(r[0]).strip().upper() == real_max_mcap_ticker][0]
            record("Summary: Highest_Market_Cap_Value согласовано",
                   num_close(hmc_val, real_mcap, tol=max(1.0, real_mcap * 1e-6)),
                   f"Summary: {hmc_val}, данные: {real_mcap}")
        except (IndexError, TypeError, ValueError):
            record("Summary: Highest_Market_Cap_Value согласовано", False,
                   "Не удалось сопоставить значение")

    lp_val = summary.get("lowest_price_value")
    if lp_val is not None and real_min_price_ticker is not None:
        try:
            real_price = [float(r[2]) for r in rows
                          if str(r[0]).strip().upper() == real_min_price_ticker][0]
            record("Summary: Lowest_Price_Value согласовано",
                   num_close(lp_val, real_price, tol=0.5),
                   f"Summary: {lp_val}, данные: {real_price}")
        except (IndexError, TypeError, ValueError):
            record("Summary: Lowest_Price_Value согласовано", False,
                   "Не удалось сопоставить значение")

    return True


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_word(args.agent_workspace)
    check_excel(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== ИТОГО ===")
    print(f"  Пройдено: {PASS_COUNT}")
    print(f"  Провалено: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILED:
        success = False
        print("  Результат: FAIL (провалена критическая проверка)")
    elif accuracy >= 70:
        success = True
        print("  Результат: PASS")
    else:
        success = False
        print("  Результат: FAIL (accuracy < 70%)")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "critical_failed": CRITICAL_FAILED,
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
