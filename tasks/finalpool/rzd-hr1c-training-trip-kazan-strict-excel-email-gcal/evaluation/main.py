"""
Strict evaluation for rzd-hr1c-training-trip-kazan-strict-excel-email-gcal.

В отличие от soft-варианта здесь проверяется *правильность* решений:
    - 5 сотрудников именно из коммерческого блока с опытом ≥ 3
    - в Travel_Plan туда только 716Г, обратно только 717Г (см. task.md)
    - бюджет ровно 45 000 ₽
    - сводное письмо содержит номера поездов / дату / бюджет
    - персональное письмо адресовано одному из участников
    - GCal: ровно одно утреннее и одно вечернее событие 17.03

PASS если ≥ 80%.
"""
import glob
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


# ─── helpers ─────────────────────────────────────────────────────────────────

def find_sheet(wb, *keywords):
    for s in wb.sheetnames:
        low = s.lower()
        if any(k in low for k in keywords):
            return wb[s]
    return None


def data_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    return [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]


def header_index(ws, *candidates):
    """Найти первый столбец, чьё имя содержит любую из подстрок (case-insensitive)."""
    hdr = next(ws.iter_rows(values_only=True), ())
    for i, cell in enumerate(hdr):
        if cell is None:
            continue
        low = str(cell).lower()
        if any(c.lower() in low for c in candidates):
            return i
    return None


# ─── DB helpers ──────────────────────────────────────────────────────────────

def fetch_eligible_employees():
    """Сотрудники, удовлетворяющие критерию (Продажи/Маркетинг + опыт ≥ 3)."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT "EMPLOYEE_NAME", "DEPARTMENT", "YEARS_EXPERIENCE", "EMAIL"
        FROM hr1c_data."HR__PUBLIC__EMPLOYEES"
        WHERE "DEPARTMENT" IN ('Продажи', 'Маркетинг')
          AND "YEARS_EXPERIENCE" >= 3
    """)
    rows = cur.fetchall()
    cur.close(); conn.close()
    return rows  # list of (name, dept, years, email)


# ─── Excel checks ────────────────────────────────────────────────────────────

def check_excel(agent_workspace):
    print("\n=== Check 1: Excel-отчёт ===")
    xlsx_files = glob.glob(os.path.join(agent_workspace, "*.xlsx"))
    train_files = [f for f in xlsx_files if any(
        kw in os.path.basename(f).lower()
        for kw in ["training", "travel", "report", "trip", "kazan"]
    )]
    if not train_files:
        record("xlsx-отчёт существует", False, f"нет подходящих xlsx: {xlsx_files}")
        return
    record("xlsx-отчёт существует", True)

    xlsx_path = train_files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel-файл читается", False, str(e))
        return
    record("Excel-файл читается", True)

    eligible = fetch_eligible_employees()
    eligible_names = {r[0] for r in eligible}

    # ── Employees ────────────────────────────────────────────────────────
    ws = find_sheet(wb, "employ", "staff", "сотрудник")
    if ws is None:
        record("лист Employees найден", False, f"листы: {wb.sheetnames}")
    else:
        record("лист Employees найден", True)
        rows = data_rows(ws)
        record("Employees ровно 5 строк", len(rows) == 5, f"строк: {len(rows)}")

        name_col = header_index(ws, "name", "фио", "имя")
        dept_col = header_index(ws, "department", "отдел")
        eligible_col = header_index(ws, "eligible", "допущен")

        all_text = " ".join(str(c) for r in rows for c in r if c is not None)
        valid_names = 0
        for r in rows:
            if name_col is not None and r[name_col] is not None \
               and str(r[name_col]).strip() in eligible_names:
                valid_names += 1
        record("все 5 ФИО присутствуют в HR-системе как валидные кандидаты",
               valid_names == 5, f"совпало {valid_names}/5")

        if dept_col is not None:
            depts_ok = all(
                r[dept_col] is not None and str(r[dept_col]).strip() in ("Продажи", "Маркетинг")
                for r in rows
            )
        else:
            depts_ok = ("Продажи" in all_text and "Маркетинг" in all_text)
        record("все 5 — из Продаж/Маркетинг", depts_ok, f"snippet: {all_text[:200]}")

        if eligible_col is not None:
            all_eligible = all(
                r[eligible_col] is not None
                and str(r[eligible_col]).strip().lower() in ("да", "yes", "true", "1")
                for r in rows
            )
        else:
            all_eligible = ("Да" in all_text or "Yes" in all_text)
        record("Training_Eligible = «Да» для всех 5", all_eligible)

    # ── Travel_Plan ──────────────────────────────────────────────────────
    ws = find_sheet(wb, "travel", "plan", "trip", "поездк")
    if ws is None:
        record("лист Travel_Plan найден", False, f"листы: {wb.sheetnames}")
    else:
        record("лист Travel_Plan найден", True)
        rows = data_rows(ws)
        record("Travel_Plan ровно 10 строк", len(rows) == 10, f"строк: {len(rows)}")

        train_col = header_index(ws, "train", "поезд", "номер")
        dir_col = header_index(ws, "direction", "направлен")
        price_col = header_index(ws, "price", "цена", "стоимост")

        def row_train(r):
            return str(r[train_col]).strip() if train_col is not None and r[train_col] is not None else ""
        def row_dir(r):
            return str(r[dir_col]).strip().lower() if dir_col is not None and r[dir_col] is not None else ""
        def row_price(r):
            try:
                return float(r[price_col]) if price_col is not None else None
            except (TypeError, ValueError):
                return None

        outbound = [r for r in rows if "туд" in row_dir(r) or "outbound" in row_dir(r)]
        ret = [r for r in rows if "обрат" in row_dir(r) or "return" in row_dir(r)
                                   or "back" in row_dir(r)]
        record("5 строк с направлением «Туда»", len(outbound) == 5, f"найдено: {len(outbound)}")
        record("5 строк с направлением «Обратно»", len(ret) == 5, f"найдено: {len(ret)}")

        outbound_trains = {row_train(r) for r in outbound}
        return_trains = {row_train(r) for r in ret}
        out_ok = outbound_trains and all(
            "716" in t and "718" not in t for t in outbound_trains
        )
        record("туда выбран ранний «Стриж» 716 (а не 718)",
               bool(out_ok), f"номера поездов туда: {outbound_trains}")
        ret_ok = return_trains and all(
            "717" in t and "719" not in t for t in return_trains
        )
        record("обратно выбран 717 — прибытие в тот же день (а не 719)",
               bool(ret_ok), f"номера поездов обратно: {return_trains}")

        prices = [row_price(r) for r in rows if row_price(r) is not None]
        record("все цены билетов = 4500 ₽",
               len(prices) == 10 and all(p == 4500 for p in prices),
               f"цены: {prices}")

    # ── Budget_Summary ───────────────────────────────────────────────────
    ws = find_sheet(wb, "budget", "summar", "бюджет")
    if ws is None:
        record("лист Budget_Summary найден", False, f"листы: {wb.sheetnames}")
    else:
        record("лист Budget_Summary найден", True)
        rows = data_rows(ws)
        record("Budget_Summary ровно 3 строки", len(rows) == 3, f"строк: {len(rows)}")
        all_vals = []
        for r in rows:
            for c in r:
                try:
                    all_vals.append(float(c))
                except (TypeError, ValueError):
                    pass
        has_45000 = 45000.0 in all_vals
        record("в Budget_Summary есть итог ровно 45 000 ₽",
               has_45000, f"числовые значения: {sorted(all_vals)}")


# ─── Calendar ────────────────────────────────────────────────────────────────

def check_gcal():
    print("\n=== Check 2: Calendar events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-17'
          AND start_datetime < '2026-03-18'
          AND summary NOT ILIKE '%кикофф%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close(); conn.close()

    record("ровно 2 события поездки на 17.03.2026", len(events) == 2,
           f"найдено {len(events)}: {[e[0] for e in events]}")
    if len(events) >= 1:
        e1 = events[0]
        morning = e1[1].hour < 7
        record("первое событие — утреннее (выезд, начало до 07:00)",
               morning, f"start: {e1[1]}")
    if len(events) >= 2:
        e2 = events[1]
        evening = e2[1].hour >= 18
        record("второе событие — вечернее (возврат, начало с 18:00)",
               evening, f"start: {e2[1]}")


# ─── Emails ──────────────────────────────────────────────────────────────────

def fetch_emails_to(pattern, exclude_from=None):
    rows = []
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        sql = ("SELECT from_addr, to_addr::text, subject, body_text FROM email.messages"
               " WHERE to_addr::text ILIKE %s")
        params = [f"%{pattern}%"]
        if exclude_from:
            sql += " AND from_addr NOT ILIKE %s"
            params.append(f"%{exclude_from}%")
        cur.execute(sql, params)
        rows += cur.fetchall()
        cur.close(); conn.close()
    except Exception:
        pass
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT from_addr, to_addr::text, subject, body_text FROM email.sent_log"
            " WHERE to_addr::text ILIKE %s",
            (f"%{pattern}%",),
        )
        rows += cur.fetchall()
        cur.close(); conn.close()
    except Exception:
        pass
    return rows


def check_emails():
    print("\n=== Check 3: Emails ===")
    hr_addr = "training@hr.company.ru"
    hr_msgs = fetch_emails_to(hr_addr, exclude_from=hr_addr)
    record("письмо отправлено на training@hr.company.ru",
           len(hr_msgs) >= 1, f"найдено: {len(hr_msgs)}")

    combined = " ".join(
        " ".join(str(x) for x in m if x) for m in hr_msgs
    ).lower()
    record("в письме HR упомянуты оба поезда (716 и 717)",
           "716" in combined and "717" in combined,
           f"snippet: {combined[:300]}")
    has_total = ("45000" in combined or "45 000" in combined)
    record("в письме HR упомянута итоговая сумма 45 000 ₽", has_total,
           f"snippet: {combined[:300]}")
    has_date = ("17.03" in combined or "17 март" in combined)
    record("в письме HR упомянута дата (17.03.2026 / 17 марта)", has_date,
           f"snippet: {combined[:300]}")

    eligible_emails = {r[3] for r in fetch_eligible_employees() if r[3]}
    employee_msgs = []
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT to_addr::text FROM email.messages"
            " WHERE to_addr::text NOT ILIKE %s AND from_addr NOT ILIKE %s",
            (f"%{hr_addr}%", f"%{hr_addr}%"),
        )
        employee_msgs += [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
    except Exception:
        pass
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT to_addr::text FROM email.sent_log WHERE to_addr::text NOT ILIKE %s",
            (f"%{hr_addr}%",),
        )
        employee_msgs += [r[0] for r in cur.fetchall()]
        cur.close(); conn.close()
    except Exception:
        pass

    addressed = any(
        any(e in to_text for e in eligible_emails) for to_text in employee_msgs
    )
    record("персональное письмо отправлено на email одного из кандидатов",
           addressed,
           f"emails to: {employee_msgs[:5]}")


# ─── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 80:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
