"""Генератор groundtruth-эталона для strict-таски.
Запускать вручную: docker run --rm -v $(pwd):/w -w /w cowork-pack:latest \
    /opt/venv/bin/python tasks/finalpool/rzd-hr1c-training-trip-kazan-strict-excel-email-gcal/groundtruth_workspace/_build_groundtruth.py
"""
import openpyxl

EMPLOYEES = [
    # ID, ФИО, Отдел, Стаж, Оклад, Email
    (10005, "Соколов Игорь Павлович",       "Продажи",   22, 3600000, "sokolov.i@company.ru"),
    (10009, "Фёдорова Ольга Владимировна",  "Продажи",   18, 2200000, "fedorova.o@company.ru"),
    (10012, "Зайцева Татьяна Александровна","Маркетинг", 17, 2700000, "zaytseva.t@company.ru"),
    (10003, "Смирнов Дмитрий Олегович",     "Продажи",   16, 2400000, "smirnov.d@company.ru"),
    (10018, "Виноградова Алёна Геннадьевна","Маркетинг", 13, 2050000, "vinogradova.a@company.ru"),
]

TICKET_PRICE = 4500

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Employees"
ws.append(["Employee_ID", "Name", "Department", "Years_Experience", "Salary", "Email", "Training_Eligible"])
for eid, name, dept, exp, sal, email in EMPLOYEES:
    ws.append([eid, name, dept, exp, sal, email, "Да"])

ws = wb.create_sheet("Travel_Plan")
ws.append(["Employee_Name", "Train_No", "Direction", "Departure_Time", "Arrival_Time", "Seat_Class", "Ticket_Price_RUB"])
for _, name, *_ in EMPLOYEES:
    ws.append([name, "716Г", "Туда",    "06:25", "11:50", "Эконом", TICKET_PRICE])
for _, name, *_ in EMPLOYEES:
    ws.append([name, "717Г", "Обратно", "18:30", "23:55", "Эконом", TICKET_PRICE])

ws = wb.create_sheet("Budget_Summary")
ws.append(["Item", "Count", "Unit_Price_RUB", "Total_RUB"])
ws.append(["Билеты туда",    5, TICKET_PRICE, 5 * TICKET_PRICE])
ws.append(["Билеты обратно", 5, TICKET_PRICE, 5 * TICKET_PRICE])
ws.append(["Итого",        "",           "",  2 * 5 * TICKET_PRICE])

out = "tasks/finalpool/rzd-hr1c-training-trip-kazan-strict-excel-email-gcal/groundtruth_workspace/Training_Travel_Report.xlsx"
wb.save(out)
print(f"groundtruth written: {out}")
