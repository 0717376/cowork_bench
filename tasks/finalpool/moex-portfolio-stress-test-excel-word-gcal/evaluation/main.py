"""Evaluation for moex-portfolio-stress-test-excel-word-gcal.

Структура: каждый чек регистрируется через record(); подмножество отмечено
как CRITICAL. Любой провал критичного чека => немедленный FAIL (sys.exit(1))
ДО проверки порога точности. Иначе PASS при accuracy >= 70%.

CRITICAL_CHECKS (семантика, ядро задачи):
  C1: Portfolio Overview — Current_Price и Sharpe_Ratio совпадают с эталоном
      (из фиксированного исторического сида moex.*) для всех 6 тикеров.
  C2: Stress Scenarios — Portfolio_Total Scenario_PnL для 'Market Crash' и
      'Historical Replay' совпадают с эталоном.
  C3: Risk Summary — Breach_Threshold и Worst_Scenario совпадают с эталоном.
  C4: Risk Summary — Portfolio_VaR_95 и Max_Historical_Drawdown_Pct в допуске.
  C5: GCal — ровно 4 события 'Stress Test Review - <Scenario>' на 2026-03-16,
      03-23, 03-30, 04-06, покрывающие все четыре сценария.

MOEX-значения берутся из эталонных файлов groundtruth_workspace, которые
регенерированы из ДЕТЕРМИНИРОВАННОГО сида moex.* (не из живого/волатильного фида).
"""
import argparse
import datetime
import json
import os
import sys

import psycopg2

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

CRITICAL_CHECKS = {
    "Portfolio Overview: Current_Price и Sharpe_Ratio совпадают с эталоном (6 тикеров MOEX)",
    "Stress Scenarios: Portfolio_Total PnL для Market Crash и Historical Replay совпадают с эталоном",
    "Risk Summary: Breach_Threshold и Worst_Scenario совпадают с эталоном",
    "Risk Summary: Portfolio_VaR_95 и Max_Historical_Drawdown_Pct в допуске",
    "GCal: 4 события 'Stress Test Review - <Scenario>' на нужные даты, все 4 сценария",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    msg = f": {detail[:300]}" if detail else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    import openpyxl

    agent_path = os.path.join(agent_workspace, "Stress_Test_Report.xlsx")
    gt_path = os.path.join(groundtruth_workspace, "Stress_Test_Report.xlsx")

    if not os.path.exists(agent_path):
        record("Excel: Stress_Test_Report.xlsx существует", False, "файл не найден")
        # Register critical checks as failed so missing file => FAIL
        record("Portfolio Overview: Current_Price и Sharpe_Ratio совпадают с эталоном (6 тикеров MOEX)", False, "нет файла")
        record("Stress Scenarios: Portfolio_Total PnL для Market Crash и Historical Replay совпадают с эталоном", False, "нет файла")
        record("Risk Summary: Breach_Threshold и Worst_Scenario совпадают с эталоном", False, "нет файла")
        record("Risk Summary: Portfolio_VaR_95 и Max_Historical_Drawdown_Pct в допуске", False, "нет файла")
        return
    record("Excel: Stress_Test_Report.xlsx существует", True)

    if not os.path.exists(gt_path):
        record("Excel: эталонный файл существует", False, "groundtruth не найден")
        return
    record("Excel: эталонный файл существует", True)

    try:
        wb_agent = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        record("Excel: файл читается", False, str(e))
        return

    # --- Sheet: Portfolio Overview ---
    agent_rows = load_sheet_rows(wb_agent, "Portfolio Overview")
    gt_rows = load_sheet_rows(wb_gt, "Portfolio Overview")
    record("Portfolio Overview: лист присутствует", agent_rows is not None)

    price_ok = True
    sharpe_ok = True
    price_detail = ""
    if agent_rows and gt_rows:
        agent_data = [r for r in agent_rows[1:] if r and r[0] is not None]
        gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]
        record("Portfolio Overview: число строк совпадает",
               len(agent_data) >= len(gt_data),
               f"{len(agent_data)} vs {len(gt_data)}")

        agent_lookup = {str(r[0]).strip().upper(): r for r in agent_data}
        gt_lookup = {str(r[0]).strip().upper(): r for r in gt_data}

        for sym, gt_row in gt_lookup.items():
            if sym not in agent_lookup:
                record(f"Portfolio Overview: тикер {sym} присутствует", False)
                price_ok = sharpe_ok = False
                continue
            record(f"Portfolio Overview: тикер {sym} присутствует", True)
            a_row = agent_lookup[sym]
            # Allocation_Pct (col 1) - structural
            record(f"{sym} Allocation_Pct", num_close(a_row[1], gt_row[1], 0.5),
                   f"{a_row[1]} vs {gt_row[1]}")
            # Monthly_Volatility_Pct (col 5) - structural/non-critical
            record(f"{sym} Monthly_Volatility_Pct", num_close(a_row[5], gt_row[5], 0.5),
                   f"{a_row[5]} vs {gt_row[5]}")
            # Current_Price (col 3) -> contributes to CRITICAL C1 (RUB scale, tol relaxed)
            if not num_close(a_row[3], gt_row[3], 5.0):
                price_ok = False
                price_detail = f"{sym} price {a_row[3]} vs {gt_row[3]}"
            # Sharpe_Ratio (col 7) -> contributes to CRITICAL C1
            if not num_close(a_row[7], gt_row[7], 0.1):
                sharpe_ok = False
                price_detail = (price_detail + f"; {sym} sharpe {a_row[7]} vs {gt_row[7]}").strip("; ")
    else:
        price_ok = sharpe_ok = False
        price_detail = "нет данных листа"

    record("Portfolio Overview: Current_Price и Sharpe_Ratio совпадают с эталоном (6 тикеров MOEX)",
           price_ok and sharpe_ok, price_detail)

    # --- Sheet: Stress Scenarios ---
    agent_rows2 = load_sheet_rows(wb_agent, "Stress Scenarios")
    gt_rows2 = load_sheet_rows(wb_gt, "Stress Scenarios")
    record("Stress Scenarios: лист присутствует", agent_rows2 is not None)

    crit_scen_ok = True
    crit_scen_detail = ""
    if agent_rows2 and gt_rows2:
        agent_data2 = [r for r in agent_rows2[1:] if r and r[0] is not None]
        gt_data2 = [r for r in gt_rows2[1:] if r and r[0] is not None]

        gt_totals = {}
        for r in gt_data2:
            if r[1] and str(r[1]).strip() == "Portfolio_Total":
                gt_totals[str(r[0]).strip()] = (r[4], r[5])
        agent_totals = {}
        for r in agent_data2:
            if r[1] and str(r[1]).strip() == "Portfolio_Total":
                agent_totals[str(r[0]).strip()] = (r[4], r[5])

        # Non-critical: every scenario total present and within tol (RUB scale tol)
        for sc_name, (gt_val, gt_pnl) in gt_totals.items():
            if sc_name not in agent_totals:
                record(f"Stress Scenarios: Portfolio_Total для {sc_name}", False, "строка отсутствует")
                if sc_name in ("Market Crash", "Historical Replay"):
                    crit_scen_ok = False
                    crit_scen_detail += f"{sc_name} отсутствует; "
                continue
            a_val, a_pnl = agent_totals[sc_name]
            ok_val = num_close(a_val, gt_val, 100.0)
            ok_pnl = num_close(a_pnl, gt_pnl, 100.0)
            record(f"Stress Scenarios: {sc_name} total Value", ok_val, f"{a_val} vs {gt_val}")
            record(f"Stress Scenarios: {sc_name} total PnL", ok_pnl, f"{a_pnl} vs {gt_pnl}")
            if sc_name in ("Market Crash", "Historical Replay") and not ok_pnl:
                crit_scen_ok = False
                crit_scen_detail += f"{sc_name} PnL {a_pnl} vs {gt_pnl}; "
        # ensure both critical scenarios were present
        for sc_name in ("Market Crash", "Historical Replay"):
            if sc_name not in gt_totals or sc_name not in agent_totals:
                if sc_name not in agent_totals and sc_name in gt_totals:
                    crit_scen_ok = False
    else:
        crit_scen_ok = False
        crit_scen_detail = "нет данных листа"

    record("Stress Scenarios: Portfolio_Total PnL для Market Crash и Historical Replay совпадают с эталоном",
           crit_scen_ok, crit_scen_detail)

    # --- Sheet: Risk Summary ---
    agent_rows3 = load_sheet_rows(wb_agent, "Risk Summary")
    gt_rows3 = load_sheet_rows(wb_gt, "Risk Summary")
    record("Risk Summary: лист присутствует", agent_rows3 is not None)

    breach_ws_ok = True
    breach_detail = ""
    var_dd_ok = True
    var_dd_detail = ""
    if agent_rows3 and gt_rows3:
        agent_data3 = [r for r in agent_rows3[1:] if r and r[0] is not None]
        gt_data3 = [r for r in gt_rows3[1:] if r and r[0] is not None]
        agent_metrics = {str(r[0]).strip().lower(): r[1] for r in agent_data3}
        gt_metrics = {str(r[0]).strip().lower(): r[1] for r in gt_data3}

        for metric, gt_val in gt_metrics.items():
            if metric not in agent_metrics:
                record(f"Risk Summary: метрика {metric} присутствует", False)
                if metric in ("breach_threshold", "worst_scenario"):
                    breach_ws_ok = False
                if metric in ("portfolio_var_95", "max_historical_drawdown_pct"):
                    var_dd_ok = False
                continue
            a_val = agent_metrics[metric]
            if metric in ("worst_scenario", "best_scenario", "breach_threshold"):
                ok = str(a_val).strip().lower() == str(gt_val).strip().lower()
                record(f"Risk Summary: {metric}", ok, f"'{a_val}' vs '{gt_val}'")
                if metric in ("breach_threshold", "worst_scenario") and not ok:
                    breach_ws_ok = False
                    breach_detail += f"{metric}: {a_val} vs {gt_val}; "
            elif metric == "total_portfolio_value":
                record(f"Risk Summary: {metric}", num_close(a_val, gt_val, 100), f"{a_val} vs {gt_val}")
            elif metric in ("portfolio_var_95", "worst_scenario_loss", "best_scenario_pnl"):
                # RUB scale: tol widened from $500 to 50000 RUB (~0.05% of 100M)
                ok = num_close(a_val, gt_val, 50000)
                record(f"Risk Summary: {metric}", ok, f"{a_val} vs {gt_val}")
                if metric == "portfolio_var_95" and not ok:
                    var_dd_ok = False
                    var_dd_detail += f"VaR {a_val} vs {gt_val}; "
            elif metric in ("max_historical_drawdown_pct", "worst_scenario_loss_pct"):
                ok = num_close(a_val, gt_val, 1.0)
                record(f"Risk Summary: {metric}", ok, f"{a_val} vs {gt_val}")
                if metric == "max_historical_drawdown_pct" and not ok:
                    var_dd_ok = False
                    var_dd_detail += f"MDD {a_val} vs {gt_val}; "
    else:
        breach_ws_ok = False
        var_dd_ok = False
        breach_detail = var_dd_detail = "нет данных листа"

    record("Risk Summary: Breach_Threshold и Worst_Scenario совпадают с эталоном",
           breach_ws_ok, breach_detail)
    record("Risk Summary: Portfolio_VaR_95 и Max_Historical_Drawdown_Pct в допуске",
           var_dd_ok, var_dd_detail)


# RU+EN keyword pairs: a requirement is satisfied if ANY variant appears.
REQUIRED_SECTIONS = [
    ["executive summary", "краткое резюме", "резюме"],
    ["portfolio composition", "структура портфеля", "состав портфеля"],
    ["stress test results", "результаты стресс"],
    ["risk metrics", "метрики риска"],
    ["scenario comparison", "сравнение сценариев"],
    ["recommendation", "рекомендац"],
]
VAR_KEYWORDS = ["var", "value at risk", "стоимость под риском"]
DRAWDOWN_KEYWORDS = ["drawdown", "просадк"]


def check_word(agent_workspace):
    try:
        from docx import Document
    except Exception as e:
        record("Word: библиотека доступна", False, str(e))
        return

    path = os.path.join(agent_workspace, "Risk_Assessment.docx")
    if not os.path.exists(path):
        record("Word: Risk_Assessment.docx существует", False, "файл не найден")
        return
    record("Word: Risk_Assessment.docx существует", True)

    try:
        doc = Document(path)
        full_text = "\n".join([p.text for p in doc.paragraphs]).lower()
    except Exception as e:
        record("Word: документ читается", False, str(e))
        return

    for variants in REQUIRED_SECTIONS:
        ok = any(v in full_text for v in variants)
        record(f"Word: раздел {variants[0]}", ok)

    record("Word: упомянут VaR / стоимость под риском",
           any(k in full_text for k in VAR_KEYWORDS))
    record("Word: упомянута просадка / drawdown",
           any(k in full_text for k in DRAWDOWN_KEYWORDS))


def check_gcal():
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"),
            port=5432,
            dbname="cowork_gym",
            user=os.environ.get("PGUSER", "eigent"),
            password=os.environ.get("PGPASSWORD", "camel"),
        )
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date, description
            FROM gcal.events
            WHERE LOWER(summary) LIKE '%stress test%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("GCal: 4 события 'Stress Test Review - <Scenario>' на нужные даты, все 4 сценария",
               False, str(e))
        return

    expected_scenarios = ["market crash", "sector rotation", "inflation shock", "historical replay"]
    found_scenarios = set()
    for summary, _d, _desc in rows:
        sl = summary.lower() if summary else ""
        for sc in expected_scenarios:
            if sc in sl:
                found_scenarios.add(sc)

    expected_dates = [
        datetime.date(2026, 3, 16),
        datetime.date(2026, 3, 23),
        datetime.date(2026, 3, 30),
        datetime.date(2026, 4, 6),
    ]
    actual_dates = sorted([r[1] for r in rows])

    # Non-critical structural sub-checks
    record("GCal: найдено >= 4 событий Stress Test", len(rows) >= 4, f"найдено {len(rows)}")
    for sc in expected_scenarios:
        record(f"GCal: событие для сценария {sc}", sc in found_scenarios)
    dates_ok = True
    for i, exp in enumerate(expected_dates):
        ok = i < len(actual_dates) and actual_dates[i] == exp
        record(f"GCal: дата встречи {i+1} = {exp}", ok,
               f"{actual_dates[i] if i < len(actual_dates) else 'нет'}")
        if not ok:
            dates_ok = False

    # CRITICAL aggregate
    crit_ok = (len(rows) >= 4
               and all(sc in found_scenarios for sc in expected_scenarios)
               and dates_ok)
    record("GCal: 4 события 'Stress Test Review - <Scenario>' на нужные даты, все 4 сценария",
           crit_ok,
           f"rows={len(rows)} scen={sorted(found_scenarios)} dates={[str(d) for d in actual_dates]}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace")

    print("  Checking Excel file...")
    check_excel(agent_ws, gt_ws)
    print("  Checking Word document...")
    check_word(agent_ws)
    print("  Checking GCal events...")
    check_gcal()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: ни один чек не выполнен.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_fails": CRITICAL_FAILS,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)

    if CRITICAL_FAILS:
        print(f"FAIL: критичные чеки провалены ({len(CRITICAL_FAILS)}): {CRITICAL_FAILS}")
        sys.exit(1)
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
