"""
Evaluation script for moex-stock-comparison-excel-gcal task.

Checks:
1. Excel file Stock_Comparison_Report.xlsx - 2 sheets with correct structure and data
2. Google Calendar has investment review event (RU title) on 2026-03-18
3. Email sent to analyst@investment.com mentioning the three MOEX tickers
4. Word document Stock_Analysis_Report.docx exists with RU/EN analysis content

MOEX prices are read live from the moex.* seed and are volatile, so the
evaluation does NOT compare absolute prices against a pre-built groundtruth.
Instead it verifies SELF-CONSISTENCY of the agent's own numbers
(Price_Change == Latest - Month_Start, Return_Pct sign/ranking) and the
core structural deliverables. Critical semantic checks gate the result.
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

SYMBOLS = ["SBER", "GAZP", "LKOH"]

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILED.append(name)


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=1.0):
    fa, fb = to_float(a), to_float(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= tol


def load_sheet_by_name(wb, name):
    for sname in wb.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            return [[cell.value for cell in row] for row in wb[sname].iter_rows()]
    return None


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace):
    print("\n=== Checking Stock_Comparison_Report.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        return False

    agent_file = os.path.join(agent_workspace, "Stock_Comparison_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

    all_ok = True

    # ----- Price History sheet -----
    a_hist = load_sheet_by_name(agent_wb, "Price History")
    record("Sheet 'Price History' exists", a_hist is not None)

    if a_hist is not None:
        a_data = [r for r in a_hist[1:] if any(v is not None for v in r)]
        record("Price History has at least 15 rows (trading days)",
               len(a_data) >= 15,
               f"Found {len(a_data)} rows")

        if a_hist and len(a_hist) > 0:
            header = [str(v).strip().lower() if v else "" for v in a_hist[0]]
            record("Price History has Date column", any("date" in h for h in header))
            record("Price History has SBER column",
                   any("sber" in h for h in header))
            record("Price History has GAZP column",
                   any("gazp" in h for h in header))
            record("Price History has LKOH column",
                   any("lkoh" in h for h in header))

    # ----- Performance Summary sheet -----
    a_summ = load_sheet_by_name(agent_wb, "Performance Summary")
    record("Sheet 'Performance Summary' exists", a_summ is not None)

    summary_rows = {}  # SYMBOL -> dict of parsed fields
    if a_summ is not None:
        a_data = [r for r in a_summ[1:] if any(v is not None for v in r)]

        # CRITICAL: exactly 3 rows for the 3 MOEX symbols
        record("Performance Summary has exactly 3 rows (SBER, GAZP, LKOH)",
               len(a_data) == 3,
               f"Found {len(a_data)} rows",
               critical=True)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().upper().split(".")[0]] = row

        # CRITICAL: all three expected symbols present
        present = [s for s in SYMBOLS if s in a_lookup]
        record("Performance Summary covers SBER, GAZP, LKOH",
               len(present) == 3,
               f"Found symbols: {sorted(a_lookup.keys())}",
               critical=True)

        # Columns: Symbol, Latest_Price, Month_Start_Price, Price_Change,
        #          Return_Pct, Min_Price, Max_Price, Volatility_Score
        for sym in SYMBOLS:
            a_row = a_lookup.get(sym)
            if a_row is None:
                continue
            latest = to_float(a_row[1]) if len(a_row) > 1 else None
            start = to_float(a_row[2]) if len(a_row) > 2 else None
            change = to_float(a_row[3]) if len(a_row) > 3 else None
            ret = to_float(a_row[4]) if len(a_row) > 4 else None
            mn = to_float(a_row[5]) if len(a_row) > 5 else None
            mx = to_float(a_row[6]) if len(a_row) > 6 else None
            vol = to_float(a_row[7]) if len(a_row) > 7 else None
            summary_rows[sym] = {
                "latest": latest, "start": start, "change": change,
                "ret": ret, "min": mn, "max": mx, "vol": vol,
            }

            record(f"{sym}: Latest_Price is a positive number",
                   latest is not None and latest > 0,
                   f"got {a_row[1] if len(a_row) > 1 else None}")
            record(f"{sym}: Volatility_Score is numeric",
                   vol is not None,
                   f"got {a_row[7] if len(a_row) > 7 else None}")

            # CRITICAL: self-consistent Price_Change = Latest - Month_Start
            if latest is not None and start is not None and change is not None:
                record(f"{sym}: Price_Change == Latest - Month_Start (self-consistent)",
                       abs(round(change - (latest - start), 2)) <= 0.01,
                       f"change={change}, latest-start={latest - start:.4f}",
                       critical=True)
            else:
                record(f"{sym}: Price_Change == Latest - Month_Start (self-consistent)",
                       False, "missing latest/start/change", critical=True)

            # Min <= Latest <= Max sanity (non-critical)
            if mn is not None and mx is not None and latest is not None:
                record(f"{sym}: Min_Price <= Latest <= Max_Price",
                       mn - 0.01 <= latest <= mx + 0.01,
                       f"min={mn}, latest={latest}, max={mx}")

    # CRITICAL: Return_Pct best/worst ranking matches Latest vs Month_Start
    if len(summary_rows) == 3 and all(
            r["latest"] is not None and r["start"] is not None and r["ret"] is not None
            for r in summary_rows.values()):
        # ranking derived from the agent's own latest vs month_start
        derived = {s: (r["latest"] - r["start"]) / r["start"] * 100
                   for s, r in summary_rows.items()}
        best_derived = max(derived, key=derived.get)
        worst_derived = min(derived, key=derived.get)
        best_reported = max(summary_rows, key=lambda s: summary_rows[s]["ret"])
        worst_reported = min(summary_rows, key=lambda s: summary_rows[s]["ret"])
        record("Return_Pct ranking matches Latest-vs-MonthStart derived ranking",
               best_derived == best_reported and worst_derived == worst_reported,
               f"derived best/worst={best_derived}/{worst_derived}, "
               f"reported best/worst={best_reported}/{worst_reported}",
               critical=True)
        # stash for email cross-check
        check_excel.best = best_reported
        check_excel.worst = worst_reported
        check_excel.returns = {s: summary_rows[s]["ret"] for s in summary_rows}
    else:
        record("Return_Pct ranking matches Latest-vs-MonthStart derived ranking",
               False, "incomplete Performance Summary data", critical=True)

    return all_ok


# ============================================================================
# Check 2: Google Calendar
# ============================================================================

def check_gcal():
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")
    record("At least 1 calendar event created", len(events) >= 1, f"Found {len(events)}")

    # RU + EN keyword alternatives for the review meeting title
    kw = ("portfolio", "investment", "review",
          "портфел", "инвест", "обзор", "ревью", "комитет")
    portfolio_events = [e for e in events
                        if e[0] and any(k in e[0].lower() for k in kw)]
    record("Investment/Portfolio review event found",
           len(portfolio_events) >= 1,
           f"Events: {[e[0] for e in events[:5]]}")

    # Event on 2026-03-18
    march18_events = [e for e in events
                      if e[1] and "2026-03-18" in str(e[1])]
    record("Event on 2026-03-18", len(march18_events) >= 1,
           f"March 18 events: {[e[0] for e in march18_events]}")

    # CRITICAL: a review-titled event exists on 2026-03-18
    crit_events = [e for e in march18_events
                   if e[0] and any(k in e[0].lower() for k in kw)]
    record("Portfolio/investment review event on 2026-03-18 exists",
           len(crit_events) >= 1,
           f"March 18 review events: {[e[0] for e in crit_events]}",
           critical=True)

    return len(portfolio_events) >= 1 and len(march18_events) >= 1


# ============================================================================
# Check 3: Email
# ============================================================================

def check_emails():
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(all_emails)} total emails.")
    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    # Find the email addressed to the analyst (recipient is the anchor)
    target = None
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = str(to_addr or "").lower()
        if "analyst@investment.com" in to_str:
            target = (subject, from_addr, to_addr, body_text)
            break

    record("Email to analyst@investment.com found", target is not None,
           f"Emails: {[(e[0], str(e[2])[:60]) for e in all_emails[:3]]}",
           critical=True)

    if target is None:
        return False

    subject, from_addr, to_addr, body_text = target
    subject_lower = (subject or "").lower()
    body_lower = (body_text or "").lower()

    record("Email subject mentions stock/performance/march",
           any(k in subject_lower for k in ("stock", "performance", "march")),
           f"Subject: {subject}")

    # CRITICAL: body mentions all three tickers (RU substrings)
    record("Email body mentions all three tickers (SBER, GAZP, LKOH)",
           all(t in body_lower for t in ("sber", "gazp", "lkoh")),
           f"Body: {body_lower[:200]}",
           critical=True)

    # Body should reference returns/percentages
    record("Email body references return percentage(s)",
           "%" in (body_text or "") or "процент" in body_lower or "доходн" in body_lower,
           "Body missing return percentages")

    # Cross-check best/worst naming if excel gave us a ranking
    best = getattr(check_excel, "best", None)
    worst = getattr(check_excel, "worst", None)
    if best and worst:
        record("Email body names the best- and worst-performing stock",
               best.lower() in body_lower and worst.lower() in body_lower,
               f"expected best={best}, worst={worst}")

    return True


# ============================================================================
# Check 4: Word document
# ============================================================================

def check_word(agent_workspace):
    print("\n=== Checking Stock_Analysis_Report.docx ===")

    docx_path = os.path.join(agent_workspace, "Stock_Analysis_Report.docx")
    if not os.path.isfile(docx_path):
        record("Word file exists", False, f"Not found: {docx_path}")
        return False
    record("Word file exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        record("Word doc has content", len(all_text.strip()) >= 100,
               f"Content length: {len(all_text.strip())}")
        # RU + EN terms / tickers
        terms = ("stock", "analysis", "акци", "анализ", "сравнен",
                 "sber", "gazp", "lkoh", "сбер", "газпром", "лукойл")
        record("Word doc mentions stock analysis content",
               any(t in all_text for t in terms),
               "Missing stock analysis content")
        return True
    except ImportError:
        size = os.path.getsize(docx_path)
        record("Word file has content (>2KB)", size > 2000, f"Size: {size} bytes")
        return size > 2000
    except Exception as e:
        record("Word file readable", False, str(e))
        return False


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    gcal_ok = check_gcal()
    email_ok = check_emails()
    word_ok = check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    # Critical checks gate the result BEFORE the accuracy threshold.
    if CRITICAL_FAILED:
        print(f"  CRITICAL checks failed: {CRITICAL_FAILED}")
        print(f"  Overall: FAIL (critical)")
        if args.res_log_file:
            with open(args.res_log_file, "w") as f:
                json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                           "accuracy": accuracy, "success": False,
                           "critical_failed": CRITICAL_FAILED}, f, indent=2)
        sys.exit(1)

    success = accuracy >= 70.0
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy, "success": success}, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
