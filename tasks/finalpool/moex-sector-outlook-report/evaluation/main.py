"""Evaluation for yf-sector-outlook-report (RU: moex-finance + teamly).

The agent fetches the firm's sector outlook JSON (RU sector labels) from the
mock portal, pulls MOEX price history for 5 holdings, builds Sector_Outlook.xlsx
(two sheets), a Sector_Report.docx narrative, and a teamly page
'Sector Outlook Report - Q1 2026' under the RESEARCH space.

Nothing volatile is hardcoded:
  * outlook / growth_forecast / risk_level are READ from the seeded
    files/mock_pages/api/sector_outlook.json the agent also reads.
  * the (deliberate) sector mapping override is encoded here (the core trick):
    MGNT.ME and MTSS.ME are forced into 'Технологии' despite their official
    sector, so they carry the Технологии outlook.
  * each Return_1Y_Pct is RECOMPUTED from moex.stock_prices the same way the
    agent reads it: earliest available close (~1y request) vs latest close.
  * Positive_Outlook_Count / High_Risk_Count / Avg / Best / Worst sectors are
    recomputed from the seeded JSON + mapping + recomputed returns.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import psycopg2

DB = dict(
    host=os.environ.get("PGHOST", "localhost"),
    port=5432,
    dbname=os.environ.get("PGDATABASE", "cowork_gym"),
    user="eigent",
    password="camel",
)

# Holdings and the (trick) sector mapping the agent must apply.
# MGNT.ME and MTSS.ME are remapped into 'Технологии' (override).
STOCK_SECTOR = {
    "SBER.ME": "Финансы",
    "GAZP.ME": "Энергетика",
    "LKOH.ME": "Энергетика",
    "MGNT.ME": "Технологии",
    "MTSS.ME": "Технологии",
}
SYMBOLS = list(STOCK_SECTOR.keys())

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "All 5 holdings present with correctly mapped sector + outlook",
    "Remapped tickers (MGNT.ME, MTSS.ME) carry the Технологии override outlook",
    "Per-stock Return_1Y_Pct matches recomputed value from moex.stock_prices",
    "Cross-Sector Summary counts + average match recomputed values",
    "Teamly page 'Sector Outlook Report - Q1 2026' exists with avg return + outlook contrast",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        d = (str(detail)[:300]) if detail else ""
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, abs_tol=0.5, rel_tol=0.03):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


# --------------------------------------------------------------------------
# Source-of-truth loaders (non-volatile recomputation)
# --------------------------------------------------------------------------
def load_outlook_json():
    here = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(here, "..", "files", "mock_pages", "api", "sector_outlook.json")
    with open(path, encoding="utf-8") as f:
        return json.load(f)["sectors"]


def load_returns():
    """Recompute each stock's return from moex.stock_prices the same way the
    agent does: earliest available close (history over ~1y) vs latest close."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    returns = {}
    prices = {}
    for sym in SYMBOLS:
        cur.execute(
            "SELECT date, close FROM moex.stock_prices WHERE symbol = %s "
            "AND close IS NOT NULL ORDER BY date ASC",
            (sym,),
        )
        rows = cur.fetchall()
        if not rows:
            continue
        first_close = float(rows[0][1])
        last_close = float(rows[-1][1])
        ret = round((last_close - first_close) / first_close * 100, 2)
        returns[sym] = ret
        prices[sym] = (first_close, last_close)
    cur.close()
    conn.close()
    return returns, prices


def expected_summary(sectors, returns):
    pos = sum(1 for s in SYMBOLS if sectors.get(STOCK_SECTOR[s], {}).get("outlook") == "Positive")
    high = sum(1 for s in SYMBOLS if sectors.get(STOCK_SECTOR[s], {}).get("risk_level") == "High")
    avg = round(sum(returns[s] for s in SYMBOLS) / len(SYMBOLS), 2)
    best_sym = max(returns, key=returns.get)
    worst_sym = min(returns, key=returns.get)
    return {
        "positive_outlook_count": pos,
        "high_risk_count": high,
        "avg_1y_return": avg,
        "best_1y_sector": STOCK_SECTOR[best_sym],
        "worst_1y_sector": STOCK_SECTOR[worst_sym],
    }


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------
def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_ws, sectors, returns):
    print("\n=== Checking Excel ===")
    import openpyxl
    path = os.path.join(agent_ws, "Sector_Outlook.xlsx")
    if not os.path.exists(path):
        check("All 5 holdings present with correctly mapped sector + outlook", False, "xlsx missing")
        check("Remapped tickers (MGNT.ME, MTSS.ME) carry the Технологии override outlook", False, "xlsx missing")
        check("Per-stock Return_1Y_Pct matches recomputed value from moex.stock_prices", False, "xlsx missing")
        check("Cross-Sector Summary counts + average match recomputed values", False, "xlsx missing")
        return
    wb = openpyxl.load_workbook(path, data_only=True)

    rows = load_sheet_rows(wb, "Sector Performance")
    perf_by_stock = {}
    if rows is None:
        check("Sheet 'Sector Performance' present", False)
    else:
        check("Sheet 'Sector Performance' present", True)
        header = [str(c).strip() if c is not None else "" for c in rows[0]] if rows else []
        # index columns by name
        idx = {h.lower(): i for i, h in enumerate(header)}
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        for r in data_rows:
            sym = str(r[idx.get("stock", 1)]).strip().upper() if "stock" in idx else (
                str(r[1]).strip().upper() if len(r) > 1 and r[1] else None)
            if sym:
                perf_by_stock[sym] = (r, idx)

    # CRITICAL: all 5 holdings present and mapped to the right sector + outlook
    mapped_ok = True
    detail = []
    for sym in SYMBOLS:
        if sym not in perf_by_stock:
            mapped_ok = False
            detail.append(f"{sym} missing")
            continue
        r, idx = perf_by_stock[sym]
        sec = str(r[idx.get("sector", 0)]).strip() if "sector" in idx else str(r[0]).strip()
        exp_sec = STOCK_SECTOR[sym]
        if sec != exp_sec:
            mapped_ok = False
            detail.append(f"{sym} sector={sec} expected {exp_sec}")
            continue
        if "outlook" in idx:
            ol = str(r[idx["outlook"]]).strip()
            exp_ol = sectors[exp_sec]["outlook"]
            if ol.lower() != exp_ol.lower():
                mapped_ok = False
                detail.append(f"{sym} outlook={ol} expected {exp_ol}")
    check("All 5 holdings present with correctly mapped sector + outlook", mapped_ok, "; ".join(detail))

    # CRITICAL: the override trick — MGNT/MTSS carry Технологии outlook (Positive)
    override_ok = True
    odet = []
    for sym in ["MGNT.ME", "MTSS.ME"]:
        if sym not in perf_by_stock:
            override_ok = False
            odet.append(f"{sym} missing")
            continue
        r, idx = perf_by_stock[sym]
        sec = str(r[idx.get("sector", 0)]).strip()
        ol = str(r[idx.get("outlook", 5)]).strip() if "outlook" in idx else ""
        if sec != "Технологии" or ol.lower() != sectors["Технологии"]["outlook"].lower():
            override_ok = False
            odet.append(f"{sym} sector={sec} outlook={ol}")
    check("Remapped tickers (MGNT.ME, MTSS.ME) carry the Технологии override outlook",
          override_ok, "; ".join(odet))

    # CRITICAL: each Return_1Y_Pct matches recomputed value
    ret_ok = True
    rdet = []
    for sym in SYMBOLS:
        if sym not in perf_by_stock or sym not in returns:
            ret_ok = False
            rdet.append(f"{sym} no data")
            continue
        r, idx = perf_by_stock[sym]
        if "return_1y_pct" not in idx:
            ret_ok = False
            rdet.append("Return_1Y_Pct column missing")
            break
        val = r[idx["return_1y_pct"]]
        if not num_close(val, returns[sym], abs_tol=0.5, rel_tol=0.05):
            ret_ok = False
            rdet.append(f"{sym} return={val} expected ~{returns[sym]}")
    check("Per-stock Return_1Y_Pct matches recomputed value from moex.stock_prices",
          ret_ok, "; ".join(rdet))

    # Non-critical structural: growth_forecast / risk_level present & correct
    gr_ok = True
    for sym in SYMBOLS:
        if sym not in perf_by_stock:
            gr_ok = False
            break
        r, idx = perf_by_stock[sym]
        sec = STOCK_SECTOR[sym]
        if "growth_forecast" in idx and not num_close(
                r[idx["growth_forecast"]], sectors[sec]["growth_forecast"], abs_tol=0.1):
            gr_ok = False
        if "risk_level" in idx and str(r[idx["risk_level"]]).strip().lower() != sectors[sec]["risk_level"].lower():
            gr_ok = False
    check("Growth_Forecast / Risk_Level looked up correctly per sector", gr_ok)

    # Cross-Sector Summary (CRITICAL for counts + avg)
    exp = expected_summary(sectors, returns)
    rows2 = load_sheet_rows(wb, "Cross-Sector Summary")
    if rows2 is None:
        check("Sheet 'Cross-Sector Summary' present", False)
        check("Cross-Sector Summary counts + average match recomputed values", False, "sheet missing")
    else:
        check("Sheet 'Cross-Sector Summary' present", True)
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        lookup = {str(r[0]).strip().lower(): r[1] for r in data2 if r[0]}
        counts_ok = True
        cdet = []
        if not num_close(lookup.get("positive_outlook_count"), exp["positive_outlook_count"], abs_tol=0):
            counts_ok = False
            cdet.append(f"Positive_Outlook_Count={lookup.get('positive_outlook_count')} exp {exp['positive_outlook_count']}")
        if not num_close(lookup.get("high_risk_count"), exp["high_risk_count"], abs_tol=0):
            counts_ok = False
            cdet.append(f"High_Risk_Count={lookup.get('high_risk_count')} exp {exp['high_risk_count']}")
        if not num_close(lookup.get("avg_1y_return"), exp["avg_1y_return"], abs_tol=0.5, rel_tol=0.05):
            counts_ok = False
            cdet.append(f"Avg_1Y_Return={lookup.get('avg_1y_return')} exp ~{exp['avg_1y_return']}")
        check("Cross-Sector Summary counts + average match recomputed values", counts_ok, "; ".join(cdet))

        # Non-critical: best/worst sector
        bw_ok = (str(lookup.get("best_1y_sector", "")).strip() == exp["best_1y_sector"]
                 and str(lookup.get("worst_1y_sector", "")).strip() == exp["worst_1y_sector"])
        check("Best_1Y_Sector / Worst_1Y_Sector correct", bw_ok,
              f"best exp {exp['best_1y_sector']}, worst exp {exp['worst_1y_sector']}; got "
              f"{lookup.get('best_1y_sector')}/{lookup.get('worst_1y_sector')}")


# --------------------------------------------------------------------------
# Word
# --------------------------------------------------------------------------
def check_word(agent_ws):
    print("\n=== Checking Word ===")
    path = os.path.join(agent_ws, "Sector_Report.docx")
    if not os.path.exists(path):
        check("Sector_Report.docx exists", False)
        return
    check("Sector_Report.docx exists", True)
    try:
        from docx import Document
        doc = Document(path)
        text = "\n".join(p.text for p in doc.paragraphs).lower()
    except Exception as e:
        check("Sector_Report.docx readable", False, str(e))
        return
    check("Sector_Report.docx has substantive narrative (>=200 chars)", len(text) >= 200,
          f"{len(text)} chars")
    # RU+EN keyword acceptance
    kw_groups = [
        ("sector", "сектор"),
        ("outlook", "прогноз"),
        ("technology", "технолог"),
        ("energy", "энергет"),
    ]
    missing = []
    for group in kw_groups:
        if not any(k in text for k in group):
            missing.append("/".join(group))
    check("Sector_Report.docx mentions sector/outlook/technology/energy (RU or EN)",
          not missing, f"missing: {missing}")


# --------------------------------------------------------------------------
# Teamly
# --------------------------------------------------------------------------
def check_teamly(sectors, returns):
    print("\n=== Checking Teamly ===")
    exp = expected_summary(sectors, returns)
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute(
            "SELECT p.title, COALESCE(p.body, '') FROM teamly.pages p "
            "JOIN teamly.spaces s ON s.id = p.space_id WHERE s.key = 'RESEARCH'"
        )
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly page 'Sector Outlook Report - Q1 2026' exists with avg return + outlook contrast",
              False, str(e))
        return

    target = [(t, b) for t, b in rows if t and "sector outlook report" in t.lower()]
    check("Teamly page titled 'Sector Outlook Report - Q1 2026' present (structural)",
          len(target) >= 1, f"RESEARCH pages: {[t for t, _ in rows]}")

    page_text = " ".join((str(t) + " " + str(b)) for t, b in target).lower()

    # avg return mentioned (accept the rounded value or its integer part)
    avg = exp["avg_1y_return"]
    avg_variants = {
        f"{avg:.2f}", f"{avg:.1f}", str(int(round(avg))),
        f"{avg:.2f}".replace(".", ","), f"{avg:.1f}".replace(".", ","),
    }
    avg_ok = any(v in page_text for v in avg_variants)

    # positive vs cautious contrast mentioned (RU or EN)
    has_pos = "positive" in page_text or "позитив" in page_text
    has_cau = "cautious" in page_text or "осторож" in page_text

    check("Teamly page 'Sector Outlook Report - Q1 2026' exists with avg return + outlook contrast",
          bool(target) and avg_ok and has_pos and has_cau,
          f"avg_ok={avg_ok} (exp {avg}), positive={has_pos}, cautious={has_cau}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    print("=" * 70)
    print("YF-SECTOR-OUTLOOK-REPORT (moex + teamly) - EVALUATION")
    print("=" * 70)

    sectors = load_outlook_json()
    returns, _ = load_returns()
    if len(returns) < len(SYMBOLS):
        print(f"[warn] only {len(returns)}/{len(SYMBOLS)} symbols have price data in moex.stock_prices")

    check_excel(agent_ws, sectors, returns)
    check_word(agent_ws)
    check_teamly(sectors, returns)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print("\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "total_passed": PASS_COUNT,
                "total_checks": total,
                "accuracy": accuracy,
                "critical_failed": critical_failed,
                "success": (not critical_failed) and accuracy >= 70,
            }, f, indent=2)

    if critical_failed:
        print("Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("Overall: PASS")
        sys.exit(0)
    print("Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
