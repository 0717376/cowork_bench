"""Evaluation for canvas-course-comparison."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


SUMMARY_METRICS = ("total_courses", "avg_students_per_course", "most_popular_course")


def summary_lookup(rows):
    """Map Metric (lower) -> Value from a Summary sheet.

    Accepts both layouts:
      - vertical Metric|Value (row[0]=metric, row[1]=value), skipping header
      - horizontal: header row holds the metric names as columns, with the
        first data row holding the values.
    task.md only names the three metrics, not a layout, so both are valid.
    """
    out = {}
    if not rows:
        return out
    header = rows[0] if rows else []
    header_keys = [str(c).strip().lower() if c is not None else None for c in header]
    # Horizontal only when the header spreads >=2 metrics across columns; a
    # single metric name in row[0] is a headerless vertical table, not a header.
    if sum(1 for k in header_keys if k in SUMMARY_METRICS) >= 2:
        value_row = rows[1] if len(rows) > 1 else []
        for idx, key in enumerate(header_keys):
            if key in SUMMARY_METRICS:
                out[key] = value_row[idx] if idx < len(value_row) else None
        return out
    start = 0 if (header_keys and header_keys[0] in SUMMARY_METRICS) else 1
    for row in rows[start:]:
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
    return out


def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Course_Comparison.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Course_Comparison.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- Load Course Comparison rows ----
    a_cmp = load_sheet_rows(agent_wb, "Course Comparison")
    g_cmp = load_sheet_rows(gt_wb, "Course Comparison")

    # ============================================================
    # CRITICAL CHECKS (semantic core deliverable).
    # Any failure => immediate FAIL, before the accuracy gate.
    # These reflect the substance: the agent must actually enumerate
    # all live Canvas courses and aggregate them correctly.
    # ============================================================
    critical_errors = []

    a_data = a_cmp[1:] if a_cmp and len(a_cmp) > 1 else []
    g_data = g_cmp[1:] if g_cmp and len(g_cmp) > 1 else []

    a_by_code = {}
    for row in a_data:
        if row and row[0] is not None:
            a_by_code[str(row[0]).strip().lower()] = row
    g_by_code = {}
    for row in g_data:
        if row and row[0] is not None:
            g_by_code[str(row[0]).strip().lower()] = row

    # CRITICAL 1: at least ~90% of groundtruth course rows present with
    # Students within tol=10 (agent enumerated all Canvas courses, not a subset).
    if not g_by_code:
        critical_errors.append("CRITICAL: groundtruth 'Course Comparison' is empty")
    else:
        present_ok = 0
        for key, g_row in g_by_code.items():
            a_row = a_by_code.get(key)
            if a_row is None:
                continue
            gv = g_row[2] if len(g_row) > 2 else None
            av = a_row[2] if len(a_row) > 2 else None
            if num_close(av, gv, 10):
                present_ok += 1
        coverage = present_ok / len(g_by_code)
        if coverage < 0.90:
            critical_errors.append(
                f"CRITICAL: only {present_ok}/{len(g_by_code)} course rows present with "
                f"correct Students (tol=10), coverage={coverage:.2f} < 0.90"
            )

    # Load Summary sheets
    a_sum = summary_lookup(load_sheet_rows(agent_wb, "Summary"))
    g_sum = summary_lookup(load_sheet_rows(gt_wb, "Summary"))

    # CRITICAL 2: Total_Courses equals number of data rows in Comparison AND groundtruth.
    a_total = to_num(a_sum.get("total_courses"))
    g_total = to_num(g_sum.get("total_courses"))
    if a_total is None:
        critical_errors.append("CRITICAL: Summary.Total_Courses missing")
    else:
        if g_total is not None and abs(a_total - g_total) > 0:
            critical_errors.append(
                f"CRITICAL: Total_Courses={a_total} != groundtruth {g_total}"
            )
        # internal consistency with actual Comparison data rows
        if abs(a_total - len(a_data)) > 0:
            critical_errors.append(
                f"CRITICAL: Total_Courses={a_total} != number of Comparison rows {len(a_data)}"
            )

    # CRITICAL 3: Most_Popular_Course equals the course with max Students in
    # the agent's own Comparison sheet AND matches groundtruth code (case-insensitive).
    a_pop = a_sum.get("most_popular_course")
    g_pop = g_sum.get("most_popular_course")
    if a_pop is None:
        critical_errors.append("CRITICAL: Summary.Most_Popular_Course missing")
    else:
        if g_pop is not None and not str_match(a_pop, g_pop):
            critical_errors.append(
                f"CRITICAL: Most_Popular_Course={a_pop} != groundtruth {g_pop}"
            )
        # internal consistency: must be the max-Students row in agent data
        best_code, best_students = None, None
        for row in a_data:
            if not row or row[0] is None:
                continue
            s = to_num(row[2] if len(row) > 2 else None)
            if s is None:
                continue
            if best_students is None or s > best_students:
                best_students, best_code = s, str(row[0]).strip()
        if best_code is not None and not str_match(a_pop, best_code):
            critical_errors.append(
                f"CRITICAL: Most_Popular_Course={a_pop} is not the max-Students course "
                f"in the Comparison sheet ({best_code} has {best_students})"
            )

    # CRITICAL 4: Avg_Students_Per_Course consistent with the agent's own rows
    # = round(sum(Students)/Total_Courses), tol<=2, and matches groundtruth tol<=2.
    a_avg = to_num(a_sum.get("avg_students_per_course"))
    g_avg = to_num(g_sum.get("avg_students_per_course"))
    if a_avg is None:
        critical_errors.append("CRITICAL: Summary.Avg_Students_Per_Course missing")
    else:
        students = [to_num(r[2]) for r in a_data if r and len(r) > 2 and to_num(r[2]) is not None]
        if students and a_total:
            expected_avg = round(sum(students) / len(students))
            if abs(a_avg - expected_avg) > 2:
                critical_errors.append(
                    f"CRITICAL: Avg_Students_Per_Course={a_avg} not consistent with "
                    f"round(sum/Total)={expected_avg} (tol=2)"
                )
        if g_avg is not None and abs(a_avg - g_avg) > 2:
            critical_errors.append(
                f"CRITICAL: Avg_Students_Per_Course={a_avg} != groundtruth {g_avg} (tol=2)"
            )

    if critical_errors:
        print("\n=== CRITICAL CHECK FAILURES ===")
        for e in critical_errors:
            print(f"  {e}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)

    # ============================================================
    # ACCURACY GATE (non-critical structural / per-cell checks).
    # PASS requires accuracy >= 70.
    # ============================================================
    total_checks = 0
    passed_checks = 0
    all_errors = []

    # Course Comparison per-cell
    print("  Checking Course Comparison...")
    if a_cmp is None:
        all_errors.append("Sheet 'Course Comparison' not found in agent output")
    elif g_cmp is None:
        all_errors.append("Sheet 'Course Comparison' not found in groundtruth")
    else:
        errors = []
        for key, g_row in g_by_code.items():
            a_row = a_by_code.get(key)
            if a_row is None:
                total_checks += 4
                errors.append(f"Missing row: {g_row[0]}")
                continue
            for idx, label, tol in [
                (2, "Students", 10), (3, "Assignments", 2),
                (4, "Quizzes", 2), (5, "Discussions", 2),
            ]:
                if len(a_row) > idx and len(g_row) > idx:
                    total_checks += 1
                    if num_close(a_row[idx], g_row[idx], tol):
                        passed_checks += 1
                    else:
                        errors.append(f"{key}.{label}: {a_row[idx]} vs {g_row[idx]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # Summary per-cell
    print("  Checking Summary...")
    for metric in ("total_courses", "avg_students_per_course", "most_popular_course"):
        if metric in g_sum:
            total_checks += 1
            tol = 10.0 if metric != "most_popular_course" else 0
            if num_close(a_sum.get(metric), g_sum.get(metric), tol):
                passed_checks += 1
            else:
                all_errors.append(f"Summary.{metric}: {a_sum.get(metric)} vs {g_sum.get(metric)}")

    accuracy = (passed_checks / total_checks * 100.0) if total_checks else 0.0
    print(f"\nAccuracy: {passed_checks}/{total_checks} = {accuracy:.1f}%")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
