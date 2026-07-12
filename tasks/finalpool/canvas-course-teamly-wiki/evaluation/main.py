"""
Evaluation script for canvas-course-notion-wiki task (Teamly knowledge base).

Checks:
1. Excel file (course_catalog.xlsx) - correct course data and summary
2. Teamly pages created for each course (knowledge base wiki)

Source of truth is the LIVE Canvas MCP instance; the expected counts below
mirror that fixed Canvas seed (codes ending in 2014J).

Critical checks (CRITICAL_CHECKS): any failure there => overall FAIL regardless
of accuracy. Otherwise pass threshold: accuracy >= 70%.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

COURSE_CODES = ["AAA-2014J", "BBB-2014J", "CCC-2014J", "DDD-2014J",
                "EEE-2014J", "FFF-2014J", "GGG-2014J"]

EXPECTED_STUDENTS = {
    "AAA-2014J": 365, "BBB-2014J": 2292, "CCC-2014J": 2498,
    "DDD-2014J": 1803, "EEE-2014J": 1188, "FFF-2014J": 2365, "GGG-2014J": 749,
}

EXPECTED_ASSIGNMENTS = {
    "AAA-2014J": 6, "BBB-2014J": 6, "CCC-2014J": 10,
    "DDD-2014J": 7, "EEE-2014J": 5, "FFF-2014J": 13, "GGG-2014J": 10,
}

# Instructor last names per course (from the fixed Canvas instance), used to
# verify Teamly page bodies actually list the course instructors.
EXPECTED_INSTRUCTORS = {
    "AAA-2014J": ["Martin", "Walker"],
    "BBB-2014J": ["Jackson", "Parker"],
    "CCC-2014J": ["Rivera"],
    "DDD-2014J": ["Wilson", "Cook"],
    "EEE-2014J": ["Jones", "Morris"],
    "FFF-2014J": ["Allen", "Lee", "Taylor", "Davis"],
    "GGG-2014J": ["Walker", "Bailey", "Adams"],
}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Excel file exists",
    "Courses: all 7 Student_Count values correct",
    "Courses: all 7 Assignment_Count values correct",
    "Summary: Total_Students correct",
    "Summary: Total_Assignments correct",
    "Summary: Largest_Course = CCC-2014J",
    "Summary: Smallest_Course = AAA-2014J",
    "Summary: Most_Assignments = FFF-2014J",
    "Teamly: at least 7 course pages created",
    "Teamly: every 2014J course code is a page title",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def int_close(a, b, tol=10):
    try:
        return abs(int(float(a)) - int(float(b))) <= tol
    except (TypeError, ValueError):
        return False


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "course_catalog.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(agent_file, data_only=True)

    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    # Sheet 1: Courses
    ws1 = get_sheet(wb, "Courses")
    if ws1 is None:
        record("Sheet 'Courses' exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Sheet 'Courses' exists", True)

    headers = [str(c.value).strip() if c.value else "" for c in ws1[1]]
    expected_headers = ["Course_Code", "Course_Name", "Start_Date", "End_Date",
                        "Student_Count", "Assignment_Count", "Instructor_Names"]
    headers_ok = all(str_match(h, e) for h, e in zip(headers, expected_headers))
    record("Courses headers match", headers_ok,
           f"Expected: {expected_headers}, Got: {headers}")

    rows = list(ws1.iter_rows(min_row=2, values_only=True))
    record("Courses has 7 rows", len(rows) == 7, f"Got {len(rows)}")

    # Check sorted by course code
    agent_codes = [str(r[0]).strip() for r in rows if r and r[0]]
    record("Courses sorted by Course_Code",
           agent_codes == sorted(agent_codes),
           f"Got: {agent_codes}")

    students_all_ok = True
    assignments_all_ok = True
    for code in COURSE_CODES:
        agent_row = None
        for r in rows:
            if r and str_match(r[0], code):
                agent_row = r
                break
        if not agent_row:
            record(f"Course {code} present", False, "Missing")
            students_all_ok = False
            assignments_all_ok = False
            continue
        record(f"Course {code} present", True)

        s_ok = int_close(agent_row[4], EXPECTED_STUDENTS[code])
        if not s_ok:
            students_all_ok = False
            print(f"    -> {code} Student_Count expected ~{EXPECTED_STUDENTS[code]}, got {agent_row[4]}")

        a_ok = int_close(agent_row[5], EXPECTED_ASSIGNMENTS[code], 2)
        if not a_ok:
            assignments_all_ok = False
            print(f"    -> {code} Assignment_Count expected ~{EXPECTED_ASSIGNMENTS[code]}, got {agent_row[5]}")

    # CRITICAL aggregate checks for per-course correctness.
    record("Courses: all 7 Student_Count values correct", students_all_ok,
           "One or more Student_Count values wrong")
    record("Courses: all 7 Assignment_Count values correct", assignments_all_ok,
           "One or more Assignment_Count values wrong")

    # Sheet 2: Summary
    ws2 = get_sheet(wb, "Summary")
    if ws2 is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Sheet 'Summary' exists", True)

    summary = {}
    for row in ws2.iter_rows(min_row=1, values_only=True):
        if row and row[0]:
            summary[str(row[0]).strip().lower()] = row[1]

    record("Summary: Total_Courses = 7",
           str(summary.get("total_courses", "")).strip() == "7",
           f"Got {summary.get('total_courses')}")

    record("Summary: Total_Students correct",
           int_close(summary.get("total_students", 0), 11260, 50),
           f"Expected ~11260, got {summary.get('total_students')}")

    record("Summary: Total_Assignments correct",
           int_close(summary.get("total_assignments", 0), 57, 3),
           f"Expected ~57, got {summary.get('total_assignments')}")

    record("Summary: Largest_Course = CCC-2014J",
           str_match(summary.get("largest_course", ""), "CCC-2014J"),
           f"Got {summary.get('largest_course')}")

    record("Summary: Smallest_Course = AAA-2014J",
           str_match(summary.get("smallest_course", ""), "AAA-2014J"),
           f"Got {summary.get('smallest_course')}")

    record("Summary: Most_Assignments = FFF-2014J",
           str_match(summary.get("most_assignments", ""), "FFF-2014J"),
           f"Got {summary.get('most_assignments')}")


# ============================================================================
# Check 2: Teamly knowledge-base pages
# ============================================================================

def check_teamly():
    print("\n=== Checking Teamly Pages ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Only user-created pages (seed pages have id <= 3).
        cur.execute("""
            SELECT title, body
            FROM teamly.pages
            WHERE id > 3
        """)
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Teamly: at least 7 course pages created", False, f"DB error: {e}")
        record("Teamly: every 2014J course code is a page title", False, f"DB error: {e}")
        return

    print(f"[check_teamly] Found {len(pages)} user-created Teamly pages.")

    record("Teamly: at least 7 course pages created", len(pages) >= 7,
           f"Found {len(pages)}")

    titles = [(t or "") for t, _b in pages]

    # Every course code must appear as a page title in "[Code] - ..." form.
    all_codes_titled = True
    for code in COURSE_CODES:
        matched = None
        for t in titles:
            tl = t.lower()
            if code.lower() in tl and "-" in t:
                matched = t
                break
        record(f"Teamly: page title for {code} ('[Code] - [Name]')",
               matched is not None,
               f"No '[{code}] - ...' title among {titles}")
        if matched is None:
            all_codes_titled = False

    record("Teamly: every 2014J course code is a page title", all_codes_titled,
           "One or more course codes missing as a page title")

    # Each course page body should contain the student enrollment count and
    # at least one of the course's instructors (read from Canvas).
    def body_for(code):
        for t, b in pages:
            if code.lower() in (t or "").lower():
                return (b or "")
        return ""

    for code in COURSE_CODES:
        body = body_for(code)
        bl = body.lower()
        students = str(EXPECTED_STUDENTS[code])
        student_ok = students in body
        instr_ok = any(ln.lower() in bl for ln in EXPECTED_INSTRUCTORS[code])
        record(f"Teamly {code}: body has student count + instructor",
               student_ok and instr_ok,
               f"students_in_body={student_ok}, instructor_in_body={instr_ok}")


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {critical_failed}")

    all_passed = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if all_passed else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "critical_failed": critical_failed,
            "success": all_passed,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        sys.exit(0)
    sys.exit(1)


if __name__ == "__main__":
    main()
