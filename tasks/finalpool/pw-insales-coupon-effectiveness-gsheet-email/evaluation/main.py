"""Evaluation script for pw-insales-coupon-effectiveness-gsheet-email."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Benchmark category -> Market_Avg_Price as shipped on the russified mock
# dashboard (http://localhost:30309). Categories are the russified store
# category NAMEs (the join key). Used to prove the agent really scraped the
# page and really joined on category name (not invented rows / filler).
BENCHMARK = {
    "Электроника": 58.07,
    "ТВ и домашний кинотеатр": 289.65,
    "Аудио": 74.15,
    "Камеры": 22.71,
    "Бытовая техника": 33.98,
    "Часы": 57.66,
}

# task.md is russified and instructs RU column names (e.g. "категория товара",
# "наша средняя цена"). The groundtruth/eval contract uses canonical English
# header identifiers. Map RU (and EN) header variants to the canonical English
# keys so a correct russified report satisfies the same column/row-key checks.
# This only aligns the contract with the russified task; CRITICAL value
# tolerances are unchanged.
HEADER_ALIASES = {
    # Category (the join key / main dimension)
    'category': 'category',
    'категория': 'category',
    'категория товара': 'category',
    'категория товаров': 'category',
    'наименование категории': 'category',
    'название категории': 'category',
    # Product_Count
    'product_count': 'product_count',
    'количество товаров': 'product_count',
    'кол-во товаров': 'product_count',
    'число товаров': 'product_count',
    # Our_Avg_Price (our internal value)
    'our_avg_price': 'our_avg_price',
    'наша средняя цена': 'our_avg_price',
    'наш средний чек': 'our_avg_price',
    'наша средняя стоимость': 'our_avg_price',
    'средняя цена (наша)': 'our_avg_price',
    # Total_Sales
    'total_sales': 'total_sales',
    'итого продаж': 'total_sales',
    'всего продаж': 'total_sales',
    'общие продажи': 'total_sales',
    'сумма продаж': 'total_sales',
    # Market_Avg_Price (external benchmark value)
    'market_avg_price': 'market_avg_price',
    'рыночная средняя цена': 'market_avg_price',
    'бенчмарк средний чек': 'market_avg_price',
    'бенчмарк средняя цена': 'market_avg_price',
    'эталонная средняя цена': 'market_avg_price',
    'средняя цена рынка': 'market_avg_price',
    # Price_Gap_Pct (gap / difference)
    'price_gap_pct': 'price_gap_pct',
    'разрыв цен, %': 'price_gap_pct',
    'разрыв цен %': 'price_gap_pct',
    'разрыв цен': 'price_gap_pct',
    'разрыв (%)': 'price_gap_pct',
    'разрыв, %': 'price_gap_pct',
    'разница цен, %': 'price_gap_pct',
    'разница цен %': 'price_gap_pct',
    'разница цен': 'price_gap_pct',
    'процент разрыва': 'price_gap_pct',
    # Metrics sheet
    'metric': 'metric',
    'метрика': 'metric',
    'показатель': 'metric',
    'value': 'value',
    'значение': 'value',
    # Recommendations sheet
    'priority': 'priority',
    'приоритет': 'priority',
    'action': 'action',
    'действие': 'action',
    'рекомендуемое действие': 'action',
    'рекомендация': 'action',
}


def normalize_header(value):
    """Lowercase/strip a header cell and map RU/EN variants to canonical EN key."""
    h = str(value).strip().lower() if value is not None else ""
    return HEADER_ALIASES.get(h, h)


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {tag}{name}: {detail_str}")
        if critical:
            CRITICAL_FAILED.append(name)

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    excel_path = os.path.join(agent_workspace, "Coupon_Effectiveness_Report.xlsx")
    check("Coupon_Effectiveness_Report.xlsx exists", os.path.exists(excel_path))

    # collected for critical checks
    da_rows = []          # list of dicts keyed by lowercased column name
    da_categories = set()

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [normalize_header(c.value) for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            for expected_col in ['Category', 'Product_Count', 'Our_Avg_Price', 'Total_Sales', 'Market_Avg_Price', 'Price_Gap_Pct']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            for raw in data_rows:
                if raw is None or all(v is None for v in raw):
                    continue
                row = {headers[i]: raw[i] for i in range(min(len(headers), len(raw)))}
                da_rows.append(row)
                cat = row.get('category')
                if cat is not None and str(cat).strip():
                    da_categories.add(str(cat).strip())

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
            headers = [normalize_header(c.value) for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            headers = [normalize_header(c.value) for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Category']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("wc_coupons_processor.py exists", os.path.exists(os.path.join(agent_workspace, "wc_coupons_processor.py")))

        # ---------------- CRITICAL CHECKS ----------------

        # C1: real category join — the russified store categories from the
        # benchmark must be present in Data_Analysis (proves a real category-NAME
        # join, not invented rows).
        matched_cats = [c for c in BENCHMARK if c in da_categories]
        check("Data_Analysis Category column carries the russified benchmark store categories",
              len(matched_cats) >= 6,
              f"matched {len(matched_cats)}/6: {sorted(matched_cats)} ; got categories={sorted(da_categories)}",
              critical=True)

        # C2: scraped benchmark values — for each benchmark category present,
        # Market_Avg_Price must equal the dashboard value (within tolerance).
        # Proves the agent actually scraped the playwright page.
        market_ok = 0
        market_checked = 0
        for row in da_rows:
            cat = str(row.get('category')).strip() if row.get('category') is not None else ""
            if cat in BENCHMARK:
                market_checked += 1
                mv = safe_float(row.get('market_avg_price'))
                if mv is not None and abs(mv - BENCHMARK[cat]) <= max(0.5, 0.01 * BENCHMARK[cat]):
                    market_ok += 1
        check("Market_Avg_Price matches scraped benchmark dashboard values",
              market_checked >= 6 and market_ok >= 6,
              f"matched {market_ok}/{market_checked} benchmark categories",
              critical=True)

        # C3: Price_Gap_Pct is a correct function of Our_Avg_Price vs
        # Market_Avg_Price (recompute per row, row-internal — no hardcoding of
        # volatile live store values). Proves real gap analysis, not constant filler.
        gap_ok = 0
        gap_checked = 0
        for row in da_rows:
            our = safe_float(row.get('our_avg_price'))
            mkt = safe_float(row.get('market_avg_price'))
            gap = safe_float(row.get('price_gap_pct'))
            if our is None or mkt is None or gap is None or mkt == 0:
                continue
            gap_checked += 1
            expected = (our - mkt) / mkt * 100.0
            if abs(gap - expected) <= max(1.0, 0.05 * abs(expected)):
                gap_ok += 1
        check("Price_Gap_Pct = (Our_Avg_Price - Market_Avg_Price)/Market_Avg_Price*100 per row",
              gap_checked >= 6 and gap_ok >= gap_checked,
              f"correct {gap_ok}/{gap_checked} rows",
              critical=True)

        # C4: email to team-lead@company.com, exact subject 'Analysis Report
        # Complete', non-empty body referencing gap/recommendation findings.
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "SELECT subject, to_addr, body_text FROM email.messages "
                "WHERE subject = %s",
                ('Analysis Report Complete',))
            rows = cur.fetchall()
            conn.close()
            ok = False
            for subject, to_addr, body in rows:
                to_s = str(to_addr or "")
                body_l = str(body or "").lower()
                addressed = 'team-lead@company.com' in to_s
                # RU or EN finding keywords
                kw = any(k in body_l for k in [
                    'gap', 'разрыв', 'разниц', 'рекоменд', 'recommend',
                    'цен', 'price', 'бенчмарк', 'benchmark', 'категор', 'categor'])
                if addressed and len(body_l.strip()) > 0 and kw:
                    ok = True
                    break
            check("Email to team-lead@company.com subject 'Analysis Report Complete' with findings body",
                  ok, f"found {len(rows)} subject-matching emails", critical=True)
        except Exception as e:
            check("Email to team-lead@company.com subject 'Analysis Report Complete' with findings body",
                  False, str(e), critical=True)

        # C5: Google Sheet 'Wc Coupons Tracker' exists AND has >= 1 data row
        # beyond the header carrying key category/gap data points.
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE title ILIKE %s",
                        ('%tracker%',))
            ss = cur.fetchall()
            tracker_ok = False
            detail = f"found {len(ss)} tracker spreadsheets"
            for ss_id, title in ss:
                cur.execute(
                    "SELECT value FROM gsheet.cells WHERE spreadsheet_id = %s",
                    (ss_id,))
                vals = [str(v[0]) for v in cur.fetchall() if v[0] is not None and str(v[0]).strip()]
                if len(vals) < 2:
                    continue
                joined = " | ".join(vals)
                # at least one benchmark category present as a data point
                has_cat = any(c in joined for c in BENCHMARK)
                if has_cat and len(vals) >= 4:  # header cells + at least one data row
                    tracker_ok = True
                    detail = f"'{title}' has {len(vals)} non-empty cells incl. category data"
                    break
            conn.close()
            check("Google Sheet 'Wc Coupons Tracker' exists with >= 1 category data row",
                  tracker_ok, detail, critical=True)
        except Exception as e:
            check("Google Sheet 'Wc Coupons Tracker' exists with >= 1 category data row",
                  False, str(e), critical=True)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0

    if CRITICAL_FAILED:
        print(f"CRITICAL CHECK(S) FAILED: {CRITICAL_FAILED}")
        print(f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%) — FAIL (critical)")
        sys.exit(1)

    passed = accuracy >= 70.0
    return passed, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
