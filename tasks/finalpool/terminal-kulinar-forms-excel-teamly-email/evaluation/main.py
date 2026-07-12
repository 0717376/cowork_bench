"""Evaluation для terminal-kulinar-forms-excel-teamly-email (RU-стек: kulinar/forms/teamly).

Проверяет:
  1. Meal_Program_Plan.xlsx (4 листа: Survey_Questions / Recipe_Selection /
     Weekly_Menu / Program_Summary).
  2. Forms (gform.*): опрос предпочтений по обедам с 5 вопросами.
  3. Teamly (teamly.pages): страница «База знаний рецептов» с записями по меню.
  4. Email на all_staff про программу обедов.
  5. menu_planner.py существует.

Критические чеки (CRITICAL_CHECKS) отражают СУТЬ задачи: любой их fail => FAIL,
даже если accuracy >= 70%. Структурные чеки (лист есть, колонка есть) — мягкие.
Порог: accuracy >= 70% И нет критических провалов.

Источники правды (не хардкод): названия/категории/сложность блюд сверяются с
канонической базой kulinar (KULINAR_RECIPES ниже, синхронной с all_recipes.json);
ставки стоимости читаются из dietary_requirements.json того же воркспейса, что и у агента.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Каноническая база блюд kulinar: name -> (category, difficulty).
# Источник: local_servers/kulinar-mcp/src/data/all_recipes.json (50 блюд).
KULINAR_RECIPES = {
    "Салат Оливье": ("салат", 2), "Винегрет": ("салат", 1),
    "Сельдь под шубой": ("салат", 2), "Салат Мимоза": ("салат", 2),
    "Крабовый салат": ("салат", 1), "Греческий салат": ("салат", 1),
    "Салат с курицей и грибами": ("салат", 2), "Холодец": ("закуска", 4),
    "Икра кабачковая": ("закуска", 2), "Грибы маринованные": ("закуска", 2),
    "Сало солёное": ("закуска", 2), "Селёдка с луком": ("закуска", 1),
    "Борщ": ("суп", 3), "Щи из квашеной капусты": ("суп", 2),
    "Солянка мясная": ("суп", 3), "Уха": ("суп", 2), "Окрошка": ("суп", 1),
    "Грибной суп": ("суп", 1), "Рассольник": ("суп", 2),
    "Куриный бульон с лапшой": ("суп", 1), "Бефстроганов": ("горячее", 2),
    "Пельмени домашние": ("горячее", 3), "Голубцы": ("горячее", 3),
    "Котлеты домашние": ("горячее", 2), "Жаркое в горшочках": ("горячее", 2),
    "Курица в сметане": ("горячее", 1), "Рыба запечённая по-русски": ("горячее", 2),
    "Цыплёнок табака": ("горячее", 3), "Гречка с тушёнкой": ("горячее", 1),
    "Плов узбекский": ("горячее", 4), "Картофельное пюре": ("гарнир", 1),
    "Гречневая каша": ("гарнир", 1), "Перловая каша": ("гарнир", 1),
    "Картофель отварной с укропом": ("гарнир", 1), "Рис отварной": ("гарнир", 1),
    "Пирожки с капустой жареные": ("выпечка", 3),
    "Пирожки с мясом печёные": ("выпечка", 3), "Блины тонкие": ("выпечка", 2),
    "Кулебяка с капустой и яйцом": ("выпечка", 4), "Расстегаи с рыбой": ("выпечка", 3),
    "Медовик": ("десерт", 3), "Наполеон": ("десерт", 4), "Сырники": ("десерт", 1),
    "Пасха творожная": ("десерт", 3), "Ватрушки с творогом": ("десерт", 2),
    "Кисель ягодный": ("десерт", 1), "Морс клюквенный": ("напиток", 1),
    "Компот из сухофруктов": ("напиток", 1), "Сбитень": ("напиток", 1),
    "Квас домашний": ("напиток", 2),
}

# Ставки стоимости по умолчанию (если dietary_requirements.json недоступен).
DEFAULT_COST = {"горячее": 8.0, "закуска": 7.0, "гарнир": 6.0, "салат": 5.0, "суп": 4.0}
SERVINGS = 50

CRITICAL_CHECKS = {
    "Recipe_Selection: >=7 строк, все блюда реальны (kulinar) с верными category/difficulty",
    "Weekly_Menu: 5 дней, нет двух подряд одной категории, сложность всех блюд <= 4",
    "Weekly_Menu/Program_Summary: дневная стоимость = ставке категории; total_weekly_cost = sum*50",
    "Forms: опрос предпочтений по обедам с >=5 вопросами",
    "Teamly: страница «База знаний рецептов» с записями по блюдам недельного меню (>=5)",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:250]}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def norm_recipe(s):
    s = (s or "").strip().lower().replace("ё", "е")
    return " ".join(s.split())


CANON = {norm_recipe(k): (cat, diff) for k, (cat, diff) in KULINAR_RECIPES.items()}


def lookup_recipe(name):
    """Вернёт (category, difficulty) из канона по названию (точное или вхождение)."""
    n = norm_recipe(name)
    if n in CANON:
        return CANON[n]
    for cn, val in CANON.items():
        if cn and (cn in n or n in cn):
            return val
    return None


def load_costs(workspace):
    path = os.path.join(workspace, "dietary_requirements.json")
    try:
        with open(path, encoding="utf-8") as f:
            d = json.load(f)
        cps = d.get("cost_per_serving") or {}
        if cps:
            return {str(k).strip().lower(): float(v) for k, v in cps.items()}
    except Exception:
        pass
    return dict(DEFAULT_COST)


def _col_idx(headers, *keys):
    for i, h in enumerate(headers):
        hl = str(h).lower() if h is not None else ""
        if any(k in hl for k in keys):
            return i
    return -1


def check_excel(workspace):
    print("\n=== Проверка 1: Meal_Program_Plan.xlsx ===")
    costs = load_costs(workspace)
    path = os.path.join(workspace, "Meal_Program_Plan.xlsx")
    if not os.path.exists(path):
        check("Excel-файл существует", False, f"Не найден: {path}")
        check("Recipe_Selection: >=7 строк, все блюда реальны (kulinar) с верными category/difficulty", False, "нет файла")
        check("Weekly_Menu: 5 дней, нет двух подряд одной категории, сложность всех блюд <= 4", False, "нет файла")
        check("Weekly_Menu/Program_Summary: дневная стоимость = ставке категории; total_weekly_cost = sum*50", False, "нет файла")
        return
    check("Excel-файл существует", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Не менее 4 листов", len(sheets) >= 4, f"Найдено {len(sheets)}: {sheets}")
    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    def find_sheet(*keys, contains_all=False):
        for i, s in enumerate(sheets_lower):
            if contains_all:
                if all(k in s for k in keys):
                    return sheets[i]
            else:
                if any(k in s for k in keys):
                    return sheets[i]
        return None

    # --- Survey_Questions ---
    sq_name = find_sheet("survey", "question")
    if sq_name:
        ws = wb[sq_name]
        rows = list(ws.iter_rows(values_only=True))
        data = [r for r in rows[1:] if any(c for c in r)]
        check("Survey_Questions: >=5 строк", len(data) >= 5, f"Найдено {len(data)}")
        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        check("Есть вопрос про категорию/предпочтение блюд",
              any(k in all_text for k in ("катег", "предпоч", "cuisine", "prefer")), all_text[:120])
        check("Есть вопрос про ограничения",
              any(k in all_text for k in ("ограничен", "restrict", "dietary")), all_text[:120])
    else:
        check("Survey_Questions: >=5 строк", False, f"Лист не найден: {sheets}")

    # --- Recipe_Selection (CRITICAL: реальность блюд + верные category/difficulty) ---
    rs_name = find_sheet("recipe", "select", contains_all=True) or find_sheet("recipe")
    recipe_ok = False
    if rs_name:
        ws = wb[rs_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0] if rows else []
        data = [r for r in rows[1:] if any(c for c in r)]
        name_i = _col_idx(headers, "recipe", "name", "блюд", "назван")
        cat_i = _col_idx(headers, "categor", "катег")
        diff_i = _col_idx(headers, "difficult", "сложн")
        if name_i < 0:
            name_i = 0
        check("Recipe_Selection: есть столбец difficulty",
              diff_i >= 0, f"Headers: {headers}")
        matched, mismatched = 0, []
        for r in data:
            nm = r[name_i] if name_i < len(r) else None
            if not nm:
                continue
            canon = lookup_recipe(str(nm))
            if not canon:
                mismatched.append(f"{nm}(не из kulinar)")
                continue
            cat_ok = True
            diff_ok = True
            if cat_i >= 0 and cat_i < len(r) and r[cat_i] not in (None, ""):
                cat_ok = norm_recipe(str(r[cat_i])) == norm_recipe(canon[0])
            if diff_i >= 0 and diff_i < len(r) and r[diff_i] not in (None, ""):
                try:
                    diff_ok = int(float(r[diff_i])) == canon[1]
                except (TypeError, ValueError):
                    diff_ok = False
            if cat_ok and diff_ok:
                matched += 1
            else:
                mismatched.append(f"{nm}(cat_ok={cat_ok},diff_ok={diff_ok})")
        recipe_ok = len(data) >= 7 and not mismatched and matched >= 7
        check("Recipe_Selection: >=7 строк, все блюда реальны (kulinar) с верными category/difficulty",
              recipe_ok, f"rows={len(data)}, matched={matched}, проблемы={mismatched[:5]}")
    else:
        check("Recipe_Selection: >=7 строк, все блюда реальны (kulinar) с верными category/difficulty",
              False, f"Лист не найден: {sheets}")

    # --- Weekly_Menu (CRITICAL: no-consecutive-category + difficulty<=4; стоимость) ---
    wm_name = find_sheet("weekly", "menu")
    menu_categories = []   # категории блюд по дням (для проверки чередования)
    daily_costs = []       # стоимость по дням (для проверки суммы)
    menu_recipe_names = []
    if wm_name:
        ws = wb[wm_name]
        rows = list(ws.iter_rows(values_only=True))
        headers = rows[0] if rows else []
        data = [r for r in rows[1:] if any(c for c in r)]
        check("Weekly_Menu: >=5 дней", len(data) >= 5, f"Найдено {len(data)}")
        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        check("Меню включает Monday", "monday" in all_text)
        check("Меню включает Friday", "friday" in all_text)

        day_i = _col_idx(headers, "day", "день")
        rec_i = _col_idx(headers, "lunch", "recipe", "блюд", "обед")
        cost_i = _col_idx(headers, "cost", "стоим", "цена")
        if rec_i < 0:
            rec_i = 1

        diffs_ok = True
        for r in data[:5]:
            nm = r[rec_i] if rec_i < len(r) else None
            if not nm:
                diffs_ok = False
                menu_categories.append(None)
                continue
            menu_recipe_names.append(str(nm).strip())
            canon = lookup_recipe(str(nm))
            if canon:
                menu_categories.append(canon[0])
                if canon[1] > 4:
                    diffs_ok = False
            else:
                menu_categories.append(None)
                diffs_ok = False
            if cost_i >= 0 and cost_i < len(r):
                daily_costs.append(r[cost_i])
            else:
                daily_costs.append(None)

        # чередование категорий
        no_consec = True
        cats5 = menu_categories[:5]
        for a, b in zip(cats5, cats5[1:]):
            if a is not None and b is not None and a == b:
                no_consec = False
        five_days = len(data) >= 5
        check("Weekly_Menu: 5 дней, нет двух подряд одной категории, сложность всех блюд <= 4",
              five_days and no_consec and diffs_ok and all(c is not None for c in cats5),
              f"days={len(data)}, no_consec={no_consec}, diffs_ok={diffs_ok}, cats={cats5}")
    else:
        check("Weekly_Menu: 5 дней, нет двух подряд одной категории, сложность всех блюд <= 4",
              False, f"Лист не найден: {sheets}")

    # --- Стоимость (CRITICAL): дневная == ставка категории; total = sum*50 ---
    cost_per_day_ok = True
    sum_daily = 0.0
    n_valid = 0
    for cat, val in zip(menu_categories[:5], daily_costs[:5]):
        if cat is None or val is None:
            cost_per_day_ok = False
            continue
        expected = costs.get(norm_recipe(cat))
        try:
            v = float(val)
        except (TypeError, ValueError):
            cost_per_day_ok = False
            continue
        sum_daily += v
        n_valid += 1
        if expected is None or abs(v - expected) > 1e-6:
            cost_per_day_ok = False

    expected_total = sum_daily * SERVINGS
    # читаем total_weekly_cost из Program_Summary
    ps_name = find_sheet("program", "summary")
    summary_total = None
    summary_text = ""
    if ps_name:
        ws = wb[ps_name]
        rows = list(ws.iter_rows(values_only=True))
        data = [r for r in rows[1:] if any(c for c in r)]
        check("Program_Summary: >=4 метрик", len(data) >= 4, f"Найдено {len(data)}")
        summary_text = " ".join(str(c) for r in rows for c in r if c).lower()
        check("Есть total_weekly_cost", "total" in summary_text and ("weekly" in summary_text or "cost" in summary_text))
        for r in data:
            if r and r[0] and "total_weekly_cost" in str(r[0]).lower():
                for c in r[1:]:
                    if isinstance(c, (int, float)):
                        summary_total = float(c)
                        break
                if summary_total is None:
                    # значение могло попасть как строка
                    for c in r[1:]:
                        try:
                            summary_total = float(str(c).replace(",", "."))
                            break
                        except (TypeError, ValueError):
                            continue
                break
    total_ok = (summary_total is not None and n_valid == 5
                and abs(summary_total - expected_total) <= 1e-6)
    check("Weekly_Menu/Program_Summary: дневная стоимость = ставке категории; total_weekly_cost = sum*50",
          cost_per_day_ok and total_ok and n_valid == 5,
          f"per_day_ok={cost_per_day_ok}, total={summary_total}, expected={expected_total}, sum_daily={sum_daily}")

    return menu_recipe_names


def check_forms():
    print("\n=== Проверка 2: Forms (опрос предпочтений) ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()

        def title_matches(title):
            t = (title or "").lower()
            ru = any(k in t for k in ("обед", "предпочт", "питани", "велнес", "программ обед"))
            en = any(k in t for k in ("lunch", "preference", "employee", "meal"))
            # исключаем явный шум
            if "стар" in t or "old" in t:
                return False
            return ru or en

        target = None
        for fid, title in forms:
            if title_matches(title):
                target = (fid, title)
                break
        check("Найдена форма-опрос предпочтений",
              target is not None, f"Формы: {[f[1] for f in forms]}")

        q_count = 0
        q_text = ""
        if target:
            cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (target[0],))
            q_count = cur.fetchone()[0]
            cur.execute("SELECT lower(title) FROM gform.questions WHERE form_id = %s", (target[0],))
            q_text = " ".join(r[0] or "" for r in cur.fetchall())
            check("Опрос покрывает категорию/остроту/ограничения (мягко)",
                  any(k in q_text for k in ("катег", "предпоч", "cuisine")) and
                  any(k in q_text for k in ("остр", "spice", "spicy")) and
                  any(k in q_text for k in ("ограничен", "restrict", "dietary")),
                  q_text[:150])

        check("Forms: опрос предпочтений по обедам с >=5 вопросами",
              target is not None and q_count >= 5, f"вопросов={q_count}")
    except Exception as e:
        check("Forms: опрос предпочтений по обедам с >=5 вопросами", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_teamly(menu_recipe_names):
    print("\n=== Проверка 3: Teamly «База знаний рецептов» ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
        pages = cur.fetchall()

        def title_matches(title):
            t = (title or "").lower()
            ru = ("баз" in t and "знан" in t and "рецепт" in t)
            en = ("recipe" in t and ("knowledge" in t or "base" in t))
            return ru or en

        candidates = [(t, b) for t, b in pages if title_matches(t)]
        check("Создана страница базы знаний рецептов",
              len(candidates) >= 1, f"Заголовки: {[t for t, _ in pages]}")

        body = "\n\n".join(b for _t, b in candidates)
        nb = norm_recipe(body)

        # сколько блюд из недельного меню упомянуто на странице
        menu_norm = [norm_recipe(n) for n in (menu_recipe_names or []) if n]
        if menu_norm:
            menu_hits = sum(1 for n in set(menu_norm) if n and n in nb)
            base_count_ok = menu_hits >= min(5, len(set(menu_norm)))
            detail = f"меню-блюд на странице={menu_hits}/{len(set(menu_norm))}"
        else:
            # запасной вариант: считаем любые каноничные блюда на странице
            menu_hits = sum(1 for cn in CANON if cn and cn in nb)
            base_count_ok = menu_hits >= 5
            detail = f"каноничных блюд на странице={menu_hits}"

        check("Teamly: страница «База знаний рецептов» с записями по блюдам недельного меню (>=5)",
              bool(candidates) and base_count_ok, detail)
    except Exception as e:
        check("Teamly: страница «База знаний рецептов» с записями по блюдам недельного меню (>=5)",
              False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Проверка 4: Письмо персоналу ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%all_staff%%'
               OR subject ILIKE '%%lunch program%%'
               OR subject ILIKE '%%lunch%%survey%%'
               OR subject ILIKE '%%employee%%lunch%%'
        """)
        emails = cur.fetchall()
        # отбросим шумовое письмо про парковку
        emails = [e for e in emails if not (e[1] and "парков" in str(e[1]).lower())]
        check("Письмо про программу обедов отправлено", len(emails) >= 1, "Подходящее письмо не найдено")
        if emails:
            subject = str(emails[0][1]).lower() if emails[0][1] else ""
            check("Тема письма про lunch/survey",
                  "lunch" in subject or "survey" in subject or "meal" in subject,
                  f"Subject: {emails[0][1]}")
            body = str(emails[0][3] or "").lower()
            check("В теле письма упомянуты меню/разнообразие и опрос",
                  any(k in body for k in ("меню", "разнообраз", "menu", "variety")) and
                  any(k in body for k in ("опрос", "survey")),
                  body[:150])
    except Exception as e:
        check("Email check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Проверка 5: menu_planner.py ===")
    path = os.path.join(workspace, "menu_planner.py")
    check("menu_planner.py существует", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Обратная валидация ===")
    path = os.path.join(workspace, "Meal_Program_Plan.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        check("Нет отрицательных значений в Excel", False,
                              f"Найдено {cell} в листе {sheet_name}")
                        return
        check("Нет отрицательных значений в Excel", True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL KULINAR FORMS EXCEL TEAMLY EMAIL - EVALUATION")
    print("=" * 70)

    menu_recipes = check_excel(args.agent_workspace)
    check_forms()
    check_teamly(menu_recipes or [])
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: проверки не выполнены.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"CRITICAL FAIL: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
