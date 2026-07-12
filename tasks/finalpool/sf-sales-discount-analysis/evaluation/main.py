"""Evaluation for sf-sales-discount-analysis (ClickHouse fork).

Two-gate evaluation:
  - CRITICAL_CHECKS: semantic core deliverables. Any critical failure =>
    immediate FAIL (sys.exit(1)), before the accuracy gate.
      * all four-band coverage in 'Discount Analysis' with Orders + Revenue
        matching the groundtruth within tolerance (any missing expected band fails);
      * Summary internal consistency: Total_Revenue == No_Discount_Revenue +
        Discounted_Revenue, and Total_Revenue matches the groundtruth total;
      * an email was actually sent to finance@company.com with subject
        'Discount Impact Analysis' (query PG email.messages).
  - Non-critical structural checks (sheets/headers present, 2-decimal rounding,
    alphabetical band order) feed an overall accuracy>=70 gate.

Discount bands are NUMERIC (No Discount / 1-10% / 11-20% / 20%+), independent of the
central russification of SALES_DW realia, so the groundtruth Excel stays English-numeric.
"""
import argparse
import json
import os
import sys

import openpyxl

try:
    import psycopg2
except ImportError:
    psycopg2 = None

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def is_2dp(v):
    f = to_float(v)
    if f is None:
        return False
    return abs(round(f, 2) - f) < 1e-6


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_to_lookup(rows):
    """key (lowercased first col) -> row list, skipping header."""
    lookup = {}
    for row in (rows[1:] if rows and len(rows) > 1 else []):
        if row and row[0] is not None:
            lookup[str(row[0]).strip().lower()] = row
    return lookup


def parse_recipients(to_addr):
    if not to_addr:
        return []
    if isinstance(to_addr, list):
        return [str(r).strip().lower() for r in to_addr]
    if isinstance(to_addr, str):
        try:
            parsed = json.loads(to_addr)
            if isinstance(parsed, list):
                return [str(r).strip().lower() for r in parsed]
        except (json.JSONDecodeError, TypeError):
            pass
        return [str(to_addr).strip().lower()]
    return [str(to_addr).strip().lower()]


def check_email_sent():
    """Return True if an email to finance@company.com with the expected subject was sent."""
    if psycopg2 is None:
        print("    (psycopg2 unavailable - cannot verify email)")
        return False
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        print(f"    (PG connect failed: {e})")
        return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr FROM email.messages")
        rows = cur.fetchall()
    except Exception as e:
        print(f"    (email query failed: {e})")
        conn.close()
        return False
    conn.close()
    for subj, to_addr in rows:
        recipients = parse_recipients(to_addr)
        subj_l = (subj or "").lower()
        if "finance@company.com" in recipients and "discount impact analysis" in subj_l:
            return True
    return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Discount_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Discount_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # checks: list of (name, passed, is_critical)
    checks = []

    def record(name, passed, critical=False):
        checks.append((name, bool(passed), critical))
        status = "PASS" if passed else "FAIL"
        tag = " [CRITICAL]" if critical else ""
        print(f"  [{status}] {name}{tag}")

    # ----- Sheet: Discount Analysis ------------------------------------
    a_rows = load_sheet_rows(agent_wb, "Discount Analysis")
    g_rows = load_sheet_rows(gt_wb, "Discount Analysis")

    record("Discount Analysis: sheet present", a_rows is not None, critical=False)

    if a_rows is None or g_rows is None:
        if g_rows is None:
            print("  (groundtruth missing Discount Analysis sheet)")
        record("Discount Analysis: all expected bands with Orders+Revenue correct",
               False, critical=True)
    else:
        # Header columns present (non-critical)
        header = [str(c).strip() if c is not None else "" for c in a_rows[0]] if a_rows else []
        record("Discount Analysis: header columns present",
               header[:4] == ["Discount_Band", "Orders", "Revenue", "Avg_Order_Value"],
               critical=False)

        a_lookup = rows_to_lookup(a_rows)
        g_lookup = rows_to_lookup(g_rows)

        # CRITICAL: every groundtruth band must be present and match Orders/Revenue.
        band_ok = True
        band_detail = []
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                band_ok = False
                band_detail.append(f"missing band {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 10):
                band_ok = False
                band_detail.append(f"{key}.Orders {a_row[1]} vs {g_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 100.0):
                band_ok = False
                band_detail.append(f"{key}.Revenue {a_row[2]} vs {g_row[2]}")
        if band_detail:
            print(f"      detail: {'; '.join(band_detail[:5])}")
        record("Discount Analysis: all expected bands with Orders+Revenue correct",
               band_ok, critical=True)

        # Non-critical: Avg_Order_Value close
        avg_ok = True
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row and len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 2.0):
                    avg_ok = False
        record("Discount Analysis: Avg_Order_Value correct", avg_ok, critical=False)

        # Non-critical: Revenue rounded to 2 decimals
        rev_2dp = all(
            is_2dp(r[2]) for r in (a_rows[1:] if len(a_rows) > 1 else [])
            if r and r[0] is not None and len(r) > 2
        )
        record("Discount Analysis: Revenue rounded to 2 decimals", rev_2dp, critical=False)

        # Non-critical: bands sorted alphabetically by Discount_Band
        band_names = [str(r[0]).strip() for r in (a_rows[1:] if len(a_rows) > 1 else [])
                      if r and r[0] is not None]
        record("Discount Analysis: bands sorted alphabetically",
               band_names == sorted(band_names, key=lambda s: s.lower()),
               critical=False)

    # ----- Sheet: Summary ----------------------------------------------
    a_srows = load_sheet_rows(agent_wb, "Summary")
    g_srows = load_sheet_rows(gt_wb, "Summary")

    record("Summary: sheet present", a_srows is not None, critical=False)

    if a_srows is None or g_srows is None:
        if g_srows is None:
            print("  (groundtruth missing Summary sheet)")
        record("Summary: Total_Revenue == No_Discount + Discounted and matches GT",
               False, critical=True)
    else:
        a_slook = rows_to_lookup(a_srows)
        g_slook = rows_to_lookup(g_srows)

        def aval(k):
            r = a_slook.get(k.lower())
            return to_float(r[1]) if r and len(r) > 1 else None

        def gval(k):
            r = g_slook.get(k.lower())
            return to_float(r[1]) if r and len(r) > 1 else None

        # Non-critical: each metric within tolerance vs GT
        metrics_ok = True
        for key, g_row in g_slook.items():
            a_row = a_slook.get(key)
            if a_row is None or len(a_row) < 2 or len(g_row) < 2:
                metrics_ok = False
                continue
            if not num_close(a_row[1], g_row[1], 200.0):
                metrics_ok = False
        record("Summary: metric values match groundtruth", metrics_ok, critical=False)

        # CRITICAL: internal consistency + total matches GT total.
        total_rev = aval("Total_Revenue")
        no_disc = aval("No_Discount_Revenue")
        disc = aval("Discounted_Revenue")
        gt_total = gval("Total_Revenue")
        internal_ok = (
            total_rev is not None and no_disc is not None and disc is not None
            and num_close(total_rev, no_disc + disc, 1.0)
            and gt_total is not None and num_close(total_rev, gt_total, 200.0)
        )
        record("Summary: Total_Revenue == No_Discount + Discounted and matches GT",
               internal_ok, critical=True)

    # ----- CRITICAL: email actually sent -------------------------------
    print("  Checking email (PG email.messages)...")
    email_ok = check_email_sent()
    record("Email sent to finance@company.com with subject 'Discount Impact Analysis'",
           email_ok, critical=True)

    # ----- Gates -------------------------------------------------------
    critical_failed = [n for n, p, c in checks if c and not p]
    total = len(checks)
    passed = sum(1 for _, p, _ in checks if p)
    accuracy = (passed / total * 100.0) if total else 0.0

    print(f"\n  Critical failures: {len(critical_failed)}")
    for n in critical_failed:
        print(f"    - {n}")
    print(f"  Accuracy: {passed}/{total} = {accuracy:.1f}%")

    if critical_failed:
        print("\n=== RESULT: FAIL (critical check failed) ===")
        sys.exit(1)

    if accuracy < 70.0:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)

    print("\n=== RESULT: PASS ===")
    sys.exit(0)


if __name__ == "__main__":
    main()
