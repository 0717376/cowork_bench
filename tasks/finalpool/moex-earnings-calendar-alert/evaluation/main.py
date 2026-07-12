"""Evaluation для yf-earnings-calendar-alert (RU / moex-finance).

Агент строит Earnings_Analysis.xlsx по данным:
  - mock-портала календаря отчётностей (http://localhost:30225/index.html);
  - MCP moex-finance (схема moex.*, годовой income_stmt за 2024 и 2025).

Тикеры портфеля: 3 эмитента с предстоящими отчётностями (SBER.ME, GAZP.ME,
TCSG.ME) и 2 holdings без отчётностей (LKOH.ME, MGNT.ME).

Эталонные значения пересчитаны от moex-сида (db/zzz_moex_after_init.sql) и
mock-страницы:

  Historical_Avg_EPS = среднее годовых Diluted EPS 2024+2025, округл. до 2 знаков
    SBER.ME: (44.352 + 50.4)/2   = 47.38
    GAZP.ME: (77.8888 + 88.51)/2 = 83.20
    TCSG.ME: (277.112 + 314.9)/2 = 296.01

  Expected_EPS (с mock-страницы): SBER 52.00, GAZP 78.00, TCSG 300.00
  Surprise_Trend (правило 5%):
    SBER: (52-47.38)/47.38   = +9.75%  -> Above
    GAZP: (78-83.20)/83.20   = -6.25%  -> Below
    TCSG: (300-296.01)/296.01 = +1.35% -> In Line

  Avg_Expected_EPS = (52+78+300)/3 = 143.33
  Stocks_Reporting = 3
  Stocks_Above_Historical = 1 (только SBER)
  Earliest_Report = 2026-04-21 (TCSG), Latest_Report = 2026-04-29 (SBER)

CRITICAL_CHECKS (семантические): любой их провал => общий FAIL независимо от
accuracy. Структурные проверки (наличие листов/столбцов, число строк) — не
критические.
"""
import argparse
import os
import sys

import psycopg2


# ----- Эталонные значения, пересчитанные от сида и mock-страницы -----
REPORTING = ["GAZP.ME", "SBER.ME", "TCSG.ME"]  # алфавитный порядок

HIST_AVG = {"SBER.ME": 47.38, "GAZP.ME": 83.20, "TCSG.ME": 296.01}
EXPECTED_EPS = {"SBER.ME": 52.00, "GAZP.ME": 78.00, "TCSG.ME": 300.00}
SURPRISE = {"SBER.ME": "above", "GAZP.ME": "below", "TCSG.ME": "in line"}

AVG_EXPECTED_EPS = 143.33
STOCKS_REPORTING = 3
STOCKS_ABOVE_HIST = 1
EARLIEST_REPORT = "2026-04-21"
LATEST_REPORT = "2026-04-29"

# Diluted EPS по периодам (из moex-сида) для листа Financial Trends
DILUTED_EPS = {
    ("SBER.ME", "2024"): 44.352, ("SBER.ME", "2025"): 50.4,
    ("GAZP.ME", "2024"): 77.8888, ("GAZP.ME", "2025"): 88.51,
    ("TCSG.ME", "2024"): 277.112, ("TCSG.ME", "2025"): 314.9,
}

EMAIL_SUBJECT = "Earnings Season Alert - Q1 2026"

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = False

CRITICAL_CHECKS = {
    "Earnings Calendar: Historical_Avg_EPS совпадает со средним годовых Diluted EPS",
    "Earnings Calendar: Surprise_Trend корректен по правилу 5%",
    "Alert Summary: Stocks_Reporting/Earliest/Latest/Avg_Expected_EPS верны",
    "Email: тема и упоминание всех 3 эмитентов",
    "Financial Trends: Diluted_EPS по периодам совпадают с moex-сидом",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        tag = " (CRITICAL)" if name in CRITICAL_CHECKS else ""
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED = True


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def norm_sym(v):
    return str(v).strip().upper()


def date_str(v):
    s = str(v).strip()
    return s[:10]


def period_year(v):
    """Извлечь год из значения Period (напр. '2025', '2025-12-31', '2025 FY')."""
    import re
    m = re.search(r"(20\d{2})", str(v))
    return m.group(1) if m else None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    import openpyxl
    path = os.path.join(agent_workspace, "Earnings_Analysis.xlsx")
    if not os.path.exists(path):
        record("Earnings_Analysis.xlsx присутствует", False, path)
        record("Earnings Calendar: Historical_Avg_EPS совпадает со средним годовых Diluted EPS", False, "нет файла")
        record("Earnings Calendar: Surprise_Trend корректен по правилу 5%", False, "нет файла")
        record("Alert Summary: Stocks_Reporting/Earliest/Latest/Avg_Expected_EPS верны", False, "нет файла")
        record("Financial Trends: Diluted_EPS по периодам совпадают с moex-сидом", False, "нет файла")
        return
    record("Earnings_Analysis.xlsx присутствует", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # ---------------- Earnings Calendar ----------------
    rows = load_sheet_rows(wb, "Earnings Calendar")
    record("Лист 'Earnings Calendar' присутствует", rows is not None)
    if rows is None:
        record("Earnings Calendar: Historical_Avg_EPS совпадает со средним годовых Diluted EPS", False, "нет листа")
        record("Earnings Calendar: Surprise_Trend корректен по правилу 5%", False, "нет листа")
    else:
        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        col = {h: i for i, h in enumerate(header)}
        record("Earnings Calendar: заголовки на месте",
               all(h in col for h in ["symbol", "name", "earnings_date",
                                      "expected_eps", "historical_avg_eps", "surprise_trend"]),
               f"header={header}")

        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        record("Earnings Calendar: ровно 3 строки эмитентов",
               len(data_rows) == 3, f"строк={len(data_rows)}")

        by_sym = {}
        for r in data_rows:
            by_sym[norm_sym(r[0])] = r

        # отсутствующие тикеры
        missing = [s for s in REPORTING if s not in by_sym]
        record("Earnings Calendar: присутствуют SBER.ME/GAZP.ME/TCSG.ME",
               not missing, f"нет: {missing}")

        # CRITICAL: Historical_Avg_EPS
        hist_errs = []
        ci = col.get("historical_avg_eps")
        for sym in REPORTING:
            r = by_sym.get(sym)
            if r is None or ci is None or len(r) <= ci:
                hist_errs.append(f"{sym}: нет значения")
                continue
            if not num_close(r[ci], HIST_AVG[sym], abs_tol=0.5, rel_tol=0.02):
                hist_errs.append(f"{sym}: {r[ci]} != {HIST_AVG[sym]}")
        record("Earnings Calendar: Historical_Avg_EPS совпадает со средним годовых Diluted EPS",
               not hist_errs, "; ".join(hist_errs))

        # CRITICAL: Surprise_Trend
        trend_errs = []
        ti = col.get("surprise_trend")
        for sym in REPORTING:
            r = by_sym.get(sym)
            if r is None or ti is None or len(r) <= ti:
                trend_errs.append(f"{sym}: нет метки")
                continue
            got = str(r[ti]).strip().lower()
            if got != SURPRISE[sym]:
                trend_errs.append(f"{sym}: '{r[ti]}' != '{SURPRISE[sym]}'")
        record("Earnings Calendar: Surprise_Trend корректен по правилу 5%",
               not trend_errs, "; ".join(trend_errs))

        # NON-critical: Expected_EPS совпадает с mock-страницей
        ei = col.get("expected_eps")
        exp_errs = []
        for sym in REPORTING:
            r = by_sym.get(sym)
            if r is None or ei is None or len(r) <= ei:
                exp_errs.append(f"{sym}: нет Expected_EPS")
                continue
            if not num_close(r[ei], EXPECTED_EPS[sym], abs_tol=0.1, rel_tol=0.02):
                exp_errs.append(f"{sym}: {r[ei]} != {EXPECTED_EPS[sym]}")
        record("Earnings Calendar: Expected_EPS совпадает с порталом",
               not exp_errs, "; ".join(exp_errs))

        # NON-critical: сортировка по Symbol A-Z
        syms_order = [norm_sym(r[0]) for r in data_rows]
        record("Earnings Calendar: отсортировано по Symbol (A-Z)",
               syms_order == sorted(syms_order), f"{syms_order}")

    # ---------------- Financial Trends ----------------
    rows2 = load_sheet_rows(wb, "Financial Trends")
    record("Лист 'Financial Trends' присутствует", rows2 is not None)
    if rows2 is None:
        record("Financial Trends: Diluted_EPS по периодам совпадают с moex-сидом", False, "нет листа")
    else:
        header2 = [str(c).strip().lower() if c is not None else "" for c in rows2[0]]
        col2 = {h: i for i, h in enumerate(header2)}
        record("Financial Trends: заголовки Symbol/Period/Diluted_EPS/Total_Revenue",
               all(h in col2 for h in ["symbol", "period", "diluted_eps", "total_revenue"]),
               f"header={header2}")
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        record("Financial Trends: ровно 6 строк (по 2 годовых на эмитента)",
               len(data2) == 6, f"строк={len(data2)}")

        si, pi, di = col2.get("symbol"), col2.get("period"), col2.get("diluted_eps")
        ft_errs = []
        seen = set()
        if None in (si, pi, di):
            ft_errs.append("нет нужных столбцов")
        else:
            for r in data2:
                sym = norm_sym(r[si]) if len(r) > si else ""
                yr = period_year(r[pi]) if len(r) > pi else None
                if sym not in HIST_AVG or yr not in ("2024", "2025"):
                    continue
                key = (sym, yr)
                seen.add(key)
                exp = DILUTED_EPS.get(key)
                got = r[di] if len(r) > di else None
                if exp is None:
                    continue
                if not num_close(got, exp, abs_tol=0.5, rel_tol=0.02):
                    ft_errs.append(f"{sym} {yr}: {got} != {exp}")
            for key in DILUTED_EPS:
                if key not in seen:
                    ft_errs.append(f"нет строки {key[0]} {key[1]}")
        record("Financial Trends: Diluted_EPS по периодам совпадают с moex-сидом",
               not ft_errs, "; ".join(ft_errs))

    # ---------------- Alert Summary ----------------
    rows3 = load_sheet_rows(wb, "Alert Summary")
    record("Лист 'Alert Summary' присутствует", rows3 is not None)
    if rows3 is None:
        record("Alert Summary: Stocks_Reporting/Earliest/Latest/Avg_Expected_EPS верны", False, "нет листа")
    else:
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        lookup = {str(r[0]).strip().lower(): (r[1] if len(r) > 1 else None) for r in data3}
        summ_errs = []
        if not num_close(lookup.get("stocks_reporting"), STOCKS_REPORTING, abs_tol=0, rel_tol=0):
            summ_errs.append(f"Stocks_Reporting={lookup.get('stocks_reporting')} != 3")
        if date_str(lookup.get("earliest_report")) != EARLIEST_REPORT:
            summ_errs.append(f"Earliest_Report={lookup.get('earliest_report')} != {EARLIEST_REPORT}")
        if date_str(lookup.get("latest_report")) != LATEST_REPORT:
            summ_errs.append(f"Latest_Report={lookup.get('latest_report')} != {LATEST_REPORT}")
        if not num_close(lookup.get("avg_expected_eps"), AVG_EXPECTED_EPS, abs_tol=0.5, rel_tol=0.02):
            summ_errs.append(f"Avg_Expected_EPS={lookup.get('avg_expected_eps')} != {AVG_EXPECTED_EPS}")
        record("Alert Summary: Stocks_Reporting/Earliest/Latest/Avg_Expected_EPS верны",
               not summ_errs, "; ".join(summ_errs))

        # NON-critical: Stocks_Above_Historical
        record("Alert Summary: Stocks_Above_Historical == 1",
               num_close(lookup.get("stocks_above_historical"), STOCKS_ABOVE_HIST, abs_tol=0, rel_tol=0),
               f"={lookup.get('stocks_above_historical')}")


def check_email():
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432,
                                dbname="cowork_gym", user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, COALESCE(body_text, ''), COALESCE(body_html, '')
            FROM email.messages
            WHERE to_addr::text ILIKE '%investment-team@company.com%'
            ORDER BY id DESC LIMIT 10
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email: тема и упоминание всех 3 эмитентов", False, f"DB error: {e}")
        return

    if not rows:
        record("Email: тема и упоминание всех 3 эмитентов", False,
               "нет письма на investment-team@company.com")
        return

    # NON-critical: письмо существует
    record("Email: письмо на investment-team@company.com отправлено", True)

    # CRITICAL: точная тема и упоминание всех 3 тикеров
    detail = []
    ok = False
    for subj, btext, bhtml in rows:
        body = f"{btext}\n{bhtml}"
        subj_ok = (str(subj).strip() == EMAIL_SUBJECT)
        tickers_ok = all(t in body for t in REPORTING) or \
            all(t.split(".")[0] in body for t in REPORTING)
        if subj_ok and tickers_ok:
            ok = True
            break
        if not subj_ok:
            detail.append(f"тема='{subj}'")
        if not tickers_ok:
            detail.append("упомянуты не все тикеры")
    record("Email: тема и упоминание всех 3 эмитентов", ok, "; ".join(detail[:4]))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or "."

    print("  Проверка Excel...")
    check_excel(agent_ws)

    print("  Проверка email...")
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    if CRITICAL_FAILED:
        print("=== RESULT: FAIL (провалена критическая проверка) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print("=== RESULT: FAIL (accuracy < 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
