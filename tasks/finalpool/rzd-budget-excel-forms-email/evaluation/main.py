"""
Evaluation for rzd-budget-excel-forms-email task (RU stack: rzd + forms).

Сценарий: 8 сотрудников едут «Сапсаном» Москва — Санкт-Петербург 10.03.2026.
Данные РЖД (seeded): поезда 752А/754А/756А/758А, классы Эконом=5500,
Эконом+=7500, Бизнес=14000 ₽.

Проверки:
1. Travel_Budget.xlsx с листами Seat_Options, Budget_Scenarios, Summary.
2. Корректные цены и итоги для 8 человек (RUB).
3. Опрос «Опрос предпочтений по командировке» с 3 вопросами.
4. Письмо на finance@company.com с правильной темой и телом.

Критические чеки (CRITICAL_CHECKS): любой их fail => задача FAIL, независимо
от общей accuracy. Это семантические чеки сути задачи (правильные числа из РЖД,
ключевой результат), а НЕ структура (наличие листа/колонки).
"""
import json
import os
import sys
import unicodedata
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# --- Эталонные значения (из rzd seed на 2026-03-10) ---
PRICE_EC = 5500.0      # Эконом
PRICE_ECP = 7500.0     # Эконом+
PRICE_BIZ = 14000.0    # Бизнес
TOTAL_BUDGET = 44000.0     # 5500 * 8
TOTAL_STANDARD = 60000.0   # 7500 * 8
TOTAL_PREMIUM = 112000.0   # 14000 * 8
PRICE_DIFF = 68000.0       # 112000 - 44000

# Критические чеки — по строке name, как в record()
CRITICAL_CHECKS = {
    "Seat_Options: цены 5500, 7500, 14000 ₽ присутствуют",
    "Budget_Scenarios: итоги 44000 / 60000 / 112000 ₽ для 8 человек",
    "Summary: разница 68000 ₽ И сценарии Budget/Premium как дешёвый/дорогой",
    "Опрос: 3 вопроса — отправление, класс, открытый текст",
    "Письмо finance@company.com: тело содержит 44000/60000/112000 и 68000",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def normalize(s):
    """NFKD + кириллица→латиница-двойники, нижний регистр.
    Только для смешанного кир/лат сопоставления ID (напр. '752А' vs '752A')."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    table = str.maketrans({
        "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H",
        "О": "O", "Р": "P", "С": "C", "Т": "T", "У": "Y", "Х": "X",
        "а": "a", "е": "e", "о": "o", "р": "p", "с": "c", "х": "x",
    })
    return s.translate(table).lower()


def collect_numerics(rows):
    out = []
    for r in rows:
        for c in r:
            try:
                out.append(float(c))
            except (TypeError, ValueError):
                pass
    return out


def has_num(vals, target, tol=1.0):
    return any(abs(v - target) < tol for v in vals)


def check_excel(agent_workspace):
    print("\n=== Check 1: Travel_Budget.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Travel_Budget.xlsx")
    if not os.path.exists(xlsx_path):
        record("Travel_Budget.xlsx exists", False, f"Not found at {xlsx_path}")
        # Критические чеки Excel также проваливаются
        record("Seat_Options: цены 5500, 7500, 14000 ₽ присутствуют", False, "no file")
        record("Budget_Scenarios: итоги 44000 / 60000 / 112000 ₽ для 8 человек", False, "no file")
        record("Summary: разница 68000 ₽ И сценарии Budget/Premium как дешёвый/дорогой", False, "no file")
        return
    record("Travel_Budget.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        record("Seat_Options: цены 5500, 7500, 14000 ₽ присутствуют", False, "unreadable")
        record("Budget_Scenarios: итоги 44000 / 60000 / 112000 ₽ для 8 человек", False, "unreadable")
        record("Summary: разница 68000 ₽ И сценарии Budget/Premium как дешёвый/дорогой", False, "unreadable")
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # ---- Seat_Options ----
    if "seat_options" not in sheet_names_lower:
        record("Seat_Options sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Seat_Options: цены 5500, 7500, 14000 ₽ присутствуют", False, "no sheet")
    else:
        record("Seat_Options sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("seat_options")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Seat_Options has at least 4 rows", len(data_rows) >= 4, f"Found {len(data_rows)}")

        # Коды поездов: 752А/754А (кириллическая А) — сопоставляем через normalize()
        all_norm = " ".join(normalize(c) for r in rows for c in r if c)
        has_codes = ("752a" in all_norm) and ("754a" in all_norm)
        record("Seat_Options содержит коды поездов 752А и 754А", has_codes, all_norm[:200])

        numeric_vals = collect_numerics(data_rows)
        has_prices = (has_num(numeric_vals, PRICE_EC) and
                      has_num(numeric_vals, PRICE_ECP) and
                      has_num(numeric_vals, PRICE_BIZ))
        # CRITICAL
        record("Seat_Options: цены 5500, 7500, 14000 ₽ присутствуют", has_prices,
               f"Numerics: {numeric_vals[:30]}")

    # ---- Budget_Scenarios ----
    if "budget_scenarios" not in sheet_names_lower:
        record("Budget_Scenarios sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Budget_Scenarios: итоги 44000 / 60000 / 112000 ₽ для 8 человек", False, "no sheet")
    else:
        record("Budget_Scenarios sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("budget_scenarios")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        record("Budget_Scenarios has 3 rows", len(data_rows2) == 3, f"Found {len(data_rows2)}")

        numeric_vals2 = collect_numerics(data_rows2)
        has_totals = (has_num(numeric_vals2, TOTAL_BUDGET) and
                      has_num(numeric_vals2, TOTAL_STANDARD) and
                      has_num(numeric_vals2, TOTAL_PREMIUM))
        # CRITICAL
        record("Budget_Scenarios: итоги 44000 / 60000 / 112000 ₽ для 8 человек", has_totals,
               f"Numerics: {numeric_vals2[:30]}")

    # ---- Summary ----
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Summary: разница 68000 ₽ И сценарии Budget/Premium как дешёвый/дорогой", False, "no sheet")
    else:
        record("Summary sheet exists", True)
        ws3 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows3 = list(ws3.iter_rows(values_only=True))

        nums3 = collect_numerics(rows3)
        has_diff = has_num(nums3, PRICE_DIFF, tol=5.0)
        record("Summary содержит разницу 68000 ₽", has_diff, f"Numerics: {nums3[:30]}")

        # Семантика Summary: дешёвый = Budget, дорогой = Premium.
        # Ищем по русскому/английскому тексту (исходный .lower(), не normalize()).
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c is not None).lower()
        cheapest_ok = ("budget" in all_text3) or ("бюджет" in all_text3) or ("эконом" in all_text3)
        premium_ok = ("premium" in all_text3) or ("премиум" in all_text3) or ("бизнес" in all_text3)
        # CRITICAL: разница верна И корректно названы самый дешёвый/дорогой сценарии
        record("Summary: разница 68000 ₽ И сценарии Budget/Premium как дешёвый/дорогой",
               has_diff and cheapest_ok and premium_ok,
               f"diff={has_diff} cheapest={cheapest_ok} premium={premium_ok}; text={all_text3[:200]}")


def check_gform():
    print("\n=== Check 2: Опрос предпочтений ===")

    crit_q = "Опрос: 3 вопроса — отправление, класс, открытый текст"
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Опрос предпочтений по командировке создан", False, str(e))
        record(crit_q, False, "no db")
        return

    cur = conn.cursor()
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE title ILIKE '%командиров%'
           OR title ILIKE '%предпочтени%'
           OR title ILIKE '%travel preference%'
           OR title ILIKE '%business trip%'
    """)
    forms = cur.fetchall()
    record("Опрос предпочтений по командировке создан", len(forms) >= 1,
           f"Found forms: {[f[1] for f in forms]}")

    if not forms:
        record(crit_q, False, "no form")
        cur.close()
        conn.close()
        return

    form_id = forms[0][0]
    cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (form_id,))
    q_count = cur.fetchone()[0]
    record("Форма содержит ровно 3 вопроса", q_count == 3, f"Found {q_count}")

    cur.execute(
        "SELECT title, question_type FROM gform.questions WHERE form_id = %s ORDER BY position",
        (form_id,),
    )
    questions = cur.fetchall()
    cur.close()
    conn.close()

    def qtitle(q):
        return (q[0] or "").lower()

    # Вопрос про время отправления (RU/EN-ключи в исходном .lower())
    has_departure_q = any(
        ("отправлени" in qtitle(q)) or ("время" in qtitle(q))
        or ("departure" in qtitle(q)) or ("time" in qtitle(q))
        for q in questions
    )
    # Вопрос про класс обслуживания / место (RU: 'класс', 'место'; EN: 'seat', 'class')
    has_seat_q = any(
        ("класс" in qtitle(q)) or ("место" in qtitle(q))
        or ("seat" in qtitle(q)) or ("class" in qtitle(q))
        for q in questions
    )
    # Открытый текстовый вопрос. RU forms MCP пишет question_type='textQuestion'.
    has_text_q = any(
        (q[1] or "") in ("textQuestion", "TEXT", "PARAGRAPH", "SHORT_ANSWER")
        for q in questions
    )

    record("Форма содержит вопрос о времени отправления", has_departure_q, f"Questions: {questions}")
    record("Форма содержит вопрос о классе обслуживания", has_seat_q, f"Questions: {questions}")
    record("Форма содержит открытый текстовый вопрос", has_text_q, f"Questions: {questions}")

    # CRITICAL
    record(crit_q, (q_count == 3 and has_departure_q and has_seat_q and has_text_q),
           f"count={q_count} dep={has_departure_q} seat={has_seat_q} text={has_text_q}")


def check_email():
    print("\n=== Check 3: Письмо на finance@company.com ===")

    crit_e = "Письмо finance@company.com: тело содержит 44000/60000/112000 и 68000"
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Сводное письмо по бюджету отправлено", False, str(e))
        record(crit_e, False, "no db")
        return

    cur = conn.cursor()
    cur.execute("""
        SELECT to_addr::text, subject, body_text FROM email.messages
        WHERE subject ILIKE '%бюджет%'
           OR subject ILIKE '%командиров%'
           OR subject ILIKE '%петербург%'
           OR subject ILIKE '%budget%'
           OR subject ILIKE '%conference%'
    """)
    messages = cur.fetchall()
    cur.close()
    conn.close()

    record("Сводное письмо по бюджету отправлено", len(messages) >= 1,
           f"Found {len(messages)} matching emails")

    if not messages:
        record("Письмо отправлено на finance@company.com", False, "no email")
        record(crit_e, False, "no email")
        return

    # Выбираем письмо, адресованное finance@company.com (а не первое попавшееся)
    target = None
    for to_raw, subj, body in messages:
        to_str = (str(to_raw).lower() if to_raw else "")
        if "finance@company.com" in to_str:
            target = (to_str, subj or "", body or "")
            break

    record("Письмо отправлено на finance@company.com", target is not None,
           f"To-addrs: {[str(m[0])[:80] for m in messages]}")

    if target is None:
        record(crit_e, False, "no finance@ recipient")
        return

    body = target[2]
    # Нормализуем тело: убираем разделители тысяч, чтобы '44 000'/'44,000' совпали с '44000'
    body_compact = (body or "").replace(" ", "").replace(" ", "").replace(",", "").replace(".", "")
    has_b = "44000" in body_compact
    has_s = "60000" in body_compact
    has_p = "112000" in body_compact
    has_d = "68000" in body_compact
    # CRITICAL: тело содержит все три итога и разницу
    record(crit_e, has_b and has_s and has_p and has_d,
           f"44000={has_b} 60000={has_s} 112000={has_p} 68000={has_d}; body={body[:200]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": CRITICAL_FAILED,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"CRITICAL FAIL: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL (accuracy < 70%)")
        sys.exit(1)


if __name__ == "__main__":
    main()
