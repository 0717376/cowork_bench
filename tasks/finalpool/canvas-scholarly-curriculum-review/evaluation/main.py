"""Evaluation for canvas-scholarly-curriculum-review.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Canvas is READ-ONLY. Course names come from the russified Canvas data
(Основы финансов, Прикладная аналитика, Осень 2013); live counts (~22 courses,
~383 students) are hardcoded with tolerances. Course-name substrings are matched
in Russian; the pedagogy paper titles stay English (real scholarly titles).
"""
import argparse
import json
import os
import sys

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": os.environ.get("PGUSER", "eigent"),
    "password": os.environ.get("PGPASSWORD", "camel"),
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
# These reflect SUBSTANCE (correct compliance logic, honest data ingestion,
# summary math consistency, correct scholarly retrieval, real report content).
CRITICAL_CHECKS = {
    "Основы финансов (Осень 2013) is Non-compliant (No)",
    "Applied Analytics (Fall 2013) student count honest (~383) and ratio == students/faculty",
    "Summary math consistent (total ~22, compliant+non-compliant==total, rate==round)",
    "Literature Support: 5 topics with titles matching injected pedagogy papers (not noise)",
    "Teamly 'Accreditation Review Report' page exists with compliance rate and a non-compliant course",
}

# The 5 injected pedagogy papers (English titles preserved; eval-grepped).
PEDAGOGY_PAPER_TITLES = [
    "active learning increases student performance",
    "classroom assessment techniques",
    "evaluation of evidence-based practices in online learning",
    "understanding by design",
    "student engagement and student learning",
]
# Noise papers that must NOT be counted as valid literature support.
NOISE_PAPER_TITLES = [
    "machine learning approaches for natural language processing",
    "blockchain technology in supply chain management",
    "climate change impact on agricultural productivity",
]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def num_close(a, b, abs_tol=1.0, rel_tol=0.05):
    a, b = safe_float(a), safe_float(b)
    if a is None or b is None:
        return False
    if abs(a - b) <= abs_tol:
        return True
    if b != 0 and abs(a - b) / abs(b) <= rel_tol:
        return True
    return False


def load_sheet_rows(wb, sheet_name):
    target = sheet_name.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_conn():
    import psycopg2

    return psycopg2.connect(**DB_CONFIG)


def check_excel(agent_workspace):
    import openpyxl

    path = os.path.join(agent_workspace, "Curriculum_Review.xlsx")
    if not os.path.exists(path):
        check("Curriculum_Review.xlsx exists", False, "not found")
        # Mark dependent critical checks as failed.
        check("Основы финансов (Осень 2013) is Non-compliant (No)", False, "no excel")
        check(
            "Applied Analytics (Fall 2013) student count honest (~383) and ratio == students/faculty",
            False, "no excel",
        )
        check(
            "Summary math consistent (total ~22, compliant+non-compliant==total, rate==round)",
            False, "no excel",
        )
        check(
            "Literature Support: 5 topics with titles matching injected pedagogy papers (not noise)",
            False, "no excel",
        )
        return None
    check("Curriculum_Review.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        check("Curriculum_Review.xlsx readable", False, str(e))
        return None
    check("Curriculum_Review.xlsx readable", True)

    # ---- Sheet 1: Course Compliance ----
    rows = load_sheet_rows(wb, "Course Compliance")
    course_count = None
    compliant_count = None
    noncompliant_count = None
    if rows is None:
        check("'Course Compliance' sheet exists", False)
        check("Основы финансов (Осень 2013) is Non-compliant (No)", False, "no sheet")
        check(
            "Applied Analytics (Fall 2013) student count honest (~383) and ratio == students/faculty",
            False, "no sheet",
        )
    else:
        check("'Course Compliance' sheet exists", True)
        data_rows = [r for r in rows[1:] if r and r[0] is not None]
        course_count = len(data_rows)
        # NON-critical structural: ~22 rows.
        check("Course Compliance has ~22 rows", len(data_rows) >= 20, f"got {len(data_rows)}")

        compliant_vals = []
        for r in data_rows:
            if len(r) > 7 and r[7]:
                compliant_vals.append(str(r[7]).strip().lower())
        compliant_count = sum(1 for v in compliant_vals if v == "yes")
        noncompliant_count = sum(1 for v in compliant_vals if v == "no")
        check("Some courses Compliant='Yes'", compliant_count >= 1, f"yes={compliant_count}")
        check("Some courses Compliant='No'", noncompliant_count >= 1, f"no={noncompliant_count}")

        # CRITICAL: Основы финансов (Осень 2013) must be Non-compliant.
        fof_rows = [
            r for r in data_rows
            if r[0] and "основы финансов" in str(r[0]).lower()
            and "осень 2013" in str(r[0]).lower()
        ]
        fof_ok = bool(fof_rows) and len(fof_rows[0]) > 7 and \
            str(fof_rows[0][7]).strip().lower() == "no"
        check("Основы финансов (Осень 2013) is Non-compliant (No)", fof_ok,
              f"row: {fof_rows[0] if fof_rows else 'missing'}")

        # CRITICAL: Applied Analytics (Fall 2013) honest student count (~383) AND
        # Student_Faculty_Ratio == Student_Count / Faculty_Count.
        aa_rows = [
            r for r in data_rows
            if r[0] and "прикладная аналитика" in str(r[0]).lower()
            and "осень 2013" in str(r[0]).lower()
        ]
        aa_ok = False
        aa_detail = "missing"
        if aa_rows and len(aa_rows[0]) > 6:
            r = aa_rows[0]
            students = safe_float(r[4])
            faculty = safe_float(r[5])
            ratio = safe_float(r[6])
            students_ok = num_close(students, 383, abs_tol=50)
            ratio_ok = False
            if students is not None and faculty not in (None, 0) and ratio is not None:
                expected_ratio = students / faculty
                ratio_ok = num_close(ratio, expected_ratio, abs_tol=1.0, rel_tol=0.05)
            aa_ok = students_ok and ratio_ok
            aa_detail = f"students={students} faculty={faculty} ratio={ratio}"
        check(
            "Applied Analytics (Fall 2013) student count honest (~383) and ratio == students/faculty",
            aa_ok, aa_detail,
        )

    # ---- Sheet 2: Literature Support ----
    rows2 = load_sheet_rows(wb, "Literature Support")
    paper_count = 0
    if rows2 is None:
        check("'Literature Support' sheet exists", False)
        check(
            "Literature Support: 5 topics with titles matching injected pedagogy papers (not noise)",
            False, "no sheet",
        )
    else:
        check("'Literature Support' sheet exists", True)
        data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
        paper_count = len(data_rows2)
        check("Literature Support has >= 5 rows", len(data_rows2) >= 5, f"got {len(data_rows2)}")

        # NON-critical: topic keyword coverage.
        topics_found = [str(r[0]).strip().lower() for r in data_rows2 if r[0]]
        for keyword in ["active", "assessment", "online", "curriculum", "engagement"]:
            check(
                f"Literature Support covers topic '{keyword}'",
                any(keyword in t for t in topics_found),
                f"topics: {topics_found}",
            )

        # CRITICAL: 5 topics covered AND each row's Paper_Title matches an injected
        # pedagogy paper (NOT a noise paper). Verifies correct scholarly retrieval.
        titles = [str(r[1]).strip().lower() if len(r) > 1 and r[1] else "" for r in data_rows2]
        topic_keywords = {"active", "assessment", "online", "curriculum", "engagement"}
        covered = {k for k in topic_keywords if any(k in t for t in topics_found)}
        pedagogy_matches = sum(
            1 for t in titles if any(p in t for p in PEDAGOGY_PAPER_TITLES)
        )
        noise_matches = sum(1 for t in titles if any(n in t for n in NOISE_PAPER_TITLES))
        lit_ok = (
            len(data_rows2) >= 5
            and len(covered) == 5
            and pedagogy_matches >= 5
            and noise_matches == 0
        )
        check(
            "Literature Support: 5 topics with titles matching injected pedagogy papers (not noise)",
            lit_ok,
            f"covered={sorted(covered)} pedagogy={pedagogy_matches} noise={noise_matches}",
        )

    # ---- Sheet 3: Summary ----
    rows3 = load_sheet_rows(wb, "Summary")
    if rows3 is None:
        check("'Summary' sheet exists", False)
        check(
            "Summary math consistent (total ~22, compliant+non-compliant==total, rate==round)",
            False, "no sheet",
        )
    else:
        check("'Summary' sheet exists", True)
        data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("Summary has >= 5 rows", len(data_rows3) >= 5, f"got {len(data_rows3)}")

        def find_metric(*needles):
            for r in data_rows3:
                label = str(r[0]).lower() if r[0] else ""
                if all(n in label for n in needles) and len(r) > 1:
                    return safe_float(r[1])
            return None

        s_total = find_metric("total", "course")
        # compliant / non-compliant rows: match on 'compliant' and (non/not) prefix.
        s_compliant = None
        s_noncompliant = None
        s_rate = None
        for r in data_rows3:
            label = str(r[0]).lower() if r[0] else ""
            val = safe_float(r[1]) if len(r) > 1 else None
            if "compliant" not in label and "соответ" not in label and "rate" not in label \
                    and "доля" not in label:
                continue
            is_rate = "rate" in label or "доля" in label or "%" in str(r[1] or "")
            is_non = "non" in label or "not " in label or "несоответ" in label or "не соответ" in label
            if is_rate:
                s_rate = val
            elif is_non:
                s_noncompliant = val
            elif "compliant" in label or "соответ" in label:
                s_compliant = val

        # CRITICAL: summary internally consistent with Course Compliance sheet.
        total_ok = num_close(s_total, 22, abs_tol=2) if s_total is not None else False
        sum_ok = False
        rate_ok = False
        if s_total is not None and s_compliant is not None and s_noncompliant is not None:
            sum_ok = num_close(s_compliant + s_noncompliant, s_total, abs_tol=0.5)
            if s_rate is not None and s_total not in (None, 0):
                expected_rate = round(s_compliant / s_total * 100)
                rate_ok = num_close(s_rate, expected_rate, abs_tol=1.0, rel_tol=0.0)
        summary_ok = total_ok and sum_ok and rate_ok
        check(
            "Summary math consistent (total ~22, compliant+non-compliant==total, rate==round)",
            summary_ok,
            f"total={s_total} comp={s_compliant} noncomp={s_noncompliant} rate={s_rate}",
        )

    return wb


def check_teamly():
    """CRITICAL: a Teamly page titled 'Accreditation Review Report' exists, and
    its body contains the overall compliance rate value AND names at least one
    non-compliant course (Основы финансов).

    Seed pages have id <= 3; the preprocess parent page ('Документы по
    аккредитации') is a container and must NOT satisfy this check. We accept the
    English title marker 'accreditation review report' OR a Russian-phrased
    title combined with the report substance. Keyword matching uses RU+EN
    alternatives in lowered ORIGINAL text (no normalize).
    """
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check(
            "Teamly 'Accreditation Review Report' page exists with compliance rate and a non-compliant course",
            False, str(e),
        )
        return

    report = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "документы по аккредитации" in tl:
            continue  # preprocess container page, not the report
        if "accreditation review report" in tl or (
            ("аккредитац" in tl or "accreditation" in tl)
            and ("отчёт" in tl or "отчет" in tl or "report" in tl or "обзор" in tl)
        ):
            report = (pid, title, body)
            break
    if report is None:
        check(
            "Teamly 'Accreditation Review Report' page exists with compliance rate and a non-compliant course",
            False, f"new pages: {[(p[0], p[1]) for p in pages]}",
        )
        return

    text = ((report[1] or "") + " " + (report[2] or "")).lower()

    # Overall compliance rate: must mention a percentage / rate phrasing (RU+EN).
    import re
    has_pct = bool(re.search(r"\d+\s*%", text)) or bool(re.search(r"\d+(\.\d+)?\s*процент", text))
    mentions_rate = (
        "compliance rate" in text or "доля соответ" in text or "процент соответ"
        in text or "уровень соответ" in text or "соответ" in text
    )
    rate_ok = has_pct and mentions_rate

    # Names at least one non-compliant course (Основы финансов is the
    # known failing course; course names come from russified Canvas data).
    names_failing = "основы финансов" in text

    body_ok = rate_ok and names_failing
    check(
        "Teamly 'Accreditation Review Report' page exists with compliance rate and a non-compliant course",
        body_ok,
        f"title='{report[1]}' has_pct={has_pct} mentions_rate={mentions_rate} names_fof={names_failing}",
    )


def run_evaluation(agent_workspace):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    print("  Checking Excel file...")
    check_excel(agent_workspace)
    print("  Checking Teamly page...")
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    return success, accuracy, critical_failed, total


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace"
    )

    success, accuracy, critical_failed, total = run_evaluation(agent_ws)

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump(
                    {
                        "total_passed": PASS_COUNT,
                        "total_checks": total,
                        "accuracy": accuracy,
                        "critical_failed": critical_failed,
                    },
                    f,
                    indent=2,
                )
        except Exception:
            pass

    if success:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        if critical_failed:
            print("\n=== RESULT: FAIL (critical check failed) ===")
        else:
            print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
