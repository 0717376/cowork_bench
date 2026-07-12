"""Evaluation for sf-hr-salary-overview (ClickHouse fork).

Two-gate evaluation:
  - CRITICAL_CHECKS: semantic core deliverables (correct per-department values
    joined from the russified DB + benchmark PDF, and the Summary aggregates).
    Any critical failure => immediate FAIL (sys.exit(1)), before the accuracy gate.
  - Non-critical structural checks (sheet/column presence, row count) feed an
    overall accuracy>=70 gate.

Department names are russified centrally (db/zzz_clickhouse_after_init.sql) and the
groundtruth Excel + benchmark PDF carry the SAME russified labels, so the lowercase
Department join key matches between agent output and groundtruth.
"""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Salary_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Salary_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # checks: list of (name, passed, is_critical)
    checks = []

    def record(name, passed, critical=False):
        checks.append((name, bool(passed), critical))
        status = "PASS" if passed else "FAIL"
        tag = " [CRITICAL]" if critical else ""
        print(f"  [{status}] {name}{tag}")

    # ----- Sheet: Department Analysis -----------------------------------
    a_rows = load_sheet_rows(agent_wb, "Department Analysis")
    g_rows = load_sheet_rows(gt_wb, "Department Analysis")

    # Structural (non-critical): sheet present
    record("Department Analysis: sheet present", a_rows is not None, critical=False)

    if a_rows is None or g_rows is None:
        if g_rows is None:
            print("  (groundtruth missing Department Analysis sheet)")
        # mark dependent critical checks as failed so we exit on critical gate
        record("Department Analysis: per-department Avg_Salary correct", False, critical=True)
        record("Department Analysis: Employee_Count + Benchmark joined from PDF", False, critical=True)
        record("Department Analysis: Variance + Variance_Pct derived", False, critical=True)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        # Structural (non-critical): row count
        record("Department Analysis: row count matches", len(a_data) == len(g_data), critical=False)

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        avg_ok = True       # critical: per-department Avg_Salary
        join_ok = True      # critical: Employee_Count + Benchmark joined from PDF
        variance_ok = True  # critical: Variance + Variance_Pct derived
        minmax_ok = True    # non-critical: Min/Max salaries

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                print(f"      Missing row: {g_row[0]}")
                avg_ok = join_ok = variance_ok = minmax_ok = False
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    print(f"      {key}.Employee_Count: {a_row[1]} vs {g_row[1]}")
                    join_ok = False
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 5.0):
                    print(f"      {key}.Avg_Salary: {a_row[2]} vs {g_row[2]}")
                    avg_ok = False
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1.0):
                    print(f"      {key}.Min_Salary: {a_row[3]} vs {g_row[3]}")
                    minmax_ok = False
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1.0):
                    print(f"      {key}.Max_Salary: {a_row[4]} vs {g_row[4]}")
                    minmax_ok = False
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1.0):
                    print(f"      {key}.Benchmark: {a_row[5]} vs {g_row[5]}")
                    join_ok = False
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 5.0):
                    print(f"      {key}.Variance: {a_row[6]} vs {g_row[6]}")
                    variance_ok = False
            if len(a_row) > 7 and len(g_row) > 7:
                if not num_close(a_row[7], g_row[7], 0.5):
                    print(f"      {key}.Variance_Pct: {a_row[7]} vs {g_row[7]}")
                    variance_ok = False

        record("Department Analysis: per-department Avg_Salary correct", avg_ok, critical=True)
        record("Department Analysis: Employee_Count + Benchmark joined from PDF", join_ok, critical=True)
        record("Department Analysis: Variance + Variance_Pct derived", variance_ok, critical=True)
        record("Department Analysis: Min/Max salaries correct", minmax_ok, critical=False)

    # ----- Sheet: Summary -----------------------------------------------
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    record("Summary: sheet present", a_rows is not None, critical=False)

    if a_rows is None or g_rows is None:
        if g_rows is None:
            print("  (groundtruth missing Summary sheet)")
        record("Summary: Total_Employees + Overall_Avg_Salary correct", False, critical=True)
        record("Summary: Departments_Above/Below_Benchmark correct", False, critical=True)
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        def summary_close(metric, tol):
            g_val = None
            for g_row in g_data:
                if g_row and g_row[0] is not None and str(g_row[0]).strip().lower() == metric.lower():
                    g_val = g_row[1] if len(g_row) > 1 else None
                    break
            a_row = a_lookup.get(metric.lower())
            if a_row is None:
                print(f"      Summary missing metric: {metric}")
                return False
            a_val = a_row[1] if len(a_row) > 1 else None
            ok = num_close(a_val, g_val, tol)
            if not ok:
                print(f"      {metric}: {a_val} vs {g_val}")
            return ok

        totals_ok = summary_close("Total_Employees", 10.0) and summary_close("Overall_Avg_Salary", 10.0)
        record("Summary: Total_Employees + Overall_Avg_Salary correct", totals_ok, critical=True)

        # Non-critical: Overall_Benchmark
        record("Summary: Overall_Benchmark correct", summary_close("Overall_Benchmark", 1.0), critical=False)

        counts_ok = (summary_close("Departments_Above_Benchmark", 0.5)
                     and summary_close("Departments_Below_Benchmark", 0.5))
        record("Summary: Departments_Above/Below_Benchmark correct", counts_ok, critical=True)

    # ----- Gates --------------------------------------------------------
    critical_failed = [n for n, p, c in checks if c and not p]
    total = len(checks)
    passed = sum(1 for _, p, _ in checks if p)
    accuracy = (passed / total * 100.0) if total else 0.0

    print(f"\n  Critical failures: {len(critical_failed)}")
    for n in critical_failed:
        print(f"    - {n}")
    print(f"  Accuracy: {passed}/{total} = {accuracy:.1f}%")

    if critical_failed:
        print("\n=== RESULT: FAIL (critical check failed) ===")
        sys.exit(1)

    if accuracy < 70.0:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)

    print("\n=== RESULT: PASS ===")
    sys.exit(0)


if __name__ == "__main__":
    main()
