"""Evaluation for insales-review-competitor-study (InSales + Teamly, russified)."""
import argparse
import json
import os
import sys

import psycopg2


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


# Russian store category name (as written by the agent / groundtruth Excel) ->
# English dashboard/groundtruth_data category. The wc seed russifies the store
# category data values centrally; this bridge lets the eval match either form.
RU2EN_CATEGORY = {
    "электроника": "Electronics",
    "камеры": "Cameras",
    "аудио": "Audio",
    "тв и домашний кинотеатр": "TV & Home Theater",
    "часы": "Watches",
    "бытовая техника": "Home Appliances",
}


def category_matches(cell_value, en_category):
    """True if an Excel Category cell refers to the given English category,
    accepting either the English name or its Russian store equivalent."""
    if cell_value is None:
        return False
    cell = str(cell_value).strip().lower()
    if en_category.lower() in cell:
        return True
    # Map the cell (RU) to EN and compare.
    mapped = RU2EN_CATEGORY.get(cell)
    if mapped and mapped.lower() == en_category.lower():
        return True
    # Also allow RU substring of a known RU label.
    for ru, en in RU2EN_CATEGORY.items():
        if en.lower() == en_category.lower() and ru in cell:
            return True
    return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_summary_value(rows, metric):
    """Return Value cell for a given Metric label in the Summary sheet."""
    for r in rows[1:]:
        if r and r[0] is not None and str(r[0]).strip().lower() == metric.strip().lower():
            return r[1] if len(r) > 1 else None
    return None


def check_excel(agent_workspace, gt_data):
    """Returns (errors, critical_errors)."""
    errors = []
    critical = []
    import openpyxl
    path = os.path.join(agent_workspace, "Review_Benchmark.xlsx")
    if not os.path.exists(path):
        return ["Review_Benchmark.xlsx not found"], ["Review_Benchmark.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # ---- Category Comparison sheet ----
        rows = load_sheet_rows(wb, "Category Comparison")
        if rows is None:
            errors.append("Sheet 'Category Comparison' not found")
            critical.append("Sheet 'Category Comparison' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_categories"]
            if abs(len(data_rows) - expected) > 1:
                errors.append(f"Category Comparison has {len(data_rows)} rows, expected {expected}")

            for gc in gt_data["comparisons"]:
                cat_rows = [r for r in data_rows if category_matches(r[0], gc["category"])]
                if not cat_rows:
                    critical.append(f"Category Comparison: row for '{gc['category']}' not found")
                    continue
                row = cat_rows[0]
                # CRITICAL: Our_Avg_Rating (col 1) within 0.15 of groundtruth.
                try:
                    val = float(row[1])
                    if abs(val - gc["our_avg_rating"]) > 0.15:
                        critical.append(
                            f"{gc['category']} Our_Avg_Rating={val}, expected ~{gc['our_avg_rating']}")
                except (ValueError, TypeError, IndexError):
                    critical.append(f"{gc['category']} Our_Avg_Rating missing/invalid")
                # CRITICAL: Competitor_Avg_Rating (col 3) within 0.1 of dashboard value.
                try:
                    cval = float(row[3])
                    if abs(cval - gc["competitor_avg_rating"]) > 0.1:
                        critical.append(
                            f"{gc['category']} Competitor_Avg_Rating={cval}, "
                            f"expected ~{gc['competitor_avg_rating']}")
                except (ValueError, TypeError, IndexError):
                    critical.append(f"{gc['category']} Competitor_Avg_Rating missing/invalid")
                # CRITICAL: Status (col 6) equals expected.
                try:
                    status = str(row[6]).strip().lower()
                    if status != gc["status"].lower():
                        critical.append(
                            f"{gc['category']} Status='{row[6]}', expected '{gc['status']}'")
                except (IndexError, AttributeError):
                    critical.append(f"{gc['category']} Status missing")

        # ---- Products Below Benchmark sheet ----
        rows2 = load_sheet_rows(wb, "Products Below Benchmark")
        if rows2 is None:
            errors.append("Sheet 'Products Below Benchmark' not found")
            critical.append("Sheet 'Products Below Benchmark' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            expected_pb = gt_data["products_below_count"]
            if abs(len(data_rows2) - expected_pb) > 3:
                critical.append(
                    f"Products Below Benchmark has {len(data_rows2)} rows, expected ~{expected_pb}")
            # CRITICAL: known below-benchmark Product_IDs present.
            present_ids = set()
            for r in data_rows2:
                try:
                    present_ids.add(int(float(r[0])))
                except (ValueError, TypeError):
                    pass
            for pid in (61, 21, 76, 25, 13):
                if pid not in present_ids:
                    critical.append(f"Products Below Benchmark missing Product_ID {pid}")

        # ---- Summary sheet ----
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")
            critical.append("Sheet 'Summary' not found")
        else:
            expected_summary = {
                "Total_Categories": gt_data["total_categories"],
                "Categories_Above_Benchmark": gt_data["categories_above"],
                "Categories_Below_Benchmark": gt_data["categories_below"],
            }
            for metric, exp in expected_summary.items():
                val = find_summary_value(rows3, metric)
                try:
                    if int(float(val)) != int(exp):
                        critical.append(f"Summary {metric}={val}, expected {exp}")
                except (ValueError, TypeError):
                    critical.append(f"Summary {metric} missing/invalid (got {val})")
            # Products_Below_Benchmark within +-3 of groundtruth.
            pbv = find_summary_value(rows3, "Products_Below_Benchmark")
            try:
                if abs(int(float(pbv)) - gt_data["products_below_count"]) > 3:
                    critical.append(
                        f"Summary Products_Below_Benchmark={pbv}, "
                        f"expected ~{gt_data['products_below_count']}")
            except (ValueError, TypeError):
                critical.append(f"Summary Products_Below_Benchmark missing/invalid (got {pbv})")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
        critical.append(f"Error reading Excel: {e}")
    return errors, critical


def check_teamly(gt_data):
    """Returns (errors, critical_errors). Accepts the English literal title
    'Review Performance Analysis Q1 2026' OR a Russian equivalent, and requires
    the body to name an underperforming category + a recommendation keyword."""
    errors = []
    critical = []
    try:
        conn = psycopg2.connect(
            host=os.environ.get("PGHOST", "localhost"), port=5432,
            dbname=os.environ.get("PGDATABASE", "cowork_gym"),
            user="eigent", password="camel",
        )
        cur = conn.cursor()
        # Only pages created by the agent (seed pages have id <= 3).
        cur.execute("SELECT id, title, body FROM teamly.pages WHERE id > 3 ORDER BY id DESC")
        rows = cur.fetchall()
        cur.close()
        conn.close()

        # The two noise pages ("База знаний", "Старый анализ за Q4 2025") were
        # injected with id > 3 in preprocess; they must not satisfy the check.
        noise_titles = {"база знаний", "старый анализ за q4 2025"}

        found = None
        for _id, title, body in rows:
            t = (title or "").strip().lower()
            if t in noise_titles:
                continue
            is_en = "review" in t and ("performance" in t or "analysis" in t or "q1" in t)
            is_ru = (("отзыв" in t or "оценк" in t) and
                     ("анализ" in t or "производительн" in t or "q1" in t or "качеств" in t))
            if is_en or is_ru:
                found = (title or "", body or "")
                break

        if found is None:
            critical.append(
                "No Teamly analysis page found (expected 'Review Performance "
                "Analysis Q1 2026' or RU equivalent)")
            return errors, critical

        body_l = found[1].lower()
        # CRITICAL: body names an underperforming category (Electronics / Электроника)
        # and contains a recommendation keyword.
        names_under = "electronics" in body_l or "электроника" in body_l
        has_reco = "рекоменд" in body_l or "улучш" in body_l or "recommend" in body_l
        if not names_under:
            critical.append("Teamly page body does not name the underperforming category")
        if not has_reco:
            critical.append("Teamly page body has no recommendation keyword (рекоменд/улучш)")

    except Exception as e:
        errors.append(f"Error checking Teamly: {e}")
        critical.append(f"Error checking Teamly: {e}")
    return errors, critical


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    all_errors = []
    critical_errors = []

    print("  Checking Excel file...")
    errs, crit = check_excel(agent_ws, gt_data)
    all_errors.extend(errs)
    critical_errors.extend(crit)
    for e in (errs + crit)[:8]:
        print(f"    ERROR: {e}")
    if not errs and not crit:
        print("    PASS")

    print("  Checking Teamly page...")
    errs, crit = check_teamly(gt_data)
    all_errors.extend(errs)
    critical_errors.extend(crit)
    for e in (errs + crit)[:5]:
        print(f"    ERROR: {e}")
    if not errs and not crit:
        print("    PASS")

    # ---- Critical gate: any critical failure => FAIL regardless of accuracy ----
    if critical_errors:
        print(f"\n=== CRITICAL CHECK FAILED ({len(critical_errors)}) ===")
        for e in critical_errors[:15]:
            print(f"  CRITICAL: {e}")
        print("=== RESULT: FAIL ===")
        sys.exit(1)

    # ---- Non-critical accuracy gate (threshold 70%) ----
    # Total non-critical structural checks tracked above.
    total_structural = 4  # row-count + 3 sheet-presence (counted via all_errors)
    failed = len([e for e in all_errors if e not in critical_errors])
    accuracy = 100.0 * (total_structural - min(failed, total_structural)) / total_structural
    print(f"\n  Structural accuracy: {accuracy:.0f}% ({failed} non-critical issue(s))")

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
