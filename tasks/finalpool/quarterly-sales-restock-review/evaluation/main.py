"""
Evaluation script for quarterly-sales-restock-review task.

Checks:
1. Excel file (Q4_2025_Business_Review.xlsx) - 4 sheets with correct data
2. Teamly knowledge-base page created with correct structure
3. Emails sent to correct suppliers (and not to wrong ones)

Gate: any CRITICAL check failure => overall FAIL regardless of accuracy.
Otherwise PASS requires accuracy >= 70%.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00" \
        --res_log_file /path/to/result.json
"""

import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical semantic checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Summary Top_Region is correct top-revenue region",
    "Summary Top_Brand == 'LG'",
    "Summary Out_Of_Stock_Count == 5",
    "Summary Products_Below_Threshold == 37",
    "All 5 out-of-stock SKUs present in Restock Alerts",
    "Restock emails sent to exactly the 3 correct suppliers with correct SKU(s)",
    "No restock emails sent to suppliers without out-of-stock products",
    "Teamly Q4 2025 page exists with required headings and checked + unchecked to-do",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Case-insensitive string comparison with whitespace normalization."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check the Excel output file against groundtruth."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Q4_2025_Business_Review.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Q4_2025_Business_Review.xlsx")

    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
        gt_wb = openpyxl.load_workbook(gt_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return False

    all_passed = True

    # Check sheet names
    expected_sheets = ["Regional Sales", "Brand Performance", "Restock Alerts", "Summary"]
    agent_sheets = agent_wb.sheetnames
    for sheet_name in expected_sheets:
        found = any(str_match(s, sheet_name) for s in agent_sheets)
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {agent_sheets}")
        if not found:
            all_passed = False

    # Helper to find sheet case-insensitively
    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if str_match(s, name):
                return wb[s]
        return None

    # --- Sheet 1: Regional Sales ---
    print("\n--- Regional Sales ---")
    agent_ws = get_sheet(agent_wb, "Regional Sales")
    gt_ws = get_sheet(gt_wb, "Regional Sales")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Regional Sales row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")
        for gt_row in gt_rows:
            region = gt_row[0]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], region):
                    matched = ar
                    break
            if matched:
                check(f"Region '{region}' Order_Count",
                      num_close(matched[1], gt_row[1], 0.5),
                      f"Expected {gt_row[1]}, got {matched[1]}")
                check(f"Region '{region}' Total_Revenue",
                      num_close(matched[2], gt_row[2], 1.0),
                      f"Expected {gt_row[2]}, got {matched[2]}")
                check(f"Region '{region}' Avg_Order_Value",
                      num_close(matched[3], gt_row[3], 0.5),
                      f"Expected {gt_row[3]}, got {matched[3]}")
            else:
                check(f"Region '{region}' found", False, "Region not in agent output")
                all_passed = False
    else:
        all_passed = False

    # --- Sheet 2: Brand Performance ---
    print("\n--- Brand Performance ---")
    agent_ws = get_sheet(agent_wb, "Brand Performance")
    gt_ws = get_sheet(gt_wb, "Brand Performance")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Brand Performance row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")
        for gt_row in gt_rows:
            brand = gt_row[0]
            matched = None
            for ar in agent_rows:
                if ar and str_match(ar[0], brand):
                    matched = ar
                    break
            if matched:
                check(f"Brand '{brand}' Total_Revenue",
                      num_close(matched[2], gt_row[2], 1.0),
                      f"Expected {gt_row[2]}, got {matched[2]}")
                check(f"Brand '{brand}' Total_Units",
                      num_close(matched[3], gt_row[3], 0.5),
                      f"Expected {gt_row[3]}, got {matched[3]}")
            else:
                check(f"Brand '{brand}' found", False, "Brand not in agent output")
                all_passed = False
    else:
        all_passed = False

    # --- Sheet 3: Restock Alerts ---
    print("\n--- Restock Alerts ---")
    agent_ws = get_sheet(agent_wb, "Restock Alerts")
    gt_ws = get_sheet(gt_wb, "Restock Alerts")
    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Restock Alerts row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}")
        # Check key rows (out-of-stock products) — SKUs stay English.
        oos_skus = [r[0] for r in gt_rows if r[2] == 0]
        missing = [sku for sku in oos_skus
                   if not any(ar and str_match(ar[0], sku) for ar in agent_rows)]
        for sku in oos_skus:
            found = any(ar and str_match(ar[0], sku) for ar in agent_rows)
            check(f"Out-of-stock SKU '{sku}' in Restock Alerts", found)
        check("All 5 out-of-stock SKUs present in Restock Alerts",
              not missing, f"Missing: {missing}")
        if missing:
            all_passed = False
    else:
        all_passed = False

    # --- Sheet 4: Summary ---
    print("\n--- Summary ---")
    agent_ws = get_sheet(agent_wb, "Summary")
    gt_ws = get_sheet(gt_wb, "Summary")
    if agent_ws and gt_ws:
        gt_data = {}
        for row in gt_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                gt_data[str(row[0]).strip().lower()] = row[1]
        agent_data = {}
        for row in agent_ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key, gt_val in gt_data.items():
            agent_val = agent_data.get(key)
            if agent_val is None:
                # Try fuzzy key match
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                ok = num_close(agent_val, gt_val, 1.0)
                check(f"Summary '{key}'", ok,
                      f"Expected {gt_val}, got {agent_val}")
            else:
                ok = str_match(agent_val, gt_val)
                check(f"Summary '{key}'", ok,
                      f"Expected '{gt_val}', got '{agent_val}'")
            if not ok:
                all_passed = False

        # --- Explicit critical Summary checks (semantic core deliverable) ---
        def _agent_summary(key):
            v = agent_data.get(key)
            if v is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        return av
            return v

        # Top_Region: gt is already the RU region string from the central
        # clickhouse map; accept EN 'Asia Pacific' fallback for robustness.
        gt_top_region = gt_data.get("top_region")
        agent_top_region = _agent_summary("top_region")
        region_ok = str_match(agent_top_region, gt_top_region) or \
            str_match(agent_top_region, "Asia Pacific") or \
            str_match(agent_top_region, "Азиатско-Тихоокеанский регион")
        check("Summary Top_Region is correct top-revenue region", region_ok,
              f"Expected '{gt_top_region}', got '{agent_top_region}'")

        check("Summary Top_Brand == 'LG'",
              str_match(_agent_summary("top_brand"), "LG"),
              f"got '{_agent_summary('top_brand')}'")

        check("Summary Out_Of_Stock_Count == 5",
              num_close(_agent_summary("out_of_stock_count"), 5, 0.5),
              f"got '{_agent_summary('out_of_stock_count')}'")

        check("Summary Products_Below_Threshold == 37",
              num_close(_agent_summary("products_below_threshold"), 37, 0.5),
              f"got '{_agent_summary('products_below_threshold')}'")
    else:
        all_passed = False

    return all_passed


def check_teamly():
    """Check that a Teamly knowledge-base page was created correctly.

    Teamly stores pages as title + markdown body in teamly.pages. Seed pages
    have id <= 3; the agent's page is a user-created page (id > 3) whose title
    carries the English marker 'Q4 2025'. Headings are matched against RU+EN
    alternatives since the agent writes Russian prose; structural markers
    (region/brand) stay English in identifiers but the region value is the
    RU central-map string, so we accept RU OR EN.
    """
    print("\n=== Checking Teamly Page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    all_passed = True

    try:
        cur.execute(
            "SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3"
        )
        pages = cur.fetchall()
    except Exception as e:
        check("Teamly Q4 2025 page exists with required headings and "
              "checked + unchecked to-do", False, str(e))
        cur.close()
        conn.close()
        return False

    found = None
    for pid, title, body in pages:
        if "q4 2025" in (title or "").lower():
            found = (pid, title or "", body or "")
            break

    check("Teamly page with 'Q4 2025' in title exists", found is not None,
          f"pages: {[(p[0], p[1]) for p in pages]}")
    if not found:
        check("Teamly Q4 2025 page exists with required headings and "
              "checked + unchecked to-do", False, "no page")
        cur.close()
        conn.close()
        return False

    text = ((found[1] or "") + "\n" + (found[2] or "")).lower()

    # Headings: agent writes RU or keeps EN heading wording — accept both.
    has_sales = any(k in text for k in ("sales highlights", "sales", "продаж", "highlights"))
    check("Teamly page has 'Sales Highlights' heading (RU/EN)", has_sales)

    has_inventory = any(k in text for k in ("inventory", "склад", "запас", "остатк"))
    check("Teamly page has 'Inventory Alerts' heading (RU/EN)", has_inventory)

    has_action = any(k in text for k in ("action item", "action", "действ", "задач"))
    check("Teamly page has 'Action Items' heading (RU/EN)", has_action)

    # To-do markers: markdown checkboxes — one checked, one unchecked.
    has_checked = ("[x]" in text) or ("[х]" in text) or ("☑" in text)
    has_unchecked = ("[ ]" in text) or ("☐" in text)
    check("Teamly page has a checked to-do item", has_checked)
    check("Teamly page has an unchecked to-do item", has_unchecked)

    # Top region (RU central-map value, EN fallback) and top brand ('LG').
    has_region = ("азиатско-тихоокеанский" in text) or ("asia pacific" in text)
    check("Teamly mentions top region (RU 'Азиатско-Тихоокеанский' or EN 'Asia Pacific')",
          has_region)
    has_brand = "lg" in text.replace(",", " ").replace(".", " ").split()
    check("Teamly mentions 'LG' (top brand)", has_brand)

    # Critical aggregate: page exists with all three headings + both to-do states.
    check("Teamly Q4 2025 page exists with required headings and "
          "checked + unchecked to-do",
          has_sales and has_inventory and has_action and has_checked and has_unchecked,
          f"sales={has_sales} inv={has_inventory} action={has_action} "
          f"checked={has_checked} unchecked={has_unchecked}")
    if not (has_sales and has_inventory and has_action and has_checked and has_unchecked):
        all_passed = False

    cur.close()
    conn.close()
    return all_passed


def check_emails():
    """Check that restock emails were sent to the correct suppliers."""
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    all_passed = True

    # Get sent emails (folder_id=2 is Sent)
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE folder_id = 2
    """)
    sent_emails = cur.fetchall()

    # Also check inbox of recipients via sent_log or messages in any folder
    # The email MCP typically stores sent emails in folder 2
    # But also check all messages for supplier recipients
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()

    # Expected suppliers to receive emails (those with out-of-stock products)
    expected_suppliers = {
        "trade@asiatech.com": {
            "name": "Asia Tech Trading",
            "skus": ["INFINITY-JBL-GLIDE-1020", "SILENCER-PANELS-FLAM-1079",
                     "30KG-DIGITAL-SCALE-1071"],
        },
        "b2b@smarthome-ws.com": {
            "name": "SmartHome Wholesale",
            "skus": ["BOXTUDIO-LIGHTBOX-TA-1039"],
        },
        "wholesale@digitaldreams.com": {
            "name": "Digital Dreams Supply",
            "skus": ["TYPEC-EARPHONE-FOR-1045"],
        },
    }

    # Should NOT receive emails (suppliers without out-of-stock products)
    should_not_receive = [
        "verkauf@euroelec.de",
        "supply@premiumgadgets.com",
        "sales@globalelec.com",
        "orders@avpartners.com",
        "orders@techworld-dist.com",
    ]

    def find_email_for_recipient(recipient):
        """Find an email addressed to this recipient across all emails."""
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    correct_supplier_ok = True
    # Check each expected supplier received an email
    for supplier_email, info in expected_suppliers.items():
        result = find_email_for_recipient(supplier_email)
        check(f"Email sent to {supplier_email}", result is not None)
        if result:
            subj, from_addr, to_addr, body = result
            # Check subject contains "Restock" (case-insensitive)
            has_restock_subject = "restock" in (subj or "").lower()
            check(f"Email to {supplier_email} subject contains 'Restock'",
                  has_restock_subject,
                  f"Subject: {(subj or '')[:100]}")
            # Check subject contains supplier name
            has_supplier_name = info["name"].lower() in (subj or "").lower()
            check(f"Email to {supplier_email} subject contains supplier name",
                  has_supplier_name,
                  f"Subject: {(subj or '')[:100]}")
            # Check body mentions at least one SKU
            body_lower = (body or "").lower()
            has_sku = any(sku.lower() in body_lower for sku in info["skus"])
            check(f"Email to {supplier_email} body mentions product SKU(s)",
                  has_sku,
                  f"Expected one of {info['skus']}")
            if not has_sku:
                correct_supplier_ok = False
        else:
            all_passed = False
            correct_supplier_ok = False

    check("Restock emails sent to exactly the 3 correct suppliers with correct SKU(s)",
          correct_supplier_ok)

    # Check suppliers that should NOT receive emails
    no_wrong = True
    for email_addr in should_not_receive:
        result = find_email_for_recipient(email_addr)
        check(f"No email sent to {email_addr} (no out-of-stock products)",
              result is None,
              f"Unexpected email found with subject: {result[0][:100] if result else ''}")
        if result:
            all_passed = False
            no_wrong = False

    check("No restock emails sent to suppliers without out-of-stock products",
          no_wrong)

    cur.close()
    conn.close()
    return all_passed


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks.

    Gate: any CRITICAL check failure => FAIL regardless of accuracy.
    Otherwise PASS requires accuracy >= 70%.
    """
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_ok = check_excel(agent_workspace, groundtruth_workspace)
    teamly_ok = check_teamly()
    email_ok = check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    success = (not critical_failed) and accuracy >= 70

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    print(f"  Failed: {FAIL_COUNT}")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "total_checks": total,
            "accuracy": accuracy,
            "critical_failed": critical_failed,
            "success": success,
            "details": {
                "excel": excel_ok,
                "teamly": teamly_ok,
                "email": email_ok,
            }
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
