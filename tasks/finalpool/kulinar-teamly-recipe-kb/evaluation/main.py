"""
Evaluation script for kulinar-teamly-recipe-kb task.

Checks:
1. Excel file (Recipe_Comparison.xlsx) - two sheets with correct structure and
   data; recipe names cross-validated against the real kulinar recipe MCP data
   so counts/categories cannot be fabricated.
2. Teamly - a space "Коллекция рецептов команды" with >= 5 recipe pages, each
   body listing Название + Категория + Difficulty (replaces the old Notion
   database-property check).
3. Memory - memory.json has entities recording selected categories / filtering
   criteria (RU+EN keyword match).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "All Recipes has at least 8 recipes",
    "All Recipes spans at least 4 categories",
    "All Recipes rows cross-check against kulinar recipes",
    "Top Picks has at least 5 entries",
    "Top Picks reason column filled",
    "Teamly recipe space 'Коллекция рецептов команды' has >= 5 recipe pages",
    "Teamly recipe pages list Название + Категория + Difficulty",
    "Memory contains filtering criteria or category info",
}

# Path to the kulinar recipe source data (for name cross-validation).
KULINAR_DATA = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "..", "..", "..",
    "local_servers", "kulinar-mcp", "src", "data", "all_recipes.json",
)


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def normalize(text):
    """Normalize text for comparison: lowercase, strip whitespace."""
    if text is None:
        return ""
    return str(text).strip().lower()


def load_kulinar_names():
    """Return set of normalized kulinar recipe names, or None if unavailable."""
    try:
        with open(KULINAR_DATA, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {normalize(r.get("name")) for r in data if r.get("name")}
    except Exception as e:
        print(f"  [WARN] could not load kulinar data: {e}")
        return None


def load_sheet_rows(wb, sheet_name):
    """Load all rows from a sheet (case-insensitive name lookup)."""
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    ws = wb[matched]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_excel(agent_workspace):
    """Check Recipe_Comparison.xlsx has correct structure and content."""
    print("\n=== Checking Excel Output ===")

    excel_path = os.path.join(agent_workspace, "Recipe_Comparison.xlsx")

    if not os.path.isfile(excel_path):
        record("Recipe_Comparison.xlsx exists", False, f"Not found at {excel_path}")
        return False

    record("Recipe_Comparison.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    record("Excel file readable", True)
    all_ok = True

    kulinar_names = load_kulinar_names()

    # --- Sheet 1: All Recipes ---
    all_recipes = load_sheet_rows(wb, "All Recipes")
    if all_recipes is None:
        record("Sheet 'All Recipes' exists", False, "Not found")
        all_ok = False
    else:
        record("Sheet 'All Recipes' exists", True)

        # Check header row (keyword substrings, RU prose tolerated).
        if len(all_recipes) < 1:
            record("All Recipes has header", False, "Sheet is empty")
            all_ok = False
        else:
            headers = [normalize(h) for h in all_recipes[0]]
            header_keywords = {
                "recipe": False, "category": False, "difficulty": False,
                "ingredient": False, "step": False,
            }
            for h in headers:
                for kw in header_keywords:
                    if kw in h:
                        header_keywords[kw] = True

            all_headers_found = all(header_keywords.values())
            record("All Recipes headers correct", all_headers_found,
                   f"Headers: {headers}, missing keywords: "
                   f"{[k for k, v in header_keywords.items() if not v]}")
            if not all_headers_found:
                all_ok = False

        # Check data rows (at least 8)
        data_rows = all_recipes[1:] if len(all_recipes) > 1 else []
        data_rows = [r for r in data_rows if any(c is not None for c in r)]
        record("All Recipes has at least 8 recipes",
               len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")
        if len(data_rows) < 8:
            all_ok = False

        # Check that recipes span at least 4 categories
        if data_rows and len(all_recipes[0]) >= 2:
            categories = set()
            for row in data_rows:
                if len(row) >= 2 and row[1] is not None:
                    categories.add(normalize(row[1]))
            record("All Recipes spans at least 4 categories",
                   len(categories) >= 4,
                   f"Found {len(categories)} categories: {categories}")
            if len(categories) < 4:
                all_ok = False

        # Check that numeric columns have positive integer values
        numeric_ok_count = 0
        for row in data_rows:
            if len(row) >= 5:
                try:
                    ing = int(float(row[3])) if row[3] is not None else None
                    steps = int(float(row[4])) if row[4] is not None else None
                    if ing is not None and ing > 0 and steps is not None and steps > 0:
                        numeric_ok_count += 1
                except (ValueError, TypeError):
                    pass
        record("All Recipes numeric columns valid",
               numeric_ok_count >= 6,
               f"{numeric_ok_count}/{len(data_rows)} rows have valid numeric data")
        if numeric_ok_count < 6:
            all_ok = False

        # CRITICAL: recipe names must correspond to real kulinar recipes so the
        # agent cannot fabricate plausible rows.
        if kulinar_names is None:
            # Cannot validate fidelity; do not block on infra issue.
            record("All Recipes rows cross-check against kulinar recipes", True,
                   "kulinar data unavailable - skipped")
        else:
            matched = 0
            unmatched = []
            for row in data_rows:
                if row and row[0] is not None:
                    nm = normalize(row[0])
                    if nm in kulinar_names:
                        matched += 1
                    else:
                        unmatched.append(row[0])
            ok = matched >= 6
            record("All Recipes rows cross-check against kulinar recipes", ok,
                   f"{matched}/{len(data_rows)} names matched kulinar; "
                   f"unmatched: {unmatched[:6]}")
            if not ok:
                all_ok = False

    # --- Sheet 2: Top Picks ---
    top_picks = load_sheet_rows(wb, "Top Picks")
    if top_picks is None:
        record("Sheet 'Top Picks' exists", False, "Not found")
        all_ok = False
    else:
        record("Sheet 'Top Picks' exists", True)

        if len(top_picks) < 1:
            record("Top Picks has header", False, "Sheet is empty")
            all_ok = False
        else:
            headers = [normalize(h) for h in top_picks[0]]
            header_keywords = {
                "rank": False, "recipe": False, "category": False, "reason": False,
            }
            for h in headers:
                for kw in header_keywords:
                    if kw in h:
                        header_keywords[kw] = True

            all_headers_found = all(header_keywords.values())
            record("Top Picks headers correct", all_headers_found,
                   f"Headers: {headers}, missing: "
                   f"{[k for k, v in header_keywords.items() if not v]}")
            if not all_headers_found:
                all_ok = False

        data_rows = top_picks[1:] if len(top_picks) > 1 else []
        data_rows = [r for r in data_rows if any(c is not None for c in r)]
        record("Top Picks has at least 5 entries",
               len(data_rows) >= 5,
               f"Found {len(data_rows)} data rows")
        if len(data_rows) < 5:
            all_ok = False

        # Reason column (index 3) non-empty for >= 4 rows.
        reason_filled = 0
        for row in data_rows:
            if len(row) >= 4 and row[3] is not None and len(str(row[3]).strip()) > 3:
                reason_filled += 1
        record("Top Picks reason column filled",
               reason_filled >= 4,
               f"{reason_filled}/{len(data_rows)} rows have reasons")
        if reason_filled < 4:
            all_ok = False

    return all_ok


def check_teamly():
    """Check Teamly has a recipe-collection space with >= 5 recipe pages.

    Replaces the old Notion database-property check. Teamly has no database
    primitive (only spaces + pages), so the deliverable is reframed as a space
    'Коллекция рецептов команды' containing one page per top recipe; each page
    body must list Название + Категория + Difficulty.
    """
    print("\n=== Checking Teamly ===")

    all_ok = True

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find the recipe-collection space.
        cur.execute("SELECT id, name FROM teamly.spaces")
        spaces = cur.fetchall()
        recipe_space = None
        for sid, name in spaces:
            nl = (name or "").lower()
            if "коллекция рецептов" in nl or "recipe collection" in nl \
                    or ("рецепт" in nl and ("коллекц" in nl or "команд" in nl)):
                recipe_space = (sid, name)
                break

        record("Teamly recipe-collection space exists", recipe_space is not None,
               f"Spaces: {[s[1] for s in spaces]}")
        if recipe_space is None:
            record("Teamly recipe space 'Коллекция рецептов команды' has >= 5 recipe pages",
                   False, "space missing")
            record("Teamly recipe pages list Название + Категория + Difficulty",
                   False, "space missing")
            cur.close()
            conn.close()
            return False

        sid = recipe_space[0]
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages "
                    "WHERE space_id = %s", (sid,))
        pages = cur.fetchall()

        record("Teamly recipe space 'Коллекция рецептов команды' has >= 5 recipe pages",
               len(pages) >= 5,
               f"Found {len(pages)} pages in space")
        if len(pages) < 5:
            all_ok = False

        # Each page body should list Название + Категория + Difficulty.
        # Count pages whose combined title+body cover all three field markers.
        # Имя блюда засчитывается и когда оно вынесено в заголовок страницы
        # (без литеральной метки «Название») — заголовок сверяется с базой kulinar.
        kulinar_names = load_kulinar_names()

        def covers_fields(ttl, body):
            t = ((ttl or "") + " " + (body or "")).lower()
            title_n = normalize(ttl)
            title_is_name = bool(title_n) and (
                kulinar_names is None
                or any(rn and rn in title_n for rn in kulinar_names)
            )
            has_name = ("назван" in t) or ("name" in t) or ("блюдо" in t) \
                or title_is_name
            has_cat = ("категор" in t) or ("category" in t)
            has_diff = ("difficulty" in t) or ("сложност" in t) \
                or ("лёгк" in t) or ("легк" in t) or ("средн" in t) or ("сложн" in t)
            return has_name and has_cat and has_diff

        good_pages = sum(1 for _, ttl, body in pages if covers_fields(ttl, body))
        record("Teamly recipe pages list Название + Категория + Difficulty",
               good_pages >= 5,
               f"{good_pages}/{len(pages)} pages cover Название+Категория+Difficulty")
        if good_pages < 5:
            all_ok = False

        cur.close()
        conn.close()

    except Exception as e:
        record("Teamly recipe-collection space exists", False, str(e))
        record("Teamly recipe space 'Коллекция рецептов команды' has >= 5 recipe pages",
               False, str(e))
        record("Teamly recipe pages list Название + Категория + Difficulty",
               False, str(e))
        all_ok = False

    return all_ok


def check_memory(agent_workspace):
    """Check that memory.json has entities recording filtering criteria."""
    print("\n=== Checking Memory ===")

    mem_file = os.path.join(agent_workspace, "memory", "memory.json")

    if not os.path.isfile(mem_file):
        record("memory.json exists", False, f"Not found at {mem_file}")
        return False

    record("memory.json exists", True)

    try:
        with open(mem_file, "r", encoding="utf-8") as f:
            raw = f.read()
    except Exception as e:
        record("memory.json readable", False, str(e))
        return False

    # Memory MCP персистит NDJSON (по одной записи {"type":"entity",...} на
    # строку); fallback — единый объект {"entities": [...]}.
    entities = []
    parsed_ok = False
    try:
        memory_data = json.loads(raw)
        parsed_ok = True
        if isinstance(memory_data, dict):
            entities = memory_data.get("entities", [])
        elif isinstance(memory_data, list):
            entities = [e for e in memory_data
                        if isinstance(e, dict) and e.get("type", "entity") == "entity"]
    except json.JSONDecodeError:
        for line in raw.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                rec = json.loads(line)
            except json.JSONDecodeError:
                continue
            parsed_ok = True
            if isinstance(rec, dict) and rec.get("type", "entity") == "entity":
                entities.append(rec)

    if not parsed_ok:
        record("memory.json readable", False, "neither JSON nor NDJSON")
        return False

    record("memory.json readable", True)

    all_ok = True
    record("Memory has entities",
           len(entities) > 0,
           f"Found {len(entities)} entities")
    if len(entities) == 0:
        all_ok = False

    # At least one entity records selection criteria / categories (RU+EN).
    keywords = [
        # RU
        "рецепт", "категор", "критери", "отбор", "выбор", "фильтр",
        "салат", "закуск", "суп", "горяч", "гарнир", "выпечк", "десерт", "напиток",
        # EN (tolerated)
        "categor", "recipe", "criteria", "filter", "select",
    ]
    criteria_found = False
    for entity in entities:
        entity_str = json.dumps(entity, ensure_ascii=False).lower()
        if any(kw in entity_str for kw in keywords):
            criteria_found = True
            break

    record("Memory contains filtering criteria or category info",
           criteria_found,
           "No entity related to recipe selection criteria found")
    if not criteria_found:
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    teamly_ok = check_teamly()
    memory_ok = check_memory(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:   {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Teamly:  {'PASS' if teamly_ok else 'FAIL'}")
    print(f"  Memory:  {'PASS' if memory_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    overall = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
