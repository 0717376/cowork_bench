"""Evaluation for sf-sla-breach-audit (ClickHouse warehouse, sf_data schema)."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "SLA_Breach_Audit.xlsx")
    gt_file = os.path.join(gt_dir, "SLA_Breach_Audit.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # checks: list of (name, passed, is_critical)
    checks = []

    def add(name, passed, critical=False):
        checks.append((name, bool(passed), critical))

    # --- Breach Analysis sheet ---
    a_rows = load_sheet_rows(agent_wb, "Breach Analysis")
    g_rows = load_sheet_rows(gt_wb, "Breach Analysis")
    add("Breach Analysis sheet exists", a_rows is not None, critical=False)

    a_lookup = {}
    if a_rows is not None and g_rows is not None:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        # structural (NON-critical): rows sorted by SLA_Target_Hours ascending
        targets = []
        for row in a_data:
            if row and row[0] is not None and len(row) > 3:
                try:
                    targets.append(float(row[3]))
                except (TypeError, ValueError):
                    pass
        add("Breach Analysis sorted by SLA_Target_Hours asc",
            targets == sorted(targets) and len(targets) > 0, critical=False)

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            present = a_row is not None
            add(f"Breach Analysis row '{g_row[0]}' present", present, critical=False)
            if not present:
                # critical computations also fail for this priority
                add(f"{key}.Breached_Count match", False, critical=True)
                add(f"{key}.Breach_Rate match", False, critical=True)
                continue

            # Total_Tickets (col 1) — structural/loose
            if len(a_row) > 1 and len(g_row) > 1:
                add(f"{key}.Total_Tickets match", num_close(a_row[1], g_row[1], 10), critical=False)
            # Avg_Response_Hours (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                add(f"{key}.Avg_Response_Hours match", num_close(a_row[2], g_row[2], 0.5), critical=False)
            # SLA_Target_Hours (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                add(f"{key}.SLA_Target_Hours match", num_close(a_row[3], g_row[3], 0.5), critical=False)
            # Breached_Count (col 4) — CRITICAL core computation, tightened tol
            if len(a_row) > 4 and len(g_row) > 4:
                add(f"{key}.Breached_Count match", num_close(a_row[4], g_row[4], 50), critical=True)
            # Breach_Rate (col 5) — CRITICAL core computation, tight tol
            if len(a_row) > 5 and len(g_row) > 5:
                add(f"{key}.Breach_Rate match", num_close(a_row[5], g_row[5], 1.0), critical=True)
    else:
        add("Breach Analysis groundtruth loaded", g_rows is not None, critical=False)

    # --- Summary sheet ---
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    add("Summary sheet exists", a_rows is not None, critical=False)

    if a_rows is not None and g_rows is not None:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        s_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                s_lookup[str(row[0]).strip().lower()] = row

        # critical metrics
        critical_metrics = {"overall_breach_rate", "worst_priority", "best_priority"}
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = s_lookup.get(key)
            is_crit = key in critical_metrics
            if a_row is None:
                add(f"Summary.{key} present", False, critical=is_crit)
                continue
            g_val = g_row[1]
            a_val = a_row[1]
            try:
                float(a_val); float(g_val)
                # Overall_Breach_Rate is the headline deliverable -> tighter tol
                tol = 0.5 if key == "overall_breach_rate" else 1.0
                ok = num_close(a_val, g_val, tol)
            except (TypeError, ValueError):
                # Worst_Priority / Best_Priority are English literals matched exactly
                ok = str_match(a_val, g_val)
            add(f"Summary.{key} match", ok, critical=is_crit)

    # --- Email sent (CRITICAL content) ---
    overall_rate_int = "79"   # Overall_Breach_Rate ~= 79.02
    worst_priority = "medium"  # English literal, lowercased for matching
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432,
                                dbname="cowork_gym", user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT m.subject, m.to_addr, m.body_text, m.from_addr FROM email.messages m")
        all_msgs = cur.fetchall()
        cur.close()
        conn.close()

        found = False
        for subj, to_addr, body, from_addr in all_msgs:
            subj_str = str(subj or "").lower()
            to_str = str(to_addr or "").lower()
            from_str = str(from_addr or "").lower()
            body_str = str(body or "").lower()
            if ("sla breach audit results" in subj_str
                    and "ops-director@company.com" in to_str
                    and "sla-audit@company.com" in from_str
                    and overall_rate_int in body_str
                    and worst_priority in body_str):
                found = True
                break
        add("Email from sla-audit to ops-director with subject+rate(79)+worst(Medium)",
            found, critical=True)
    except Exception as e:
        print(f"Email check error: {e}")
        add("Email check executed", False, critical=True)

    # --- Scoring ---
    total = len(checks)
    passed = sum(1 for _, p, _ in checks if p)
    accuracy = (passed / total * 100.0) if total else 0.0
    critical_fail = [n for n, p, c in checks if c and not p]

    print("=== CHECKS ===")
    for name, p, c in checks:
        tag = "CRIT" if c else "    "
        print(f"  [{'PASS' if p else 'FAIL'}] {tag} {name}")
    print(f"\nAccuracy: {passed}/{total} = {accuracy:.1f}%")

    if critical_fail:
        print(f"=== RESULT: FAIL (critical checks failed: {critical_fail}) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print(f"=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
