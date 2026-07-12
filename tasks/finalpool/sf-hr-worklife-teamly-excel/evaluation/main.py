"""Evaluation for sf-hr-worklife task (ClickHouse + Teamly).

Named checks. Any CRITICAL_CHECKS failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Department names in the ClickHouse DWH are russified centrally
(Engineering->Инженерия, Finance->Финансы, HR->Кадры, Operations->Операции,
R&D->НИОКР, Sales->Продажи, Support->Поддержка). Matching accepts both the RU
and EN department keys so seed/groundtruth/agent output stay in sync.
"""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Bilingual department map (EN <-> RU). Used to canonicalize dept keys so that a
# row keyed by the RU name matches a groundtruth row keyed by EN (or vice versa).
DEPT_EN_RU = {
    "engineering": "инженерия",
    "finance": "финансы",
    "hr": "кадры",
    "operations": "операции",
    "r&d": "ниокр",
    "sales": "продажи",
    "support": "поддержка",
}
DEPT_CANON = {}
for _en, _ru in DEPT_EN_RU.items():
    DEPT_CANON[_en] = _en
    DEPT_CANON[_ru] = _en  # canonical = EN form


def dept_key(name):
    """Canonicalize a department name to its EN form (lowercased), accepting RU or EN."""
    if name is None:
        return None
    k = str(name).strip().lower()
    return DEPT_CANON.get(k, k)


# Best-performing department per the data (raw, unrounded averages on the
# HR_ANALYTICS__PUBLIC__EMPLOYEES seed, 50000 rows):
#   - Highest Avg_WLB: Операции / Operations (raw 4.5434 > Финансы 4.5404).
#   - Highest Avg_Job_Satisfaction: Финансы / Finance (raw 6.5932 > Продажи 6.5874).
# At 2-dp rounding Операции and Финансы tie on WLB (both 4.54), so an agent that
# names either (or a comma-joined tie) is accepted as long as the true raw max is
# present among the listed departments.
BEST_WLB_DEPT_CANON = "operations"
BEST_JS_DEPT_CANON = "finance"

PASS_COUNT = 0
TOTAL = 0
FAILED = []

CRITICAL_CHECKS = {
    "excel_exists",
    "dept_avg_metrics",
    "findings_totals",
    "findings_best_dept",
    "teamly_dashboard",
    "email_to_director",
}


def check(name, ok, detail=""):
    global PASS_COUNT, TOTAL
    TOTAL += 1
    if ok:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAILED.append(name)
        print(f"  [FAIL] {name} {('- ' + detail) if detail else ''}")
    return ok


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    import openpyxl

    # --- Check 1: Excel file + Department Analysis metrics ---
    print("Checking Excel file...")
    agent_file = os.path.join(agent_ws, "WL_Balance_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WL_Balance_Report.xlsx")

    agent_wb = None
    if not check("excel_exists", os.path.exists(agent_file), "WL_Balance_Report.xlsx not found"):
        pass
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # --- Department Analysis sheet ---
        a_rows = load_sheet_rows(agent_wb, "Department Analysis")
        g_rows = load_sheet_rows(gt_wb, "Department Analysis")
        check("dept_sheet_present", a_rows is not None, "Sheet 'Department Analysis' missing")

        if a_rows is not None:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            check("dept_row_count", len(a_data) == 7, f"got {len(a_data)} rows, expected 7")

            # Index agent rows by canonical (bilingual) department key.
            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[dept_key(row[0])] = row

            # CRITICAL: every department's Avg_WLB & Avg_Job_Satisfaction match GT.
            metrics_ok = True
            metrics_detail = []
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = dept_key(g_row[0])
                a_row = a_lookup.get(key)
                if a_row is None:
                    metrics_ok = False
                    metrics_detail.append(f"missing dept {g_row[0]}")
                    continue
                if not num_close(a_row[1], g_row[1], 1):
                    metrics_ok = False
                    metrics_detail.append(f"{g_row[0]} Employee_Count {a_row[1]}!={g_row[1]}")
                if not num_close(a_row[2], g_row[2], 0.05):
                    metrics_ok = False
                    metrics_detail.append(f"{g_row[0]} Avg_WLB {a_row[2]}!={g_row[2]}")
                if not num_close(a_row[3], g_row[3], 0.05):
                    metrics_ok = False
                    metrics_detail.append(f"{g_row[0]} Avg_JS {a_row[3]}!={g_row[3]}")
            check("dept_avg_metrics", metrics_ok, "; ".join(metrics_detail[:5]))

        # --- Findings sheet ---
        a_rows2 = load_sheet_rows(agent_wb, "Findings")
        g_rows2 = load_sheet_rows(gt_wb, "Findings")
        check("findings_sheet_present", a_rows2 is not None, "Sheet 'Findings' missing")

        if a_rows2 is not None:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            a_lookup2 = {}
            for row in a_data2:
                if row and row[0] is not None:
                    a_lookup2[str(row[0]).strip().lower()] = row
            g_data2 = g_rows2[1:] if len(g_rows2) > 1 else []
            g_lookup2 = {}
            for row in g_data2:
                if row and row[0] is not None:
                    g_lookup2[str(row[0]).strip().lower()] = row

            def gv(metric):
                r = g_lookup2.get(metric)
                return r[1] if r else None

            # CRITICAL: Total_Employees ~50000, Departments_Analyzed == 7,
            # Overall_Avg_WLB / Overall_Avg_Job_Satisfaction match GT.
            te = a_lookup2.get("total_employees")
            da = a_lookup2.get("departments_analyzed")
            owlb = a_lookup2.get("overall_avg_wlb")
            ojs = a_lookup2.get("overall_avg_job_satisfaction")
            totals_ok = True
            totals_detail = []
            if te is None or not num_close(te[1], 50000, 1000):
                totals_ok = False
                totals_detail.append(f"Total_Employees={te[1] if te else None}")
            if da is None or not num_close(da[1], 7, 0):
                totals_ok = False
                totals_detail.append(f"Departments_Analyzed={da[1] if da else None}")
            if owlb is None or not num_close(owlb[1], gv("overall_avg_wlb"), 0.05):
                totals_ok = False
                totals_detail.append(f"Overall_Avg_WLB={owlb[1] if owlb else None}")
            if ojs is None or not num_close(ojs[1], gv("overall_avg_job_satisfaction"), 0.05):
                totals_ok = False
                totals_detail.append(f"Overall_Avg_JS={ojs[1] if ojs else None}")
            check("findings_totals", totals_ok, "; ".join(totals_detail))

            # CRITICAL: Best_WLB_Department -> Операции/Operations (true raw max),
            # Best_JS_Department -> Финансы/Finance (true raw max). A tie/comma-joined
            # value passes if the expected canonical dept appears among its parts.
            def dept_matches(cell, expected_canon):
                if cell is None or cell[1] is None:
                    return False
                parts = str(cell[1]).replace("/", ",").split(",")
                return any(dept_key(p) == expected_canon for p in parts if p.strip())

            bwlb = a_lookup2.get("best_wlb_department")
            bjs = a_lookup2.get("best_js_department")
            best_ok = (
                dept_matches(bwlb, BEST_WLB_DEPT_CANON)
                and dept_matches(bjs, BEST_JS_DEPT_CANON)
            )
            check(
                "findings_best_dept", best_ok,
                f"Best_WLB={bwlb[1] if bwlb else None} (expected Операции/Operations), "
                f"Best_JS={bjs[1] if bjs else None} (expected Финансы/Finance)",
            )

    # --- Check 2: Teamly dashboard page ---
    print("Checking Teamly page...")
    teamly_ok = False
    teamly_detail = ""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Agent-created pages live above the seeded id range (id > 3).
        cur.execute(
            "SELECT title, COALESCE(body, '') FROM teamly.pages WHERE id > 3"
        )
        pages = cur.fetchall()
        cur.close()
        conn.close()

        # Title must reference the wellbeing dashboard (RU or EN).
        def has_any(text, options):
            t = text.lower()
            return any(o in t for o in options)

        title_opts = ["hr wellbeing dashboard", "wellbeing", "благополуч", "work-life", "work life", "баланс"]
        wlb_opts = ["work-life", "work life", "wlb", "баланс"]
        js_opts = ["job satisfaction", "job-satisfaction", "satisfaction", "удовлетвор"]
        best_opts = ["финансы", "finance"]

        for title, body in pages:
            combined = f"{title}\n{body}"
            if not has_any(title, title_opts):
                continue
            # Page content must mention WLB metric, job-satisfaction metric,
            # and name the best-performing department.
            if (has_any(combined, wlb_opts)
                    and has_any(combined, js_opts)
                    and has_any(combined, best_opts)):
                teamly_ok = True
                break
        if not teamly_ok:
            teamly_detail = f"no dashboard page with WLB+JS+best-dept content among {len(pages)} agent pages"
    except Exception as e:
        teamly_detail = f"error: {e}"
    check("teamly_dashboard", teamly_ok, teamly_detail)

    # --- Check 3: Email to HR director (recipient + subject + body) ---
    print("Checking email...")
    email_ok = False
    email_detail = ""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(subject,''), COALESCE(body_text,'')
            FROM email.messages
            WHERE to_addr::text ILIKE '%hr_director@company.com%'
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        subj_opts = ["wellbeing", "well-being", "work-life", "work life",
                     "благополуч", "баланс", "удовлетвор"]
        body_opts = ["work-life", "work life", "wlb", "баланс",
                     "satisfaction", "удовлетвор"]
        for subject, body in rows:
            s = subject.lower()
            b = body.lower()
            if any(o in s for o in subj_opts) and any(o in b for o in body_opts):
                email_ok = True
                break
        if not email_ok:
            email_detail = (f"{len(rows)} message(s) to director, but none with a "
                            f"wellbeing subject + findings body")
    except Exception as e:
        email_detail = f"error: {e}"
    check("email_to_director", email_ok, email_detail)

    # --- Final result ---
    accuracy = PASS_COUNT / TOTAL * 100 if TOTAL else 0
    critical_failed = [n for n in FAILED if n in CRITICAL_CHECKS]
    print(f"\n=== Results: {PASS_COUNT}/{TOTAL} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {critical_failed}")
        print("=== RESULT: FAIL ===")
        sys.exit(1)
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print(f"=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
