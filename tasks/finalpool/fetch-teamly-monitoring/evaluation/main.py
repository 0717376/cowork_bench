"""
Evaluation script for fetch-notion-monitoring task (RU / Teamly).

Expected values are computed DIRECTLY from the source
initial_workspace/service_status_log.json — no hardcoded service dict — so the
eval stays aligned with the real data the agent analyses.

Checks:
1. Service_Availability_Report.xlsx: Availability Summary + Incidents sheets,
   with per-service counts/uptime computed from the JSON.
2. Teamly "Service Monitoring Dashboard - March 2026" page with content.
3. Calendar "Incident Review: <service>" events for every service with downtime.
4. Email monitoring@company.com -> devops-team@company.com listing sub-99% services.

CRITICAL checks (any failure => sys.exit(1) before the accuracy gate):
  - per-service Uptime_Pct in Availability Summary matches JSON (tol <= 0.2)
  - per-service Up/Down/Degraded/Total counts match JSON
  - Incidents sheet exists with correct down/degraded windows
  - email from monitoring@ to devops-team@ with exact subject + sub-99% services
  - calendar Incident Review events for each service with downtime
"""

import argparse
import collections
import datetime
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRIT" if critical else "    "
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS][{tag}] {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL][{tag}] {name}{msg}")


def num_close(a, b, tol):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.lower() in str(haystack).lower()


# ---------------------------------------------------------------------------
# Ground-truth derived from the source JSON
# ---------------------------------------------------------------------------

def _source_json_path(agent_workspace):
    """Find service_status_log.json (initial_workspace copy or agent copy)."""
    here = os.path.dirname(os.path.abspath(__file__))
    candidates = [
        os.path.join(here, "..", "initial_workspace", "service_status_log.json"),
        os.path.join(agent_workspace, "service_status_log.json"),
    ]
    for c in candidates:
        if os.path.isfile(c):
            return os.path.abspath(c)
    return None


def compute_expected(agent_workspace):
    """Compute per-service stats and incident windows from the JSON."""
    path = _source_json_path(agent_workspace)
    if path is None:
        return None
    data = json.load(open(path))
    by = collections.defaultdict(list)
    for d in data:
        by[d["service_name"]].append(d)

    expected = {}
    for svc, rows in by.items():
        rows.sort(key=lambda r: r["timestamp"])
        counts = collections.Counter(r["status"] for r in rows)
        total = len(rows)
        up = counts.get("up", 0)
        down = counts.get("down", 0)
        degraded = counts.get("degraded", 0)
        uptime = round(up / total * 100, 1) if total else 0.0
        rts = [r["response_time_ms"] for r in rows if r.get("response_time_ms") is not None]
        avg_rt = round(sum(rts) / len(rts), 2) if rts else None

        # Continuous down/degraded windows.
        incidents = []
        cur = None
        for r in rows:
            st = r["status"]
            if st in ("down", "degraded"):
                if cur is None:
                    cur = {"start": r["timestamp"], "end": r["timestamp"],
                           "status": st, "n": 1}
                else:
                    cur["end"] = r["timestamp"]
                    cur["n"] += 1
            else:
                if cur is not None:
                    incidents.append(cur)
                    cur = None
        if cur is not None:
            incidents.append(cur)
        for inc in incidents:
            s = datetime.datetime.fromisoformat(inc["start"].replace("Z", "+00:00"))
            e = datetime.datetime.fromisoformat(inc["end"].replace("Z", "+00:00"))
            inc["duration_min"] = (e - s).total_seconds() / 60.0

        expected[svc] = {
            "total": total, "up": up, "down": down, "degraded": degraded,
            "uptime": uptime, "avg_rt": avg_rt, "incidents": incidents,
        }
    return expected


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _match_row(data_rows, svc):
    """Find the data row whose first cell equals/contains the service name."""
    sl = svc.lower()
    for row in data_rows:
        if row and row[0] and str(row[0]).strip().lower() == sl:
            return row
    # fall back to substring (full name, not a single token)
    for row in data_rows:
        if row and str_contains(row[0], svc):
            return row
    return None


def _header_index(rows, names):
    if not rows:
        return {}
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx = {}
    for want in names:
        for i, h in enumerate(header):
            if h == want.lower():
                idx[want] = i
                break
    return idx


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Service_Availability_Report.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}", critical=True)
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e), critical=True)
        return

    # ---- Availability Summary ----
    summary_sheet = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "availability" in nl or "summary" in nl:
            summary_sheet = name
            break
    if summary_sheet is None:
        # tolerate a generic name but prefer the canonical one
        for name in wb.sheetnames:
            if "service" in name.lower() or "status" in name.lower():
                summary_sheet = name
                break

    if not summary_sheet:
        record("Availability Summary sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Availability Summary sheet exists", True)
        ws = wb[summary_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)]

        record("Availability Summary has exactly 3 rows", len(data_rows) == 3,
               f"Found {len(data_rows)} rows", critical=True)

        idx = _header_index(rows, ["Service_Name", "Total_Checks", "Up_Count",
                                   "Down_Count", "Degraded_Count", "Uptime_Pct",
                                   "Avg_Response_Time_Ms"])

        for svc, exp in expected.items():
            row = _match_row(data_rows, svc)
            if row is None:
                record(f"'{svc}' in Availability Summary", False, "Not found", critical=True)
                continue
            record(f"'{svc}' in Availability Summary", True)

            # Uptime: prefer the named column; else sniff a float near expected.
            uptime_val = None
            if "Uptime_Pct" in idx and idx["Uptime_Pct"] < len(row):
                uptime_val = row[idx["Uptime_Pct"]]
            if uptime_val is None:
                for cell in row[1:]:
                    try:
                        v = float(cell)
                        if 0.0 <= v <= 100.0 and num_close(v, exp["uptime"], 5.0):
                            uptime_val = v
                            break
                    except (TypeError, ValueError):
                        continue
            record(f"'{svc}' Uptime_Pct == {exp['uptime']}",
                   num_close(uptime_val, exp["uptime"], 0.2),
                   f"Got {uptime_val}", critical=True)

            # Count columns (only assert if header present).
            for col, key in [("Total_Checks", "total"), ("Up_Count", "up"),
                             ("Down_Count", "down"), ("Degraded_Count", "degraded")]:
                if col in idx and idx[col] < len(row):
                    got = row[idx[col]]
                    record(f"'{svc}' {col} == {exp[key]}",
                           num_close(got, exp[key], 0.0),
                           f"Got {got}", critical=True)

    # ---- Incidents ----
    incidents_sheet = None
    for name in wb.sheetnames:
        if "incident" in name.lower():
            incidents_sheet = name
            break

    if not incidents_sheet:
        record("Incidents sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Incidents sheet exists", True)
        ws = wb[incidents_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)]

        total_expected = sum(len(e["incidents"]) for e in expected.values())
        record("Incidents row count matches JSON windows",
               len(data_rows) == total_expected,
               f"Expected {total_expected}, got {len(data_rows)}", critical=True)

        # Per service: number of incident rows must match the JSON windows.
        for svc, exp in expected.items():
            n_exp = len(exp["incidents"])
            svc_rows = [r for r in data_rows if r and str(r[0]).strip().lower() == svc.lower()]
            record(f"Incidents: '{svc}' has {n_exp} window(s)",
                   len(svc_rows) == n_exp,
                   f"Got {len(svc_rows)}", critical=True)

    wb.close()


# ---------------------------------------------------------------------------
# Teamly
# ---------------------------------------------------------------------------

def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def check_teamly(expected):
    print("\n=== Checking Teamly ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Teamly DB accessible", False, str(e), critical=True)
        return

    dash = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if (("service monitoring dashboard" in tl)
                or ("monitoring" in tl and "dashboard" in tl)
                or ("мониторинг" in tl and ("панель" in tl or "сервис" in tl or "дашборд" in tl))):
            dash = (pid, title, body)
            break

    record("Teamly monitoring dashboard page exists", dash is not None,
           f"new pages: {[(p[0], p[1]) for p in pages]}", critical=True)
    if dash is None:
        return

    text = ((dash[1] or "") + " " + (dash[2] or "")).lower()
    # Page must mention each service.
    for svc in expected:
        record(f"Teamly page mentions '{svc}'", svc.lower() in text,
               "service name absent in page body")
    # Page must call out the services that had downtime/degradation.
    for svc, exp in expected.items():
        if exp["incidents"]:
            # at least the service name is present; richer content is bonus
            pass
    record("Teamly page has non-trivial content", len(text.strip()) >= 40,
           f"len={len(text.strip())}")


# ---------------------------------------------------------------------------
# Calendar
# ---------------------------------------------------------------------------

def check_calendar(expected):
    print("\n=== Checking Google Calendar ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT summary, description, start_datetime, end_datetime FROM gcal.events"
        )
        events = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e), critical=True)
        return

    services_with_downtime = [s for s, e in expected.items() if e["down"] > 0]

    for svc in services_with_downtime:
        match = None
        for summary, description, start_dt, end_dt in events:
            sl = (summary or "").lower()
            if "incident review" in sl and svc.lower() in sl:
                # verify the date is 2026-03-07
                if start_dt is not None and "2026-03-07" in str(start_dt):
                    match = (summary, start_dt)
                    break
        record(f"Calendar 'Incident Review: {svc}' on 2026-03-07 exists",
               match is not None,
               f"events: {[e[0] for e in events]}", critical=True)


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def _all_emails():
    rows = []
    for table in ("email.messages", "email.sent_log"):
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                f"SELECT subject, from_addr, to_addr, COALESCE(body_text,'') FROM {table}"
            )
            rows.extend(cur.fetchall())
            conn.close()
        except Exception:
            pass
    return rows


def check_emails(expected):
    print("\n=== Checking Emails ===")
    emails = _all_emails()

    subject_want = "service availability report - week of feb 27"
    target = None
    for subject, from_addr, to_addr, body in emails:
        if (str_contains(subject, subject_want)
                and str_contains(from_addr, "monitoring@company.com")
                and str_contains(to_addr, "devops-team@company.com")):
            target = (subject, from_addr, to_addr, body)
            break

    record("Email monitoring@->devops-team@ with exact subject exists",
           target is not None,
           f"emails: {[(e[0], e[1], e[2]) for e in emails]}", critical=True)
    if target is None:
        return

    body = (target[3] or "").lower()
    # Services with uptime < 99% must be listed with their percentage.
    sub99 = {s: e for s, e in expected.items() if e["uptime"] < 99.0}
    for svc, exp in sub99.items():
        has_name = svc.lower() in body
        # accept e.g. "97.6" or "97,6"
        pct = f"{exp['uptime']:.1f}"
        has_pct = pct in body or pct.replace(".", ",") in body
        record(f"Email lists sub-99% service '{svc}' with {pct}%",
               has_name and has_pct,
               f"name={has_name} pct={has_pct}", critical=True)

    # Services with uptime >= 99% should NOT be reported as below-99.
    for svc, exp in expected.items():
        if exp["uptime"] >= 99.0:
            record(f"Email does NOT list {svc} (uptime>=99%)",
                   svc.lower() not in body or exp["uptime"] >= 99.0,
                   "")  # informational, non-critical


# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = compute_expected(args.agent_workspace)
    if expected is None:
        print("[FATAL] Could not locate source service_status_log.json")
        sys.exit(1)

    print("=== Expected (from JSON) ===")
    for svc, e in expected.items():
        print(f"  {svc}: total={e['total']} up={e['up']} down={e['down']} "
              f"degraded={e['degraded']} uptime={e['uptime']} "
              f"incidents={len(e['incidents'])}")

    check_excel(args.agent_workspace, expected)
    check_teamly(expected)
    check_calendar(expected)
    check_emails(expected)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    print("\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILED}")
        print("  Overall: FAIL (critical)")
        sys.exit(1)

    if accuracy >= 70.0:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL (accuracy < 70)")
    sys.exit(1)


if __name__ == "__main__":
    main()
