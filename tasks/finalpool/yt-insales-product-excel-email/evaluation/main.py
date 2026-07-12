"""
Evaluation for yt-insales-product-excel-email task.

Checks:
1. Marketing_Opportunity_Report.xlsx exists with 3 sheets
2. Video_Topics sheet has 10 rows (Rank 1-10) with required columns
3. Product_Matches sheet has required columns
4. Summary sheet has 3 rows with correct labels
5. Email to marketing@company.com with correct subject and content
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}{msg}")
        if critical:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Marketing_Opportunity_Report.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Marketing_Opportunity_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Marketing_Opportunity_Report.xlsx exists", False, f"Not found at {xlsx_path}", critical=True)
        return
    record("Marketing_Opportunity_Report.xlsx exists", True, critical=True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e), critical=True)
        return
    record("Excel file readable", True, critical=True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Video_Topics sheet
    vt_idx = next((i for i, s in enumerate(sheet_names_lower)
                   if "video_topic" in s or "video topic" in s), None)
    if vt_idx is None:
        record("Video_Topics sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Video_Topics sheet exists", True, critical=True)
        ws = wb[wb.sheetnames[vt_idx]]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in rows[0]] if rows else []
        has_rank = any("rank" in h for h in headers)
        has_title = any("title" in h for h in headers)
        has_view = any("view" in h for h in headers)
        has_topic = any("topic" in h for h in headers)
        record("Video_Topics has required columns (Rank, Title, View_Count, Main_Topic)",
               has_rank and has_title and has_view and has_topic,
               f"Headers: {rows[0] if rows else 'empty'}")

        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Video_Topics has at least 8 rows (top videos)", len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")

        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        has_ai_topic = "ai" in all_text
        has_deepseek = "deepseek" in all_text
        record("Video_Topics contains AI-related videos", has_ai_topic or has_deepseek,
               "No AI topic found")

        # CRITICAL semantic: the #1 video by view count (sorted by Rank) must be the
        # real top Fireship 2024+ video ("Big Tech in panic mode... DeepSeek R1 ...").
        # A non-doer / preseeded sheet will not reproduce the correct ordering.
        rank1_text = ""
        for r in data_rows:
            cells = [str(c).strip() if c is not None else "" for c in r]
            # find the row whose Rank column equals 1
            if cells and (cells[0] == "1" or cells[0] == "1.0"):
                rank1_text = " ".join(cells).lower()
                break
        if not rank1_text and data_rows:
            rank1_text = " ".join(str(c) for c in data_rows[0] if c).lower()
        record("Video_Topics rank-1 video is the real top Fireship video (DeepSeek R1)",
               "deepseek" in rank1_text and ("panic" in rank1_text or "bubble" in rank1_text),
               f"Rank-1 row text: {rank1_text[:200]}", critical=True)

    # Check Product_Matches sheet
    pm_idx = next((i for i, s in enumerate(sheet_names_lower)
                   if "product_match" in s or "product match" in s), None)
    if pm_idx is None:
        record("Product_Matches sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Product_Matches sheet exists", True)
        ws_pm = wb[wb.sheetnames[pm_idx]]
        pm_rows = list(ws_pm.iter_rows(values_only=True))
        pm_headers = [str(c).strip().lower() if c else "" for c in pm_rows[0]] if pm_rows else []
        has_product = any("product" in h for h in pm_headers)
        has_match_kw = any("keyword" in h or "match" in h for h in pm_headers)
        record("Product_Matches has required columns (Product_Name, Match_Keyword)",
               has_product and has_match_kw,
               f"Headers: {pm_rows[0] if pm_rows else 'empty'}")

    # Check Summary sheet
    sum_idx = next((i for i, s in enumerate(sheet_names_lower) if "summary" in s), None)
    if sum_idx is None:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Summary sheet exists", True)
        ws_sum = wb[wb.sheetnames[sum_idx]]
        sum_rows = list(ws_sum.iter_rows(values_only=True))
        sum_text = " ".join(str(c) for r in sum_rows for c in r if c).lower()
        has_total_videos = "total_videos_analyzed" in sum_text or "total videos" in sum_text
        has_matches = "total_product_matches" in sum_text or "total product" in sum_text
        has_topic = "most_common_topic" in sum_text or "most common" in sum_text
        record("Summary has 3 required labels",
               has_total_videos and has_matches and has_topic,
               f"Summary text: {sum_text[:300]}")

        # CRITICAL semantic: Most_Common_Topic across the real top-10 Fireship 2024+
        # videos must be "AI" (5 of 10 GT videos classify as AI). Also Total_Videos = 10.
        mct_value = None
        videos_value = None
        for r in sum_rows:
            cells = [str(c).strip() if c is not None else "" for c in r]
            joined = " ".join(cells).lower()
            if "most_common_topic" in joined or "most common" in joined:
                rest = [c for c in cells if "most" not in c.lower()]
                mct_value = " ".join(rest).strip().lower()
            if "total_videos_analyzed" in joined or "total videos" in joined:
                rest = [c for c in cells if c.strip().isdigit()]
                videos_value = rest[0] if rest else None
        record("Summary Most_Common_Topic is 'AI'",
               mct_value is not None and mct_value == "ai",
               f"Most_Common_Topic value: {mct_value}", critical=True)
        record("Summary Total_Videos_Analyzed is 10",
               videos_value == "10",
               f"Total_Videos_Analyzed value: {videos_value}")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Marketing_Opportunity_Report.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_email():
    print("\n=== Check 2: Email to marketing@company.com ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT to_addr, subject, body_text FROM email.messages
        WHERE to_addr::text ILIKE '%marketing@company.com%'
        AND subject ILIKE '%marketing%'
        ORDER BY id DESC LIMIT 5
    """)
    emails = cur.fetchall()
    cur.close()
    conn.close()

    record("Email to marketing@company.com with marketing subject",
           len(emails) > 0, "No matching email found", critical=True)

    if emails:
        to_addr, subject, body = emails[0]
        record("Email subject is 'Tech Video Marketing Opportunities'",
               "tech video marketing opportunities" in subject.lower(),
               f"Subject: {subject}", critical=True)

        body_lower = (body or "").lower()
        # Body should mention product-video matches
        has_product = any(kw in body_lower for kw in
                          ["laptop", "tv", "headphone", "monitor", "adapter", "usb", "hub"])
        has_video = any(kw in body_lower for kw in
                        ["deepseek", "linux", "microsoft", "windows", "vibe"])
        record("Email body describes video-product matches",
               has_product or has_video,
               f"Body excerpt: {body_lower[:300]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # CRITICAL gate: any failed critical check => hard FAIL regardless of accuracy.
    if CRITICAL_FAILED:
        print(f"\nCRITICAL CHECK(S) FAILED: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
