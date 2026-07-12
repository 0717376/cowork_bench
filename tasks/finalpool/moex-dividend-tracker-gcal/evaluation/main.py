"""
Evaluation script for the MOEX dividend-tracker-gcal task.

Data source: moex.stock_info (RU MOEX tickers). The agent must read MOEX
field names, which differ from Yahoo's:
  - lastDividendValue           -> Dividend_Rate
  - fiveYearAvgDividendYield    -> Dividend_Yield (percent)
  - exDividendDate (unix epoch) -> Ex_Dividend_Date (YYYY-MM-DD, UTC)
  - payoutRatio                 -> Payout_Ratio
A stock is "dividend-paying" iff lastDividendValue > 0 AND exDividendDate is set.

Checks:
1. Excel file (Dividend_Tracker.xlsx) - 'Dividend Stocks' and 'Summary' sheets.
2. Google Calendar all-day events per dividend stock on the ex-dividend date.
3. Google Sheet titled 'Dividend Watch List' with an 'Overview' sheet replicating the data.

Scoring: accuracy >= 70% AND no CRITICAL check failed => PASS.
"""

import argparse
import datetime
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

TICKERS = ("SBER.ME", "GAZP.ME", "LKOH.ME", "MGNT.ME", "MTSS.ME")

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{detail_str}")


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def epoch_to_iso(epoch):
    return datetime.datetime.fromtimestamp(int(epoch), datetime.timezone.utc).strftime("%Y-%m-%d")


def get_expected_dividends():
    """Expected dividend data from moex.stock_info using MOEX field names.

    A stock pays a dividend iff lastDividendValue > 0 AND exDividendDate is set.
    Returns rows: (symbol, name, div_rate, div_yield, ex_div_iso, payout_ratio)
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT symbol,
               data->>'shortName' as name,
               (data->>'lastDividendValue')::float as div_rate,
               (data->>'fiveYearAvgDividendYield')::float as div_yield,
               (data->>'exDividendDate')::bigint as ex_div_epoch,
               (data->>'payoutRatio')::float as payout_ratio
        FROM moex.stock_info
        WHERE symbol IN %s
          AND (data->>'lastDividendValue') IS NOT NULL
          AND (data->>'lastDividendValue')::float > 0
          AND (data->>'exDividendDate') IS NOT NULL
        ORDER BY symbol
        """,
        (TICKERS,),
    )
    rows = []
    for sym, name, rate, yld, epoch, payout in cur.fetchall():
        rows.append((sym, name, rate, yld, epoch_to_iso(epoch), payout))
    cur.close()
    conn.close()
    return rows


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def check_excel(agent_workspace, expected_rows):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Dividend_Tracker.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        check("Excel deliverable produced", False, "missing file", critical=True)
        return

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        check("Excel deliverable produced", False, str(e), critical=True)
        return

    ws = get_sheet(wb, "Dividend Stocks")
    check("Sheet 'Dividend Stocks' exists", ws is not None, f"Sheets: {wb.sheetnames}")
    if ws is None:
        check("Dividend Stocks sheet present", False, "missing", critical=True)
        return

    # Map header -> column index
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(header)}

    data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
    num_expected = len(expected_rows)
    check(
        f"Dividend Stocks has >= {num_expected} rows",
        len(data_rows) >= num_expected,
        f"Got {len(data_rows)} rows, expected >= {num_expected}",
    )

    # Index excel rows by ticker
    by_ticker = {}
    for r in data_rows:
        t = str(r[0]).strip().upper()
        by_ticker[t] = r

    expected_tickers = {r[0] for r in expected_rows}
    for t in expected_tickers:
        check(f"Ticker {t} in Dividend Stocks", t in by_ticker, f"Found: {set(by_ticker)}")

    # CRITICAL: real per-ticker values (Dividend_Rate and Ex_Dividend_Date) correct.
    rate_ok = 0
    date_ok = 0
    has_rate_col = "Dividend_Rate" in col
    has_date_col = "Ex_Dividend_Date" in col
    for sym, name, exp_rate, exp_yield, exp_iso, payout in expected_rows:
        row = by_ticker.get(sym)
        if not row:
            continue
        if has_rate_col:
            got = _num(row[col["Dividend_Rate"]])
            if got is not None and abs(got - exp_rate) <= max(0.01, abs(exp_rate) * 0.02):
                rate_ok += 1
        if has_date_col:
            cell = row[col["Ex_Dividend_Date"]]
            cell_s = str(cell)
            if isinstance(cell, (datetime.datetime, datetime.date)):
                cell_s = cell.strftime("%Y-%m-%d")
            if exp_iso in cell_s:
                date_ok += 1
    check(
        "Dividend_Rate values match source (lastDividendValue)",
        has_rate_col and rate_ok == num_expected,
        f"matched {rate_ok}/{num_expected}, col_present={has_rate_col}",
        critical=True,
    )
    check(
        "Ex_Dividend_Date values match source (exDividendDate epoch->date)",
        has_date_col and date_ok == num_expected,
        f"matched {date_ok}/{num_expected}, col_present={has_date_col}",
        critical=True,
    )

    # Summary sheet
    ws2 = get_sheet(wb, "Summary")
    check("Sheet 'Summary' exists", ws2 is not None, f"Sheets: {wb.sheetnames}")
    if ws2 is None:
        check("Summary sheet present", False, "missing", critical=True)
    else:
        summary = {}
        for row in ws2.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                key = str(row[0]).strip().lower().replace(" ", "_")
                summary[key] = row[1]

        def find(*subs):
            for k, v in summary.items():
                if all(s in k for s in subs):
                    return v
            return None

        total_v = find("total", "dividend")
        avg_v = find("avg", "yield")
        high_v = find("highest", "yield")

        check("Summary has Total_Dividend_Stocks", total_v is not None, f"Keys: {list(summary)}")
        check("Summary has Avg_Yield", avg_v is not None, f"Keys: {list(summary)}")
        check("Summary has Highest_Yield_Ticker", high_v is not None, f"Keys: {list(summary)}")

        # CRITICAL: computed Summary values are correct.
        exp_total = num_expected
        exp_high = max(expected_rows, key=lambda r: r[3])[0] if expected_rows else None
        gt = _num(total_v)
        check(
            "Total_Dividend_Stocks equals count of dividend payers",
            gt is not None and int(round(gt)) == exp_total,
            f"got {total_v}, expected {exp_total}",
            critical=True,
        )
        high_s = str(high_v).strip().upper() if high_v is not None else ""
        check(
            "Highest_Yield_Ticker equals max-yield ticker",
            exp_high is not None and exp_high in high_s,
            f"got {high_v!r}, expected {exp_high}",
            critical=True,
        )


def check_calendar(expected_rows):
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime FROM gcal.events")
    events = cur.fetchall()
    cur.close()
    conn.close()
    print(f"[check_calendar] Found {len(events)} events.")

    dividend_events = [
        e for e in events if e[0] and ("dividend" in e[0].lower() or "ex-dividend" in e[0].lower())
    ]
    check(
        "At least 3 calendar events with dividend in summary",
        len(dividend_events) >= 3,
        f"Found {len(dividend_events)} dividend events",
    )

    # CRITICAL: one all-day event per dividend stock, titled 'Ex-Dividend: <TICKER>',
    # whose start date equals that ticker's ex-dividend date.
    matched = 0
    for sym, name, rate, yld, exp_iso, payout in expected_rows:
        for summ, desc, start in events:
            s = (summ or "")
            start_s = str(start)
            if sym.upper() in s.upper() and "dividend" in s.lower() and exp_iso in start_s:
                matched += 1
                break
    check(
        "Each dividend stock has an Ex-Dividend event on its ex-div date",
        matched == len(expected_rows),
        f"matched {matched}/{len(expected_rows)}",
        critical=True,
    )


def check_gsheet(expected_rows):
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    spreadsheets = cur.fetchall()

    dividend_sheets = [s for s in spreadsheets if s[1] and "dividend" in s[1].lower()]
    check(
        "Google Sheet with 'dividend' in title exists",
        len(dividend_sheets) > 0,
        f"Titles: {[s[1] for s in spreadsheets]}",
    )

    # Gather all cell text across the dividend spreadsheet(s) and verify each
    # dividend ticker is actually present in the data (data replicated, not just title).
    found_tickers = set()
    overview_present = False
    if dividend_sheets:
        ids = tuple(s[0] for s in dividend_sheets)
        cur.execute("SELECT id, spreadsheet_id, title FROM gsheet.sheets WHERE spreadsheet_id IN %s", (ids,))
        sub_sheets = cur.fetchall()
        overview_present = any(t and "overview" in t.lower() for _, _, t in sub_sheets)
        if sub_sheets:
            sheet_ids = tuple(s[0] for s in sub_sheets)
            cur.execute("SELECT value FROM gsheet.cells WHERE sheet_id IN %s", (sheet_ids,))
            blob = " ".join(str(v[0]) for v in cur.fetchall() if v[0] is not None).upper()
            for sym, *_ in expected_rows:
                if sym.upper() in blob:
                    found_tickers.add(sym)

    cur.close()
    conn.close()

    check("Google Sheet has an 'Overview' sheet", overview_present, "no Overview tab found")
    check(
        "Overview replicates all dividend tickers",
        len(found_tickers) == len(expected_rows),
        f"found {sorted(found_tickers)} of {len(expected_rows)}",
        critical=True,
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_rows = get_expected_dividends()
    print(f"Expected {len(expected_rows)} dividend-paying MOEX stocks from DB: "
          f"{[r[0] for r in expected_rows]}")

    if len(expected_rows) < 3:
        # Data-layer sanity guard: the deliverable must be non-trivial.
        print("[eval] WARNING: fewer than 3 expected dividend payers — check seed data.")

    check_excel(args.agent_workspace, expected_rows)
    check_calendar(expected_rows)
    check_gsheet(expected_rows)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_ok = len(CRITICAL_FAILED) == 0
    success = critical_ok and accuracy >= 70.0

    print("\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILED:
        print(f"  CRITICAL failures: {CRITICAL_FAILED}")
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "critical_failed": CRITICAL_FAILED,
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if not critical_ok:
        sys.exit(1)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
