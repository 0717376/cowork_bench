"""
Evaluation for insales-sales-category-excel-email task.

Checks:
1. Excel file Category_Performance.xlsx with correct data
2. Email sent to sales-team@shop.com mentioning highest/lowest revenue categories
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Semantic critical checks (any fail => hard FAIL regardless of accuracy)
CRITICAL_CHECKS = {
    "CRITICAL: highest-revenue category is 'ТВ и домашний кинотеатр' (top row)",
    "CRITICAL: lowest-revenue category is 'Бытовая техника' (bottom row)",
    "CRITICAL: Category Summary has exactly 6 categories sorted by Total_Revenue descending",
    "CRITICAL: 'ТВ и домашний кинотеатр' Total_Revenue matches groundtruth (tight tol)",
    "CRITICAL: Email sent to sales-team@shop.com with subject 'Category Performance Report'",
    "CRITICAL: Email body names highest-revenue category (ТВ и домашний кинотеатр)",
    "CRITICAL: Email body names lowest-revenue category (Бытовая техника)",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL] {name}: {d}")


def _to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel File ===")
    agent_file = os.path.join(agent_workspace, "Category_Performance.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Category_Performance.xlsx")

    if not os.path.exists(agent_file):
        check("Excel file exists", False, f"Not found: {agent_file}")
        return
    check("Excel file exists", True)

    if not os.path.exists(gt_file):
        check("Groundtruth file exists", False, f"Not found: {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Category Summary sheet
    print("  Checking Category Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Category Summary")
    g_rows = load_sheet_rows(gt_wb, "Category Summary")

    if a_rows is None:
        check("Sheet 'Category Summary' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth sheet exists", False, "Not found")
    else:
        check("Sheet 'Category Summary' exists", True)
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        g_data = [r for r in (g_rows[1:] if len(g_rows) > 1 else []) if r and r[0] is not None]

        check("Category row count matches", len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        # --- Semantic critical checks on Category Summary ---
        EXPECTED_CATS = {
            "тв и домашний кинотеатр", "электроника", "аудио",
            "камеры", "часы", "бытовая техника",
        }
        a_cats = [str(r[0]).strip().lower() for r in a_data]

        check("CRITICAL: Category Summary has exactly 6 categories sorted by Total_Revenue descending",
              set(a_cats) == EXPECTED_CATS and len(a_data) == 6
              and all(_to_float(a_data[i][3]) is not None for i in range(len(a_data)))
              and all(_to_float(a_data[i][3]) >= _to_float(a_data[i + 1][3]) - 1e-6
                      for i in range(len(a_data) - 1)),
              f"cats={a_cats}, revenues={[_to_float(r[3]) for r in a_data]}")

        check("CRITICAL: highest-revenue category is 'ТВ и домашний кинотеатр' (top row)",
              len(a_cats) > 0 and a_cats[0] == "тв и домашний кинотеатр",
              f"top row = {a_cats[0] if a_cats else None}")

        check("CRITICAL: lowest-revenue category is 'Бытовая техника' (bottom row)",
              len(a_cats) > 0 and a_cats[-1] == "бытовая техника",
              f"bottom row = {a_cats[-1] if a_cats else None}")

        # Tight tolerance on the headline (highest-revenue) category's Total_Revenue
        a_head = next((r for r in a_data if str(r[0]).strip().lower() == "тв и домашний кинотеатр"), None)
        g_head = next((r for r in g_data if str(r[0]).strip().lower() == "тв и домашний кинотеатр"), None)
        gt_rev = _to_float(g_head[3]) if g_head else None
        head_rev = _to_float(a_head[3]) if a_head else None
        head_tol = max(1.0, abs(gt_rev) * 0.01) if gt_rev is not None else 1.0
        check("CRITICAL: 'ТВ и домашний кинотеатр' Total_Revenue matches groundtruth (tight tol)",
              a_head is not None and g_head is not None
              and num_close(head_rev, gt_rev, head_tol),
              f"{head_rev} vs {gt_rev} (tol {head_tol})")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Row '{g_row[0]}'", False, "Missing")
                continue

            # Product_Count (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"{key}.Product_Count",
                      num_close(a_row[1], g_row[1], 2),
                      f"{a_row[1]} vs {g_row[1]}")

            # Total_Units_Sold (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Total_Units_Sold",
                      num_close(a_row[2], g_row[2], 50),
                      f"{a_row[2]} vs {g_row[2]}")

            # Total_Revenue (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Total_Revenue",
                      num_close(a_row[3], g_row[3], 500),
                      f"{a_row[3]} vs {g_row[3]}")

            # Avg_Price (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"{key}.Avg_Price",
                      num_close(a_row[4], g_row[4], 5.0),
                      f"{a_row[4]} vs {g_row[4]}")

    # Check Top Products sheet
    print("  Checking Top Products sheet...")
    a_rows = load_sheet_rows(agent_wb, "Top Products")
    g_rows = load_sheet_rows(gt_wb, "Top Products")

    if a_rows is None:
        check("Sheet 'Top Products' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth Top Products sheet exists", False, "Not found")
    else:
        check("Sheet 'Top Products' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("Top Products row count", len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        # Check that top product per category matches
        g_by_cat = {}
        for row in g_data:
            if row and row[0]:
                cat = str(row[0]).strip().lower()
                if cat not in g_by_cat:
                    g_by_cat[cat] = row

        a_by_cat = {}
        for row in a_data:
            if row and row[0]:
                cat = str(row[0]).strip().lower()
                if cat not in a_by_cat:
                    a_by_cat[cat] = row

        for cat, g_row in g_by_cat.items():
            a_row = a_by_cat.get(cat)
            if a_row is None:
                check(f"Top product for '{cat}'", False, "Category missing")
            else:
                check(f"Top product for '{cat}' matches",
                      str(a_row[1] or "").strip()[:30].lower() == str(g_row[1] or "").strip()[:30].lower(),
                      f"Got '{str(a_row[1])[:50]}' vs '{str(g_row[1])[:50]}'")


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for email check", False, str(e))
        return

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    # Find email to sales-team@shop.com
    target = "sales-team@shop.com"
    found = None
    for subj, from_addr, to_addr, body in all_emails:
        to_str = str(to_addr or "").lower()
        if target in to_str:
            found = (subj, from_addr, to_addr, body)
            break

    if found:
        subj, from_addr, to_addr, body = found
        body_lower = (body or "").lower()
        subj_lower = (subj or "").lower()

        check("CRITICAL: Email sent to sales-team@shop.com with subject 'Category Performance Report'",
              "category performance report" in subj_lower,
              f"Subject: {(subj or '')[:100]}")

        check("CRITICAL: Email body names highest-revenue category (ТВ и домашний кинотеатр)",
              ("кинотеатр" in body_lower)
              or ("home theater" in body_lower)
              or ("тв и домашний" in body_lower),
              f"highest category not found in body: {body_lower[:200]}")

        check("CRITICAL: Email body names lowest-revenue category (Бытовая техника)",
              "бытовая техника" in body_lower or "home appliances" in body_lower,
              f"lowest category not found in body: {body_lower[:200]}")
    else:
        check("CRITICAL: Email sent to sales-team@shop.com with subject 'Category Performance Report'",
              False, f"No email to {target}; found {len(all_emails)} total emails")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_email()

    total_pass = PASS_COUNT
    total_fail = FAIL_COUNT
    total = total_pass + total_fail
    accuracy = (total_pass / total * 100.0) if total else 0.0
    critical_ok = len(CRITICAL_FAILED) == 0
    all_ok = critical_ok and accuracy >= 70.0

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILED:
        print(f"  CRITICAL FAILURES ({len(CRITICAL_FAILED)}):")
        for c in CRITICAL_FAILED:
            print(f"    - {c}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "passed": total_pass,
                "failed": total_fail,
                "accuracy": accuracy,
                "critical_failed": CRITICAL_FAILED,
                "success": all_ok,
            }, f, indent=2)

    if not critical_ok:
        print("  -> Critical check failed: forcing FAIL.")
        sys.exit(1)

    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
