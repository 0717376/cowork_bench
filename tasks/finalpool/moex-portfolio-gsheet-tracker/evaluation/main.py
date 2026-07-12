"""
Evaluation for moex-portfolio-gsheet-tracker task.
Checks:
1. Local file: portfolio_summary.xlsx exists with correct Holdings and Summary sheets
2. Google Sheet: gsheet.cells in DB have been updated with correct values
All numeric comparisons use a 2% relative tolerance.

A set of CRITICAL semantic checks must all pass; any critical failure causes an
immediate FAIL (sys.exit(1)) before the accuracy gate is evaluated.
"""
import os
import sys
import json
from argparse import ArgumentParser
from datetime import datetime

import openpyxl
import psycopg2


DB_CONN = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

SPREADSHEET_ID = "sp_portfolio_tracker_q1_2026"
SHEET_ID = 1
TOLERANCE = 0.02  # 2% relative tolerance

# ---------- Latest prices from moex.stock_prices ----------
def get_latest_prices():
    """Query actual latest closing prices from MOEX Finance DB."""
    conn = psycopg2.connect(**DB_CONN)
    cur = conn.cursor()
    cur.execute("""
        SELECT symbol, close FROM moex.stock_prices
        WHERE symbol IN ('SBER.ME','GAZP.ME','LKOH.ME','MGNT.ME','MTSS.ME')
          AND date = (SELECT MAX(date) FROM moex.stock_prices WHERE symbol = 'SBER.ME')
        ORDER BY symbol;
    """)
    prices = {row[0]: float(row[1]) for row in cur.fetchall()}
    cur.close()
    conn.close()
    return prices


# ---------- Compute expected values ----------
# Portfolio holdings match initial_workspace/portfolio_holdings.xlsx and the
# PREFILLED_DATA list in preprocess/main.py. If the input file changes, these
# must be updated together. Purchase prices are in RUB.
HOLDINGS = [
    ("SBER.ME", 2000, 110.50),
    ("GAZP.ME", 1500, 225.00),
    ("LKOH.ME", 150, 3400.00),
    ("MGNT.ME", 120, 3950.00),
    ("MTSS.ME", 900, 240.00),
]
# Risk ratings come from the mock web portal (files/mock_pages/index.html).
# Cannot be queried from PostgreSQL at evaluation time. Labels MUST stay in sync
# (case-insensitive string match) with the portal HTML.
RISK_RATINGS = {
    "SBER.ME": "Средний",
    "GAZP.ME": "Средне-высокий",
    "LKOH.ME": "Средне-высокий",
    "MGNT.ME": "Средний",
    "MTSS.ME": "Низкий",
}


def compute_expected(prices):
    """Compute all expected values from latest prices."""
    rows = []
    for sym, shares, pp in HOLDINGS:
        cp = prices[sym]
        mv = round(shares * cp, 2)
        gl = round((cp - pp) * shares, 2)
        gl_pct = round(((cp - pp) / pp) * 100, 2)
        rows.append({
            "Symbol": sym, "Shares": shares, "Purchase_Price": pp,
            "Current_Price": cp, "Market_Value": mv,
            "Gain_Loss": gl, "Gain_Loss_Pct": gl_pct,
            "Risk_Rating": RISK_RATINGS[sym],
        })

    total_mv = sum(r["Market_Value"] for r in rows)
    for r in rows:
        r["Allocation_Pct"] = round((r["Market_Value"] / total_mv) * 100, 2)
        r["Compliance_Status"] = "OK" if r["Allocation_Pct"] <= 30 else "OVER_LIMIT"

    total_cost = round(sum(s * pp for _, s, pp in HOLDINGS), 2)
    total_gl = round(total_mv - total_cost, 2)
    total_gl_pct = round((total_gl / total_cost) * 100, 2)
    gl_pcts = {r["Symbol"]: r["Gain_Loss_Pct"] for r in rows}
    highest = max(gl_pcts, key=gl_pcts.get)
    lowest = min(gl_pcts, key=gl_pcts.get)
    compliance_issues = sum(1 for r in rows if r["Compliance_Status"] == "OVER_LIMIT")

    summary = {
        "Total_Market_Value": round(total_mv, 2),
        "Total_Cost_Basis": total_cost,
        "Total_Gain_Loss": total_gl,
        "Total_Gain_Loss_Pct": total_gl_pct,
        "Highest_Gainer": highest,
        "Lowest_Gainer": lowest,
        "Compliance_Issues": compliance_issues,
    }
    return rows, summary


# ---------- Comparison helpers ----------
def nums_close(expected, actual, tol=TOLERANCE):
    """Check if two numeric values are within relative tolerance."""
    try:
        e = float(expected)
        a = float(actual)
    except (ValueError, TypeError):
        return False
    if abs(e) < 1e-9:
        return abs(a) < 0.01
    return abs(e - a) / abs(e) <= tol


def val_match(expected, actual, tol=TOLERANCE):
    """Check if expected and actual values match (numeric or string)."""
    if expected is None and actual is None:
        return True
    e_str = str(expected).strip()
    a_str = str(actual).strip()
    if e_str.lower() == a_str.lower():
        return True
    return nums_close(expected, actual, tol)


# ---------- Check 1: Local Excel file ----------
def check_local_excel(workspace, expected_rows, expected_summary):
    path = os.path.join(workspace, "portfolio_summary.xlsx")
    if not os.path.exists(path):
        print(f"FAIL: portfolio_summary.xlsx not found at {path}")
        return 0, 0

    wb = openpyxl.load_workbook(path)
    total_checks = 0
    passed_checks = 0

    # Check Holdings sheet exists
    holdings_name = None
    for name in wb.sheetnames:
        if name.lower() == "holdings":
            holdings_name = name
            break
    if not holdings_name:
        print("FAIL: 'Holdings' sheet not found in portfolio_summary.xlsx")
        return 0, 1

    ws_h = wb[holdings_name]
    rows_data = list(ws_h.iter_rows(min_row=2, values_only=True))

    # Check Holdings data (5 rows x 10 columns)
    COLS = ["Symbol", "Shares", "Purchase_Price", "Current_Price", "Market_Value",
            "Gain_Loss", "Gain_Loss_Pct", "Allocation_Pct", "Risk_Rating", "Compliance_Status"]

    for i, exp_row in enumerate(expected_rows):
        if i >= len(rows_data):
            print(f"FAIL: Missing row {i+1} in Holdings sheet")
            total_checks += len(COLS)
            continue
        actual_row = rows_data[i]
        for j, col in enumerate(COLS):
            total_checks += 1
            exp_val = exp_row[col]
            act_val = actual_row[j] if j < len(actual_row) else None
            if val_match(exp_val, act_val):
                passed_checks += 1
            else:
                print(f"  Holdings mismatch row {i+1} col '{col}': expected={exp_val}, actual={act_val}")

    # Check Summary sheet exists
    summary_name = None
    for name in wb.sheetnames:
        if name.lower() == "summary":
            summary_name = name
            break
    if not summary_name:
        print("FAIL: 'Summary' sheet not found in portfolio_summary.xlsx")
        total_checks += len(expected_summary)
        return passed_checks, total_checks

    ws_s = wb[summary_name]
    summary_data = {}
    for row in ws_s.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            summary_data[str(row[0]).strip()] = row[1]

    for metric, exp_val in expected_summary.items():
        total_checks += 1
        act_val = summary_data.get(metric)
        if val_match(exp_val, act_val):
            passed_checks += 1
        else:
            print(f"  Summary mismatch '{metric}': expected={exp_val}, actual={act_val}")

    wb.close()
    return passed_checks, total_checks


# ---------- Check 2: Google Sheet in DB ----------
def check_gsheet(expected_rows):
    conn = psycopg2.connect(**DB_CONN)
    cur = conn.cursor()

    # Fetch all cells for this spreadsheet/sheet
    cur.execute("""
        SELECT row_index, col_index, value FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s
        ORDER BY row_index, col_index;
    """, (SPREADSHEET_ID, SHEET_ID))

    cells = {}
    for row_idx, col_idx, value in cur.fetchall():
        cells[(row_idx, col_idx)] = value

    cur.close()
    conn.close()

    if not cells:
        print("FAIL: No cells found in gsheet for the spreadsheet")
        return 0, 1

    # Columns that the agent should have filled: col indices 3-9
    FILL_COLS = {
        3: "Current_Price",
        4: "Market_Value",
        5: "Gain_Loss",
        6: "Gain_Loss_Pct",
        7: "Allocation_Pct",
        8: "Risk_Rating",
        9: "Compliance_Status",
    }

    total_checks = 0
    passed_checks = 0

    for i, exp_row in enumerate(expected_rows):
        row_idx = i + 1  # row 0 is header
        for col_idx, col_name in FILL_COLS.items():
            total_checks += 1
            exp_val = exp_row[col_name]
            act_val = cells.get((row_idx, col_idx))
            if act_val is None:
                print(f"  GSheet missing cell ({row_idx},{col_idx}) for {exp_row['Symbol']}.{col_name}")
                continue
            if val_match(exp_val, act_val):
                passed_checks += 1
            else:
                print(f"  GSheet mismatch ({row_idx},{col_idx}) {exp_row['Symbol']}.{col_name}: expected={exp_val}, actual={act_val}")

    return passed_checks, total_checks


# ---------- CRITICAL semantic checks ----------
def get_gsheet_cells():
    conn = psycopg2.connect(**DB_CONN)
    cur = conn.cursor()
    cur.execute("""
        SELECT row_index, col_index, value FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s;
    """, (SPREADSHEET_ID, SHEET_ID))
    cells = {(r, c): v for r, c, v in cur.fetchall()}
    cur.close()
    conn.close()
    return cells


def load_summary_sheet(workspace):
    path = os.path.join(workspace, "portfolio_summary.xlsx")
    if not os.path.exists(path):
        return None, None
    wb = openpyxl.load_workbook(path)
    h_name = next((n for n in wb.sheetnames if n.lower() == "holdings"), None)
    s_name = next((n for n in wb.sheetnames if n.lower() == "summary"), None)
    holdings = list(wb[h_name].iter_rows(min_row=2, values_only=True)) if h_name else None
    summary = {}
    if s_name:
        for row in wb[s_name].iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip()] = row[1]
    wb.close()
    return holdings, summary


def run_critical_checks(workspace, expected_rows, expected_summary, prices):
    """Semantic gate. Returns list of (name, ok, detail). Any False => FAIL."""
    results = []
    xlsx_holdings, xlsx_summary = load_summary_sheet(workspace)
    gcells = get_gsheet_cells()

    # col index map (gsheet) and column position map (xlsx Holdings)
    COLS = ["Symbol", "Shares", "Purchase_Price", "Current_Price", "Market_Value",
            "Gain_Loss", "Gain_Loss_Pct", "Allocation_Pct", "Risk_Rating", "Compliance_Status"]
    xlsx_by_sym = {}
    if xlsx_holdings:
        for r in xlsx_holdings:
            if r and r[0]:
                xlsx_by_sym[str(r[0]).strip()] = r

    # CRITICAL 1: Current_Price for every holding matches latest moex close (2% tol)
    # in BOTH the Google Sheet and the xlsx Holdings sheet.
    ok = True
    detail = []
    for i, exp in enumerate(expected_rows):
        sym = exp["Symbol"]
        exp_cp = prices[sym]
        gs_cp = gcells.get((i + 1, 3))
        xrow = xlsx_by_sym.get(sym)
        xl_cp = xrow[3] if xrow and len(xrow) > 3 else None
        if not nums_close(exp_cp, gs_cp):
            ok = False; detail.append(f"{sym} gsheet Current_Price={gs_cp} != {exp_cp}")
        if not nums_close(exp_cp, xl_cp):
            ok = False; detail.append(f"{sym} xlsx Current_Price={xl_cp} != {exp_cp}")
    results.append(("Current_Price matches latest MOEX close (gsheet + xlsx)", ok, "; ".join(detail)))

    # CRITICAL 2: Risk_Rating for every symbol exactly matches the portal value.
    ok = True
    detail = []
    for i, exp in enumerate(expected_rows):
        sym = exp["Symbol"]
        exp_rr = RISK_RATINGS[sym]
        gs_rr = gcells.get((i + 1, 8))
        xrow = xlsx_by_sym.get(sym)
        xl_rr = xrow[8] if xrow and len(xrow) > 8 else None
        if not (gs_rr and str(gs_rr).strip().lower() == exp_rr.lower()):
            ok = False; detail.append(f"{sym} gsheet Risk_Rating={gs_rr} != {exp_rr}")
        if not (xl_rr and str(xl_rr).strip().lower() == exp_rr.lower()):
            ok = False; detail.append(f"{sym} xlsx Risk_Rating={xl_rr} != {exp_rr}")
    results.append(("Risk_Rating exactly matches compliance portal", ok, "; ".join(detail)))

    # CRITICAL 3: Compliance_Status correctly derived from 30% single-stock limit.
    ok = True
    detail = []
    for i, exp in enumerate(expected_rows):
        sym = exp["Symbol"]
        exp_cs = exp["Compliance_Status"]
        gs_cs = gcells.get((i + 1, 9))
        if not (gs_cs and str(gs_cs).strip().lower() == exp_cs.lower()):
            ok = False; detail.append(f"{sym} Compliance_Status={gs_cs} != {exp_cs}")
    results.append(("Compliance_Status derived from 30% allocation limit", ok, "; ".join(detail)))

    # CRITICAL 4: Summary totals internally consistent with per-row Holdings (2% tol).
    ok = True
    detail = []
    for metric in ("Total_Market_Value", "Total_Cost_Basis", "Total_Gain_Loss"):
        act = (xlsx_summary or {}).get(metric)
        if not nums_close(expected_summary[metric], act):
            ok = False; detail.append(f"{metric}={act} != {expected_summary[metric]}")
    results.append(("Summary totals consistent with Holdings", ok, "; ".join(detail)))

    # CRITICAL 5: Highest_Gainer / Lowest_Gainer equal argmax/argmin of Gain_Loss_Pct.
    ok = True
    detail = []
    for metric in ("Highest_Gainer", "Lowest_Gainer"):
        act = (xlsx_summary or {}).get(metric)
        exp_v = expected_summary[metric]
        if not (act and str(act).strip().lower() == str(exp_v).lower()):
            ok = False; detail.append(f"{metric}={act} != {exp_v}")
    results.append(("Highest/Lowest_Gainer = argmax/argmin Gain_Loss_Pct", ok, "; ".join(detail)))

    return results


# ---------- Main ----------
def main(args):
    print("Fetching latest stock prices from MOEX Finance DB ...")
    prices = get_latest_prices()
    print(f"  Prices: {prices}")

    expected_rows, expected_summary = compute_expected(prices)
    print(f"  Expected summary: {expected_summary}")

    # ---- CRITICAL semantic gate ----
    print("\n--- CRITICAL checks ---")
    critical_failed = False
    for name, ok, detail in run_critical_checks(
        args.agent_workspace, expected_rows, expected_summary, prices
    ):
        status = "PASS" if ok else "FAIL"
        print(f"  [{status}] {name}" + (f" -- {detail}" if (not ok and detail) else ""))
        if not ok:
            critical_failed = True
    if critical_failed:
        print("\nFAIL: one or more CRITICAL checks failed.")
        sys.exit(1)

    total_passed = 0
    total_checks = 0

    # Check 1: Local Excel
    print("\n--- Check 1: Local portfolio_summary.xlsx ---")
    p, t = check_local_excel(args.agent_workspace, expected_rows, expected_summary)
    print(f"  Local Excel: {p}/{t} checks passed")
    total_passed += p
    total_checks += t

    # Check 2: Google Sheet
    print("\n--- Check 2: Google Sheet in DB ---")
    p, t = check_gsheet(expected_rows)
    print(f"  Google Sheet: {p}/{t} checks passed")
    total_passed += p
    total_checks += t

    # Overall
    if total_checks == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = total_passed / total_checks * 100
    print(f"\nOverall: {total_passed}/{total_checks} checks passed ({accuracy:.1f}%)")

    if args.output_file:
        report = {
            "total_passed": total_passed,
            "total_checks": total_checks,
            "accuracy": accuracy,
            "timestamp": datetime.now().isoformat(),
        }
        with open(args.output_file, "w") as f:
            json.dump(report, f, indent=2)
        print(f"Report saved to {args.output_file}")

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--output_file", required=False)
    args = parser.parse_args()
    main(args)
