"""Evaluation for insales-product-review-excel-notion (InSales + Teamly).

Category names are russified centrally (wc.* seed). Eval therefore matches
categories by their STABLE NUMERIC SIGNATURE (review_count / avg_rating /
product_count) and tolerates both Russian and English category names — it must
NOT rely on English-only name literals.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Expected category data (language-independent numeric signatures from wc.*).
# en/ru names kept only as a tolerance aid; matching is by review_count.
EXPECTED_CATEGORIES = {
    "audio": {"product_count": 15, "review_count": 78, "avg_rating": 4.47, "ru": "аудио"},
    "cameras": {"product_count": 10, "review_count": 49, "avg_rating": 4.69, "ru": "камеры"},
    "electronics": {"product_count": 30, "review_count": 149, "avg_rating": 4.57, "ru": "электроника"},
    "headphones": {"product_count": 10, "review_count": 50, "avg_rating": 4.52, "ru": "наушники"},
    "home appliances": {"product_count": 8, "review_count": 29, "avg_rating": 4.72, "ru": "бытовая техника"},
    "speakers": {"product_count": 5, "review_count": 28, "avg_rating": 4.39, "ru": "колонки"},
    "tv & home theater": {"product_count": 13, "review_count": 63, "avg_rating": 4.78, "ru": "домашний кинотеатр"},
    "watches": {"product_count": 6, "review_count": 28, "avg_rating": 4.43, "ru": "часы"},
}

# Numeric signatures used for name-agnostic matching of key categories.
ELECTRONICS_RC = 149          # Electronics / Электроника
ELECTRONICS_AVG = 4.57
TV_RC = 63                    # TV & Home Theater / ТВ и домашний кинотеатр (highest avg)
TV_AVG = 4.78
SPEAKERS_AVG = 4.39          # lowest avg category (Speakers / Колонки)

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Category Analysis has 8 data rows (8 categories)",
    "Electronics category (review_count=149) has avg_rating ~4.57",
    "Highest-rated category (review_count=63) has avg_rating ~4.78",
    "Top Products top row avg_rating = 5.0 and all rows have >=3 reviews",
    "Top Products sorted by Avg_Rating desc then Review_Count desc",
    "Teamly 'Product Review Insights' page exists with category numeric signatures",
    "Email subject 'Product Review Analysis Report' and body names highest+lowest category",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.1):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", ".").replace("%", "").replace("$", "").strip())
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    target = sheet_name.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Product_Review_Analysis.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Product_Review_Analysis.xlsx exists", False, f"Not found: {xlsx_path}")
        return
    check("Product_Review_Analysis.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # ---- Category Analysis sheet ----
    cat_rows = load_sheet_rows(wb, "Category Analysis")
    if cat_rows is None:
        check("Sheet 'Category Analysis' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Category Analysis' exists", True)
        data_rows = [r for r in cat_rows[1:] if r and any(c is not None for c in r)]
        check("Category Analysis has 8 data rows (8 categories)",
              len(data_rows) == 8,
              f"Found {len(data_rows)} rows")

        # Header columns (English identifiers pinned in task.md).
        header = cat_rows[0] if cat_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col_name in ["category", "product_count", "review_count", "avg_rating",
                         "five_star_count", "five_star_rate"]:
            check(f"Column '{col_name}' present",
                  any(col_name.replace("_", "") in h.replace("_", "") or col_name in h
                      for h in header_lower),
                  f"Header: {header}")

        # Name-agnostic match: the row whose Review_Count=149 is Electronics.
        electronics_row = None
        tv_row = None
        for row in data_rows:
            if len(row) < 4:
                continue
            rc = safe_float(row[2])
            if rc is not None and num_close(rc, ELECTRONICS_RC, 0.5):
                electronics_row = row
            if rc is not None and num_close(rc, TV_RC, 0.5):
                tv_row = row

        check("Electronics category (review_count=149) has avg_rating ~4.57",
              electronics_row is not None and num_close(electronics_row[3], ELECTRONICS_AVG, 0.05),
              f"Row: {electronics_row}")

        check("Highest-rated category (review_count=63) has avg_rating ~4.78",
              tv_row is not None and num_close(tv_row[3], TV_AVG, 0.05),
              f"Row: {tv_row}")

        # The highest Avg_Rating across all rows should be ~4.78 (sanity on aggregation).
        avgs = [safe_float(r[3]) for r in data_rows if len(r) > 3 and safe_float(r[3]) is not None]
        if avgs:
            check("Max category avg_rating is ~4.78",
                  num_close(max(avgs), TV_AVG, 0.05),
                  f"Max found {max(avgs)}")
            check("Min category avg_rating is ~4.39",
                  num_close(min(avgs), SPEAKERS_AVG, 0.05),
                  f"Min found {min(avgs)}")

        # Sorted alphabetically by Category (RU or EN ordering both acceptable: just
        # check non-empty names are present; ordering is volatile across collations).
        names = [str(r[0]).strip() for r in data_rows if r and r[0]]
        check("Category column populated for all 8 rows", len(names) == 8, f"names: {names}")

    # ---- Top Products sheet ----
    top_rows = load_sheet_rows(wb, "Top Products")
    if top_rows is None:
        check("Sheet 'Top Products' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Top Products' exists", True)
        header = top_rows[0] if top_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        # Find Review_Count and Avg_Rating column indices.
        rc_i = next((i for i, h in enumerate(header_lower) if "review_count" in h or "reviewcount" in h), 2)
        ar_i = next((i for i, h in enumerate(header_lower) if "avg_rating" in h or "avgrating" in h), 3)

        data_rows = [r for r in top_rows[1:] if r and any(c is not None for c in r)]
        check("Top Products has 1..10 rows",
              1 <= len(data_rows) <= 10,
              f"Found {len(data_rows)} rows")

        ratings = [safe_float(r[ar_i]) if len(r) > ar_i else None for r in data_rows]
        counts = [safe_float(r[rc_i]) if len(r) > rc_i else None for r in data_rows]

        top_ok = bool(data_rows) and num_close(ratings[0], 5.0, 0.05)
        min3_ok = all(c is not None and c >= 3 for c in counts) if counts else False
        check("Top Products top row avg_rating = 5.0 and all rows have >=3 reviews",
              top_ok and min3_ok,
              f"top_avg={ratings[0] if ratings else None}, counts={counts}")

        # Sort: Avg_Rating desc, then Review_Count desc.
        sorted_ok = True
        for i in range(len(data_rows) - 1):
            a_r, a_c = ratings[i], counts[i]
            b_r, b_c = ratings[i + 1], counts[i + 1]
            if a_r is None or b_r is None:
                sorted_ok = False
                break
            if a_r < b_r - 1e-6:
                sorted_ok = False
                break
            if num_close(a_r, b_r, 1e-6) and a_c is not None and b_c is not None and a_c < b_c - 1e-6:
                sorted_ok = False
                break
        check("Top Products sorted by Avg_Rating desc then Review_Count desc",
              sorted_ok,
              f"ratings={ratings}, counts={counts}")

    return wb


def _extract_numbers(text):
    """Return the set of integers and 2-decimal floats appearing in text."""
    import re
    ints = set(int(x) for x in re.findall(r"(?<![\d.])\d{1,4}(?![\d.])", text))
    floats = set(round(float(x), 2) for x in re.findall(r"\d+\.\d+", text))
    return ints, floats


def check_teamly():
    print("\n=== Checking Teamly 'Product Review Insights' ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        # User-created pages have id > 3 (seed pages are id <= 3).
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Product Review Insights' page exists with category numeric signatures",
              False, str(e))
        return

    target = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "product review" in tl or "review insight" in tl \
                or ("отзыв" in tl and ("товар" in tl or "категор" in tl or "продукт" in tl)) \
                or ("product review insights" in tl):
            target = (pid, title, body)
            break

    if target is None:
        check("Teamly 'Product Review Insights' page exists with category numeric signatures",
              False,
              f"new pages: {[(p[0], p[1]) for p in pages]}")
        return

    body = (target[2] or "")
    ints, floats = _extract_numbers(body)

    # Verify the table carries the category numeric signatures: how many of the
    # 8 categories' review_count AND avg_rating appear in the body.
    matched = 0
    for cat, sig in EXPECTED_CATEGORIES.items():
        rc_ok = sig["review_count"] in ints
        avg_ok = any(num_close(f, sig["avg_rating"], 0.05) for f in floats)
        if rc_ok and avg_ok:
            matched += 1
    check("Teamly 'Product Review Insights' page exists with category numeric signatures",
          matched >= 6,
          f"categories with matching review_count+avg in body: {matched}/8")

    # Non-critical: title pinned in English.
    check("Teamly page title contains 'Product Review Insights'",
          "product review insights" in (target[1] or "").lower(),
          f"title: {target[1]}")


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            """
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%product_team@store.com%'
               OR subject ILIKE '%Product Review Analysis Report%'
               OR subject ILIKE '%product%analysis%'
            """
        )
        emails = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Email subject 'Product Review Analysis Report' and body names highest+lowest category",
              False, str(e))
        return

    check("Email sent to product_team@store.com", len(emails) >= 1, "No matching email found")
    if not emails:
        check("Email subject 'Product Review Analysis Report' and body names highest+lowest category",
              False, "no email")
        return

    # Prefer the email actually addressed to product_team@store.com.
    email = next((e for e in emails if e[2] and "product_team@store.com" in str(e[2]).lower()), emails[0])
    subject = str(email[1]) if email[1] else ""
    body = str(email[3]) if email[3] else ""
    body_l = body.lower()

    subject_ok = "product review analysis report" in subject.lower()
    # Highest-rated category: TV & Home Theater / ТВ и домашний кинотеатр (4.78).
    # Lowest-rated category: Speakers / Колонки (4.39).
    highest_named = (
        "домашний кинотеатр" in body_l or "tv & home theater" in body_l
        or "tv and home theater" in body_l or "тв и домашний" in body_l
        or "4.78" in body_l
    )
    lowest_named = (
        "колонки" in body_l or "speakers" in body_l or "4.39" in body_l
    )
    check("Email subject 'Product Review Analysis Report' and body names highest+lowest category",
          subject_ok and highest_named and lowest_named,
          f"subject_ok={subject_ok}, highest={highest_named}, lowest={lowest_named}")

    # Non-critical sanity.
    check("Email body has a content summary", len(body) > 40, f"Body length: {len(body)}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    if success:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
