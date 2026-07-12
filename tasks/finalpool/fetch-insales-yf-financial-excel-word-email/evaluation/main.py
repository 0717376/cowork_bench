"""Evaluation script for fetch-insales-yf-financial-excel-word-email (InSales fork).

Structural (soft / NON-critical) checks gate on accuracy >= 70:
  sheet existence, RU+EN column-name presence, row-count thresholds,
  file existence, Word has content, an email with a report/analysis subject.

CRITICAL semantic checks (any fail => sys.exit(1) BEFORE the accuracy gate).
These verify SUBSTANCE from the ACTUAL sources the agent reads, never from the
frozen groundtruth (which mirrors the benchmark into Our_Avg_Price and is not a
reliable value oracle):
  C1  Market_Avg_Price in Data_Analysis matches the fetched data.json benchmark
      per benchmarked category (Электроника=61.13, Аудио=78.05, Камеры=23.91,
      Бытовая техника=35.77, ТВ и домашний кинотеатр=304.9; tolerance) — proves
      the agent fetched+joined the external API.
  C2  Our_Avg_Price per benchmarked category equals the AVG regular_price of that
      category's products in the live InSales (wc.*) store (tolerance) — proves
      the agent read live store data, not fabricated numbers.
  C3  Price_Gap_Pct == round((Our_Avg_Price - Market_Avg_Price)/Market_Avg_Price
      *100, 1) per benchmarked row — core gap computation.
  C4  Data_Analysis Category column is sorted alphabetically (per task.md) with
      >= 3 recognized store-category rows (canonicalized EN-or-RU, consistent with
      C1/C2/C3; the live InSales store is Russified by db/zzz_wc_after_init.sql:
      Аудио/Камеры/Электроника...).
  C5  Email with subject exactly 'Analysis Report Complete' to team-lead@company.com
      with a non-trivial body summarizing findings.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PASS_COUNT = 0
FAIL_COUNT = 0

# --- Category language bridge -------------------------------------------------
# data.json benchmark keys are ENGLISH; the InSales store category names are
# Russified centrally by db/zzz_wc_after_init.sql. Map EN benchmark keys -> the
# Russified store names, and accept either spelling for any sheet cell.
EN_TO_RU_CATEGORY = {
    "Audio": "Аудио", "Cameras": "Камеры", "Electronics": "Электроника",
    "Headphones": "Наушники", "Home Appliances": "Бытовая техника",
    "Speakers": "Колонки", "TV & Home Theater": "ТВ и домашний кинотеатр",
    "Watches": "Часы",
}
RU_TO_EN_CATEGORY = {ru: en for en, ru in EN_TO_RU_CATEGORY.items()}

# RU+EN header tokens accepted as a match for each expected (English) column.
RU_HEADER_TOKENS = {
    "Category": ["category", "категория", "категории"],
    "Product_Count": ["product_count", "product count", "productcount",
                      "количество товаров", "кол-во товаров", "число товаров", "количество"],
    "Our_Avg_Price": ["our_avg_price", "our avg price", "ouravgprice",
                      "наша средняя цена", "наша ср. цена", "средняя цена магазина"],
    "Total_Sales": ["total_sales", "total sales", "totalsales",
                    "суммарные продажи", "общие продажи", "продажи"],
    "Market_Avg_Price": ["market_avg_price", "market avg price", "marketavgprice",
                         "рыночная средняя цена", "средняя рыночная цена", "бенчмарк"],
    "Price_Gap_Pct": ["price_gap_pct", "price gap pct", "pricegappct",
                      "разрыв цен", "отклонение цены", "разрыв, %", "ценовой разрыв"],
    "Metric": ["metric", "метрика", "показатель"],
    "Value": ["value", "значение"],
    "Priority": ["priority", "приоритет"],
    "Action": ["action", "действие", "мера"],
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '.').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


# --- helpers -----------------------------------------------------------------
def _norm_header(h):
    return str(h).strip().lower() if h is not None else ""


def build_col_map(headers, expected_cols):
    """Map each expected English column to a header index, RU+EN tolerant.
    Returns {expected_col: index}; missing columns are simply absent.
    """
    col_map = {}
    norm_headers = [_norm_header(h) for h in headers]
    for ec in expected_cols:
        # 1) exact normalized match against EN variant (ignore _/space)
        target = ec.lower().replace("_", "").replace(" ", "")
        for idx, h in enumerate(headers):
            if h and target == str(h).strip().lower().replace("_", "").replace(" ", ""):
                col_map[ec] = idx
                break
        # 2) substring match against EN variant
        if ec not in col_map:
            for idx, h in enumerate(headers):
                if h and ec.lower().replace("_", " ") in str(h).strip().lower().replace("_", " "):
                    col_map[ec] = idx
                    break
        # 3) RU+EN token match
        if ec not in col_map:
            for token in RU_HEADER_TOKENS.get(ec, []):
                for idx, nh in enumerate(norm_headers):
                    if nh and (nh == token or token in nh):
                        col_map[ec] = idx
                        break
                if ec in col_map:
                    break
    return col_map


def _category_canon(cell):
    """Return the canonical RU category name for a sheet cell (RU or EN), or None."""
    if cell is None:
        return None
    v = str(cell).strip()
    if v in RU_TO_EN_CATEGORY:        # already RU
        return v
    if v in EN_TO_RU_CATEGORY:        # EN -> RU
        return EN_TO_RU_CATEGORY[v]
    return None


def load_benchmark():
    """Read the fetched mock benchmark; return {RU_category: avg_price}."""
    for rel in ("tmp/mock_pages/api/data.json", "files/data.json"):
        p = os.path.join(TASK_ROOT, rel)
        if os.path.exists(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception:
                continue
            out = {}
            for row in data.get("market_data", []):
                en = row.get("category")
                ru = EN_TO_RU_CATEGORY.get(en, en)
                ap = safe_float(row.get("avg_price"))
                if ru and ap is not None:
                    out[ru] = ap
            if out:
                return out
    return {}


def get_store_category_avg():
    """Live InSales store: {RU_category: (product_count, avg_regular_price, total_sales)}.
    products.categories jsonb carries names; map EN->RU so keys match the sheet.
    """
    out = {}
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT jsonb_array_elements(categories)->>'name' AS cat,
                   COUNT(*),
                   ROUND(AVG(regular_price::float)::numeric, 2),
                   COALESCE(SUM(total_sales), 0)
            FROM wc.products
            GROUP BY cat
        """)
        for cat, cnt, avg, sales in cur.fetchall():
            ru = EN_TO_RU_CATEGORY.get(cat, cat)
            out[ru] = (int(cnt), float(avg), int(sales))
        conn.close()
    except Exception as e:
        print(f"  [warn] store query failed: {e}")
    return out


def read_data_analysis(agent_workspace):
    """Return (headers, col_map, rows-as-dicts keyed by expected col, raw cat list)."""
    path = os.path.join(agent_workspace, "Yf_Financial_Report.xlsx")
    if not os.path.exists(path):
        return None
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return None
    if "Data_Analysis" not in wb.sheetnames:
        return None
    ws = wb["Data_Analysis"]
    headers = [c.value for c in ws[1]]
    cols = ["Category", "Product_Count", "Our_Avg_Price", "Total_Sales",
            "Market_Avg_Price", "Price_Gap_Pct"]
    cmap = build_col_map(headers, cols)
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if "Category" in cmap and r[cmap["Category"]] is not None:
            rows.append(r)
    return headers, cmap, rows


# --- CRITICAL checks ----------------------------------------------------------
def critical_market_benchmark(agent_workspace):
    """C1: Market_Avg_Price per benchmarked category == data.json benchmark."""
    errors = []
    parsed = read_data_analysis(agent_workspace)
    if not parsed:
        return False, ["C1: Yf_Financial_Report.xlsx / Data_Analysis missing"]
    _, cmap, rows = parsed
    if "Market_Avg_Price" not in cmap or "Category" not in cmap:
        return False, ["C1: Category/Market_Avg_Price column not found"]
    benchmark = load_benchmark()
    if not benchmark:
        return False, ["C1: benchmark data.json not loadable"]
    by_cat = {}
    for r in rows:
        canon = _category_canon(r[cmap["Category"]])
        if canon:
            by_cat[canon] = r
    for ru_cat, bench in benchmark.items():
        if ru_cat not in by_cat:
            errors.append(f"C1: benchmarked category '{ru_cat}' missing from Data_Analysis")
            continue
        v = safe_float(by_cat[ru_cat][cmap["Market_Avg_Price"]])
        if v is None or abs(v - bench) > 0.5:
            errors.append(f"C1: '{ru_cat}' Market_Avg_Price {v} != benchmark {bench}")
    return (not errors), errors


def critical_store_avg(agent_workspace):
    """C2: Our_Avg_Price per benchmarked category == live store AVG(regular_price)."""
    errors = []
    parsed = read_data_analysis(agent_workspace)
    if not parsed:
        return False, ["C2: Data_Analysis missing"]
    _, cmap, rows = parsed
    if "Our_Avg_Price" not in cmap or "Category" not in cmap:
        return False, ["C2: Category/Our_Avg_Price column not found"]
    store = get_store_category_avg()
    if not store:
        return False, ["C2: live store data unavailable"]
    benchmark = load_benchmark()
    by_cat = {}
    for r in rows:
        canon = _category_canon(r[cmap["Category"]])
        if canon:
            by_cat[canon] = r
    # Verify against the benchmarked categories (the ones the task centers on).
    for ru_cat in benchmark:
        if ru_cat not in store:
            continue
        exp_avg = store[ru_cat][1]
        if ru_cat not in by_cat:
            errors.append(f"C2: category '{ru_cat}' missing from Data_Analysis")
            continue
        v = safe_float(by_cat[ru_cat][cmap["Our_Avg_Price"]])
        if v is None or abs(v - exp_avg) > 1.0:
            errors.append(f"C2: '{ru_cat}' Our_Avg_Price {v} != store avg {exp_avg}")
    return (not errors), errors


def critical_gap_math(agent_workspace):
    """C3: Price_Gap_Pct == round((Our-Market)/Market*100, 1) per benchmarked row."""
    errors = []
    parsed = read_data_analysis(agent_workspace)
    if not parsed:
        return False, ["C3: Data_Analysis missing"]
    _, cmap, rows = parsed
    for need in ("Our_Avg_Price", "Market_Avg_Price", "Price_Gap_Pct", "Category"):
        if need not in cmap:
            return False, [f"C3: column {need} not found"]
    benchmark = load_benchmark()
    checked = 0
    for r in rows:
        canon = _category_canon(r[cmap["Category"]])
        if canon not in benchmark:
            continue
        our = safe_float(r[cmap["Our_Avg_Price"]])
        mkt = safe_float(r[cmap["Market_Avg_Price"]])
        gap = safe_float(r[cmap["Price_Gap_Pct"]])
        if our is None or mkt is None or gap is None or mkt == 0:
            errors.append(f"C3: '{canon}' missing numeric values (our={our}, mkt={mkt}, gap={gap})")
            continue
        expected = round((our - mkt) / mkt * 100, 1)
        if abs(gap - expected) > 0.3:
            errors.append(f"C3: '{canon}' Price_Gap_Pct {gap} != expected {expected}")
        checked += 1
    if checked < 3:
        errors.append(f"C3: only {checked} benchmarked rows verified (need >= 3)")
    return (not errors), errors


def critical_sorted_categories(agent_workspace):
    """C4: Category column sorted alphabetically with >= 3 recognized store categories.

    Category names are canonicalized via _category_canon (accepts the now-Russian
    InSales spelling OR the English benchmark spelling), consistent with C1/C2/C3.
    A correct agent reads the live store, whose categories are Russified by
    db/zzz_wc_after_init.sql, so it writes >= 3 known categories; a non-doing agent
    still fails this gate.
    """
    errors = []
    parsed = read_data_analysis(agent_workspace)
    if not parsed:
        return False, ["C4: Data_Analysis missing"]
    _, cmap, rows = parsed
    if "Category" not in cmap:
        return False, ["C4: Category column not found"]
    cats = [str(r[cmap["Category"]]).strip() for r in rows if r[cmap["Category"]] is not None]
    if len(cats) < 3:
        errors.append(f"C4: only {len(cats)} category rows (need >= 3)")
    if cats != sorted(cats):
        errors.append(f"C4: Category column not sorted alphabetically: {cats}")
    known_rows = [c for c in cats if _category_canon(c) is not None]
    if len(known_rows) < 3:
        errors.append(f"C4: only {len(known_rows)} recognized store-category rows (need >= 3): {cats}")
    return (not errors), errors


def critical_email():
    """C5: email subject exactly 'Analysis Report Complete' to team-lead@company.com
    with a non-trivial body."""
    errors = []
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text
            FROM email.messages
            WHERE subject = %s
        """, ('Analysis Report Complete',))
        emails = cur.fetchall()
        conn.close()
    except Exception as e:
        return False, [f"C5: email query failed: {e}"]
    if not emails:
        return False, ["C5: no email with subject exactly 'Analysis Report Complete'"]
    matched = None
    for subject, to_addr, body in emails:
        to_str = str(to_addr).lower() if to_addr else ""
        if "team-lead@company.com" in to_str:
            matched = (subject, body)
            break
    if matched is None:
        return False, ["C5: 'Analysis Report Complete' email not addressed to team-lead@company.com"]
    body = (matched[1] or "").strip()
    if len(body) < 50:
        errors.append(f"C5: email body too short ({len(body)} chars)")
    return (not errors), errors


# --- structural (soft) checks -------------------------------------------------
def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    excel_path = os.path.join(agent_workspace, "Yf_Financial_Report.xlsx")
    check("Yf_Financial_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")
            headers = [c.value for c in ws[1]]
            cmap = build_col_map(headers, ['Category', 'Product_Count', 'Our_Avg_Price',
                                           'Total_Sales', 'Market_Avg_Price', 'Price_Gap_Pct'])
            for ec in ['Category', 'Product_Count', 'Our_Avg_Price', 'Total_Sales',
                       'Market_Avg_Price', 'Price_Gap_Pct']:
                check(f"Data_Analysis has {ec} column", ec in cmap, f"headers: {headers[:8]}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")
            headers = [c.value for c in ws[1]]
            cmap = build_col_map(headers, ['Metric', 'Value'])
            for ec in ['Metric', 'Value']:
                check(f"Metrics has {ec} column", ec in cmap, f"headers: {headers[:8]}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            headers = [c.value for c in ws[1]]
            cmap = build_col_map(headers, ['Priority', 'Action', 'Category'])
            for ec in ['Priority', 'Action', 'Category']:
                check(f"Recommendations has {ec} column", ec in cmap, f"headers: {headers[:8]}")

        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT subject FROM email.messages WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s",
                        ('%report%', '%analysis%', '%отчёт%', '%анализ%'))
            emails = cur.fetchall()
            check("Analysis email sent", len(emails) >= 1, f"found {len(emails)} matching emails")
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        import glob as globmod
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content", len(text) > 50, f"text length: {len(text)}")

        check("wc_yf_finance_processor.py exists", os.path.exists(os.path.join(agent_workspace, "wc_yf_finance_processor.py")))

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # ---- CRITICAL semantic gate (any fail => exit 1, before accuracy) ----
    critical = [
        ("C1 Market_Avg_Price == data.json benchmark", lambda: critical_market_benchmark(args.agent_workspace)),
        ("C2 Our_Avg_Price == live store avg", lambda: critical_store_avg(args.agent_workspace)),
        ("C3 Price_Gap_Pct math", lambda: critical_gap_math(args.agent_workspace)),
        ("C4 Data_Analysis sorted + Russified categories", lambda: critical_sorted_categories(args.agent_workspace)),
        ("C5 'Analysis Report Complete' email to team-lead", critical_email),
    ]
    critical_failed = False
    for label, fn in critical:
        try:
            ok, errs = fn()
        except Exception as e:
            ok, errs = False, [f"{label}: exception {e}"]
        if ok:
            print(f"[CRITICAL PASS] {label}")
        else:
            critical_failed = True
            print(f"[CRITICAL FAIL] {label}")
            for e in errs:
                print(f"  - {e}")
    if critical_failed:
        print("\nCritical semantic check(s) failed. => FAIL")
        sys.exit(1)

    # ---- Accuracy gate (structural / soft checks), threshold >= 70 ----
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    total = PASS_COUNT + FAIL_COUNT
    accuracy = 100.0 * PASS_COUNT / total if total else 0.0
    print(message)
    print(f"Accuracy: {accuracy:.1f}% ({PASS_COUNT}/{total})")
    if accuracy >= 70:
        print("All critical checks passed and accuracy >= 70. => PASS")
        sys.exit(0)
    else:
        print("Accuracy below threshold. => FAIL")
        sys.exit(1)

if __name__ == "__main__":
    main()
