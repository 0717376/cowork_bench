"""Evaluation for insales-customer-loyalty-tier-excel-email."""
import argparse
import os
import re
import sys
import psycopg2
import openpyxl


DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def get_tier_data_from_db():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT first_name, last_name, email, total_spent, orders_count FROM wc.customers ORDER BY total_spent DESC")
    customers = cur.fetchall()
    cur.close()
    conn.close()

    # (name, email, total_spent, orders_count)
    vip = [(f"{r[0]} {r[1]}", r[2], float(r[3]), r[4]) for r in customers if float(r[3]) >= 2000]
    gold = [(f"{r[0]} {r[1]}", r[2], float(r[3]), r[4]) for r in customers if 500 <= float(r[3]) < 2000]
    std = [(f"{r[0]} {r[1]}", r[2], float(r[3]), r[4]) for r in customers if float(r[3]) < 500]

    vip_rev = round(sum(c[2] for c in vip), 2)
    gold_rev = round(sum(c[2] for c in gold), 2)
    std_rev = round(sum(c[2] for c in std), 2)

    # email -> (total_spent, expected_tier)
    by_email = {}
    for name, email, ts, oc in vip:
        by_email[(email or "").strip().lower()] = (ts, "vip")
    for name, email, ts, oc in gold:
        by_email[(email or "").strip().lower()] = (ts, "gold")
    for name, email, ts, oc in std:
        by_email[(email or "").strip().lower()] = (ts, "standard")

    return {
        "vip": vip,
        "gold": gold,
        "std": std,
        "vip_rev": vip_rev,
        "gold_rev": gold_rev,
        "std_rev": std_rev,
        "vip_avg": round(vip_rev / len(vip), 2) if vip else 0,
        "gold_avg": round(gold_rev / len(gold), 2) if gold else 0,
        "std_avg": round(std_rev / len(std), 2) if std else 0,
        "by_email": by_email,
        "vip_emails": set((c[1] or "").strip().lower() for c in vip),
        "vip_amounts": {(c[1] or "").strip().lower(): c[2] for c in vip},
    }


# ---------------------------------------------------------------------------
# Non-critical structural checks
# ---------------------------------------------------------------------------

def check_gsheet_exists():
    """Spreadsheet exists AND its cells contain the three tier rows + numbers."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%customer loyalty%' OR LOWER(title) LIKE '%loyalty program%'")
        rows = cur.fetchall()
        if not rows:
            cur.close()
            conn.close()
            return False
        ok = False
        for ss_id, _title in rows:
            cur.execute("SELECT value FROM gsheet.cells WHERE spreadsheet_id = %s", (ss_id,))
            vals = " ".join((r[0] or "") for r in cur.fetchall()).lower()
            if "vip" in vals and "gold" in vals and "standard" in vals and re.search(r"\d", vals):
                ok = True
                break
        cur.close()
        conn.close()
        return ok
    except Exception:
        return False


def check_teamly_page_exists():
    """Teamly page (id > 3) with the loyalty title AND tier statistics in body."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception:
        return False
    for _pid, title, body in pages:
        tl = (title or "").lower()
        if "customer loyalty program 2026" in tl or "loyalty program 2026" in tl or "2026" in tl:
            bl = ((title or "") + " " + (body or "")).lower()
            if "vip" in bl and "gold" in bl and "standard" in bl and re.search(r"\d", bl):
                return True
    return False


# ---------------------------------------------------------------------------
# Critical semantic checks
# ---------------------------------------------------------------------------

def crit_tier_summary(agent_wb, tier_data, errs):
    """VIP/Gold/Standard Customer_Count exact + revenue within tol vs live DB."""
    a_rows = load_sheet_rows(agent_wb, "Tier Summary")
    if a_rows is None:
        errs.append("CRITICAL: Tier Summary sheet missing")
        return
    a_data = [r for r in a_rows[1:] if r and r[0] is not None]
    a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
    for tier_name, members, exp_rev in [
        ("vip", tier_data["vip"], tier_data["vip_rev"]),
        ("gold", tier_data["gold"], tier_data["gold_rev"]),
        ("standard", tier_data["std"], tier_data["std_rev"]),
    ]:
        a_row = a_lookup.get(tier_name)
        if a_row is None:
            errs.append(f"CRITICAL: Tier Summary missing row '{tier_name}'")
            continue
        if not (len(a_row) > 1 and num_close(a_row[1], len(members), 0)):
            errs.append(f"CRITICAL: {tier_name}.Customer_Count {a_row[1] if len(a_row)>1 else None} != {len(members)}")
        if not (len(a_row) > 2 and num_close(a_row[2], exp_rev, 2.0)):
            errs.append(f"CRITICAL: {tier_name}.Total_Revenue {a_row[2] if len(a_row)>2 else None} vs {exp_rev} (tol=2.0)")


def crit_customer_tiers(agent_wb, tier_data, errs):
    """Every customer present exactly once, correct tier, sorted by Total_Spent desc."""
    a_rows = load_sheet_rows(agent_wb, "Customer Tiers")
    if a_rows is None:
        errs.append("CRITICAL: Customer Tiers sheet missing")
        return
    data_rows = [r for r in a_rows[1:] if r and any(c is not None for c in r)]
    by_email = tier_data["by_email"]

    seen = {}
    prev_spent = None
    sort_ok = True
    for row in data_rows:
        if len(row) < 5:
            continue
        email = (str(row[1]).strip().lower() if row[1] is not None else "")
        tier = (str(row[4]).strip().lower() if row[4] is not None else "")
        spent = to_float(row[2])
        seen[email] = seen.get(email, 0) + 1
        # tier correctness
        exp = by_email.get(email)
        if exp is not None and tier != exp[1]:
            errs.append(f"CRITICAL: {email} tier '{tier}' != expected '{exp[1]}'")
        # sort order
        if spent is not None and prev_spent is not None and spent > prev_spent + 0.01:
            sort_ok = False
        if spent is not None:
            prev_spent = spent

    # every DB customer present exactly once
    missing = [e for e in by_email if e not in seen]
    dupes = [e for e, n in seen.items() if n > 1 and e in by_email]
    extra = [e for e in seen if e and e not in by_email]
    if missing:
        errs.append(f"CRITICAL: {len(missing)} customers missing from Customer Tiers (e.g. {missing[:3]})")
    if dupes:
        errs.append(f"CRITICAL: {len(dupes)} customers duplicated in Customer Tiers (e.g. {dupes[:3]})")
    if extra:
        errs.append(f"CRITICAL: {len(extra)} unexpected customers in Customer Tiers (e.g. {extra[:3]})")
    if not sort_ok:
        errs.append("CRITICAL: Customer Tiers not sorted by Total_Spent descending")


def crit_vip_emails(tier_data, errs):
    """Exactly one email per VIP to that VIP's address, translated subject marker,
    body mentions that customer's own total amount (numeric)."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT to_addr, subject, COALESCE(body_text, '') FROM email.messages")
        msgs = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        errs.append(f"CRITICAL: cannot read sent emails: {e}")
        return

    vip_emails = tier_data["vip_emails"]
    vip_amounts = tier_data["vip_amounts"]

    # subject marker: keep matchable to the (translated) 'Welcome to Our VIP Program'.
    def is_vip_subject(s):
        sl = (s or "").lower()
        return "vip" in sl

    recip_to_msgs = {}
    for to_addr, subject, body in msgs:
        if not is_vip_subject(subject):
            continue
        # email.messages.to_addr is jsonb (a list); psycopg2 returns a Python list.
        if isinstance(to_addr, list):
            to_addr = ", ".join(str(x) for x in to_addr)
        addr = (to_addr or "").strip().lower()
        recip_to_msgs.setdefault(addr, []).append((subject, body))

    recip_set = set(recip_to_msgs.keys())
    if recip_set != vip_emails:
        missing = vip_emails - recip_set
        extra = recip_set - vip_emails
        errs.append(f"CRITICAL: VIP email recipients mismatch. missing={sorted(missing)[:3]} extra={sorted(extra)[:3]}")

    # each VIP must get exactly one VIP-subject email mentioning their own amount.
    for email in vip_emails:
        msgs_for = recip_to_msgs.get(email, [])
        if len(msgs_for) != 1:
            errs.append(f"CRITICAL: VIP {email} received {len(msgs_for)} VIP emails (expected 1)")
        amt = vip_amounts.get(email, 0.0)
        whole = int(round(amt))
        body_join = " ".join(b for _s, b in msgs_for)
        # strip whitespace + thousands separators so "4 586.91"/"4,586.91"/"4586.91" match.
        compact = re.sub(r"[\s,]", "", body_join)
        amt_variants = [f"{amt:.2f}", f"{amt:.1f}", str(whole)]
        if not any(v in compact for v in amt_variants):
            errs.append(f"CRITICAL: VIP {email} body does not mention own total amount ~{amt:.2f}")


def crit_teamly_page(errs):
    if not check_teamly_page_exists():
        errs.append("CRITICAL: Teamly page 'Customer Loyalty Program 2026' with tier statistics not found")


def crit_gsheet(errs):
    if not check_gsheet_exists():
        errs.append("CRITICAL: Google Sheet 'Customer Loyalty Program' with tier-summary data not found")


def check_vip_emails_sent(vip_count):
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM email.messages WHERE LOWER(subject) LIKE '%vip%'")
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
        return cnt >= vip_count
    except Exception:
        return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Customer_Loyalty_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Customer_Loyalty_Analysis.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    all_errors = []
    critical_errors = []

    # Tiers recomputed live from wc.customers (honest, not pre-seeded).
    try:
        tier_data = get_tier_data_from_db()
    except Exception as e:
        print(f"FAIL: Could not query DB for tier data: {e}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ----- CRITICAL semantic checks -----
    print("== CRITICAL checks ==")
    crit_tier_summary(agent_wb, tier_data, critical_errors)
    crit_customer_tiers(agent_wb, tier_data, critical_errors)
    crit_vip_emails(tier_data, critical_errors)
    crit_teamly_page(critical_errors)
    crit_gsheet(critical_errors)

    if critical_errors:
        print(f"\n=== RESULT: FAIL ({len(critical_errors)} CRITICAL errors) ===")
        for e in critical_errors:
            print(f"  {e}")
        sys.exit(1)
    print("  All critical checks PASS")

    # ----- Non-critical structural checks (accuracy gate) -----
    # Customer Tiers (structural)
    print("  Checking Customer Tiers sheet (structural)...")
    a_rows = load_sheet_rows(agent_wb, "Customer Tiers")
    if a_rows is None:
        all_errors.append("Sheet 'Customer Tiers' not found in agent output")
    else:
        data_rows = [r for r in a_rows[1:] if r and any(c is not None for c in r)]
        if len(data_rows) < 10:
            all_errors.append(f"Customer Tiers has too few rows: {len(data_rows)} (expected >= 10)")
        else:
            print(f"    PASS ({len(data_rows)} data rows)")

    # Tier Summary (structural existence already validated by critical)
    print("  Checking Tier Summary sheet (structural)...")
    a_rows = load_sheet_rows(agent_wb, "Tier Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Tier Summary' not found in agent output")
    else:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        if len(a_data) < 3:
            all_errors.append(f"Tier Summary has {len(a_data)} rows, expected 3")
        else:
            # avg within tol (revenue/count already critical)
            a_lookup = {str(r[0]).strip().lower(): r for r in a_data}
            for tier_name, exp_avg in [
                ("vip", tier_data["vip_avg"]),
                ("gold", tier_data["gold_avg"]),
                ("standard", tier_data["std_avg"]),
            ]:
                a_row = a_lookup.get(tier_name)
                if a_row is not None and len(a_row) > 3:
                    if not num_close(a_row[3], exp_avg, 5.0):
                        all_errors.append(f"{tier_name}.Avg_Spent: {a_row[3]} vs {exp_avg} (tol=5.0)")
            print("    PASS")

    # GSheet
    print("  Checking Google Sheet...")
    if check_gsheet_exists():
        print("    PASS")
    else:
        all_errors.append("Google Sheet 'Customer Loyalty Program' not found in DB")

    # VIP emails count
    print("  Checking VIP emails sent...")
    vip_count = len(tier_data["vip"])
    if check_vip_emails_sent(vip_count):
        print(f"    PASS (at least {vip_count} VIP emails)")
    else:
        all_errors.append(f"Expected >= {vip_count} VIP emails, not found in email.messages")

    # Teamly page
    print("  Checking Teamly page...")
    if check_teamly_page_exists():
        print("    PASS")
    else:
        all_errors.append("Teamly page 'Customer Loyalty Program 2026' not found")

    # Accuracy gate
    total_checks = 6
    passed = total_checks - len(all_errors)
    accuracy = 100.0 * passed / total_checks
    print(f"\nNon-critical accuracy: {accuracy:.1f}% ({passed}/{total_checks})")
    if all_errors:
        for e in all_errors[:10]:
            print(f"  {e}")
    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
