"""Evaluation for insales-review-summary.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Data values (product names/SKUs, review bodies) are russified CENTRALLY and the
groundtruth xlsx is already in sync — this script does NOT hand-translate any
wc.* realia literals.
"""
import argparse
import os
import sys
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Review Summary: filter rule >=3 reviews applied exactly",
    "Review Summary: every listed product has Reviews>=3",
    "Review Summary: rows sorted by Avg_Rating descending",
    "Summary: Products_With_Reviews and Total_Reviews exact",
    "Summary: Overall_Avg_Rating weighted average correct",
    "Teamly 'Product Review Dashboard' page exists with top/lowest products",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


# ---------------------------------------------------------------------------
# Ground-truth derived directly from the wc.* source of truth (DB).
# ---------------------------------------------------------------------------
def db_review_facts():
    """Return (set_of_product_names_with>=3_reviews, total_reviews,
    weighted_overall_avg, top_rated_name) computed from wc.product_reviews."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT p.name, COUNT(r.id) AS cnt, AVG(r.rating)::float AS avg_r
        FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        HAVING COUNT(r.id) >= 3
        """
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    names = {str(n).strip().lower() for n, _, _ in rows}
    total = sum(c for _, c, _ in rows)
    weighted = (sum(c * a for _, c, a in rows) / total) if total else 0.0
    # Top rated: highest avg, ties broken alphabetically CASE-INSENSITIVELY
    # (matches the groundtruth, which uses a case-insensitive min).
    top = sorted(rows, key=lambda x: (-x[2], x[0].lower()))[0][0] if rows else None
    return names, total, weighted, top


# ---------------------------------------------------------------------------
def check_review_summary(agent_wb, gt_wb, db_facts):
    a_rows = load_sheet_rows(agent_wb, "Review Summary")
    g_rows = load_sheet_rows(gt_wb, "Review Summary")
    if a_rows is None:
        check("Review Summary sheet present", False, "missing in agent output")
        # Cascade-fail the critical filter/sort checks too.
        check("Review Summary: filter rule >=3 reviews applied exactly", False, "no sheet")
        check("Review Summary: every listed product has Reviews>=3", False, "no sheet")
        check("Review Summary: rows sorted by Avg_Rating descending", False, "no sheet")
        return
    check("Review Summary sheet present", True)

    a_data = [r for r in a_rows[1:] if r and r[0] is not None]
    g_data = [r for r in (g_rows[1:] if g_rows else []) if r and r[0] is not None]

    a_lookup = {str(r[0]).strip().lower(): r for r in a_data}

    # Per-row numeric accuracy vs groundtruth (non-critical, lenient).
    for g_row in g_data:
        key = str(g_row[0]).strip().lower()
        a_row = a_lookup.get(key)
        if a_row is None:
            check(f"row present: {str(g_row[0])[:40]}", False, "missing")
            continue
        check(f"row present: {str(g_row[0])[:40]}", True)
        if len(a_row) > 1 and len(g_row) > 1:
            check(f"{key[:30]}.Reviews", num_close(a_row[1], g_row[1], 1),
                  f"{a_row[1]} vs {g_row[1]}")
        if len(a_row) > 2 and len(g_row) > 2:
            check(f"{key[:30]}.Avg_Rating", num_close(a_row[2], g_row[2], 0.1),
                  f"{a_row[2]} vs {g_row[2]}")
        if len(a_row) > 3 and len(g_row) > 3:
            check(f"{key[:30]}.Min_Rating", num_close(a_row[3], g_row[3], 0.1),
                  f"{a_row[3]} vs {g_row[3]}")
        if len(a_row) > 4 and len(g_row) > 4:
            check(f"{key[:30]}.Max_Rating", num_close(a_row[4], g_row[4], 0.1),
                  f"{a_row[4]} vs {g_row[4]}")

    # CRITICAL: core filter rule (>=3 reviews) validated directly against DB.
    db_names = db_facts[0]
    agent_names = set(a_lookup.keys())
    missing = db_names - agent_names      # products that SHOULD be listed but aren't
    extra = agent_names - db_names        # products listed that do NOT qualify
    check("Review Summary: filter rule >=3 reviews applied exactly",
          (not missing) and (not extra),
          f"missing={len(missing)} extra={len(extra)} "
          f"(e.g. missing={list(missing)[:2]}, extra={list(extra)[:2]})")

    # CRITICAL: every listed product's Reviews cell is >= 3.
    bad_counts = []
    for key, a_row in a_lookup.items():
        cnt = to_float(a_row[1]) if len(a_row) > 1 else None
        if cnt is None or cnt < 3:
            bad_counts.append((key[:30], cnt))
    check("Review Summary: every listed product has Reviews>=3",
          not bad_counts, f"violations={bad_counts[:3]}")

    # CRITICAL: rows sorted by Avg_Rating descending.
    avgs = [to_float(r[2]) for r in a_data if len(r) > 2 and to_float(r[2]) is not None]
    sorted_ok = all(avgs[i] >= avgs[i + 1] - 1e-9 for i in range(len(avgs) - 1))
    check("Review Summary: rows sorted by Avg_Rating descending",
          sorted_ok and len(avgs) >= 2, f"avg sequence head={avgs[:5]}")


def check_summary(agent_wb, gt_wb, db_facts):
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        check("Summary sheet present", False, "missing in agent output")
        check("Summary: Products_With_Reviews and Total_Reviews exact", False, "no sheet")
        check("Summary: Overall_Avg_Rating weighted average correct", False, "no sheet")
        return
    check("Summary sheet present", True)

    a_metrics = {}
    for r in a_rows[1:]:
        if r and r[0] is not None:
            a_metrics[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

    db_names, db_total, db_weighted, db_top = db_facts

    # CRITICAL: integer counts EXACT (tol=0) against DB-derived truth.
    pwr = a_metrics.get("products_with_reviews")
    tr = a_metrics.get("total_reviews")
    pwr_ok = num_close(pwr, len(db_names), 0)
    tr_ok = num_close(tr, db_total, 0)
    check("Summary: Products_With_Reviews and Total_Reviews exact",
          pwr_ok and tr_ok,
          f"PWR {pwr} vs {len(db_names)}; TR {tr} vs {db_total}")

    # CRITICAL: weighted overall average within tol=0.05.
    oar = a_metrics.get("overall_avg_rating")
    check("Summary: Overall_Avg_Rating weighted average correct",
          num_close(oar, round(db_weighted, 2), 0.05),
          f"{oar} vs {round(db_weighted, 2)}")

    # NON-critical: Top_Rated_Product. GT stores a 50-char truncation; the agent
    # may write the full name. Accept prefix match in either direction (RU/EN
    # names are kept English by policy).
    top = a_metrics.get("top_rated_product")
    top_s = str(top).strip().lower() if top is not None else ""
    db_top_s = str(db_top).strip().lower() if db_top else ""
    prefix_ok = bool(top_s) and bool(db_top_s) and (
        top_s.startswith(db_top_s[:40]) or db_top_s.startswith(top_s[:40])
    )
    check("Summary: Top_Rated_Product matches top-rated product",
          prefix_ok, f"{top_s[:50]} vs {db_top_s[:50]}")


def check_teamly(db_facts):
    """CRITICAL: a 'Product Review Dashboard' page exists in Teamly and its body
    references both the top-rated and a lowest-rated product."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages")
        pages = cur.fetchall()
        # Lowest-rated product (min avg, >=3 reviews) for content cross-check.
        cur.execute(
            """
            SELECT p.name
            FROM wc.product_reviews r
            JOIN wc.products p ON r.product_id = p.id
            GROUP BY p.name
            HAVING COUNT(r.id) >= 3
            ORDER BY AVG(r.rating) ASC, p.name ASC
            LIMIT 1
            """
        )
        low_row = cur.fetchone()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly 'Product Review Dashboard' page exists with top/lowest products",
              False, str(e))
        return

    dash = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "архив старых заметок" in tl:
            continue  # noise leftover
        if "product review dashboard" in tl or ("review" in tl and ("dashboard" in tl or "дашборд" in tl or "панель" in tl)):
            dash = (pid, title, body)
            break

    if dash is None:
        check("Teamly 'Product Review Dashboard' page exists with top/lowest products",
              False, f"pages={[(p[0], p[1]) for p in pages]}")
        return

    text = ((dash[1] or "") + " " + (dash[2] or "")).lower()
    top_name = (db_facts[3] or "").lower()
    low_name = (low_row[0] if low_row else "").lower()

    # Names are kept English; match on a discriminating prefix of each.
    top_ok = bool(top_name) and top_name[:30] in text
    low_ok = bool(low_name) and low_name[:30] in text
    check("Teamly 'Product Review Dashboard' page exists with top/lowest products",
          top_ok and low_ok,
          f"top_ref={top_ok} low_ref={low_ok}")


def run_evaluation(agent_workspace, groundtruth_workspace):
    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(agent_workspace, "WC_Review_Summary.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Review_Summary.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    db_facts = db_review_facts()

    print("  Checking Review Summary...")
    check_review_summary(agent_wb, gt_wb, db_facts)
    print("  Checking Summary...")
    check_summary(agent_wb, gt_wb, db_facts)
    print("  Checking Teamly dashboard...")
    check_teamly(db_facts)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    print(f"\n=== RESULT: {'PASS' if success else 'FAIL'} ===")
    return success


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success = run_evaluation(args.agent_workspace, args.groundtruth_workspace)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
