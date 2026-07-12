"""Evaluation script for fetch-sf-sales-competitor-excel-teamly.

Groundtruth is computed at runtime: ClickHouse aggregates (sf_data schema) are
joined against the external benchmark (data.json, region names russified to the
SAME central map as db/zzz_clickhouse_after_init.sql). Structural checks confirm
the workbook shape; CRITICAL_CHECKS confirm the numbers are actually correct.

Pass rule: NO critical failure AND accuracy >= 70%.
"""
import argparse
import json
import os
import sys
import tarfile

import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic checks: any failure here => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Data_Analysis: Revenue/Order_Count match ClickHouse per region",
    "Data_Analysis: Market_Size_M matches benchmark per region",
    "Data_Analysis: Market_Penetration_Pct == Revenue/Market_Size_M*100",
    "Teamly dashboard page exists with region + gap conclusion",
    "Recommendations: top priority targets lowest-penetration region",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:300] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def load_benchmark():
    """Load region -> market_size_m from the task's benchmark (russified regions).

    Reads data.json from the packaged tarball (always present), falling back to
    the extracted tmp copy. Region values must match the central ClickHouse map.
    """
    tar_path = os.path.join(TASK_ROOT, "files", "mock_pages.tar.gz")
    raw = None
    try:
        with tarfile.open(tar_path, "r:gz") as tar:
            member = tar.getmember("mock_pages/api/data.json")
            raw = tar.extractfile(member).read().decode("utf-8")
    except Exception:
        alt = os.path.join(TASK_ROOT, "tmp", "mock_pages", "api", "data.json")
        if os.path.exists(alt):
            with open(alt, encoding="utf-8") as f:
                raw = f.read()
    if raw is None:
        return {}
    data = json.loads(raw)
    return {b["region"]: float(b["market_size_m"]) for b in data.get("benchmarks", [])}


def compute_gt_aggregates():
    """Per-REGION (russified) Order_Count and Revenue from ClickHouse (sf_data)."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        '''
        SELECT c."REGION", COUNT(*) AS order_count, SUM(o."TOTAL_AMOUNT") AS revenue
        FROM sf_data."SALES_DW__PUBLIC__ORDERS" o
        JOIN sf_data."SALES_DW__PUBLIC__CUSTOMERS" c
          ON o."CUSTOMER_ID" = c."CUSTOMER_ID"
        GROUP BY c."REGION"
        '''
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {r[0]: {"order_count": int(r[1]), "revenue": float(r[2])} for r in rows}


def read_data_analysis(ws):
    """Return list of dicts keyed by header name (lowercased)."""
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return out


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    benchmark = load_benchmark()  # russified region -> market_size_m
    try:
        gt_agg = compute_gt_aggregates()  # russified region -> {order_count, revenue}
    except Exception as e:
        gt_agg = {}
        print(f"  [WARN] could not compute ClickHouse groundtruth: {e}")

    excel_path = os.path.join(agent_workspace, "Sales_Competitor_Report.xlsx")
    check("Sales_Competitor_Report.xlsx exists", os.path.exists(excel_path))

    da_rows = []
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ----- Data_Analysis structural -----
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_rows = read_data_analysis(ws)
            check("Data_Analysis has >= 5 rows", len(da_rows) >= 5, f"got {len(da_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Region', 'Order_Count', 'Revenue', 'Market_Size_M', 'Market_Penetration_Pct']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Region sort (ascending) — structural, non-critical.
            regions_in_order = [str(r.get("region") or "").strip() for r in da_rows]
            check("Data_Analysis sorted by Region (asc)",
                  regions_in_order == sorted(regions_in_order),
                  f"order: {regions_in_order}")

        # ----- Metrics structural -----
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            mrows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len([r for r in mrows if any(v is not None for v in r)]) >= 4,
                  f"got {len(mrows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in mrows:
                if r and r[0] is not None:
                    metrics_map[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        # ----- Recommendations structural -----
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        rec_rows = []
        rec_headers = []
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            rec_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            rec_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                        if any(v is not None for v in r)]
            check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")
            for expected_col in ['Priority', 'Action', 'Region']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in rec_headers, f"headers: {rec_headers[:8]}")

        # ===== CRITICAL semantic checks =====

        # Index agent rows by region for value comparison.
        agent_by_region = {}
        for r in da_rows:
            reg = str(r.get("region") or "").strip()
            if reg:
                agent_by_region[reg] = r

        # 1) Revenue & Order_Count match ClickHouse aggregates per region.
        if gt_agg:
            ok_agg = True
            details = []
            for reg, agg in gt_agg.items():
                ar = agent_by_region.get(reg)
                if ar is None:
                    ok_agg = False
                    details.append(f"missing region {reg!r}")
                    continue
                oc = safe_float(ar.get("order_count"))
                rev = safe_float(ar.get("revenue"))
                if oc is None or int(round(oc)) != agg["order_count"]:
                    ok_agg = False
                    details.append(f"{reg}: Order_Count {oc} != {agg['order_count']}")
                if rev is None or abs(rev - agg["revenue"]) > max(1.0, abs(agg["revenue"]) * 0.005):
                    ok_agg = False
                    details.append(f"{reg}: Revenue {rev} != {agg['revenue']:.2f}")
            check("Data_Analysis: Revenue/Order_Count match ClickHouse per region",
                  ok_agg, "; ".join(details[:5]))
        else:
            check("Data_Analysis: Revenue/Order_Count match ClickHouse per region",
                  False, "ClickHouse aggregates unavailable")

        # 2) Market_Size_M matches benchmark (data.json) per region.
        if benchmark:
            ok_ms = True
            details = []
            for reg, ms in benchmark.items():
                ar = agent_by_region.get(reg)
                if ar is None:
                    ok_ms = False
                    details.append(f"missing region {reg!r}")
                    continue
                got = safe_float(ar.get("market_size_m"))
                if got is None or abs(got - ms) > 0.5:
                    ok_ms = False
                    details.append(f"{reg}: Market_Size_M {got} != {ms}")
            check("Data_Analysis: Market_Size_M matches benchmark per region",
                  ok_ms, "; ".join(details[:5]))
        else:
            check("Data_Analysis: Market_Size_M matches benchmark per region",
                  False, "benchmark unavailable")

        # 3) Market_Penetration_Pct == Revenue/Market_Size_M*100 per row (cross-source join succeeded).
        ok_pen = len(da_rows) >= 5
        details = []
        for r in da_rows:
            rev = safe_float(r.get("revenue"))
            ms = safe_float(r.get("market_size_m"))
            pen = safe_float(r.get("market_penetration_pct"))
            if rev is None or ms is None or pen is None or ms == 0:
                ok_pen = False
                details.append(f"{r.get('region')}: non-numeric/null")
                continue
            expected = rev / ms * 100.0
            if abs(pen - expected) > max(0.5, abs(expected) * 0.005):
                ok_pen = False
                details.append(f"{r.get('region')}: {pen} != {expected:.2f}")
        check("Data_Analysis: Market_Penetration_Pct == Revenue/Market_Size_M*100",
              ok_pen, "; ".join(details[:5]))

        # 5) Recommendations: top priority targets the lowest-penetration region.
        if rec_rows and 'priority' in rec_headers and 'region' in rec_headers and da_rows:
            pi = rec_headers.index('priority')
            ri = rec_headers.index('region')
            # lowest penetration region from Data_Analysis
            pen_pairs = []
            for r in da_rows:
                pen = safe_float(r.get("market_penetration_pct"))
                reg = str(r.get("region") or "").strip()
                if pen is not None and reg:
                    pen_pairs.append((pen, reg))
            min_region = min(pen_pairs)[1] if pen_pairs else None
            # top priority row = smallest Priority value
            def _pr(v):
                f = safe_float(v[pi] if pi < len(v) else None)
                return f if f is not None else 1e9
            top_row = min(rec_rows, key=_pr)
            top_region = str(top_row[ri]).strip() if ri < len(top_row) and top_row[ri] is not None else ""
            check("Recommendations: top priority targets lowest-penetration region",
                  min_region is not None and top_region == min_region,
                  f"top={top_region!r} expected={min_region!r}")
        else:
            check("Recommendations: top priority targets lowest-penetration region",
                  False, "missing Recommendations data/columns")

        check("sf_competitor_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "sf_competitor_processor.py")))
    else:
        # No workbook -> the value/semantic critical checks cannot pass.
        for n in ["Data_Analysis: Revenue/Order_Count match ClickHouse per region",
                  "Data_Analysis: Market_Size_M matches benchmark per region",
                  "Data_Analysis: Market_Penetration_Pct == Revenue/Market_Size_M*100",
                  "Recommendations: top priority targets lowest-penetration region"]:
            check(n, False, "workbook missing")

    # ----- Teamly dashboard (replaces the old notion.pages check) -----
    russ_regions = set(benchmark.keys()) if benchmark else {
        "Азиатско-Тихоокеанский регион", "Европа", "Латинская Америка",
        "Ближний Восток", "Северная Америка",
    }
    try:
        conn = get_conn()
        cur = conn.cursor()
        # New user pages only (seeds have id <= 3); exclude the seeded noise page.
        cur.execute(
            """
            SELECT p.title, p.body
            FROM teamly.pages p
            JOIN teamly.spaces s ON p.space_id = s.id
            WHERE p.id > 3
              AND s.key = 'SFCOMP'
              AND p.title NOT ILIKE '%Старые заметки проекта%'
            """
        )
        rows = cur.fetchall()
        conn.close()
        dash_ok = False
        for title, body in rows:
            tl = (title or "").lower()
            text = ((title or "") + "\n" + (body or "")).lower()
            if "dashboard" not in tl and "дашборд" not in tl:
                continue
            has_region = any(reg.lower() in text for reg in russ_regions)
            has_gap = any(k in text for k in [
                "проникнов", "разрыв", "gap", "penetration", "потенциал", "отстава"])
            if has_region and has_gap and (body or "").strip():
                dash_ok = True
                break
        check("Teamly dashboard page exists with region + gap conclusion",
              dash_ok, f"scanned {len(rows)} SFCOMP pages")
    except Exception as e:
        check("Teamly dashboard page exists with region + gap conclusion", False, str(e))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({
                "total_passed": PASS_COUNT,
                "total_checks": total,
                "accuracy": accuracy,
                "critical_failed": critical_failed,
            }, f, indent=2)

    if critical_failed:
        return False, f"FAIL (critical): {PASS_COUNT}/{total} checks"
    success = accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
