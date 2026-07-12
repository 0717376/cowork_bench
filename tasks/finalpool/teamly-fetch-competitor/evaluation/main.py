"""
Evaluation script for the competitor-analysis task (Teamly variant).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Checks:
1. Excel file (Competitor_Analysis.xlsx) - two sheets with correct product data.
2. Teamly page exists with the competitor analysis summary.
3. Memory file (memory/memory.json) has been updated with analysis notes.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

EXPECTED_PRODUCTS = [
    {"name": "Gamma Ultra", "price": 499, "rating": 4.8},
    {"name": "Zeta Max", "price": 599, "rating": 4.7},
    {"name": "Iota Edge", "price": 399, "rating": 4.6},
    {"name": "Alpha Pro", "price": 299, "rating": 4.5},
    {"name": "Epsilon Core", "price": 349, "rating": 4.3},
    {"name": "Beta Suite", "price": 199, "rating": 4.2},
    {"name": "Theta Plus", "price": 249, "rating": 4.1},
    {"name": "Kappa Flex", "price": 159, "rating": 4.0},
    {"name": "Delta Lite", "price": 89, "rating": 3.9},
    {"name": "Eta Basic", "price": 49, "rating": 3.5},
]

DETAILED_PRODUCTS = ["Alpha Pro", "Gamma Ultra", "Zeta Max"]

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Product Listing data matches expected",
    "Detailed Analysis has 3 rows",
    "Teamly page with 'Competitor' in title exists",
    "Teamly page mentions 10 products",
    "Teamly page mentions average price ~289",
    "Teamly page mentions Gamma Ultra",
    "Memory file has been updated with notes",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def str_match(a, b):
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Competitor_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        record("Product Listing data matches expected", False, "no excel")
        record("Detailed Analysis has 3 rows", False, "no excel")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        record("Product Listing data matches expected", False, "unreadable")
        record("Detailed Analysis has 3 rows", False, "unreadable")
        return False

    all_ok = True

    # Check Product Listing sheet
    sheet_name = None
    for name in wb.sheetnames:
        if "product" in name.lower() and "listing" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        for name in wb.sheetnames:
            if "listing" in name.lower() or "product" in name.lower():
                sheet_name = name
                break

    listing_ok = True
    if not sheet_name:
        record("Sheet 'Product Listing' exists", False, f"Sheets found: {wb.sheetnames}")
        all_ok = False
        listing_ok = False
        record("Product Listing data matches expected", False, "no sheet")
    else:
        record("Sheet 'Product Listing' exists", True)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        if len(data_rows) != 10:
            record("Product Listing has 10 rows", False, f"Found {len(data_rows)} rows")
            all_ok = False
            listing_ok = False
        else:
            record("Product Listing has 10 rows", True)
            # task.md does NOT mandate any row ordering (only "one row per
            # product"). Match each expected product to a row BY NAME so an
            # agent that outputs API order (or any order) is not penalized.
            expected_by_name = {p["name"].strip().lower(): p for p in EXPECTED_PRODUCTS}
            actual_by_name = {}
            for row in data_rows:
                pname = str(row[0]).strip().lower() if row[0] else ""
                if pname:
                    actual_by_name[pname] = row
            for expected in EXPECTED_PRODUCTS:
                key = expected["name"].strip().lower()
                row = actual_by_name.get(key)
                if row is None:
                    record(f"Product '{expected['name']}' present", False,
                           "Missing from Product Listing")
                    listing_ok = False
                    continue
                price_ok = num_close(row[1], expected["price"], tol=5.0)
                if not price_ok:
                    record(f"{expected['name']} price", False,
                           f"Got {row[1]}, expected {expected['price']}")
                    listing_ok = False
                rating_ok = num_close(row[2], expected["rating"], tol=0.2)
                if not rating_ok:
                    record(f"{expected['name']} rating", False,
                           f"Got {row[2]}, expected {expected['rating']}")
                    listing_ok = False

        # CRITICAL: semantic correctness of the fetched 10-product data.
        record("Product Listing data matches expected", listing_ok)
        if not listing_ok:
            all_ok = False

        # Reverse: check no extra rows beyond 10
        if len(data_rows) > 10:
            record("Product Listing has no extra rows", False,
                   f"Found {len(data_rows)} rows, expected exactly 10")
            all_ok = False

    # Check Detailed Analysis sheet
    detail_sheet = None
    for name in wb.sheetnames:
        if "detail" in name.lower():
            detail_sheet = name
            break
    if not detail_sheet:
        record("Sheet 'Detailed Analysis' exists", False, f"Sheets found: {wb.sheetnames}")
        all_ok = False
        record("Detailed Analysis has 3 rows", False, "no sheet")
    else:
        record("Sheet 'Detailed Analysis' exists", True)
        ws = wb[detail_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        detail_rows_ok = len(data_rows) == 3
        record("Detailed Analysis has 3 rows", detail_rows_ok,
               "" if detail_rows_ok else f"Found {len(data_rows)} rows")
        if not detail_rows_ok:
            all_ok = False
        if data_rows:
            found_products = set()
            for row in data_rows:
                pname = str(row[0]).strip() if row[0] else ""
                for dp in DETAILED_PRODUCTS:
                    if str_match(pname, dp):
                        found_products.add(dp.lower())
            for dp in DETAILED_PRODUCTS:
                if dp.lower() in found_products:
                    record(f"Detailed Analysis contains {dp}", True)
                else:
                    record(f"Detailed Analysis contains {dp}", False, "Not found")
                    all_ok = False

        # Reverse: check no extra rows beyond 3
        if len(data_rows) > 3:
            record("Detailed Analysis has no extra rows", False,
                   f"Found {len(data_rows)} rows, expected exactly 3")
            all_ok = False

    # Reverse validation: no unexpected sheets
    expected_sheets = {"product listing", "detailed analysis"}
    actual_sheets = set()
    for s in wb.sheetnames:
        normalized = s.strip().lower()
        actual_sheets.add(normalized)
    extra = actual_sheets - expected_sheets
    if extra:
        record("No unexpected extra sheets", False, f"Extra sheets: {extra}")
        all_ok = False
    else:
        record("No unexpected extra sheets", True)

    wb.close()
    return all_ok


def _avg_price_present(text):
    """Accept the average price ~289 in EN or RU-formatted variants."""
    candidates = ["289", "288", "290", "289.0", "289,0", "289.40", "289,40"]
    return any(c in text for c in candidates)


def check_teamly():
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # User-created pages only (global seed pages have id <= 3).
        cur.execute(
            "SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3"
        )
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Teamly DB accessible", False, str(e))
        record("Teamly page with 'Competitor' in title exists", False, str(e))
        record("Teamly page mentions 10 products", False, str(e))
        record("Teamly page mentions average price ~289", False, str(e))
        record("Teamly page mentions Gamma Ultra", False, str(e))
        return False

    all_ok = True

    # Identify the competitor analysis page by title (EN or RU marker).
    page = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "competitor" in tl or "конкурент" in tl:
            page = (pid, title, body)
            break

    if page is None:
        record("Teamly page with 'Competitor' in title exists", False,
               f"Found {len(pages)} pages but none match: "
               f"{[(p[0], p[1]) for p in pages]}")
        record("Teamly page mentions 10 products", False, "no page")
        record("Teamly page mentions average price ~289", False, "no page")
        record("Teamly page mentions Gamma Ultra", False, "no page")
        return False

    record("Teamly page with 'Competitor' in title exists", True)
    all_text = ((page[1] or "") + " " + str(page[2] or "")).lower()

    has_count = "10" in all_text or "десят" in all_text
    record("Teamly page mentions 10 products", has_count)
    if not has_count:
        all_ok = False

    has_avg_price = _avg_price_present(all_text)
    record("Teamly page mentions average price ~289", has_avg_price)
    if not has_avg_price:
        all_ok = False

    has_highest = "gamma" in all_text
    record("Teamly page mentions Gamma Ultra", has_highest)
    if not has_highest:
        all_ok = False

    # Reverse: no extra competitor pages beyond the expected one.
    extra_pages = []
    for pid, title, body in pages:
        if pid == page[0]:
            continue
        tl = (title or "").lower()
        if "competitor" in tl or "конкурент" in tl:
            extra_pages.append(title)
    if extra_pages:
        record("No unexpected Teamly competitor pages", False,
               f"Found {len(extra_pages)} extra page(s): {extra_pages[:3]}")
        all_ok = False
    else:
        record("No unexpected Teamly competitor pages", True)

    return all_ok


def check_memory(agent_workspace):
    print("\n=== Checking Memory ===")
    memory_path = os.path.join(agent_workspace, "memory", "memory.json")
    if not os.path.isfile(memory_path):
        record("Memory file exists", False, f"Not found: {memory_path}")
        record("Memory file has been updated with notes", False, "no file")
        return False
    record("Memory file exists", True)

    try:
        with open(memory_path, "r") as f:
            data = json.load(f)
    except Exception as e:
        record("Memory file is valid JSON", False, str(e))
        record("Memory file has been updated with notes", False, "invalid json")
        return False
    record("Memory file is valid JSON", True)

    content = json.dumps(data, ensure_ascii=False).lower()
    has_notes = len(content) > 30
    record("Memory file has been updated with notes", has_notes)
    has_analysis = (
        "analysis" in content or "competitor" in content or "complete" in content
        or "анализ" in content or "конкурент" in content or "заверш" in content
    )
    record("Memory mentions analysis/competitor/complete", has_analysis)
    return has_notes and has_analysis


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()
    check_memory(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    print(f"  Overall:  {'PASS' if success else 'FAIL'}")
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
