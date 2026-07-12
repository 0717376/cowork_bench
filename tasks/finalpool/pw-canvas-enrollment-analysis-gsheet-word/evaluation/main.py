"""Evaluation script for pw-canvas-enrollment-analysis-gsheet-word.

Структурные проверки (NON-critical): наличие файлов/листов/столбцов и
минимальное число строк. Они НЕ должны хардкодить волатильные значения.

CRITICAL-проверки (семантические): любой провал => sys.exit(1) ДО порога точности.
Ожидаемые значения берутся из committed groundtruth xlsx
(сгенерирован из того же глобального сида Canvas), а НЕ хардкодятся числами,
зависящими от сида. Сравнение — с допуском.

Порог: точность (PASS / total) >= 70  И  нет ни одного провала critical => PASS.
"""
import os
import argparse, json, os, sys
import glob as globmod
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# допуски для семантических проверок
TOL_ENROLLMENT = 0          # суммы целочисленные, должны совпадать точно
TOL_SCORE = 1.5             # средние баллы/доли — небольшой float-допуск
TOL_ROW = 2                 # числовые значения по строкам


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{'[CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL]{'[CRIT]' if critical else ''} {name}: {detail_str}")


def norm_rate(val):
    """Нормализация шкалы pass-rate: доля 0–1 -> проценты 0–100."""
    if val is None:
        return None
    return val * 100 if val <= 1.0 else val


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _row_dict(ws):
    """Вернуть (headers_lower, list[dict]) для листа Data_Analysis."""
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    hlow = [h.lower() for h in headers]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(dict(zip(hlow, r)))
    return hlow, rows


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    excel_path = os.path.join(agent_workspace, "Enrollment_Analysis_Report.xlsx")
    check("Enrollment_Analysis_Report.xlsx exists", os.path.exists(excel_path))

    gt_path = os.path.join(groundtruth_workspace, "Enrollment_Analysis_Report.xlsx")
    gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

    # ----- groundtruth-производные ожидаемые значения -----
    gt_da = {}   # code(lower) -> dict(enrollment, avg_score, pass_rate)
    gt_metrics = {}
    gt_worst_course = None  # курс с наибольшим отрицательным разрывом (по groundtruth Recommendations: Priority 1)
    if gt_wb is not None:
        if "Data_Analysis" in gt_wb.sheetnames:
            _, gt_rows = _row_dict(gt_wb["Data_Analysis"])
            for r in gt_rows:
                code = str(r.get("code", "")).strip().lower()
                if code:
                    gt_da[code] = {
                        "enrollment": safe_float(r.get("enrollment")),
                        "avg_score": safe_float(r.get("avg_score")),
                        "pass_rate": norm_rate(safe_float(r.get("pass_rate"))),
                    }
        if "Metrics" in gt_wb.sheetnames:
            for r in gt_wb["Metrics"].iter_rows(min_row=2, values_only=True):
                if r and r[0]:
                    gt_metrics[str(r[0]).strip().lower()] = safe_float(r[1])
        if "Recommendations" in gt_wb.sheetnames:
            rec_rows = [r for r in gt_wb["Recommendations"].iter_rows(min_row=2, values_only=True)
                        if r and any(v is not None for v in r)]
            # первая строка (Priority 1) — курс с наибольшим отрицательным разрывом
            if rec_rows:
                gt_worst_course = str(rec_rows[0][-1]).strip().lower()

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ================= STRUCTURAL (non-critical) =================
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        da_rows = []
        da_headers = []
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_headers, da_rows = _row_dict(ws)
            check("Data_Analysis has >= 6 rows", len(da_rows) >= 6, f"got {len(da_rows)}")
            for expected_col in ['Course', 'Code', 'Enrollment', 'Avg_Score', 'Pass_Rate']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in da_headers, f"headers: {da_headers[:8]}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            mrows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len([r for r in mrows if r and r[0]]) >= 4,
                  f"got {len(mrows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in mrows:
                if r and r[0]:
                    metrics_map[str(r[0]).strip().lower()] = safe_float(r[1])

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        rec_rows = []
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            rec_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                        if r and any(v is not None for v in r)]
            check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Course']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ================= CRITICAL (semantic) =================

        # 1) Data_Analysis: курсы соответствуют живому сиду Canvas (>=6 курсов,
        #    коды вида *-2013J/*-2014J/*-2013B/*-2014B), значения числовые и в
        #    пределах допуска от groundtruth-производных (не выдуманные).
        codes = [str(r.get("code", "")).strip() for r in da_rows]
        codes_low = [c.lower() for c in codes if c]
        has_seed_codes = sum(1 for c in codes_low
                             if any(suf in c for suf in ("-2013j", "-2014j", "-2013b", "-2014b"))) >= 4
        check("CRITICAL: Data_Analysis содержит коды курсов из сида Canvas (>=4 вида *-2013J/*-2014J/...)",
              len(da_rows) >= 6 and has_seed_codes, f"codes={codes_low}", critical=True)

        # значения строк против groundtruth по коду
        if gt_da:
            matched, within = 0, 0
            for r in da_rows:
                code = str(r.get("code", "")).strip().lower()
                if code in gt_da:
                    matched += 1
                    g = gt_da[code]
                    en = safe_float(r.get("enrollment"))
                    sc = safe_float(r.get("avg_score"))
                    pr = norm_rate(safe_float(r.get("pass_rate")))
                    ok = (en is not None and g["enrollment"] is not None
                          and abs(en - g["enrollment"]) <= TOL_ROW
                          and sc is not None and g["avg_score"] is not None
                          and abs(sc - g["avg_score"]) <= TOL_SCORE
                          and pr is not None and g["pass_rate"] is not None
                          and abs(pr - g["pass_rate"]) <= TOL_SCORE)
                    if ok:
                        within += 1
            check("CRITICAL: значения Enrollment/Avg_Score/Pass_Rate совпадают с источником (groundtruth) в пределах допуска",
                  matched >= 6 and within >= max(5, matched - 1),
                  f"matched={matched}, within_tol={within}", critical=True)
        else:
            check("CRITICAL: groundtruth доступен для сверки значений", False,
                  "groundtruth xlsx not found", critical=True)

        # 2) Data_Analysis отсортирован по алфавиту по столбцу Course
        course_vals = [str(r.get("course", "")) for r in da_rows]
        check("CRITICAL: Data_Analysis отсортирован по алфавиту по столбцу Course",
              len(course_vals) >= 6 and course_vals == sorted(course_vals, key=lambda s: s.lower()),
              f"order={course_vals}", critical=True)

        # 3) Metrics семантика: Total_Courses == число строк Data_Analysis;
        #    Total_Enrollment == сумма Enrollment; средние в пределах допуска от groundtruth
        agent_total_enr = sum(safe_float(r.get("enrollment"), 0) or 0 for r in da_rows)
        tc = metrics_map.get("total_courses")
        te = metrics_map.get("total_enrollment")
        ok_tc = tc is not None and abs(tc - len(da_rows)) < 0.5
        ok_te = te is not None and abs(te - agent_total_enr) <= TOL_ENROLLMENT
        check("CRITICAL: Metrics Total_Courses == число строк Data_Analysis и Total_Enrollment == сумма Enrollment",
              ok_tc and ok_te,
              f"Total_Courses={tc} rows={len(da_rows)}; Total_Enrollment={te} sum={agent_total_enr}",
              critical=True)

        if gt_metrics:
            checks_avg = []
            for key in ("avg_score", "avg_pass_rate"):
                if key in gt_metrics and gt_metrics[key] is not None:
                    av = metrics_map.get(key)
                    gv = gt_metrics[key]
                    if key == "avg_pass_rate":
                        av, gv = norm_rate(av), norm_rate(gv)
                    checks_avg.append(av is not None and abs(av - gv) <= TOL_SCORE)
            check("CRITICAL: Metrics Avg_Score/Avg_Pass_Rate в пределах допуска от groundtruth",
                  len(checks_avg) >= 1 and all(checks_avg),
                  f"agent={ {k: metrics_map.get(k) for k in ('avg_score','avg_pass_rate')} } gt={gt_metrics}",
                  critical=True)

        # 4) Recommendations gap-логика: худший курс (наибольший отрицательный разрыв)
        #    присутствует среди приоритетных действий.
        if gt_worst_course:
            rec_text = " ".join(str(v) for r in rec_rows for v in r if v is not None).lower()
            # сопоставляем по ключевому фрагменту названия курса
            frag = gt_worst_course
            present = frag in rec_text
            if not present:
                # запасное сопоставление по «значимым» словам названия
                words = [w for w in frag.replace("(", " ").replace(")", " ").split()
                         if len(w) > 3]
                present = sum(1 for w in words if w in rec_text) >= max(1, len(words) - 1)
            check("CRITICAL: худший курс (наибольший отрицательный разрыв) есть в Recommendations",
                  present, f"worst='{gt_worst_course}' rec='{rec_text[:160]}'", critical=True)

        # 5) Провенанс: course_enrollment_processor.py существует И
        #    course_enrollment_results.json реально создан и является валидным JSON.
        proc_path = os.path.join(agent_workspace, "course_enrollment_processor.py")
        json_path = os.path.join(agent_workspace, "course_enrollment_results.json")
        proc_ok = os.path.exists(proc_path)
        json_ok = False
        if os.path.exists(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                json_ok = data is not None and (isinstance(data, (list, dict)) and len(data) > 0)
            except Exception:
                json_ok = False
        check("CRITICAL: course_enrollment_processor.py существует И course_enrollment_results.json — валидный непустой JSON",
              proc_ok and json_ok,
              f"processor={proc_ok}, results_json={json_ok}", critical=True)

        # ===== прочие структурные =====
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT title FROM gsheet.spreadsheets WHERE title ILIKE %s", ('%tracker%',))
            sheets = cur.fetchall()
            check("Google Sheet created", len(sheets) >= 1, f"found {len(sheets)} sheets")
            conn.close()
        except Exception as e:
            check("GSheet check", False, str(e))

        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content", len(text) > 50, f"text length: {len(text)}")

        check("course_enrollment_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "course_enrollment_processor.py")))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    success = (not CRITICAL_FAILED) and accuracy >= 70
    msg = f"Passed {PASS_COUNT}/{total} checks (accuracy={accuracy:.1f}%)"
    if CRITICAL_FAILED:
        msg += f" | CRITICAL FAILED: {CRITICAL_FAILED}"
    return success, msg


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)

    # Любой провал critical => немедленный FAIL (до порога точности).
    if CRITICAL_FAILED:
        sys.exit(1)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
