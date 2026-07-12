"""Evaluation for canvas-enrollment-overview."""
import argparse
import os
import sys
import openpyxl

try:
    import psycopg2
except Exception:
    psycopg2 = None

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_to_lookup(rows):
    """rows without header -> {key_lower: row}."""
    lookup = {}
    for row in rows:
        if row and row[0] is not None:
            lookup[str(row[0]).strip().lower()] = row
    return lookup


def check_email_deliverable():
    """CRITICAL: письмо на registrar@openuniversity.ac.uk с точной темой
    'Canvas Enrollment Overview Report' и телом, упоминающим итоговые числа."""
    recipient = "registrar@openuniversity.ac.uk"
    subject_exact = "canvas enrollment overview report"
    if psycopg2 is None:
        return False, "psycopg2 недоступен"
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        return False, f"нет подключения к БД: {e}"
    try:
        cur = conn.cursor()
        try:
            cur.execute(
                """
                SELECT m.subject, m.body_text
                  FROM email.messages m
                  JOIN email.sent_log s ON s.message_id = m.id
                 WHERE m.to_addr::text ILIKE %s
                """,
                (f"%{recipient}%",),
            )
            rows = cur.fetchall()
        except Exception:
            rows = []
        if not rows:
            cur.execute(
                """
                SELECT subject, body_text FROM email.messages
                 WHERE to_addr::text ILIKE %s
                """,
                (f"%{recipient}%",),
            )
            rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()

    if not rows:
        return False, f"нет письма на {recipient}"

    subj_ok = any(str(s or "").strip().lower() == subject_exact for s, _b in rows)
    if not subj_ok:
        return False, f"тема не совпадает точно с 'Canvas Enrollment Overview Report'"

    # Тело должно упоминать итоговые числа (любая из агрегатных величин).
    body = " ".join(str(b or "") for _s, b in rows).lower()
    has_number = any(tok in body for tok in ("32605", "32 605", "23", "41"))
    if not has_number:
        return False, "в теле письма нет упоминания итоговых чисел"
    return True, f"письмо найдено ({len(rows)} шт.)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Enrollment_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Enrollment_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_failures = []

    # -------- Load sheets --------
    a_enr = load_sheet_rows(agent_wb, "Enrollment Data")
    g_enr = load_sheet_rows(gt_wb, "Enrollment Data")
    a_sum = load_sheet_rows(agent_wb, "Summary")
    g_sum = load_sheet_rows(gt_wb, "Summary")

    # -------- Non-critical: Enrollment Data per-cell diff --------
    print(f"  Checking Enrollment Data...")
    if a_enr is None:
        all_errors.append("Sheet 'Enrollment Data' not found in agent output")
        critical_failures.append("Лист 'Enrollment Data' отсутствует в выводе агента")
    elif g_enr is None:
        all_errors.append("Sheet 'Enrollment Data' not found in groundtruth")
    else:
        a_data = a_enr[1:] if len(a_enr) > 1 else []
        g_data = g_enr[1:] if len(g_enr) > 1 else []
        a_lookup = rows_to_lookup(a_data)
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 5):
                    errors.append(f"{key}.Students: {a_row[2]} vs {g_row[2]} (tol=5)")
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1):
                    errors.append(f"{key}.Teachers: {a_row[3]} vs {g_row[3]} (tol=1)")
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1):
                    errors.append(f"{key}.TAs: {a_row[4]} vs {g_row[4]} (tol=1)")
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 5):
                    errors.append(f"{key}.Total_Enrolled: {a_row[5]} vs {g_row[5]} (tol=5)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # -------- Non-critical: Summary per-cell diff --------
    print(f"  Checking Summary...")
    if a_sum is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
        critical_failures.append("Лист 'Summary' отсутствует в выводе агента")
    elif g_sum is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_sum[1:] if len(a_sum) > 1 else []
        g_data = g_sum[1:] if len(g_sum) > 1 else []
        a_lookup = rows_to_lookup(a_data)
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 10.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=10.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # ================= CRITICAL CHECKS =================
    print("\n  === CRITICAL CHECKS ===")

    a_enr_lookup = rows_to_lookup(a_enr[1:]) if a_enr and len(a_enr) > 1 else {}
    g_enr_lookup = rows_to_lookup(g_enr[1:]) if g_enr and len(g_enr) > 1 else {}
    a_sum_lookup = rows_to_lookup(a_sum[1:]) if a_sum and len(a_sum) > 1 else {}
    g_sum_lookup = rows_to_lookup(g_sum[1:]) if g_sum and len(g_sum) > 1 else {}

    # CRITICAL 1: полнота — строка для каждого Course_Code из GT.
    missing = [g_enr_lookup[k][0] for k in g_enr_lookup if k not in a_enr_lookup]
    if missing:
        critical_failures.append(f"Отсутствуют строки курсов: {missing}")
    print(f"    [crit] Полнота списка курсов: missing={len(missing)}")

    # CRITICAL 2: Total_Courses == GT (22).
    g_tc = g_sum_lookup.get("total_courses")
    a_tc = a_sum_lookup.get("total_courses")
    gt_tc = g_tc[1] if g_tc and len(g_tc) > 1 else None
    if not (a_tc and len(a_tc) > 1 and num_close(a_tc[1], gt_tc, 0)):
        critical_failures.append(
            f"Summary.Total_Courses != {gt_tc} (получено {a_tc[1] if a_tc and len(a_tc) > 1 else None})"
        )
    print(f"    [crit] Total_Courses ожидается {gt_tc}")

    # CRITICAL 3: Largest_Course / Smallest_Course — точное совпадение строк с GT.
    for metric in ("largest_course", "smallest_course"):
        g_m = g_sum_lookup.get(metric)
        a_m = a_sum_lookup.get(metric)
        gt_v = g_m[1] if g_m and len(g_m) > 1 else None
        a_v = a_m[1] if a_m and len(a_m) > 1 else None
        if not str_match(a_v, gt_v):
            critical_failures.append(
                f"Summary.{metric}: {a_v} != {gt_v} (требуется точное совпадение)"
            )
        print(f"    [crit] {metric} ожидается {gt_v}")

    # CRITICAL 4: агрегаты Total_Students (tol=10) и Total_Teachers (tol=1).
    for metric, tol in (("total_students", 10.0), ("total_teachers", 1.0)):
        g_m = g_sum_lookup.get(metric)
        a_m = a_sum_lookup.get(metric)
        gt_v = g_m[1] if g_m and len(g_m) > 1 else None
        a_v = a_m[1] if a_m and len(a_m) > 1 else None
        if not num_close(a_v, gt_v, tol):
            critical_failures.append(
                f"Summary.{metric}: {a_v} vs {gt_v} (tol={tol})"
            )
        print(f"    [crit] {metric} ожидается {gt_v} (tol={tol})")

    # CRITICAL 5: внутренняя согласованность Total_Enrolled = Students+Teachers+TAs.
    inconsistent = []
    for key, a_row in a_enr_lookup.items():
        if key not in g_enr_lookup:
            continue  # лишние строки не штрафуем
        try:
            s = float(a_row[2]); t = float(a_row[3]); ta = float(a_row[4]); tot = float(a_row[5])
        except (TypeError, ValueError, IndexError):
            inconsistent.append(key)
            continue
        if abs((s + t + ta) - tot) > 0.5:
            inconsistent.append(f"{key}({s}+{t}+{ta}!={tot})")
    if inconsistent:
        critical_failures.append(
            f"Total_Enrolled != Students+Teachers+TAs для: {inconsistent[:5]}"
        )
    print(f"    [crit] Согласованность Total_Enrolled: нарушений={len(inconsistent)}")

    # CRITICAL 6: письмо-результат (тема точная, тело упоминает итоги).
    email_ok, email_msg = check_email_deliverable()
    if not email_ok:
        critical_failures.append(f"Письмо: {email_msg}")
    print(f"    [crit] Письмо на registrar@openuniversity.ac.uk: {email_msg}")

    # ================= GATES =================
    if critical_failures:
        print(f"\n=== CRITICAL FAILURES ({len(critical_failures)}) ===")
        for c in critical_failures:
            print(f"  CRITICAL: {c}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
