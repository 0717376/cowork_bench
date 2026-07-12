"""
Evaluation script for sf-sales-forecast-notion task (Teamly + ClickHouse).

Checks:
1. Teamly page "Sales Performance Dashboard" exists with both
   "Monthly Revenue" and "Regional Performance" sections.
2. Regional Performance: every region (russified REGION values from sf_data,
   read dynamically via load_expected) appears with its exact revenue and
   order count in BOTH the Excel Regional sheet and the Teamly page body.
3. Monthly Revenue: representative complete months (first / middle / last)
   appear with correct Order Count AND Revenue in BOTH the Excel Monthly sheet
   and the Teamly page; the trailing partial month (2026-03) is excluded;
   monthly entry count in Excel equals len(expected_monthly) exactly.
4. Excel workbook has the two required sheets with the correct header rows.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy
(sys.exit(1) before the accuracy gate). Otherwise PASS requires accuracy >= 70%.

REGION data VALUES are russified centrally (db/zzz_clickhouse_after_init.sql);
this script reads them dynamically from sf_data, so RU values flow through
without any hand-translated literals here.
"""
import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Trailing partial month that must be excluded from the dashboard.
PARTIAL_MONTH = "2026-03"

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Teamly 'Sales Performance Dashboard' page exists",
    "Teamly page has both 'Monthly Revenue' and 'Regional Performance' sections",
    "Teamly: all regions appear with exact revenue",
    "Teamly partial month 2026-03 excluded",
    "Excel: all regions appear with exact revenue and order count",
    "Excel: representative complete months appear with exact Order Count and Revenue",
    "Excel: monthly entry count equals number of complete months",
    "Excel partial month 2026-03 excluded",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = f": {str(detail)[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def number_in_text(value, text):
    text = str(text)
    val_str = str(value)
    if val_str in text:
        return True
    try:
        f2 = f"{float(value):.2f}"
        if f2 in text:
            return True
    except (ValueError, TypeError):
        pass
    return False


def load_expected():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT TO_CHAR("ORDER_DATE"::timestamp, 'YYYY-MM') as month,
               COUNT(*) as orders,
               ROUND(SUM("TOTAL_AMOUNT")::numeric, 2) as revenue
        FROM sf_data."SALES_DW__PUBLIC__ORDERS"
        GROUP BY TO_CHAR("ORDER_DATE"::timestamp, 'YYYY-MM')
        ORDER BY month
    """)
    # Complete months only: the trailing partial month has far fewer orders.
    monthly = [(m, int(o), float(r)) for m, o, r in cur.fetchall() if int(o) >= 500]

    cur.execute("""
        SELECT c."REGION",
               COUNT(o."ORDER_ID"),
               ROUND(SUM(o."TOTAL_AMOUNT")::numeric, 2)
        FROM sf_data."SALES_DW__PUBLIC__ORDERS" o
        JOIN sf_data."SALES_DW__PUBLIC__CUSTOMERS" c ON o."CUSTOMER_ID" = c."CUSTOMER_ID"
        GROUP BY c."REGION"
        ORDER BY c."REGION"
    """)
    regions = [(r, int(o), float(rev)) for r, o, rev in cur.fetchall()]

    cur.close()
    conn.close()
    return monthly, regions


def representative_months(expected_monthly):
    """First, middle and last complete month (de-duplicated, order preserved)."""
    if not expected_monthly:
        return []
    idxs = sorted({0, len(expected_monthly) // 2, len(expected_monthly) - 1})
    return [expected_monthly[i] for i in idxs]


def check_teamly(expected_monthly, expected_regions):
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        for n in ("Teamly 'Sales Performance Dashboard' page exists",
                  "Teamly page has both 'Monthly Revenue' and 'Regional Performance' sections",
                  "Teamly: all regions appear with exact revenue",
                  "Teamly partial month 2026-03 excluded"):
            check(n, False, str(e))
        return

    # Find the dashboard page (English title marker preserved per task.md).
    dashboard = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "sales performance dashboard" in tl or ("sales" in tl and "dashboard" in tl) \
                or "панель продаж" in tl:
            dashboard = (pid, title, body)
            break
    check("Teamly 'Sales Performance Dashboard' page exists", dashboard is not None,
          f"pages: {[(p[0], p[1]) for p in pages]}")

    if dashboard is None:
        for n in ("Teamly page has both 'Monthly Revenue' and 'Regional Performance' sections",
                  "Teamly: all regions appear with exact revenue",
                  "Teamly partial month 2026-03 excluded"):
            check(n, False, "no dashboard page")
        return

    text = ((dashboard[1] or "") + " " + (dashboard[2] or ""))
    text_l = text.lower()

    check("Teamly page has both 'Monthly Revenue' and 'Regional Performance' sections",
          "monthly revenue" in text_l and "regional performance" in text_l,
          "section heading(s) missing")

    # CRITICAL: every region appears with its exact revenue.
    all_regions_ok = True
    for region, orders, revenue in expected_regions:
        ok = region.lower() in text_l and number_in_text(revenue, text)
        if not ok:
            all_regions_ok = False
        check(f"Teamly Regional: '{region}' with revenue {revenue}", ok,
              "region+revenue not found together")
    check("Teamly: all regions appear with exact revenue", all_regions_ok)

    # Representative months present with Order Count and Revenue (non-critical
    # individually; the Excel sheet carries the exact-value critical gate).
    for month, orders, revenue in representative_months(expected_monthly):
        check(f"Teamly Monthly contains '{month}' with revenue {revenue}",
              month in text and number_in_text(revenue, text),
              "month+revenue not found")

    # CRITICAL: trailing partial month excluded.
    check("Teamly partial month 2026-03 excluded", PARTIAL_MONTH not in text,
          f"'{PARTIAL_MONTH}' present in dashboard page")


def check_excel(agent_workspace, expected_monthly, expected_regions):
    print("\n=== Checking Excel File ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(agent_workspace, "Sales_Dashboard_Backup.xlsx")
    check("Excel file exists", os.path.isfile(xlsx_path), f"Expected {xlsx_path}")
    if not os.path.isfile(xlsx_path):
        for n in ("Excel: all regions appear with exact revenue and order count",
                  "Excel: representative complete months appear with exact Order Count and Revenue",
                  "Excel: monthly entry count equals number of complete months",
                  "Excel partial month 2026-03 excluded"):
            check(n, False, "no excel file")
        return

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        for n in ("Excel: all regions appear with exact revenue and order count",
                  "Excel: representative complete months appear with exact Order Count and Revenue",
                  "Excel: monthly entry count equals number of complete months",
                  "Excel partial month 2026-03 excluded"):
            check(n, False, "unreadable")
        return

    def find_sheet(keywords):
        for s in wb.sheetnames:
            sl = s.lower()
            if all(k in sl for k in keywords):
                return wb[s]
        return None

    def sheet_text(ws):
        txt = ""
        for row in ws.iter_rows(values_only=True):
            txt += " ".join(str(c) for c in row if c is not None) + " "
        return txt

    def header_row(ws):
        return [str(c).strip().lower() for c in next(ws.iter_rows(values_only=True)) if c is not None]

    # ── Monthly Revenue sheet ────────────────────────────────────────────────
    ws_m = find_sheet(["monthly"]) or find_sheet(["revenue"])
    check("Monthly Revenue sheet exists", ws_m is not None, f"Sheets: {wb.sheetnames}")

    if ws_m is not None:
        hdr = header_row(ws_m)
        check("Monthly sheet has Month/Order Count/Revenue headers",
              "month" in hdr and "order count" in hdr and "revenue" in hdr,
              f"headers: {hdr}")

        rows = [r for r in ws_m.iter_rows(min_row=2, values_only=True)
                if any(c is not None for c in r)]
        # Map month -> (orders, revenue) using header positions.
        m_i = hdr.index("month") if "month" in hdr else 0
        o_i = hdr.index("order count") if "order count" in hdr else 1
        r_i = hdr.index("revenue") if "revenue" in hdr else 2
        found_months = {}
        for r in rows:
            if m_i < len(r) and r[m_i] is not None:
                found_months[str(r[m_i]).strip()] = (
                    r[o_i] if o_i < len(r) else None,
                    r[r_i] if r_i < len(r) else None,
                )

        # CRITICAL: representative complete months with exact Order Count + Revenue.
        months_ok = True
        for month, orders, revenue in representative_months(expected_monthly):
            cell = found_months.get(month)
            ok = (cell is not None
                  and number_in_text(orders, str(cell[0]))
                  and number_in_text(revenue, str(cell[1])))
            if not ok:
                months_ok = False
            check(f"Excel Monthly row {month}: {orders} orders / {revenue}", ok,
                  f"got {cell}")
        check("Excel: representative complete months appear with exact Order Count and Revenue",
              months_ok)

        # CRITICAL: exact entry count.
        check("Excel: monthly entry count equals number of complete months",
              len(rows) == len(expected_monthly),
              f"got {len(rows)} rows, expected {len(expected_monthly)}")

        # CRITICAL: partial month excluded.
        check("Excel partial month 2026-03 excluded",
              PARTIAL_MONTH not in found_months,
              "partial month present")
    else:
        for n in ("Excel: representative complete months appear with exact Order Count and Revenue",
                  "Excel: monthly entry count equals number of complete months",
                  "Excel partial month 2026-03 excluded"):
            check(n, False, "no monthly sheet")

    # ── Regional Performance sheet ───────────────────────────────────────────
    ws_r = find_sheet(["regional"]) or find_sheet(["region"])
    check("Regional Performance sheet exists", ws_r is not None, f"Sheets: {wb.sheetnames}")

    if ws_r is not None:
        hdr = header_row(ws_r)
        check("Regional sheet has Region/Order Count/Revenue headers",
              "region" in hdr and "order count" in hdr and "revenue" in hdr,
              f"headers: {hdr}")
        txt = sheet_text(ws_r)
        all_regions_ok = True
        for region, orders, revenue in expected_regions:
            ok = (region.lower() in txt.lower()
                  and number_in_text(revenue, txt)
                  and number_in_text(orders, txt))
            if not ok:
                all_regions_ok = False
            check(f"Excel Regional '{region}': {orders} orders / {revenue}", ok)
        check("Excel: all regions appear with exact revenue and order count", all_regions_ok)
    else:
        check("Excel: all regions appear with exact revenue and order count", False,
              "no regional sheet")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected_monthly, expected_regions = load_expected()

    check_teamly(expected_monthly, expected_regions)
    check_excel(args.agent_workspace, expected_monthly, expected_regions)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    print(f"  Failed: {FAIL_COUNT}")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    # CRITICAL gate first.
    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    success = accuracy >= 70
    print(f"  Overall: {'PASS' if success else 'FAIL'}")
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
