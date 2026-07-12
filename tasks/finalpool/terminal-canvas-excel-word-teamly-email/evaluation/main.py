"""Evaluation for terminal-canvas-excel-word-teamly-email.

The agent must, for course 16 (Основы финансов Fall 2013) and course 17
(Основы финансов Fall 2014):
  1. Build Student_Risk_Analysis.xlsx (4 sheets: Course_Overview,
     Risk_Distribution, At_Risk_Students, Intervention_Plan).
  2. Write Intervention_Plan.docx.
  3. Create a Teamly page "Student Risk Tracker" with one table row per course
     (Student Course / Risk Level / Student Count / Average Score / Pass Rate).
  4. Email academic_advisors@university.edu (subject "Student Retention Risk
     Analysis - Action Required") with the total number of High-risk students.
  5. Write & run risk_scorer.py.

All expected quantitative values are recomputed LIVE from canvas.* (course ids
16 and 17) — nothing is hardcoded — so the eval stays honest if the seed
changes. Classification follows scoring_model.json: arithmetic mean of all
graded submissions per student per course, excluding null scores; cutoffs
<60 High, 60-74.99 Medium, >=75 Low.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

COURSE_IDS = (16, 17)

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "Course_Overview avg_score correct for both courses",
    "Risk_Distribution combined High/Medium/Low counts correct",
    "At_Risk_Students per-course high/medium/low counts correct",
    "Teamly 'Student Risk Tracker' page exists with correct per-course Average Score & Pass Rate",
    "Email contains correct total High-risk student count",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def body_has_int(text, n):
    """True if integer n appears as a standalone token in text."""
    pat = r"(?<!\d)" + re.escape(str(int(n))) + r"(?!\d)"
    return re.search(pat, text) is not None


def get_expected():
    """Recompute per-course quantitative ground truth live from canvas.*.

    Returns dict keyed by course_id with:
      name, enrollment_count, avg_score (mean of all graded submissions),
      students (# with >=1 graded submission), high/medium/low counts,
      pass_rate (% of those students with avg >= 60).
    Also overall high/medium/low totals.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    exp = {}
    for cid in COURSE_IDS:
        cur.execute("SELECT name FROM canvas.courses WHERE id=%s", (cid,))
        row = cur.fetchone()
        name = row[0] if row else f"course {cid}"

        # Total enrolled students.
        cur.execute("""
            SELECT COUNT(*) FROM canvas.enrollments
            WHERE course_id=%s AND type ILIKE '%%student%%'
        """, (cid,))
        enrollment_count = cur.fetchone()[0]

        # Mean of ALL graded submissions (course avg_score).
        cur.execute("""
            SELECT AVG(s.score)
            FROM canvas.submissions s JOIN canvas.assignments a ON a.id=s.assignment_id
            WHERE a.course_id=%s AND s.score IS NOT NULL
        """, (cid,))
        avg_score = float(cur.fetchone()[0])

        # Per-student averages -> risk classification.
        cur.execute("""
            WITH stu AS (
              SELECT s.user_id, AVG(s.score) AS avg_score
              FROM canvas.submissions s JOIN canvas.assignments a ON a.id=s.assignment_id
              WHERE a.course_id=%s AND s.score IS NOT NULL
              GROUP BY s.user_id
            )
            SELECT
              COUNT(*),
              SUM(CASE WHEN avg_score < 60 THEN 1 ELSE 0 END),
              SUM(CASE WHEN avg_score >= 60 AND avg_score < 75 THEN 1 ELSE 0 END),
              SUM(CASE WHEN avg_score >= 75 THEN 1 ELSE 0 END)
            FROM stu
        """, (cid,))
        students, high, medium, low = cur.fetchone()
        students = int(students or 0)
        high, medium, low = int(high or 0), int(medium or 0), int(low or 0)
        pass_rate = round(100.0 * (medium + low) / students, 2) if students else 0.0

        exp[cid] = {
            "name": name, "enrollment_count": enrollment_count,
            "avg_score": avg_score, "students": students,
            "high": high, "medium": medium, "low": low,
            "pass_rate": pass_rate,
        }
    conn.close()

    exp["overall"] = {
        "high": sum(exp[c]["high"] for c in COURSE_IDS),
        "medium": sum(exp[c]["medium"] for c in COURSE_IDS),
        "low": sum(exp[c]["low"] for c in COURSE_IDS),
        "students": sum(exp[c]["students"] for c in COURSE_IDS),
    }
    return exp


def check_excel(workspace, exp):
    print("\n=== Check 1: Student_Risk_Analysis.xlsx ===")
    path = os.path.join(workspace, "Student_Risk_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")
    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # ---- Course_Overview ----
    co_idx = next((i for i, s in enumerate(sheets_lower) if "course" in s and "overview" in s), 0)
    ws1 = wb[sheets[co_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c is not None and str(c) != "" for c in r)]
    check("Course_Overview has 2 rows", len(data1) >= 2, f"Found {len(data1)}")

    all_text1 = " ".join(str(c) for r in rows1 for c in r if c).lower()
    check("Contains Основы финансов", "finance" in all_text1, f"Text: {all_text1[:120]}")

    headers1 = [str(c).lower() if c else "" for c in (rows1[0] if rows1 else [])]
    avg_idx = next((i for i, h in enumerate(headers1) if "avg" in h and "score" in h), -1)
    cid_idx = next((i for i, h in enumerate(headers1) if h.strip() == "course_id"), -1)

    # CRITICAL: avg_score per course matches live mean (tol 1.0).
    avg_ok = False
    if avg_idx >= 0 and len(data1) >= 2:
        expected_means = sorted(exp[c]["avg_score"] for c in COURSE_IDS)
        got = []
        for r in data1:
            try:
                got.append(float(r[avg_idx]))
            except (TypeError, ValueError):
                pass
        got_sorted = sorted(got)
        avg_ok = (len(got_sorted) >= 2 and
                  num_close(got_sorted[0], expected_means[0], 1.0) and
                  num_close(got_sorted[-1], expected_means[-1], 1.0))
        detail = f"expected~{[round(m,2) for m in expected_means]}, got {[round(g,2) for g in got_sorted]}"
    else:
        detail = "avg_score column not found"
    check("Course_Overview avg_score correct for both courses", avg_ok, detail)

    # ---- Risk_Distribution ----
    rd_idx = next((i for i, s in enumerate(sheets_lower) if "risk" in s and "dist" in s), None)
    if rd_idx is not None:
        ws2 = wb[sheets[rd_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c is not None and str(c) != "" for c in r)]
        check("Risk_Distribution has 3 rows", len(data2) >= 3, f"Found {len(data2)}")
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Risk_Distribution has High/Medium/Low",
              "high" in all_text2 and "medium" in all_text2 and "low" in all_text2,
              all_text2[:120])

        # CRITICAL: combined per-level student_count correct.
        headers2 = [str(c).lower() if c else "" for c in (rows2[0] if rows2 else [])]
        lvl_idx = next((i for i, h in enumerate(headers2) if "risk" in h and "level" in h), 0)
        cnt_idx = next((i for i, h in enumerate(headers2) if ("student" in h and "count" in h) or h.strip() == "count"), -1)
        got_counts = {}
        if cnt_idx >= 0:
            for r in data2:
                lvl = str(r[lvl_idx]).strip().lower() if lvl_idx < len(r) and r[lvl_idx] else ""
                try:
                    got_counts[lvl] = int(float(r[cnt_idx]))
                except (TypeError, ValueError):
                    pass
        ov = exp["overall"]
        dist_ok = (got_counts.get("high") == ov["high"] and
                   got_counts.get("medium") == ov["medium"] and
                   got_counts.get("low") == ov["low"])
        check("Risk_Distribution combined High/Medium/Low counts correct", dist_ok,
              f"expected H={ov['high']} M={ov['medium']} L={ov['low']}, got {got_counts}")
    else:
        check("Risk_Distribution has 3 rows", False, "Risk_Distribution sheet not found")
        check("Risk_Distribution combined High/Medium/Low counts correct", False, "sheet missing")

    # ---- At_Risk_Students ----
    ar_idx = next((i for i, s in enumerate(sheets_lower)
                   if "at_risk" in s or "risk_student" in s or ("risk" in s and "student" in s)), None)
    if ar_idx is not None:
        ws3 = wb[sheets[ar_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c is not None and str(c) != "" for c in r)]
        check("At_Risk_Students has 2 course rows", len(data3) >= 2, f"Found {len(data3)}")

        headers3 = [str(c).lower() if c else "" for c in (rows3[0] if rows3 else [])]
        h_idx = next((i for i, h in enumerate(headers3) if "high" in h), -1)
        m_idx = next((i for i, h in enumerate(headers3) if "medium" in h), -1)
        l_idx = next((i for i, h in enumerate(headers3) if h.startswith("low") or ("low" in h and "count" in h)), -1)
        # Match each data row to a course by name, compare counts.
        matched = 0
        ok_courses = 0
        for cid in COURSE_IDS:
            cname_short = exp[cid]["name"].split("(")[0].strip().lower()
            target = None
            for r in data3:
                rtext = " ".join(str(c) for c in r if c).lower()
                # disambiguate Fall 2013 vs 2014 by year token
                year = "2013" if cid == 16 else "2014"
                if "finance" in rtext and year in rtext:
                    target = r
                    break
            if target is None:
                continue
            matched += 1
            try:
                gh = int(float(target[h_idx])) if h_idx >= 0 else None
                gm = int(float(target[m_idx])) if m_idx >= 0 else None
                gl = int(float(target[l_idx])) if l_idx >= 0 else None
            except (TypeError, ValueError):
                gh = gm = gl = None
            if (gh == exp[cid]["high"] and gm == exp[cid]["medium"] and gl == exp[cid]["low"]):
                ok_courses += 1
        check("At_Risk_Students per-course high/medium/low counts correct",
              ok_courses == 2,
              f"matched {matched}/2 rows, {ok_courses}/2 correct; "
              f"expected 16:H{exp[16]['high']}/M{exp[16]['medium']}/L{exp[16]['low']} "
              f"17:H{exp[17]['high']}/M{exp[17]['medium']}/L{exp[17]['low']}")
    else:
        check("At_Risk_Students has 2 course rows", False, "sheet not found")
        check("At_Risk_Students per-course high/medium/low counts correct", False, "sheet missing")

    # ---- Intervention_Plan sheet ----
    ip_idx = next((i for i, s in enumerate(sheets_lower) if "intervention" in s), None)
    if ip_idx is not None:
        ws4 = wb[sheets[ip_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(c is not None and str(c) != "" for c in r)]
        check("Intervention_Plan has 3 rows", len(data4) >= 3, f"Found {len(data4)}")
        all_text4 = " ".join(str(c) for r in rows4 for c in r if c).lower()
        check("Intervention mentions advising", "advis" in all_text4, all_text4[:150])
        check("Intervention mentions tutoring", "tutor" in all_text4, all_text4[:150])
    else:
        check("Intervention_Plan has 3 rows", False, "sheet not found")


def check_word(workspace, exp):
    print("\n=== Check 2: Intervention_Plan.docx ===")
    path = os.path.join(workspace, "Intervention_Plan.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs)
    low = full_text.lower()
    check("Document mentions retention or intervention",
          "retention" in low or "intervention" in low, low[:120])
    check("Document mentions high risk",
          "high risk" in low or "high-risk" in low or ("high" in low and "риск" in low),
          low[:120])
    check("Document mentions both courses", "2013" in low and "2014" in low, low[:150])
    check("Document has substantial content", len(full_text) > 200, f"Length: {len(full_text)}")


def check_teamly(exp):
    print("\n=== Check 3: Teamly 'Student Risk Tracker' page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE '%%student risk tracker%%'
               OR title ILIKE '%%risk tracker%%'
        """)
        pages = cur.fetchall()
        if not pages:
            cur.execute("SELECT COUNT(*) FROM teamly.pages")
            total = cur.fetchone()[0]
            check("Student Risk Tracker page exists in Teamly", False,
                  f"Found {total} pages, none matching 'Student Risk Tracker'")
            check("Teamly 'Student Risk Tracker' page exists with correct per-course Average Score & Pass Rate",
                  False, "page missing")
            return
        check("Student Risk Tracker page exists in Teamly", True)

        body = "\n".join(str(b) for _, _, b in pages)
        body_lower = body.lower()
        check("Tracker page has non-trivial body", len(body) >= 100, f"len={len(body)}")
        check("Tracker mentions Average Score & Pass Rate column headers",
              "average score" in body_lower and "pass rate" in body_lower, body_lower[:200])

        # CRITICAL: per-course Average Score (tol 1.0) AND Pass Rate present.
        # Pass Rate may be rendered with or without rounding/% -> accept the
        # rounded integer or one-decimal forms as standalone tokens.
        ok_courses = 0
        details = []
        for cid in COURSE_IDS:
            year = "2013" if cid == 16 else "2014"
            avg = exp[cid]["avg_score"]
            pr = exp[cid]["pass_rate"]
            # find avg as a number close to expected anywhere in body
            avg_ok = any(num_close(m, avg, 1.0) for m in _all_numbers(body))
            # pass rate variants
            pr_variants = {str(round(pr)), f"{pr:.1f}", f"{pr:.2f}",
                           str(int(pr)), f"{round(pr,1)}"}
            pr_ok = any(re.search(r"(?<!\d)" + re.escape(v) + r"(?!\d)", body)
                        for v in pr_variants)
            if avg_ok and pr_ok:
                ok_courses += 1
            else:
                details.append(f"{cid}: avg~{round(avg,2)}({avg_ok}) pr~{pr}({pr_ok})")
        check("Teamly 'Student Risk Tracker' page exists with correct per-course Average Score & Pass Rate",
              ok_courses == 2, f"{ok_courses}/2 ok; {details}")
    except Exception as e:
        check("Student Risk Tracker page exists in Teamly", False, str(e))
        check("Teamly 'Student Risk Tracker' page exists with correct per-course Average Score & Pass Rate",
              False, str(e))
    finally:
        cur.close()
        conn.close()


def _all_numbers(text):
    """All decimal numbers in text as floats."""
    out = []
    for m in re.findall(r"\d+(?:\.\d+)?", text):
        try:
            out.append(float(m))
        except ValueError:
            pass
    return out


def check_email(exp):
    print("\n=== Check 4: Email to Academic Advisors ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%academic_advisors%%'
               OR subject ILIKE '%%retention%%risk%%'
               OR subject ILIKE '%%risk%%analysis%%'
        """)
        emails = cur.fetchall()
        check("Email sent about risk analysis", len(emails) >= 1, "No matching email found")

        target = None
        for e in emails:
            subj = (e[1] or "").lower()
            to = str(e[2]).lower()
            if "academic_advisors@university.edu" in to:
                target = e
                break
        if target is None and emails:
            target = emails[0]

        if target:
            subj = str(target[1] or "")
            to = str(target[2] or "")
            check("Email recipient and subject correct",
                  "academic_advisors@university.edu" in to.lower()
                  and "student retention risk analysis - action required" in subj.lower(),
                  f"subject={subj!r}, to={to}")

            body = str(target[3] or "")
            # CRITICAL: total High-risk students across both courses appears.
            total_high = exp["overall"]["high"]
            check("Email contains correct total High-risk student count",
                  body_has_int(body, total_high),
                  f"expected total High={total_high} in body; body[:200]={body[:200]!r}")
        else:
            check("Email recipient and subject correct", False, "no email")
            check("Email contains correct total High-risk student count", False, "no email")
    except Exception as e:
        check("Email check", False, str(e))
        check("Email contains correct total High-risk student count", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: risk_scorer.py ===")
    path = os.path.join(workspace, "risk_scorer.py")
    check("risk_scorer.py exists", os.path.exists(path))


def check_reverse_validation():
    print("\n=== Reverse Validation ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT to_addr FROM email.messages
            WHERE subject ILIKE '%%risk%%' OR subject ILIKE '%%retention%%'
        """)
        emails = cur.fetchall()
        noise_recipients = ["all-staff@university.edu", "it@university.edu",
                            "facilities@university.edu"]
        for email_row in emails:
            to_str = str(email_row[0]).lower()
            for noise in noise_recipients:
                if noise in to_str:
                    check("No risk emails sent to wrong recipients", False,
                          f"Sent to noise recipient: {noise}")
                    cur.close(); conn.close()
                    return
        check("No risk emails sent to wrong recipients", True)
        cur.close(); conn.close()
    except Exception as e:
        check("Reverse validation", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    exp = get_expected()
    print(f"[eval] Live ground truth: "
          f"16(avg={exp[16]['avg_score']:.2f},H{exp[16]['high']}/M{exp[16]['medium']}/L{exp[16]['low']},pr={exp[16]['pass_rate']}) "
          f"17(avg={exp[17]['avg_score']:.2f},H{exp[17]['high']}/M{exp[17]['medium']}/L{exp[17]['low']},pr={exp[17]['pass_rate']}) "
          f"overall H={exp['overall']['high']}")

    check_excel(args.agent_workspace, exp)
    check_word(args.agent_workspace, exp)
    check_teamly(exp)
    check_email(exp)
    check_script(args.agent_workspace)
    check_reverse_validation()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and (accuracy >= 70)
    result = {"total_passed": PASS_COUNT, "total_checks": total,
              "accuracy": accuracy, "critical_failed": critical_failed,
              "success": success}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
