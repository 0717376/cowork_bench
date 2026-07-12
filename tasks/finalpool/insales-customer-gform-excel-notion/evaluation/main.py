"""Evaluation для insales-customer-gform-excel-notion (RU-стек: insales/forms/teamly).

Деливераблы:
  - terminal/Python: customer_segmenter.py + customer_segments.json.
  - Excel Customer_Insights_Report.xlsx: листы Customer_Segments / Top_Customers / Segment_Strategy.
  - Forms (gform.*): опрос «Customer Experience Survey» с >=4 вопросами нужных типов.
  - Teamly (teamly.pages): страница «Customer Intelligence Hub» с дашбордом сегментации.

Реальные данные о клиентах агент читает живьём из wc.* (InSales, русифицированы централизованно),
поэтому НЕ хардкодим конкретные значения realia — критические чеки ПЕРЕСЧИТЫВАЮТ уровни
из wc.customers, а сравнение ячеек Excel — только структурно/по толерантности к эталону.

Критические чеки (CRITICAL_CHECKS): любой их fail => задача FAIL независимо от accuracy.
Порог: accuracy >= 70% И нет проваленных критических чеков => PASS.
"""
import os
import argparse, json, os, sys
import openpyxl


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

# Имена критических чеков (по строке из check())
CRITICAL_CHECKS = {
    "CRITICAL: уровни Customer_Segments соответствуют порогам wc (VIP>500/Regular 100-500/New<100)",
    "CRITICAL: опрос Customer Experience Survey с нужными типами вопросов",
    "CRITICAL: Top_Customers — 10 строк, Total_Spent по убыванию, email из wc",
    "CRITICAL: страница Teamly Customer Intelligence Hub с дашбордом сегментации",
    "CRITICAL: customer_segmenter.py и customer_segments.json созданы",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


# ---------------------------------------------------------------------------
# Excel (структурное сравнение с эталоном + извлечение листов для критчеков)
# ---------------------------------------------------------------------------
def check_excel(agent_workspace, groundtruth_workspace):
    """Возвращает (wb, sheets_dict) или (None, {}) если файла нет."""
    excel_path = os.path.join(agent_workspace, "Customer_Insights_Report.xlsx")
    check("Customer_Insights_Report.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return None

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    gt_path = os.path.join(groundtruth_workspace, "Customer_Insights_Report.xlsx")
    gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

    if gt_wb:
        for sheet_name in gt_wb.sheetnames:
            check(f"{sheet_name} sheet exists", sheet_name in wb.sheetnames)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                gt_ws = gt_wb[sheet_name]
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                for h in gt_headers:
                    if h:
                        check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                data_rows = list(ws.iter_rows(min_row=2, values_only=True))
                min_rows = max(1, len(gt_rows) - 2)
                check(f"{sheet_name} has >= {min_rows} data rows", len(data_rows) >= min_rows, f"got {len(data_rows)}")
                # NB: эталонный Customer_Insights_Report.xlsx содержит СИНТЕТИЧЕСКИЕ
                # placeholder-значения (Customer A / a@example.com / VIP=15 / 12000 и т.п.),
                # НЕ отражающие живые данные wc.customers. Сравнивать ячейки агента с этими
                # placeholder'ами нельзя — корректный агент (читающий realia живьём) их завалит.
                # Поэтому здесь — ТОЛЬКО структура (листы/колонки/число строк), а реальные
                # значения проверяют критические чеки, пересчитывающие уровни из wc.customers.
    return wb


def sheet_rows(wb, name):
    """Список словарей {header_lower: value} по строкам данных листа name, либо []."""
    if wb is None or name not in wb.sheetnames:
        return []
    ws = wb[name]
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row):
            continue
        d = {}
        for h, v in zip(headers, row):
            if h:
                d[h.strip().lower()] = v
        out.append(d)
    return out


# ---------------------------------------------------------------------------
# Критический чек: уровни сегментации пересчитываются из wc.customers
# ---------------------------------------------------------------------------
def check_segments_critical(wb):
    rows = sheet_rows(wb, "Customer_Segments")
    if not rows:
        check("CRITICAL: уровни Customer_Segments соответствуют порогам wc (VIP>500/Regular 100-500/New<100)",
              False, "лист Customer_Segments отсутствует/пуст")
        return

    # Реальные уровни из wc.customers
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT total_spent FROM wc.customers")
        spents = [safe_float(r[0], 0.0) or 0.0 for r in cur.fetchall()]
        conn.close()
    except Exception as e:
        check("CRITICAL: уровни Customer_Segments соответствуют порогам wc (VIP>500/Regular 100-500/New<100)",
              False, f"wc недоступен: {e}")
        return

    exp = {"vip": 0, "regular": 0, "new": 0}
    for s in spents:
        if s > 500:
            exp["vip"] += 1
        elif s >= 100:
            exp["regular"] += 1
        else:
            exp["new"] += 1
    total_customers = len(spents)

    # Парсим строки Excel по сегментам
    by_seg = {}
    for r in rows:
        seg = str(r.get("segment", "")).strip().lower()
        if seg in ("vip", "regular", "new"):
            by_seg[seg] = r

    has_all_tiers = all(t in by_seg for t in ("vip", "regular", "new"))
    check("Customer_Segments содержит уровни VIP/Regular/New",
          has_all_tiers, f"сегменты: {list(by_seg.keys())}")

    # Сумма Customer_Count == всего клиентов (толеранс ±2 на пограничные округления)
    cnt_sum = sum(safe_float(by_seg[t].get("customer_count"), 0) or 0
                  for t in by_seg if t in ("vip", "regular", "new"))
    count_ok = abs(cnt_sum - total_customers) <= 2
    check("Customer_Segments: сумма Customer_Count ≈ числу клиентов wc",
          count_ok, f"сумма={cnt_sum}, клиентов wc={total_customers}")

    # Avg_Spend каждого уровня в своей полосе (с небольшим допуском)
    band_ok = True
    detail = []
    if "vip" in by_seg:
        v = safe_float(by_seg["vip"].get("avg_spend"))
        ok = v is not None and v > 450
        band_ok = band_ok and ok
        detail.append(f"VIP avg_spend={v}")
    else:
        band_ok = False
    if "regular" in by_seg:
        v = safe_float(by_seg["regular"].get("avg_spend"))
        ok = v is not None and 90 <= v <= 520
        band_ok = band_ok and ok
        detail.append(f"Regular avg_spend={v}")
    else:
        band_ok = False
    if "new" in by_seg:
        v = safe_float(by_seg["new"].get("avg_spend"))
        ok = v is not None and v < 120
        band_ok = band_ok and ok
        detail.append(f"New avg_spend={v}")
    else:
        band_ok = False
    check("Customer_Segments: Avg_Spend уровней в своих полосах", band_ok, "; ".join(detail))

    # Сортировка по Total_Revenue по убыванию
    revs = [safe_float(r.get("total_revenue")) for r in rows
            if str(r.get("segment", "")).strip().lower() in ("vip", "regular", "new")]
    revs = [x for x in revs if x is not None]
    sorted_ok = all(revs[i] >= revs[i + 1] - 1e-6 for i in range(len(revs) - 1)) if len(revs) >= 2 else True
    check("Customer_Segments отсортирован по Total_Revenue по убыванию", sorted_ok, f"revs={revs}")

    # CRITICAL: уровни корректны (все три + счётчики совпадают с порогами wc по уровням)
    tier_counts_ok = has_all_tiers
    for t in ("vip", "regular", "new"):
        if t in by_seg:
            c = safe_float(by_seg[t].get("customer_count"), -1)
            # допуск ±2 клиента на пограничные округления total_spent
            if c is None or abs(c - exp[t]) > 2:
                tier_counts_ok = False
    check("CRITICAL: уровни Customer_Segments соответствуют порогам wc (VIP>500/Regular 100-500/New<100)",
          has_all_tiers and band_ok and count_ok and tier_counts_ok,
          f"exp={exp}, excel={[(t, by_seg[t].get('customer_count')) for t in by_seg]}")


# ---------------------------------------------------------------------------
# Критический чек: Top_Customers
# ---------------------------------------------------------------------------
def check_top_customers_critical(wb):
    rows = sheet_rows(wb, "Top_Customers")
    if not rows:
        check("CRITICAL: Top_Customers — 10 строк, Total_Spent по убыванию, email из wc",
              False, "лист Top_Customers отсутствует/пуст")
        return

    has_10 = len(rows) >= 10
    check("Top_Customers: >= 10 строк данных", has_10, f"got {len(rows)}")

    spents = [safe_float(r.get("total_spent")) for r in rows[:10]]
    spents = [x for x in spents if x is not None]
    mono = all(spents[i] >= spents[i + 1] - 1e-6 for i in range(len(spents) - 1)) if len(spents) >= 2 else False
    check("Top_Customers: Total_Spent монотонно не возрастает", mono, f"spents={spents}")

    # email из реальных клиентов wc
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT LOWER(email) FROM wc.customers")
        wc_emails = {r[0] for r in cur.fetchall() if r[0]}
        conn.close()
    except Exception as e:
        wc_emails = set()
        check("Top_Customers: доступ к wc.customers", False, str(e))

    excel_emails = [str(r.get("email", "")).strip().lower() for r in rows[:10] if r.get("email")]
    matched = sum(1 for e in excel_emails if e in wc_emails)
    emails_ok = bool(wc_emails) and matched >= 8
    check("Top_Customers: email совпадают с реальными клиентами wc", emails_ok,
          f"совпало {matched}/{len(excel_emails)}")

    check("CRITICAL: Top_Customers — 10 строк, Total_Spent по убыванию, email из wc",
          has_10 and mono and emails_ok,
          f"rows={len(rows)}, mono={mono}, emails matched={matched}")


# ---------------------------------------------------------------------------
# Критический чек: Forms (gform) опрос
# ---------------------------------------------------------------------------
def check_forms_critical():
    try:
        conn = get_conn()
        cur = conn.cursor()
    except Exception as e:
        check("CRITICAL: опрос Customer Experience Survey с нужными типами вопросов", False, str(e))
        return

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    check("Forms: создана хотя бы одна форма", len(forms) >= 1, f"найдено {len(forms)}")

    def title_matches(t):
        tl = (t or "").lower()
        en = "customer experience survey" in tl
        ru = ("опрос" in tl or "удовлетвор" in tl or "клиент" in tl) and "канцеляр" not in tl
        return en or ru

    target = None
    for fid, title in forms:
        if title_matches(title):
            target = fid
            break

    q_count = 0
    types = []
    titles = []
    if target is not None:
        cur.execute("SELECT title, question_type FROM gform.questions WHERE form_id = %s", (target,))
        qs = cur.fetchall()
        q_count = len(qs)
        types = [(qt or "").upper() for _t, qt in qs]
        titles = [(_t or "").lower() for _t, qt in qs]
    conn.close()

    check("Forms: найден опрос Customer Experience Survey",
          target is not None, f"заголовки: {[t for _f, t in forms]}")
    check("Forms: опрос содержит >= 4 вопросов", q_count >= 4, f"вопросов={q_count}")

    # Типы: ожидаем как минимум один radio/choice и один текстовый вопрос
    radio_like = {"RADIO", "MULTIPLE_CHOICE", "CHOICE", "SCALE", "LINEAR_SCALE", "CHOICEQUESTION"}
    text_like = {"TEXT", "PARAGRAPH", "SHORT_ANSWER", "LONG_ANSWER", "TEXTQUESTION"}
    has_radio = any(t in radio_like for t in types)
    has_text = any(t in text_like for t in types)
    check("Forms: есть вопрос-шкала/выбор", has_radio, f"types={types}")
    check("Forms: есть текстовый вопрос (предложения)", has_text, f"types={types}")

    check("CRITICAL: опрос Customer Experience Survey с нужными типами вопросов",
          target is not None and q_count >= 4 and has_radio and has_text,
          f"target={target}, q_count={q_count}, types={types}")


# ---------------------------------------------------------------------------
# Критический чек: Teamly страница
# ---------------------------------------------------------------------------
def check_teamly_critical():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("CRITICAL: страница Teamly Customer Intelligence Hub с дашбордом сегментации", False, str(e))
        return

    check("Teamly: создана хотя бы одна пользовательская страница", len(pages) >= 1,
          f"найдено {len(pages)} (id>3)")

    def title_matches(t):
        tl = (t or "").lower()
        en = "customer intelligence hub" in tl
        ru = ("клиент" in tl and ("интеллект" in tl or "инсайт" in tl or "хаб" in tl or "база" in tl))
        return en or ru

    candidates = [(t, b) for t, b in pages if title_matches(t)]
    check("Teamly: страница Customer Intelligence Hub найдена", len(candidates) >= 1,
          f"заголовки: {[t for t, _b in pages]}")

    body = "\n\n".join(b for _t, b in candidates)
    bl = body.lower()
    has_heading = ("customer segmentation dashboard" in bl) or ("сегментац" in bl and ("панель" in bl or "дашборд" in bl))
    has_segments = all(seg in body.upper() for seg in ("VIP", "REGULAR", "NEW")) or \
                   all(seg in bl for seg in ("vip", "regular", "new"))
    check("Teamly: заголовок Customer Segmentation Dashboard", has_heading, f"body[:120]={body[:120]}")
    check("Teamly: упомянуты все три сегмента", has_segments, f"body[:200]={body[:200]}")

    check("CRITICAL: страница Teamly Customer Intelligence Hub с дашбордом сегментации",
          len(candidates) >= 1 and has_heading and has_segments,
          f"candidates={len(candidates)}, heading={has_heading}, segments={has_segments}")


# ---------------------------------------------------------------------------
# Критический чек: артефакты Python (terminal)
# ---------------------------------------------------------------------------
def check_python_critical(agent_workspace):
    try:
        files = os.listdir(agent_workspace)
    except Exception:
        files = []
    has_script = "customer_segmenter.py" in files or any(f.endswith(".py") for f in files)
    has_segments_json = "customer_segments.json" in files
    check("Python: customer_segmenter.py существует",
          "customer_segmenter.py" in files or any(f.endswith(".py") for f in files),
          f"py files: {[f for f in files if f.endswith('.py')]}")
    check("Python: customer_segments.json создан", has_segments_json, f"files: {files[:20]}")
    check("CRITICAL: customer_segmenter.py и customer_segments.json созданы",
          has_script and has_segments_json,
          f"script={has_script}, segments_json={has_segments_json}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    print("\n=== Excel Customer_Insights_Report.xlsx ===")
    wb = check_excel(agent_workspace, groundtruth_workspace)

    print("\n=== Критические чеки сегментации ===")
    check_segments_critical(wb)
    check_top_customers_critical(wb)

    print("\n=== Python (terminal) ===")
    check_python_critical(agent_workspace)

    print("\n=== Forms (опрос) ===")
    check_forms_critical()

    print("\n=== Teamly (страница) ===")
    check_teamly_critical()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nPassed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)")

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAIL: {CRITICAL_FAILED}")
        return False, f"CRITICAL FAIL: {CRITICAL_FAILED}; {PASS_COUNT}/{total}"

    overall = accuracy >= 70.0
    return overall, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
