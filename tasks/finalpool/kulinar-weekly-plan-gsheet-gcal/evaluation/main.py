"""
Evaluation for kulinar-weekly-plan-gsheet-gcal task (russified -> kulinar).
Checks: GSheet weekly meal plan, Excel file with 2 sheets, GCal events, email.

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Otherwise pass threshold: accuracy >= 70%.

Notes:
- Meal_Type may be Russian (Завтрак/Обед/Ужин) or English (Breakfast/Lunch/
  Dinner); both are accepted.
- Recipe names / categories are chosen by the agent from kulinar's 50 RU
  recipes (non-deterministic). We therefore do NOT pin specific dishes; we only
  verify that the chosen recipes come from the kulinar database and span enough
  distinct kulinar categories for variety. The old hardcoded ground-truth XLSX
  row-value comparison (Chinese-derived recipes) was removed.
"""
import argparse
import os
import sys

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# kulinar's 50 RU recipes: name -> category. Used for membership/category-span
# checks. (Source: local_servers/kulinar-mcp/src/data/all_recipes.json)
KULINAR_RECIPES = {
    "Салат Оливье": "салат",
    "Винегрет": "салат",
    "Сельдь под шубой": "салат",
    "Салат Мимоза": "салат",
    "Крабовый салат": "салат",
    "Греческий салат": "салат",
    "Салат с курицей и грибами": "салат",
    "Холодец": "закуска",
    "Икра кабачковая": "закуска",
    "Грибы маринованные": "закуска",
    "Сало солёное": "закуска",
    "Селёдка с луком": "закуска",
    "Борщ": "суп",
    "Щи из квашеной капусты": "суп",
    "Солянка мясная": "суп",
    "Уха": "суп",
    "Окрошка": "суп",
    "Грибной суп": "суп",
    "Рассольник": "суп",
    "Куриный бульон с лапшой": "суп",
    "Бефстроганов": "горячее",
    "Пельмени домашние": "горячее",
    "Голубцы": "горячее",
    "Котлеты домашние": "горячее",
    "Жаркое в горшочках": "горячее",
    "Курица в сметане": "горячее",
    "Рыба запечённая по-русски": "горячее",
    "Цыплёнок табака": "горячее",
    "Гречка с тушёнкой": "горячее",
    "Плов узбекский": "горячее",
    "Картофельное пюре": "гарнир",
    "Гречневая каша": "гарнир",
    "Перловая каша": "гарнир",
    "Картофель отварной с укропом": "гарнир",
    "Рис отварной": "гарнир",
    "Пирожки с капустой жареные": "выпечка",
    "Пирожки с мясом печёные": "выпечка",
    "Блины тонкие": "выпечка",
    "Кулебяка с капустой и яйцом": "выпечка",
    "Расстегаи с рыбой": "выпечка",
    "Медовик": "десерт",
    "Наполеон": "десерт",
    "Сырники": "десерт",
    "Пасха творожная": "десерт",
    "Ватрушки с творогом": "десерт",
    "Кисель ягодный": "десерт",
    "Морс клюквенный": "напиток",
    "Компот из сухофруктов": "напиток",
    "Сбитень": "напиток",
    "Квас домашний": "напиток",
}
KULINAR_NAMES_LOWER = {n.lower(): c for n, c in KULINAR_RECIPES.items()}
KULINAR_CATEGORIES = set(KULINAR_RECIPES.values())

# Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "GSheet 'Weekly Meal Plan March 9-15 2026' exists",
    "GSheet has at least 21 meal data rows",
    "GSheet covers all three meal types (Завтрак/Обед/Ужин)",
    "GSheet recipes come from the kulinar database",
    "GSheet recipes span at least 3 distinct kulinar categories",
    "Excel has 'Daily Plan' sheet",
    "Excel has 'Category Distribution' sheet",
    "GCal has weekly meal prep planning event",
    "GCal has at least 7 daily meal reminder events",
    "Email to family@home.com with meal plan subject",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def _kulinar_hits(all_text):
    """Return (set of matched kulinar names, set of matched categories) found as
    substrings in the lowercased joined cell text."""
    names = set()
    cats = set()
    for name_lower, cat in KULINAR_NAMES_LOWER.items():
        if name_lower in all_text:
            names.add(name_lower)
            cats.add(cat)
    return names, cats


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()

        target_ss = None
        for sid, title in spreadsheets:
            if title and "meal plan" in title.lower() and ("march" in title.lower() or "2026" in title.lower()):
                target_ss = sid
                break

        record("GSheet 'Weekly Meal Plan March 9-15 2026' exists",
               target_ss is not None,
               f"Found sheets: {[t for _, t in spreadsheets]}")

        if target_ss is None:
            conn.close()
            return

        cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id = %s", (target_ss,))
        sheets = cur.fetchall()
        record("GSheet has at least one sheet", len(sheets) > 0)

        if not sheets:
            conn.close()
            return

        sheet_id = sheets[0][0]
        cur.execute("""
            SELECT COUNT(DISTINCT row_index) FROM gsheet.cells
            WHERE spreadsheet_id = %s AND sheet_id = %s AND row_index > 0
        """, (target_ss, sheet_id))
        data_rows = cur.fetchone()[0]
        record("GSheet has at least 21 meal data rows", data_rows >= 21,
               f"Found {data_rows} data rows")

        # Collect all cell text
        cur.execute("""
            SELECT LOWER(value) FROM gsheet.cells
            WHERE spreadsheet_id = %s AND sheet_id = %s
        """, (target_ss, sheet_id))
        cell_values = [row[0] for row in cur.fetchall() if row[0]]
        all_text = " ".join(cell_values)

        # Meal types: accept Russian OR English.
        has_breakfast = "завтрак" in all_text or "breakfast" in all_text
        has_lunch = "обед" in all_text or "lunch" in all_text
        has_dinner = "ужин" in all_text or "dinner" in all_text
        record("GSheet contains breakfast entries (Завтрак/Breakfast)", has_breakfast)
        record("GSheet contains lunch entries (Обед/Lunch)", has_lunch)
        record("GSheet contains dinner entries (Ужин/Dinner)", has_dinner)
        record("GSheet covers all three meal types (Завтрак/Обед/Ужин)",
               has_breakfast and has_lunch and has_dinner)

        # Recipes must come from the kulinar database and span >=3 categories.
        names, cats = _kulinar_hits(all_text)
        record("GSheet recipes come from the kulinar database", len(names) >= 3,
               f"Matched {len(names)} kulinar recipes: {sorted(names)[:10]}")
        record("GSheet recipes span at least 3 distinct kulinar categories",
               len(cats) >= 3, f"Matched categories: {sorted(cats)}")

        conn.close()
    except Exception as e:
        record("GSheet connection", False, str(e))


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel File ===")
    xl_path = os.path.join(agent_workspace, "Meal_Plan_Summary.xlsx")
    if not os.path.isfile(xl_path):
        record("Excel file Meal_Plan_Summary.xlsx exists", False, f"Not found at: {xl_path}")
        return
    record("Excel file Meal_Plan_Summary.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xl_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names = [s.lower() for s in wb.sheetnames]
    has_daily = any("daily" in s for s in sheet_names)
    has_category = any("category" in s for s in sheet_names)
    record("Excel has 'Daily Plan' sheet", has_daily, f"Found sheets: {wb.sheetnames}")
    record("Excel has 'Category Distribution' sheet", has_category, f"Found sheets: {wb.sheetnames}")

    # Check Daily Plan has >= 21 rows of data
    daily_sheet = None
    for sname in wb.sheetnames:
        if "daily" in sname.lower():
            daily_sheet = wb[sname]
            break

    daily_rows = 0
    if daily_sheet:
        daily_rows = sum(1 for row in daily_sheet.iter_rows(min_row=2, values_only=True)
                         if any(cell is not None and str(cell).strip() != "" for cell in row))
        record("Excel Daily Plan sheet has at least 21 data rows", daily_rows >= 21,
               f"Found {daily_rows} data rows")

    # Check Category Distribution has >= 3 categories
    cat_sheet = None
    for sname in wb.sheetnames:
        if "category" in sname.lower():
            cat_sheet = wb[sname]
            break

    if cat_sheet:
        cat_data_rows = list(cat_sheet.iter_rows(min_row=2, values_only=True))
        cat_rows = [r for r in cat_data_rows
                    if any(cell is not None and str(cell).strip() != "" for cell in r)]
        record("Excel Category Distribution has at least 3 categories", len(cat_rows) >= 3,
               f"Found {len(cat_rows)} category rows")

        # Category counts should sum to the number of Daily Plan rows.
        total_count = 0
        count_ok = False
        for r in cat_rows:
            for cell in r:
                try:
                    total_count += int(float(cell))
                    break
                except (TypeError, ValueError):
                    continue
        if daily_rows:
            count_ok = abs(total_count - daily_rows) <= 1
        record("Excel category counts sum to Daily Plan row count", count_ok,
               f"Sum of category counts={total_count}, Daily Plan rows={daily_rows}")


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT summary FROM gcal.events
            WHERE LOWER(summary) LIKE '%meal prep%' OR LOWER(summary) LIKE '%prep planning%'
        """)
        prep_events = cur.fetchall()
        record("GCal has weekly meal prep planning event", len(prep_events) > 0,
               f"Found: {prep_events}")

        cur.execute("""
            SELECT summary FROM gcal.events
            WHERE LOWER(summary) LIKE '%daily meal reminder%' OR LOWER(summary) LIKE '%meal reminder%'
        """)
        daily_events = cur.fetchall()
        record("GCal has at least 7 daily meal reminder events", len(daily_events) >= 7,
               f"Found {len(daily_events)} daily reminder events")

        cur.execute("SELECT COUNT(*) FROM gcal.events")
        total_events = cur.fetchone()[0]
        record("GCal has at least 8 events total (1 prep + 7 daily)", total_events >= 8,
               f"Found {total_events} total events")

        conn.close()
    except Exception as e:
        record("GCal connection", False, str(e))


def check_email():
    print("\n=== Checking Email ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE LOWER(subject) LIKE '%meal plan%'
        """)
        emails = cur.fetchall()
        record("Email with 'Meal Plan' in subject sent", len(emails) > 0,
               f"Found {len(emails)} matching emails")

        target_found = False
        for subject, to_addr in emails:
            to_str = str(to_addr).lower() if to_addr else ""
            if "family@home.com" in to_str:
                target_found = True
                break
        record("Email to family@home.com with meal plan subject", target_found,
               f"Recipients: {[e[1] for e in emails]}")

        conn.close()
    except Exception as e:
        record("Email connection", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_gsheet()
    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES ({len(critical_failed)}): {critical_failed}")
        print("FAIL: critical check(s) failed.")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print(f"FAIL: accuracy {accuracy:.1f}% < 70%")
        sys.exit(1)


if __name__ == "__main__":
    main()
