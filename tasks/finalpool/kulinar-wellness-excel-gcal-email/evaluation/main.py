"""Evaluation для kulinar-wellness-excel-gcal-email.

Критические проверки (CRITICAL_CHECKS): любой провал => общий FAIL независимо от
accuracy. Иначе порог прохождения: accuracy >= 70%.
"""
import argparse
import json
import os
import re
import sys
from datetime import datetime, timedelta

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Критические (семантические) проверки: любой провал => общий FAIL.
CRITICAL_CHECKS = {
    "Recipe_Name значения — реальные блюда из базы kulinar",
    "Summary: Unique_Categories соответствует столбцу Category и >= 2",
    "Все 5 'Healthy Lunch Break' начинаются в 12:00 и длятся 1 час",
    "5 'Healthy Lunch Break' приходятся на Пн-Пт недели запуска задачи",
    "Письмо all-staff от wellness@company.example.com с темой и RU-целями",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


# ---------------------------------------------------------------------------
# Эталонная база kulinar (источник правды для подлинности рецептов)
# ---------------------------------------------------------------------------

def _norm_name(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def load_kulinar_recipes():
    """Возвращает список рецептов kulinar или None, если база недоступна."""
    candidates = []
    env = os.environ.get("KULINAR_RECIPES_JSON")
    if env:
        candidates.append(env)

    here = os.path.abspath(__file__)
    cur = os.path.dirname(here)
    rel = os.path.join("local_servers", "kulinar-mcp", "src", "data", "all_recipes.json")
    for _ in range(12):
        candidates.append(os.path.join(cur, rel))
        parent = os.path.dirname(cur)
        if parent == cur:
            break
        cur = parent

    for path in candidates:
        try:
            if path and os.path.exists(path):
                with open(path, encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list) and data:
                    return data
        except Exception:
            continue
    return None


KULINAR = load_kulinar_recipes()
KULINAR_NAMES = {_norm_name(r["name"]) for r in KULINAR} if KULINAR else set()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Excel meal plan file."""
    print("\n=== Checking Excel File ===")
    try:
        import openpyxl
    except ImportError:
        check("openpyxl available", False, "openpyxl not installed")
        return

    agent_file = os.path.join(agent_workspace, "Wellness_Meal_Plan.xlsx")

    check("Wellness_Meal_Plan.xlsx exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    # Check Meal Plan sheet
    print("\n--- Meal Plan Sheet ---")
    mp_ws = get_sheet(agent_wb, "Meal Plan")
    check("Sheet 'Meal Plan' exists", mp_ws is not None, f"Found: {agent_wb.sheetnames}")

    distinct_categories = set()
    recipe_names = []

    if mp_ws:
        headers = [c.value for c in list(mp_ws.rows)[0]] if mp_ws.max_row > 0 else []
        check("Has Day column", any("day" in str(h).lower() for h in headers if h), f"Headers: {headers}")
        check("Has Meal_Type column",
              any("meal" in str(h).lower() or "type" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")
        check("Has Recipe_Name column",
              any("recipe" in str(h).lower() for h in headers if h),
              f"Headers: {headers}")

        data_rows = [row for row in mp_ws.iter_rows(min_row=2, values_only=True)
                     if any(v is not None for v in row)]
        check("Meal Plan has at least 15 rows", len(data_rows) >= 15,
              f"Found {len(data_rows)} rows")

        # Check all 5 days covered (Day column stays English Monday..Friday)
        days_found = set()
        for row in data_rows:
            if row and row[0]:
                day = str(row[0]).strip().lower()
                for d in ["monday", "tuesday", "wednesday", "thursday", "friday"]:
                    if d in day:
                        days_found.add(d)
        check("All 5 weekdays covered", len(days_found) == 5,
              f"Found days: {days_found}")

        # Check meal types
        meal_types_found = set()
        col_idx = 1  # Meal_Type is 2nd column (0-indexed)
        for row in data_rows:
            if row and len(row) > col_idx and row[col_idx]:
                mt = str(row[col_idx]).strip().lower()
                if mt in ["breakfast", "lunch", "dinner"]:
                    meal_types_found.add(mt)
        check("All 3 meal types present (Breakfast, Lunch, Dinner)",
              len(meal_types_found) == 3,
              f"Found: {meal_types_found}")

        # Collect Recipe_Name (col 2) and Category (col 3) values
        for row in data_rows:
            if row and len(row) > 2 and row[2] is not None and str(row[2]).strip():
                recipe_names.append(str(row[2]).strip())
            if row and len(row) > 3 and row[3] is not None and str(row[3]).strip():
                distinct_categories.add(str(row[3]).strip().lower())

        # CRITICAL: Recipe names must be real kulinar dishes
        if KULINAR_NAMES:
            matched = [n for n in recipe_names if _norm_name(n) in KULINAR_NAMES]
            ok = len(recipe_names) > 0 and len(matched) >= max(1, (len(recipe_names) + 1) // 2)
            check("Recipe_Name значения — реальные блюда из базы kulinar", ok,
                  f"matched {len(matched)}/{len(recipe_names)}")
        else:
            check("Recipe_Name значения — реальные блюда из базы kulinar", False,
                  "kulinar all_recipes.json не найден")

    # Check Summary sheet
    print("\n--- Summary Sheet ---")
    sum_ws = get_sheet(agent_wb, "Summary")
    check("Sheet 'Summary' exists", sum_ws is not None, f"Found: {agent_wb.sheetnames}")

    if sum_ws:
        summary_data = {}
        for row in sum_ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary_data[str(row[0]).strip().lower()] = row[1]

        check("Summary has Days_Covered = 5",
              any(("days" in k) and num_close(v, 5, 0.1) for k, v in summary_data.items()),
              f"Data: {summary_data}")
        check("Summary has Total_Recipes >= 15",
              any(("total" in k and "recipe" in k) and v is not None and float(v) >= 15
                  for k, v in summary_data.items() if v is not None),
              f"Data: {summary_data}")

        # CRITICAL: Unique_Categories internal consistency
        uc_val = None
        for k, v in summary_data.items():
            if "unique" in k and ("categ" in k or "categor" in k):
                uc_val = v
                break
        expected_uc = len(distinct_categories)
        ok = (uc_val is not None and expected_uc >= 2 and num_close(uc_val, expected_uc, 0.1))
        check("Summary: Unique_Categories соответствует столбцу Category и >= 2", ok,
              f"Summary={uc_val}, distinct in Category={expected_uc}")
    else:
        check("Summary: Unique_Categories соответствует столбцу Category и >= 2", False,
              "лист Summary отсутствует")


def _launch_week_monday(launch_time):
    """Возвращает date понедельника недели запуска или None.

    Если запуск приходится на выходной (Сб/Вс), переносим на понедельник
    следующей рабочей недели: планировать обеды на уже прошедшие будни недели
    бессмысленно.
    """
    if not launch_time:
        return None

    def _monday(dt):
        # Запуск в выходной (5=Сб, 6=Вс) -> понедельник следующей недели.
        if dt.weekday() >= 5:
            return (dt + timedelta(days=7 - dt.weekday())).date()
        return (dt - timedelta(days=dt.weekday())).date()

    s = str(launch_time).strip().replace("T", " ").replace("Z", "")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s[:len(fmt) + 2].strip()[:19], fmt)
            return _monday(dt)
        except ValueError:
            continue
    try:
        dt = datetime.fromisoformat(s)
        return _monday(dt)
    except ValueError:
        return None


def check_gcal(launch_time):
    """Check Google Calendar events."""
    print("\n=== Checking Google Calendar Events ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, start_datetime, end_datetime, start_timezone FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    conn.close()

    check("At least 5 calendar events created", len(events) >= 5,
          f"Found {len(events)} events")

    lunch_events = [e for e in events if "healthy lunch break" in (e[1] or "").lower()]
    check("5 'Healthy Lunch Break' events created",
          len(lunch_events) >= 5,
          f"Healthy Lunch Break events: {len(lunch_events)}")

    def _to_dt(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v
        try:
            return datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        except ValueError:
            return None

    # CRITICAL: each event starts at 12:00 and lasts exactly 1 hour
    noon_hour_ok = []
    for e in lunch_events:
        sdt = _to_dt(e[2])
        edt = _to_dt(e[3])
        if sdt is None or edt is None:
            noon_hour_ok.append(False)
            continue
        # "полдень" is timezone-unspecified in task.md: noon counts in the event's
        # OWN declared zone (start_timezone), falling back to Europe/Moscow.
        tzname = e[4] if len(e) > 4 and e[4] else "Europe/Moscow"
        try:
            from zoneinfo import ZoneInfo
            local = sdt.astimezone(ZoneInfo(tzname)) if sdt.tzinfo else sdt
        except Exception:
            local = sdt
        starts_noon = local.hour == 12 and local.minute == 0
        one_hour = abs((edt - sdt).total_seconds() - 3600) <= 60
        noon_hour_ok.append(starts_noon and one_hour)
    ok_noon = len(lunch_events) >= 5 and sum(noon_hour_ok) >= 5
    check("Все 5 'Healthy Lunch Break' начинаются в 12:00 и длятся 1 час", ok_noon,
          f"valid noon+1h events: {sum(noon_hour_ok)} of {len(lunch_events)}")

    # CRITICAL: events fall on Mon-Fri of the launch week, one per weekday
    monday = _launch_week_monday(launch_time)
    if monday is None:
        check("5 'Healthy Lunch Break' приходятся на Пн-Пт недели запуска задачи", False,
              f"не удалось разобрать launch_time={launch_time}")
    else:
        # Принимаем будни (Пн-Пт) недели запуска ЛИБО сразу следующей рабочей
        # недели: если запуск близок к концу недели, корректно планировать
        # обеды на ближайшие предстоящие будни, а не на уже прошедшие.
        next_monday = monday + timedelta(days=7)
        candidate_weeks = [
            {monday + timedelta(days=i) for i in range(5)},        # Mon..Fri недели запуска
            {next_monday + timedelta(days=i) for i in range(5)},   # Mon..Fri следующей недели
        ]
        best = (set(), monday)
        for wk_monday, week_days in ((monday, candidate_weeks[0]), (next_monday, candidate_weeks[1])):
            found = set()
            for e in lunch_events:
                sdt = _to_dt(e[2])
                if sdt is not None and sdt.date() in week_days:
                    found.add(sdt.date())
            if len(found) > len(best[0]):
                best = (found, wk_monday)
        found_days, matched_monday = best
        ok_week = len(found_days) == 5
        check("5 'Healthy Lunch Break' приходятся на Пн-Пт недели запуска задачи", ok_week,
              f"week Mon={matched_monday}, distinct weekdays matched={len(found_days)}")


def check_emails():
    """Check that wellness email was sent."""
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    conn.close()

    def find_email_for_recipient(recipient):
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    result = find_email_for_recipient("all-staff@company.example.com")

    # CRITICAL: combined email check (recipient + from + subject + RU goals)
    crit_ok = False
    detail = f"Total emails: {len(all_emails)}"
    if result:
        subj, from_addr, to_addr, body = result
        subj_l = (subj or "").lower()
        body_l = (body or "").lower()
        subject_ok = "wellness" in subj_l and "5" in (subj or "")
        from_ok = "wellness@company.example.com" in (from_addr or "").lower()
        goals_ok = any(kw in body_l for kw in ["стресс", "фокус", "концентрац", "stress", "focus"])
        crit_ok = subject_ok and from_ok and goals_ok
        detail = f"subject_ok={subject_ok}, from_ok={from_ok}, goals_ok={goals_ok}, subj={subj}"
    check("Письмо all-staff от wellness@company.example.com с темой и RU-целями",
          crit_ok, detail)

    # Non-critical granular checks
    check("Email sent to all-staff@company.example.com", result is not None,
          f"Total emails: {len(all_emails)}")

    if result:
        subj, from_addr, to_addr, body = result
        check("Email subject contains '5-Day Corporate Wellness Meal Plan'",
              "wellness" in (subj or "").lower() and "5" in (subj or ""),
              f"Subject: {subj}")
        check("Email from wellness@company.example.com",
              "wellness@company.example.com" in (from_addr or "").lower(),
              f"From: {from_addr}")
        body_lower = (body or "").lower()
        check("Email mentions wellness goals (stress or focus)",
              any(kw in body_lower for kw in
                  ["стресс", "фокус", "концентрац", "stress", "focus"]),
              "Expected wellness goals mentioned")
        check("Email mentions 5 days or 15 meals",
              "5" in (body or "") or "15" in (body or "")
              or "пять" in body_lower or "five" in body_lower,
              "Expected meal plan scope mentioned")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("KULINAR WELLNESS EXCEL GCAL EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, gt_dir)
    check_gcal(args.launch_time)
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
