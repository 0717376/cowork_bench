"""Evaluation for canvas-quiz-report.

Two layers:
  * Structural diff (non-critical): per-row/value comparison of the agent XLSX
    against the pre-built groundtruth XLSX (Canvas data is read live, so the GT
    file encodes the values served at build time). These feed the accuracy gate.
  * CRITICAL_CHECKS (semantic): core deliverable invariants. Any critical
    failure => overall FAIL regardless of accuracy. These are derived from the
    GT file itself so they stay in sync with whatever Canvas served.

PASS requires: no critical failure AND accuracy >= 70%.
"""
import argparse
import os
import sys

import openpyxl


PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "All 'Quiz Performance' rows have Submissions > 100",
    "Pass_Rate_Pct == round(Avg_Score / Points_Possible * 100, 1) for each row",
    "Summary.Total_Quizzes equals row count in 'Quiz Performance'",
    "Summary.Hardest_Quiz is the lowest-pass-rate quiz title",
    "Summary.Easiest_Quiz is the highest-pass-rate quiz title",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", "").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def structural_diff(agent_wb, gt_wb):
    """Non-critical per-value diff; returns list of error strings."""
    all_errors = []

    # Quiz Performance: join on Quiz_Title (col index 1), lowercased.
    a_rows = load_sheet_rows(agent_wb, "Quiz Performance")
    g_rows = load_sheet_rows(gt_wb, "Quiz Performance")
    if a_rows is None:
        all_errors.append("Sheet 'Quiz Performance' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Quiz Performance' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[1] is not None:
                a_lookup[str(row[1]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[1] is None:
                continue
            key = str(g_row[1]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row: {g_row[1]}")
                continue
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 1.0):
                all_errors.append(f"{key}.Points_Possible: {a_row[2]} vs {g_row[2]} (tol=1.0)")
            if len(a_row) > 4 and len(g_row) > 4 and not num_close(a_row[4], g_row[4], 5):
                all_errors.append(f"{key}.Submissions: {a_row[4]} vs {g_row[4]} (tol=5)")
            if len(a_row) > 5 and len(g_row) > 5 and not num_close(a_row[5], g_row[5], 1.0):
                all_errors.append(f"{key}.Avg_Score: {a_row[5]} vs {g_row[5]} (tol=1.0)")
            if len(a_row) > 6 and len(g_row) > 6 and not num_close(a_row[6], g_row[6], 1.0):
                all_errors.append(f"{key}.Pass_Rate_Pct: {a_row[6]} vs {g_row[6]} (tol=1.0)")

    # Summary: join on Metric (col index 0).
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing Summary row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 5.0):
                all_errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=5.0)")

    return all_errors


def summary_lookup(wb):
    rows = load_sheet_rows(wb, "Summary")
    out = {}
    if not rows:
        return out
    for row in rows[1:]:
        if row and row[0] is not None and len(row) > 1:
            out[str(row[0]).strip().lower()] = row[1]
    return out


def critical_checks(agent_wb, gt_wb):
    """Semantic invariants checked against the agent output, with expected
    titles derived from the GT (kept in sync with live Canvas)."""
    a_rows = load_sheet_rows(agent_wb, "Quiz Performance")
    a_data = a_rows[1:] if a_rows and len(a_rows) > 1 else []
    a_data = [r for r in a_data if r and r[1] is not None]

    # 1. Submissions > 100 for every reported quiz.
    bad_sub = []
    for r in a_data:
        sub = safe_float(r[4]) if len(r) > 4 else None
        if sub is None or sub <= 100:
            bad_sub.append((r[1], sub))
    check(
        "All 'Quiz Performance' rows have Submissions > 100",
        len(a_data) > 0 and not bad_sub,
        f"violations={bad_sub[:5]}",
    )

    # 2. Pass_Rate_Pct formula consistency per row.
    bad_pr = []
    for r in a_data:
        pts = safe_float(r[2]) if len(r) > 2 else None
        avg = safe_float(r[5]) if len(r) > 5 else None
        pr = safe_float(r[6]) if len(r) > 6 else None
        if pts in (None, 0) or avg is None or pr is None:
            bad_pr.append((r[1], "missing"))
            continue
        expected = round(avg / pts * 100, 1)
        if abs(expected - pr) > 1.0:
            bad_pr.append((r[1], f"{pr} vs {expected}"))
    check(
        "Pass_Rate_Pct == round(Avg_Score / Points_Possible * 100, 1) for each row",
        len(a_data) > 0 and not bad_pr,
        f"violations={bad_pr[:5]}",
    )

    a_summary = summary_lookup(agent_wb)

    # 3. Total_Quizzes equals actual row count.
    total = safe_float(a_summary.get("total_quizzes"))
    check(
        "Summary.Total_Quizzes equals row count in 'Quiz Performance'",
        total is not None and abs(total - len(a_data)) < 0.5,
        f"Total_Quizzes={total} rows={len(a_data)}",
    )

    # 4 & 5. Hardest/Easiest titles: derive expected set from GT extremes by
    # pass rate. Accept agent value if it matches the GT-declared extreme OR an
    # actual extreme computed over the agent's own rows (ties tolerated).
    g_summary = summary_lookup(gt_wb)
    gt_hardest = str(g_summary.get("hardest_quiz", "")).strip().lower()
    gt_easiest = str(g_summary.get("easiest_quiz", "")).strip().lower()

    pr_pairs = []
    for r in a_data:
        pr = safe_float(r[6]) if len(r) > 6 else None
        if pr is not None:
            pr_pairs.append((str(r[1]).strip().lower(), pr))
    agent_min_titles = set()
    agent_max_titles = set()
    if pr_pairs:
        min_pr = min(p for _, p in pr_pairs)
        max_pr = max(p for _, p in pr_pairs)
        agent_min_titles = {t for t, p in pr_pairs if abs(p - min_pr) <= 1.0}
        agent_max_titles = {t for t, p in pr_pairs if abs(p - max_pr) <= 1.0}

    a_hardest = str(a_summary.get("hardest_quiz", "")).strip().lower()
    a_easiest = str(a_summary.get("easiest_quiz", "")).strip().lower()
    check(
        "Summary.Hardest_Quiz is the lowest-pass-rate quiz title",
        bool(a_hardest) and (a_hardest == gt_hardest or a_hardest in agent_min_titles),
        f"agent={a_hardest} gt={gt_hardest}",
    )
    check(
        "Summary.Easiest_Quiz is the highest-pass-rate quiz title",
        bool(a_easiest) and (a_easiest == gt_easiest or a_easiest in agent_max_titles),
        f"agent={a_easiest} gt={gt_easiest}",
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Quiz_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Quiz_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # --- Critical semantic checks ---
    print("Critical checks:")
    critical_checks(agent_wb, gt_wb)

    # --- Structural diff (non-critical, feeds accuracy) ---
    print("Structural diff:")
    struct_errors = structural_diff(agent_wb, gt_wb)
    check(
        "Structural diff against groundtruth (per-row values)",
        not struct_errors,
        f"{len(struct_errors)} mismatch(es): " + "; ".join(struct_errors[:5]),
    )

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if critical_failed:
        print("=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print("=== RESULT: FAIL (accuracy < 70) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
