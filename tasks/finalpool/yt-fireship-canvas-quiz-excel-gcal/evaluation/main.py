"""
Evaluation for yt-fireship-canvas-quiz-excel-gcal task.

Checks:
1. TypeScript_Course_Tracker.xlsx exists with "Videos" sheet having >= 4 rows
2. "Quiz_Questions" sheet exists with >= 8 rows
3. "Grade_Template" sheet exists
4. Canvas quiz exists (canvas.quizzes WHERE title ILIKE '%TypeScript%')
5. Canvas quiz has >= 5 questions (canvas.quiz_questions)
6. GCal has TypeScript/Quiz event in March/April 2026
7. Email sent to students@webdev.edu
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Semantic CRITICAL checks. Each maps to a boolean gate that MUST pass.
# A correct RU agent satisfies all of these; a non-doer fails every one
# (preprocess seeds NO xlsx, NO canvas quiz, NO sent email).
CRITICAL = {
    "excel_three_sheets": False,   # xlsx with Videos>=4, Quiz_Questions>=8, Grade_Template
    "canvas_quiz": False,          # TypeScript quiz with >=5 questions
    "gcal_two_events": False,      # >=2 TypeScript/Quiz events in Mar/Apr 2026
    "email_to_students": False,    # email to students@webdev.edu w/ quiz content
}


def critical(key, passed):
    if passed:
        CRITICAL[key] = True


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-3: TypeScript_Course_Tracker.xlsx ===")
    xlsx_path = None
    for fname in os.listdir(agent_workspace):
        if fname.lower().endswith(".xlsx") and ("typescript" in fname.lower() or "tracker" in fname.lower() or "course" in fname.lower()):
            xlsx_path = os.path.join(agent_workspace, fname)
            break
    if not xlsx_path:
        for fname in os.listdir(agent_workspace):
            if fname.lower().endswith(".xlsx"):
                xlsx_path = os.path.join(agent_workspace, fname)
                break

    record("TypeScript_Course_Tracker.xlsx exists", xlsx_path is not None,
           f"No matching xlsx in {agent_workspace}")

    if not xlsx_path:
        for chk in ["Videos sheet has >= 4 rows", "Quiz_Questions sheet has >= 8 rows", "Grade_Template sheet exists"]:
            record(chk, False, "xlsx not found")
        return

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path)

        # Require the three EXACT sheet names (case-insensitive), not just any sheet.
        sheet_map = {s.strip().lower(): s for s in wb.sheetnames}
        videos_ok = quiz_ok = grade_ok = False

        # Videos sheet (must be named "Videos")
        if "videos" in sheet_map:
            videos_sheet = wb[sheet_map["videos"]]
            data_rows = [r for r in videos_sheet.iter_rows(min_row=2, values_only=True)
                         if any(c is not None for c in r)]
            videos_ok = len(data_rows) >= 4
            record("Videos sheet has >= 4 data rows", videos_ok,
                   f"Found {len(data_rows)} rows")
        else:
            record("Videos sheet has >= 4 data rows", False, f"No 'Videos' sheet: {wb.sheetnames}")

        # Quiz_Questions sheet (must be named "Quiz_Questions")
        if "quiz_questions" in sheet_map:
            qsheet = wb[sheet_map["quiz_questions"]]
            record("Quiz_Questions sheet exists", True, "")
            data_rows = [r for r in qsheet.iter_rows(min_row=2, values_only=True)
                         if any(c is not None for c in r)]
            quiz_ok = len(data_rows) >= 8
            record("Quiz_Questions sheet has >= 8 data rows", quiz_ok,
                   f"Found {len(data_rows)} rows")
        else:
            record("Quiz_Questions sheet exists", False, f"Sheets: {wb.sheetnames}")
            record("Quiz_Questions sheet has >= 8 data rows", False, "Sheet not found")

        # Grade_Template sheet (must be named "Grade_Template")
        grade_ok = "grade_template" in sheet_map
        record("Grade_Template sheet exists", grade_ok, f"Sheets: {wb.sheetnames}")

        # CRITICAL: all three correctly-named sheets present with required rows.
        critical("excel_three_sheets", videos_ok and quiz_ok and grade_ok)

        # --- Groundtruth XLSX value comparison ---
        gt_path = os.path.join(groundtruth_workspace, "TypeScript_Course_Tracker.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = wb[asn]; break
                if a_ws is None:
                    record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                       f"Expected {len(gt_rows)}, got {len(a_rows)}")
                for ri in range(min(3, len(gt_rows))):
                    if ri >= len(a_rows): break
                    ok = True
                    for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                        gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                        if gv is None: continue
                        if isinstance(gv, (int, float)):
                            if not num_close(av, gv, max(abs(gv)*0.1, 1.0)): ok = False; break
                        else:
                            if not str_match(av, gv): ok = False; break
                    record(f"GT '{gt_sname}' row {ri+1} values", ok,
                           f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
            gt_wb.close()

    except Exception as e:
        for chk in ["Videos sheet has >= 4 rows", "Quiz_Questions sheet has >= 8 rows", "Grade_Template sheet exists"]:
            record(chk, False, str(e))


def check_canvas():
    print("\n=== Check 4-5: Canvas Quiz ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT id, title FROM canvas.quizzes
            WHERE title ILIKE '%typescript%' OR title ILIKE '%type script%'
            ORDER BY id DESC LIMIT 1
        """)
        quiz = cur.fetchone()
        record("Canvas quiz with TypeScript in title exists", quiz is not None,
               "No TypeScript quiz found in canvas.quizzes")

        if quiz:
            quiz_id = quiz[0]
            cur.execute("""
                SELECT COUNT(*) FROM canvas.quiz_questions WHERE quiz_id = %s
            """, (quiz_id,))
            q_count = cur.fetchone()[0]
            record("Canvas quiz has >= 5 questions", q_count >= 5,
                   f"Found {q_count} questions for quiz {quiz_id}")
            critical("canvas_quiz", q_count >= 5)
        else:
            record("Canvas quiz has >= 5 questions", False, "No quiz found")
    except Exception as e:
        record("Canvas quiz with TypeScript in title exists", False, str(e))
        record("Canvas quiz has >= 5 questions", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 6: GCal TypeScript Events in Mar/Apr 2026 ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-01' AND start_datetime < '2026-05-01'
        AND (summary ILIKE '%typescript%' OR summary ILIKE '%quiz%')
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    record("GCal has >= 1 TypeScript/Quiz event in Mar/Apr 2026",
           len(events) >= 1, f"Found {len(events)} matching events")
    if events:
        record("GCal has >= 2 TypeScript/Quiz events (deadline + review)",
               len(events) >= 2, f"Found {len(events)} events")
    critical("gcal_two_events", len(events) >= 2)


def check_email():
    print("\n=== Check 7: Email to students@webdev.edu ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else to_addr.lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "students@webdev.edu" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    record("Email sent to students@webdev.edu", matching is not None,
           f"Total messages: {len(messages)}")
    if matching:
        body = (matching[0] or "") + " " + (matching[3] or "")
        has_content = any(k in body.lower() for k in ["typescript", "quiz", "deadline"])
        record("Email mentions TypeScript quiz content", has_content, f"Subject: {matching[0]}")
        # CRITICAL: an outbound announcement to students@webdev.edu about the quiz.
        critical("email_to_students", has_content)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_canvas()
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    # --- CRITICAL semantic gate: every critical check must pass. ---
    print("\n=== CRITICAL gate ===")
    crit_failed = [k for k, v in CRITICAL.items() if not v]
    for k, v in CRITICAL.items():
        print(f"  [{'PASS' if v else 'FAIL'}] CRITICAL: {k}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_passed": len(CRITICAL) - len(crit_failed),
        "critical_total": len(CRITICAL),
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if crit_failed:
        print(f"\nFAIL: critical checks failed: {crit_failed}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
