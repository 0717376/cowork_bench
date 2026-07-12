"""Evaluation for canvas-exam-prep-scheduler (RU-localized, keep-foreign archetype).

Все исходные значения (курсы, средние баллы, пороговые тесты) читаются из БД
(canvas.*), а не захардкожены — имена курсов берутся из глобального seed Canvas.

Критические проверки (CRITICAL_CHECKS): любой их провал => общий FAIL,
независимо от accuracy. Структурные проверки (файл/лист/заголовок существует)
неприоритетны. Порог прохождения: accuracy >= 70 И нет провалов критических.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

FOCUS_THRESHOLD = 80  # из prep_guidelines.json: focus_threshold
COORDINATOR_EMAIL = "academic-coordinator@university.edu"
EMAIL_SUBJECT = "Exam Review Sessions Scheduled"  # сохранён как английский маркер

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Критические (семантические) проверки. Имена должны точно совпадать со
# строками, переданными в check(...).
CRITICAL_CHECKS = {
    "Below-threshold: помеченные тесты совпадают с расчётом по БД",
    "Summary: Total_Quizzes_Analyzed и Below_Threshold_Quizzes верны",
    "Review Schedule: по одной сессии на нужный курс, будни, дата >= 2026-03-16, дневной слот",
    "Calendar: события на нужные курсы с упоминанием среднего балла",
    "Email: письмо координатору с верной темой и содержанием",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def is_below_yes(v):
    """Принимаем английские Yes/True/Y и русские Да/Истина."""
    s = str(v).strip().lower()
    return s in ("yes", "true", "y", "да", "истина", "1")


def is_below_no(v):
    s = str(v).strip().lower()
    return s in ("no", "false", "n", "нет", "ложь", "0")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_quiz_data():
    """Средние баллы по 100-балльным тестам из БД."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name, q.title, ROUND(AVG(qs.score), 1) as avg_score
        FROM canvas.quizzes q
        JOIN canvas.courses c ON c.id = q.course_id
        LEFT JOIN canvas.quiz_submissions qs ON qs.quiz_id = q.id
        WHERE q.points_possible = 100
        GROUP BY c.name, q.title
        ORDER BY c.name, q.title
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_excel(agent_workspace, expected, below_thr, courses_needing):
    print("\n=== Проверка Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Exam_Prep.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Exam_Prep.xlsx exists", False, f"Не найден: {xlsx_path}")
        return
    check("Exam_Prep.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    n_courses_needing = len(courses_needing)

    # --- Quiz Performance ---
    qp_rows = load_sheet_rows(wb, "Quiz Performance")
    if qp_rows is None:
        check("Sheet 'Quiz Performance' exists", False, f"Есть: {wb.sheetnames}")
    else:
        check("Sheet 'Quiz Performance' exists", True)
        data_rows = [r for r in qp_rows[1:] if r and r[0]] if len(qp_rows) > 1 else []
        check(f"Quiz Performance has {len(expected)} rows (100-pt quizzes)",
              abs(len(data_rows) - len(expected)) <= 2,
              f"Найдено {len(data_rows)}, ожидалось {len(expected)}")

        header = qp_rows[0] if qp_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "quiz", "avg_score", "below_threshold"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Заголовок: {header}")

        # СЕМАНТИКА (критично): помечены именно те тесты, что < порога по БД.
        below_titles = {str(r[1]).strip().lower() for r in below_thr}
        mismatches = []
        for row in data_rows:
            if not row or row[1] is None:
                continue
            title = str(row[1]).strip().lower()
            should_be_below = any(bt in title or title in bt for bt in below_titles)
            flag = row[3] if len(row) > 3 else None
            if should_be_below and not is_below_yes(flag):
                mismatches.append(f"{row[1]}=>{flag} (ожид. Yes)")
            if (not should_be_below) and is_below_yes(flag):
                # допускаем, только если балл реально < порога
                if not (row[2] is not None and float(row[2]) < FOCUS_THRESHOLD - 0.05):
                    mismatches.append(f"{row[1]}=>{flag} (ожид. No)")
        check("Below-threshold: помеченные тесты совпадают с расчётом по БД",
              len(mismatches) == 0, f"Несовпадений: {mismatches[:5]}")

        # Спот-проверка значения Avg_Score для одного тест ниже порога.
        if below_thr:
            sample = below_thr[0]
            for row in data_rows:
                if row and row[1] and sample[1].lower() in str(row[1]).lower():
                    check(f"Avg_Score теста '{sample[1]}' ~{sample[2]}",
                          num_close(row[2], float(sample[2])), f"Получено {row[2]}")
                    break

    # --- Review Schedule ---
    rs_rows = load_sheet_rows(wb, "Review Schedule")
    if rs_rows is None:
        check("Sheet 'Review Schedule' exists", False, f"Есть: {wb.sheetnames}")
    else:
        check("Sheet 'Review Schedule' exists", True)
        data_rows = [r for r in rs_rows[1:] if r and r[0]] if len(rs_rows) > 1 else []
        check(f"Review Schedule has ~{n_courses_needing} sessions",
              abs(len(data_rows) - n_courses_needing) <= 2,
              f"Найдено {len(data_rows)}, ожидалось ~{n_courses_needing}")

        # СЕМАНТИКА (критично): по одной строке на каждый нужный курс,
        # дата ISO >= 2026-03-16 и будний день, время — дневной слот 16:00.
        import datetime as _dt
        sched_courses = set()
        date_ok = True
        time_ok = True
        for row in data_rows:
            if not row or not row[0]:
                continue
            sched_courses.add(str(row[0]).strip().lower())
            dval = row[3] if len(row) > 3 else None
            tval = row[4] if len(row) > 4 else None
            d = None
            if isinstance(dval, _dt.datetime):
                d = dval.date()
            elif isinstance(dval, _dt.date):
                d = dval
            else:
                try:
                    d = _dt.date.fromisoformat(str(dval)[:10])
                except Exception:
                    d = None
            if d is None or d < _dt.date(2026, 3, 16) or d.weekday() >= 5:
                date_ok = False
            tstr = str(tval) if tval is not None else ""
            if "16:00" not in tstr and "16" not in tstr.split(":")[0]:
                time_ok = False
        needing_lc = {c.lower() for c in courses_needing}
        coverage_ok = needing_lc.issubset(sched_courses) if needing_lc else (len(data_rows) >= 0)
        check("Review Schedule: по одной сессии на нужный курс, будни, дата >= 2026-03-16, дневной слот",
              coverage_ok and date_ok and time_ok and len(sched_courses) == len(needing_lc),
              f"courses={sched_courses} needing={needing_lc} date_ok={date_ok} time_ok={time_ok}")

    # --- Summary ---
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Есть: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        total_ok = num_close(lookup.get("total_quizzes_analyzed"), len(expected))
        below_ok = num_close(lookup.get("below_threshold_quizzes"), len(below_thr))
        check("Summary: Total_Quizzes_Analyzed и Below_Threshold_Quizzes верны",
              total_ok and below_ok,
              f"total={lookup.get('total_quizzes_analyzed')} (ожид {len(expected)}), "
              f"below={lookup.get('below_threshold_quizzes')} (ожид {len(below_thr)})")

        # Неприоритетно: наличие двух дополнительных строк Summary.
        check("Summary: строка Courses_Needing_Review присутствует",
              "courses_needing_review" in lookup, f"Ключи: {list(lookup.keys())}")
        check("Summary: строка Review_Sessions_Scheduled присутствует",
              "review_sessions_scheduled" in lookup, f"Ключи: {list(lookup.keys())}")
        check("Summary: Courses_Needing_Review корректно",
              num_close(lookup.get("courses_needing_review"), n_courses_needing),
              f"Получено {lookup.get('courses_needing_review')}, ожид {n_courses_needing}")


def check_calendar(courses_needing, below_avgs_by_course):
    print("\n=== Проверка событий календаря ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, COALESCE(description, ''), start_datetime
            FROM gcal.events
            WHERE summary ILIKE '%%review%%' OR summary ILIKE '%%разбор%%'
               OR summary ILIKE '%%quiz%%' OR summary ILIKE '%%повтор%%'
               OR summary ILIKE '%%exam%%' OR summary ILIKE '%%экзамен%%'
        """)
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar: события созданы", False, str(e))
        return

    # Неприоритетно: вообще есть события разбора.
    check("Calendar: события созданы", len(events) >= 1, f"Найдено {len(events)} событий")

    # СЕМАНТИКА (критично): по событию на каждый нужный курс, и в описании
    # упомянут средний балл, послуживший поводом.
    covered = 0
    desc_score_ok = 0
    for course in courses_needing:
        cl = course.lower()
        matched = [e for e in events if cl in (e[0] or "").lower()]
        if not matched:
            continue
        covered += 1
        avgs = below_avgs_by_course.get(course) or []
        if not avgs:
            continue
        # task.md требует упоминания среднего балла теста, ПОСЛУЖИВШЕГО поводом
        # для разбора. У курса может быть несколько тестов ниже порога — корректно
        # сослаться на ЛЮБОЙ из них, не обязательно минимальный. Поэтому принимаем
        # любое из допустимых текстовых представлений любого below-порогового балла.
        cands = set()
        for avg in avgs:
            cands.add(str(avg))
            cands.add(str(int(round(float(avg)))))
            cands.add(f"{float(avg):.1f}")
            cands.add(f"{float(avg):.1f}".replace(".", ","))
        for _s, desc, _d in matched:
            d = desc or ""
            if any(c in d for c in cands):
                desc_score_ok += 1
                break
    check("Calendar: события на нужные курсы с упоминанием среднего балла",
          covered == len(courses_needing) and desc_score_ok == len(courses_needing) and len(courses_needing) > 0,
          f"covered={covered}/{len(courses_needing)}, с баллом={desc_score_ok}")


def check_email(courses_needing, n_sessions):
    print("\n=== Проверка письма ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr::text, COALESCE(body_text, '')
            FROM email.messages
            WHERE to_addr::text ILIKE %s
        """, (f"%{COORDINATOR_EMAIL}%",))
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Email: письмо координатору с верной темой и содержанием", False, str(e))
        return

    # Неприоритетно: вообще есть письмо координатору.
    check("Email: письмо координатору отправлено", len(emails) >= 1,
          f"Получателей {COORDINATOR_EMAIL}: {len(emails)}")

    # СЕМАНТИКА (критично): тема совпадает + тело упоминает число сессий и курсы.
    ok = False
    for subject, _to, body in emails:
        subj_ok = EMAIL_SUBJECT.lower() in (subject or "").lower()
        b = (body or "").lower()
        # число сессий упомянуто
        count_ok = str(n_sessions) in b
        # хотя бы про "курс"/"course"/"сесси"/"session" и упомянут >=1 нужный курс
        kw_ok = any(k in b for k in ("курс", "course", "сесси", "session", "разбор", "review"))
        course_ok = any(c.lower() in b for c in courses_needing) if courses_needing else True
        if subj_ok and count_ok and kw_ok and course_ok:
            ok = True
            break
    check("Email: письмо координатору с верной темой и содержанием", ok,
          f"тема='{EMAIL_SUBJECT}', нужно число {n_sessions} и курсы {courses_needing}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = get_expected_quiz_data()
    below_thr = [r for r in expected if r[2] is not None and float(r[2]) < FOCUS_THRESHOLD]
    courses_needing = sorted(set(r[0] for r in below_thr))
    # ВСЕ below-пороговые средние баллы по каждому нужному курсу. У курса может
    # быть несколько тестов ниже порога; task.md разрешает сослаться в описании
    # события на любой из них (тот, что послужил поводом для разбора).
    below_avgs_by_course = {}
    for name, _title, avg in below_thr:
        if avg is None:
            continue
        below_avgs_by_course.setdefault(name, []).append(avg)

    check_excel(args.agent_workspace, expected, below_thr, courses_needing)
    check_calendar(courses_needing, below_avgs_by_course)
    check_email(courses_needing, len(courses_needing))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Итог: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"КРИТИЧЕСКИЕ ПРОВАЛЫ: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (провалена критическая проверка)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
