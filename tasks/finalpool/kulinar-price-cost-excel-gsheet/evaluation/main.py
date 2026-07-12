"""
Evaluation for kulinar-price-cost-excel-gsheet.

The agent picks 5 kulinar dishes from 5 distinct categories, matches each
recipe's ingredients to grocery_prices.json (RU names, RUB prices), and builds
Recipe_Cost_Analysis.xlsx with 3 sheets (Dish Costs / Ingredient Prices /
Budget Summary) plus a Google Sheet "Cafeteria Menu Cost Analysis".

We do NOT hardcode which dishes the agent chose. Every semantic check
RECOMPUTES the expected value from the agent's own rows / the price list /
the kulinar recipe set, so any valid choice of 5 dishes passes.

Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Source of truth for kulinar recipes (globally seeded MCP data).
KULINAR_JSON = os.path.join(
    os.path.dirname(__file__),
    "..", "..", "..", "..",
    "local_servers", "kulinar-mcp", "src", "data", "all_recipes.json",
)

CRITICAL_CHECKS = {
    "Excel file exists",
    "Dish Costs sheet exists",
    "Dish Costs has 5 data rows",
    "Dishes are real kulinar recipes from >=3 categories",
    "Cost_Per_Serving == round(Total/4) for every dish",
    "Estimated_Total_Cost matches sum of matched ingredient prices",
    "Budget Summary aggregates correct",
    "Ingredient Prices: Price_Per_Unit matches grocery_prices.json",
    "Google Sheet 'Cafeteria Menu Cost Analysis' exists",
    "Google Sheet reproduces Dish Costs dish names",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def norm_name(s):
    return str(s).strip().lower() if s is not None else ""


def load_kulinar():
    with open(os.path.normpath(KULINAR_JSON), encoding="utf-8") as f:
        recipes = json.load(f)
    by_name = {r["name"]: r for r in recipes}
    name_to_cat = {r["name"]: r["category"] for r in recipes}
    return by_name, name_to_cat


def load_prices(agent_workspace):
    """Load grocery_prices.json (prefer the copy in the agent workspace)."""
    candidates = [
        os.path.join(agent_workspace, "grocery_prices.json"),
        os.path.join(os.path.dirname(__file__), "..", "initial_workspace", "grocery_prices.json"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            with open(p, encoding="utf-8") as f:
                rows = json.load(f)["prices"]
            return {r["ingredient"]: r for r in rows}
    return {}


def sheet_rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = [r for r in rows[1:] if r and any(c is not None for c in r)]
    return header, data


def col_idx(header, *aliases):
    low = [h.lower() for h in header]
    for a in aliases:
        if a.lower() in low:
            return low.index(a.lower())
    return None


def find_sheet(wb, *keywords):
    for name in wb.sheetnames:
        nl = name.lower()
        if all(k in nl for k in keywords):
            return name
    return None


def check_excel(agent_workspace):
    print("\n=== Check 1: Excel File ===")
    by_name, name_to_cat = load_kulinar()
    price_map = load_prices(agent_workspace)

    excel_path = os.path.join(agent_workspace, "Recipe_Cost_Analysis.xlsx")
    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)
    record("Excel has 3 sheets", len(wb.sheetnames) >= 3, f"Sheets: {wb.sheetnames}")

    # ---------------- Dish Costs ----------------
    dc_name = find_sheet(wb, "dish") or find_sheet(wb, "cost")
    dishes = []  # (name, ingredient_list[str], total, cps)
    if not dc_name:
        record("Dish Costs sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Dish Costs sheet exists", True)
        ws = wb[dc_name]
        header, data = sheet_rows(ws)
        record("Dish Costs has 5 data rows", len(data) == 5, f"Found {len(data)} rows")

        ci_name = col_idx(header, "Dish_Name")
        ci_ing = col_idx(header, "Ingredient_List")
        ci_tot = col_idx(header, "Estimated_Total_Cost")
        ci_cps = col_idx(header, "Cost_Per_Serving")
        record("Dish Costs has required columns",
               None not in (ci_name, ci_ing, ci_tot, ci_cps),
               f"Header: {header}")

        if None not in (ci_name, ci_ing, ci_tot, ci_cps):
            for r in data[:5]:
                dishes.append((r[ci_name], r[ci_ing], r[ci_tot], r[ci_cps]))

            # (a) dishes are real kulinar recipes from >=3 distinct categories
            real = [d[0] for d in dishes if d[0] in by_name]
            cats = {name_to_cat[n] for n in real}
            record("Dishes are real kulinar recipes from >=3 categories",
                   len(real) == 5 and len(cats) >= 3,
                   f"real={len(real)}/5, categories={sorted(cats)}")

            # (b) Cost_Per_Serving == round(total/4, 2) recomputed from the row
            cps_ok = all(num_close(cps, round(float(tot) / 4, 2), 0.5)
                         for (_, _, tot, cps) in dishes
                         if tot is not None and cps is not None)
            record("Cost_Per_Serving == round(Total/4) for every dish", cps_ok,
                   f"rows={[(d[0], d[2], d[3]) for d in dishes]}")

            # (c) Estimated_Total_Cost == sum of matched ingredient prices.
            # Recompute from the recipe + price list; matching = recipe ingredient
            # whose name is a price-list key. Tolerant to small rounding.
            if price_map:
                tot_ok = True
                detail = ""
                for (dn, _, tot, _) in dishes:
                    if dn not in by_name or tot is None:
                        continue
                    expected = round(sum(
                        price_map[i["name"]]["price_per_unit"]
                        for i in by_name[dn]["ingredients"]
                        if i["name"] in price_map
                    ), 2)
                    if not num_close(tot, expected, 1.0):
                        tot_ok = False
                        detail = f"{dn}: got {tot}, expected {expected}"
                        break
                record("Estimated_Total_Cost matches sum of matched ingredient prices",
                       tot_ok, detail)
            else:
                record("Estimated_Total_Cost matches sum of matched ingredient prices",
                       False, "price list not found")

    # ---------------- Ingredient Prices ----------------
    ip_name = find_sheet(wb, "ingredient") or find_sheet(wb, "price")
    if not ip_name:
        record("Ingredient Prices sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Ingredient Prices sheet exists", True)
        ws = wb[ip_name]
        header, data = sheet_rows(ws)
        record("Ingredient Prices has >= 10 rows", len(data) >= 10, f"Found {len(data)} rows")

        ci_in = col_idx(header, "Ingredient_Name")
        ci_pr = col_idx(header, "Price_Per_Unit")
        if None in (ci_in, ci_pr) or not price_map:
            record("Ingredient Prices: Price_Per_Unit matches grocery_prices.json",
                   False, f"header={header}, prices_loaded={bool(price_map)}")
        else:
            ok = True
            detail = ""
            checked = 0
            for r in data:
                iname = str(r[ci_in]).strip() if r[ci_in] is not None else ""
                if iname in price_map:
                    checked += 1
                    if not num_close(r[ci_pr], price_map[iname]["price_per_unit"], 0.01):
                        ok = False
                        detail = f"{iname}: got {r[ci_pr]}, expected {price_map[iname]['price_per_unit']}"
                        break
            record("Ingredient Prices: Price_Per_Unit matches grocery_prices.json",
                   ok and checked >= 5, detail or f"checked {checked} matched ingredients")

    # ---------------- Budget Summary ----------------
    bs_name = find_sheet(wb, "budget") or find_sheet(wb, "summary")
    if not bs_name:
        record("Budget Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Budget Summary sheet exists", True)
        ws = wb[bs_name]
        _, data = sheet_rows(ws)
        kv = {}
        for r in data:
            if len(r) >= 2 and r[0] is not None:
                kv[str(r[0]).strip()] = r[1]
        needed = {"Total_Cost", "Average_Cost_Per_Dish", "Cheapest_Dish", "Most_Expensive_Dish"}
        record("Budget Summary has all 4 metrics", needed.issubset(kv.keys()),
               f"keys={list(kv.keys())}")

        if dishes and needed.issubset(kv.keys()):
            try:
                costs = [(d[0], float(d[2])) for d in dishes if d[2] is not None]
            except (TypeError, ValueError):
                costs = []
            if len(costs) == 5:
                exp_total = round(sum(c for _, c in costs), 2)
                exp_avg = round(exp_total / 5, 2)
                exp_cheap = min(costs, key=lambda x: x[1])[0]
                exp_exp = max(costs, key=lambda x: x[1])[0]
                ok = (num_close(kv["Total_Cost"], exp_total, 1.0)
                      and num_close(kv["Average_Cost_Per_Dish"], exp_avg, 0.5)
                      and norm_name(kv["Cheapest_Dish"]) == norm_name(exp_cheap)
                      and norm_name(kv["Most_Expensive_Dish"]) == norm_name(exp_exp))
                record("Budget Summary aggregates correct", ok,
                       f"got total={kv['Total_Cost']} avg={kv['Average_Cost_Per_Dish']} "
                       f"cheap={kv['Cheapest_Dish']} exp={kv['Most_Expensive_Dish']}; "
                       f"expected total={exp_total} avg={exp_avg} cheap={exp_cheap} exp={exp_exp}")
            else:
                record("Budget Summary aggregates correct", False,
                       "could not parse 5 dish costs")
        else:
            record("Budget Summary aggregates correct", False,
                   "missing metrics or no dish rows")

    wb.close()
    return dishes


def check_gsheet(dishes):
    print("\n=== Check 2: Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()

        if not spreadsheets:
            record("Google Sheet 'Cafeteria Menu Cost Analysis' exists", False, "No spreadsheets")
            cur.close(); conn.close()
            return

        target = None
        for ss_id, ss_title in spreadsheets:
            tl = (ss_title or "").lower()
            if "cafeteria" in tl or ("cost" in tl and "analysis" in tl):
                target = (ss_id, ss_title)
                break
        record("Google Sheet 'Cafeteria Menu Cost Analysis' exists", target is not None,
               f"Titles: {[s[1] for s in spreadsheets]}")
        if not target:
            target = spreadsheets[0]

        cur.execute("""
            SELECT c.value
            FROM gsheet.cells c
            JOIN gsheet.sheets s ON c.sheet_id = s.id
            WHERE s.spreadsheet_id = %s
        """, (target[0],))
        cells = [row[0] for row in cur.fetchall()]
        record("Google Sheet has data", len(cells) >= 5, f"Found {len(cells)} cells")

        # Cross-source consistency: the 5 dish names from Dish Costs must appear
        # among the Google Sheet cells.
        all_text = " ".join(str(c) for c in cells if c is not None)
        if dishes:
            found = sum(1 for d in dishes
                        if d[0] and str(d[0]).strip() and str(d[0]).strip() in all_text)
            record("Google Sheet reproduces Dish Costs dish names", found >= 5,
                   f"{found}/5 dish names found in gsheet")
        else:
            record("Google Sheet reproduces Dish Costs dish names", False,
                   "no dish rows parsed from Excel")

        cur.close(); conn.close()
    except Exception as e:
        record("Google Sheet 'Cafeteria Menu Cost Analysis' exists", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    dishes = check_excel(args.agent_workspace)
    check_gsheet(dishes or [])

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
