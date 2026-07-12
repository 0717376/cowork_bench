"""Evaluation for insales-product-sales-ranking (InSales)."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def fail_critical(msg):
    print(f"CRITICAL FAIL: {msg}")
    print("\n=== RESULT: FAIL (critical check) ===")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Product_Rankings.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Product_Rankings.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    total_checks = 0
    passed_checks = 0

    # ---- Load Product Rankings ----
    a_pr = load_sheet_rows(agent_wb, "Product Rankings")
    g_pr = load_sheet_rows(gt_wb, "Product Rankings")

    a_pr_data = a_pr[1:] if a_pr and len(a_pr) > 1 else []
    g_pr_data = [r for r in (g_pr[1:] if g_pr and len(g_pr) > 1 else []) if r and r[0] is not None]

    # ================= CRITICAL CHECKS =================
    # CRITICAL 1: Product Rankings must be the correct top-20 set ranked by
    # Units_Sold DESCENDING. Verify the top-1 / top-3 product names sit in the
    # correct ROW positions (not just set membership).
    if a_pr is None:
        fail_critical("Sheet 'Product Rankings' not found in agent output")
    a_pr_clean = [r for r in a_pr_data if r and r[0] is not None]
    if len(a_pr_clean) < 20:
        fail_critical(f"Product Rankings has fewer than 20 product rows ({len(a_pr_clean)})")

    def name_prefix(s, n=40):
        return str(s).strip().lower()[:n]

    gt_top3 = [name_prefix(g_pr_data[i][0]) for i in range(3)]
    agent_top3 = [name_prefix(a_pr_clean[i][0]) for i in range(3)]
    if agent_top3 != gt_top3:
        fail_critical(
            "Top-3 rows are not in correct descending-by-Units_Sold order. "
            f"agent={agent_top3} gt={gt_top3}"
        )

    # CRITICAL 2: Units_Sold for the top-5 products must match GT within tol=2
    # (ranking signal must come from live wc.* sales data, not fabricated).
    for i in range(5):
        g_units = to_float(g_pr_data[i][3])
        a_units = to_float(a_pr_clean[i][3]) if len(a_pr_clean[i]) > 3 else None
        if not num_close(a_units, g_units, 2):
            fail_critical(
                f"Top-{i+1} Units_Sold mismatch: agent={a_units} vs gt={g_units} (tol=2)"
            )

    # CRITICAL 3: Est_Revenue must equal round(Price*Units_Sold, 2) recomputed
    # from that row's OWN Price and Units_Sold (proves agent computed, not copied)
    # — sample the first 5 agent rows.
    for i in range(min(5, len(a_pr_clean))):
        row = a_pr_clean[i]
        price = to_float(row[1]) if len(row) > 1 else None
        units = to_float(row[3]) if len(row) > 3 else None
        rev = to_float(row[4]) if len(row) > 4 else None
        if price is None or units is None or rev is None:
            fail_critical(f"Row {i+1} missing Price/Units_Sold/Est_Revenue for revenue recompute")
        expected = round(price * units, 2)
        if abs(rev - expected) > 1.0:
            fail_critical(
                f"Row {i+1} Est_Revenue not consistent with own Price*Units_Sold: "
                f"got {rev}, expected {expected}"
            )

    # CRITICAL 4: Summary totals must match within tight tolerance.
    a_sum = load_sheet_rows(agent_wb, "Summary")
    g_sum = load_sheet_rows(gt_wb, "Summary")
    if a_sum is None:
        fail_critical("Sheet 'Summary' not found in agent output")

    def sum_lookup(rows):
        d = {}
        for r in (rows[1:] if rows and len(rows) > 1 else []):
            if r and r[0] is not None:
                d[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None
        return d

    a_sl = sum_lookup(a_sum)
    g_sl = sum_lookup(g_sum)

    crit_summary = {
        "top_20_total_units": 2.0,
        "top_20_total_revenue": 5.0,
        "avg_price_top20": 1.0,
    }
    for key, tol in crit_summary.items():
        if key not in g_sl:
            continue
        if key not in a_sl:
            fail_critical(f"Summary missing row '{key}'")
        if not num_close(a_sl[key], g_sl[key], tol):
            fail_critical(
                f"Summary {key}: agent={a_sl[key]} vs gt={g_sl[key]} (tol={tol})"
            )

    # CRITICAL 5: Sales_Rankings_Brief.docx must exist and contain real content.
    # (Naming the actual top-3 products is checked as a NON-critical accuracy
    # signal below: the frozen groundtruth docx is a generic stub, so requiring
    # the product names as CRITICAL would false-fail the GT-vs-GT invariant.)
    docx_path = os.path.join(args.agent_workspace, "Sales_Rankings_Brief.docx")
    if not os.path.exists(docx_path):
        fail_critical("Sales_Rankings_Brief.docx not found")

    # ================= NON-CRITICAL (accuracy) CHECKS =================
    # Product Rankings: per-row field accuracy keyed by product name.
    print("  Checking Product Rankings (accuracy)...")
    a_lookup = {}
    for row in a_pr_clean:
        a_lookup[str(row[0]).strip().lower()] = row
    for g_row in g_pr_data:
        key = str(g_row[0]).strip().lower()
        a_row = a_lookup.get(key)
        total_checks += 4  # Price, Stock, Units_Sold, Est_Revenue
        if a_row is None:
            all_errors.append(f"Missing row: {g_row[0]}")
            continue
        if len(a_row) > 1 and len(g_row) > 1:
            if num_close(a_row[1], g_row[1], 0.5):
                passed_checks += 1
            else:
                all_errors.append(f"{key}.Price: {a_row[1]} vs {g_row[1]} (tol=0.5)")
        if len(a_row) > 2 and len(g_row) > 2:
            if num_close(a_row[2], g_row[2], 2):
                passed_checks += 1
            else:
                all_errors.append(f"{key}.Stock: {a_row[2]} vs {g_row[2]} (tol=2)")
        if len(a_row) > 3 and len(g_row) > 3:
            if num_close(a_row[3], g_row[3], 2):
                passed_checks += 1
            else:
                all_errors.append(f"{key}.Units_Sold: {a_row[3]} vs {g_row[3]} (tol=2)")
        if len(a_row) > 4 and len(g_row) > 4:
            if num_close(a_row[4], g_row[4], 5.0):
                passed_checks += 1
            else:
                all_errors.append(f"{key}.Est_Revenue: {a_row[4]} vs {g_row[4]} (tol=5.0)")

    # Summary: per-metric accuracy (including Highest_Revenue_Product str-match).
    print("  Checking Summary (accuracy)...")
    for key, gval in g_sl.items():
        total_checks += 1
        aval = a_sl.get(key)
        if aval is None:
            all_errors.append(f"Summary missing row: {key}")
            continue
        if num_close(aval, gval, 50.0):
            passed_checks += 1
        else:
            all_errors.append(f"{key}.Value: {aval} vs {gval}")

    # docx (non-critical): some descriptive text present + names the top-3 products.
    if os.path.exists(docx_path):
        total_checks += 1  # structural: enough text
        try:
            from docx import Document as _DocCheck2
            _doc2 = _DocCheck2(docx_path)
            _text2 = " ".join(p.text for p in _doc2.paragraphs).strip()
            if len(_text2) >= 50:
                passed_checks += 1
            else:
                all_errors.append("Sales_Rankings_Brief.docx has too little text (< 50 chars)")
            # top-3 product naming (substring on distinctive prefixes)
            total_checks += 1
            _lt = _text2.lower()
            top3_tokens = ["nihara", "jbl flip 4", "limbani"]
            present = [t for t in top3_tokens if t in _lt]
            if len(present) >= 2:
                passed_checks += 1
            else:
                all_errors.append(
                    f"Sales_Rankings_Brief.docx does not name the top-3 products "
                    f"(found {present} of {top3_tokens})"
                )
        except ImportError:
            if os.path.getsize(docx_path) >= 100:
                passed_checks += 1
            else:
                all_errors.append("Sales_Rankings_Brief.docx too small")
        except Exception as _e:
            all_errors.append(f"Error reading Sales_Rankings_Brief.docx: {_e}")

    # ================= ACCURACY GATE =================
    accuracy = (passed_checks / total_checks * 100.0) if total_checks else 0.0
    print(f"\nAccuracy: {passed_checks}/{total_checks} = {accuracy:.1f}%")
    if all_errors:
        print(f"Non-critical errors ({len(all_errors)}):")
        for e in all_errors[:10]:
            print(f"  {e}")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
