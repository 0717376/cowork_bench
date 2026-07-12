"""
Evaluation script for gcal-canvas-office-hours task (RU stack: forms MCP).

Checks:
1. Excel file Office_Hours_Schedule.xlsx - Bookings sheet (6 rows) and Summary sheet
   (incl. per-date Total_Bookings and distinct Time_Slots).
2. Google Calendar events - one per unique (date, time_slot) titled
   "Office Hours: <topic>" with the booking students' names in the description.
3. Emails sent - per-recipient confirmation email with that student's own
   date + time slot in the body (RU or EN date form accepted).

Source of truth: the "Office Hours Booking" form responses (gform.* schema,
backed by the RU forms MCP). The form carries Russian student names with
ASCII-translit emails; dates/times/topics are English identifiers.

CRITICAL_CHECKS: any critical failure => task FAIL (sys.exit(1)) regardless of
accuracy. Non-critical checks must reach accuracy >= 70 to PASS.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")
        if critical:
            CRITICAL_FAILED.append(name)


# Map an English date string ("March 9, 2026") -> list of acceptable substrings
# the agent may have written (English literal, RU "9 марта", ISO 2026-03-09).
EN_MONTH_TO_RU = {"march": "марта"}
DATE_ISO = {
    "March 9, 2026": "2026-03-09",
    "March 10, 2026": "2026-03-10",
    "March 11, 2026": "2026-03-11",
    "March 12, 2026": "2026-03-12",
    "March 13, 2026": "2026-03-13",
}
# Time slot -> acceptable body substrings (12h and 24h).
TIME_VARIANTS = {
    "9:00 AM": ["9:00", "09:00"],
    "10:00 AM": ["10:00"],
    "11:00 AM": ["11:00"],
    "2:00 PM": ["2:00", "14:00"],
}


def date_variants(date_en):
    """Acceptable lowercase substrings for a given English date string."""
    variants = [date_en.lower()]
    # "march 9" without year
    parts = date_en.replace(",", "").split()  # [March, 9, 2026]
    if len(parts) >= 2:
        month, day = parts[0].lower(), parts[1]
        variants.append(f"{month} {day}")              # march 9
        ru_month = EN_MONTH_TO_RU.get(month, month)
        variants.append(f"{day} {ru_month}")           # 9 марта
    iso = DATE_ISO.get(date_en)
    if iso:
        variants.append(iso)
    return variants


def body_has_date(body_lower, date_en):
    return any(v in body_lower for v in date_variants(date_en))


def body_has_time(body_lower, time_slot):
    return any(v in body_lower for v in TIME_VARIANTS.get(time_slot, [time_slot.lower()]))


def get_form_responses():
    """Read all form responses from the database."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT q.id, q.title
        FROM gform.questions q
        JOIN gform.forms f ON q.form_id = f.id
        WHERE LOWER(f.title) LIKE '%office hours%booking%'
        ORDER BY q.position
    """)
    question_map = {row[0]: row[1] for row in cur.fetchall()}

    cur.execute("""
        SELECT r.answers
        FROM gform.responses r
        JOIN gform.forms f ON r.form_id = f.id
        WHERE LOWER(f.title) LIKE '%office hours%booking%'
    """)
    raw_responses = cur.fetchall()

    cur.close()
    conn.close()

    q_name_id = q_email_id = q_date_id = q_time_id = q_topic_id = None
    for qid, qtitle in question_map.items():
        tl = qtitle.lower()
        if "name" in tl:
            q_name_id = qid
        elif "email" in tl:
            q_email_id = qid
        elif "date" in tl:
            q_date_id = qid
        elif "time" in tl or "slot" in tl:
            q_time_id = qid
        elif "topic" in tl:
            q_topic_id = qid

    responses = []
    for (answers_raw,) in raw_responses:
        answers = json.loads(answers_raw) if isinstance(answers_raw, str) else answers_raw
        responses.append({
            "name": answers.get(q_name_id, ""),
            "email": answers.get(q_email_id, ""),
            "date": answers.get(q_date_id, ""),
            "time_slot": answers.get(q_time_id, ""),
            "topic": answers.get(q_topic_id, ""),
        })

    return responses


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, responses):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Office_Hours_Schedule.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        # Cannot validate substance without the file.
        check("Bookings sheet has all 6 form respondents", False,
              "Excel file missing", critical=True)
        check("Summary sheet has correct per-date totals and slots", False,
              "Excel file missing", critical=True)
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        check("Bookings sheet has all 6 form respondents", False, str(e), critical=True)
        check("Summary sheet has correct per-date totals and slots", False, str(e), critical=True)
        return

    check("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    check("Sheet 'Bookings' exists", any("booking" in s for s in sheet_names_lower),
          f"Found: {wb.sheetnames}")
    check("Sheet 'Summary' exists", any("summary" in s for s in sheet_names_lower),
          f"Found: {wb.sheetnames}")

    # --- Bookings sheet ---
    print("\n--- Bookings Sheet ---")
    ws_bookings = next((wb[s] for s in wb.sheetnames if "booking" in s.lower()), None)

    bookings_ok = True
    if ws_bookings is None:
        bookings_ok = False
    else:
        data_rows = [r for r in ws_bookings.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in r)]
        check("Bookings has 6 data rows", len(data_rows) == 6, f"Got {len(data_rows)}")
        if len(data_rows) != 6:
            bookings_ok = False

        headers = [str(c.value).strip().lower() if c.value else "" for c in ws_bookings[1]]
        for eh in ["student_name", "student_email", "preferred_date",
                   "preferred_time_slot", "topic"]:
            found = any(eh.replace("_", "") in h.replace("_", "").replace(" ", "")
                        for h in headers)
            check(f"Header '{eh}' present", found, f"Headers: {headers}")
            if not found:
                bookings_ok = False

        # CRITICAL: every respondent's name + email + date + time + topic appears
        # somewhere in the Bookings rows (substance of the schedule).
        row_blobs = []
        for row in data_rows:
            row_blobs.append(" ".join(str(c) for c in row if c is not None).lower())
        all_blob = " ".join(row_blobs)

        for resp in responses:
            ok = (
                resp["name"].strip().lower() in all_blob
                and resp["email"].strip().lower() in all_blob
                and resp["date"].lower() in all_blob
                and resp["time_slot"].lower() in all_blob
                and resp["topic"].strip().lower() in all_blob
            )
            if not ok:
                bookings_ok = False

    check("Bookings sheet has all 6 form respondents", bookings_ok,
          "Some respondent name/email/date/time/topic missing from Bookings",
          critical=True)

    # --- Summary sheet ---
    print("\n--- Summary Sheet ---")
    ws_summary = next((wb[s] for s in wb.sheetnames if "summary" in s.lower()), None)

    # Expected per-date totals and distinct slots.
    date_counts = {}
    date_slots = {}
    for r in responses:
        date_counts[r["date"]] = date_counts.get(r["date"], 0) + 1
        date_slots.setdefault(r["date"], set()).add(r["time_slot"])
    unique_dates = set(date_counts)

    summary_ok = True
    if ws_summary is None:
        summary_ok = False
    else:
        data_rows = [r for r in ws_summary.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in r)]
        check("Summary has correct number of date rows",
              len(data_rows) == len(unique_dates),
              f"Expected {len(unique_dates)}, got {len(data_rows)}")
        if len(data_rows) != len(unique_dates):
            summary_ok = False

        headers = [str(c.value).strip().lower() if c.value else "" for c in ws_summary[1]]
        for eh in ["date", "total", "slot"]:
            found = any(eh in h for h in headers)
            check(f"Summary header contains '{eh}'", found, f"Headers: {headers}")

        # Validate each expected date row: total count + distinct slots.
        matched_dates = set()
        for row in data_rows:
            if not row or row[0] is None:
                continue
            row_blob = " ".join(str(c) for c in row if c is not None).lower()
            date_val = str(row[0]).strip().lower()
            matched = None
            for d in date_counts:
                if d.lower() in row_blob or any(v in date_val for v in date_variants(d)):
                    matched = d
                    break
            if not matched:
                summary_ok = False
                continue
            matched_dates.add(matched)
            # total count present and correct
            total_ok = False
            for c in row[1:]:
                try:
                    if int(float(str(c))) == date_counts[matched]:
                        total_ok = True
                        break
                except (ValueError, TypeError):
                    continue
            if not total_ok:
                summary_ok = False
            # all distinct slots for that date listed in the row text
            for slot in date_slots[matched]:
                if slot.lower() not in row_blob:
                    summary_ok = False

        if matched_dates != unique_dates:
            summary_ok = False

    check("Summary sheet has correct per-date totals and slots", summary_ok,
          "Per-date Total_Bookings or Time_Slots incorrect/missing", critical=True)


# ============================================================================
# Check 2: Google Calendar events
# ============================================================================

def check_gcal(responses):
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")
    for ev in events:
        print(f"  Event: {ev[0]} | {ev[2]} - {ev[3]}")

    # Build expected per-slot data: (date, time_slot) -> {topics, names}
    slots = {}
    for r in responses:
        key = (r["date"], r["time_slot"])
        s = slots.setdefault(key, {"topics": set(), "names": []})
        s["topics"].add(r["topic"])
        s["names"].append(r["name"])
    unique_slots = set(slots)

    valid_dates = {"2026-03-09", "2026-03-10", "2026-03-11", "2026-03-12", "2026-03-13"}

    # Non-critical structural checks.
    oh_events = [e for e in events if "office hours:" in (e[0] or "").lower()]
    check("Events titled 'Office Hours:'", len(oh_events) >= 1,
          f"{len(oh_events)} events start with 'Office Hours:'")

    in_range = 0
    dur_ok = 0
    for summary, description, start_dt, end_dt in events:
        if start_dt and start_dt.strftime("%Y-%m-%d") in valid_dates:
            in_range += 1
        if start_dt and end_dt:
            mins = (end_dt - start_dt).total_seconds() / 60
            if 25 <= mins <= 35:
                dur_ok += 1
    check("Events fall on March 9-13, 2026", in_range >= len(unique_slots),
          f"{in_range} in range, expected {len(unique_slots)}")
    check("Events have ~30 minute duration", dur_ok >= len(unique_slots),
          f"{dur_ok} events ~30min")

    check("Exactly one event per unique (date, time slot)",
          len(events) == len(unique_slots),
          f"Found {len(events)}, expected {len(unique_slots)}")

    # CRITICAL: for each unique slot there is an event whose title is
    # "Office Hours: <one of the slot's topics>" AND whose description lists the
    # booking student name(s) for that slot.
    gcal_ok = True
    for (date_en, time_slot), info in slots.items():
        matched = False
        for summary, description, start_dt, end_dt in events:
            title = (summary or "").lower()
            desc = (description or "").lower()
            if "office hours:" not in title:
                continue
            # event must be on the right date
            if not start_dt or start_dt.strftime("%Y-%m-%d") != DATE_ISO.get(date_en):
                continue
            # title carries one of this slot's topics
            if not any(t.lower() in title for t in info["topics"]):
                continue
            # description lists every booking student name for this slot
            if all(n.lower() in desc for n in info["names"]):
                matched = True
                break
        if not matched:
            gcal_ok = False
            print(f"    [gcal] no matching event for {date_en} {time_slot} "
                  f"(topics={info['topics']}, names={info['names']})")

    check("Each unique slot has an event with correct topic in title and "
          "student names in description", gcal_ok,
          "A slot is missing its topic-titled event or student names in description",
          critical=True)


# ============================================================================
# Check 3: Emails
# ============================================================================

def check_emails(responses):
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(all_emails)} total emails.")

    def to_list(to_addr):
        out = []
        if isinstance(to_addr, list):
            out = [str(a).lower() for a in to_addr]
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                out = [str(a).lower() for a in parsed] if isinstance(parsed, list) \
                    else [str(to_addr).lower()]
            except (json.JSONDecodeError, TypeError):
                out = [str(to_addr).lower()]
        return out

    confirmation = [e for e in all_emails if "confirmation" in (e[0] or "").lower()]
    check("Emails have 'Confirmation' in subject", len(confirmation) >= len(responses),
          f"{len(confirmation)} confirmation-subject emails")

    ta = [e for e in confirmation if "ta@university.edu" in (str(e[1]) or "").lower()]
    check("Emails sent from ta@university.edu", len(ta) >= len(responses),
          f"{len(ta)} from ta@university.edu")

    # CRITICAL: each booked student gets a confirmation email (from ta@university.edu,
    # subject contains "confirmation") whose body contains that student's own
    # booked date AND time slot (RU or EN date form accepted).
    emails_ok = True
    for resp in responses:
        match = False
        for subject, from_addr, to_addr, body_text in all_emails:
            if "confirmation" not in (subject or "").lower():
                continue
            if "ta@university.edu" not in (str(from_addr) or "").lower():
                continue
            if resp["email"].lower() not in " ".join(to_list(to_addr)):
                continue
            body = (body_text or "").lower()
            if body_has_date(body, resp["date"]) and body_has_time(body, resp["time_slot"]):
                match = True
                break
        if not match:
            emails_ok = False
            print(f"    [email] no correct confirmation for {resp['name']} "
                  f"<{resp['email']}> ({resp['date']} {resp['time_slot']})")

    check("Each student got a confirmation email with their own date and time slot",
          emails_ok,
          "A student is missing a confirmation email or it lacks their date/time",
          critical=True)


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    responses = get_form_responses()
    check("Form responses available (6)", len(responses) == 6,
          f"Got {len(responses)}", critical=True)
    if not responses:
        print("\nNo form responses -> cannot evaluate.")
        if args.res_log_file:
            with open(args.res_log_file, "w") as f:
                json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                           "accuracy": 0.0, "success": False}, f, indent=2)
        sys.exit(1)

    check_excel(args.agent_workspace, responses)
    check_gcal(responses)
    check_emails(responses)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = 100.0 * PASS_COUNT / total if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILED:
        print(f"  CRITICAL FAILED: {CRITICAL_FAILED}")

    success = (not CRITICAL_FAILED) and accuracy >= 70.0
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "passed": PASS_COUNT,
                "failed": FAIL_COUNT,
                "accuracy": round(accuracy, 1),
                "critical_failed": CRITICAL_FAILED,
                "success": success,
            }, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
