"""Evaluation for moex-portfolio-allocation-excel-gcal (russified)."""
import os
import argparse, os, sys
import psycopg2


# Per-ticker weights fixed in task.md / Portfolio_Guidelines.pdf (sum = 100).
WEIGHTS = {
    "SBER.ME": 20.0,
    "LKOH.ME": 20.0,
    "TCSG.ME": 20.0,
    "MGNT.ME": 15.0,
    "GAZP.ME": 15.0,
    "MTSS.ME": 10.0,
}
# Dominant analyst bucket (period '0m') from moex.recommendations grid, mapped to literals.
RECS = {
    "SBER.ME": "buy",
    "LKOH.ME": "buy",
    "TCSG.ME": "buy",
    "MGNT.ME": "buy",
    "GAZP.ME": "hold",
    "MTSS.ME": "hold",
}
TICKERS = set(WEIGHTS.keys())
STRONG_BUY_COUNT = sum(1 for v in RECS.values() if v == "strong_buy")  # 0
BUY_COUNT = sum(1 for v in RECS.values() if v == "buy")                 # 4


def load_prices():
    """Prices the MOEX tool actually returns: latest close per ticker from
    moex.stock_prices (pg_adapter.info overrides currentPrice/regularMarketPrice
    with float(latest close)). Recomputed live so GT tracks the tool, not the
    stale stock_info JSONB values."""
    conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
                            user="eigent", password="camel")
    cur = conn.cursor()
    prices = {}
    for t in WEIGHTS:
        cur.execute("SELECT close FROM moex.stock_prices WHERE symbol = %s ORDER BY date DESC LIMIT 1", (t,))
        r = cur.fetchone()
        if r and r[0] is not None:
            prices[t] = float(r[0])
    cur.close(); conn.close()
    return prices


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def norm_sym(s):
    return str(s).strip().upper()


def check_pdf(agent_workspace):
    errors = []
    pdf_path = os.path.join(agent_workspace, "Portfolio_Guidelines.pdf")
    if not os.path.exists(pdf_path):
        errors.append("Portfolio_Guidelines.pdf not found in agent workspace")
    return errors


def check_excel(agent_workspace):
    """Returns (errors, critical_errors)."""
    errors, critical = [], []
    import openpyxl
    try:
        PRICES = load_prices()
    except Exception as e:
        return [f"Could not load MOEX prices: {e}"], [f"Could not load MOEX prices: {e}"]
    WEIGHTED_AVG_PRICE = round(sum(PRICES[s] * WEIGHTS[s] / 100 for s in PRICES), 2)
    path = os.path.join(agent_workspace, "Portfolio_Allocation.xlsx")
    if not os.path.exists(path):
        return ["Portfolio_Allocation.xlsx not found"], ["Portfolio_Allocation.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # ---- Stock Analysis sheet ----
        rows = load_sheet_rows(wb, "Stock Analysis")
        if rows is None:
            errors.append("Sheet 'Stock Analysis' not found")
            critical.append("Sheet 'Stock Analysis' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            header = [str(c).strip() if c is not None else "" for c in rows[0]] if rows else []
            # column index map (by header name, case-insensitive)
            hmap = {h.strip().lower(): i for i, h in enumerate(header)}

            symbols = {norm_sym(r[0]) for r in data_rows}
            # Non-critical: exactly the tracked tickers present
            if symbols != TICKERS:
                errors.append(f"Stock Analysis symbols={sorted(symbols)}, expected {sorted(TICKERS)}")

            # Build per-symbol row lookup
            by_sym = {norm_sym(r[0]): r for r in data_rows}

            # CRITICAL: each tracked ticker present with correct weight; weights sum to 100
            w_idx = hmap.get("allocated_weight_pct", 4)
            total_w = 0.0
            for sym, exp_w in WEIGHTS.items():
                r = by_sym.get(sym)
                if not r:
                    critical.append(f"Stock Analysis missing ticker {sym}")
                    continue
                w = r[w_idx] if len(r) > w_idx else None
                if w is None or not num_close(w, exp_w, 0.6):
                    critical.append(f"{sym} Allocated_Weight_Pct={w}, expected {exp_w}")
                else:
                    try:
                        total_w += float(w)
                    except (TypeError, ValueError):
                        pass
            if not num_close(total_w, 100.0, 0.6):
                critical.append(f"Allocated_Weight_Pct sum={total_w}, expected 100")

            # CRITICAL: Recommendation per ticker == dominant analyst bucket
            rec_idx = hmap.get("recommendation", 3)
            for sym, exp_rec in RECS.items():
                r = by_sym.get(sym)
                if not r:
                    continue
                val = r[rec_idx] if len(r) > rec_idx else None
                if val is None or str(val).strip().lower() != exp_rec:
                    critical.append(f"{sym} Recommendation={val}, expected {exp_rec}")

            # Non-critical: Price column reasonably matches seed prices
            p_idx = hmap.get("price", 1)
            for sym, exp_p in PRICES.items():
                r = by_sym.get(sym)
                if not r:
                    continue
                p = r[p_idx] if len(r) > p_idx else None
                if p is None or not num_close(p, exp_p, max(0.5, exp_p * 0.02)):
                    errors.append(f"{sym} Price={p}, expected ~{exp_p}")

            # Non-critical: alphabetical sort by Symbol
            ordered = [norm_sym(r[0]) for r in data_rows]
            if ordered != sorted(ordered):
                errors.append(f"Stock Analysis not sorted alphabetically by Symbol: {ordered}")

        # ---- Allocation Summary sheet ----
        rows2 = load_sheet_rows(wb, "Allocation Summary")
        if rows2 is None:
            errors.append("Sheet 'Allocation Summary' not found")
            critical.append("Sheet 'Allocation Summary' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows2 if r[0]}

            # CRITICAL: Total_Invested
            if "total_invested" in lookup:
                if not num_close(lookup["total_invested"], 100000, 1.0):
                    critical.append(f"Total_Invested={lookup['total_invested']}, expected 100000")
            else:
                critical.append("Total_Invested not found in Allocation Summary")

            # CRITICAL: Strong_Buy_Count and Buy_Count from MOEX dominant-bucket rule
            if "strong_buy_count" in lookup:
                if not num_close(lookup["strong_buy_count"], STRONG_BUY_COUNT, 0):
                    critical.append(f"Strong_Buy_Count={lookup['strong_buy_count']}, expected {STRONG_BUY_COUNT}")
            else:
                critical.append("Strong_Buy_Count not found in Allocation Summary")

            if "buy_count" in lookup:
                if not num_close(lookup["buy_count"], BUY_COUNT, 0):
                    critical.append(f"Buy_Count={lookup['buy_count']}, expected {BUY_COUNT}")
            else:
                critical.append("Buy_Count not found in Allocation Summary")

            # CRITICAL: Weighted_Avg_Price computed from MOEX seed prices & weights
            if "weighted_avg_price" in lookup:
                if not num_close(lookup["weighted_avg_price"], WEIGHTED_AVG_PRICE, 1.0):
                    critical.append(f"Weighted_Avg_Price={lookup['weighted_avg_price']}, expected {WEIGHTED_AVG_PRICE}")
            else:
                critical.append("Weighted_Avg_Price not found in Allocation Summary")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
        critical.append(f"Error reading Excel: {e}")
    return errors, critical


def check_gcal():
    """Returns (errors, critical_errors)."""
    errors, critical = [], []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime FROM gcal.events
            WHERE start_datetime::date = '2026-04-30'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No GCal event found on 2026-04-30")
            critical.append("No GCal event found on 2026-04-30")
        else:
            summaries = [(r[0] or "").lower() for r in rows]
            # Accept RU ('ребаланс'/'портфел') or EN ('portfolio'/'rebalanc')
            ev = None
            for r in rows:
                s = (r[0] or "").lower()
                if ("portfolio" in s or "rebalanc" in s or "ребаланс" in s or "портфел" in s):
                    ev = r
                    break
            if ev is None:
                errors.append(f"No portfolio rebalancing event on 2026-04-30 (found: {[r[0] for r in rows]})")
                critical.append("No portfolio rebalancing GCal event on 2026-04-30")
            else:
                # CRITICAL: description mentions tracked tickers + 100000 total
                desc = (ev[1] or "")
                desc_l = desc.lower()
                missing = [t for t in TICKERS if t.lower() not in desc_l]
                if missing:
                    critical.append(f"GCal description missing tickers: {sorted(missing)}")
                if "100000" not in desc.replace(",", "").replace(" ", "") and "100 000" not in desc and "100,000" not in desc:
                    critical.append("GCal description does not mention 100000 portfolio total")
    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
        critical.append(f"Error checking GCal: {e}")
    return errors, critical


def check_email():
    """Returns (errors, critical_errors)."""
    errors, critical = [], []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr FROM email.messages
            WHERE to_addr::text ILIKE '%portfolio_manager@wealth.com%'
            ORDER BY id DESC LIMIT 10
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No email found to portfolio_manager@wealth.com")
            critical.append("No email found to portfolio_manager@wealth.com")
        else:
            subjects = [(r[0] or "").strip() for r in rows]
            exact = "Portfolio Allocation Plan - April 2026"
            if not any(s == exact for s in subjects):
                critical.append(f"No email with exact subject '{exact}' (found: {subjects})")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
        critical.append(f"Error checking email: {e}")
    return errors, critical


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []
    critical_errors = []

    print("  Checking PDF in workspace...")
    errs = check_pdf(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Excel file...")
    errs, crit = check_excel(agent_ws)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")
    critical_errors.extend(crit)

    print("  Checking GCal event...")
    errs, crit = check_gcal()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")
    critical_errors.extend(crit)

    print("  Checking email...")
    errs, crit = check_email()
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]: print(f"    ERROR: {e}")
    else:
        print("    PASS")
    critical_errors.extend(crit)

    # ---- CRITICAL gate: any semantic failure => hard FAIL before accuracy ----
    if critical_errors:
        print(f"\n=== CRITICAL FAIL ({len(critical_errors)} critical errors) ===")
        for e in critical_errors[:10]: print(f"  CRITICAL: {e}")
        sys.exit(1)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
