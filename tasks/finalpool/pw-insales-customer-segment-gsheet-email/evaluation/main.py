"""Evaluation script for pw-insales-customer-segment-gsheet-email."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Groundtruth benchmark Market_Avg_Price per (russified) category, scraped from the
# mock benchmark page at localhost:30338 (Product_Category column is russified via
# scripts/wc_relabel_map.py so the wc RU categories join the benchmark rows).
BENCHMARK_MARKET_PRICE = {
    "Электроника": 58.07,
    "ТВ и домашний кинотеатр": 289.65,
    "Аудио": 74.15,
    "Камеры": 22.71,
    "Бытовая техника": 33.98,
    "Часы": 57.66,
}
# Expected russified wc category set (product_categories.name after central seed).
EXPECTED_CATEGORIES = {
    "Аудио", "Камеры", "Электроника", "Бытовая техника",
    "ТВ и домашний кинотеатр", "Часы",
}
# Groundtruth Total_Products = sum of per-category Product_Count over the 6 rows.
EXPECTED_TOTAL_PRODUCTS = 82

CRITICAL_FAILURES = []

def critical_fail(name, detail=""):
    detail_str = str(detail)[:200] if detail else ""
    CRITICAL_FAILURES.append(name)
    print(f"  [CRITICAL-FAIL] {name}: {detail_str}")

def critical_pass(name):
    print(f"  [CRITICAL-PASS] {name}")

def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def _norm_cat(v):
    return str(v).strip() if v is not None else ""

def read_data_analysis_rows(ws):
    """Return list of dicts keyed by Category from the Data_Analysis sheet, mapping
    header name (lowercased) -> cell value, regardless of column order/sorting."""
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    idx = {h.lower(): i for i, h in enumerate(headers)}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in r):
            continue
        d = {}
        for h, i in idx.items():
            d[h] = r[i] if i < len(r) else None
        rows.append(d)
    return rows, idx

def run_critical_checks(agent_workspace, wb):
    """Five SEMANTIC critical checks. Any failure => task FAIL (sys.exit(1)) before
    the accuracy gate. These prove the agent read real wc product data, scraped+joined
    the benchmark, computed the gap, and sent the specific notification email."""

    # --- Excel-derived critical checks (need Data_Analysis) -------------------
    da_rows = None
    if wb is not None and "Data_Analysis" in wb.sheetnames:
        da_rows, _idx = read_data_analysis_rows(wb["Data_Analysis"])

    # 1) Total_Products sums to 82 (groundtruth) over the category rows.
    if da_rows is None:
        critical_fail("Total_Products sums to 82", "Data_Analysis sheet missing")
    else:
        total = 0.0
        ok_counts = True
        for d in da_rows:
            pc = safe_float(d.get("product_count"))
            if pc is None:
                ok_counts = False
            else:
                total += pc
        if ok_counts and abs(total - EXPECTED_TOTAL_PRODUCTS) < 0.5:
            critical_pass("Total_Products sums to 82")
        else:
            critical_fail("Total_Products sums to 82",
                          f"sum of Product_Count = {total} (expected {EXPECTED_TOTAL_PRODUCTS})")

    # 2) At least one row's Market_Avg_Price equals the benchmark value for its
    #    matching russified category (within tolerance) -> benchmark was scraped+joined.
    if da_rows is None:
        critical_fail("Market_Avg_Price matches benchmark", "Data_Analysis sheet missing")
    else:
        matched = 0
        for d in da_rows:
            cat = _norm_cat(d.get("category"))
            mp = safe_float(d.get("market_avg_price"))
            exp = BENCHMARK_MARKET_PRICE.get(cat)
            if exp is not None and mp is not None and abs(mp - exp) <= max(0.05, exp * 0.02):
                matched += 1
        if matched >= 1:
            critical_pass("Market_Avg_Price matches benchmark")
        else:
            critical_fail("Market_Avg_Price matches benchmark",
                          "no row's Market_Avg_Price matched the benchmark page values")

    # 3) Category column contains the russified wc category names (set overlap).
    if da_rows is None:
        critical_fail("Category column uses russified wc category names", "Data_Analysis sheet missing")
    else:
        cats = {_norm_cat(d.get("category")) for d in da_rows}
        overlap = cats & EXPECTED_CATEGORIES
        if len(overlap) >= 4:
            critical_pass("Category column uses russified wc category names")
        else:
            critical_fail("Category column uses russified wc category names",
                          f"overlap with expected RU categories = {sorted(overlap)}")

    # 4) Price_Gap_Pct ~= (Our_Avg_Price - Market_Avg_Price)/Market_Avg_Price*100
    #    for each row (within tolerance) -> gap actually computed, not stubbed.
    if da_rows is None:
        critical_fail("Price_Gap_Pct internally consistent", "Data_Analysis sheet missing")
    else:
        checked = 0
        bad = 0
        for d in da_rows:
            our = safe_float(d.get("our_avg_price"))
            mkt = safe_float(d.get("market_avg_price"))
            gap = safe_float(d.get("price_gap_pct"))
            if our is None or mkt is None or gap is None or mkt == 0:
                continue
            checked += 1
            expected_gap = (our - mkt) / mkt * 100.0
            if abs(gap - expected_gap) > 1.0:  # 1 percentage-point tolerance
                bad += 1
        if checked >= 1 and bad == 0:
            critical_pass("Price_Gap_Pct internally consistent")
        else:
            critical_fail("Price_Gap_Pct internally consistent",
                          f"checked={checked} inconsistent={bad}")

    # 5) An email with subject EXACTLY 'Analysis Report Complete' sent to team-lead@company.com.
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT to_addr FROM email.messages WHERE TRIM(subject) = %s",
            ('Analysis Report Complete',))
        rows = cur.fetchall()
        conn.close()
        found = False
        for (to_addr,) in rows:
            blob = json.dumps(to_addr) if not isinstance(to_addr, str) else to_addr
            if 'team-lead@company.com' in blob:
                found = True
                break
        if found:
            critical_pass("Email 'Analysis Report Complete' to team-lead@company.com")
        else:
            critical_fail("Email 'Analysis Report Complete' to team-lead@company.com",
                          f"{len(rows)} subject-matching emails, none to team-lead@company.com")
    except Exception as e:
        critical_fail("Email 'Analysis Report Complete' to team-lead@company.com", str(e))

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    excel_path = os.path.join(agent_workspace, "Customer_Segment_Report.xlsx")
    check("Customer_Segment_Report.xlsx exists", os.path.exists(excel_path))
    wb = None
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Category', 'Product_Count', 'Our_Avg_Price', 'Total_Sales', 'Market_Avg_Price', 'Price_Gap_Pct']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            # Check headers
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Category']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT subject FROM email.messages WHERE subject ILIKE %s OR subject ILIKE %s",
                        ('%report%', '%analysis%'))
            emails = cur.fetchall()
            check("Analysis email sent", len(emails) >= 1, f"found {len(emails)} matching emails")
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT title FROM gsheet.spreadsheets WHERE title ILIKE %s", ('%tracker%',))
            sheets = cur.fetchall()
            check("Google Sheet created", len(sheets) >= 1, f"found {len(sheets)} sheets")
            conn.close()
        except Exception as e:
            check("GSheet check", False, str(e))

        check("wc_segment_processor.py exists", os.path.exists(os.path.join(agent_workspace, "wc_segment_processor.py")))

    # --- CRITICAL semantic checks -------------------------------------------
    print("--- Critical checks ---")
    run_critical_checks(agent_workspace, wb)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0

    if CRITICAL_FAILURES:
        print(f"CRITICAL FAILURE(S): {CRITICAL_FAILURES}")
        print(f"Passed {PASS_COUNT}/{total} structural checks; accuracy {accuracy:.1f}%")
        sys.exit(1)

    success = accuracy >= 70.0
    return success, f"Passed {PASS_COUNT}/{total} checks; accuracy {accuracy:.1f}%"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
