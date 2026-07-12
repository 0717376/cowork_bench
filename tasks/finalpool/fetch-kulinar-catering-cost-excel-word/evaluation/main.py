"""Evaluation for fetch-kulinar-catering-cost-excel-word (russified -> kulinar).

Проверяет Excel-отчёт Catering_Cost_Report.xlsx (3 листа), Word-документ и
наличие cook_catering_processor.py.

Семантические (CRITICAL) проверки — любой их провал => общий FAIL независимо
от accuracy:
  - рецепты в Data_Analysis действительно берутся из базы kulinar
    (названия совпадают с all_recipes.json) — агент реально обращался к
    свопнутому MCP;
  - столбец Meets_Guidelines внутренне согласован с Protein_g относительно
    дневной нормы по белку (Protein daily_g), потреблённой из data.json;
  - сводные показатели на листе Metrics пересчитаны из листа Data_Analysis
    (Avg_Calories ~ среднее Calories; Recipes_Meeting_Guidelines == число
    строк со значением «Да»/Yes), а не захардкожены.

Прочие (структурные) проверки — некритичные. Порог прохождения: accuracy >= 70%.
"""
import argparse
import glob as globmod
import json
import os
import re
import sys
import urllib.request

import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Дневная норма по белку из источника нормативов питания (data.json).
# Используется fallback, если эндпоинт недоступен в момент оценки.
PROTEIN_DAILY_G_DEFAULT = 50
DATA_URL = "http://localhost:30319/api/data.json"

CRITICAL_CHECKS = {
    "Catering_Cost_Report.xlsx exists",
    "Recipes originate from kulinar DB",
    "Meets_Guidelines consistent with Protein_g vs daily norm",
    "Metrics: Avg_Calories recomputed from Data_Analysis",
    "Metrics: Recipes_Meeting_Guidelines == count of 'Yes' rows",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def _norm_name(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


def yes_no(val):
    """Возвращает True для «Да»/Yes, False для «Нет»/No, None — иначе."""
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("да", "yes", "y", "true", "1", "+"):
        return True
    if s in ("нет", "no", "n", "false", "0", "-"):
        return False
    return None


def load_kulinar_names():
    """Множество нормализованных названий рецептов kulinar или пустое множество."""
    candidates = []
    env = os.environ.get("KULINAR_RECIPES_JSON")
    if env:
        candidates.append(env)
    here = os.path.abspath(__file__)
    cur = os.path.dirname(here)
    rel = os.path.join("local_servers", "kulinar-mcp", "src", "data", "all_recipes.json")
    for _ in range(12):
        candidates.append(os.path.join(cur, rel))
        parent = os.path.dirname(cur)
        if parent == cur:
            break
        cur = parent
    for path in candidates:
        try:
            if path and os.path.exists(path):
                with open(path, encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list) and data:
                    return {_norm_name(r["name"]) for r in data}
        except Exception:
            continue
    return set()


def fetch_protein_daily_g():
    """Дневная норма по белку из data.json; fallback на дефолт."""
    try:
        with urllib.request.urlopen(DATA_URL, timeout=4) as resp:
            data = json.loads(resp.read().decode("utf-8"))
        for g in data.get("nutrition_guidelines", []):
            if str(g.get("nutrient", "")).strip().lower() == "protein":
                v = safe_float(g.get("daily_g"))
                if v is not None:
                    return v
    except Exception:
        pass
    return PROTEIN_DAILY_G_DEFAULT


def col_index(headers_lower, name):
    name = name.lower()
    for i, h in enumerate(headers_lower):
        if h == name:
            return i
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES.clear()

    kulinar_names = load_kulinar_names()
    protein_norm = fetch_protein_daily_g()

    excel_path = os.path.join(agent_workspace, "Catering_Cost_Report.xlsx")
    check("Catering_Cost_Report.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return _finalize(res_log_file)

    wb = openpyxl.load_workbook(excel_path, data_only=True)

    # ----- Data_Analysis -----
    da_rows = []          # list of dicts
    check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
    if "Data_Analysis" in wb.sheetnames:
        ws = wb["Data_Analysis"]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        raw_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                    if any(c is not None for c in r)]
        check("Data_Analysis has >= 5 rows", len(raw_rows) >= 5, f"got {len(raw_rows)}")
        for expected_col in ['Recipe', 'Category', 'Calories', 'Protein_g', 'Meets_Guidelines']:
            check(f"Data_Analysis has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")

        i_recipe = col_index(headers, 'recipe')
        i_cal = col_index(headers, 'calories')
        i_prot = col_index(headers, 'protein_g')
        i_meets = col_index(headers, 'meets_guidelines')
        for r in raw_rows:
            da_rows.append({
                "recipe": r[i_recipe] if i_recipe is not None and i_recipe < len(r) else None,
                "calories": safe_float(r[i_cal]) if i_cal is not None and i_cal < len(r) else None,
                "protein": safe_float(r[i_prot]) if i_prot is not None and i_prot < len(r) else None,
                "meets": yes_no(r[i_meets]) if i_meets is not None and i_meets < len(r) else None,
            })

        # CRITICAL: рецепты из базы kulinar
        if kulinar_names and da_rows:
            agent_names = {_norm_name(d["recipe"]) for d in da_rows if d["recipe"]}
            unknown = agent_names - kulinar_names
            ok = len(da_rows) >= 5 and bool(agent_names) and not unknown
            check("Recipes originate from kulinar DB", ok,
                  f"неизвестные названия: {sorted(unknown)[:5]}")
        else:
            check("Recipes originate from kulinar DB", False,
                  "не удалось загрузить базу kulinar или нет строк")

        # CRITICAL: Meets_Guidelines согласован с Protein_g vs нормой по белку
        if da_rows:
            consistent = True
            checked_any = False
            for d in da_rows:
                if d["protein"] is None or d["meets"] is None:
                    consistent = False
                    break
                checked_any = True
                expected = d["protein"] >= protein_norm
                if d["meets"] != expected:
                    consistent = False
                    break
            check("Meets_Guidelines consistent with Protein_g vs daily norm",
                  consistent and checked_any,
                  f"норма белка={protein_norm}")
        else:
            check("Meets_Guidelines consistent with Protein_g vs daily norm", False, "нет строк")

    # ----- Metrics -----
    metrics = {}
    check("Metrics sheet exists", "Metrics" in wb.sheetnames)
    if "Metrics" in wb.sheetnames:
        ws = wb["Metrics"]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        m_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                  if any(c is not None for c in r)]
        check("Metrics has >= 4 rows", len(m_rows) >= 4, f"got {len(m_rows)}")
        for expected_col in ['Metric', 'Value']:
            check(f"Metrics has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")
        for r in m_rows:
            if r and r[0] is not None:
                metrics[str(r[0]).strip().lower().replace(" ", "_")] = r[1] if len(r) > 1 else None

    # CRITICAL: Avg_Calories пересчитан из Data_Analysis
    cals = [d["calories"] for d in da_rows if d["calories"] is not None]
    if cals and "avg_calories" in metrics:
        expected_avg = round(sum(cals) / len(cals))
        got = safe_float(metrics.get("avg_calories"))
        ok = got is not None and abs(got - expected_avg) <= max(1.0, expected_avg * 0.02)
        check("Metrics: Avg_Calories recomputed from Data_Analysis", ok,
              f"expected ~{expected_avg}, got {got}")
    else:
        check("Metrics: Avg_Calories recomputed from Data_Analysis", False,
              f"keys={list(metrics.keys())}, cals={len(cals)}")

    # CRITICAL: Recipes_Meeting_Guidelines == число строк «Да»/Yes
    meet_key = None
    for k in metrics:
        if "meeting" in k or ("recipes" in k and "guidelin" in k):
            meet_key = k
            break
    if meet_key and da_rows:
        expected_cnt = sum(1 for d in da_rows if d["meets"] is True)
        got = safe_float(metrics.get(meet_key))
        ok = got is not None and int(got) == expected_cnt
        check("Metrics: Recipes_Meeting_Guidelines == count of 'Yes' rows", ok,
              f"expected {expected_cnt}, got {got}")
    else:
        check("Metrics: Recipes_Meeting_Guidelines == count of 'Yes' rows", False,
              f"keys={list(metrics.keys())}")

    # ----- Recommendations -----
    check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
    if "Recommendations" in wb.sheetnames:
        ws = wb["Recommendations"]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        rec_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                    if any(c is not None for c in r)]
        check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")
        for expected_col in ['Priority', 'Action']:
            check(f"Recommendations has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")

    # ----- Word -----
    word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
    check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
    if word_files:
        from docx import Document
        doc = Document(word_files[0])
        text = " ".join(p.text for p in doc.paragraphs)
        low = text.lower()
        check("Word has content", len(text) > 50, f"text length: {len(text)}")
        # Три раздела: принимаем RU и EN маркеры (некритично)
        sections = {
            "summary": ("резюме", "summary", "обзор", "overview"),
            "findings": ("вывод", "findings", "находк", "результат"),
            "recommendations": ("рекомендац", "recommendation"),
        }
        missing = [s for s, kws in sections.items() if not any(k in low for k in kws)]
        check("Word has summary/findings/recommendations sections", not missing,
              f"не найдено: {missing}")

    check("cook_catering_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "cook_catering_processor.py")))

    return _finalize(res_log_file)


def _finalize(res_log_file):
    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        return False, "No checks performed"
    accuracy = PASS_COUNT / total * 100
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump(result, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
