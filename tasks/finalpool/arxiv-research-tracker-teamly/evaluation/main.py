"""
Evaluation script for arxiv-research-tracker-teamly task.

Checks:
1. Excel file (Transformer_Research_Tracker.xlsx) exists and is readable
2. "Paper Comparison" sheet has EXACTLY the 5 target papers (no noise),
   correct Titles, Citation_Counts (tight tol), sorted by Citation_Count desc
3. "Statistics" sheet has required metrics with correct values
4. Teamly knowledge-base page exists with substantive research-tracker content

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Pass threshold otherwise: accuracy >= 70%.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
# These reflect the task's substance (correct papers/values + real KB page),
# not mere structure.
CRITICAL_CHECKS = {
    "Paper Comparison contains exactly the 5 target papers (no noise)",
    "Citation counts exact for all 5 papers",
    "Statistics: Most_Cited_Paper is LoRA",
    "Statistics: Top_Venue is ICLR",
    "Statistics: Total_Papers = 5",
    "Paper Comparison sorted by Citation_Count descending",
    "Teamly tracker page has substantive paper/citation content",
}

# Expected target paper data
EXPECTED_PAPERS = {
    "2402.10001": {
        "title": "FlashAttention-3: Fast and Memory-Efficient Attention with IO-Awareness",
        "citation_count": 520,
        "venue": "NeurIPS",
        "primary_category": "cs.LG",
    },
    "2402.10002": {
        "title": "Efficient Transformers via Token Merging and Pruning",
        "citation_count": 180,
        "venue": "ICLR",
        "primary_category": "cs.CV",
    },
    "2402.10003": {
        "title": "LoRA: Low-Rank Adaptation of Large Language Models",
        "citation_count": 1200,
        "venue": "ICLR",
        "primary_category": "cs.CL",
    },
    "2402.10004": {
        "title": "Quantization-Aware Training for Efficient Transformer Inference",
        "citation_count": 95,
        "venue": "ICML",
        "primary_category": "cs.LG",
    },
    "2402.10005": {
        "title": "Sparse Mixture of Experts for Scalable Transformer Models",
        "citation_count": 340,
        "venue": "JMLR",
        "primary_category": "cs.AI",
    },
}

EXPECTED_PAPER_IDS = set(EXPECTED_PAPERS.keys())

# Noise paper IDs that should NOT appear
NOISE_PAPER_IDS = {"2402.10010", "2402.10011"}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=50):
    """Compare two numeric values with tolerance."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    """Load all rows from a sheet (case-insensitive name lookup)."""
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            matched = name
            break
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            matched = name
            break
    if matched is None:
        return None
    ws = wb[matched]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def find_column_index(header_row, possible_names):
    """Find column index by trying multiple possible header names (case-insensitive)."""
    if header_row is None:
        return None
    for i, cell_val in enumerate(header_row):
        if cell_val is None:
            continue
        cell_lower = str(cell_val).strip().lower().replace(" ", "_")
        for name in possible_names:
            if name.lower().replace(" ", "_") == cell_lower:
                return i
    return None


def check_excel(agent_workspace):
    """Check Transformer_Research_Tracker.xlsx."""
    print("\n=== Checking Excel Output ===")

    excel_path = os.path.join(agent_workspace, "Transformer_Research_Tracker.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    record("Excel file readable", True)

    all_ok = True

    # ── Check Paper Comparison sheet ─────────────────────────────────────────
    pc_rows = load_sheet_rows(wb, "Paper Comparison")
    if pc_rows is None:
        pc_rows = load_sheet_rows(wb, "Paper_Comparison")

    if pc_rows is None:
        record("Sheet 'Paper Comparison' exists", False, f"Available sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Paper Comparison' exists", True)

        header = pc_rows[0] if pc_rows else []
        data_rows = pc_rows[1:] if len(pc_rows) > 1 else []

        # Find column indices
        id_col = find_column_index(header, ["Paper_ID", "Paper ID", "PaperID", "ID", "paper_id"])
        title_col = find_column_index(header, ["Title", "title", "Paper_Title"])
        citation_col = find_column_index(header, [
            "Citation_Count", "Citation Count", "Citations", "citation_count", "CitationCount"
        ])
        venue_col = find_column_index(header, ["Venue", "venue", "Publication_Venue"])
        category_col = find_column_index(header, [
            "Primary_Category", "Primary Category", "Category", "primary_category"
        ])

        # Check row count
        record("Paper Comparison has 5 data rows",
               len(data_rows) == 5,
               f"Found {len(data_rows)} data rows, expected 5")
        if len(data_rows) != 5:
            all_ok = False

        # Check Paper IDs — exact set of the 5 targets, no noise (CRITICAL)
        if id_col is not None:
            found_ids = set()
            id_order = []
            for row in data_rows:
                if id_col < len(row) and row[id_col] is not None:
                    pid = str(row[id_col]).strip()
                    found_ids.add(pid)
                    id_order.append(pid)

            exact_set = (found_ids == EXPECTED_PAPER_IDS)
            record("Paper Comparison contains exactly the 5 target papers (no noise)",
                   exact_set,
                   f"Found IDs: {sorted(found_ids)}, expected: {sorted(EXPECTED_PAPER_IDS)}")
            if not exact_set:
                all_ok = False

            # Per-id presence (non-critical, informational)
            for expected_id in EXPECTED_PAPER_IDS:
                present = expected_id in found_ids
                record(f"Paper ID {expected_id} present", present,
                       f"Not found. Found IDs: {found_ids}")
                if not present:
                    all_ok = False

            for noise_id in NOISE_PAPER_IDS:
                absent = noise_id not in found_ids
                record(f"Noise paper {noise_id} absent", absent,
                       "Noise paper should not be in the tracker")
                if not absent:
                    all_ok = False
        else:
            record("Paper_ID column found", False, f"Header: {header}")
            record("Paper Comparison contains exactly the 5 target papers (no noise)",
                   False, "Paper_ID column missing")
            all_ok = False

        # Check titles
        if title_col is not None:
            found_titles = set()
            for row in data_rows:
                if title_col < len(row) and row[title_col] is not None:
                    found_titles.add(str(row[title_col]).strip().lower())

            for pid, pdata in EXPECTED_PAPERS.items():
                title_lower = pdata["title"].lower()
                present = any(title_lower in t or t in title_lower for t in found_titles)
                record(f"Title present: {pdata['title'][:50]}...", present,
                       "Title not found in Paper Comparison sheet")
                if not present:
                    all_ok = False
        else:
            record("Title column found", False, f"Header: {header}")
            all_ok = False

        # Check citation counts — tight tolerance, source is deterministic.
        if citation_col is not None and id_col is not None:
            all_cites_ok = True
            cite_seq = []  # (paper_id, citation) in row order, for sort check
            for row in data_rows:
                if id_col < len(row) and row[id_col] is not None:
                    paper_id = str(row[id_col]).strip()
                    actual_cite = row[citation_col] if citation_col < len(row) else None
                    if paper_id in EXPECTED_PAPERS:
                        expected_cite = EXPECTED_PAPERS[paper_id]["citation_count"]
                        ok = num_close(actual_cite, expected_cite, tol=5)
                        record(f"Citation count for {paper_id}", ok,
                               f"Got {actual_cite}, expected {expected_cite} (tol=5)")
                        if not ok:
                            all_ok = False
                            all_cites_ok = False
                    try:
                        cite_seq.append((paper_id, float(actual_cite)))
                    except (TypeError, ValueError):
                        cite_seq.append((paper_id, None))

            # CRITICAL aggregate: every citation value correct
            record("Citation counts exact for all 5 papers", all_cites_ok,
                   "One or more citation counts deviate from source")
            if not all_cites_ok:
                all_ok = False

            # CRITICAL: rows sorted by Citation_Count strictly descending
            vals = [c for _, c in cite_seq if c is not None]
            sorted_ok = (len(vals) == len(cite_seq) and
                         all(vals[i] >= vals[i + 1] for i in range(len(vals) - 1)))
            record("Paper Comparison sorted by Citation_Count descending",
                   sorted_ok,
                   f"Citation order: {[c for _, c in cite_seq]}")
            if not sorted_ok:
                all_ok = False
        else:
            record("Citation_Count column found", False, f"Header: {header}")
            record("Citation counts exact for all 5 papers", False,
                   "Citation_Count column missing")
            record("Paper Comparison sorted by Citation_Count descending", False,
                   "Citation_Count column missing")
            all_ok = False

    # ── Check Statistics sheet ───────────────────────────────────────────────
    stats_rows = load_sheet_rows(wb, "Statistics")
    if stats_rows is None:
        record("Sheet 'Statistics' exists", False, f"Available sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Statistics' exists", True)

        # Build metrics dict
        metrics = {}
        for row in stats_rows[1:]:  # skip header
            if row and row[0] is not None:
                key = str(row[0]).strip().lower().replace(" ", "_")
                val = row[1] if len(row) > 1 else None
                metrics[key] = val

        # Check Total_Papers
        total_key = None
        for k in metrics:
            if "total" in k and "paper" in k:
                total_key = k
                break
        if total_key:
            ok = num_close(metrics[total_key], 5, tol=0)
            record("Statistics: Total_Papers = 5", ok,
                   f"Got {metrics[total_key]}")
            if not ok:
                all_ok = False
        else:
            record("Statistics: Total_Papers metric exists", False,
                   f"Available metrics: {list(metrics.keys())}")
            all_ok = False

        # Check Avg_Citations
        avg_key = None
        for k in metrics:
            if "avg" in k and "cit" in k:
                avg_key = k
                break
        expected_avg = (520 + 180 + 1200 + 95 + 340) / 5  # 467.0
        if avg_key:
            ok = num_close(metrics[avg_key], expected_avg, tol=1)
            record("Statistics: Avg_Citations", ok,
                   f"Got {metrics[avg_key]}, expected ~{expected_avg} (tol=1)")
            if not ok:
                all_ok = False
        else:
            record("Statistics: Avg_Citations metric exists", False,
                   f"Available metrics: {list(metrics.keys())}")
            all_ok = False

        # Check Most_Cited_Paper
        most_cited_key = None
        for k in metrics:
            if "most" in k and "cit" in k:
                most_cited_key = k
                break
        if most_cited_key:
            val = str(metrics[most_cited_key]).lower() if metrics[most_cited_key] else ""
            ok = "lora" in val or "low-rank" in val or "low rank" in val
            record("Statistics: Most_Cited_Paper is LoRA", ok,
                   f"Got '{metrics[most_cited_key]}'")
            if not ok:
                all_ok = False
        else:
            record("Statistics: Most_Cited_Paper metric exists", False,
                   f"Available metrics: {list(metrics.keys())}")
            all_ok = False

        # Check Top_Venue
        top_venue_key = None
        for k in metrics:
            if "top" in k and "venue" in k:
                top_venue_key = k
                break
        if top_venue_key:
            val = str(metrics[top_venue_key]).lower() if metrics[top_venue_key] else ""
            ok = "iclr" in val  # ICLR appears twice
            record("Statistics: Top_Venue is ICLR", ok,
                   f"Got '{metrics[top_venue_key]}'")
            if not ok:
                all_ok = False
        else:
            record("Statistics: Top_Venue metric exists", False,
                   f"Available metrics: {list(metrics.keys())}")
            all_ok = False

        # Check Date_Range exists
        date_key = None
        for k in metrics:
            if "date" in k and "range" in k:
                date_key = k
                break
        if date_key:
            val = str(metrics[date_key]) if metrics[date_key] else ""
            ok = len(val) > 5  # just check it has some content
            record("Statistics: Date_Range has content", ok,
                   f"Got '{val}'")
            if not ok:
                all_ok = False
        else:
            record("Statistics: Date_Range metric exists", False,
                   f"Available metrics: {list(metrics.keys())}")
            all_ok = False

    return all_ok


def check_teamly():
    """Check that a Teamly knowledge-base page exists with substantive,
    research-tracker content."""
    print("\n=== Checking Teamly Page ===")

    all_ok = True

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        # Find the tracker page by title (English title is preserved per task.md).
        cur.execute("""
            SELECT id, title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE '%%efficient transformers%%'
               OR title ILIKE '%%research%%tracker%%'
               OR title ILIKE '%%трекер%%'
        """)
        pages = cur.fetchall()

        if not pages:
            cur.execute("SELECT COUNT(*) FROM teamly.pages")
            total = cur.fetchone()[0]
            record("Teamly tracker page exists", False,
                   f"Found {total} total pages but none matching the expected title")
            record("Teamly tracker page has substantive paper/citation content", False,
                   "No matching tracker page")
            cur.close()
            conn.close()
            return False

        record("Teamly tracker page exists", True)

        # Combined title + body text of matching pages (original case -> lower).
        combined_text = " ".join(
            (str(t) + " " + str(b)).lower() for _, t, b in pages
        )

        # Structural: page should not be near-empty.
        max_len = max(len(str(b)) for _, _, b in pages)
        record("Teamly tracker page has non-trivial body",
               max_len >= 100,
               f"Longest matching page body is {max_len} chars")
        if max_len < 100:
            all_ok = False

        # CRITICAL semantic: must mention several of the 5 paper titles AND
        # carry citation/venue info — not just the word 'transformer'.
        paper_title_keywords = [
            "flashattention",
            "token merging",
            "lora",
            "low-rank",
            "quantization",
            "mixture of experts",
        ]
        title_hits = sum(1 for kw in paper_title_keywords if kw in combined_text)

        venue_keywords = ["neurips", "iclr", "icml", "jmlr"]
        has_venue = any(v in combined_text for v in venue_keywords)
        # Citation evidence: a known citation count or the word цитир/citation.
        cite_keywords = ["цитир", "citation", "1200", "520", "340", "180", "95"]
        has_citation = any(c in combined_text for c in cite_keywords)

        substantive = (title_hits >= 3) and (has_venue or has_citation)
        record("Teamly tracker page has substantive paper/citation content",
               substantive,
               f"paper-title hits={title_hits} (need >=3), "
               f"venue={has_venue}, citation={has_citation}")
        if not substantive:
            all_ok = False

        cur.close()
        conn.close()

    except Exception as e:
        record("Teamly database accessible", False, str(e))
        record("Teamly tracker page has substantive paper/citation content", False, str(e))
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
