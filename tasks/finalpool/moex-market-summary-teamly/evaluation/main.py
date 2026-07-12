"""
Evaluation script for moex-market-summary-teamly task.

Checks:
1. Excel Market_Overview.xlsx with Stock Summary sheet (5 RU MOEX tickers,
   live values computed from moex.stock_prices / moex.stock_info)
2. Teamly page titled "Market Dashboard" exists AND its summary text names
   the best/worst 30d tickers with their return percentages.

Critical checks (see CRITICAL_CHECKS): any failure => overall FAIL regardless
of accuracy. Otherwise pass threshold: accuracy >= 70%.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

SYMBOLS = ['GAZP.ME', 'LKOH.ME', 'MGNT.ME', 'MTSS.ME', 'SBER.ME']

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic critical checks: any failure => overall FAIL regardless of accuracy.
# Built dynamically (per-ticker) in run_evaluation, plus the Teamly summary checks.
CRITICAL_CHECKS = set()


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        d = (str(detail)[:300]) if detail else ""
        print(f"  [FAIL] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    symbols = SYMBOLS
    in_clause = "('GAZP.ME','LKOH.ME','MGNT.ME','MTSS.ME','SBER.ME')"

    # Latest prices (use the global max date across the tracked tickers)
    cur.execute(f"""
        SELECT symbol, close FROM moex.stock_prices
        WHERE date = (SELECT MAX(date) FROM moex.stock_prices WHERE symbol IN {in_clause})
        AND symbol IN {in_clause} ORDER BY symbol
    """)
    latest = {r[0]: float(r[1]) for r in cur.fetchall()}

    # 30 trading days ago (rn = 31)
    cur.execute(f"""
        WITH ranked AS (
            SELECT symbol, date, close, ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY date DESC) as rn
            FROM moex.stock_prices WHERE symbol IN {in_clause}
        )
        SELECT symbol, close FROM ranked WHERE rn = 31 ORDER BY symbol
    """)
    price_30d = {r[0]: float(r[1]) for r in cur.fetchall()}

    # 60 trading days ago (rn = 61) -- within the ~65-row moex history window
    cur.execute(f"""
        WITH ranked AS (
            SELECT symbol, date, close, ROW_NUMBER() OVER(PARTITION BY symbol ORDER BY date DESC) as rn
            FROM moex.stock_prices WHERE symbol IN {in_clause}
        )
        SELECT symbol, close FROM ranked WHERE rn = 61 ORDER BY symbol
    """)
    price_60d = {r[0]: float(r[1]) for r in cur.fetchall()}

    # Stock info
    cur.execute(f"""
        SELECT symbol, data->>'shortName', data->>'sector', data->>'longName'
        FROM moex.stock_info
        WHERE symbol IN {in_clause} ORDER BY symbol
    """)
    info = {r[0]: (r[1], r[2], r[3]) for r in cur.fetchall()}

    conn.close()

    rows = []
    long_names = {}
    for sym in symbols:
        lp = latest.get(sym, 0)
        p30 = price_30d.get(sym, 0)
        p60 = price_60d.get(sym, 0)
        ret_30 = round((lp - p30) / p30 * 100, 2) if p30 else 0
        ret_60 = round((lp - p60) / p60 * 100, 2) if p60 else 0
        name, sector, long_name = info.get(sym, ("", "", ""))
        long_names[sym] = long_name
        rows.append((sym, name, sector, lp, p30, ret_30, p60, ret_60))

    # Best/worst 30d
    best = max(rows, key=lambda r: r[5])
    worst = min(rows, key=lambda r: r[5])

    return {"stocks": rows, "best_30d": best, "worst_30d": worst,
            "long_names": long_names}


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Market_Overview.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        return

    try:
        wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return

    check("Sheet 'Stock Summary' exists", get_sheet(wb, "Stock Summary") is not None,
          f"Found: {wb.sheetnames}")

    ws = get_sheet(wb, "Stock Summary")
    if ws and expected:
        agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
        exp = expected["stocks"]
        check("Stock Summary row count", len(agent_rows) == len(exp),
              f"Expected {len(exp)}, got {len(agent_rows)}")

        agent_by_sym = {}
        for row in agent_rows:
            if row and row[0]:
                agent_by_sym[str(row[0]).strip().upper()] = row

        # Exactly the 5 chosen RU tickers present, no missing/extra (critical, structural-semantic)
        check("Exactly the 5 RU tickers present (no missing/extra)",
              set(agent_by_sym.keys()) == set(SYMBOLS),
              f"Got: {sorted(agent_by_sym.keys())}")

        for exp_row in exp:
            sym = exp_row[0]
            agent_row = agent_by_sym.get(sym)
            if agent_row:
                long_name = expected.get("long_names", {}).get(sym)
                check(f"'{sym}' Company_Name",
                      str_match(agent_row[1], exp_row[1])
                      or str_match(agent_row[1], long_name),
                      f"Expected {exp_row[1]} or {long_name}, got {agent_row[1]}")
                check(f"'{sym}' Sector",
                      str_match(agent_row[2], exp_row[2]),
                      f"Expected {exp_row[2]}, got {agent_row[2]}")
                check(f"'{sym}' Latest_Price",
                      num_close(agent_row[3], exp_row[3], 2.0),
                      f"Expected {exp_row[3]}, got {agent_row[3]}")
                check(f"'{sym}' Price_30d_Ago",
                      num_close(agent_row[4], exp_row[4], 2.0),
                      f"Expected {exp_row[4]}, got {agent_row[4]}")
                check(f"'{sym}' Return_30d_Pct",
                      num_close(agent_row[5], exp_row[5], 1.0),
                      f"Expected {exp_row[5]}, got {agent_row[5]}")
                check(f"'{sym}' Price_60d_Ago",
                      num_close(agent_row[6], exp_row[6], 2.0),
                      f"Expected {exp_row[6]}, got {agent_row[6]}")
                check(f"'{sym}' Return_60d_Pct",
                      num_close(agent_row[7], exp_row[7], 1.5),
                      f"Expected {exp_row[7]}, got {agent_row[7]}")
            else:
                check(f"'{sym}' found in output", False, "Not in agent output")

        # Check sort order (alphabetical by symbol)
        if len(agent_rows) >= 2:
            syms = [str(r[0]).strip().upper() for r in agent_rows if r and r[0]]
            check("Sorted by Symbol alphabetically",
                  syms == sorted(syms),
                  f"Got: {syms}")


def check_teamly(expected):
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for teamly check", False, str(e))
        return

    cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
    pages = cur.fetchall()
    conn.close()

    check("At least one Teamly page exists", len(pages) > 0,
          f"Found {len(pages)} pages")

    # Find the Market Dashboard page (title or body mentions it)
    dashboard = None
    for title, body in pages:
        blob = f"{title or ''} {body or ''}".lower()
        if "market dashboard" in blob:
            dashboard = (title or "", body or "")
            break

    check("Teamly page 'Market Dashboard' found", dashboard is not None,
          f"Pages: {[t for t, _ in pages][:10]}")

    if dashboard is None or not expected:
        return

    text = f"{dashboard[0]} {dashboard[1]}".lower()

    best = expected["best_30d"]
    worst = expected["worst_30d"]
    best_sym = best[0].lower()
    worst_sym = worst[0].lower()
    best_pct = best[5]
    worst_pct = worst[5]

    def pct_present(blob, pct):
        # Accept the percentage with 1 or 2 decimals, with optional sign.
        cands = {f"{pct:.2f}", f"{pct:.1f}", f"{abs(pct):.2f}", f"{abs(pct):.1f}"}
        return any(c in blob for c in cands)

    # Critical: best ticker symbol + its return percentage present in summary
    check("Teamly summary names best-30d ticker",
          best_sym in text,
          f"Expected '{best_sym}' in summary")
    check("Teamly summary contains best-30d return percentage",
          pct_present(text, best_pct),
          f"Expected '{best_pct}' in summary")

    # Critical: worst ticker symbol + its return percentage present in summary
    check("Teamly summary names worst-30d ticker",
          worst_sym in text,
          f"Expected '{worst_sym}' in summary")
    check("Teamly summary contains worst-30d return percentage",
          pct_present(text, worst_pct),
          f"Expected '{worst_pct}' in summary")


def check_excel_gt(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel (vs groundtruth) ===")
    agent_file = os.path.join(agent_workspace, "Market_Overview.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Market_Overview.xlsx")
    check("Excel file exists", os.path.isfile(agent_file))
    check("Groundtruth file exists", os.path.isfile(gt_file))
    if not os.path.isfile(agent_file) or not os.path.isfile(gt_file):
        return
    agent_wb = openpyxl.load_workbook(agent_file)
    gt_wb = openpyxl.load_workbook(gt_file)
    check("Sheet 'Stock Summary' exists", get_sheet(agent_wb, "Stock Summary") is not None)
    a_ws = get_sheet(agent_wb, "Stock Summary")
    g_ws = get_sheet(gt_wb, "Stock Summary")
    if a_ws and g_ws:
        a_rows = list(a_ws.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_ws.iter_rows(min_row=2, values_only=True))
        check("Row count matches", len(a_rows) == len(g_rows),
              f"Expected {len(g_rows)}, got {len(a_rows)}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()

    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        # Build dynamic critical checks (live financial deliverables)
        for sym in SYMBOLS:
            CRITICAL_CHECKS.add(f"'{sym}' Latest_Price")
            CRITICAL_CHECKS.add(f"'{sym}' Price_30d_Ago")
            CRITICAL_CHECKS.add(f"'{sym}' Return_30d_Pct")
            CRITICAL_CHECKS.add(f"'{sym}' Sector")
        CRITICAL_CHECKS.add("Exactly the 5 RU tickers present (no missing/extra)")
        CRITICAL_CHECKS.add("Teamly page 'Market Dashboard' found")
        CRITICAL_CHECKS.add("Teamly summary names best-30d ticker")
        CRITICAL_CHECKS.add("Teamly summary contains best-30d return percentage")
        CRITICAL_CHECKS.add("Teamly summary names worst-30d ticker")
        CRITICAL_CHECKS.add("Teamly summary contains worst-30d return percentage")
        check_excel(agent_workspace, expected)
    else:
        print("INFO: Falling back to groundtruth Excel")
        check_excel_gt(agent_workspace, groundtruth_workspace)

    check_teamly(expected)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {critical_failed}")
    all_ok = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "critical_failed": critical_failed,
            "success": all_ok,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return all_ok, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
