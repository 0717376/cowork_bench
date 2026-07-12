"""Evaluation for insales-inventory-reorder-forecast (InSales, RU).

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Pass threshold otherwise: accuracy >= 70%.
"""
import argparse
import datetime
import json
import os
import re
import sys

import psycopg2

# Russified store category -> supplier lead time (days), matching the
# re-keyed mock_pages/api/lead_times.json that the agent must fetch and join.
CATEGORY_LEAD_TIME = {
    "Аудио": 14,
    "Электроника": 21,
    "Камеры": 7,
    "ТВ и домашний кинотеатр": 28,
    "Часы": 10,
    "Бытовая техника": 18,
    "Колонки": 14,
    "Наушники": 14,
}

# Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Lead_Time per row matches russified category->lead_time map",
    "Reorder Schedule == products with Needs_Reorder=Yes (set equality)",
    "Every Reorder_Date is tomorrow-or-future and matches its gcal event start",
    "Each reorder gcal event summary/description links to real store data",
    "Avg_Daily_Sales consistent with Current_Stock & Days_Until_Safety_Stock",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v).strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
    if not m:
        return None
    try:
        return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    except ValueError:
        return None


def get_gcal_events():
    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432,
        dbname=os.environ.get("PGDATABASE", "cowork_gym"),
        user="eigent", password="camel",
    )
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        WHERE summary ILIKE '%reorder%'
        ORDER BY start_datetime
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def evaluate(agent_ws, gt_data, today):
    import openpyxl

    path = os.path.join(agent_ws, "Inventory_Forecast.xlsx")
    if not os.path.exists(path):
        record("Inventory_Forecast.xlsx exists", False, "file not found")
        # everything else depends on the file; record critical fails so we bail
        for c in CRITICAL_CHECKS:
            record(c, False, "no workbook")
        return
    record("Inventory_Forecast.xlsx exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # ---- Stock Analysis sheet ----
    sa = load_sheet_rows(wb, "Stock Analysis")
    record("Sheet 'Stock Analysis' present", sa is not None)
    stock_rows = []
    if sa:
        header = [str(c).strip() if c is not None else "" for c in sa[0]]
        idx = {h: i for i, h in enumerate(header)}
        needed = ["Product_ID", "Product_Name", "Category", "Current_Stock",
                  "Avg_Daily_Sales", "Days_Until_Safety_Stock", "Lead_Time",
                  "Needs_Reorder", "Reorder_Date"]
        record("Stock Analysis has required columns",
               all(n in idx for n in needed),
               f"missing: {[n for n in needed if n not in idx]}")
        for r in sa[1:]:
            if not r or r[0] is None:
                continue
            stock_rows.append(r)

        # (structural) total product count tolerance
        expected = gt_data["total_products"]
        record(f"Stock Analysis row count ~{expected} (+/-5)",
               abs(len(stock_rows) - expected) <= 5,
               f"got {len(stock_rows)}")

        # (structural) reorder=Yes count tolerance
        yes_rows = [r for r in stock_rows
                    if str(r[idx["Needs_Reorder"]]).strip().lower() == "yes"]
        exp_re = gt_data["reorder_count"]
        record(f"Needs_Reorder=Yes count ~{exp_re} (+/-3)",
               abs(len(yes_rows) - exp_re) <= 3, f"got {len(yes_rows)}")
    else:
        idx = {}
        yes_rows = []

    # ---- CRITICAL 1: Lead_Time per row matches russified category map ----
    if sa and idx:
        bad = []
        cats_checked = set()
        for r in stock_rows:
            cat = r[idx["Category"]]
            lt = r[idx["Lead_Time"]]
            cat_s = str(cat).strip() if cat is not None else ""
            if cat_s in CATEGORY_LEAD_TIME:
                cats_checked.add(cat_s)
                if _to_float(lt) != float(CATEGORY_LEAD_TIME[cat_s]):
                    bad.append((cat_s, lt))
        ok = (len(bad) == 0) and (len(cats_checked) >= 3)
        record("Lead_Time per row matches russified category->lead_time map",
               ok, f"mismatches={bad[:5]} distinct_cats={len(cats_checked)}")
    else:
        record("Lead_Time per row matches russified category->lead_time map",
               False, "no Stock Analysis rows")

    # ---- Reorder Schedule sheet ----
    rs = load_sheet_rows(wb, "Reorder Schedule")
    record("Sheet 'Reorder Schedule' present", rs is not None)
    sched_rows = []
    sidx = {}
    if rs:
        sheader = [str(c).strip() if c is not None else "" for c in rs[0]]
        sidx = {h: i for i, h in enumerate(sheader)}
        sneeded = ["Product_Name", "Category", "Current_Stock",
                   "Reorder_Date", "Lead_Time_Days"]
        record("Reorder Schedule has required columns",
               all(n in sidx for n in sneeded),
               f"missing: {[n for n in sneeded if n not in sidx]}")
        for r in rs[1:]:
            if not r or r[0] is None:
                continue
            sched_rows.append(r)
        record(f"Reorder Schedule row count ~{gt_data['reorder_count']} (+/-3)",
               abs(len(sched_rows) - gt_data["reorder_count"]) <= 3,
               f"got {len(sched_rows)}")

    # ---- CRITICAL 2: Reorder Schedule == products with Needs_Reorder=Yes ----
    # Names in the two sheets may differ in truncation (the store API returns
    # truncated names on one path), so match prefix-/substring-tolerantly but
    # still require a one-to-one correspondence (no fabricated or dropped rows).
    if sa and idx and rs and sidx:
        yes_names = [str(r[idx["Product_Name"]]).strip() for r in yes_rows]
        sched_names = [str(r[sidx["Product_Name"]]).strip() for r in sched_rows]

        def _name_match(a, b):
            a, b = a.strip(), b.strip()
            if not a or not b:
                return False
            return a == b or a.startswith(b) or b.startswith(a) or a in b or b in a

        unmatched_sched = []
        remaining = list(yes_names)
        for sn in sched_names:
            hit = next((y for y in remaining if _name_match(sn, y)), None)
            if hit is None:
                unmatched_sched.append(sn)
            else:
                remaining.remove(hit)
        ok = (len(yes_names) > 0
              and len(unmatched_sched) == 0
              and len(remaining) == 0
              and len(sched_names) == len(yes_names))
        record("Reorder Schedule == products with Needs_Reorder=Yes (set equality)",
               ok,
               f"sched_unmatched={unmatched_sched[:3]} yes_unmatched={remaining[:3]} "
               f"|yes|={len(yes_names)} |sched|={len(sched_names)}")
    else:
        record("Reorder Schedule == products with Needs_Reorder=Yes (set equality)",
               False, "missing a sheet")

    # ---- Fetch gcal events once ----
    events = get_gcal_events()
    record(f"Reorder calendar events ~{gt_data['reorder_count']} (+/-3) and > 0",
           len(events) > 0 and abs(len(events) - gt_data["reorder_count"]) <= 3,
           f"got {len(events)}")

    # event lookup by product name token from summary "Reorder: <name>"
    ev_by_name = {}
    for summary, desc, sdt, edt in events:
        if summary is None:
            continue
        m = re.match(r"\s*Reorder:\s*(.+)$", str(summary), re.IGNORECASE)
        name = m.group(1).strip() if m else str(summary).strip()
        ev_by_name[name] = (str(summary), desc, sdt, edt)

    # ---- CRITICAL 3: Reorder_Date future-or-tomorrow AND == event start ----
    if rs and sidx:
        problems = []
        matched = 0
        for r in sched_rows:
            pname = str(r[sidx["Product_Name"]]).strip()
            d = _parse_date(r[sidx["Reorder_Date"]])
            if d is None:
                problems.append((pname, "bad date"))
                continue
            if d <= today:
                problems.append((pname, f"date {d} not after today {today}"))
                continue
            # match to a gcal event (substring-tolerant, names may be truncated)
            ev = ev_by_name.get(pname)
            if ev is None:
                for en, e in ev_by_name.items():
                    if en and (en in pname or pname in en):
                        ev = e
                        break
            if ev is None:
                problems.append((pname, "no matching gcal event"))
                continue
            sdt = ev[2]
            sdate = sdt.date() if isinstance(sdt, datetime.datetime) else _parse_date(sdt)
            if sdate != d:
                problems.append((pname, f"event start {sdate} != reorder_date {d}"))
                continue
            matched += 1
        ok = (len(problems) == 0) and (matched >= max(1, len(sched_rows) - 1))
        record("Every Reorder_Date is tomorrow-or-future and matches its gcal event start",
               ok, f"problems={problems[:4]} matched={matched}/{len(sched_rows)}")
    else:
        record("Every Reorder_Date is tomorrow-or-future and matches its gcal event start",
               False, "no Reorder Schedule")

    # ---- CRITICAL 4: each event summary/description links to real store data ----
    if rs and sidx and events:
        problems = []
        ok_count = 0
        # map scheduled product name -> (current_stock, lead_time)
        sched_info = {}
        for r in sched_rows:
            pname = str(r[sidx["Product_Name"]]).strip()
            sched_info[pname] = (r[sidx["Current_Stock"]], r[sidx["Lead_Time_Days"]])
        for summary, desc, sdt, edt in events:
            s = str(summary or "")
            if not re.match(r"\s*Reorder:\s*\S", s, re.IGNORECASE):
                problems.append((s[:40], "summary not 'Reorder: <name>'"))
                continue
            ev_name = re.match(r"\s*Reorder:\s*(.+)$", s, re.IGNORECASE).group(1).strip()
            info = sched_info.get(ev_name)
            if info is None:
                for pn, v in sched_info.items():
                    if pn and (pn in ev_name or ev_name in pn):
                        info = v
                        break
            d = str(desc or "")
            if info is None:
                problems.append((ev_name[:30], "no matching scheduled product"))
                continue
            stock, lead = info
            # 09:00 -> 10:00 window
            shour = sdt.hour if isinstance(sdt, datetime.datetime) else None
            ehour = edt.hour if isinstance(edt, datetime.datetime) else None
            has_stock = str(int(stock)) in d if _to_float(stock) is not None else False
            has_lead = str(int(lead)) in d if _to_float(lead) is not None else False
            if has_stock and has_lead and shour == 9 and ehour == 10:
                ok_count += 1
            else:
                problems.append((ev_name[:30],
                                 f"stock_in_desc={has_stock} lead_in_desc={has_lead} "
                                 f"start_h={shour} end_h={ehour}"))
        ok = (len(problems) == 0) and (ok_count > 0)
        record("Each reorder gcal event summary/description links to real store data",
               ok, f"problems={problems[:4]} ok={ok_count}/{len(events)}")
    else:
        record("Each reorder gcal event summary/description links to real store data",
               False, "no events or schedule")

    # ---- CRITICAL 5: Avg_Daily_Sales consistent with stock & days-until ----
    # Policy: Days_Until_Safety_Stock == max(Current_Stock-5,0)/Avg_Daily_Sales
    if sa and idx:
        checked = 0
        good = 0
        for r in stock_rows:
            if str(r[idx["Needs_Reorder"]]).strip().lower() != "yes":
                continue
            adv = _to_float(r[idx["Avg_Daily_Sales"]])
            cs = _to_float(r[idx["Current_Stock"]])
            du = _to_float(r[idx["Days_Until_Safety_Stock"]])
            if adv is None or cs is None or du is None or adv <= 0:
                continue
            checked += 1
            eff = max(cs - gt_data["safety_stock"], 0)
            expected_du = eff / adv
            tol = max(1.0, 0.05 * expected_du)
            if abs(expected_du - du) <= tol:
                good += 1
            if good >= 3:
                break
        record("Avg_Daily_Sales consistent with Current_Stock & Days_Until_Safety_Stock",
               good >= 3, f"consistent={good} checked={checked}")
    else:
        record("Avg_Daily_Sales consistent with Current_Stock & Days_Until_Safety_Stock",
               False, "no Stock Analysis rows")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    # "today" = preprocess/launch date, computed dynamically (never hardcoded).
    today = None
    if args.launch_time:
        today = _parse_date(args.launch_time)
    if today is None:
        today = datetime.date.today()
    print(f"  Using reference date (today) = {today}")

    evaluate(agent_ws, gt_data, today)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "accuracy": accuracy,
                    "passed": PASS_COUNT,
                    "total": total,
                    "critical_failed": critical_failed,
                }, f, ensure_ascii=False, indent=2)
        except Exception as e:
            print(f"  (could not write res_log_file: {e})")

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
