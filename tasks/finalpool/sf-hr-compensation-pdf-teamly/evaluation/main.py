"""Evaluation for sf-hr-compensation-pdf-clickhouse-teamly.

Verifies the Excel deliverable (Compensation_Data.xlsx) against the
russified-warehouse groundtruth, the PDF deliverable, and the Teamly
knowledge-base page summarizing key findings.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Pass threshold otherwise: accuracy >= 70%.
"""
import argparse
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Headline findings derived from the russified groundtruth warehouse.
TOP_PAID_DEPARTMENT = "инженерия"          # highest average salary
MOST_COMMON_EDUCATION = "бакалавр"          # most employees overall

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Excel file exists",
    "Department Summary: department rows match (Employees+Avg tight)",
    "Department Summary: Median_Salary matches",
    "Education Breakdown: department x education rows match (Count tight)",
    "Teamly: page 'Compensation Analysis 2026' with both headline findings",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:400]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, gt_dir):
    agent_file = os.path.join(agent_workspace, "Compensation_Data.xlsx")
    gt_file = os.path.join(gt_dir, "Compensation_Data.xlsx")

    if not os.path.exists(agent_file):
        record("Excel file exists", False, f"not found: {agent_file}")
        return
    record("Excel file exists", True)
    if not os.path.exists(gt_file):
        record("Groundtruth Excel exists", False, f"not found: {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    EXPECTED_DEP_HDR = ["Department", "Employees", "Avg_Salary", "Min_Salary", "Max_Salary", "Median_Salary"]
    EXPECTED_EDU_HDR = ["Department", "Education_Level", "Count", "Avg_Salary", "Min_Salary", "Max_Salary"]

    # ---------------- Department Summary ----------------
    a_rows = load_sheet_rows(agent_wb, "Department Summary")
    g_rows = load_sheet_rows(gt_wb, "Department Summary")

    if a_rows is None:
        record("Department Summary sheet exists", False, "missing in agent output")
        record("Department Summary headers exact", False, "no sheet")
        record("Department Summary: department rows match (Employees+Avg tight)", False, "no sheet")
        record("Department Summary: Median_Salary matches", False, "no sheet")
    elif g_rows is None:
        record("Department Summary sheet exists", False, "missing in groundtruth")
    else:
        record("Department Summary sheet exists", True)
        a_hdr = [str(c).strip() if c is not None else None for c in (a_rows[0] if a_rows else [])]
        record("Department Summary headers exact", a_hdr[:6] == EXPECTED_DEP_HDR,
               f"got {a_hdr[:6]}")

        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        rows_ok = True
        median_ok = True
        details = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                rows_ok = False
                median_ok = False
                details.append(f"Missing department: {g_row[0]}")
                continue
            # Employees count: tight (tol<=2)
            if len(a_row) > 1 and not num_close(a_row[1], g_row[1], 2):
                rows_ok = False
                details.append(f"{key}.Employees {a_row[1]} vs {g_row[1]}")
            # Avg_Salary: tight (tol<=50)
            if len(a_row) > 2 and not num_close(a_row[2], g_row[2], 50):
                rows_ok = False
                details.append(f"{key}.Avg_Salary {a_row[2]} vs {g_row[2]}")
            # Min_Salary / Max_Salary: looser structural confirmation
            if len(a_row) > 3 and not num_close(a_row[3], g_row[3], 100):
                details.append(f"{key}.Min_Salary {a_row[3]} vs {g_row[3]}")
            if len(a_row) > 4 and not num_close(a_row[4], g_row[4], 100):
                details.append(f"{key}.Max_Salary {a_row[4]} vs {g_row[4]}")
            # Median_Salary: tight (tol<=50) — task explicitly requires it
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 50):
                    median_ok = False
                    details.append(f"{key}.Median_Salary {a_row[5]} vs {g_row[5]}")
            else:
                median_ok = False
                details.append(f"{key}: Median_Salary column missing")

        record("Department Summary: department rows match (Employees+Avg tight)", rows_ok,
               "; ".join(details[:8]))
        record("Department Summary: Median_Salary matches", median_ok,
               "; ".join([d for d in details if "Median" in d][:8]))

    # ---------------- Education Breakdown ----------------
    a_rows = load_sheet_rows(agent_wb, "Education Breakdown")
    g_rows = load_sheet_rows(gt_wb, "Education Breakdown")

    if a_rows is None:
        record("Education Breakdown sheet exists", False, "missing in agent output")
        record("Education Breakdown headers exact", False, "no sheet")
        record("Education Breakdown: department x education rows match (Count tight)", False, "no sheet")
    elif g_rows is None:
        record("Education Breakdown sheet exists", False, "missing in groundtruth")
    else:
        record("Education Breakdown sheet exists", True)
        a_hdr = [str(c).strip() if c is not None else None for c in (a_rows[0] if a_rows else [])]
        record("Education Breakdown headers exact", a_hdr[:6] == EXPECTED_EDU_HDR,
               f"got {a_hdr[:6]}")

        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None and row[1] is not None:
                k = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                a_lookup[k] = row

        rows_ok = True
        details = []
        for g_row in g_data:
            if not g_row or g_row[0] is None or g_row[1] is None:
                continue
            k = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(k)
            if a_row is None:
                rows_ok = False
                details.append(f"Missing edu row: {g_row[0]} / {g_row[1]}")
                continue
            # Count: tight (tol<=2)
            if len(a_row) > 2 and not num_close(a_row[2], g_row[2], 2):
                rows_ok = False
                details.append(f"{k}.Count {a_row[2]} vs {g_row[2]}")
            # Avg_Salary: moderate
            if len(a_row) > 3 and not num_close(a_row[3], g_row[3], 100):
                details.append(f"{k}.Avg_Salary {a_row[3]} vs {g_row[3]}")

        record("Education Breakdown: department x education rows match (Count tight)", rows_ok,
               "; ".join(details[:8]))


def check_pdf(agent_workspace):
    pdf_path = os.path.join(agent_workspace, "Compensation_Report.pdf")
    if not os.path.exists(pdf_path):
        record("PDF Compensation_Report.pdf exists", False, "not found")
        return
    size = os.path.getsize(pdf_path)
    record("PDF Compensation_Report.pdf exists", size >= 500, f"{size} bytes")


def check_teamly():
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432,
                                dbname="cowork_gym", user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT to_regclass('teamly.pages')")
        if cur.fetchone()[0] is None:
            record("Teamly: page 'Compensation Analysis 2026' with both headline findings",
                   False, "teamly.pages not found")
            cur.close(); conn.close()
            return
        # User-created pages only (seeds have id <= 3).
        cur.execute("SELECT title, body FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        cur.close(); conn.close()
    except Exception as e:
        record("Teamly: page 'Compensation Analysis 2026' with both headline findings",
               False, f"DB error: {e}")
        return

    # Locate the target page by title; fall back to any page mentioning compensation.
    target = None
    for title, body in pages:
        if "compensation analysis 2026" in (title or "").lower():
            target = (title, body)
            break
    if target is None:
        for title, body in pages:
            t = ((title or "") + " " + (body or "")).lower()
            if "compensation" in t or "оплат" in t or "зарплат" in t:
                target = (title, body)
                break

    record("Teamly: target page exists", target is not None,
           f"{len(pages)} user pages found")

    text = (((target[0] if target else "") or "") + " " + ((target[1] if target else "") or "")).lower()
    has_top = TOP_PAID_DEPARTMENT in text
    has_edu = MOST_COMMON_EDUCATION in text
    record("Teamly: page 'Compensation Analysis 2026' with both headline findings",
           target is not None and has_top and has_edu,
           f"top_dept={has_top} edu={has_edu}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=== Checking Excel ===")
    check_excel(args.agent_workspace, gt_dir)
    print("=== Checking PDF ===")
    check_pdf(args.agent_workspace)
    print("=== Checking Teamly ===")
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")
        print("=== RESULT: FAIL (critical check failed) ===")
        sys.exit(1)

    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print("=== RESULT: FAIL (accuracy below threshold) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
