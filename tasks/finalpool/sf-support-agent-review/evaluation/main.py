"""Evaluation for sf-support-agent-review.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Data-value realia (AGENT_NAME, ISSUE_TYPE) are russified centrally by the
ClickHouse relabel map; PRIORITY/TEAM/SKILL_LEVEL stay English. get_expected()
reads these LIVE from the DB, so the eval auto-syncs with whatever the map
produced — no English literals are hardcoded for those volatile values.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")
PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Priority High count",
    "Priority Medium count",
    "Priority Low count",
    "Priority High SLA",
    "Priority Medium SLA",
    "Priority Low SLA",
    "Total_Agents",
    "Total_Tickets",
    "Agent Roster names match live",
    "GSheet Support Performance Summary exists with Metric rows",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1; print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1; FAILED_NAMES.append(name); print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except: return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_expected():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute('SELECT "AGENT_NAME","TEAM","SKILL_LEVEL" FROM sf_data."SUPPORT_CENTER__PUBLIC__AGENTS" ORDER BY "AGENT_NAME"')
    agents = [{"name": r[0], "team": r[1], "skill": r[2]} for r in cur.fetchall()]
    cur.execute('SELECT "PRIORITY",COUNT(*),ROUND(AVG("RESPONSE_TIME_HOURS")::numeric,2),ROUND(AVG("CUSTOMER_SATISFACTION")::numeric,2) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" GROUP BY "PRIORITY"')
    pstats = [{"priority": r[0], "count": r[1], "avg_resp": float(r[2]), "avg_sat": float(r[3])} for r in cur.fetchall()]
    cur.execute('SELECT t."PRIORITY",COUNT(*),COUNT(CASE WHEN t."RESPONSE_TIME_HOURS"<=s."RESPONSE_TARGET_HOURS" THEN 1 END) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" t JOIN sf_data."SUPPORT_CENTER__PUBLIC__SLA_POLICIES" s ON s."PRIORITY"=t."PRIORITY" GROUP BY t."PRIORITY"')
    sla = {r[0]: round(r[2]/r[1]*100,2) if r[1]>0 else 0 for r in cur.fetchall()}
    for p in pstats: p["sla_pct"] = sla.get(p["priority"], 0)
    cur.execute('SELECT "ISSUE_TYPE",COUNT(*),ROUND(AVG("RESPONSE_TIME_HOURS")::numeric,2),ROUND(AVG("CUSTOMER_SATISFACTION")::numeric,2) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS" GROUP BY "ISSUE_TYPE" ORDER BY COUNT(*) DESC')
    istats = [{"type": r[0], "count": r[1], "avg_resp": float(r[2]), "avg_sat": float(r[3])} for r in cur.fetchall()]
    cur.execute('SELECT COUNT(*),ROUND(AVG("RESPONSE_TIME_HOURS")::numeric,2),ROUND(AVG("CUSTOMER_SATISFACTION")::numeric,2) FROM sf_data."SUPPORT_CENTER__PUBLIC__TICKETS"')
    o = cur.fetchone()
    conn.close()
    return {"agents": agents, "priority_stats": pstats, "issue_stats": istats,
            "overall": {"total": o[0], "avg_resp": float(o[1]), "avg_sat": float(o[2])}, "total_agents": len(agents)}


def sheet_dicts(wb, name):
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2: return []
            hdrs = [str(h).strip() if h else "" for h in rows[0]]
            return [{hdrs[i]: row[i] for i in range(len(hdrs))} for row in rows[1:] if not all(v is None for v in row)]
    return None


def check_excel(ws_path, exp):
    print("\n=== Checking Excel ===")
    p = os.path.join(ws_path, "Support_Analysis.xlsx")
    if not os.path.isfile(p):
        record("Excel file exists", False, p); return
    record("Excel file exists", True)
    wb = openpyxl.load_workbook(p, data_only=True)
    # Agent Roster
    d = sheet_dicts(wb, "Agent Roster")
    if d is None: record("Sheet Agent Roster", False, str(wb.sheetnames))
    else:
        record("Sheet Agent Roster", True)
        record("Agent count", len(d) == len(exp["agents"]), f"{len(d)} vs {len(exp['agents'])}")
        # Semantic: every live agent present with correct Team + Skill_Level pairing.
        all_ok = True
        miss = []
        for a in exp["agents"]:
            m = next((r for r in d if str_match(r.get("Agent_Name"), a["name"])), None)
            if m is None or not str_match(m.get("Team"), a["team"]) or not str_match(m.get("Skill_Level"), a["skill"]):
                all_ok = False; miss.append(a["name"])
        record("Agent Roster names match live", all_ok, f"Mismatched/missing: {miss}")
        # Sort-order: agents sorted by name (uses live ordering by AGENT_NAME).
        names = [str(r.get("Agent_Name")).strip() for r in d if r.get("Agent_Name") is not None]
        record("Agent Roster sorted by name", names == sorted(names), f"{names}")
    # Priority Analysis
    d = sheet_dicts(wb, "Priority Analysis")
    if d is None: record("Sheet Priority Analysis", False, str(wb.sheetnames))
    else:
        record("Sheet Priority Analysis", True)
        for e in exp["priority_stats"]:
            m = next((r for r in d if str_match(r.get("Priority"), e["priority"])), None)
            if not m: record(f"Priority {e['priority']} count", False, "Missing"); continue
            record(f"Priority {e['priority']} count", num_close(m.get("Ticket_Count"), e["count"], 50),
                   f"{m.get('Ticket_Count')} vs {e['count']}")
            record(f"Priority {e['priority']} resp", num_close(m.get("Avg_Response_Hours"), e["avg_resp"], 1.0),
                   f"{m.get('Avg_Response_Hours')} vs {e['avg_resp']}")
            record(f"Priority {e['priority']} sat", num_close(m.get("Avg_Satisfaction"), e["avg_sat"], 0.5),
                   f"{m.get('Avg_Satisfaction')} vs {e['avg_sat']}")
            # Central business rule: SLA compliance % from JOIN to SLA_POLICIES (tol ~2pp).
            record(f"Priority {e['priority']} SLA", num_close(m.get("SLA_Compliance_Pct"), e["sla_pct"], 2.0),
                   f"{m.get('SLA_Compliance_Pct')} vs {e['sla_pct']}")
    # Issue Type Analysis
    d = sheet_dicts(wb, "Issue Type Analysis")
    if d is None: record("Sheet Issue Type Analysis", False, str(wb.sheetnames))
    else:
        record("Sheet Issue Type Analysis", True)
        for e in exp["issue_stats"]:
            m = next((r for r in d if str_match(r.get("Issue_Type"), e["type"])), None)
            if not m: record(f"Issue {e['type']} count", False, "Missing"); continue
            record(f"Issue {e['type']} count", num_close(m.get("Ticket_Count"), e["count"], 50),
                   f"{m.get('Ticket_Count')} vs {e['count']}")
        # Sort-order: issue types sorted by ticket count descending.
        counts = [r.get("Ticket_Count") for r in d if r.get("Ticket_Count") is not None]
        try:
            cf = [float(c) for c in counts]
            record("Issue Type sorted by count desc", cf == sorted(cf, reverse=True), f"{cf}")
        except Exception as e2:
            record("Issue Type sorted by count desc", False, str(e2))
    # Summary
    d = sheet_dicts(wb, "Summary")
    if d is None: record("Sheet Summary", False, str(wb.sheetnames))
    else:
        record("Sheet Summary", True)
        ms = {str(r.get("Metric","")).strip(): r.get("Value") for r in d}
        record("Total_Agents", num_close(ms.get("Total_Agents"), exp["total_agents"], 0), f"{ms.get('Total_Agents')}")
        record("Total_Tickets", num_close(ms.get("Total_Tickets"), exp["overall"]["total"], 100), f"{ms.get('Total_Tickets')}")
        record("Avg_Resp", num_close(ms.get("Overall_Avg_Response_Hours"), exp["overall"]["avg_resp"], 1.0), f"{ms.get('Overall_Avg_Response_Hours')}")
        record("Avg_Sat", num_close(ms.get("Overall_Avg_Satisfaction"), exp["overall"]["avg_sat"], 0.5), f"{ms.get('Overall_Avg_Satisfaction')}")
    wb.close()


def check_gsheet(exp):
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    ss = cur.fetchall()
    if not ss:
        record("GSheet Support Performance Summary exists with Metric rows", False, "No spreadsheet found")
        conn.close(); return
    tid = None
    for sid, t in ss:
        if "support" in str(t).lower() and "summary" in str(t).lower():
            tid = sid; break
    if tid is None:
        tid = ss[0][0]
        record("GSheet title matches Support Performance Summary", False, f"Found: {[s[1] for s in ss]}")
    else:
        record("GSheet title matches Support Performance Summary", True)
    cur.execute("SELECT id FROM gsheet.sheets WHERE spreadsheet_id=%s", (tid,))
    sheets = cur.fetchall()
    record("GSheet has Summary sheet", len(sheets) >= 1, f"{len(sheets)}")
    vals = []
    for (sheet_id,) in sheets:
        cur.execute("SELECT value FROM gsheet.cells WHERE sheet_id=%s", (sheet_id,))
        vals += [str(r[0]).strip().lower() for r in cur.fetchall() if r[0] is not None]
    has_agents = any("total_agents" in v for v in vals)
    has_tickets = any("total_tickets" in v for v in vals)
    has_resp = any("overall_avg_response_hours" in v for v in vals)
    has_sat = any("overall_avg_satisfaction" in v for v in vals)
    # Critical: the shared spreadsheet exists AND carries the four Metric rows.
    record("GSheet Support Performance Summary exists with Metric rows",
           has_agents and has_tickets and has_resp and has_sat,
           f"agents={has_agents} tickets={has_tickets} resp={has_resp} sat={has_sat}")
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", default=".")
    parser.add_argument("--groundtruth_workspace", default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    exp = get_expected()
    check_excel(args.agent_workspace, exp)
    check_gsheet(exp)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy, "critical_failed": critical_failed,
                       "success": success}, f)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
