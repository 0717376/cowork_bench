"""
Evaluation для rzd-insales-elektrosnab-supplier-visit-spb-kzn-excel-email-gcal.

Что проверяем:
1. Supplier_Visit_Plan.xlsx существует, 3 листа (Products, Travel_Plan, Visit_Schedule)
2. Products: >=5 строк, есть нужные колонки; распределение по городам — 3 СПб + 2 Казань;
   приоритеты только из {High, Medium, Low}
3. Travel_Plan: 2 строки; есть Сапсан 752 (СПб) и Стриж 716 (Казань); Бизнес класс
4. Visit_Schedule: >=5 строк; даты только 2026-03-10 / 2026-03-17 (ISO!);
   время после буфера; статус Scheduled
5. GCal: >=2 события — по одному на 03-10 и 03-17, в summary упоминается город
6. Письма: 3 отправленных на spb_supplier@partner.ru, kzn_supplier@partner.ru,
   procurement@elektrosnab.ru
"""
import json
import os
import re
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []  # имена провалившихся обязательных проверок

# Чеки, провал которых означает содержательное невыполнение задачи.
# Если такой чек FAIL — итог всей задачи FAIL, независимо от accuracy.
CRITICAL_CHECKS = {
    "Products: Priority соответствует stock по правилу гайда",
    "Products: хотя бы 1 поставщик High приоритета (критичный stock)",
    "Travel_Plan: цена СПб Бизнес ≈ 14000₽ (±500)",
    "Travel_Plan: цена Казань Бизнес ≈ 11000₽ (±500)",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def normalize_train(s):
    """Кириллица → латиница для номеров типа 752А, 716Г, 016У."""
    if s is None:
        return ""
    s = str(s).strip().upper()
    cyr_to_lat = str.maketrans({
        "А": "A", "В": "B", "Е": "E", "К": "K", "М": "M",
        "Н": "H", "О": "O", "Р": "P", "С": "C", "Т": "T",
        "У": "Y", "Х": "X", "Г": "G",
    })
    return s.translate(cyr_to_lat)


def normalize_city(s):
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = s.replace("ё", "е")
    return s


def is_iso_date(s):
    """Проверка формата YYYY-MM-DD."""
    if s is None:
        return False
    s = str(s).strip()
    # допустим и datetime/date объект (openpyxl может вернуть)
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return True
    return False


def parse_hhmm(s):
    """Извлечь HH:MM из строки/времени."""
    if s is None:
        return None
    s = str(s).strip()
    m = re.search(r"(\d{1,2}):(\d{2})", s)
    if not m:
        return None
    return int(m.group(1)) * 60 + int(m.group(2))


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Excel Supplier_Visit_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Supplier_Visit_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Supplier_Visit_Plan.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Supplier_Visit_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    has_products = any("product" in s for s in sheet_names_lower)
    has_travel = any("travel" in s for s in sheet_names_lower)
    has_schedule = any("schedule" in s or "visit" in s for s in sheet_names_lower)

    record("Excel has Products sheet", has_products, f"Sheets: {wb.sheetnames}")
    record("Excel has Travel_Plan sheet", has_travel, f"Sheets: {wb.sheetnames}")
    record("Excel has Visit_Schedule sheet", has_schedule, f"Sheets: {wb.sheetnames}")

    # ---- Products ----
    if has_products:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "product" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and c != "" for c in r)]
        record("Products: >= 5 строк", len(data_rows) >= 5, f"Found {len(data_rows)} rows")

        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            has_id = any("id" in h or "product" in h for h in headers)
            has_supplier_name = any("supplier" in h and "name" in h for h in headers)
            has_supplier_city = any("supplier" in h and "city" in h for h in headers)
            has_priority = any("priority" in h for h in headers)
            record("Products: есть колонка Product_ID", has_id, f"Headers: {rows[0]}")
            record("Products: есть колонка Supplier_Name", has_supplier_name, f"Headers: {rows[0]}")
            record("Products: есть колонка Supplier_City", has_supplier_city, f"Headers: {rows[0]}")
            record("Products: есть колонка Priority", has_priority, f"Headers: {rows[0]}")

            # Распределение по городам — индексы колонок
            try:
                city_col = next(i for i, h in enumerate(headers) if "supplier" in h and "city" in h)
                priority_col = next(i for i, h in enumerate(headers) if "priority" in h)
                stock_col = next(i for i, h in enumerate(headers) if "stock" in h and "status" not in h)

                cities = [normalize_city(r[city_col]) for r in data_rows
                          if city_col < len(r) and r[city_col]]
                spb_count = sum(1 for c in cities if "петербург" in c or "санкт" in c or "спб" in c)
                kzn_count = sum(1 for c in cities if "казан" in c)
                record("Products: 3 поставщика в СПб", spb_count >= 3,
                       f"Found {spb_count} (cities={cities})")
                record("Products: 2 поставщика в Казани", kzn_count >= 2,
                       f"Found {kzn_count} (cities={cities})")

                # Приоритеты — только из {High, Medium, Low}
                priorities = [str(r[priority_col]).strip().lower() for r in data_rows
                              if priority_col < len(r) and r[priority_col]]
                valid_prio = {"high", "medium", "low"}
                ok_prio = all(p in valid_prio for p in priorities)
                record("Products: Priority только High/Medium/Low",
                       ok_prio and len(priorities) >= 5,
                       f"Got: {priorities}")

                # Priority строго соответствует stock (правило гайда):
                #   stock < 5  → High
                #   5 ≤ stock ≤ 15 → Medium
                #   stock > 15 → Low
                def expected_prio(stk):
                    try:
                        s = int(stk)
                    except (TypeError, ValueError):
                        return None
                    if s < 5:
                        return "high"
                    if s <= 15:
                        return "medium"
                    return "low"

                mismatches = []
                high_count = 0
                for r in data_rows:
                    if stock_col >= len(r) or priority_col >= len(r):
                        continue
                    stk = r[stock_col]
                    pr = str(r[priority_col]).strip().lower() if r[priority_col] else ""
                    exp = expected_prio(stk)
                    if pr == "high":
                        high_count += 1
                    if exp and pr != exp:
                        mismatches.append(f"stock={stk}, prio={pr}, expected={exp}")

                record("Products: Priority соответствует stock по правилу гайда",
                       len(mismatches) == 0,
                       f"Mismatches: {mismatches}")
                # Critical items в WC реально есть (stock 0/1/2) — High обязателен
                record("Products: хотя бы 1 поставщик High приоритета (критичный stock)",
                       high_count >= 1,
                       f"High count: {high_count}")
            except StopIteration:
                pass

    # ---- Travel_Plan ----
    if has_travel:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_names_lower) if "travel" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and c != "" for c in r)]
        record("Travel_Plan: 2 строки", len(data_rows) == 2, f"Found {len(data_rows)} rows")

        all_text_norm = " ".join(
            normalize_train(c) for row in rows for c in row if c
        )
        record("Travel_Plan: Сапсан 752 (СПб маршрут)",
               "752" in all_text_norm,
               f"Content: {all_text_norm[:200]}")
        record("Travel_Plan: Стриж 716 (Казань маршрут)",
               "716" in all_text_norm,
               f"Content: {all_text_norm[:200]}")

        # Бизнес-класс
        all_text_raw = " ".join(str(c).lower() for row in rows for c in row if c)
        record("Travel_Plan: класс Бизнес",
               "бизнес" in all_text_raw or "business" in all_text_raw,
               f"Content: {all_text_raw[:200]}")

        # Города
        record("Travel_Plan: упомянут Санкт-Петербург",
               "петербург" in all_text_raw or "спб" in all_text_raw,
               f"Content: {all_text_raw[:200]}")
        record("Travel_Plan: упомянута Казань",
               "казан" in all_text_raw,
               f"Content: {all_text_raw[:200]}")

        # Цены билетов из системы РЖД (rzd.train_seats, класс Бизнес):
        # СПб (752А) Бизнес 14000₽, Казань (716Г) Бизнес 11000₽.
        # ±500₽ допуск на округление/округлённые источники
        if rows:
            headers_tp = [str(c).lower() if c else "" for c in rows[0]]
            try:
                city_col_tp = next(i for i, h in enumerate(headers_tp) if "city" in h)
                price_col_tp = next(i for i, h in enumerate(headers_tp)
                                    if "price" in h or "rub" in h or "ticket" in h)
                spb_price = None
                kzn_price = None
                for r in data_rows:
                    if city_col_tp >= len(r) or price_col_tp >= len(r):
                        continue
                    c = normalize_city(r[city_col_tp])
                    try:
                        p = float(str(r[price_col_tp]).replace(" ", "").replace(",", "."))
                    except (TypeError, ValueError):
                        continue
                    if "петербург" in c or "санкт" in c or "спб" in c:
                        spb_price = p
                    elif "казан" in c:
                        kzn_price = p
                record("Travel_Plan: цена СПб Бизнес ≈ 14000₽ (±500)",
                       spb_price is not None and 13500 <= spb_price <= 14500,
                       f"spb_price={spb_price}")
                record("Travel_Plan: цена Казань Бизнес ≈ 11000₽ (±500)",
                       kzn_price is not None and 10500 <= kzn_price <= 11500,
                       f"kzn_price={kzn_price}")
            except StopIteration:
                pass

    # ---- Visit_Schedule ----
    if has_schedule:
        idx = next(i for i, s in enumerate(sheet_names_lower) if "schedule" in s or "visit" in s)
        ws_name = wb.sheetnames[idx]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and c != "" for c in r)]
        record("Visit_Schedule: >= 5 строк", len(data_rows) >= 5, f"Found {len(data_rows)} rows")

        if rows:
            headers = [str(c).lower() if c else "" for c in rows[0]]
            try:
                date_col = next(i for i, h in enumerate(headers) if "date" in h)
                city_col = next(i for i, h in enumerate(headers) if h == "city" or "city" in h)
                time_col = next(i for i, h in enumerate(headers) if "time" in h)
                status_col = next(i for i, h in enumerate(headers) if "status" in h)

                # Даты в ISO и только 2026-03-10 / 2026-03-17
                allowed_dates = {"2026-03-10", "2026-03-17"}
                iso_count = 0
                allowed_count = 0
                for r in data_rows:
                    if date_col >= len(r):
                        continue
                    raw = r[date_col]
                    s = str(raw).strip() if raw is not None else ""
                    # openpyxl может вернуть datetime — нормализуем
                    if hasattr(raw, "strftime"):
                        s = raw.strftime("%Y-%m-%d")
                    if is_iso_date(s):
                        iso_count += 1
                    if s[:10] in allowed_dates:
                        allowed_count += 1
                record("Visit_Schedule: все даты в ISO формате YYYY-MM-DD",
                       iso_count == len(data_rows),
                       f"ISO {iso_count}/{len(data_rows)}")
                record("Visit_Schedule: только разрешённые даты (10.03 или 17.03)",
                       allowed_count == len(data_rows),
                       f"Allowed {allowed_count}/{len(data_rows)}")

                # Все статусы Scheduled
                statuses = [str(r[status_col]).strip().lower() for r in data_rows
                            if status_col < len(r) and r[status_col]]
                all_scheduled = all(s == "scheduled" for s in statuses)
                record("Visit_Schedule: все Status = Scheduled",
                       all_scheduled and len(statuses) >= 5,
                       f"Got: {statuses}")

                # Буфер времени — встречи в СПб >= 11:20, в Казани >= 12:20
                buffer_ok = True
                buffer_details = []
                for r in data_rows:
                    if city_col >= len(r) or time_col >= len(r):
                        continue
                    city = normalize_city(r[city_col])
                    mins = parse_hhmm(r[time_col])
                    if mins is None:
                        continue
                    if "петербург" in city or "санкт" in city or "спб" in city:
                        if mins < 11 * 60 + 20:
                            buffer_ok = False
                            buffer_details.append(f"СПб {r[time_col]}")
                    elif "казан" in city:
                        if mins < 12 * 60 + 20:
                            buffer_ok = False
                            buffer_details.append(f"Казань {r[time_col]}")
                record("Visit_Schedule: соблюдён буфер времени",
                       buffer_ok,
                       f"Violations: {buffer_details}")
            except StopIteration:
                pass


def check_gcal():
    print("\n=== Check 2: GCal supplier visit events ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime::date IN ('2026-03-10', '2026-03-17')
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    record("GCal: >= 2 события на 10.03 или 17.03", len(events) >= 2,
           f"Found {len(events)} events")

    summaries = [str(e[0]).lower() for e in events]
    dates = [e[1].date().isoformat() if e[1] else "" for e in events]

    has_spb_event = any(
        ("2026-03-10" == d) and (
            "петербург" in s or "санкт" in s or "спб" in s
            or "supplier" in s or "visit" in s or "meeting" in s or "поставщик" in s
        )
        for s, d in zip(summaries, dates)
    )
    has_kzn_event = any(
        ("2026-03-17" == d) and (
            "казан" in s
            or "supplier" in s or "visit" in s or "meeting" in s or "поставщик" in s
        )
        for s, d in zip(summaries, dates)
    )
    record("GCal: событие про СПб на 2026-03-10", has_spb_event,
           f"Summaries: {list(zip(summaries, dates))}")
    record("GCal: событие про Казань на 2026-03-17", has_kzn_event,
           f"Summaries: {list(zip(summaries, dates))}")


def check_emails():
    print("\n=== Check 3: Emails sent ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    def to_addresses(to_addr):
        if isinstance(to_addr, list):
            return " ".join(str(r).lower() for r in to_addr)
        elif to_addr:
            try:
                parsed = json.loads(str(to_addr))
                return " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                return str(to_addr).lower()
        return ""

    target_addrs = {
        "spb_supplier@partner.ru",
        "kzn_supplier@partner.ru",
        "procurement@elektrosnab.ru",
    }

    to_spb = [m for m in messages if "spb_supplier@partner.ru" in to_addresses(m[2])]
    to_kzn = [m for m in messages if "kzn_supplier@partner.ru" in to_addresses(m[2])]
    to_proc = [m for m in messages if "procurement@elektrosnab.ru" in to_addresses(m[2])]

    record("Email: отправлено на spb_supplier@partner.ru", len(to_spb) >= 1,
           f"Total messages: {len(messages)}")
    record("Email: отправлено на kzn_supplier@partner.ru", len(to_kzn) >= 1,
           f"Total messages: {len(messages)}")
    record("Email: отправлено на procurement@elektrosnab.ru", len(to_proc) >= 1,
           f"Total messages: {len(messages)}")

    outgoing = [m for m in messages if any(a in to_addresses(m[2]) for a in target_addrs)]
    record("Email: всего >= 3 целевых писем", len(outgoing) >= 3,
           f"Outgoing: {len(outgoing)} / total: {len(messages)}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"Critical fails: {CRITICAL_FAILS}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_fails": CRITICAL_FAILS,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"FAIL: критичные чеки провалены ({len(CRITICAL_FAILS)})")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
