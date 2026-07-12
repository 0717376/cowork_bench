"""Evaluation for insales-order-refund-analysis."""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        detail_str = f": {detail[:200]}" if detail else ""
        tag = " [CRITICAL]" if critical else ""
        print(f"  [FAIL]{tag} {name}{detail_str}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Excel output against groundtruth."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Refund_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Refund_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}", critical=True)
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e), critical=True)
        return False

    def get_sheet(wb, name):
        for s in wb.sheetnames:
            if s.strip().lower() == name.strip().lower():
                return wb[s]
        return None

    # Check Refund Details sheet
    print("\n--- Refund Details ---")
    agent_ws = get_sheet(agent_wb, "Refund Details")
    gt_ws = get_sheet(gt_wb, "Refund Details")
    check("Sheet 'Refund Details' exists", agent_ws is not None,
          f"Found: {agent_wb.sheetnames}", critical=True)

    if agent_ws and gt_ws:
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        check("Refund Details row count", len(agent_rows) == len(gt_rows),
              f"Expected {len(gt_rows)}, got {len(agent_rows)}", critical=True)

        for gt_row in gt_rows:
            oid, cname, total, date_str = gt_row
            matched = None
            for ar in agent_rows:
                if ar and num_close(ar[0], oid, 0):
                    matched = ar
                    break
            if matched:
                check(f"Order {oid} Customer_Name",
                      str_match(matched[1], cname),
                      f"Expected '{cname}', got '{matched[1]}'", critical=True)
                check(f"Order {oid} Total",
                      num_close(matched[2], total, 0.5),
                      f"Expected {total}, got {matched[2]}", critical=True)
            else:
                check(f"Order {oid} found", False, critical=True)

    # Check Summary sheet
    print("\n--- Summary ---")
    agent_sum = get_sheet(agent_wb, "Summary")
    gt_sum = get_sheet(gt_wb, "Summary")
    check("Sheet 'Summary' exists", agent_sum is not None,
          f"Found: {agent_wb.sheetnames}", critical=True)

    if agent_sum and gt_sum:
        gt_data = {}
        for row in gt_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                gt_data[str(row[0]).strip().lower()] = row[1]

        agent_data = {}
        for row in agent_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                agent_data[str(row[0]).strip().lower()] = row[1]

        for key, gt_val in gt_data.items():
            agent_val = agent_data.get(key)
            if agent_val is None:
                for ak, av in agent_data.items():
                    if key.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                check(f"Summary '{key}'",
                      num_close(agent_val, gt_val, 1.0),
                      f"Expected {gt_val}, got {agent_val}", critical=True)
            else:
                check(f"Summary '{key}'",
                      str_match(agent_val, gt_val),
                      f"Expected '{gt_val}', got '{agent_val}'")

    return True


def read_gt_summary(gt_file):
    """Read expected refund count and total amount from groundtruth Refund_Report.xlsx."""
    wb = openpyxl.load_workbook(gt_file, data_only=True)
    ws = None
    for s in wb.sheetnames:
        if s.strip().lower() == "summary":
            ws = wb[s]
            break
    count, amount = None, None
    if ws:
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            key = str(row[0]).strip().lower()
            if key == "total_refunds":
                count = row[1]
            elif key == "total_amount":
                amount = row[1]
    return count, amount


def body_has_number(body_lower, value):
    """Substring-match a number in email body, tolerant of RU formatting.

    Accepts plain integers, RU thousands separators (space / non-breaking
    space / comma), and matches on the integer part so '4256', '4,256',
    '4 256', '4 256,31', '4256.31' all satisfy a 4256 expectation.
    """
    if value is None:
        return False
    int_part = str(int(round(float(value))))
    # Strip common thousands separators and decimal markers from body so a
    # formatted RU number collapses to its bare digit run.
    collapsed = body_lower
    for sep in (" ", " ", " ", ",", ".", "'"):
        collapsed = collapsed.replace(sep, "")
    return int_part in collapsed or int_part in body_lower


def check_emails(gt_file):
    """Check that summary email was sent to manager.

    Expected refund count and total amount are derived from the groundtruth
    Refund_Report.xlsx (not hardcoded), so this stays in sync with the wc seed.
    """
    print("\n=== Checking Emails ===")

    gt_count, gt_amount = read_gt_summary(gt_file)

    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    conn.close()

    def find_email_for_recipient(recipient):
        for subj, from_addr, to_addr, body in all_emails:
            if to_addr:
                recipients = []
                if isinstance(to_addr, list):
                    recipients = [str(r).strip().lower() for r in to_addr]
                elif isinstance(to_addr, str):
                    try:
                        parsed = json.loads(to_addr)
                        if isinstance(parsed, list):
                            recipients = [str(r).strip().lower() for r in parsed]
                        else:
                            recipients = [str(to_addr).strip().lower()]
                    except (json.JSONDecodeError, TypeError):
                        recipients = [str(to_addr).strip().lower()]
                if recipient.lower() in recipients:
                    return subj, from_addr, to_addr, body
        return None

    result = find_email_for_recipient("manager@store.example.com")
    check("Summary email sent to manager", result is not None, critical=True)

    if result:
        subj, from_addr, to_addr, body = result
        has_refund_subject = "refund" in (subj or "").lower()
        check("Email subject contains 'Refund'", has_refund_subject,
              f"Subject: {(subj or '')[:100]}", critical=True)

        body_lower = (body or "").lower()
        count_int = str(int(gt_count)) if gt_count is not None else "9"
        check(f"Email body mentions total refunds ({count_int})",
              body_has_number(body_lower, gt_count),
              f"Expected mention of {count_int} refunds", critical=True)
        check("Email body mentions total amount",
              body_has_number(body_lower, gt_amount),
              f"Expected mention of total amount ~{gt_amount}", critical=True)

    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("WC ORDER REFUND ANALYSIS - EVALUATION")
    print("=" * 70)

    gt_file = os.path.join(gt_dir, "Refund_Report.xlsx")

    check_excel(args.agent_workspace, gt_dir)
    check_emails(gt_file)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL checks failed: {CRITICAL_FAILED}")
        print(f"  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
