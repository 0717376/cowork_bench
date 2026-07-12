"""Evaluation для yf-financial-metrics-notion-email (RU: moex + teamly).

Проверки:
1. Excel Financial_Health_Report.xlsx — листы Key Metrics и Summary с корректными
   значениями (эталон пересчитан из moex.financial_statements, годовой период
   2025-12-31).
2. Teamly: страница «Investment Portfolio Financial Analysis 2026» с финансовой
   сводкой по портфелю.
3. Word Financial_Analysis_Report.docx с заголовком и анализом по компаниям.
4. Письмо на portfolio.manager@investment.com с финансовой оценкой.

CRITICAL_CHECKS: любой провал => общий FAIL независимо от accuracy.
Иначе PASS требует accuracy >= 70%.

Инструменты moex-finance имеют те же имена; тикеры — RU-аналоги:
  OZON.ME (Ozon, наибольшая выручка), YNDX.ME (Yandex, наибольшая маржа),
  PHOR.ME (PhosAgro, наименьшая выручка). Все суммы — в рублях.
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Тикеры задачи и их человекочитаемые имена (для поиска в тексте).
COMPANY_TERMS = ["ozon.me", "yndx.me", "phor.me", "ozon", "yandex", "phosagro", "фосагро"]

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Семантические критические проверки: их провал = общий FAIL.
CRITICAL_CHECKS = {
    "Excel: лист Key Metrics содержит 3 строки",
    "Excel: Highest_Revenue_Company == OZON.ME",
    "Excel: Most_Profitable_Company == YNDX.ME",
    "Teamly: страница с финансовой сводкой по портфелю",
    "Письмо на portfolio.manager@investment.com с финансовой оценкой",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_by_name(wb, name):
    for sname in wb.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            return [[cell.value for cell in row] for row in wb[sname].iter_rows()]
    return None


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Проверка Financial_Health_Report.xlsx ===")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        return False

    agent_file = os.path.join(agent_workspace, "Financial_Health_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Financial_Health_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        record("Excel: лист Key Metrics содержит 3 строки", False, "файл отсутствует")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth file exists", False, f"Not found: {gt_file}")
        return False

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_ok = True

    # Check Key Metrics sheet
    a_km = load_sheet_by_name(agent_wb, "Key Metrics")
    g_km = load_sheet_by_name(gt_wb, "Key Metrics")
    record("Sheet 'Key Metrics' exists", a_km is not None)

    a_data = []
    if a_km is not None and g_km is not None:
        a_data = [r for r in a_km[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_km[1:] if any(v is not None for v in r)]
        record("Excel: лист Key Metrics содержит 3 строки",
               len(a_data) == 3,
               f"Found {len(a_data)} rows")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().upper()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            sym = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(sym)
            if a_row is None:
                record(f"{sym} row in Key Metrics", False, "Not found")
                all_ok = False
                continue
            record(f"{sym} row exists", True)

            # Total_Revenue в млрд (col 1)
            if len(g_row) > 1 and len(a_row) > 1:
                record(f"{sym}: Total_Revenue in expected range",
                       num_close(a_row[1], g_row[1], 20.0),
                       f"got {a_row[1]}B, expected ~{g_row[1]}B")
            # Net_Income в млрд (col 2)
            if len(g_row) > 2 and len(a_row) > 2:
                record(f"{sym}: Net_Income in expected range",
                       num_close(a_row[2], g_row[2], 10.0),
                       f"got {a_row[2]}B, expected ~{g_row[2]}B")
            # Total_Assets в млрд (col 3)
            if len(g_row) > 3 and len(a_row) > 3:
                record(f"{sym}: Total_Assets in expected range",
                       num_close(a_row[3], g_row[3], 25.0),
                       f"got {a_row[3]}B, expected ~{g_row[3]}B")
            # Profit_Margin_Pct (col 4)
            if len(g_row) > 4 and len(a_row) > 4:
                record(f"{sym}: Profit_Margin_Pct correct",
                       num_close(a_row[4], g_row[4], 3.0),
                       f"got {a_row[4]}, expected ~{g_row[4]}")

        # Сортировка по Total_Revenue по убыванию (non-critical, structural)
        revs = [r[1] for r in a_data if len(r) > 1 and r[1] is not None]
        try:
            record("Key Metrics sorted by Total_Revenue desc",
                   revs == sorted(revs, reverse=True), str(revs))
        except TypeError:
            record("Key Metrics sorted by Total_Revenue desc", False, str(revs))
    else:
        record("Excel: лист Key Metrics содержит 3 строки", False, "лист отсутствует")

    # Check Summary sheet
    a_summ = load_sheet_by_name(agent_wb, "Summary")
    g_summ = load_sheet_by_name(gt_wb, "Summary")
    record("Sheet 'Summary' exists", a_summ is not None)

    a_lookup = {}
    if a_summ is not None and g_summ is not None:
        a_sdata = [r for r in a_summ[1:] if any(v is not None for v in r)]
        g_sdata = [r for r in g_summ[1:] if any(v is not None for v in r)]

        for row in a_sdata:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_sdata:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary row: {g_row[0]}", False, "Not found")
                continue
            record(f"Summary row: {g_row[0]}", True)

            if key == "avg_profit_margin":
                record("Avg_Profit_Margin correct",
                       num_close(a_row[1], g_row[1], 3.0),
                       f"got {a_row[1]}, expected ~{g_row[1]}")

    # CRITICAL: Highest_Revenue_Company == OZON.ME (наибольшая выручка)
    hrc = a_lookup.get("highest_revenue_company")
    record("Excel: Highest_Revenue_Company == OZON.ME",
           hrc is not None and hrc[1] is not None and "ozon" in str(hrc[1]).lower(),
           f"got {hrc[1] if hrc else None}")

    # CRITICAL: Most_Profitable_Company == YNDX.ME (наибольшая маржа)
    mpc = a_lookup.get("most_profitable_company")
    record("Excel: Most_Profitable_Company == YNDX.ME",
           mpc is not None and mpc[1] is not None and (
               "yndx" in str(mpc[1]).lower() or "yandex" in str(mpc[1]).lower()),
           f"got {mpc[1] if mpc else None}")

    return all_ok


# ============================================================================
# Check 2: Teamly page
# ============================================================================

def check_teamly():
    print("\n=== Проверка Teamly ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Только пользовательские страницы агента (сид имеет id <= 3).
    cur.execute("""
        SELECT id, title, COALESCE(body, '')
        FROM teamly.pages
        WHERE id > 3
    """)
    pages = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_teamly] Найдено {len(pages)} пользовательских страниц.")
    record("Создана хотя бы одна страница Teamly", len(pages) >= 1)

    found = False
    for _pid, title, body in pages:
        text = ((title or "") + " " + (body or "")).lower()
        title_ok = ("investment" in text or "financial" in text or
                    "portfolio" in text or "портфел" in text or "финанс" in text)
        company_ok = any(term in text for term in COMPANY_TERMS)
        if title_ok and company_ok:
            found = True
            break

    record("Teamly: страница с финансовой сводкой по портфелю", found,
           f"страниц: {len(pages)}")
    return found


# ============================================================================
# Check 3: Word document
# ============================================================================

def check_word(agent_workspace):
    print("\n=== Проверка Financial_Analysis_Report.docx ===")

    docx_path = os.path.join(agent_workspace, "Financial_Analysis_Report.docx")
    if not os.path.isfile(docx_path):
        record("Word file exists", False, f"Not found: {docx_path}")
        return False
    record("Word file exists", True)

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        record("Word doc has content", len(all_text.strip()) >= 100,
               f"Content length: {len(all_text.strip())}")
        record("Word doc mentions financial health or Q1",
               any(term in all_text for term in
                   ["financial", "health", "q1", "2026", "revenue", "выручк", "финанс"]),
               "Missing financial content")
        record("Word doc mentions companies",
               any(term in all_text for term in COMPANY_TERMS),
               "Missing company names")
        return True
    except ImportError:
        size = os.path.getsize(docx_path)
        record("Word file has content (>2KB)", size > 2000, f"Size: {size} bytes")
        return size > 2000
    except Exception as e:
        record("Word file readable", False, str(e))
        return False


# ============================================================================
# Check 4: Email
# ============================================================================

def check_emails():
    print("\n=== Проверка почты ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Найдено {len(all_emails)} писем.")
    record("Отправлено хотя бы одно письмо", len(all_emails) >= 1, f"Found {len(all_emails)}")

    found_email = False
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = str(to_addr or "").lower()
        subject_lower = (subject or "").lower()
        if "portfolio.manager@investment.com" not in to_str:
            continue
        body_lower = (body_text or "").lower()
        subj_ok = any(term in subject_lower for term in
                      ["financial", "q1", "health", "assessment"])
        body_ok = any(term in body_lower for term in
                      ["revenue", "profit", "выручк", "прибыл", "маржа"] + COMPANY_TERMS)
        record("Email subject mentions financial assessment", subj_ok,
               f"Subject: {subject}")
        record("Email body mentions companies and metrics", body_ok,
               "Body missing company/metric content")
        if subj_ok and body_ok:
            found_email = True
        break

    record("Письмо на portfolio.manager@investment.com с финансовой оценкой",
           found_email,
           f"Emails: {[(e[0], str(e[2])[:60]) for e in all_emails[:3]]}")
    return found_email


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_teamly()
    check_word(args.agent_workspace)
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    print(f"  Failed: {FAIL_COUNT}")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "total_passed": PASS_COUNT,
            "total_checks": total,
            "accuracy": accuracy,
            "critical_failed": critical_failed,
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
