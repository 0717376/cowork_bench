"""Оценка для sf-hr-salary-growth (форк ClickHouse).

Данные HR в sf_data (HR_ANALYTICS) русифицированы централизованно через
db/zzz_clickhouse_after_init.sql, поэтому агент читает РУССКИЕ названия отделов
(Операции/Кадры/НИОКР/Поддержка/Финансы/Продажи/Инженерия) из БД и записывает их
в Excel в качестве ключей строк листа "Salary Growth" и в значение
Highest_Avg_Increase_Dept на листе "Summary". Эталонный groundtruth
HR_Salary_Growth.xlsx переразмечен через scripts/clickhouse_relabel_map.FLAT_VALUE_MAP,
поэтому его ячейки названий отделов уже содержат те же русские значения. Имена
колонок, имена листов и метки Metric (Total_Salary_Changes/Highest_Avg_Increase_Dept/
Overall_Avg_Increase) остаются английскими токенами.

Оценка: сначала применяется гейт CRITICAL_CHECKS (любой провал критической проверки
=> FAIL), затем гейт accuracy>=70 по всем (критическим и структурным) проверкам.
"""
import argparse
import os
import sys
import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        d = (str(detail)[:300] + "...") if len(str(detail)) > 300 else detail
        msg = f": {d}" if d else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Salary_Growth.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Salary_Growth.xlsx")

    if not os.path.exists(agent_file):
        record("Файл HR_Salary_Growth.xlsx существует", False, agent_file, critical=True)
        print(f"FAIL: вывод агента не найден: {agent_file}")
        print(f"\n=== RESULT: FAIL (критические проверки: {CRITICAL_FAILS}) ===")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: эталон не найден: {gt_file}")
        sys.exit(1)

    record("Файл HR_Salary_Growth.xlsx существует", True)
    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # === Лист: Salary Growth ===
    print("\n=== Проверка: Salary Growth ===")
    a_rows = load_sheet_rows(agent_wb, "Salary Growth")
    g_rows = load_sheet_rows(gt_wb, "Salary Growth")
    if a_rows is None:
        record("Лист 'Salary Growth' присутствует", False, "нет листа", critical=True)
    elif g_rows is None:
        record("Лист 'Salary Growth' (эталон)", False, "нет листа в эталоне")
    else:
        record("Лист 'Salary Growth' присутствует", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            dept = g_row[0]
            a_row = a_lookup.get(key)
            # Наличие строки отдела + Total_Changes — КРИТИЧНО (ключ = русское название отдела из БД)
            if a_row is None:
                record(f"Строка отдела '{dept}' присутствует", False, "строка отсутствует", critical=True)
                record(f"{dept}.Total_Changes корректно", False, "нет строки", critical=True)
                record(f"{dept}.Employees_With_Changes корректно", False, "нет строки", critical=True)
                continue
            record(f"Строка отдела '{dept}' присутствует", True)

            if len(a_row) > 1 and len(g_row) > 1:
                record(f"{dept}.Employees_With_Changes корректно",
                       num_close(a_row[1], g_row[1], 2),
                       f"{a_row[1]} vs {g_row[1]} (tol=2)", critical=True)

            if len(a_row) > 2 and len(g_row) > 2:
                record(f"{dept}.Total_Changes корректно",
                       num_close(a_row[2], g_row[2], 2),
                       f"{a_row[2]} vs {g_row[2]} (tol=2)", critical=True)

            if len(a_row) > 3 and len(g_row) > 3:
                record(f"{dept}.Avg_Increase корректно",
                       num_close(a_row[3], g_row[3], 50.0),
                       f"{a_row[3]} vs {g_row[3]} (tol=50.0)")

            if len(a_row) > 4 and len(g_row) > 4:
                record(f"{dept}.Avg_Pct_Change корректно",
                       num_close(a_row[4], g_row[4], 2.0),
                       f"{a_row[4]} vs {g_row[4]} (tol=2.0)")

    # === Лист: Summary ===
    print("\n=== Проверка: Summary ===")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        record("Лист 'Summary' присутствует", False, "нет листа", critical=True)
    elif g_rows is None:
        record("Лист 'Summary' (эталон)", False, "нет листа в эталоне")
    else:
        record("Лист 'Summary' присутствует", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        # Tightened tolerances per metric; всё КРИТИЧНО (ядро задачи)
        metric_tol = {
            "total_salary_changes": 2.0,
            "highest_avg_increase_dept": 0.0,   # строковое сравнение (num_close fallback)
            "overall_avg_increase": 5.0,        # ужесточено с 50.0 -> проверяет логику взвешивания
        }
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            metric = g_row[0]
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary.{metric} присутствует", False, "строка отсутствует", critical=True)
                continue
            tol = metric_tol.get(key, 50.0)
            if len(a_row) > 1 and len(g_row) > 1:
                record(f"Summary.{metric} корректно",
                       num_close(a_row[1], g_row[1], tol),
                       f"{a_row[1]} vs {g_row[1]} (tol={tol})", critical=True)

    # === Итог ===
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nПройдено: {PASS_COUNT}/{total} ({accuracy:.1f}%)")

    if CRITICAL_FAILS:
        print(f"\n=== RESULT: FAIL (провалены критические проверки: {CRITICAL_FAILS}) ===")
        sys.exit(1)

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
