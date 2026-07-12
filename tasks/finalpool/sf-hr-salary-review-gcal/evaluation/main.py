"""
Evaluation script for sf-hr-salary-review-gcal task (ClickHouse-backed HR analytics).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Department names are read LIVE from the ClickHouse-backed sf_data schema, which
is russified centrally (Инженерия/Финансы/Кадры/Операции/НИОКР/Продажи/Поддержка).
We never hardcode dept-name realia here; expected names come from the DB.

The calendar event summary prefix 'Salary Review:' and the email subject
'Quarterly Salary Band Analysis' are kept English by the task; greps are
broadened to RU+EN to be robust.
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic checks that gate PASS regardless of accuracy. Per-dept value checks
# and the extreme-dept checks are added dynamically via critical_extra=True.
CRITICAL_CHECKS = {
    "Total_Employees == DB count",
    "Overall_Avg_Salary matches DB",
    "Highest_Avg_Dept matches DB",
    "Lowest_Avg_Dept matches DB",
    "Exactly 7 salary review events, one per day Mar 10-16",
    "Email present (subject/from/to correct)",
}


def record(name, passed, detail="", critical_extra=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")
    if critical_extra:
        CRITICAL_CHECKS.add(name)


def num_close(a, b, rel_tol=0.05, abs_tol=0.5):
    try:
        return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def get_expected_stats():
    """Query actual salary stats from the ClickHouse-backed sf_data schema."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT "DEPARTMENT",
               MIN("SALARY"),
               PERCENTILE_CONT(0.25) WITHIN GROUP (ORDER BY "SALARY"),
               PERCENTILE_CONT(0.50) WITHIN GROUP (ORDER BY "SALARY"),
               PERCENTILE_CONT(0.75) WITHIN GROUP (ORDER BY "SALARY"),
               MAX("SALARY"),
               AVG("SALARY"),
               COUNT(*)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    dept_stats = cur.fetchall()

    cur.execute('SELECT COUNT(*) FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    total_emp = cur.fetchone()[0]

    cur.execute("""
        SELECT "DEPARTMENT", AVG("SALARY") as avg_sal
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY avg_sal DESC
    """)
    dept_avgs = cur.fetchall()
    highest_avg_dept = dept_avgs[0][0]
    lowest_avg_dept = dept_avgs[-1][0]

    cur.execute('SELECT AVG("SALARY") FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    overall_avg = float(cur.fetchone()[0])

    cur.close()
    conn.close()

    return dept_stats, total_emp, highest_avg_dept, lowest_avg_dept, overall_avg


def check_excel(workspace):
    """Check Excel file against live DB values."""
    from openpyxl import load_workbook

    xlsx_path = os.path.join(workspace, "Salary_Band_Analysis.xlsx")
    if not os.path.exists(xlsx_path):
        record("Salary_Band_Analysis.xlsx exists", False, "file not found")
        return
    record("Salary_Band_Analysis.xlsx exists", True)

    dept_stats, total_emp, highest_avg_dept, lowest_avg_dept, overall_avg = get_expected_stats()
    # Map department -> (min, p25, median, p75, max, avg, count)
    expected_by_dept = {r[0]: r for r in dept_stats}

    wb = load_workbook(xlsx_path)
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # ---- Department Bands sheet ----
    if "department bands" not in sheet_names_lower:
        record("Department Bands sheet present", False, f"Found: {wb.sheetnames}")
    else:
        record("Department Bands sheet present", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("department bands")]]
        data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value is not None)
        record("Department Bands has 7 data rows", data_rows >= 7, f"got {data_rows}")

        headers = [str(cell.value).lower().replace(" ", "_") if cell.value else "" for cell in ws[1]]

        def find_col(*keys):
            for idx, h in enumerate(headers):
                if all(k in h for k in keys):
                    return idx
            return None

        dept_col = find_col("department")
        min_col = find_col("min", "salary")
        p25_col = find_col("p25")
        med_col = find_col("median")
        p75_col = find_col("p75")
        max_col = find_col("max", "salary")
        avg_col = find_col("avg", "salary")
        count_col = find_col("employee", "count")
        if count_col is None:
            count_col = find_col("count")

        for label, col in [("department", dept_col), ("min_salary", min_col),
                           ("p25_salary", p25_col), ("median_salary", med_col),
                           ("p75_salary", p75_col), ("max_salary", max_col),
                           ("avg_salary", avg_col), ("employee_count", count_col)]:
            record(f"Header present: {label}", col is not None,
                   f"headers={headers}")

        # Per-department value correctness (CRITICAL for avg + median).
        if dept_col is not None:
            for row in ws.iter_rows(min_row=2):
                if row[dept_col].value is None:
                    continue
                dept_name = str(row[dept_col].value).strip()
                exp = None
                for d, r in expected_by_dept.items():
                    if str_match(d, dept_name):
                        exp = r
                        break
                if exp is None:
                    record(f"Dept '{dept_name}' recognized", False,
                           "unknown dept name", critical_extra=True)
                    continue
                e_min, e_p25, e_med, e_p75, e_max, e_avg, e_cnt = (
                    float(exp[1]), float(exp[2]), float(exp[3]), float(exp[4]),
                    float(exp[5]), float(exp[6]), int(exp[7]))

                if avg_col is not None:
                    ok = num_close(row[avg_col].value, e_avg)
                    record(f"Avg_Salary[{dept_name}] correct", ok,
                           f"got {row[avg_col].value}, expected ~{e_avg:.0f}",
                           critical_extra=True)
                if med_col is not None:
                    ok = num_close(row[med_col].value, e_med)
                    record(f"Median_Salary[{dept_name}] correct", ok,
                           f"got {row[med_col].value}, expected ~{e_med:.0f}",
                           critical_extra=True)
                # Non-critical structural value checks for the other percentiles.
                if min_col is not None:
                    record(f"Min_Salary[{dept_name}] correct",
                           num_close(row[min_col].value, e_min),
                           f"got {row[min_col].value}, expected ~{e_min:.0f}")
                if p25_col is not None:
                    record(f"P25_Salary[{dept_name}] correct",
                           num_close(row[p25_col].value, e_p25),
                           f"got {row[p25_col].value}, expected ~{e_p25:.0f}")
                if p75_col is not None:
                    record(f"P75_Salary[{dept_name}] correct",
                           num_close(row[p75_col].value, e_p75),
                           f"got {row[p75_col].value}, expected ~{e_p75:.0f}")
                if max_col is not None:
                    record(f"Max_Salary[{dept_name}] correct",
                           num_close(row[max_col].value, e_max),
                           f"got {row[max_col].value}, expected ~{e_max:.0f}")
                if count_col is not None:
                    record(f"Employee_Count[{dept_name}] correct",
                           num_close(row[count_col].value, e_cnt, rel_tol=0.0, abs_tol=0.5),
                           f"got {row[count_col].value}, expected {e_cnt}")

    # ---- Summary sheet ----
    if "summary" not in sheet_names_lower:
        record("Summary sheet present", False, f"Found: {wb.sheetnames}")
    else:
        record("Summary sheet present", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        summary = {}
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                key = str(row[0].value).lower().replace(" ", "_")
                summary[key] = row[1].value if len(row) > 1 else None

        def get_metric(*keys):
            for k, v in summary.items():
                if all(part in k for part in keys):
                    return v
            return None

        # Total_Employees (CRITICAL)
        total_val = get_metric("total", "emp")
        if total_val is None:
            record("Total_Employees == DB count", False, "row missing")
        else:
            ok = num_close(total_val, total_emp, rel_tol=0.0, abs_tol=0.5)
            record("Total_Employees == DB count", ok,
                   f"got {total_val}, expected {total_emp}")

        # Overall_Avg_Salary (CRITICAL)
        overall_val = get_metric("overall", "avg")
        if overall_val is None:
            record("Overall_Avg_Salary matches DB", False, "row missing")
        else:
            record("Overall_Avg_Salary matches DB", num_close(overall_val, overall_avg),
                   f"got {overall_val}, expected ~{overall_avg:.0f}")

        # Highest_Avg_Dept (CRITICAL) — dept name read live from DB, auto-russified.
        high_val = get_metric("highest", "dept")
        record("Highest_Avg_Dept matches DB", str_match(high_val, highest_avg_dept),
               f"got {high_val}, expected {highest_avg_dept}")

        # Lowest_Avg_Dept (CRITICAL)
        low_val = get_metric("lowest", "dept")
        record("Lowest_Avg_Dept matches DB", str_match(low_val, lowest_avg_dept),
               f"got {low_val}, expected {lowest_avg_dept}")


def check_gcal(cur):
    """Check for 7 salary review events, one per day Mar 10-16 2026, 10:00-11:00 ET."""
    cur.execute("""
        SELECT id, summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE (LOWER(summary) LIKE '%salary review%'
               OR LOWER(summary) LIKE '%пересмотр зарплат%'
               OR LOWER(summary) LIKE '%обзор зарплат%')
          AND start_datetime >= '2026-03-10T00:00:00'
          AND start_datetime < '2026-03-17T00:00:00'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()

    record("At least 7 salary review events", len(events) >= 7,
           f"found {len(events)}")

    # Distinct days Mar 10-16.
    days = set()
    for ev in events:
        sd = ev[2]
        sd_str = sd.isoformat() if hasattr(sd, "isoformat") else str(sd)
        days.add(sd_str[:10])
    expected_days = {f"2026-03-{d:02d}" for d in range(10, 17)}
    covered = expected_days & days
    record("Exactly 7 salary review events, one per day Mar 10-16",
           len(events) == 7 and covered == expected_days,
           f"events={len(events)}, days covered={sorted(covered)}")

    # Each event names a distinct department (the 7 russified dept names live).
    cur.execute('SELECT DISTINCT "DEPARTMENT" FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    db_depts = [r[0] for r in cur.fetchall()]
    named = set()
    for ev in events:
        summ = (ev[1] or "")
        for d in db_depts:
            if d.lower() in summ.lower():
                named.add(d)
    record("All 7 departments named across events", len(named) == len(db_depts),
           f"named {sorted(named)}")


def check_email(cur):
    """Check the summary email: subject, from, to, and body mentions extremes."""
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%quarterly salary band analysis%'
        ORDER BY id DESC
    """)
    emails = cur.fetchall()

    if not emails:
        record("Email present (subject/from/to correct)", False,
               "no email with subject 'Quarterly Salary Band Analysis'")
        record("Email body mentions highest & lowest dept", False, "no email")
        return

    email = emails[0]
    from_str = str(email[1] or "").lower()
    to_str = str(email[2] or "").lower()
    body = str(email[3] or "").lower()

    ok = ("hr-analytics@company.com" in from_str
          and "hr-director@company.com" in to_str)
    record("Email present (subject/from/to correct)", ok,
           f"from={email[1]}, to={email[2]}")

    # Body must mention both highest and lowest avg-salary departments.
    cur.execute("""
        SELECT "DEPARTMENT", AVG("SALARY") avg_sal
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT" ORDER BY avg_sal DESC
    """)
    avgs = cur.fetchall()
    highest = avgs[0][0]
    lowest = avgs[-1][0]
    body_ok = highest.lower() in body and lowest.lower() in body
    record("Email body mentions highest & lowest dept", body_ok,
           f"expected mentions of '{highest}' and '{lowest}'")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("\n=== Checking Excel ===")
    try:
        check_excel(args.agent_workspace)
    except Exception as e:
        record("Excel check ran", False, f"exception: {e}")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        print("\n=== Checking GCal Events ===")
        check_gcal(cur)

        print("\n=== Checking Email ===")
        check_email(cur)

        cur.close()
        conn.close()
    except Exception as e:
        record("DB checks ran", False, f"exception: {e}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "total": total,
            "accuracy": accuracy,
            "critical_failed": critical_failed,
            "failed": FAILED_NAMES,
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
