"""
Evaluation for sf-hr-satisfaction-excel-gcal task.

Checks:
1. Excel file Employee_Satisfaction.xlsx with correct data
2. Google Calendar events for 2 lowest-satisfaction departments
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0

# Russified department names from the central ClickHouse substitution map
# (db/zzz_clickhouse_after_init.sql). The agent reads these russified names
# from the DB and writes them into Excel / calendar events, so the eval and
# groundtruth must key on the SAME RU values. Do NOT hand-translate elsewhere.
RU_DEPT_LOWEST = "ниокр"        # R&D  -> НИОКР  (lowest avg job satisfaction)
RU_DEPT_SECOND_LOWEST = "поддержка"  # Support -> Поддержка (2nd lowest)
RU_DEPT_HIGHEST = "финансы"     # Finance -> Финансы (highest avg job satisfaction)

# Critical checks: any failure here => immediate FAIL (sys.exit(1)) regardless
# of the overall accuracy gate. These reflect the core analytical deliverable.
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{'[CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        d = (detail[:300]) if len(detail) > 300 else detail
        print(f"  [FAIL]{'[CRITICAL]' if critical else ''} {name}: {d}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


# Epsilon for statistical ties on Avg_Job_Satisfaction. The deliverable rounds
# this metric to 2 decimals (granularity 0.01), so two departments whose
# *rounded* Avg_Job_Satisfaction are equal are indistinguishable in the
# deliverable: the agent cannot know which one is the true argmax. TIE_EPS is
# half the rounding step (0.005) so that only departments sharing the same
# rounded value are treated as tied. For the argmax selections (highest /
# 2nd-lowest department) any such tied department is an acceptable answer; we
# MUST NOT force the agent to reproduce the hidden-groundtruth winner of a tie
# that is invisible at the deliverable's own precision. This does NOT broaden to
# departments that are genuinely lower (e.g. 6.58 vs 6.59 are NOT tied).
TIE_EPS = 0.005


def _gt_dept_satisfaction(g_wb):
    """Return {ru_dept_lower: avg_job_satisfaction(float)} from the groundtruth
    'Department Satisfaction' sheet (col index 2)."""
    rows = load_sheet_rows(g_wb, "Department Satisfaction") or []
    out = {}
    for r in rows[1:]:
        if not r or r[0] is None or len(r) <= 2:
            continue
        try:
            out[str(r[0]).strip().lower()] = float(r[2])
        except (TypeError, ValueError):
            continue
    return out


def _tie_set(dept_sat, target_dept):
    """All dept keys whose Avg_Job_Satisfaction is within TIE_EPS of the
    groundtruth target department's value (the target itself included)."""
    base = dept_sat.get(str(target_dept).strip().lower())
    if base is None:
        return {str(target_dept).strip().lower()}
    return {d for d, v in dept_sat.items() if abs(v - base) <= TIE_EPS}


def _highest_tie_set(dept_sat):
    if not dept_sat:
        return set()
    top = max(dept_sat.values())
    return {d for d, v in dept_sat.items() if abs(v - top) <= TIE_EPS}


def _second_lowest_tie_set(dept_sat):
    """Tie set for the 2nd-lowest department: the lowest dept is excluded, then
    every remaining dept within TIE_EPS of the next-lowest value is accepted."""
    if len(dept_sat) < 2:
        return set(dept_sat.keys())
    ordered = sorted(dept_sat.items(), key=lambda kv: kv[1])
    lowest_key = ordered[0][0]
    rest = {d: v for d, v in dept_sat.items() if d != lowest_key}
    second = min(rest.values())
    return {d for d, v in rest.items() if abs(v - second) <= TIE_EPS}


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Excel File ===")
    agent_file = os.path.join(agent_workspace, "Employee_Satisfaction.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Employee_Satisfaction.xlsx")

    if not os.path.exists(agent_file):
        check("Excel file exists", False, f"Not found: {agent_file}")
        return
    check("Excel file exists", True)

    if not os.path.exists(gt_file):
        check("Groundtruth file exists", False, f"Not found: {gt_file}")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # Check Department Satisfaction sheet
    print("  Checking Department Satisfaction sheet...")
    a_rows = load_sheet_rows(agent_wb, "Department Satisfaction")
    g_rows = load_sheet_rows(gt_wb, "Department Satisfaction")

    if a_rows is None:
        check("Sheet 'Department Satisfaction' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth sheet exists", False, "Not found")
    else:
        check("Sheet 'Department Satisfaction' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        check("Row count matches", len(a_data) == len(g_data),
              f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Row '{g_row[0]}'", False, "Missing")
                continue

            # Employee_Count (col 1)
            if len(a_row) > 1 and len(g_row) > 1:
                check(f"{key}.Employee_Count",
                      num_close(a_row[1], g_row[1], 5),
                      f"{a_row[1]} vs {g_row[1]}")

            # Avg_Job_Satisfaction (col 2)
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Avg_Job_Satisfaction",
                      num_close(a_row[2], g_row[2], 0.1),
                      f"{a_row[2]} vs {g_row[2]}")

            # Avg_Work_Life_Balance (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Avg_Work_Life_Balance",
                      num_close(a_row[3], g_row[3], 0.1),
                      f"{a_row[3]} vs {g_row[3]}")

            # Low_Satisfaction_Count (col 4)
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"{key}.Low_Satisfaction_Count",
                      num_close(a_row[4], g_row[4], 10),
                      f"{a_row[4]} vs {g_row[4]}")

            # Low_Satisfaction_Pct (col 5)
            if len(a_row) > 5 and len(g_row) > 5:
                check(f"{key}.Low_Satisfaction_Pct",
                      num_close(a_row[5], g_row[5], 1.0),
                      f"{a_row[5]} vs {g_row[5]}")

    # Check Summary sheet
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")

    if a_rows is None:
        check("Sheet 'Summary' exists", False, "Not found")
    elif g_rows is None:
        check("Groundtruth Summary sheet exists", False, "Not found")
    else:
        check("Sheet 'Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Summary: '{g_row[0]}'", False, "Missing")
                continue
            g_val = g_row[1] if len(g_row) > 1 else None
            a_val = a_row[1] if len(a_row) > 1 else None
            # Try numeric comparison first, then string
            is_numeric = True
            try:
                float(g_val)
            except (TypeError, ValueError):
                is_numeric = False
            if is_numeric:
                check(f"Summary: {key}",
                      num_close(a_val, g_val, 5),
                      f"{a_val} vs {g_val}")
            else:
                check(f"Summary: {key}",
                      str_match(a_val, g_val),
                      f"{a_val} vs {g_val}")


def check_excel_critical(agent_workspace, groundtruth_workspace):
    """Semantic CRITICAL checks on the Excel deliverable.

    The dept names below are the central-map RU values; groundtruth was
    regenerated through the same substitution, so both key on RU names.
    """
    print("\n=== CRITICAL: Excel semantics ===")
    agent_file = os.path.join(agent_workspace, "Employee_Satisfaction.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Employee_Satisfaction.xlsx")
    if not (os.path.exists(agent_file) and os.path.exists(gt_file)):
        check("CRITICAL Excel files present", False,
              f"agent={os.path.exists(agent_file)} gt={os.path.exists(gt_file)}",
              critical=True)
        return

    a_wb = openpyxl.load_workbook(agent_file, data_only=True)
    g_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # --- Summary critical values ---
    a_sum = load_sheet_rows(a_wb, "Summary") or []
    g_sum = load_sheet_rows(g_wb, "Summary") or []
    a_smap = {str(r[0]).strip().lower(): (r[1] if len(r) > 1 else None)
              for r in a_sum[1:] if r and r[0] is not None}
    g_smap = {str(r[0]).strip().lower(): (r[1] if len(r) > 1 else None)
              for r in g_sum[1:] if r and r[0] is not None}

    # Lowest/highest satisfaction department names (core deliverable)
    check("CRITICAL Summary.Department_Lowest_Satisfaction == lowest dept",
          str_match(a_smap.get("department_lowest_satisfaction"),
                    g_smap.get("department_lowest_satisfaction")),
          f"{a_smap.get('department_lowest_satisfaction')} vs "
          f"{g_smap.get('department_lowest_satisfaction')} (expect RU '{RU_DEPT_LOWEST}')",
          critical=True)
    # Highest-satisfaction department is a statistical tie at the rounding
    # granularity (e.g. Финансы 6.59320 vs Продажи 6.58742, both round to 6.59).
    # Accept any department within TIE_EPS of the groundtruth maximum.
    g_dept_sat = _gt_dept_satisfaction(g_wb)
    highest_ties = _highest_tie_set(g_dept_sat)
    a_highest = a_smap.get("department_highest_satisfaction")
    a_highest_key = str(a_highest).strip().lower() if a_highest is not None else None
    check("CRITICAL Summary.Department_Highest_Satisfaction == highest dept (tie-aware)",
          a_highest_key in highest_ties,
          f"{a_highest} not in highest tie-set {sorted(highest_ties)} "
          f"(gt='{g_smap.get('department_highest_satisfaction')}')",
          critical=True)

    # Full-dataset confirmation: Total_Employees and Employees_At_Risk
    check("CRITICAL Summary.Total_Employees == 50000 (full dataset)",
          num_close(a_smap.get("total_employees"),
                    g_smap.get("total_employees"), 0),
          f"{a_smap.get('total_employees')} vs {g_smap.get('total_employees')}",
          critical=True)
    check("CRITICAL Summary.Employees_At_Risk matches groundtruth",
          num_close(a_smap.get("employees_at_risk"),
                    g_smap.get("employees_at_risk"), 50),
          f"{a_smap.get('employees_at_risk')} vs {g_smap.get('employees_at_risk')}",
          critical=True)

    # --- Satisfaction ranking: the two lowest rows' Avg_Job_Satisfaction ---
    a_dept = load_sheet_rows(a_wb, "Department Satisfaction") or []
    g_dept = load_sheet_rows(g_wb, "Department Satisfaction") or []
    a_dmap = {str(r[0]).strip().lower(): r for r in a_dept[1:]
              if r and r[0] is not None}
    g_dmap = {str(r[0]).strip().lower(): r for r in g_dept[1:]
              if r and r[0] is not None}
    for ru_key in (RU_DEPT_LOWEST, RU_DEPT_SECOND_LOWEST):
        a_r = a_dmap.get(ru_key)
        g_r = g_dmap.get(ru_key)
        ok = (a_r is not None and g_r is not None
              and len(a_r) > 2 and len(g_r) > 2
              and num_close(a_r[2], g_r[2], 0.1))
        check(f"CRITICAL '{ru_key}'.Avg_Job_Satisfaction correct",
              ok,
              f"{(a_r[2] if a_r and len(a_r) > 2 else None)} vs "
              f"{(g_r[2] if g_r and len(g_r) > 2 else None)}",
              critical=True)


def check_gcal(second_lowest_ties=None):
    print("\n=== Checking Google Calendar ===")
    if not second_lowest_ties:
        second_lowest_ties = {RU_DEPT_SECOND_LOWEST}
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("DB connection for gcal check", False, str(e))
        return

    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        WHERE LOWER(summary) LIKE '%%wellness review%%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    check("At least 2 wellness review events", len(events) >= 2,
          f"Found {len(events)}")

    # The lowest-satisfaction department (НИОКР) is unambiguous and CRITICAL.
    found_lowest = any(
        RU_DEPT_LOWEST in (e[0] or "").lower() or RU_DEPT_LOWEST in (e[1] or "").lower()
        for e in events)
    check(f"CRITICAL Wellness Review event for department '{RU_DEPT_LOWEST}'",
          found_lowest, f"Not found among {len(events)} events", critical=True)

    # The 2nd-lowest department is a statistical tie at the rounding granularity
    # (e.g. Поддержка 6.52924 vs Операции 6.53205, both round to 6.53). Accept a
    # wellness-review event for ANY department in the 2nd-lowest tie set.
    found_second = any(
        any(dept in (e[0] or "").lower() or dept in (e[1] or "").lower()
            for dept in second_lowest_ties)
        for e in events)
    check("CRITICAL Wellness Review event for 2nd-lowest dept (tie-aware)",
          found_second,
          f"None of {sorted(second_lowest_ties)} found among {len(events)} events",
          critical=True)

    # Exactly the two lowest-satisfaction departments should have events
    # (no extra departments scheduled).
    check("Exactly 2 wellness review events", len(events) == 2,
          f"Found {len(events)}")

    # Check dates: lowest dept on the earlier date (2026-03-13), second on 03-14.
    if len(events) >= 2:
        d1 = events[0][2]
        d2 = events[1][2]
        e1_text = ((events[0][0] or "") + " " + (events[0][1] or "")).lower()
        check("CRITICAL First wellness-review event on 2026-03-13 is lowest dept",
              d1 is not None and "2026-03-13" in str(d1)
              and RU_DEPT_LOWEST in e1_text,
              f"date={d1}, text={e1_text!r}", critical=True)
        check("Second event on 2026-03-14",
              d2 is not None and "2026-03-14" in str(d2),
              f"Got {d2}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_excel_critical(args.agent_workspace, gt_dir)

    # Derive the 2nd-lowest-satisfaction tie set from the groundtruth Excel so
    # the calendar check accepts any statistically-tied department.
    second_lowest_ties = {RU_DEPT_SECOND_LOWEST}
    gt_file = os.path.join(gt_dir, "Employee_Satisfaction.xlsx")
    if os.path.exists(gt_file):
        try:
            g_wb = openpyxl.load_workbook(gt_file, data_only=True)
            second_lowest_ties = _second_lowest_tie_set(_gt_dept_satisfaction(g_wb))
        except Exception as e:
            print(f"  [WARN] could not derive 2nd-lowest tie set: {e}")
    check_gcal(second_lowest_ties)

    total_pass = PASS_COUNT
    total_fail = FAIL_COUNT
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (100.0 * PASS_COUNT / total) if total else 0.0

    critical_ok = len(CRITICAL_FAILED) == 0
    all_ok = critical_ok and accuracy >= 70.0

    print(f"\n=== SUMMARY ===")
    print(f"  Total checks - Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}% (threshold 70%)")
    if not critical_ok:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILED}")
    print(f"  Overall: {'PASS' if all_ok else 'FAIL'}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": total_pass, "failed": total_fail, "success": all_ok}, f, indent=2)

    sys.exit(0 if all_ok else 1)


if __name__ == "__main__":
    main()
