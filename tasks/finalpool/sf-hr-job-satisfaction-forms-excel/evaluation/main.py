"""Evaluation for sf-hr-job-satisfaction-gform-excel.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Note: department data-values (Финансы/Инженерия/...) are russified centrally in
db/zzz_clickhouse_after_init.sql and read here via a LIVE query to sf_data, so
seed<->eval stay in sync automatically. Do NOT hardcode department literals.
"""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Employee_Satisfaction_Analysis.xlsx exists",
    "Department Scores: per-department Avg/Combined match DB",
    "Summary: Highest/Lowest_Satisfaction_Dept correct",
    "Summary: Overall_Avg_Satisfaction and Overall_Avg_WLB match DB",
    "Summary: Total_Employees matches DB count",
    "Forms: Employee Wellbeing Survey 2026 has 4 questions (2 text + 2 choice)",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def num_close(a, b, tol):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_satisfaction():
    """Query actual satisfaction data from the live DB (auto-syncs RU names)."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               COUNT(*) as emp_count,
               AVG("JOB_SATISFACTION") as avg_js,
               AVG("WORK_LIFE_BALANCE") as avg_wlb
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    rows = cur.fetchall()
    cur.execute('SELECT COUNT(*), AVG("JOB_SATISFACTION"), AVG("WORK_LIFE_BALANCE") FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"')
    totals = cur.fetchone()
    cur.close()
    conn.close()
    return rows, totals


def check_excel(agent_workspace):
    try:
        import openpyxl
    except ImportError:
        check("Employee_Satisfaction_Analysis.xlsx exists", False, "openpyxl not installed")
        return

    agent_file = os.path.join(agent_workspace, "Employee_Satisfaction_Analysis.xlsx")
    exists = os.path.exists(agent_file)
    check("Employee_Satisfaction_Analysis.xlsx exists", exists, agent_file)
    if not exists:
        check("Department Scores: per-department Avg/Combined match DB", False, "no file")
        check("Summary: Highest/Lowest_Satisfaction_Dept correct", False, "no file")
        check("Summary: Overall_Avg_Satisfaction and Overall_Avg_WLB match DB", False, "no file")
        check("Summary: Total_Employees matches DB count", False, "no file")
        return

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)

    try:
        dept_rows, totals = get_expected_satisfaction()
    except Exception as e:
        check("Department Scores: per-department Avg/Combined match DB", False, f"DB error: {e}")
        check("Summary: Highest/Lowest_Satisfaction_Dept correct", False, f"DB error: {e}")
        check("Summary: Overall_Avg_Satisfaction and Overall_Avg_WLB match DB", False, f"DB error: {e}")
        check("Summary: Total_Employees matches DB count", False, f"DB error: {e}")
        return

    total_emp = int(totals[0])
    overall_avg_js = round(float(totals[1]), 2)
    overall_avg_wlb = round(float(totals[2]), 2)

    dept_data = {}
    for r in dept_rows:
        dept = str(r[0]).strip().lower()
        dept_data[dept] = {
            "count": int(r[1]),
            "avg_js": round(float(r[2]), 2),
            "avg_wlb": round(float(r[3]), 2),
            "combined": round((float(r[2]) + float(r[3])) / 2, 2),
        }

    sorted_depts = sorted(dept_data.items(), key=lambda x: x[1]["combined"], reverse=True)
    highest_dept = sorted_depts[0][0]
    lowest_dept = sorted_depts[-1][0]

    # ---- Department Scores sheet ----
    a_rows = load_sheet_rows(agent_wb, "Department Scores")
    dept_errors = []
    if a_rows is None:
        dept_errors.append("sheet 'Department Scores' missing")
    else:
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r[0]}
        for dept_key, expected in dept_data.items():
            a_row = a_lookup.get(dept_key)
            if a_row is None:
                dept_errors.append(f"missing department {dept_key}")
                continue
            # Required columns must be present (no silent skip): row must have >=5 cols.
            if len(a_row) < 5:
                dept_errors.append(f"{dept_key}: missing columns (row len {len(a_row)})")
                continue
            if not num_close(a_row[2], expected["avg_js"], 0.05):
                dept_errors.append(f"{dept_key} Avg_Job_Satisfaction got {a_row[2]} exp {expected['avg_js']}")
            if not num_close(a_row[3], expected["avg_wlb"], 0.05):
                dept_errors.append(f"{dept_key} Avg_Work_Life_Balance got {a_row[3]} exp {expected['avg_wlb']}")
            if not num_close(a_row[4], expected["combined"], 0.05):
                dept_errors.append(f"{dept_key} Combined_Score got {a_row[4]} exp {expected['combined']}")
    check("Department Scores: per-department Avg/Combined match DB", not dept_errors,
          "; ".join(dept_errors[:6]))

    # Non-critical: Employee_Count per department.
    if a_rows is not None:
        cnt_errors = []
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data if r[0]}
        for dept_key, expected in dept_data.items():
            a_row = a_lookup.get(dept_key)
            if a_row is not None and len(a_row) > 1 and not num_close(a_row[1], expected["count"], 1):
                cnt_errors.append(f"{dept_key} count got {a_row[1]} exp {expected['count']}")
        check("Department Scores: Employee_Count per department matches DB", not cnt_errors,
              "; ".join(cnt_errors[:6]))

    # ---- Summary sheet ----
    a_sum = load_sheet_rows(agent_wb, "Summary")
    if a_sum is None:
        check("Summary: Highest/Lowest_Satisfaction_Dept correct", False, "Summary sheet missing")
        check("Summary: Overall_Avg_Satisfaction and Overall_Avg_WLB match DB", False, "Summary sheet missing")
        check("Summary: Total_Employees matches DB count", False, "Summary sheet missing")
        return

    a_sum_data = {str(r[0]).strip().lower(): r[1]
                  for r in (a_sum[1:] if len(a_sum) > 1 else []) if r and r[0]}

    # Highest / Lowest by combined score (exact, case-insensitive)
    hsd = a_sum_data.get("highest_satisfaction_dept")
    lsd = a_sum_data.get("lowest_satisfaction_dept")
    hl_ok = (hsd is not None and str(hsd).strip().lower() == highest_dept
             and lsd is not None and str(lsd).strip().lower() == lowest_dept)
    check("Summary: Highest/Lowest_Satisfaction_Dept correct", hl_ok,
          f"got high='{hsd}' low='{lsd}', exp high='{highest_dept}' low='{lowest_dept}'")

    # Overall_Avg_Satisfaction AND Overall_Avg_WLB (closes desync with task.md)
    oas = a_sum_data.get("overall_avg_satisfaction")
    owlb = a_sum_data.get("overall_avg_wlb")
    avg_ok = num_close(oas, overall_avg_js, 0.05) and num_close(owlb, overall_avg_wlb, 0.05)
    check("Summary: Overall_Avg_Satisfaction and Overall_Avg_WLB match DB", avg_ok,
          f"got sat={oas} wlb={owlb}, exp sat={overall_avg_js} wlb={overall_avg_wlb}")

    # Total_Employees == COUNT(*) with tight tolerance
    te = a_sum_data.get("total_employees")
    check("Summary: Total_Employees matches DB count", num_close(te, total_emp, 1),
          f"got {te}, exp {total_emp}")


def check_forms():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT f.id, f.title
            FROM gform.forms f
            WHERE LOWER(f.title) LIKE '%%wellbeing%%'
               OR LOWER(f.title) LIKE '%%employee%%survey%%'
               OR LOWER(f.title) LIKE '%%благополуч%%'
               OR (LOWER(f.title) LIKE '%%опрос%%' AND LOWER(f.title) LIKE '%%сотрудник%%')
        """)
        forms = cur.fetchall()
        if not forms:
            check("Forms: Employee Wellbeing Survey 2026 has 4 questions (2 text + 2 choice)",
                  False, "form not found")
            cur.close()
            conn.close()
            return
        form_id = forms[0][0]
        cur.execute("SELECT title, question_type, required FROM gform.questions WHERE form_id = %s ORDER BY position", (form_id,))
        qs = cur.fetchall()
        cur.close()
        conn.close()

        def is_text(t):
            return (t or "") in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH")

        def is_choice(t):
            return (t or "") in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE", "CHOICE", "CHECKBOX")

        types = [q[1] for q in qs]
        text_count = sum(1 for t in types if is_text(t))
        choice_count = sum(1 for t in types if is_choice(t))
        # department text + improvement text = 2 text; two 1-5 rating = 2 choice.
        struct_ok = (len(qs) >= 4 and text_count >= 2 and choice_count >= 2)
        check("Forms: Employee Wellbeing Survey 2026 has 4 questions (2 text + 2 choice)",
              struct_ok, f"n={len(qs)} text={text_count} choice={choice_count} types={types}")

        # Non-critical: at least 3 required questions (department + 2 ratings).
        req_count = sum(1 for q in qs if q[2])
        check("Forms: at least 3 required questions", req_count >= 3,
              f"required={req_count}")
    except Exception as e:
        check("Forms: Employee Wellbeing Survey 2026 has 4 questions (2 text + 2 choice)",
              False, f"DB error: {e}")


def check_email():
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, COALESCE(body_text, '')
            FROM email.messages
            WHERE LOWER(subject) LIKE '%%satisfaction%%'
               OR LOWER(subject) LIKE '%%wellbeing%%'
               OR LOWER(subject) LIKE '%%удовлетвор%%'
               OR LOWER(subject) LIKE '%%благополуч%%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Email to hr.team@company.com about satisfaction analysis", False, f"DB error: {e}")
        return

    found = None
    for subj, to_addr, body in emails:
        addrs = to_addr
        if isinstance(addrs, str):
            try:
                addrs = json.loads(addrs)
            except json.JSONDecodeError:
                addrs = [addrs]
        if not isinstance(addrs, list):
            addrs = [addrs]
        if any("hr.team" in str(a).lower() for a in addrs):
            found = (subj, body)
            break
    check("Email to hr.team@company.com about satisfaction analysis", found is not None,
          f"matching emails: {len(emails)}")
    # Non-critical: body mentions the Excel report.
    if found is not None:
        body = found[1].lower()
        mentions_excel = any(k in body for k in
                             ("excel", "employee_satisfaction_analysis", ".xlsx",
                              "отчёт", "отчет", "таблиц", "файл"))
        check("Email body references the Excel report", mentions_excel, found[1][:120])


def check_teamly():
    """Page titled 'Employee Satisfaction Q1 2026' (EN or RU) created in Teamly."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Seed sample pages have id <= 3; the agent's page is a new one.
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly page 'Employee Satisfaction Q1 2026' exists", False, f"DB error: {e}")
        return

    found = None
    for pid, title, body in pages:
        blob = ((title or "") + " " + (body or "")).lower()
        if ("employee satisfaction" in blob or "satisfaction q1" in blob
                or ("удовлетвор" in blob and ("q1" in blob or "сотрудник" in blob))):
            found = (pid, title)
            break
    check("Teamly page 'Employee Satisfaction Q1 2026' exists", found is not None,
          f"new pages: {[(p[0], p[1]) for p in pages]}")


def run_evaluation(agent_workspace, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    print("\n=== Checking Excel ===")
    check_excel(agent_workspace)
    print("\n=== Checking Forms ===")
    check_forms()
    print("\n=== Checking Email ===")
    check_email()
    print("\n=== Checking Teamly Page ===")
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                    "success": (not critical_failed) and accuracy >= 70,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(args.agent_workspace, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
