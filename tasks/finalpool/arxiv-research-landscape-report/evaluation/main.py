"""
Evaluation script for arxiv-research-landscape-report task.

Checks:
1. Excel file (Research_Landscape.xlsx) with 3 sheets
2. Paper Analysis has 5 target papers with correct IDs and citation counts
3. Conference Fit has 3 conferences
4. Summary has required metrics (Total_Papers, Total_Conferences,
   Best_Conference_Fit, Highest_Cited_Paper, Average_Citations)
5. Word document (Landscape_Report.docx) exists with substantive content

Scoring:
- A subset of checks is marked CRITICAL (the core semantic deliverables).
  Any critical failure => immediate FAIL (sys.exit(1)).
- Otherwise PASS requires accuracy (passed / total) >= 70%.
"""
import argparse
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

EXPECTED_PAPERS = {
    "2401.00001": {"title": "Efficient Transformers for NLP", "citation_count": 350},
    "2401.00002": {"title": "Deep RL with Human Feedback", "citation_count": 520},
    "2401.00003": {"title": "Generative Models for Code", "citation_count": 280},
    "2401.00004": {"title": "Knowledge Graph Embeddings", "citation_count": 190},
    "2401.00005": {"title": "Optimization in Deep Learning", "citation_count": 150},
}

NOISE_IDS = {"2401.00006", "2401.00007", "2401.00008"}

CONFERENCES = ["NeurIPS 2026", "ICML 2026", "AAAI 2026"]


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")


def num_close(a, b, tol=50):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def check_excel(workspace):
    print("\n=== Checking Excel Output ===")
    path = os.path.join(workspace, "Research_Landscape.xlsx")
    # CRITICAL precondition: the core Excel deliverable must exist. A missing
    # file must fail closed (sys.exit(1)) instead of skipping the 6 Excel
    # critical checks below (which only fire when the file is present).
    if not os.path.isfile(path):
        record("Excel file exists", False, f"Not found: {path}", critical=True)
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e), critical=True)
        return
    record("Excel readable", True)

    # --- Paper Analysis sheet ---
    pa_rows = load_sheet_rows(wb, "Paper Analysis") or load_sheet_rows(wb, "Paper_Analysis")
    if pa_rows is None:
        record("Sheet 'Paper Analysis' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Paper Analysis' exists", True)
        header = pa_rows[0] if pa_rows else []
        data = pa_rows[1:]

        id_col = find_col(header, ["Paper_ID", "Paper ID", "ID", "paper_id"])
        cite_col = find_col(header, ["Citation_Count", "Citation Count", "Citations", "citation_count"])

        record("Paper Analysis has 5 data rows", len(data) == 5, f"Found {len(data)}")

        if id_col is not None:
            found_ids = {str(r[id_col]).strip() for r in data if id_col < len(r) and r[id_col]}
            # CRITICAL: exactly the 5 target IDs and none of the 3 noise IDs
            targets_present = all(eid in found_ids for eid in EXPECTED_PAPERS)
            noise_absent = all(nid not in found_ids for nid in NOISE_IDS)
            record(
                "Paper Analysis contains exactly the 5 target IDs and no noise IDs",
                targets_present and noise_absent,
                f"Found: {sorted(found_ids)}",
                critical=True,
            )
        else:
            record("Paper_ID column found", False, f"Header: {header}", critical=True)

        if cite_col is not None and id_col is not None:
            for row in data:
                pid = str(row[id_col]).strip() if id_col < len(row) and row[id_col] else ""
                if pid in EXPECTED_PAPERS:
                    ok = num_close(row[cite_col] if cite_col < len(row) else None,
                                   EXPECTED_PAPERS[pid]["citation_count"], tol=50)
                    record(f"Citation count for {pid}", ok,
                           f"Got {row[cite_col] if cite_col < len(row) else None}, expected {EXPECTED_PAPERS[pid]['citation_count']}")

    # --- Conference Fit sheet ---
    cf_rows = load_sheet_rows(wb, "Conference Fit") or load_sheet_rows(wb, "Conference_Fit")
    if cf_rows is None:
        record("Sheet 'Conference Fit' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Conference Fit' exists", True)
        data = cf_rows[1:]
        record("Conference Fit has 3 rows", len(data) == 3, f"Found {len(data)}")

    # --- Summary sheet ---
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Summary' exists", True)
        metrics = {}
        for row in sum_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        # CRITICAL: Total Papers == 5
        tp_key = next((k for k in metrics if "total" in k and "paper" in k), None)
        if tp_key:
            record("Summary: Total_Papers = 5", num_close(metrics[tp_key], 5, tol=0),
                   f"Got {metrics[tp_key]}", critical=True)
        else:
            record("Summary: Total_Papers exists", False, critical=True)

        # CRITICAL: Total Conferences == 3
        tc_key = next((k for k in metrics if "total" in k and ("conf" in k or "venue" in k)), None)
        if tc_key:
            record("Summary: Total_Conferences = 3", num_close(metrics[tc_key], 3, tol=0),
                   f"Got {metrics[tc_key]}", critical=True)
        else:
            record("Summary: Total_Conferences exists", False,
                   f"Metrics: {list(metrics)}", critical=True)

        # CRITICAL: Average Citations ~= 298.0
        avg_key = next((k for k in metrics if "avg" in k or "average" in k), None)
        expected_avg = (350 + 520 + 280 + 190 + 150) / 5  # 298.0
        if avg_key:
            record("Summary: Average_Citations ~= 298.0",
                   num_close(metrics[avg_key], expected_avg, tol=5),
                   f"Got {metrics[avg_key]}, expected ~{expected_avg}", critical=True)
        else:
            record("Summary: Average_Citations exists", False, critical=True)

        # CRITICAL: Highest Cited Paper -> Deep RL / Human Feedback (2401.00002)
        hc_key = next((k for k in metrics if "highest" in k or "most" in k), None)
        if hc_key:
            val = str(metrics[hc_key]).lower() if metrics[hc_key] else ""
            ok = "deep rl" in val or "human feedback" in val or "2401.00002" in val or "rl with" in val
            record("Summary: Highest_Cited is Deep RL / Human Feedback paper", ok,
                   f"Got: {metrics[hc_key]}", critical=True)
        else:
            record("Summary: Highest_Cited_Paper exists", False, critical=True)

        # CRITICAL: Best Conference Fit -> a tied-max conference.
        # All 3 conferences have exactly 2 matching papers (a genuine 3-way
        # tie), so any of them is a correct tie-break. Accept any scheduled
        # conference name; reject empty/unknown/noise values.
        bf_key = next((k for k in metrics if "best" in k and ("conf" in k or "fit" in k or "venue" in k)), None)
        if bf_key:
            val = str(metrics[bf_key]).lower() if metrics[bf_key] else ""
            ok = any(c in val for c in ["neurips", "icml", "aaai"])
            record("Summary: Best_Conference_Fit is a tied-max conference (NeurIPS/ICML/AAAI)", ok,
                   f"Got: {metrics[bf_key]}", critical=True)
        else:
            record("Summary: Best_Conference_Fit exists", False,
                   f"Metrics: {list(metrics)}", critical=True)


def check_word(workspace):
    print("\n=== Checking Word Document ===")
    path = os.path.join(workspace, "Landscape_Report.docx")
    if not os.path.isfile(path):
        record("Word document exists", False, f"Not found: {path}", critical=True)
        return
    record("Word document exists", True)

    try:
        from docx import Document
        doc = Document(path)
        # Use ORIGINAL lowercased text (NOT normalize) for RU+EN keyword checks.
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
    except Exception as e:
        record("Word document readable", False, str(e), critical=True)
        return

    has_content = len(full_text) > 200
    # Conference name (venue names stay English in the report as identifiers).
    mentions_conf = any(c in full_text for c in ["neurips", "icml", "aaai", "конференц"])
    # Topic keyword: RU OR EN alternatives.
    topic_kws = [
        "transformer", "трансформер",
        "reinforcement", "обучение с подкреплением", "подкреплени",
        "optimization", "оптимизац",
        "knowledge graph", "граф знаний", "графов знаний",
        "generative", "генеративн",
        "nlp", "обработк",  # natural language processing / обработка языка
    ]
    mentions_topic = any(kw in full_text for kw in topic_kws)

    record("Document has substantial content", has_content, f"Only {len(full_text)} chars")
    record("Mentions research/landscape",
           any(kw in full_text for kw in ["research", "landscape", "ландшафт", "исследован", "обзор"]),
           "Missing research/landscape keywords")

    # CRITICAL: substantive content AND a conference name AND a topic keyword (RU or EN).
    record(
        "Report is substantive and mentions a conference and a topic keyword (RU/EN)",
        has_content and mentions_conf and mentions_topic,
        f"content={has_content}, conf={mentions_conf}, topic={mentions_topic}",
        critical=True,
    )


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Total: {total}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL checks failed: {CRITICAL_FAILED}")
        print("  Result: FAIL (critical check failed)")
        sys.exit(1)

    if accuracy >= 70:
        print("  Result: PASS")
        sys.exit(0)
    else:
        print("  Result: FAIL (accuracy < 70%)")
        sys.exit(1)


if __name__ == "__main__":
    main()
