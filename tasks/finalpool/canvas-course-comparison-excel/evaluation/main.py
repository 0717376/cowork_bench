"""Evaluation for canvas-course-comparison-excel."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.lower() in str(haystack).lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_dir = args.agent_workspace or os.path.join(task_root, "initial_workspace")

    agent_file = os.path.join(agent_dir, "Course_Comparison_Fall2013.xlsx")
    gt_file = os.path.join(gt_dir, "Course_Comparison_Fall2013.xlsx")

    file_errors = []
    db_errors = []
    critical_failures = []

    if not os.path.exists(agent_file):
        file_errors.append(f"Agent output not found: {agent_file}")
        critical_failures.append(f"Agent output not found: {agent_file}")
    if not os.path.exists(gt_file):
        file_errors.append(f"Groundtruth not found: {gt_file}")

    # Will hold parsed sheet data for critical checks
    a_course_lookup = {}
    g_course_data = []
    a_summary_lookup = {}
    g_summary_lookup = {}

    if not file_errors:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Course Stats sheet
        print("  Checking Course Stats...")
        a_rows = load_sheet_rows(agent_wb, "Course Stats")
        g_rows = load_sheet_rows(gt_wb, "Course Stats")
        if a_rows is None:
            file_errors.append("Sheet 'Course Stats' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Course Stats' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            if len(a_data) != len(g_data):
                file_errors.append(f"Course Stats row count: agent {len(a_data)} vs gt {len(g_data)}")

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row
            a_course_lookup = a_lookup
            g_course_data = g_data

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing course: {g_row[0]}")
                    continue

                # Course_Code (col 1)
                if len(a_row) > 1 and len(g_row) > 1:
                    if not str_match(a_row[1], g_row[1]):
                        file_errors.append(f"{g_row[0]}: code {a_row[1]} vs {g_row[1]}")

                # Student_Count (col 2)
                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 5):
                        file_errors.append(f"{g_row[0]}: students {a_row[2]} vs {g_row[2]}")

                # Avg_Final_Score (col 3)
                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 1.0):
                        file_errors.append(f"{g_row[0]}: avg_score {a_row[3]} vs {g_row[3]}")

                # Assignment_Count (col 4)
                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 1):
                        file_errors.append(f"{g_row[0]}: assignments {a_row[4]} vs {g_row[4]}")

                # Quiz_Count (col 5)
                if len(a_row) > 5 and len(g_row) > 5:
                    if not num_close(a_row[5], g_row[5], 1):
                        file_errors.append(f"{g_row[0]}: quizzes {a_row[5]} vs {g_row[5]}")

        # Check Summary sheet
        print("  Checking Summary...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            file_errors.append("Sheet 'Summary' not found in agent output")
        elif g_rows is None:
            file_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row
            a_summary_lookup = a_lookup
            for row in g_data:
                if row and row[0] is not None:
                    g_summary_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing summary metric: {g_row[0]}")
                    continue

                if len(a_row) > 1 and len(g_row) > 1:
                    g_val = g_row[1]
                    a_val = a_row[1]
                    try:
                        fa, fb = float(a_val), float(g_val)
                        if abs(fa - fb) > 5:
                            file_errors.append(f"Summary {key}: {a_val} vs {g_val}")
                    except (TypeError, ValueError):
                        if not str_contains(str(a_val), str(g_val)[:20]):
                            file_errors.append(f"Summary {key}: '{a_val}' vs '{g_val}'")

    # Check email sent (DB check) -- now CRITICAL: recipient AND subject AND body content.
    print("  Checking email...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Require correct recipient AND correct subject (AND, not OR).
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%academic-affairs@university.edu%'
              AND subject ILIKE '%Fall 2013 Course Performance Summary%'
            LIMIT 5
        """)
        email_rows = cur.fetchall()
        if not email_rows:
            cur.execute("SELECT COUNT(*) FROM email.messages")
            total = cur.fetchone()[0]
            critical_failures.append(
                f"No email to academic-affairs@university.edu with subject "
                f"'Fall 2013 Course Performance Summary' found (total messages: {total})"
            )
        else:
            # Body must mention the total course count AND the best/worst course names.
            # Body is russified prose -> use .lower() on ORIGINAL text (no normalize()).
            best_name = "Экологическая экономика и этика"
            worst_name = "Проектирование на основе данных"
            body_ok = False
            for subj, to_addr, body in email_rows:
                body_l = (body or "").lower()
                has_count = ("6" in body_l) or ("шесть" in body_l)
                has_best = best_name.lower() in body_l
                has_worst = worst_name.lower() in body_l
                if has_count and has_best and has_worst:
                    body_ok = True
                    break
            if not body_ok:
                critical_failures.append(
                    "Email body must mention the total course count (6) and the "
                    "best/worst courses (Экологическая экономика и этика / Проектирование на основе данных)"
                )
        cur.close()
        conn.close()
    except Exception as e:
        critical_failures.append(f"Email check error: {e}")

    # ---- CRITICAL semantic checks (gate before final accuracy decision) ----
    print("  Running critical checks...")

    # CRITICAL 1 & 2: per-course core data accuracy (code exact, students tol5,
    # avg tol0.2, assignments+quizzes exact).
    if g_course_data:
        for g_row in g_course_data:
            if not g_row or g_row[0] is None:
                continue
            name = g_row[0]
            key = str(name).strip().lower()
            a_row = a_course_lookup.get(key)
            if a_row is None:
                critical_failures.append(f"CRITICAL: missing Fall 2013 course '{name}'")
                continue
            if len(a_row) > 1 and len(g_row) > 1 and not str_match(a_row[1], g_row[1]):
                critical_failures.append(f"CRITICAL: {name} Course_Code {a_row[1]} != {g_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 5):
                critical_failures.append(f"CRITICAL: {name} Student_Count {a_row[2]} != {g_row[2]}")
            if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 0.2):
                critical_failures.append(f"CRITICAL: {name} Avg_Final_Score {a_row[3]} != {g_row[3]}")
            if len(a_row) > 4 and len(g_row) > 4 and not num_close(a_row[4], g_row[4], 0):
                critical_failures.append(f"CRITICAL: {name} Assignment_Count {a_row[4]} != {g_row[4]}")
            if len(a_row) > 5 and len(g_row) > 5 and not num_close(a_row[5], g_row[5], 0):
                critical_failures.append(f"CRITICAL: {name} Quiz_Count {a_row[5]} != {g_row[5]}")
    else:
        critical_failures.append("CRITICAL: Course Stats data could not be read")

    # CRITICAL 3: Summary analytical conclusion.
    if g_summary_lookup:
        tc_a = a_summary_lookup.get("total_courses")
        if not (tc_a and len(tc_a) > 1 and num_close(tc_a[1], 6, 0)):
            critical_failures.append(f"CRITICAL: Summary Total_Courses must be 6 (got {tc_a[1] if tc_a and len(tc_a) > 1 else None})")
        ts_a = a_summary_lookup.get("total_students")
        ts_g = g_summary_lookup.get("total_students")
        if ts_g and len(ts_g) > 1:
            if not (ts_a and len(ts_a) > 1 and num_close(ts_a[1], ts_g[1], 5)):
                critical_failures.append(f"CRITICAL: Summary Total_Students {ts_a[1] if ts_a and len(ts_a) > 1 else None} != {ts_g[1]}")
        for metric, expected in (("highest_avg_score_course", "Экологическая экономика и этика"),
                                 ("lowest_avg_score_course", "Проектирование на основе данных")):
            m_a = a_summary_lookup.get(metric)
            val = m_a[1] if (m_a and len(m_a) > 1) else None
            if not str_contains(val, expected):
                critical_failures.append(f"CRITICAL: Summary {metric} must name '{expected}' (got '{val}')")
    else:
        critical_failures.append("CRITICAL: Summary sheet data could not be read")

    # Final result
    print(f"\n=== SUMMARY ===")
    print(f"  File errors:        {len(file_errors)}")
    print(f"  Critical failures:  {len(critical_failures)}")
    if critical_failures:
        for e in critical_failures[:15]:
            print(f"    [CRITICAL] {e}")
    if file_errors:
        for e in file_errors[:15]:
            print(f"    [FILE] {e}")

    # Any critical failure => FAIL (gate before structural/accuracy decision).
    if critical_failures:
        print(f"  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if file_errors:
        print(f"  Overall: FAIL")
        sys.exit(1)
    print(f"  Overall: PASS")
    sys.exit(0)


if __name__ == "__main__":
    main()
