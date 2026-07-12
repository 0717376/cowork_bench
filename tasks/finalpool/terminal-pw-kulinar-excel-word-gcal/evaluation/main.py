"""Evaluation for terminal-pw-kulinar-excel-word-gcal.

Structural checks (sheet/column existence, row counts, sort, weekday presence)
are NON-critical. The substance of the task is verified by CRITICAL_CHECKS:
any critical failure => overall FAIL via sys.exit(1) regardless of accuracy.
Otherwise pass threshold: accuracy >= 70%.

Store names FreshMart / ValueGrocer / BulkBarn are preserved English literals
(they map to the *_Price column identifiers and the gcal title prefix
"Meal Prep:" is matched verbatim). Recipe names / prose are Russian.
"""
import os
import argparse
import json
import os
import re
import sys
import openpyxl
from zoneinfo import ZoneInfo


def local_hm(dt, tzname):
    """gcal stores start_datetime as UTC; a naive dateTime + timeZone is
    converted to UTC and the IANA zone kept in start_timezone (Google Calendar
    semantics). Convert back to the event's local zone before reading the
    hour/minute so both '08:00 + Europe/Moscow' and a verbatim '08:00 UTC' read
    as (8, 0). Returns (hour, minute) or (-1, -1)."""
    if dt is None:
        return (-1, -1)
    if tzname and dt.tzinfo is not None:
        try:
            dt = dt.astimezone(ZoneInfo(tzname))
        except Exception:
            pass
    return (dt.hour, dt.minute)


def local_date(dt, tzname):
    """ISO date (YYYY-MM-DD) in the event's local zone, mirroring local_hm."""
    if dt is None:
        return ""
    if tzname and dt.tzinfo is not None:
        try:
            dt = dt.astimezone(ZoneInfo(tzname))
        except Exception:
            pass
    return dt.strftime("%Y-%m-%d")

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

# Canonical cheapest store per grocery item, recomputed from the mock price
# page (tmp/mock_pages/grocery_prices.html). Keyed by lowercase item name.
# Used to verify the agent's Cheapest_Store column is genuinely computed.
GT_CHEAPEST = {
    "говядина": "ValueGrocer",
    "свинина": "ValueGrocer",
    "баранина": "BulkBarn",
    "фарш мясной": "BulkBarn",
    "куриные бёдра": "ValueGrocer",
    "цыплёнок": "FreshMart",
    "филе рыбы": "ValueGrocer",
    "тушёнка говяжья": "ValueGrocer",
    "сметана": "ValueGrocer",
    "сыр твёрдый": "FreshMart",
    "яйцо куриное": "ValueGrocer",
    "мука пшеничная": "ValueGrocer",
    "рис круглый": "ValueGrocer",
    "крупа гречневая": "FreshMart",
    "картофель": "ValueGrocer",
    "морковь": "ValueGrocer",
    "лук репчатый": "ValueGrocer",
    "капуста белокочанная": "ValueGrocer",
    "чеснок": "FreshMart",
    "масло растительное": "BulkBarn",
    "масло сливочное": "BulkBarn",
    "паста томатная": "BulkBarn",
}

STORE_TOKENS = ["freshmart", "valuegrocer", "valuegr", "bulkbarn"]

# Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Cheapest_Store correct for >=80% of items",
    "Cost_Per_Serving == round(Total_Cost/4, 2)",
    "5 Meal Prep events on weekdays 16-20 Mar 2026, 08:00-09:30",
    "Meal Prep event titles match Weekly_Meal_Plan Lunch recipe",
    "Meal Prep descriptions name ingredients and a store",
    "Word Budget Overview cites store names, budget and a number",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _norm_store(val):
    s = str(val or "").lower()
    if "freshmart" in s:
        return "FreshMart"
    if "valuegr" in s:
        return "ValueGrocer"
    if "bulkbarn" in s:
        return "BulkBarn"
    return str(val or "").strip()


# weekday date -> name, for 16-20 March 2026
WEEKDAY_DATE = {
    "2026-03-16": "monday",
    "2026-03-17": "tuesday",
    "2026-03-18": "wednesday",
    "2026-03-19": "thursday",
    "2026-03-20": "friday",
}


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0

    lunch_by_day = {}  # day(lower) -> recipe name (from Weekly_Meal_Plan)

    # ---- Excel ----
    excel_path = os.path.join(agent_workspace, "Weekly_Meal_Budget.xlsx")
    check("Weekly_Meal_Budget.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # --- Grocery_Prices (structural) ---
        check("Grocery_Prices sheet exists", "Grocery_Prices" in wb.sheetnames)
        gp_rows = []
        gp_headers = []
        if "Grocery_Prices" in wb.sheetnames:
            ws = wb["Grocery_Prices"]
            gp_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Grocery_Prices has >= 15 rows", len(gp_rows) >= 15, f"got {len(gp_rows)}")
            gp_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Item', 'FreshMart_Price', 'ValueGrocer_Price', 'BulkBarn_Price', 'Cheapest_Store']:
                check(f"Grocery_Prices has {col} column", col.lower() in gp_headers, f"headers: {gp_headers[:6]}")
            if len(gp_rows) >= 2:
                items = [str(r[0]).lower() for r in gp_rows if r[0]]
                check("Grocery_Prices sorted alphabetically", items == sorted(items), f"first items: {items[:5]}")

        # --- CRITICAL: Cheapest_Store correctness ---
        try:
            item_idx = gp_headers.index("item") if "item" in gp_headers else 0
            cs_idx = gp_headers.index("cheapest_store") if "cheapest_store" in gp_headers else None
            matched = 0
            considered = 0
            if cs_idx is not None:
                for r in gp_rows:
                    item = str(r[item_idx] or "").strip().lower()
                    if item not in GT_CHEAPEST:
                        continue
                    considered += 1
                    if _norm_store(r[cs_idx]) == GT_CHEAPEST[item]:
                        matched += 1
            ratio = (matched / considered) if considered else 0.0
            check("Cheapest_Store correct for >=80% of items",
                  considered >= 10 and ratio >= 0.8,
                  f"matched {matched}/{considered} = {ratio:.0%}")
        except Exception as e:
            check("Cheapest_Store correct for >=80% of items", False, str(e))

        # --- Recipe_Cost_Analysis (structural) ---
        check("Recipe_Cost_Analysis sheet exists", "Recipe_Cost_Analysis" in wb.sheetnames)
        rc_rows = []
        rc_headers = []
        if "Recipe_Cost_Analysis" in wb.sheetnames:
            ws = wb["Recipe_Cost_Analysis"]
            rc_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recipe_Cost_Analysis has >= 5 rows", len(rc_rows) >= 5, f"got {len(rc_rows)}")
            rc_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Recipe', 'Total_Cost', 'Cost_Per_Serving']:
                check(f"Recipe_Cost_Analysis has {col} column", col.lower() in rc_headers, f"headers: {rc_headers[:6]}")

        # --- CRITICAL: Cost_Per_Serving == round(Total_Cost/4, 2) ---
        try:
            tc_idx = rc_headers.index("total_cost")
            cps_idx = rc_headers.index("cost_per_serving")
            ok = 0
            tot = 0
            for r in rc_rows:
                try:
                    tc = float(r[tc_idx])
                    cps = float(r[cps_idx])
                except (TypeError, ValueError):
                    continue
                tot += 1
                if abs(cps - round(tc / 4.0, 2)) <= 0.02 and tc > 0:
                    ok += 1
            check("Cost_Per_Serving == round(Total_Cost/4, 2)",
                  tot >= 5 and ok == tot,
                  f"{ok}/{tot} rows correct")
        except Exception as e:
            check("Cost_Per_Serving == round(Total_Cost/4, 2)", False, str(e))

        # --- Weekly_Meal_Plan (structural) ---
        check("Weekly_Meal_Plan sheet exists", "Weekly_Meal_Plan" in wb.sheetnames)
        if "Weekly_Meal_Plan" in wb.sheetnames:
            ws = wb["Weekly_Meal_Plan"]
            wp_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Weekly_Meal_Plan has >= 10 rows", len(wp_rows) >= 10, f"got {len(wp_rows)}")
            wp_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for col in ['Day', 'Meal', 'Recipe', 'Estimated_Cost']:
                check(f"Weekly_Meal_Plan has {col} column", col.lower() in wp_headers, f"headers: {wp_headers[:5]}")
            days = set(str(r[0]).lower() for r in wp_rows if r[0])
            for day in ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']:
                check(f"Weekly_Meal_Plan has {day}", day in days, f"found days: {days}")
            # Build day -> lunch recipe map
            try:
                d_idx = wp_headers.index("day")
                m_idx = wp_headers.index("meal")
                r_idx = wp_headers.index("recipe")
                for r in wp_rows:
                    day = str(r[d_idx] or "").strip().lower()
                    meal = str(r[m_idx] or "").strip().lower()
                    rec = str(r[r_idx] or "").strip()
                    if meal == "lunch" and day:
                        lunch_by_day[day] = rec
            except Exception:
                pass

    # ---- Word ----
    word_path = os.path.join(agent_workspace, "Cafeteria_Meal_Proposal.docx")
    check("Cafeteria_Meal_Proposal.docx exists", os.path.exists(word_path))
    if os.path.exists(word_path):
        from docx import Document
        doc = Document(word_path)
        text = " ".join(p.text for p in doc.paragraphs)
        text_l = text.lower()
        check("Word has substantial content", len(text_l) > 200, f"text length: {len(text_l)}")
        # CRITICAL: budget overview references stores + budget word + a number
        has_store = any(tok in text_l for tok in STORE_TOKENS)
        has_budget = ("budget" in text_l) or ("бюджет" in text_l)
        has_number = bool(re.search(r"\d{2,}", text))
        check("Word Budget Overview cites store names, budget and a number",
              has_store and has_budget and has_number,
              f"store={has_store} budget={has_budget} number={has_number}")
    else:
        check("Word Budget Overview cites store names, budget and a number", False, "no docx")

    # ---- terminal script ----
    check("meal_cost_optimizer.py exists",
          os.path.exists(os.path.join(agent_workspace, "meal_cost_optimizer.py")))

    # ---- Calendar ----
    events = []
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT summary, description, start_datetime, end_datetime, "
            "start_timezone, end_timezone "
            "FROM gcal.events WHERE summary ILIKE %s", ('%meal prep%',))
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar query", False, str(e))

    check("5 meal prep calendar events", len(events) >= 5, f"found {len(events)}")

    # CRITICAL: 5 events on weekdays 16-20 Mar 2026, 08:00-09:30
    good_window = 0
    title_match = 0
    desc_ok = 0
    for summ, desc, start, end, s_tz, e_tz in events:
        # Normalize to the event's local zone (gcal stores UTC; a naive
        # dateTime + IANA timeZone is converted to UTC on write). Both
        # '08:00 Europe/Moscow' (stored 05:00Z) and a verbatim '08:00 UTC'
        # must read as 08:00 local.
        date_part = local_date(start, s_tz)
        s_hm = local_hm(start, s_tz)
        e_hm = local_hm(end, e_tz) if end else (-1, -1)
        if date_part in WEEKDAY_DATE and s_hm == (8, 0) and e_hm == (9, 30):
            good_window += 1
            # title recipe vs lunch recipe
            m = re.match(r"\s*meal prep\s*:\s*(.+)", str(summ), re.IGNORECASE)
            if m:
                title_rec = m.group(1).strip().lower()
                day = WEEKDAY_DATE[date_part]
                lunch_rec = lunch_by_day.get(day, "").strip().lower()
                if lunch_rec and (title_rec == lunch_rec
                                  or title_rec in lunch_rec or lunch_rec in title_rec):
                    title_match += 1
        d = str(desc or "")
        if len(d) > 20 and any(tok in d.lower() for tok in STORE_TOKENS):
            desc_ok += 1

    check("5 Meal Prep events on weekdays 16-20 Mar 2026, 08:00-09:30",
          good_window >= 5, f"{good_window} in window")
    check("Meal Prep event titles match Weekly_Meal_Plan Lunch recipe",
          title_match >= 4, f"{title_match}/5 match lunch recipe")
    check("Meal Prep descriptions name ingredients and a store",
          desc_ok >= 4, f"{desc_ok} descriptions name a store")

    accuracy = PASS_COUNT / (PASS_COUNT + FAIL_COUNT) if (PASS_COUNT + FAIL_COUNT) else 0.0
    return accuracy


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in output."""
    import psycopg2
    print("\n=== Reverse Validation ===")

    path = os.path.join(workspace, "Weekly_Meal_Budget.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        check("No negative cost values in Excel", not has_negative, "Found negative cost value")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE summary ILIKE '%%meal prep%%'
              AND EXTRACT(DOW FROM start_datetime) IN (0, 6)
        """)
        weekend_count = cur.fetchone()[0]
        check("No meal prep events on weekends", weekend_count == 0,
              f"Found {weekend_count} weekend events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    accuracy = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    check_reverse_validation(args.agent_workspace)

    # Critical gate: any critical check failed => FAIL regardless of accuracy.
    failed_critical = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    print(f"\nPassed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks (accuracy={accuracy:.0%})")
    if failed_critical:
        print("CRITICAL checks failed: " + "; ".join(failed_critical))
        sys.exit(1)

    sys.exit(0 if accuracy >= 0.70 else 1)


if __name__ == "__main__":
    main()
