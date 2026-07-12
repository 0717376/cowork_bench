"""Evaluation for sf-hr-attrition-gcal (ClickHouse HR analytics).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Department names are read LIVE from the ClickHouse-backed sf_data schema, which
is russified centrally (Инженерия/Финансы/Кадры/Операции/НИОКР/Продажи/Поддержка).
We never hardcode dept-name realia here.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")
PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic checks that gate PASS regardless of accuracy.
CRITICAL_CHECKS = {
    "Department count is 7",
    "Department Scores sorted ascending by Avg_Satisfaction",
    "Action Items has 3 rows",
    "Exactly 3 HR Review events on 2026-03-16",
}


def record(name, passed, detail="", critical_extra=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1; print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1; FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")
    # Per-dept value correctness is added dynamically to CRITICAL_CHECKS by callers.
    if critical_extra:
        CRITICAL_CHECKS.add(name)


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except: return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_expected():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute('''SELECT "DEPARTMENT", COUNT(*),
        ROUND(AVG("JOB_SATISFACTION")::numeric,2),
        ROUND(AVG("WORK_LIFE_BALANCE")::numeric,2),
        ROUND(AVG("PERFORMANCE_RATING")::numeric,2)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT" ORDER BY AVG("JOB_SATISFACTION")''')
    depts = [{"dept": r[0], "count": r[1], "avg_sat": float(r[2]),
              "avg_wlb": float(r[3]), "avg_perf": float(r[4])} for r in cur.fetchall()]
    conn.close()
    return {"all_depts": depts, "low_sat": depts[:3]}


def sheet_dicts(wb, name):
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2: return []
            hdrs = [str(h).strip() if h else "" for h in rows[0]]
            return [{hdrs[i]: row[i] for i in range(len(hdrs))} for row in rows[1:] if not all(v is None for v in row)]
    return None


def check_excel(ws_path, exp):
    print("\n=== Checking Excel ===")
    p = os.path.join(ws_path, "Satisfaction_Report.xlsx")
    if not os.path.isfile(p):
        record("Excel file exists", False, p); return
    record("Excel file exists", True)
    wb = openpyxl.load_workbook(p, data_only=True)

    d = sheet_dicts(wb, "Department Scores")
    if d is None:
        record("Sheet Department Scores", False, str(wb.sheetnames))
    else:
        record("Sheet Department Scores", True)
        record("Department count is 7", len(d) == 7, f"Got {len(d)}")
        for e in exp["all_depts"]:
            m = next((r for r in d if str_match(r.get("Department"), e["dept"])), None)
            if not m:
                # Missing dept is a critical correctness failure.
                record(f"Dept {e['dept']} present", False, "Missing", critical_extra=True)
                continue
            # Tightened tolerance so per-dept value is actually verified
            # (depts cluster ~6.51-6.59; loose tol=0.5 let any mis-assignment pass).
            record(f"Dept {e['dept']} satisfaction",
                   num_close(m.get("Avg_Satisfaction"), e["avg_sat"], 0.05),
                   f"{m.get('Avg_Satisfaction')} vs {e['avg_sat']}", critical_extra=True)
            # Tightened from tol=100; counts are exact GROUP BY counts.
            record(f"Dept {e['dept']} count",
                   num_close(m.get("Employee_Count"), e["count"], 1),
                   f"{m.get('Employee_Count')} vs {e['count']}")

        # Assert ascending sort order on the deliverable's core ranking column.
        sat_vals = []
        for r in d:
            try: sat_vals.append(float(r.get("Avg_Satisfaction")))
            except: sat_vals.append(None)
        monotonic = all(
            sat_vals[i] is not None and sat_vals[i + 1] is not None and sat_vals[i] <= sat_vals[i + 1] + 1e-9
            for i in range(len(sat_vals) - 1)
        ) if len(sat_vals) >= 2 else False
        record("Department Scores sorted ascending by Avg_Satisfaction", monotonic, f"Values: {sat_vals}")

    d = sheet_dicts(wb, "Action Items")
    if d is None:
        record("Sheet Action Items", False, str(wb.sheetnames))
    else:
        record("Sheet Action Items", True)
        record("Action Items has 3 rows", len(d) == 3, f"Got {len(d)}")
        for e in exp["low_sat"]:
            m = next((r for r in d if str_match(r.get("Department"), e["dept"])), None)
            if not m:
                record(f"Action {e['dept']} present", False, "Missing", critical_extra=True)
                continue
            record(f"Action {e['dept']} status",
                   str_match(m.get("Status"), "Needs Review"),
                   f"Got: {m.get('Status')}")
    wb.close()


def check_gcal(exp):
    print("\n=== Checking Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()

    # Exactly 3 HR Review events on 2026-03-16.
    hr_events = [
        (s, d, sd, ed) for (s, d, sd, ed) in events
        if s and "hr review" in s.lower() and sd and sd.strftime("%Y-%m-%d") == "2026-03-16"
    ]
    record("Exactly 3 HR Review events on 2026-03-16", len(hr_events) == 3, f"Found {len(hr_events)}")

    expected_starts = {"09:00", "10:30", "12:00"}
    found_starts = set()

    for dept_info in exp["low_sat"]:
        dept = dept_info["dept"]
        match = None
        for summary, desc, start_dt, end_dt in hr_events:
            if dept.lower() in (summary or "").lower():
                match = (summary, desc, start_dt, end_dt)
                break
        if not match:
            record(f"Event for {dept} exists", False, "Not found", critical_extra=True)
            continue
        record(f"Event for {dept} exists", True)
        summary, desc, start_dt, end_dt = match

        # 1-hour duration.
        if start_dt and end_dt:
            dur_min = (end_dt - start_dt).total_seconds() / 60.0
            record(f"Event {dept} is 1 hour", abs(dur_min - 60) < 1e-6, f"Got {dur_min} min")
        else:
            record(f"Event {dept} is 1 hour", False, "Missing start/end")

        # Track start time slot.
        if start_dt:
            found_starts.add(start_dt.strftime("%H:%M"))

        # Strict: the dept's average satisfaction score must appear in the description.
        sat = dept_info["avg_sat"]
        desc_text = desc or ""
        sat_variants = {str(sat), f"{sat:.2f}", f"{sat:.1f}", str(sat).replace(".", ",")}
        record(f"Event {dept} description mentions satisfaction score",
               any(v in desc_text for v in sat_variants),
               f"Score {sat} not in desc: {desc_text[:150]}", critical_extra=True)

    # All three NY start slots present across the 3 flagged events.
    record("Event start times are 09:00/10:30/12:00",
           expected_starts.issubset(found_starts),
           f"Found starts: {sorted(found_starts)}")

    cur.close()
    conn.close()


def check_email(exp):
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
    emails = cur.fetchall()
    record("At least 1 email sent", len(emails) >= 1, f"Found {len(emails)}")

    found_summary = False
    for subj, to, body in emails:
        subj_lower = (subj or "").lower()
        if "satisfaction" in subj_lower or "review" in subj_lower:
            found_summary = True
            to_str = json.dumps(to).lower() if isinstance(to, list) else str(to).lower()
            record("Email to hr-director",
                   "hr-director@company.example.com" in to_str, f"To: {to}")
            body_lower = (body or "").lower()
            # All 7 departments must be listed.
            for dept_info in exp["all_depts"]:
                record(f"Email lists {dept_info['dept']}",
                       dept_info["dept"].lower() in body_lower,
                       "Not found in body")
            # The 3 flagged low-satisfaction depts must be present (subset of above,
            # but kept explicit as the core flagging deliverable).
            for dept_info in exp["low_sat"]:
                record(f"Email flags {dept_info['dept']}",
                       dept_info["dept"].lower() in body_lower,
                       "Not found in body")
            break
    if not found_summary:
        record("Summary email found", False, f"Subjects: {[e[0] for e in emails]}")

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", default=".")
    parser.add_argument("--groundtruth_workspace", default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    exp = get_expected()
    check_excel(args.agent_workspace, exp)
    check_gcal(exp)
    check_email(exp)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy, "critical_failed": critical_failed,
                       "success": success}, f)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
