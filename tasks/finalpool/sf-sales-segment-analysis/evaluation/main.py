"""Evaluation for sf-sales-segment-analysis (ClickHouse / RU).

Hardened:
 - CRITICAL_CHECKS gate (sys.exit(1) on any critical failure) before accuracy>=70.
 - Tightened numeric tolerances (Revenue tol<=1.0, Customers/Orders exact).
 - Google Sheet 'Segment Dashboard' / 'Overview' deliverable verified against the
   xlsx 'Segment Analysis' rows.
 - RU value matching uses .lower() on ORIGINAL text (no normalize()); segment names
   are the centrally-russified SALES_DW values (Частные клиенты / Корпоративный /
   Государственный / Малый и средний бизнес).
"""
import argparse
import os
import sys
import openpyxl

try:
    import psycopg2
except Exception:
    psycopg2 = None

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_to_lookup(rows):
    data = rows[1:] if rows and len(rows) > 1 else []
    lk = {}
    for row in data:
        if row and row[0] is not None:
            lk[str(row[0]).strip().lower()] = row
    return lk


def fetch_gsheet_overview():
    """Return list of row-lists for sheet 'Overview' in spreadsheet 'Segment Dashboard',
    or None if not found / DB unavailable."""
    if psycopg2 is None:
        return None
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        print(f"  [gsheet] DB connect failed: {e}")
        return None
    try:
        cur = conn.cursor()
        cur.execute(
            "SELECT id FROM gsheet.spreadsheets WHERE lower(trim(title))=lower(%s)",
            ("Segment Dashboard",),
        )
        sp = cur.fetchone()
        if not sp:
            return None
        spid = sp[0]
        cur.execute(
            "SELECT id FROM gsheet.sheets WHERE spreadsheet_id=%s AND lower(trim(title))=lower(%s)",
            (spid, "Overview"),
        )
        sh = cur.fetchone()
        if not sh:
            return None
        shid = sh[0]
        cur.execute(
            "SELECT row_index, col_index, value FROM gsheet.cells "
            "WHERE spreadsheet_id=%s AND sheet_id=%s",
            (spid, shid),
        )
        cells = cur.fetchall()
        if not cells:
            return []
        max_r = max(c[0] for c in cells)
        max_c = max(c[1] for c in cells)
        grid = [[None] * (max_c + 1) for _ in range(max_r + 1)]
        for r, c, v in cells:
            grid[r][c] = v
        return grid
    finally:
        conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Segment_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Segment_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- collect checks ----
    checks = []           # (name, passed, is_critical)
    all_errors = []

    def record(name, passed, critical=False, err=None):
        checks.append((name, passed, critical))
        if not passed and err:
            all_errors.append(err)

    # ---------- Segment Analysis ----------
    print("  Checking Segment Analysis...")
    a_rows = load_sheet_rows(agent_wb, "Segment Analysis")
    g_rows = load_sheet_rows(gt_wb, "Segment Analysis")
    a_seg = rows_to_lookup(a_rows) if a_rows else {}
    g_seg = rows_to_lookup(g_rows) if g_rows else {}

    if a_rows is None:
        record("seg_sheet_present", False, critical=True,
               err="Sheet 'Segment Analysis' not found in agent output")
    else:
        record("seg_sheet_present", True)
        for key, g_row in g_seg.items():
            a_row = a_seg.get(key)
            present = a_row is not None
            record(f"seg_row_present[{key}]", present, critical=True,
                   err=None if present else f"Missing segment row: {g_row[0]}")
            if not present:
                continue
            # Customers (exact-ish, tol 1)
            if len(g_row) > 1:
                ok = num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], 1)
                record(f"seg_customers[{key}]", ok,
                       err=None if ok else f"{key}.Customers: {a_row[1]} vs {g_row[1]} (tol=1)")
            # Orders (exact-ish, tol 1)
            if len(g_row) > 2:
                ok = num_close(a_row[2] if len(a_row) > 2 else None, g_row[2], 1)
                record(f"seg_orders[{key}]", ok,
                       err=None if ok else f"{key}.Orders: {a_row[2]} vs {g_row[2]} (tol=1)")
            # Revenue (tight tol 1.0, CRITICAL)
            if len(g_row) > 3:
                ok = num_close(a_row[3] if len(a_row) > 3 else None, g_row[3], 1.0)
                record(f"seg_revenue[{key}]", ok, critical=True,
                       err=None if ok else f"{key}.Revenue: {a_row[3]} vs {g_row[3]} (tol=1.0)")
            # Avg_Order_Value
            if len(g_row) > 4:
                ok = num_close(a_row[4] if len(a_row) > 4 else None, g_row[4], 0.5)
                record(f"seg_aov[{key}]", ok,
                       err=None if ok else f"{key}.Avg_Order_Value: {a_row[4]} vs {g_row[4]} (tol=0.5)")
            # Revenue_Share_Pct
            if len(g_row) > 5:
                ok = num_close(a_row[5] if len(a_row) > 5 else None, g_row[5], 0.2)
                record(f"seg_share[{key}]", ok,
                       err=None if ok else f"{key}.Revenue_Share_Pct: {a_row[5]} vs {g_row[5]} (tol=0.2)")

        # Revenue_Share_Pct internal consistency: share == round(Revenue/Total*100,1)
        try:
            a_total_rev = sum(float(r[3]) for r in a_seg.values() if r[3] is not None)
            share_ok = a_total_rev > 0
            if share_ok:
                for key, r in a_seg.items():
                    exp = round(float(r[3]) / a_total_rev * 100.0, 1)
                    got = r[5] if len(r) > 5 else None
                    if got is None or abs(float(got) - exp) > 0.2:
                        share_ok = False
                        all_errors.append(
                            f"{key}.Revenue_Share_Pct internal: {got} vs computed {exp} (tol=0.2)")
                        break
            record("seg_share_internal_consistency", share_ok, critical=True,
                   err=None if share_ok else "Revenue_Share_Pct not consistent with Revenue/Total")
        except (TypeError, ValueError):
            record("seg_share_internal_consistency", False, critical=True,
                   err="Could not compute Revenue_Share_Pct consistency")

    # ---------- Summary ----------
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    a_sum = rows_to_lookup(a_rows) if a_rows else {}
    g_sum = rows_to_lookup(g_rows) if g_rows else {}

    if a_rows is None:
        record("summary_sheet_present", False, critical=True,
               err="Sheet 'Summary' not found in agent output")
    else:
        record("summary_sheet_present", True)
        for key, g_row in g_sum.items():
            a_row = a_sum.get(key)
            present = a_row is not None
            record(f"summary_row_present[{key}]", present,
                   err=None if present else f"Missing summary row: {g_row[0]}")
            if not present:
                continue
            if key == "top_segment":
                ok = str_match(a_row[1] if len(a_row) > 1 else None, g_row[1])
                record("summary_top_segment", ok, critical=True,
                       err=None if ok else f"Top_Segment: {a_row[1]} vs {g_row[1]}")
            elif key == "total_revenue":
                ok = num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], 1.0)
                record("summary_total_revenue", ok, critical=True,
                       err=None if ok else f"Total_Revenue: {a_row[1]} vs {g_row[1]} (tol=1.0)")
            else:
                ok = num_close(a_row[1] if len(a_row) > 1 else None, g_row[1], 1.0)
                record(f"summary_value[{key}]", ok,
                       err=None if ok else f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=1.0)")

        # CRITICAL: Top_Segment must be the russified top segment AND have the max Revenue
        try:
            top_name = a_sum.get("top_segment", [None, None])[1]
            if top_name is not None and a_seg:
                top_key = str(top_name).strip().lower()
                top_row = a_seg.get(top_key)
                max_rev = max(float(r[3]) for r in a_seg.values() if r[3] is not None)
                ok = (top_row is not None and top_row[3] is not None
                      and abs(float(top_row[3]) - max_rev) <= 1.0)
                record("top_segment_is_max_revenue", ok, critical=True,
                       err=None if ok else "Top_Segment in Summary is not the max-Revenue segment")
            else:
                record("top_segment_is_max_revenue", False, critical=True,
                       err="Top_Segment missing or Segment Analysis empty")
        except (TypeError, ValueError):
            record("top_segment_is_max_revenue", False, critical=True,
                   err="Could not verify Top_Segment is max revenue")

        # CRITICAL: Total_Revenue == sum of segment Revenues (cross-sheet consistency)
        try:
            tr = a_sum.get("total_revenue", [None, None])[1]
            seg_sum = sum(float(r[3]) for r in a_seg.values() if r[3] is not None)
            ok = tr is not None and abs(float(tr) - seg_sum) <= 1.0
            record("total_revenue_matches_segment_sum", ok, critical=True,
                   err=None if ok else f"Total_Revenue {tr} != sum(segment Revenue) {round(seg_sum,2)}")
        except (TypeError, ValueError):
            record("total_revenue_matches_segment_sum", False, critical=True,
                   err="Could not verify Total_Revenue cross-sheet consistency")

    # ---------- Google Sheet: Segment Dashboard / Overview ----------
    print("  Checking Google Sheet 'Segment Dashboard'/'Overview'...")
    grid = fetch_gsheet_overview()
    if grid is None:
        record("gsheet_overview_present", False, critical=True,
               err="Google Sheet 'Segment Dashboard' with sheet 'Overview' not found")
    else:
        record("gsheet_overview_present", True)
        # Build segment->revenue lookup from the gsheet grid (skip header row).
        gs_lookup = {}
        for r in grid[1:]:
            if r and r[0] is not None and str(r[0]).strip() != "":
                gs_lookup[str(r[0]).strip().lower()] = r
        # Every groundtruth segment row must appear with matching Revenue (col index 3).
        gs_ok = True
        for key, g_row in g_seg.items():
            r = gs_lookup.get(key)
            if r is None or len(r) < 4 or not num_close(r[3], g_row[3], 1.0):
                gs_ok = False
                got = (r[3] if r is not None and len(r) > 3 else None)
                all_errors.append(f"gsheet Overview segment '{key}' Revenue: {got} vs {g_row[3]}")
                break
        record("gsheet_overview_data_matches", gs_ok, critical=True,
               err=None if gs_ok else "Google Sheet 'Overview' rows do not match Segment Analysis")

    # ---- scoring ----
    critical_failed = [n for (n, p, c) in checks if c and not p]
    total = len(checks)
    passed = sum(1 for (_, p, _) in checks if p)
    accuracy = (passed / total * 100.0) if total else 0.0

    print(f"\nChecks passed: {passed}/{total} ({accuracy:.1f}%)")
    if all_errors:
        print("Errors:")
        for e in all_errors[:15]:
            print(f"  {e}")

    if critical_failed:
        print(f"\n=== RESULT: FAIL (critical checks failed: {critical_failed}) ===")
        sys.exit(1)

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
