"""
Evaluation for rzd-hr1c-training-trip-kazan-excel-email-gcal.

Checks:
  Excel (Training_Travel_Report.xlsx):
    1. файл существует
    2. читается openpyxl
    3. лист Employees + ≥ 4 строк данных
    4. Employees содержит «Продажи» или «Маркетинг»
    5. лист Travel_Plan + ≥ 8 строк данных
    6. лист Travel_Plan содержит хотя бы один из выбранных номеров поезда (716Г/718Г/717Г/719Г)
    7. лист Budget_Summary + ≥ 2 строк данных
    8. в Budget_Summary встречается итоговая сумма ~45000 ₽ (±5000)
  Calendar:
    9. ≥ 2 новых события на 17.03.2026 (помимо «кикофф»)
  Email:
   10. отправлено письмо на training@hr.company.ru
   11. отправлено ≥ 1 дополнительное письмо (не на training@hr.company.ru)
PASS если ≥ 70% чеков.
"""
import glob
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def check_excel(agent_workspace):
    print("\n=== Check 1: Training_Travel_Report.xlsx ===")
    xlsx_files = glob.glob(os.path.join(agent_workspace, "*.xlsx"))
    train_files = [f for f in xlsx_files if any(
        kw in os.path.basename(f).lower()
        for kw in ["training", "travel", "report", "trip", "kazan"]
    )]
    if not train_files:
        record("xlsx-отчёт существует", False, f"нет подходящих xlsx в {agent_workspace}: {xlsx_files}")
        return
    record("xlsx-отчёт существует", True)

    xlsx_path = train_files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel-файл читается", False, str(e))
        return
    record("Excel-файл читается", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # ── Employees sheet
    emp_match = [s for s in sheet_names_lower if "employ" in s or "staff" in s]
    if not emp_match:
        record("лист Employees найден", False, f"листы: {wb.sheetnames}")
    else:
        record("лист Employees найден", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(emp_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Employees содержит ≥ 4 строк данных", len(data_rows) >= 4,
               f"найдено строк: {len(data_rows)}")
        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        has_dept = "продажи" in all_text or "маркетинг" in all_text
        record("Employees содержит «Продажи» или «Маркетинг»", has_dept,
               f"первые 200 символов: {all_text[:200]}")

    # ── Travel_Plan sheet
    plan_match = [s for s in sheet_names_lower if "travel" in s or "plan" in s or "trip" in s]
    if not plan_match:
        record("лист Travel_Plan найден", False, f"листы: {wb.sheetnames}")
    else:
        record("лист Travel_Plan найден", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(plan_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel_Plan содержит ≥ 8 строк данных", len(data_rows) >= 8,
               f"найдено строк: {len(data_rows)}")
        all_text = " ".join(str(c) for r in rows for c in r if c)
        has_train = any(tn in all_text for tn in ("716Г", "718Г", "717Г", "719Г",
                                                   "716G", "718G", "717G", "719G"))
        record("Travel_Plan упоминает один из «Стрижей» (716/717/718/719)", has_train,
               f"snippet: {all_text[:200]}")

    # ── Budget_Summary sheet
    budget_match = [s for s in sheet_names_lower if "budget" in s or "summar" in s
                    or "cost" in s or "бюджет" in s]
    if not budget_match:
        record("лист Budget_Summary найден", False, f"листы: {wb.sheetnames}")
    else:
        record("лист Budget_Summary найден", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(budget_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Budget_Summary содержит ≥ 2 строк данных", len(data_rows) >= 2,
               f"найдено строк: {len(data_rows)}")

        # 5 × 4500 × 2 = 45000 ₽; ±5000 покрывает класс или мелкие сборы
        all_vals = []
        for r in rows:
            for c in r:
                try:
                    all_vals.append(float(c))
                except Exception:
                    pass
        has_total = any(40000 <= v <= 50000 for v in all_vals)
        record("Budget_Summary содержит итог ~45 000 ₽ (±5000)", has_total,
               f"числовые значения: {sorted(all_vals)}")


def check_gcal():
    print("\n=== Check 2: Calendar events ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE start_datetime >= '2026-03-17'
          AND start_datetime < '2026-03-18'
          AND summary NOT ILIKE '%кикофф%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()
    record("≥ 2 новых события на 17.03.2026", len(events) >= 2,
           f"найдено {len(events)}: {[e[0] for e in events]}")


def _count_email(to_pattern, exclude_from_pattern=None):
    cnt = sent = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        if exclude_from_pattern:
            cur.execute(
                "SELECT COUNT(*) FROM email.messages"
                " WHERE to_addr::text ILIKE %s AND from_addr NOT ILIKE %s",
                (f"%{to_pattern}%", f"%{exclude_from_pattern}%"),
            )
        else:
            cur.execute(
                "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s",
                (f"%{to_pattern}%",),
            )
        cnt = cur.fetchone()[0]
        cur.close(); conn.close()
    except Exception:
        pass
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COUNT(*) FROM email.sent_log WHERE to_addr::text ILIKE %s",
            (f"%{to_pattern}%",),
        )
        sent = cur2.fetchone()[0]
        cur2.close(); conn2.close()
    except Exception:
        pass
    return cnt, sent


def check_emails():
    print("\n=== Check 3: Emails ===")
    hr_cnt, hr_sent = _count_email("training@hr.company.ru", "training@hr.company.ru")
    record("письмо отправлено на training@hr.company.ru", hr_cnt >= 1 or hr_sent >= 1,
           f"messages: {hr_cnt}, sent_log: {hr_sent}")

    emp_cnt = emp_sent = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM email.messages"
            " WHERE to_addr::text NOT ILIKE '%training@hr.company.ru%'"
            "   AND from_addr NOT ILIKE '%training@hr.company.ru%'"
        )
        emp_cnt = cur.fetchone()[0]
        cur.close(); conn.close()
    except Exception:
        pass
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COUNT(*) FROM email.sent_log"
            " WHERE to_addr::text NOT ILIKE '%training@hr.company.ru%'"
        )
        emp_sent = cur2.fetchone()[0]
        cur2.close(); conn2.close()
    except Exception:
        pass
    record("отправлено ≥ 1 дополнительное письмо (сотруднику)",
           emp_cnt >= 1 or emp_sent >= 1,
           f"messages: {emp_cnt}, sent_log: {emp_sent}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
