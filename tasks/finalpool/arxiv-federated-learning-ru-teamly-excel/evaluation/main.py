"""
Evaluation for arxiv-federated-learning-ru-teamly-excel task.

Checks Teamly knowledge base (hub page + one child page per paper) and the
Excel spreadsheet (Paper Details + Summary sheets).

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Pass threshold otherwise: accuracy >= 80%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# FL papers in arxiv.papers (IDs from preprocess)
EXPECTED_FL_PAPERS = {
    "1602.05629": "Communication-Efficient Learning of Deep Networks from Decentralized Data",
    "1812.06127": "Federated Optimization in Heterogeneous Networks",
    "1908.07873": "Federated Learning: Challenges, Methods, and Future Directions",
    "1912.04977": "Advances and Open Problems in Federated Learning",
}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Excel file exists",
    "Teamly hub page exists",
    "Teamly: >= 4 paper entries",
    "Summary: Total_Papers between 4-6",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=50):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def _fl_mention(text):
    t = (text or "").lower()
    return "federat" in t or "федератив" in t


def check_teamly():
    """Check Teamly hub page + child pages, one per paper."""
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Only user-created pages (seeds have id <= 3).
        cur.execute("SELECT id, title, body, parent_id FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Teamly hub page exists", False, f"DB error: {e}")
        record("Teamly: >= 4 paper entries", False, f"DB error: {e}")
        return

    # Hub page: title mentions FL и выглядит как раздел/база знаний.
    hub_id = None
    for pid, title, body, parent_id in pages:
        tl = (title or "").lower()
        if _fl_mention(tl) and any(k in tl for k in ["база знани", "knowledge", "hub", "research", "учени"]):
            hub_id = pid
            break
    # Fallback: любая новая страница, заголовок которой явно про FL и не является
    # названием конкретной статьи (без года/конкретики) — берём первую FL-страницу.
    if hub_id is None:
        for pid, title, body, parent_id in pages:
            if _fl_mention(title):
                hub_id = pid
                break

    record("Teamly hub page exists", hub_id is not None,
           f"Найдено {len(pages)} новых страниц, хаб не опознан")

    # Paper entries: дочерние страницы хаба, либо (fallback) FL-страницы кроме хаба.
    child_pages = [p for p in pages if p[3] == hub_id and p[0] != hub_id]
    if len(child_pages) >= 4:
        entries = child_pages
    else:
        entries = [p for p in pages if p[0] != hub_id and (_fl_mention(p[1]) or _fl_mention(p[2]))]
    record("Teamly: >= 4 paper entries", len(entries) >= 4,
           f"Найдено {len(entries)} записей (дочерних: {len(child_pages)})")

    # Покрытие конкретных статей в телах/заголовках teamly-страниц.
    all_text = " ".join(((p[1] or "") + " " + (p[2] or "")) for p in pages).lower()
    covered = sum(1 for t in EXPECTED_FL_PAPERS.values()
                  if t.lower()[:40] in all_text)
    record("Teamly: >= 3 target paper titles present", covered >= 3,
           f"Найдено {covered}/4 названий статей")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            ws = wb[name]
            return [[cell.value for cell in row] for row in ws.iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        c = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == c:
                return i
    return None


def check_excel(agent_workspace):
    """Check Excel spreadsheet."""
    print("\n=== Checking Excel ===")
    excel_path = os.path.join(agent_workspace, "Federated_Learning_Papers.xlsx")

    if not os.path.isfile(excel_path):
        record("Excel file exists", False, f"Not found: {excel_path}")
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    record("Excel readable", True)

    # Check Paper Details sheet
    details_rows = load_sheet_rows(wb, "Paper Details")
    if details_rows is None:
        details_rows = load_sheet_rows(wb, "Paper_Details")
    if details_rows is None:
        record("Sheet 'Paper Details' exists", False, f"Available: {wb.sheetnames}")
        return

    record("Sheet 'Paper Details' exists", True)
    header = details_rows[0] if details_rows else []
    data_rows = details_rows[1:] if len(details_rows) > 1 else []

    record("Paper Details has >= 4 rows", len(data_rows) >= 4,
           f"Found {len(data_rows)} data rows")

    title_col = find_col(header, ["Title", "title"])
    abstract_len_col = find_col(header, ["Abstract_Length", "abstract_length", "Abstract Length"])

    if title_col is not None:
        found_titles = []
        for row in data_rows:
            if title_col < len(row) and row[title_col]:
                found_titles.append(str(row[title_col]).strip().lower())

        for pid, expected_title in EXPECTED_FL_PAPERS.items():
            found = any(expected_title.lower() in t or t in expected_title.lower()
                        for t in found_titles)
            record(f"Has paper: {expected_title[:50]}...", found)

    record("Abstract_Length column exists", abstract_len_col is not None,
           f"Header: {header}")

    # Check Summary sheet
    summary_rows = load_sheet_rows(wb, "Summary")
    if summary_rows is None:
        record("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
        return

    record("Sheet 'Summary' exists", True)

    metrics = {}
    for row in summary_rows:
        if row and row[0] is not None:
            key = str(row[0]).strip().lower().replace(" ", "_")
            val = row[1] if len(row) > 1 else None
            metrics[key] = val

    # Total_Papers
    total_key = None
    for k in metrics:
        if "total" in k and "paper" in k:
            total_key = k
            break
    if total_key:
        val = metrics[total_key]
        ok = val is not None and 4 <= int(float(val)) <= 6
        record("Summary: Total_Papers between 4-6", ok, f"Got {val}")
    else:
        record("Summary: Total_Papers between 4-6", False, f"Keys: {list(metrics.keys())}")

    # Avg_Abstract_Length
    avg_key = None
    for k in metrics:
        if "avg" in k and "abstract" in k:
            avg_key = k
            break
    if avg_key:
        val = metrics[avg_key]
        ok = val is not None and float(val) > 100
        record("Summary: Avg_Abstract_Length > 100", ok, f"Got {val}")
    else:
        record("Summary: Avg_Abstract_Length exists", False, f"Keys: {list(metrics.keys())}")

    # Earliest/Latest Year
    for label, expected in [("earliest", 2016), ("latest", 2019)]:
        year_key = None
        for k in metrics:
            if label in k and "year" in k:
                year_key = k
                break
        if year_key:
            ok = num_close(metrics[year_key], expected, tol=1)
            record(f"Summary: {label.title()}_Year ~ {expected}", ok,
                   f"Got {metrics[year_key]}")
        else:
            record(f"Summary: {label.title()}_Year exists", False,
                   f"Keys: {list(metrics.keys())}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--res_log_file", required=False)
    parser.add_argument("--launch_time", required=False)
    args = parser.parse_args()

    check_teamly()
    check_excel(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 80:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
