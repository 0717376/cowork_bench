"""Evaluation for insales-review-sentiment-gsheet (InSales / wc.* schema)."""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def norm_name(s):
    return str(s or "").strip().lower()


def get_expected_data():
    """Compute expected data from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Total reviews and avg rating
    cur.execute("SELECT COUNT(*), ROUND(AVG(rating)::numeric, 2) FROM wc.product_reviews")
    total_reviews, overall_avg = cur.fetchone()

    # Rating distribution
    cur.execute("""
        SELECT rating, COUNT(*),
               ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM wc.product_reviews), 1)
        FROM wc.product_reviews
        GROUP BY rating ORDER BY rating
    """)
    rating_dist = cur.fetchall()

    # Top 20 most-reviewed products
    cur.execute("""
        SELECT p.name, COUNT(r.id) as review_count,
               ROUND(AVG(r.rating)::numeric, 2) as avg_rating,
               SUM(CASE WHEN r.rating=5 THEN 1 ELSE 0 END) as five_star,
               SUM(CASE WHEN r.rating=1 THEN 1 ELSE 0 END) as one_star
        FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        ORDER BY COUNT(r.id) DESC, p.name ASC
        LIMIT 20
    """)
    top_products = cur.fetchall()

    # Most reviewed product (alphabetically first if tied)
    cur.execute("""
        SELECT p.name FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        ORDER BY COUNT(r.id) DESC, p.name ASC
        LIMIT 1
    """)
    most_reviewed = cur.fetchone()[0]

    # Highest rated (min 3 reviews)
    cur.execute("""
        SELECT p.name FROM wc.product_reviews r
        JOIN wc.products p ON r.product_id = p.id
        GROUP BY p.name
        HAVING COUNT(r.id) >= 3
        ORDER BY AVG(r.rating) DESC, p.name ASC
        LIMIT 1
    """)
    highest_rated = cur.fetchone()[0]

    cur.close()
    conn.close()
    return {
        "total_reviews": total_reviews,
        "overall_avg": float(overall_avg),
        "rating_dist": rating_dist,
        "top_products": top_products,
        "most_reviewed": most_reviewed,
        "highest_rated": highest_rated,
    }


def _find_sheet(wb, *needles_groups):
    """Find a sheet whose lowercased name contains ALL needles in any group."""
    for name in wb.sheetnames:
        ln = name.lower()
        for group in needles_groups:
            if all(n in ln for n in group):
                return wb[name]
    return None


def check_excel(agent_workspace, expected):
    """Check Review_Sentiment_Analysis.xlsx (non-critical structural)."""
    print("\n=== Checking Review_Sentiment_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Review_Sentiment_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return False

    all_ok = True

    # Check Product Reviews sheet
    pr_sheet = _find_sheet(wb, ("product", "review"))
    if pr_sheet is None:
        record("Sheet 'Product Reviews' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Product Reviews' exists", True)
        rows = list(pr_sheet.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if r and r[0]]
        record("Product Reviews has 20 rows", len(rows) == 20, f"Got {len(rows)}")
        if len(rows) != 20:
            all_ok = False

    # Check Rating Distribution sheet
    rd_sheet = _find_sheet(wb, ("rating", "distribution"))
    if rd_sheet is None:
        record("Sheet 'Rating Distribution' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Rating Distribution' exists", True)
        rows = list(rd_sheet.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if r and r[0] is not None]
        record("Rating Distribution has 5 rows", len(rows) == 5, f"Got {len(rows)}")

        for exp_r in expected["rating_dist"]:
            rating_val = exp_r[0]
            found = False
            for r in rows:
                if r and r[0] is not None:
                    try:
                        rv = int(r[0])
                    except (TypeError, ValueError):
                        continue
                    if rv == rating_val:
                        found = True
                        ok_count = num_close(r[1], exp_r[1], 1)
                        record(f"Rating {rating_val} count", ok_count,
                               f"Expected {exp_r[1]}, got {r[1]}")
                        if not ok_count:
                            all_ok = False
                        break
            if not found:
                record(f"Rating {rating_val} found", False, "Missing")
                all_ok = False

    # Check Summary sheet
    sum_sheet = _find_sheet(wb, ("summary",))
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Summary' exists", True)
        summary = {}
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

        for key, val in summary.items():
            if "total" in key and "review" in key:
                ok = num_close(val, expected["total_reviews"], 5)
                record("Summary Total_Reviews", ok,
                       f"Expected {expected['total_reviews']}, got {val}")
                if not ok:
                    all_ok = False
            elif key == "avg_rating" or ("avg" in key and "rating" in key):
                ok = num_close(val, expected["overall_avg"], 0.1)
                record("Summary Avg_Rating", ok,
                       f"Expected {expected['overall_avg']}, got {val}")
                if not ok:
                    all_ok = False

    wb.close()
    return all_ok


def check_word(agent_workspace):
    """Check Review_Sentiment_Report.docx (non-critical, RU+EN keywords)."""
    print("\n=== Checking Review_Sentiment_Report.docx ===")
    from docx import Document

    docx_file = os.path.join(agent_workspace, "Review_Sentiment_Report.docx")
    if not os.path.isfile(docx_file):
        record("Word file exists", False, f"Not found: {docx_file}")
        return False
    record("Word file exists", True)

    try:
        doc = Document(docx_file)
    except Exception as e:
        record("Word readable", False, str(e))
        return False

    all_text = " ".join(p.text.lower() for p in doc.paragraphs)
    record("Word mentions sentiment (RU/EN)",
           any(k in all_text for k in ["sentiment", "тональност", "настроени"]),
           "No mention of sentiment/тональность/настроение")
    record("Word mentions review (RU/EN)",
           any(k in all_text for k in ["review", "отзыв"]),
           "No mention of review/отзыв")

    return True


def _gsheet_rating_distribution_cells():
    """Return list of (rating, count) parsed from the GSheet 'Rating Distribution'
    sheet of the 'Review Sentiment Dashboard' spreadsheet, or None if not found."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM gsheet.spreadsheets "
        "WHERE title ILIKE '%%review sentiment dashboard%%' "
        "OR title ILIKE '%%review%%' OR title ILIKE '%%sentiment%%'"
    )
    ss = cur.fetchall()
    if not ss:
        cur.close()
        conn.close()
        return None, None
    # find a 'Rating Distribution' sheet under any matching spreadsheet
    for (ssid,) in ss:
        cur.execute(
            "SELECT id FROM gsheet.sheets WHERE spreadsheet_id=%s "
            "AND title ILIKE '%%rating%%' AND title ILIKE '%%distribution%%'",
            (ssid,),
        )
        sh = cur.fetchone()
        if not sh:
            continue
        sheet_id = sh[0]
        cur.execute(
            "SELECT row_index, col_index, value FROM gsheet.cells "
            "WHERE spreadsheet_id=%s AND sheet_id=%s",
            (ssid, sheet_id),
        )
        cells = cur.fetchall()
        cur.close()
        conn.close()
        return ssid, cells
    cur.close()
    conn.close()
    return ss[0][0], None


def check_gsheet():
    """Check Google Sheet 'Review Sentiment Dashboard' exists (non-critical)."""
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM gsheet.spreadsheets "
        "WHERE title ILIKE '%%review%%' OR title ILIKE '%%sentiment%%'"
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()
    found = len(rows) > 0
    record("GSheet with 'review'/'sentiment' in title", found, "No matching spreadsheet found")
    return found


# ---------------------------------------------------------------------------
# CRITICAL semantic checks: any failure => hard FAIL (sys.exit(1)).
# ---------------------------------------------------------------------------
def check_critical(agent_workspace, expected):
    print("\n=== CRITICAL CHECKS ===")
    critical_ok = True

    def crit(name, passed, detail=""):
        nonlocal critical_ok
        record(name, passed, detail)
        if not passed:
            critical_ok = False

    agent_file = os.path.join(agent_workspace, "Review_Sentiment_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        crit("[CRITICAL] Excel file exists", False, f"Not found: {agent_file}")
        return critical_ok
    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        crit("[CRITICAL] Excel readable", False, str(e))
        return critical_ok

    # ---- Summary: Most_Reviewed_Product & Highest_Rated_Product ----
    sum_sheet = _find_sheet(wb, ("summary",))
    summary = {}
    if sum_sheet is not None:
        for row in sum_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                summary[str(row[0]).strip().lower()] = row[1]

    most_val = None
    high_val = None
    for k, v in summary.items():
        if "most" in k and "review" in k:
            most_val = v
        if "highest" in k and "rat" in k:
            high_val = v

    crit("[CRITICAL] Summary Most_Reviewed_Product matches expected",
         most_val is not None and norm_name(most_val) == norm_name(expected["most_reviewed"]),
         f"Expected {expected['most_reviewed']!r}, got {most_val!r}")
    crit("[CRITICAL] Summary Highest_Rated_Product matches expected",
         high_val is not None and norm_name(high_val) == norm_name(expected["highest_rated"]),
         f"Expected {expected['highest_rated']!r}, got {high_val!r}")

    # ---- Product Reviews: top-20 set + order + per-row metrics ----
    pr_sheet = _find_sheet(wb, ("product", "review"))
    if pr_sheet is None:
        crit("[CRITICAL] 'Product Reviews' sheet present", False)
    else:
        rows = list(pr_sheet.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if r and r[0]]
        exp = expected["top_products"]  # name, count, avg, five, one
        # Same ordered set of names
        got_names = [norm_name(r[0]) for r in rows]
        exp_names = [norm_name(p[0]) for p in exp]
        crit("[CRITICAL] Product Reviews top-20 names+order match expected",
             got_names == exp_names,
             f"got={got_names[:5]}... exp={exp_names[:5]}...")
        # Per-row metric correctness for the rows that line up
        metric_ok = True
        bad = None
        for r, p in zip(rows, exp):
            if norm_name(r[0]) != norm_name(p[0]):
                metric_ok = False
                bad = (r[0], "name mismatch")
                break
            if not (num_close(r[1], p[1], 1) and num_close(r[2], p[2], 0.05)
                    and num_close(r[3], p[3], 1) and num_close(r[4], p[4], 1)):
                metric_ok = False
                bad = (r[0], f"count={r[1]}/{p[1]} avg={r[2]}/{p[2]} 5*={r[3]}/{p[3]} 1*={r[4]}/{p[4]}")
                break
        crit("[CRITICAL] Product Reviews per-row Count/Avg/Five_Star/One_Star correct",
             metric_ok, f"bad row: {bad}")

    # ---- Rating Distribution exact counts ----
    rd_sheet = _find_sheet(wb, ("rating", "distribution"))
    if rd_sheet is None:
        crit("[CRITICAL] 'Rating Distribution' sheet present", False)
    else:
        rows = list(rd_sheet.iter_rows(min_row=2, values_only=True))
        got = {}
        for r in rows:
            if r and r[0] is not None:
                try:
                    got[int(r[0])] = r[1]
                except (TypeError, ValueError):
                    pass
        dist_ok = True
        detail = []
        for rating_val, cnt, pct in expected["rating_dist"]:
            g = got.get(rating_val)
            if not num_close(g, cnt, 1):
                dist_ok = False
                detail.append(f"r{rating_val}: exp {cnt} got {g}")
        crit("[CRITICAL] Rating Distribution counts match all 5 ratings (tol<=1)",
             dist_ok, "; ".join(detail))

    wb.close()

    # ---- GSheet Rating Distribution cells mirror Excel data ----
    ssid, cells = _gsheet_rating_distribution_cells()
    if ssid is None:
        crit("[CRITICAL] GSheet 'Review Sentiment Dashboard' exists", False)
    elif cells is None:
        crit("[CRITICAL] GSheet 'Rating Distribution' sheet exists with cells", False)
    else:
        # Build per-row text; locate a numeric rating (1..5) -> its count on same row.
        by_row = {}
        for ri, ci, val in cells:
            by_row.setdefault(ri, {})[ci] = val
        # expected count per rating
        exp_counts = {int(rv): int(cnt) for rv, cnt, pct in expected["rating_dist"]}
        matched = {}
        for ri, cols in by_row.items():
            ordered = [cols[c] for c in sorted(cols)]
            nums = []
            for v in ordered:
                try:
                    nums.append(int(float(v)))
                except (TypeError, ValueError):
                    nums.append(None)
            # find a cell that is a rating 1..5, with a count cell after it
            for idx, n in enumerate(nums):
                if n in exp_counts:
                    # count = next numeric cell on the row
                    for n2 in nums[idx + 1:]:
                        if n2 is not None:
                            matched[n] = n2
                            break
                    break
        gsheet_ok = len(matched) >= 5 and all(
            num_close(matched.get(rv), cnt, 1) for rv, cnt in exp_counts.items()
        )
        crit("[CRITICAL] GSheet Rating Distribution counts mirror Excel (all 5 ratings)",
             gsheet_ok, f"matched={matched} expected={exp_counts}")

    return critical_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    expected = get_expected_data()

    excel_ok = check_excel(args.agent_workspace, expected)
    word_ok = check_word(args.agent_workspace)

    db_fail_before = FAIL_COUNT
    gsheet_ok = check_gsheet()
    db_failures = FAIL_COUNT - db_fail_before

    # SEMANTIC critical checks: any failure => hard FAIL regardless of accuracy.
    critical_ok = check_critical(args.agent_workspace, expected)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Word:     {'PASS' if word_ok else 'FAIL'}")
    print(f"  GSheet:   {'PASS' if gsheet_ok else 'FAIL'} (non-blocking)")
    print(f"  Critical: {'PASS' if critical_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if db_failures > 0:
        print(f"  NOTE: {db_failures} GSheet checks failed (non-blocking for accuracy gate)")

    if not critical_ok:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = (accuracy >= 70) and excel_ok and word_ok
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
