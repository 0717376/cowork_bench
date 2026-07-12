"""Evaluation for sf-sales-brand-performance (ClickHouse fork).

The sf_data SALES_DW data is russified centrally (db/zzz_clickhouse_after_init.sql).
Brand-name literals (LG/Microsoft/Bose/...) are proper nouns and stay EXACTLY English,
so the static groundtruth Brand_Performance.xlsx remains value-consistent with the
clickhouse-served data and brand lookup keys match.

Structure: structural/value checks are NON-critical (loose tolerances). A small set of
CRITICAL semantic checks (Top_Brand, total revenue, top-3 sort order, share consistency,
docx names top brands) hard-FAIL before the accuracy>=70 gate.
"""
import argparse
import os
import sys
import openpyxl


PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Brand_Performance.xlsx")
    gt_file = os.path.join(gt_dir, "Brand_Performance.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ----- Brand Rankings (non-critical per-cell, loose tolerances) -----
    print("  Checking Brand Rankings...")
    a_rows = load_sheet_rows(agent_wb, "Brand Rankings")
    g_rows = load_sheet_rows(gt_wb, "Brand Rankings")

    a_brand_data = []  # list of (brand, products, orders, revenue, aov, share)
    g_brand_data = []

    if a_rows is None:
        record("Sheet 'Brand Rankings' present", False, "missing in agent output")
    elif g_rows is None:
        record("Sheet 'Brand Rankings' present in groundtruth", False)
    else:
        record("Sheet 'Brand Rankings' present", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        for row in a_data:
            if row and row[0] is not None:
                a_brand_data.append(row)
        for row in g_data:
            if row and row[0] is not None:
                g_brand_data.append(row)

        a_lookup = {str(r[0]).strip().lower(): r for r in a_brand_data}
        for g_row in g_brand_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Brand row present: {g_row[0]}", False, "Missing row")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                record(f"{key}.Products", num_close(a_row[1], g_row[1], 1),
                       f"{a_row[1]} vs {g_row[1]} (tol=1)")
            if len(a_row) > 2 and len(g_row) > 2:
                record(f"{key}.Orders", num_close(a_row[2], g_row[2], 5),
                       f"{a_row[2]} vs {g_row[2]} (tol=5)")
            if len(a_row) > 3 and len(g_row) > 3:
                record(f"{key}.Revenue", num_close(a_row[3], g_row[3], 50.0),
                       f"{a_row[3]} vs {g_row[3]} (tol=50.0)")
            if len(a_row) > 4 and len(g_row) > 4:
                record(f"{key}.Avg_Order_Value", num_close(a_row[4], g_row[4], 2.0),
                       f"{a_row[4]} vs {g_row[4]} (tol=2.0)")
            if len(a_row) > 5 and len(g_row) > 5:
                record(f"{key}.Revenue_Share_Pct", num_close(a_row[5], g_row[5], 0.5),
                       f"{a_row[5]} vs {g_row[5]} (tol=0.5)")

    # ----- Summary (non-critical per-cell) -----
    print("  Checking Summary...")
    a_sum_rows = load_sheet_rows(agent_wb, "Summary")
    g_sum_rows = load_sheet_rows(gt_wb, "Summary")
    a_summary = {}
    g_summary = {}
    if a_sum_rows is None:
        record("Sheet 'Summary' present", False, "missing in agent output")
    elif g_sum_rows is None:
        record("Sheet 'Summary' present in groundtruth", False)
    else:
        record("Sheet 'Summary' present", True)
        for row in (a_sum_rows[1:] if len(a_sum_rows) > 1 else []):
            if row and row[0] is not None:
                a_summary[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
        for row in (g_sum_rows[1:] if len(g_sum_rows) > 1 else []):
            if row and row[0] is not None:
                g_summary[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
        for key, g_val in g_summary.items():
            a_val = a_summary.get(key)
            if a_val is None and key not in a_summary:
                record(f"Summary.{key} present", False, "Missing row")
                continue
            record(f"Summary.{key}", num_close(a_val, g_val, 50.0),
                   f"{a_val} vs {g_val} (tol=50.0)")

    # ----- docx (RU+EN tolerant) -----
    docx_path = os.path.join(args.agent_workspace, "Brand_Summary.docx")
    doc_text = ""
    if not os.path.exists(docx_path):
        record("Brand_Summary.docx present", False, "not found")
    else:
        record("Brand_Summary.docx present", True)
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            doc_text = " ".join(p.text for p in _doc.paragraphs)
            record("Brand_Summary.docx has content (>=50 chars)",
                   len(doc_text.strip()) >= 50, f"len={len(doc_text.strip())}")
        except ImportError:
            record("Brand_Summary.docx non-trivial size",
                   os.path.getsize(docx_path) >= 100)
        except Exception as e:
            record("Brand_Summary.docx readable", False, str(e))

    # ===================== CRITICAL SEMANTIC CHECKS =====================
    # Derive groundtruth ordering from the groundtruth sheet (revenue desc).
    gt_sorted = sorted(
        [(str(r[0]).strip(), to_float(r[3]) if len(r) > 3 else None) for r in g_brand_data
         if len(r) > 3 and to_float(r[3]) is not None],
        key=lambda t: t[1], reverse=True)

    # Agent ordering as written (do NOT re-sort) to verify Revenue-descending.
    a_order = [(str(r[0]).strip(), to_float(r[3]) if len(r) > 3 else None)
               for r in a_brand_data if len(r) > 3 and to_float(r[3]) is not None]

    gt_top_brand = gt_sorted[0][0] if gt_sorted else None
    gt_total_rev = sum(v for _, v in gt_sorted) if gt_sorted else None

    # CRITICAL 1: Top_Brand in Summary equals max-revenue brand and groundtruth Top_Brand.
    a_top = a_summary.get("top_brand")
    top_ok = (a_top is not None and gt_top_brand is not None
              and str(a_top).strip().lower() == gt_top_brand.lower())
    record("Top_Brand equals max-revenue brand (groundtruth)", top_ok,
           f"agent={a_top}, expected={gt_top_brand}", critical=True)

    # CRITICAL 2: Total_Revenue equals sum of per-brand Revenue and groundtruth total.
    a_total = to_float(a_summary.get("total_revenue"))
    total_ok = (a_total is not None and gt_total_rev is not None
                and abs(a_total - gt_total_rev) <= 50.0)
    record("Total_Revenue matches sum of brand revenues (groundtruth)", total_ok,
           f"agent={a_total}, expected={gt_total_rev}", critical=True)

    # CRITICAL 3: Top-3 brands by Revenue descending match groundtruth ordering.
    gt_top3 = [b.lower() for b, _ in gt_sorted[:3]]
    a_top3 = [b.lower() for b, _ in a_order[:3]]
    top3_ok = (len(a_top3) >= 3 and a_top3 == gt_top3)
    record("Top-3 brands by Revenue descending match groundtruth", top3_ok,
           f"agent={a_top3}, expected={gt_top3}", critical=True)

    # CRITICAL 4: Top_Brand_Share_Pct internally consistent with Revenue/Total.
    a_top_rev = to_float(a_summary.get("top_brand_revenue"))
    a_top_share = to_float(a_summary.get("top_brand_share_pct"))
    if a_top_rev is not None and a_total not in (None, 0) and a_top_share is not None:
        expected_share = a_top_rev / a_total * 100.0
        share_ok = abs(expected_share - a_top_share) <= 0.5
    else:
        share_ok = False
    record("Top_Brand_Share_Pct consistent with Top_Brand_Revenue/Total_Revenue",
           share_ok, f"reported={a_top_share}, computed={a_top_rev}/{a_total}",
           critical=True)

    # CRITICAL 5: docx names the actual top-3 brand names (proper nouns stay English).
    dl = doc_text.lower()
    named = sum(1 for b, _ in gt_sorted[:3] if b.lower() in dl)
    record("Brand_Summary.docx names the top-3 brands by revenue",
           named >= 3, f"named {named}/3 of {gt_top3}; text: {doc_text[:200]}",
           critical=True)

    # ===================== GATE =====================
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (threshold accuracy>=70)")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
