"""Evaluation for sf-sales-monthly-trends.

Critical checks (CRITICAL_CHECKS): любой провал критического чека => общий FAIL
независимо от accuracy. Иначе PASS требует accuracy >= 70%.

Идентификаторы (имена файлов, листов, столбцов, ключи метрик Total_Months/
Peak_Month/..., значение Peak_Month '2024-12-01') — это литералы вывода,
сравниваемые напрямую, и ДОЛЖНЫ оставаться английскими/числовыми. Не переводить.
"""
import argparse
import os
import sys
import openpyxl


PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "Monthly Trends: все месяцы из groundtruth присутствуют, Revenue в пределах допуска",
    "Summary.Peak_Month совпадает с groundtruth",
    "Summary.Peak_Revenue совпадает с groundtruth",
    "Summary.Total_Revenue совпадает с groundtruth",
    "Summary.Total_Months совпадает с groundtruth",
    "Trends_Summary.docx: текст >=50 символов и есть RU/EN ключевое слово о динамике/сводке",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def summary_lookup(rows):
    """Return {metric_key_lower: value} from a Metric/Value sheet (skip header)."""
    out = {}
    for row in (rows[1:] if rows and len(rows) > 1 else []):
        if row and row[0] is not None:
            out[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Sales_Monthly_Trends.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Monthly_Trends.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- Monthly Trends sheet ----
    print("  Checking Monthly Trends...")
    a_rows = load_sheet_rows(agent_wb, "Monthly Trends")
    g_rows = load_sheet_rows(gt_wb, "Monthly Trends")
    if a_rows is None:
        check("Sheet 'Monthly Trends' present in agent output", False, "sheet missing")
    elif g_rows is None:
        check("Sheet 'Monthly Trends' present in groundtruth", False, "sheet missing")
    else:
        check("Sheet 'Monthly Trends' present in agent output", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        all_months_present = True
        all_revenue_ok = True
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_months_present = False
                all_revenue_ok = False
                # non-critical per-row visibility checks
                check(f"Monthly Trends row present: {g_row[0]}", False, "missing row")
                continue
            check(f"Monthly Trends row present: {g_row[0]}", True)

            if len(a_row) > 1 and len(g_row) > 1:
                ok = num_close(a_row[1], g_row[1], 10.0)
                all_revenue_ok = all_revenue_ok and ok
                check(f"{key}.Revenue", ok, f"{a_row[1]} vs {g_row[1]} (tol=10.0)")
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"{key}.Order_Count", num_close(a_row[2], g_row[2], 2),
                      f"{a_row[2]} vs {g_row[2]} (tol=2)")
            if len(a_row) > 3 and len(g_row) > 3:
                check(f"{key}.Unique_Customers", num_close(a_row[3], g_row[3], 2),
                      f"{a_row[3]} vs {g_row[3]} (tol=2)")
            if len(a_row) > 4 and len(g_row) > 4:
                check(f"{key}.Avg_Order_Value", num_close(a_row[4], g_row[4], 1.0),
                      f"{a_row[4]} vs {g_row[4]} (tol=1.0)")

        # CRITICAL: no missing months and all revenue within tolerance
        check("Monthly Trends: все месяцы из groundtruth присутствуют, Revenue в пределах допуска",
              all_months_present and all_revenue_ok,
              "missing rows or revenue mismatch")

    # ---- Summary sheet ----
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        check("Sheet 'Summary' present in agent output", False, "sheet missing")
    elif g_rows is None:
        check("Sheet 'Summary' present in groundtruth", False, "sheet missing")
    else:
        check("Sheet 'Summary' present in agent output", True)
        a_sum = summary_lookup(a_rows)
        g_sum = summary_lookup(g_rows)

        # Generic non-critical per-metric numeric check (matches old behaviour, tol=100)
        for key, gv in g_sum.items():
            av = a_sum.get(key)
            check(f"Summary[{key}].Value", num_close(av, gv, 100.0),
                  f"{av} vs {gv} (tol=100.0)")

        # CRITICAL semantic checks
        check("Summary.Peak_Month совпадает с groundtruth",
              str_match(a_sum.get("peak_month"), g_sum.get("peak_month")),
              f"{a_sum.get('peak_month')} vs {g_sum.get('peak_month')}")
        check("Summary.Peak_Revenue совпадает с groundtruth",
              num_close(a_sum.get("peak_revenue"), g_sum.get("peak_revenue"), 10.0),
              f"{a_sum.get('peak_revenue')} vs {g_sum.get('peak_revenue')}")
        check("Summary.Total_Revenue совпадает с groundtruth",
              num_close(a_sum.get("total_revenue"), g_sum.get("total_revenue"), 100.0),
              f"{a_sum.get('total_revenue')} vs {g_sum.get('total_revenue')}")
        check("Summary.Total_Months совпадает с groundtruth",
              num_close(a_sum.get("total_months"), g_sum.get("total_months"), 0),
              f"{a_sum.get('total_months')} vs {g_sum.get('total_months')}")

    # ---- Word narrative ----
    print("  Checking Trends_Summary.docx...")
    docx_path = os.path.join(args.agent_workspace, "Trends_Summary.docx")
    if not os.path.exists(docx_path):
        check("Trends_Summary.docx exists", False, "file missing")
        check("Trends_Summary.docx: текст >=50 символов и есть RU/EN ключевое слово о динамике/сводке",
              False, "file missing")
    else:
        check("Trends_Summary.docx exists", True)
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            # ORIGINAL text lowercased (NOT normalized) for RU keyword matching
            _text = " ".join(p.text for p in _doc.paragraphs).lower()
            _len_ok = len(_text.strip()) >= 50
            # RU + EN trend/summary keywords
            _kws = ["динамик", "тренд", "сводк", "итог", "trends", "summary"]
            _kw_ok = any(k in _text for k in _kws)
            check("Trends_Summary.docx: текст >=50 символов и есть RU/EN ключевое слово о динамике/сводке",
                  _len_ok and _kw_ok,
                  f"len_ok={_len_ok} kw_ok={_kw_ok}")
        except ImportError:
            ok = os.path.getsize(docx_path) >= 100
            check("Trends_Summary.docx: текст >=50 символов и есть RU/EN ключевое слово о динамике/сводке",
                  ok, "python-docx unavailable; size fallback")
        except Exception as _e:
            check("Trends_Summary.docx: текст >=50 символов и есть RU/EN ключевое слово о динамике/сводке",
                  False, f"error reading docx: {_e}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    print("\n=== RESULT: PASS ===" if success else "\n=== RESULT: FAIL ===")
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
