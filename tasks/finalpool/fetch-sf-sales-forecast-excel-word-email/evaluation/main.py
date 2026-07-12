"""Evaluation script for fetch-sf-sales-forecast-excel-word-email task (ClickHouse fork).

Region literals are CYRILLIC: the sf_data SALES_DW REGION values are russified centrally
(db/zzz_clickhouse_after_init.sql), and the task-local mock market API JSON
(files/mock_pages.tar.gz) is russified in lockstep, so the agent's region-name join
(DWH REGION <-> market_projections.region) still matches and the Excel Region columns
carry the same Cyrillic labels the eval keys on.

Numeric values are unchanged from the English source (totals 3048998.33 -> 3219098.33,
increase 170100, avg 5.6).
"""
import argparse
import json
import os
import sys

import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Cyrillic region labels (must match db/zzz_clickhouse_after_init.sql and the mock API JSON).
RU_REGIONS = {
    "Asia Pacific": "Азиатско-Тихоокеанский регион",
    "Europe": "Европа",
    "Latin America": "Латинская Америка",
    "Middle East": "Ближний Восток",
    "North America": "Северная Америка",
}
HIGHEST_GROWTH_REGION = RU_REGIONS["Asia Pacific"]   # 8.5
LOWEST_GROWTH_REGION = RU_REGIONS["Europe"]          # 3.2

# Expected per-region figures (unchanged from English source).
EXPECTED = {
    RU_REGIONS["Asia Pacific"]:   {"orders": 4198, "current": 642644.81, "growth": 8.5, "forecast": 697269.62, "share": 21.1},
    RU_REGIONS["Europe"]:         {"orders": 4100, "current": 648798.47, "growth": 3.2, "forecast": 669560.02, "share": 21.3},
    RU_REGIONS["Latin America"]:  {"orders": 3697, "current": 549129.15, "growth": 6.8, "forecast": 586469.93, "share": 18.0},
    RU_REGIONS["Middle East"]:    {"orders": 3872, "current": 602107.55, "growth": 5.4, "forecast": 634621.36, "share": 19.7},
    RU_REGIONS["North America"]:  {"orders": 4133, "current": 606318.35, "growth": 4.1, "forecast": 631177.40, "share": 19.9},
}
TOTAL_CURRENT = 3048998.33
TOTAL_FORECAST = 3219098.33
TOTAL_INCREASE = 170100.0
AVG_GROWTH = 5.6


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def num_close(a, b, tol=1.0):
    a = safe_float(a)
    b = safe_float(b)
    if a is None or b is None:
        return False
    return abs(a - b) <= tol


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def sheet_rows_as_dicts(ws):
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return out


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    excel_path = os.path.join(agent_workspace, "Sales_Forecast_Q2_2026.xlsx")
    record("Sales_Forecast_Q2_2026.xlsx exists", os.path.exists(excel_path), critical=True)

    wb = None
    if os.path.exists(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path)
        except Exception as e:
            record("Excel file readable", False, str(e), critical=True)

    # ---------------- Regional_Forecast ----------------
    rf_rows = []
    if wb is not None:
        ok = "Regional_Forecast" in wb.sheetnames
        record("Regional_Forecast sheet exists", ok, f"sheets: {wb.sheetnames}")
        if ok:
            ws = wb["Regional_Forecast"]
            rf_rows = sheet_rows_as_dicts(ws)
            record("Regional_Forecast has >= 5 rows", len(rf_rows) >= 5, f"got {len(rf_rows)}")
            headers_lower = [h.lower() for h in (rf_rows[0].keys() if rf_rows else [])]
            for col in ["Region", "Current_Revenue", "Forecasted_Revenue"]:
                record(f"Regional_Forecast has {col} column", col.lower() in headers_lower,
                       f"headers: {list(rf_rows[0].keys()) if rf_rows else []}")

    # build region -> row map (case-insensitive header lookup)
    def get(row, key):
        for k, v in row.items():
            if k.lower() == key.lower():
                return v
        return None

    rf_by_region = {}
    for r in rf_rows:
        reg = get(r, "Region")
        if reg is not None:
            rf_by_region[str(reg).strip()] = r

    # CRITICAL 1: per-region Forecasted_Revenue == Current*(1+growth/100), growth join correct
    region_formula_ok = True
    formula_detail = ""
    for reg, exp in EXPECTED.items():
        row = rf_by_region.get(reg)
        if row is None:
            region_formula_ok = False
            formula_detail = f"missing region {reg}; have {list(rf_by_region.keys())}"
            break
        cur = safe_float(get(row, "Current_Revenue"))
        gr = safe_float(get(row, "Growth_Rate_Pct"))
        fc = safe_float(get(row, "Forecasted_Revenue"))
        if cur is None or gr is None or fc is None:
            region_formula_ok = False
            formula_detail = f"{reg}: non-numeric cur/gr/fc"
            break
        if not num_close(gr, exp["growth"], tol=0.05):
            region_formula_ok = False
            formula_detail = f"{reg}: growth {gr} != {exp['growth']} (market-API join failed)"
            break
        expected_fc = round(cur * (1 + gr / 100.0), 2)
        if not num_close(fc, expected_fc, tol=1.0):
            region_formula_ok = False
            formula_detail = f"{reg}: forecast {fc} != {expected_fc}"
            break
    record("Per-region Forecasted_Revenue applies growth-rate formula (market-API join)",
           region_formula_ok, formula_detail, critical=True)

    # ---------------- Forecast_Summary ----------------
    summary = {}
    if wb is not None and "Forecast_Summary" in wb.sheetnames:
        record("Forecast_Summary sheet exists", True)
        ws = wb["Forecast_Summary"]
        rows = sheet_rows_as_dicts(ws)
        record("Forecast_Summary has >= 5 rows", len(rows) >= 5, f"got {len(rows)}")
        hl = [h.lower() for h in (rows[0].keys() if rows else [])]
        for col in ["Metric", "Value"]:
            record(f"Forecast_Summary has {col} column", col.lower() in hl)
        for r in rows:
            m = get(r, "Metric")
            v = get(r, "Value")
            if m is not None:
                summary[str(m).strip()] = v
    else:
        record("Forecast_Summary sheet exists", False,
               f"sheets: {wb.sheetnames if wb else None}")

    # CRITICAL 2: Highest/Lowest growth region + Avg_Growth_Rate
    hi = summary.get("Highest_Growth_Region")
    lo = summary.get("Lowest_Growth_Region")
    avg = summary.get("Avg_Growth_Rate")
    record("Highest_Growth_Region == Азиатско-Тихоокеанский регион",
           hi is not None and str(hi).strip() == HIGHEST_GROWTH_REGION,
           f"got {hi!r}", critical=True)
    record("Lowest_Growth_Region == Европа",
           lo is not None and str(lo).strip() == LOWEST_GROWTH_REGION,
           f"got {lo!r}", critical=True)
    record("Avg_Growth_Rate == 5.6", num_close(avg, AVG_GROWTH, tol=0.05),
           f"got {avg!r}", critical=True)

    # CRITICAL 3: cross-sheet consistency of totals
    tot_cur = summary.get("Total_Current_Revenue")
    tot_fc = summary.get("Total_Forecasted_Revenue")
    tot_inc = summary.get("Total_Revenue_Increase")
    record("Total_Current_Revenue == 3048998.33", num_close(tot_cur, TOTAL_CURRENT, tol=2.0),
           f"got {tot_cur!r}")
    sum_region_fc = sum(safe_float(get(r, "Forecasted_Revenue"), 0.0) for r in rf_rows) if rf_rows else None
    cross_ok = (
        num_close(tot_fc, TOTAL_FORECAST, tol=2.0)
        and (sum_region_fc is not None and num_close(tot_fc, sum_region_fc, tol=2.0))
        and num_close(tot_inc, TOTAL_INCREASE, tol=2.0)
        and num_close(safe_float(tot_fc, 0) - safe_float(tot_cur, 0), TOTAL_INCREASE, tol=2.0)
    )
    record("Total_Forecasted == sum(region forecasts) and increase == forecast-current",
           cross_ok,
           f"tot_fc={tot_fc} sum_region={sum_region_fc} tot_cur={tot_cur} tot_inc={tot_inc}",
           critical=True)

    # ---------------- Growth_Ranking ----------------
    gr_rows = []
    if wb is not None and "Growth_Ranking" in wb.sheetnames:
        record("Growth_Ranking sheet exists", True)
        ws = wb["Growth_Ranking"]
        gr_rows = sheet_rows_as_dicts(ws)
        record("Growth_Ranking has >= 5 rows", len(gr_rows) >= 5, f"got {len(gr_rows)}")
        hl = [h.lower() for h in (gr_rows[0].keys() if gr_rows else [])]
        for col in ["Rank", "Region", "Growth_Rate_Pct"]:
            record(f"Growth_Ranking has {col} column", col.lower() in hl)
    else:
        record("Growth_Ranking sheet exists", False,
               f"sheets: {wb.sheetnames if wb else None}")

    # CRITICAL 4: ranking sorted by growth desc, Rank 1..N, Market_Share_Pct correct
    ranking_ok = True
    rank_detail = ""
    if len(gr_rows) >= 5:
        prev_growth = None
        for idx, r in enumerate(gr_rows[:5], start=1):
            rk = safe_float(get(r, "Rank"))
            reg = str(get(r, "Region")).strip()
            g = safe_float(get(r, "Growth_Rate_Pct"))
            share = safe_float(get(r, "Market_Share_Pct"))
            if rk is None or int(rk) != idx:
                ranking_ok = False
                rank_detail = f"row {idx}: Rank={rk}"
                break
            if g is None or (prev_growth is not None and g > prev_growth + 1e-6):
                ranking_ok = False
                rank_detail = f"row {idx}: growth {g} not descending (prev {prev_growth})"
                break
            prev_growth = g
            exp = EXPECTED.get(reg)
            if exp is None:
                ranking_ok = False
                rank_detail = f"row {idx}: unknown region {reg!r}"
                break
            if not num_close(share, exp["share"], tol=0.15):
                ranking_ok = False
                rank_detail = f"{reg}: Market_Share_Pct {share} != {exp['share']}"
                break
    else:
        ranking_ok = False
        rank_detail = f"only {len(gr_rows)} ranking rows"
    record("Growth_Ranking sorted desc, Rank 1..5, Market_Share_Pct correct",
           ranking_ok, rank_detail, critical=True)

    # ---------------- Word report ----------------
    word_path = os.path.join(agent_workspace, "Q2_Forecast_Report.docx")
    record("Q2_Forecast_Report.docx exists", os.path.exists(word_path))
    if os.path.exists(word_path):
        try:
            from docx import Document
            doc = Document(word_path)
            text = " ".join(p.text for p in doc.paragraphs)
            text_lower = text.lower()
            # Russian prose (or English heading) keyword checks.
            forecast_kw = any(k in text_lower for k in ["прогноз", "forecast"])
            growth_kw = any(k in text_lower for k in ["рост", "темп", "growth"])
            record("Word report mentions forecast (прогноз/forecast)", forecast_kw)
            record("Word report mentions growth (рост/темп/growth)", growth_kw)
            # references a concrete figure (top forecasted total)
            record("Word report references a concrete forecast figure",
                   any(s in text for s in ["3,219,098", "3219098", "697,269", "697269"]),
                   text[:200])
        except Exception as e:
            record("Word report readable", False, str(e))

    # ---------------- forecast_builder.py ----------------
    record("forecast_builder.py exists",
           os.path.exists(os.path.join(agent_workspace, "forecast_builder.py")))

    # ---------------- Email ----------------
    # CRITICAL 5: email to sales-team@company.com with the forecast subject + summary content
    email_ok = False
    email_detail = ""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, to_addr, body_text FROM email.messages WHERE subject ILIKE %s",
            ("%forecast%",),
        )
        emails = cur.fetchall()
        conn.close()
        record("Forecast email exists", len(emails) >= 1, f"found {len(emails)}")
        for subject, to_addr, body in emails:
            to_s = str(to_addr).lower()
            body_s = str(body or "")
            body_l = body_s.lower()
            to_ok = "sales-team@company.com" in to_s
            # body references total forecasted revenue figure and top growth region
            figure_ok = any(s in body_s for s in ["3,219,098", "3219098", "3 219 098"])
            region_ok = (HIGHEST_GROWTH_REGION.lower() in body_l) or ("азиатско" in body_l)
            if to_ok and figure_ok and region_ok:
                email_ok = True
                break
            email_detail = f"to_ok={to_ok} figure_ok={figure_ok} region_ok={region_ok} subj={subject!r}"
        if not email_ok and not email_detail and emails:
            email_detail = "no email matched all sub-conditions"
    except Exception as e:
        email_detail = str(e)
    record("Email to sales-team with total forecast figure and top growth region",
           email_ok, email_detail, critical=True)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILS:
        print(f"  Critical failures: {CRITICAL_FAILS}")
        return False, f"FAIL (critical): {CRITICAL_FAILS}"

    overall = accuracy >= 70
    return overall, f"Passed {PASS_COUNT}/{total} checks (accuracy {accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
