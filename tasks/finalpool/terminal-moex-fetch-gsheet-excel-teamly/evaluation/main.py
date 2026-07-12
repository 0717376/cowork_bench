"""Evaluation for terminal-moex-fetch-gsheet-excel-teamly.

The agent reads a PDF strategy + market_params.json, fetches macro indicators
from a local mock HTTP feed, pulls SBER.ME/GAZP.ME/MGNT.ME data from the
moex-finance MCP (name/sector/latest close/30d return), runs market_analyzer.py
to compute BUY/SELL/HOLD signals + qualitative correlations, builds a 4-sheet
Excel report (Stock_Overview, Economic_Indicators, Correlation_Matrix,
Portfolio_Signals), publishes Stock_Overview to a cloud spreadsheet
("Market Analysis Live Data"), and creates teamly knowledge-base pages
("Investment Research Log", one per stock).

Gate: any CRITICAL check failure => overall FAIL (sys.exit(1)) regardless of
accuracy. Otherwise pass threshold: accuracy >= 70%.

Signals are derived from honestly-read MOEX prices (volatile) — the critical
signal check RE-DERIVES the expected signal from the agent's own reported 30d
return plus the known macro feed (consumer confidence 98.5 > 95, inflation
3.4 > 3.0), it does NOT hardcode an expected BUY/SELL per ticker.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

# Post-swap MOEX tickers (must match task.md + market_params.json).
SYMBOLS = ["SBER.ME", "GAZP.ME", "MGNT.ME"]

# Known macro feed values (from the mock indicators server; rules depend on them).
CONSUMER_CONFIDENCE = 98.5  # > 95 => satisfies BUY confidence condition
INFLATION = 3.4             # > 3.0 => satisfies SELL inflation condition

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Stock_Overview contains exactly the 3 required MOEX tickers with price>0 and numeric return",
    "Economic_Indicators: GDP=2.1 Favorable AND Inflation=3.4 Unfavorable",
    "Portfolio_Signals match BUY/SELL/HOLD rules re-derived from reported 30d returns",
    "Cloud spreadsheet 'Market Analysis Live Data' contains the 3 stock rows",
    "Teamly 'Investment Research Log': 3 stock pages, each naming its symbol and signal",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_float(v):
    try:
        return float(str(v).replace("%", "").replace(",", "").strip())
    except (TypeError, ValueError, AttributeError):
        return None


def expected_signal(ret_30d):
    """Re-derive the documented signal from the agent's reported 30d return and
    the known macro feed. Mirrors market_params.json rules:
      BUY  if return > 0 and consumer_confidence > 95
      SELL if return < 0 and inflation > 3.0
      else HOLD
    """
    if ret_30d is None:
        return None
    if ret_30d > 0 and CONSUMER_CONFIDENCE > 95:
        return "buy"
    if ret_30d < 0 and INFLATION > 3.0:
        return "sell"
    return "hold"


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def check_excel(agent_ws):
    print("\n=== Checking Market_Analysis_Report.xlsx ===")
    agent_file = os.path.join(agent_ws, "Market_Analysis_Report.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # ---- Sheet 1: Stock_Overview (CRITICAL: 3 required tickers, price>0, numeric return) ----
    print("  Checking Stock_Overview...")
    ws1 = get_sheet(awb, "Stock_Overview")
    check("Sheet Stock_Overview exists", ws1 is not None, f"Sheets: {awb.sheetnames}")
    overview = {}  # symbol -> Return_30d_Pct (float)
    if ws1:
        rows = list(ws1.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in rows if r and r[0]]
        check("Stock_Overview has 3 rows", len(data_rows) == 3, f"Got {len(data_rows)}")

        ok_overview = True
        details = []
        symbols = {str(r[0]).strip().upper() for r in data_rows if r and r[0]}
        for sym in SYMBOLS:
            if sym.upper() not in symbols:
                ok_overview = False
                details.append(f"missing {sym}")
        for r in data_rows:
            if not (r and len(r) >= 5 and r[0]):
                continue
            sym = str(r[0]).strip().upper()
            price = to_float(r[3])
            ret = to_float(r[4])
            if price is None or price <= 0:
                ok_overview = False
                details.append(f"{sym} bad price {r[3]}")
            if ret is None:
                ok_overview = False
                details.append(f"{sym} non-numeric return {r[4]}")
            else:
                overview[sym] = ret
        check("Stock_Overview contains exactly the 3 required MOEX tickers with price>0 and numeric return",
              ok_overview and symbols == {s.upper() for s in SYMBOLS},
              f"symbols={symbols}, issues={details}")

    # ---- Sheet 2: Economic_Indicators (CRITICAL: GDP/Inflation value+trend, un-gated) ----
    print("  Checking Economic_Indicators...")
    ws2 = get_sheet(awb, "Economic_Indicators")
    check("Sheet Economic_Indicators exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    if ws2:
        rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        data_rows2 = [r for r in rows2 if r and r[0]]
        check("Economic_Indicators has 5 rows", len(data_rows2) == 5, f"Got {len(data_rows2)}")

        ind_lookup = {}
        for r in data_rows2:
            if r and r[0]:
                ind_lookup[str(r[0]).strip().lower()] = r

        gdp = ind_lookup.get("gdp growth rate") or ind_lookup.get("gdp growth") \
            or ind_lookup.get("gdp")
        infl = ind_lookup.get("inflation rate") or ind_lookup.get("inflation")

        gdp_ok = bool(gdp) and num_close(gdp[1], 2.1, 0.2) and gdp[2] \
            and "favorable" in str(gdp[2]).lower() and "unfavorable" not in str(gdp[2]).lower()
        infl_ok = bool(infl) and num_close(infl[1], 3.4, 0.2) and infl[2] \
            and "unfavorable" in str(infl[2]).lower()
        check("Economic_Indicators: GDP=2.1 Favorable AND Inflation=3.4 Unfavorable",
              gdp_ok and infl_ok,
              f"gdp_row={gdp}, infl_row={infl}")

    # ---- Sheet 3: Correlation_Matrix (structural: 3 rows, valid value set) ----
    print("  Checking Correlation_Matrix...")
    ws3 = get_sheet(awb, "Correlation_Matrix")
    check("Sheet Correlation_Matrix exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        data_rows3 = [r for r in rows3 if r and r[0]]
        check("Correlation_Matrix has 3 rows", len(data_rows3) == 3, f"Got {len(data_rows3)}")

        valid_values = {"positive", "negative", "neutral"}
        all_valid = True
        for r in data_rows3:
            if r and len(r) >= 4:
                for col_idx in [1, 2, 3]:
                    val = str(r[col_idx]).strip().lower() if r[col_idx] else ""
                    if val not in valid_values:
                        all_valid = False
        check("All correlation values are valid", all_valid)

    # ---- Sheet 4: Portfolio_Signals (CRITICAL: signals re-derived from returns) ----
    print("  Checking Portfolio_Signals...")
    ws4 = get_sheet(awb, "Portfolio_Signals")
    check("Sheet Portfolio_Signals exists", ws4 is not None, f"Sheets: {awb.sheetnames}")
    if ws4:
        rows4 = list(ws4.iter_rows(min_row=2, values_only=True))
        data_rows4 = [r for r in rows4 if r and r[0]]
        check("Portfolio_Signals has 3 rows", len(data_rows4) == 3, f"Got {len(data_rows4)}")

        valid_signals = {"buy", "sell", "hold"}
        all_have_rationale = True
        signals_match = True
        details = []
        for r in data_rows4:
            if not (r and len(r) >= 3):
                continue
            sym = str(r[0]).strip().upper()
            sig = str(r[1]).strip().lower() if r[1] else ""
            if sig not in valid_signals:
                signals_match = False
                details.append(f"{sym} invalid signal {r[1]}")
            if not r[2]:
                all_have_rationale = False
            # Re-derive expected signal from the agent's own reported 30d return.
            if sym in overview:
                exp = expected_signal(overview[sym])
                if exp is not None and sig != exp:
                    signals_match = False
                    details.append(f"{sym}: got {sig}, expected {exp} (ret={overview[sym]})")
        check("All signals have rationale", all_have_rationale)
        check("Portfolio_Signals match BUY/SELL/HOLD rules re-derived from reported 30d returns",
              signals_match, "; ".join(details))


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute(
            "SELECT id FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%market analysis%'"
        )
        ss_ids = [r[0] for r in cur.fetchall()]
        check("Google Sheet 'Market Analysis Live Data' exists", len(ss_ids) >= 1,
              f"Found {len(ss_ids)}")

        # CRITICAL: the cloud sheet must actually contain the 3 stock rows (symbols).
        found_symbols = set()
        if ss_ids:
            cur.execute(
                "SELECT LOWER(value) FROM gsheet.cells WHERE spreadsheet_id = ANY(%s)",
                (ss_ids,),
            )
            cell_text = " ".join(r[0] for r in cur.fetchall() if r[0])
            for sym in SYMBOLS:
                base = sym.split(".")[0].lower()
                if sym.lower() in cell_text or base in cell_text:
                    found_symbols.add(sym)
        check("Cloud spreadsheet 'Market Analysis Live Data' contains the 3 stock rows",
              len(found_symbols) == 3, f"Found symbols: {found_symbols}")
        cur.close()
        conn.close()
    except Exception as e:
        check("Cloud spreadsheet 'Market Analysis Live Data' contains the 3 stock rows",
              False, str(e))


def check_teamly():
    print("\n=== Checking Teamly Knowledge Base ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly 'Investment Research Log': 3 stock pages, each naming its symbol and signal",
              False, str(e))
        return

    # Deliverable pages: titled with "Investment Research Log".
    log_pages = [(t, b) for t, b in rows
                 if t and "investment research log" in t.lower()]
    check("At least 3 teamly 'Investment Research Log' pages created",
          len(log_pages) >= 3, f"Found {len(log_pages)} (total pages {len(rows)})")

    # CRITICAL: exactly 3 pages, each (title+body) naming its symbol AND a valid signal.
    valid_signals = ("buy", "sell", "hold")
    matched = 0
    matched_syms = set()
    for t, b in log_pages:
        text = f"{t} {b}".lower()
        sym = None
        for s in SYMBOLS:
            base = s.split(".")[0].lower()
            if s.lower() in text or base in text:
                sym = s
                break
        if sym and any(sig in text for sig in valid_signals):
            matched += 1
            matched_syms.add(sym)
    check("Teamly 'Investment Research Log': 3 stock pages, each naming its symbol and signal",
          len(log_pages) == 3 and matched == 3 and matched_syms == set(SYMBOLS),
          f"pages={len(log_pages)}, matched={matched}, syms={matched_syms}")


def check_reverse_validation(workspace):
    """Verify things that should NOT exist in the output (non-critical)."""
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Market_Analysis_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        expected_keywords = {"stock", "overview", "economic", "indicator", "correlation",
                             "matrix", "portfolio", "signal"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0,
              f"Unexpected: {unexpected}")

        ws1 = get_sheet(wb, "Stock_Overview")
        if ws1:
            no_neg = True
            for row in ws1.iter_rows(min_row=2, values_only=True):
                if row and len(row) >= 4 and row[3] is not None:
                    price = to_float(row[3])
                    if price is not None and price < 0:
                        no_neg = False
            check("No negative stock prices", no_neg)

    # Teamly: no duplicate Investment Research Log pages beyond the 3 deliverables.
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM teamly.pages WHERE lower(title) LIKE '%investment research log%'"
        )
        n = cur.fetchone()[0]
        check("No duplicate Investment Research Log pages", n <= 3,
              f"Found {n} matching pages")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL MOEX FETCH GSHEET EXCEL TEAMLY - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_gsheet()
    check_teamly()
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "total_passed": PASS_COUNT,
                "total_checks": total,
                "accuracy": accuracy,
                "critical_failed": critical_failed,
                "success": (not critical_failed) and accuracy >= 70,
            }, f, indent=2)

    if critical_failed:
        print("Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("Overall: PASS")
        sys.exit(0)
    print("Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
