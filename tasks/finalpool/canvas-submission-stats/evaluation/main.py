"""Evaluation for canvas-submission-stats.

The agent must:
  1. Produce Canvas_Submissions.xlsx with a "Submission Stats" sheet (per-course
     Course_Code / Submissions / Avg_Score / Late_Count / Late_Pct) and a
     "Summary" sheet (Total_Submissions / Overall_Avg_Score / Total_Late /
     Overall_Late_Pct).
  2. Create a Teamly page titled "Canvas Submission Analysis" with the key
     findings (replaces the old, never-graded Notion deliverable).

The per-course numbers are checked against a committed groundtruth xlsx computed
from the canonical 22-course Canvas seed.

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
  - All groundtruth courses present in 'Submission Stats' -- the per-course
    deliverable must actually exist (otherwise the per-course value checks would
    pass vacuously and a non-doing agent could slip through).
  - Per-course Avg_Score correct (tol=2.0) -- the core weighted-mean deliverable.
  - Per-course Late_Pct correct (tol=1.0) -- the key late rule.
  - Summary Overall_Avg_Score / Overall_Late_Pct correct (weighted, not naive).
  - Summary Total_Submissions / Total_Late internally consistent with rows.
  - Teamly "Canvas Submission Analysis" page exists with a non-trivial body.
Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "All courses present in 'Submission Stats'",
    "Per-course Avg_Score correct for all courses (tol=2.0)",
    "Per-course Late_Pct correct for all courses (tol=1.0)",
    "Summary Overall_Avg_Score correct (tol=1.0)",
    "Summary Overall_Late_Pct correct (tol=1.0)",
    "Summary Total_Submissions internally consistent with rows",
    "Summary Total_Late internally consistent with rows",
    "Teamly 'Canvas Submission Analysis' page exists with body",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def num_close(a, b, tol):
    fa, fb = to_float(a), to_float(b)
    if fa is None and fb is None:
        return True
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= tol


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_to_lookup(rows):
    data = rows[1:] if rows and len(rows) > 1 else []
    lookup = {}
    for row in data:
        if row and row[0] is not None:
            lookup[str(row[0]).strip().lower()] = row
    return lookup


def check_xlsx(agent_file, gt_file):
    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- Submission Stats ----
    print("\n=== Checking 'Submission Stats' ===")
    a_rows = load_sheet_rows(agent_wb, "Submission Stats")
    g_rows = load_sheet_rows(gt_wb, "Submission Stats")
    record("Sheet 'Submission Stats' present", a_rows is not None,
           "sheet not found in agent output")

    if a_rows is None:
        # Fail the completeness CRITICAL closed: no rows => no per-course
        # deliverable, and the per-course value checks would pass vacuously.
        record("All courses present in 'Submission Stats'", False,
               "'Submission Stats' sheet/rows missing")

    sum_submissions = sum_late = 0
    weighted_score = 0.0
    if a_rows is not None and g_rows is not None:
        a_lookup = rows_to_lookup(a_rows)
        g_lookup = rows_to_lookup(g_rows)

        missing = [k for k in g_lookup if k not in a_lookup]
        record("All courses present in 'Submission Stats'", not missing,
               f"missing rows: {missing}")

        # Structural: row sorted by Course_Code (ascending).
        a_keys = [str(r[0]).strip() for r in a_rows[1:] if r and r[0] is not None]
        record("Rows sorted by Course_Code", a_keys == sorted(a_keys),
               f"agent order != sorted order")

        sub_ok = score_ok = late_ok = pct_ok = True
        sub_err = score_err = late_err = pct_err = []
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                continue
            # Submissions (tol relative-ish, conservative absolute)
            if len(a_row) > 1 and len(g_row) > 1:
                fa = to_float(a_row[1])
                if fa is not None:
                    sum_submissions += fa
                tol = max(5.0, 0.01 * (to_float(g_row[1]) or 0))
                if not num_close(a_row[1], g_row[1], tol):
                    sub_ok = False
                    sub_err.append(f"{key}: {a_row[1]} vs {g_row[1]}")
            # Avg_Score (CRITICAL)
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 2.0):
                    score_ok = False
                    score_err.append(f"{key}: {a_row[2]} vs {g_row[2]}")
            # Late_Count
            if len(a_row) > 3 and len(g_row) > 3:
                fl = to_float(a_row[3])
                if fl is not None:
                    sum_late += fl
                tol = max(5.0, 0.02 * (to_float(g_row[3]) or 0))
                if not num_close(a_row[3], g_row[3], tol):
                    late_ok = False
                    late_err.append(f"{key}: {a_row[3]} vs {g_row[3]}")
            # Late_Pct (CRITICAL)
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1.0):
                    pct_ok = False
                    pct_err.append(f"{key}: {a_row[4]} vs {g_row[4]}")
            # accumulate weighted score for summary consistency
            s = to_float(a_row[2]) if len(a_row) > 2 else None
            n = to_float(a_row[1]) if len(a_row) > 1 else None
            if s is not None and n is not None:
                weighted_score += s * n

        record("Per-course Submissions correct", sub_ok, f"{sub_err[:3]}")
        record("Per-course Avg_Score correct for all courses (tol=2.0)", score_ok, f"{score_err[:5]}")
        record("Per-course Late_Count correct", late_ok, f"{late_err[:3]}")
        record("Per-course Late_Pct correct for all courses (tol=1.0)", pct_ok, f"{pct_err[:5]}")

    # ---- Summary ----
    print("\n=== Checking 'Summary' ===")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    record("Sheet 'Summary' present", a_rows is not None, "sheet not found")

    if a_rows is not None and g_rows is not None:
        a_lookup = rows_to_lookup(a_rows)
        g_lookup = rows_to_lookup(g_rows)

        def gv(d, k):
            r = d.get(k.lower())
            return r[1] if r and len(r) > 1 else None

        # Total_Submissions vs groundtruth AND internal consistency with rows.
        a_ts = gv(a_lookup, "Total_Submissions")
        g_ts = gv(g_lookup, "Total_Submissions")
        record("Summary Total_Submissions matches groundtruth",
               num_close(a_ts, g_ts, 50.0), f"{a_ts} vs {g_ts}")
        # internal consistency: declared total ~ sum of per-course Submissions.
        # Fail closed: if the agent omitted the per-course rows (row-sum 0) the
        # declared total cannot be reconciled, so this must FAIL, not pass.
        record("Summary Total_Submissions internally consistent with rows",
               num_close(a_ts, sum_submissions,
                         max(5.0, 0.005 * (to_float(a_ts) or 1))),
               f"declared {a_ts} vs row-sum {sum_submissions}")

        a_tl = gv(a_lookup, "Total_Late")
        g_tl = gv(g_lookup, "Total_Late")
        record("Summary Total_Late matches groundtruth",
               num_close(a_tl, g_tl, 50.0), f"{a_tl} vs {g_tl}")
        # Fail closed (groundtruth Total_Late is large/non-zero): a 0 row-sum
        # means the per-course rows were omitted and cannot be reconciled.
        record("Summary Total_Late internally consistent with rows",
               num_close(a_tl, sum_late,
                         max(5.0, 0.01 * (to_float(a_tl) or 1))),
               f"declared {a_tl} vs row-sum {sum_late}")

        # Overall_Avg_Score must be weighted by submission count (CRITICAL).
        a_oas = gv(a_lookup, "Overall_Avg_Score")
        g_oas = gv(g_lookup, "Overall_Avg_Score")
        record("Summary Overall_Avg_Score correct (tol=1.0)",
               num_close(a_oas, g_oas, 1.0), f"{a_oas} vs {g_oas}")

        a_olp = gv(a_lookup, "Overall_Late_Pct")
        g_olp = gv(g_lookup, "Overall_Late_Pct")
        record("Summary Overall_Late_Pct correct (tol=1.0)",
               num_close(a_olp, g_olp, 1.0), f"{a_olp} vs {g_olp}")


def check_teamly():
    print("\n=== Checking Teamly 'Canvas Submission Analysis' page ===")
    try:
        conn = psycopg2.connect(**DB)
    except Exception as e:
        record("Teamly 'Canvas Submission Analysis' page exists with body", False,
               f"cannot connect to DB: {e}")
        return
    cur = conn.cursor()
    cur.execute("SELECT to_regclass('teamly.pages')")
    if cur.fetchone()[0] is None:
        record("Teamly 'Canvas Submission Analysis' page exists with body", False,
               "teamly.pages not found")
        conn.close()
        return

    # English title 'Canvas Submission Analysis' is preserved per task.md.
    cur.execute("""
        SELECT title, COALESCE(body, '')
        FROM teamly.pages
        WHERE title ILIKE '%canvas submission analysis%'
           OR title ILIKE '%submission analysis%'
    """)
    pages = cur.fetchall()
    if not pages:
        cur.execute("SELECT COUNT(*) FROM teamly.pages")
        total = cur.fetchone()[0]
        record("Teamly 'Canvas Submission Analysis' page exists with body", False,
               f"found {total} pages, none titled 'Canvas Submission Analysis'")
        conn.close()
        return

    body = "\n".join(str(b) for _, b in pages)
    record("Teamly 'Canvas Submission Analysis' page exists with body",
           len(body.strip()) >= 50, f"body is {len(body)} chars")
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Submissions.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Submissions.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    check_xlsx(agent_file, gt_file)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY: {PASS_COUNT} passed, {FAIL_COUNT} failed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and (accuracy >= 70)
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                       "accuracy": accuracy,
                       "critical_failed": critical_failed,
                       "success": success}, f)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
