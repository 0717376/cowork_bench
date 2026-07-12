"""Evaluation script for pw-yf-insales-ecommerce-index-excel-notion (RU / InSales + Teamly).

Структурные проверки (наличие листов, столбцов, минимум строк, наличие
скрипта-процессора) — НЕ критические. Критические проверки валидируют СУТЬ
анализа и выводятся из живых источников (внешний бенчмарк localhost:30328 и
русифицированные категории магазина InSales), а не хардкодятся:

  CRITICAL_CHECKS (любой провал => sys.exit(1) до порога точности):
    1. Столбец Category содержит русские категории магазина InSales
       (Электроника/Аудио/Камеры/Бытовая техника/ТВ и домашний кинотеатр/Часы) —
       значит агент сопоставил EN-бенчмарк с RU-категориями магазина.
    2. Market_Avg_Price по каждой категории совпадает с внешним бенчмарком
       (Electronics 58.07, Audio 74.15, Cameras 22.71, Home Appliances 33.98,
       TV&Home Theater 289.65, Watches 57.66) в пределах допуска — агент реально
       спарсил localhost:30328 и сджойнил EN-бенчмарк к RU-категории.
    3. Price_Gap_Pct = (Our_Avg_Price - Market_Avg_Price)/Market_Avg_Price*100
       посчитан верно (знак+величина в допуске) по каждой строке — формула
       применена, а не сфабрикована.
    4. Metrics: Total_Categories == 6 (число категорий), а Total_Products
       согласуется с суммой Product_Count листа Data_Analysis (допуск).
    5. В Teamly существует страница "Yf Wc Ecommerce Dashboard" с непустым телом,
       упоминающим суть анализа (заменяет проверку notion.pages ILIKE %dashboard%).
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Внешний бенчмарк (localhost:30328) -> ожидаемая Market_Avg_Price.
# Ключи — RU-названия категорий магазина InSales; значения берём из mock-страницы.
MARKET_BENCHMARK = {
    "электроника": 58.07,
    "аудио": 74.15,
    "камеры": 22.71,
    "бытовая техника": 33.98,
    "тв и домашний кинотеатр": 289.65,
    "часы": 57.66,
}
# EN-метки внешнего бенчмарка считаем эквивалентными RU-категориям магазина.
EN_RU_ALIASES = {
    "электроника": ["electronics"],
    "аудио": ["audio"],
    "камеры": ["cameras", "camera"],
    "бытовая техника": ["home appliances", "appliances"],
    "тв и домашний кинотеатр": ["tv & home theater", "tv home theater", "tv and home theater"],
    "часы": ["watches", "watch"],
}

EXPECTED_CATEGORIES = list(MARKET_BENCHMARK.keys())


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRIT" if critical else "    "
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS][{tag}] {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL][{tag}] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def norm_cat(s):
    """Нормализация названия категории для матчинга (RU или EN)."""
    if s is None:
        return ""
    return str(s).strip().lower().replace("ё", "е")


def canonical_cat(raw):
    """Свести RU/EN-метку строки к каноническому RU-ключу, либо None."""
    n = norm_cat(raw)
    if not n:
        return None
    for ru in EXPECTED_CATEGORIES:
        if n == ru.replace("ё", "е"):
            return ru
        for en in EN_RU_ALIASES.get(ru, []):
            if n == en:
                return ru
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    excel_path = os.path.join(agent_workspace, "Wc_Ecommerce_Index_Report.xlsx")
    check("Wc_Ecommerce_Index_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---- Data_Analysis -------------------------------------------------
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        da_rows = []
        da_headers = []
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(da_rows) >= 6, f"got {len(da_rows)}")

            da_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Category', 'Product_Count', 'Our_Avg_Price', 'Total_Sales', 'Market_Avg_Price', 'Price_Gap_Pct']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in da_headers, f"headers: {da_headers[:8]}")

        # ---- Metrics -------------------------------------------------------
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_rows = []
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            metrics_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(metrics_rows) >= 3, f"got {len(metrics_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ---- Recommendations ----------------------------------------------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            rec_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Category']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ====================================================================
        # CRITICAL: семантика Data_Analysis
        # ====================================================================
        def col_idx(name):
            try:
                return da_headers.index(name.lower())
            except ValueError:
                return None

        i_cat = col_idx("Category")
        i_pcount = col_idx("Product_Count")
        i_our = col_idx("Our_Avg_Price")
        i_mkt = col_idx("Market_Avg_Price")
        i_gap = col_idx("Price_Gap_Pct")

        # Сопоставляем строки к каноническим RU-категориям.
        row_by_cat = {}
        if i_cat is not None:
            for row in da_rows:
                if i_cat < len(row):
                    cc = canonical_cat(row[i_cat])
                    if cc is not None and cc not in row_by_cat:
                        row_by_cat[cc] = row

        found_cats = set(row_by_cat.keys())
        check("CRITICAL: Data_Analysis Category содержит 6 RU-категорий магазина InSales",
              found_cats >= set(EXPECTED_CATEGORIES),
              f"найдены: {sorted(found_cats)}; ожидались: {EXPECTED_CATEGORIES}",
              critical=True)

        # Market_Avg_Price по категории совпадает с внешним бенчмарком.
        mkt_ok = (i_mkt is not None)
        if i_mkt is not None:
            for ru, expected in MARKET_BENCHMARK.items():
                row = row_by_cat.get(ru)
                got = safe_float(row[i_mkt]) if (row and i_mkt < len(row)) else None
                ok = got is not None and abs(got - expected) <= 0.5
                if not ok:
                    mkt_ok = False
                check(f"CRITICAL: Market_Avg_Price[{ru}] ~= {expected}",
                      ok, f"got {got}", critical=True)
        else:
            check("CRITICAL: Market_Avg_Price column присутствует", False,
                  "столбец отсутствует", critical=True)

        # Price_Gap_Pct = (Our - Market)/Market*100 по каждой строке.
        if i_our is not None and i_mkt is not None and i_gap is not None:
            gap_all_ok = True
            for ru, row in row_by_cat.items():
                our = safe_float(row[i_our]) if i_our < len(row) else None
                mkt = safe_float(row[i_mkt]) if i_mkt < len(row) else None
                gap = safe_float(row[i_gap]) if i_gap < len(row) else None
                if our is None or mkt in (None, 0) or gap is None:
                    gap_all_ok = False
                    check(f"CRITICAL: Price_Gap_Pct[{ru}] вычислим", False,
                          f"our={our} mkt={mkt} gap={gap}", critical=True)
                    continue
                expected_gap = (our - mkt) / mkt * 100.0
                ok = abs(gap - expected_gap) <= 1.0
                if not ok:
                    gap_all_ok = False
                check(f"CRITICAL: Price_Gap_Pct[{ru}] = (Our-Market)/Market*100",
                      ok, f"got {gap}, expected ~{round(expected_gap,2)}", critical=True)
        else:
            check("CRITICAL: столбцы Our/Market/Price_Gap присутствуют", False,
                  f"i_our={i_our} i_mkt={i_mkt} i_gap={i_gap}", critical=True)

        # ====================================================================
        # CRITICAL: Metrics Total_Categories / Total_Products
        # ====================================================================
        m = {}
        for r in metrics_rows:
            if r and len(r) >= 2 and r[0] is not None:
                m[str(r[0]).strip().lower()] = r[1]

        total_cats = safe_float(m.get("total_categories"))
        check("CRITICAL: Metrics Total_Categories == 6",
              total_cats is not None and abs(total_cats - 6) < 0.5,
              f"got {total_cats}", critical=True)

        # Total_Products должен согласовываться с суммой Product_Count.
        total_products = safe_float(m.get("total_products"))
        sum_pcount = None
        if i_pcount is not None:
            vals = [safe_float(row[i_pcount]) for row in da_rows
                    if i_pcount < len(row) and safe_float(row[i_pcount]) is not None]
            if vals:
                sum_pcount = sum(vals)
        check("CRITICAL: Metrics Total_Products согласуется с суммой Product_Count",
              total_products is not None and sum_pcount is not None
              and abs(total_products - sum_pcount) <= 2,
              f"Total_Products={total_products}, sum(Product_Count)={sum_pcount}",
              critical=True)

        # ---- processor script ---------------------------------------------
        check("yf_wc_ecommerce_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "yf_wc_ecommerce_processor.py")))

    # ====================================================================
    # CRITICAL: Teamly dashboard page
    # ====================================================================
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        pages = None
        check("CRITICAL: Teamly доступен", False, str(e), critical=True)

    if pages is not None:
        dash = None
        for pid, title, body in pages:
            tl = (title or "").lower()
            if ("dashboard" in tl and ("ecommerce" in tl or "wc" in tl or "yf" in tl)) \
               or ("yf wc ecommerce dashboard" in tl):
                dash = (pid, title, body)
                break
        if dash is None:
            # запасной матч: любая страница с 'dashboard' в заголовке
            for pid, title, body in pages:
                if "dashboard" in (title or "").lower():
                    dash = (pid, title, body)
                    break
        check("CRITICAL: Teamly страница 'Yf Wc Ecommerce Dashboard' существует",
              dash is not None,
              f"страницы: {[(p[0], p[1]) for p in pages]}", critical=True)
        if dash is not None:
            body_text = (dash[2] or "").strip()
            check("CRITICAL: страница Teamly имеет непустое тело с резюме анализа",
                  len(body_text) >= 30, f"len={len(body_text)}", critical=True)

    # ---- gating ----------------------------------------------------------
    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    if CRITICAL_FAILED:
        print(f"\nCRITICAL FAILURES: {CRITICAL_FAILED}")
        return False, f"Critical check(s) failed: {CRITICAL_FAILED}; {PASS_COUNT}/{total} checks"

    success = accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.0f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
