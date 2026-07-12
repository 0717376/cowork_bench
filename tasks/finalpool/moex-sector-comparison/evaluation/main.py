"""Evaluation for yf-sector-comparison (MOEX)."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "YF_Sector_Comparison.xlsx")
    gt_file = os.path.join(gt_dir, "YF_Sector_Comparison.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []          # non-critical (structural) issues -> lower accuracy
    critical_failures = []   # any of these -> immediate FAIL
    checks_total = 0
    checks_passed = 0

    def record(ok, label):
        nonlocal checks_total, checks_passed
        checks_total += 1
        if ok:
            checks_passed += 1
        else:
            all_errors.append(label)
        return ok

    # ----- Sheet: Sector Comparison -----
    print("  Checking Sector Comparison...")
    a_rows = load_sheet_rows(agent_wb, "Sector Comparison")
    g_rows = load_sheet_rows(gt_wb, "Sector Comparison")
    if a_rows is None:
        critical_failures.append("Sheet 'Sector Comparison' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Sector Comparison' not found in groundtruth")
    else:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]

        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}

        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                # CRITICAL: a groundtruth sector is missing entirely
                critical_failures.append(f"Sector Comparison missing row: {g_row[0]}")
                continue

            # CRITICAL: Stock_Count must match EXACTLY (real grouping)
            if not (len(a_row) > 1 and len(g_row) > 1 and num_close(a_row[1], g_row[1], 0)):
                av = a_row[1] if len(a_row) > 1 else None
                critical_failures.append(
                    f"{g_row[0]}.Stock_Count: {av} vs {g_row[1]} (must match exactly)"
                )

            # CRITICAL: Avg_Market_Cap_B within tolerance (RUB billions; loose tol)
            if not (len(a_row) > 2 and len(g_row) > 2 and num_close(a_row[2], g_row[2], 50.0)):
                av = a_row[2] if len(a_row) > 2 else None
                critical_failures.append(
                    f"{g_row[0]}.Avg_Market_Cap_B: {av} vs {g_row[2]} (tol=50.0)"
                )

            # CRITICAL: Avg_PE_Ratio present, numeric and within tolerance
            gpe = g_row[3] if len(g_row) > 3 else None
            ape = a_row[3] if len(a_row) > 3 else None
            if gpe is not None:
                ok_pe = False
                try:
                    ok_pe = ape is not None and abs(float(ape) - float(gpe)) <= 0.5
                except (TypeError, ValueError):
                    ok_pe = False
                if not ok_pe:
                    critical_failures.append(
                        f"{g_row[0]}.Avg_PE_Ratio: {ape} vs {gpe} (tol=0.5, must be numeric)"
                    )

        # CRITICAL: rows sorted by Avg_Market_Cap_B descending
        try:
            caps = [float(r[2]) for r in a_data if len(r) > 2 and r[2] is not None]
            if caps != sorted(caps, reverse=True):
                critical_failures.append(
                    "Sector Comparison not sorted by Avg_Market_Cap_B descending"
                )
            else:
                record(True, "")
        except (TypeError, ValueError):
            critical_failures.append("Avg_Market_Cap_B column not numeric (cannot verify sort)")

    # ----- Sheet: Summary -----
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        critical_failures.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = [r for r in a_rows[1:] if r and r[0] is not None]
        g_data = [r for r in g_rows[1:] if r and r[0] is not None]
        a_lookup = {str(r[0]).strip().lower(): r for r in a_data}

        for g_row in g_data:
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                critical_failures.append(f"Summary missing row: {g_row[0]}")
                continue
            a_val = a_row[1] if len(a_row) > 1 else None
            g_val = g_row[1] if len(g_row) > 1 else None

            if key in ("total_sectors", "total_stocks"):
                # CRITICAL: exact counts (tightened from tol=5.0)
                if not num_close(a_val, g_val, 0):
                    critical_failures.append(
                        f"{g_row[0]}.Value: {a_val} vs {g_val} (must match exactly)"
                    )
            elif key == "largest_sector":
                # CRITICAL: sector with max Avg_Market_Cap_B
                if not str_match(a_val, g_val):
                    critical_failures.append(
                        f"Largest_Sector: {a_val} vs {g_val}"
                    )
            else:
                if not num_close(a_val, g_val, 5.0):
                    all_errors.append(f"{g_row[0]}.Value: {a_val} vs {g_val} (tol=5.0)")

    # ----- Sector_Analysis.docx (RU prose) -----
    docx_path = os.path.join(args.agent_workspace, "Sector_Analysis.docx")
    if not os.path.exists(docx_path):
        critical_failures.append("Sector_Analysis.docx not found")
    else:
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs)
            _low = _text.lower()

            # CRITICAL: substantive prose
            if len(_text.strip()) < 200:
                critical_failures.append(
                    f"Sector_Analysis.docx too short ({len(_text.strip())} chars, need >=200)"
                )
            # CRITICAL: mentions at least one real sector name (EN identifiers kept literal)
            sector_names = [
                "energy", "financial services", "communication services",
                "consumer defensive", "unknown",
            ]
            if not any(s in _low for s in sector_names):
                critical_failures.append(
                    "Sector_Analysis.docx does not mention any real sector name"
                )
            # CRITICAL: valuation / market-cap observation (RU+EN keywords)
            val_kws = [
                "оценк", "капитализац", "p/e", "p\\e", "pe", "мультипликат",
                "valuation", "market cap", "market-cap", "прибыл",
            ]
            if not any(k in _low for k in val_kws):
                critical_failures.append(
                    "Sector_Analysis.docx lacks a valuation / market-cap observation"
                )
            # NON-critical: topical keywords (RU + EN)
            topic_kws = ["сектор", "отрасл", "sector", "анализ", "analysis"]
            record(any(k in _low for k in topic_kws),
                   "Sector_Analysis.docx missing topical keyword (сектор/отрасл/анализ)")
        except ImportError:
            if os.path.getsize(docx_path) < 100:
                critical_failures.append("Sector_Analysis.docx too small")
        except Exception as _e:
            all_errors.append(f"Error reading Sector_Analysis.docx: {_e}")

    # ----- CRITICAL gate -----
    if critical_failures:
        print(f"\n=== CRITICAL FAIL ({len(critical_failures)}) ===")
        for e in critical_failures[:15]:
            print(f"  CRITICAL: {e}")
        sys.exit(1)

    # ----- Accuracy gate (>=70) -----
    accuracy = 100.0 if checks_total == 0 else (checks_passed / checks_total) * 100.0
    print(f"\nNon-critical accuracy: {accuracy:.1f}% ({checks_passed}/{checks_total})")
    for e in all_errors[:10]:
        print(f"  {e}")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
