"""Evaluation for insales-tax-compliance-excel-gcal-gform (russified, InSales + forms).

Структура проверки:
- Накапливаем обычные ошибки (структурные / мягкие пороги) -> влияют на accuracy.
- Накапливаем КРИТИЧЕСКИЕ ошибки (CRITICAL_CHECKS: верность ключевых значений из
  замороженного эталона). Любой критический провал => немедленный FAIL (sys.exit(1))
  до проверки порога точности.
- Порог: accuracy >= 70 И отсутствие критических провалов => PASS.

ВАЖНО: groundtruth_workspace/Tax_Compliance_Report.xlsx заморожен и map-патчен
(scripts/wc_patch_groundtruth.py). Эталонные значения читаются из него, а не
хардкодятся, поэтому реалии wc.* здесь не дублируются.
"""
import os
import argparse, os, sys
import psycopg2

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)



DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def safe_float(v, default=None):
    if v is None:
        return default
    try:
        return float(v)
    except (ValueError, TypeError):
        return default


def norm_yes_no(v):
    """Нормализует флаг Yes/No (англ.) и RU-эквиваленты Да/Нет -> 'yes'/'no'."""
    s = str(v).strip().lower() if v is not None else ""
    if s in ("yes", "y", "да", "true", "1"):
        return "yes"
    if s in ("no", "n", "нет", "false", "0"):
        return "no"
    return s


def check_excel(agent_workspace, groundtruth_workspace):
    """Возвращает (errors, critical_errors)."""
    errors = []
    critical = []
    import openpyxl

    agent_path = os.path.join(agent_workspace, "Tax_Compliance_Report.xlsx")
    if not os.path.exists(agent_path):
        return (["Tax_Compliance_Report.xlsx not found"], ["Tax_Compliance_Report.xlsx not found"])

    gt_path = os.path.join(groundtruth_workspace, "Tax_Compliance_Report.xlsx")
    if not os.path.exists(gt_path):
        return (["Groundtruth Tax_Compliance_Report.xlsx not found"], [])

    try:
        wb_agent = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)

        # --- Sheet 1: Order Tax Audit ---
        agent_rows = load_sheet_rows(wb_agent, "Order Tax Audit")
        gt_rows = load_sheet_rows(wb_gt, "Order Tax Audit")
        if agent_rows is None:
            errors.append("Sheet 'Order Tax Audit' not found")
            critical.append("Sheet 'Order Tax Audit' not found")
        elif gt_rows is None:
            errors.append("Groundtruth sheet 'Order Tax Audit' not found")
        else:
            agent_data = [r for r in agent_rows[1:] if r and r[0] is not None]
            gt_data = [r for r in gt_rows[1:] if r and r[0] is not None]
            if len(agent_data) != len(gt_data):
                errors.append(f"Order Tax Audit: {len(agent_data)} rows, expected {len(gt_data)}")
                critical.append(f"Order Tax Audit: {len(agent_data)} rows, expected {len(gt_data)}")
            else:
                # Build lookup by Order_ID
                gt_lookup = {}
                for r in gt_data:
                    oid = int(r[0]) if r[0] else None
                    if oid:
                        gt_lookup[oid] = r

                # Мягкий счётчик строковых несоответствий (как раньше)
                mismatches = 0
                # КРИТИЧЕСКИЙ счётчик: верность Expected_Tax + Status по Order_ID
                core_mismatches = 0
                for r in agent_data:
                    oid = int(r[0]) if r[0] else None
                    if oid not in gt_lookup:
                        mismatches += 1
                        core_mismatches += 1
                        continue
                    gt_r = gt_lookup[oid]
                    row_bad = False
                    # Check Order_Total (col 1), tolerance 0.5
                    a_total = safe_float(r[1])
                    g_total = safe_float(gt_r[1])
                    if a_total is not None and g_total is not None and abs(a_total - g_total) > 0.5:
                        mismatches += 1
                        continue
                    # Check Applicable_Rate (col 3), tolerance 0.001
                    a_rate = safe_float(r[3])
                    g_rate = safe_float(gt_r[3])
                    if a_rate is not None and g_rate is not None and abs(a_rate - g_rate) > 0.001:
                        mismatches += 1
                        continue
                    # Check Expected_Tax (col 4), tolerance 0.5 (CRITICAL-relevant)
                    a_exp = safe_float(r[4])
                    g_exp = safe_float(gt_r[4])
                    if a_exp is not None and g_exp is not None and abs(a_exp - g_exp) > 0.5:
                        mismatches += 1
                        row_bad = True
                    # Check Status (col 7) (CRITICAL-relevant)
                    a_status = str(r[7]).strip().lower() if r[7] else ""
                    g_status = str(gt_r[7]).strip().lower() if gt_r[7] else ""
                    if a_status != g_status:
                        mismatches += 1
                        row_bad = True
                    if row_bad:
                        core_mismatches += 1

                if mismatches > 5:
                    errors.append(f"Order Tax Audit: {mismatches} row mismatches (>5 threshold)")
                # CRITICAL: Expected_Tax + Status должны совпадать почти для всех заказов
                if core_mismatches > 1:
                    critical.append(
                        f"CRITICAL: Order Tax Audit Expected_Tax/Status неверны для {core_mismatches} заказов (>1)")

        # --- Sheet 2: State Summary ---
        agent_ss = load_sheet_rows(wb_agent, "State Summary")
        gt_ss = load_sheet_rows(wb_gt, "State Summary")
        if agent_ss is None:
            errors.append("Sheet 'State Summary' not found")
            critical.append("Sheet 'State Summary' not found")
        elif gt_ss is None:
            errors.append("Groundtruth sheet 'State Summary' not found")
        else:
            agent_ss_data = [r for r in agent_ss[1:] if r and r[0] is not None]
            gt_ss_data = [r for r in gt_ss[1:] if r and r[0] is not None]
            if abs(len(agent_ss_data) - len(gt_ss_data)) > 2:
                errors.append(f"State Summary: {len(agent_ss_data)} rows, expected ~{len(gt_ss_data)}")
            gt_state_lookup = {str(r[0]).strip().upper(): r for r in gt_ss_data}
            agent_state_lookup = {str(r[0]).strip().upper(): r for r in agent_ss_data}
            ss_mismatches = 0
            for r in agent_ss_data:
                state = str(r[0]).strip().upper() if r[0] else ""
                if state not in gt_state_lookup:
                    ss_mismatches += 1
                    continue
                gt_r = gt_state_lookup[state]
                # Check Order_Count (col 1)
                a_count = safe_float(r[1])
                g_count = safe_float(gt_r[1])
                if a_count is not None and g_count is not None and abs(a_count - g_count) > 0:
                    ss_mismatches += 1
                    continue
                # Check compliance rate (col 6), tolerance 5
                a_comp = safe_float(r[6])
                g_comp = safe_float(gt_r[6])
                if a_comp is not None and g_comp is not None and abs(a_comp - g_comp) > 5.0:
                    ss_mismatches += 1
            if ss_mismatches > 3:
                errors.append(f"State Summary: {ss_mismatches} state mismatches (>3 threshold)")

            # CRITICAL: Order_Count и Requires_Filing (>$100 actual tax) корректны
            # для каждого региона из эталона.
            REQ_COL = 7  # Requires_Filing
            state_core_bad = 0
            for st, gt_r in gt_state_lookup.items():
                a_r = agent_state_lookup.get(st)
                if a_r is None:
                    state_core_bad += 1
                    continue
                g_count = safe_float(gt_r[1])
                a_count = safe_float(a_r[1])
                if g_count is not None and a_count is not None and abs(a_count - g_count) > 0:
                    state_core_bad += 1
                    continue
                # Requires_Filing flag (если в эталоне столбец присутствует)
                if len(gt_r) > REQ_COL and len(a_r) > REQ_COL:
                    g_req = norm_yes_no(gt_r[REQ_COL])
                    a_req = norm_yes_no(a_r[REQ_COL])
                    if g_req in ("yes", "no") and a_req != g_req:
                        state_core_bad += 1
            if state_core_bad > 1:
                critical.append(
                    f"CRITICAL: State Summary Order_Count/Requires_Filing неверны для {state_core_bad} регионов (>1)")

        # --- Sheet 3: Compliance Overview ---
        agent_co = load_sheet_rows(wb_agent, "Compliance Overview")
        gt_co = load_sheet_rows(wb_gt, "Compliance Overview")
        if agent_co is None:
            errors.append("Sheet 'Compliance Overview' not found")
            critical.append("Sheet 'Compliance Overview' not found")
        elif gt_co is None:
            errors.append("Groundtruth sheet 'Compliance Overview' not found")
        else:
            agent_co_data = {str(r[0]).strip().lower(): r[1] for r in agent_co[1:] if r and r[0]}
            gt_co_data = {str(r[0]).strip().lower(): r[1] for r in gt_co[1:] if r and r[0]}

            # CRITICAL: total orders должно совпадать точно (без допуска)
            a_total = safe_float(agent_co_data.get("total_orders_audited"))
            g_total = safe_float(gt_co_data.get("total_orders_audited"))
            if a_total is None:
                errors.append("Compliance Overview: Total_Orders_Audited отсутствует")
                critical.append("CRITICAL: Total_Orders_Audited отсутствует")
            elif g_total is not None and abs(a_total - g_total) > 0:
                errors.append(f"Total_Orders_Audited: {a_total}, expected {g_total}")
                critical.append(f"CRITICAL: Total_Orders_Audited {a_total}, expected {g_total}")

            # Check compliant orders (tolerance 5) — мягкая
            a_comp = safe_float(agent_co_data.get("compliant_orders"))
            g_comp = safe_float(gt_co_data.get("compliant_orders"))
            if a_comp is not None and g_comp is not None and abs(a_comp - g_comp) > 5:
                errors.append(f"Compliant_Orders: {a_comp}, expected {g_comp}")

            # CRITICAL: overall compliance rate в пределах 1.0 от эталона
            a_rate = safe_float(agent_co_data.get("overall_compliance_rate"))
            g_rate = safe_float(gt_co_data.get("overall_compliance_rate"))
            if a_rate is None:
                errors.append("Compliance Overview: Overall_Compliance_Rate отсутствует")
                critical.append("CRITICAL: Overall_Compliance_Rate отсутствует")
            elif g_rate is not None and abs(a_rate - g_rate) > 1.0:
                errors.append(f"Overall_Compliance_Rate: {a_rate}, expected {g_rate}")
                critical.append(f"CRITICAL: Overall_Compliance_Rate {a_rate}, expected {g_rate} (допуск 1.0)")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
        critical.append(f"CRITICAL: Error reading Excel: {e}")
    return (errors, critical)


def check_gcal():
    """Возвращает (errors, critical_errors)."""
    errors = []
    critical = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Расширено на RU+EN: tax filing / tax deadline / налог + (декларац/отчётност/срок/подач).
        cur.execute("""
            SELECT summary, start_datetime::date FROM gcal.events
            WHERE summary ILIKE '%tax filing%'
               OR summary ILIKE '%tax deadline%'
               OR (summary ILIKE '%налог%'
                   AND (summary ILIKE '%декларац%' OR summary ILIKE '%отчётност%'
                        OR summary ILIKE '%отчетност%' OR summary ILIKE '%срок%'
                        OR summary ILIKE '%подач%'))
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        if len(rows) < 4:
            errors.append(f"Expected 4 tax filing deadline events in GCal, found {len(rows)}")
            critical.append(f"CRITICAL: найдено {len(rows)} ежеквартальных событий о налоговой отчётности, ожидалось 4")
        else:
            # Check that Q1-Q4 labels are present
            summaries = " ".join(str(r[0]) for r in rows).lower()
            missing_q = [q for q in ["q1", "q2", "q3", "q4"] if q not in summaries]
            for q in missing_q:
                errors.append(f"Missing '{q}' in calendar event summaries")
            if missing_q:
                critical.append(f"CRITICAL: в названиях событий отсутствуют маркеры {missing_q}")

            # Check dates (2026)
            dates = [r[1] for r in rows]
            from datetime import date
            expected_dates = [date(2026, 4, 15), date(2026, 7, 15), date(2026, 10, 15), date(2026, 1, 15)]
            missing_dates = []
            for ed in expected_dates:
                if ed not in dates:
                    # Allow +/- 1 day tolerance
                    close = any(abs((d - ed).days) <= 1 for d in dates)
                    if not close:
                        missing_dates.append(ed)
            for ed in missing_dates:
                errors.append(f"Missing calendar event for date {ed}")
            if missing_dates:
                critical.append(f"CRITICAL: отсутствуют события на даты {missing_dates}")

    except Exception as e:
        errors.append(f"Error checking GCal: {e}")
        critical.append(f"CRITICAL: Error checking GCal: {e}")
    return (errors, critical)


def check_gform():
    """Возвращает (errors, critical_errors)."""
    errors = []
    critical = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Название формы: RU+EN. EN-маркер 'Vendor Tax Information' фиксирован в task.md.
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%vendor%'
               OR title ILIKE '%tax information%'
               OR title ILIKE '%поставщик%'
               OR (title ILIKE '%налог%' AND title ILIKE '%сведен%')
               OR (title ILIKE '%налог%' AND title ILIKE '%информац%')
            ORDER BY created_at DESC LIMIT 5
        """)
        forms = cur.fetchall()

        if not forms:
            cur.close()
            conn.close()
            return (["No GForm found matching vendor tax form"],
                    ["CRITICAL: не найдена форма для налоговых сведений поставщиков"])

        form_id = forms[0][0]

        cur.execute("""
            SELECT title, question_type FROM gform.questions
            WHERE form_id = %s ORDER BY position
        """, (form_id,))
        questions = cur.fetchall()
        cur.close()
        conn.close()

        if len(questions) < 5:
            errors.append(f"Vendor tax form has {len(questions)} questions, expected 5")
            critical.append(f"CRITICAL: в форме {len(questions)} вопросов, ожидалось 5")

        # Ключевые темы вопросов: каждая группа = список RU+EN синонимов (любой подходит).
        q_titles = " ".join(str(q[0]) for q in questions).lower()
        q_titles_nospace = q_titles.replace(" ", "")
        topic_groups = {
            "vendor name": ["vendor", "поставщик", "наименование", "название"],
            "tax id": ["tax id", "taxid", "инн", "налоговый идентиф", "налоговый номер", "идентификационный"],
            "state of registration": ["state", "регион", "штат", "регистрац"],
            "tax-exempt": ["exempt", "освобожд"],
        }
        for topic, kws in topic_groups.items():
            hit = any((kw in q_titles) or (kw.replace(" ", "") in q_titles_nospace) for kw in kws)
            if not hit:
                errors.append(f"Missing question about '{topic}' in vendor form")
                critical.append(f"CRITICAL: в форме нет вопроса про '{topic}'")

    except Exception as e:
        errors.append(f"Error checking GForm: {e}")
        critical.append(f"CRITICAL: Error checking GForm: {e}")
    return (errors, critical)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []
    critical_errors = []
    # Счётчик «секций» для грубой оценки accuracy: каждая проверка либо чистая, либо нет.
    section_results = []

    print("  Checking Excel file...")
    errs, crit = check_excel(agent_ws, gt_ws)
    all_errors.extend(errs)
    critical_errors.extend(crit)
    section_results.append(("excel", not errs))
    if errs:
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GCal events...")
    errs, crit = check_gcal()
    all_errors.extend(errs)
    critical_errors.extend(crit)
    section_results.append(("gcal", not errs))
    if errs:
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking GForm...")
    errs, crit = check_gform()
    all_errors.extend(errs)
    critical_errors.extend(crit)
    section_results.append(("gform", not errs))
    if errs:
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    # --- Критические проверки: любой провал => немедленный FAIL до порога точности ---
    if critical_errors:
        print(f"\n=== CRITICAL FAILURE ({len(critical_errors)}) ===")
        for e in critical_errors[:10]:
            print(f"  {e}")
        print("=== RESULT: FAIL ===")
        sys.exit(1)

    # --- Порог точности: доля чистых секций >= 70% ---
    passed = sum(1 for _, ok in section_results if ok)
    accuracy = 100.0 * passed / len(section_results)
    print(f"\nAccuracy: {accuracy:.1f}% ({passed}/{len(section_results)} sections clean)")

    if accuracy >= 70 and not critical_errors:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
