"""Evaluation script for arxiv-research-pipeline-teamly-excel.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Injected papers the agent must catalog (English identifiers preserved).
EXPECTED_PAPERS = {
    "2301.00001": "LLM Reasoning Survey",
    "2302.00002": "Prompt Engineering Guide",
    "2303.00003": "In-Context Learning Theory",
}

# Priority enum, English + Russian equivalents accepted.
PRIORITY_VALUES = {
    "critical", "important", "nice-to-have", "nice to have",
    "критический", "критичный", "важный", "желательный", "необязательный",
}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Research_Knowledge_Base.xlsx exists",
    "Paper_Catalog contains all 3 injected papers by ID+Title",
    "Paper_Catalog sorted by Citation_Count descending",
    "Teamly LLM Research Hub page exists (not the noise page)",
    "Teamly page covers all 4 required content sections",
    "Research_Gaps has >= 4 rows with valid Priority",
    "research_synthesis.json exists and is valid JSON with paper IDs",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_sheet(wb, name):
    target = name.strip().lower().replace(" ", "_")
    for n in wb.sheetnames:
        if n.strip().lower().replace(" ", "_") == target:
            return wb[n]
    return None


def check_paper_catalog(wb):
    """Critical: all 3 injected papers present + sorted by Citation_Count desc."""
    ws = get_sheet(wb, "Paper_Catalog")
    if ws is None:
        check("Paper_Catalog contains all 3 injected papers by ID+Title", False, "no sheet")
        check("Paper_Catalog sorted by Citation_Count descending", False, "no sheet")
        return
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    hmap = {h: i for i, h in enumerate(headers)}
    id_i = hmap.get("paper_id")
    title_i = hmap.get("title")
    cit_i = hmap.get("citation_count")
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

    all_present = True
    if id_i is None or title_i is None:
        all_present = False
    else:
        for pid, ptitle in EXPECTED_PAPERS.items():
            found = False
            for r in rows:
                rid = str(r[id_i]).strip() if id_i < len(r) and r[id_i] is not None else ""
                rtitle = str(r[title_i]).strip().lower() if title_i < len(r) and r[title_i] is not None else ""
                if pid in rid and ptitle.lower() in rtitle:
                    found = True
                    break
            if not found:
                all_present = False
    check("Paper_Catalog contains all 3 injected papers by ID+Title", all_present,
          f"rows: {len(rows)}")

    # Sorted by Citation_Count descending.
    if cit_i is not None:
        cits = [safe_float(r[cit_i]) for r in rows if cit_i < len(r)]
        cits = [c for c in cits if c is not None]
        sorted_ok = all(cits[i] >= cits[i + 1] for i in range(len(cits) - 1)) and len(cits) >= 3
        check("Paper_Catalog sorted by Citation_Count descending", sorted_ok, f"cits: {cits}")
    else:
        check("Paper_Catalog sorted by Citation_Count descending", False, "no Citation_Count col")


def check_research_gaps(wb):
    """Critical: >= 4 data rows with a valid Priority value (RU/EN)."""
    ws = get_sheet(wb, "Research_Gaps")
    if ws is None:
        check("Research_Gaps has >= 4 rows with valid Priority", False, "no sheet")
        return
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    prio_i = {h: i for i, h in enumerate(headers)}.get("priority")
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
    if prio_i is None:
        check("Research_Gaps has >= 4 rows with valid Priority", False, "no Priority col")
        return
    valid = 0
    for r in rows:
        if prio_i < len(r) and r[prio_i] is not None:
            if str(r[prio_i]).strip().lower() in PRIORITY_VALUES:
                valid += 1
    check("Research_Gaps has >= 4 rows with valid Priority", valid >= 4,
          f"valid priority rows: {valid}")


def check_excel(agent_workspace, groundtruth_workspace):
    excel_path = os.path.join(agent_workspace, "Research_Knowledge_Base.xlsx")
    check("Research_Knowledge_Base.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        # Mark dependent critical checks as failed.
        check("Paper_Catalog contains all 3 injected papers by ID+Title", False, "no excel")
        check("Paper_Catalog sorted by Citation_Count descending", False, "no excel")
        check("Research_Gaps has >= 4 rows with valid Priority", False, "no excel")
        return

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    gt_path = os.path.join(groundtruth_workspace, "Research_Knowledge_Base.xlsx")
    gt_wb = openpyxl.load_workbook(gt_path) if os.path.exists(gt_path) else None

    if gt_wb:
        for sheet_name in gt_wb.sheetnames:
            ws = get_sheet(wb, sheet_name)
            check(f"{sheet_name} sheet exists", ws is not None)
            if ws is not None:
                gt_ws = gt_wb[sheet_name]
                gt_headers = [str(c.value).strip().lower() if c.value else "" for c in gt_ws[1]]
                headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
                for h in gt_headers:
                    if h:
                        check(f"{sheet_name} has {h} column", h in headers, f"headers: {headers[:10]}")
                gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                             if any(c is not None for c in r)]
                min_rows = max(1, len(gt_rows) - 2)
                check(f"{sheet_name} has >= {min_rows} data rows", len(data_rows) >= min_rows,
                      f"got {len(data_rows)}")

    # Critical content checks.
    check_paper_catalog(wb)
    check_research_gaps(wb)


def check_synthesis(agent_workspace):
    """Critical: research_synthesis.json exists, is valid JSON, references paper IDs."""
    path = os.path.join(agent_workspace, "research_synthesis.json")
    if not os.path.exists(path):
        check("research_synthesis.json exists and is valid JSON with paper IDs", False, "missing")
        return
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        check("research_synthesis.json exists and is valid JSON with paper IDs", False, str(e))
        return
    text = json.dumps(data, ensure_ascii=False)
    has_ids = sum(1 for pid in EXPECTED_PAPERS if pid in text) >= 2
    check("research_synthesis.json exists and is valid JSON with paper IDs", has_ids,
          f"ids found: {[p for p in EXPECTED_PAPERS if p in text]}")


def check_script(agent_workspace):
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")


def check_teamly():
    """Critical: an LLM Research Hub page exists with all 4 content sections.

    Seed pages have id <= 3; the noise page ('Архив протоколов совещаний') is a
    user-leftover and must NOT satisfy the hub check.
    """
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly LLM Research Hub page exists (not the noise page)", False, str(e))
        check("Teamly page covers all 4 required content sections", False, str(e))
        return

    # Identify the hub page: title carries the English marker, and it is not the
    # noise page ('архив протоколов').
    hub = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "архив протоколов" in tl:
            continue
        if "llm research hub" in tl or ("research" in tl and "hub" in tl) \
                or ("llm" in tl and ("исследован" in tl or "research" in tl)):
            hub = (pid, title, body)
            break
    check("Teamly LLM Research Hub page exists (not the noise page)", hub is not None,
          f"new pages: {[(p[0], p[1]) for p in pages]}")

    # Content sections: landscape overview, key papers, methodology comparison,
    # research gaps — keyword presence in original-case-lowered text (RU or EN).
    if hub is None:
        check("Teamly page covers all 4 required content sections", False, "no hub page")
        return
    text = ((hub[1] or "") + " " + (hub[2] or "")).lower()
    sections = [
        ("landscape overview", ["landscape", "overview", "ландшафт", "обзор"]),
        ("key papers", ["key paper", "paper", "статьи", "статей", "ключев"]),
        ("methodology comparison", ["methodolog", "method", "comparison", "методолог", "метод", "сравнен"]),
        ("research gaps", ["gap", "future work", "пробел", "будущ", "рекомендац"]),
    ]
    present = sum(1 for _, kws in sections if any(k in text for k in kws))
    check("Teamly page covers all 4 required content sections", present >= 4,
          f"sections found: {present}/4")
    # Dashboard heading marker (non-critical).
    check("Teamly page includes the dashboard heading marker",
          "large language model research dashboard" in text or "dashboard" in text
          or "дашборд" in text or "панель" in text,
          "heading marker absent")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    check_excel(agent_workspace, groundtruth_workspace)
    check_script(agent_workspace)
    check_synthesis(agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
