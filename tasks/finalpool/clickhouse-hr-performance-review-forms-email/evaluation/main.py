"""
Evaluation for sf-hr-performance-review-gform-email task (russified, ClickHouse + forms).

CRITICAL_CHECKS reflect the task's substance: a single critical failure => overall
FAIL (sys.exit(1)) regardless of accuracy. The accuracy>=70 gate applies afterward.

Checks:
1. Performance_Review_Setup.xlsx exists with "Performance Analysis" sheet and
   exactly 7 russified department rows.
2. Инженерия (Engineering) row matches the live re-query of
   sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" (Employee_Count and Avg_Performance),
   falling back to known constants (7096, ~3.21) if the DB is unavailable.
3. Agent-created form (NOT pre-seeded) with >=6 questions including a Department
   radio whose options are the 7 russified departments and a rating radio with the
   5 rating options.
4. Email to hr_director@company.com with the exact (kept-English) subject and a
   body summarizing the per-department distribution.
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Russified department values (must match db/zzz_clickhouse_after_init.sql).
EXPECTED_DEPARTMENTS = {"Инженерия", "Финансы", "Кадры", "Операции", "НИОКР", "Продажи", "Поддержка"}
ENGINEERING_RU = "Инженерия"

# Fallbacks if the live re-query is unavailable (relabel is value-only; row
# counts & ratings are unchanged by the snowflake->clickhouse swap).
ENG_COUNT_FALLBACK = 7096
ENG_AVG_FALLBACK = 3.21

EXACT_SUBJECT = "Annual Performance Review Process Setup Complete"

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        d = (detail[:300] + "...") if len(detail) > 300 else detail
        msg = f": {d}" if d else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def query_engineering_stats():
    """Live re-query of the Инженерия row from sf_data. Returns (count, avg) or None."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            'SELECT COUNT(*), ROUND(AVG("PERFORMANCE_RATING")::numeric, 2) '
            'FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" '
            'WHERE "DEPARTMENT" = %s',
            (ENGINEERING_RU,),
        )
        row = cur.fetchone()
        cur.close()
        conn.close()
        if row and row[0]:
            return int(row[0]), float(row[1])
    except Exception as e:
        print(f"  [info] live re-query unavailable, using fallback constants ({e})")
    return None


def _option_values(config):
    """Extract option text values from a question config (RU forms-mcp shape)."""
    vals = []
    if not config:
        return vals
    cfg = config
    if isinstance(cfg, str):
        try:
            cfg = json.loads(cfg)
        except Exception:
            return vals
    opts = cfg.get("options") if isinstance(cfg, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    vals.append(str(v))
            else:
                vals.append(str(o))
    return vals


def check_excel(agent_workspace):
    print("\n=== Проверка 1: Excel Performance_Review_Setup.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Performance_Review_Setup.xlsx")
    if not os.path.exists(xlsx_path):
        record("Performance_Review_Setup.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Лист содержит ровно 7 русифицированных отделов", False, "no file", critical=True)
        record("Строка Инженерия: Employee_Count и Avg_Performance корректны", False, "no file", critical=True)
        return
    record("Performance_Review_Setup.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e))
        record("Лист содержит ровно 7 русифицированных отделов", False, "unreadable", critical=True)
        record("Строка Инженерия: Employee_Count и Avg_Performance корректны", False, "unreadable", critical=True)
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    perf_idx = next((i for i, s in enumerate(sheet_names_lower) if "performance" in s or "analysis" in s), None)
    if perf_idx is None:
        perf_idx = 0

    ws = wb[wb.sheetnames[perf_idx]]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        record("Sheet has data", False, "Sheet is empty")
        record("Лист содержит ровно 7 русифицированных отделов", False, "empty sheet", critical=True)
        record("Строка Инженерия: Employee_Count и Avg_Performance корректны", False, "empty sheet", critical=True)
        return

    data_rows = [r for r in rows[1:] if any(c for c in r)]

    headers = [str(c).lower().strip() if c else "" for c in rows[0]]
    has_dept = any("dept" in h or "department" in h for h in headers)
    has_count = any("count" in h or "employee" in h for h in headers)
    has_avg = any("avg" in h or "average" in h or "performance" in h for h in headers)
    record("Has Department column", has_dept, f"Headers: {rows[0]}")
    record("Has Employee_Count column", has_count, f"Headers: {rows[0]}")
    record("Has Avg_Performance column", has_avg, f"Headers: {rows[0]}")

    # CRITICAL: exactly the 7 russified departments present as data rows.
    found_depts = set()
    for row in data_rows:
        for cell in row:
            if cell and str(cell).strip() in EXPECTED_DEPARTMENTS:
                found_depts.add(str(cell).strip())
    record("Лист содержит ровно 7 русифицированных отделов",
           found_depts == EXPECTED_DEPARTMENTS and len(data_rows) == 7,
           f"data_rows={len(data_rows)}, found={found_depts}, missing={EXPECTED_DEPARTMENTS - found_depts}",
           critical=True)

    # CRITICAL: Инженерия row matches live (or fallback) Engineering stats.
    dept_col = next((i for i, h in enumerate(headers) if "dept" in h or "department" in h), 0)
    count_col = next((i for i, h in enumerate(headers) if "count" in h or "employee" in h), 1)
    avg_col = next((i for i, h in enumerate(headers) if "avg" in h or "average" in h), 2)

    engineering_row = None
    for row in data_rows:
        if dept_col < len(row) and row[dept_col] and str(row[dept_col]).strip() == ENGINEERING_RU:
            engineering_row = row
            break

    live = query_engineering_stats()
    exp_count, exp_avg = (live if live else (ENG_COUNT_FALLBACK, ENG_AVG_FALLBACK))

    if not engineering_row:
        record("Строка Инженерия: Employee_Count и Avg_Performance корректны", False,
               "Инженерия row not found", critical=True)
    else:
        emp_count = engineering_row[count_col] if count_col < len(engineering_row) else None
        avg_perf = engineering_row[avg_col] if avg_col < len(engineering_row) else None
        ok = True
        detail = []
        try:
            emp_count_val = int(emp_count) if emp_count is not None else -1
            if emp_count_val != exp_count:
                ok = False
            detail.append(f"count={emp_count_val} (exp {exp_count})")
        except (TypeError, ValueError):
            ok = False
            detail.append(f"count unparseable: {emp_count}")
        try:
            avg_val = float(avg_perf) if avg_perf is not None else -1.0
            if abs(avg_val - exp_avg) > 0.05:
                ok = False
            detail.append(f"avg={avg_val} (exp ~{exp_avg})")
        except (TypeError, ValueError):
            ok = False
            detail.append(f"avg unparseable: {avg_perf}")
        record("Строка Инженерия: Employee_Count и Avg_Performance корректны", ok,
               "; ".join(detail), critical=True)


def check_gform():
    print("\n=== Проверка 2: Форма ежегодной оценки (forms / gform.*) ===")

    crit_form = "Форма оценки эффективности создана агентом (6 вопросов: Отдел + Оценка)"
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Хотя бы одна форма существует", False, str(e))
        record(crit_form, False, "no db", critical=True)
        return
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    record("Хотя бы одна форма существует", len(forms) > 0, "No forms in gform.forms")

    review_form = None
    for form_id, title in forms:
        t = (title or "").lower()
        if ("performance review" in t or "annual performance" in t
                or "оценк" in t or "эффектив" in t or "аттестац" in t):
            review_form = (form_id, title)
            break
    if review_form is None and forms:
        review_form = forms[0]

    record("Форма ежегодной оценки существует", review_form is not None,
           f"Forms: {[f[1] for f in forms]}")

    if review_form is None:
        record(crit_form, False, "no form", critical=True)
        cur.close()
        conn.close()
        return

    form_id, title = review_form
    cur.execute("SELECT title, question_type, config FROM gform.questions "
                "WHERE form_id = %s ORDER BY position", (form_id,))
    questions = cur.fetchall()
    cur.close()
    conn.close()

    q_titles_lower = [(q[0] or "").lower() for q in questions]
    q_count = len(questions)

    record("Форма содержит не менее 6 вопросов", q_count >= 6, f"Found {q_count} questions")

    has_dept_q = any("department" in t or "отдел" in t for t in q_titles_lower)
    has_rating_q = any("rating" in t or "performance" in t or "оценк" in t or "эффектив" in t
                       for t in q_titles_lower)
    has_achievement_q = any("achievement" in t or "accomplishment" in t or "достижен" in t
                            for t in q_titles_lower)
    record("Вопрос про отдел присутствует", has_dept_q, f"Questions: {[q[0] for q in questions]}")
    record("Вопрос про оценку эффективности присутствует", has_rating_q,
           f"Questions: {[q[0] for q in questions]}")
    record("Вопрос про ключевые достижения присутствует", has_achievement_q,
           f"Questions: {[q[0] for q in questions]}")

    # CRITICAL: dept radio options == 7 russified depts; rating radio has 5 options.
    dept_opts = set()
    rating_opts_count = 0
    for q_title, q_type, q_config in questions:
        tl = (q_title or "").lower()
        opts = _option_values(q_config)
        if "department" in tl or "отдел" in tl:
            dept_opts = {o.strip() for o in opts}
        if "rating" in tl or "оценк" in tl or "эффектив" in tl:
            rating_opts_count = max(rating_opts_count, len(opts))

    dept_ok = EXPECTED_DEPARTMENTS.issubset(dept_opts)
    rating_ok = rating_opts_count >= 5
    record(crit_form, dept_ok and rating_ok and q_count >= 6,
           f"dept_options={dept_opts}; rating_options_count={rating_opts_count}; q_count={q_count}",
           critical=True)


def check_email():
    print("\n=== Проверка 3: Письмо на hr_director@company.com ===")

    crit_email = "Письмо HR-директору: точная тема + распределение по отделам в теле"
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Email sent to hr_director@company.com", False, str(e))
        record(crit_email, False, "no db", critical=True)
        return
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "hr_director@company.com" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    record("Email sent to hr_director@company.com", matching is not None,
           f"Messages found: {len(messages)}")

    if matching is None:
        record(crit_email, False, "no matching email", critical=True)
        return

    subject, _, _, body_text = matching
    subj = (subject or "")
    body = (body_text or "")

    # Exact (kept-English) subject.
    subject_ok = subj.strip() == EXACT_SUBJECT
    record("Email subject exact match", subject_ok, f"Subject: {subj!r}")

    # Body must summarize the per-department distribution: mention >=4 russified
    # department names. Check the ORIGINAL (non-normalized) text.
    body_l = body.lower()
    mentioned = {d for d in EXPECTED_DEPARTMENTS if d.lower() in body_l}
    has_summary = len(mentioned) >= 4

    record("Email mentions performance-review content",
           ("оценк" in body_l or "эффектив" in body_l or "performance" in subj.lower()
            or "annual" in subj.lower() or len(mentioned) > 0),
           f"Subject: {subj}")

    record(crit_email, subject_ok and has_summary,
           f"subject_ok={subject_ok}; depts_in_body={mentioned}", critical=True)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILS,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILS}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
