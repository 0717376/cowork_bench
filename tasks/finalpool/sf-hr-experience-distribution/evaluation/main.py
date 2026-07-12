"""Evaluation for sf-hr-experience-distribution.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Band labels ('0-2 years', '3-5 years', '6-9 years', '10+ years') are output value
literals written by the agent and MUST stay English (compared via str_match and used
as row keys). Do NOT translate them.
"""
import argparse
import os
import sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

CRITICAL_CHECKS = {
    "Experience Analysis: Employee_Count matches groundtruth for all bands",
    "Experience Analysis: Avg_Salary within tolerance for all bands",
    "Summary.Total_Employees == 50000",
    "Summary.Most_Common_Band == '10+ years'",
    "Summary.Highest_Salary_Band == '10+ years'",
    "Teamly 'HR Experience Distribution' page exists with non-empty findings paragraph",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_experience_analysis(agent_wb, gt_wb):
    a_rows = load_sheet_rows(agent_wb, "Experience Analysis")
    g_rows = load_sheet_rows(gt_wb, "Experience Analysis")
    if a_rows is None or g_rows is None:
        check("Experience Analysis sheet present", False, "sheet missing")
        check("Experience Analysis: Employee_Count matches groundtruth for all bands", False, "sheet missing")
        check("Experience Analysis: Avg_Salary within tolerance for all bands", False, "sheet missing")
        check("Experience Analysis: Avg_Rating within tolerance for all bands", False, "sheet missing")
        check("Experience Analysis: Avg_Satisfaction within tolerance for all bands", False, "sheet missing")
        return
    check("Experience Analysis sheet present", True)

    a_data = a_rows[1:] if len(a_rows) > 1 else []
    g_data = g_rows[1:] if len(g_rows) > 1 else []
    a_lookup = {}
    for row in a_data:
        if row and row[0] is not None:
            a_lookup[str(row[0]).strip().lower()] = row

    count_errs, salary_errs, rating_errs, sat_errs = [], [], [], []
    for g_row in g_data:
        if not g_row or g_row[0] is None:
            continue
        key = str(g_row[0]).strip().lower()
        a_row = a_lookup.get(key)
        if a_row is None:
            count_errs.append(f"missing band: {g_row[0]}")
            salary_errs.append(f"missing band: {g_row[0]}")
            continue
        if len(a_row) > 1 and len(g_row) > 1 and not num_close(a_row[1], g_row[1], 5):
            count_errs.append(f"{key}.Employee_Count: {a_row[1]} vs {g_row[1]}")
        if len(a_row) > 2 and len(g_row) > 2 and not num_close(a_row[2], g_row[2], 10.0):
            salary_errs.append(f"{key}.Avg_Salary: {a_row[2]} vs {g_row[2]}")
        if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 0.1):
            rating_errs.append(f"{key}.Avg_Rating: {a_row[3]} vs {g_row[3]}")
        if len(a_row) > 4 and len(g_row) > 4 and not num_close(a_row[4], g_row[4], 0.1):
            sat_errs.append(f"{key}.Avg_Satisfaction: {a_row[4]} vs {g_row[4]}")

    check("Experience Analysis: Employee_Count matches groundtruth for all bands", not count_errs, count_errs)
    check("Experience Analysis: Avg_Salary within tolerance for all bands", not salary_errs, salary_errs)
    check("Experience Analysis: Avg_Rating within tolerance for all bands", not rating_errs, rating_errs)
    check("Experience Analysis: Avg_Satisfaction within tolerance for all bands", not sat_errs, sat_errs)


def check_summary(agent_wb, gt_wb):
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None or g_rows is None:
        check("Summary sheet present", False, "sheet missing")
        check("Summary.Total_Employees == 50000", False, "sheet missing")
        check("Summary.Most_Common_Band == '10+ years'", False, "sheet missing")
        check("Summary.Highest_Salary_Band == '10+ years'", False, "sheet missing")
        return
    check("Summary sheet present", True)

    a_data = a_rows[1:] if len(a_rows) > 1 else []
    g_data = g_rows[1:] if len(g_rows) > 1 else []
    a_lookup = {}
    for row in a_data:
        if row and row[0] is not None:
            a_lookup[str(row[0]).strip().lower()] = row
    g_lookup = {}
    for row in g_data:
        if row and row[0] is not None:
            g_lookup[str(row[0]).strip().lower()] = row

    def cmp(metric, label):
        g_row = g_lookup.get(metric.lower())
        a_row = a_lookup.get(metric.lower())
        gv = g_row[1] if g_row and len(g_row) > 1 else None
        av = a_row[1] if a_row and len(a_row) > 1 else None
        check(label, str_match(av, gv), f"'{av}' vs '{gv}'")

    cmp("Total_Employees", "Summary.Total_Employees == 50000")
    cmp("Most_Common_Band", "Summary.Most_Common_Band == '10+ years'")
    cmp("Highest_Salary_Band", "Summary.Highest_Salary_Band == '10+ years'")


def check_teamly():
    """Critical: an 'HR Experience Distribution' page exists with a non-empty
    findings paragraph. Seed pages have id <= 3 and must NOT satisfy the check.
    """
    name = "Teamly 'HR Experience Distribution' page exists with non-empty findings paragraph"
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception:
        # Fallback for schemas that use 'content' instead of 'body'.
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT id, title, COALESCE(content, '') FROM teamly.pages WHERE id > 3")
            pages = cur.fetchall()
            conn.close()
        except Exception as e:
            check(name, False, str(e))
            return

    target = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "hr experience distribution" in tl or ("experience" in tl and "distribut" in tl) \
                or ("опыт" in tl and "распределен" in tl):
            target = (pid, title, body)
            break
    # Non-empty findings paragraph: at least ~40 chars of body text.
    ok = target is not None and len((target[2] or "").strip()) >= 40
    check(name, ok, f"new pages: {[(p[0], p[1]) for p in pages]}")


def run_evaluation(agent_workspace, groundtruth_workspace):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(agent_workspace, "HR_Experience_Report.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Experience_Report.xlsx")

    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        return False, "groundtruth missing"

    if not os.path.exists(agent_file):
        check("HR_Experience_Report.xlsx exists", False, agent_file)
    else:
        check("HR_Experience_Report.xlsx exists", True)
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
        check_experience_analysis(agent_wb, gt_wb)
        check_summary(agent_wb, gt_wb)

    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(args.agent_workspace, args.groundtruth_workspace)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
