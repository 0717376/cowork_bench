"""
Evaluation script for canvas-grade-distribution-forms task (RU).

Gate: accuracy >= 70% AND no CRITICAL check failed => PASS.
Any CRITICAL failure => immediate FAIL (sys.exit(1)) regardless of accuracy.

Checks:
1. Excel "Course Grades" + "Summary" verified AGAINST live canvas DB
   (Avg_Grade / Pass_Rate / Median / Total — agent must compute, not fabricate).
2. Forms survey (RU forms fork, schema gform.*) with >=4 questions,
   verified by question TYPE (choice vs text), not just count.
3. Email subject "Spring 2014 Course Feedback Survey" from registrar@university.edu
   to students@university.edu (both endpoints required).
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "[CRIT]" if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")


def num_close(a, b, tol=0.5):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_course_data():
    """Query actual grade stats from canvas DB (read live, never hardcoded).

    Returns list of dicts keyed by course_code with avg/median/pass_rate/total.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.course_code, c.name,
               AVG(s.score)                                                    AS avg_grade,
               PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY s.score)            AS median_grade,
               100.0 * COUNT(*) FILTER (WHERE s.score >= 60) / COUNT(*)        AS pass_rate,
               COUNT(DISTINCT s.user_id)                                       AS total_students
        FROM canvas.courses c
        JOIN canvas.assignments a ON a.course_id = c.id
        JOIN canvas.submissions s ON s.assignment_id = a.id
        WHERE (c.name LIKE '%%Spring 2014%%' OR c.name LIKE '%%Весна 2014%%'
               OR c.course_code LIKE '%%-2014B')
          AND s.score IS NOT NULL
        GROUP BY c.course_code, c.name
        ORDER BY c.name
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    stats = []
    for r in rows:
        stats.append({
            "code": str(r[0]).strip().upper(),
            "name": str(r[1]),
            "avg": float(r[2]) if r[2] is not None else None,
            "median": float(r[3]) if r[3] is not None else None,
            "pass_rate": float(r[4]) if r[4] is not None else None,
            "total": int(r[5]) if r[5] is not None else None,
        })
    return stats


def find_col(headers, *keys):
    for idx, h in enumerate(headers):
        hn = h.replace("_", "")
        for k in keys:
            kn = k.replace("_", "")
            if k in h or kn in hn:
                return idx
    return None


def check_excel(workspace):
    print("\n=== Checking Excel ===")
    from openpyxl import load_workbook

    xlsx_path = os.path.join(workspace, "Grade_Distribution_Report.xlsx")
    if not os.path.exists(xlsx_path):
        check("Grade_Distribution_Report.xlsx exists", False, xlsx_path, critical=True)
        return
    check("Grade_Distribution_Report.xlsx exists", True)

    try:
        course_stats = get_expected_course_data()
    except Exception as e:
        check("Canvas DB readable for groundtruth", False, str(e), critical=True)
        return
    check("Canvas Spring 2014 courses found", len(course_stats) > 0,
          f"found {len(course_stats)} courses")
    stats_by_code = {cs["code"]: cs for cs in course_stats}

    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e), critical=True)
        return
    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # ---- Course Grades sheet ----
    if "course grades" not in sheet_names_lower:
        check("'Course Grades' sheet exists", False, f"Found: {wb.sheetnames}", critical=True)
    else:
        check("'Course Grades' sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("course grades")]]
        headers = [str(c.value).lower().replace(" ", "_") if c.value else "" for c in ws[1]]

        for rh in ["course_code", "course_name", "avg_grade", "total_students"]:
            check(f"Course Grades header '{rh}'",
                  find_col(headers, rh) is not None,
                  f"headers={headers}")

        code_col = find_col(headers, "course_code", "code")
        avg_col = find_col(headers, "avg_grade", "average")
        med_col = find_col(headers, "median_grade", "median")
        pass_col = find_col(headers, "pass_rate", "pass")
        tot_col = find_col(headers, "total_students", "students")

        agent_rows = {}
        for row in ws.iter_rows(min_row=2):
            if code_col is not None and row[code_col].value:
                agent_rows[str(row[code_col].value).strip().upper()] = row

        check("Course Grades has one row per Spring 2014 course",
              len(agent_rows) >= len(course_stats),
              f"rows={len(agent_rows)}, expected>={len(course_stats)}")

        # All expected course codes present (structural)
        missing_codes = set(stats_by_code) - set(agent_rows)
        check("All Spring 2014 course codes present", not missing_codes,
              f"missing={missing_codes}")

        # CRITICAL: Avg_Grade matches DB-computed AVG(score) within +/-0.5
        avg_errors = []
        if avg_col is not None:
            for code, cs in stats_by_code.items():
                row = agent_rows.get(code)
                if row is None:
                    avg_errors.append(f"{code}: no row")
                    continue
                v = row[avg_col].value
                if not num_close(v, cs["avg"], 0.5):
                    avg_errors.append(f"{code}: Avg_Grade {v} vs DB {cs['avg']:.2f}")
        else:
            avg_errors.append("Avg_Grade column not found")
        check("Course Grades Avg_Grade matches canvas DB (+/-0.5)",
              not avg_errors, "; ".join(avg_errors[:6]), critical=True)

        # CRITICAL: Pass_Rate matches DB-computed pass rate within +/-0.5
        pass_errors = []
        if pass_col is not None:
            for code, cs in stats_by_code.items():
                row = agent_rows.get(code)
                if row is None:
                    pass_errors.append(f"{code}: no row")
                    continue
                v = row[pass_col].value
                if not num_close(v, cs["pass_rate"], 0.5):
                    pass_errors.append(f"{code}: Pass_Rate {v} vs DB {cs['pass_rate']:.2f}")
        else:
            pass_errors.append("Pass_Rate column not found")
        check("Course Grades Pass_Rate matches canvas DB (+/-0.5)",
              not pass_errors, "; ".join(pass_errors[:6]), critical=True)

        # Median (non-critical structural-value check)
        med_errors = []
        if med_col is not None:
            for code, cs in stats_by_code.items():
                row = agent_rows.get(code)
                if row is None:
                    continue
                if not num_close(row[med_col].value, cs["median"], 0.5):
                    med_errors.append(f"{code}: Median {row[med_col].value} vs {cs['median']:.2f}")
        check("Course Grades Median_Grade matches canvas DB (+/-0.5)",
              not med_errors, "; ".join(med_errors[:6]))

        # Total_Students (non-critical structural-value check)
        tot_errors = []
        if tot_col is not None:
            for code, cs in stats_by_code.items():
                row = agent_rows.get(code)
                if row is None:
                    continue
                if not num_close(row[tot_col].value, cs["total"], 0):
                    tot_errors.append(f"{code}: Total {row[tot_col].value} vs {cs['total']}")
        check("Course Grades Total_Students matches canvas DB",
              not tot_errors, "; ".join(tot_errors[:6]))

    # ---- Summary sheet ----
    if "summary" not in sheet_names_lower:
        check("'Summary' sheet exists", False, f"Found: {wb.sheetnames}", critical=True)
    else:
        check("'Summary' sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        summary = {}
        for row in ws.iter_rows(min_row=2):
            if row[0].value:
                key = str(row[0].value).lower().replace(" ", "_")
                summary[key] = row[1].value if len(row) > 1 else None

        def get_summary(*keys):
            for k, v in summary.items():
                kn = k.replace("_", "")
                for key in keys:
                    if all(part in k for part in key.split("+")) or \
                       all(part in kn for part in key.replace("_", "").split("+")):
                        return v
            return None

        # Total_Courses == number of Spring 2014 courses
        total_val = get_summary("total+course")
        ok_total = False
        try:
            ok_total = int(float(total_val)) == len(course_stats)
        except (TypeError, ValueError):
            ok_total = False
        check("Summary Total_Courses correct", ok_total,
              f"got {total_val}, expected {len(course_stats)}")

        # Derive highest/lowest/overall from DB
        valid = [cs for cs in course_stats if cs["avg"] is not None]
        if valid:
            highest = max(valid, key=lambda c: c["avg"])
            lowest = min(valid, key=lambda c: c["avg"])
            overall = sum(c["avg"] for c in valid) / len(valid)

            hi_val = get_summary("highest+avg", "highest")
            lo_val = get_summary("lowest+avg", "lowest")
            ov_val = get_summary("overall+avg", "overall")

            def name_match(cell, cs):
                if cell is None:
                    return False
                s = str(cell).strip().lower()
                return cs["name"].lower() in s or s in cs["name"].lower() \
                    or cs["code"].lower() in s

            # CRITICAL: highest/lowest course identification + overall avg
            hi_ok = name_match(hi_val, highest)
            lo_ok = name_match(lo_val, lowest)
            ov_ok = num_close(ov_val, overall, 0.5)
            check("Summary Highest/Lowest/Overall match canvas DB",
                  hi_ok and lo_ok and ov_ok,
                  f"highest='{hi_val}' (exp {highest['name']}) "
                  f"lowest='{lo_val}' (exp {lowest['name']}) "
                  f"overall={ov_val} (exp {overall:.2f})",
                  critical=True)


def _qtype_is_text(t):
    t = (t or "")
    return t in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH")


def _qtype_is_choice(t):
    t = (t or "")
    return t in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE", "CHOICE", "CHECKBOX")


def check_gform(cur):
    print("\n=== Checking Forms survey (forms RU fork, gform schema) ===")

    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE LOWER(title) LIKE '%%spring 2014%%'
          AND (LOWER(title) LIKE '%%survey%%' OR LOWER(title) LIKE '%%feedback%%'
               OR LOWER(title) LIKE '%%опрос%%' OR LOWER(title) LIKE '%%обратн%%')
        ORDER BY created_at DESC LIMIT 1
    """)
    form_row = cur.fetchone()
    if not form_row:
        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE LOWER(title) LIKE '%%spring 2014%%'
            ORDER BY created_at DESC LIMIT 1
        """)
        form_row = cur.fetchone()

    check("Spring 2014 feedback survey form exists", form_row is not None,
          critical=True)
    if not form_row:
        return

    form_id = form_row[0]
    cur.execute("""
        SELECT title, question_type, required
        FROM gform.questions
        WHERE form_id = %s ORDER BY position ASC
    """, (form_id,))
    questions = cur.fetchall()
    types = [q[1] for q in questions]

    check("Form has >=4 questions", len(questions) >= 4,
          f"found {len(questions)}: {[q[0] for q in questions]}")

    choice_count = sum(1 for t in types if _qtype_is_choice(t))
    text_count = sum(1 for t in types if _qtype_is_text(t))
    check("Form has >=3 choice questions", choice_count >= 3,
          f"choice_count={choice_count}, types={types}")
    check("Form has >=1 text question", text_count >= 1,
          f"text_count={text_count}, types={types}")

    # Non-required text question (suggestions, not required)
    has_optional_text = any(
        _qtype_is_text(q[1]) and (q[2] is False or q[2] is None or str(q[2]).lower() in ("false", "0", "f"))
        for q in questions
    )
    check("Form has a non-required text question (suggestions)", has_optional_text,
          f"questions={[(q[0], q[1], q[2]) for q in questions]}")

    # CRITICAL: structured question types, not just count
    check("Survey has >=4 questions incl. >=3 choice + >=1 non-required text",
          len(questions) >= 4 and choice_count >= 3 and text_count >= 1 and has_optional_text,
          f"q={len(questions)} choice={choice_count} text={text_count} opt_text={has_optional_text}",
          critical=True)


def check_email(cur):
    print("\n=== Checking Email ===")
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%%spring 2014%%'
          AND (LOWER(subject) LIKE '%%survey%%' OR LOWER(subject) LIKE '%%feedback%%')
    """)
    emails = cur.fetchall()
    check("Email with Spring 2014 survey/feedback subject exists",
          len(emails) >= 1, critical=True)
    if not emails:
        return

    # Pick the message that best satisfies both endpoints.
    def from_ok(m):
        return "registrar@university.edu" in str(m[1]).lower()

    def to_ok(m):
        return "students@university.edu" in str(m[2]).lower()

    any_to = any(to_ok(m) for m in emails)
    any_from = any(from_ok(m) for m in emails)
    both = any(from_ok(m) and to_ok(m) for m in emails)

    check("Email sent to students@university.edu", any_to,
          f"to_addrs={[m[2] for m in emails]}")
    check("Email sent from registrar@university.edu", any_from,
          f"from_addrs={[m[1] for m in emails]}")
    # CRITICAL: both required endpoints on the same message
    check("Survey email: from registrar@university.edu to students@university.edu",
          both,
          f"endpoints={[(m[1], m[2]) for m in emails]}",
          critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        check_gform(cur)
        check_email(cur)
        cur.close()
        conn.close()
    except Exception as e:
        check("DB checks", False, str(e), critical=True)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "critical_failed": CRITICAL_FAILED,
            "success": not CRITICAL_FAILED and accuracy >= 70,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print("CRITICAL checks failed:")
        for c in CRITICAL_FAILED:
            print(f"  - {c}")
        print("=> FAIL (critical)")
        sys.exit(1)

    if accuracy >= 70:
        print("=> PASS (accuracy >= 70% and no critical failure)")
        sys.exit(0)
    print(f"=> FAIL (accuracy {accuracy:.1f}% < 70%)")
    sys.exit(1)


if __name__ == "__main__":
    main()
