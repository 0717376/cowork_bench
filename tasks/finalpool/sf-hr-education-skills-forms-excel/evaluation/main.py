"""Evaluation for sf-hr-education-skills-gform-excel (ClickHouse + forms RU fork).

Gate: accuracy >= 70% AND no CRITICAL check failed => PASS.
Any CRITICAL failure => immediate FAIL (sys.exit(1)) regardless of accuracy.

CRITICAL checks (substance):
1. Education Breakdown: every (department, education_level) russified row matches
   groundtruth Employee_Count (+/-1) AND Avg_Performance (+/-0.05).
2. Department Summary: per-department High_School_Count/Bachelor_Count/Masters_Count/
   PhD_Count match groundtruth exactly AND Total_Employees matches — enforces the
   "Diploma counted in Total but excluded from level columns" rule.
3. Department Summary: Higher_Ed_Pct matches groundtruth within +/-0.5 per department.
4. Google Form (gform.*): "Training Interest Survey" with >=4 questions incl. >=2
   choice questions and questions referencing department + education level.
5. Email to training@company.com with non-empty subject AND body containing RU
   education/analysis keywords.

Dept/edu data values are RUSSIAN (russified centrally in sf_data + groundtruth xlsx).
The agent reads them from ClickHouse and writes them verbatim; row-matching keys on
both sides are Russian.
"""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "[CRIT]" if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def int_eq(a, b):
    try:
        return int(round(float(a))) == int(round(float(b)))
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_ws, gt_dir):
    import openpyxl

    agent_file = os.path.join(agent_ws, "HR_Education_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Education_Analysis.xlsx")

    print("\n=== Checking Excel file ===")
    if not os.path.exists(agent_file):
        check("HR_Education_Analysis.xlsx exists", False, "not found", critical=True)
        return
    check("HR_Education_Analysis.xlsx exists", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # --- Education Breakdown sheet ---
    print("  -- Education Breakdown --")
    a_rows = load_sheet_rows(agent_wb, "Education Breakdown")
    g_rows = load_sheet_rows(gt_wb, "Education Breakdown")
    if a_rows is None:
        check("Sheet 'Education Breakdown' present", False, critical=True)
    else:
        check("Sheet 'Education Breakdown' present", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Education Breakdown has 35 rows", len(a_data) == 35,
              f"got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] and row[1]:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower())
                a_lookup[key] = row

        cnt_ok = perf_ok = True
        cnt_bad = perf_bad = ""
        for g_row in g_data:
            if not g_row or not g_row[0]:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower())
            a_row = a_lookup.get(key)
            if a_row is None:
                cnt_ok = perf_ok = False
                cnt_bad = f"missing {g_row[0]}/{g_row[1]}"
                continue
            if not num_close(a_row[2], g_row[2], 1):
                cnt_ok = False
                cnt_bad = f"{g_row[0]}/{g_row[1]}: {a_row[2]} vs {g_row[2]}"
            if not num_close(a_row[3], g_row[3], 0.05):
                perf_ok = False
                perf_bad = f"{g_row[0]}/{g_row[1]}: {a_row[3]} vs {g_row[3]}"
        # CRITICAL: source-of-truth aggregation
        check("Education Breakdown Employee_Count matches all 35 rows (+/-1)",
              cnt_ok, cnt_bad, critical=True)
        check("Education Breakdown Avg_Performance matches all 35 rows (+/-0.05)",
              perf_ok, perf_bad, critical=True)

    # --- Department Summary sheet ---
    print("  -- Department Summary --")
    a_rows2 = load_sheet_rows(agent_wb, "Department Summary")
    g_rows2 = load_sheet_rows(gt_wb, "Department Summary")
    if a_rows2 is None:
        check("Sheet 'Department Summary' present", False, critical=True)
    else:
        check("Sheet 'Department Summary' present", True)
        a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
        g_data2 = g_rows2[1:] if len(g_rows2) > 1 else []
        check("Department Summary has 7 rows", len(a_data2) == 7,
              f"got {len(a_data2)}")

        a_lookup2 = {}
        for row in a_data2:
            if row and row[0]:
                a_lookup2[str(row[0]).strip().lower()] = row

        total_ok = levels_ok = pct_ok = True
        total_bad = levels_bad = pct_bad = ""
        for g_row in g_data2:
            if not g_row or not g_row[0]:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup2.get(key)
            if a_row is None:
                total_ok = levels_ok = pct_ok = False
                total_bad = f"missing dept {g_row[0]}"
                continue
            # Total_Employees (index 1)
            if not num_close(a_row[1], g_row[1], 1):
                total_ok = False
                total_bad = f"{g_row[0]} Total: {a_row[1]} vs {g_row[1]}"
            # Per-level counts (indices 2..5): High_School, Bachelor, Masters, PhD
            for idx, label in ((2, "High_School"), (3, "Bachelor"),
                               (4, "Masters"), (5, "PhD")):
                av = a_row[idx] if len(a_row) > idx else None
                gv = g_row[idx] if len(g_row) > idx else None
                if not int_eq(av, gv):
                    levels_ok = False
                    levels_bad = f"{g_row[0]} {label}_Count: {av} vs {gv}"
            # Higher_Ed_Pct (index 6)
            av6 = a_row[6] if len(a_row) > 6 else None
            gv6 = g_row[6] if len(g_row) > 6 else None
            if not num_close(av6, gv6, 0.5):
                pct_ok = False
                pct_bad = f"{g_row[0]} Higher_Ed_Pct: {av6} vs {gv6}"

        # CRITICAL: Diploma-exclusion rule enforced via per-level counts + total
        check("Department Summary Total_Employees matches all depts",
              total_ok, total_bad, critical=True)
        check("Department Summary per-level counts (HS/Bach/Mast/PhD) match exactly "
              "(Diploma excluded from level cols, counted in Total)",
              levels_ok, levels_bad, critical=True)
        check("Department Summary Higher_Ed_Pct matches all depts (+/-0.5)",
              pct_ok, pct_bad, critical=True)


def is_choice(qtype):
    t = (qtype or "").strip()
    return t in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE", "CHOICE", "CHECKBOX")


def check_gform(cur):
    print("\n=== Checking Forms survey (forms RU fork, gform schema) ===")
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE LOWER(title) LIKE '%%training interest%%'
        ORDER BY created_at DESC LIMIT 1
    """)
    row = cur.fetchone()
    if not row:
        check("Form 'Training Interest Survey' exists", False,
              "no form titled like 'training interest'", critical=True)
        return
    check("Form 'Training Interest Survey' exists", True)
    form_id = row[0]

    cur.execute("""
        SELECT title, question_type, COALESCE(config::text, '')
        FROM gform.questions
        WHERE form_id = %s
        ORDER BY position
    """, (form_id,))
    questions = cur.fetchall()
    titles = [(q[0] or "").lower() for q in questions]
    blob = " ".join(titles)
    choice_count = sum(1 for q in questions if is_choice(q[1]))

    has_dept = any(("отдел" in t or "департамент" in t or "department" in t)
                   for t in titles)
    has_edu = any(("образован" in t or "education" in t) for t in titles)
    # training-interest / format choice questions
    has_interest = any(
        is_choice(q[1]) and ("интерес" in (q[0] or "").lower()
                             or "сфер" in (q[0] or "").lower()
                             or "interest" in (q[0] or "").lower())
        for q in questions)
    has_format = any(
        is_choice(q[1]) and ("формат" in (q[0] or "").lower()
                            or "format" in (q[0] or "").lower())
        for q in questions)

    # CRITICAL: real question substance, not just title of an empty form
    check("Form has >=4 questions", len(questions) >= 4,
          f"found {len(questions)}: {[q[0] for q in questions]}", critical=True)
    check("Form references department + education level questions",
          has_dept and has_edu,
          f"dept={has_dept} edu={has_edu} titles={titles}", critical=True)
    check("Form has >=2 choice questions (training interest + format)",
          choice_count >= 2 and (has_interest or has_format),
          f"choice={choice_count} interest={has_interest} format={has_format} "
          f"blob={blob[:200]}", critical=True)


def check_email(cur):
    print("\n=== Checking email to training@company.com ===")
    cur.execute("""
        SELECT subject, body_text
        FROM email.messages
        WHERE to_addr::text ILIKE '%%training@company.com%%'
        ORDER BY id DESC
    """)
    rows = cur.fetchall()
    if not rows:
        check("Email to training@company.com exists", False,
              "no message", critical=True)
        return
    check("Email to training@company.com exists", True)

    # non-empty subject AND body with RU education/analysis keywords
    KW = ("образован", "обучени", "отдел", "анализ", "сотрудник")
    best = None
    for subj, body in rows:
        subj = (subj or "").strip()
        body = (body or "").strip()
        text = (subj + " " + body).lower()
        kw_hit = sum(1 for k in KW if k in text)
        if subj and body and kw_hit >= 1:
            best = (subj, body, kw_hit)
            break
        if best is None:
            best = (subj, body, kw_hit)

    subj, body, kw_hit = best
    check("Email has non-empty subject", bool(subj), f"subject={subj[:80]!r}")
    # CRITICAL: substance — non-empty body referencing the analysis topic
    check("Email body non-empty with education/analysis RU keywords",
          bool(body) and kw_hit >= 1,
          f"kw_hit={kw_hit} body[:120]={body[:120]!r}", critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    check_excel(agent_ws, gt_dir)

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        check_gform(cur)
        check_email(cur)
        cur.close()
        conn.close()
    except Exception as e:
        check("DB checks ran", False, str(e), critical=True)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (100.0 * PASS_COUNT / total) if total else 0.0
    print(f"\n=== Summary: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if CRITICAL_FAILED:
        print("CRITICAL checks failed:")
        for c in CRITICAL_FAILED:
            print(f"  - {c}")

    if CRITICAL_FAILED:
        print("=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    print(f"=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
    sys.exit(1)


if __name__ == "__main__":
    main()
