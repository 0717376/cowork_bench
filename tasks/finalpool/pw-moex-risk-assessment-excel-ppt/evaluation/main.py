"""Evaluation for pw-moex-risk-assessment-excel-ppt.

Структурные (НЕ критичные) чеки: наличие файлов/листов/колонок, минимальное
число строк, наличие .pptx и yf_risk_processor.py.

CRITICAL чеки (любой провал => немедленный FAIL вне зависимости от accuracy):
  C1: Data_Analysis — каждый Symbol это реальный тикер MOEX, а Current_Price и
      Sector совпадают с живыми данными moex.stock_info (читаются из БД,
      сверка с допуском). Если БД недоступна — чек пропускается, не фейлит.
  C2: Data_Analysis — для каждой строки Target_Price ≈ Current_Price*1.10 и
      Upside ≈ Target_Price - Current_Price (с допуском). Проверка реального
      расчёта, а не выдуманных чисел.
  C3: Data_Analysis отсортирован по алфавиту по колонке Symbol.
  C4: Metrics — Total_Stocks == числу строк Data_Analysis И Avg_Upside ==
      среднему колонки Upside (кросс-листовая согласованность, с допуском).
  C5: Recommendations — хотя бы одна строка ссылается на Symbol, реально
      присутствующий в Data_Analysis.

PASS, если нет провалов критичных чеков И accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []
CRITICAL_CHECKS = {
    "C1: Data_Analysis — Symbol/Current_Price/Sector соответствуют живым данным MOEX",
    "C2: Data_Analysis — Target_Price=Current*1.10 и Upside=Target-Current (расчёт)",
    "C3: Data_Analysis отсортирован по алфавиту по Symbol",
    "C4: Metrics согласован с Data_Analysis (Total_Stocks и Avg_Upside)",
    "C5: Recommendations ссылается на реальный Symbol из Data_Analysis",
}

VALID_TICKERS = {"SBER.ME", "GAZP.ME", "LKOH.ME", "TCSG.ME", "MGNT.ME", "MTSS.ME"}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('₽', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def load_moex_info():
    """Возвращает {symbol: {currentPrice, sector}} из живой БД MOEX.

    currentPrice берём из последней moex.stock_prices.close — это единый источник
    правды, который MCP-адаптер (get_stock_info) подставляет в поле currentPrice
    (см. moex-finance-mcp/pg_adapter.py). JSONB data->>'currentPrice' статичен и
    может расходиться с close, поэтому для сверки с агентом используем close.

    При недоступности БД возвращает None (DB-зависимые чеки тогда пропускаются)."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT symbol, data FROM moex.stock_info WHERE symbol LIKE '%%.ME'")
        out = {}
        for sym, data in cur.fetchall():
            if isinstance(data, str):
                data = json.loads(data)
            out[sym] = {
                "currentPrice": safe_float(data.get("currentPrice")),
                "sector": (data.get("sector") or "").strip(),
            }
        # Override currentPrice with latest stock_prices.close (tool's source of truth)
        for sym in out:
            cur.execute(
                "SELECT close FROM moex.stock_prices WHERE symbol = %s "
                "ORDER BY date DESC LIMIT 1", (sym,))
            r = cur.fetchone()
            if r and r[0] is not None:
                out[sym]["currentPrice"] = safe_float(r[0])
        cur.close()
        conn.close()
        return out
    except Exception as e:
        print(f"  [WARN] MOEX БД недоступна, чеки с живыми ценами пропущены: {e}")
        return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    moex = load_moex_info()

    excel_path = os.path.join(agent_workspace, "Risk_Assessment_Report.xlsx")
    check("Risk_Assessment_Report.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return FAIL_COUNT == 0 and not CRITICAL_FAILS, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"

    wb = openpyxl.load_workbook(excel_path)

    # ---------------- Data_Analysis ----------------
    da_rows = []          # list of dicts keyed by lowercased header
    headers = []
    check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
    if "Data_Analysis" in wb.sheetnames:
        ws = wb["Data_Analysis"]
        headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for expected_col in ['Symbol', 'Name', 'Sector', 'Current_Price', 'Target_Price', 'Upside']:
            check(f"Data_Analysis has {expected_col} column",
                  expected_col.lower() in headers, f"headers: {headers[:8]}")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(c is None for c in row):
                continue
            da_rows.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
        check("Data_Analysis has >= 5 rows", len(da_rows) >= 5, f"got {len(da_rows)}")

    def col(d, name):
        return d.get(name.lower())

    # ---- C1: live MOEX price/sector match ----
    if da_rows and moex is not None:
        ok, bad = True, []
        for d in da_rows:
            sym = str(col(d, "Symbol")).strip()
            if sym not in VALID_TICKERS:
                ok = False; bad.append(f"{sym}: не тикер MOEX"); continue
            info = moex.get(sym)
            if not info:
                ok = False; bad.append(f"{sym}: нет в БД"); continue
            cur_price = safe_float(col(d, "Current_Price"))
            live = info["currentPrice"]
            if cur_price is None or live is None or abs(cur_price - live) > max(0.5, 0.01 * live):
                ok = False; bad.append(f"{sym}: Current_Price={cur_price} vs live={live}")
            sec = (str(col(d, "Sector")).strip() if col(d, "Sector") else "")
            live_sec = info["sector"]
            if live_sec and sec.lower() != live_sec.lower():
                ok = False; bad.append(f"{sym}: Sector={sec!r} vs live={live_sec!r}")
        check("C1: Data_Analysis — Symbol/Current_Price/Sector соответствуют живым данным MOEX",
              ok, "; ".join(bad))
    elif da_rows and moex is None:
        print("  [SKIP] C1 (БД недоступна)")
    else:
        check("C1: Data_Analysis — Symbol/Current_Price/Sector соответствуют живым данным MOEX",
              False, "нет строк Data_Analysis")

    # ---- C2: Target = Current*1.10, Upside = Target - Current ----
    if da_rows:
        ok, bad = True, []
        for d in da_rows:
            sym = str(col(d, "Symbol")).strip()
            cur_p = safe_float(col(d, "Current_Price"))
            tgt = safe_float(col(d, "Target_Price"))
            up = safe_float(col(d, "Upside"))
            if cur_p is None or tgt is None or up is None:
                ok = False; bad.append(f"{sym}: пустые значения"); continue
            exp_tgt = cur_p * 1.10
            if abs(tgt - exp_tgt) > max(0.5, 0.02 * exp_tgt):
                ok = False; bad.append(f"{sym}: Target={tgt} != Current*1.10={exp_tgt:.2f}")
            if abs(up - (tgt - cur_p)) > max(0.5, 0.02 * abs(tgt - cur_p) if tgt != cur_p else 0.5):
                ok = False; bad.append(f"{sym}: Upside={up} != Target-Current={tgt-cur_p:.2f}")
        check("C2: Data_Analysis — Target_Price=Current*1.10 и Upside=Target-Current (расчёт)",
              ok, "; ".join(bad))
    else:
        check("C2: Data_Analysis — Target_Price=Current*1.10 и Upside=Target-Current (расчёт)",
              False, "нет строк")

    # ---- C3: alphabetical sort by Symbol ----
    if da_rows:
        syms = [str(col(d, "Symbol")).strip() for d in da_rows]
        check("C3: Data_Analysis отсортирован по алфавиту по Symbol",
              syms == sorted(syms), f"got {syms}")
    else:
        check("C3: Data_Analysis отсортирован по алфавиту по Symbol", False, "нет строк")

    # ---------------- Metrics ----------------
    metrics = {}
    check("Metrics sheet exists", "Metrics" in wb.sheetnames)
    if "Metrics" in wb.sheetnames:
        ws = wb["Metrics"]
        mheaders = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for expected_col in ['Metric', 'Value']:
            check(f"Metrics has {expected_col} column",
                  expected_col.lower() in mheaders, f"headers: {mheaders[:8]}")
        m_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        check("Metrics has >= 3 rows", len(m_rows) >= 3, f"got {len(m_rows)}")
        for r in m_rows:
            if r and r[0] is not None:
                metrics[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

    # ---- C4: cross-sheet consistency ----
    if da_rows and metrics:
        ok, bad = True, []
        total = safe_float(metrics.get("total_stocks"))
        if total is None or int(total) != len(da_rows):
            ok = False; bad.append(f"Total_Stocks={metrics.get('total_stocks')} != rows={len(da_rows)}")
        ups = [safe_float(col(d, "Upside")) for d in da_rows]
        ups = [u for u in ups if u is not None]
        if ups:
            mean_up = sum(ups) / len(ups)
            avg = safe_float(metrics.get("avg_upside"))
            if avg is None or abs(avg - mean_up) > max(0.5, 0.02 * abs(mean_up) if mean_up else 0.5):
                ok = False; bad.append(f"Avg_Upside={metrics.get('avg_upside')} != mean={mean_up:.2f}")
        else:
            ok = False; bad.append("нет значений Upside")
        check("C4: Metrics согласован с Data_Analysis (Total_Stocks и Avg_Upside)", ok, "; ".join(bad))
    else:
        check("C4: Metrics согласован с Data_Analysis (Total_Stocks и Avg_Upside)",
              False, "нет Data_Analysis/Metrics")

    # ---------------- Recommendations ----------------
    rec_symbols = []
    check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
    if "Recommendations" in wb.sheetnames:
        ws = wb["Recommendations"]
        rheaders = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        for expected_col in ['Priority', 'Action', 'Symbol']:
            check(f"Recommendations has {expected_col} column",
                  expected_col.lower() in rheaders, f"headers: {rheaders[:8]}")
        sym_idx = rheaders.index("symbol") if "symbol" in rheaders else None
        r_rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        check("Recommendations has >= 2 rows", len(r_rows) >= 2, f"got {len(r_rows)}")
        if sym_idx is not None:
            for r in r_rows:
                if sym_idx < len(r) and r[sym_idx] is not None:
                    rec_symbols.append(str(r[sym_idx]).strip())

    # ---- C5: recommendation references a real Symbol ----
    da_syms = {str(col(d, "Symbol")).strip() for d in da_rows}
    check("C5: Recommendations ссылается на реальный Symbol из Data_Analysis",
          bool(rec_symbols) and any(s in da_syms for s in rec_symbols),
          f"rec={rec_symbols} da={sorted(da_syms)}")

    # ---------------- PPTX + processor ----------------
    import glob as globmod
    pptx_files = globmod.glob(os.path.join(agent_workspace, "*.pptx"))
    check("PowerPoint exists", len(pptx_files) >= 1, f"found {len(pptx_files)} pptx files")
    check("yf_risk_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "yf_risk_processor.py")))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0
    no_critical = len(CRITICAL_FAILS) == 0
    success = no_critical and accuracy >= 70
    msg = f"Passed {PASS_COUNT}/{total} checks (accuracy {accuracy:.1f}%), critical_fails={CRITICAL_FAILS}"
    return success, msg, no_critical


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message, no_critical = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    if not no_critical:
        print(f"FAIL: провалены критичные чеки: {CRITICAL_FAILS}")
        sys.exit(1)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
