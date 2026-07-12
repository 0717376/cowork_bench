"""Evaluation for terminal-canvas-teamly-study-gcal-excel.

The agent must:
  1. Build Study_Plan_Report.xlsx with three sheets (Course_Analysis,
     Weekly_Schedule, Priority_Matrix).
  2. Create a "Student Study Planner" page in the Teamly knowledge base whose
     body lists the 5 courses with their numeric Weekly Hours.
  3. Schedule 5 weekday study-session events (Mon-Fri of 2026-03-09..03-13).
  4. Produce a study_planner.py script.

CRITICAL_CHECKS (semantic) — any failure => overall FAIL regardless of
accuracy. The central deliverable is the computed Weekly_Hours
(3 + 0.5*Assignments + 1*Quizzes) and the derived Priority
(High>8 / Medium 5-8 inclusive / Low<5). These are verified INTRA-ROW
(each row's own Assignments/Quizzes/Weekly_Hours/Priority cells), so the
checks stay honest even though Canvas course data is live/volatile and the
exact enrollments / top-5 mapping cannot be hardcoded.

Otherwise pass threshold: accuracy >= 70%.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Course names are read live from Canvas and carry a "(Term Year)" suffix in
# the report (e.g. "Основы финансов (Осень 2014)"). We match on the base
# name substring only; the trailing 'finance' fallback keeps the structural
# (non-critical) name check lenient. NOT used for any critical value check.
EXPECTED_COURSE_HINTS = [
    "креативные вычисления",
    "основы финансов",
    "финанс",
    "биохими",
]

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical semantic checks: substance, not structure.
CRITICAL_CHECKS = {
    "Course_Analysis Weekly_Hours = 3 + 0.5*Assignments + 1*Quizzes (>=3 rows)",
    "Course_Analysis Priority matches threshold rule (>=3 rows)",
    "Exactly 5 weekday study sessions Mon-Fri of 2026-03-09..03-13",
    "Teamly 'Student Study Planner' page lists 5 courses with Weekly Hours",
    "Priority_Matrix Course_Count/Total_Weekly_Hours consistent with Course_Analysis",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        s = str(val).replace(",", ".").replace("$", "").strip()
        # strip a trailing comma->dot artefact like "16.5." gracefully
        return float(s)
    except Exception:
        return default


def priority_from_hours(h):
    if h is None:
        return None
    if h > 8:
        return "high"
    if h >= 5:  # 5..8 inclusive
        return "medium"
    return "low"


def num_variants(n):
    """Integer and (if whole) ".0"-style renderings; tolerate RU comma decimal."""
    out = set()
    if float(n) == int(n):
        out.add(str(int(n)))
    # decimal forms: dot and comma
    out.add(("%g" % n))
    out.add(("%g" % n).replace(".", ","))
    return out


def body_has_number(text, n):
    """True if any rendering of n appears as a standalone token in text."""
    for v in num_variants(n):
        pat = r"(?<![\d.,])" + re.escape(v) + r"(?![\d])"
        if re.search(pat, text):
            return True
    return False


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

CA_ROWS = []  # populated by check_excel for cross-sheet critical checks


def _find_sheet(sheets, *keywords):
    low = [s.lower() for s in sheets]
    for i, s in enumerate(low):
        if all(k in s for k in keywords):
            return i
    for i, s in enumerate(low):
        if any(k in s for k in keywords):
            return i
    return None


def check_excel(workspace):
    global CA_ROWS
    print("\n=== Check 1: Study_Plan_Report.xlsx ===")
    path = os.path.join(workspace, "Study_Plan_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    # --- Course_Analysis ---
    ca_idx = _find_sheet(sheets, "course", "analysis")
    if ca_idx is None:
        ca_idx = _find_sheet(sheets, "course") or _find_sheet(sheets, "analysis") or 0
    ws = wb[sheets[ca_idx]]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).lower() if c else "" for c in rows[0]] if rows else []
    data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
    check("Course_Analysis has >=5 course rows", len(data_rows) >= 5, f"Found {len(data_rows)}")

    all_text = " ".join(str(c) for r in rows for c in r if c is not None).lower()
    check("Contains Creative Computing course", "креативные вычисления" in all_text,
          f"Text: {all_text[:120]}")
    check("Contains Основы финансов course",
          "основы финансов" in all_text or "финанс" in all_text,
          f"Text: {all_text[:120]}")

    # Column index resolution by header name.
    def col(*subs):
        for i, h in enumerate(headers):
            if any(s in h for s in subs):
                return i
        return None

    i_assign = col("assignment")
    i_quiz = col("quiz")
    i_weekly = col("weekly", "hours")
    i_prio = col("priority")

    check("Has Weekly_Hours column", i_weekly is not None, f"Headers: {rows[0] if rows else None}")
    check("Has Priority column", i_prio is not None, f"Headers: {rows[0] if rows else None}")
    check("Has Assignments column", i_assign is not None, f"Headers: {rows[0] if rows else None}")
    check("Has Quizzes column", i_quiz is not None, f"Headers: {rows[0] if rows else None}")

    # Stash parsed rows for cross-sheet checks and critical checks.
    CA_ROWS = []
    if None not in (i_assign, i_quiz, i_weekly, i_prio):
        for r in data_rows:
            a = safe_float(r[i_assign]) if i_assign < len(r) else None
            q = safe_float(r[i_quiz]) if i_quiz < len(r) else None
            w = safe_float(r[i_weekly]) if i_weekly < len(r) else None
            p = str(r[i_prio]).strip().lower() if i_prio < len(r) and r[i_prio] is not None else None
            CA_ROWS.append({"a": a, "q": q, "w": w, "p": p})

    # CRITICAL: Weekly_Hours formula intra-row.
    formula_ok = 0
    formula_total = 0
    fdetails = []
    for row in CA_ROWS:
        if None in (row["a"], row["q"], row["w"]):
            continue
        formula_total += 1
        expected_w = 3 + 0.5 * row["a"] + 1 * row["q"]
        if abs(row["w"] - expected_w) < 0.01:
            formula_ok += 1
        else:
            fdetails.append(f"a={row['a']},q={row['q']},w={row['w']} (exp {expected_w})")
    check("Course_Analysis Weekly_Hours = 3 + 0.5*Assignments + 1*Quizzes (>=3 rows)",
          formula_ok >= 3 and formula_total >= 3,
          f"matched {formula_ok}/{formula_total}; bad: {fdetails}")

    # CRITICAL: Priority threshold rule intra-row.
    prio_ok = 0
    prio_total = 0
    pdetails = []
    for row in CA_ROWS:
        if row["w"] is None or not row["p"]:
            continue
        prio_total += 1
        if row["p"] == priority_from_hours(row["w"]):
            prio_ok += 1
        else:
            pdetails.append(f"w={row['w']} -> {row['p']} (exp {priority_from_hours(row['w'])})")
    check("Course_Analysis Priority matches threshold rule (>=3 rows)",
          prio_ok >= 3 and prio_total >= 3,
          f"matched {prio_ok}/{prio_total}; bad: {pdetails}")

    # --- Weekly_Schedule ---
    ws_idx = _find_sheet(sheets, "weekly", "schedule")
    if ws_idx is None:
        ws_idx = _find_sheet(sheets, "schedule") or _find_sheet(sheets, "weekly")
    if ws_idx is not None:
        ws2 = wb[sheets[ws_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c is not None and str(c).strip() for c in r)]
        check("Weekly_Schedule has 5 session rows", len(data_rows2) >= 5, f"Found {len(data_rows2)}")
        sched_text = " ".join(str(c) for r in rows2 for c in r if c is not None).lower()
        check("Weekly_Schedule mentions Study Session",
              "study session" in sched_text, f"Text: {sched_text[:120]}")

    # --- Priority_Matrix ---
    pm_idx = _find_sheet(sheets, "priority", "matrix")
    if pm_idx is None:
        pm_idx = _find_sheet(sheets, "matrix") or _find_sheet(sheets, "priority")
    if pm_idx is not None:
        ws3 = wb[sheets[pm_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        pm_headers = [str(c).lower() if c else "" for c in rows3[0]] if rows3 else []
        all_text3 = " ".join(str(c) for r in rows3 for c in r if c is not None).lower()
        check("Priority_Matrix has High level", "high" in all_text3, f"Text: {all_text3[:120]}")
        check("Priority_Matrix has Medium level", "medium" in all_text3, f"Text: {all_text3[:120]}")

        # CRITICAL: Priority_Matrix consistency with Course_Analysis.
        check_priority_matrix_consistency(rows3, pm_headers)


def check_priority_matrix_consistency(pm_rows, pm_headers):
    """Course_Count must sum to len(CA_ROWS) (5) and per-level Total_Weekly_Hours
    must match the sum of Course_Analysis Weekly_Hours for that level."""
    if not CA_ROWS:
        check("Priority_Matrix Course_Count/Total_Weekly_Hours consistent with Course_Analysis",
              False, "No Course_Analysis rows parsed")
        return

    def pmcol(*subs):
        for i, h in enumerate(pm_headers):
            if any(s in h for s in subs):
                return i
        return None

    i_level = pmcol("level", "priority")
    i_count = pmcol("count")
    i_total = pmcol("total")
    if None in (i_level, i_count, i_total):
        check("Priority_Matrix Course_Count/Total_Weekly_Hours consistent with Course_Analysis",
              False, f"Missing PM columns headers={pm_headers}")
        return

    # Expected aggregates from Course_Analysis (intra-data, not hardcoded).
    exp = {"high": [0, 0.0], "medium": [0, 0.0], "low": [0, 0.0]}
    for row in CA_ROWS:
        lvl = row["p"]
        if lvl in exp and row["w"] is not None:
            exp[lvl][0] += 1
            exp[lvl][1] += row["w"]

    got = {}
    for r in pm_rows[1:]:
        if i_level >= len(r) or r[i_level] is None:
            continue
        lvl = str(r[i_level]).strip().lower()
        if lvl not in ("high", "medium", "low"):
            continue
        cnt = safe_float(r[i_count]) if i_count < len(r) else None
        tot = safe_float(r[i_total]) if i_total < len(r) else None
        got[lvl] = (cnt, tot)

    total_count = sum(v[0] for v in got.values() if v[0] is not None)
    matches = 0
    details = []
    for lvl in ("high", "medium", "low"):
        e_cnt, e_tot = exp[lvl]
        g = got.get(lvl)
        if g is None:
            details.append(f"{lvl}: missing row")
            continue
        cnt_ok = g[0] is not None and abs(g[0] - e_cnt) < 0.01
        tot_ok = g[1] is not None and abs(g[1] - e_tot) < 0.01
        if cnt_ok and tot_ok:
            matches += 1
        else:
            details.append(f"{lvl}: got={g} exp=({e_cnt},{e_tot})")

    ok = matches == 3 and abs(total_count - len(CA_ROWS)) < 0.01
    check("Priority_Matrix Course_Count/Total_Weekly_Hours consistent with Course_Analysis",
          ok, f"matched {matches}/3, counts sum={total_count} vs {len(CA_ROWS)}; {details}")


# ---------------------------------------------------------------------------
# Teamly
# ---------------------------------------------------------------------------

def check_teamly():
    print("\n=== Check 2: Teamly Student Study Planner page ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT id, title, COALESCE(body, '')
        FROM teamly.pages
        WHERE title ILIKE '%student study planner%'
           OR title ILIKE '%study planner%'
           OR (title ILIKE '%study%' AND title ILIKE '%plan%')
    """)
    pages = cur.fetchall()
    check("Study Planner page exists in Teamly", len(pages) > 0,
          "No matching teamly.pages title found")

    # Use the page with the richest body.
    body = ""
    if pages:
        body = max((p[2] or "" for p in pages), key=len)
    body_lower = body.lower()

    # Structural: at least 3 of the named course hints appear.
    name_hits = sum(1 for h in EXPECTED_COURSE_HINTS[:3] if h in body_lower)
    check("Page body mentions known course names", name_hits >= 2,
          f"matched {name_hits} of named-course hints in body")

    # CRITICAL: page lists 5 courses WITH their numeric Weekly Hours.
    # We verify against Course_Analysis (CA_ROWS) Weekly_Hours values: every
    # weekly-hours value present in the report must also appear in the page
    # body, and there must be >=5 course-bearing data points.
    weekly_values = [row["w"] for row in CA_ROWS if row["w"] is not None]
    if len(weekly_values) >= 5:
        present = sum(1 for w in weekly_values if body_has_number(body, w))
        check("Teamly 'Student Study Planner' page lists 5 courses with Weekly Hours",
              present >= 5,
              f"{present}/{len(weekly_values)} Weekly_Hours values found in page body")
    else:
        # Fall back: cannot cross-check without Course_Analysis; require the
        # body to carry at least 5 numeric tokens plus course names.
        nums = re.findall(r"\d+(?:[.,]\d+)?", body)
        check("Teamly 'Student Study Planner' page lists 5 courses with Weekly Hours",
              name_hits >= 2 and len(nums) >= 5,
              f"name_hits={name_hits}, numeric tokens={len(nums)} (no CA cross-check)")

    cur.close()
    conn.close()


# ---------------------------------------------------------------------------
# Google Calendar
# ---------------------------------------------------------------------------

def check_gcal():
    print("\n=== Check 3: Google Calendar Study Sessions ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # All study-session events.
    cur.execute("""
        SELECT summary, start_datetime,
               EXTRACT(DOW FROM start_datetime) AS dow,
               start_datetime::date AS d
        FROM gcal.events
        WHERE lower(summary) LIKE '%%study session%%'
           OR lower(summary) LIKE '%%study%%'
    """)
    rows = cur.fetchall()
    check("At least 5 study session events", len(rows) >= 5, f"Found {len(rows)} events")

    summaries = " ".join(str(r[0]) for r in rows).lower()
    check("Events mention Study Session", "study session" in summaries or "study" in summaries,
          f"Summaries: {summaries[:160]}")
    check("Events mention a known course name",
          any(h in summaries for h in EXPECTED_COURSE_HINTS),
          f"Summaries: {summaries[:160]}")

    # CRITICAL: exactly 5 weekday sessions, one per weekday of the study week,
    # none on a weekend.
    week_days = {"2026-03-09", "2026-03-10", "2026-03-11", "2026-03-12", "2026-03-13"}
    in_week = [r for r in rows if str(r[3]) in week_days]
    weekend = [r for r in rows if int(r[2]) in (0, 6)]
    distinct_days = {str(r[3]) for r in in_week}
    ok = (len(in_week) == 5 and len(weekend) == 0 and len(distinct_days) == 5)
    check("Exactly 5 weekday study sessions Mon-Fri of 2026-03-09..03-13",
          ok,
          f"in_week={len(in_week)}, distinct_days={len(distinct_days)}, weekend={len(weekend)}")

    cur.close()
    conn.close()


# ---------------------------------------------------------------------------
# Script
# ---------------------------------------------------------------------------

def check_script(workspace):
    print("\n=== Check 4: study_planner.py ===")
    path = os.path.join(workspace, "study_planner.py")
    check("study_planner.py exists", os.path.exists(path))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()
    check_gcal()
    check_script(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
        "critical_failed": critical_failed,
        "success": (not critical_failed) and accuracy >= 70,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
