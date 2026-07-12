"""Evaluation for insales-shipping-analysis (InSales / RU)."""
import argparse
import json
import os
import sys
import openpyxl

try:
    import psycopg2
except Exception:
    psycopg2 = None

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432,
      "dbname": "cowork_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def num_exact(a, b):
    """Exact integer/numeric equality (no tolerance)."""
    try:
        return float(a) == float(b)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def summary_lookup(rows, keys):
    """Header/orientation-tolerant {metric_lower: value} map for a Summary sheet.

    Scans ALL cells; a metric's value is the cell to its right, or the cell
    below when the right neighbour is itself another metric key (horizontal
    layout). A 'Metric/Value' header never collides with metric keys.
    """
    keys = {str(k).strip().lower() for k in keys}
    out = {}
    rows = rows or []

    def _is_key(v):
        return v is not None and str(v).strip().lower() in keys

    for r, row in enumerate(rows):
        for c, cell in enumerate(row or []):
            if cell is None:
                continue
            k = str(cell).strip().lower()
            if k not in keys or k in out:
                continue
            right = row[c + 1] if c + 1 < len(row) else None
            below = None
            if r + 1 < len(rows) and rows[r + 1] and c < len(rows[r + 1]):
                below = rows[r + 1][c]
            if right is not None and not _is_key(right):
                out[k] = right
            elif below is not None and not _is_key(below):
                out[k] = below
    return out


def recipients_of(to_addr):
    """Normalize email.messages.to_addr (str / json list / list) to lowercase strings."""
    if to_addr is None:
        return []
    if isinstance(to_addr, list):
        return [str(r).strip().lower() for r in to_addr]
    if isinstance(to_addr, str):
        s = to_addr.strip()
        try:
            parsed = json.loads(s)
            if isinstance(parsed, list):
                return [str(r).strip().lower() for r in parsed]
        except Exception:
            pass
        return [p.strip().lower() for p in s.replace(";", ",").split(",") if p.strip()]
    return [str(to_addr).strip().lower()]


def find_shipping_email():
    """Return list of (subject, to_recipients, body) for matching sent emails."""
    if psycopg2 is None:
        return None
    try:
        conn = psycopg2.connect(**DB)
    except Exception:
        return None
    out = []
    try:
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        for subject, to_addr, body in cur.fetchall():
            out.append((subject or "", recipients_of(to_addr), body or ""))
        cur.close()
    finally:
        conn.close()
    return out


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Shipping_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Shipping_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_failures = []

    def critical(name, condition, detail=""):
        if condition:
            print(f"  [CRITICAL PASS] {name}")
        else:
            critical_failures.append(name)
            print(f"  [CRITICAL FAIL] {name}: {detail}")

    # ------------------------------------------------------------------
    # Build groundtruth zone map dynamically (do NOT hand-code RU literals;
    # zone names are russified centrally and the GT xlsx is map-patched).
    # ------------------------------------------------------------------
    g_zone_rows = load_sheet_rows(gt_wb, "Shipping Zones")
    gt_zones = {}  # lower zone name -> method count
    if g_zone_rows:
        for gr in g_zone_rows[1:]:
            if gr and gr[0] is not None:
                gt_zones[str(gr[0]).strip().lower()] = gr[1]

    g_summary_rows = load_sheet_rows(gt_wb, "Summary")
    gt_summary = {}  # lower metric -> value
    if g_summary_rows:
        for gr in g_summary_rows[1:]:
            if gr and gr[0] is not None:
                gt_summary[str(gr[0]).strip().lower()] = gr[1]

    # ------------------------------------------------------------------
    # Structural (non-critical) checks — accuracy gate, tol=1 retained.
    # ------------------------------------------------------------------
    print("  Checking Shipping Zones...")
    a_rows = load_sheet_rows(agent_wb, "Shipping Zones")
    g_rows = g_zone_rows
    a_zone_data = []
    if a_rows is None:
        all_errors.append("Sheet 'Shipping Zones' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Shipping Zones' not found in groundtruth")
    else:
        errors = []
        a_zone_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = {}
        for row in a_zone_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Shipping_Methods: {a_row[1]} vs {g_row[1]} (tol=1)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    print("  Checking Summary...")
    a_sum_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = g_summary_rows
    if a_sum_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        a_lookup = summary_lookup(a_sum_rows, gt_summary.keys())
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            if key not in a_lookup:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(g_row) > 1:
                if not num_close(a_lookup[key], g_row[1], 1.0):
                    errors.append(f"{key}.Value: {a_lookup[key]} vs {g_row[1]} (tol=1.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # ------------------------------------------------------------------
    # CRITICAL checks — any failure => FAIL regardless of accuracy.
    # All expected values are read dynamically from groundtruth (no hand
    # literals); integer counts use tol=0.
    # ------------------------------------------------------------------
    print("\n  === CRITICAL CHECKS ===")

    a_zone_lookup = {}
    for row in a_zone_data:
        if row and row[0] is not None:
            a_zone_lookup[str(row[0]).strip().lower()] = row

    # 1. Exact zone set + exact method counts (tol=0).
    zone_ok = True
    zone_detail = []
    for zkey, gcount in gt_zones.items():
        arow = a_zone_lookup.get(zkey)
        if arow is None:
            zone_ok = False
            zone_detail.append(f"missing zone '{zkey}'")
        elif len(arow) < 2 or not num_exact(arow[1], gcount):
            zone_ok = False
            zone_detail.append(f"'{zkey}' methods {arow[1] if len(arow) > 1 else None} != {gcount}")
    critical(f"Shipping Zones has exactly {len(gt_zones)} zones with exact method counts",
             zone_ok, "; ".join(zone_detail))

    # 2. No spurious/extra zone rows beyond groundtruth set.
    extra = [k for k in a_zone_lookup if k not in gt_zones]
    critical("No extra/spurious zone rows", len(extra) == 0, f"extra zones: {extra}")

    # 3. Summary exact integer match (tol=0) for Total_Zones / Total_Methods.
    a_sum_lookup = summary_lookup(a_sum_rows, gt_summary.keys())
    sum_ok = True
    sum_detail = []
    for mkey, gval in gt_summary.items():
        aval = a_sum_lookup.get(mkey)
        if aval is None or not num_exact(aval, gval):
            sum_ok = False
            sum_detail.append(f"{mkey}={aval} != {gval}")
    critical("Summary Total_Zones/Total_Methods exact match (tol=0)",
             sum_ok, "; ".join(sum_detail))

    # 4. Shipping Zones rows sorted by Zone name (as required).
    a_zone_names = [str(r[0]).strip() for r in a_zone_data if r and r[0] is not None]
    sorted_ok = a_zone_names == sorted(a_zone_names, key=lambda s: s.lower())
    critical("Shipping Zones sorted by Zone name", sorted_ok, f"order: {a_zone_names}")

    # 5. Email sent to logistics@shop.com with required subject, listing zones.
    emails = find_shipping_email()
    if emails is None:
        critical("Email sent to logistics@shop.com with subject 'Shipping Zone Configuration Review'",
                 False, "could not query email.messages (psycopg2 unavailable / DB error)")
    else:
        matched = None
        for subject, recips, body in emails:
            subj_ok = "shipping zone configuration review" in subject.strip().lower()
            to_ok = any("logistics@shop.com" == r or "logistics@shop.com" in r for r in recips)
            if subj_ok and to_ok:
                matched = (subject, recips, body)
                break
        critical("Email sent to logistics@shop.com with subject 'Shipping Zone Configuration Review'",
                 matched is not None,
                 f"sent emails: {[(s, r) for s, r, _ in emails][:5]}")
        if matched:
            body_l = (matched[2] or "").lower()
            zones_listed = sum(1 for zkey in gt_zones if zkey in body_l)
            critical("Email body lists the shipping zones",
                     len(gt_zones) > 0 and zones_listed >= len(gt_zones),
                     f"{zones_listed}/{len(gt_zones)} zone names found in body")

    # ------------------------------------------------------------------
    # Gate: critical first, then accuracy.
    # ------------------------------------------------------------------
    if critical_failures:
        print(f"\n=== RESULT: FAIL (critical) — {len(critical_failures)} critical check(s) failed ===")
        for c in critical_failures:
            print(f"  CRITICAL: {c}")
        sys.exit(1)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)

    print("\n=== RESULT: PASS ===")
    sys.exit(0)


if __name__ == "__main__":
    main()
