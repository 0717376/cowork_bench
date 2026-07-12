"""
Evaluation for yt-top-videos-excel-gcal task.

Checks:
1. Top_Videos_Watchlist.xlsx exists with Watchlist and Summary sheets
2. Watchlist has 10 rows, correct columns, correct top video
3. Summary has correct combined view count
4. GCal has 10 learning sessions on Tuesdays April-June 2026
5. Email sent to team@company.com with correct subject

CRITICAL_CHECKS (semantic) gate before the accuracy>=70 gate: any failure => FAIL,
regardless of overall accuracy. These ensure a correct RU agent passes while a
non-doer (stub file, no events, no email) fails.
"""
import json
import os
import sys
from argparse import ArgumentParser
from datetime import date, datetime, timedelta

import psycopg2
import openpyxl

def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Ground-truth constants (verified against youtube.* read-only data).
TOP_VIDEO_ID = "Nl7aCUsWykg"          # Rank 1, most-viewed Fireship video (DeepSeek R1)
TOP_VIDEO_VIEWS = 3878491
COMBINED_VIEWS = 25315767             # sum of top-10 view counts
EXPECTED_SESSIONS = 10
FIRST_SESSION = date(2026, 4, 7)      # Tuesday

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "[CRITICAL] " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILURES.append(name)


def _to_float(x):
    try:
        return float(str(x).replace(",", "").strip())
    except Exception:
        return None


def check_excel(agent_workspace):
    print("\n=== Check 1: Top_Videos_Watchlist.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Top_Videos_Watchlist.xlsx")
    if not os.path.exists(xlsx_path):
        record("Top_Videos_Watchlist.xlsx exists", False, f"Not found at {xlsx_path}", critical=True)
        return
    record("Top_Videos_Watchlist.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e), critical=True)
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # ---- Watchlist sheet ----
    if "watchlist" not in sheet_names_lower:
        record("Watchlist sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Watchlist sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("watchlist")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Watchlist has 10 rows", len(data_rows) == 10, f"Found {len(data_rows)}")

        # Headers
        headers = [str(c).lower() if c else "" for c in rows[0]] if rows else []
        has_rank = any("rank" in h for h in headers)
        has_vid = any("video_id" in h or h == "video_id" for h in headers)
        has_dur = any("duration" in h or "dur" in h for h in headers)
        has_like_rate = any("like_rate" in h or "rate" in h for h in headers)
        record("Has Rank, Video_ID, Duration_Min, Like_Rate columns",
               has_rank and has_vid and has_dur and has_like_rate, f"Headers: {rows[0] if rows else None}")

        # CRITICAL: the Rank-1 (first/top) data row must specifically be the
        # most-viewed video (DeepSeek R1, Nl7aCUsWykg) with its real view count.
        # This is a row-specific check (not a workbook-wide substring), so a
        # stub or wrongly-sorted file fails.
        top_ok = False
        top_detail = "No data rows"
        if data_rows:
            first_row = data_rows[0]
            row_text = " ".join(str(c) for c in first_row if c is not None).lower()
            id_ok = TOP_VIDEO_ID.lower() in row_text
            # View_Count cell on the rank-1 row should equal the GT (tolerant).
            view_ok = any(
                (_to_float(c) is not None and num_close(_to_float(c), TOP_VIDEO_VIEWS))
                for c in first_row
            )
            top_ok = id_ok and view_ok
            top_detail = f"First row: {first_row[:5]}"
        record("Rank-1 row is most-viewed video (Nl7aCUsWykg, ~3,878,491 views)",
               top_ok, top_detail, critical=True)

    # ---- Summary sheet ----
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Summary sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows2 = list(ws2.iter_rows(values_only=True))

        # Build a Label -> Value map from the two-column Summary sheet.
        summ = {}
        for r in rows2:
            if r and len(r) >= 2 and r[0] is not None:
                summ[str(r[0]).strip().lower()] = r[1]

        # CRITICAL: Combined_Views must equal the real GT sum (exact value).
        cv = summ.get("combined_views")
        cv_ok = cv is not None and _to_float(cv) is not None and int(round(_to_float(cv))) == COMBINED_VIEWS
        record("Summary Combined_Views == 25315767", cv_ok,
               f"Combined_Views={cv}", critical=True)

        ts = summ.get("total_selected")
        ts_ok = ts is not None and _to_float(ts) is not None and int(round(_to_float(ts))) == 10
        record("Summary Total_Selected == 10", ts_ok, f"Total_Selected={ts}")


def check_gcal():
    print("\n=== Check 2: GCal learning sessions ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE summary ILIKE '%team learning%'
        AND start_datetime >= '2026-04-07' AND start_datetime < '2026-07-01'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    # CRITICAL: exactly the 10 sessions must exist (a non-doer creates none).
    record("Exactly 10 'Team Learning' events scheduled", len(events) == EXPECTED_SESSIONS,
           f"Found {len(events)} events", critical=True)

    if not events:
        return

    # First event on Tuesday 2026-04-07.
    first_start = events[0][1]
    first_date = first_start.date() if hasattr(first_start, "date") else first_start
    record("First session on 2026-04-07 (Tuesday)", first_date == FIRST_SESSION,
           f"First event date: {first_date}")

    # Sessions are 1 hour long.
    s0, e0 = events[0][1], events[0][2]
    if s0 and e0:
        dur_h = (e0 - s0).total_seconds() / 3600
        record("Sessions are 1 hour long", abs(dur_h - 1.0) < 0.1, f"Duration: {dur_h:.2f}h")

    # All on Tuesdays.
    all_tuesdays = all(e[1].weekday() == 1 for e in events if e[1])
    record("All sessions scheduled on Tuesdays", all_tuesdays,
           f"Weekdays: {[e[1].weekday() for e in events if e[1]]}")

    # CRITICAL: dates are consecutive Tuesdays starting 2026-04-07
    # (April 7, 14, 21, 28, ...). Guards against dumping 10 events on one day.
    if len(events) == EXPECTED_SESSIONS:
        expected = {FIRST_SESSION + timedelta(weeks=i) for i in range(EXPECTED_SESSIONS)}
        got = {e[1].date() if hasattr(e[1], "date") else e[1] for e in events if e[1]}
        record("Sessions on 10 consecutive Tuesdays from 2026-04-07",
               got == expected, f"Got dates: {sorted(str(d) for d in got)}", critical=True)


def check_email():
    print("\n=== Check 3: Email sent ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT m.to_addr, m.subject FROM email.messages m
        JOIN email.sent_log sl ON sl.message_id = m.id
        WHERE m.to_addr::text ILIKE %s
        ORDER BY sl.sent_at DESC LIMIT 5
    """, ("%team@company.com%",))
    emails = cur.fetchall()

    if not emails:
        cur.execute("""
            SELECT to_addr, subject FROM email.messages
            WHERE to_addr::text ILIKE %s
            ORDER BY date DESC LIMIT 5
        """, ("%team@company.com%",))
        emails = cur.fetchall()

    cur.close()
    conn.close()

    # CRITICAL: a confirmation email to team@company.com with the right subject
    # must have been sent (a non-doer sends none).
    sent = len(emails) >= 1
    subj_ok = False
    if sent:
        subject = str(emails[0][1]).lower() if emails[0][1] else ""
        subj_ok = "learning" in subject and "scheduled" in subject
    record("Email to team@company.com with subject 'Learning Sessions Scheduled'",
           sent and subj_ok, f"Found: {emails}", critical=True)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILURES,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # CRITICAL gate: any critical failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILURES:
        print(f"\nFAIL: {len(CRITICAL_FAILURES)} critical check(s) failed: {CRITICAL_FAILURES}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
