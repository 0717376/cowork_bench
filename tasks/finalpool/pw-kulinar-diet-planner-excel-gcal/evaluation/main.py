"""Evaluation script for pw-kulinar-diet-planner-excel-gcal.

Структурные проверки (НЕ критические):
  - Diet_Planner_Report.xlsx существует и читается
  - Лист Data_Analysis: >= 5 строк, колонки Recipe/Category/Calories/Protein_g/Meets_Guidelines
  - Лист Metrics: >= 4 строк, колонки Metric/Value
  - Лист Recommendations: >= 2 строк, колонки Priority/Action
  - cook_diet_processor.py существует
  - В календаре есть хотя бы одно событие-обзор

CRITICAL (любой провал => немедленный FAIL до порога accuracy):
  C1. Событие 'Analysis Review' существует и начинается 2026-03-14 14:00 UTC, длится до 15:00 UTC
  C2. cook_diet_results.json существует и содержит результат анализа (расхождения/сравнение), не пустой
  C3. Колонка Meets_Guidelines согласована с правилом <= 800 ккал на порцию (не захардкожена Yes)
  C4. Метрики на листе Metrics численно согласованы с данными Data_Analysis
       (Total_Recipes / Avg_Calories / Avg_Protein пересчитываются и сверяются с допуском)
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:300] if detail else ""
        print(f"  [FAIL] {tag}{name}: {detail_str}")
        if critical:
            CRITICAL_FAILED.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    excel_path = os.path.join(agent_workspace, "Diet_Planner_Report.xlsx")
    check("Diet_Planner_Report.xlsx exists", os.path.exists(excel_path))

    data_rows_parsed = []   # list of dicts: recipe, category, calories, protein, meets
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers_raw = [c.value for c in ws[1]]
            headers = [str(c).strip().lower() if c is not None else "" for c in headers_raw]
            data_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                         if any(c is not None and str(c).strip() for c in r)]
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            col_idx = {}
            for expected_col in ['Recipe', 'Category', 'Calories', 'Protein_g', 'Meets_Guidelines']:
                present = expected_col.lower() in headers
                check(f"Data_Analysis has {expected_col} column", present, f"headers: {headers[:8]}")
                if present:
                    col_idx[expected_col] = headers.index(expected_col.lower())

            if all(k in col_idx for k in ['Recipe', 'Calories', 'Protein_g', 'Meets_Guidelines']):
                for r in data_rows:
                    cal = safe_float(r[col_idx['Calories']])
                    prot = safe_float(r[col_idx['Protein_g']])
                    meets = str(r[col_idx['Meets_Guidelines']]).strip().lower() if r[col_idx['Meets_Guidelines']] is not None else ""
                    data_rows_parsed.append({
                        "recipe": r[col_idx['Recipe']],
                        "calories": cal, "protein": prot, "meets": meets,
                    })

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            m_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None and str(c).strip() for c in r)]
            check("Metrics has >= 4 rows", len(m_rows) >= 4, f"got {len(m_rows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in m_rows:
                if r and r[0] is not None:
                    metrics_map[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            rec_rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                        if any(c is not None and str(c).strip() for c in r)]
            check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")
            for expected_col in ['Priority', 'Action']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # -------- CRITICAL C3: Meets_Guidelines consistent with <= 800 kcal rule --------
        if data_rows_parsed and all(d["calories"] is not None for d in data_rows_parsed):
            yes_words = {"yes", "да", "true", "1"}
            no_words = {"no", "нет", "false", "0"}
            consistent = 0
            classifiable = 0
            for d in data_rows_parsed:
                expected_meets = d["calories"] <= 800
                m = d["meets"]
                if m in yes_words:
                    actual = True
                elif m in no_words:
                    actual = False
                else:
                    continue
                classifiable += 1
                if actual == expected_meets:
                    consistent += 1
            # хотя бы 80% строк должны соответствовать правилу, и не все хардкод-Yes без единого No
            distinct_vals = {d["meets"] for d in data_rows_parsed}
            not_all_same = not (len(distinct_vals) == 1 and any(d["calories"] > 800 for d in data_rows_parsed))
            ratio = consistent / classifiable if classifiable else 0
            check("Meets_Guidelines применяет правило <= 800 ккал (не захардкожено)",
                  classifiable >= 5 and ratio >= 0.8 and not_all_same,
                  f"consistent={consistent}/{classifiable}, distinct={distinct_vals}",
                  critical=True)
        else:
            check("Meets_Guidelines применяет правило <= 800 ккал (не захардкожено)",
                  False, "не удалось распарсить Calories/Meets_Guidelines", critical=True)

        # -------- CRITICAL C4: Metrics numerically consistent with Data_Analysis --------
        if data_rows_parsed and metrics_map:
            cals = [d["calories"] for d in data_rows_parsed if d["calories"] is not None]
            prots = [d["protein"] for d in data_rows_parsed if d["protein"] is not None]
            n = len(data_rows_parsed)
            ok_parts = []
            detail_parts = []

            def find_metric(*keys):
                for k, v in metrics_map.items():
                    if all(part in k for part in keys):
                        return safe_float(v)
                return None

            total_v = find_metric("total")
            if total_v is not None:
                ok = abs(total_v - n) < 0.5
                ok_parts.append(ok); detail_parts.append(f"total={total_v} vs {n}:{ok}")

            if cals:
                avg_cal_v = find_metric("avg", "cal") or find_metric("avg", "калор") or find_metric("сред", "калор")
                if avg_cal_v is not None:
                    exp = sum(cals) / len(cals)
                    ok = abs(avg_cal_v - exp) <= max(1.0, exp * 0.05)
                    ok_parts.append(ok); detail_parts.append(f"avg_cal={avg_cal_v} vs {exp:.1f}:{ok}")

            if prots:
                avg_prot_v = find_metric("avg", "prot") or find_metric("avg", "белк") or find_metric("сред", "белк")
                if avg_prot_v is not None:
                    exp = sum(prots) / len(prots)
                    ok = abs(avg_prot_v - exp) <= max(0.5, exp * 0.05)
                    ok_parts.append(ok); detail_parts.append(f"avg_prot={avg_prot_v} vs {exp:.1f}:{ok}")

            check("Metrics численно согласованы с Data_Analysis",
                  len(ok_parts) >= 2 and all(ok_parts),
                  f"checked={len(ok_parts)}; {detail_parts}", critical=True)
        else:
            check("Metrics численно согласованы с Data_Analysis",
                  False, "нет данных для сверки метрик", critical=True)

    # -------- CRITICAL C2: cook_diet_results.json exists and has analysis content --------
    results_path = os.path.join(agent_workspace, "cook_diet_results.json")
    if os.path.exists(results_path):
        try:
            with open(results_path, encoding="utf-8") as f:
                raw = f.read()
            parsed = json.loads(raw)
            non_empty = bool(parsed) and (len(raw.strip()) > 2)
            low = raw.lower()
            has_analysis = any(kw in low for kw in
                               ["gap", "расхожд", "calor", "калор", "protein", "белк",
                                "recipe", "блюд", "benchmark", "норм", "meets", "guideline"])
            check("cook_diet_results.json содержит результат анализа",
                  non_empty and has_analysis,
                  f"non_empty={non_empty}, has_analysis={has_analysis}", critical=True)
        except Exception as e:
            check("cook_diet_results.json содержит результат анализа", False, str(e), critical=True)
    else:
        check("cook_diet_results.json содержит результат анализа", False,
              "файл отсутствует", critical=True)

    check("cook_diet_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "cook_diet_processor.py")))

    # -------- Calendar --------
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Структурная (нестрогая) проверка: вообще есть событие-обзор
        cur.execute("SELECT summary FROM gcal.events WHERE summary ILIKE %s OR summary ILIKE %s",
                    ('%review%', '%обзор%'))
        events = cur.fetchall()
        check("Событие-обзор создано", len(events) >= 1, f"found {len(events)} events")

        # CRITICAL C1: точный заголовок 'Analysis Review' + дата/время 2026-03-14 14:00-15:00 UTC
        cur.execute("""
            SELECT summary, start_datetime, end_datetime
              FROM gcal.events
             WHERE summary ILIKE %s
               AND start_datetime::date = %s
        """, ('%analysis review%', '2026-03-14'))
        rows = cur.fetchall()
        matched = False
        detail = []
        for summ, sdt, edt in rows:
            sh = sdt.hour if sdt is not None else None
            sm = sdt.minute if sdt is not None else None
            eh = edt.hour if edt is not None else None
            detail.append(f"{summ} {sdt}->{edt}")
            if sh == 14 and sm == 0 and (eh == 15 or eh is None):
                matched = True
                break
        check("Событие 'Analysis Review' на 2026-03-14 14:00-15:00 UTC",
              matched, f"candidates: {detail}", critical=True)
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e), critical=True)

    return FAIL_COUNT == 0, f"Passed {PASS_COUNT}/{PASS_COUNT + FAIL_COUNT} checks"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_failed": CRITICAL_FAILED}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"CRITICAL FAILED: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
