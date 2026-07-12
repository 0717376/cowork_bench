"""Evaluation for terminal-sf-arxiv-excel-word-teamly.
Checks:
1. Retention_Strategy.xlsx with 3 sheets (Department_Analysis, Research_Summary, Action_Plan)
2. Retention_Strategy_Report.docx
3. Teamly knowledge base page "Retention Action Items" with 7 departments
4. flight_risk_analysis.py and synthesis.py scripts exist
5. flight_risk_analysis.json and synthesis.json outputs exist

ClickHouse (sf_data) data VALUES are russified centrally by a deterministic map;
this eval treats the English UPPERCASE column/identifier names and the logical
DB name HR_ANALYTICS as fixed. The DEPARTMENT column VALUES are russified too, so
the live DB returns Russian department names; this eval translates them back to
the English identifiers (RU_TO_EN_DEPARTMENT) before aggregating, then keys all
checks on the English names.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

DEPARTMENTS = ["Engineering", "Finance", "HR", "Operations", "R&D", "Sales", "Support"]

# ClickHouse (sf_data) russifies the DEPARTMENT column VALUES via the central
# deterministic map (scripts/clickhouse_relabel_map.py: DEPARTMENTS). The live
# DB therefore returns Russian department names, while this eval keys everything
# on the English identifiers. This is the inverse of that map (RU -> EN); the 7
# RU values are distinct, so the inverse is unambiguous. We use it to translate
# the GROUP BY result keys back to English before aggregating.
RU_TO_EN_DEPARTMENT = {
    "Инженерия": "Engineering", "Финансы": "Finance", "Кадры": "HR",
    "Операции": "Operations", "НИОКР": "R&D", "Продажи": "Sales", "Поддержка": "Support",
}


def canonical_priority(pct):
    """The SINGLE canonical threshold rule shared by task.md, methodology.md,
    the Excel Action_Plan check and the Teamly check:
        > 8.3%          -> High
        7.9% .. 8.3%    -> Medium  (inclusive of both bounds)
        < 7.9%          -> Low
    """
    if pct is None:
        return None
    if pct > 8.3:
        return "High"
    if pct >= 7.9:
        return "Medium"
    return "Low"


# Hardcoded fallback flight risk data (sat<=4 AND perf>=4); priority derived
# from the canonical rule above (used only if the DB is unreachable).
_FALLBACK_EXPECTED_DATA = {
    "Engineering": {"headcount": 7096, "flight_risk": 566, "pct": 7.98, "priority": "Medium"},
    "Finance":     {"headcount": 7148, "flight_risk": 598, "pct": 8.37, "priority": "High"},
    "HR":          {"headcount": 7077, "flight_risk": 594, "pct": 8.39, "priority": "High"},
    "Operations":  {"headcount": 7120, "flight_risk": 564, "pct": 7.92, "priority": "Medium"},
    "R&D":         {"headcount": 7083, "flight_risk": 576, "pct": 8.13, "priority": "Medium"},
    "Sales":       {"headcount": 7232, "flight_risk": 596, "pct": 8.24, "priority": "Medium"},
    "Support":     {"headcount": 7244, "flight_risk": 537, "pct": 7.41, "priority": "Low"},
}


def _get_expected_data_from_db():
    """Query sf_data schema to compute department headcounts and flight risk
    counts dynamically, then apply the canonical priority rule."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        try:
            # Get headcount per department
            cur.execute("""
                SELECT "DEPARTMENT", COUNT(*) as headcount
                FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
                GROUP BY "DEPARTMENT"
            """)
            # Translate russified DEPARTMENT values back to the English names this
            # eval keys on; unmapped values (already English) pass through.
            headcounts = {RU_TO_EN_DEPARTMENT.get(r[0], r[0]): r[1]
                          for r in cur.fetchall()}

            # Get flight risk count per department (satisfaction<=4 AND performance>=4)
            cur.execute("""
                SELECT "DEPARTMENT", COUNT(*) as flight_risk
                FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
                WHERE "JOB_SATISFACTION" <= 4 AND "PERFORMANCE_RATING" >= 4
                GROUP BY "DEPARTMENT"
            """)
            flight_risks = {RU_TO_EN_DEPARTMENT.get(r[0], r[0]): r[1]
                            for r in cur.fetchall()}

            result = {}
            for dept in DEPARTMENTS:
                hc = headcounts.get(dept, 0)
                fr = flight_risks.get(dept, 0)
                pct = round(fr / hc * 100, 2) if hc > 0 else 0
                result[dept] = {"headcount": hc, "flight_risk": fr,
                                "pct": pct, "priority": canonical_priority(pct)}
            return result
        finally:
            cur.close()
            conn.close()
    except Exception:
        return _FALLBACK_EXPECTED_DATA


EXPECTED_DATA = _get_expected_data_from_db()

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}: {str(detail)[:200]}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(workspace):
    print("\n=== Check 1: Retention_Strategy.xlsx ===")
    path = os.path.join(workspace, "Retention_Strategy.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    # Sheet 1: Department_Analysis
    da_idx = next((i for i, s in enumerate(sheets_lower) if "department" in s and "analysis" in s), 0)
    ws1 = wb[sheets[da_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Department_Analysis has 7 rows", len(data1) >= 7, f"Found {len(data1)}")

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        check("Has flight_risk_pct column",
              any("flight" in h and "pct" in h for h in headers) or any("risk" in h and "%" in h for h in headers) or any("flight_risk_pct" in h for h in headers),
              f"Headers: {rows1[0]}")
        check("Has avg_satisfaction column",
              any("satisfaction" in h for h in headers),
              f"Headers: {rows1[0]}")

    # Check actual values
    dept_col = next((i for i, h in enumerate(headers) if "department" in h or "dept" in h), 0) if rows1 else 0
    hc_col = next((i for i, h in enumerate(headers) if "headcount" in h or "head_count" in h), 1) if rows1 else 1
    fr_col = next((i for i, h in enumerate(headers) if "flight_risk_count" in h or ("flight" in h and "count" in h)), 2) if rows1 else 2
    pct_col = next((i for i, h in enumerate(headers) if "pct" in h or "percent" in h), 3) if rows1 else 3

    found_depts = 0
    by_dept = {}
    for row in data1:
        dept_name = str(row[dept_col]).strip() if dept_col < len(row) and row[dept_col] else ""
        if dept_name in EXPECTED_DATA:
            by_dept[dept_name] = row
    for dept_name in DEPARTMENTS:
        row = by_dept.get(dept_name)
        exp = EXPECTED_DATA[dept_name]
        if row is None:
            check(f"{dept_name} headcount correct", False, "department row missing", critical=True)
            check(f"{dept_name} flight_risk_count correct", False, "department row missing", critical=True)
            continue
        found_depts += 1
        # CRITICAL: headcount must match live DB value within a tight tolerance
        # (no silent skip if the cell is empty/zero).
        hc_val = row[hc_col] if hc_col < len(row) else None
        check(f"{dept_name} headcount correct",
              num_close(hc_val, exp["headcount"], 2),
              f"Got {hc_val}, expected {exp['headcount']}", critical=True)
        # CRITICAL: flight_risk_count must match live DB value (sat<=4 AND perf>=4).
        fr_val = row[fr_col] if fr_col < len(row) else None
        check(f"{dept_name} flight_risk_count correct",
              num_close(fr_val, exp["flight_risk"], 2),
              f"Got {fr_val}, expected {exp['flight_risk']}", critical=True)
    check("All 7 departments found in Department_Analysis", found_depts >= 7, f"Found {found_depts}",
          critical=True)

    # Sheet 2: Research_Summary
    rs_idx = next((i for i, s in enumerate(sheets_lower) if "research" in s), 1)
    if rs_idx < len(sheets):
        ws2 = wb[sheets[rs_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Research_Summary has 3 rows", len(data2) >= 3, f"Found {len(data2)}")
        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            check("Has applicability_score column",
                  any("applicability" in h or "score" in h for h in headers2),
                  f"Headers: {rows2[0]}")
        # Check that the 3 relevant retention papers are included and the 2 noise
        # papers (autonomous-vehicle, quantum protein-folding) are excluded.
        # Paper titles stay English, so substring checks remain English.
        all_text2 = " ".join(str(c) for r in rows2 for c in r if c).lower()
        check("Contains retention-related paper",
              "retention" in all_text2 or "turnover" in all_text2 or "employee" in all_text2,
              critical=True)
        check("Excludes both noise papers (autonomous-vehicle, protein folding)",
              "autonomous vehicle" not in all_text2 and "protein folding" not in all_text2
              and "autonomous-vehicle" not in all_text2,
              "Noise paper included in research summary", critical=True)

    # Sheet 3: Action_Plan
    ap_idx = next((i for i, s in enumerate(sheets_lower) if "action" in s or "plan" in s), 2)
    if ap_idx < len(sheets):
        ws3 = wb[sheets[ap_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Action_Plan has 7 rows", len(data3) >= 7, f"Found {len(data3)}")

        if rows3:
            headers3 = [str(c).lower() if c else "" for c in rows3[0]]
            pri_col = next((i for i, h in enumerate(headers3) if "priority" in h), None)
            dept_col3 = next((i for i, h in enumerate(headers3) if "department" in h or "dept" in h), 0)

            if pri_col is not None:
                ap_by_dept = {}
                for row in data3:
                    dn = str(row[dept_col3]).strip() if dept_col3 < len(row) and row[dept_col3] else ""
                    if dn in EXPECTED_DATA:
                        ap_by_dept[dn] = row
                for dept_name in DEPARTMENTS:
                    exp_pri = EXPECTED_DATA[dept_name]["priority"]
                    row = ap_by_dept.get(dept_name)
                    got_pri = str(row[pri_col]).strip() if row is not None and pri_col < len(row) and row[pri_col] else ""
                    gl = got_pri.lower()
                    # Accept EN priority and the RU equivalents (Высок/Средн/Низк).
                    pri_ok = {
                        "High": gl in ("high", "высокий", "высокая", "высок"),
                        "Medium": gl in ("medium", "средний", "средняя", "средн"),
                        "Low": gl in ("low", "низкий", "низкая", "низк"),
                    }[exp_pri]
                    # CRITICAL: priority must match the canonical rule applied to live pct.
                    check(f"Action_Plan {dept_name} priority is {exp_pri}",
                          pri_ok, f"Got '{got_pri}', expected '{exp_pri}'", critical=True)


def check_word(workspace):
    print("\n=== Check 2: Retention_Strategy_Report.docx ===")
    path = os.path.join(workspace, "Retention_Strategy_Report.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    # Use .lower() on the ORIGINAL text (no normalize()) so RU keyword checks work.
    full_text = " ".join(p.text for p in doc.paragraphs).lower()

    def has_any(*subs):
        return any(s in full_text for s in subs)

    # Title is the English "Employee Retention Strategy Report" per task.md, but
    # accept the RU prose equivalents too.
    check("Has title mentioning retention",
          (("retention" in full_text or "удержан" in full_text)
           and ("strategy" in full_text or "report" in full_text
                or "стратег" in full_text or "отчёт" in full_text or "отчет" in full_text)))
    check("Mentions flight risk",
          has_any("flight risk", "flight-risk", "риск ухода", "риска ухода", "отток", "оттока"))
    check("Mentions executive summary",
          has_any("executive summary", "summary", "резюме", "краткое резюме", "сводк"))
    check("Mentions research findings",
          (has_any("research", "исследован", "статья", "статьи", "статей")
           and has_any("finding", "paper", "вывод", "статья", "статьи", "статей")))
    check("Mentions recommendations",
          has_any("recommend", "рекоменд"))
    check("Mentions specific departments", sum(1 for d in DEPARTMENTS if d.lower() in full_text) >= 5,
          f"Found {sum(1 for d in DEPARTMENTS if d.lower() in full_text)} departments")
    check("Has substantial content", len(full_text) > 500, f"Length: {len(full_text)}")
    check("Mentions priority levels",
          (has_any("high", "высок") and has_any("medium", "low", "средн", "низк")))


def _priority_in(text, priority):
    """True if the RU/EN spelling of `priority` appears in `text` (already lower)."""
    variants = {
        "High": ("high", "высок"),
        "Medium": ("medium", "средн"),
        "Low": ("low", "низк"),
    }[priority]
    return any(v in text for v in variants)


def check_teamly():
    print("\n=== Check 3: Teamly 'Retention Action Items' ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        check("Teamly 'Retention Action Items' page exists", False, str(e), critical=True)
        return
    try:
        # Agent-created pages have id > 3 (seed pages are id <= 3). The page title
        # carries the English marker "Retention Action Items"; accept RU prose too.
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        new_pages = cur.fetchall()

        matched = []
        for pid, title, body in new_pages:
            tl = (title or "").lower()
            if ("retention action items" in tl
                    or ("retention" in tl and "action" in tl)
                    or ("удержан" in tl and ("действи" in tl or "план" in tl or "items" in tl))):
                matched.append((pid, title, body))

        check("Teamly 'Retention Action Items' page exists", len(matched) >= 1,
              f"new pages: {[(p[0], p[1]) for p in new_pages]}", critical=True)
        if not matched:
            return

        # Combined title+body of the matching page(s) (original case -> lower).
        combined = " ".join((t or "") + " " + (b or "") for _, t, b in matched).lower()

        # CRITICAL: all 7 departments present.
        depts_found = sum(1 for d in DEPARTMENTS if d.lower() in combined)
        check("Teamly page covers all 7 departments", depts_found >= 7,
              f"Found {depts_found}/7", critical=True)

        # CRITICAL: priority distribution matches the canonical-rule expectation.
        # Count substring occurrences of each priority token (RU+EN) and compare
        # to the expected distribution derived dynamically from EXPECTED_DATA.
        exp_dist = {"High": 0, "Medium": 0, "Low": 0}
        for d in DEPARTMENTS:
            exp_dist[EXPECTED_DATA[d]["priority"]] += 1

        def count_token(prio):
            toks = {"High": ("high", "высок"), "Medium": ("medium", "средн"),
                    "Low": ("low", "низк")}[prio]
            return sum(combined.count(t) for t in toks)

        got_dist = {p: count_token(p) for p in ("High", "Medium", "Low")}
        dist_ok = all(got_dist[p] >= exp_dist[p] for p in exp_dist if exp_dist[p] > 0)
        check("Teamly priority distribution matches canonical rule",
              dist_ok, f"expected>={exp_dist}, got {got_dist}", critical=True)

        # CRITICAL: every page reports Status 'Not Started' (RU 'Не начато'/'Не начат').
        # Require one occurrence per department.
        status_hits = combined.count("not started") + combined.count("не начат")
        check("Teamly Status is 'Not Started' for all departments",
              status_hits >= 7, f"Found {status_hits} Not-Started markers", critical=True)

        # NON-critical: owner (department head) names present.
        owners = ["чернова", "тарасов", "панкова", "ким", "фомина", "мартынов", "ванина"]
        owners_found = sum(1 for o in owners if o in combined)
        check("Teamly page lists department-head owners",
              owners_found >= 5, f"Found {owners_found}/7 owners")

        # NON-critical structural: Priority/Status/Owner/Strategy labels present.
        check("Teamly page mentions Priority/Приоритет",
              "priority" in combined or "приоритет" in combined)
        check("Teamly page mentions Strategy/Стратегия",
              "strategy" in combined or "стратег" in combined)
    except Exception as e:
        check("Teamly check", False, str(e), critical=True)
    finally:
        cur.close()
        conn.close()


def check_scripts(workspace):
    print("\n=== Check 4: Python Scripts ===")
    check("flight_risk_analysis.py exists",
          os.path.exists(os.path.join(workspace, "flight_risk_analysis.py")))
    check("synthesis.py exists",
          os.path.exists(os.path.join(workspace, "synthesis.py")))


def check_json_outputs(workspace):
    print("\n=== Check 5: JSON Outputs ===")
    fr_path = os.path.join(workspace, "flight_risk_analysis.json")
    if os.path.exists(fr_path):
        check("flight_risk_analysis.json exists", True)
        try:
            with open(fr_path) as f:
                fr_data = json.load(f)
            check("flight_risk_analysis.json is valid JSON", True)
            # CRITICAL: all 7 departments present with numeric flight-risk metrics.
            fr_text = json.dumps(fr_data).lower()
            depts = sum(1 for d in DEPARTMENTS if d.lower() in fr_text)
            check("flight_risk_analysis.json reports all 7 departments",
                  depts >= 7, f"Found {depts}/7", critical=True)
            # The dumped JSON must carry the headcount/flight-risk numbers we expect
            # (numbers appear as substrings regardless of nesting/field naming).
            hc_hits = sum(1 for d in DEPARTMENTS
                          if str(EXPECTED_DATA[d]["headcount"]) in fr_text)
            fr_hits = sum(1 for d in DEPARTMENTS
                          if str(EXPECTED_DATA[d]["flight_risk"]) in fr_text)
            check("flight_risk_analysis.json carries live headcount/flight-risk numbers",
                  hc_hits >= 6 and fr_hits >= 6,
                  f"headcount hits={hc_hits}/7, flight_risk hits={fr_hits}/7", critical=True)
        except (json.JSONDecodeError, Exception) as e:
            check("flight_risk_analysis.json is valid JSON", False, str(e), critical=True)
    else:
        check("flight_risk_analysis.json exists", False, critical=True)

    syn_path = os.path.join(workspace, "synthesis.json")
    if os.path.exists(syn_path):
        check("synthesis.json exists", True)
        try:
            with open(syn_path) as f:
                syn_data = json.load(f)
            check("synthesis.json is valid JSON", True)
        except (json.JSONDecodeError, Exception) as e:
            check("synthesis.json is valid JSON", False, str(e))
    else:
        check("synthesis.json exists", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_teamly()
    check_scripts(args.agent_workspace)
    check_json_outputs(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # CRITICAL gate: any critical failure => hard FAIL before the accuracy gate.
    if CRITICAL_FAILS:
        print(f"\nFAIL: {len(CRITICAL_FAILS)} critical check(s) failed: {CRITICAL_FAILS}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
