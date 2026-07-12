"""
Evaluation for grocery-budget-planner task (RU / kulinar).
Checks Excel, Google Sheet, and Calendar events.

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL,
regardless of accuracy. Structural / soft checks are non-critical.
"""
import argparse
import json
import os
import re
import sys
from datetime import timezone

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

BUDGET = 6000.0  # ₽

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic checks that gate PASS. Any failure here => overall FAIL.
CRITICAL_CHECKS = {
    "No shellfish in plan (allergy rule)",
    "Total Weekly Plan cost within budget",
    "Budget_Remaining arithmetic consistent",
    "Weekly Plan has 21 meal rows with name+cost",
    "Two grocery events with correct dates/times",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    # Strip currency/formatting (₽, NBSP, thousands separators) the task itself uses.
    s = str(v).replace("₽", "").replace(" ", "").replace(" ", "").replace(",", "").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group())
    except (TypeError, ValueError):
        return None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


# Shellfish keywords in RU + EN. Scanned against lowercased ORIGINAL text.
SHELLFISH_KEYWORDS = [
    "креветк", "краб", "лобстер", "мидии", "мидия", "моллюск", "лангуст", "устриц",
    "shrimp", "prawn", "crab", "lobster", "clam", "mussel", "oyster", "scallop",
]


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Meal_Budget.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    weekly_total = None

    # --- Weekly Plan ---
    wp_rows = load_sheet_rows(wb, "Weekly Plan") or load_sheet_rows(wb, "Weekly_Plan")
    if wp_rows is None:
        record("Sheet 'Weekly Plan' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Weekly Plan' exists", True)
        data = [r for r in wp_rows[1:] if r and r[0] is not None]

        name_col = find_col(wp_rows[0], ["Recipe_Name", "Recipe Name", "Recipe"])
        ing_col = find_col(wp_rows[0], ["Ingredients", "Ingredient"])
        cost_col = find_col(wp_rows[0], ["Cost", "Estimated_Cost"])

        # CRITICAL: 21 meal rows, each with non-empty Recipe_Name and numeric Cost.
        rows_ok = len(data) >= 21
        complete = rows_ok and name_col is not None and cost_col is not None
        if complete:
            good = 0
            for r in data:
                nm = r[name_col] if name_col < len(r) else None
                cv = r[cost_col] if cost_col < len(r) else None
                if nm is not None and str(nm).strip() and to_num(cv) is not None:
                    good += 1
            complete = good >= 21
            detail = f"{good} valid rows of {len(data)}"
        else:
            detail = f"rows={len(data)}, name_col={name_col}, cost_col={cost_col}"
        record("Weekly Plan has 21 meal rows with name+cost", complete, detail)

        # CRITICAL: no shellfish in Recipe_Name + Ingredients (original text).
        if name_col is not None:
            blob_parts = []
            for r in data:
                if name_col < len(r) and r[name_col] is not None:
                    blob_parts.append(str(r[name_col]))
                if ing_col is not None and ing_col < len(r) and r[ing_col] is not None:
                    blob_parts.append(str(r[ing_col]))
            blob = " ".join(blob_parts).lower()
            hits = [k for k in SHELLFISH_KEYWORDS if k in blob]
            record("No shellfish in plan (allergy rule)", len(hits) == 0,
                   f"Found keywords: {hits}")

        # CRITICAL: total cost within budget.
        if cost_col is not None:
            total = 0.0
            for r in data:
                if cost_col < len(r):
                    v = to_num(r[cost_col])
                    if v is not None:
                        total += v
            weekly_total = total
            record("Total Weekly Plan cost within budget", total <= BUDGET,
                   f"Total: {total:.2f} ₽ (budget {BUDGET:.0f})")

    # --- Shopping List ---
    sl_rows = load_sheet_rows(wb, "Shopping List") or load_sheet_rows(wb, "Shopping_List")
    if sl_rows is None:
        record("Sheet 'Shopping List' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Shopping List' exists", True)
        data = [r for r in sl_rows[1:] if r and r[0] is not None]
        record("Shopping List has >= 5 ingredients", len(data) >= 5, f"Found {len(data)}")

    # --- Budget Summary ---
    bs_rows = load_sheet_rows(wb, "Budget Summary") or load_sheet_rows(wb, "Budget_Summary")
    if bs_rows is None:
        record("Sheet 'Budget Summary' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Budget Summary' exists", True)
        metrics = {}
        for row in bs_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        twc_key = next((k for k in metrics if "total" in k and "cost" in k), None)
        summary_total = to_num(metrics.get(twc_key)) if twc_key else None
        if twc_key:
            record("Total_Weekly_Cost <= budget",
                   summary_total is not None and summary_total <= BUDGET,
                   f"Got {metrics.get(twc_key)}")
            # Summary total should match summed Weekly Plan total (non-critical, soft).
            if summary_total is not None and weekly_total is not None:
                record("Summary total ~ Weekly Plan total",
                       num_close(summary_total, weekly_total, max(weekly_total * 0.05, 5.0)),
                       f"summary={summary_total}, plan={weekly_total:.2f}")

        br_key = next((k for k in metrics if "budget" in k and "remain" in k), None)
        rem = to_num(metrics.get(br_key)) if br_key else None
        # CRITICAL: Budget_Remaining numeric, >=0, == budget - total within tol.
        if rem is not None and summary_total is not None:
            arith_ok = rem >= 0 and num_close(rem, BUDGET - summary_total, max(summary_total * 0.05, 5.0))
            record("Budget_Remaining arithmetic consistent", arith_ok,
                   f"remaining={rem}, expected={BUDGET - summary_total:.2f}")
        else:
            record("Budget_Remaining arithmetic consistent", False,
                   f"remaining={metrics.get(br_key)}, total={metrics.get(twc_key)}")

    # --- Groundtruth XLSX comparison (NON-critical; reference shape only) ---
    gt_path = os.path.join(groundtruth_workspace, "Meal_Budget.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            # Soft: agent has at least as many rows as the reference plan.
            record(f"GT '{gt_sname}' has comparable row count", len(a_rows) >= max(1, len(gt_rows) - 3),
                   f"reference {len(gt_rows)}, agent {len(a_rows)}")
        gt_wb.close()

    return True


def check_gsheet():
    print("\n=== Checking Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    sheets = cur.fetchall()
    record("Google Sheet created", len(sheets) >= 1, f"Found {len(sheets)}")

    if sheets:
        has_meal = any("meal" in str(s[1]).lower() for s in sheets)
        record("Sheet title mentions meal plan", has_meal,
               f"Titles: {[s[1] for s in sheets]}")

        cur.execute("SELECT COUNT(*) FROM gsheet.cells")
        count = cur.fetchone()[0]
        record("Google Sheet has data", count >= 10, f"Found {count} cells")

    cur.close()
    conn.close()
    return True


def check_calendar():
    print("\n=== Checking Calendar ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, start_datetime, end_datetime, description FROM gcal.events")
    events = cur.fetchall()
    record("Calendar events created", len(events) >= 2, f"Found {len(events)}")

    if events:
        summaries = [str(e[1]).lower() for e in events]
        has_shopping = any("shopping" in s or "grocery" in s for s in summaries)
        record("Shopping events present", has_shopping, f"Summaries: {summaries}")

    # CRITICAL: exactly the two required grocery trips with correct dates/times.
    # Trip 1: 2026-03-14 10:00, 1h.  Trip 2: 2026-03-18 17:00, 1h.
    # start_datetime is timestamptz; psycopg2 returns it tz-aware. The calendar MCC
    # converts a naive local time + IANA timeZone (e.g. Europe/Moscow, UTC+3) to UTC
    # on write, so 10:00 MSK is stored as 07:00Z. Normalize to UTC and accept the
    # literal hour OR its Europe/Moscow-UTC equivalent (local - 3h). Both 10:00 and
    # 17:00 MSK stay on the same calendar day in UTC, so the date stays comparable.
    def find_event(date_str, hour):
        utc_equiv = (hour - 3) % 24
        for e in events:
            sdt = e[2]
            if sdt is None:
                continue
            if sdt.tzinfo is not None:
                sdt = sdt.astimezone(timezone.utc).replace(tzinfo=None)
            if str(sdt.date()) == date_str and sdt.hour in (hour, utc_equiv):
                return e
        return None

    e1 = find_event("2026-03-14", 10)
    e2 = find_event("2026-03-18", 17)

    def dur_ok(e):
        if not e or e[2] is None or e[3] is None:
            return False
        secs = (e[3] - e[2]).total_seconds()
        return abs(secs - 3600) <= 600  # ~1 hour

    both_ok = e1 is not None and e2 is not None and dur_ok(e1) and dur_ok(e2)
    record("Two grocery events with correct dates/times", both_ok,
           f"e1={e1[1:4] if e1 else None}, e2={e2[1:4] if e2 else None}")

    # Soft: budget figure present in a description.
    descs = " ".join(str(e[4]) for e in events if e[4])
    record("Budget mentioned in event description",
           "6000" in descs.replace(" ", "").replace(",", ""),
           f"Descriptions snippet: {descs[:120]}")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gsheet()
    check_calendar()

    total = PASS_COUNT + FAIL_COUNT
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    if total == 0:
        print("FAIL: no checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"  Accuracy: {accuracy:.1f}%")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
