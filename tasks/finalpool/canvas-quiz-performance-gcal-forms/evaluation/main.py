"""
Evaluation for canvas-quiz-performance-gcal-gform task (RU stack: forms + canvas +
excel + google_calendar + emails).

The agent must produce, from live Canvas data:
1. Quiz_Performance_Report.xlsx — sheets 'Quiz Stats' and 'Course Summary'.
2. A form 'Quiz Improvement Feedback' (RU forms-mcp backend, schema gform.*) with a
   required multiple-choice topic question (4 options), a required 1..5 confidence
   "scale" question, and an optional text resources question.
3. Three 'CCC Spring 2014 Tutoring Session N' calendar events on 2026-03-12/19/26,
   each from 15:00 to 17:00 (a 2-hour window).
4. An email to ccc.instructor@university.edu with subject 'Creative Computing Quiz
   Performance Report' summarizing quiz performance and the tutoring sessions.

Note on the RU forms MCP (local_servers/forms-mcp): it stores questions as
question_type in ('textQuestion', 'choiceQuestion') only — there is NO scale/rating/
linear_scale type (DB CHECK constraint). The 1..5 confidence question is therefore a
choiceQuestion whose options are {1,2,3,4,5}; it is validated by its options, not by a
scale type literal. Choice options live in config JSONB as
{'type':'RADIO','options':[{'value':...}]}.

Gate: accuracy >= 70% AND no CRITICAL check failed => PASS.
Any CRITICAL failure => immediate FAIL (sys.exit(1)) regardless of accuracy.

All preserved English identifiers (sheet/column names, quiz titles CMA 2428x, form
title, MC option strings, session titles, email subject/recipient) are grepped on the
ORIGINAL text (lowercased), NOT normalized.
"""

import argparse
import json
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_by_name(wb, name):
    for sname in wb.sheetnames:
        if sname.strip().lower() == name.strip().lower():
            return [[cell.value for cell in row] for row in wb[sname].iter_rows()]
    return None


# ============================================================================
# Check 1: Excel file
# ============================================================================

def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Quiz_Performance_Report.xlsx ===")

    crit_quiz = ("Excel Quiz Stats: все 4 теста Canvas (CMA 24286/24287/24288/24289) "
                 "присутствуют с Avg_Score в пределах tol 2.0 от эталона")
    crit_summary = ("Excel Course Summary: Lowest_Avg_Quiz=='CMA 24286', "
                    "Highest_Avg_Quiz=='CMA 24287', Overall_Avg_Score ~ 69.43")

    try:
        import openpyxl
    except ImportError:
        record("openpyxl available", False, "pip install openpyxl")
        record(crit_quiz, False, "openpyxl missing", critical=True)
        record(crit_summary, False, "openpyxl missing", critical=True)
        return False

    agent_file = os.path.join(agent_workspace, "Quiz_Performance_Report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Quiz_Performance_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        record(crit_quiz, False, "no xlsx", critical=True)
        record(crit_summary, False, "no xlsx", critical=True)
        return False
    record("Excel file exists", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # --- Quiz Stats sheet ---
    a_quiz = load_sheet_by_name(agent_wb, "Quiz Stats")
    g_quiz = load_sheet_by_name(gt_wb, "Quiz Stats")
    record("Sheet 'Quiz Stats' exists", a_quiz is not None)

    quiz_crit_ok = True
    quiz_crit_detail = []
    if a_quiz is not None and g_quiz is not None:
        a_data = [r for r in a_quiz[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_quiz[1:] if any(v is not None for v in r)]
        record("Quiz Stats row count matches", len(a_data) == len(g_data),
               f"Expected {len(g_data)}, got {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            title = str(g_row[0]).strip()
            key = title.lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Quiz row exists: {title}", False, "Row not found")
                quiz_crit_ok = False
                quiz_crit_detail.append(f"{title}: missing")
                continue
            record(f"Quiz row exists: {title}", True)

            # Avg_Score (col index 2)
            avg_ok = (len(g_row) > 2 and len(a_row) > 2
                      and num_close(a_row[2], g_row[2], 2.0))
            record(f"{title}: Avg_Score correct", avg_ok,
                   f"got {a_row[2] if len(a_row) > 2 else None}, expected {g_row[2] if len(g_row) > 2 else None}")
            if not avg_ok:
                quiz_crit_ok = False
                quiz_crit_detail.append(
                    f"{title}: avg got {a_row[2] if len(a_row) > 2 else None} exp {g_row[2] if len(g_row) > 2 else None}")

            # Total_Submissions (col index 1)
            if len(g_row) > 1 and len(a_row) > 1:
                record(f"{title}: Total_Submissions correct",
                       num_close(a_row[1], g_row[1], 5),
                       f"got {a_row[1]}, expected {g_row[1]}")
            # Pass_Rate_Pct (col index 5)
            if len(g_row) > 5 and len(a_row) > 5:
                record(f"{title}: Pass_Rate_Pct correct",
                       num_close(a_row[5], g_row[5], 2.0),
                       f"got {a_row[5]}, expected {g_row[5]}")
    else:
        quiz_crit_ok = False
        quiz_crit_detail.append("Quiz Stats sheet missing")

    record(crit_quiz, quiz_crit_ok, "; ".join(quiz_crit_detail), critical=True)

    # --- Course Summary sheet ---
    a_summ = load_sheet_by_name(agent_wb, "Course Summary")
    g_summ = load_sheet_by_name(gt_wb, "Course Summary")
    record("Sheet 'Course Summary' exists", a_summ is not None)

    summ_crit_ok = True
    summ_crit_detail = []
    if a_summ is not None and g_summ is not None:
        a_data = [r for r in a_summ[1:] if any(v is not None for v in r)]
        g_data = [r for r in g_summ[1:] if any(v is not None for v in r)]

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                record(f"Summary row: {g_row[0]}", False, "Row not found")
                summ_crit_ok = False
                summ_crit_detail.append(f"{g_row[0]}: missing")
                continue
            record(f"Summary row: {g_row[0]}", True)

            if key == "total_quizzes":
                record("Total_Quizzes value",
                       num_close(a_row[1], g_row[1], 0),
                       f"got {a_row[1]}, expected {g_row[1]}")
            elif key == "total_quiz_submissions":
                record("Total_Quiz_Submissions value",
                       num_close(a_row[1], g_row[1], 10),
                       f"got {a_row[1]}, expected {g_row[1]}")
            elif key == "overall_avg_score":
                ok = num_close(a_row[1], g_row[1], 2.0)
                record("Overall_Avg_Score value", ok,
                       f"got {a_row[1]}, expected {g_row[1]}")
                if not ok:
                    summ_crit_ok = False
                    summ_crit_detail.append(
                        f"overall_avg got {a_row[1]} exp {g_row[1]}")
            elif key == "lowest_avg_quiz":
                ok = str_match(a_row[1], g_row[1])
                record("Lowest_Avg_Quiz value", ok,
                       f"got {a_row[1]}, expected {g_row[1]}")
                if not ok:
                    summ_crit_ok = False
                    summ_crit_detail.append(
                        f"lowest got {a_row[1]} exp {g_row[1]}")
            elif key == "highest_avg_quiz":
                ok = str_match(a_row[1], g_row[1])
                record("Highest_Avg_Quiz value", ok,
                       f"got {a_row[1]}, expected {g_row[1]}")
                if not ok:
                    summ_crit_ok = False
                    summ_crit_detail.append(
                        f"highest got {a_row[1]} exp {g_row[1]}")
    else:
        summ_crit_ok = False
        summ_crit_detail.append("Course Summary sheet missing")

    record(crit_summary, summ_crit_ok, "; ".join(summ_crit_detail), critical=True)

    return quiz_crit_ok and summ_crit_ok


# ============================================================================
# Check 2: Form (RU forms-mcp, gform.*)
# ============================================================================

def _option_values(config):
    """Extract option text values from a question config (RU forms-mcp shape)."""
    vals = []
    if not config:
        return vals
    cfg = config if isinstance(config, dict) else (
        json.loads(config) if config else {})
    opts = cfg.get("options") if isinstance(cfg, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    vals.append(str(v))
            else:
                vals.append(str(o))
    return vals


def check_gform():
    print("\n=== Checking Form 'Quiz Improvement Feedback' (forms / gform.*) ===")

    crit_form = ("Форма 'Quiz Improvement Feedback': обязательный MC-вопрос про тему "
                 "(4 варианта) + обязательный вопрос-шкала 1..5 + необязательный "
                 "текстовый вопрос")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("forms DB reachable", False, str(e))
        record(crit_form, False, "no db", critical=True)
        return False

    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()
    print(f"[check_gform] Found {len(forms)} forms.")
    record("At least 1 form created", len(forms) >= 1, f"Found {len(forms)}")

    target_form = None
    for form_id, title in forms:
        t = (title or "").lower()
        if "quiz" in t and ("improvement" in t or "feedback" in t):
            target_form = (form_id, title)
            break

    if target_form is None:
        record("Form 'Quiz Improvement Feedback' found", False,
               f"Forms: {[(f[0], f[1]) for f in forms]}")
        record(crit_form, False, "form not found", critical=True)
        cur.close()
        conn.close()
        return False

    record("Form 'Quiz Improvement Feedback' found", True)
    form_id = target_form[0]

    cur.execute(
        "SELECT title, question_type, required, config FROM gform.questions "
        "WHERE form_id=%s ORDER BY position", (form_id,))
    questions = cur.fetchall()
    cur.close()
    conn.close()

    record("Form has at least 3 questions", len(questions) >= 3,
           f"Found {len(questions)} questions")

    parsed = []
    for q_title, q_type, q_required, q_config in questions:
        opts = _option_values(q_config)
        parsed.append({
            "title": (q_title or "").lower(),
            "type": q_type,
            "required": bool(q_required),
            "options": opts,
            "options_lower": [v.strip().lower() for v in opts],
        })

    # --- MC topic question: choiceQuestion, required, with the 4 topic options ---
    topic_opts = ["programming fundamentals", "data structures",
                  "algorithms", "user interface design"]
    mc_q = None
    for q in parsed:
        if q["type"] != "choiceQuestion":
            continue
        n = sum(1 for k in topic_opts if any(k in o for o in q["options_lower"]))
        if n >= 3:  # this is the topic question
            mc_q = q
            break
    mc_ok = False
    if mc_q is not None:
        present = [k for k in topic_opts if any(k in o for o in mc_q["options_lower"])]
        mc_ok = (len(present) == 4 and mc_q["required"])
        record("Form has required MC topic question with 4 options", mc_ok,
               f"present={present} required={mc_q['required']} options={mc_q['options']}")
    else:
        record("Form has required MC topic question with 4 options", False,
               f"no topic choiceQuestion; questions={[(q['title'], q['type'], q['options']) for q in parsed]}")

    # --- Confidence "scale" question: choiceQuestion with options {1,2,3,4,5} ---
    # The RU forms fork has no scale type; a 1..5 question is a choiceQuestion whose
    # options are the digits 1..5 (validated by options, not by type literal).
    scale_q = None
    for q in parsed:
        if q["type"] != "choiceQuestion":
            continue
        digits = {o for o in q["options_lower"] if o in {"1", "2", "3", "4", "5"}}
        if len(digits) >= 5:
            scale_q = q
            break
    scale_ok = False
    if scale_q is not None:
        scale_ok = scale_q["required"]
        record("Form has required 1..5 scale question", scale_ok,
               f"required={scale_q['required']} options={scale_q['options']}")
    else:
        record("Form has required 1..5 scale question", False,
               f"no 1..5 choiceQuestion; questions={[(q['title'], q['type'], q['options']) for q in parsed]}")

    # --- Optional text resources question ---
    text_q = None
    for q in parsed:
        if q["type"] == "textQuestion":
            text_q = q
            break
    text_ok = text_q is not None
    record("Form has text (resources) question", text_ok,
           f"types={[q['type'] for q in parsed]}")

    crit_ok = mc_ok and scale_ok and text_ok
    record(crit_form, crit_ok,
           f"mc={mc_ok} scale={scale_ok} text={text_ok}", critical=True)

    return crit_ok


# ============================================================================
# Check 3: Google Calendar (kept-foreign infra; gcal.* data layer)
# ============================================================================

def check_gcal():
    print("\n=== Checking Google Calendar ===")

    crit_gcal = ("Календарь: 3 репетиторские сессии на 2026-03-12/19/26, каждая в "
                 "окне 15:00-17:00 (2 часа)")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("gcal DB reachable", False, str(e))
        record(crit_gcal, False, "no db", critical=True)
        return False

    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_gcal] Found {len(events)} calendar events.")
    record("At least 3 calendar events created", len(events) >= 3, f"Found {len(events)}")

    tutoring_events = [e for e in events
                       if e[0] and ("tutoring" in e[0].lower() or "session" in e[0].lower()
                                    or "ccc" in e[0].lower())]
    record("3 tutoring/session events found", len(tutoring_events) >= 3,
           f"Found {len(tutoring_events)}: {[e[0] for e in tutoring_events]}")

    march_events = [e for e in events
                    if e[1] and "2026-03" in str(e[1])]
    record("Events scheduled in March 2026", len(march_events) >= 3,
           f"Found {len(march_events)} March 2026 events")

    # CRITICAL: each of 12/19/26 March 2026 covered by a tutoring event with a 2h
    # (15:00-17:00) window. start_datetime is tz-aware (timestamptz); we bucket by
    # the stored calendar day and require exact 120-minute duration (TZ-independent).
    expected_days = {12, 19, 26}
    by_day = {}
    for summ, sdt, edt in tutoring_events:
        if sdt is None or edt is None:
            continue
        # Must be a March 2026 event.
        if not (sdt.year == 2026 and sdt.month == 3):
            continue
        dur_min = round((edt - sdt).total_seconds() / 60.0)
        if sdt.day in expected_days:
            by_day.setdefault(sdt.day, []).append(dur_min)

    days_covered = set(by_day.keys())
    all_days = (days_covered == expected_days)
    durations_ok = all(
        any(abs(dm - 120) <= 1 for dm in by_day.get(d, []))
        for d in expected_days
    )
    crit_ok = all_days and durations_ok
    record(crit_gcal, crit_ok,
           f"days_covered={sorted(days_covered)} expected={sorted(expected_days)} "
           f"durations(min)_by_day={by_day}", critical=True)

    return crit_ok


# ============================================================================
# Check 4: Email
# ============================================================================

def check_emails():
    print("\n=== Checking Emails ===")

    crit_email = ("Письмо на ccc.instructor@university.edu (получатель обязателен) с "
                  "темой 'Creative Computing Quiz Performance Report' и телом про "
                  "успеваемость по тестам и репетиторские сессии")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("email DB reachable", False, str(e))
        record(crit_email, False, "no db", critical=True)
        return False

    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
    """)
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    print(f"[check_emails] Found {len(all_emails)} total emails.")
    record("At least 1 email sent", len(all_emails) >= 1, f"Found {len(all_emails)}")

    # Recipient is enforced via AND (not OR with subject).
    target = None
    for subject, from_addr, to_addr, body_text in all_emails:
        to_str = str(to_addr or "").lower()
        if "ccc.instructor@university.edu" in to_str:
            target = (subject, from_addr, to_addr, body_text)
            break

    if target is None:
        record("Email to ccc.instructor@university.edu found", False,
               f"Emails: {[(e[0], e[2]) for e in all_emails[:3]]}")
        record(crit_email, False, "no email to required recipient", critical=True)
        return False

    record("Email to ccc.instructor@university.edu found", True)
    subject, from_addr, to_addr, body_text = target
    subject_lower = (subject or "").lower()
    body_lower = (body_text or "").lower()

    # Subject: prefer exact RU+EN tolerant; the canonical subject is English.
    subj_ok = ("quiz" in subject_lower or "performance" in subject_lower
               or "тест" in subject_lower or "успеваемост" in subject_lower)
    record("Email subject mentions quiz performance", subj_ok,
           f"Subject: {subject}")

    # Body: RU+EN keywords — quiz/performance/score (EN) OR тест/успеваемост/отчёт/
    # балл (RU) AND a mention of tutoring/sessions (репетит/session/сесси).
    body_quiz = any(k in body_lower for k in
                    ("quiz", "performance", "score", "тест", "успеваемост",
                     "отчёт", "отчет", "балл"))
    body_tutoring = any(k in body_lower for k in
                        ("tutoring", "session", "репетит", "сесси", "занят"))
    body_ok = body_quiz and body_tutoring
    record("Email body mentions quiz performance and tutoring sessions", body_ok,
           f"quiz_kw={body_quiz} tutoring_kw={body_tutoring}")

    crit_ok = subj_ok and body_ok
    record(crit_email, crit_ok,
           f"subject_ok={subj_ok} body_ok={body_ok}", critical=True)

    return crit_ok


# ============================================================================
# Main
# ============================================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    print(f"  Critical failures: {CRITICAL_FAILS}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILS,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILS}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
