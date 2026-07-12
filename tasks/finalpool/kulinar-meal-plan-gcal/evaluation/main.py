"""
Evaluation script for kulinar-meal-plan-gcal task (kulinar recipes).

Semantic CRITICAL checks (any failure => overall FAIL regardless of accuracy):
  - >=21 cooking events (RU 'готов'/'приготов' OR EN 'cook') within March 9-15 2026,
    spanning all 7 distinct dates with >=3 events per date (breakfast/lunch/dinner).
  - Cooking event start times match the three meal slots (07:00, 11:30, 18:00
    America/New_York local) and each event lasts 1 hour.
  - 'Meal Plan' sheet has >=21 data rows; Prep_Time populated and numeric.
  - Recipe names in 'Meal Plan' are >=18 distinct AND appear in the calendar
    event summaries (Excel<->GCal consistency).
  - 'Shopping List' aggregates >=5 unique ingredients with Ingredient + Quantity.

Structural (non-critical) checks: sheet/header existence, ISO date span.
Pass threshold: accuracy >= 70% AND no critical failure.

RU keyword note: the cooking-keyword match runs on the ORIGINAL lowercased
summary text (NOT normalized), so Cyrillic 'готов'/'приготов' is preserved.
"""

import argparse
import json
import os
import re
import sys
from datetime import timezone

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# America/New_York local meal slots. March 9-15 2026 is after DST start
# (2026-03-08), so the zone is EDT = UTC-4:
#   07:00 local -> 11:00 UTC, 11:30 local -> 15:30 UTC, 18:00 local -> 22:00 UTC.
MEAL_SLOTS_UTC = {(11, 0), (15, 30), (22, 0)}

CRITICAL_CHECKS = {
    ">=21 cooking events spanning 7 days with >=3 per date",
    "Cooking events match meal slots (07:00/11:30/18:00 local) and 1h duration",
    "Meal Plan sheet has >=21 data rows with numeric Prep_Time",
    "Recipe names >=18 distinct and consistent Excel<->GCal",
    "Shopping List aggregates >=5 ingredients with Ingredient + Quantity",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


COOK_KEYWORDS = ("готов", "приготов", "cook")


def is_cooking_summary(summary):
    """Match RU 'готов'/'приготов' or EN 'cook' on ORIGINAL lowercased text."""
    if not summary:
        return False
    low = str(summary).lower()
    return any(kw in low for kw in COOK_KEYWORDS)


def recipe_from_summary(summary):
    """Extract recipe name after the 'Готовка:' / 'Cook:' prefix, lowercased."""
    if not summary:
        return ""
    s = str(summary).strip()
    s = re.sub(r"^\s*(готовка|приготовление|cook)\s*[:\-—]\s*", "", s,
               flags=re.IGNORECASE)
    return s.strip().lower()


def fetch_cooking_events(cur):
    cur.execute("""
        SELECT summary, start_datetime, end_datetime, start_timezone
        FROM gcal.events
        WHERE start_datetime >= '2026-03-09T00:00:00+00:00'
          AND start_datetime < '2026-03-16T23:59:59+00:00'
    """)
    rows = cur.fetchall()
    return [r for r in rows if is_cooking_summary(r[0])]


def check_gcal(cur):
    events = fetch_cooking_events(cur)

    # --- Critical: count, day coverage, >=3 per date ---
    days = {}
    for summary, start_dt, _end, _tz in events:
        if start_dt:
            d = start_dt.astimezone(timezone.utc).date() if hasattr(start_dt, "astimezone") else str(start_dt)[:10]
            days.setdefault(d, 0)
            days[d] += 1
    enough = len(events) >= 21 and len(days) >= 7 and all(v >= 3 for v in days.values())
    record(">=21 cooking events spanning 7 days with >=3 per date", enough,
           f"events={len(events)}, days={len(days)}, per_day={sorted(days.values())}")

    # --- Critical: meal-slot times + 1h duration ---
    slot_ok = 0
    dur_ok = 0
    checked = 0
    for summary, start_dt, end_dt, _tz in events:
        if not (start_dt and end_dt):
            continue
        checked += 1
        su = start_dt.astimezone(timezone.utc)
        if (su.hour, su.minute) in MEAL_SLOTS_UTC:
            slot_ok += 1
        if abs((end_dt - start_dt).total_seconds() - 3600) <= 60:
            dur_ok += 1
    # Require the vast majority of events to honor the slots/duration.
    times_ok = checked >= 21 and slot_ok >= 21 and dur_ok >= 21
    record("Cooking events match meal slots (07:00/11:30/18:00 local) and 1h duration",
           times_ok, f"checked={checked}, slot_ok={slot_ok}, dur_ok={dur_ok}")

    # --- Non-critical structural: ISO date strings present ---
    record("GCal events carry ISO start datetimes",
           all(s for _, s, _, _ in events) and len(events) > 0,
           f"events={len(events)}")

    # Return recipe names extracted from summaries for cross-checking.
    return {recipe_from_summary(s) for s, _, _, _ in events if recipe_from_summary(s)}


def load_sheet(wb, target):
    low = [s.lower() for s in wb.sheetnames]
    if target.lower() in low:
        return wb[wb.sheetnames[low.index(target.lower())]]
    return None


def check_excel(workspace, gcal_recipes):
    from openpyxl import load_workbook

    xlsx_path = os.path.join(workspace, "Weekly_Meal_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Weekly_Meal_Plan.xlsx exists", False, xlsx_path)
        # Cannot run the dependent critical checks; mark them failed.
        record("Meal Plan sheet has >=21 data rows with numeric Prep_Time", False, "no file")
        record("Recipe names >=18 distinct and consistent Excel<->GCal", False, "no file")
        record("Shopping List aggregates >=5 ingredients with Ingredient + Quantity", False, "no file")
        return
    record("Weekly_Meal_Plan.xlsx exists", True)

    wb = load_workbook(xlsx_path)

    # --- Meal Plan sheet ---
    ws = load_sheet(wb, "meal plan")
    record("'Meal Plan' sheet exists", ws is not None, wb.sheetnames)

    recipe_names = []
    prep_numeric = 0
    plan_rows = 0
    if ws is not None:
        headers = [str(c.value).lower() if c.value else "" for c in ws[1]]
        for rh in ["day", "meal_type", "recipe_name", "prep_time"]:
            record(f"Meal Plan header '{rh}'",
                   any(rh.replace("_", " ") in h or rh in h for h in headers), headers)
        # locate columns
        def col_idx(key):
            for i, h in enumerate(headers):
                if key.replace("_", " ") in h or key in h:
                    return i
            return None
        ci_recipe = col_idx("recipe_name")
        ci_prep = col_idx("prep_time")
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            plan_rows += 1
            if ci_recipe is not None and ci_recipe < len(row):
                v = row[ci_recipe].value
                if v is not None and str(v).strip():
                    recipe_names.append(str(v).strip())
            if ci_prep is not None and ci_prep < len(row):
                pv = row[ci_prep].value
                try:
                    float(str(pv).replace(",", ".").split()[0])
                    prep_numeric += 1
                except (TypeError, ValueError, IndexError):
                    pass

    # Critical: >=21 rows with numeric prep time
    record("Meal Plan sheet has >=21 data rows with numeric Prep_Time",
           plan_rows >= 21 and prep_numeric >= 21,
           f"plan_rows={plan_rows}, prep_numeric={prep_numeric}")

    # Critical: variety + Excel<->GCal consistency
    distinct = {r.lower() for r in recipe_names}
    # consistency: most plan recipes should be present among gcal summaries
    matched = 0
    for r in distinct:
        if any(r in g or g in r for g in gcal_recipes):
            matched += 1
    consistent = len(distinct) >= 18 and gcal_recipes and matched >= max(1, int(0.7 * len(distinct)))
    record("Recipe names >=18 distinct and consistent Excel<->GCal", consistent,
           f"distinct={len(distinct)}, gcal_recipes={len(gcal_recipes)}, matched={matched}")

    # --- Shopping List sheet ---
    ws2 = load_sheet(wb, "shopping list")
    record("'Shopping List' sheet exists", ws2 is not None, wb.sheetnames)
    ing_ok = False
    if ws2 is not None:
        headers2 = [str(c.value).lower() if c.value else "" for c in ws2[1]]
        has_ing = any("ingredient" in h for h in headers2)
        has_qty = any("quantity" in h for h in headers2)
        record("Shopping List header 'Ingredient'", has_ing, headers2)
        record("Shopping List header 'Quantity'", has_qty, headers2)
        data_rows = sum(1 for row in ws2.iter_rows(min_row=2) if row[0].value is not None)
        ing_ok = has_ing and has_qty and data_rows >= 5
    record("Shopping List aggregates >=5 ingredients with Ingredient + Quantity",
           ing_ok)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("\n=== Checking Google Calendar Events ===")
    gcal_recipes = set()
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        gcal_recipes = check_gcal(cur)
        cur.close()
        conn.close()
    except Exception as e:
        record(">=21 cooking events spanning 7 days with >=3 per date", False, str(e))
        record("Cooking events match meal slots (07:00/11:30/18:00 local) and 1h duration", False, str(e))

    print("\n=== Checking Excel ===")
    check_excel(args.agent_workspace, gcal_recipes)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
        "success": (not critical_failed) and accuracy >= 70,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
