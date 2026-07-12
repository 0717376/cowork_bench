"""
Evaluation script for sf-training-needs-assessment (ClickHouse fork).

The HR_ANALYTICS departments in sf_data are russified centrally
(db/zzz_clickhouse_after_init.sql): Engineering->Инженерия, Finance->Финансы,
HR->Кадры, Operations->Операции, R&D->НИОКР, Sales->Продажи, Support->Поддержка.
The agent reads the live ClickHouse-backed DWH, so the DEPARTMENT value it writes
into Training_Needs.xlsx and matches against department_heads.json is the Cyrillic
name. The training catalog (titles, dept tags) stays ENGLISH, so COURSE_MAP titles
and course costs are unchanged. Numeric figures (counts, costs, budget) are NOT
translated.

Checks:
1. Excel file Training_Needs.xlsx with 3 sheets (structure + exact semantic values).
2. Emails sent to each department head with the correct body content.

CRITICAL gate: any critical-check failure => hard FAIL before the accuracy gate.
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Cheapest department-specific course per (Cyrillic) department, title kept ENGLISH.
COURSE_MAP = {
    "Инженерия": ("Technical Deep Dive", 1800),
    "Финансы": ("Data Analysis Fundamentals", 900),
    "Кадры": ("People Management", 600),
    "Операции": ("Process Optimization", 1100),
    "НИОКР": ("Innovation Workshop", 1400),
    "Продажи": ("Negotiation Skills", 950),
    "Поддержка": ("Customer Service Mastery", 700),
}

DEPT_HEADS = {
    "Инженерия": "eng-head@company.com",
    "Продажи": "sales-head@company.com",
    "Кадры": "hr-head@company.com",
    "Финансы": "finance-head@company.com",
    "НИОКР": "rd-head@company.com",
    "Операции": "ops-head@company.com",
    "Поддержка": "support-head@company.com",
}


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}{detail_str}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, abs_tol=1.0, rel_tol=0.0):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_values():
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Count low performers (PERFORMANCE_RATING = 1) by department (Cyrillic values).
    cur.execute("""
        SELECT "DEPARTMENT", COUNT(*)
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        WHERE "PERFORMANCE_RATING" = 1
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    dept_counts = {}
    for dept, cnt in cur.fetchall():
        dept_counts[dept] = cnt

    total_employees = sum(dept_counts.values())

    recommended = []
    total_est_cost = 0
    for dept in sorted(COURSE_MAP.keys()):
        title, cost = COURSE_MAP[dept]
        count = dept_counts.get(dept, 0)
        total = cost * count
        total_est_cost += total
        recommended.append({
            "Department": dept,
            "Course_Title": title,
            "Course_Cost": cost,
            "Employee_Count": count,
            "Total_Cost": total,
        })

    summary = {
        "Total_Eligible_Employees": total_employees,
        "Total_Departments": len([d for d in dept_counts if dept_counts[d] > 0]),
        "Estimated_Total_Cost": total_est_cost,
        "Available_Budget": 500000,
        "Budget_Remaining": 500000 - total_est_cost,
        "Within_Budget": "Yes" if total_est_cost <= 500000 else "No",
    }

    cur.close()
    conn.close()

    return {
        "dept_counts": dept_counts,
        "recommended": recommended,
        "summary": summary,
        "total_employees": total_employees,
    }


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def check_excel(agent_workspace, expected):
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Training_Needs.xlsx")
    check("Excel file exists", os.path.isfile(agent_file), f"Expected {agent_file}")
    if not os.path.isfile(agent_file):
        # Cascade: the core deliverable is missing -> mark the semantic criticals failed.
        check("Low Performers per-dept counts exact", False, "no xlsx", critical=True)
        check("Recommended Courses correct", False, "no xlsx", critical=True)
        check("Budget Summary exact", False, "no xlsx", critical=True)
        return

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        check("Low Performers per-dept counts exact", False, "unreadable xlsx", critical=True)
        check("Recommended Courses correct", False, "unreadable xlsx", critical=True)
        check("Budget Summary exact", False, "unreadable xlsx", critical=True)
        return
    check("Excel file readable", True)

    for sn in ["Low Performers", "Recommended Courses", "Budget Summary"]:
        found = any(str_match(s, sn) for s in wb.sheetnames)
        check(f"Sheet '{sn}' exists", found, f"Found: {wb.sheetnames}")

    # --- Low Performers ---
    print("\n--- Low Performers ---")
    ws = get_sheet(wb, "Low Performers")
    lp_counts_ok = True
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Low Performers has data", len(rows) > 0)
        # Total row count must equal exact DB total (non-critical structural confirm).
        check("Low Performers count exact",
              len(rows) == expected["total_employees"],
              f"Expected {expected['total_employees']}, got {len(rows)}")

        # Department distribution (Cyrillic dept values).
        dept_dist = {}
        for r in rows:
            if r and r[1]:
                d = str(r[1]).strip()
                dept_dist[d] = dept_dist.get(d, 0) + 1
        for dept, exp_cnt in expected["dept_counts"].items():
            actual = dept_dist.get(dept, 0)
            ok = (actual == exp_cnt)
            if not ok:
                lp_counts_ok = False
            check(f"Dept '{dept}' count exact", ok,
                  f"Expected {exp_cnt}, got {actual}")
    else:
        lp_counts_ok = False

    # CRITICAL: Low Performers per-department row counts exactly match DB dept_counts.
    check("Low Performers per-dept counts exact", lp_counts_ok,
          "one or more department buckets differ from DB", critical=True)

    # --- Recommended Courses ---
    print("\n--- Recommended Courses ---")
    ws = get_sheet(wb, "Recommended Courses")
    rec_ok = True
    if ws:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        check("Recommended Courses row count", len(rows) == 7,
              f"Expected 7, got {len(rows)}")

        for exp_row in expected["recommended"]:
            dept = exp_row["Department"]
            matched = None
            for r in rows:
                if r and str_match(r[0], dept):
                    matched = r
                    break
            if matched:
                title_ok = str_match(matched[1], exp_row["Course_Title"])
                cost_ok = num_close(matched[2], exp_row["Course_Cost"], 0)
                count_ok = num_close(matched[3], exp_row["Employee_Count"], 0)
                total_ok = num_close(matched[4], exp_row["Total_Cost"], 0)
                check(f"{dept} Course_Title", title_ok,
                      f"Expected '{exp_row['Course_Title']}', got '{matched[1]}'")
                check(f"{dept} Course_Cost", cost_ok,
                      f"Expected {exp_row['Course_Cost']}, got {matched[2]}")
                check(f"{dept} Employee_Count", count_ok,
                      f"Expected {exp_row['Employee_Count']}, got {matched[3]}")
                check(f"{dept} Total_Cost", total_ok,
                      f"Expected {exp_row['Total_Cost']}, got {matched[4]}")
                if not (title_ok and cost_ok and count_ok and total_ok):
                    rec_ok = False
            else:
                check(f"Dept '{dept}' found", False, "Not in output")
                rec_ok = False
    else:
        rec_ok = False

    # CRITICAL: each department has correct cheapest dept-specific course title + exact
    # cost from catalog, and Total_Cost == Course_Cost * Employee_Count exactly.
    check("Recommended Courses correct", rec_ok,
          "title/cost/count/total mismatch", critical=True)

    # --- Budget Summary ---
    print("\n--- Budget Summary ---")
    ws = get_sheet(wb, "Budget Summary")
    bs_ok = True
    if ws:
        data = {}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                data[str(row[0]).strip().lower().replace(" ", "_")] = row[1]

        for key, gt_val in expected["summary"].items():
            key_lower = key.lower()
            agent_val = data.get(key_lower)
            if agent_val is None:
                for ak, av in data.items():
                    if key_lower.replace("_", "") in ak.replace("_", ""):
                        agent_val = av
                        break
            if isinstance(gt_val, (int, float)):
                ok = num_close(agent_val, gt_val, 0)
            else:
                ok = str_match(agent_val, gt_val)
            check(f"Summary '{key}'", ok, f"Expected {gt_val}, got {agent_val}")
            if not ok:
                bs_ok = False
    else:
        bs_ok = False

    # CRITICAL: Total_Eligible_Employees, Estimated_Total_Cost, Budget_Remaining and
    # Within_Budget computed exactly and self-consistent.
    check("Budget Summary exact", bs_ok,
          "eligible/estimated/remaining/within-budget mismatch", critical=True)


def check_emails(expected):
    print("\n=== Checking Emails ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("""
        SELECT subject, to_addr, body_text
        FROM email.messages
        WHERE LOWER(subject) LIKE '%training%'
        ORDER BY subject
    """)
    all_emails = list(cur.fetchall())

    try:
        cur.execute("""
            SELECT m.subject, m.to_addr, m.body_text
            FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
            WHERE LOWER(m.subject) LIKE '%training%'
            ORDER BY m.subject
        """)
        all_emails += list(cur.fetchall())
    except Exception:
        pass

    check("Training emails sent", len(all_emails) >= 7,
          f"Found {len(all_emails)} training emails")

    # Per-department: recipient present AND body mentions the dept's low-performer count
    # AND its recommended course title.
    rec_by_dept = {r["Department"]: r for r in expected["recommended"]}
    body_ok_all = True

    for dept, email_addr in DEPT_HEADS.items():
        rec = rec_by_dept.get(dept, {})
        count = rec.get("Employee_Count", expected["dept_counts"].get(dept, 0))
        title = rec.get("Course_Title", "")

        # find emails addressed to this head
        addressed = [e for e in all_emails
                     if e[1] and email_addr.lower() in str(e[1]).lower()]
        check(f"Email to {dept} head ({email_addr})", len(addressed) > 0)

        # subject format "Training Plan for [Department]"
        subj_ok = any(
            e[0] and "training plan for" in str(e[0]).lower()
            for e in addressed
        )
        check(f"{dept} subject format 'Training Plan for ...'", subj_ok,
              f"subjects: {[e[0] for e in addressed]}")

        # body mentions count + course title
        bodies = " \n ".join(str(e[2]) for e in addressed if e[2])
        bl = bodies.lower()
        count_in_body = bool(re.search(r"(?<!\d)%d(?!\d)" % int(count), bodies)) if count is not None else False
        title_in_body = title.lower() in bl if title else False
        this_ok = count_in_body and title_in_body
        if not this_ok:
            body_ok_all = False
        check(f"{dept} body mentions count={count} and course '{title}'",
              this_ok,
              f"count_in_body={count_in_body}, title_in_body={title_in_body}")

    # CRITICAL: each dept-head email body mentions that dept's exact low-performer count
    # AND its recommended course title.
    check("Email bodies contain count + course title per dept", body_ok_all,
          "one or more dept emails missing count/title in body", critical=True)

    cur.close()
    conn.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=== Computing Expected Values ===")
    try:
        expected = compute_expected_values()
        print(f"  Total eligible: {expected['total_employees']}")
        print(f"  Departments: {len(expected['dept_counts'])}")
        print(f"  Estimated cost: {expected['summary']['Estimated_Total_Cost']}")
    except Exception as e:
        print(f"  ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    check_excel(args.agent_workspace, expected)
    check_emails(expected)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")

    # CRITICAL gate: any critical failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILS:
        success = False
        print("  Overall: FAIL (critical check failed)")
    else:
        success = accuracy >= 70
        print(f"  Overall: {'PASS' if success else 'FAIL'} (threshold accuracy>=70)")

    if args.res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": round(accuracy, 1),
            "critical_fails": CRITICAL_FAILS,
            "success": success,
        }
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
