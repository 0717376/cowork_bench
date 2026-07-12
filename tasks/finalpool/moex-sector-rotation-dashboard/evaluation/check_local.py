"""
Local check for moex-sector-rotation-dashboard task.
Compares agent-produced sector_rotation_report.xlsx against expected values
computed dynamically from PostgreSQL (moex.stock_prices, moex.financial_statements).

Sector benchmarks and analyst ratings come from the mock research portal
(static HTML served during the task). These are hardcoded constants since
they are defined by the task itself and do not change.

Falls back to static groundtruth Excel if PostgreSQL is unavailable.

The check produces a list of granular results so the evaluator can apply an
accuracy threshold AND a CRITICAL-checks gate. Each result is a dict:
    {"name": str, "passed": bool, "critical": bool, "msg": str}
"""
import os
import openpyxl


# --- Constants from the mock research portal (http://localhost:30145) ---
STOCKS = ['SBER.ME', 'GAZP.ME', 'LKOH.ME', 'MGNT.ME', 'MTSS.ME']
SECTORS = {
    'SBER.ME': 'Financial Services',
    'GAZP.ME': 'Energy',
    'LKOH.ME': 'Energy',
    'MGNT.ME': 'Consumer Defensive',
    'MTSS.ME': 'Communication Services',
}
BENCHMARKS = {
    'Financial Services': -3.0,
    'Energy': 5.0,
    'Consumer Defensive': -2.0,
    'Communication Services': -4.0,
}
ANALYST_RATINGS = {
    'SBER.ME': 'Buy',
    'GAZP.ME': 'Overweight',
    'LKOH.ME': 'Hold',
    'MGNT.ME': 'Neutral',
    'MTSS.ME': 'Outperform',
}
TARGET_PRICES = {
    'SBER.ME': 140,
    'GAZP.ME': 220,
    'LKOH.ME': 4000,
    'MGNT.ME': 4500,
    'MTSS.ME': 280,
}

# In-range comparison window for moex daily data (spans 2026-02-25..2026-05-26).
BASELINE_DATE = '2026-02-25'
CURRENT_DATE = '2026-05-26'

DB_CONFIG = dict(
    host=os.environ.get('PGHOST', 'localhost'),
    port=int(os.environ.get('PGPORT', 5432)),
    database=os.environ.get('PGDATABASE', 'cowork_gym'),
    user=os.environ.get('PGUSER', 'postgres'),
    password=os.environ.get('PGPASSWORD', 'postgres'),
)


def _str_match(a, b):
    """Case-insensitive string comparison after stripping."""
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected_from_db():
    """Compute all expected values from PostgreSQL (moex schema)."""
    try:
        import psycopg2
    except ImportError:
        return None

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    perf = {}  # sym -> dict
    fin = {}   # sym -> dict

    for sym in STOCKS:
        sector = SECTORS[sym]
        bench = BENCHMARKS[sector]

        # Baseline price and current price
        cur.execute("SELECT close FROM moex.stock_prices WHERE symbol=%s AND date=%s",
                    (sym, BASELINE_DATE))
        row = cur.fetchone()
        if not row:
            conn.close()
            return None
        p1y = round(float(row[0]), 4)

        cur.execute("SELECT close FROM moex.stock_prices WHERE symbol=%s AND date=%s",
                    (sym, CURRENT_DATE))
        row = cur.fetchone()
        if not row:
            conn.close()
            return None
        pcur = round(float(row[0]), 2)

        ret_pct = round((pcur - p1y) / p1y * 100, 2)
        alpha = round(ret_pct - bench, 2)
        target = TARGET_PRICES[sym]
        upside = round((target - pcur) / pcur * 100, 2)

        perf[sym] = {
            'Symbol': sym,
            'Sector': sector,
            'Price_1Y_Ago': p1y,
            'Current_Price': pcur,
            'Return_Pct': ret_pct,
            'Benchmark_Return_Pct': bench,
            'Alpha': alpha,
            'Analyst_Rating': ANALYST_RATINGS[sym],
            'Target_Price': target,
            'Upside_Pct': upside,
        }

        # Financial data: latest ANNUAL income statement (no quarterly rows in moex)
        cur.execute(
            "SELECT data FROM moex.financial_statements "
            "WHERE symbol=%s AND stmt_type='income_stmt' AND freq='annual' "
            "ORDER BY period_end DESC LIMIT 1", (sym,))
        row = cur.fetchone()
        if not row:
            conn.close()
            return None
        d = row[0]
        rev_raw = d.get('Total Revenue')
        ni_raw = d.get('Net Income')
        if rev_raw is None or ni_raw is None:
            conn.close()
            return None
        rev = round(float(rev_raw) / 1e6, 2)
        ni = round(float(ni_raw) / 1e6, 2)
        margin = round(ni / rev * 100, 2)

        fin[sym] = {
            'Symbol': sym,
            'Revenue_Latest_Q': rev,
            'Net_Income_Latest_Q': ni,
            'Profit_Margin_Pct': margin,
        }

    conn.close()

    # Summary
    returns = {s: perf[s]['Return_Pct'] for s in STOCKS}
    alphas = {s: perf[s]['Alpha'] for s in STOCKS}
    best = max(returns, key=returns.get)
    worst = min(returns, key=returns.get)
    avg_alpha = round(sum(alphas.values()) / len(alphas), 2)
    above = sum(1 for a in alphas.values() if a > 0)
    below = sum(1 for a in alphas.values() if a < 0)

    summary_rows = [
        ('Best_Performer', best),
        ('Worst_Performer', worst),
        ('Avg_Alpha', avg_alpha),
        ('Stocks_Above_Benchmark', above),
        ('Stocks_Below_Benchmark', below),
    ]

    return {'performance': perf, 'financials': fin, 'summary': summary_rows}


def _load_expected_from_gt(gt_wb):
    """Build the same expected structure from the static groundtruth workbook."""
    perf = {}
    fin = {}
    perf_headers = ['Symbol', 'Sector', 'Price_1Y_Ago', 'Current_Price', 'Return_Pct',
                    'Benchmark_Return_Pct', 'Alpha', 'Analyst_Rating', 'Target_Price', 'Upside_Pct']
    fin_headers = ['Symbol', 'Revenue_Latest_Q', 'Net_Income_Latest_Q', 'Profit_Margin_Pct']

    gp = gt_wb["Performance"]
    for row_idx in range(2, 2 + len(STOCKS)):
        vals = [gp.cell(row=row_idx, column=c).value for c in range(1, 11)]
        sym = vals[0]
        perf[sym] = {h: v for h, v in zip(perf_headers, vals)}

    gf = gt_wb["Financials"]
    for row_idx in range(2, 2 + len(STOCKS)):
        vals = [gf.cell(row=row_idx, column=c).value for c in range(1, 5)]
        sym = vals[0]
        fin[sym] = {h: v for h, v in zip(fin_headers, vals)}

    gs = gt_wb["Summary"]
    summary_rows = []
    for row_idx in range(1, 6):
        summary_rows.append((gs.cell(row=row_idx, column=1).value,
                             gs.cell(row=row_idx, column=2).value))
    return {'performance': perf, 'financials': fin, 'summary': summary_rows}


def run_checks(agent_workspace: str, groundtruth_workspace: str):
    """Return (results, fatal_error).

    results: list of {"name","passed","critical","msg"}.
    fatal_error: str if the workbook could not be opened at all, else None.
    """
    agent_file = os.path.join(agent_workspace, "sector_rotation_report.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "sector_rotation_report.xlsx")

    if not os.path.exists(agent_file):
        return [], f"Missing agent file: {agent_file}"

    try:
        agent_wb = openpyxl.load_workbook(agent_file)
    except Exception as e:
        return [], f"Error loading agent workbook: {e}"

    expected = compute_expected_from_db()
    if expected is not None:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
    else:
        print("INFO: Falling back to static groundtruth Excel file")
        try:
            gt_wb = openpyxl.load_workbook(gt_file)
        except Exception as e:
            return [], f"Error loading groundtruth workbook: {e}"
        expected = _load_expected_from_gt(gt_wb)

    return _build_results(agent_wb, expected), None


def _num_close(a, b, abs_tol=1.0):
    try:
        return abs(float(a) - float(b)) <= abs_tol
    except (TypeError, ValueError):
        return False


def _num_close_rel(a, b, rel_tol=0.01):
    try:
        a = float(a); b = float(b)
    except (TypeError, ValueError):
        return False
    if b == 0:
        return abs(a) <= 1.0
    return abs(a - b) / abs(b) <= rel_tol


def _build_results(agent_wb, expected):
    results = []

    def add(name, passed, critical, msg=""):
        results.append({"name": name, "passed": bool(passed), "critical": critical, "msg": msg})

    # --- Structural: sheets present (NON-critical) ---
    required_sheets = ["Performance", "Financials", "Summary"]
    agent_sheet_map = {s.strip().lower(): s for s in agent_wb.sheetnames}
    for sheet_name in required_sheets:
        present = sheet_name.lower() in agent_sheet_map
        add(f"sheet_present:{sheet_name}", present, critical=False,
            msg="" if present else f"Missing sheet: {sheet_name}")

    perf = expected['performance']
    fin = expected['financials']
    summary = expected['summary']

    perf_headers = ['Symbol', 'Sector', 'Price_1Y_Ago', 'Current_Price', 'Return_Pct',
                    'Benchmark_Return_Pct', 'Alpha', 'Analyst_Rating', 'Target_Price', 'Upside_Pct']
    fin_headers = ['Symbol', 'Revenue_Latest_Q', 'Net_Income_Latest_Q', 'Profit_Margin_Pct']

    # --- Performance headers (NON-critical, structural) ---
    if "performance" in agent_sheet_map:
        agent_perf = agent_wb[agent_sheet_map["performance"]]
        agent_headers = [c.value for c in agent_perf[1]]
        hdr_ok = (len(agent_headers) == len(perf_headers) and
                  all(_str_match(e, a) for e, a in zip(perf_headers, agent_headers)))
        add("performance_headers", hdr_ok, critical=False,
            msg="" if hdr_ok else f"expected {perf_headers}, got {agent_headers}")

        # Per-symbol value checks
        for row_idx, sym in enumerate(STOCKS, start=2):
            gt_vals = perf[sym]
            arow = {h: agent_perf.cell(row=row_idx, column=ci + 1).value
                    for ci, h in enumerate(perf_headers)}

            # CRITICAL: Return_Pct (live moex prices)
            add(f"perf_return_pct:{sym}",
                _num_close(arow['Return_Pct'], gt_vals['Return_Pct'], 0.05),
                critical=True,
                msg=f"Return_Pct expected {gt_vals['Return_Pct']}, got {arow['Return_Pct']}")

            # CRITICAL: Alpha = Return_Pct - Benchmark (portal benchmark)
            add(f"perf_alpha:{sym}",
                _num_close(arow['Alpha'], gt_vals['Alpha'], 0.05),
                critical=True,
                msg=f"Alpha expected {gt_vals['Alpha']}, got {arow['Alpha']}")

            # CRITICAL: Analyst_Rating + Target_Price (scraped from portal)
            add(f"perf_rating:{sym}",
                _str_match(arow['Analyst_Rating'], gt_vals['Analyst_Rating']),
                critical=True,
                msg=f"Analyst_Rating expected {gt_vals['Analyst_Rating']}, got {arow['Analyst_Rating']}")
            add(f"perf_target_price:{sym}",
                _num_close(arow['Target_Price'], gt_vals['Target_Price'], 0.5),
                critical=True,
                msg=f"Target_Price expected {gt_vals['Target_Price']}, got {arow['Target_Price']}")

            # NON-critical: structural / derived columns
            add(f"perf_symbol:{sym}", _str_match(arow['Symbol'], gt_vals['Symbol']), False,
                msg=f"Symbol expected {gt_vals['Symbol']}, got {arow['Symbol']}")
            add(f"perf_sector:{sym}", _str_match(arow['Sector'], gt_vals['Sector']), False,
                msg=f"Sector expected {gt_vals['Sector']}, got {arow['Sector']}")
            add(f"perf_benchmark:{sym}",
                _num_close(arow['Benchmark_Return_Pct'], gt_vals['Benchmark_Return_Pct'], 0.05), False,
                msg=f"Benchmark expected {gt_vals['Benchmark_Return_Pct']}, got {arow['Benchmark_Return_Pct']}")
            add(f"perf_price_baseline:{sym}",
                _num_close(arow['Price_1Y_Ago'], gt_vals['Price_1Y_Ago'], 1.0), False,
                msg=f"Price_1Y_Ago expected {gt_vals['Price_1Y_Ago']}, got {arow['Price_1Y_Ago']}")
            add(f"perf_current_price:{sym}",
                _num_close(arow['Current_Price'], gt_vals['Current_Price'], 1.0), False,
                msg=f"Current_Price expected {gt_vals['Current_Price']}, got {arow['Current_Price']}")
            add(f"perf_upside:{sym}",
                _num_close(arow['Upside_Pct'], gt_vals['Upside_Pct'], 0.5), False,
                msg=f"Upside_Pct expected {gt_vals['Upside_Pct']}, got {arow['Upside_Pct']}")
    else:
        add("performance_headers", False, critical=False, msg="Performance sheet missing")

    # --- Financials (CRITICAL: revenue/net income/margin from annual income_stmt) ---
    if "financials" in agent_sheet_map:
        agent_fin = agent_wb[agent_sheet_map["financials"]]
        agent_fin_headers = [c.value for c in agent_fin[1]]
        fhdr_ok = (len(agent_fin_headers) == len(fin_headers) and
                   all(_str_match(e, a) for e, a in zip(fin_headers, agent_fin_headers)))
        add("financials_headers", fhdr_ok, critical=False,
            msg="" if fhdr_ok else f"expected {fin_headers}, got {agent_fin_headers}")

        for row_idx, sym in enumerate(STOCKS, start=2):
            gt_vals = fin[sym]
            arow = {h: agent_fin.cell(row=row_idx, column=ci + 1).value
                    for ci, h in enumerate(fin_headers)}
            add(f"fin_revenue:{sym}",
                _num_close_rel(arow['Revenue_Latest_Q'], gt_vals['Revenue_Latest_Q'], 0.01),
                critical=True,
                msg=f"Revenue expected {gt_vals['Revenue_Latest_Q']}, got {arow['Revenue_Latest_Q']}")
            add(f"fin_net_income:{sym}",
                _num_close_rel(arow['Net_Income_Latest_Q'], gt_vals['Net_Income_Latest_Q'], 0.01),
                critical=True,
                msg=f"Net_Income expected {gt_vals['Net_Income_Latest_Q']}, got {arow['Net_Income_Latest_Q']}")
            add(f"fin_margin:{sym}",
                _num_close(arow['Profit_Margin_Pct'], gt_vals['Profit_Margin_Pct'], 0.1), False,
                msg=f"Profit_Margin expected {gt_vals['Profit_Margin_Pct']}, got {arow['Profit_Margin_Pct']}")
    else:
        add("financials_headers", False, critical=False, msg="Financials sheet missing")

    # --- Summary (CRITICAL: best/worst performer; rest non-critical) ---
    if "summary" in agent_sheet_map:
        agent_sum = agent_wb[agent_sheet_map["summary"]]
        smap = {}
        for row_idx in range(1, 6):
            lbl = agent_sum.cell(row=row_idx, column=1).value
            val = agent_sum.cell(row=row_idx, column=2).value
            if lbl is not None:
                smap[str(lbl).strip().lower()] = val

        for gt_label, gt_value in summary:
            avalue = smap.get(str(gt_label).strip().lower())
            is_critical = gt_label in ("Best_Performer", "Worst_Performer")
            if isinstance(gt_value, str):
                ok = _str_match(avalue, gt_value)
            else:
                ok = _num_close(avalue, gt_value, 1.0)
            add(f"summary:{gt_label}", ok, critical=is_critical,
                msg=f"{gt_label} expected {gt_value}, got {avalue}")
    else:
        add("summary:present", False, critical=False, msg="Summary sheet missing")

    return results


# Backwards-compatible boolean entrypoint (returns (pass, error)).
def check_local(agent_workspace: str, groundtruth_workspace: str):
    results, fatal = run_checks(agent_workspace, groundtruth_workspace)
    if fatal:
        return False, fatal
    failed = [r for r in results if not r["passed"]]
    if failed:
        return False, "; ".join(f"{r['name']}: {r['msg']}" for r in failed[:10])
    return True, "All checks passed."
