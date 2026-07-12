"""Evaluation for canvas-assignment-workload-excel."""
import argparse
import json
import os
import sys
import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": "cowork_gym", "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}{msg}")
        if critical:
            CRITICAL_FAILED.append(name)


def load_gt_data(gt_dir):
    """Load derived groundtruth values from _data.json (no hardcoding)."""
    with open(os.path.join(gt_dir, "_data.json"), encoding="utf-8") as f:
        return json.load(f)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def int_close(a, b, tol=2):
    try:
        return abs(int(float(a)) - int(float(b))) <= tol
    except (TypeError, ValueError):
        return False


def check_excel(agent_workspace, gt_dir, gt_data):
    """Check Workload_Report.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Workload_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Workload_Report.xlsx")

    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False
    record("Excel file exists", True)

    if not os.path.isfile(gt_file):
        record("Groundtruth Excel exists", False, f"Not found: {gt_file}")
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        record("Excel files readable", False, str(e))
        return False

    all_ok = True

    def get_sheet(wb, target):
        for name in wb.sheetnames:
            if name.strip().lower() == target.strip().lower():
                return wb[name]
        return None

    # Sheet 1: Course Workload
    agent_ws = get_sheet(agent_wb, "Course Workload")
    gt_ws = get_sheet(gt_wb, "Course Workload")

    if agent_ws is None:
        record("Sheet 'Course Workload' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Course Workload' exists", True)
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))
        gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))

        record("Course Workload row count", abs(len(agent_rows) - len(gt_rows)) <= 2,
               f"Expected {len(gt_rows)}, got {len(agent_rows)}")

        gt_lookup = {}
        for r in gt_rows:
            if r and r[0]:
                gt_lookup[str(r[0]).strip().lower()] = r

        agent_lookup = {}
        for r in agent_rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        # Check top 5 courses
        for gt_row in gt_rows[:5]:
            if not gt_row or not gt_row[0]:
                continue
            key = str(gt_row[0]).strip().lower()
            a_row = agent_lookup.get(key)
            if a_row is None:
                record(f"Course '{gt_row[0][:40]}' present", False, "Missing")
                all_ok = False
                continue

            errors = []
            # Total_Assignments (col 1)
            if len(a_row) > 1 and len(gt_row) > 1:
                if not int_close(a_row[1], gt_row[1], 2):
                    errors.append(f"Assignments: {a_row[1]} vs {gt_row[1]}")

            # Avg_Submissions (col 2)
            if len(a_row) > 2 and len(gt_row) > 2:
                if not num_close(a_row[2], gt_row[2], 50.0):
                    errors.append(f"Avg_Submissions: {a_row[2]} vs {gt_row[2]}")

            if errors:
                record(f"Course '{gt_row[0][:40]}' data", False, "; ".join(errors))
                all_ok = False
            else:
                record(f"Course '{gt_row[0][:40]}' data", True)

        # CRITICAL: top course by Total_Assignments + its Avg_Submissions
        gt_cw = gt_data["course_workload"]
        top_course = gt_cw[0]
        top_key = top_course["course"].strip().lower()
        a_top = agent_lookup.get(top_key)
        if a_top is None:
            record(f"Top course '{top_course['course'][:40]}' present", False,
                   "Missing", critical=True)
            all_ok = False
        else:
            ok_cnt = len(a_top) > 1 and int_close(a_top[1], top_course["assignments"], 1)
            record(f"Top course Total_Assignments == {top_course['assignments']}",
                   ok_cnt, f"Got {a_top[1] if len(a_top) > 1 else None}", critical=True)
            if not ok_cnt:
                all_ok = False
            ok_avg = len(a_top) > 2 and num_close(a_top[2], top_course["avg_subs"], 5.0)
            record(f"Top course Avg_Submissions ~= {top_course['avg_subs']}",
                   ok_avg, f"Got {a_top[2] if len(a_top) > 2 else None}", critical=True)
            if not ok_avg:
                all_ok = False

    # Sheet 2: Busiest Periods
    agent_ws2 = get_sheet(agent_wb, "Busiest Periods")
    gt_ws2 = get_sheet(gt_wb, "Busiest Periods")

    if agent_ws2 is None:
        record("Sheet 'Busiest Periods' exists", False, f"Sheets: {agent_wb.sheetnames}")
        all_ok = False
    else:
        record("Sheet 'Busiest Periods' exists", True)
        agent_rows2 = list(agent_ws2.iter_rows(min_row=2, values_only=True))
        gt_rows2 = list(gt_ws2.iter_rows(min_row=2, values_only=True))

        record("Busiest Periods has data", len(agent_rows2) >= 5,
               f"Got {len(agent_rows2)} rows")

        # Check top 3 busiest months
        gt_months = {}
        for r in gt_rows2:
            if r and r[0]:
                gt_months[str(r[0]).strip()] = int(r[1]) if r[1] else 0

        agent_months = {}
        for r in agent_rows2:
            if r and r[0]:
                agent_months[str(r[0]).strip()] = int(float(r[1])) if r[1] else 0

        # CRITICAL: top-3 busiest months and counts, derived from groundtruth
        for bp in gt_data["busiest_periods"][:3]:
            month = str(bp["month"]).strip()
            a_count = agent_months.get(month)
            if a_count is None:
                record(f"Busiest month '{month}' present", False, "Missing",
                       critical=True)
                all_ok = False
            else:
                ok = int_close(a_count, bp["count"], 1)
                record(f"Busiest month '{month}' count == {bp['count']}", ok,
                       f"Got {a_count} vs {bp['count']}", critical=True)
                if not ok:
                    all_ok = False

    return all_ok


def check_gcal(gt_data):
    """Check calendar study sessions."""
    print("\n=== Checking Google Calendar ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT summary, description FROM gcal.events ORDER BY summary")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")

    study_events = [e for e in events if "study session" in (e[0] or "").lower()]
    record("At least 3 'Study Session' events found", len(study_events) >= 3,
           f"Found {len(study_events)} study session events", critical=True)

    # Derive the actual top-3 busiest months from groundtruth (no hardcoding)
    heavy_months = [str(bp["month"]).strip() for bp in gt_data["busiest_periods"][:3]]
    count_by_month = {str(bp["month"]).strip(): bp["count"]
                      for bp in gt_data["busiest_periods"]}

    all_ok = len(study_events) >= 3
    for month in heavy_months:
        # The month token must appear in the event title or description
        ev = next((e for e in events
                   if month in (e[0] or "") or month in (e[1] or "")), None)
        if ev is None:
            record(f"Study Session event for month {month}", False,
                   "Missing", critical=True)
            all_ok = False
            continue
        # Description must include the numeric assignment count for that month
        desc = (ev[1] or "")
        cnt_ok = str(count_by_month[month]) in desc
        record(f"Study Session {month} description mentions count "
               f"{count_by_month[month]}", cnt_ok, f"Desc: {desc[:80]}",
               critical=True)
        if not cnt_ok:
            all_ok = False

    return all_ok


def check_emails(gt_data):
    """Check email was sent to advisor."""
    print("\n=== Checking Emails ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    all_emails = cur.fetchall()
    cur.close()
    conn.close()

    # Derive expected values from groundtruth (no hardcoded literals)
    course_count = len(gt_data["course_workload"])
    heavy_months = [str(bp["month"]).strip() for bp in gt_data["busiest_periods"][:3]]

    found = False
    all_ok = False
    for subject, from_addr, to_addr, body in all_emails:
        subj_lower = (subject or "").lower()
        if "workload" in subj_lower or "assignment" in subj_lower:
            to_str = str(to_addr or "").lower()
            if "academic.advisor@university.example.com" not in to_str:
                continue
            found = True
            record("Email sent to academic.advisor@university.example.com",
                   True, f"To: {to_addr}", critical=True)

            body_text = body or ""
            cnt_ok = str(course_count) in body_text
            record(f"Email body mentions course count {course_count}", cnt_ok,
                   critical=True)
            months_ok = all(m in body_text for m in heavy_months)
            record(f"Email body mentions top-3 months {heavy_months}", months_ok,
                   critical=True)
            all_ok = cnt_ok and months_ok
            break

    if not found:
        record("Workload email to advisor found", False,
               f"Checked {len(all_emails)} emails", critical=True)
    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    if not os.path.isfile(os.path.join(gt_dir, "_data.json")):
        gt_dir = os.path.join(task_root, "groundtruth_workspace")

    gt_data = load_gt_data(gt_dir)

    excel_ok = check_excel(args.agent_workspace, gt_dir, gt_data)
    gcal_ok = check_gcal(gt_data)
    email_ok = check_emails(gt_data)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL checks failed: {CRITICAL_FAILED}")
        print("  Overall: FAIL (critical)")
        sys.exit(1)

    overall = accuracy >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
