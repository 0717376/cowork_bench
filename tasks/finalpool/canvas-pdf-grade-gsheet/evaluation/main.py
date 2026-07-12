"""
Evaluation script for canvas-pdf-grade-gsheet task.

Policy:
  - A set of CRITICAL_CHECKS encodes the core semantic deliverables (correct
    course averages + letter grades from the PDF scale, instructor-resolution
    rule, distinction/probation flags, the exact teacher emails, and the
    Summary statistics). ANY critical failure => overall FAIL via sys.exit(1),
    regardless of accuracy.
  - Otherwise PASS requires overall accuracy >= 70% across all checks
    (critical + structural soft checks from check_local / check_gsheet /
    check_email).

All identifiers (sheet/column names, course codes, emails, letter grades,
'Yes'/'No'/'N/A', subject template) are English on purpose and must stay so
to match the Canvas-provisioned LMS data and the substring-based checks.

Usage:
    python -m evaluation.main \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --res_log_file /path/to/result.json \
        --launch_time "2026-03-06 10:00:00"
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

from .check_local import check_local
from .check_gsheet import check_gsheet
from .check_email import check_email

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Groundtruth (Canvas is read-only / fixed-seed, so these are stable).
# Course_Code -> (Class_Average, Letter_Grade, Lead_Instructor, Instructor_Email)
GT_COURSES = {
    "BBB-2014B": (77.35, "C", "N/A", "N/A"),
    "CCC-2014B": (62.77, "D", "Д-р Кирилл Мельников", "caleb.miller@openuniversity.ac.uk"),
    "DDD-2014B": (66.47, "D", "Д-р Иван Мельников", "evan.miller@openuniversity.ac.uk"),
    "EEE-2014B": (78.80, "C", "Д-р Эмма Плотникова", "emma.wright@openuniversity.ac.uk"),
    "FFF-2014B": (74.67, "C", "Д-р Кирилл Моргунов", "caleb.morgan@openuniversity.ac.uk"),
    "GGG-2014B": (77.38, "C", "Д-р Лилия Карпова", "lily.carter@openuniversity.ac.uk"),
}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Excel Course Grades: 6 2014B courses with correct Class_Average + Letter_Grade",
    "Excel Course Grades: instructor resolution (BBB N/A; teachers + emails)",
    "Distinction/Probation all 'No' in Excel and Google Sheet",
    "Emails: exactly the 5 teacher courses to the correct instructors",
    "Summary statistics correct",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}" + (f" :: {detail}" if detail else ""))


def _num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def _s(x):
    return str(x).strip() if x is not None else ""


def _load_course_grades(agent_workspace):
    """Return {Course_Code: dict-of-columns} from the agent's Excel, or None."""
    agent_file = os.path.join(agent_workspace, "semester_grade_report.xlsx")
    if not os.path.isfile(agent_file):
        return None, None
    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception:
        return None, None
    ws = None
    for nm in wb.sheetnames:
        if nm.strip().lower() == "course grades":
            ws = wb[nm]
            break
    if ws is None:
        return None, wb
    cols = ["Course_Code", "Course_Name", "Lead_Instructor", "Instructor_Email",
            "Students_Scored", "Class_Average", "Letter_Grade", "Distinction", "Probation"]
    rows = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        code = _s(r[0]).upper()
        rows[code] = {c: (r[i] if i < len(r) else None) for i, c in enumerate(cols)}
    return rows, wb


def critical_checks(agent_workspace):
    """Run the strict semantic CRITICAL checks against the agent's Excel + DB."""
    print("\n=== CRITICAL CHECKS ===")
    rows, wb = _load_course_grades(agent_workspace)

    # --- CRITICAL 1: 6 courses with correct average + letter grade ---
    ok1 = rows is not None and len(rows) == 6
    detail1 = ""
    if rows is None:
        detail1 = "Excel 'Course Grades' sheet missing"
    else:
        for code, (gt_avg, gt_grade, _, _) in GT_COURSES.items():
            ck = code.upper()
            if ck not in rows:
                ok1 = False
                detail1 = f"{code} missing"
                break
            if not _num_close(rows[ck]["Class_Average"], gt_avg, 0.5):
                ok1 = False
                detail1 = f"{code} avg expected {gt_avg}, got {rows[ck]['Class_Average']}"
                break
            if _s(rows[ck]["Letter_Grade"]).upper() != gt_grade:
                ok1 = False
                detail1 = f"{code} grade expected {gt_grade}, got {rows[ck]['Letter_Grade']}"
                break
    check("Excel Course Grades: 6 2014B courses with correct Class_Average + Letter_Grade",
          ok1, detail1)

    # --- CRITICAL 2: instructor resolution rule ---
    ok2 = rows is not None
    detail2 = "" if rows is not None else "Course Grades missing"
    if rows is not None:
        for code, (_, _, gt_inst, gt_email) in GT_COURSES.items():
            ck = code.upper()
            if ck not in rows:
                ok2 = False
                detail2 = f"{code} missing"
                break
            if _s(rows[ck]["Lead_Instructor"]).lower() != gt_inst.lower():
                ok2 = False
                detail2 = f"{code} instructor expected '{gt_inst}', got '{rows[ck]['Lead_Instructor']}'"
                break
            if _s(rows[ck]["Instructor_Email"]).lower() != gt_email.lower():
                ok2 = False
                detail2 = f"{code} email expected '{gt_email}', got '{rows[ck]['Instructor_Email']}'"
                break
    check("Excel Course Grades: instructor resolution (BBB N/A; teachers + emails)",
          ok2, detail2)

    # --- CRITICAL 3: all Distinction/Probation == 'No' (Excel + Google Sheet) ---
    ok3 = rows is not None
    detail3 = "" if rows is not None else "Course Grades missing"
    if rows is not None:
        for code in GT_COURSES:
            ck = code.upper()
            if _s(rows[ck]["Distinction"]).lower() != "no" or _s(rows[ck]["Probation"]).lower() != "no":
                ok3 = False
                detail3 = f"{code} dist/prob not both 'No'"
                break
    if ok3:
        # verify in the Google Sheet too: no 'Yes' in the Grades sheet cells
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            c = conn.cursor()
            c.execute("""
                SELECT lower(value) FROM gsheet.cells cc
                JOIN gsheet.sheets s ON cc.sheet_id = s.id
                WHERE lower(s.title) = 'grades'
            """)
            vals = [r[0] for r in c.fetchall()]
            c.close()
            conn.close()
            if any(v == "yes" for v in vals):
                ok3 = False
                detail3 = "Google Sheet 'Grades' contains a 'Yes' distinction/probation value"
        except Exception as e:
            ok3 = False
            detail3 = f"gsheet query error: {e}"
    check("Distinction/Probation all 'No' in Excel and Google Sheet", ok3, detail3)

    # --- CRITICAL 4: exactly the 5 teacher emails to correct instructors ---
    teacher_emails = {
        code: email for code, (_, _, inst, email) in GT_COURSES.items() if inst != "N/A"
    }
    ok4 = True
    detail4 = ""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        c = conn.cursor()
        c.execute("SELECT subject, from_addr, to_addr FROM email.messages")
        msgs = c.fetchall()
        c.close()
        conn.close()
    except Exception as e:
        msgs = []
        ok4 = False
        detail4 = f"email query error: {e}"

    def _recips(to_addr):
        if isinstance(to_addr, list):
            return [str(x).lower() for x in to_addr]
        s = str(to_addr or "")
        try:
            p = json.loads(s)
            if isinstance(p, list):
                return [str(x).lower() for x in p]
        except Exception:
            pass
        return [s.lower()]

    if ok4:
        # BBB-2014B (no teacher) must NOT have an email
        for subj, _fr, _to in msgs:
            if "bbb-2014b" in str(subj or "").lower():
                ok4 = False
                detail4 = "email sent for BBB-2014B which has no teacher"
                break
    if ok4:
        for code, email in teacher_emails.items():
            matched = 0
            for subj, _fr, to_addr in msgs:
                sl = str(subj or "").lower()
                if code.lower() in sl and "end-of-semester grade report" in sl:
                    if any(email in r for r in _recips(to_addr)):
                        matched += 1
            if matched < 1:
                ok4 = False
                detail4 = f"no correct email for {code} -> {email}"
                break
    check("Emails: exactly the 5 teacher courses to the correct instructors", ok4, detail4)

    # --- CRITICAL 5: Summary statistics ---
    ok5 = wb is not None
    detail5 = "" if wb is not None else "workbook missing"
    if wb is not None:
        sws = None
        for nm in wb.sheetnames:
            if nm.strip().lower() == "summary":
                sws = wb[nm]
                break
        if sws is None:
            ok5 = False
            detail5 = "Summary sheet missing"
        else:
            summ = {}
            for r in sws.iter_rows(min_row=2, values_only=True):
                if r and r[0]:
                    summ[_s(r[0]).lower()] = r[1]
            expected = {
                "total_courses": (6, "int"),
                "avg_class_average": (72.91, "num"),
                "highest_average_course": ("EEE-2014B", "str"),
                "lowest_average_course": ("CCC-2014B", "str"),
                "distinction_count": (0, "int"),
                "probation_count": (0, "int"),
            }
            for k, (exp, kind) in expected.items():
                if k not in summ:
                    ok5 = False
                    detail5 = f"missing {k}"
                    break
                v = summ[k]
                if kind == "int":
                    if not _num_close(v, exp, 0.01):
                        ok5 = False; detail5 = f"{k} expected {exp}, got {v}"; break
                elif kind == "num":
                    if not _num_close(v, exp, 0.5):
                        ok5 = False; detail5 = f"{k} expected {exp}, got {v}"; break
                else:
                    if _s(v).lower() != exp.lower():
                        ok5 = False; detail5 = f"{k} expected {exp}, got {v}"; break
    check("Summary statistics correct", ok5, detail5)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run critical + soft checks; combine into accuracy + critical gate."""
    global PASS_COUNT, FAIL_COUNT

    # 0. CRITICAL semantic checks
    critical_checks(agent_workspace)

    # 1. Soft / structural checks (Excel groundtruth diff)
    print("\n=== Checking Excel Output (structural) ===")
    p, f, errs = check_local(agent_workspace, groundtruth_workspace)
    PASS_COUNT += p
    FAIL_COUNT += f
    print(f"  Excel: {p} passed, {f} failed")
    for e in errs:
        print(f"    [FAIL] {e[:200]}")

    # 2. Google Sheet
    print("\n=== Checking Google Sheet ===")
    p, f, errs = check_gsheet()
    PASS_COUNT += p
    FAIL_COUNT += f
    print(f"  Google Sheet: {p} passed, {f} failed")
    for e in errs:
        print(f"    [FAIL] {e[:200]}")

    # 3. Emails
    print("\n=== Checking Emails ===")
    p, f, errs = check_email()
    PASS_COUNT += p
    FAIL_COUNT += f
    print(f"  Emails: {p} passed, {f} failed")
    for e in errs:
        print(f"    [FAIL] {e[:200]}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if res_log_file:
        result = {
            "total_passed": PASS_COUNT,
            "total_checks": total,
            "accuracy": accuracy,
            "success": success,
            "critical_failed": critical_failed,
        }
        with open(res_log_file, "w") as fh:
            json.dump(result, fh, indent=2)

    return success, f"Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
