"""
Evaluation for rzd-team-novgorod-conference-excel-gcal-email-teamly.

Checks:
1. Conference_Travel_Plan.xlsx existsand readable
2. "Outbound" sheet has at least 4 data rows
3. "Outbound" sheet has Train_No column containing 818А (московский) или 822А (питерский)
4. "Return" sheet has at least 4 data rows
5. "Coordination_Notes" sheet has at least 3 rows
6. GCal has at least 2 new travel events (помимо самой конференции)
7. Email sent to moscow_team@uni.ru
8. Email sent to spb_team@uni.ru
9. Teamly: создана страница в пространстве TRIPS про поездку
10. Groundtruth row-by-row сравнение
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Excel: Conference_Travel_Plan.xlsx ===")
    import glob

    pattern = os.path.join(agent_workspace, "*.xlsx")
    all_xlsx = glob.glob(pattern)
    conf_files = [f for f in all_xlsx if any(
        kw in os.path.basename(f).lower()
        for kw in ["conference", "travel", "novgorod", "новгород", "plan", "план"]
    )]

    if not conf_files:
        record("Conference travel xlsx существует", False,
               f"Нет подходящего xlsx в {agent_workspace}")
        return
    record("Conference travel xlsx существует", True)

    xlsx_path = conf_files[0]
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel читается", False, str(e))
        return
    record("Excel читается", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Outbound
    if "outbound" not in sheet_names_lower:
        record("Лист Outbound существует", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Лист Outbound существует", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("outbound")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Outbound ≥ 4 строк данных", len(data_rows) >= 4,
               f"Найдено {len(data_rows)} строк")

        all_text = " ".join(str(c) for r in rows for c in r if c).lower()
        # Принимаем оба варианта: с кириллицей и латиницей
        has_train = any(t in all_text for t in ["818а", "818a", "822а", "822a"])
        record("Outbound содержит номер 818А или 822А", has_train,
               f"Text head: {all_text[:200]}")

    # Return
    if "return" not in sheet_names_lower:
        record("Лист Return существует", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Лист Return существует", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("return")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Return ≥ 4 строк данных", len(data_rows) >= 4,
               f"Найдено {len(data_rows)} строк")

    # Coordination_Notes
    coord_match = [s for s in sheet_names_lower if "coord" in s or "координ" in s or "note" in s]
    if not coord_match:
        record("Лист Coordination_Notes существует", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Лист Coordination_Notes существует", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index(coord_match[0])]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Coordination_Notes ≥ 3 строк", len(data_rows) >= 3,
               f"Найдено {len(data_rows)} строк")

    # --- Groundtruth row-by-row сравнение (нестрогое) ---
    gt_path = os.path.join(groundtruth_workspace, "Conference_Travel_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx существует", False, gt_path)
        return

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT-лист '{gt_sheet_name}' есть у агента", False, f"Доступно: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        record(f"GT '{gt_sheet_name}' число строк", len(agent_rows) == len(gt_rows),
               f"Ожидаем {len(gt_rows)}, получено {len(agent_rows)}")

        # Для Outbound/Return — проверяем только колонку Train_No (col 3, индекс 2) в первых 2 строках.
        # Для Coordination_Notes — проверяем колонку Check_Point (col 1) — менее строго (str_match).
        check_indices = list(range(min(2, len(gt_rows))))
        for idx in check_indices:
            gt_row = gt_rows[idx]
            if idx < len(agent_rows):
                a_row = agent_rows[idx]
                if gt_sheet_name.lower() in ("outbound", "return"):
                    cols_to_check = [2]
                else:
                    cols_to_check = [0]
                row_ok = True
                for col_idx in cols_to_check:
                    if col_idx >= len(gt_row) or col_idx >= len(a_row):
                        continue
                    gt_val = gt_row[col_idx]
                    a_val = a_row[col_idx]
                    if gt_val is None:
                        continue
                    if cols_to_check == [2]:
                        # train_no — нечёткое сравнение: убираем кириллицу/латиницу
                        norm_gt = str(gt_val).strip().lower().replace('а', 'a')
                        norm_a = str(a_val).strip().lower().replace('а', 'a') if a_val else ''
                        ok = norm_a == norm_gt
                    else:
                        # Нестрогое сравнение Check_Point: совпадение по 1-2 ключевым словам
                        gt_words = [w for w in str(gt_val).lower().split() if len(w) > 3]
                        a_text = str(a_val or '').lower()
                        hits = sum(1 for w in gt_words if w in a_text)
                        ok = str_match(a_val, gt_val) or hits >= 1
                    if not ok:
                        record(f"GT '{gt_sheet_name}' строка {idx+1} col {col_idx+1}",
                               False, f"Ожидаем {gt_val}, получено {a_val}")
                        row_ok = False
                        break
                if row_ok:
                    record(f"GT '{gt_sheet_name}' строка {idx+1} значения совпали", True)
            else:
                record(f"GT '{gt_sheet_name}' строка {idx+1} существует", False, "Строка отсутствует у агента")
    gt_wb.close()


def check_gcal():
    print("\n=== Calendar: события поездки ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
        WHERE (start_datetime >= '2026-03-12' AND start_datetime < '2026-03-16')
          AND summary NOT ILIKE '%древнерусск%'
          AND summary NOT ILIKE '%конференц%'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()
    record("Не менее 2 новых событий поездки", len(events) >= 2,
           f"Найдено {len(events)} событий: {[e[0] for e in events]}")


def _query_email_count(addr_pattern):
    cnt = 0
    sent = 0
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT COUNT(*) FROM email.messages WHERE to_addr::text ILIKE %s"
            " AND from_addr NOT ILIKE %s",
            (f"%{addr_pattern}%", f"%{addr_pattern}%"),
        )
        cnt = cur.fetchone()[0]
        cur.close()
        conn.close()
    except Exception:
        pass
    try:
        conn2 = psycopg2.connect(**DB_CONFIG)
        cur2 = conn2.cursor()
        cur2.execute(
            "SELECT COUNT(*) FROM email.sent_log WHERE to_addr::text ILIKE %s",
            (f"%{addr_pattern}%",),
        )
        sent = cur2.fetchone()[0]
        cur2.close()
        conn2.close()
    except Exception:
        pass
    return cnt, sent


def check_emails():
    print("\n=== Email: письма группам ===")

    moscow_cnt, moscow_sent = _query_email_count("moscow_team@uni.ru")
    record("Письмо отправлено на moscow_team@uni.ru", moscow_cnt >= 1 or moscow_sent >= 1,
           f"messages: {moscow_cnt}, sent_log: {moscow_sent}")

    spb_cnt, spb_sent = _query_email_count("spb_team@uni.ru")
    record("Письмо отправлено на spb_team@uni.ru", spb_cnt >= 1 or spb_sent >= 1,
           f"messages: {spb_cnt}, sent_log: {spb_sent}")


def check_teamly():
    print("\n=== Teamly: страница в TRIPS ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT p.id, p.title FROM teamly.pages p
            JOIN teamly.spaces s ON s.id = p.space_id
            WHERE s.key = 'TRIPS'
              AND (p.title ILIKE '%новгород%' OR p.title ILIKE '%конференц%'
                   OR p.body ILIKE '%818%' OR p.body ILIKE '%822%')
              AND p.id > 3
            ORDER BY p.id DESC
        """)
        pages = cur.fetchall()
        cur.close()
        conn.close()
        record("Создана страница в Teamly/TRIPS", len(pages) >= 1,
               f"Найдено {len(pages)} страниц: {[p[1] for p in pages]}")
    except Exception as e:
        record("Создана страница в Teamly/TRIPS", False, f"DB error: {e}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_gcal()
    check_emails()
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
