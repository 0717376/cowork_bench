"""Evaluation for insales-product-inventory-report."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "WC_Inventory_Report.xlsx")
    gt_file = os.path.join(gt_dir, "WC_Inventory_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # ---------------------------------------------------------------
    # CRITICAL semantic checks: any failure => immediate FAIL.
    # These verify the core deliverable (live aggregates, stock policy,
    # explicit descending-sort requirement, sheet structure).
    # ---------------------------------------------------------------
    critical_failures = []

    inv_rows = load_sheet_rows(agent_wb, "Inventory Report")
    sum_rows = load_sheet_rows(agent_wb, "Summary")

    # Structural: both required sheets present
    if inv_rows is None:
        critical_failures.append("CRITICAL: sheet 'Inventory Report' missing")
    if sum_rows is None:
        critical_failures.append("CRITICAL: sheet 'Summary' missing")

    # Summary aggregates (computed from live store data) must match groundtruth.
    if sum_rows is not None:
        gt_sum = load_sheet_rows(gt_wb, "Summary") or []
        gt_metrics = {str(r[0]).strip().lower(): r[1] for r in gt_sum[1:]
                      if r and r[0] is not None}
        a_metrics = {str(r[0]).strip().lower(): r[1] for r in sum_rows[1:]
                     if r and r[0] is not None}
        for metric, tol in (("total_products", 0), ("on_sale_count", 0),
                            ("out_of_stock", 0), ("avg_price", 1.0)):
            if metric not in a_metrics:
                critical_failures.append(f"CRITICAL: Summary metric '{metric}' missing")
            elif not num_close(a_metrics.get(metric), gt_metrics.get(metric), tol):
                critical_failures.append(
                    f"CRITICAL: Summary {metric}={a_metrics.get(metric)} "
                    f"!= expected {gt_metrics.get(metric)} (tol={tol})")

    if inv_rows is not None:
        a_inv = inv_rows[1:] if len(inv_rows) > 1 else []

        # Descending sort by Regular_Price (column index 2) — explicit requirement.
        prices = [r[2] for r in a_inv if r and r[2] is not None]
        try:
            fp = [float(p) for p in prices]
            if fp != sorted(fp, reverse=True):
                critical_failures.append(
                    "CRITICAL: 'Inventory Report' not sorted by Regular_Price descending")
        except (TypeError, ValueError):
            critical_failures.append("CRITICAL: non-numeric Regular_Price values")

        # Stock policy: every Stock_Qty==0 row must be classified 'outofstock'.
        for r in a_inv:
            if not r or len(r) < 6:
                continue
            qty, status = r[4], r[5]
            try:
                is_zero = qty is not None and float(qty) == 0
            except (TypeError, ValueError):
                is_zero = False
            if is_zero and not str_match(status, "outofstock"):
                critical_failures.append(
                    f"CRITICAL: Stock_Qty==0 but Stock_Status='{status}' "
                    f"(expected 'outofstock') for product '{r[0]}'")
                break

    if critical_failures:
        print("\n=== CRITICAL CHECKS FAILED ===")
        for c in critical_failures:
            print(f"  {c}")
        print("=== RESULT: FAIL ===")
        sys.exit(1)

    total_checks = 0

    # Check sheet: Inventory Report
    print(f"  Checking Inventory Report...")
    a_rows = load_sheet_rows(agent_wb, "Inventory Report")
    g_rows = load_sheet_rows(gt_wb, "Inventory Report")
    if a_rows is None:
        all_errors.append("Sheet 'Inventory Report' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Inventory Report' not found in groundtruth")
    else:
        sheet_name = "Inventory Report"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                total_checks += 1
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 2 and len(g_row) > 2:
                total_checks += 1
                if not num_close(a_row[2], g_row[2], 1.0):
                    errors.append(f"{key}.Regular_Price: {a_row[2]} vs {g_row[2]} (tol=1.0)")

            if len(a_row) > 4 and len(g_row) > 4:
                total_checks += 1
                if not num_close(a_row[4], g_row[4], 2):
                    errors.append(f"{key}.Stock_Qty: {a_row[4]} vs {g_row[4]} (tol=2)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                total_checks += 1
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                total_checks += 1
                if not num_close(a_row[1], g_row[1], 5.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=5.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    

    # Accuracy gate: critical checks already passed above. Tolerate minor
    # per-row noise on the loosely-checked columns; require >=70% correct.
    passed = max(0, total_checks - len(all_errors))
    accuracy = (passed / total_checks * 100.0) if total_checks else 100.0
    print(f"\nAccuracy: {passed}/{total_checks} = {accuracy:.1f}% (threshold 70%)")
    if all_errors:
        print(f"Errors ({len(all_errors)}):")
        for e in all_errors[:10]:
            print(f"  {e}")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
