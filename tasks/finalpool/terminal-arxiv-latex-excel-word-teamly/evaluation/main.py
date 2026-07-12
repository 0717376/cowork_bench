"""Evaluation for terminal-arxiv-latex-excel-word-teamly.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Checks:
1. Literature_Review_Matrix.xlsx with 3 sheets (Paper_Catalog semantic content,
   Methodology_Comparison, Citation_Network edges)
2. Literature_Review_Draft.docx (RU or EN prose accepted)
3. Teamly knowledge base "Transformer Research Papers"
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# The 5 transformer papers the agent must catalog (English identifiers preserved).
EXPECTED = {
    "1706.03762": {"year": 2017, "cat": "cs.CL", "author": "vaswani"},
    "1810.04805": {"year": 2018, "cat": "cs.CL", "author": "devlin"},
    "2005.14165": {"year": 2020, "cat": "cs.CL", "author": "brown"},
    "2010.11929": {"year": 2020, "cat": "cs.CV", "author": "dosovitskiy"},
    "2301.07041": {"year": 2023, "cat": "cs.LG", "author": "kaplan"},
}
# Noise papers that MUST be filtered out of the catalog.
NOISE_IDS = {"1901.02860", "2002.05709", "2106.09685"}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Paper_Catalog: 5 transformer IDs present and 3 noise IDs absent",
    "Paper_Catalog: Year matches source for all 5 papers",
    "Paper_Catalog: Primary_Category matches source for all 5 papers",
    "Paper_Catalog: Authors holds correct first-author last name",
    "Citation_Network: BERT->Transformer and GPT-3->Transformer edges present",
    "Teamly 'Transformer Research Papers' KB has >= 5 entries with year+category",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def _hmap(ws):
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return {h: i for i, h in enumerate(headers)}


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Literature_Review_Matrix.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Literature_Review_Matrix.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Literature_Review_Matrix.xlsx")

    exists = os.path.isfile(agent_file)
    check("Excel file exists", exists, agent_file)
    if not exists:
        # mark dependent critical checks failed
        check("Paper_Catalog: 5 transformer IDs present and 3 noise IDs absent", False, "no excel")
        check("Paper_Catalog: Year matches source for all 5 papers", False, "no excel")
        check("Paper_Catalog: Primary_Category matches source for all 5 papers", False, "no excel")
        check("Paper_Catalog: Authors holds correct first-author last name", False, "no excel")
        check("Citation_Network: BERT->Transformer and GPT-3->Transformer edges present", False, "no excel")
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True) if os.path.isfile(gt_file) else None
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # --- Paper_Catalog ---
    print("  Checking Paper_Catalog...")
    a_sheet = get_sheet(agent_wb, "Paper_Catalog")
    check("Sheet 'Paper_Catalog' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet is not None:
        hm = _hmap(a_sheet)
        id_i = hm.get("arxiv_id", 0)
        title_i = hm.get("title", 1)
        auth_i = hm.get("authors", 2)
        year_i = hm.get("year", 3)
        cat_i = hm.get("primary_category", 4)
        a_rows = [r for r in a_sheet.iter_rows(min_row=2, values_only=True)
                  if r and any(c is not None for c in r)]
        check("Paper_Catalog has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")

        a_lookup = {}
        for r in a_rows:
            if id_i < len(r) and r[id_i] is not None:
                a_lookup[str(r[id_i]).strip()] = r
        a_ids = set(a_lookup.keys())

        # CRITICAL: 5 expected present, 3 noise absent.
        present_ok = all(eid in a_ids for eid in EXPECTED)
        noise_ok = all(nid not in a_ids for nid in NOISE_IDS)
        check("Paper_Catalog: 5 transformer IDs present and 3 noise IDs absent",
              present_ok and noise_ok,
              f"present={present_ok}, noise_absent={noise_ok}, ids={sorted(a_ids)}")

        # CRITICAL: Year matches source.
        year_ok = True
        for eid, meta in EXPECTED.items():
            r = a_lookup.get(eid)
            if not r or year_i >= len(r) or not num_close(r[year_i], meta["year"], 0):
                year_ok = False
                break
        check("Paper_Catalog: Year matches source for all 5 papers", year_ok,
              "year mismatch")

        # CRITICAL: Primary_Category matches source.
        cat_ok = True
        for eid, meta in EXPECTED.items():
            r = a_lookup.get(eid)
            val = str(r[cat_i]).strip().lower() if r and cat_i < len(r) and r[cat_i] is not None else ""
            if val != meta["cat"].lower():
                cat_ok = False
                break
        check("Paper_Catalog: Primary_Category matches source for all 5 papers", cat_ok,
              "category mismatch")

        # CRITICAL: Authors first-author last name.
        auth_ok = True
        for eid, meta in EXPECTED.items():
            r = a_lookup.get(eid)
            val = str(r[auth_i]).strip().lower() if r and auth_i < len(r) and r[auth_i] is not None else ""
            if meta["author"] not in val:
                auth_ok = False
                break
        check("Paper_Catalog: Authors holds correct first-author last name", auth_ok,
              "author mismatch")
    else:
        check("Paper_Catalog: 5 transformer IDs present and 3 noise IDs absent", False, "no sheet")
        check("Paper_Catalog: Year matches source for all 5 papers", False, "no sheet")
        check("Paper_Catalog: Primary_Category matches source for all 5 papers", False, "no sheet")
        check("Paper_Catalog: Authors holds correct first-author last name", False, "no sheet")

    # --- Methodology_Comparison ---
    print("  Checking Methodology_Comparison...")
    a_sheet = get_sheet(agent_wb, "Methodology_Comparison")
    check("Sheet 'Methodology_Comparison' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet is not None:
        hm = _hmap(a_sheet)
        id_i = hm.get("arxiv_id", 0)
        m_i = hm.get("method_name", 1)
        ap_i = hm.get("approach", 2)
        a_rows = [r for r in a_sheet.iter_rows(min_row=2, values_only=True)
                  if r and any(c is not None for c in r)]
        check("Methodology_Comparison has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")
        a_ids = {str(r[id_i]).strip() for r in a_rows if id_i < len(r) and r[id_i]}
        for eid in EXPECTED:
            check(f"Method for '{eid}'", eid in a_ids, f"Missing from {a_ids}")
        for r in a_rows:
            rid = str(r[id_i]).strip() if id_i < len(r) and r[id_i] else ""
            if rid in EXPECTED:
                has_method = m_i < len(r) and r[m_i] is not None and len(str(r[m_i]).strip()) > 0
                has_approach = ap_i < len(r) and r[ap_i] is not None and len(str(r[ap_i]).strip()) > 0
                check(f"'{rid}' has method and approach", has_method and has_approach,
                      f"method={r[m_i] if m_i < len(r) else None}")

    # --- Citation_Network ---
    print("  Checking Citation_Network...")
    a_sheet = get_sheet(agent_wb, "Citation_Network")
    check("Sheet 'Citation_Network' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet is not None:
        hm = _hmap(a_sheet)
        cing_i = hm.get("citing_paper", 0)
        cited_i = hm.get("cited_paper", 1)
        a_rows = [r for r in a_sheet.iter_rows(min_row=2, values_only=True)
                  if r and any(c is not None for c in r)]
        check("Citation_Network has >= 5 rows", len(a_rows) >= 5, f"Got {len(a_rows)}")
        citations = set()
        for r in a_rows:
            if cing_i < len(r) and cited_i < len(r) and r[cing_i] and r[cited_i]:
                citations.add((str(r[cing_i]).strip(), str(r[cited_i]).strip()))
        # CRITICAL: the two methodological edges grounded in seeded LaTeX.
        edges_ok = (("1810.04805", "1706.03762") in citations
                    and ("2005.14165", "1706.03762") in citations)
        check("Citation_Network: BERT->Transformer and GPT-3->Transformer edges present",
              edges_ok, f"citations: {sorted(citations)[:8]}")
    else:
        check("Citation_Network: BERT->Transformer and GPT-3->Transformer edges present",
              False, "no sheet")


def check_word(agent_workspace):
    print("\n=== Checking Literature_Review_Draft.docx ===")
    docx_path = os.path.join(agent_workspace, "Literature_Review_Draft.docx")
    check("Literature_Review_Draft.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        # Content bar raised above the trivial len>300 threshold.
        check("Document has substantial content", len(text) > 800, f"Length: {len(text)}")
        # transformer/attention are standard EN terms in RU ML prose; accept RU alts.
        check("Contains transformer reference",
              "transformer" in text or "трансформер" in text)
        check("Contains attention reference",
              "attention" in text or "self-attention" in text or "внимани" in text)
        check("Contains BERT or GPT reference",
              "bert" in text or "gpt" in text)
        check("Contains methodology comparison",
              any(k in text for k in ("method", "approach", "architecture",
                                      "метод", "подход", "архитектур")))
    except ImportError:
        check("python-docx available", False)
    except Exception as e:
        check("Word document readable", False, str(e))


def check_teamly():
    print("\n=== Checking Teamly Knowledge Base ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # New (agent-created) pages have id > 3; seed pages are id <= 3.
        cur.execute("SELECT id, title, COALESCE(body,'') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Transformer Research Papers' KB has >= 5 entries with year+category",
              False, str(e))
        return

    # Exclude the seeded noise page.
    pages = [p for p in pages if "архив протоколов" not in (p[1] or "").lower()]

    def is_tr(title):
        t = (title or "").lower()
        en = ("transformer" in t and ("research" in t or "paper" in t))
        ru = ("трансформер" in t and ("стат" in t or "исследован" in t))
        return en or ru

    # The "knowledge base" is realized as a hub/database page plus per-paper
    # entries, OR a set of per-paper pages whose titles carry the paper title.
    # Identify entries: pages that carry a paper title/ID and expose year+category.
    paper_titles = {
        "1706.03762": "attention is all you need",
        "1810.04805": "bert",
        "2005.14165": ("gpt-3", "few-shot"),
        "2010.11929": ("16x16", "image"),
        "2301.07041": "scaling laws",
    }

    def matches_paper(text):
        for pid, marker in paper_titles.items():
            markers = marker if isinstance(marker, tuple) else (marker,)
            if pid in text or any(m in text for m in markers):
                return True
        return False

    entries = []
    for pid, title, body in pages:
        text = ((title or "") + " " + (body or "")).lower()
        # An entry must reference a specific paper AND expose year + category.
        has_year = any(str(y) in text for y in (2017, 2018, 2020, 2023))
        has_cat = any(c in text for c in ("cs.cl", "cs.cv", "cs.lg"))
        if matches_paper(text) and has_year and has_cat:
            entries.append((pid, title))

    # A 'Transformer Research Papers' container page should also exist (non-critical).
    has_container = any(is_tr(t) for _, t, _ in pages)
    check("Teamly 'Transformer Research Papers' container page exists",
          has_container,
          f"new pages: {[(p[0], p[1]) for p in pages][:10]}")

    # CRITICAL: >= 5 entries each exposing non-empty year + category.
    check("Teamly 'Transformer Research Papers' KB has >= 5 entries with year+category",
          len(entries) >= 5,
          f"valid entries: {len(entries)} -> {entries}; all new pages: {[(p[0], p[1]) for p in pages][:10]}")


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Literature_Review_Matrix.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            expected_sheets = {"paper_catalog", "methodology_comparison", "citation_network"}
            actual_sheets = {s.strip().lower().replace(" ", "_") for s in wb.sheetnames}
            unexpected = actual_sheets - expected_sheets
            check("No unexpected sheets in Excel", len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_teamly()
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    overall = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
