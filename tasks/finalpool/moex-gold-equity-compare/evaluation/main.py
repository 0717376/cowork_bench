"""Evaluation for yf-gold-equity-compare (MOEX re-theme).

CRITICAL_CHECKS: semantic correctness of the core deliverable. Any critical
failure => sys.exit(1) before the accuracy gate. Structural checks are
non-critical. PASS requires no critical failure AND accuracy >= 70.
"""
import argparse
import json
import os
import sys

import openpyxl

try:
    import psycopg2
except Exception:
    psycopg2 = None

DB = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# The 6 globally-seeded MOEX blue-chip tickers.
EXPECTED_TICKERS = {"SBER.ME", "GAZP.ME", "LKOH.ME", "TCSG.ME", "MGNT.ME", "MTSS.ME"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


SUMMARY_KEYS = {"total_assets", "widest_range", "narrowest_range"}


def summary_value(rows, metric):
    """Header/orientation-tolerant lookup of a Summary metric value.

    Scans ALL cells (no header-row assumption); the value is the cell to the
    right, or the cell below when the right neighbour is itself another metric
    key (horizontal layout). A 'Metric/Value' header never collides with keys.
    """
    if not rows:
        return None
    key = metric.strip().lower()

    def _is_key(v):
        return v is not None and str(v).strip().lower() in SUMMARY_KEYS

    for r, row in enumerate(rows):
        for c, cell in enumerate(row or []):
            if cell is None or str(cell).strip().lower() != key:
                continue
            right = row[c + 1] if c + 1 < len(row) else None
            below = None
            if r + 1 < len(rows) and rows[r + 1] and c < len(rows[r + 1]):
                below = rows[r + 1][c]
            if right is not None and not _is_key(right):
                return right
            if below is not None and not _is_key(below):
                return below
    return None


def check_email(widest, narrowest, critical, results):
    """Verify an email to portfolio@investment.com exists with the required
    subject and a body that names the correct widest/narrowest tickers."""
    if psycopg2 is None:
        results.append(("Email check (psycopg2 unavailable)", False, ""))
        critical.append(False)
        return
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        messages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        results.append((f"Email check (DB error: {e})", False, ""))
        critical.append(False)
        return

    def to_addresses(to_addr):
        if isinstance(to_addr, list):
            return " ".join(str(r).lower() for r in to_addr)
        if to_addr:
            try:
                parsed = json.loads(str(to_addr))
                if isinstance(parsed, list):
                    return " ".join(str(r).lower() for r in parsed)
                return str(to_addr).lower()
            except Exception:
                return str(to_addr).lower()
        return ""

    recips = [m for m in messages if "portfolio@investment.com" in to_addresses(m[2])]
    ok_recipient = len(recips) > 0
    results.append(("Email to portfolio@investment.com exists", ok_recipient,
                    f"matching messages: {len(recips)} / total: {len(messages)}"))
    critical.append(ok_recipient)
    if not ok_recipient:
        return

    subj = (recips[0][0] or "")
    body = (recips[0][3] or "")
    ok_subject = "gold vs equity comparison" in subj.strip().lower()
    results.append(("Email subject == 'Gold vs Equity Comparison'", ok_subject, f"subject: {subj!r}"))
    critical.append(ok_subject)

    content = (subj + " " + body).lower()
    # Tickers may be written with or without the .ME suffix; match the base symbol.
    w_base = widest.lower().replace(".me", "")
    n_base = narrowest.lower().replace(".me", "")
    ok_widest = w_base in content
    ok_narrowest = n_base in content
    results.append((f"Email body names widest ticker ({widest})", ok_widest, ""))
    results.append((f"Email body names narrowest ticker ({narrowest})", ok_narrowest, ""))
    critical.append(ok_widest)
    critical.append(ok_narrowest)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "YF_Gold_Equity_Compare.xlsx")
    gt_file = os.path.join(gt_dir, "YF_Gold_Equity_Compare.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL (CRITICAL): Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    results = []          # (label, ok, detail) — non-critical structural/accuracy checks
    critical = []         # list of bools; any False => hard fail

    # --- Load groundtruth Summary to know the correct widest/narrowest ---
    g_summary = load_sheet_rows(gt_wb, "Summary")
    gt_widest = str(summary_value(g_summary, "Widest_Range") or "").strip()
    gt_narrowest = str(summary_value(g_summary, "Narrowest_Range") or "").strip()

    # ================= Asset Comparison =================
    a_rows = load_sheet_rows(agent_wb, "Asset Comparison")
    g_rows = load_sheet_rows(gt_wb, "Asset Comparison")
    if a_rows is None:
        results.append(("Sheet 'Asset Comparison' present", False, ""))
        critical.append(False)
    elif g_rows is None:
        print("FAIL: Sheet 'Asset Comparison' not found in groundtruth")
        sys.exit(1)
    else:
        results.append(("Sheet 'Asset Comparison' present", True, ""))
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        a_symbols = []
        for row in a_data:
            if row and row[0] is not None:
                key = str(row[0]).strip().upper()
                a_lookup[key] = row
                a_symbols.append(key)

        # CRITICAL: exactly the 6 seeded MOEX tickers present.
        present = set(a_lookup.keys())
        tickers_ok = present == EXPECTED_TICKERS
        results.append(("Exactly the 6 seeded MOEX tickers present", tickers_ok,
                        f"got: {sorted(present)}"))
        critical.append(tickers_ok)

        # NON-critical structural: sorted by Symbol.
        sorted_ok = a_symbols == sorted(a_symbols)
        results.append(("Asset Comparison sorted by Symbol", sorted_ok, f"order: {a_symbols}"))

        # Per-ticker numeric checks (tightened tolerances; non-critical accuracy signal).
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().upper()
            a_row = a_lookup.get(key)
            if a_row is None:
                results.append((f"{key} row present", False, ""))
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                ok = num_close(a_row[1], g_row[1], 5.0)
                results.append((f"{key}.Avg_Price", ok, f"{a_row[1]} vs {g_row[1]} (tol=5.0)"))
            if len(a_row) > 2 and len(g_row) > 2:
                ok = num_close(a_row[2], g_row[2], 0.5)
                results.append((f"{key}.Price_Range_Pct", ok, f"{a_row[2]} vs {g_row[2]} (tol=0.5)"))
            if len(a_row) > 3 and len(g_row) > 3:
                ok = num_close(a_row[3], g_row[3], 2)
                results.append((f"{key}.Data_Points", ok, f"{a_row[3]} vs {g_row[3]} (tol=2)"))

    # ================= Summary =================
    a_summary = load_sheet_rows(agent_wb, "Summary")
    if a_summary is None:
        results.append(("Sheet 'Summary' present", False, ""))
        critical.append(False)
    else:
        results.append(("Sheet 'Summary' present", True, ""))

        a_total = summary_value(a_summary, "Total_Assets")
        total_ok = num_close(a_total, 6, 0)
        results.append(("Summary.Total_Assets == 6", total_ok, f"got: {a_total}"))
        critical.append(total_ok)

        a_widest = str(summary_value(a_summary, "Widest_Range") or "").strip()
        a_narrowest = str(summary_value(a_summary, "Narrowest_Range") or "").strip()

        # CRITICAL: exact widest/narrowest ticker strings.
        widest_ok = str_match(a_widest, gt_widest)
        narrowest_ok = str_match(a_narrowest, gt_narrowest)
        results.append((f"Summary.Widest_Range == {gt_widest}", widest_ok, f"got: {a_widest!r}"))
        results.append((f"Summary.Narrowest_Range == {gt_narrowest}", narrowest_ok, f"got: {a_narrowest!r}"))
        critical.append(widest_ok)
        critical.append(narrowest_ok)

    # ================= Email (CRITICAL) =================
    check_email(gt_widest, gt_narrowest, critical, results)

    # ===== Report =====
    print("=== Checks ===")
    for label, ok, detail in results:
        print(f"  [{'PASS' if ok else 'FAIL'}] {label}" + (f"  ({detail})" if detail else ""))

    critical_failed = critical.count(False)
    total = len(results)
    passed = sum(1 for _, ok, _ in results if ok)
    accuracy = (passed / total * 100) if total else 0.0
    print(f"\nCritical checks: {len(critical) - critical_failed}/{len(critical)} passed")
    print(f"Accuracy: {passed}/{total} = {accuracy:.1f}%")

    if critical_failed > 0:
        print(f"\n=== RESULT: FAIL ({critical_failed} CRITICAL check(s) failed) ===")
        sys.exit(1)
    if accuracy < 70:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)

    print("\n=== RESULT: PASS ===")
    sys.exit(0)


if __name__ == "__main__":
    main()
