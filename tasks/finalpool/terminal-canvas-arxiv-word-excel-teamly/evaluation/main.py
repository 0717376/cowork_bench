"""Evaluation for terminal-canvas-arxiv-word-excel-teamly.

Gate: any CRITICAL_CHECKS failure => overall FAIL regardless of accuracy.
Otherwise PASS requires accuracy >= 70%.

The Teamly tracker is a Confluence-style page (teamly.pages: title + body),
not a typed Notion database. The agent renders the tracker as a markdown table
in the page body; checks read the page text (RU or EN keywords accepted).
"""
import argparse
import os
import sys
import unicodedata

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Course subjects come from the live Canvas seed, which is now Russified. Each
# entry is a DISCRIMINATING lowercase substring of the Russified subject name
# (uniquely identifies one course). The live set is queried below; this list is
# the fallback / prose-grep vocabulary.
EXPECTED_SUBJECTS = [
    "аналитик", "биохими", "креативн",
    "на основе данных", "эколог", "финанс",
    "геополит",
]

# Seeded relevant arxiv paper titles (English identifiers preserved). The agent
# must pull these from arxiv_local rather than inventing papers.
EXPECTED_PAPER_TITLES = [
    "deep learning for time series analytics",
    "reinforcement learning in financial markets",
    "data-driven design optimization using neural networks",
    "computational methods for environmental economics modeling",
    "machine learning algorithms for geopolitical risk assessment",
]

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Gap_Matrix Gap_Type is consistent with Relevance_Score",
    "Research_Frontiers contains >= 5 seeded arxiv paper titles",
    "Course_Topics subjects equal the distinct Canvas course subjects",
    "Teamly Curriculum Review Tracker page exists",
    "Teamly tracker covers every Canvas course subject",
    "Teamly tracker Priority follows the gap-count thresholds",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def normalize(s):
    """NFKD + cyrillic->latin translit. ONLY for mixed cyr/lat identifier
    matching (e.g. course names that may appear with either alphabet)."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    table = {
        "а": "a", "в": "b", "е": "e", "к": "k", "м": "m", "н": "h", "о": "o",
        "р": "p", "с": "c", "т": "t", "у": "y", "х": "x",
    }
    return "".join(table.get(c, c) for c in s.lower())


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", ".").replace("%", "").strip())
    except Exception:
        return None


def get_canvas_subjects():
    """Distinct Canvas course subjects, normalized the same way the original
    eval did (strip a trailing ' (...)' qualifier).

    Scope to this task's courses: id 9991 ('Изучение культурного наследия
    России') is seeded by the unrelated rzd-canvas-fieldtrip task on the shared
    persistent Canvas DB and is not part of this curriculum-review scope. It is
    excluded so the live subject set equals the 7 intended subjects matching
    EXPECTED_SUBJECTS."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute(
        "SELECT DISTINCT regexp_replace(name, ' \\(.*\\)$', '') "
        "FROM canvas.courses WHERE id != 9991"
    )
    subjects = {r[0].strip() for r in cur.fetchall() if r[0]}
    cur.close()
    conn.close()
    return subjects


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    fpath = os.path.join(agent_workspace, "Curriculum_Gap_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return
    record("Excel file exists", True)

    wb = openpyxl.load_workbook(fpath, data_only=True)

    # --- Sheet 1: Course_Topics ---
    ct_sheet = None
    for name in wb.sheetnames:
        if "course" in name.lower() and "topic" in name.lower():
            ct_sheet = name
            break
    try:
        canvas_subjects = get_canvas_subjects()
    except Exception:
        canvas_subjects = set()

    if not ct_sheet:
        record("Course_Topics sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Course_Topics subjects equal the distinct Canvas course subjects",
               False, "no Course_Topics sheet")
    else:
        record("Course_Topics sheet exists", True)
        ws = wb[ct_sheet]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r and r[0] not in (None, "")]
        expected_n = len(canvas_subjects) if canvas_subjects else 7
        record(f"Course_Topics has {expected_n} rows",
               len(rows) == expected_n, f"Got {len(rows)}")

        # CRITICAL: set-equality of subjects against the live Canvas seed.
        if canvas_subjects:
            sheet_subjects = {normalize(r[0]).strip() for r in rows}
            canvas_norm = {normalize(s).strip() for s in canvas_subjects}
            record("Course_Topics subjects equal the distinct Canvas course subjects",
                   sheet_subjects == canvas_norm,
                   f"sheet={sorted(sheet_subjects)} canvas={sorted(canvas_norm)}")
        else:
            # Canvas unavailable: fall back to the known subject list. Both the
            # sheet values and the expected substrings are normalized so the
            # cyrillic->latin translit is applied consistently on both sides.
            sheet_subjects = {normalize(r[0]).strip() for r in rows}
            record("Course_Topics subjects equal the distinct Canvas course subjects",
                   all(any(normalize(es) in s for s in sheet_subjects)
                       for es in EXPECTED_SUBJECTS),
                   f"sheet={sorted(sheet_subjects)}")

    # --- Sheet 2: Research_Frontiers ---
    rf_sheet = None
    for name in wb.sheetnames:
        if "research" in name.lower() or "frontier" in name.lower():
            rf_sheet = name
            break
    if not rf_sheet:
        record("Research_Frontiers sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Research_Frontiers contains >= 5 seeded arxiv paper titles",
               False, "no Research_Frontiers sheet")
    else:
        record("Research_Frontiers sheet exists", True)
        ws = wb[rf_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Research_Frontiers has >= 5 rows", len(rows) >= 5, f"Got {len(rows)}")

        # CRITICAL: the listed titles must be the seeded arxiv papers, not invented.
        rf_text = " ".join(
            str(c) for r in rows for c in (r or []) if c is not None
        ).lower()
        matched = sum(1 for t in EXPECTED_PAPER_TITLES if t in rf_text)
        record("Research_Frontiers contains >= 5 seeded arxiv paper titles",
               matched >= 5, f"matched {matched}/5")

    # --- Sheet 3: Gap_Matrix ---
    gm_sheet = None
    for name in wb.sheetnames:
        if "gap" in name.lower() and "matrix" in name.lower():
            gm_sheet = name
            break
    if not gm_sheet:
        record("Gap_Matrix sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Gap_Matrix Gap_Type is consistent with Relevance_Score",
               False, "no Gap_Matrix sheet")
    else:
        record("Gap_Matrix sheet exists", True)
        ws = wb[gm_sheet]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        record("Gap_Matrix has >= 20 rows", len(rows) >= 20, f"Got {len(rows)}")

        gap_types = set()
        # CRITICAL: per-row consistency between Relevance_Score and Gap_Type
        # per review_criteria.json: >=7 Covered, 4..6 Partial, <4 Gap.
        # Columns: Course_Subject, Paper_Title, Relevance_Score, Gap_Type.
        checked = 0
        inconsistent = []
        for row in rows:
            if not row or len(row) < 4:
                continue
            score = safe_float(row[2])
            gtype = str(row[3]).strip().lower() if row[3] is not None else ""
            if gtype:
                gap_types.add(gtype)
            if score is None or not gtype:
                continue
            checked += 1
            if score >= 7:
                expected = "covered"
            elif score >= 4:
                expected = "partial"
            else:
                expected = "gap"
            if gtype != expected:
                inconsistent.append((row[0], score, gtype, expected))

        record("Gap_Matrix has Covered type", "covered" in gap_types,
               f"Found types: {gap_types}")
        record("Gap_Matrix has Gap type", "gap" in gap_types,
               f"Found types: {gap_types}")
        record("Gap_Matrix Gap_Type is consistent with Relevance_Score",
               checked > 0 and not inconsistent,
               f"checked={checked} inconsistent={inconsistent[:5]}")

    wb.close()


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    fpath = os.path.join(agent_workspace, "Curriculum_Enhancement_Proposal.docx")
    if not os.path.isfile(fpath):
        record("Word document exists", False, f"Not found: {fpath}")
        return
    record("Word document exists", True)

    from docx import Document
    doc = Document(fpath)
    full_text = " ".join(p.text for p in doc.paragraphs).lower()
    # Accept RU or EN framing terms (prose is Russified; the title is English).
    record("Document mentions curriculum / учебн",
           "curriculum" in full_text or "учебн" in full_text or "программ" in full_text)
    record("Document mentions gap / пробел",
           "gap" in full_text or "пробел" in full_text)
    # Subject names are Russified Canvas course names; match discriminating
    # cyrillic substrings in the (lowercased) Russified prose.
    mentioned = sum(1 for s in EXPECTED_SUBJECTS if s in full_text)
    record("Document mentions >= 3 course subjects", mentioned >= 3, f"Found {mentioned}")


def check_terminal_output(agent_workspace):
    print("\n=== Checking Terminal Output ===")
    fpath = os.path.join(agent_workspace, "curriculum_gap_output.txt")
    if not os.path.isfile(fpath):
        record("curriculum_gap_output.txt exists", False)
        return
    record("curriculum_gap_output.txt exists", True)
    with open(fpath) as f:
        content = f.read().lower()
    record("Output mentions relevance / gap / score / релевант / пробел",
           any(k in content for k in
               ("relevance", "gap", "score", "релевант", "пробел", "оценк")))


def _get_tracker_page():
    """Return (title, body) of the agent-created Curriculum Review Tracker page,
    or None. Seed pages have id <= 3; the tracker is a fresh page identified by
    the English title marker 'Curriculum Review Tracker' (RU variants accepted)."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute(
        "SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3"
    )
    pages = cur.fetchall()
    cur.close()
    conn.close()
    for _pid, title, body in pages:
        tl = (title or "").lower()
        if "curriculum review tracker" in tl \
                or ("curriculum" in tl and "tracker" in tl) \
                or ("учебн" in tl and ("трекер" in tl or "обзор" in tl or "программ" in tl)):
            return title, body
    return None


def check_teamly():
    print("\n=== Checking Teamly Tracker ===")
    try:
        tracker = _get_tracker_page()
    except Exception as e:
        record("Teamly Curriculum Review Tracker page exists", False, str(e))
        record("Teamly tracker covers every Canvas course subject", False, str(e))
        record("Teamly tracker Priority follows the gap-count thresholds", False, str(e))
        return

    record("Teamly Curriculum Review Tracker page exists", tracker is not None,
           "no tracker page with id > 3")
    if tracker is None:
        record("Teamly tracker covers every Canvas course subject", False, "no page")
        record("Teamly tracker Priority follows the gap-count thresholds", False, "no page")
        return

    text = ((tracker[0] or "") + "\n" + (tracker[1] or ""))
    text_l = text.lower()
    text_n = normalize(text)

    # Required columns / status & priority vocabulary present (non-critical).
    record("Teamly tracker mentions Review Status values",
           any(k in text_l for k in ("not started", "не начато", "не начат")))
    record("Teamly tracker mentions Priority values",
           any(k in text_l for k in ("high", "medium", "low", "высок", "средн", "низк")))

    # CRITICAL: every distinct Canvas course subject appears in the page body.
    try:
        subjects = get_canvas_subjects()
    except Exception:
        subjects = set(EXPECTED_SUBJECTS)
    missing = [s for s in subjects if normalize(s) not in text_n]
    record("Teamly tracker covers every Canvas course subject",
           bool(subjects) and not missing, f"missing={missing}")

    # CRITICAL: Priority follows gap-count thresholds (>=3 gaps High, >=1 Medium,
    # 0 Low). We recompute gap counts per subject from the agent's own Gap_Matrix
    # (internal consistency) and verify each subject's row in the tracker carries
    # the implied Priority. Fall back gracefully if the Excel is unreadable.
    ok, detail = _check_priority_consistency(text_l, normalize)
    record("Teamly tracker Priority follows the gap-count thresholds", ok, detail)


def _check_priority_consistency(tracker_text_l, norm):
    """Cross-check tracker Priority against gap counts derived from the agent's
    Gap_Matrix sheet. Returns (ok, detail). Uses internal consistency (the
    agent's own scores) rather than absolute scores to avoid false fails."""
    # Locate the Excel produced by the agent via the shared workspace argument.
    gm_path = _AGENT_WS and os.path.join(_AGENT_WS, "Curriculum_Gap_Analysis.xlsx")
    if not gm_path or not os.path.isfile(gm_path):
        return False, "Gap_Matrix Excel unavailable for gap-count derivation"
    try:
        wb = openpyxl.load_workbook(gm_path, data_only=True)
    except Exception as e:
        return False, f"cannot open Excel: {e}"
    gm_sheet = next((n for n in wb.sheetnames
                     if "gap" in n.lower() and "matrix" in n.lower()), None)
    if not gm_sheet:
        wb.close()
        return False, "no Gap_Matrix sheet"
    ws = wb[gm_sheet]
    gap_counts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 4 or row[0] in (None, ""):
            continue
        subj = norm(row[0]).strip()
        gtype = str(row[3]).strip().lower() if row[3] is not None else ""
        gap_counts.setdefault(subj, 0)
        if gtype == "gap":
            gap_counts[subj] += 1
    wb.close()

    if not gap_counts:
        return False, "no rows in Gap_Matrix to derive gap counts"

    # For each subject, expected priority and presence near subject in tracker.
    # Tracker is a markdown table; we scan line by line for the subject row.
    lines = tracker_text_l.splitlines()
    mismatches = []
    verified = 0
    for subj, gaps in gap_counts.items():
        if gaps >= 3:
            expected = {"high", "высок"}
        elif gaps >= 1:
            expected = {"medium", "средн"}
        else:
            expected = {"low", "низк"}
        # Find a tracker line mentioning this subject (normalized contains).
        subj_line = None
        for ln in lines:
            if subj and subj in norm(ln):
                subj_line = ln
                break
        if subj_line is None:
            continue
        verified += 1
        if not any(e in subj_line or e in norm(subj_line) for e in expected):
            mismatches.append((subj, gaps, subj_line.strip()[:60]))

    if verified == 0:
        return False, "could not match any subject row in tracker"
    return (not mismatches), f"verified={verified} mismatches={mismatches[:3]}"


_AGENT_WS = None


def main():
    global _AGENT_WS
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    _AGENT_WS = args.agent_workspace

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_terminal_output(args.agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print("\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}  ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES ({len(critical_failed)}):")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            import json
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if success else 'FAIL'}")
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
