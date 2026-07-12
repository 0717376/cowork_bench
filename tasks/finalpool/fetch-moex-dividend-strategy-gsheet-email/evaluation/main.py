"""Evaluation script for fetch-moex-dividend-strategy-gsheet-email.

Структурные чеки (НЕ критичные): xlsx существует, 3 листа, заголовки колонок,
минимальное число строк, наличие письма / Google-таблицы / скрипта.

CRITICAL-чеки (семантические; любой провал => FAIL независимо от accuracy):
  #1  Для каждого тикера из засеянного /api/data.json строка Data_Analysis
      содержит Target_Price, равный target_price из JSON (per-symbol, tol ~0.01).
      Доказывает реальный fetch+join бенчмарка.
  #2  Колонка Upside согласована построчно: совпадает либо с процентной формулой
      (Target-Current)/Current*100, либо с абсолютной разностью (Target-Current),
      в пределах допуска. Доказывает реальный расчёт, а не копию шаблона.
  #3  Лист Metrics: Total_Stocks == числу тикеров в data.json (5) И
      Avg_Upside ~= среднему по колонке Upside листа Data_Analysis (с допуском).
  #4  В email.messages есть письмо с темой, содержащей маркер
      "Analysis Report Complete", адресованное team-lead@company.com,
      с непустым телом. Отличает результат от засеянного шума-рассылки.
  #5  Google-таблица с названием, содержащим "Tracker", существует И в ней
      >= 5 строк данных, где первая колонка содержит тикеры MOEX из data.json.

Порог: accuracy >= 70 И нет провала CRITICAL => PASS.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []
CRITICAL_CHECKS = {
    "CRITICAL: Target_Price совпадает с data.json по каждому тикеру",
    "CRITICAL: Upside согласован построчно с Current/Target",
    "CRITICAL: Metrics Total_Stocks и Avg_Upside соответствуют данным",
    "CRITICAL: письмо 'Analysis Report Complete' на team-lead@company.com",
    "CRITICAL: Google-таблица 'Tracker' заполнена тикерами MOEX",
}

MOCK_DATA_JSON = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "tmp", "mock_pages", "api", "data.json"
)


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL]{marker} {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def load_benchmark():
    """Читает засеянный data.json (источник, который агент обязан был fetch'нуть)."""
    try:
        with open(MOCK_DATA_JSON) as f:
            data = json.load(f)
        out = {}
        for item in data.get("market_indicators", []):
            sym = str(item.get("symbol", "")).strip()
            if sym:
                out[sym] = {
                    "target_price": safe_float(item.get("target_price")),
                    "analyst_rating": item.get("analyst_rating"),
                }
        return out
    except Exception:
        return {}

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    benchmark = load_benchmark()
    bench_symbols = set(benchmark.keys())

    excel_path = os.path.join(agent_workspace, "Dividend_Strategy_Report.xlsx")
    check("Dividend_Strategy_Report.xlsx exists", os.path.exists(excel_path))

    # Накопим данные Data_Analysis для семантических чеков
    da_rows_dict = {}   # symbol -> dict(current, target, upside)
    da_upsides = []

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Symbol', 'Name', 'Sector', 'Current_Price', 'Target_Price', 'Upside']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Индексы колонок
            def col_idx(name):
                try:
                    return headers.index(name.lower())
                except ValueError:
                    return None
            i_sym = col_idx('Symbol'); i_cur = col_idx('Current_Price')
            i_tgt = col_idx('Target_Price'); i_ups = col_idx('Upside')
            for row in data_rows:
                if i_sym is None or i_sym >= len(row):
                    continue
                sym = str(row[i_sym]).strip() if row[i_sym] is not None else ""
                if not sym:
                    continue
                cur = safe_float(row[i_cur]) if (i_cur is not None and i_cur < len(row)) else None
                tgt = safe_float(row[i_tgt]) if (i_tgt is not None and i_tgt < len(row)) else None
                ups = safe_float(row[i_ups]) if (i_ups is not None and i_ups < len(row)) else None
                da_rows_dict[sym] = {"current": cur, "target": tgt, "upside": ups}
                if ups is not None:
                    da_upsides.append(ups)

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_dict = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for row in data_rows:
                if row and len(row) >= 2 and row[0] is not None:
                    metrics_dict[str(row[0]).strip()] = row[1]

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Symbol']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ---------- CRITICAL #1: Target_Price совпадает с data.json ----------
        if benchmark:
            ok_tgt = True
            mism = []
            for sym, b in benchmark.items():
                bt = b["target_price"]
                row = da_rows_dict.get(sym)
                if row is None or row["target"] is None or bt is None:
                    ok_tgt = False; mism.append(f"{sym}:missing"); continue
                if abs(row["target"] - bt) > 0.01 + 1e-6:
                    ok_tgt = False; mism.append(f"{sym}:{row['target']}!={bt}")
            check("CRITICAL: Target_Price совпадает с data.json по каждому тикеру",
                  ok_tgt, f"mismatches={mism}")
        else:
            check("CRITICAL: Target_Price совпадает с data.json по каждому тикеру",
                  False, "benchmark data.json not found")

        # ---------- CRITICAL #2: Upside согласован построчно ----------
        ok_ups = True
        bad = []
        consistent_rows = 0
        for sym, row in da_rows_dict.items():
            cur, tgt, ups = row["current"], row["target"], row["upside"]
            if cur is None or tgt is None or ups is None or cur == 0:
                continue
            pct = (tgt - cur) / cur * 100.0
            absd = tgt - cur
            # допускаем процентную ИЛИ абсолютную трактовку Upside
            if abs(ups - pct) <= max(0.5, abs(pct) * 0.02) or abs(ups - absd) <= max(0.5, abs(absd) * 0.02):
                consistent_rows += 1
            else:
                ok_ups = False
                bad.append(f"{sym}:ups={ups},pct={round(pct,2)},abs={round(absd,2)}")
        # требуем, чтобы были проверяемые строки и все они были согласованы
        check("CRITICAL: Upside согласован построчно с Current/Target",
              ok_ups and consistent_rows >= 5, f"consistent={consistent_rows} bad={bad}")

        # ---------- CRITICAL #3: Metrics соответствуют данным ----------
        ok_metrics = True
        mdetail = []
        ts = safe_float(metrics_dict.get("Total_Stocks"))
        if not (ts is not None and bench_symbols and int(ts) == len(bench_symbols)):
            ok_metrics = False; mdetail.append(f"Total_Stocks={ts} expected={len(bench_symbols)}")
        avg = safe_float(metrics_dict.get("Avg_Upside"))
        if da_upsides:
            mean_ups = sum(da_upsides) / len(da_upsides)
            if avg is None or abs(avg - mean_ups) > max(0.5, abs(mean_ups) * 0.05):
                ok_metrics = False; mdetail.append(f"Avg_Upside={avg} expected~={round(mean_ups,2)}")
        else:
            ok_metrics = False; mdetail.append("no upsides to average")
        check("CRITICAL: Metrics Total_Stocks и Avg_Upside соответствуют данным",
              ok_metrics, f"{mdetail}")

        # ---------- CRITICAL #4: письмо-результат ----------
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "SELECT subject, to_addr, body_text FROM email.messages "
                "WHERE subject ILIKE %s", ('%Analysis Report Complete%',))
            rows = cur.fetchall()
            conn.close()
            found = False
            for subj, to_addr, body in rows:
                to_s = str(to_addr) if to_addr is not None else ""
                if "team-lead@company.com" in to_s and body and str(body).strip():
                    found = True; break
            check("CRITICAL: письмо 'Analysis Report Complete' на team-lead@company.com",
                  found, f"matching subjects={len(rows)}")
        except Exception as e:
            check("CRITICAL: письмо 'Analysis Report Complete' на team-lead@company.com",
                  False, str(e))

        # Структурный (НЕ критичный) чек письма: тема содержит report/analysis (RU+EN)
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "SELECT subject FROM email.messages WHERE "
                "subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s",
                ('%report%', '%analysis%', '%отчёт%', '%анализ%'))
            emails = cur.fetchall()
            check("Analysis email sent", len(emails) >= 1, f"found {len(emails)} matching emails")
            conn.close()
        except Exception as e:
            check("Email check", False, str(e))

        # ---------- CRITICAL #5: Google-таблица заполнена тикерами ----------
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE title ILIKE %s", ('%tracker%',))
            sheets = cur.fetchall()
            check("Google Sheet created", len(sheets) >= 1, f"found {len(sheets)} sheets")

            populated_ok = False
            pdetail = ""
            for sid, title in sheets:
                cur.execute(
                    "SELECT value FROM gsheet.cells c JOIN gsheet.sheets s ON c.sheet_id = s.id "
                    "WHERE s.spreadsheet_id = %s", (sid,))
                vals = [str(v[0]).strip() for v in cur.fetchall() if v[0] is not None]
                matched = bench_symbols & set(vals)
                if len(matched) >= 5:
                    populated_ok = True; break
                pdetail = f"matched_symbols={sorted(matched)}"
            check("CRITICAL: Google-таблица 'Tracker' заполнена тикерами MOEX",
                  populated_ok, pdetail or "no tracker sheet with >=5 MOEX symbols")
            conn.close()
        except Exception as e:
            check("CRITICAL: Google-таблица 'Tracker' заполнена тикерами MOEX",
                  False, str(e))

        check("yf_dividend_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "yf_dividend_processor.py")))

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        return False, "FAIL: No checks performed."
    accuracy = PASS_COUNT / total * 100

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_fails": CRITICAL_FAILS,
                }, f, indent=2)
        except Exception:
            pass

    if CRITICAL_FAILS:
        return False, f"FAIL: критичные чеки провалены ({len(CRITICAL_FAILS)}): {CRITICAL_FAILS}"
    return accuracy >= 70, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
