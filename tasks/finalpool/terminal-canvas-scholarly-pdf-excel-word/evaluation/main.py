"""Evaluation for terminal-canvas-scholarly-pdf-excel-word.

Pass gate: any CRITICAL check failure => overall FAIL regardless of accuracy.
Otherwise PASS requires accuracy >= 70%.

Canvas is kept as a foreign (English) MCP: course name and assignments are read
live from canvas.assignments, so the eval must NOT hardcode their values. The
Word report may be written in Russian, so topical greps accept RU+EN signal.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# CRITICAL checks: any failure => overall FAIL regardless of accuracy.
# These reflect the substance of the deliverables (correct source-derived
# values, key rules, core outputs), not structural presence.
CRITICAL_CHECKS = {
    "Course_Assignments: Course_ID/Points match live Canvas for all GT assignments",
    "Course_Assignments row count equals live Canvas count",
    "Alignment_Matrix row count equals live Canvas count",
    "Related_Papers: every title matches an injected scholarly paper",
    "Related_Papers: >= 4 relevant papers with valid Relevance_Score 1-10",
    "Alignment_Matrix: every Matched_Paper is in Related_Papers and scores 0-100",
    "Alignment_Matrix sorted by Alignment_Score descending",
    "Word report exists with substantial content and RU/EN topical signal",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:300]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower().replace(" ", "_") == name.strip().lower().replace(" ", "_"):
            return wb[s]
    return None


def _canvas_count():
    """Live count of assignments for course_id IN (1,2)."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id IN (1, 2)")
        n = cur.fetchone()[0]
        cur.close(); conn.close()
        return n
    except Exception:
        return 12


def _scholar_titles():
    """Set of injected scholarly paper titles (lowercased)."""
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT title FROM scholarly.scholar_papers")
        titles = {str(r[0]).strip().lower() for r in cur.fetchall() if r[0]}
        cur.close(); conn.close()
        return titles
    except Exception:
        return set()


def check_excel(agent_ws, gt_dir):
    print("\n=== Checking Curriculum_Research_Alignment.xlsx ===")
    agent_file = os.path.join(agent_ws, "Curriculum_Research_Alignment.xlsx")
    gt_file = os.path.join(gt_dir, "Curriculum_Research_Alignment.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        awb = openpyxl.load_workbook(agent_file, data_only=True)
        gwb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    expected_count = _canvas_count()

    # Sheet 1: Course_Assignments
    print("  Checking Course_Assignments...")
    ws1 = get_sheet(awb, "Course_Assignments")
    gws1 = get_sheet(gwb, "Course_Assignments")
    check("Sheet Course_Assignments exists", ws1 is not None, f"Sheets: {awb.sheetnames}")
    if ws1 and gws1:
        a_rows = list(ws1.iter_rows(min_row=2, values_only=True))
        g_rows = list(gws1.iter_rows(min_row=2, values_only=True))

        # CRITICAL: row count equals live Canvas count (all assignments covered).
        check("Course_Assignments row count equals live Canvas count",
              len(a_rows) == expected_count, f"Got {len(a_rows)}, expected {expected_count}")

        # Build lookup by assignment name (names come from Canvas -> stay English).
        a_lookup = {}
        for r in a_rows:
            if r and len(r) >= 4 and r[2]:
                a_lookup[str(r[2]).strip().lower()] = r

        # CRITICAL: every GT assignment matches live Course_ID/Points.
        all_match = True
        mismatch_detail = ""
        for g_row in g_rows:
            if not g_row or not g_row[2]:
                continue
            key = str(g_row[2]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_match = False
                mismatch_detail = f"Missing assignment '{g_row[2]}'"
                break
            if not num_close(a_row[0], g_row[0], 0):
                all_match = False
                mismatch_detail = f"'{g_row[2]}' Course_ID expected {g_row[0]}, got {a_row[0]}"
                break
            if not num_close(a_row[3], g_row[3], 1):
                all_match = False
                mismatch_detail = f"'{g_row[2]}' Points expected {g_row[3]}, got {a_row[3]}"
                break
        check("Course_Assignments: Course_ID/Points match live Canvas for all GT assignments",
              all_match, mismatch_detail)

    # Sheet 2: Related_Papers
    print("  Checking Related_Papers...")
    ws2 = get_sheet(awb, "Related_Papers")
    check("Sheet Related_Papers exists", ws2 is not None, f"Sheets: {awb.sheetnames}")
    related_titles = set()
    if ws2:
        a_rows2 = list(ws2.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in a_rows2 if r and r[0]]
        for r in data_rows:
            related_titles.add(str(r[0]).strip().lower())

        # CRITICAL: >= 4 papers AND all relevance scores in 1-10.
        scores_ok = True
        for r in data_rows:
            if r and len(r) >= 4 and r[3] is not None:
                try:
                    score = float(r[3])
                except (TypeError, ValueError):
                    scores_ok = False
                    break
                if score < 1 or score > 10:
                    scores_ok = False
                    break
        check("Related_Papers: >= 4 relevant papers with valid Relevance_Score 1-10",
              len(data_rows) >= 4 and scores_ok,
              f"Count {len(data_rows)}, scores_ok={scores_ok}")

        # CRITICAL: every listed title is a real injected scholarly paper.
        scholar = _scholar_titles()
        if scholar:
            unknown = [r[0] for r in data_rows
                       if str(r[0]).strip().lower() not in scholar]
            check("Related_Papers: every title matches an injected scholarly paper",
                  len(unknown) == 0, f"Not in scholarly DB: {unknown}")
        else:
            check("Related_Papers: every title matches an injected scholarly paper",
                  False, "Could not load scholarly.scholar_papers")

    # Sheet 3: Alignment_Matrix
    print("  Checking Alignment_Matrix...")
    ws3 = get_sheet(awb, "Alignment_Matrix")
    check("Sheet Alignment_Matrix exists", ws3 is not None, f"Sheets: {awb.sheetnames}")
    if ws3:
        a_rows3 = list(ws3.iter_rows(min_row=2, values_only=True))
        data_rows3 = [r for r in a_rows3 if r and r[0]]

        # CRITICAL: row count equals live Canvas count.
        check("Alignment_Matrix row count equals live Canvas count",
              len(data_rows3) == expected_count,
              f"Got {len(data_rows3)}, expected {expected_count}")

        # CRITICAL: every Matched_Paper appears in Related_Papers AND score 0-100.
        matrix_ok = True
        m_detail = ""
        for r in data_rows3:
            if not (r and len(r) >= 3 and r[1] and r[2] is not None):
                matrix_ok = False
                m_detail = f"Incomplete row: {r}"
                break
            mp = str(r[1]).strip().lower()
            if related_titles and mp not in related_titles:
                matrix_ok = False
                m_detail = f"Matched_Paper not in Related_Papers: {r[1]}"
                break
            try:
                sc = float(r[2])
            except (TypeError, ValueError):
                matrix_ok = False
                m_detail = f"Non-numeric Alignment_Score: {r[2]}"
                break
            if sc < 0 or sc > 100:
                matrix_ok = False
                m_detail = f"Score out of range: {sc}"
                break
        check("Alignment_Matrix: every Matched_Paper is in Related_Papers and scores 0-100",
              matrix_ok, m_detail)

        # CRITICAL: sorted by Alignment_Score descending.
        scores = []
        sortable = True
        for r in data_rows3:
            if r and len(r) >= 3 and r[2] is not None:
                try:
                    scores.append(float(r[2]))
                except (TypeError, ValueError):
                    sortable = False
                    break
        check("Alignment_Matrix sorted by Alignment_Score descending",
              sortable and scores == sorted(scores, reverse=True),
              f"First few scores: {scores[:5]}")


def check_word(agent_ws):
    print("\n=== Checking Curriculum_Review_Report.docx ===")
    docx_path = os.path.join(agent_ws, "Curriculum_Review_Report.docx")
    check("Word file exists", os.path.isfile(docx_path), docx_path)
    if not os.path.isfile(docx_path):
        check("Word report exists with substantial content and RU/EN topical signal",
              False, "File missing")
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()

        has_len = len(text) > 500
        # Topic signal accepts EN + RU.
        topic = any(k in text for k in (
            "curriculum", "alignment", "review", "analytics", "algorithm", "course",
            "учебн", "программ", "соответств", "анализ", "курс", "задани"))
        research = any(k in text for k in (
            "paper", "research", "literature", "статья", "стате", "исследован", "литератур"))
        recommend = any(k in text for k in (
            "recommend", "suggestion", "improve", "рекоменд", "улучш", "предложен"))

        # Non-critical granular sub-checks (for accuracy signal).
        check("Document has substantial content", has_len, f"Length: {len(text)}")
        check("Contains curriculum/alignment reference (RU/EN)", topic, "Missing topic reference")
        check("Contains paper/research reference (RU/EN)", research, "Missing research reference")
        check("Contains recommendation (RU/EN)", recommend, "Missing recommendations")

        # CRITICAL: combined report-quality signal.
        check("Word report exists with substantial content and RU/EN topical signal",
              has_len and topic and research and recommend,
              f"len={len(text)} topic={topic} research={research} recommend={recommend}")
    except ImportError:
        check("python-docx available", False, "Cannot verify Word content")
        check("Word report exists with substantial content and RU/EN topical signal",
              False, "python-docx unavailable")
    except Exception as e:
        check("Word document readable", False, str(e))
        check("Word report exists with substantial content and RU/EN topical signal",
              False, str(e))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Curriculum_Research_Alignment.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            expected_sheets = {"course_assignments", "related_papers", "alignment_matrix"}
            actual_sheets = {s.strip().lower().replace(" ", "_") for s in wb.sheetnames}
            unexpected = actual_sheets - expected_sheets
            check("No unexpected sheets in Excel",
                  len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")

            ws = get_sheet(wb, "Course_Assignments")
            if ws:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row and len(row) > 3 and row[3] is not None:
                        try:
                            val = float(row[3])
                            if val < 0:
                                check("No negative points in Course_Assignments", False,
                                      f"Found negative points: {val}")
                                break
                        except (TypeError, ValueError):
                            pass
                else:
                    check("No negative points in Course_Assignments", True)
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    print(f"  Failed: {FAIL_COUNT}")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    overall = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
