"""Evaluation для terminal-clickhouse-teamly-forms-excel-email (RU-стек).

Проверяет:
  1. Employee_Engagement_Report.xlsx (3 листа: Department_Scores / Survey_Design /
     Action_Items).
  2. engagement_analysis_output.txt (вывод скрипта: наивысший/наименьший отдел).
  3. Forms (gform.*): «Опрос вовлечённости сотрудников» с 5 вопросами и верными
     наборами вариантов.
  4. Teamly (teamly.pages): страница-дашборд с 7 отделами, числом сотрудников и
     индексами вовлечённости.
  5. Email на hr-leadership@company.com с отчётом по вовлечённости.

Критические чеки (CRITICAL_CHECKS) отражают СУТЬ задачи: любой их fail => FAIL,
даже если accuracy >= 70%. Структурные чеки (лист есть, файл есть) — мягкие.
Порог: accuracy >= 70% И нет критических провалов.

Названия отделов русифицируются ЦЕНТРАЛЬНО в sf_data (db/zzz_clickhouse_after_init.sql:
Engineering->Инженерия, Finance->Финансы, HR->Кадры, Operations->Операции,
R&D->НИОКР, Sales->Продажи, Support->Поддержка). Эталон GT и engagement_targets.json
синхронизированы с русифицированными названиями. Числовые метрики (удовлетворённость/
производительность/баланс) маппинг не меняет, поэтому средние в GT остаются валидными.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

CRITICAL_CHECKS = {
    "Department_Scores: все 7 отделов — Avg_Satisfaction, Avg_Performance И Avg_Work_Life_Balance совпадают с эталоном",
    "Engagement_Index = round(0.40*Satisfaction + 0.35*Work_Life_Balance + 0.25*Performance, 2) для каждого отдела",
    "Action_Items: приоритет High для отделов ниже среднего по компании, Medium — для выше",
    "Forms: «Опрос вовлечённости сотрудников» — ровно 5 вопросов с верными наборами вариантов",
    "Teamly: страница-дашборд содержит все 7 отделов с числом сотрудников и индексами",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=0.05):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


# Результат проверки Excel, нужный другим проверкам (teamly).
DEPT_SUMMARY = []  # list of (dept, employee_count, engagement_index)


def check_excel(agent_workspace, gt_workspace):
    print("\n=== Проверка 1: Employee_Engagement_Report.xlsx ===")
    global DEPT_SUMMARY

    fpath = os.path.join(agent_workspace, "Employee_Engagement_Report.xlsx")
    gt_path = os.path.join(gt_workspace, "Employee_Engagement_Report.xlsx")

    if not os.path.isfile(fpath):
        record("Excel-файл существует", False, f"Не найден: {fpath}")
        record("Department_Scores: все 7 отделов — Avg_Satisfaction, Avg_Performance И Avg_Work_Life_Balance совпадают с эталоном", False, "нет файла")
        record("Engagement_Index = round(0.40*Satisfaction + 0.35*Work_Life_Balance + 0.25*Performance, 2) для каждого отдела", False, "нет файла")
        record("Action_Items: приоритет High для отделов ниже среднего по компании, Medium — для выше", False, "нет файла")
        return

    record("Excel-файл существует", True)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)

    # --- Лист 1: Department_Scores ---
    ds_sheet = None
    for name in wb.sheetnames:
        if "department" in name.lower() and "score" in name.lower():
            ds_sheet = name
            break
    if not ds_sheet:
        record("Department_Scores: лист существует", False, f"Листы: {wb.sheetnames}")
        record("Department_Scores: все 7 отделов — Avg_Satisfaction, Avg_Performance И Avg_Work_Life_Balance совпадают с эталоном", False, "нет листа")
        record("Engagement_Index = round(0.40*Satisfaction + 0.35*Work_Life_Balance + 0.25*Performance, 2) для каждого отдела", False, "нет листа")
        record("Action_Items: приоритет High для отделов ниже среднего по компании, Medium — для выше", False, "нет листа")
        wb.close(); gt_wb.close()
        return
    record("Department_Scores: лист существует", True)

    ws = wb[ds_sheet]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows = [r for r in rows if r and r[0]]
    record("Department_Scores: 7 строк", len(rows) == 7, f"Получено {len(rows)}")

    gt_ws = gt_wb["Department_Scores"]
    gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if r and r[0]]

    a_lookup = {str(r[0]).strip().lower(): r for r in rows}

    # CRITICAL 1: все 7 отделов, все три средних метрики совпадают с GT.
    metrics_ok = True
    metrics_detail = []
    for gt_row in gt_rows:
        key = str(gt_row[0]).strip().lower()
        a_row = a_lookup.get(key)
        if a_row is None:
            metrics_ok = False
            metrics_detail.append(f"{gt_row[0]}: отсутствует")
            continue
        # idx1=Satisfaction, idx2=Performance, idx3=WorkLifeBalance
        for idx, label in ((1, "Sat"), (2, "Perf"), (3, "WLB")):
            if not (len(a_row) > idx and num_close(a_row[idx], gt_row[idx], 0.05)):
                metrics_ok = False
                got = a_row[idx] if len(a_row) > idx else None
                metrics_detail.append(f"{gt_row[0]}.{label}={got} vs {gt_row[idx]}")
    record("Department_Scores: все 7 отделов — Avg_Satisfaction, Avg_Performance И Avg_Work_Life_Balance совпадают с эталоном",
           metrics_ok and len(rows) == 7, "; ".join(metrics_detail[:8]))

    # CRITICAL 2: индекс пересчитан по политике из собственных метрик строки.
    formula_ok = True
    formula_detail = []
    for r in rows:
        try:
            sat, perf, wlb, idx = float(r[1]), float(r[2]), float(r[3]), float(r[5])
        except (TypeError, ValueError, IndexError):
            formula_ok = False
            formula_detail.append(f"{r[0]}: нечисловые значения")
            continue
        expected = round(0.40 * sat + 0.35 * wlb + 0.25 * perf, 2)
        if abs(expected - idx) > 0.02:
            formula_ok = False
            formula_detail.append(f"{r[0]}: idx={idx} vs формула={expected}")
    record("Engagement_Index = round(0.40*Satisfaction + 0.35*Work_Life_Balance + 0.25*Performance, 2) для каждого отдела",
           formula_ok and len(rows) == 7, "; ".join(formula_detail[:6]))

    # Сохраняем сводку для проверки teamly (отдел, число сотрудников, индекс).
    DEPT_SUMMARY = []
    for r in rows:
        try:
            DEPT_SUMMARY.append((str(r[0]).strip(), int(float(r[4])), float(r[5])))
        except (TypeError, ValueError, IndexError):
            pass

    # --- Лист 2: Survey_Design ---
    sd_sheet = None
    for name in wb.sheetnames:
        if "survey" in name.lower():
            sd_sheet = name
            break
    if not sd_sheet:
        record("Survey_Design: лист существует", False, f"Листы: {wb.sheetnames}")
    else:
        record("Survey_Design: лист существует", True)
        ws2 = wb[sd_sheet]
        rows2 = [r for r in ws2.iter_rows(min_row=2, values_only=True) if r and any(c for c in r)]
        record("Survey_Design: 5 строк", len(rows2) == 5, f"Получено {len(rows2)}")

    # --- Лист 3: Action_Items (CRITICAL 3: приоритеты по политике) ---
    ai_sheet = None
    for name in wb.sheetnames:
        if "action" in name.lower():
            ai_sheet = name
            break
    if not ai_sheet:
        record("Action_Items: лист существует", False, f"Листы: {wb.sheetnames}")
        record("Action_Items: приоритет High для отделов ниже среднего по компании, Medium — для выше", False, "нет листа")
    else:
        record("Action_Items: лист существует", True)
        ws3 = wb[ai_sheet]
        rows3 = [r for r in ws3.iter_rows(min_row=2, values_only=True) if r and r[0]]
        record("Action_Items: 7 строк", len(rows3) == 7, f"Получено {len(rows3)}")

        # Среднее по компании считаем из ИНДЕКСОВ САМОГО АГЕНТА (самосогласованно).
        idx_by_dept = {d.lower(): i for d, _c, i in DEPT_SUMMARY}
        company_mean = sum(idx_by_dept.values()) / len(idx_by_dept) if idx_by_dept else None

        priority_ok = True
        prio_detail = []
        if company_mean is None:
            priority_ok = False
            prio_detail.append("нет индексов из Department_Scores")
        else:
            for r in rows3:
                dept = str(r[1]).strip().lower() if len(r) > 1 and r[1] else ""
                prio = str(r[0]).strip().lower() if r[0] else ""
                dept_idx = idx_by_dept.get(dept)
                if dept_idx is None:
                    priority_ok = False
                    prio_detail.append(f"{r[1]}: отдел не найден в Department_Scores")
                    continue
                # ниже среднего => High; выше => Medium. Граница (==) принимаем любой.
                if dept_idx < company_mean - 1e-9:
                    if "high" not in prio:
                        priority_ok = False
                        prio_detail.append(f"{r[1]}: idx={dept_idx}<{company_mean:.4f} ожидался High, получено '{r[0]}'")
                elif dept_idx > company_mean + 1e-9:
                    if "medium" not in prio:
                        priority_ok = False
                        prio_detail.append(f"{r[1]}: idx={dept_idx}>{company_mean:.4f} ожидался Medium, получено '{r[0]}'")
        record("Action_Items: приоритет High для отделов ниже среднего по компании, Medium — для выше",
               priority_ok, "; ".join(prio_detail[:6]))

    wb.close()
    gt_wb.close()


def check_terminal_output(agent_workspace):
    print("\n=== Проверка 2: engagement_analysis_output.txt ===")
    fpath = os.path.join(agent_workspace, "engagement_analysis_output.txt")
    if not os.path.isfile(fpath):
        record("engagement_analysis_output.txt существует", False)
        return
    record("engagement_analysis_output.txt существует", True)
    with open(fpath, encoding="utf-8") as f:
        content = f.read().lower()
    highest_kw = ("highest", "best", "top", "наивысш", "высш", "максим", "наибольш")
    lowest_kw = ("lowest", "worst", "bottom", "наименьш", "низш", "миним", "наимень")
    record("В выводе указан отдел с наивысшим индексом",
           any(k in content for k in highest_kw), f"Превью: {content[:200]}")
    record("В выводе указан отдел с наименьшим индексом",
           any(k in content for k in lowest_kw), f"Превью: {content[:200]}")


def check_forms():
    print("\n=== Проверка 3: Forms «Опрос вовлечённости сотрудников» ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("Forms: «Опрос вовлечённости сотрудников» — ровно 5 вопросов с верными наборами вариантов", False, str(e))
        return
    try:
        cur.execute("SELECT id, title FROM gform.forms")
        forms = cur.fetchall()

        def title_matches(title):
            t = (title or "").lower()
            # исключаем шум (архивная форма прошлого года)
            if "архив" in t or "прошл" in t or "old" in t or "стар" in t:
                return False
            ru = "вовлеч" in t or ("опрос" in t and "сотрудник" in t)
            en = "engagement" in t and ("survey" in t or "employee" in t)
            return ru or en

        target = None
        for fid, title in forms:
            if title_matches(title):
                target = (fid, title)
                break
        record("Forms: найдена форма опроса вовлечённости",
               target is not None, f"Формы: {[f[1] for f in forms]}")

        ok_5_and_options = False
        detail = "форма не найдена"
        if target:
            cur.execute(
                "SELECT question_type, config FROM gform.questions WHERE form_id = %s ORDER BY position",
                (target[0],),
            )
            qrows = cur.fetchall()
            n = len(qrows)
            # подсчёт числа вариантов по каждому вопросу (для choiceQuestion)
            opt_counts = []
            for qtype, config in qrows:
                if isinstance(config, str):
                    try:
                        config = json.loads(config)
                    except Exception:
                        config = {}
                opts = (config or {}).get("options") or []
                opt_counts.append(len(opts))
            # ожидаемые размеры наборов: {10,5,4,5,3} (порядок не важен).
            # Q1 шкала 1-10 -> 10, Q2 -> 5, Q3 -> 4, Q4 шкала 1-5 -> 5, Q5 -> 3.
            expected_multiset = sorted([10, 5, 4, 5, 3])
            got_multiset = sorted([c for c in opt_counts if c > 0])
            ok_5_and_options = (n == 5 and got_multiset == expected_multiset)
            detail = f"вопросов={n}, размеры наборов={sorted(opt_counts)}"
        record("Forms: «Опрос вовлечённости сотрудников» — ровно 5 вопросов с верными наборами вариантов",
               ok_5_and_options, detail)
    except Exception as e:
        record("Forms: «Опрос вовлечённости сотрудников» — ровно 5 вопросов с верными наборами вариантов", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_teamly():
    print("\n=== Проверка 4: Teamly страница-дашборд ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("Teamly: страница-дашборд содержит все 7 отделов с числом сотрудников и индексами", False, str(e))
        return
    try:
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
        pages = cur.fetchall()

        def title_matches(title):
            t = (title or "").lower()
            ru = "вовлеч" in t and ("панел" in t or "дашборд" in t or "персонал" in t)
            en = "engagement" in t and ("dashboard" in t or "hr" in t)
            return ru or en

        candidates = [(t, b) for t, b in pages if title_matches(t)]
        record("Teamly: страница-дашборд создана",
               len(candidates) >= 1, f"Заголовки: {[t for t, _ in pages]}")

        body = "\n\n".join(b for _t, b in candidates).lower()

        # CRITICAL: на странице есть все 7 отделов И их число сотрудников И индексы.
        if not DEPT_SUMMARY:
            record("Teamly: страница-дашборд содержит все 7 отделов с числом сотрудников и индексами",
                   False, "нет сводки из Excel (Department_Scores)")
        else:
            depts_ok = sum(1 for d, _c, _i in DEPT_SUMMARY if d.lower() in body)
            # число сотрудников ищем без разделителей разрядов («7 096» / «7 096» / «7,096»)
            body_digits = body.replace(" ", "").replace(" ", "").replace(" ", "").replace(",", "")
            counts_ok = sum(1 for _d, c, _i in DEPT_SUMMARY if str(c) in body_digits)
            # индекс ищем как «5.02» / «5,02» / «5.0»
            idx_ok = 0
            for _d, _c, i in DEPT_SUMMARY:
                variants = {f"{i:.2f}", f"{i:.2f}".replace(".", ","), str(i), str(i).replace(".", ",")}
                if any(v in body for v in variants):
                    idx_ok += 1
            ok = bool(candidates) and depts_ok == 7 and counts_ok >= 7 and idx_ok >= 7
            record("Teamly: страница-дашборд содержит все 7 отделов с числом сотрудников и индексами",
                   ok, f"отделов={depts_ok}/7, чисел={counts_ok}/7, индексов={idx_ok}/7")

        # Обратная валидация: нет дубликатов страницы-дашборда.
        record("Teamly: нет дубликатов страницы-дашборда", len(candidates) <= 1,
               f"Найдено {len(candidates)} подходящих страниц")
    except Exception as e:
        record("Teamly: страница-дашборд содержит все 7 отделов с числом сотрудников и индексами", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email():
    print("\n=== Проверка 5: Письмо руководству персонала ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        record("Письмо на hr-leadership@company.com про вовлечённость отправлено", False, str(e))
        return
    try:
        cur.execute("""
            SELECT subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%%hr-leadership@company.com%%'
        """)
        rows = cur.fetchall()
        matched = []
        for subj, body in rows:
            s = (subj or "").lower()
            if "вовлеч" in s or "engagement" in s:
                matched.append((subj, body))
        record("Письмо на hr-leadership@company.com про вовлечённость отправлено",
               len(matched) >= 1, f"Темы: {[r[0] for r in rows]}")
        if matched:
            body = (matched[0][1] or "").lower()
            # тело упоминает наивысший и наименьший отдел (мягко)
            high_kw = ("наивысш", "высш", "максим", "наибольш", "highest", "top")
            low_kw = ("наименьш", "низш", "миним", "наимень", "lowest", "bottom")
            record("В теле письма есть выводы по наивысшему/наименьшему отделу",
                   any(k in body for k in high_kw) and any(k in body for k in low_kw),
                   f"Превью: {body[:200]}")
    except Exception as e:
        record("Письмо на hr-leadership@company.com про вовлечённость отправлено", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_reverse_validation(workspace):
    print("\n=== Обратная валидация ===")
    path = os.path.join(workspace, "Employee_Engagement_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        record("Excel: не более 5 листов", len(wb.sheetnames) <= 5,
               f"Найдено {len(wb.sheetnames)}: {wb.sheetnames}")
        has_negative = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        has_negative = True
                        break
                if has_negative:
                    break
            if has_negative:
                break
        record("Нет отрицательных значений в Excel", not has_negative,
               "Найдено отрицательное значение")
        wb.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    if not os.path.isfile(os.path.join(gt_dir, "Employee_Engagement_Report.xlsx")):
        gt_dir = os.path.join(task_root, "groundtruth_workspace")

    print("=" * 70)
    print("TERMINAL CLICKHOUSE TEAMLY FORMS EXCEL EMAIL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, gt_dir)
    check_terminal_output(args.agent_workspace)
    check_forms()
    check_teamly()
    check_email()
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: проверки не выполнены.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== ИТОГО ===")
    print(f"  Пройдено {PASS_COUNT}/{total} ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAIL: {CRITICAL_FAILED}")
        print("  Overall: FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    else:
        print("  Overall: FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
