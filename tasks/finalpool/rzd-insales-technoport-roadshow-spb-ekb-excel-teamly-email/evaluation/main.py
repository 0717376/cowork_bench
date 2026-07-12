"""
Evaluation для rzd-insales-technoport-roadshow-spb-ekb-excel-teamly-email.

Что проверяем:
  1. Roadshow_Plan.xlsx существует и читается
  2. Лист Products: ≥4 строки, есть колонки Name и Price
  3. Лист Travel_Itinerary: ≥2 строки, содержит номера 752А и 016У
     (нормализуем кириллицу↔латиницу: 752А≈752A, 016У≈016Y/016U)
  4. Лист Roadshow_Schedule: ≥3 строки
  5. Teamly: создана страница в пространстве TRIPS со словами
     «road show» / «Санкт-Петербург» / «Екатеринбург» / номерами поездов
  6. Письма: отправлены на spb_dist@partner.ru, ekb_dist@partner.ru, manager@company.ru
  7. Groundtruth row-by-row сравнение листов (нестрогое: train_no — fuzzy,
     текстовые ячейки — точное; числа — ±10%)
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def normalize_date(s):
    """ISO YYYY-MM-DD остаётся как есть. DD.MM.YYYY и DD/MM/YYYY переводим в ISO.
    Если не парсится — возвращаем оригинал."""
    if s is None:
        return ""
    t = str(s).strip()
    import re
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", t)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{1,2})[./](\d{1,2})[./](\d{4})", t)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return t


def normalize_train(s):
    """Нормализуем номер поезда — приводим к нижнему регистру и кириллицу→латиница."""
    if s is None:
        return ""
    t = str(s).strip().lower()
    table = str.maketrans({
        "а": "a", "б": "b", "у": "y", "к": "k", "е": "e",
        "о": "o", "р": "p", "с": "c", "т": "t", "х": "x", "м": "m", "н": "h",
    })
    return t.translate(table)


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Excel: Roadshow_Plan.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Roadshow_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Roadshow_Plan.xlsx существует", False, f"Не найден в {xlsx_path}")
        return
    record("Roadshow_Plan.xlsx существует", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel читается", False, str(e))
        return
    record("Excel читается", True)

    # Products
    prod_sheet = None
    for name in wb.sheetnames:
        if "product" in name.lower() or "товар" in name.lower():
            prod_sheet = wb[name]
            break
    if prod_sheet is None:
        record("Лист Products существует", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Лист Products существует", True)
        rows = list(prod_sheet.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        has_name = any("name" in h or "назв" in h for h in headers)
        has_price = any("price" in h or "цен" in h for h in headers)
        record("Products: есть колонки Name и Price", has_name and has_price,
               f"Заголовки: {rows[0] if rows else []}")
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Products: ≥4 строк данных", len(data_rows) >= 4,
               f"Найдено {len(data_rows)} строк")

    # Travel_Itinerary
    travel_sheet = None
    for name in wb.sheetnames:
        if "travel" in name.lower() or "itinerary" in name.lower() or "маршр" in name.lower():
            travel_sheet = wb[name]
            break
    if travel_sheet is None:
        record("Лист Travel_Itinerary существует", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Лист Travel_Itinerary существует", True)
        rows = list(travel_sheet.iter_rows(values_only=True))
        all_text_raw = " ".join(str(c) for r in rows for c in r if c)
        all_text_norm = normalize_train(all_text_raw)
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Travel_Itinerary: ≥2 строк данных", len(data_rows) >= 2,
               f"Найдено {len(data_rows)} строк")
        has_752 = any(t in all_text_norm for t in ["752a", "752а"])
        has_016 = any(t in all_text_norm for t in ["016y", "016u", "016у"])
        record("Travel_Itinerary содержит 752А (Сапсан)", has_752,
               f"Текст: {all_text_raw[:200]}")
        record("Travel_Itinerary содержит 016У (Урал)", has_016,
               f"Текст: {all_text_raw[:200]}")

    # Roadshow_Schedule
    sched_sheet = None
    for name in wb.sheetnames:
        if "schedule" in name.lower() or "roadshow" in name.lower() or "распис" in name.lower():
            sched_sheet = wb[name]
            break
    if sched_sheet is None:
        record("Лист Roadshow_Schedule существует", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Лист Roadshow_Schedule существует", True)
        rows = list(sched_sheet.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Roadshow_Schedule: ≥3 строк данных", len(data_rows) >= 3,
               f"Найдено {len(data_rows)} строк")

    # --- Groundtruth row-by-row сравнение ---
    gt_path = os.path.join(groundtruth_workspace, "Roadshow_Plan.xlsx")
    if not os.path.isfile(gt_path):
        record("Groundtruth xlsx существует", False, gt_path)
        return

    # Колонки, которые НЕ сверяем с GT (агент берёт реальные данные из WC / придумывает venue).
    # GT для них чисто демонстративный — сравнение бессмысленно.
    #   - Products: всё кроме Stock_Status (4) — все остальные колонки реальные из WC
    #   - Customer_Regions: Customer_Count (1), Revenue_Share_Pct (2)
    #   - Roadshow_Schedule: Venue (2), Meeting_Type (3), Status (5) — творчество модели
    SKIP_COLS = {
        "products": {0, 1, 2, 3, 5},
        "customer_regions": {1, 2},
        "roadshow_schedule": {2, 3, 5},
    }

    gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    for gt_sheet_name in gt_wb.sheetnames:
        gt_ws = gt_wb[gt_sheet_name]
        agent_ws = None
        for asn in wb.sheetnames:
            if asn.strip().lower() == gt_sheet_name.strip().lower():
                agent_ws = wb[asn]
                break
        if agent_ws is None:
            record(f"GT-лист '{gt_sheet_name}' есть у агента", False, f"Доступно: {wb.sheetnames}")
            continue

        gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        agent_rows = [r for r in agent_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]

        sheet_lower = gt_sheet_name.strip().lower()

        # Customer_Regions проверяем по своим правилам, без row-count и row-by-row
        if sheet_lower == "customer_regions":
            all_text = " ".join(str(c) for r in agent_rows for c in r if c is not None).lower()
            record("Customer_Regions: ≥2 строк", len(agent_rows) >= 2,
                   f"Найдено {len(agent_rows)} строк")
            record("Customer_Regions: упоминается Санкт-Петербург",
                   "петербург" in all_text or "спб" in all_text or "saint" in all_text or "petersburg" in all_text,
                   f"Текст: {all_text[:200]}")
            record("Customer_Regions: упоминается Екатеринбург",
                   "екатеринбург" in all_text or "екб" in all_text or "yekaterinburg" in all_text or "ekaterinburg" in all_text,
                   f"Текст: {all_text[:200]}")
            continue

        # Для остальных — мягкое row-count (≥ GT)
        record(f"GT '{gt_sheet_name}' число строк ≥ {len(gt_rows)}",
               len(agent_rows) >= len(gt_rows),
               f"Ожидаем ≥{len(gt_rows)}, получено {len(agent_rows)}")

        skip_cols = SKIP_COLS.get(sheet_lower, set())

        # Сравниваем первые 2 строки
        check_indices = list(range(min(2, len(gt_rows))))
        for idx in check_indices:
            gt_row = gt_rows[idx]
            if idx >= len(agent_rows):
                record(f"GT '{gt_sheet_name}' строка {idx+1} существует", False, "Строка отсутствует у агента")
                continue
            a_row = agent_rows[idx]
            row_ok = True
            for col_idx in range(min(len(gt_row), len(a_row) if a_row else 0)):
                if col_idx in skip_cols:
                    continue
                gt_val = gt_row[col_idx]
                a_val = a_row[col_idx]
                if gt_val is None:
                    continue
                if isinstance(gt_val, (int, float)):
                    ok = num_close(a_val, gt_val, max(abs(gt_val) * 0.1, 1.0))
                else:
                    gt_str = str(gt_val).strip()
                    a_str = str(a_val or "").strip()
                    # Колонка Train_No — fuzzy
                    if any(k in gt_str for k in ("752", "016", "754", "074")):
                        ok = normalize_train(a_str) == normalize_train(gt_str)
                    # Колонка даты — нормализуем формат, но требуем семантического совпадения
                    elif gt_str[:4].isdigit() and len(gt_str) >= 8 and ("-" in gt_str or "." in gt_str):
                        ok = normalize_date(a_str) == normalize_date(gt_str)
                    else:
                        # Нестрогое: точное ИЛИ ≥1 ключевое слово
                        if str_match(a_val, gt_val):
                            ok = True
                        else:
                            gt_words = [w for w in gt_str.lower().split() if len(w) > 3]
                            a_low = a_str.lower()
                            hits = sum(1 for w in gt_words if w in a_low)
                            ok = hits >= 1 if gt_words else (gt_str.lower() in a_low or a_low in gt_str.lower())
                if not ok:
                    record(f"GT '{gt_sheet_name}' строка {idx+1} col {col_idx+1}",
                           False, f"Ожидаем {gt_val}, получено {a_val}")
                    row_ok = False
                    break
            if row_ok:
                record(f"GT '{gt_sheet_name}' строка {idx+1} значения совпали", True)
    gt_wb.close()


def check_teamly():
    print("\n=== Teamly: страница в TRIPS ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT p.id, p.title, p.body FROM teamly.pages p
            JOIN teamly.spaces s ON s.id = p.space_id
            WHERE s.key = 'TRIPS'
              AND p.id > 3
              AND (
                p.title ILIKE '%road show%' OR p.title ILIKE '%roadshow%'
                OR p.title ILIKE '%петербург%' OR p.title ILIKE '%екатеринбург%'
                OR p.body  ILIKE '%752%' OR p.body  ILIKE '%016%'
                OR p.body  ILIKE '%петербург%' OR p.body ILIKE '%екатеринбург%'
              )
            ORDER BY p.id DESC
        """)
        pages = cur.fetchall()
        cur.close()
        conn.close()
        record("Создана страница в Teamly/TRIPS", len(pages) >= 1,
               f"Найдено {len(pages)} страниц: {[p[1] for p in pages]}")

        if not pages:
            return

        # Берём любую подходящую страницу — у которой максимум вхождений
        best_body = ""
        best_score = -1
        for _id, _title, body in pages:
            body_l = (body or "").lower()
            score = sum(1 for k in ("752", "016", "петербург", "екатеринбург") if k in body_l)
            if score > best_score:
                best_score = score
                best_body = body_l

        record("Teamly: страница упоминает поезд 752А",
               "752" in best_body,
               f"body head: {best_body[:200]}")
        record("Teamly: страница упоминает поезд 016У",
               "016" in best_body,
               f"body head: {best_body[:200]}")
        record("Teamly: страница упоминает Санкт-Петербург",
               "петербург" in best_body or "спб" in best_body,
               f"body head: {best_body[:200]}")
        record("Teamly: страница упоминает Екатеринбург",
               "екатеринбург" in best_body or "екб" in best_body,
               f"body head: {best_body[:200]}")
    except Exception as e:
        record("Создана страница в Teamly/TRIPS", False, f"DB error: {e}")


def check_emails_sent():
    print("\n=== Email: отправленные письма ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT m.to_addr FROM email.messages m
            JOIN email.folders f ON m.folder_id = f.id
            WHERE UPPER(f.name) = 'SENT'
        """)
        sent_rows = cur.fetchall()
        cur.execute("""
            SELECT m.to_addr FROM email.sent_log sl
            JOIN email.messages m ON sl.message_id = m.id
        """)
        sent_rows += cur.fetchall()
        sent_text = " ".join(str(row[0]) for row in sent_rows).lower()

        record("Письмо отправлено на spb_dist@partner.ru",
               "spb_dist@partner.ru" in sent_text,
               f"Sent entries: {len(sent_rows)}")
        record("Письмо отправлено на ekb_dist@partner.ru",
               "ekb_dist@partner.ru" in sent_text,
               f"Sent entries: {len(sent_rows)}")
        record("Письмо отправлено на manager@company.ru",
               "manager@company.ru" in sent_text,
               f"Sent entries: {len(sent_rows)}")
    except Exception as e:
        record("Email sent check", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_teamly()
    check_emails_sent()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: ни одной проверки не выполнено.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
