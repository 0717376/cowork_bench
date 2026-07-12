"""
Evaluation for sf-sales-shipping-word (ClickHouse fork).

The sf_data SALES_DW ORDERS data values are russified centrally by
db/zzz_clickhouse_after_init.sql:
  SHIP_MODE: Economy->Эконом, Express->Экспресс, Next Day->На следующий день,
             Standard->Стандарт
  STATUS:    Delivered->Доставлен, Shipped->Отправлен
The agent reads these Russian values honestly from the DB. compute_expected()
self-syncs by recomputing from the live DB, so the per-mode names flow through
as Cyrillic on both the expected and the agent side. The only hand-edit needed
is the STATUS filter literal ('Доставлен' instead of 'Delivered').

Structure:
  * Soft checks feed an accuracy ratio (threshold >= 70%).
  * CRITICAL semantic checks (per-mode counts, total revenue, avg delivery days,
    sort order / mode presence) => sys.exit(1) on ANY failure before the
    accuracy gate.

UPPERCASE column/identifier names, the logical DB name SALES_DW, output FILE
names and sheet/column header names stay English (substring eval depends).

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""
import argparse
import json
import os
import re
import sys

import openpyxl
import psycopg2


def normalize_ru_numbers(text):
    """RU number normalization for substring/regex checks: collapse digit-group
    separators (space/NBSP/NNBSP/dot/comma before a 3-digit group) and turn
    decimal commas into dots ("31 588" -> "31588", "4 586,91" -> "4586.91")."""
    t = str(text or "")
    t = re.sub(r"(?<=\d)[ \xa0\u202f\u2009.,](?=\d{3}\b)", "", t)
    return re.sub(r"(?<=\d),(?=\d)", ".", t)

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Soft accuracy bookkeeping
TOTAL_CHECKS = 0
PASSED_CHECKS = 0
SOFT_ERRORS = []
CRITICAL_FAIL = []


def soft(name, condition, detail=""):
    """Non-critical check: contributes to the accuracy ratio."""
    global TOTAL_CHECKS, PASSED_CHECKS
    TOTAL_CHECKS += 1
    if condition:
        PASSED_CHECKS += 1
        print(f"  [PASS] {name}")
    else:
        d = (detail[:200] + "...") if len(detail) > 200 else detail
        SOFT_ERRORS.append(f"{name}: {d}")
        print(f"  [FAIL] {name}: {d}")


def crit(name, condition, detail=""):
    """Critical check: any failure => overall FAIL regardless of accuracy."""
    if condition:
        print(f"  [PASS-CRIT] {name}")
    else:
        d = (detail[:200] + "...") if len(detail) > 200 else detail
        CRITICAL_FAIL.append(f"{name}: {d}")
        print(f"  [FAIL-CRIT] {name}: {d}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_float(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    cur.execute("""
        SELECT "SHIP_MODE", COUNT(*) as orders,
               ROUND(AVG(("SHIP_DATE"::date - "ORDER_DATE"::date))::numeric, 1) as avg_days,
               ROUND(SUM("TOTAL_AMOUNT")::numeric, 2) as revenue,
               ROUND(AVG("TOTAL_AMOUNT")::numeric, 2) as avg_order
        FROM sf_data."SALES_DW__PUBLIC__ORDERS"
        WHERE "STATUS" = 'Доставлен'
        GROUP BY "SHIP_MODE" ORDER BY COUNT(*) DESC
    """)
    shipping_rows = cur.fetchall()

    cur.execute("""
        SELECT COUNT(*), ROUND(SUM("TOTAL_AMOUNT")::numeric, 2)
        FROM sf_data."SALES_DW__PUBLIC__ORDERS"
        WHERE "STATUS" = 'Доставлен'
    """)
    totals = cur.fetchone()
    conn.close()

    return {
        "shipping": shipping_rows,
        "total_orders": totals[0],
        "total_revenue": float(totals[1]),
    }


def check_excel(agent_workspace, expected):
    """Returns (agent_by_mode, agent_rows) for downstream critical checks."""
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Shipping_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        CRITICAL_FAIL.append(f"Excel file missing: {agent_file}")
        print(f"  [FAIL-CRIT] Excel file exists: {agent_file}")
        return {}, []
    print("  [PASS-CRIT] Excel file exists")

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        CRITICAL_FAIL.append(f"Excel file not readable: {e}")
        return {}, []

    ws = get_sheet(wb, "By Ship Mode")
    soft("Sheet 'By Ship Mode' exists", ws is not None, f"Found: {wb.sheetnames}")
    if ws is None:
        CRITICAL_FAIL.append("Sheet 'By Ship Mode' not found")
        return {}, []

    agent_rows = list(ws.iter_rows(min_row=2, values_only=True))
    agent_rows = [r for r in agent_rows if r and r[0] not in (None, "")]
    exp = expected["shipping"]

    soft("By Ship Mode row count", len(agent_rows) == len(exp),
         f"Expected {len(exp)}, got {len(agent_rows)}")

    agent_by_mode = {}
    for row in agent_rows:
        if row and row[0]:
            agent_by_mode[str(row[0]).strip().lower()] = row

    for exp_row in exp:
        mode = str(exp_row[0])
        agent_row = agent_by_mode.get(mode.lower())
        if agent_row:
            soft(f"'{mode}' Order_Count",
                 num_close(agent_row[1], exp_row[1], 5),
                 f"Expected {exp_row[1]}, got {agent_row[1]}")
            soft(f"'{mode}' Avg_Delivery_Days",
                 num_close(agent_row[2], float(exp_row[2]), 0.5),
                 f"Expected {exp_row[2]}, got {agent_row[2]}")
            soft(f"'{mode}' Total_Revenue",
                 num_close(agent_row[3], float(exp_row[3]), 500),
                 f"Expected {exp_row[3]}, got {agent_row[3]}")
            soft(f"'{mode}' Avg_Order_Value",
                 num_close(agent_row[4], float(exp_row[4]), 5),
                 f"Expected {exp_row[4]}, got {agent_row[4]}")
        else:
            soft(f"'{mode}' found in output", False, "Not in agent output")

    # Sort order (by Order_Count descending) -- non-critical structural
    if len(agent_rows) >= 2:
        counts = [to_float(r[1]) for r in agent_rows if r and r[1] is not None]
        counts = [c for c in counts if c is not None]
        soft("Sorted by Order_Count descending",
             all(counts[i] >= counts[i + 1] for i in range(len(counts) - 1)),
             f"Counts: {counts}")

    return agent_by_mode, agent_rows


def check_word(agent_workspace, expected):
    print("\n=== Checking Word Document ===")
    agent_file = os.path.join(agent_workspace, "Shipping_Report.docx")
    if not os.path.isfile(agent_file):
        CRITICAL_FAIL.append(f"Word file missing: {agent_file}")
        print(f"  [FAIL-CRIT] Word file exists: {agent_file}")
        return ""
    print("  [PASS-CRIT] Word file exists")

    try:
        from docx import Document
        doc = Document(agent_file)
    except ImportError:
        soft("python-docx available", False, "Cannot parse Word doc")
        return ""
    except Exception as e:
        soft("Word file readable", False, str(e))
        return ""

    full_text = "\n".join([p.text for p in doc.paragraphs])
    soft("Document contains 'Shipping Performance Report'",
         "shipping performance report" in full_text.lower(),
         f"Text preview: {full_text[:100]}")

    for exp_row in expected["shipping"]:
        mode = str(exp_row[0])
        soft(f"Document mentions '{mode}'",
             mode.lower() in full_text.lower(),
             f"'{mode}' not found in document text")

    total_str = str(expected["total_orders"])
    soft("Document mentions total order count",
         total_str in normalize_ru_numbers(full_text),
         f"Expected '{total_str}' in text")

    return full_text


def run_critical_checks(expected, agent_by_mode, agent_rows, word_text):
    """Semantic gating. Any failure => sys.exit(1) before the accuracy gate."""
    print("\n=== CRITICAL semantic constraints ===")
    exp = expected["shipping"]

    # C1: ALL russified ship-mode rows present with tight Order_Count (tol<=2).
    for exp_row in exp:
        mode = str(exp_row[0])
        a = agent_by_mode.get(mode.lower())
        if a is None:
            crit(f"[C1] mode '{mode}' present", False, "row missing in Excel")
            continue
        crit(f"[C1] '{mode}' Order_Count (tol<=2)",
             num_close(a[1], exp_row[1], 2),
             f"Expected {exp_row[1]}, got {a[1]}")

    # C2: per-mode Avg_Delivery_Days correct within 0.2 (verifies the metric
    #     was actually computed, not copied).
    for exp_row in exp:
        mode = str(exp_row[0])
        a = agent_by_mode.get(mode.lower())
        if a is None:
            continue
        crit(f"[C2] '{mode}' Avg_Delivery_Days (tol<=0.2)",
             num_close(a[2], float(exp_row[2]), 0.2),
             f"Expected {exp_row[2]}, got {a[2]}")

    # C3: combined total order count and total revenue (core deliverable) in the
    #     Word summary text and consistent with the DB aggregate.
    word_text_norm = normalize_ru_numbers(word_text)
    crit("[C3] Word summary contains total delivered order count",
         str(expected["total_orders"]) in word_text_norm,
         f"Expected '{expected['total_orders']}' in Word text")
    # Revenue separators are collapsed by normalize_ru_numbers; match the
    # rounded OR truncated integer part (true value has .66 decimals).
    rev_int = str(int(round(expected["total_revenue"])))
    rev_int_trunc = str(int(expected["total_revenue"]))
    crit("[C3] Word summary contains combined total revenue",
         any(s in word_text_norm for s in (rev_int, rev_int_trunc)),
         f"Expected revenue ~{rev_int} in Word text")

    # C4: Excel sorted by Order_Count descending AND all modes present.
    counts = []
    for r in agent_rows:
        v = to_float(r[1]) if r and len(r) > 1 else None
        if v is not None:
            counts.append(v)
    if len(counts) < len(exp):
        crit("[C4] all ship-mode rows present & sorted", False,
             f"Got {len(counts)} numeric rows, expected {len(exp)}")
    else:
        crit("[C4] Excel sorted by Order_Count descending",
             all(counts[i] >= counts[i + 1] for i in range(len(counts) - 1)),
             f"Counts: {counts}")


def check_excel_gt(agent_workspace, groundtruth_workspace):
    """Fallback when DB is unreachable: row-count consistency only."""
    print("\n=== Checking Excel (vs groundtruth) ===")
    agent_file = os.path.join(agent_workspace, "Shipping_Analysis.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Shipping_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        CRITICAL_FAIL.append(f"Excel file missing: {agent_file}")
        return
    if not os.path.isfile(gt_file):
        CRITICAL_FAIL.append(f"Groundtruth file missing: {gt_file}")
        return
    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    a_ws = get_sheet(agent_wb, "By Ship Mode")
    g_ws = get_sheet(gt_wb, "By Ship Mode")
    soft("Sheet 'By Ship Mode' exists", a_ws is not None)
    if a_ws and g_ws:
        a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        g_rows = [r for r in g_ws.iter_rows(min_row=2, values_only=True) if r and r[0]]
        soft("Row count matches groundtruth", len(a_rows) == len(g_rows),
             f"Expected {len(g_rows)}, got {len(a_rows)}")
        # russified ship-mode names must all appear (gt is keyed by them)
        a_modes = {str(r[0]).strip().lower() for r in a_rows}
        for g in g_rows:
            crit(f"[GT] mode '{g[0]}' present",
                 str(g[0]).strip().lower() in a_modes,
                 f"'{g[0]}' not in agent output")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    expected = compute_expected()
    if expected:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
        agent_by_mode, agent_rows = check_excel(agent_workspace, expected)
        word_text = check_word(agent_workspace, expected)
        run_critical_checks(expected, agent_by_mode, agent_rows, word_text)
    else:
        print("INFO: Falling back to groundtruth (DB unreachable)")
        check_excel_gt(agent_workspace, groundtruth_workspace)
        # Word: only the heading is verifiable without DB numbers.
        wf = os.path.join(agent_workspace, "Shipping_Report.docx")
        if not os.path.isfile(wf):
            CRITICAL_FAIL.append(f"Word file missing: {wf}")

    accuracy = (PASSED_CHECKS / TOTAL_CHECKS * 100.0) if TOTAL_CHECKS else 0.0

    print("\n=== SUMMARY ===")
    print(f"  Soft accuracy: {PASSED_CHECKS}/{TOTAL_CHECKS} = {accuracy:.1f}%")
    if SOFT_ERRORS:
        print(f"  Soft errors ({len(SOFT_ERRORS)}):")
        for e in SOFT_ERRORS[:10]:
            print(f"    {e}")

    if CRITICAL_FAIL:
        print(f"\n=== CRITICAL FAIL ({len(CRITICAL_FAIL)}) ===")
        for e in CRITICAL_FAIL[:15]:
            print(f"  {e}")
        success = False
    else:
        success = accuracy >= 70.0

    print(f"\n  Overall: {'PASS' if success else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASSED_CHECKS,
            "failed": TOTAL_CHECKS - PASSED_CHECKS,
            "accuracy": accuracy,
            "critical_fail": CRITICAL_FAIL,
            "success": success,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)

    return success, f"Accuracy: {accuracy:.1f}%, Critical fails: {len(CRITICAL_FAIL)}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
