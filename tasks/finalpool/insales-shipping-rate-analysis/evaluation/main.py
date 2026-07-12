"""Evaluation for insales-shipping-rate-analysis."""
import argparse
import json
import os
import sys

import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _to_int(v):
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def check_excel(agent_workspace, gt_data):
    """Soft structural checks: sheets exist, row counts roughly match."""
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Shipping_Audit.xlsx")
    if not os.path.exists(path):
        return ["Shipping_Audit.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # Check Rate Comparison sheet
        rows = load_sheet_rows(wb, "Rate Comparison")
        if rows is None:
            errors.append("Sheet 'Rate Comparison' not found")
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            expected = gt_data["total_orders"]
            if abs(len(data_rows) - expected) > 5:
                errors.append(f"Rate Comparison has {len(data_rows)} rows, expected ~{expected}")

        # Check Undercharged Orders sheet
        rows2 = load_sheet_rows(wb, "Undercharged Orders")
        if rows2 is None:
            errors.append("Sheet 'Undercharged Orders' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            expected_uc = gt_data["undercharged_count"]
            if abs(len(data_rows2) - expected_uc) > 5:
                errors.append(f"Undercharged Orders has {len(data_rows2)} rows, expected ~{expected_uc}")

        # Check Summary sheet
        rows3 = load_sheet_rows(wb, "Summary")
        if rows3 is None:
            errors.append("Sheet 'Summary' not found")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def read_audit_xlsx(agent_workspace):
    """Return parsed audit data for critical checks, or (None, error)."""
    import openpyxl
    path = os.path.join(agent_workspace, "Shipping_Audit.xlsx")
    if not os.path.exists(path):
        return None, "Shipping_Audit.xlsx not found"
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {"summary": {}, "undercharged_ids": set()}

    rows3 = load_sheet_rows(wb, "Summary")
    if rows3:
        for r in rows3[1:]:
            if r and r[0] is not None:
                out["summary"][str(r[0]).strip().lower()] = r[1]

    rows2 = load_sheet_rows(wb, "Undercharged Orders")
    if rows2:
        for r in rows2[1:]:
            if r and r[0] is not None:
                oid = _to_int(r[0])
                if oid is not None:
                    out["undercharged_ids"].add(oid)
    return out, None


def _summary_lookup(summary, *needles):
    for sk, sv in summary.items():
        if all(n in sk for n in needles):
            return sv
    return None


def critical_excel_checks(agent_workspace, gt_data):
    """CRITICAL: semantic correctness of the audit numbers and order set."""
    errors = []
    try:
        data, err = read_audit_xlsx(agent_workspace)
    except Exception as e:
        return [f"CRITICAL: cannot read Shipping_Audit.xlsx: {e}"]
    if err:
        return [f"CRITICAL: {err}"]

    summary = data["summary"]

    # Undercharged_Count within 1, Total_Orders_Analyzed within 2.
    uc_val = _to_int(_summary_lookup(summary, "undercharged", "count"))
    if uc_val is None:
        errors.append("CRITICAL: Summary missing Undercharged_Count")
    elif abs(uc_val - gt_data["undercharged_count"]) > 1:
        errors.append(
            f"CRITICAL: Undercharged_Count={uc_val}, expected {gt_data['undercharged_count']}"
        )

    tot_val = _to_int(_summary_lookup(summary, "total", "orders"))
    if tot_val is None:
        errors.append("CRITICAL: Summary missing Total_Orders_Analyzed")
    elif abs(tot_val - gt_data["total_orders"]) > 2:
        errors.append(
            f"CRITICAL: Total_Orders_Analyzed={tot_val}, expected {gt_data['total_orders']}"
        )

    # Total_Undercharged_Amount within ~5% (proves Difference computed correctly).
    amt = _summary_lookup(summary, "total", "amount")
    try:
        amt = float(amt)
        gt_amt = float(gt_data["total_undercharged_amount"])
        if abs(amt - gt_amt) > max(0.05 * gt_amt, 1.0):
            errors.append(
                f"CRITICAL: Total_Undercharged_Amount={amt}, expected ~{gt_amt}"
            )
    except (ValueError, TypeError):
        errors.append("CRITICAL: Summary Total_Undercharged_Amount missing or non-numeric")

    # Undercharged order-ID set must overlap gt by >=90% (correct carrier mapping).
    gt_ids = set(gt_data["undercharged_order_ids"])
    agent_ids = data["undercharged_ids"]
    if not agent_ids:
        errors.append("CRITICAL: 'Undercharged Orders' sheet has no order IDs")
    elif gt_ids:
        overlap = len(agent_ids & gt_ids) / len(gt_ids)
        if overlap < 0.90:
            errors.append(
                f"CRITICAL: undercharged order-ID overlap with groundtruth is "
                f"{overlap:.0%} (<90%); carrier-rate mapping likely wrong"
            )
    return errors


def _fetch_email():
    conn = psycopg2.connect(
        host=os.environ.get("PGHOST", "localhost"), port=5432,
        dbname=os.environ.get("PGDATABASE", "cowork_gym"),
        user="eigent", password="camel",
    )
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, to_addr, body_text FROM email.messages
        WHERE to_addr::text ILIKE '%logistics@company.com%'
        ORDER BY id DESC LIMIT 5
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return rows


def check_email(gt_data):
    """Soft check: an email exists and mentions the audit (RU or EN)."""
    errors = []
    try:
        rows = _fetch_email()
        if not rows:
            errors.append("No email found sent to logistics@company.com")
        else:
            body = (rows[0][2] or "").lower()
            kws = ("undercharg", "audit", "недозаряд", "недозаряж", "аудит", "доплат")
            if not any(k in body for k in kws):
                errors.append("Email body does not mention undercharged orders or audit")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def critical_email_checks(gt_data):
    """CRITICAL: email to logistics with correct subject, count, amount, order IDs."""
    import re
    errors = []
    try:
        rows = _fetch_email()
    except Exception as e:
        return [f"CRITICAL: cannot query email: {e}"]
    if not rows:
        return ["CRITICAL: no email sent to logistics@company.com"]

    subject = (rows[0][0] or "")
    body = (rows[0][2] or "")
    body_l = body.lower()

    if "shipping cost audit alert" not in subject.lower():
        errors.append(f"CRITICAL: email subject is '{subject}', expected 'Shipping Cost Audit Alert'")

    # Body must mention undercharge/audit (RU or EN).
    kws = ("undercharg", "audit", "недозаряд", "недозаряж", "аудит", "доплат")
    if not any(k in body_l for k in kws):
        errors.append("CRITICAL: email body lacks undercharge/audit wording")

    # Body must state the undercharged count.
    uc = gt_data["undercharged_count"]
    nums_in_body = set(int(n) for n in re.findall(r"\d+", body))
    if not any(abs(uc - n) <= 1 for n in nums_in_body):
        errors.append(f"CRITICAL: email body does not state the undercharged count (~{uc})")

    # Body must state the total amount (within ~5%).
    gt_amt = float(gt_data["total_undercharged_amount"])
    floats_in_body = [float(x) for x in re.findall(r"\d+\.\d+", body)]
    if not any(abs(f - gt_amt) <= max(0.05 * gt_amt, 1.0) for f in floats_in_body):
        errors.append(f"CRITICAL: email body does not state total undercharged amount (~{gt_amt})")

    # Body must list at least ~10 of the undercharged order IDs.
    gt_ids = set(gt_data["undercharged_order_ids"])
    listed = nums_in_body & gt_ids
    if len(listed) < 10:
        errors.append(
            f"CRITICAL: email lists only {len(listed)} undercharged order IDs, need >=10"
        )
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")

    with open(os.path.join(task_root, "files", "groundtruth_data.json")) as f:
        gt_data = json.load(f)

    # === CRITICAL semantic checks (any failure => immediate FAIL) ===
    critical_errors = []
    print("  Running CRITICAL checks...")
    critical_errors.extend(critical_excel_checks(agent_ws, gt_data))
    critical_errors.extend(critical_email_checks(gt_data))
    if critical_errors:
        print(f"\n=== RESULT: FAIL (CRITICAL: {len(critical_errors)}) ===")
        for e in critical_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    print("    CRITICAL PASS")

    all_errors = []

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking email...")
    errs = check_email(gt_data)
    if errs:
        all_errors.extend(errs)
        for e in errs[:3]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
