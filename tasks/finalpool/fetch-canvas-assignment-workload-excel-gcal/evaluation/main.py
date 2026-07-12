"""Оценка для fetch-canvas-assignment-workload-excel-gcal.

Критические проверки (CRITICAL_CHECKS): провал любой из них => общий FAIL
независимо от accuracy. В остальном PASS требует accuracy >= 70%.

Семантика проверяется по собственным данным агента (внутренняя согласованность
отчёта) и по событию календаря с точной датой/временем. Волатильные данные курсов
читаются «вживую» из Canvas самим агентом и НЕ захардкожены здесь.
"""
import os
import argparse, json, os, sys
from datetime import datetime
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Критические (семантические) проверки: провал любой => общий FAIL.
CRITICAL_CHECKS = {
    "Event 'Analysis Review' on 2026-03-14 14:00-15:00 UTC",
    "course_workload_results.json exists and is valid JSON",
    "Data_Analysis sorted alphabetically by Course",
    "Metrics Total_Courses matches Data_Analysis row count",
    "Metrics Total_Enrollment matches sum of Enrollment column",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def col_index(headers, name):
    """Индекс колонки по имени (без учёта регистра), либо None."""
    name = name.lower()
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES.clear()

    excel_path = os.path.join(agent_workspace, "Assignment_Workload_Report.xlsx")
    check("Assignment_Workload_Report.xlsx exists", os.path.exists(excel_path))

    da_headers = []
    da_rows = []
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(da_rows) >= 6, f"got {len(da_rows)}")

            da_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Course', 'Code', 'Enrollment', 'Avg_Score', 'Pass_Rate']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in da_headers, f"headers: {da_headers[:8]}")

            # Мягкая (некритическая) проверка: эталон/разрыв упомянуты в task.md,
            # но groundtruth их не содержит -> только поощряем, не валим.
            has_benchmark = any('benchmark' in h or 'эталон' in h for h in da_headers)
            has_gap = any('gap' in h or 'разрыв' in h or 'diff' in h for h in da_headers)
            check("Data_Analysis has benchmark column (soft)", has_benchmark, f"headers: {da_headers[:8]}")
            check("Data_Analysis has gap/difference column (soft)", has_gap, f"headers: {da_headers[:8]}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            m_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(m_rows) >= 4, f"got {len(m_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in m_rows:
                if r and r[0] is not None:
                    metrics_map[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            rec_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(rec_rows) >= 2, f"got {len(rec_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Course']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # --- КРИТИЧЕСКАЯ: сортировка Data_Analysis по Course (алфавит) ---
        ci_course = col_index(da_headers, 'course')
        if ci_course is not None and da_rows:
            course_vals = [str(r[ci_course]).strip() for r in da_rows if r and r[ci_course] is not None]
            sorted_ok = course_vals == sorted(course_vals, key=lambda s: s.lower())
            check("Data_Analysis sorted alphabetically by Course", sorted_ok,
                  f"got: {course_vals[:6]}")
        else:
            check("Data_Analysis sorted alphabetically by Course", False,
                  "no Course column / no rows")

        # --- КРИТИЧЕСКАЯ: внутренняя согласованность Metrics (из данных агента) ---
        ci_enr = col_index(da_headers, 'enrollment')
        total_courses_val = safe_float(metrics_map.get('total_courses'))
        check("Metrics Total_Courses matches Data_Analysis row count",
              total_courses_val is not None and abs(total_courses_val - len(da_rows)) < 0.5,
              f"Total_Courses={total_courses_val}, rows={len(da_rows)}")

        if ci_enr is not None and da_rows:
            enr_sum = sum(safe_float(r[ci_enr], 0) or 0 for r in da_rows if r)
            total_enr_val = safe_float(metrics_map.get('total_enrollment'))
            check("Metrics Total_Enrollment matches sum of Enrollment column",
                  total_enr_val is not None and abs(total_enr_val - enr_sum) < 1.0,
                  f"Total_Enrollment={total_enr_val}, sum={enr_sum}")
        else:
            check("Metrics Total_Enrollment matches sum of Enrollment column", False,
                  "no Enrollment column / no rows")

    else:
        # xlsx отсутствует -> явно валим зависимые критические проверки,
        # а не пропускаем их молча.
        check("Data_Analysis sorted alphabetically by Course", False, "no xlsx")
        check("Metrics Total_Courses matches Data_Analysis row count", False, "no xlsx")
        check("Metrics Total_Enrollment matches sum of Enrollment column", False, "no xlsx")

    # --- КРИТИЧЕСКАЯ: course_workload_results.json существует и валиден ---
    results_path = os.path.join(agent_workspace, "course_workload_results.json")
    valid_json = False
    if os.path.exists(results_path):
        try:
            with open(results_path, "r", encoding="utf-8") as f:
                json.load(f)
            valid_json = True
        except Exception as e:
            valid_json = False
    check("course_workload_results.json exists and is valid JSON", valid_json,
          f"path exists={os.path.exists(results_path)}")

    check("course_workload_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "course_workload_processor.py")))

    # --- КРИТИЧЕСКАЯ: событие 'Analysis Review' 14.03.2026 14:00-15:00 UTC ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""SELECT summary, start_datetime, end_datetime
                         FROM gcal.events
                        WHERE summary ILIKE %s""", ('%Analysis Review%',))
        events = cur.fetchall()
        conn.close()

        def matches(ev):
            s, start, end = ev
            if start is None or end is None:
                return False
            return (start.year == 2026 and start.month == 3 and start.day == 14
                    and start.hour == 14 and start.minute == 0
                    and end.hour == 15 and end.minute == 0)

        ok = any(matches(e) for e in events)
        check("Event 'Analysis Review' on 2026-03-14 14:00-15:00 UTC", ok,
              f"events: {[(e[0], str(e[1]), str(e[2])) for e in events]}")
    except Exception as e:
        check("Event 'Analysis Review' on 2026-03-14 14:00-15:00 UTC", False, str(e))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Результат: {PASS_COUNT}/{total} пройдено ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"КРИТИЧЕСКИЕ ПРОВАЛЫ: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")
        return False, f"FAIL (критические провалы): {PASS_COUNT}/{total}"

    if accuracy >= 70:
        return True, f"PASS: {PASS_COUNT}/{total} ({accuracy:.1f}%)"
    return False, f"FAIL: accuracy {accuracy:.1f}% < 70%"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
