"""Evaluation для yf-portfolio-gsheet-pdf-gcal-email (RU, MCP moex-finance).

Модель оценки:
  * Набор обычных (структурных) чеков + множество CRITICAL_CHECKS.
  * Любой провал CRITICAL => немедленный FAIL (sys.exit(1)) ДО порога точности.
  * Иначе PASS, если точность (доля пройденных чеков) >= 70%.

Ключевой принцип: НЕ хардкодим волатильные цены. Ожидаемые доли/отклонения/
действия ПЕРЕСЧИТЫВАЕМ из живого источника moex.stock_prices / moex.stock_info,
исходя из равных вложений 10000 RUB на бумагу и цены покупки = закрытие на
PURCHASE_DATE. Правило ребалансировки: |drift| <= 3 п.п. => Hold; drift > 3 =>
Sell (перевес); drift < -3 => Buy (недовес). Целевая доля — 20% на бумагу.

Если все бумаги в пределах порога (пустой кейс), письмо Drift Alert и описание
события ДОЛЖНЫ корректно указывать, что ребалансировка не требуется — это
валидный и проверяемый исход.
"""
import argparse
import json
import os
import sys

import openpyxl

TICKERS = ["SBER.ME", "GAZP.ME", "LKOH.ME", "MGNT.ME", "MTSS.ME"]
INVEST_PER = 10000.0
TARGET = 20.0
DRIFT_THRESHOLD = 3.0
PURCHASE_DATE = "2026-03-06"

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []
CRITICAL_CHECKS = {
    "Логика ребалансировки в Rebalancing согласована с правилом 3 п.п. (Action из drift)",
    "Holdings содержит ровно 5 тикеров MOEX с корректными RU-названиями и секторами; сумма долей ~100%",
    "Письмо Portfolio Drift Alert корректно перечисляет бумаги > 3 п.п. (или указывает, что таких нет)",
    "Событие Portfolio Rebalancing Review на 2026-03-17 14:00-15:00 UTC с корректным описанием",
    "Шум сохранён: шумовые письма не в Sent, шумовые события календаря не удалены",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL]{marker} {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def compute_expected(cur):
    """Пересчитываем ожидаемые значения из живого источника moex.*.

    Возвращает dict[ticker] -> {company, sector, current, purchase, shares,
    market_value, alloc, drift, action, ret, status} и список тикеров, которым
    нужна ребалансировка (|drift| > 3 п.п.).
    """
    rows = {}
    for t in TICKERS:
        cur.execute("SELECT data FROM moex.stock_info WHERE symbol = %s", (t,))
        r = cur.fetchone()
        info = r[0] if r else {}
        if isinstance(info, str):
            info = json.loads(info)
        current = info.get("currentPrice") or info.get("regularMarketPrice")
        # цена покупки = закрытие на PURCHASE_DATE (или ближайшее <= даты)
        cur.execute(
            "SELECT close FROM moex.stock_prices WHERE symbol = %s AND date <= %s "
            "ORDER BY date DESC LIMIT 1", (t, PURCHASE_DATE))
        pr = cur.fetchone()
        purchase = float(pr[0]) if pr and pr[0] is not None else None
        rows[t] = {
            "company": info.get("longName", ""),
            "sector": info.get("sector", ""),
            "current": float(current) if current is not None else None,
            "purchase": purchase,
        }
    # рыночная стоимость / доли
    total = 0.0
    for t in TICKERS:
        d = rows[t]
        if d["current"] and d["purchase"]:
            d["shares"] = INVEST_PER / d["purchase"]
            d["market_value"] = d["shares"] * d["current"]
        else:
            d["shares"] = None
            d["market_value"] = None
        if d["market_value"]:
            total += d["market_value"]
    need_rebal = []
    for t in TICKERS:
        d = rows[t]
        if d["market_value"] and total:
            d["alloc"] = d["market_value"] / total * 100.0
            d["drift"] = d["alloc"] - TARGET
            if d["drift"] > DRIFT_THRESHOLD:
                d["action"] = "Sell"
            elif d["drift"] < -DRIFT_THRESHOLD:
                d["action"] = "Buy"
            else:
                d["action"] = "Hold"
            if abs(d["drift"]) > DRIFT_THRESHOLD:
                need_rebal.append(t)
        else:
            d["alloc"] = d["drift"] = None
            d["action"] = None
        if d["current"] and d["purchase"]:
            d["ret"] = (d["current"] - d["purchase"]) / d["purchase"] * 100.0
            d["status"] = "Gain" if d["ret"] >= 0 else "Loss"
        else:
            d["ret"] = None
            d["status"] = None
    return rows, need_rebal


def sheet_to_dicts(ws):
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        out.append({headers[i]: row[i] for i in range(len(headers)) if i < len(row)})
    return headers, out


def check_xlsx(agent_workspace, expected):
    excel_path = os.path.join(agent_workspace, "Portfolio_Dashboard_Reference.xlsx")
    check("Portfolio_Dashboard_Reference.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        return
    wb = openpyxl.load_workbook(excel_path)
    for sn in ["Holdings", "Performance", "Rebalancing"]:
        check(f"{sn} sheet exists", sn in wb.sheetnames)

    # --- Holdings (CRITICAL) ---
    if "Holdings" in wb.sheetnames:
        headers, rows = sheet_to_dicts(wb["Holdings"])
        for col in ["Symbol", "Company", "Sector", "Current_Price",
                    "Shares_Held", "Market_Value", "Allocation_Pct"]:
            check(f"Holdings has {col} column", col in headers, f"headers: {headers}")
        by_sym = {str(r.get("Symbol", "")).strip().upper(): r for r in rows}
        present = [t for t in TICKERS if t in by_sym]
        alloc_sum = sum(safe_float(by_sym[t].get("Allocation_Pct"), 0) or 0 for t in present)
        names_ok = True
        sectors_ok = True
        for t in present:
            exp = expected[t]
            got_co = str(by_sym[t].get("Company", "")).strip().lower()
            got_sec = str(by_sym[t].get("Sector", "")).strip().lower()
            if exp["company"] and exp["company"].strip().lower() not in got_co \
                    and got_co not in exp["company"].strip().lower():
                names_ok = False
            if exp["sector"] and exp["sector"].strip().lower() != got_sec:
                sectors_ok = False
        check(
            "Holdings содержит ровно 5 тикеров MOEX с корректными RU-названиями и секторами; сумма долей ~100%",
            len(present) == 5 and names_ok and sectors_ok and abs(alloc_sum - 100.0) <= 1.5,
            f"present={present} names_ok={names_ok} sectors_ok={sectors_ok} alloc_sum={alloc_sum:.2f}")
        # мягкая проверка значений доли (recompute, не хардкод)
        for t in present:
            exp = expected[t]
            got = safe_float(by_sym[t].get("Allocation_Pct"))
            if exp["alloc"] is not None and got is not None:
                check(f"Holdings {t} Allocation_Pct ~{exp['alloc']:.1f}",
                      abs(got - exp["alloc"]) <= 1.5, f"got {got}")

    # --- Performance (мягко) ---
    if "Performance" in wb.sheetnames:
        headers, rows = sheet_to_dicts(wb["Performance"])
        for col in ["Symbol", "Purchase_Price", "Current_Price", "Return_Pct", "Status"]:
            check(f"Performance has {col} column", col in headers, f"headers: {headers}")
        by_sym = {str(r.get("Symbol", "")).strip().upper(): r for r in rows}
        for t in [x for x in TICKERS if x in by_sym]:
            exp = expected[t]
            got_status = str(by_sym[t].get("Status", "")).strip().lower()
            if exp["status"]:
                check(f"Performance {t} Status = {exp['status']}",
                      got_status in (exp["status"].lower(), ""),
                      f"got {got_status}")

    # --- Rebalancing (CRITICAL: согласованность правила 3 п.п.) ---
    if "Rebalancing" in wb.sheetnames:
        headers, rows = sheet_to_dicts(wb["Rebalancing"])
        for col in ["Symbol", "Current_Allocation", "Target_Allocation",
                    "Drift_Pct", "Action"]:
            check(f"Rebalancing has {col} column", col in headers, f"headers: {headers}")
        by_sym = {str(r.get("Symbol", "")).strip().upper(): r for r in rows}
        rule_ok = True
        detail = []
        for t in [x for x in TICKERS if x in by_sym]:
            r = by_sym[t]
            drift = safe_float(r.get("Drift_Pct"))
            action = str(r.get("Action", "")).strip().lower()
            if drift is None or action == "":
                rule_ok = False
                detail.append(f"{t}: missing drift/action")
                continue
            # Action должен соответствовать ЗАЯВЛЕННОМУ агентом drift
            if drift > DRIFT_THRESHOLD:
                expected_act = "sell"
            elif drift < -DRIFT_THRESHOLD:
                expected_act = "buy"
            else:
                expected_act = "hold"
            if action != expected_act:
                rule_ok = False
                detail.append(f"{t}: drift={drift} action={action} expected={expected_act}")
        # И drift должен согласоваться с пересчитанным из источника (с допуском)
        src_ok = True
        for t in [x for x in TICKERS if x in by_sym]:
            exp = expected[t]
            drift = safe_float(by_sym[t].get("Drift_Pct"))
            if exp["drift"] is not None and drift is not None:
                if abs(drift - exp["drift"]) > 1.5:
                    src_ok = False
                    detail.append(f"{t}: drift={drift} src={exp['drift']:.2f}")
        check(
            "Логика ребалансировки в Rebalancing согласована с правилом 3 п.п. (Action из drift)",
            rule_ok and src_ok and len([x for x in TICKERS if x in by_sym]) == 5,
            "; ".join(detail))


def check_db(expected, need_rebal):
    try:
        conn = get_conn()
        cur = conn.cursor()
    except Exception as e:
        check("DB connection", False, str(e))
        return

    # Google Sheet существование (структурно)
    cur.execute("SELECT COUNT(*) FROM gsheet.spreadsheets")
    check("Google Sheet created", cur.fetchone()[0] >= 1)
    cur.execute("SELECT COUNT(*) FROM gsheet.cells")
    check("Sheet has data", cur.fetchone()[0] >= 10)

    # --- Письмо Drift Alert (CRITICAL) ---
    cur.execute(
        "SELECT subject, to_addr, body_text FROM email.messages "
        "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
        "AND subject ILIKE '%portfolio%'")
    email_row = cur.fetchone()
    check("Email with correct subject sent", email_row is not None)
    if email_row:
        check("Email has recipient",
              email_row[1] is not None and "investment-team@company.com" in str(email_row[1]),
              f"to_addr: {email_row[1]}")
    body = (email_row[2] or "") if email_row else ""
    body_low = body.lower()
    email_ok = False
    if email_row:
        if need_rebal:
            # должны быть перечислены ИМЕННО эти тикеры
            base = [t.split(".")[0] for t in need_rebal]
            email_ok = all((t in body) or (b in body) for t, b in zip(need_rebal, base))
        else:
            # пустой кейс: явно указано, что ребалансировка не требуется / нет бумаг
            empty_markers = ["не требуется", "не превыш", "нет бумаг", "ни одна",
                             "no rebal", "none", "within", "no stock"]
            email_ok = any(m in body_low for m in empty_markers)
    check(
        "Письмо Portfolio Drift Alert корректно перечисляет бумаги > 3 п.п. (или указывает, что таких нет)",
        email_ok, f"need_rebal={need_rebal} body[:160]={body[:160]!r}")

    # --- Событие Rebalancing Review (CRITICAL) ---
    cur.execute(
        "SELECT summary, start_datetime, end_datetime, description "
        "FROM gcal.events WHERE summary ILIKE '%rebalanc%' OR summary ILIKE '%portfolio%'")
    ev = cur.fetchone()
    check("Calendar event exists", ev is not None)
    event_ok = False
    if ev:
        summary, sdt, edt, desc = ev
        desc = desc or ""
        desc_low = desc.lower()
        sdt_s = str(sdt)
        edt_s = str(edt)
        date_ok = "2026-03-17" in sdt_s and "14:00" in sdt_s and "15:00" in edt_s
        if need_rebal:
            base = [t.split(".")[0] for t in need_rebal]
            desc_ok = all((t in desc) or (b in desc) for t, b in zip(need_rebal, base))
        else:
            empty_markers = ["не требуется", "не превыш", "нет бумаг", "ни одна",
                             "no rebal", "none", "within", "no stock"]
            desc_ok = any(m in desc_low for m in empty_markers)
        event_ok = date_ok and desc_ok
        check("Event date/time 2026-03-17 14:00-15:00", date_ok, f"start={sdt_s} end={edt_s}")
    check(
        "Событие Portfolio Rebalancing Review на 2026-03-17 14:00-15:00 UTC с корректным описанием",
        event_ok, f"need_rebal={need_rebal}")

    # --- Сохранение шума (CRITICAL) ---
    cur.execute(
        "SELECT COUNT(*) FROM email.messages WHERE folder_id = "
        "(SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
        "AND (subject ILIKE '%newsletter%' OR subject ILIKE '%maintenance%')")
    noise_sent = cur.fetchone()[0]
    cur.execute(
        "SELECT COUNT(*) FROM gcal.events WHERE summary ILIKE '%standup%' OR summary ILIKE '%lunch%'")
    noise_events = cur.fetchone()[0]
    check(
        "Шум сохранён: шумовые письма не в Sent, шумовые события календаря не удалены",
        noise_sent == 0 and noise_events >= 2,
        f"noise_sent={noise_sent} noise_events={noise_events}")

    conn.close()


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    # Пересчёт ожидаемого из живого источника moex.*
    try:
        conn = get_conn()
        cur = conn.cursor()
        expected, need_rebal = compute_expected(cur)
        conn.close()
    except Exception as e:
        check("moex source recompute", False, str(e))
        expected, need_rebal = {t: {} for t in TICKERS}, []

    check_xlsx(agent_workspace, expected)

    # Python-скрипт (использование терминала)
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")] \
        if os.path.isdir(agent_workspace) else []
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")

    check_db(expected, need_rebal)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        return False, "FAIL: чеки не выполнялись."
    accuracy = PASS_COUNT / total * 100
    msg = f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

    result = {
        "total_passed": PASS_COUNT, "total_checks": total,
        "accuracy": accuracy, "critical_fails": CRITICAL_FAILS,
    }
    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        return False, f"FAIL: критичные чеки провалены ({len(CRITICAL_FAILS)}): {CRITICAL_FAILS}. {msg}"
    return accuracy >= 70, msg


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file)
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
