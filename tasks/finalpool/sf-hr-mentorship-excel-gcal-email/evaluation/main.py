"""Evaluation for sf-hr-mentorship-excel-gcal-email (ClickHouse fork).

The mentor/mentee selection is RECOMPUTED LIVE from the russified ClickHouse DWH
schema sf_data (HR_ANALYTICS__PUBLIC__EMPLOYEES). Employee/department literals are
russified CENTRALLY by db/zzz_clickhouse_after_init.sql, so the agent legitimately
writes Russian names; this eval reads the expected set dynamically and never
hardcodes russified literals.

CRITICAL semantic checks (any fail => sys.exit(1) before the accuracy gate):
  - >= 8 of the agent's Pairs.Mentor_Name match the live-recomputed mentor set
  - Total_Pairs == 10 exactly AND Pairs sheet has exactly 10 data rows
  - Avg_Mentor_Rating == avg rating of recomputed mentors within 0.01
  - Avg_Mentee_Experience == avg experience of recomputed mentees within 0.01
  - Email hr@company.example.com -> program@hr.example.com, subject contains
    'Mentorship Program Launch', body mentions the total pair count (10)
Structural checks (sheet/column presence, ISO date, generic event presence) are
NON-critical.
"""
import argparse
import os
import re
import sys
import unicodedata
from datetime import datetime, timedelta

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = " [CRIT]" if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL]{tag} {name}: {str(detail)[:200]}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def norm_name(s):
    """Case/space-insensitive key for matching Russian names (NFKD-folded)."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.strip().lower().split())


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def get_conn():
    return psycopg2.connect(**DB)


def load_expected_from_db():
    """Recompute mentor/mentee sets live from the russified DWH."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute(
        'SELECT "EMPLOYEE_NAME", "PERFORMANCE_RATING", "YEARS_EXPERIENCE" '
        'FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" '
        'WHERE "YEARS_EXPERIENCE" >= 10 AND "PERFORMANCE_RATING" >= 4 '
        'ORDER BY "PERFORMANCE_RATING" DESC, "YEARS_EXPERIENCE" DESC LIMIT 10'
    )
    mentors = cur.fetchall()
    cur.execute(
        'SELECT "EMPLOYEE_NAME", "PERFORMANCE_RATING", "YEARS_EXPERIENCE" '
        'FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES" '
        'WHERE "YEARS_EXPERIENCE" <= 2 AND "PERFORMANCE_RATING" >= 3 '
        'ORDER BY "PERFORMANCE_RATING" DESC LIMIT 10'
    )
    mentees = cur.fetchall()
    cur.close()
    conn.close()
    return mentors, mentees


def check_excel(agent_workspace, mentors, mentees):
    print("\n=== Checking Mentorship_Pairs.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Mentorship_Pairs.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        check("CRITICAL At least 8 mentor names match recomputed set", False, "no file", critical=True)
        check("CRITICAL Total_Pairs == 10 and 10 data rows", False, "no file", critical=True)
        check("CRITICAL Avg_Mentor_Rating matches recomputed", False, "no file", critical=True)
        check("CRITICAL Avg_Mentee_Experience matches recomputed", False, "no file", critical=True)
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return False

    all_ok = True

    # Expected values from the live DB
    exp_mentor_keys = {norm_name(m[0]) for m in mentors}
    exp_avg_rating = round(sum(m[1] for m in mentors) / len(mentors), 2) if mentors else None
    exp_avg_exp = round(sum(me[2] for me in mentees) / len(mentees), 2) if mentees else None

    # ---- Pairs sheet ----
    agent_pairs = get_sheet(agent_wb, "Pairs")
    check("Sheet 'Pairs' exists", agent_pairs is not None, f"Sheets: {agent_wb.sheetnames}")
    n_rows = 0
    if agent_pairs is None:
        all_ok = False
        check("CRITICAL At least 8 mentor names match recomputed set", False, "no Pairs sheet", critical=True)
    else:
        a_rows = [r for r in agent_pairs.iter_rows(min_row=2, values_only=True)
                  if r and any(c is not None for c in r)]
        n_rows = len(a_rows)
        check("Pairs has 10 data rows", n_rows == 10, f"Got {n_rows}")

        agent_mentor_keys = {norm_name(r[0]) for r in a_rows if r and r[0]}
        matched = len(exp_mentor_keys & agent_mentor_keys)
        check("CRITICAL At least 8 mentor names match recomputed set", matched >= 8,
              f"Matched {matched}/{len(exp_mentor_keys)}; expected sample={list(exp_mentor_keys)[:3]}",
              critical=True)

        # structural: column presence (>=5 of 6 non-null in first row)
        if a_rows:
            non_null = len([v for v in a_rows[0] if v is not None])
            check("Pairs has 6 columns", non_null >= 5, f"Got {non_null}")

    # ---- Program_Summary sheet ----
    agent_summary = get_sheet(agent_wb, "Program_Summary")
    check("Sheet 'Program_Summary' exists", agent_summary is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_summary is None:
        all_ok = False
        check("CRITICAL Total_Pairs == 10 and 10 data rows", False, "no summary", critical=True)
        check("CRITICAL Avg_Mentor_Rating matches recomputed", False, "no summary", critical=True)
        check("CRITICAL Avg_Mentee_Experience matches recomputed", False, "no summary", critical=True)
    else:
        a_summary = {}
        for row in agent_summary.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        tp = a_summary.get("total_pairs")
        check("CRITICAL Total_Pairs == 10 and 10 data rows",
              num_close(tp, 10, 0) and n_rows == 10,
              f"Total_Pairs={tp}, rows={n_rows}", critical=True)

        amr = a_summary.get("avg_mentor_rating")
        check("CRITICAL Avg_Mentor_Rating matches recomputed",
              exp_avg_rating is not None and num_close(amr, exp_avg_rating, 0.01),
              f"Got {amr}, expected {exp_avg_rating}", critical=True)

        ame = a_summary.get("avg_mentee_experience")
        check("CRITICAL Avg_Mentee_Experience matches recomputed",
              exp_avg_exp is not None and num_close(ame, exp_avg_exp, 0.01),
              f"Got {ame}, expected {exp_avg_exp}", critical=True)

    return all_ok


def check_gcal(launch_time_str):
    print("\n=== Checking Google Calendar ===")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")
    check("At least 1 calendar event created", len(events) >= 1, f"Found {len(events)}")

    # Title may be the literal English 'Mentorship Kickoff Meeting' OR a Russian rendering.
    def is_kickoff(summ):
        s = (summ or "").lower()
        en = "mentorship" in s and "kickoff" in s
        ru = ("наставнич" in s or "наставник" in s) and ("установоч" in s or "старт" in s or "kickoff" in s)
        return en or ru

    kickoff_events = [e for e in events if is_kickoff(e[0])]
    check("Mentorship Kickoff Meeting event exists", len(kickoff_events) >= 1,
          f"Events: {[e[0] for e in events]}")

    if launch_time_str and kickoff_events:
        try:
            # launch_time may carry a trailing weekday (e.g. "2026-06-05 19:45:51 Friday");
            # keep only the "YYYY-MM-DD HH:MM:SS" prefix so fromisoformat doesn't throw.
            _parts = launch_time_str.split()
            _iso = " ".join(_parts[:2]) if len(_parts) >= 2 else launch_time_str
            launch_dt = datetime.fromisoformat(_iso)
            expected_dt = launch_dt + timedelta(days=7)
            ev_dt = kickoff_events[0][2]
            if ev_dt is not None:
                diff = abs((ev_dt.replace(tzinfo=None) - expected_dt).total_seconds())
                # Tighten to within ~1 hour of the exact +7 days mark.
                check("Kickoff meeting exactly 7 days from launch", diff <= 3600,
                      f"Expected {expected_dt}, got {ev_dt} (diff {diff:.0f}s)")
        except Exception as e:
            print(f"  [INFO] Could not verify date: {e}")

    return len(kickoff_events) >= 1


def check_email():
    print("\n=== Checking Email ===")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%mentorship program launch%'
        ORDER BY date DESC
    """)
    emails = cur.fetchall()
    cur.close()
    conn.close()

    subj_ok = len(emails) >= 1
    check("Email with subject 'Mentorship Program Launch' exists", subj_ok,
          f"Found {len(emails)} matching emails")

    from_ok = to_ok = body_ok = False
    if emails:
        e = emails[0]
        to_str = str(e[2] or "").lower()
        from_str = str(e[1] or "").lower()
        body_str = str(e[3] or "")
        to_ok = "program@hr.example.com" in to_str
        from_ok = "hr@company.example.com" in from_str
        # body must mention the total pair count (10) as a standalone number.
        body_ok = re.search(r"(?<!\d)10(?!\d)", body_str) is not None

        check("Email sent to program@hr.example.com", to_ok, f"to_addr: {e[2]}")
        check("Email from hr@company.example.com", from_ok, f"from_addr: {e[1]}")
        check("Email body mentions total pair count (10)", body_ok, f"body[:120]: {body_str[:120]}")

    # Single consolidated CRITICAL gate on the email deliverable.
    check("CRITICAL Email from hr@company.example.com to program@hr.example.com, "
          "subject 'Mentorship Program Launch', body mentions 10 pairs",
          subj_ok and from_ok and to_ok and body_ok,
          f"subj={subj_ok} from={from_ok} to={to_ok} body={body_ok}", critical=True)

    return subj_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    mentors, mentees = load_expected_from_db()
    print(f"  Recomputed {len(mentors)} mentors, {len(mentees)} mentees from live sf_data")

    check_excel(args.agent_workspace, mentors, mentees)
    check_gcal(args.launch_time)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
