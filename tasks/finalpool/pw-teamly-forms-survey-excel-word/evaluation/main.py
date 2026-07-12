"""Evaluation script for survey-excel-word (teamly + forms, RU)."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Критические чеки: семантика (значения с источника, правило Gap, сводные метрики,
# дашборд-страница в teamly, реальная форма). Любой fail => FAIL до проверки accuracy.
CRITICAL_CHECKS = {
    "Data_Analysis: эталон с дашборда совпадает (>=2 категории)",
    "Data_Analysis: Gap == Internal_Value - External_Benchmark для всех строк",
    "Metrics: Total_Items и Avg_Gap корректны",
    "Teamly: страница Dashboard с русским резюме создана",
    "Forms: форма Feedback с >=1 вопросом создана",
}

# Эталон с дашборда http://localhost:30330 (Typical_Duration_Min).
# Принимаем RU или EN названия категорий.
BENCHMARK = {
    90: ["стратегическ", "strategy review"],
    15: ["стендап", "ежедневн", "team standup", "standup"],
    30: ["1-на-1", "1-on-1", "личная встреча", "one-on-one"],
    60: ["общее собрание", "all-hands", "all hands"],
    120: ["обучение", "training"],
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _match_benchmark(item_str, ext_val):
    """Вернёт True, если название категории item_str соответствует одному из
    эталонных значений и его External_Benchmark == ожидаемой длительности."""
    s = (item_str or "").strip().lower()
    for dur, aliases in BENCHMARK.items():
        if any(a in s for a in aliases):
            return ext_val is not None and abs(ext_val - dur) < 0.5
    return False


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    excel_path = os.path.join(agent_workspace, "Survey_Report.xlsx")
    check("Survey_Report.xlsx exists", os.path.exists(excel_path))

    da_gaps = []          # значения Gap из листа Data_Analysis для проверки Metrics
    da_data_rows_count = 0

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            # отбрасываем полностью пустые строки
            data_rows = [r for r in data_rows if any(v is not None and str(v).strip() != "" for v in r)]
            da_data_rows_count = len(data_rows)
            check("Data_Analysis has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            for expected_col in ['Item', 'Internal_Value', 'External_Benchmark', 'Gap']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            def col_idx(name):
                try:
                    return headers.index(name.lower())
                except ValueError:
                    return None

            i_item = col_idx('Item')
            i_int = col_idx('Internal_Value')
            i_ext = col_idx('External_Benchmark')
            i_gap = col_idx('Gap')

            # CRITICAL: эталонные значения реально взяты с дашборда (RU или EN категории)
            matched = 0
            if i_item is not None and i_ext is not None:
                for r in data_rows:
                    item_v = r[i_item] if i_item < len(r) else None
                    ext_v = safe_float(r[i_ext]) if i_ext < len(r) else None
                    if _match_benchmark(str(item_v) if item_v is not None else "", ext_v):
                        matched += 1
            check("Data_Analysis: эталон с дашборда совпадает (>=2 категории)",
                  matched >= 2, f"matched={matched}")

            # CRITICAL: Gap == Internal_Value - External_Benchmark для всех строк
            gap_ok = True
            checked_any = False
            if None not in (i_int, i_ext, i_gap):
                for r in data_rows:
                    iv = safe_float(r[i_int]) if i_int < len(r) else None
                    ev = safe_float(r[i_ext]) if i_ext < len(r) else None
                    gv = safe_float(r[i_gap]) if i_gap < len(r) else None
                    if None in (iv, ev, gv):
                        gap_ok = False
                        continue
                    checked_any = True
                    da_gaps.append(gv)
                    if abs(gv - (iv - ev)) > 0.5:
                        gap_ok = False
            else:
                gap_ok = False
            check("Data_Analysis: Gap == Internal_Value - External_Benchmark для всех строк",
                  gap_ok and checked_any, f"gap_ok={gap_ok}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [r for r in data_rows if any(v is not None and str(v).strip() != "" for v in r)]
            check("Metrics has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # CRITICAL: Total_Items == число строк Data_Analysis и Avg_Gap == среднее Gap
            metric_map = {}
            for r in data_rows:
                if len(r) >= 2 and r[0] is not None:
                    metric_map[str(r[0]).strip().lower()] = r[1]
            total_items_v = None
            avg_gap_v = None
            for k, v in metric_map.items():
                if "total_items" in k or "total items" in k:
                    total_items_v = safe_float(v)
                if "avg_gap" in k or "avg gap" in k or "средн" in k:
                    avg_gap_v = safe_float(v)
            total_ok = (total_items_v is not None and da_data_rows_count > 0
                        and abs(total_items_v - da_data_rows_count) < 0.5)
            if da_gaps:
                expected_avg = sum(da_gaps) / len(da_gaps)
                avg_ok = (avg_gap_v is not None and abs(avg_gap_v - expected_avg) <= 0.1)
            else:
                avg_ok = False
            check("Metrics: Total_Items и Avg_Gap корректны",
                  total_ok and avg_ok,
                  f"total_items={total_items_v} (rows={da_data_rows_count}), avg_gap={avg_gap_v}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows = [r for r in data_rows if any(v is not None and str(v).strip() != "" for v in r)]
            check("Recommendations has >= 1 rows", len(data_rows) >= 1, f"got {len(data_rows)}")
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # --- Teamly: страница Dashboard (CRITICAL) ---
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT title, body FROM teamly.pages WHERE title ILIKE %s", ('%dashboard%',))
            pages = cur.fetchall()
            conn.close()
            # русское непустое резюме в теле
            def has_cyrillic(s):
                return any('а' <= ch.lower() <= 'я' or ch.lower() == 'ё' for ch in (s or ""))
            ok = any(len((b or "").strip()) > 20 and has_cyrillic(b) for _t, b in pages)
            check("Teamly: страница Dashboard с русским резюме создана",
                  len(pages) >= 1 and ok, f"found {len(pages)} pages")
        except Exception as e:
            check("Teamly: страница Dashboard с русским резюме создана", False, str(e))

        # --- Forms: форма Feedback с вопросом (CRITICAL) ---
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("""
                SELECT f.id, COUNT(q.id)
                FROM gform.forms f
                LEFT JOIN gform.questions q ON q.form_id = f.id
                WHERE f.title ILIKE %s
                GROUP BY f.id
            """, ('%feedback%',))
            rows = cur.fetchall()
            conn.close()
            ok = any(cnt >= 1 for _id, cnt in rows)
            check("Forms: форма Feedback с >=1 вопросом создана",
                  len(rows) >= 1 and ok, f"forms={len(rows)}")
        except Exception as e:
            check("Forms: форма Feedback с >=1 вопросом создана", False, str(e))

        # Check Word document
        import glob as globmod
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content", len(text) > 50, f"text length: {len(text)}")

        check("notion_survey_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "notion_survey_processor.py")))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = 100.0 * PASS_COUNT / total if total else 0.0
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if CRITICAL_FAILED:
        print(f"CRITICAL FAIL: {CRITICAL_FAILED}")
    success = (not CRITICAL_FAILED) and accuracy >= 70.0
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
