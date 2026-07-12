"""
Скрипт оценки задачи kulinar-scholarly-health-study.

Проверки:
1. Health_Diet_Analysis.xlsx с тремя листами (Recipe Nutrition, Research Summary, Combined Analysis)
2. Chinese_Cuisine_Health_Report.docx с обязательными разделами

CRITICAL_CHECKS (любой провал => немедленный FAIL до порога точности):
- блюда на листе Recipe Nutrition взяты из кулинарной базы kulinar (>=3 категорий),
- калории рассчитаны (ненулевые, не константа),
- Health_Rating ∈ {Low, Medium, High} и коррелирует с калорийной плотностью,
- Research Summary содержит внедрённые статьи о питании с верным Citation_Count,
- Combined Analysis связывает блюда с исследованиями из Research Summary,
- документ Word содержит все 4 раздела (RU+EN).
"""

import argparse
import json
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Внедрённые preprocess статьи о питании: title -> citation_count.
INJECTED_PAPERS = {
    "структура рациона питания и риск хронических заболеваний": 500,
    "традиционная русская кухня и сердечно-сосудистое здоровье": 800,
    "растительные рационы питания: обзор доказательной базы": 350,
    "калорийная плотность блюд и контроль массы тела": 420,
}
# Подстроки-маркеры названий статей (для нечёткого совпадения в Research Summary / Word).
PAPER_MARKERS = [
    "рацион", "хроническ", "русская кухня", "сердечно",
    "растительн", "калорийн", "питани",
]


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "[CRIT]" if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def norm(s):
    # ВАЖНО: НЕ применяем NFKD — он разлагает кириллические «й/ё» на базовую
    # букву + комбинирующий знак, ломая поиск русских подстрок (калорийн и т.п.).
    # Для русских ключевых слов используем обычный lower() оригинального текста.
    if s is None:
        return ""
    return str(s).strip().lower()


def load_kulinar_recipes():
    """Загрузить названия рецептов и ингредиентов из источника kulinar MCP."""
    here = os.path.abspath(__file__)
    # tasks/finalpool/<task>/evaluation/main.py -> подняться до cowork_gym
    root = here
    for _ in range(5):
        root = os.path.dirname(root)
    candidates = [
        os.path.join(root, "local_servers", "kulinar-mcp", "src", "data", "all_recipes.json"),
        os.path.join(os.path.dirname(root), "local_servers", "kulinar-mcp", "src", "data", "all_recipes.json"),
    ]
    for p in candidates:
        if os.path.isfile(p):
            with open(p, encoding="utf-8") as f:
                return json.load(f)
    return []


def load_reference_calories():
    """calorie_per_100g по имени ингредиента из initial_workspace/nutrition_reference.json."""
    here = os.path.abspath(__file__)
    task_root = os.path.dirname(os.path.dirname(here))
    ref = os.path.join(task_root, "initial_workspace", "nutrition_reference.json")
    out = {}
    if os.path.isfile(ref):
        with open(ref, encoding="utf-8") as f:
            data = json.load(f)
        for ing in data.get("ingredients", []):
            out[norm(ing.get("name"))] = ing.get("calories_per_100g", 0)
    return out


def get_injected_citation_counts():
    """Прочитать citation_count внедрённых статей из scholarly (если БД доступна)."""
    try:
        import psycopg2
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT title, citation_count FROM scholarly.scholar_papers")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        return {norm(t): c for (t, c) in rows}
    except Exception as e:
        print(f"  [info] scholarly DB unavailable, using fallback citations: {e}")
        return dict(INJECTED_PAPERS)


def find_sheet(wb, *keyword_groups):
    """Найти лист: сначала по совпадению всех ключевых слов группы, затем по любому."""
    for kws in keyword_groups:
        for name in wb.sheetnames:
            ln = name.lower()
            if all(k in ln for k in kws):
                return name
    flat = [k for grp in keyword_groups for k in grp]
    for name in wb.sheetnames:
        ln = name.lower()
        if any(k in ln for k in flat):
            return name
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")
    agent_file = os.path.join(agent_workspace, "Health_Diet_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        record("Recipe Nutrition: dishes from kulinar", False, "no xlsx", critical=True)
        record("Research Summary: injected diet papers", False, "no xlsx", critical=True)
        record("Combined Analysis links dishes to research", False, "no xlsx", critical=True)
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return

    kulinar = load_kulinar_recipes()
    kulinar_names = {norm(r.get("name")) for r in kulinar}
    ref_cal = load_reference_calories()

    # ---------- Лист 1: Recipe Nutrition ----------
    rn = find_sheet(wb, ("recipe", "nutrition"), ("recipe",), ("nutrition",))
    rn_dishes = []          # (name, calories, rating)
    if not rn:
        record("Sheet 'Recipe Nutrition' exists", False, f"Sheets: {wb.sheetnames}")
        record("Recipe Nutrition: dishes from kulinar (>=3 categories)", False,
               "no sheet", critical=True)
        record("Health_Rating correlates with calorie density", False,
               "no sheet", critical=True)
    else:
        record("Sheet 'Recipe Nutrition' exists", True)
        ws = wb[rn]
        rows = list(ws.iter_rows(values_only=True))
        header = [norm(c) for c in rows[0]] if rows else []
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)]

        def col(*names):
            for nm in names:
                if nm in header:
                    return header.index(nm)
            return None

        c_name = col("dish_name", "dish name", "name")
        c_cat = col("category")
        c_cal = col("estimated_calories", "estimated calories", "calories")
        c_rate = col("health_rating", "health rating", "rating")
        # Запасные позиции, если заголовки не распознаны.
        if c_name is None:
            c_name = 0
        if c_cat is None:
            c_cat = 1
        if c_cal is None:
            c_cal = 2
        if c_rate is None:
            c_rate = len(header) - 1 if header else 6

        record(f"Recipe Nutrition has >= 5 rows ({len(data_rows)} found)",
               len(data_rows) >= 5)

        cats = set()
        nonzero_cal = 0
        cal_values = []
        for r in data_rows:
            nm = r[c_name] if c_name < len(r) else None
            if nm:
                rn_dishes.append([norm(nm),
                                  r[c_cal] if c_cal < len(r) else None,
                                  r[c_rate] if c_rate < len(r) else None])
            if c_cat < len(r) and r[c_cat]:
                cats.add(norm(r[c_cat]))
            if c_cal < len(r):
                try:
                    cv = float(r[c_cal]) if r[c_cal] is not None else 0
                    if cv > 0:
                        nonzero_cal += 1
                        cal_values.append(cv)
                except (TypeError, ValueError):
                    pass

        record(f"At least 3 rows have non-zero calories ({nonzero_cal} found)",
               nonzero_cal >= 3)

        # CRITICAL: блюда из kulinar и >=3 категорий.
        matched = sum(1 for d in rn_dishes if d[0] in kulinar_names)
        # запасное частичное совпадение по подстроке
        if matched < 5:
            for d in rn_dishes:
                if d[0] in kulinar_names:
                    continue
                if any(d[0] and (d[0] in kn or kn in d[0]) for kn in kulinar_names):
                    matched += 1
        from_kulinar_ok = matched >= 5 and len(cats) >= 3
        record(
            f"Recipe Nutrition: >=5 dishes from kulinar across >=3 categories "
            f"(matched={matched}, categories={len(cats)})",
            from_kulinar_ok,
            f"dishes={[d[0] for d in rn_dishes]}",
            critical=True,
        )

        # CRITICAL: калории не константа и правдоподобны.
        not_constant = len(set(round(v) for v in cal_values)) >= 3
        record("Estimated_Calories are computed (not a constant placeholder)",
               not_constant and len(cal_values) >= 3,
               f"distinct calorie values: {sorted(set(round(v) for v in cal_values))}",
               critical=True)

        # CRITICAL: Health_Rating ∈ {Low,Medium,High} и коррелирует с калорийностью.
        valid_ratings = {"low", "medium", "high"}
        ratings = [(d[1], norm(d[2])) for d in rn_dishes if d[1] is not None]
        all_valid = bool(ratings) and all(rt in valid_ratings for (_, rt) in ratings)
        corr_ok = True
        num = []
        for (cal, rt) in ratings:
            if rt in valid_ratings:
                try:
                    num.append((float(cal), rt))
                except (TypeError, ValueError):
                    pass
        if len(num) >= 3:
            # Направление связи не задано в task.md ("в зависимости от калорийной
            # плотности") и в рамках исследователя здорового питания инверсия
            # (низкая калорийность = High health) столь же валидна. Поэтому
            # требуем монотонность в ЛЮБУЮ сторону, но отвергаем противоречивые
            # (немонотонные) разметки.
            rank = {"low": 0, "medium": 1, "high": 2}
            pairs = [(c, rank[rt]) for (c, rt) in num]  # (калории, ранг рейтинга)
            direct_ok = True   # больше калорий -> выше рейтинг
            inverse_ok = True  # больше калорий -> ниже рейтинг
            for i in range(len(pairs)):
                for j in range(len(pairs)):
                    ci, ri = pairs[i]
                    cj, rj = pairs[j]
                    if ci < cj:
                        if ri > rj:
                            direct_ok = False
                        if ri < rj:
                            inverse_ok = False
            corr_ok = direct_ok or inverse_ok
        else:
            corr_ok = False
        record("Health_Rating ∈ {Low,Medium,High} and correlates with calorie density",
               bool(all_valid and corr_ok),
               f"ratings={[(round(c), r) for c, r in num]}",
               critical=True)

    # ---------- Лист 2: Research Summary ----------
    rs = find_sheet(wb, ("research", "summary"), ("research",), ("summary",))
    rs_titles = []
    if not rs:
        record("Sheet 'Research Summary' exists", False, f"Sheets: {wb.sheetnames}")
        record("Research Summary: injected diet papers with correct citations",
               False, "no sheet", critical=True)
    else:
        record("Sheet 'Research Summary' exists", True)
        ws = wb[rs]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)]

        record(f"Research Summary has >= 3 rows ({len(data_rows)} found)",
               len(data_rows) >= 3)

        rows_with_title = sum(1 for r in data_rows if r and r[0] and str(r[0]).strip())
        record(f"Research Summary has >= 3 non-empty titles ({rows_with_title} found)",
               rows_with_title >= 3)
        rows_with_authors = sum(
            1 for r in data_rows if r and len(r) > 1 and r[1] and str(r[1]).strip())
        record(f"Research Summary has >= 3 non-empty authors ({rows_with_authors} found)",
               rows_with_authors >= 3)

        # CRITICAL: статьи соответствуют внедрённым статьям о питании + верный Citation_Count.
        injected_cit = get_injected_citation_counts()
        full_text_rows = []
        for r in data_rows:
            joined = norm(" ".join(str(c) for c in r if c is not None))
            rs_titles.append(joined)
            full_text_rows.append((joined, r))

        # сколько строк выглядят как внедрённые статьи о питании
        diet_like = sum(
            1 for (txt, _) in full_text_rows
            if sum(1 for m in PAPER_MARKERS if m in txt) >= 1
        )
        # хотя бы одна строка имеет верный citation_count для своей статьи
        cit_ok = False
        cit_values = set(injected_cit.values())
        for (txt, r) in full_text_rows:
            # к какой внедрённой статье относится строка
            for title_norm, cc in injected_cit.items():
                key_words = [w for w in title_norm.split() if len(w) > 4][:3]
                if key_words and all(w in txt for w in key_words):
                    # ищем cc где-то в строке
                    for cell in r:
                        try:
                            if cell is not None and abs(float(cell) - float(cc)) < 0.5:
                                cit_ok = True
                        except (TypeError, ValueError):
                            continue
            if cit_ok:
                break
        # запасной вариант: любое из ожидаемых значений цитирований присутствует
        if not cit_ok:
            for (_, r) in full_text_rows:
                for cell in r:
                    try:
                        if cell is not None and any(
                                abs(float(cell) - float(v)) < 0.5 for v in cit_values):
                            cit_ok = True
                            break
                    except (TypeError, ValueError):
                        continue
                if cit_ok:
                    break

        record(
            f"Research Summary: >=3 injected diet papers (diet_like={diet_like}) "
            f"with >=1 correct Citation_Count",
            diet_like >= 3 and cit_ok,
            f"diet_like={diet_like}, cit_ok={cit_ok}",
            critical=True,
        )

    # ---------- Лист 3: Combined Analysis ----------
    ca = find_sheet(wb, ("combined", "analysis"), ("combined",), ("analysis",))
    if not ca:
        record("Sheet 'Combined Analysis' exists", False, f"Sheets: {wb.sheetnames}")
        record("Combined Analysis links dishes to research", False,
               "no sheet", critical=True)
    else:
        record("Sheet 'Combined Analysis' exists", True)
        ws = wb[ca]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and any(c is not None for c in r)]
        record(f"Combined Analysis has >= 3 rows ({len(data_rows)} found)",
               len(data_rows) >= 3)

        # CRITICAL: каждая строка связывает блюдо (из листа 1) с исследованием (из листа 2).
        dish_names = {d[0] for d in rn_dishes}
        linked = 0
        for r in data_rows:
            if not r:
                continue
            dish = norm(r[0]) if len(r) > 0 else ""
            support = norm(" ".join(str(c) for c in r[1:] if c is not None))
            dish_ok = (dish in dish_names) or any(
                dish and (dish in dn or dn in dish) for dn in dish_names)
            # поддержка ссылается на исследование (по маркеру или совпадению с Research Summary)
            supp_ok = any(m in support for m in PAPER_MARKERS) or any(
                support and support[:15] in t for t in rs_titles)
            if dish_ok and supp_ok and support.strip():
                linked += 1
        record(
            f"Combined Analysis links >=3 dishes to research references (linked={linked})",
            linked >= 3,
            critical=True,
        )

    wb.close()


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    doc_file = os.path.join(agent_workspace, "Chinese_Cuisine_Health_Report.docx")
    if not os.path.isfile(doc_file):
        record("Word document exists", False, f"Not found: {doc_file}")
        record("Word doc has all 4 sections + literature reference", False,
               "no docx", critical=True)
        return
    record("Word document exists", True)

    try:
        from docx import Document
    except ImportError:
        file_size = os.path.getsize(doc_file)
        record("Word document has content (size > 1KB, python-docx not available)",
               file_size > 1000, f"File size: {file_size} bytes")
        record("Word doc has all 4 sections + literature reference", file_size > 1000,
               "python-docx unavailable; size fallback", critical=True)
        return

    try:
        doc = Document(doc_file)
        full_text = " ".join(p.text for p in doc.paragraphs)
        text_lower = full_text.lower()
    except Exception as e:
        record("Word document readable", False, str(e))
        record("Word doc has all 4 sections + literature reference", False,
               str(e), critical=True)
        return

    record(f"Document has >= 500 chars ({len(full_text)} found)", len(full_text) >= 500)

    has_title = any(k in text_lower for k in [
        "анализ пищевой ценности", "русской кухни", "русская кухня",
        "nutritional analysis", "chinese cuisine",
    ])
    record("Document contains title keywords", has_title)

    # Разделы: принимаем RU и EN варианты.
    section_alts = [
        ("Introduction", ["introduction", "введение"]),
        ("Recipe Analysis", ["recipe analysis", "анализ рецептов"]),
        ("Literature Review", ["literature review", "обзор литературы"]),
        ("Conclusions", ["conclusions", "выводы", "заключение"]),
    ]
    sections_ok = True
    for label, alts in section_alts:
        found = any(a in text_lower for a in alts)
        record(f"Document has '{label}' section", found)
        if not found:
            sections_ok = False

    # CRITICAL: все 4 раздела присутствуют И обзор литературы упоминает внедрённую статью.
    lit_ref = any(m in text_lower for m in PAPER_MARKERS) and any(
        a in text_lower for a in ["обзор литературы", "literature review"])
    record("Word doc has all 4 sections + literature references a diet paper",
           sections_ok and lit_ref,
           f"sections_ok={sections_ok}, lit_ref={lit_ref}",
           critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (threshold 70%)")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
