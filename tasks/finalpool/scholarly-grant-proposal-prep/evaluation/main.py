"""
Оценка для задачи scholarly-grant-proposal-prep.
Проверяет Excel (Grant_Prep.xlsx), отправленные письма и события календаря.

Критические проверки (CRITICAL_CHECKS): любой провал => общий FAIL независимо
от accuracy. Иначе PASS требует accuracy >= 70%.
Текст писем/заголовки могут быть на русском, поэтому грепы ключевых слов
расширены на RU+EN.
"""
import argparse
import datetime
import json
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

AGENCIES = ["NSF", "DARPA", "NIH"]
COLLABORATOR_EMAILS = ["jpark@mit.edu", "sthompson@stanford.edu", "mchen@mayo.edu", "lwang@berkeley.edu"]

# Источник истины из funding.json (агентство -> программа, дедлайн).
FUNDING = {
    "NSF": {"program": "AI Research Institutes", "deadline": "2026-06-15"},
    "DARPA": {"program": "Explainable AI", "deadline": "2026-05-01"},
    "NIH": {"program": "AI for Health", "deadline": "2026-07-01"},
}

# Ожидаемое число цитирований для каждой целевой статьи (из scholarly).
PAPER_CITATIONS = {
    "Safe Reinforcement Learning for Language Model Alignment": 310,
    "Interpretable Neural Networks via Concept Bottleneck Layers": 180,
    "Adversarially Robust Vision Transformers": 145,
    "Deep Learning for Medical Image Segmentation: A Comprehensive Review": 420,
    "Graph Neural Networks for Drug-Target Interaction Prediction": 250,
}

# Корректное сопоставление соавтор -> агентство гранта (по экспертизе).
COLLABORATOR_GRANT = {
    "jpark@mit.edu": "NSF",
    "sthompson@stanford.edu": "DARPA",
    "mchen@mayo.edu": "NIH",
    "lwang@berkeley.edu": "NIH",
}

# Критические проверки: любой провал => общий FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Funding Opportunities: агентства/программы/дедлайны верны (funding.json)",
    "Supporting Literature: >=5 статей с верным Citation_Count",
    "Collaborator Matrix: все 4 соавтора сопоставлены с корректным грантом",
    "Письма соавторам по существу (грант + дедлайн в теле)",
    "Календарь: дедлайны и сессии написания за ~2 недели до дедлайна",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def norm_title(s):
    return re.sub(r"[^a-z0-9]+", " ", str(s or "").lower()).strip()


def check_excel(workspace, groundtruth_workspace="."):
    print("\n=== Проверка Excel ===")
    path = os.path.join(workspace, "Grant_Prep.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        # Зависимые критические проверки помечаем как провал.
        record("Funding Opportunities: агентства/программы/дедлайны верны (funding.json)", False, "нет файла")
        record("Supporting Literature: >=5 статей с верным Citation_Count", False, "нет файла")
        record("Collaborator Matrix: все 4 соавтора сопоставлены с корректным грантом", False, "нет файла")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Funding Opportunities ---
    fo_rows = load_sheet_rows(wb, "Funding Opportunities") or load_sheet_rows(wb, "Funding_Opportunities")
    funding_ok = False
    if fo_rows is None:
        record("Sheet 'Funding Opportunities' exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Sheet 'Funding Opportunities' exists", True)
        data = [r for r in fo_rows[1:] if any(c is not None for c in r)]
        record("Funding Opportunities has 3 rows", len(data) == 3, f"Found {len(data)}")

        agency_col = find_col(fo_rows[0], ["Agency", "agency"])
        prog_col = find_col(fo_rows[0], ["Program", "program"])
        dl_col = find_col(fo_rows[0], ["Deadline", "deadline"])

        if agency_col is not None:
            agencies = {str(r[agency_col]).strip() for r in data if agency_col < len(r) and r[agency_col]}
            for a in AGENCIES:
                record(f"Agency '{a}' present", a in agencies, f"Found: {agencies}")

        score_col = find_col(fo_rows[0], ["Relevance_Score", "Relevance Score", "Score"])
        if score_col is not None:
            for r in data:
                if score_col < len(r) and r[score_col] is not None:
                    try:
                        s = float(r[score_col])
                        record(f"Score {s} in range 1-10", 1 <= s <= 10, f"Got {s}")
                    except (TypeError, ValueError):
                        record("Score is numeric", False, f"Got {r[score_col]}")

        # КРИТИЧЕСКАЯ: агентство -> программа -> дедлайн строго по funding.json.
        if agency_col is not None and prog_col is not None and dl_col is not None:
            mapping = {}
            for r in data:
                if agency_col >= len(r) or not r[agency_col]:
                    continue
                ag = str(r[agency_col]).strip().upper()
                prog = str(r[prog_col]).strip() if prog_col < len(r) and r[prog_col] else ""
                dl = str(r[dl_col]).strip() if dl_col < len(r) and r[dl_col] else ""
                mapping[ag] = (prog, dl)
            funding_ok = True
            for ag, exp in FUNDING.items():
                prog, dl = mapping.get(ag, ("", ""))
                prog_ok = exp["program"].lower() in prog.lower()
                dl_ok = exp["deadline"] in dl
                if not (prog_ok and dl_ok):
                    funding_ok = False
        record("Funding Opportunities: агентства/программы/дедлайны верны (funding.json)",
               funding_ok, f"mapping={mapping if 'mapping' in dir() else None}")

    # --- Supporting Literature ---
    sl_rows = load_sheet_rows(wb, "Supporting Literature") or load_sheet_rows(wb, "Supporting_Literature")
    lit_ok = False
    if sl_rows is None:
        record("Sheet 'Supporting Literature' exists", False, f"Sheets: {wb.sheetnames}")
        record("Supporting Literature: >=5 статей с верным Citation_Count", False, "нет листа")
    else:
        record("Sheet 'Supporting Literature' exists", True)
        data = [r for r in sl_rows[1:] if any(c is not None for c in r)]
        record("Supporting Literature has >= 5 rows", len(data) >= 5, f"Found {len(data)}")

        title_col = find_col(sl_rows[0], ["Paper_Title", "Paper Title", "Title"])
        cite_col = find_col(sl_rows[0], ["Citation_Count", "Citation Count", "Citations"])
        if title_col is not None and cite_col is not None:
            by_title = {}
            for r in data:
                if title_col < len(r) and r[title_col]:
                    by_title[norm_title(r[title_col])] = r[cite_col] if cite_col < len(r) else None
            matched = 0
            for t, exp in PAPER_CITATIONS.items():
                nt = norm_title(t)
                # Сопоставляем по префиксу/подстроке: агент мог записать как полное
                # название статьи, так и его усечённый вариант (без подзаголовка).
                got = by_title.get(nt)
                if got is None:
                    for rt, rv in by_title.items():
                        if rt and (rt in nt or nt in rt):
                            got = rv
                            break
                if got is not None and num_close(got, exp, max(abs(exp) * 0.1, 1.0)):
                    matched += 1
            lit_ok = matched >= 5
            record("Supporting Literature: >=5 статей с верным Citation_Count",
                   lit_ok, f"matched={matched}/5")
        else:
            record("Supporting Literature: >=5 статей с верным Citation_Count",
                   False, "нет столбцов Paper_Title/Citation_Count")

    # --- Collaborator Matrix ---
    cm_rows = load_sheet_rows(wb, "Collaborator Matrix") or load_sheet_rows(wb, "Collaborator_Matrix")
    collab_ok = False
    if cm_rows is None:
        record("Sheet 'Collaborator Matrix' exists", False, f"Sheets: {wb.sheetnames}")
        record("Collaborator Matrix: все 4 соавтора сопоставлены с корректным грантом", False, "нет листа")
    else:
        record("Sheet 'Collaborator Matrix' exists", True)
        data = [r for r in cm_rows[1:] if any(c is not None for c in r)]
        record("Collaborator Matrix has 4 rows", len(data) == 4, f"Found {len(data)}")

        email_col = find_col(cm_rows[0], ["Email", "email"])
        grant_col = find_col(cm_rows[0], ["Recommended_Grant", "Recommended Grant", "Grant"])
        if email_col is not None and grant_col is not None:
            by_email = {}
            for r in data:
                if email_col < len(r) and r[email_col]:
                    by_email[str(r[email_col]).strip().lower()] = (
                        str(r[grant_col]).strip() if grant_col < len(r) and r[grant_col] else "")
            collab_ok = True
            for em, exp_agency in COLLABORATOR_GRANT.items():
                rec = by_email.get(em, "")
                exp_prog = FUNDING[exp_agency]["program"]
                if not (exp_agency.lower() in rec.lower() or exp_prog.lower() in rec.lower()):
                    collab_ok = False
            record("Collaborator Matrix: все 4 соавтора сопоставлены с корректным грантом",
                   collab_ok, f"mapping={by_email}")
        else:
            record("Collaborator Matrix: все 4 соавтора сопоставлены с корректным грантом",
                   False, "нет столбцов Email/Recommended_Grant")

    return True


def check_emails():
    print("\n=== Проверка писем ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Тема может содержать английские ключевые слова (агентство/программа) или
    # русские (грант/заявк/финансир/сотрудничеств).
    cur.execute("""
        SELECT id, subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%%grant%%' OR subject ILIKE '%%collaboration%%'
           OR subject ILIKE '%%nsf%%' OR subject ILIKE '%%darpa%%' OR subject ILIKE '%%nih%%'
           OR subject ILIKE '%%proposal%%' OR subject ILIKE '%%funding%%'
           OR subject ILIKE '%%грант%%' OR subject ILIKE '%%заявк%%'
           OR subject ILIKE '%%финансир%%' OR subject ILIKE '%%сотрудничеств%%'
    """)
    emails = cur.fetchall()

    record("Grant-related emails sent", len(emails) >= 3, f"Found {len(emails)} emails")

    # Для каждого соавтора найдём адресованное ему письмо и проверим, что в теле
    # упомянут корректный грант (агентство/программа) И конкретная дата дедлайна.
    def to_list(to):
        if isinstance(to, str):
            try:
                to = json.loads(to)
            except Exception:
                pass
        return str(to).lower()

    substance_ok = 0
    for em, agency in COLLABORATOR_GRANT.items():
        prog = FUNDING[agency]["program"]
        deadline = FUNDING[agency]["deadline"]
        # Возможные форматы даты дедлайна.
        dt = datetime.date.fromisoformat(deadline)
        date_variants = {
            deadline,
            dt.strftime("%d.%m.%Y"),
            dt.strftime("%d.%m.%y"),
            f"{dt.day}.{dt.month:02d}.{dt.year}",
            dt.strftime("%m/%d/%Y"),
            f"{dt.day} ",  # день месяца как отдельное число (для русских дат)
        }
        found = False
        for e in emails:
            if em not in to_list(e[3]):
                continue
            body = (str(e[4] or "") + " " + str(e[1] or "")).lower()
            grant_ok = (agency.lower() in body) or (prog.lower() in body)
            date_ok = any(dv.lower() in body for dv in date_variants)
            if grant_ok and date_ok:
                found = True
                break
        if found:
            substance_ok += 1

    record("Письма соавторам по существу (грант + дедлайн в теле)",
           substance_ok >= 3, f"{substance_ok}/4 писем по существу")

    cur.close()
    conn.close()
    return True


def check_calendar():
    print("\n=== Проверка календаря ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, summary, start_datetime FROM gcal.events")
    events = cur.fetchall()

    record("Calendar events created", len(events) >= 3, f"Found {len(events)} events")

    summaries = [str(e[1]).lower() for e in events if e[1]]
    has_deadline = any(
        "deadline" in s or "nsf" in s or "darpa" in s or "nih" in s
        or "дедлайн" in s or "срок" in s for s in summaries)
    record("Deadline events present", has_deadline, f"Summaries: {summaries[:6]}")

    has_writing = any(
        "writing" in s or "proposal" in s or "написани" in s or "заявк" in s
        for s in summaries)
    record("Writing session events present", has_writing, f"Summaries: {summaries[:6]}")

    # КРИТИЧЕСКАЯ: для каждого агентства найти событие-дедлайн в дату дедлайна
    # и событие-сессию написания ровно за ~2 недели (14 дней, допуск +/-1) до него.
    def parse_date(v):
        if v is None:
            return None
        if isinstance(v, (datetime.datetime, datetime.date)):
            return v.date() if isinstance(v, datetime.datetime) else v
        s = str(v)
        m = re.search(r"(\d{4})-(\d{2})-(\d{2})", s)
        if m:
            try:
                return datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            except ValueError:
                return None
        return None

    ev = [(str(e[1] or "").lower(), parse_date(e[2])) for e in events]
    pairs_ok = 0
    for agency, info in FUNDING.items():
        deadline = datetime.date.fromisoformat(info["deadline"])
        prog = info["program"].lower()
        key = agency.lower()
        # событие-дедлайн в дату дедлайна
        deadline_ev = any(
            d == deadline and (key in s or prog in s)
            for s, d in ev)
        # сессия написания за ~2 недели до дедлайна
        writing_ev = any(
            d is not None and abs((deadline - d).days - 14) <= 1
            and (key in s or prog in s or "написани" in s or "writing" in s
                 or "proposal" in s or "заявк" in s)
            for s, d in ev)
        if deadline_ev and writing_ev:
            pairs_ok += 1

    record("Календарь: дедлайны и сессии написания за ~2 недели до дедлайна",
           pairs_ok >= 3, f"{pairs_ok}/3 пар (дедлайн + сессия) корректны")

    cur.close()
    conn.close()
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_emails()
    check_calendar()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT,
                           "accuracy": accuracy, "critical_failed": critical_failed},
                          f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
