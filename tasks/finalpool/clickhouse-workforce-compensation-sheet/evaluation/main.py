"""Evaluation for sf-workforce-compensation-sheet (ClickHouse / RU).

Источник данных russified централизованно (db/zzz_clickhouse_after_init.sql):
названия подразделений в HR_ANALYTICS на русском (Инженерия, Финансы, Кадры,
Операции, НИОКР, Продажи, Поддержка). Groundtruth Compensation_Report.xlsx
сгенерирован с этими русскими названиями. Числа берутся из реальной агрегации.

CRITICAL_CHECKS: любой провал => общий FAIL независимо от accuracy.
Порог иначе: accuracy >= 70%.
"""
import argparse
import json
import os
import sys
import openpyxl

# Семь русифицированных подразделений (детерминированы центральной картой).
RU_DEPARTMENTS = ["Инженерия", "Финансы", "Кадры", "Операции", "НИОКР", "Продажи", "Поддержка"]

# Критические проверки: провал любой => общий FAIL.
CRITICAL_CHECKS = {
    "dept_all_present",          # все 7 русских подразделений с верным Headcount
    "dept_avg_salary",           # Avg_Salary совпадает с эталоном (реальная агрегация)
    "summary_total_employees",   # Total_Employees == 50000
    "summary_total_payroll",     # Total_Payroll совпадает с эталоном
    "summary_high_low",          # Highest=Инженерия, Lowest=Операции
    "gsheet_departments",        # Google Sheet содержит русские подразделения как значения
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def record(name, ok, msg=""):
    global PASS_COUNT, FAIL_COUNT
    if ok:
        PASS_COUNT += 1
        print(f"  PASS: {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  FAIL: {name} {msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_department_sheet(agent_wb, gt_wb):
    a_rows = load_sheet_rows(agent_wb, "Department Compensation")
    g_rows = load_sheet_rows(gt_wb, "Department Compensation")
    if a_rows is None:
        record("dept_all_present", False, "Sheet 'Department Compensation' not found in agent output")
        record("dept_avg_salary", False, "no sheet")
        return
    if g_rows is None:
        record("dept_all_present", False, "Sheet 'Department Compensation' not found in groundtruth")
        record("dept_avg_salary", False, "no gt sheet")
        return

    a_data = a_rows[1:] if len(a_rows) > 1 else []
    g_data = g_rows[1:] if len(g_rows) > 1 else []

    a_lookup = {}
    for row in a_data:
        if row and row[0] is not None:
            a_lookup[str(row[0]).strip().lower()] = row
    g_lookup = {}
    for row in g_data:
        if row and row[0] is not None:
            g_lookup[str(row[0]).strip().lower()] = row

    # --- CRITICAL: все 7 русских подразделений присутствуют + верный Headcount ---
    missing = []
    headcount_bad = []
    for dep in RU_DEPARTMENTS:
        key = dep.strip().lower()
        a_row = a_lookup.get(key)
        g_row = g_lookup.get(key)
        if a_row is None:
            missing.append(dep)
            continue
        if g_row is not None and len(a_row) > 1 and len(g_row) > 1:
            if not num_close(a_row[1], g_row[1], 0):
                headcount_bad.append(f"{dep}: {a_row[1]} vs {g_row[1]}")
    record("dept_all_present", not missing and not headcount_bad,
           f"missing={missing} headcount={headcount_bad}")

    # --- CRITICAL: Avg_Salary совпадает с эталоном (жёсткий допуск +-5.0) ---
    # подтверждает реальную агрегацию SALARY из HR_ANALYTICS, а не выдуманные числа.
    avg_bad = []
    for dep in RU_DEPARTMENTS:
        key = dep.strip().lower()
        a_row = a_lookup.get(key)
        g_row = g_lookup.get(key)
        if a_row is None or g_row is None:
            continue
        if len(a_row) > 2 and len(g_row) > 2:
            if not num_close(a_row[2], g_row[2], 5.0):
                avg_bad.append(f"{dep}: {a_row[2]} vs {g_row[2]}")
    record("dept_avg_salary", not avg_bad, f"avg mismatch: {avg_bad}")

    # --- НЕкритические: Min/Max/Total по подразделениям (умеренные допуски) ---
    min_bad, max_bad, total_bad = [], [], []
    for dep in RU_DEPARTMENTS:
        key = dep.strip().lower()
        a_row = a_lookup.get(key)
        g_row = g_lookup.get(key)
        if a_row is None or g_row is None:
            continue
        if len(a_row) > 3 and len(g_row) > 3 and not num_close(a_row[3], g_row[3], 500.0):
            min_bad.append(f"{dep}: {a_row[3]} vs {g_row[3]}")
        if len(a_row) > 4 and len(g_row) > 4 and not num_close(a_row[4], g_row[4], 5000.0):
            max_bad.append(f"{dep}: {a_row[4]} vs {g_row[4]}")
        if len(a_row) > 5 and len(g_row) > 5 and not num_close(a_row[5], g_row[5], 100000.0):
            total_bad.append(f"{dep}: {a_row[5]} vs {g_row[5]}")
    record("dept_min_salary", not min_bad, f"min mismatch: {min_bad}")
    record("dept_max_salary", not max_bad, f"max mismatch: {max_bad}")
    record("dept_total_payroll", not total_bad, f"total mismatch: {total_bad}")


def check_summary_sheet(agent_wb, gt_wb):
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        record("summary_total_employees", False, "Sheet 'Summary' not found in agent output")
        record("summary_total_payroll", False, "no sheet")
        record("summary_high_low", False, "no sheet")
        return
    if g_rows is None:
        record("summary_total_employees", False, "no gt sheet")
        record("summary_total_payroll", False, "no gt sheet")
        record("summary_high_low", False, "no gt sheet")
        return

    a_data = a_rows[1:] if len(a_rows) > 1 else []
    a_lookup = {}
    for row in a_data:
        if row and row[0] is not None:
            a_lookup[str(row[0]).strip().lower()] = row

    def aval(metric):
        r = a_lookup.get(metric.strip().lower())
        return r[1] if r and len(r) > 1 else None

    # --- CRITICAL: Total_Employees == 50000 ---
    record("summary_total_employees", num_close(aval("Total_Employees"), 50000, 0),
           f"got {aval('Total_Employees')}")

    # --- CRITICAL: Total_Payroll совпадает с эталоном (жёсткий допуск +-1000) ---
    record("summary_total_payroll", num_close(aval("Total_Payroll"), 2919806955, 1000.0),
           f"got {aval('Total_Payroll')}")

    # --- CRITICAL: Highest=Инженерия, Lowest=Операции (русские, выведены из данных) ---
    high_ok = str_match(aval("Highest_Avg_Department"), "Инженерия")
    low_ok = str_match(aval("Lowest_Avg_Department"), "Операции")
    record("summary_high_low", high_ok and low_ok,
           f"high={aval('Highest_Avg_Department')} low={aval('Lowest_Avg_Department')}")

    # --- НЕкритическая: Company_Avg_Salary ---
    record("summary_company_avg", num_close(aval("Company_Avg_Salary"), 58396.14, 5.0),
           f"got {aval('Company_Avg_Salary')}")


def check_google_sheet():
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432,
                                dbname="cowork_gym", user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gsheet.spreadsheets "
                    "WHERE LOWER(title) LIKE '%compensation%' OR LOWER(title) LIKE '%workforce%'")
        sheets = cur.fetchall()
        if not sheets:
            record("gsheet_title", False, "No Google Sheet with 'Compensation'/'Workforce' in title")
            record("gsheet_departments", False, "no spreadsheet")
            cur.close(); conn.close()
            return
        record("gsheet_title", True)

        sid = sheets[0][0]
        cur.execute("SELECT value FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
        cell_values = [str(r[0]).strip() for r in cur.fetchall() if r[0] is not None]
        cur.close(); conn.close()

        # НЕкритическая: достаточно ячеек.
        record("gsheet_cell_count", len(cell_values) >= 20, f"cells={len(cell_values)}")

        # CRITICAL: реальные русские названия подразделений присутствуют как значения ячеек.
        present = [d for d in RU_DEPARTMENTS if any(d.strip().lower() == v.strip().lower() for v in cell_values)]
        record("gsheet_departments", len(present) == len(RU_DEPARTMENTS),
               f"found {len(present)}/{len(RU_DEPARTMENTS)}: {present}")
    except Exception as e:
        record("gsheet_title", False, f"error: {e}")
        record("gsheet_departments", False, f"error: {e}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Compensation_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Compensation_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    print("  Checking Department Compensation sheet...")
    check_department_sheet(agent_wb, gt_wb)
    print("  Checking Summary sheet...")
    check_summary_sheet(agent_wb, gt_wb)
    print("  Checking Google Sheet...")
    check_google_sheet()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("\n=== RESULT: FAIL (critical check failed) ===")
        sys.exit(1)
    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
