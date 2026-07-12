"""Evaluation script for pw-sf-support-escalation-excel-email-gcal."""
import os
import argparse, json, os, sys
from datetime import datetime
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Benchmark values published on the mock dashboard (http://localhost:30332).
# Priority keys stay English: they are dimension keys joined against the
# English benchmark page (eval + agent depend on them).
BENCHMARK_INDUSTRY_AVG = {"High": 5.0, "Medium": 9.8, "Low": 20.6}

# Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
# They reflect the SUBSTANCE of the deliverable: real internal-vs-benchmark
# comparison, internally-consistent Metrics, correct email routing/subject,
# correct calendar event, and the two named artifacts.
CRITICAL_CHECKS = {
    "Data_Analysis: Industry_Avg matches benchmark for all priorities",
    "Data_Analysis: Gap == Our_Avg_Response - Industry_Avg for all priorities",
    "Metrics: Total_Tickets equals sum of Ticket_Count",
    "Email: 'Analysis Report Complete' to team-lead@company.com",
    "Calendar: 'Analysis Review' on 2026-03-14 14:00-15:00 UTC",
    "Artifacts: sf_escalation_processor.py and sf_escalation_results.json exist",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def parse_dt(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip().replace("Z", "+00:00").replace(" ", "T", 1) if " " in str(val) else str(val).strip().replace("Z", "+00:00")
    try:
        return datetime.fromisoformat(s)
    except Exception:
        for fmt in ("%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d %H:%M:%S%z", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(str(val).strip(), fmt)
            except Exception:
                continue
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Support_Escalation_Report.xlsx")
    check("Support_Escalation_Report.xlsx exists", os.path.exists(excel_path))

    da_rows = []  # list of dicts keyed by header.lower()
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---- Data_Analysis ----
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 3 rows", len(data_rows) >= 3, f"got {len(data_rows)}")

            for expected_col in ['Priority', 'Ticket_Count', 'Our_Avg_Response', 'Industry_Avg', 'Gap', 'CSAT']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            for r in data_rows:
                if not r or all(v is None for v in r):
                    continue
                da_rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})

            # CRITICAL: Industry_Avg matches the published benchmark per priority.
            ind_ok = True
            ind_detail = []
            for row in da_rows:
                prio = str(row.get('priority', '')).strip()
                bench = BENCHMARK_INDUSTRY_AVG.get(prio)
                if bench is None:
                    continue
                ind = safe_float(row.get('industry_avg'))
                if ind is None or abs(ind - bench) > 0.11:
                    ind_ok = False
                    ind_detail.append(f"{prio}: got {ind}, expected {bench}")
            covered = {str(row.get('priority', '')).strip() for row in da_rows}
            if not set(BENCHMARK_INDUSTRY_AVG).issubset(covered):
                ind_ok = False
                ind_detail.append(f"missing priorities: {set(BENCHMARK_INDUSTRY_AVG) - covered}")
            check("Data_Analysis: Industry_Avg matches benchmark for all priorities",
                  ind_ok, "; ".join(ind_detail))

            # CRITICAL: Gap == Our_Avg_Response - Industry_Avg (per analysis guide).
            gap_ok = True
            gap_detail = []
            for row in da_rows:
                prio = str(row.get('priority', '')).strip()
                if prio not in BENCHMARK_INDUSTRY_AVG:
                    continue
                our = safe_float(row.get('our_avg_response'))
                ind = safe_float(row.get('industry_avg'))
                gap = safe_float(row.get('gap'))
                if our is None or ind is None or gap is None:
                    gap_ok = False
                    gap_detail.append(f"{prio}: missing value")
                    continue
                if abs(gap - (our - ind)) > 0.05:
                    gap_ok = False
                    gap_detail.append(f"{prio}: gap={gap}, our-ind={round(our - ind, 2)}")
            check("Data_Analysis: Gap == Our_Avg_Response - Industry_Avg for all priorities",
                  gap_ok and len(da_rows) >= 3, "; ".join(gap_detail))

        # ---- Metrics ----
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            m_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 3 rows", len(m_rows) >= 3, f"got {len(m_rows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in m_rows:
                if r and r[0] is not None:
                    metrics_map[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

        # CRITICAL: Total_Tickets equals the sum of Ticket_Count across Data_Analysis.
        sum_tickets = sum(safe_float(row.get('ticket_count'), 0) or 0 for row in da_rows)
        total_val = None
        for k, v in metrics_map.items():
            if 'total' in k and 'ticket' in k:
                total_val = safe_float(v)
                break
        check("Metrics: Total_Tickets equals sum of Ticket_Count",
              total_val is not None and sum_tickets > 0 and abs(total_val - sum_tickets) <= 1.0,
              f"Total_Tickets={total_val}, sum(Ticket_Count)={sum_tickets}")

        # ---- Recommendations ----
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            # Priority/Action headers; agent may legitimately write Russian Action free text.
            for expected_col in ['Priority', 'Action']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

    # ---- Artifacts (CRITICAL) ----
    proc_exists = os.path.exists(os.path.join(agent_workspace, "sf_escalation_processor.py"))
    json_exists = os.path.exists(os.path.join(agent_workspace, "sf_escalation_results.json"))
    check("sf_escalation_processor.py exists", proc_exists)
    check("sf_escalation_results.json exists", json_exists)
    check("Artifacts: sf_escalation_processor.py and sf_escalation_results.json exist",
          proc_exists and json_exists)

    # ---- Email (CRITICAL: exact subject + recipient; body may be Russian) ----
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""SELECT subject, to_addr FROM email.messages
            WHERE subject = %s AND to_addr::text ILIKE %s""",
            ('Analysis Report Complete', '%team-lead@company.com%'))
        rows = cur.fetchall()
        check("Email: 'Analysis Report Complete' to team-lead@company.com",
              len(rows) >= 1, f"found {len(rows)} matching")
        conn.close()
    except Exception as e:
        check("Email: 'Analysis Report Complete' to team-lead@company.com", False, str(e))

    # ---- Calendar (CRITICAL: exact summary + date + 14:00-15:00 UTC) ----
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""SELECT summary, start_datetime, end_datetime FROM gcal.events
            WHERE summary = %s""", ('Analysis Review',))
        events = cur.fetchall()
        cal_ok = False
        cal_detail = f"found {len(events)} 'Analysis Review' events"
        for summ, sdt, edt in events:
            sd = parse_dt(sdt)
            ed = parse_dt(edt)
            if not sd or not ed:
                continue
            # Compare wall-clock as UTC (storage is naive UTC in gcal.events).
            sd_n = sd.replace(tzinfo=None)
            ed_n = ed.replace(tzinfo=None)
            if (sd_n.year, sd_n.month, sd_n.day) == (2026, 3, 14) and \
               sd_n.hour == 14 and sd_n.minute == 0 and \
               ed_n.hour == 15 and ed_n.minute == 0:
                cal_ok = True
                break
        check("Calendar: 'Analysis Review' on 2026-03-14 14:00-15:00 UTC", cal_ok, cal_detail)
        conn.close()
    except Exception as e:
        check("Calendar: 'Analysis Review' on 2026-03-14 14:00-15:00 UTC", False, str(e))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump(result, f, indent=2)
        except Exception:
            pass

    if critical_failed:
        return False, f"FAIL (critical): {len(critical_failed)} critical checks failed; {PASS_COUNT}/{total}"
    if accuracy >= 70:
        return True, f"PASS: {PASS_COUNT}/{total} ({accuracy:.1f}%)"
    return False, f"FAIL: {PASS_COUNT}/{total} ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
