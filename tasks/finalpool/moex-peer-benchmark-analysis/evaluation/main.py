"""Evaluation for yf-peer-benchmark-analysis (MOEX-russified).

Expected values are recomputed at eval time from the LIVE moex-finance source
(the same data the agent's MCP serves) combined with the static RU consensus
portal (files/mock_pages/index.html):
  - Current price  = latest moex.stock_prices close per symbol
        (matches pg_adapter PgTicker.info: ORDER BY date DESC LIMIT 1)
  - Trailing PE    = moex.stock_info.data->>'trailingPE'
  - Target price / sector PE benchmark = portal constants below
This avoids the stale-seed mismatch that previously hard-coded prices
(SBER 133.3 ...) against a seed whose latest close had diverged (SBER 120.8171 ...).
"""
import os
import argparse
import sys
import psycopg2

EXP_SYMBOLS = ["GAZP.ME", "LKOH.ME", "MGNT.ME", "MTSS.ME", "SBER.ME"]

# ---- static portal consensus data (files/mock_pages/index.html) ----
PORTAL_TARGET = {
    "SBER.ME": 160.00, "GAZP.ME": 240.00, "LKOH.ME": 3500.00,
    "MGNT.ME": 6200.00, "MTSS.ME": 250.00,
}
PORTAL_PE_BENCHMARK = {
    "SBER.ME": 5.00, "GAZP.ME": 4.00, "LKOH.ME": 3.00,
    "MGNT.ME": 7.50, "MTSS.ME": 10.00,
}


def _moex_conn():
    return psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432,
                            dbname="cowork_gym", user="eigent", password="camel")


def compute_expected():
    """Recompute expected metrics from live moex source + portal constants.

    Mirrors pg_adapter: currentPrice = latest stock_prices close per symbol;
    trailingPE from stock_info JSONB.
    """
    conn = _moex_conn()
    cur = conn.cursor()
    price, pe = {}, {}
    for sym in EXP_SYMBOLS:
        cur.execute(
            "SELECT close FROM moex.stock_prices WHERE symbol=%s ORDER BY date DESC LIMIT 1",
            (sym,))
        r = cur.fetchone()
        price[sym] = float(r[0])
        cur.execute(
            "SELECT (data->>'trailingPE')::float FROM moex.stock_info WHERE symbol=%s",
            (sym,))
        pe[sym] = float(cur.fetchone()[0])
    cur.close()
    conn.close()

    upside = {s: round((PORTAL_TARGET[s] - price[s]) / price[s] * 100, 2)
              for s in EXP_SYMBOLS}
    pe_vs = {s: ("Premium" if pe[s] > PORTAL_PE_BENCHMARK[s] else "Discount")
             for s in EXP_SYMBOLS}
    avg_upside = round(sum(upside.values()) / len(EXP_SYMBOLS), 2)
    stocks_above = sum(1 for s in EXP_SYMBOLS if price[s] > PORTAL_TARGET[s])
    stocks_below = sum(1 for s in EXP_SYMBOLS if price[s] < PORTAL_TARGET[s])
    avg_pe_premium = round(
        sum(pe[s] - PORTAL_PE_BENCHMARK[s] for s in EXP_SYMBOLS) / len(EXP_SYMBOLS), 2)
    most_undervalued = max(EXP_SYMBOLS, key=lambda s: upside[s])
    return dict(upside=upside, pe_vs=pe_vs, avg_upside=avg_upside,
                stocks_above=stocks_above, stocks_below=stocks_below,
                avg_pe_premium=avg_pe_premium,
                most_undervalued=norm_sym(most_undervalued))


# Populated in main() from compute_expected().
EXP_UPSIDE = {}
EXP_PE_VS = {}
EXP_AVG_UPSIDE = None
EXP_STOCKS_ABOVE = None
EXP_STOCKS_BELOW = None
EXP_AVG_PE_PREMIUM = None
EXP_MOST_UNDERVALUED = None


def norm_sym(s):
    """Normalize a symbol for comparison; tolerate missing/extra .ME suffix and case."""
    s = str(s).strip().upper()
    if s.endswith(".ME"):
        s = s[:-3]
    return s


def num_close(a, b, abs_tol=0.5, rel_tol=0.02):
    try:
        a_f, b_f = float(a), float(b)
        return abs(a_f - b_f) <= max(abs_tol, abs(b_f) * rel_tol)
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def _consensus_lookup(rows):
    """Build {NORMSYM: {col: value}} from Consensus Analysis rows, using the header."""
    header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    idx = {h: i for i, h in enumerate(header)}
    out = {}
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        sym = norm_sym(r[0])
        rec = {}
        for col in ("upside_pct", "pe_vs_benchmark", "consensus",
                    "current_price", "target_price", "trailing_pe", "pe_benchmark"):
            if col in idx and idx[col] < len(r):
                rec[col] = r[idx[col]]
        out[sym] = rec
    return out


def check_excel(agent_workspace):
    """Returns (errors, critical_errors)."""
    errors, critical = [], []
    import openpyxl
    path = os.path.join(agent_workspace, "Peer_Benchmark.xlsx")
    if not os.path.exists(path):
        return (["Peer_Benchmark.xlsx not found"], ["Peer_Benchmark.xlsx not found"])
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        rows = load_sheet_rows(wb, "Consensus Analysis")
        cons = {}
        if rows is None:
            errors.append("Sheet 'Consensus Analysis' not found")
            critical.append("Sheet 'Consensus Analysis' not found")
        else:
            cons = _consensus_lookup(rows)
            # structural: all 5 symbols present
            for sym in EXP_SYMBOLS:
                if norm_sym(sym) not in cons:
                    errors.append(f"Symbol {sym} missing from Consensus Analysis")

            # CRITICAL: every row's Upside_Pct correct
            for sym in EXP_SYMBOLS:
                k = norm_sym(sym)
                if k not in cons:
                    continue
                got = cons[k].get("upside_pct")
                if not num_close(got, EXP_UPSIDE[sym], abs_tol=0.5):
                    critical.append(
                        f"{sym} Upside_Pct={got}, expected ~{EXP_UPSIDE[sym]}")

            # CRITICAL: PE_vs_Benchmark (Premium/Discount) rule on every row
            for sym in EXP_SYMBOLS:
                k = norm_sym(sym)
                if k not in cons:
                    continue
                got = str(cons[k].get("pe_vs_benchmark", "")).strip().lower()
                exp = EXP_PE_VS[sym].lower()
                if got != exp:
                    critical.append(
                        f"{sym} PE_vs_Benchmark={cons[k].get('pe_vs_benchmark')}, expected {EXP_PE_VS[sym]}")

        rows2 = load_sheet_rows(wb, "Summary")
        if rows2 is None:
            errors.append("Sheet 'Summary' not found")
            critical.append("Sheet 'Summary' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            lookup = {str(r[0]).strip().lower(): r[1] for r in data_rows2 if r[0]}

            # CRITICAL: Avg_Upside_Pct
            if "avg_upside_pct" in lookup:
                if not num_close(lookup["avg_upside_pct"], EXP_AVG_UPSIDE, abs_tol=0.5):
                    critical.append(
                        f"Avg_Upside_Pct={lookup['avg_upside_pct']}, expected ~{EXP_AVG_UPSIDE}")
            else:
                errors.append("Avg_Upside_Pct not found")
                critical.append("Avg_Upside_Pct not found")

            # CRITICAL: Stocks_Above_Target exact
            if "stocks_above_target" in lookup:
                if not num_close(lookup["stocks_above_target"], EXP_STOCKS_ABOVE, abs_tol=0):
                    critical.append(
                        f"Stocks_Above_Target={lookup['stocks_above_target']}, expected {EXP_STOCKS_ABOVE}")
            else:
                errors.append("Stocks_Above_Target not found")

            # non-critical: Stocks_Below_Target
            if "stocks_below_target" in lookup:
                if not num_close(lookup["stocks_below_target"], EXP_STOCKS_BELOW, abs_tol=0):
                    errors.append(
                        f"Stocks_Below_Target={lookup['stocks_below_target']}, expected {EXP_STOCKS_BELOW}")

            # non-critical: Avg_PE_Premium
            if "avg_pe_premium" in lookup:
                if not num_close(lookup["avg_pe_premium"], EXP_AVG_PE_PREMIUM, abs_tol=0.3):
                    errors.append(
                        f"Avg_PE_Premium={lookup['avg_pe_premium']}, expected ~{EXP_AVG_PE_PREMIUM}")

            # CRITICAL: Most_Undervalued
            if "most_undervalued" in lookup:
                if norm_sym(lookup["most_undervalued"]) != EXP_MOST_UNDERVALUED:
                    critical.append(
                        f"Most_Undervalued={lookup['most_undervalued']}, expected {EXP_MOST_UNDERVALUED}")
            else:
                errors.append("Most_Undervalued not found")
                critical.append("Most_Undervalued not found")

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
        critical.append(f"Error reading Excel: {e}")
    return errors, critical


def _gsheet_conn():
    return psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432,
                            dbname="cowork_gym", user="eigent", password="camel")


def check_gsheet():
    """Returns (errors, critical_errors). Cross-checks 'Benchmark' sheet content."""
    errors, critical = [], []
    try:
        conn = _gsheet_conn()
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM gsheet.spreadsheets")
        count = cur.fetchone()[0]
        if count == 0:
            errors.append("No Google Sheet spreadsheet created")
            critical.append("No Google Sheet spreadsheet created")
            cur.close()
            conn.close()
            return errors, critical

        # Gather all cell text from the gsheet store.
        cur.execute("SELECT value FROM gsheet.cells")
        cell_vals = [str(r[0]) for r in cur.fetchall() if r[0] is not None]
        cell_count = len(cell_vals)
        if cell_count < 5:
            errors.append(f"Google Sheet has only {cell_count} cells, expected at least 5")
            critical.append(f"Google Sheet has only {cell_count} cells")

        joined = " ".join(cell_vals).upper()

        # CRITICAL: the 5 swapped RU symbols must actually appear in the gsheet.
        missing = [s for s in EXP_SYMBOLS if norm_sym(s) not in joined and s.upper() not in joined]
        if missing:
            critical.append(
                f"Google Sheet 'Benchmark' missing symbols: {missing}")

        # CRITICAL: at least 3 of the per-stock Upside_Pct values must appear somewhere
        # in the sheet (matches the Excel Consensus Analysis numbers). Tolerance via
        # string presence of the rounded value or near value among numeric cells.
        numeric = []
        for v in cell_vals:
            try:
                numeric.append(float(str(v).replace(",", ".")))
            except (TypeError, ValueError):
                pass
        matched = 0
        for sym in EXP_SYMBOLS:
            exp = EXP_UPSIDE[sym]
            if any(abs(n - exp) <= 0.5 for n in numeric):
                matched += 1
        if matched < 3:
            critical.append(
                f"Google Sheet 'Benchmark' Upside_Pct values do not match Excel "
                f"(only {matched}/5 of expected upside numbers present)")

        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking GSheet: {e}")
        critical.append(f"Error checking GSheet: {e}")
    return errors, critical


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    global EXP_UPSIDE, EXP_PE_VS, EXP_AVG_UPSIDE, EXP_STOCKS_ABOVE
    global EXP_STOCKS_BELOW, EXP_AVG_PE_PREMIUM, EXP_MOST_UNDERVALUED
    exp = compute_expected()
    EXP_UPSIDE = exp["upside"]
    EXP_PE_VS = exp["pe_vs"]
    EXP_AVG_UPSIDE = exp["avg_upside"]
    EXP_STOCKS_ABOVE = exp["stocks_above"]
    EXP_STOCKS_BELOW = exp["stocks_below"]
    EXP_AVG_PE_PREMIUM = exp["avg_pe_premium"]
    EXP_MOST_UNDERVALUED = exp["most_undervalued"]

    all_errors = []
    critical_errors = []

    print("  Checking Excel file...")
    errs, crit = check_excel(agent_ws)
    all_errors.extend(errs)
    critical_errors.extend(crit)
    if errs:
        for e in errs[:8]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    print("  Checking Google Sheet...")
    errs, crit = check_gsheet()
    all_errors.extend(errs)
    critical_errors.extend(crit)
    if errs:
        for e in errs[:5]:
            print(f"    ERROR: {e}")
    else:
        print("    PASS")

    # Critical gate: any critical failure => immediate FAIL.
    if critical_errors:
        print(f"\n=== CRITICAL FAILURE ({len(critical_errors)}) ===")
        for e in critical_errors[:10]:
            print(f"  CRITICAL: {e}")
        sys.exit(1)

    # Accuracy gate (kept binary here: all non-critical checks must also pass).
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
