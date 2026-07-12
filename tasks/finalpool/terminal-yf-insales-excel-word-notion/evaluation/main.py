"""Evaluation для terminal-yf-insales-excel-word-notion (RU: moex + teamly).

Проверки:
1. Commodity_Impact_Analysis.xlsx с 4 листами (Stock_Trends / Product_Margins /
   Correlation_Analysis / Strategic_Recommendations).
2. Pricing_Strategy_Memo.docx.
3. Teamly: страница/раздел «Market Research Dashboard» с двумя записями
   (золото + потребительская уверенность).
4. Скрипт correlation_analysis.py существует.

Семантические значения (цены/изменения/волатильность, рекомендации по марже)
выводятся из живой схемы moex.* через build_groundtruth.compute_finance /
compute_products / target_margin_and_action — НЕ зашиты в коде.

CRITICAL_CHECKS: любой провал => общий FAIL независимо от accuracy.
Иначе PASS требует accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2
from docx import Document

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_groundtruth import (
    compute_finance, compute_products, gold_significant,
    target_margin_and_action, COMMODITY, CONSUMER,
)

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Семантические критические проверки: их провал = общий FAIL.
CRITICAL_CHECKS = {
    "Stock_Trends содержит GLDRUB_TOM и OZON.ME",
    "Stock_Trends avg_price/price_change_pct совпадают с moex-данными",
    "Strategic_Recommendations: целевая маржа соответствует порогу золота (rescaled)",
    "Teamly 'Market Research Dashboard' существует с 2 записями (золото + спрос)",
    "correlation_analysis.py существует",
}

PRICE_TOL = 5.0      # рубли (масштаб ~4000-5000)
PCT_TOL = 1.0        # проценты


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def find_number_in_text(text, target, tol):
    """True, если target (с точностью tol) встречается среди чисел текста."""
    import re
    for tok in re.findall(r"-?\d+(?:[.,]\d+)?", text or ""):
        try:
            if abs(float(tok.replace(",", ".")) - target) <= tol:
                return True
        except ValueError:
            continue
    return False


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def check_excel(workspace, fin, prods):
    print("\n=== Check 1: Commodity_Impact_Analysis.xlsx ===")
    path = os.path.join(workspace, "Commodity_Impact_Analysis.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        check("Stock_Trends содержит GLDRUB_TOM и OZON.ME", False, "no file")
        check("Stock_Trends avg_price/price_change_pct совпадают с moex-данными", False, "no file")
        check("Strategic_Recommendations: целевая маржа соответствует порогу золота (rescaled)", False, "no file")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")
    sheets_lower = [s.lower().replace(" ", "_") for s in sheets]

    g = fin[COMMODITY]
    c = fin[CONSUMER]

    # --- Stock_Trends ---
    st_idx = next((i for i, s in enumerate(sheets_lower) if "stock" in s or "trend" in s), 0)
    ws1 = wb[sheets[st_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data1 = [r for r in rows1[1:] if any(x for x in r)]
    check("Stock_Trends has 2 rows", len(data1) >= 2, f"Found {len(data1)}")
    all_text1 = " ".join(str(x) for r in rows1 for x in r if x).lower()
    # CRITICAL: оба новых инструмента присутствуют
    has_gold = "gldrub_tom" in all_text1 or "gold" in all_text1 or "золот" in all_text1
    has_ozon = "ozon.me" in all_text1 or "ozon" in all_text1 or "озон" in all_text1
    check("Stock_Trends содержит GLDRUB_TOM и OZON.ME", has_gold and has_ozon,
          f"gold={has_gold} ozon={has_ozon}; text={all_text1[:120]}")

    if rows1:
        headers = [str(x).lower() if x else "" for x in rows1[0]]
        check("Has volatility column",
              any("volatil" in h for h in headers) or any("std" in h for h in headers),
              f"Headers: {rows1[0]}")

    # CRITICAL: значения avg_price и price_change_pct совпадают с moex-данными.
    # Сопоставляем строки по символу.
    by_sym = {}
    for r in data1:
        key = " ".join(str(x) for x in r if x).lower()
        if "gldrub_tom" in key or "gold" in key or "золот" in key:
            by_sym["gold"] = r
        elif "ozon" in key or "озон" in key:
            by_sym["ozon"] = r
    val_ok = True
    detail = ""
    for tag, sym in (("gold", COMMODITY), ("ozon", CONSUMER)):
        r = by_sym.get(tag)
        if r is None:
            val_ok = False
            detail = f"{tag} row missing"
            break
        nums = [x for x in r if isinstance(x, (int, float))]
        exp = fin[sym]
        ap_ok = any(num_close(n, exp["avg_price"], PRICE_TOL) for n in nums)
        pc_ok = any(num_close(n, exp["price_change_pct"], PCT_TOL) for n in nums)
        if not (ap_ok and pc_ok):
            val_ok = False
            detail = f"{tag}: avg_ok={ap_ok} pct_ok={pc_ok} nums={nums} exp_avg={exp['avg_price']} exp_pct={exp['price_change_pct']}"
            break
    check("Stock_Trends avg_price/price_change_pct совпадают с moex-данными", val_ok, detail)

    # --- Product_Margins ---
    pm_idx = next((i for i, s in enumerate(sheets_lower) if "product" in s or "margin" in s), 1)
    if pm_idx < len(sheets):
        ws2 = wb[sheets[pm_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if any(x for x in r)]
        check("Product_Margins has category rows", len(data2) >= 3, f"Found {len(data2)}")
        all_text2 = " ".join(str(x) for r in rows2 for x in r if x).lower()
        check("Has Электроника category", "электроника" in all_text2 or "electronics" in all_text2,
              f"Text: {all_text2[:120]}")

    # --- Correlation_Analysis ---
    ca_idx = next((i for i, s in enumerate(sheets_lower) if "correlation" in s), 2)
    if ca_idx < len(sheets):
        ws3 = wb[sheets[ca_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if any(x for x in r)]
        check("Correlation_Analysis has entries", len(data3) >= 2, f"Found {len(data3)}")
        all_text3 = " ".join(str(x) for r in rows3 for x in r if x).lower()
        check("Mentions gold in correlation",
              "gold" in all_text3 or "золот" in all_text3 or "gldrub" in all_text3,
              f"Text: {all_text3[:120]}")

    # --- Strategic_Recommendations ---
    sr_idx = next((i for i, s in enumerate(sheets_lower) if "strategic" in s or "recommend" in s), 3)
    tgt, action = target_margin_and_action()
    if sr_idx < len(sheets):
        ws4 = wb[sheets[sr_idx]]
        rows4 = list(ws4.iter_rows(values_only=True))
        data4 = [r for r in rows4[1:] if any(x for x in r)]
        check("Strategic_Recommendations has entries", len(data4) >= 3, f"Found {len(data4)}")
        all_text4 = " ".join(str(x) for r in rows4 for x in r if x).lower()
        # CRITICAL: целевая маржа и действие соответствуют порогу золота (rescaled).
        tgt_present = str(tgt) in all_text4
        action_present = action.lower() in all_text4
        check("Strategic_Recommendations: целевая маржа соответствует порогу золота (rescaled)",
              tgt_present and action_present,
              f"gold_change={g['price_change_pct']}% (>50%? {gold_significant()}) "
              f"expect target={tgt} action='{action}'; tgt_present={tgt_present} action_present={action_present}")
    else:
        check("Strategic_Recommendations: целевая маржа соответствует порогу золота (rescaled)",
              False, "sheet missing")


def check_word(workspace, fin):
    print("\n=== Check 2: Pricing_Strategy_Memo.docx ===")
    path = os.path.join(workspace, "Pricing_Strategy_Memo.docx")
    if not os.path.exists(path):
        check("Word document exists", False, f"Not found at {path}")
        return
    check("Word document exists", True)

    doc = Document(path)
    full_text = " ".join(p.text for p in doc.paragraphs)
    low = full_text.lower()
    check("Mentions gold or commodity",
          "gold" in low or "золот" in low or "gldrub" in low or "commodity" in low or "сырь" in low)
    check("Mentions OZON",
          "ozon" in low or "озон" in low)
    check("Mentions margin", "margin" in low or "маржа" in low or "маржу" in low or "маржи" in low)
    check("Has substantial content", len(full_text) > 200, f"Length: {len(full_text)}")
    # Конкретные числа золота присутствуют (хотя бы изменение или средняя цена).
    g = fin[COMMODITY]
    num_ok = (find_number_in_text(full_text, g["price_change_pct"], PCT_TOL)
              or find_number_in_text(full_text, g["avg_price"], PRICE_TOL))
    check("Word memo содержит конкретные числа по золоту", num_ok,
          f"expect ~{g['price_change_pct']}% или ~{g['avg_price']}")


# --------------------------------------------------------------------------- #
# Teamly
# --------------------------------------------------------------------------- #
def check_teamly(fin):
    print("\n=== Check 3: Teamly — Market Research Dashboard ===")
    g = fin[COMMODITY]
    c = fin[CONSUMER]
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Кандидаты: пользовательские страницы (id>3), плюс пространства с маркером.
        cur.execute("""
            SELECT p.id, COALESCE(p.title,''), COALESCE(p.body,''),
                   COALESCE(s.name,''), COALESCE(s.key,'')
            FROM teamly.pages p
            LEFT JOIN teamly.spaces s ON s.id = p.space_id
            WHERE p.id > 3
        """)
        pages = cur.fetchall()
        cur.execute("SELECT id, key, name, COALESCE(description,'') FROM teamly.spaces")
        spaces = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly 'Market Research Dashboard' существует с 2 записями (золото + спрос)", False, str(e))
        return

    def is_dashboard(text):
        t = (text or "").lower()
        return ("market" in t and ("research" in t or "dashboard" in t)) or \
               ("маркет" in t and "дашборд" in t) or ("исследован" in t and "рынк" in t)

    # Хаб-страница или пространство c названием Market Research Dashboard.
    dash_pages = [p for p in pages if is_dashboard(p[1]) or is_dashboard(p[3])]
    dash_spaces = [s for s in spaces if is_dashboard(s[2]) or is_dashboard(s[3]) or is_dashboard(s[1])]
    dash_exists = bool(dash_pages) or bool(dash_spaces)

    # Записи: ищем содержимое про золото и про потребительскую уверенность/OZON
    # среди страниц агента (заголовок+тело) или страниц в dashboard-пространстве.
    candidate = pages
    blob = " ".join((p[1] + " " + p[2]) for p in candidate).lower()

    gold_entry = ("gold" in blob or "золот" in blob or "gldrub_tom" in blob) and \
                 find_number_in_text(blob, g["price_change_pct"], PCT_TOL + 0.5)
    consumer_entry = ("ozon" in blob or "озон" in blob or "consumer" in blob or "потребит" in blob or "уверенност" in blob) and \
                     find_number_in_text(blob, c["price_change_pct"], PCT_TOL + 0.5)

    # Считаем число «записей» как минимум по наличию двух тематических страниц.
    n_entries = 0
    for p in candidate:
        txt = (p[1] + " " + p[2]).lower()
        if "gold" in txt or "золот" in txt or "ozon" in txt or "озон" in txt or "consumer" in txt or "потребит" in txt:
            n_entries += 1

    # CRITICAL: дашборд существует И содержит обе тематические записи с числами.
    check("Teamly 'Market Research Dashboard' существует с 2 записями (золото + спрос)",
          dash_exists and gold_entry and consumer_entry,
          f"dash={dash_exists} gold_entry={gold_entry} consumer_entry={consumer_entry} n={n_entries}; blob={blob[:160]}")

    # NON-critical структурный: маркер Trend (Up/Down/Stable) присутствует.
    trend_ok = any(w in blob for w in ("up", "down", "stable", "рост", "падение", "стабиль"))
    check("Teamly: присутствует индикатор тренда (Up/Down/Stable)", trend_ok, "no trend marker")


def check_script(workspace):
    print("\n=== Check 4: correlation_analysis.py ===")
    path = os.path.join(workspace, "correlation_analysis.py")
    check("correlation_analysis.py существует", os.path.exists(path), f"not at {path}")


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Commodity_Impact_Analysis.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        expected_keywords = {"stock", "trend", "product", "margin", "correlation", "strategic", "recommend"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0, f"Unexpected: {unexpected}")

    # Teamly: не должно быть дублей хаба Market Research Dashboard.
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM teamly.pages
            WHERE (title ILIKE '%market%' AND (title ILIKE '%research%' OR title ILIKE '%dashboard%'))
        """)
        n = cur.fetchone()[0]
        cur.close()
        conn.close()
        check("No duplicate Market Research Dashboard hub pages", n <= 1, f"Found {n} hub pages")
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    fin = compute_finance()
    prods = compute_products()

    check_excel(args.agent_workspace, fin, prods)
    check_word(args.agent_workspace, fin)
    check_teamly(fin)
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    result = {"total_passed": PASS_COUNT, "total_checks": total,
              "accuracy": accuracy, "critical_failed": critical_failed}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("FAIL (critical)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
