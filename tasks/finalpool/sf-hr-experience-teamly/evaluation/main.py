"""Evaluation for sf-hr-experience (ClickHouse + Teamly).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.
Expected values are computed LIVE from the read-only DB so eval stays in sync
with the centrally-russified department names.
"""
import argparse
import os
import re
import sys

import openpyxl
import psycopg2


def normalize_ru_numbers(text):
    """RU number normalization for substring/regex checks: collapse digit-group
    separators (space/NBSP/NNBSP/dot/comma before a 3-digit group) and turn
    decimal commas into dots ("31 588" -> "31588", "4 586,91" -> "4586.91")."""
    t = str(text or "")
    t = re.sub(r"(?<=\d)[ \xa0\u202f\u2009.,](?=\d{3}\b)", "", t)
    return re.sub(r"(?<=\d),(?=\d)", ".", t)

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Any failure here => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Excel file exists",
    "Per-department Avg_Tenure/Employee_Count/Min/Max match DB",
    "Summary Total_Employees matches DB exactly",
    "Summary Department_Count matches COUNT(DISTINCT DEPARTMENT)",
    "Summary Overall_Avg_Tenure matches DB",
    "Summary Highest_Avg_Tenure_Dept matches DB-derived top department",
    "Teamly 'Employee Tenure Dashboard' page profiles every department + summary",
    "Email to chro@company.com from hr-analytics@company.com, exact subject, correct body",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_expected_data():
    """Compute expected tenure data from read-only DB."""
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT "DEPARTMENT",
               ROUND(AVG("YEARS_EXPERIENCE")::numeric, 2) as avg_tenure,
               MIN("YEARS_EXPERIENCE") as min_tenure,
               MAX("YEARS_EXPERIENCE") as max_tenure,
               COUNT(*) as emp_count
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
        ORDER BY "DEPARTMENT"
    """)
    rows = cur.fetchall()

    cur.execute("""
        SELECT ROUND(AVG("YEARS_EXPERIENCE")::numeric, 2),
               COUNT(*),
               COUNT(DISTINCT "DEPARTMENT")
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
    """)
    overall = cur.fetchone()  # (overall_avg, total_employees, dept_count)
    cur.close()
    conn.close()
    # Top department by avg tenure (then by name for determinism).
    top_dept = max(rows, key=lambda r: (float(r[1]), )) [0] if rows else None
    return rows, overall, top_dept


def check_excel(agent_workspace):
    """Check Tenure_Analysis.xlsx against DB-derived expected values."""
    print("\n=== Checking Tenure_Analysis.xlsx ===")

    agent_file = os.path.join(agent_workspace, "Tenure_Analysis.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return
    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file exists", False, f"unreadable: {e}")
        return

    dept_data, overall, top_dept = get_expected_data()
    overall_avg, total_emp, dept_count = overall

    # --- Department Tenure sheet ---
    dept_sheet = None
    for name in wb.sheetnames:
        if "department" in name.lower() and "tenure" in name.lower():
            dept_sheet = wb[name]
            break
    dept_all_ok = True
    if dept_sheet is None:
        record("Sheet 'Department Tenure' exists", False, f"Sheets: {wb.sheetnames}")
        dept_all_ok = False
    else:
        record("Sheet 'Department Tenure' exists", True)
        rows = list(dept_sheet.iter_rows(min_row=2, values_only=True))
        record("Department Tenure has one row per department",
               len(rows) == len(dept_data), f"Got {len(rows)}, expected {len(dept_data)}")

        agent_lookup = {}
        for r in rows:
            if r and r[0]:
                agent_lookup[str(r[0]).strip().lower()] = r

        for exp in dept_data:
            dept, e_avg, e_min, e_max, e_cnt = exp
            a_row = agent_lookup.get(str(dept).strip().lower())
            if a_row is None:
                record(f"Department '{dept}' present", False, "Missing")
                dept_all_ok = False
                continue
            # a_row columns: Department, Avg_Tenure, Min_Tenure, Max_Tenure, Employee_Count
            ok_avg = num_close(a_row[1], e_avg, 0.05)
            ok_min = num_close(a_row[2], e_min, 0)
            ok_max = num_close(a_row[3], e_max, 0)
            ok_cnt = num_close(a_row[4], e_cnt, 0)
            ok = ok_avg and ok_min and ok_max and ok_cnt
            record(f"Department '{dept}' tenure values", ok,
                   f"avg exp={e_avg} got={a_row[1]} | min exp={e_min} got={a_row[2]} | "
                   f"max exp={e_max} got={a_row[3]} | cnt exp={e_cnt} got={a_row[4]}")
            if not ok:
                dept_all_ok = False
    record("Per-department Avg_Tenure/Employee_Count/Min/Max match DB", dept_all_ok)

    # --- Summary sheet ---
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = wb[name]
            break
    if sum_sheet is None:
        record("Sheet 'Summary' exists", False, f"Sheets: {wb.sheetnames}")
        record("Summary Total_Employees matches DB exactly", False, "no Summary sheet")
        record("Summary Department_Count matches COUNT(DISTINCT DEPARTMENT)", False, "no Summary sheet")
        record("Summary Overall_Avg_Tenure matches DB", False, "no Summary sheet")
        record("Summary Highest_Avg_Tenure_Dept matches DB-derived top department", False, "no Summary sheet")
        return
    record("Sheet 'Summary' exists", True)
    summary = {}
    for row in sum_sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            summary[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

    def find_val(*needles):
        for k, v in summary.items():
            if all(n in k for n in needles):
                return v
        return None

    v_total = find_val("total", "employee")
    record("Summary Total_Employees matches DB exactly", num_close(v_total, total_emp, 0),
           f"Expected {total_emp}, got {v_total}")

    v_deptcount = find_val("department", "count")
    record("Summary Department_Count matches COUNT(DISTINCT DEPARTMENT)",
           num_close(v_deptcount, dept_count, 0), f"Expected {dept_count}, got {v_deptcount}")

    v_overall = find_val("overall", "avg")
    record("Summary Overall_Avg_Tenure matches DB", num_close(v_overall, overall_avg, 0.05),
           f"Expected {overall_avg}, got {v_overall}")

    v_top = find_val("highest", "dept")
    if v_top is None:
        v_top = find_val("highest")
    top_ok = v_top is not None and str(v_top).strip().lower() == str(top_dept).strip().lower()
    record("Summary Highest_Avg_Tenure_Dept matches DB-derived top department", top_ok,
           f"Expected '{top_dept}', got '{v_top}'")


def check_teamly():
    """Critical: 'Employee Tenure Dashboard' page profiles every department + summary."""
    print("\n=== Checking Teamly Page ===")
    name = "Teamly 'Employee Tenure Dashboard' page profiles every department + summary"
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        dept_data, overall, top_dept = get_expected_data()
        cur.close()
        conn.close()
    except Exception as e:
        record(name, False, str(e))
        return

    # Find the dashboard page by title marker (English identifier preserved).
    page = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "employee tenure dashboard" in tl or ("tenure" in tl and ("dashboard" in tl or "дашборд" in tl or "панель" in tl)):
            page = (pid, title, body)
            break
    if page is None:
        record(name, False, f"new pages: {[(p[0], p[1]) for p in pages]}")
        return

    text = ((page[1] or "") + "\n" + (page[2] or ""))
    text_l = text.lower()

    # Every department name must appear in the page body.
    missing = [d[0] for d in dept_data if str(d[0]).lower() not in text_l]
    deps_ok = not missing

    # Concluding summary naming the top department (already covered by dept presence;
    # require a summary cue + the top dept appearing).
    summary_ok = str(top_dept).lower() in text_l and any(
        k in text_l for k in ["наиболее опыт", "наибольш", "most experienced", "highest", "в среднем", "итог", "вывод", "summary"]
    )

    record(name, deps_ok and summary_ok,
           f"missing departments: {missing}; summary_ok={summary_ok}")


def check_email():
    """Critical: HR tenure email to CHRO with exact subject + correct body."""
    print("\n=== Checking Email ===")
    name = "Email to chro@company.com from hr-analytics@company.com, exact subject, correct body"
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        rows = cur.fetchall()
        dept_data, overall, top_dept = get_expected_data()
        cur.close()
        conn.close()
    except Exception as e:
        record(name, False, str(e))
        return

    overall_avg, total_emp, dept_count = overall

    match = None
    for subject, from_addr, to_addr, body in rows:
        subj = str(subject or "")
        to_s = str(to_addr or "").lower()
        if subj.strip() == "Employee Tenure Analysis Report" and "chro@company.com" in to_s:
            match = (subject, from_addr, to_addr, body)
            break
    if match is None:
        record(name, False,
               f"no email with exact subject to chro@company.com; got {[(str(r[0]), str(r[2])) for r in rows]}")
        return

    subject, from_addr, to_addr, body = match
    from_ok = "hr-analytics@company.com" in str(from_addr or "").lower()
    body_l = str(body or "").lower()
    body_n = normalize_ru_numbers(body_l)

    total_ok = str(total_emp) in body_n
    dept_ok = str(dept_count) in body_n and any(
        k in body_l for k in ["отдел", "department"]
    )
    top_ok = str(top_dept).lower() in body_l

    record(name, from_ok and total_ok and dept_ok and top_ok,
           f"from_ok={from_ok} total_ok={total_ok}(exp {total_emp}) "
           f"dept_ok={dept_ok}(exp {dept_count}) top_ok={top_ok}(exp '{top_dept}')")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
