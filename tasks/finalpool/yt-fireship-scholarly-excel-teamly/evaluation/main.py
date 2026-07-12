"""
Evaluation for yt-fireship-scholarly-excel-teamly task.

Checks:
1. Tech_Research_Map.xlsx exists with Video_Paper_Mapping and All_Papers sheets
2. Video_Paper_Mapping has the 8 top Fireship videos (correct titles + views,
   rank 1 = the DeepSeek R1 bubble video), AI topic -> DeepSeek R1 top paper
3. All_Papers sheet lists the 10 topical papers with correct citation counts
4. Teamly page 'Tech Content Research Bridge' exists with substantive content
   (bridges several real video titles to real paper titles)

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Pass threshold otherwise: accuracy >= 70%.
"""
import json
import os
import sys
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks reflect the task's substance (correct ranking, correct
# top-paper-by-citation mapping, real bridge page), not mere structure.
CRITICAL_CHECKS = {
    "Rank 1 video is the DeepSeek R1 bubble video",
    "Video_Paper_Mapping has all 8 expected video titles in order",
    "AI topic Top_Paper is DeepSeek R1 (highest citations)",
    "All_Papers citation counts correct for key papers",
    "Teamly bridge page has substantive video<->paper content",
}

# Top-8 most-viewed Fireship videos (from youtube.videos, read-only source),
# ordered by view_count desc. (title, view_count)
EXPECTED_VIDEOS = [
    ("Big Tech in panic mode... Did DeepSeek R1 just pop the AI bubble?", 3878491),
    ("This free Chinese AI just crushed OpenAI's $200 o1 model...", 3115193),
    ("100+ Linux Things you Need to Know", 2891861),
    ("25 crazy software bugs explained", 2496715),
    ("Some bad code just broke a billion Windows machines", 2467547),
    ("DeepSeek stole our tech... says OpenAI", 2436569),
    ("Microsoft's new chip looks like science fiction", 2329914),
    ("Hackers are destroying the Internet's history book right now", 1927279),
]

# Highest-cited paper per topic (Top_Paper).
TOP_PAPER_BY_TOPIC = {
    "ai and large language models": ("DeepSeek R1: Incentivizing Reasoning Capability in LLMs via Reinforcement Learning", 1204),
    "linux and systems": ("Performance Analysis of Linux Kernel Scheduler Algorithms in Cloud Environments", 245),
    "software engineering": ("Automated Detection and Classification of Software Bugs Using Deep Learning", 312),
    "hardware and computing": ("Next-Generation Quantum-Classical Hybrid Computing Architectures", 423),
    "security": ("Supply Chain Security Vulnerabilities in Open Source Software: A Systematic Review", 334),
}

# Per-paper citation source-of-truth (title-keyword -> citation_count).
PAPER_CITATIONS = {
    "deepseek r1": 1204,
    "large language models for automated reasoning": 892,
    "performance analysis of linux kernel scheduler": 245,
    "ebpf-based observability": 178,
    "automated detection and classification of software bugs": 312,
    "root cause analysis in distributed systems": 189,
    "next-generation quantum-classical": 423,
    "energy-efficient neural processing units": 267,
    "supply chain security vulnerabilities": 334,
    "internet archive attacks": 156,
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def check_excel(agent_workspace):
    print("\n=== Check 1: Tech_Research_Map.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Tech_Research_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Research_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        # Critical checks that depend on the file fail too.
        record("Rank 1 video is the DeepSeek R1 bubble video", False, "no xlsx")
        record("Video_Paper_Mapping has all 8 expected video titles in order", False, "no xlsx")
        record("AI topic Top_Paper is DeepSeek R1 (highest citations)", False, "no xlsx")
        record("All_Papers citation counts correct for key papers", False, "no xlsx")
        return
    record("Tech_Research_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    record("Has Video_Paper_Mapping sheet",
           any("video" in s and "paper" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")
    record("Has All_Papers sheet",
           any("all" in s and "paper" in s for s in sheet_names_lower) or
           any("papers" in s for s in sheet_names_lower),
           f"Sheets: {wb.sheetnames}")

    # --- Video_Paper_Mapping ---
    mapping_sheet = None
    for name in wb.sheetnames:
        if "video" in name.lower() and "paper" in name.lower():
            mapping_sheet = wb[name]; break
        if "mapping" in name.lower():
            mapping_sheet = wb[name]; break

    map_rows = []
    if mapping_sheet:
        all_rows = list(mapping_sheet.iter_rows(values_only=True))
        header = all_rows[0] if all_rows else ()
        map_rows = [r for r in all_rows[1:] if any(c is not None and str(c).strip() for c in r)]
        record("Video_Paper_Mapping has 8 rows", len(map_rows) >= 8,
               f"Found {len(map_rows)} rows")

        headers_text = " ".join(norm(c) for c in header if c)
        record("Has Video_Rank, Video_Title, Video_Views columns",
               "rank" in headers_text and "title" in headers_text and "view" in headers_text,
               f"Headers: {header}")

        # Locate columns
        hidx = {norm(c): i for i, c in enumerate(header) if c is not None}
        def col(*cands):
            for cand in cands:
                for h, i in hidx.items():
                    if cand in h:
                        return i
            return None
        rank_i = col("rank")
        title_i = col("video_title", "title")
        views_i = col("view")
        topic_i = col("tech_topic", "topic")
        toppaper_i = col("top_paper_title")

        # CRITICAL: rank 1 is the DeepSeek R1 bubble video
        rank1_ok = False
        if title_i is not None:
            # find row with rank == 1, else first data row
            row1 = None
            if rank_i is not None:
                for r in map_rows:
                    if num_close(r[rank_i] if rank_i < len(r) else None, 1, 0.5):
                        row1 = r; break
            if row1 is None and map_rows:
                row1 = map_rows[0]
            if row1 is not None and title_i < len(row1):
                t = norm(row1[title_i])
                rank1_ok = "deepseek" in t and "bubble" in t
        record("Rank 1 video is the DeepSeek R1 bubble video", rank1_ok,
               "Rank 1 row is not the DeepSeek R1 bubble video")

        # CRITICAL: all 8 expected titles present, in view-desc order
        order_ok = False
        if title_i is not None and len(map_rows) >= 8:
            got_titles = [norm(r[title_i]) for r in map_rows[:8] if title_i < len(r)]
            order_ok = True
            for idx, (exp_title, _) in enumerate(EXPECTED_VIDEOS):
                et = norm(exp_title)
                if idx >= len(got_titles) or et not in got_titles[idx]:
                    order_ok = False
                    break
        record("Video_Paper_Mapping has all 8 expected video titles in order", order_ok,
               "Titles missing or not ordered by views desc")

        # Views match for first 3 rows (structural sanity)
        if views_i is not None and len(map_rows) >= 3:
            views_ok = all(
                num_close(map_rows[i][views_i] if views_i < len(map_rows[i]) else None,
                          EXPECTED_VIDEOS[i][1], max(EXPECTED_VIDEOS[i][1] * 0.02, 1))
                for i in range(3)
            )
            record("Video_Views match source for top 3", views_ok,
                   "View counts deviate from source")

        # CRITICAL: AI topic Top_Paper is DeepSeek R1
        ai_ok = False
        if topic_i is not None and toppaper_i is not None:
            for r in map_rows:
                if topic_i < len(r) and "ai and large language" in norm(r[topic_i]):
                    tp = norm(r[toppaper_i]) if toppaper_i < len(r) else ""
                    if "deepseek r1" in tp:
                        ai_ok = True; break
        record("AI topic Top_Paper is DeepSeek R1 (highest citations)", ai_ok,
               "AI topic's Top_Paper is not DeepSeek R1")
    else:
        record("Rank 1 video is the DeepSeek R1 bubble video", False, "no mapping sheet")
        record("Video_Paper_Mapping has all 8 expected video titles in order", False, "no mapping sheet")
        record("AI topic Top_Paper is DeepSeek R1 (highest citations)", False, "no mapping sheet")

    # --- All_Papers ---
    papers_sheet = None
    for name in wb.sheetnames:
        if "all" in name.lower() and "paper" in name.lower():
            papers_sheet = wb[name]; break
        if "paper" in name.lower() and "mapping" not in name.lower() and "video" not in name.lower():
            papers_sheet = wb[name]; break

    cites_ok = False
    if papers_sheet:
        all_rows = list(papers_sheet.iter_rows(values_only=True))
        # Locate the header row robustly (it may not be at index 0 if the sheet
        # has leading blank rows): the row that mentions both a title and citations.
        header_idx = 0
        for i, r in enumerate(all_rows):
            cells = " ".join(norm(c) for c in r if c is not None)
            if "citation" in cells and "title" in cells:
                header_idx = i
                break
        header = all_rows[header_idx] if all_rows else ()
        data_rows = [r for r in all_rows[header_idx + 1:]
                     if any(c is not None and str(c).strip() for c in r)]
        record("All_Papers has at least 10 papers", len(data_rows) >= 10,
               f"Found {len(data_rows)} data rows")

        hidx = {norm(c): i for i, c in enumerate(header) if c is not None}
        title_i = next((i for h, i in hidx.items() if "paper_title" in h or h == "title" or "title" in h), None)
        cite_i = next((i for h, i in hidx.items() if "citation" in h), None)

        # CRITICAL: citation counts correct for the key papers present
        if title_i is not None and cite_i is not None:
            matched = 0
            wrong = []
            for r in data_rows:
                t = norm(r[title_i]) if title_i < len(r) else ""
                for kw, exp_c in PAPER_CITATIONS.items():
                    if kw in t:
                        matched += 1
                        actual = r[cite_i] if cite_i < len(r) else None
                        if not num_close(actual, exp_c, 1):
                            wrong.append((kw, actual, exp_c))
                        break
            cites_ok = (matched >= 8 and not wrong)
            record("All_Papers citation counts correct for key papers", cites_ok,
                   f"matched={matched} (need >=8), wrong={wrong[:3]}")
        else:
            record("All_Papers citation counts correct for key papers", False,
                   f"Missing Paper_Title/Citations column; header={header}")
    else:
        record("All_Papers citation counts correct for key papers", False, "no All_Papers sheet")


def check_teamly():
    print("\n=== Check 2: Teamly page 'Tech Content Research Bridge' ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Teamly database accessible", False, str(e))
        record("Teamly bridge page has substantive video<->paper content", False, str(e))
        return
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE '%%tech content research bridge%%'
               OR title ILIKE '%%research bridge%%'
        """)
        pages = cur.fetchall()
        record("Teamly page 'Tech Content Research Bridge' exists", len(pages) >= 1,
               f"Found {len(pages)} matching pages")

        if not pages:
            record("Teamly bridge page has substantive video<->paper content", False,
                   "No matching page")
            return

        combined = " ".join((str(t) + " " + str(b)).lower() for _, t, b in pages)
        max_len = max(len(str(b)) for _, _, b in pages)
        record("Teamly bridge page has non-trivial body", max_len >= 100,
               f"Longest matching body is {max_len} chars")

        # CRITICAL semantic: bridges real video titles to real paper titles.
        video_keywords = ["deepseek", "linux", "software bug", "windows", "chip",
                          "hackers", "internet", "openai"]
        paper_keywords = ["deepseek r1", "linux kernel scheduler",
                          "software bugs", "quantum-classical", "supply chain"]
        video_hits = sum(1 for kw in video_keywords if kw in combined)
        paper_hits = sum(1 for kw in paper_keywords if kw in combined)
        substantive = video_hits >= 4 and paper_hits >= 2
        record("Teamly bridge page has substantive video<->paper content", substantive,
               f"video_hits={video_hits} (need >=4), paper_hits={paper_hits} (need >=2)")
    except Exception as e:
        record("Teamly page check", False, str(e))
        record("Teamly bridge page has substantive video<->paper content", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
