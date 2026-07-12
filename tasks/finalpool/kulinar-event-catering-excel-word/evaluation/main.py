"""Evaluation для kulinar-event-catering-excel-word (RU-стек: kulinar/forms).

Проверяет:
  - Excel Catering_Plan.xlsx: лист Menu (>=8 строк, требуемые колонки,
    Servings_For_30 числовой >=1) и лист Ingredients (колонки + ссылочная
    целостность Dish_Name на Menu).
  - Блюда из Menu — реальные блюда из базы kulinar (не выдуманы).
  - Word Catering_Proposal.docx: резюме / обзор меню / график подготовки
    (RU+EN ключевые слова).
  - Forms (gform.*): опрос «Menu Approval Survey» с >=3 вопросами, созданный
    агентом (preprocess НЕ пред-засевает форму).
  - Email на client@corporate.com с нужной темой и упоминанием меню/предложения/формы.

Критические чеки (CRITICAL_CHECKS): любой их fail => задача FAIL, даже если общая
accuracy >= 70%. Структурные чеки — мягкие. Порог: accuracy >= 70% И нет
критических провалов.
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# Канонический набор блюд kulinar (источник: local_servers/kulinar-mcp/src/data/all_recipes.json).
KULINAR_RECIPES = {
    "Салат Оливье", "Винегрет", "Сельдь под шубой", "Салат Мимоза", "Крабовый салат",
    "Греческий салат", "Салат с курицей и грибами", "Холодец", "Икра кабачковая",
    "Грибы маринованные", "Сало солёное", "Селёдка с луком", "Борщ",
    "Щи из квашеной капусты", "Солянка мясная", "Уха", "Окрошка", "Грибной суп",
    "Рассольник", "Куриный бульон с лапшой", "Бефстроганов", "Пельмени домашние",
    "Голубцы", "Котлеты домашние", "Жаркое в горшочках", "Курица в сметане",
    "Рыба запечённая по-русски", "Цыплёнок табака", "Гречка с тушёнкой",
    "Плов узбекский", "Картофельное пюре", "Гречневая каша", "Перловая каша",
    "Картофель отварной с укропом", "Рис отварной", "Пирожки с капустой жареные",
    "Пирожки с мясом печёные", "Блины тонкие", "Кулебяка с капустой и яйцом",
    "Расстегаи с рыбой", "Медовик", "Наполеон", "Сырники", "Пасха творожная",
    "Ватрушки с творогом", "Кисель ягодный", "Морс клюквенный",
    "Компот из сухофруктов", "Сбитень", "Квас домашний",
}

# Критические чеки по имени record()
CRITICAL_CHECKS = {
    "Excel Menu: >=8 строк с требуемыми колонками и числовым Servings_For_30",
    "Excel Ingredients: колонки и ссылочная целостность Dish_Name на Menu",
    "Блюда Menu — реальные блюда из базы kulinar",
    "Word: резюме + обзор меню + график подготовки",
    "Forms: опрос «Menu Approval Survey» с >=3 вопросами",
    "Email: на client@corporate.com с нужной темой и упоминанием меню/предложения/формы",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def is_num_ge(v, lo):
    try:
        return float(v) >= lo
    except (TypeError, ValueError):
        return False


def norm_recipe(s: str) -> str:
    """Нормализация названия блюда (регистр, ё/е, пробелы) для сравнения с каноном."""
    s = (s or "").strip().lower().replace("ё", "е")
    return " ".join(s.split())


CANON_NORM = {norm_recipe(r) for r in KULINAR_RECIPES}


def recipe_is_real(name: str) -> bool:
    n = norm_recipe(name)
    if not n:
        return False
    return n in CANON_NORM or any(c in n or n in c for c in CANON_NORM)


def check_excel(agent_workspace):
    print("\n=== Проверка Excel Catering_Plan.xlsx ===")
    menu_dishes = []
    try:
        import openpyxl
    except ImportError:
        record("openpyxl доступен", False, "openpyxl не установлен")
        record("Excel Menu: >=8 строк с требуемыми колонками и числовым Servings_For_30", False, "нет openpyxl")
        record("Excel Ingredients: колонки и ссылочная целостность Dish_Name на Menu", False, "нет openpyxl")
        return menu_dishes

    xlsx_path = os.path.join(agent_workspace, "Catering_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Catering_Plan.xlsx существует", False, f"Не найден: {xlsx_path}")
        record("Excel Menu: >=8 строк с требуемыми колонками и числовым Servings_For_30", False, "нет файла")
        record("Excel Ingredients: колонки и ссылочная целостность Dish_Name на Menu", False, "нет файла")
        return menu_dishes
    record("Catering_Plan.xlsx существует", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel читается", False, str(e))
        record("Excel Menu: >=8 строк с требуемыми колонками и числовым Servings_For_30", False, "файл не читается")
        record("Excel Ingredients: колонки и ссылочная целостность Dish_Name на Menu", False, "файл не читается")
        return menu_dishes
    record("Excel читается", True)

    def get_sheet(substr):
        for s in wb.sheetnames:
            if substr in s.strip().lower():
                return wb[s]
        return None

    menu_ws = get_sheet("menu")
    ingr_ws = get_sheet("ingredient")
    record("Лист Menu существует", menu_ws is not None, f"Листы: {wb.sheetnames}")
    record("Лист Ingredients существует", ingr_ws is not None, f"Листы: {wb.sheetnames}")

    # --- Menu ---
    menu_dish_set = set()
    if menu_ws is not None:
        rows = list(menu_ws.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c is not None else "" for c in (rows[0] if rows else [])]
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]

        def col_idx(*keys):
            for i, h in enumerate(headers):
                if any(k in h for k in keys):
                    return i
            return None

        i_dish = col_idx("dish", "name")
        i_cat = col_idx("category")
        i_serv = col_idx("serving", "servings_for")
        i_prep = col_idx("prep", "time")
        i_cost = col_idx("cost", "price")

        has_all_cols = all(x is not None for x in (i_dish, i_cat, i_serv, i_prep, i_cost))
        record("Menu: колонка Dish_Name", i_dish is not None, f"Headers: {headers}")
        record("Menu: колонка Category", i_cat is not None, f"Headers: {headers}")
        record("Menu: колонка Servings_For_30", i_serv is not None, f"Headers: {headers}")
        record("Menu: колонка Prep_Time_Minutes", i_prep is not None, f"Headers: {headers}")
        record("Menu: колонка Estimated_Cost_USD", i_cost is not None, f"Headers: {headers}")

        # числовой Servings_For_30 >=1 в каждой строке
        servings_ok = i_serv is not None and len(data_rows) > 0
        if servings_ok:
            for r in data_rows:
                if i_serv >= len(r) or not is_num_ge(r[i_serv], 1):
                    servings_ok = False
                    break

        # собрать названия блюд
        if i_dish is not None:
            for r in data_rows:
                if i_dish < len(r) and r[i_dish] is not None and str(r[i_dish]).strip():
                    d = str(r[i_dish]).strip()
                    menu_dishes.append(d)
                    menu_dish_set.add(norm_recipe(d))

        record("Excel Menu: >=8 строк с требуемыми колонками и числовым Servings_For_30",
               has_all_cols and len(data_rows) >= 8 and servings_ok,
               f"cols={has_all_cols}, rows={len(data_rows)}, servings_ok={servings_ok}")
    else:
        record("Excel Menu: >=8 строк с требуемыми колонками и числовым Servings_For_30",
               False, "нет листа Menu")

    # --- Ingredients ---
    if ingr_ws is not None:
        rows = list(ingr_ws.iter_rows(values_only=True))
        headers = [str(c).strip().lower() if c is not None else "" for c in (rows[0] if rows else [])]
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]

        def col_idx(*keys):
            for i, h in enumerate(headers):
                if any(k in h for k in keys):
                    return i
            return None

        i_dish = col_idx("dish", "name")
        i_ingr = col_idx("ingredient")
        i_qty = col_idx("quantity", "qty")
        i_unit = col_idx("unit")
        has_all_cols = all(x is not None for x in (i_dish, i_ingr, i_qty, i_unit))
        record("Ingredients: требуемые колонки присутствуют", has_all_cols, f"Headers: {headers}")

        # ссылочная целостность: каждый Dish_Name в Ingredients есть в Menu
        ref_ok = has_all_cols and len(data_rows) > 0 and bool(menu_dish_set)
        if ref_ok:
            for r in data_rows:
                if i_dish >= len(r):
                    ref_ok = False
                    break
                d = norm_recipe(str(r[i_dish]) if r[i_dish] is not None else "")
                if not d:
                    continue
                if d not in menu_dish_set and not any(d in m or m in d for m in menu_dish_set):
                    ref_ok = False
                    break

        record("Excel Ingredients: колонки и ссылочная целостность Dish_Name на Menu",
               has_all_cols and ref_ok, f"cols={has_all_cols}, ref_ok={ref_ok}")
    else:
        record("Excel Ingredients: колонки и ссылочная целостность Dish_Name на Menu",
               False, "нет листа Ingredients")

    return menu_dishes


def check_recipes(menu_dishes):
    print("\n=== Проверка блюд против базы kulinar ===")
    if not menu_dishes:
        record("Блюда Menu — реальные блюда из базы kulinar", False, "нет блюд в Menu")
        return
    matched = sum(1 for d in menu_dishes if recipe_is_real(d))
    unknown = [d for d in menu_dishes if not recipe_is_real(d)]
    # Допускаем небольшую долю шума, но подавляющее большинство должно быть реальным.
    ratio_ok = matched >= max(8, int(0.8 * len(menu_dishes)))
    record("Блюда Menu — реальные блюда из базы kulinar",
           ratio_ok, f"совпало {matched}/{len(menu_dishes)}; неизвестные: {unknown[:5]}")


def check_word_doc(agent_workspace):
    print("\n=== Проверка Word Catering_Proposal.docx ===")
    try:
        from docx import Document
    except ImportError:
        record("python-docx доступен", False, "python-docx не установлен")
        record("Word: резюме + обзор меню + график подготовки", False, "нет python-docx")
        return

    docx_path = os.path.join(agent_workspace, "Catering_Proposal.docx")
    if not os.path.exists(docx_path):
        record("Catering_Proposal.docx существует", False, f"Не найден: {docx_path}")
        record("Word: резюме + обзор меню + график подготовки", False, "нет файла")
        return
    record("Catering_Proposal.docx существует", True)

    try:
        doc = Document(docx_path)
    except Exception as e:
        record("Word документ читается", False, str(e))
        record("Word: резюме + обзор меню + график подготовки", False, "файл не читается")
        return
    record("Word документ читается", True)

    all_text = "\n".join(p.text for p in doc.paragraphs).lower()

    has_catering = any(k in all_text for k in ("catering", "menu", "кейтеринг", "общепит", "меню"))
    has_summary = any(k in all_text for k in ("summary", "overview", "executive", "резюме", "обзор"))
    has_timeline = any(k in all_text for k in ("timeline", "preparation", "schedule",
                                               "график", "подготовк", "расписан"))

    record("Word: упоминает кейтеринг/меню (RU+EN)", has_catering, "нет catering/меню")
    record("Word: раздел резюме/обзора (RU+EN)", has_summary, "нет summary/резюме")
    record("Word: раздел графика подготовки (RU+EN)", has_timeline, "нет timeline/график")

    record("Word: резюме + обзор меню + график подготовки",
           has_catering and has_summary and has_timeline,
           f"catering={has_catering}, summary={has_summary}, timeline={has_timeline}")


def check_forms():
    print("\n=== Проверка Forms (Menu Approval Survey) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Forms: опрос «Menu Approval Survey» с >=3 вопросами", False, str(e))
        return
    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    def title_matches(title):
        t = (title or "").lower()
        en = any(k in t for k in ("menu", "approval", "survey"))
        ru = any(k in t for k in ("меню", "одобрен", "опрос", "согласован"))
        return en or ru

    target = None
    for fid, title in forms:
        if title_matches(title):
            target = (fid, title)
            break

    record("Forms: найдена форма с подходящим заголовком",
           target is not None, f"Заголовки: {[t for _f, t in forms]}")

    q_count = 0
    if target is not None:
        cur.execute("SELECT COUNT(*) FROM gform.questions WHERE form_id = %s", (target[0],))
        q_count = cur.fetchone()[0]

    record("Forms: опрос «Menu Approval Survey» с >=3 вопросами",
           target is not None and q_count >= 3, f"вопросов={q_count}")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Проверка Email на client@corporate.com ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Email: на client@corporate.com с нужной темой и упоминанием меню/предложения/формы",
               False, str(e))
        return
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()
    cur.close()
    conn.close()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "client@corporate.com" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    record("Email отправлен на client@corporate.com", matching is not None,
           f"Сообщений: {len(messages)}")

    subj_ok = False
    body_ok = False
    if matching:
        subject, _, _, body_text = matching
        subj_l = (subject or "").lower()
        subj_ok = "catering proposal for team building event" in subj_l
        all_text = ((subject or "") + " " + (body_text or "")).lower()
        body_ok = any(k in all_text for k in (
            "catering", "menu", "proposal", "form",
            "кейтеринг", "меню", "предложен", "форм"))
        record("Email: тема корректна", subj_ok, f"Subject: {subject}")
        record("Email: упоминает меню/предложение/форму (RU+EN)", body_ok, "нет ключевых слов")

    record("Email: на client@corporate.com с нужной темой и упоминанием меню/предложения/формы",
           matching is not None and subj_ok and body_ok,
           f"matching={matching is not None}, subj={subj_ok}, body={body_ok}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("KULINAR-EVENT-CATERING (RU: kulinar/forms) - EVALUATION")
    print("=" * 70)

    menu_dishes = check_excel(args.agent_workspace)
    check_recipes(menu_dishes or [])
    check_word_doc(args.agent_workspace)
    check_forms()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": CRITICAL_FAILED,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"CRITICAL FAIL: {CRITICAL_FAILED}")
        print("FAIL")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
