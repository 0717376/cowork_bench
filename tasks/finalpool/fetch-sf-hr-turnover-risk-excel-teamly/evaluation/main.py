"""Evaluation script for fetch-sf-hr-turnover-risk-excel (clickhouse + teamly swap).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Value-level checks recompute Risk_Level / Satisfaction_Gap / Estimated_Turnover_Cost
/ summary counts from the agent's OWN Risk_Overview rows tied to the benchmark
JSON thresholds (turnover_rate / risk_threshold_satisfaction), so garbage numbers
no longer pass. Department labels are accepted in both RU and EN form.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Benchmark thresholds keyed by BOTH the russified warehouse dept name (produced
# by the central clickhouse map) and the original English name. These are the
# turnover_rate / risk_threshold_satisfaction the agent must read from
# http://localhost:30302/api/turnover_benchmarks.json.
BENCHMARKS = {
    "Engineering": (13.2, 2.5), "Инженерия": (13.2, 2.5),
    "Finance": (11.8, 2.8),     "Финансы": (11.8, 2.8),
    "HR": (10.5, 2.5),          "Кадры": (10.5, 2.5),
    "Operations": (12.1, 2.6),  "Операции": (12.1, 2.6),
    "R&D": (14.0, 2.4),         "НИОКР": (14.0, 2.4),
    "Sales": (15.3, 2.7),       "Продажи": (15.3, 2.7),
    "Support": (16.8, 2.5),     "Поддержка": (16.8, 2.5),
}

# RU<->EN alias groups so the same department maps to one canonical bucket.
DEPT_CANON = {
    "engineering": "ENG", "инженерия": "ENG",
    "finance": "FIN", "финансы": "FIN",
    "hr": "HR", "кадры": "HR",
    "operations": "OPS", "операции": "OPS",
    "r&d": "RND", "ниокр": "RND",
    "sales": "SAL", "продажи": "SAL",
    "support": "SUP", "поддержка": "SUP",
}

CRITICAL_CHECKS = {
    "Risk_Overview: Risk_Level recomputed from Avg_Satisfaction vs benchmark threshold",
    "Risk_Overview: Industry_Turnover_Rate & Risk_Threshold match benchmark JSON",
    "Detailed_Metrics: Satisfaction_Gap and Estimated_Turnover_Cost correct",
    "Risk_Summary: counts consistent with recomputed Risk_Overview levels",
    "Risk_Summary: Highest_Risk_Department = min Satisfaction_Gap department",
    "Teamly 'Turnover Risk Dashboard' page exists with non-empty analysis content",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_sheet(wb, name):
    target = name.strip().lower().replace(" ", "_")
    for n in wb.sheetnames:
        if n.strip().lower().replace(" ", "_") == target:
            return wb[n]
    return None


def _header_map(ws):
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def _dept_key(label):
    return DEPT_CANON.get(str(label or "").strip().lower())


def _level_from(sat, thr):
    if sat is None or thr is None:
        return None
    if sat < thr:
        return "high"
    if sat <= thr + 0.5:
        return "medium"
    return "low"


def check_structure(wb):
    """NON-critical structural checks (sheets / headers / row counts)."""
    for sheet, cols, minrows in [
        ("Risk_Overview", ['Department', 'Employee_Count', 'Avg_Salary', 'Risk_Level'], 7),
        ("Risk_Summary", ['Metric', 'Value'], 5),
        ("Detailed_Metrics", ['Department', 'Satisfaction_Gap', 'Estimated_Turnover_Cost'], 7),
    ]:
        ws = get_sheet(wb, sheet)
        check(f"{sheet} sheet exists", ws is not None)
        if ws is None:
            continue
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        check(f"{sheet} has >= {minrows} rows", len(data_rows) >= minrows, f"got {len(data_rows)}")
        headers, _ = _header_map(ws)
        for col in cols:
            check(f"{sheet} has {col} column", col.lower() in headers, f"headers: {headers[:8]}")


def parse_overview(wb):
    """Return {canon_key: dict(label, count, salary, sat, rate, thr, level)} from Risk_Overview."""
    ws = get_sheet(wb, "Risk_Overview")
    if ws is None:
        return {}
    headers, hm = _header_map(ws)
    def col(*names):
        for n in names:
            if n in hm:
                return hm[n]
        return None
    ci = {
        "dept": col("department"), "count": col("employee_count"),
        "salary": col("avg_salary"), "sat": col("avg_satisfaction"),
        "rate": col("industry_turnover_rate"), "thr": col("risk_threshold"),
        "level": col("risk_level"),
    }
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        label = row[ci["dept"]] if ci["dept"] is not None else None
        key = _dept_key(label)
        if key is None:
            continue
        out[key] = {
            "label": label,
            "count": safe_float(row[ci["count"]]) if ci["count"] is not None else None,
            "salary": safe_float(row[ci["salary"]]) if ci["salary"] is not None else None,
            "sat": safe_float(row[ci["sat"]]) if ci["sat"] is not None else None,
            "rate": safe_float(row[ci["rate"]]) if ci["rate"] is not None else None,
            "thr": safe_float(row[ci["thr"]]) if ci["thr"] is not None else None,
            "level": str(row[ci["level"]]).strip().lower() if ci["level"] is not None and row[ci["level"]] is not None else None,
        }
    return out


LEVEL_ALIASES = {
    "high": {"high", "высокий"},
    "medium": {"medium", "средний"},
    "low": {"low", "низкий"},
}


def _level_matches(agent_level, expected):
    if agent_level is None:
        return False
    return agent_level in LEVEL_ALIASES.get(expected, {expected})


def check_overview_values(ov):
    """CRITICAL: benchmark match + Risk_Level recompute for all departments."""
    if not ov:
        check("Risk_Overview: Industry_Turnover_Rate & Risk_Threshold match benchmark JSON",
              False, "no parseable Risk_Overview rows")
        check("Risk_Overview: Risk_Level recomputed from Avg_Satisfaction vs benchmark threshold",
              False, "no parseable Risk_Overview rows")
        return

    bench_ok = True
    bench_detail = []
    level_ok = True
    level_detail = []
    for key, r in ov.items():
        bench = BENCHMARKS.get(str(r["label"]).strip())
        if bench is None:
            bench_ok = False
            bench_detail.append(f"{r['label']}: no benchmark")
            continue
        exp_rate, exp_thr = bench
        if r["rate"] is None or abs(r["rate"] - exp_rate) > 0.05:
            bench_ok = False
            bench_detail.append(f"{r['label']}: rate {r['rate']}!={exp_rate}")
        if r["thr"] is None or abs(r["thr"] - exp_thr) > 0.01:
            bench_ok = False
            bench_detail.append(f"{r['label']}: thr {r['thr']}!={exp_thr}")
        expected_level = _level_from(r["sat"], exp_thr)
        if not _level_matches(r["level"], expected_level):
            level_ok = False
            level_detail.append(f"{r['label']}: sat={r['sat']} thr={exp_thr} -> {expected_level}, got {r['level']}")

    check("Risk_Overview: Industry_Turnover_Rate & Risk_Threshold match benchmark JSON",
          bench_ok, "; ".join(bench_detail))
    check("Risk_Overview: Risk_Level recomputed from Avg_Satisfaction vs benchmark threshold",
          level_ok, "; ".join(level_detail))


def check_detailed_values(wb, ov):
    """CRITICAL: Satisfaction_Gap & Estimated_Turnover_Cost for >=3 departments."""
    ws = get_sheet(wb, "Detailed_Metrics")
    if ws is None or not ov:
        check("Detailed_Metrics: Satisfaction_Gap and Estimated_Turnover_Cost correct",
              False, "missing sheet or overview")
        return
    headers, hm = _header_map(ws)
    def col(*names):
        for n in names:
            if n in hm:
                return hm[n]
        return None
    ci_dept = col("department")
    ci_gap = col("satisfaction_gap")
    ci_cost = col("estimated_turnover_cost")
    if ci_dept is None or ci_gap is None or ci_cost is None:
        check("Detailed_Metrics: Satisfaction_Gap and Estimated_Turnover_Cost correct",
              False, f"headers: {headers[:8]}")
        return

    checked = 0
    bad = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or all(c is None for c in row):
            continue
        key = _dept_key(row[ci_dept])
        if key is None or key not in ov:
            continue
        r = ov[key]
        bench = BENCHMARKS.get(str(r["label"]).strip())
        if bench is None or r["sat"] is None or r["count"] is None or r["salary"] is None:
            continue
        exp_rate, exp_thr = bench
        exp_gap = round(r["sat"] - exp_thr, 2)
        exp_cost = round(r["count"] * r["salary"] * exp_rate / 100)
        gap = safe_float(row[ci_gap])
        cost = safe_float(row[ci_cost])
        checked += 1
        if gap is None or abs(gap - exp_gap) > 0.02:
            bad.append(f"{r['label']}: gap {gap}!={exp_gap}")
        # cost tolerance: rounding + agent may round salary; allow 0.5%
        tol = max(2.0, abs(exp_cost) * 0.005)
        if cost is None or abs(cost - exp_cost) > tol:
            bad.append(f"{r['label']}: cost {cost}!={exp_cost}")

    check("Detailed_Metrics: Satisfaction_Gap and Estimated_Turnover_Cost correct",
          checked >= 3 and not bad, f"checked={checked} bad={bad[:6]}")


def check_summary_values(wb, ov):
    """CRITICAL: summary counts + Highest_Risk_Department from recomputed levels."""
    ws = get_sheet(wb, "Risk_Summary")
    if ws is None or not ov:
        check("Risk_Summary: counts consistent with recomputed Risk_Overview levels",
              False, "missing sheet or overview")
        check("Risk_Summary: Highest_Risk_Department = min Satisfaction_Gap department",
              False, "missing sheet or overview")
        return

    # Build metric->value map (Metric/Value layout).
    metrics = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        metrics[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

    # Recompute levels + gaps from benchmark.
    levels = {}
    gaps = {}
    for key, r in ov.items():
        bench = BENCHMARKS.get(str(r["label"]).strip())
        if bench is None or r["sat"] is None:
            continue
        _, exp_thr = bench
        levels[key] = _level_from(r["sat"], exp_thr)
        gaps[key] = round(r["sat"] - exp_thr, 2)

    exp_total = len(ov)
    exp_high = sum(1 for v in levels.values() if v == "high")
    exp_med = sum(1 for v in levels.values() if v == "medium")
    exp_low = sum(1 for v in levels.values() if v == "low")
    exp_atrisk = sum(int(ov[k]["count"]) for k, v in levels.items()
                     if v in ("high", "medium") and ov[k]["count"] is not None)

    def mget(name):
        return safe_float(metrics.get(name.lower()))

    counts_ok = True
    cdetail = []
    for label, exp in [
        ("Total_Departments", exp_total), ("High_Risk_Count", exp_high),
        ("Medium_Risk_Count", exp_med), ("Low_Risk_Count", exp_low),
        ("Total_At_Risk_Employees", exp_atrisk),
    ]:
        got = mget(label)
        if got is None or abs(got - exp) > 0.5:
            counts_ok = False
            cdetail.append(f"{label}: {got}!={exp}")
    check("Risk_Summary: counts consistent with recomputed Risk_Overview levels",
          counts_ok, "; ".join(cdetail))

    # Highest_Risk_Department = department with minimum gap.
    hrd_ok = False
    hrd_detail = ""
    if gaps:
        min_key = min(gaps, key=lambda k: gaps[k])
        raw = metrics.get("highest_risk_department")
        agent_key = _dept_key(raw)
        # accept ties on the minimum gap
        min_gap = gaps[min_key]
        tie_keys = {k for k, g in gaps.items() if abs(g - min_gap) < 1e-9}
        hrd_ok = agent_key in tie_keys
        hrd_detail = f"agent={raw} expected one of {[ov[k]['label'] for k in tie_keys]}"
    check("Risk_Summary: Highest_Risk_Department = min Satisfaction_Gap department",
          hrd_ok, hrd_detail)


def check_files(agent_workspace):
    check("risk_scorer.py exists",
          os.path.exists(os.path.join(agent_workspace, "risk_scorer.py")))
    check("risk_assessment.json exists",
          os.path.exists(os.path.join(agent_workspace, "risk_assessment.json")))


def check_teamly():
    """CRITICAL: 'Turnover Risk Dashboard' page exists (non-seed, id>3) with
    non-empty analysis content referencing departments / metrics. Independent of Excel."""
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, COALESCE(title,''), COALESCE(body,'') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Turnover Risk Dashboard' page exists with non-empty analysis content", False, str(e))
        return

    dash = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "план спринта" in tl:  # noise page
            continue
        if "turnover" in tl or "dashboard" in tl or "дашборд" in tl or \
           ("риск" in tl and "текуч" in tl):
            dash = (pid, title, body)
            break

    if dash is None:
        check("Teamly 'Turnover Risk Dashboard' page exists with non-empty analysis content",
              False, f"new pages: {[(p[0], p[1]) for p in pages]}")
        return

    text = ((dash[1] or "") + " " + (dash[2] or "")).lower()
    refs = any(k in text for k in [
        "risk", "риск", "high", "высок", "department", "подраздел",
        "turnover", "текуч", "recommend", "рекоменд", "satisfaction", "удовлетвор",
    ])
    non_empty = len((dash[2] or "").strip()) >= 20
    check("Teamly 'Turnover Risk Dashboard' page exists with non-empty analysis content",
          non_empty and refs, f"body_len={len((dash[2] or '').strip())} refs={refs}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Turnover_Risk_Assessment.xlsx")
    check("Turnover_Risk_Assessment.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        check_structure(wb)
        ov = parse_overview(wb)
        check_overview_values(ov)
        check_detailed_values(wb, ov)
        check_summary_values(wb, ov)
        check_files(agent_workspace)

    # Teamly check runs independently of Excel existence.
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
