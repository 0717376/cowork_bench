"""Evaluation for canvas-enrollment-overview-excel-email."""
import argparse
import os
import re
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def normalize_ru_numbers(text):
    """Collapse RU digit-group separators ('7 804'/'7.804'/'7,804' -> '7804') and decimal commas ('0,5' -> '0.5')."""
    t = text or ""
    t = re.sub(r"(?<=\d)[\u00a0\u202f ](?=\d)", "", t)
    t = re.sub(r"(?<=\d)[.,](?=\d{3}(?!\d))", "", t)
    t = re.sub(r"(?<=\d),(?=\d)", ".", t)
    return t


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Spring2014_Enrollment.xlsx")
    gt_file = os.path.join(gt_dir, "Spring2014_Enrollment.xlsx")

    file_errors = []
    db_errors = []
    # CRITICAL_CHECKS: any failure here => hard FAIL regardless of accuracy.
    critical_failures = []

    if not os.path.exists(agent_file):
        file_errors.append(f"Agent output not found: {agent_file}")
        critical_failures.append("Spring2014_Enrollment.xlsx not produced")
    if not os.path.exists(gt_file):
        file_errors.append(f"Groundtruth not found: {gt_file}")

    if not file_errors:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Enrollment sheet
        print("  Checking Enrollment...")
        a_rows = load_sheet_rows(agent_wb, "Enrollment")
        g_rows = load_sheet_rows(gt_wb, "Enrollment")
        if a_rows is None:
            file_errors.append("Sheet 'Enrollment' not found in agent output")
            critical_failures.append("Sheet 'Enrollment' missing")
        elif g_rows is None:
            file_errors.append("Sheet 'Enrollment' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row

            # CRITICAL: every GT course must be present with a tight Student_Count match.
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing row: {g_row[0]}")
                    critical_failures.append(f"Course missing in Enrollment: {g_row[0]}")
                    continue

                # Student_Count (index 2) — TIGHT tolerance (CRITICAL)
                if len(a_row) > 2 and len(g_row) > 2:
                    if not num_close(a_row[2], g_row[2], 2):
                        file_errors.append(f"{key}.Student_Count: {a_row[2]} vs {g_row[2]}")
                        critical_failures.append(f"{key}.Student_Count off: {a_row[2]} vs {g_row[2]}")
                # Teacher_Count (index 3)
                if len(a_row) > 3 and len(g_row) > 3:
                    if not num_close(a_row[3], g_row[3], 1):
                        file_errors.append(f"{key}.Teacher_Count: {a_row[3]} vs {g_row[3]}")
                # TA_Count (index 4)
                if len(a_row) > 4 and len(g_row) > 4:
                    if not num_close(a_row[4], g_row[4], 1):
                        file_errors.append(f"{key}.TA_Count: {a_row[4]} vs {g_row[4]}")
                # Student_Teacher_Ratio (index 5) — validate against GT (non-critical)
                if len(a_row) > 5 and len(g_row) > 5:
                    if not num_close(a_row[5], g_row[5], 1.0):
                        file_errors.append(f"{key}.Student_Teacher_Ratio: {a_row[5]} vs {g_row[5]}")

        # Check Summary sheet
        print("  Checking Summary...")
        a_rows = load_sheet_rows(agent_wb, "Summary")
        g_rows = load_sheet_rows(gt_wb, "Summary")
        if a_rows is None:
            file_errors.append("Sheet 'Summary' not found in agent output")
            critical_failures.append("Sheet 'Summary' missing")
        elif g_rows is None:
            file_errors.append("Sheet 'Summary' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().lower()] = row
            # Metrics that must match exactly/tightly (CRITICAL)
            critical_metrics = {
                "total_courses", "total_students", "avg_students_per_course",
                "smallest_course", "largest_course",
            }
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    file_errors.append(f"Missing summary row: {g_row[0]}")
                    if key in critical_metrics:
                        critical_failures.append(f"Summary metric missing: {g_row[0]}")
                    continue
                if len(a_row) > 1 and len(g_row) > 1:
                    if isinstance(g_row[1], (int, float)):
                        # TIGHT numeric tolerance for summary values
                        if not num_close(a_row[1], g_row[1], 1):
                            file_errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
                            if key in critical_metrics:
                                critical_failures.append(f"{key}.Value off: {a_row[1]} vs {g_row[1]}")
                    else:
                        if not str_match(a_row[1], g_row[1]):
                            file_errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
                            if key in critical_metrics:
                                critical_failures.append(f"{key}.Value off: {a_row[1]} vs {g_row[1]}")

    # Check email (DB check) — now ENFORCED (critical)
    print("  Checking email...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute(
            "SELECT m.subject, m.body_text FROM email.messages m "
            "JOIN email.folders f ON f.id = m.folder_id WHERE LOWER(f.name) = 'sent'"
        )
        rows = cur.fetchall()
        subj_ok = False
        body_ok = False
        for subj, body in rows:
            s = (subj or "").lower()
            if ("enrollment" in s or "spring 2014" in s or "зачислен" in s):
                subj_ok = True
                b = normalize_ru_numbers((body or "").lower())
                # Body should reference the total student count (7804) and total course count (6).
                if "7804" in b and "6" in b:
                    body_ok = True
        if not subj_ok:
            db_errors.append("No email about enrollment/Spring 2014 found in sent folder")
            critical_failures.append("Required summary email not sent")
        elif not body_ok:
            db_errors.append("Email body does not mention total students (7804) and total courses (6)")
            critical_failures.append("Email body missing required totals (7804 students, 6 courses)")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Email check: {e}")
        critical_failures.append(f"Email check error: {e}")

    # Check calendar event (DB check) — now ENFORCED with date/time assertion (critical)
    print("  Checking calendar event...")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute(
            "SELECT summary, start_datetime, end_datetime FROM gcal.events "
            "WHERE LOWER(summary) LIKE '%enrollment%review%' OR LOWER(summary) LIKE '%spring 2014%'"
        )
        events = cur.fetchall()
        if not events:
            db_errors.append("No calendar event with 'Enrollment Review' found")
            critical_failures.append("Required calendar event not created")
        else:
            time_ok = False
            for summary, start_dt, end_dt in events:
                start_s = str(start_dt) if start_dt is not None else ""
                # Required: 2026-03-14, 14:00 UTC start
                if "2026-03-14" in start_s and "14:00" in start_s:
                    time_ok = True
            if not time_ok:
                db_errors.append("Calendar event has wrong date/time (expected 2026-03-14 14:00 UTC)")
                critical_failures.append("Calendar event date/time incorrect")
        cur.close()
        conn.close()
    except Exception as e:
        db_errors.append(f"Calendar check: {e}")
        critical_failures.append(f"Calendar check error: {e}")

    # Final result
    print(f"\n=== SUMMARY ===")
    print(f"  File errors: {len(file_errors)}")
    print(f"  DB errors:   {len(db_errors)}")
    print(f"  Critical failures: {len(critical_failures)}")
    if db_errors:
        for e in db_errors[:10]:
            print(f"    [DB] {e}")
    if file_errors:
        for e in file_errors[:10]:
            print(f"    [FILE] {e}")

    # CRITICAL gate: any critical failure => hard FAIL before any accuracy consideration.
    if critical_failures:
        print("  Critical checks FAILED:")
        for e in critical_failures[:10]:
            print(f"    [CRITICAL] {e}")
        print(f"  Overall: FAIL")
        sys.exit(1)

    # Accuracy gate: structural/non-critical issues. Keep threshold >= 70%.
    total_checks = 7  # 6 courses + summary block as a coarse structural proxy
    accuracy = 100.0 if not file_errors else max(0.0, 100.0 * (1 - len(file_errors) / max(total_checks, len(file_errors))))
    print(f"  Accuracy: {accuracy:.1f}%")
    if accuracy >= 70:
        print(f"  Overall: PASS")
        sys.exit(0)
    else:
        print(f"  Overall: FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
