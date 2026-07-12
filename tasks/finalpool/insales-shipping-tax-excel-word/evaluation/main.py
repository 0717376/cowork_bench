"""Evaluation for insales-shipping-tax-excel-word."""
import os
import argparse, os, sys
import psycopg2


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol=1.0):
    try: return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError): return False


def _norm(v):
    # Слаги вида 'reduced-rate' эквивалентны 'Reduced Rate'
    return str(v).strip().lower().replace("-", " ")


def str_match(a, b):
    if a is None or b is None: return a is None and b is None
    return _norm(a) == _norm(b)


def check_excel(agent_workspace, groundtruth_workspace="."):
    errors = []
    import openpyxl
    path = os.path.join(agent_workspace, "Shipping_Tax_Analysis.xlsx")
    if not os.path.exists(path):
        return ["Shipping_Tax_Analysis.xlsx not found"]
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        rows1 = load_sheet_rows(wb, "Shipping Zones")
        if rows1 is None:
            errors.append("Sheet 'Shipping Zones' not found")
        else:
            data_rows = [r for r in rows1[1:] if r and r[0] is not None]
            if len(data_rows) < 3:
                errors.append(f"Shipping Zones has {len(data_rows)} rows, expected at least 3")
            zone_names = {str(r[0]).strip().lower() for r in data_rows}
            # Зоны русифицированы централизованным wc-сидом (db/zzz_wc_after_init.sql
            # через scripts/wc_relabel_map.py): Domestic US -> Доставка по РФ,
            # California -> Москва, International -> Международная. Принимаем RU и EN.
            for expected_alts in (
                ["доставка по рф", "domestic us"],
                ["москва", "california"],
                ["международная", "international"],
            ):
                if not any(alt in z for z in zone_names for alt in expected_alts):
                    errors.append(f"Expected zone (one of {expected_alts}) not found in Shipping Zones")

        rows2 = load_sheet_rows(wb, "Tax Rates")
        if rows2 is None:
            errors.append("Sheet 'Tax Rates' not found")
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 6:
                errors.append(f"Tax Rates has {len(data_rows2)} rows, expected at least 6")

        # --- Groundtruth XLSX value comparison ---
        gt_path = os.path.join(groundtruth_workspace, "Shipping_Tax_Analysis.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in wb.sheetnames:
                    if asn.strip().lower() == gt_sname.strip().lower():
                        a_ws = wb[asn]; break
                if a_ws is None:
                    errors.append(f"GT sheet '{gt_sname}' not found in agent xlsx (available: {wb.sheetnames})")
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
                # Порядок строк в GT произволен (id/locale-зависим) — сравниваем
                # нечувствительно к порядку, сортируя обе стороны по каноническому ключу.
                def _row_key(r):
                    return tuple(_norm(c) if c is not None else "" for c in r)
                gt_rows = sorted(gt_rows, key=_row_key)
                a_rows = sorted(a_rows, key=_row_key)
                if len(a_rows) != len(gt_rows):
                    errors.append(f"GT '{gt_sname}' row count: expected {len(gt_rows)}, got {len(a_rows)}")
                for ri in range(min(3, len(gt_rows))):
                    if ri >= len(a_rows): break
                    for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                        gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                        if gv is None: continue
                        if isinstance(gv, (int, float)):
                            if not num_close(av, gv, max(abs(gv)*0.1, 1.0)):
                                errors.append(f"GT '{gt_sname}' row {ri+1} col {ci+1}: expected {gv}, got {av}")
                                break
                        else:
                            if not str_match(av, gv):
                                errors.append(f"GT '{gt_sname}' row {ri+1} col {ci+1}: expected '{gv}', got '{av}'")
                                break
            gt_wb.close()
    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
    return errors


def check_word_doc(agent_workspace):
    errors = []
    doc_path = os.path.join(agent_workspace, "Operations_Report.docx")
    if not os.path.exists(doc_path):
        return ["Operations_Report.docx not found"]
    try:
        from docx import Document
        doc = Document(doc_path)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
        # Текст может быть на русском или английском.
        if not any(k in full_text for k in ("shipping", "достав")):
            errors.append("Word doc does not mention shipping (shipping/достав)")
        if not any(k in full_text for k in ("tax", "налог")):
            errors.append("Word doc does not mention tax (tax/налог)")
        if len(full_text.strip()) < 100:
            errors.append("Word doc content is too short (less than 100 chars)")
    except Exception as e:
        errors.append(f"Error reading Word doc: {e}")
    return errors


def check_email():
    errors = []
    try:
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%operations@store.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close(); conn.close()
        if not rows:
            errors.append("No email found to operations@store.com")
        else:
            subj_ok = any(r[0] and "shipping and tax configuration report" in str(r[0]).lower() for r in rows)
            if not subj_ok:
                errors.append("Email subject 'Shipping and Tax Configuration Report' not found")
            body_ok = any(r[2] and len(str(r[2]).strip()) >= 50 for r in rows)
            if not body_ok:
                errors.append("Email to operations@store.com has no substantive body (>=50 chars)")
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    return errors


def run_critical_checks(agent_ws, gt_ws):
    """Семантические критические проверки. Любой провал => немедленный FAIL.
    Зоны/значения сверяются с замороженным groundtruth_workspace (русифицирован
    централизованным wc-сидом). Имена налоговых классов и коды штатов (Standard/
    Reduced Rate/Zero Rate, CA/FL/NY/TX/WA) НЕ русифицируются картой — остаются EN."""
    import openpyxl
    crit = []
    # --- Excel ---
    xlsx = os.path.join(agent_ws, "Shipping_Tax_Analysis.xlsx")
    if not os.path.exists(xlsx):
        crit.append("CRITICAL: Shipping_Tax_Analysis.xlsx отсутствует")
        return crit
    try:
        wb = openpyxl.load_workbook(xlsx, data_only=True)
    except Exception as e:
        crit.append(f"CRITICAL: не удалось открыть Excel: {e}")
        return crit

    # 1) Лист Shipping Zones содержит все 3 русифицированные зоны (RU или EN).
    rows1 = load_sheet_rows(wb, "Shipping Zones")
    if not rows1:
        crit.append("CRITICAL: лист 'Shipping Zones' отсутствует")
    else:
        drows = [r for r in rows1[1:] if r and r[0] is not None]
        znames = {str(r[0]).strip().lower() for r in drows}
        for alts in (["доставка по рф", "domestic us"],
                     ["москва", "california"],
                     ["международная", "international"]):
            if not any(alt in z for z in znames for alt in alts):
                crit.append(f"CRITICAL: зона (одна из {alts}) отсутствует на листе Shipping Zones")
        # 5) Methods_Count по зоне совпадает с groundtruth (Москва=2, Доставка по РФ=3,
        # Международная=1) в пределах допуска.
        gt_counts = {("москва", "california"): 2,
                     ("доставка по рф", "domestic us"): 3,
                     ("международная", "international"): 1}
        # столбец Methods_Count — индекс 1
        zone_to_count = {}
        for r in drows:
            try:
                zone_to_count[str(r[0]).strip().lower()] = r[1]
            except IndexError:
                pass
        for alts, expected in gt_counts.items():
            matched = None
            for zn, cnt in zone_to_count.items():
                if any(a in zn for a in alts):
                    matched = cnt; break
            if matched is None:
                crit.append(f"CRITICAL: не найдена зона {alts} для проверки Methods_Count")
            elif not num_close(matched, expected, 0.5):
                crit.append(f"CRITICAL: Methods_Count для {alts}: ожидалось {expected}, получено {matched}")

    # 2) Лист Tax Rates: >=6 строк И присутствуют краевые записи
    # (Zero Rate с пустыми state+rate; Reduced Rate).
    rows2 = load_sheet_rows(wb, "Tax Rates")
    if not rows2:
        crit.append("CRITICAL: лист 'Tax Rates' отсутствует")
    else:
        drows2 = [r for r in rows2[1:] if r and r[0] is not None]
        if len(drows2) < 6:
            crit.append(f"CRITICAL: Tax Rates содержит {len(drows2)} строк, ожидалось >=6")
        # woo_tax_rates_list возвращает слаги ('reduced-rate') — нормализуем дефисы
        classes = {_norm(r[0]) for r in drows2}
        if "reduced rate" not in classes:
            crit.append("CRITICAL: отсутствует налоговый класс 'Reduced Rate'")
        # краевая запись Zero Rate с пустыми state и rate
        zero_edge = any(
            _norm(r[0]) == "zero rate"
            and (len(r) < 2 or r[1] in (None, "", "None"))
            and (len(r) < 3 or r[2] in (None, "", "None"))
            for r in drows2
        )
        if not zero_edge:
            crit.append("CRITICAL: отсутствует краевая запись 'Zero Rate' с пустыми State и Rate")

    # --- Email: тема + содержательное тело ---
    eerr = check_email()
    crit.extend("CRITICAL: " + e for e in eerr)

    # --- Word: упоминание доставки И налогов, существенный объём ---
    werr = check_word_doc(agent_ws)
    crit.extend("CRITICAL: " + e for e in werr)

    return crit


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(os.path.dirname(__file__), "..", "groundtruth_workspace")

    # === КРИТИЧЕСКИЕ СЕМАНТИЧЕСКИЕ ПРОВЕРКИ (любой провал => немедленный FAIL) ===
    print("  Running CRITICAL checks...")
    crit_errors = run_critical_checks(agent_ws, gt_ws)
    if crit_errors:
        print(f"\n=== RESULT: FAIL (critical) ===")
        for e in crit_errors[:10]: print(f"  {e}")
        sys.exit(1)
    print("    CRITICAL PASS")

    # === СТРУКТУРНЫЕ ПРОВЕРКИ (порог точности >= 70%) ===
    checks = []  # (name, errors)

    print("  Checking Excel file...")
    errs = check_excel(agent_ws, gt_ws)
    checks.append(("excel", errs))
    for e in errs[:3]: print(f"    ERROR: {e}")
    if not errs: print("    PASS")

    print("  Checking Word document...")
    errs = check_word_doc(agent_ws)
    checks.append(("word", errs))
    for e in errs[:3]: print(f"    ERROR: {e}")
    if not errs: print("    PASS")

    print("  Checking email...")
    errs = check_email()
    checks.append(("email", errs))
    for e in errs[:3]: print(f"    ERROR: {e}")
    if not errs: print("    PASS")

    passed = sum(1 for _, e in checks if not e)
    total = len(checks)
    accuracy = 100.0 * passed / total if total else 0.0
    all_errors = [e for _, errs in checks for e in errs]
    print(f"\n  Accuracy: {passed}/{total} = {accuracy:.1f}%")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]: print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
