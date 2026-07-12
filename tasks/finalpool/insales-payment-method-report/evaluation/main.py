"""Evaluation for insales-payment-method-report."""
import argparse
import os
import re
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Payment_Method_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Payment_Method_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_failures = []

    # --- Build groundtruth lookups for critical checks ---
    def rows_to_lookup(wb, sheet):
        rows = load_sheet_rows(wb, sheet)
        if not rows or len(rows) < 2:
            return {}
        return {str(r[0]).strip().lower(): r for r in rows[1:] if r and r[0] is not None}

    gt_pm = rows_to_lookup(gt_wb, "Payment Methods")   # russified payment titles -> row
    gt_sum = rows_to_lookup(gt_wb, "Summary")
    a_pm = rows_to_lookup(agent_wb, "Payment Methods")
    a_sum = rows_to_lookup(agent_wb, "Summary")

    def gt_summary_val(key):
        r = gt_sum.get(key.lower())
        return r[1] if r else None

    def a_summary_val(key):
        r = a_sum.get(key.lower())
        return r[1] if r else None

    # --- Check Payment Methods sheet ---
    print("  Checking Payment Methods sheet...")
    a_rows = load_sheet_rows(agent_wb, "Payment Methods")
    g_rows = load_sheet_rows(gt_wb, "Payment Methods")
    if a_rows is None:
        all_errors.append("Sheet 'Payment Methods' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Payment Methods' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row: {g_row[0]}")
                continue

            # Order_Count (col 1) — exact (deterministic seed)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 0):
                    all_errors.append(f"{key}.Order_Count: {a_row[1]} vs {g_row[1]}")

            # Total_Revenue (col 2) — tightened tol
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 2.0):
                    all_errors.append(f"{key}.Total_Revenue: {a_row[2]} vs {g_row[2]}")

            # Avg_Order_Value (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5.0):
                    all_errors.append(f"{key}.Avg_Order_Value: {a_row[3]} vs {g_row[3]}")

        if not [e for e in all_errors if "Payment Methods" in e or "Missing row" in e]:
            print("    PASS")

    # --- Check Summary sheet ---
    print("  Checking Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Summary: {g_row[0]}")
                continue

            g_val = g_row[1]
            a_val = a_row[1]

            try:
                float(a_val); float(g_val)
                if not num_close(a_val, g_val, 50.0):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val} (tol=50.0)")
            except (TypeError, ValueError):
                if not str_match(a_val, g_val):
                    all_errors.append(f"Summary.{key}: {a_val} vs {g_val}")

        if not [e for e in all_errors if "Summary" in e]:
            print("    PASS")

    # --- Check email sent ---
    print("  Checking email...")
    email_rows = None
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        cur.execute("SELECT m.subject, m.to_addr, m.body_text FROM email.messages m")
        email_rows = cur.fetchall()

        found_email = False
        for subj, to_addr, body in email_rows:
            subj_str = str(subj or "").lower()
            to_str = str(to_addr or "").lower()
            if "payment" in subj_str and "finance-lead" in to_str:
                found_email = True
                break

        if not found_email:
            all_errors.append("No email with 'Payment' in subject sent to finance-lead@company.com")
        else:
            print("    PASS")

        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Email check error: {e}")

    # --- CRITICAL CHECKS (semantic; derived from groundtruth, not hardcoded) ---
    print("  Running CRITICAL checks...")

    # C1: Most_Used_Method matches groundtruth russified title (exact str_match).
    gt_most = gt_summary_val("Most_Used_Method")
    a_most = a_summary_val("Most_Used_Method")
    if gt_most is not None and not str_match(a_most, gt_most):
        critical_failures.append(f"CRITICAL Most_Used_Method: {a_most!r} != {gt_most!r}")

    # C2: Highest_Revenue_Method matches groundtruth (exact str_match).
    gt_high = gt_summary_val("Highest_Revenue_Method")
    a_high = a_summary_val("Highest_Revenue_Method")
    if gt_high is not None and not str_match(a_high, gt_high):
        critical_failures.append(f"CRITICAL Highest_Revenue_Method: {a_high!r} != {gt_high!r}")

    # C3: Total_Orders and Number_Of_Methods exact integer match (tol=0).
    gt_tot = gt_summary_val("Total_Orders")
    a_tot = a_summary_val("Total_Orders")
    if gt_tot is not None and not num_close(a_tot, gt_tot, 0):
        critical_failures.append(f"CRITICAL Total_Orders: {a_tot} != {gt_tot}")
    gt_nm = gt_summary_val("Number_Of_Methods")
    a_nm = a_summary_val("Number_Of_Methods")
    if gt_nm is not None and not num_close(a_nm, gt_nm, 0):
        critical_failures.append(f"CRITICAL Number_Of_Methods: {a_nm} != {gt_nm}")

    # C4: All groundtruth payment-method rows present with exact Order_Count and tight Total_Revenue.
    for key, g_row in gt_pm.items():
        a_row = a_pm.get(key)
        if a_row is None:
            critical_failures.append(f"CRITICAL missing payment-method row: {g_row[0]!r}")
            continue
        if len(a_row) > 1 and not num_close(a_row[1], g_row[1], 0):
            critical_failures.append(f"CRITICAL {key} Order_Count: {a_row[1]} != {g_row[1]}")
        if len(a_row) > 2 and not num_close(a_row[2], g_row[2], 2.0):
            critical_failures.append(f"CRITICAL {key} Total_Revenue: {a_row[2]} != {g_row[2]}")

    # C5: Email to finance-lead with 'Payment' subject AND body contains most-used method + total revenue.
    # Derive expected substrings from groundtruth (no per-seed hardcoding).
    gt_total_rev = gt_summary_val("Total_Revenue")
    def revenue_substrings(val):
        # integer part of the revenue as a tolerant body substring (e.g. 61712 from 61712.04)
        try:
            return str(int(float(val)))
        except (TypeError, ValueError):
            return None
    rev_sub = revenue_substrings(gt_total_rev)
    # most-used method substring: use a salient token (the part before any parenthesis), lowercased
    most_token = None
    if gt_most is not None:
        most_token = str(gt_most).split("(")[0].strip().lower()

    if email_rows is None:
        critical_failures.append("CRITICAL email body check: could not read email.messages")
    else:
        body_ok = False
        for subj, to_addr, body in email_rows:
            subj_str = str(subj or "").lower()
            to_str = str(to_addr or "").lower()
            body_str = str(body or "").lower()
            if "payment" not in subj_str or "finance-lead" not in to_str:
                continue
            # Normalize away locale thousands separators (regular space, non-breaking
            #  , narrow no-break  ) so a Russian-rendered total like
            # "61 712,04 RUB" still matches the integer-part digit run "61712".
            body_digits = re.sub(r"[\s  ]", "", body_str)
            method_ok = (most_token is None) or (most_token in body_str)
            rev_ok = (rev_sub is None) or (rev_sub in body_str) or (rev_sub in body_digits)
            if method_ok and rev_ok:
                body_ok = True
                break
        if not body_ok:
            critical_failures.append(
                f"CRITICAL email body must mention most-used method ({most_token!r}) "
                f"and total revenue ({rev_sub!r})"
            )

    if critical_failures:
        print(f"\n=== CRITICAL FAILURES ({len(critical_failures)}) ===")
        for e in critical_failures:
            print(f"  {e}")
        print("=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    print("    CRITICAL checks PASS")

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
