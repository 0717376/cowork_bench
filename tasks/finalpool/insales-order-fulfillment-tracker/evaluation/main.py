"""Evaluation for insales-order-fulfillment-tracker."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Fulfillment_Report.xlsx")
    gt_file = os.path.join(gt_dir, "Fulfillment_Report.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_errors = []

    # --- Check Order Status sheet ---
    print("  Checking Order Status sheet...")
    a_rows = load_sheet_rows(agent_wb, "Order Status")
    g_rows = load_sheet_rows(gt_wb, "Order Status")
    if a_rows is None:
        all_errors.append("Sheet 'Order Status' not found in agent output")
        critical_errors.append("Sheet 'Order Status' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Order Status' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                # CRITICAL: status labels must be lowercase as required
                raw_status = str(row[0]).strip()
                if raw_status != raw_status.lower():
                    critical_errors.append(f"Order Status label not lowercase: {raw_status!r}")
                a_lookup[raw_status.lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing status: {g_row[0]}")
                # CRITICAL: all 7 status rows must be present
                critical_errors.append(f"Missing status row: {g_row[0]}")
                continue

            # Order_Count (col 1) - CRITICAL exact match (tol 0)
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 2):
                    all_errors.append(f"{key}.Order_Count: {a_row[1]} vs {g_row[1]}")
                if not num_close(a_row[1], g_row[1], 0):
                    critical_errors.append(f"{key}.Order_Count exact: {a_row[1]} vs {g_row[1]}")

            # Total_Revenue (col 2) - CRITICAL tight tolerance
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 50.0):
                    all_errors.append(f"{key}.Total_Revenue: {a_row[2]} vs {g_row[2]}")
                if not num_close(a_row[2], g_row[2], 1.0):
                    critical_errors.append(f"{key}.Total_Revenue tight: {a_row[2]} vs {g_row[2]}")

            # Avg_Order_Value (col 3)
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 5.0):
                    all_errors.append(f"{key}.Avg_Order_Value: {a_row[3]} vs {g_row[3]}")

        if not [e for e in all_errors if "Order Status" in e or "Missing status" in e]:
            print("    PASS")

    # --- Check Fulfillment Summary sheet ---
    print("  Checking Fulfillment Summary sheet...")
    a_rows = load_sheet_rows(agent_wb, "Fulfillment Summary")
    g_rows = load_sheet_rows(gt_wb, "Fulfillment Summary")
    # CRITICAL summary metrics: name -> (expected, critical_tolerance)
    CRIT_SUMMARY = {
        "total_orders": (150, 0),
        "completed_orders": (72, 0),
        "fulfillment_rate": (48.0, 0.5),
        "pending_processing_count": (56, 0),
        "cancellation_count": (13, 0),
        "cancellation_rate": (8.67, 0.5),
    }
    if a_rows is None:
        all_errors.append("Sheet 'Fulfillment Summary' not found in agent output")
        critical_errors.append("Sheet 'Fulfillment Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Fulfillment Summary' not found in groundtruth")
    else:
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row

        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                all_errors.append(f"Missing row in Fulfillment Summary: {g_row[0]}")
                if key in CRIT_SUMMARY:
                    critical_errors.append(f"Missing critical Fulfillment Summary row: {g_row[0]}")
                continue

            g_val = g_row[1]
            a_val = a_row[1]

            try:
                float(a_val); float(g_val)
                if not num_close(a_val, g_val, 2.0):
                    all_errors.append(f"Fulfillment Summary.{key}: {a_val} vs {g_val}")
                if key in CRIT_SUMMARY:
                    exp, ctol = CRIT_SUMMARY[key]
                    if not num_close(a_val, exp, ctol):
                        critical_errors.append(f"Fulfillment Summary.{key} (critical): {a_val} vs {exp} (tol {ctol})")
            except (TypeError, ValueError):
                if not str_match(a_val, g_val):
                    all_errors.append(f"Fulfillment Summary.{key}: {a_val} vs {g_val}")
                if key in CRIT_SUMMARY:
                    critical_errors.append(f"Fulfillment Summary.{key} (critical): non-numeric {a_val!r}")

        if not [e for e in all_errors if "Fulfillment Summary" in e]:
            print("    PASS")

    # --- Check Google Sheet ---
    print("  Checking Google Sheet...")
    try:
        import psycopg2
        conn = psycopg2.connect(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym",
                                user="eigent", password="camel")
        cur = conn.cursor()
        # Require the specific dashboard title, not any title containing 'order'
        cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE LOWER(title) LIKE '%fulfillment dashboard%' OR LOWER(title) LIKE '%order fulfillment%'")
        sheets = cur.fetchall()
        if not sheets:
            all_errors.append("No Google Sheet titled 'Order Fulfillment Dashboard' found")
            critical_errors.append("No Google Sheet titled 'Order Fulfillment Dashboard' found")
        else:
            sid = sheets[0][0]
            cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
            cell_count = cur.fetchone()[0]
            if cell_count < 20:
                all_errors.append(f"Google Sheet has too few cells: {cell_count}")

            # Collect all cell values (lowercased strings) for content assertions
            cur.execute("SELECT value FROM gsheet.cells WHERE spreadsheet_id = %s", (sid,))
            cell_vals = [r[0] for r in cur.fetchall() if r[0] is not None]
            cell_str_set = set(str(v).strip().lower() for v in cell_vals)

            # CRITICAL: all 7 status labels must actually appear in the gsheet
            statuses = ["completed", "processing", "pending", "on-hold",
                        "cancelled", "failed", "refunded"]
            missing_status = [s for s in statuses if s not in cell_str_set]
            if missing_status:
                msg = f"Google Sheet missing status labels: {missing_status}"
                all_errors.append(msg)
                critical_errors.append(msg)

            # CRITICAL: per-status Order_Count values must be present in the gsheet
            gt_counts = {"completed": 72, "processing": 28, "pending": 13,
                         "on-hold": 15, "cancelled": 8, "failed": 5, "refunded": 9}
            numeric_cells = set()
            for v in cell_vals:
                try:
                    numeric_cells.add(int(float(str(v).strip())))
                except (TypeError, ValueError):
                    pass
            missing_counts = [k for k, c in gt_counts.items() if c not in numeric_cells]
            if missing_counts:
                msg = f"Google Sheet missing Order_Count values for: {missing_counts}"
                all_errors.append(msg)
                critical_errors.append(msg)

            if not missing_status and not missing_counts and cell_count >= 20:
                print("    PASS")
        cur.close()
        conn.close()
    except Exception as e:
        all_errors.append(f"Google Sheet check error: {e}")

    # --- Critical gate ---
    if critical_errors:
        print(f"\n=== CRITICAL FAILURE ({len(critical_errors)} critical errors) ===")
        for e in critical_errors[:10]:
            print(f"  CRITICAL: {e}")
        sys.exit(1)

    # --- Final result ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
