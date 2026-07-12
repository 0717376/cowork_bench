"""Evaluation для terminal-insales-excel-gsheet-gcal-email.

Проверяем:
1. Inventory_Lifecycle_Report.xlsx с 4 листами (структура + СЕМАНТИКА).
2. Google Sheet "Inventory Dashboard".
3. Событие в календаре по пересмотру пополнения.
4. ОТПРАВЛЕННОЕ письмо на purchasing@company.com.
5. demand_forecast.py существует.

Семантические CRITICAL-чеки сверяют вывод агента с замороженным groundtruth-файлом
(groundtruth_workspace/Inventory_Lifecycle_Report.xlsx), который централизованно
пропатчен под русификацию wc-данных (scripts/wc_relabel_map.py) и потому всегда
синхронен с живой схемой wc.*. Ключ сопоставления строк — product_name (frozen English,
уникален). Это валидирует расчёты demand_forecast (rate, days_of_supply, reorder_point,
urgency, quantity) без хардкода волатильных значений в eval.

Категории в groundtruth регенерированы из ЖИВОЙ wc.products.categories[0].name
(русифицированы по scripts/wc_relabel_map.py): 'Аудио', 'Камеры', 'Электроника',
'Бытовая техника', 'Часы', 'ТВ и домашний кинотеатр' — 6 категорий. total_value
= sum(price * stock_quantity) по той же таксономии. Для устойчивости (Электроника/
Аудио/Камеры) дополнительно принимаем RU+EN-синонимы.
"""
import argparse
import json
import math
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Чеки, провал которых означает содержательное невыполнение задачи.
# Любой такой FAIL => итог FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Category_Summary: набор категорий и total_value совпадают с эталоном",
    "Restock_Schedule: reorder_date/quantity/supplier совпадают с эталоном",
    "Reorder_Alerts: только Critical/Out_of_Stock, Out_of_Stock сортируются первыми",
    "Product_Inventory: строки по товарам, сортировка по days_of_supply, значения совпадают с эталоном",
    "Email: реальное письмо на purchasing@company.com с корректными счётчиками",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}: {str(detail)[:200]}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


# Принимаемые синонимы категорий (RU из карты + EN-оригинал + замороженные).
CATEGORY_SYNONYMS = {
    "электроника": {"электроника", "electronics"},
    "аудио": {"аудио", "audio"},
    "камеры": {"камеры", "cameras"},
}


def _norm_cat(c):
    return str(c).strip().lower() if c is not None else ""


def _cat_match(a, b):
    """Совпадение категорий с учётом RU/EN синонимов."""
    na, nb = _norm_cat(a), _norm_cat(b)
    if na == nb:
        return True
    for syn in CATEGORY_SYNONYMS.values():
        if na in syn and nb in syn:
            return True
    return False


def _load_sheet(wb, *keywords_groups):
    """Найти лист, чьё имя (lower, пробел->_) содержит все ключевые слова одной из групп."""
    sheets = wb.sheetnames
    sl = [s.lower().replace(" ", "_") for s in sheets]
    for kws in keywords_groups:
        for i, s in enumerate(sl):
            if all(k in s for k in kws):
                return wb[sheets[i]]
    return None


def _rows(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = rows[0]
    data = [r for r in rows[1:] if any(c is not None and str(c) != "" for c in r)]
    return header, data


# ----------------------------------------------------------------------------
# STRUCTURAL (non-critical) Excel checks
# ----------------------------------------------------------------------------
def check_excel_structure(workspace):
    print("\n=== Check 1: Inventory_Lifecycle_Report.xlsx (структура) ===")
    path = os.path.join(workspace, "Inventory_Lifecycle_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return None
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = wb.sheetnames
    check("Has at least 4 sheets", len(sheets) >= 4, f"Found {len(sheets)}: {sheets}")

    ws1 = _load_sheet(wb, ("product", "inventory"))
    if ws1 is not None:
        h1, d1 = _rows(ws1)
        check("Product_Inventory has 70+ product rows", len(d1) >= 70, f"Found {len(d1)}")
        headers = [str(c).lower() if c else "" for c in h1]
        check("Has days_of_supply column", any("days" in h and "supply" in h for h in headers),
              f"Headers: {h1}")
        check("Has total_sales column", any("sales" in h for h in headers), f"Headers: {h1}")

    ws2 = _load_sheet(wb, ("reorder",), ("alert",))
    if ws2 is not None:
        h2, d2 = _rows(ws2)
        check("Reorder_Alerts has entries", len(d2) >= 5, f"Found {len(d2)}")

    ws3 = _load_sheet(wb, ("category",), ("summary",))
    if ws3 is not None:
        h3, d3 = _rows(ws3)
        check("Category_Summary has category rows", len(d3) >= 3, f"Found {len(d3)}")
        all_text3 = " ".join(str(c) for r in d3 for c in r if c is not None).lower()
        # Broaden RU+EN per central relabel map: Электроника/Electronics
        check("Has Electronics category (RU/EN)",
              "электроника" in all_text3 or "electronics" in all_text3, f"Text: {all_text3[:120]}")
        # 'TV' category stays frozen English
        check("Has TV category", "tv" in all_text3, f"Text: {all_text3[:120]}")

    ws4 = _load_sheet(wb, ("restock",), ("schedule",))
    if ws4 is not None:
        h4, d4 = _rows(ws4)
        check("Restock_Schedule has entries", len(d4) >= 5, f"Found {len(d4)}")
        all_text4 = " ".join(str(c) for r in d4 for c in r if c is not None).lower()
        check("Has supplier info", "supplier" in all_text4 or "primary" in all_text4,
              f"Text: {all_text4[:120]}")

    return wb


# ----------------------------------------------------------------------------
# CRITICAL semantic checks — сверка с эталоном (groundtruth)
# ----------------------------------------------------------------------------
def _index_by_name(ws, name_col=0):
    _, data = _rows(ws)
    idx = {}
    for r in data:
        key = str(r[name_col]).strip() if r and r[name_col] is not None else None
        if key:
            idx[key] = r
    return idx, data


def check_excel_semantics(agent_wb, gt_path):
    print("\n=== Check 1b: Inventory_Lifecycle_Report.xlsx (семантика, сверка с эталоном) ===")
    if agent_wb is None:
        # все критические Excel-чеки провалены — файл отсутствует
        for nm in ["Category_Summary: набор категорий и total_value совпадают с эталоном",
                   "Restock_Schedule: reorder_date/quantity/supplier совпадают с эталоном",
                   "Reorder_Alerts: только Critical/Out_of_Stock, Out_of_Stock сортируются первыми",
                   "Product_Inventory: строки по товарам, сортировка по days_of_supply, значения совпадают с эталоном"]:
            check(nm, False, "agent Excel отсутствует")
        return
    if not os.path.exists(gt_path):
        check("Groundtruth Excel доступен", False, f"Not found at {gt_path}")
        return
    gt = openpyxl.load_workbook(gt_path, data_only=True)

    # --- Product_Inventory: сортировка + значения days_of_supply ---
    a_pi = _load_sheet(agent_wb, ("product", "inventory"))
    g_pi = _load_sheet(gt, ("product", "inventory"))
    if a_pi is not None and g_pi is not None:
        a_idx, a_data = _index_by_name(a_pi)
        g_idx, _ = _index_by_name(g_pi)
        # columns by header
        a_h, _ = _rows(a_pi)
        ah = [str(c).lower() if c else "" for c in a_h]
        try:
            dos_col = next(i for i, h in enumerate(ah) if "days" in h and "supply" in h)
        except StopIteration:
            dos_col = 6
        dos_seq = []
        for r in a_data:
            try:
                dos_seq.append(float(r[dos_col]))
            except (TypeError, ValueError):
                dos_seq.append(float("inf"))
        sorted_ok = all(dos_seq[i] <= dos_seq[i + 1] + 1e-6 for i in range(len(dos_seq) - 1))
        # значения days_of_supply совпадают с эталоном по выборке товаров
        g_h, _ = _rows(g_pi)
        gh = [str(c).lower() if c else "" for c in g_h]
        try:
            g_dos_col = next(i for i, h in enumerate(gh) if "days" in h and "supply" in h)
        except StopIteration:
            g_dos_col = 6
        common = [k for k in a_idx if k in g_idx]
        val_ok = len(common) >= 70
        mism = 0
        for k in common:
            try:
                av = float(a_idx[k][dos_col]); gv = float(g_idx[k][g_dos_col])
            except (TypeError, ValueError):
                mism += 1; continue
            tol = max(1.0, abs(gv) * 0.02)
            if abs(av - gv) > tol:
                mism += 1
        val_ok = val_ok and mism <= max(2, int(len(common) * 0.05))
        check("Product_Inventory: строки по товарам, сортировка по days_of_supply, значения совпадают с эталоном",
              sorted_ok and val_ok and len(a_data) >= 70,
              f"rows={len(a_data)} sorted={sorted_ok} common={len(common)} mismatch={mism}")
    else:
        check("Product_Inventory: строки по товарам, сортировка по days_of_supply, значения совпадают с эталоном",
              False, "лист Product_Inventory не найден")

    # --- Category_Summary: набор категорий + total_value ---
    a_cs = _load_sheet(agent_wb, ("category",), ("summary",))
    g_cs = _load_sheet(gt, ("category",), ("summary",))
    if a_cs is not None and g_cs is not None:
        _, a_rows = _rows(a_cs)
        _, g_rows = _rows(g_cs)
        # total_value — последняя числовая колонка (category, product_count, avg_stock, total_value)
        def cat_map(rows):
            m = {}
            for r in rows:
                if not r or r[0] is None:
                    continue
                tv = None
                for c in reversed(r):
                    try:
                        tv = float(c); break
                    except (TypeError, ValueError):
                        continue
                m[_norm_cat(r[0])] = (r[0], tv)
            return m
        am = cat_map(a_rows)
        gm = cat_map(g_rows)
        ok = True
        details = []
        for gcat_norm, (gcat_label, gtv) in gm.items():
            # найти соответствующую категорию агента (с учётом синонимов)
            match = None
            for acat_norm, (alabel, atv) in am.items():
                if _cat_match(gcat_label, alabel):
                    match = (alabel, atv); break
            if match is None:
                ok = False; details.append(f"нет категории ~ {gcat_label}")
                continue
            atv = match[1]
            if gtv is None or atv is None:
                ok = False; details.append(f"{gcat_label}: нет total_value")
                continue
            tol = max(1.0, abs(gtv) * 0.02)
            if abs(atv - gtv) > tol:
                ok = False; details.append(f"{gcat_label}: total_value {atv} != {gtv}")
        # количество категорий совпадает
        if len(am) != len(gm):
            ok = False; details.append(f"category count {len(am)} != {len(gm)}")
        check("Category_Summary: набор категорий и total_value совпадают с эталоном",
              ok, "; ".join(details))
    else:
        check("Category_Summary: набор категорий и total_value совпадают с эталоном",
              False, "лист Category_Summary не найден")

    # --- Reorder_Alerts: только Critical/Out_of_Stock, Out_of_Stock первыми ---
    a_ra = _load_sheet(agent_wb, ("reorder",), ("alert",))
    g_ra = _load_sheet(gt, ("reorder",), ("alert",))
    if a_ra is not None and g_ra is not None:
        a_h, a_data = _rows(a_ra)
        ah = [str(c).lower() if c else "" for c in a_h]
        try:
            urg_col = next(i for i, h in enumerate(ah) if "urgen" in h)
        except StopIteration:
            urg_col = 3
        urg_seq = [str(r[urg_col]).strip().lower() if r[urg_col] is not None else "" for r in a_data]
        allowed = {"critical", "out_of_stock", "out of stock", "outofstock"}
        only_allowed = all(u in allowed for u in urg_seq)
        # Out_of_Stock сортируются перед Critical
        def is_oos(u):
            return "out" in u and "stock" in u
        seen_critical = False
        order_ok = True
        for u in urg_seq:
            if u == "critical":
                seen_critical = True
            elif is_oos(u) and seen_critical:
                order_ok = False; break
        # число строк совпадает с эталоном (только Critical/OOS)
        _, g_data = _rows(g_ra)
        count_ok = abs(len(a_data) - len(g_data)) <= 1
        check("Reorder_Alerts: только Critical/Out_of_Stock, Out_of_Stock сортируются первыми",
              only_allowed and order_ok and count_ok,
              f"only_allowed={only_allowed} order_ok={order_ok} a={len(a_data)} g={len(g_data)}")
    else:
        check("Reorder_Alerts: только Critical/Out_of_Stock, Out_of_Stock сортируются первыми",
              False, "лист Reorder_Alerts не найден")

    # --- Restock_Schedule: reorder_date/quantity/supplier ---
    a_rs = _load_sheet(agent_wb, ("restock",), ("schedule",))
    g_rs = _load_sheet(gt, ("restock",), ("schedule",))
    if a_rs is not None and g_rs is not None:
        a_h, a_data = _rows(a_rs)
        ah = [str(c).lower() if c else "" for c in a_h]
        def col(*kw):
            for i, h in enumerate(ah):
                if all(k in h for k in kw):
                    return i
            return None
        c_date = col("reorder", "date") or col("date")
        c_qty = col("quantity") or col("qty")
        c_sup = col("supplier")
        if c_date is None: c_date = 1
        if c_qty is None: c_qty = 2
        if c_sup is None: c_sup = 3
        # эталонные quantity по product_name
        g_idx, g_data = _index_by_name(g_rs)
        g_h, _ = _rows(g_rs)
        gh = [str(c).lower() if c else "" for c in g_h]
        def gcol(*kw):
            for i, h in enumerate(gh):
                if all(k in h for k in kw):
                    return i
            return None
        gc_qty = gcol("quantity") or gcol("qty") or 2
        date_ok = True; sup_ok = True; qty_ok = True
        qty_mism = 0
        for r in a_data:
            d = str(r[c_date]).strip() if r[c_date] is not None else ""
            if "2026-03-10" not in d:
                date_ok = False
            s = str(r[c_sup]).strip() if r[c_sup] is not None else ""
            if s != "Primary Supplier":
                sup_ok = False
            name = str(r[0]).strip() if r[0] is not None else ""
            if name in g_idx:
                try:
                    av = float(r[c_qty]); gv = float(g_idx[name][gc_qty])
                    if abs(av - gv) > 0.5:
                        qty_mism += 1
                except (TypeError, ValueError):
                    qty_mism += 1
        qty_ok = qty_mism <= max(1, int(len(a_data) * 0.05))
        check("Restock_Schedule: reorder_date/quantity/supplier совпадают с эталоном",
              date_ok and sup_ok and qty_ok and len(a_data) >= 5,
              f"date_ok={date_ok} sup_ok={sup_ok} qty_mism={qty_mism} rows={len(a_data)}")
    else:
        check("Restock_Schedule: reorder_date/quantity/supplier совпадают с эталоном",
              False, "лист Restock_Schedule не найден")


def _count_oos_critical_from_gt(gt_path):
    """Из эталонного Reorder_Alerts вернуть (out_of_stock_count, critical_count)
    и топ-3 самых срочных позиций (по days_of_supply из Product_Inventory)."""
    if not os.path.exists(gt_path):
        return None, None, []
    gt = openpyxl.load_workbook(gt_path, data_only=True)
    ra = _load_sheet(gt, ("reorder",), ("alert",))
    h, data = _rows(ra)
    ah = [str(c).lower() if c else "" for c in h]
    try:
        urg_col = next(i for i, x in enumerate(ah) if "urgen" in x)
    except StopIteration:
        urg_col = 3
    oos = sum(1 for r in data if r[urg_col] and "out" in str(r[urg_col]).lower() and "stock" in str(r[urg_col]).lower())
    crit = sum(1 for r in data if str(r[urg_col]).strip().lower() == "critical")
    return oos, crit, []


def check_gsheet():
    print("\n=== Check 2: Google Sheet Inventory Dashboard ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT id, title FROM gsheet.spreadsheets")
        spreadsheets = cur.fetchall()
        dashboard = None
        for ss_id, title in spreadsheets:
            if title and "inventory" in title.lower():
                dashboard = (ss_id, title)
                break
        check("Inventory Dashboard spreadsheet exists", dashboard is not None,
              f"Spreadsheets: {[s[1] for s in spreadsheets]}")
        if dashboard:
            cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (dashboard[0],))
            cell_count = cur.fetchone()[0]
            check("Dashboard has data cells", cell_count >= 5, f"Found {cell_count} cells")
    except Exception as e:
        check("Gsheet check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_gcal():
    print("\n=== Check 3: Calendar Restock Meeting ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        cur.execute("SELECT summary, description, start_datetime FROM gcal.events")
        events = cur.fetchall()
        restock_event = None
        for summary, desc, start in events:
            s = (str(summary) + " " + str(desc)).lower()
            if "restock" in s or "inventory" in s:
                restock_event = (summary, desc, start)
                break
        check("Restock review meeting exists", restock_event is not None,
              f"Events: {[e[0] for e in events]}")
        if restock_event:
            check("Meeting mentions restock or inventory",
                  "restock" in str(restock_event[0]).lower() or "inventory" in str(restock_event[0]).lower())
    except Exception as e:
        check("Gcal check", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_email(gt_path):
    print("\n=== Check 4: Email to Purchasing Team (ОТПРАВЛЕНО) ===")
    oos, crit, _ = _count_oos_critical_from_gt(gt_path)
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    try:
        # ОТПРАВЛЕННОЕ письмо (email.messages), не черновик
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%purchasing@company.com%%'
        """)
        emails = cur.fetchall()
        check("Письмо отправлено на purchasing@company.com (не черновик)", len(emails) >= 1,
              "Нет отправленного письма на purchasing@company.com")

        target = None
        for e in emails:
            subj = str(e[1] or "").lower()
            if "critical" in subj and "inventory" in subj and ("restock" in subj or "alert" in subj):
                target = e; break
        if target is None and emails:
            target = emails[0]

        if target:
            subject = str(target[1] or "")
            check("Email subject = Critical Inventory Alert - Restock Required",
                  subject.strip() == "Critical Inventory Alert - Restock Required",
                  f"Subject: {subject}")
            body = str(target[3] or "")
            # CRITICAL: тело содержит корректные счётчики (число out-of-stock и critical)
            counts_ok = False
            if oos is not None and crit is not None:
                counts_ok = (str(oos) in body) and (str(crit) in body)
            check("Email: реальное письмо на purchasing@company.com с корректными счётчиками",
                  len(emails) >= 1 and counts_ok,
                  f"oos={oos} crit={crit} present_in_body={counts_ok}; body[:120]={body[:120]}")
        else:
            check("Email: реальное письмо на purchasing@company.com с корректными счётчиками",
                  False, "письмо не найдено")
    except Exception as e:
        check("Email check", False, str(e))
        check("Email: реальное письмо на purchasing@company.com с корректными счётчиками", False, str(e))
    finally:
        cur.close()
        conn.close()


def check_script(workspace):
    print("\n=== Check 5: demand_forecast.py ===")
    path = os.path.join(workspace, "demand_forecast.py")
    check("demand_forecast.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Inventory_Lifecycle_Report.xlsx")
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=True)
        expected_keywords = {"product", "inventory", "reorder", "alert", "category", "summary",
                             "restock", "schedule"}
        unexpected = [s for s in wb.sheetnames
                      if not any(kw in s.lower() for kw in expected_keywords)]
        check("No unexpected sheets in Excel", len(unexpected) == 0, f"Unexpected: {unexpected}")

        ws = _load_sheet(wb, ("product", "inventory"))
        if ws is not None:
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            has_negative_stock = False
            for r in rows:
                if r and len(r) >= 4:
                    try:
                        stock = float(r[3]) if r[3] is not None else None
                    except (ValueError, TypeError):
                        stock = None
                    if stock is not None and stock < 0:
                        has_negative_stock = True
                        break
            check("No negative stock quantities", not has_negative_stock, "Found negative stock value")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE (summary ILIKE '%%restock%%' OR summary ILIKE '%%inventory%%')
              AND start_datetime < '2026-03-01'
        """)
        old_events = cur.fetchone()[0]
        check("No restock events before March 2026", old_events == 0, f"Found {old_events} old events")
        cur.close()
        conn.close()
    except Exception:
        pass


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    gt_path = os.path.join(args.groundtruth_workspace, "Inventory_Lifecycle_Report.xlsx")

    agent_wb = check_excel_structure(args.agent_workspace)
    check_excel_semantics(agent_wb, gt_path)
    check_gsheet()
    check_gcal()
    check_email(gt_path)
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"Critical fails: {CRITICAL_FAILS}")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy,
              "critical_fails": CRITICAL_FAILS}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"FAIL: критичные чеки провалены ({len(CRITICAL_FAILS)})")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
