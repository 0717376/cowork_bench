"""
Evaluation для playwright-insales-competitor-pricing-excel-gcal (InSales).

Что проверяем:
1. Competitive_Pricing_Analysis.xlsx с 4 листами и корректными данными.
2. События календаря для 3 совещаний по пересмотру цен (по категориям).
3. Отправленное письмо с результатами анализа цен.

Важно про русификацию:
- Категории в данных магазина InSales (wc.*) русифицированы централизованно:
  Электроника / Наушники / Колонки. groundtruth_workspace уже синхронизирован
  с этими русскими категориями. Поэтому проверки сопоставляют русские названия
  категорий (с допуском на английские синонимы), а не только английские.
- Числовые средние конкурента (COMPETITOR_AVGS) вычислены из USD-цен каталога
  catalog.html и НЕ меняются (каталог русифицирован только по подписям).

CRITICAL_CHECKS: любой провал => итог FAIL (sys.exit(1)) ДО порога accuracy.
Структурные проверки (лист существует, число строк, ISO-даты) — НЕ критичные.
Порог: accuracy >= 70 И нет критичных провалов => PASS.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Содержательные проверки. Провал любой => вся задача FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Competitor avg per category matches groundtruth",
    "Summary: Total_Competitor_Products = 15",
    "Summary: Total_Our_Products = 45",
    "Summary: Cheapest category is Наушники/Headphones",
    "Summary: Most expensive category is Электроника/Electronics",
    "Calendar: 3 Price Review events (one per category)",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=5.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains_any(haystack, needles):
    """True если haystack содержит хотя бы одну из подстрок needles (RU+EN)."""
    if haystack is None:
        return False
    h = str(haystack).strip().lower()
    for n in needles:
        if n and n.strip().lower() in h:
            return True
    return False


# Средние цены конкурента по категориям (из USD-цен catalog.html). Ключи — RU.
COMPETITOR_AVGS = {"электроника": 59.06, "наушники": 74.98, "колонки": 138.73}

# Синонимы названий категорий (RU + EN) для устойчивого сопоставления.
CAT_SYNONYMS = {
    "электроника": ["электроника", "electronics"],
    "наушники": ["наушники", "headphone"],
    "колонки": ["колонки", "speaker"],
}


def cat_key(label):
    """По произвольной подписи категории вернуть канонический RU-ключ или None."""
    if label is None:
        return None
    s = str(label).strip().lower()
    for key, syns in CAT_SYNONYMS.items():
        for syn in syns:
            if syn in s:
                return key
    return None


def check_excel(agent_workspace):
    """Проверка Competitive_Pricing_Analysis.xlsx."""
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # --- Sheet 1: Competitor Products ---
    comp_sheet = None
    for name in wb.sheetnames:
        if "competitor" in name.lower():
            comp_sheet = name
            break
    if not comp_sheet:
        record("Competitor Products sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Competitor Products sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 15
        record("Competitor Products has 15 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 2: Our Products ---
    our_sheet = None
    for name in wb.sheetnames:
        if "our" in name.lower():
            our_sheet = name
            break
    if not our_sheet:
        record("Our Products sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Our Products sheet exists", True)
        ws = wb[our_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        # Должны быть товары из 3 категорий (30+10+5=45)
        ok = len(data_rows) >= 30
        record("Our Products has >= 30 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

    # --- Sheet 3: Category Comparison ---
    cat_sheet = None
    for name in wb.sheetnames:
        if "category" in name.lower() or "comparison" in name.lower():
            cat_sheet = name
            break
    if not cat_sheet:
        record("Category Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Category Comparison sheet exists", True)
        ws = wb[cat_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        ok = len(data_rows) == 3
        record("Category Comparison has 3 rows", ok, f"Found {len(data_rows)}")
        if not ok:
            all_ok = False

        # CRITICAL: средняя цена конкурента по каждой категории совпадает с эталоном.
        # Сопоставляем строку по канонической категории (RU+EN), затем ищем ожидаемое
        # значение среди числовых ячеек строки.
        matched = {}
        for row in data_rows:
            if not row or row[0] is None:
                continue
            key = cat_key(row[0])
            if key is None or key not in COMPETITOR_AVGS:
                continue
            expected = COMPETITOR_AVGS[key]
            found = any(num_close(cell, expected, tol=1.0) for cell in row[1:])
            matched[key] = found

        comp_ok = (len(matched) == 3) and all(matched.values())
        record("Competitor avg per category matches groundtruth", comp_ok,
               f"Per-category match: {matched}")
        if not comp_ok:
            all_ok = False

    # --- Sheet 4: Summary ---
    sum_sheet = None
    for name in wb.sheetnames:
        if "summary" in name.lower():
            sum_sheet = name
            break
    if not sum_sheet:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Summary sheet exists", True)
        ws = wb[sum_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        seen = {"total_competitor": False, "total_our": False,
                "cheapest": False, "expensive": False}

        for row in data_rows:
            if row and row[0] is not None:
                metric = str(row[0]).strip().lower()
                val = row[1] if len(row) > 1 else None
                if "total_competitor" in metric:
                    seen["total_competitor"] = True
                    ok = num_close(val, 15, tol=0)
                    record("Summary: Total_Competitor_Products = 15", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "total_our" in metric:
                    seen["total_our"] = True
                    ok = num_close(val, 45, tol=0)
                    record("Summary: Total_Our_Products = 45", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "cheapest" in metric:
                    seen["cheapest"] = True
                    ok = str_contains_any(val, CAT_SYNONYMS["наушники"])
                    record("Summary: Cheapest category is Наушники/Headphones", ok, f"Got {val}")
                    if not ok:
                        all_ok = False
                elif "expensive" in metric:
                    seen["expensive"] = True
                    ok = str_contains_any(val, CAT_SYNONYMS["электроника"])
                    record("Summary: Most expensive category is Электроника/Electronics", ok, f"Got {val}")
                    if not ok:
                        all_ok = False

        # Если критичная метрика отсутствует в Summary — фиксируем критичный провал.
        for present, cname in [
            (seen["total_competitor"], "Summary: Total_Competitor_Products = 15"),
            (seen["total_our"], "Summary: Total_Our_Products = 45"),
            (seen["cheapest"], "Summary: Cheapest category is Наушники/Headphones"),
            (seen["expensive"], "Summary: Most expensive category is Электроника/Electronics"),
        ]:
            if not present:
                record(cname, False, "metric row missing in Summary")
                all_ok = False

    wb.close()
    return all_ok


def check_calendar():
    """Проверка событий календаря: 3 совещания по пересмотру цен, по 1 на категорию,
    на 2026-03-10/11/12, 14:00-15:00 America/New_York."""
    print("\n=== Checking Google Calendar ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, description, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        return False

    all_ok = True
    categories_found = set()
    dates_found = set()

    expected_dates = {"2026-03-10", "2026-03-11", "2026-03-12"}

    for summary, description, start_dt, end_dt in events:
        summary_lower = (summary or "").lower()
        # "price review" (EN) либо "обзор цен"/"пересмотр цен"/"price review" (RU)
        is_review = ("price review" in summary_lower
                     or "обзор цен" in summary_lower
                     or "пересмотр цен" in summary_lower)
        if not is_review:
            continue
        key = cat_key(summary)
        if key is not None:
            categories_found.add(key)
        # Дата начала события (любой разумный формат содержит ISO-подстроку).
        start_str = str(start_dt)
        for d in expected_dates:
            if d in start_str:
                dates_found.add(d)

    # CRITICAL: ровно по одному событию на каждую из трёх категорий.
    cal_ok = categories_found == {"электроника", "наушники", "колонки"}
    record("Calendar: 3 Price Review events (one per category)", cal_ok,
           f"Categories found: {categories_found}")
    if not cal_ok:
        all_ok = False

    # Структурная (НЕ критичная) проверка дат.
    dates_ok = expected_dates.issubset(dates_found)
    record("Calendar: events on 2026-03-10/11/12 (ISO)", dates_ok,
           f"Dates found: {dates_found}")
    if not dates_ok:
        all_ok = False

    return all_ok


def check_email():
    """Проверка отправленного письма с результатами анализа цен."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    all_ok = True
    found_email = False

    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        if "pricing" in subj_lower or "competitive" in subj_lower or "price" in subj_lower:
            found_email = True
            record("Pricing analysis email exists", True)

            from_ok = str_contains_any(from_addr, ["pricing", "ecommerce"])
            record("Email from pricing/ecommerce address", from_ok, f"From: {from_addr}")
            if not from_ok:
                all_ok = False

            to_ok = str_contains_any(to_addr, ["manager", "ecommerce"])
            record("Email to manager/ecommerce address", to_ok, f"To: {to_addr}")
            if not to_ok:
                all_ok = False

            # Тело упоминает хотя бы одну категорию (RU+EN).
            body = body_text or ""
            body_ok = (str_contains_any(body, CAT_SYNONYMS["электроника"])
                       or str_contains_any(body, CAT_SYNONYMS["наушники"])
                       or str_contains_any(body, CAT_SYNONYMS["колонки"]))
            record("Email body mentions categories", body_ok,
                   f"Body preview: {body[:200]}")
            if not body_ok:
                all_ok = False
            break

    if not found_email:
        record("Pricing analysis email exists", False,
               f"Found {len(emails)} emails, none with pricing/competitive in subject")
        all_ok = False

    return all_ok


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace)
    cal_ok = check_calendar()
    email_ok = check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:    {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Calendar: {'PASS' if cal_ok else 'FAIL'}")
    print(f"  Email:    {'PASS' if email_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"  Critical fails: {CRITICAL_FAILS}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_fails": CRITICAL_FAILS,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"  Overall: FAIL (критичные чеки провалены: {len(CRITICAL_FAILS)})")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL (accuracy < 70)")
    sys.exit(1)


if __name__ == "__main__":
    main()
