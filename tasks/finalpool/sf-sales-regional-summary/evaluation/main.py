"""Оценка для sf-sales-regional-summary (форк ClickHouse).

Регионы в sf_data русифицированы центральным отображением
(scripts/clickhouse_relabel_map.REGIONS). Эталон
groundtruth_workspace/Regional_Sales.xlsx переразмечен ТЕМ ЖЕ отображением,
поэтому seed <-> eval <-> groundtruth остаются согласованными; числовые значения
заморожены. Идентификаторы колонок/листов/метрик остаются английскими.
"""
import argparse
import os
import sys
import openpyxl

# Русифицированные имена регионов (из scripts/clickhouse_relabel_map.REGIONS).
RU_REGIONS = [
    "Северная Америка",
    "Европа",
    "Азиатско-Тихоокеанский регион",
    "Латинская Америка",
    "Ближний Восток",
]
GT_TOTAL_REVENUE = 2177149.66
GT_TOTAL_REGIONS = 5


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_to_lookup(rows):
    """Первая колонка -> строка (без заголовка), ключ в lower/strip."""
    data = rows[1:] if rows and len(rows) > 1 else []
    lookup = {}
    for row in data:
        if row and row[0] is not None:
            lookup[str(row[0]).strip().lower()] = row
    return lookup


def to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Regional_Sales.xlsx")
    gt_file = os.path.join(gt_dir, "Regional_Sales.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---------------------------------------------------------------
    # КРИТИЧЕСКИЕ ПРОВЕРКИ: любой провал => немедленный выход (sys.exit(1))
    # до подсчёта итоговой точности. Проверяют СУТЬ результата.
    # ---------------------------------------------------------------
    critical_errors = []

    a_rb = load_sheet_rows(agent_wb, "Regional Breakdown")
    a_sm = load_sheet_rows(agent_wb, "Summary")
    rb_lookup = rows_to_lookup(a_rb) if a_rb is not None else {}
    sm_lookup = rows_to_lookup(a_sm) if a_sm is not None else {}

    if a_rb is None:
        critical_errors.append("CRITICAL: лист 'Regional Breakdown' отсутствует")
    if a_sm is None:
        critical_errors.append("CRITICAL: лист 'Summary' отсутствует")

    # CRITICAL 1: все 5 русифицированных регионов присутствуют в Regional Breakdown.
    for region in RU_REGIONS:
        if region.strip().lower() not in rb_lookup:
            critical_errors.append(f"CRITICAL: отсутствует регион '{region}' в Regional Breakdown")

    # Собираем per-region Revenue из агентских данных (если регион найден).
    region_revenue = {}
    for region in RU_REGIONS:
        row = rb_lookup.get(region.strip().lower())
        if row and len(row) > 3:
            region_revenue[region] = to_float(row[3])

    sum_region_revenue = sum(v for v in region_revenue.values() if v is not None)

    # CRITICAL 2: Total_Revenue близок к эталону И равен сумме региональных Revenue.
    tr_row = sm_lookup.get("total_revenue")
    if tr_row is None or len(tr_row) < 2:
        critical_errors.append("CRITICAL: метрика Total_Revenue отсутствует в Summary")
    else:
        tr_val = to_float(tr_row[1])
        if tr_val is None or abs(tr_val - GT_TOTAL_REVENUE) > 1.0:
            critical_errors.append(
                f"CRITICAL: Total_Revenue={tr_row[1]} != эталон {GT_TOTAL_REVENUE} (tol=1.0)"
            )
        elif sum_region_revenue and abs(tr_val - sum_region_revenue) > 1.0:
            critical_errors.append(
                f"CRITICAL: Total_Revenue={tr_val} != сумма региональных Revenue "
                f"{round(sum_region_revenue, 2)} (внутренняя несогласованность)"
            )

    # CRITICAL 3: Total_Regions == 5.
    treg_row = sm_lookup.get("total_regions")
    if treg_row is None or len(treg_row) < 2:
        critical_errors.append("CRITICAL: метрика Total_Regions отсутствует в Summary")
    else:
        treg_val = to_float(treg_row[1])
        if treg_val is None or int(treg_val) != GT_TOTAL_REGIONS:
            critical_errors.append(
                f"CRITICAL: Total_Regions={treg_row[1]} != {GT_TOTAL_REGIONS}"
            )

    # CRITICAL 4: Top_Region == регион с наибольшей выручкой
    # (эталон: 'Азиатско-Тихоокеанский регион'). Проверяем по агентским данным,
    # что Top_Region действительно совпадает с регионом максимальной выручки.
    top_row = sm_lookup.get("top_region")
    if top_row is None or len(top_row) < 2:
        critical_errors.append("CRITICAL: метрика Top_Region отсутствует в Summary")
    else:
        top_val = str(top_row[1]).strip().lower() if top_row[1] is not None else ""
        # Ожидаемый top по эталону.
        gt_rb = load_sheet_rows(gt_wb, "Regional Breakdown")
        gt_lookup = rows_to_lookup(gt_rb) if gt_rb is not None else {}
        gt_top = None
        gt_max = None
        for key, row in gt_lookup.items():
            rev = to_float(row[3]) if len(row) > 3 else None
            if rev is not None and (gt_max is None or rev > gt_max):
                gt_max = rev
                gt_top = key
        if gt_top is not None and top_val != gt_top:
            critical_errors.append(
                f"CRITICAL: Top_Region='{top_row[1]}' != регион с макс. выручкой "
                f"(эталон '{gt_top}')"
            )

    # CRITICAL 5: каждый Revenue_Share_Pct согласован с Revenue/Total_Revenue*100.
    if sum_region_revenue:
        for region in RU_REGIONS:
            row = rb_lookup.get(region.strip().lower())
            if not row or len(row) < 6:
                continue
            rev = to_float(row[3])
            share = to_float(row[5])
            if rev is None or share is None:
                continue
            expected = rev / sum_region_revenue * 100.0
            if abs(share - expected) > 0.3:
                critical_errors.append(
                    f"CRITICAL: Revenue_Share_Pct для '{region}'={share} != "
                    f"{round(expected, 1)} (Revenue/Total*100, tol=0.3)"
                )

    if critical_errors:
        print("=== КРИТИЧЕСКИЙ ПРОВАЛ ===")
        for e in critical_errors:
            print(f"  {e}")
        sys.exit(1)

    # ---------------------------------------------------------------
    # НЕ-КРИТИЧЕСКИЕ (структурные) проверки -> итоговая точность.
    # ---------------------------------------------------------------
    all_errors = []
    total_checks = 0
    passed_checks = 0

    # Check sheet: Regional Breakdown
    print(f"  Checking Regional Breakdown...")
    a_rows = a_rb
    g_rows = load_sheet_rows(gt_wb, "Regional Breakdown")
    if a_rows is None:
        all_errors.append("Sheet 'Regional Breakdown' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Regional Breakdown' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                total_checks += 1
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                total_checks += 1
                if not num_close(a_row[1], g_row[1], 5):
                    errors.append(f"{key}.Orders: {a_row[1]} vs {g_row[1]} (tol=5)")
                else:
                    passed_checks += 1

            if len(a_row) > 2 and len(g_row) > 2:
                total_checks += 1
                if not num_close(a_row[2], g_row[2], 5):
                    errors.append(f"{key}.Customers: {a_row[2]} vs {g_row[2]} (tol=5)")
                else:
                    passed_checks += 1

            if len(a_row) > 3 and len(g_row) > 3:
                total_checks += 1
                if not num_close(a_row[3], g_row[3], 100.0):
                    errors.append(f"{key}.Revenue: {a_row[3]} vs {g_row[3]} (tol=100.0)")
                else:
                    passed_checks += 1

            if len(a_row) > 4 and len(g_row) > 4:
                total_checks += 1
                if not num_close(a_row[4], g_row[4], 2.0):
                    errors.append(f"{key}.Avg_Order_Value: {a_row[4]} vs {g_row[4]} (tol=2.0)")
                else:
                    passed_checks += 1

            if len(a_row) > 5 and len(g_row) > 5:
                total_checks += 1
                if not num_close(a_row[5], g_row[5], 0.5):
                    errors.append(f"{key}.Revenue_Share_Pct: {a_row[5]} vs {g_row[5]} (tol=0.5)")
                else:
                    passed_checks += 1
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = a_sm
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                total_checks += 1
                errors.append(f"Missing row: {g_row[0]}")
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                total_checks += 1
                # Top_Region — текстовая метрика: сравнение без учёта регистра.
                if key == "top_region":
                    if not str_match(a_row[1], g_row[1]):
                        errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]}")
                    else:
                        passed_checks += 1
                elif not num_close(a_row[1], g_row[1], 100.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=100.0)")
                else:
                    passed_checks += 1
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    accuracy = (passed_checks / total_checks * 100.0) if total_checks else 0.0
    print(f"\nТочность: {passed_checks}/{total_checks} = {accuracy:.1f}%")

    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
