"""
Evaluation script for playwright-insales-review-sentiment-excel-email task.

Checks:
1. Review_Comparison_Report.xlsx with Product Comparison and Category Summary sheets
2. Email with quality alert for flagged products
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=0.3):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


def check_excel(agent_workspace, groundtruth_workspace):
    """Check Review_Comparison_Report.xlsx."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Review_Comparison_Report.xlsx")
    if not os.path.isfile(agent_file):
        record("Excel file exists", False, f"Not found: {agent_file}")
        return False

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return False

    all_ok = True

    # Check Product Comparison sheet
    comp_sheet = None
    for name in wb.sheetnames:
        if "comparison" in name.lower() or "product" in name.lower():
            comp_sheet = name
            break

    if not comp_sheet:
        record("Product Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Product Comparison sheet exists", True)
        ws = wb[comp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if r and r[0]] if len(rows) > 1 else []

        record(
            "Product Comparison has >= 12 products",
            len(data_rows) >= 12,
            f"Found {len(data_rows)} data rows",
        )
        if len(data_rows) < 12:
            all_ok = False

        # Check that alert flags exist
        alert_col = None
        if rows:
            header = [str(h).lower() if h else "" for h in rows[0]]
            for i, h in enumerate(header):
                if "alert" in h or "flag" in h:
                    alert_col = i
                    break

        if alert_col is not None:
            flagged = [r for r in data_rows if r[alert_col] and str(r[alert_col]).lower() == "yes"]
            record(
                "At least 2 products flagged",
                len(flagged) >= 2,
                f"Found {len(flagged)} flagged products",
            )
            if len(flagged) < 2:
                all_ok = False
        else:
            record("Alert flag column exists", False, f"Headers: {rows[0] if rows else 'none'}")
            all_ok = False

        # Check rating difference column exists
        diff_col = None
        if rows:
            for i, h in enumerate(header):
                if "difference" in h or "diff" in h:
                    diff_col = i
                    break

        if diff_col is not None:
            record("Rating difference column exists", True)
            # Spot check: NIHARA should have diff ~ 0.6
            for r in data_rows:
                if r[0] and "nihara" in str(r[0]).lower():
                    ok = num_close(r[diff_col], 0.6, tol=0.3)
                    record(
                        "NIHARA rating diff ~0.6",
                        ok,
                        f"Got {r[diff_col]}",
                    )
                    if not ok:
                        all_ok = False
                    break
        else:
            record("Rating difference column exists", False)
            all_ok = False

    # Check Category Summary sheet
    cat_sheet = None
    for name in wb.sheetnames:
        if "category" in name.lower() or "summary" in name.lower():
            cat_sheet = name
            break

    if not cat_sheet:
        record("Category Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        all_ok = False
    else:
        record("Category Summary sheet exists", True)
        ws2 = wb[cat_sheet]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if r and r[0]] if len(rows2) > 1 else []

        record(
            "Category Summary has 3 categories",
            len(data_rows2) >= 3,
            f"Found {len(data_rows2)} rows",
        )
        if len(data_rows2) < 3:
            all_ok = False

        # Check categories present (RU labels after russification, EN fallback)
        cats = [str(r[0]).lower() for r in data_rows2 if r[0]]
        has_electronics = any("electron" in c or "электрон" in c for c in cats)
        has_cameras = any("camera" in c or "камер" in c for c in cats)
        has_appliances = any(
            "appliance" in c or "техник" in c or "бытов" in c for c in cats
        )
        record(
            "All 3 categories present",
            has_electronics and has_cameras and has_appliances,
            f"Categories: {cats}",
        )
        if not (has_electronics and has_cameras and has_appliances):
            all_ok = False

    wb.close()
    return all_ok


def check_email():
    """Check quality alert email."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute(
            "SELECT subject, from_addr, to_addr, body_text FROM email.messages"
        )
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e))
        return False

    found = False
    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        if "quality" in subj_lower or "alert" in subj_lower or "discrepan" in subj_lower:
            found = True
            record("Quality alert email exists", True)

            # Check recipient
            to_str = str(to_addr).lower() if to_addr else ""
            record(
                "Email to product-team",
                "product" in to_str,
                f"To: {to_addr}",
            )

            # Check body mentions flagged products
            body_lower = (body_text or "").lower()
            mentions_product = any(
                kw in body_lower
                for kw in ["nihara", "craftwings", "limbani", "vacuum", "laptop", "fan"]
            )
            record(
                "Email body mentions flagged products",
                mentions_product,
                f"Body length: {len(body_lower)}",
            )

            # Check body has rating info (RU + EN)
            has_rating = any(
                kw in body_lower
                for kw in [
                    "rating", "internal", "external", "difference", "diff",
                    "рейтинг", "внутр", "внешн", "разниц", "расхожд",
                ]
            )
            record("Email body has rating information", has_rating)
            break

    if not found:
        record(
            "Quality alert email exists",
            False,
            f"Found {len(emails)} emails but none about quality/alert/discrepancy",
        )

    return found


def _cat_label(c):
    """Map a category cell (RU or EN) to a canonical key, or None."""
    c = str(c or "").lower()
    if "electron" in c or "электрон" in c:
        return "electronics"
    if "camera" in c or "камер" in c:
        return "cameras"
    if "appliance" in c or "техник" in c or "бытов" in c:
        return "appliances"
    return None


def check_critical(agent_workspace):
    """SEMANTIC critical checks. Any failure => hard FAIL (sys.exit(1)).

    Accepts Russian category labels (Электроника/Камеры/Бытовая техника) and
    English fallback, since wc.* data is russified centrally while external
    product names stay English.
    """
    print("\n=== CRITICAL CHECKS ===")
    critical_ok = True

    def crit(name, passed, detail=""):
        nonlocal critical_ok
        record(name, passed, detail)
        if not passed:
            critical_ok = False

    agent_file = os.path.join(agent_workspace, "Review_Comparison_Report.xlsx")
    if not os.path.isfile(agent_file):
        crit("[CRITICAL] Excel file exists", False, f"Not found: {agent_file}")
        return critical_ok
    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        crit("[CRITICAL] Excel readable", False, str(e))
        return critical_ok

    # locate sheets
    comp_sheet = next(
        (n for n in wb.sheetnames if "comparison" in n.lower() or "product" in n.lower()),
        None,
    )
    cat_sheet = next(
        (n for n in wb.sheetnames if "category" in n.lower() or "summary" in n.lower()),
        None,
    )

    # ---- Product Comparison: header map + rows ----
    comp_rows = []
    hmap = {}
    if comp_sheet:
        rows = list(wb[comp_sheet].iter_rows(values_only=True))
        if rows:
            header = [str(h).lower() if h else "" for h in rows[0]]

            def col(*keys):
                for i, h in enumerate(header):
                    if any(k in h for k in keys):
                        return i
                return None

            hmap = {
                "name": col("product_name", "name"),
                "cat": col("category"),
                "internal": col("internal_avg", "internal_rating"),
                "external": col("external_avg", "external_rating"),
                "diff": col("difference", "diff"),
                "flag": col("alert", "flag"),
            }
            comp_rows = [r for r in rows[1:] if r and r[0]]

    def find_prod(substr):
        for r in comp_rows:
            if hmap.get("name") is not None and substr in str(r[hmap["name"]] or "").lower():
                return r
        return None

    def is_yes(v):
        return str(v or "").strip().lower() in ("yes", "да")

    # CRITICAL 1: NIHARA diff ~0.6 AND flagged
    r = find_prod("nihara")
    if r is None:
        crit("[CRITICAL] NIHARA row present", False)
    else:
        di = hmap.get("diff")
        fi = hmap.get("flag")
        ok = (di is not None and num_close(r[di], 0.6, tol=0.3)) and (
            fi is not None and is_yes(r[fi])
        )
        crit(
            "[CRITICAL] NIHARA Rating_Difference ~0.6 AND Alert_Flag=Yes",
            ok,
            f"diff={r[di] if di is not None else '?'} flag={r[fi] if fi is not None else '?'}",
        )

    # CRITICAL 2: all 3 groundtruth-flagged products are Yes with matching diff
    expected = [("nihara", 0.6), ("craftwings", 0.7), ("limbani", 0.51)]
    for substr, exp_diff in expected:
        rr = find_prod(substr)
        di = hmap.get("diff")
        fi = hmap.get("flag")
        ok = (
            rr is not None
            and di is not None
            and fi is not None
            and num_close(rr[di], exp_diff, tol=0.3)
            and is_yes(rr[fi])
        )
        crit(
            f"[CRITICAL] {substr} flagged with diff ~{exp_diff}",
            ok,
            ""
            if rr is None
            else f"diff={rr[di] if di is not None else '?'} flag={rr[fi] if fi is not None else '?'}",
        )

    # ---- Category Summary: per-category flagged counts ----
    if not cat_sheet:
        crit("[CRITICAL] Category Summary sheet present", False)
    else:
        rows2 = list(wb[cat_sheet].iter_rows(values_only=True))
        header2 = [str(h).lower() if h else "" for h in (rows2[0] if rows2 else [])]
        flag_col = next(
            (i for i, h in enumerate(header2) if "flag" in h and "count" in h), None
        )
        if flag_col is None:
            flag_col = next((i for i, h in enumerate(header2) if "flag" in h), None)
        data2 = [r for r in rows2[1:] if r and r[0]] if len(rows2) > 1 else []
        got = {}
        for r in data2:
            key = _cat_label(r[0])
            if key and flag_col is not None:
                try:
                    got[key] = int(float(r[flag_col]))
                except (TypeError, ValueError):
                    got[key] = None
        exp_counts = {"electronics": 1, "cameras": 0, "appliances": 2}
        ok = all(got.get(k) == v for k, v in exp_counts.items())
        crit(
            "[CRITICAL] Products_Flagged_Count per category (Электроника=1, Камеры=0, Бытовая техника=2)",
            ok,
            f"Got: {got}",
        )

    wb.close()
    return critical_ok


def check_email_critical():
    """CRITICAL: quality alert email to product-team naming >=2 flagged products
    with rating info (RU or EN)."""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("[CRITICAL] Email DB accessible", False, str(e))
        return False

    for subject, from_addr, to_addr, body_text in emails:
        subj_lower = (subject or "").lower()
        if not ("quality alert" in subj_lower or "discrepan" in subj_lower):
            continue
        to_str = str(to_addr).lower() if to_addr else ""
        body_lower = (body_text or "").lower()
        prod_kw = ["nihara", "craftwings", "limbani", "vacuum", "laptop", "fan"]
        n_prod = sum(1 for kw in prod_kw if kw in body_lower)
        has_rating = any(
            kw in body_lower
            for kw in [
                "rating", "internal", "external", "difference", "diff",
                "рейтинг", "внутр", "внешн", "разниц", "расхожд",
            ]
        )
        ok = ("product" in to_str) and (n_prod >= 2) and has_rating
        record(
            "[CRITICAL] Quality alert email to product-team names >=2 flagged products + rating info",
            ok,
            f"to={to_addr} n_prod={n_prod} has_rating={has_rating}",
        )
        return ok

    record(
        "[CRITICAL] Quality alert email exists",
        False,
        f"Found {len(emails)} emails but none with 'Quality Alert'/'discrepancy' subject",
    )
    return False


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    excel_ok = check_excel(args.agent_workspace, args.groundtruth_workspace)
    email_ok = check_email()

    # SEMANTIC critical checks: any failure => hard FAIL regardless of accuracy.
    crit_excel_ok = check_critical(args.agent_workspace)
    crit_email_ok = check_email_critical()
    critical_ok = crit_excel_ok and crit_email_ok

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Excel:  {'PASS' if excel_ok else 'FAIL'}")
    print(f"  Email:  {'PASS' if email_ok else 'FAIL'}")
    print(f"  Critical:  {'PASS' if critical_ok else 'FAIL'}")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")

    if not critical_ok:
        print("  Overall:  FAIL (critical check failed)")
        sys.exit(1)

    overall = (accuracy >= 70) and excel_ok and email_ok
    print(f"  Overall:  {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
