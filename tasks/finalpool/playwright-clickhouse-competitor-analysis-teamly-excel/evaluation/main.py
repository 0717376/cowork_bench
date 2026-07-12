"""Evaluation for playwright-clickhouse-competitor-analysis-teamly-excel.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Notes on the swap:
- snowflake -> clickhouse: sf_data region/segment DATA VALUES are russified
  centrally (db/zzz_clickhouse_after_init.sql). The groundtruth xlsx mirrors the
  russified region/segment names. Region/segment Revenue is verified by matching
  the agent's per-row numeric Revenue against the groundtruth aggregate
  (language-agnostic) -- this verifies the actual warehouse query.
- notion -> teamly: the 'Competitive Intelligence Tracker' deliverable maps to a
  Teamly parent page/space with five competitor child pages (one per brand), each
  page body carrying the Market Share number and the Competitive Position class.
"""
import argparse
import os
import re
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Competitor brands + expected market share + derived position (source of truth
# for the classification rule: Leader>=20, Challenger 10-19.9, Niche<10).
COMPETITORS = [
    ("AlphaTech Solutions", 22.5, "Leader"),
    ("BetaWave Industries", 16.4, "Challenger"),
    ("Gamma Digital Corp", 14.8, "Challenger"),
    ("DeltaForce Tech", 10.3, "Challenger"),
    ("Epsilon Dynamics", 6.3, "Niche"),
]

# Required SWOT factors (matched by keyword, RU or EN).
SWOT_FACTORS = [
    ("Total Revenue", ["total revenue", "суммарн", "общая выруч", "итого выруч"]),
    ("Market Breadth", ["market breadth", "breadth", "регион", "широта", "охват"]),
    ("Customer Base", ["customer base", "orders", "заказ", "клиентск", "база клиент"]),
    ("Satisfaction", ["satisfaction", "удовлетвор"]),
]

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Internal Performance: per-region Revenue matches warehouse aggregate",
    "Internal Performance: per-segment Revenue matches warehouse aggregate",
    "Competitor Profiles: 5 brands with correct Market_Share and Position",
    "SWOT Summary: 4 required factors present and consistent",
    "Teamly: 5 competitor pages under 'Competitive Intelligence Tracker'",
    "Teamly: competitor pages carry Market Share + Competitive Position",
    "Email to board@company.com with required subject and substantive body",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d.\-]", "", str(v))
    try:
        return float(s)
    except ValueError:
        return None


def derive_position(ms):
    if ms is None:
        return None
    if ms >= 20:
        return "leader"
    if ms >= 10:
        return "challenger"
    return "niche"


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


# ---------------------------------------------------------------------------
# Excel checks
# ---------------------------------------------------------------------------
def check_excel(agent_ws, gt_dir):
    import openpyxl

    agent_file = os.path.join(agent_ws, "Competitive_Analysis.xlsx")
    gt_file = os.path.join(gt_dir, "Competitive_Analysis.xlsx")

    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        record("Competitive_Analysis.xlsx exists in agent workspace", False)
        for n in ("Competitor Profiles: 5 brands with correct Market_Share and Position",
                  "Internal Performance: per-region Revenue matches warehouse aggregate",
                  "Internal Performance: per-segment Revenue matches warehouse aggregate",
                  "SWOT Summary: 4 required factors present and consistent"):
            record(n, False, "workbook missing")
        return
    record("Competitive_Analysis.xlsx exists in agent workspace", True)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    _check_competitor_profiles(agent_wb)
    _check_internal_performance(agent_wb, gt_wb)
    _check_swot(agent_wb)


def _check_competitor_profiles(agent_wb):
    a_rows = load_sheet_rows(agent_wb, "Competitor Profiles")
    if a_rows is None:
        record("Sheet 'Competitor Profiles' present", False)
        record("Competitor Profiles: 5 brands with correct Market_Share and Position", False, "sheet missing")
        return
    record("Sheet 'Competitor Profiles' present", True)

    data = a_rows[1:] if len(a_rows) > 1 else []
    lookup = {}
    for row in data:
        if row and row[0]:
            lookup[str(row[0]).strip().lower()] = row

    errors = []
    for brand, exp_ms, exp_pos in COMPETITORS:
        a_row = lookup.get(brand.strip().lower())
        if a_row is None:
            errors.append(f"missing {brand}")
            continue
        ms = to_float(a_row[2]) if len(a_row) > 2 else None
        if not num_close(ms, exp_ms, 0.5):
            errors.append(f"{brand} Market_Share {ms} vs {exp_ms}")
        pos = str(a_row[5]).strip().lower() if len(a_row) > 5 and a_row[5] else ""
        if pos != exp_pos.lower():
            errors.append(f"{brand} Position {pos!r} vs {exp_pos}")
        # Position must also be self-consistent with the agent's own market share.
        derived = derive_position(ms)
        if derived and pos and derived != pos:
            errors.append(f"{brand} Position {pos!r} inconsistent with its Market_Share {ms}")

    record("Competitor Profiles: 5 brands with correct Market_Share and Position",
           not errors, "; ".join(errors))


def _aggregate_targets(gt_wb):
    """Read region & segment (Orders, Revenue) targets from the groundtruth.

    Returns (region_targets, segment_targets) as lists of (orders, revenue).
    The groundtruth is the russified warehouse aggregate kept in lockstep with the
    central russification map.
    """
    rows = load_sheet_rows(gt_wb, "Internal Performance")
    region_targets, segment_targets = [], []
    section = None
    for r in rows:
        if not r or r[0] is None:
            continue
        head = str(r[0]).strip().lower()
        if head == "region":
            section = "region"
            continue
        if head == "segment":
            section = "segment"
            continue
        orders = to_float(r[1]) if len(r) > 1 else None
        rev = to_float(r[2]) if len(r) > 2 else None
        if rev is None:
            continue
        if section == "region":
            region_targets.append((orders, rev))
        elif section == "segment":
            segment_targets.append((orders, rev))
    return region_targets, segment_targets


def _collect_numeric_rows(rows):
    """All rows with a non-empty label and a numeric value in the Revenue (3rd) col."""
    out = []
    for r in rows:
        if not r or r[0] is None:
            continue
        orders = to_float(r[1]) if len(r) > 1 else None
        rev = to_float(r[2]) if len(r) > 2 else None
        if rev is not None:
            out.append((orders, rev))
    return out


def _check_internal_performance(agent_wb, gt_wb):
    a_rows = load_sheet_rows(agent_wb, "Internal Performance")
    if a_rows is None:
        record("Sheet 'Internal Performance' present", False)
        record("Internal Performance: per-region Revenue matches warehouse aggregate", False, "sheet missing")
        record("Internal Performance: per-segment Revenue matches warehouse aggregate", False, "sheet missing")
        return
    record("Sheet 'Internal Performance' present", True)

    region_targets, segment_targets = _aggregate_targets(gt_wb)
    agent_numeric = _collect_numeric_rows(a_rows)
    agent_revs = [rev for _, rev in agent_numeric]

    def match_all(targets, label):
        # Language-agnostic: each target Revenue (and Orders) must appear among the
        # agent's numeric rows within tolerance. This verifies the warehouse query
        # regardless of whether region/segment labels are RU or EN.
        missing = []
        used = [False] * len(agent_numeric)
        for t_orders, t_rev in targets:
            hit = -1
            for i, (a_orders, a_rev) in enumerate(agent_numeric):
                if used[i]:
                    continue
                if num_close(a_rev, t_rev, 2.0):
                    if t_orders is None or num_close(a_orders, t_orders, 2.0):
                        hit = i
                        break
            if hit < 0:
                missing.append(f"rev~{t_rev}")
            else:
                used[hit] = True
        record(label, not missing,
               f"unmatched targets: {missing}; agent revenues: {[round(x,1) for x in agent_revs]}")

    if region_targets:
        match_all(region_targets, "Internal Performance: per-region Revenue matches warehouse aggregate")
    else:
        record("Internal Performance: per-region Revenue matches warehouse aggregate", False, "no GT region targets")
    if segment_targets:
        match_all(segment_targets, "Internal Performance: per-segment Revenue matches warehouse aggregate")
    else:
        record("Internal Performance: per-segment Revenue matches warehouse aggregate", False, "no GT segment targets")


def _check_swot(agent_wb):
    a_rows = load_sheet_rows(agent_wb, "SWOT Summary")
    if a_rows is None:
        record("Sheet 'SWOT Summary' present", False)
        record("SWOT Summary: 4 required factors present and consistent", False, "sheet missing")
        return
    record("Sheet 'SWOT Summary' present", True)

    data = a_rows[1:] if len(a_rows) > 1 else []
    factor_rows = {}
    for r in data:
        if not r or r[0] is None:
            continue
        label = str(r[0]).strip().lower()
        iv = to_float(r[1]) if len(r) > 1 else None
        ca = to_float(r[2]) if len(r) > 2 else None
        factor_rows[label] = (iv, ca)

    errors = []
    found = 0
    avg_comp_rev = sum(c[1] for c in [(0, 4250), (0, 3100), (0, 2800), (0, 1950), (0, 1200)]) / 5  # 2660 ($M)
    avg_comp_prod = (340 + 285 + 210 + 175 + 120) / 5  # 226
    avg_comp_sat = (4.3 + 4.1 + 4.5 + 3.9 + 4.0) / 5  # 4.16
    for canon, kws in SWOT_FACTORS:
        match = None
        for label, vals in factor_rows.items():
            if any(k in label for k in kws):
                match = vals
                break
        if match is None:
            errors.append(f"missing factor {canon}")
            continue
        found += 1
        iv, ca = match
        # Consistency: Competitor_Avg must reflect the source competitor dataset.
        if canon == "Total Revenue" and ca is not None and not num_close(ca, avg_comp_rev, 60):
            errors.append(f"{canon} Competitor_Avg {ca} vs ~{avg_comp_rev}")
        if canon == "Market Breadth" and ca is not None and not num_close(ca, avg_comp_prod, 12):
            errors.append(f"{canon} Competitor_Avg {ca} vs ~{avg_comp_prod}")
        if canon == "Satisfaction":
            if iv is not None and not num_close(iv, 4.2, 0.3):
                errors.append(f"{canon} Internal_Value {iv} vs 4.2")
            if ca is not None and not num_close(ca, avg_comp_sat, 0.3):
                errors.append(f"{canon} Competitor_Avg {ca} vs ~{avg_comp_sat}")

    record("SWOT Summary: 4 required factors present and consistent",
           found >= 4 and not errors,
           f"found {found}/4; {('; '.join(errors)) if errors else ''}")


# ---------------------------------------------------------------------------
# Teamly checks (notion -> teamly)
# ---------------------------------------------------------------------------
def check_teamly():
    print("Checking Teamly...")
    try:
        conn = get_conn()
        cur = conn.cursor()
        # User-created pages only (global seed pages have id <= 3).
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Teamly: 5 competitor pages under 'Competitive Intelligence Tracker'", False, str(e))
        record("Teamly: competitor pages carry Market Share + Competitive Position", False, str(e))
        return

    # Noise page ('Трекер проектов' / 'Project Tracker') must NOT count.
    def is_noise(title):
        tl = (title or "").lower()
        return "трекер проектов" in tl or "project tracker" in tl

    # Competitor pages: title contains a competitor brand and is not the noise page.
    brand_pages = []
    for pid, title, body in pages:
        if is_noise(title):
            continue
        tl = (title or "").lower()
        for brand, _, _ in COMPETITORS:
            if brand.lower() in tl:
                brand_pages.append((brand, title, body))
                break

    found_brands = {b for b, _, _ in brand_pages}
    record("Teamly: 5 competitor pages under 'Competitive Intelligence Tracker'",
           len(found_brands) >= 5,
           f"found brands: {sorted(found_brands)}; pages: {[(p[0], p[1]) for p in pages]}")

    # Tracker container exists (parent page or space) carrying the EN marker.
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM teamly.pages WHERE id > 3
            AND lower(title) LIKE '%competitive%intelligence%'
        """)
        cont_pages = cur.fetchone()[0]
        cur.execute("""
            SELECT COUNT(*) FROM teamly.spaces
            WHERE lower(name) LIKE '%competitive%intelligence%'
        """)
        cont_spaces = cur.fetchone()[0]
        conn.close()
    except Exception:
        cont_pages = cont_spaces = 0
    record("Teamly: 'Competitive Intelligence Tracker' container present",
           (cont_pages + cont_spaces) > 0,
           "no page/space titled 'Competitive Intelligence Tracker'")

    # Each competitor page body carries its Market Share number and a correct
    # Competitive Position class consistent with the Leader/Challenger/Niche rule.
    ms_map = {b.lower(): (ms, pos) for b, ms, pos in COMPETITORS}
    prop_errors = []
    for brand, title, body in brand_pages:
        text = ((title or "") + " " + (body or "")).lower()
        exp_ms, exp_pos = ms_map[brand.lower()]
        # market share numeric presence (e.g. 22.5 or 22,5)
        ms_str = str(exp_ms)
        ms_alt = ms_str.replace(".", ",")
        if ms_str not in text and ms_alt not in text:
            prop_errors.append(f"{brand}: market share {exp_ms} absent")
        if exp_pos.lower() not in text:
            prop_errors.append(f"{brand}: position '{exp_pos}' absent")
    if not brand_pages:
        prop_errors.append("no competitor pages")
    record("Teamly: competitor pages carry Market Share + Competitive Position",
           not prop_errors, "; ".join(prop_errors))


# ---------------------------------------------------------------------------
# Email check
# ---------------------------------------------------------------------------
def check_email():
    print("Checking email...")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, COALESCE(body_text, '') FROM email.messages
            WHERE to_addr::text ILIKE '%board@company.com%'
            AND subject ILIKE '%competitive%'
        """)
        rows = cur.fetchall()
        conn.close()
    except Exception as e:
        record("Email to board@company.com with required subject and substantive body", False, str(e))
        return

    if not rows:
        record("Email to board@company.com with required subject and substantive body",
               False, "no email to board@company.com with 'competitive' subject")
        return

    # Substantive body: must name the leading competitor (AlphaTech Solutions) and
    # reference total internal revenue (~3.0M / 3,048,998) and the strongest region
    # (Европа / Europe). RU+EN keyword alternatives accepted.
    ok = False
    detail = ""
    for subject, body in rows:
        text = (str(subject) + " " + str(body)).lower()
        has_leader = "alphatech" in text
        has_region = ("европ" in text or "europe" in text)
        # total revenue ~3.0M: accept variants of the digits or a "3.0 / 3,0 млн"
        has_revenue = bool(re.search(r"3[\s.,]?0\d{0,2}", text)) or "3 048" in text \
            or "3048" in text or "3,0" in text or "3.0" in text or "3 млн" in text \
            or "3 million" in text
        if has_leader and has_region and has_revenue:
            ok = True
            break
        detail = f"leader={has_leader} region={has_region} revenue={has_revenue}"
    record("Email to board@company.com with required subject and substantive body",
           ok, detail)


# ---------------------------------------------------------------------------
def run_evaluation(agent_ws, gt_dir):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    check_excel(agent_ws, gt_dir)
    check_teamly()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    return (not critical_failed) and accuracy >= 70


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    success = run_evaluation(agent_ws, gt_dir)
    if success:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
