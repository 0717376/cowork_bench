"""Evaluation for sf-sales-quarterly-review-gcal (ClickHouse fork)."""
import argparse
import json
import os
import sys

import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def qkey(v):
    """Normalize a quarter label to a year-agnostic key, e.g. 'Q1 2025'->'q1', 'q1'->'q1'."""
    s = str(v).strip().lower()
    return s.split()[0] if s else s


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


# ---------------------------------------------------------------------------
# Excel: returns (non_critical_errors, critical_results)
# critical_results is a list of (passed: bool, message: str).
# ---------------------------------------------------------------------------
def check_excel(agent_workspace, gt_dir):
    errors = []
    crit = []

    def C(ok, msg):
        crit.append((bool(ok), msg))

    try:
        import openpyxl
    except ImportError:
        errors.append("openpyxl not installed")
        C(False, "Excel: openpyxl available")
        return errors, crit

    agent_file = os.path.join(agent_workspace, "Sales_Quarterly_2025.xlsx")
    gt_file = os.path.join(gt_dir, "Sales_Quarterly_2025.xlsx")

    if not os.path.exists(agent_file):
        errors.append("Sales_Quarterly_2025.xlsx not found in agent workspace")
        C(False, "Excel: Sales_Quarterly_2025.xlsx exists")
        return errors, crit

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    # ---- Quarterly Performance sheet ----
    a_rows = load_sheet_rows(agent_wb, "Quarterly Performance")
    g_rows = load_sheet_rows(gt_wb, "Quarterly Performance")
    quarterly_ok = True
    if a_rows is None:
        errors.append("Sheet 'Quarterly Performance' not found in agent output")
        quarterly_ok = False
    else:
        a_data = [r for r in (a_rows[1:] if len(a_rows) > 1 else []) if r and r[0] is not None]
        g_data = [r for r in (g_rows[1:] if g_rows and len(g_rows) > 1 else []) if r and r[0] is not None]

        if len(a_data) < 4:
            errors.append(f"Quarterly Performance: expected 4 data rows, got {len(a_data)}")
            quarterly_ok = False
        else:
            a_lookup = {qkey(r[0]): r for r in a_data if r[0]}
            for g_row in g_data:
                key = qkey(g_row[0])
                a_row = a_lookup.get(key)
                if a_row is None:
                    errors.append(f"Missing quarter row: {g_row[0]}")
                    quarterly_ok = False
                    continue
                # Revenue col 1 (also gates CRITICAL)
                if len(a_row) > 1 and not num_close(a_row[1], g_row[1], 2.0):
                    errors.append(f"{g_row[0]} Revenue: got {a_row[1]}, expected {g_row[1]} (tol=2.0)")
                    quarterly_ok = False
                # Order_Count col 2
                if len(a_row) > 2 and not num_close(a_row[2], g_row[2], 5):
                    errors.append(f"{g_row[0]} Order_Count: got {a_row[2]}, expected {g_row[2]} (tol=5)")
                # Avg_Order_Value col 3
                if len(a_row) > 3 and not num_close(a_row[3], g_row[3], 1.0):
                    errors.append(f"{g_row[0]} Avg_Order_Value: got {a_row[3]}, expected {g_row[3]} (tol=1.0)")
                # QoQ_Change_Pct col 4 (skip Q1 which is None); gates CRITICAL
                if len(a_row) > 4 and g_row[4] is not None:
                    if not num_close(a_row[4], g_row[4], 0.5):
                        errors.append(f"{g_row[0]} QoQ_Change_Pct: got {a_row[4]}, expected {g_row[4]} (tol=0.5)")
                        quarterly_ok = False

    # CRITICAL 1: all 4 quarter rows present with correct Revenue and QoQ_Change_Pct.
    C(quarterly_ok,
      "Quarterly Performance: all 4 quarters present with correct Revenue (tol=2.0) and QoQ_Change_Pct (tol=0.5)")

    # ---- Annual Summary sheet ----
    a_sum = load_sheet_rows(agent_wb, "Annual Summary")
    g_sum = load_sheet_rows(gt_wb, "Annual Summary")
    annual_ok = True
    if a_sum is None:
        errors.append("Sheet 'Annual Summary' not found in agent output")
        annual_ok = False
    else:
        a_sum_data = {str(r[0]).strip().lower(): r[1] for r in (a_sum[1:] if len(a_sum) > 1 else []) if r and r[0]}
        g_sum_data = {str(r[0]).strip().lower(): r[1] for r in (g_sum[1:] if g_sum and len(g_sum) > 1 else []) if r and r[0]}

        # Annual_Revenue
        if "annual_revenue" in g_sum_data:
            av = a_sum_data.get("annual_revenue")
            if av is None:
                errors.append("Annual Summary missing Annual_Revenue row")
                annual_ok = False
            elif not num_close(av, g_sum_data["annual_revenue"], 2.0):
                errors.append(f"Annual_Revenue: got {av}, expected {g_sum_data['annual_revenue']} (tol=2.0)")
                annual_ok = False

        # Best_Quarter
        if "best_quarter" in g_sum_data:
            bq = a_sum_data.get("best_quarter")
            if bq is None:
                errors.append("Annual Summary missing Best_Quarter row")
                annual_ok = False
            elif qkey(bq) != qkey(g_sum_data["best_quarter"]):
                errors.append(f"Best_Quarter: got '{bq}', expected '{g_sum_data['best_quarter']}'")
                annual_ok = False

        # Worst_Quarter
        if "worst_quarter" in g_sum_data:
            wq = a_sum_data.get("worst_quarter")
            if wq is None:
                errors.append("Annual Summary missing Worst_Quarter row")
                annual_ok = False
            elif qkey(wq) != qkey(g_sum_data["worst_quarter"]):
                errors.append(f"Worst_Quarter: got '{wq}', expected '{g_sum_data['worst_quarter']}'")
                annual_ok = False

        # Avg_Quarterly_Revenue (non-critical)
        if "avg_quarterly_revenue" in g_sum_data:
            aq = a_sum_data.get("avg_quarterly_revenue")
            if aq is None:
                errors.append("Annual Summary missing Avg_Quarterly_Revenue row")
            elif not num_close(aq, g_sum_data["avg_quarterly_revenue"], 1.0):
                errors.append(f"Avg_Quarterly_Revenue: got {aq}, expected {g_sum_data['avg_quarterly_revenue']} (tol=1.0)")

        # Total_Orders
        if "total_orders" in g_sum_data:
            to = a_sum_data.get("total_orders")
            if to is None:
                errors.append("Annual Summary missing Total_Orders row")
                annual_ok = False
            elif not num_close(to, g_sum_data["total_orders"], 5):
                errors.append(f"Total_Orders: got {to}, expected {g_sum_data['total_orders']} (tol=5)")
                annual_ok = False

    # CRITICAL 2: headline annual metrics correct.
    C(annual_ok,
      "Annual Summary: Annual_Revenue, Best_Quarter='Q4 2025', Worst_Quarter='Q3 2025', Total_Orders all correct")

    return errors, crit


# ---------------------------------------------------------------------------
# GCal: exactly the four planning meetings on the required dates, 10:00, 2h.
# ---------------------------------------------------------------------------
def check_gcal():
    errors = []
    crit = []
    expected_dates = {"2026-04-01", "2026-07-01", "2026-10-01", "2027-01-04"}
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT id, summary, start_datetime, end_datetime
            FROM gcal.events
            WHERE (LOWER(summary) LIKE '%planning%'
                   OR LOWER(summary) LIKE '%sales%'
                   OR LOWER(summary) LIKE '%планир%'
                   OR LOWER(summary) LIKE '%продаж%')
              AND start_datetime >= '2026-01-01T00:00:00'
            ORDER BY start_datetime
        """)
        events = cur.fetchall()
        cur.close()
        conn.close()

        if len(events) < 4:
            errors.append(f"Expected at least 4 planning/sales gcal events, found {len(events)}")

        # Map each required date -> a matching event with 10:00 start and ~2h duration.
        matched = {}
        for ev in events:
            sd, ed = ev[2], ev[3]
            if sd is None:
                continue
            sd_str = sd.isoformat() if hasattr(sd, "isoformat") else str(sd)
            day = sd_str[:10]
            if day not in expected_dates:
                continue
            # start hour == 10
            hour_ok = getattr(sd, "hour", None) == 10
            # duration ~2h (allow small slack)
            dur_ok = False
            if ed is not None and hasattr(ed, "isoformat"):
                try:
                    dur_h = (ed - sd).total_seconds() / 3600.0
                    dur_ok = abs(dur_h - 2.0) <= 0.1
                except Exception:
                    dur_ok = False
            if hour_ok and dur_ok:
                matched[day] = ev

        missing = sorted(expected_dates - set(matched.keys()))
        if missing:
            errors.append(
                f"GCal: missing valid planning meeting(s) (10:00 start, 2h) on dates: {missing}")
        gcal_ok = len(missing) == 0
    except Exception as e:
        errors.append(f"GCal DB check error: {e}")
        gcal_ok = False

    # CRITICAL 3: the four exact planning meetings (dates + 10:00 + 2h).
    crit.append((gcal_ok,
                 "GCal: four planning meetings on 2026-04-01/07-01/10-01 & 2027-01-04 at 10:00 lasting 2h"))
    return errors, crit


# ---------------------------------------------------------------------------
# Email: subject contains '2025' + 'sales', recipient sales.leadership,
# and a non-trivial body summarizing findings.
# ---------------------------------------------------------------------------
def check_email():
    errors = []
    crit = []
    email_ok = False
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, from_addr, to_addr, body_text
            FROM email.messages
            WHERE LOWER(subject) LIKE '%2025%'
              AND LOWER(subject) LIKE '%sales%'
            ORDER BY id DESC
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
        if not emails:
            errors.append("No email with '2025' and 'sales' in subject found")
        else:
            for em in emails:
                subject = str(em[0] or "")
                to_str = str(em[2] or "").lower()
                body = str(em[3] or "")
                to_ok = "sales.leadership" in to_str
                subj_ok = "2025 annual sales performance report" in subject.lower()
                body_ok = len(body.strip()) >= 50
                if to_ok and subj_ok and body_ok:
                    email_ok = True
                    break
            if not email_ok:
                # provide specific diagnostics on best candidate
                em = emails[0]
                to_str = str(em[2] or "").lower()
                if "sales.leadership" not in to_str:
                    errors.append("No email sent to sales.leadership@company.com")
                if "2025 annual sales performance report" not in str(em[0] or "").lower():
                    errors.append("Email subject missing '2025 Annual Sales Performance Report'")
                if len(str(em[3] or "").strip()) < 50:
                    errors.append("Email body too short to summarize key findings")
    except Exception as e:
        errors.append(f"Email DB check error: {e}")

    # CRITICAL 4: email to sales.leadership with the English subject literal + non-empty body.
    crit.append((email_ok,
                 "Email to sales.leadership@company.com with subject '2025 Annual Sales Performance Report' and a non-trivial body"))
    return errors, crit


def check_word(agent_workspace):
    errors = []
    docx_path = os.path.join(agent_workspace, "Sales_Executive_Summary.docx")
    if not os.path.exists(docx_path):
        errors.append("Sales_Executive_Summary.docx not found")
        return errors
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        if len(text.strip()) < 50:
            errors.append("Sales_Executive_Summary.docx has too little content")
        # Quarter labels are language-neutral; require Q4 and Q1 mentioned.
        for kw in ["q4", "q1"]:
            if kw not in text:
                errors.append(f"Sales_Executive_Summary.docx missing keyword: {kw}")
    except ImportError:
        if os.path.getsize(docx_path) < 100:
            errors.append("Sales_Executive_Summary.docx too small")
    except Exception as e:
        errors.append(f"Error reading Sales_Executive_Summary.docx: {e}")
    return errors


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    all_errors = []
    critical_results = []

    print("\n=== Checking Excel ===")
    excel_errors, excel_crit = check_excel(args.agent_workspace, gt_dir)
    critical_results.extend(excel_crit)
    if excel_errors:
        for e in excel_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(excel_errors)
    else:
        print("  [PASS] Excel check passed")

    print("\n=== Checking GCal Events ===")
    gcal_errors, gcal_crit = check_gcal()
    critical_results.extend(gcal_crit)
    if gcal_errors:
        for e in gcal_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(gcal_errors)
    else:
        print("  [PASS] GCal check passed")

    print("\n=== Checking Email ===")
    email_errors, email_crit = check_email()
    critical_results.extend(email_crit)
    if email_errors:
        for e in email_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(email_errors)
    else:
        print("  [PASS] Email check passed")

    print("\n=== Checking Word Document ===")
    word_errors = check_word(args.agent_workspace)
    if word_errors:
        for e in word_errors:
            print(f"  [FAIL] {e}")
        all_errors.extend(word_errors)
    else:
        print("  [PASS] Word check passed")

    # ---- CRITICAL gate (semantic substance). Any critical failure => hard FAIL. ----
    print("\n=== CRITICAL checks ===")
    failed_critical = [m for ok, m in critical_results if not ok]
    for ok, m in critical_results:
        print(f"  [{'PASS' if ok else 'FAIL'}] {m}")

    # Total checks for accuracy: every individual error counts against a base set.
    total_checks = len(critical_results) + 6  # 4 critical-backed sections + word(2) + buffer
    accuracy = max(0.0, 100.0 * (total_checks - len(all_errors)) / total_checks) if total_checks else 0.0

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump({
                "errors": all_errors,
                "critical_failed": failed_critical,
                "accuracy": accuracy,
                "success": (not failed_critical) and accuracy >= 70.0,
            }, f, indent=2)

    if failed_critical:
        print(f"\n=== RESULT: FAIL (critical check failed: {len(failed_critical)}) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)

    print(f"\n  Accuracy: {accuracy:.1f}% ({len(all_errors)} non-critical errors of {total_checks} checks)")
    if accuracy >= 70.0:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
