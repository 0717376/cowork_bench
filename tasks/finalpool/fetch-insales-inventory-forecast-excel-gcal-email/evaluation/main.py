"""Evaluation script for fetch-insales-inventory-forecast-excel-gcal-email."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}: {detail_str}")
        if critical:
            CRITICAL_FAILED.append(name)


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def load_supplier_catalog():
    """Read the mock supplier catalog JSON that the agent fetched (single source of truth)."""
    path = os.path.join(TASK_ROOT, "tmp", "mock_pages", "api", "supplier_catalog.json")
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {s["name"]: s for s in data.get("suppliers", [])}
    except Exception as e:
        print(f"  [WARN] could not load supplier catalog: {e}")
        return {}


def header_index(ws):
    return {str(c.value).strip().lower(): i for i, c in enumerate(ws[1]) if c.value}


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    excel_path = os.path.join(agent_workspace, "Inventory_Forecast_Report.xlsx")
    check("Inventory_Forecast_Report.xlsx exists", os.path.exists(excel_path))

    stock_rows_parsed = []   # list of dicts for critical computation checks
    summary_map = {}

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---------- Stock_Status (structural) ----------
        check("Stock_Status sheet exists", "Stock_Status" in wb.sheetnames)
        if "Stock_Status" in wb.sheetnames:
            ws = wb["Stock_Status"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Stock_Status has >= 8 rows", len(data_rows) >= 8, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Product', 'Current_Stock', 'Total_Sales', 'Daily_Rate', 'Days_Remaining', 'Needs_Restock']:
                check(f"Stock_Status has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            hi = header_index(ws)
            need_cols = ['product', 'current_stock', 'total_sales', 'daily_rate', 'days_remaining', 'needs_restock']
            if all(c in hi for c in need_cols):
                for r in data_rows:
                    if r[hi['product']] is None:
                        continue
                    stock_rows_parsed.append({
                        'product': r[hi['product']],
                        'current_stock': safe_float(r[hi['current_stock']]),
                        'total_sales': safe_float(r[hi['total_sales']]),
                        'daily_rate': safe_float(r[hi['daily_rate']]),
                        'days_remaining': safe_float(r[hi['days_remaining']]),
                        'needs_restock': str(r[hi['needs_restock']]).strip().lower() if r[hi['needs_restock']] is not None else "",
                    })

        # ---------- Supplier_Info (structural) ----------
        supplier_rows = []
        check("Supplier_Info sheet exists", "Supplier_Info" in wb.sheetnames)
        if "Supplier_Info" in wb.sheetnames:
            ws = wb["Supplier_Info"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Supplier_Info has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Supplier', 'Lead_Time_Days', 'Min_Order_Qty', 'Reliability_Score']:
                check(f"Supplier_Info has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            hi = header_index(ws)
            if all(c in hi for c in ['supplier', 'lead_time_days', 'min_order_qty', 'reliability_score']):
                for r in data_rows:
                    if r[hi['supplier']] is None:
                        continue
                    supplier_rows.append(r)
                _supplier_hi = hi
            else:
                _supplier_hi = {}
        else:
            _supplier_hi = {}

        # ---------- Restock_Summary (structural) ----------
        check("Restock_Summary sheet exists", "Restock_Summary" in wb.sheetnames)
        if "Restock_Summary" in wb.sheetnames:
            ws = wb["Restock_Summary"]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Restock_Summary has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Restock_Summary has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for r in data_rows:
                if r and r[0] is not None:
                    summary_map[str(r[0]).strip().lower()] = r[1]

        # ================= CRITICAL SEMANTIC CHECKS =================

        # CRITICAL 1: Supplier_Info rows match the mock supplier catalog exactly.
        catalog = load_supplier_catalog()
        if catalog and _supplier_hi:
            hi = _supplier_hi
            present = {}
            for r in supplier_rows:
                present[str(r[hi['supplier']]).strip()] = (
                    safe_float(r[hi['lead_time_days']]),
                    safe_float(r[hi['min_order_qty']]),
                    safe_float(r[hi['reliability_score']]),
                )
            all_match = True
            mismatches = []
            for name, s in catalog.items():
                got = present.get(name)
                exp = (float(s['lead_time_days']), float(s['min_order_qty']), float(s['reliability_score']))
                if got is None or got != exp:
                    all_match = False
                    mismatches.append(f"{name}: got {got} expected {exp}")
            check("CRITICAL: Supplier_Info matches supplier_catalog.json",
                  all_match, "; ".join(mismatches), critical=True)
        else:
            check("CRITICAL: Supplier_Info matches supplier_catalog.json",
                  False, "catalog or supplier sheet unavailable", critical=True)

        # CRITICAL 2: per-row Daily_Rate / Days_Remaining / Needs_Restock formulas.
        if stock_rows_parsed:
            formula_ok = True
            bad = []
            for row in stock_rows_parsed:
                ts, cs = row['total_sales'], row['current_stock']
                dr, drem, nr = row['daily_rate'], row['days_remaining'], row['needs_restock']
                if None in (ts, cs, dr, drem):
                    formula_ok = False
                    bad.append(f"{row['product']}: missing numeric")
                    continue
                exp_dr = round(ts / 90, 2)
                if abs(dr - exp_dr) > 0.05:
                    formula_ok = False
                    bad.append(f"{row['product']}: Daily_Rate {dr} != {exp_dr}")
                    continue
                exp_drem = round(cs / dr, 1) if dr else None
                if exp_drem is None or abs(drem - exp_drem) > 0.2:
                    formula_ok = False
                    bad.append(f"{row['product']}: Days_Remaining {drem} != {exp_drem}")
                    continue
                exp_nr = (drem < 30)
                got_nr = nr in ("yes", "да")
                if got_nr != exp_nr:
                    formula_ok = False
                    bad.append(f"{row['product']}: Needs_Restock {nr} (days={drem})")
            check("CRITICAL: Stock_Status formulas (Daily_Rate/Days_Remaining/Needs_Restock)",
                  formula_ok, "; ".join(bad[:5]), critical=True)
        else:
            check("CRITICAL: Stock_Status formulas (Daily_Rate/Days_Remaining/Needs_Restock)",
                  False, "no parseable Stock_Status rows", critical=True)

        # CRITICAL 3: Restock_Summary internally consistent with Stock_Status.
        if stock_rows_parsed and summary_map:
            analyzed = len(stock_rows_parsed)
            need = sum(1 for r in stock_rows_parsed if r['needs_restock'] in ("yes", "да"))
            healthy = analyzed - need
            rems = [r['days_remaining'] for r in stock_rows_parsed if r['days_remaining'] is not None]
            avg = round(sum(rems) / len(rems), 1) if rems else None

            def gv(*keys):
                for k in keys:
                    if k in summary_map:
                        return safe_float(summary_map[k])
                return None

            consistent = True
            d = []
            if gv('total_products_analyzed') != analyzed:
                consistent = False; d.append(f"analyzed={gv('total_products_analyzed')}!={analyzed}")
            if gv('products_need_restock') != need:
                consistent = False; d.append(f"need={gv('products_need_restock')}!={need}")
            if gv('products_healthy') != healthy:
                consistent = False; d.append(f"healthy={gv('products_healthy')}!={healthy}")
            gavg = gv('avg_days_remaining')
            if avg is None or gavg is None or abs(gavg - avg) > 0.2:
                consistent = False; d.append(f"avg={gavg}!={avg}")
            check("CRITICAL: Restock_Summary consistent with Stock_Status",
                  consistent, "; ".join(d), critical=True)
        else:
            check("CRITICAL: Restock_Summary consistent with Stock_Status",
                  False, "missing summary or stock rows", critical=True)

        # inventory_forecaster.py exists (structural)
        check("inventory_forecaster.py exists",
              os.path.exists(os.path.join(agent_workspace, "inventory_forecaster.py")))

    # ================= DB CHECKS =================
    try:
        conn = get_conn()
        cur = conn.cursor()

        # Structural: an inventory-review event exists (RU or EN summary).
        cur.execute("""SELECT summary, start_datetime, end_datetime, description
                       FROM gcal.events
                       WHERE summary ILIKE %s OR summary ILIKE %s OR summary ILIKE %s""",
                    ('%inventory%', '%запас%', '%пополнен%'))
        ev_rows = cur.fetchall()
        check("Inventory review event created", len(ev_rows) > 0,
              f"found {len(ev_rows)}")

        # CRITICAL 4: event on 2026-03-12, start 09:00, end 10:00.
        ev_ok = False
        ev_detail = []
        for summary, sdt, edt, desc in ev_rows:
            if sdt is None:
                continue
            if sdt.date().isoformat() == "2026-03-12" and sdt.hour == 9 and sdt.minute == 0:
                if edt is not None and edt.hour == 10 and edt.minute == 0:
                    ev_ok = True
                    break
            ev_detail.append(f"{summary}: {sdt} -> {edt}")
        check("CRITICAL: calendar event 2026-03-12 09:00-10:00 UTC",
              ev_ok, "; ".join(ev_detail[:4]), critical=True)

        # Structural: a restock email exists (RU or EN subject).
        cur.execute("""SELECT subject, to_addr::text, body_text
                       FROM email.messages
                       WHERE subject ILIKE %s OR subject ILIKE %s OR subject ILIKE %s""",
                    ('%restock%', '%пополнен%', '%запас%'))
        mail_rows = cur.fetchall()
        check("Restock email sent", len(mail_rows) > 0, f"found {len(mail_rows)}")

        # CRITICAL 5: email to procurement@company.com, restock subject, body lists a product.
        prod_names = [r['product'] for r in stock_rows_parsed
                      if r['needs_restock'] in ("yes", "да") and r['product']]
        mail_ok = False
        mail_detail = []
        for subject, to_addr, body in mail_rows:
            to_l = (to_addr or "").lower()
            if "procurement@company.com" not in to_l:
                mail_detail.append(f"to={to_addr}")
                continue
            body_l = (body or "").lower()
            if not body_l.strip():
                mail_detail.append("empty body")
                continue
            # body must list at least one restock product (or contain a token from one)
            if prod_names:
                listed = any(
                    (str(p).lower()[:25] in body_l) or
                    any(tok for tok in str(p).lower().split() if len(tok) > 4 and tok in body_l)
                    for p in prod_names
                )
            else:
                # fallback: non-trivial body if we couldn't parse products
                listed = len(body_l) > 20
            if listed:
                mail_ok = True
                break
            mail_detail.append("body lists no restock product")
        check("CRITICAL: email -> procurement@company.com lists restock products",
              mail_ok, "; ".join(mail_detail[:4]), critical=True)

        conn.close()
    except Exception as e:
        check("DB checks", False, str(e))
        check("CRITICAL: calendar event 2026-03-12 09:00-10:00 UTC", False, str(e), critical=True)
        check("CRITICAL: email -> procurement@company.com lists restock products", False, str(e), critical=True)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    # Critical gate: any critical failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILED:
        print(f"\nCRITICAL checks failed: {CRITICAL_FAILED}")
        print(f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%) but CRITICAL gate failed")
        return False, f"CRITICAL FAIL: {CRITICAL_FAILED} | {PASS_COUNT}/{total}"

    success = accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
