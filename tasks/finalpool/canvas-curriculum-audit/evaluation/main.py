"""Evaluation for canvas-curriculum-audit (russified, keep-foreign canvas)."""
import argparse
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAIL = False


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAIL
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAIL = True
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def is_yes(v):
    return str(v).strip().lower() in ("yes", "y", "true", "да", "1")


def is_no(v):
    return str(v).strip().lower() in ("no", "n", "false", "нет", "0")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_compliance():
    """Get expected compliance data live from Canvas DB."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name,
               (SELECT COUNT(*) FROM canvas.assignments a WHERE a.course_id = c.id) as asgn,
               (SELECT COUNT(*) FROM canvas.quizzes q WHERE q.course_id = c.id) as quiz,
               (SELECT COUNT(*) FROM canvas.modules m WHERE m.course_id = c.id) as mods,
               CASE WHEN c.syllabus_body IS NOT NULL AND c.syllabus_body != '' THEN true ELSE false END as has_syl
        FROM canvas.courses c ORDER BY c.name
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    result = {}
    for name, asgn, quiz, mods, syl in rows:
        compliant = asgn >= 8 and quiz >= 3 and mods >= 4 and syl
        result[name.lower()] = {
            "assignments": asgn, "quizzes": quiz, "modules": mods,
            "syllabus": syl, "compliant": compliant,
        }
    return result


def find_col(header_lower, *keys):
    for i, h in enumerate(header_lower):
        if any(k in h for k in keys):
            return i
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Curriculum_Audit.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Curriculum_Audit.xlsx exists", False, f"Not found: {xlsx_path}", critical=True)
        return
    check("Curriculum_Audit.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e), critical=True)
        return
    check("Excel readable", True)

    expected = get_expected_compliance()
    compliant_expected = sum(1 for v in expected.values() if v["compliant"])
    non_compliant_expected = sum(1 for v in expected.values() if not v["compliant"])
    rate_expected = round(compliant_expected / 22 * 100)

    # --- Compliance Matrix sheet ---
    cm_rows = load_sheet_rows(wb, "Compliance Matrix")
    if cm_rows is None:
        check("Sheet 'Compliance Matrix' exists", False, f"Available: {wb.sheetnames}", critical=True)
    else:
        check("Sheet 'Compliance Matrix' exists", True)
        data_rows = [r for r in cm_rows[1:] if r and r[0]] if len(cm_rows) > 1 else []
        check("Compliance Matrix has 22 rows", len(data_rows) == 22, f"Found {len(data_rows)}")

        header = cm_rows[0] if cm_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["course", "assignments_count", "quizzes_count", "modules_count",
                    "has_syllabus", "compliant_yn", "issues"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

        # Resolve column indices (fallback to fixed order if headers unclear)
        c_asgn = find_col(header_lower, "assignments_count", "assignments") or 1
        c_quiz = find_col(header_lower, "quizzes_count", "quizzes") or 2
        c_mods = find_col(header_lower, "modules_count", "modules") or 3
        c_syl = find_col(header_lower, "has_syllabus", "syllabus") or 4
        c_comp = find_col(header_lower, "compliant_yn", "compliant") or 5
        c_iss = find_col(header_lower, "issues")

        # CRITICAL: full per-row recompute against live Canvas DB
        matched = 0
        counts_ok = True
        compliant_flag_ok = True
        bad_count = []
        bad_flag = []
        for row in data_rows:
            course_key = str(row[0]).strip().lower()
            if course_key not in expected:
                continue
            matched += 1
            exp = expected[course_key]
            ok_counts = (num_close(row[c_asgn], exp["assignments"]) and
                         num_close(row[c_quiz], exp["quizzes"]) and
                         num_close(row[c_mods], exp["modules"]) and
                         (is_yes(row[c_syl]) == exp["syllabus"]))
            if not ok_counts:
                counts_ok = False
                if len(bad_count) < 5:
                    bad_count.append(
                        f"{course_key}: got a={row[c_asgn]} q={row[c_quiz]} "
                        f"m={row[c_mods]} syl={row[c_syl]} | exp a={exp['assignments']} "
                        f"q={exp['quizzes']} m={exp['modules']} syl={exp['syllabus']}")
            # Compliant_YN must equal the threshold rule
            flag_val = is_yes(row[c_comp]) if (is_yes(row[c_comp]) or is_no(row[c_comp])) else None
            if flag_val is None or flag_val != exp["compliant"]:
                compliant_flag_ok = False
                if len(bad_flag) < 5:
                    bad_flag.append(f"{course_key}: got {row[c_comp]} exp {exp['compliant']}")

        check("All 22 courses matched to live Canvas data", matched == 22,
              f"Matched {matched}/22", critical=True)
        check("Counts (assignments/quizzes/modules/syllabus) match live Canvas for every row",
              counts_ok, "; ".join(bad_count), critical=True)
        check("Compliant_YN equals (asgn>=8 AND quiz>=3 AND mods>=4 AND has_syllabus) for every row",
              compliant_flag_ok, "; ".join(bad_flag), critical=True)

        # CRITICAL: Issues column reflects real deficiencies for non-compliant courses
        if c_iss is not None:
            issues_ok = True
            bad_iss = []
            for row in data_rows:
                course_key = str(row[0]).strip().lower()
                if course_key not in expected:
                    continue
                exp = expected[course_key]
                iss_text = str(row[c_iss] or "").lower()
                if exp["compliant"]:
                    continue
                # Build the set of failing requirements
                fails = []
                if exp["assignments"] < 8:
                    fails.append(("assign", "задани"))
                if exp["quizzes"] < 3:
                    fails.append(("quiz", "тест"))
                if exp["modules"] < 4:
                    fails.append(("modul", "модул"))
                if not exp["syllabus"]:
                    fails.append(("syllab", "учебн", "план"))
                # Each failing requirement should be mentioned (EN or RU keyword)
                if iss_text.strip() in ("", "none", "нет"):
                    issues_ok = False
                    if len(bad_iss) < 5:
                        bad_iss.append(f"{course_key}: empty Issues but non-compliant")
                    continue
                for kw_group in fails:
                    if not any(k in iss_text for k in kw_group):
                        issues_ok = False
                        if len(bad_iss) < 5:
                            bad_iss.append(f"{course_key}: missing {kw_group} in '{iss_text[:80]}'")
                        break
            check("Issues column documents the actual failing requirement(s) for non-compliant courses",
                  issues_ok, "; ".join(bad_iss), critical=True)

        # Alphabetical sort (non-critical)
        names = [str(r[0]).strip() for r in data_rows]
        check("Courses sorted alphabetically by name",
              names == sorted(names, key=lambda s: s.lower()),
              f"First few: {names[:3]}")

    # --- Summary sheet ---
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}", critical=True)
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        check("Total_Courses = 22", num_close(lookup.get("total_courses"), 22),
              f"Got {lookup.get('total_courses')}")
        # CRITICAL: counts derived live from DB, sum to 22, and rate matches
        check(f"Compliant_Courses = {compliant_expected} (live)",
              num_close(lookup.get("compliant_courses"), compliant_expected),
              f"Got {lookup.get('compliant_courses')}", critical=True)
        check(f"Non_Compliant_Courses = {non_compliant_expected} (live)",
              num_close(lookup.get("non_compliant_courses"), non_compliant_expected),
              f"Got {lookup.get('non_compliant_courses')}", critical=True)
        # Compliance_Rate as percentage; accept value with/without % sign, round or floor
        rate_raw = lookup.get("compliance_rate")
        rate_num = None
        if rate_raw is not None:
            m = re.search(r"[-+]?\d*\.?\d+", str(rate_raw))
            if m:
                rate_num = float(m.group())
        import math
        floor_rate = math.floor(compliant_expected / 22 * 100)
        check("Compliance_Rate matches compliant/22 (round or floor)",
              rate_num is not None and (num_close(rate_num, rate_expected) or
                                        num_close(rate_num, floor_rate)),
              f"Got {rate_raw}, expected ~{rate_expected}")


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Audit_Report.docx")
    if not os.path.isfile(docx_path):
        check("Audit_Report.docx exists", False, f"Not found: {docx_path}", critical=True)
        return
    check("Audit_Report.docx exists", True)
    check("Word doc has content (> 1KB)", os.path.getsize(docx_path) > 1000,
          f"Size: {os.path.getsize(docx_path)}")

    try:
        from docx import Document
        doc = Document(docx_path)
        all_text = " ".join(p.text for p in doc.paragraphs).lower()
        has_compliance = ("compli" in all_text) or ("соответств" in all_text)
        has_recommend = ("recommend" in all_text) or ("рекоменд" in all_text)
        has_rate = re.search(r"\d{1,3}\s*%", all_text) is not None
        check("Report mentions compliance / соответствие", has_compliance,
              f"Sample: {all_text[:200]}")
        check("Report mentions recommendations / рекомендации", has_recommend,
              f"Sample: {all_text[:200]}")
        # CRITICAL: narrative is non-trivial and surfaces a numeric compliance rate
        check("Audit report is a non-trivial narrative with compliance/соответствие, "
              "recommendations/рекомендации AND a numeric % rate",
              has_compliance and has_recommend and has_rate and len(all_text) > 400,
              f"len={len(all_text)} rate={has_rate}", critical=True)
    except ImportError:
        check("python-docx available", False, "pip install python-docx")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    if CRITICAL_FAIL:
        print("CRITICAL check failed -> FAIL regardless of accuracy")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS (accuracy >= 70 and no critical failures)")
        sys.exit(0)
    else:
        print(f"FAIL (accuracy {accuracy:.1f}% < 70)")
        sys.exit(1)


if __name__ == "__main__":
    main()
