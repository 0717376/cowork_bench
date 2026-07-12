"""Evaluation script for fetch-arxiv-lit-review-word-gsheet.

Критические проверки (CRITICAL_CHECKS): любой провал => общий FAIL,
независимо от accuracy. Иначе PASS требует accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Реальные релевантные статьи arxiv, которые должны попасть в анализ
# (английские идентификаторы сохраняются).
REAL_PAPER_IDS = {"2301.01234", "2302.05678", "2303.09012", "2304.03456"}
# Шумовая статья (quant-ph), которую агент обязан ИСКЛЮЧИТЬ.
NOISE_PAPER_ID = "9999.99999"

# Критические проверки: любой провал => общий FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Lit_Review_Report.xlsx exists",
    "Data_Analysis содержит все 4 релевантных Paper_ID и исключает шумовую статью",
    "Data_Analysis отсортирован по возрастанию основного измерения",
    "Metrics внутренне согласован (Total_Papers и Avg_Citations)",
    "Google Sheet 'Arxiv Litreview Tracker' содержит >= 2 реальных Paper_ID",
    "Word-документ содержит три обязательных раздела (резюме/выводы/рекомендации)",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Lit_Review_Report.xlsx")
    check("Lit_Review_Report.xlsx exists", os.path.exists(excel_path))

    da_rows = []        # строки данных Data_Analysis
    da_headers = []
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # --- Data_Analysis ---
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 4 rows", len(da_rows) >= 4, f"got {len(da_rows)}")

            da_headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Paper_ID', 'Title', 'Area', 'Citations', 'Relevance_Score']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in da_headers, f"headers: {da_headers[:8]}")

            # --- КРИТИЧЕСКАЯ: правильное множество Paper_ID ---
            def col_idx(name):
                try:
                    return da_headers.index(name.lower())
                except ValueError:
                    return None
            pid_idx = col_idx('Paper_ID')
            present_ids = set()
            if pid_idx is not None:
                for r in da_rows:
                    if pid_idx < len(r) and r[pid_idx] is not None:
                        present_ids.add(str(r[pid_idx]).strip())
            has_all_real = REAL_PAPER_IDS.issubset(present_ids)
            excludes_noise = NOISE_PAPER_ID not in present_ids
            check("Data_Analysis содержит все 4 релевантных Paper_ID и исключает шумовую статью",
                  has_all_real and excludes_noise,
                  f"present={sorted(present_ids)}")

            # --- КРИТИЧЕСКАЯ: сортировка по возрастанию основного измерения ---
            # Принимается сортировка по возрастанию по Area ИЛИ по Paper_ID
            # (задача допускает любое из основных измерений).
            def is_col_sorted(idx):
                if idx is None:
                    return False
                keys = [str(r[idx]).strip().lower()
                        for r in da_rows if idx < len(r) and r[idx] is not None]
                return bool(keys) and keys == sorted(keys)
            is_sorted = is_col_sorted(col_idx('Area')) or is_col_sorted(pid_idx)
            check("Data_Analysis отсортирован по возрастанию основного измерения",
                  is_sorted, "по столбцу Area или Paper_ID")

            # Внутренняя согласованность Metrics опирается на эти столбцы.
            cit_idx = col_idx('Citations')
            citations = []
            if cit_idx is not None:
                for r in da_rows:
                    if cit_idx < len(r):
                        v = safe_float(r[cit_idx])
                        if v is not None:
                            citations.append(v)
        else:
            check("Data_Analysis содержит все 4 релевантных Paper_ID и исключает шумовую статью",
                  False, "нет листа Data_Analysis")
            check("Data_Analysis отсортирован по возрастанию основного измерения",
                  False, "нет листа Data_Analysis")
            citations = []

        # --- Metrics ---
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            m_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(m_rows) >= 4, f"got {len(m_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            for r in m_rows:
                if r and r[0] is not None and len(r) >= 2:
                    metrics_map[str(r[0]).strip().lower()] = r[1]

        # --- КРИТИЧЕСКАЯ: внутренняя согласованность Metrics ---
        n_data = len(da_rows)
        total_papers = safe_float(metrics_map.get('total_papers'))
        avg_cit = safe_float(metrics_map.get('avg_citations'))
        expected_avg = (sum(citations) / len(citations)) if citations else None
        total_ok = (total_papers is not None and abs(total_papers - n_data) <= 1)
        avg_ok = (avg_cit is not None and expected_avg is not None
                  and abs(avg_cit - expected_avg) <= 1)
        check("Metrics внутренне согласован (Total_Papers и Avg_Citations)",
              total_ok and avg_ok,
              f"Total_Papers={total_papers} vs {n_data}; Avg_Citations={avg_cit} vs {expected_avg}")

        # --- Recommendations ---
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            r_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(r_rows) >= 2, f"got {len(r_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Area']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # --- Google Sheet ---
        gsheet_ids = set()
        sheets = []
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute("SELECT id, title FROM gsheet.spreadsheets WHERE title ILIKE %s",
                        ('%tracker%',))
            sheets = cur.fetchall()
            check("Google Sheet created (название содержит 'tracker')",
                  len(sheets) >= 1, f"found {len(sheets)} sheets")

            all_values = ""
            if sheets:
                ss_id = sheets[0][0]
                cur.execute("""
                    SELECT c.value FROM gsheet.cells c
                    JOIN gsheet.sheets s ON c.spreadsheet_id = s.spreadsheet_id AND c.sheet_id = s.id
                    WHERE c.spreadsheet_id = %s
                """, (ss_id,))
                cells = cur.fetchall()
                all_values = " ".join(str(x[0]) for x in cells if x[0] is not None)
            n_ref = sum(1 for pid in REAL_PAPER_IDS if pid in all_values)
            check("Google Sheet 'Arxiv Litreview Tracker' содержит >= 2 реальных Paper_ID",
                  n_ref >= 2, f"найдено {n_ref} Paper_ID в ячейках")
            conn.close()
        except Exception as e:
            check("Google Sheet created (название содержит 'tracker')", False, str(e))
            check("Google Sheet 'Arxiv Litreview Tracker' содержит >= 2 реальных Paper_ID",
                  False, str(e))

        # --- Word ---
        import glob as globmod
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        word_text = ""
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            word_text = " ".join(p.text for p in doc.paragraphs).lower()
            check("Word has content", len(word_text) > 50, f"text length: {len(word_text)}")

        # --- КРИТИЧЕСКАЯ: три обязательных раздела (RU+EN) ---
        required_sections = [
            ["executive summary", "краткое резюме", "резюме", "overview"],
            ["key findings", "ключевые выводы", "выводы"],
            ["recommendations", "рекомендации"],
        ]
        sections_ok = all(any(alt in word_text for alt in group) for group in required_sections)
        check("Word-документ содержит три обязательных раздела (резюме/выводы/рекомендации)",
              sections_ok, "проверка по заголовкам разделов")

        # --- Скрипт обработки ---
        check("arxiv_litreview_processor.py exists",
              os.path.exists(os.path.join(agent_workspace, "arxiv_litreview_processor.py")))
    else:
        # Без xlsx критические проверки, зависящие от него, отмечаем как провал.
        for n in ["Data_Analysis содержит все 4 релевантных Paper_ID и исключает шумовую статью",
                  "Data_Analysis отсортирован по возрастанию основного измерения",
                  "Metrics внутренне согласован (Total_Papers и Avg_Citations)",
                  "Google Sheet 'Arxiv Litreview Tracker' содержит >= 2 реальных Paper_ID",
                  "Word-документ содержит три обязательных раздела (резюме/выводы/рекомендации)"]:
            check(n, False, "нет Lit_Review_Report.xlsx")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({"passed": PASS_COUNT, "total": total,
                           "accuracy": accuracy, "critical_failed": critical_failed},
                          f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
