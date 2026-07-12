"""
Evaluation for yt-veritasium-gcal-gform task.

Checks:
1. Viewing_Schedule.xlsx with Sessions sheet: 5 rows, correct columns
2. GCal has 5 Science Enrichment sessions on consecutive Wednesdays in April 2026
3. GForm 'Science Enrichment Program Feedback' with 4 questions
4. Email to students@school.edu with correct subject listing 5 dates
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []

EXPECTED_DATES = {"2026-04-01", "2026-04-08", "2026-04-15", "2026-04-22", "2026-04-29"}


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILURES.append(name)


def check_excel(agent_workspace):
    print("\n=== Check 1: Viewing_Schedule.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Viewing_Schedule.xlsx")
    if not os.path.exists(xlsx_path):
        record("Viewing_Schedule.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Viewing_Schedule.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    sess_idx = next((i for i, s in enumerate(sheet_names_lower) if "session" in s), None)
    if sess_idx is None:
        record("Sessions sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Sessions sheet exists", True)

    ws = wb[wb.sheetnames[sess_idx]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        record("Has data", False, "Sheet is empty")
        return

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    has_session_num = any("session" in h or "num" in h for h in headers)
    has_date = any("date" in h for h in headers)
    has_title = any("title" in h for h in headers)
    has_view = any("view" in h for h in headers)
    has_duration = any("duration" in h for h in headers)
    record("Sessions sheet has required columns (Session_Num, Date, Title, View_Count, Duration_Min)",
           has_session_num and has_date and has_title and has_view and has_duration,
           f"Headers: {rows[0]}")

    data_rows = [r for r in rows[1:] if any(c for c in r)]
    record("Sessions sheet has exactly 5 data rows", len(data_rows) == 5,
           f"Found {len(data_rows)} data rows", critical=True)

    # The five expected top-5 Veritasium 2025 video IDs (by view count) must appear.
    all_text = " ".join(str(c) for r in rows for c in r if c)
    expected_video_ids = {"SC2eSujzrUY", "88bMVbx1dzM", "Q56PMJbCFXQ", "qJZ1Ez28C-A", "Ni82f1-cAXg"}
    ids_found = sum(1 for vid in expected_video_ids if vid in all_text)
    record("Sessions data references the correct top-5 Veritasium 2025 videos (>=4 video IDs)",
           ids_found >= 4,
           f"Found {ids_found}/5 expected video IDs", critical=True)

    # Check dates are present
    dates_in_sheet = set()
    for expected_date in EXPECTED_DATES:
        if expected_date in " ".join(str(c) for r in rows for c in r if c):
            dates_in_sheet.add(expected_date)
    record("Sessions sheet contains expected April 2026 dates (>=4)",
           len(dates_in_sheet) >= 4,
           f"Found dates: {dates_in_sheet}")


def check_gcal():
    print("\n=== Check 2: Google Calendar - Science Enrichment sessions ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, description, start_datetime, end_datetime
        FROM gcal.events
        WHERE start_datetime >= '2026-04-01' AND start_datetime < '2026-05-01'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    science_events = [e for e in events if "science enrichment" in (e[0] or "").lower()]
    record("At least 5 Science Enrichment events in April 2026",
           len(science_events) >= 5,
           f"Found {len(science_events)} out of {len(events)} total April events",
           critical=True)

    if science_events:
        # Check duration 14:00-15:30 = 90 minutes
        summary, desc, start_dt, end_dt = science_events[0]
        if start_dt and end_dt:
            duration_min = (end_dt - start_dt).total_seconds() / 60
            record("Events run 14:00-15:30 (90 minutes)", 85 <= duration_min <= 95,
                   f"Duration: {duration_min:.0f} min")

        # Check dates are correct Wednesdays
        event_dates = set(e[2].strftime("%Y-%m-%d") for e in science_events if e[2])
        dates_match = len(event_dates & EXPECTED_DATES)
        record(f"Events are on correct Wednesdays (>=4 match expected dates)",
               dates_match >= 4,
               f"Found: {event_dates}, Expected: {EXPECTED_DATES}",
               critical=True)

        # Check descriptions mention view count (a number) or duration
        desc_text = " ".join((e[1] or "") for e in science_events)
        has_numbers = any(c.isdigit() for c in desc_text)
        record("Event descriptions contain numeric data (view count or duration)",
               has_numbers, "No numeric data in event descriptions")


def check_gform():
    print("\n=== Check 3: Google Form - Science Enrichment Program Feedback ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT id, title FROM gform.forms
        WHERE title ILIKE '%science enrichment%'
        ORDER BY id DESC LIMIT 5
    """)
    forms = cur.fetchall()
    record("'Science Enrichment Program Feedback' form exists", len(forms) > 0,
           f"No matching form found", critical=True)

    if forms:
        form_id = forms[0][0]
        cur.execute("""
            SELECT title, question_type, config FROM gform.questions
            WHERE form_id = %s ORDER BY position
        """, (form_id,))
        questions = cur.fetchall()
        record("Form has exactly 4 questions", len(questions) == 4,
               f"Found {len(questions)} questions", critical=True)

        if questions:
            q_titles = [q[0].lower() for q in questions]
            has_topic_q = any("topic" in t or "interested" in t for t in q_titles)
            has_difficulty_q = any("difficult" in t or "level" in t for t in q_titles)
            has_frequency_q = any("often" in t or "frequently" in t or "how often" in t for t in q_titles)
            has_text_q = any("other" in t or "explore" in t for t in q_titles)
            record("Form has required question topics (topic, difficulty, frequency, open-text)",
                   has_topic_q and has_difficulty_q and has_frequency_q and has_text_q,
                   f"Questions: {q_titles}")

            # Check text question type. The forms MCP stores question_type as
            # 'textQuestion' (open-ended) and 'choiceQuestion' (multiple choice).
            q_types = [q[1] for q in questions]
            has_text_type = any(t in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH")
                                for t in q_types)
            has_choice_type = any(t in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE")
                                  for t in q_types)
            record("Form has at least one text/open-ended question (textQuestion)", has_text_type,
                   f"Question types: {q_types}", critical=True)
            record("Form has multiple-choice questions (choiceQuestion)", has_choice_type,
                   f"Question types: {q_types}")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 4: Email to students@school.edu ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT to_addr, subject, body_text FROM email.messages
        WHERE to_addr::text ILIKE '%students@school.edu%'
        ORDER BY id DESC LIMIT 5
    """)
    emails = cur.fetchall()
    cur.close()
    conn.close()

    record("Email to students@school.edu exists", len(emails) > 0,
           "No email to students@school.edu found", critical=True)

    if emails:
        to_addr, subject, body = emails[0]
        record("Email subject is 'Science Enrichment Sessions Scheduled'",
               "science enrichment sessions scheduled" in subject.lower(),
               f"Subject: {subject}", critical=True)

        body_lower = (body or "").lower()
        # Body should list dates
        dates_mentioned = sum(1 for d in ["april 1", "april 8", "april 15", "april 22", "april 29",
                                          "2026-04-01", "2026-04-08", "2026-04-15", "2026-04-22", "2026-04-29"]
                              if d in body_lower)
        record("Email body lists at least 4 session dates", dates_mentioned >= 4,
               f"Dates found: {dates_mentioned}")

        has_videos = any(kw in body_lower for kw in
                         ["planet", "quantum", "glue", "skyscraper", "zooming", "veritasium"])
        record("Email body references video titles", has_videos,
               f"Body excerpt: {body_lower[:300]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_gform()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    # CRITICAL gate: core semantic deliverables must all be present, regardless
    # of overall accuracy. A non-doer (or partial attempt) fails immediately.
    if CRITICAL_FAILURES:
        print(f"\nFAIL: {len(CRITICAL_FAILURES)} critical check(s) failed: {CRITICAL_FAILURES}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
