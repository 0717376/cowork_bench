"""
Evaluation script for fetch-sf-hr-benchmark-excel-word task (ClickHouse fork).

Checks:
1. HR_Salary_Benchmark.xlsx with 3 sheets and correct data
2. Salary_Benchmark_Report.docx with executive summary content
3. Email sent with benchmark analysis subject

Department literals are CYRILLIC: the sf_data HR_ANALYTICS departments are russified
centrally (db/zzz_clickhouse_after_init.sql) and the task-local mock benchmark JSON is
russified in lockstep, so the benchmark sheet, the Internal Data sheet (read live from
the DWH) and the eval keys all share the same Cyrillic labels.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRIT]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL]{' [CRIT]' if critical else ''} {name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=500.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains(haystack, needle):
    if haystack is None or needle is None:
        return False
    return needle.strip().lower() in str(haystack).strip().lower()


# Expected data based on actual DWH (sf_data, ClickHouse fork) values and benchmark API.
# Keys are the russified department labels returned by the DWH and the mock JSON.
BENCHMARK_AVGS = {
    "инженерия": 62000.00,
    "финансы": 55000.00,
    "кадры": 54000.00,
    "операции": 56000.00,
    "ниокр": 61000.00,
    "продажи": 57000.00,
    "поддержка": 52000.00,
}

INTERNAL_AVGS = {
    "инженерия": 58991.61,
    "финансы": 57878.19,
    "кадры": 58920.45,
    "операции": 57808.74,
    "ниокр": 57905.93,
    "продажи": 58864.79,
    "поддержка": 58400.48,
}

EXPECTED_HEADCOUNTS = {
    "инженерия": 7096,
    "финансы": 7148,
    "кадры": 7077,
    "операции": 7120,
    "ниокр": 7083,
    "продажи": 7232,
    "поддержка": 7244,
}

# Largest positive variance: Поддержка (Support) ~ +6400.48
# Largest negative variance: НИОКР (R&D) ~ -3094.07
TOP_POS_DEPT = "поддержка"
TOP_NEG_DEPT = "ниокр"


def check_excel(agent_workspace):
    """Check HR_Salary_Benchmark.xlsx."""
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "HR_Salary_Benchmark.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}", critical=True)
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e), critical=True)
        return

    # --- Sheet 1: Industry Benchmarks ---
    bench_sheet = None
    for name in wb.sheetnames:
        if "benchmark" in name.lower() or "industry" in name.lower():
            bench_sheet = name
            break
    if not bench_sheet:
        record("Industry Benchmarks sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Industry Benchmarks sheet exists", True)
        ws = wb[bench_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        record("Benchmarks sheet has 7 rows", len(data_rows) == 7, f"Found {len(data_rows)}")

        for row in data_rows:
            if row and row[0]:
                dept = str(row[0]).strip().lower()
                if dept in BENCHMARK_AVGS:
                    expected = BENCHMARK_AVGS[dept]
                    found = any(num_close(cell, expected, tol=100) for cell in row[1:])
                    if not found:
                        record(f"Benchmark avg for {dept}", False,
                               f"Expected ~{expected}, row: {row[:5]}")

    # --- Sheet 2: Internal Data --- (CRITICAL: live DWH avg + headcount)
    int_sheet = None
    for name in wb.sheetnames:
        if "internal" in name.lower():
            int_sheet = name
            break
    if not int_sheet:
        record("Internal Data sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Internal Data sheet exists", True)
        ws = wb[int_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        record("Internal Data sheet has 7 rows", len(data_rows) == 7, f"Found {len(data_rows)}")

        seen_avg = {}
        seen_headcount = {}
        for row in data_rows:
            if row and row[0]:
                dept = str(row[0]).strip().lower()
                if dept in INTERNAL_AVGS:
                    seen_avg[dept] = any(
                        num_close(cell, INTERNAL_AVGS[dept], tol=500) for cell in row[1:]
                    )
                    seen_headcount[dept] = any(
                        num_close(cell, EXPECTED_HEADCOUNTS[dept], tol=50) for cell in row[1:]
                    )

        # CRITICAL: every department's internal avg matches the live DWH value
        avg_all_ok = len(seen_avg) == 7 and all(seen_avg.values())
        record("Internal_Avg_Salary correct for all 7 departments", avg_all_ok,
               f"Per-dept avg matches: {seen_avg}", critical=True)

        # CRITICAL: headcounts match the DWH (currently grouped by russified dept)
        hc_all_ok = len(seen_headcount) == 7 and all(seen_headcount.values())
        record("Headcount correct for all 7 departments", hc_all_ok,
               f"Per-dept headcount matches: {seen_headcount}", critical=True)

    # --- Sheet 3: Variance Analysis --- (CRITICAL: Support / R&D variance)
    var_sheet = None
    for name in wb.sheetnames:
        if "variance" in name.lower() or "analysis" in name.lower():
            var_sheet = name
            break
    if not var_sheet:
        record("Variance Analysis sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Variance Analysis sheet exists", True)
        ws = wb[var_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        record("Variance sheet has 7 rows", len(data_rows) == 7, f"Found {len(data_rows)}")

        pos_amount = 6400.48   # Поддержка (Support), largest positive
        neg_amount = -3094.07  # НИОКР (R&D), largest negative
        pos_pct = pos_amount / BENCHMARK_AVGS[TOP_POS_DEPT] * 100   # ~12.31
        neg_pct = neg_amount / BENCHMARK_AVGS[TOP_NEG_DEPT] * 100   # ~-5.07

        pos_amount_ok = neg_amount_ok = False
        pos_pct_ok = neg_pct_ok = False
        for row in data_rows:
            if row and row[0]:
                dept = str(row[0]).strip().lower()
                if dept == TOP_POS_DEPT:
                    pos_amount_ok = any(num_close(c, pos_amount, tol=800) for c in row[1:])
                    pos_pct_ok = any(num_close(c, pos_pct, tol=2.0) for c in row[1:])
                elif dept == TOP_NEG_DEPT:
                    neg_amount_ok = any(num_close(c, neg_amount, tol=800) for c in row[1:])
                    neg_pct_ok = any(num_close(c, neg_pct, tol=2.0) for c in row[1:])

        record("Support (поддержка) variance amount correct (largest positive)",
               pos_amount_ok, f"Expected ~{pos_amount}", critical=True)
        record("R&D (ниокр) variance amount correct (largest negative)",
               neg_amount_ok, f"Expected ~{neg_amount}", critical=True)
        record("Support (поддержка) variance pct correct", pos_pct_ok,
               f"Expected ~{pos_pct:.2f}")
        record("R&D (ниокр) variance pct correct", neg_pct_ok,
               f"Expected ~{neg_pct:.2f}")

    wb.close()


def check_word(agent_workspace):
    """Check Salary_Benchmark_Report.docx."""
    print("\n=== Checking Word Output ===")

    fpath = os.path.join(agent_workspace, "Salary_Benchmark_Report.docx")
    if not os.path.isfile(fpath):
        record("Word file exists", False, f"Not found: {fpath}", critical=True)
        return

    record("Word file exists", True)

    try:
        from docx import Document
        doc = Document(fpath)
        full_text = "\n".join(p.text for p in doc.paragraphs).lower()
    except Exception as e:
        record("Word file readable", False, str(e), critical=True)
        return

    # Department-name mentions (Cyrillic). 'кадры' is HR; accept 'hr' too just in case.
    checks = [
        ("Mentions Инженерия", "инженерия" in full_text),
        ("Mentions Финансы", "финанс" in full_text),
        ("Mentions Кадры (HR)", "кадры" in full_text or "hr" in full_text),
        ("Mentions Поддержка (Support)", "поддержка" in full_text or "support" in full_text),
        ("Mentions НИОКР (R&D)", "ниокр" in full_text or "r&d" in full_text),
        ("Mentions benchmark/эталон/бенчмарк",
         any(k in full_text for k in ("benchmark", "эталон", "бенчмарк"))),
        ("Mentions variance/comparison/отклонение/сравнение",
         any(k in full_text for k in ("variance", "comparison", "compared", "difference",
                                      "отклонен", "сравнен", "разниц"))),
    ]
    for name, cond in checks:
        record(name, cond)

    # CRITICAL: above/below classification names BOTH top depts correctly.
    above_below_terms = any(k in full_text for k in ("выше", "ниже", "above", "below"))
    pos_named = TOP_POS_DEPT in full_text or "support" in full_text
    neg_named = TOP_NEG_DEPT in full_text or "r&d" in full_text
    record("Word names above/below classification for Support and R&D",
           above_below_terms and pos_named and neg_named,
           f"above/below_terms={above_below_terms}, support={pos_named}, r&d={neg_named}",
           critical=True)


def check_email():
    """Check email was sent with benchmark analysis."""
    print("\n=== Checking Email ===")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Email DB accessible", False, str(e), critical=True)
        return

    target = None
    for subject, from_addr, to_addr, body_text in emails:
        subj = (subject or "")
        if "annual salary benchmark analysis report" in subj.lower():
            target = (subject, from_addr, to_addr, body_text)
            break
    # Fallback: any benchmark/salary subject (non-critical leniency for matching the row)
    if target is None:
        for subject, from_addr, to_addr, body_text in emails:
            sl = (subject or "").lower()
            if "benchmark" in sl or "salary" in sl:
                target = (subject, from_addr, to_addr, body_text)
                break

    if target is None:
        record("Benchmark analysis email exists", False,
               f"Found {len(emails)} emails, none matching report subject", critical=True)
        return

    subject, from_addr, to_addr, body_text = target
    record("Benchmark analysis email exists", True)

    # CRITICAL: exact subject literal
    subj_ok = "annual salary benchmark analysis report" in (subject or "").lower()
    record("Email subject is 'Annual Salary Benchmark Analysis Report'", subj_ok,
           f"Subject: {subject}", critical=True)

    # From address is non-critical: the mail MCP may rewrite it to the authenticated
    # account regardless of the requested sender. Keep a lenient HR check only.
    from_ok = str_contains(from_addr, "hr")
    record("Email from HR-related address", from_ok, f"From: {from_addr}")

    # CRITICAL: addressed to vp-hr@company.com
    to_ok = str_contains(to_addr, "vp-hr@company.com")
    record("Email to vp-hr@company.com", to_ok, f"To: {to_addr}", critical=True)

    # CRITICAL: body names BOTH the largest-positive (Поддержка) and
    # largest-negative (НИОКР) variance departments.
    body_lower = (body_text or "").lower()
    pos_named = TOP_POS_DEPT in body_lower or "support" in body_lower
    neg_named = TOP_NEG_DEPT in body_lower or "r&d" in body_lower
    record("Email body names both Support (largest +) and R&D (largest -)",
           pos_named and neg_named,
           f"support={pos_named}, r&d={neg_named}; body: {(body_text or '')[:200]}",
           critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if CRITICAL_FAILS:
        print(f"  CRITICAL FAILURES: {CRITICAL_FAILS}")

    # CRITICAL gate: any critical failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILS:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70
    print(f"  Overall: {'PASS' if overall else 'FAIL'} (threshold accuracy>=70)")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
