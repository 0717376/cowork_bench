"""
Evaluation for arxiv-team-reading-gform-gcal task (RU stack: forms + google_calendar).

The agent must produce, from scratch:
1. Reading_List.xlsx (sheet Papers) with EXACTLY 7 data rows; exactly 5 papers
   classified Topic=LLM_Reasoning and exactly 2 classified Topic=Other (the two
   noise papers: Vision Transformers and Medical Image Segmentation).
2. A form 'Reading Group Paper Selection' (gform.* / RU forms-mcp backend) with
   a paper-priority multiple-choice question listing the 5 LLM reasoning titles
   (NOT the 2 noise papers), an aspects question (Reasoning Methods / Implementation
   / Evaluation / Theory), and an availability question (Monday..Friday).
3. Four 'Reading Group Session' calendar events on 2026-04-07/14/21/28, each
   lasting 1.5 hours.

CRITICAL_CHECKS reflect the task's substance: a single critical failure => overall
FAIL (sys.exit(1)) regardless of accuracy. The accuracy>=70 gate applies afterward.

Note on the RU forms MCP (local_servers/forms-mcp): it stores choice questions as
question_type='choiceQuestion' with config={'type':'RADIO','options':[{'value':...}]}.
There is no separate CHECKBOX type, so the availability question is validated by its
options (Monday..Friday), not by a checkbox type literal.
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# The two noise papers that MUST be classified as Other (not LLM_Reasoning).
NOISE_KEYS = [
    ["vision transformer"],            # Efficient Training of Vision Transformers
    ["medical image", "segmentation"], # Medical Image Segmentation with Deep Learning
]
# The five LLM reasoning paper titles (substrings) the priority question must offer.
LLM_TITLE_KEYS = [
    "chain-of-thought",
    "self-consistency",
    "tree of thoughts",
    "least-to-most",
    "react",
]


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def check_excel(agent_workspace):
    print("\n=== Check 1: Excel Reading_List.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Reading_List.xlsx")
    if not os.path.exists(xlsx_path):
        record("Reading_List.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Excel: ровно 7 строк, 5 LLM_Reasoning + 2 Other", False,
               "no file", critical=True)
        record("Excel: 2 шумовые статьи помечены как Other", False,
               "no file", critical=True)
        return
    record("Reading_List.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel file readable", False, str(e))
        record("Excel: ровно 7 строк, 5 LLM_Reasoning + 2 Other", False,
               "unreadable", critical=True)
        record("Excel: 2 шумовые статьи помечены как Other", False,
               "unreadable", critical=True)
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    if "papers" not in sheet_names_lower:
        record("Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Excel: ровно 7 строк, 5 LLM_Reasoning + 2 Other", False,
               "no Papers sheet", critical=True)
        record("Excel: 2 шумовые статьи помечены как Other", False,
               "no Papers sheet", critical=True)
        return
    record("Papers sheet exists", True)

    ws = wb[wb.sheetnames[sheet_names_lower.index("papers")]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        record("Has data rows", False, "Sheet is empty")
        record("Excel: ровно 7 строк, 5 LLM_Reasoning + 2 Other", False,
               "empty", critical=True)
        record("Excel: 2 шумовые статьи помечены как Other", False,
               "empty", critical=True)
        return

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    record("Has required columns (ArXiv_ID and Title)",
           any("arxiv" in h for h in headers) and any("title" in h for h in headers),
           f"Headers: {rows[0]}")

    # Locate Title and Topic columns by header.
    def find_col(pred):
        for i, h in enumerate(headers):
            if pred(h):
                return i
        return None

    title_col = find_col(lambda h: h == "title" or "title" in h)
    topic_col = find_col(lambda h: h == "topic" or "topic" in h)
    record("Has Topic column", topic_col is not None, f"Headers: {rows[0]}")

    data_rows = [r for r in rows[1:] if any(c for c in r)]
    record("Has at least 5 data rows (papers)", len(data_rows) >= 5,
           f"Found {len(data_rows)} data rows")

    # Substring keyword sanity (non-critical).
    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    has_cot = "chain-of-thought" in all_text or "chain of thought" in all_text
    has_react = "react" in all_text or "synergizing" in all_text
    has_self = "self-consistency" in all_text or "self consistency" in all_text
    papers_found = sum([has_cot, has_react, has_self])
    record("Contains key LLM reasoning paper keywords", papers_found >= 2,
           f"CoT:{has_cot}, ReAct:{has_react}, SelfConsistency:{has_self}")

    # --- CRITICAL: exact 7 rows, 5/2 Topic split ---
    n_llm = 0
    n_other = 0
    if topic_col is not None:
        for r in data_rows:
            val = (str(r[topic_col]).strip().lower()
                   if topic_col < len(r) and r[topic_col] is not None else "")
            if "llm_reasoning" in val or "llm reasoning" in val:
                n_llm += 1
            elif "other" in val:
                n_other += 1
    seven_rows = (len(data_rows) == 7)
    split_ok = (n_llm == 5 and n_other == 2)
    record("Excel: ровно 7 строк, 5 LLM_Reasoning + 2 Other",
           seven_rows and split_ok,
           f"rows={len(data_rows)} LLM_Reasoning={n_llm} Other={n_other}",
           critical=True)

    # --- CRITICAL: the 2 noise papers are classified as Other ---
    noise_ok = True
    noise_detail = []
    if title_col is not None and topic_col is not None:
        for keys in NOISE_KEYS:
            found_row = None
            for r in data_rows:
                t = (str(r[title_col]).lower()
                     if title_col < len(r) and r[title_col] is not None else "")
                if all(k in t for k in keys):
                    found_row = r
                    break
            if found_row is None:
                noise_ok = False
                noise_detail.append(f"missing paper {keys}")
                continue
            topic_val = (str(found_row[topic_col]).strip().lower()
                         if topic_col < len(found_row) and found_row[topic_col] is not None
                         else "")
            if "other" not in topic_val:
                noise_ok = False
                noise_detail.append(f"{keys} -> Topic={topic_val!r} (expected Other)")
    else:
        noise_ok = False
        noise_detail.append("Title/Topic column not found")
    record("Excel: 2 шумовые статьи помечены как Other", noise_ok,
           "; ".join(noise_detail), critical=True)


# ---------------------------------------------------------------------------
# Forms (RU forms-mcp, schema gform.*)
# ---------------------------------------------------------------------------
def _option_values(config):
    """Extract option text values from a question config (RU forms-mcp shape)."""
    vals = []
    if not config:
        return vals
    opts = config.get("options") if isinstance(config, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    vals.append(str(v))
            else:
                vals.append(str(o))
    return vals


def check_gform():
    print("\n=== Check 2: Форма 'Reading Group Paper Selection' (forms / gform.*) ===")

    crit_priority = ("Форма: вопрос-приоритет содержит 5 названий статей LLM reasoning "
                     "(без 2 шумовых)")
    crit_aspects = ("Форма: вопросы про аспекты (Reasoning Methods/Implementation/"
                    "Evaluation/Theory) и доступность (Monday..Friday)")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Reading Group Paper Selection form exists", False, str(e))
        record(crit_priority, False, "no db", critical=True)
        record(crit_aspects, False, "no db", critical=True)
        return

    cur = conn.cursor()
    cur.execute("SELECT id, title FROM gform.forms")
    forms = cur.fetchall()

    reading_form = None
    for form_id, title in forms:
        t = (title or "").lower()
        if "reading group" in t or "paper selection" in t:
            reading_form = (form_id, title)
            break

    record("Reading Group Paper Selection form exists", reading_form is not None,
           f"Forms found: {[f[1] for f in forms]}")

    if not reading_form:
        record(crit_priority, False, "no form", critical=True)
        record(crit_aspects, False, "no form", critical=True)
        cur.close()
        conn.close()
        return

    form_id, _title = reading_form
    cur.execute(
        "SELECT title, question_type, config FROM gform.questions "
        "WHERE form_id = %s ORDER BY position",
        (form_id,),
    )
    questions = cur.fetchall()
    cur.close()
    conn.close()

    record("Form has at least 3 questions", len(questions) >= 3,
           f"Found {len(questions)} questions")

    # Each question's normalized title + its option set.
    parsed = []
    for q_title, q_type, q_config in questions:
        cfg = q_config if isinstance(q_config, dict) else (
            json.loads(q_config) if q_config else {})
        parsed.append({
            "title": (q_title or "").lower(),
            "type": q_type,
            "options": _option_values(cfg),
            "options_lower": [v.lower() for v in _option_values(cfg)],
        })

    # --- CRITICAL: priority question lists the 5 LLM titles, none of the 2 noise ---
    priority_q = None
    for q in parsed:
        joined = " ".join(q["options_lower"])
        n_titles = sum(1 for k in LLM_TITLE_KEYS if k in joined)
        if n_titles >= 4:  # this is the paper-selection question
            priority_q = q
            break
    if priority_q is None:
        # fall back: a question whose title hints at reading first / priority
        for q in parsed:
            if "first" in q["title"] or "priorit" in q["title"] or "read" in q["title"] \
               or "приоритет" in q["title"] or "перв" in q["title"] or "читать" in q["title"]:
                priority_q = q
                break

    if priority_q is None:
        record(crit_priority, False,
               f"no paper-selection question; questions={[q['title'] for q in parsed]}",
               critical=True)
    else:
        joined = " ".join(priority_q["options_lower"])
        present = [k for k in LLM_TITLE_KEYS if k in joined]
        # the 2 noise papers must NOT be among options
        noise_present = []
        if "vision transformer" in joined:
            noise_present.append("vision transformer")
        if "medical image" in joined or "segmentation" in joined:
            noise_present.append("medical image segmentation")
        ok = (len(present) == 5 and not noise_present)
        record(crit_priority, ok,
               f"LLM titles present={present}; noise present={noise_present}; "
               f"options={priority_q['options']}",
               critical=True)

    # --- CRITICAL: aspects question + availability question ---
    aspects_keys = ["reasoning methods", "implementation", "evaluation", "theory"]
    avail_keys = ["monday", "tuesday", "wednesday", "thursday", "friday"]

    aspects_ok = any(
        sum(1 for k in aspects_keys if any(k in o for o in q["options_lower"])) >= 4
        for q in parsed
    )
    avail_ok = any(
        sum(1 for k in avail_keys if any(k in o for o in q["options_lower"])) >= 5
        for q in parsed
    )
    record("Форма содержит вопрос про аспекты (4 варианта)", aspects_ok,
           f"questions={[(q['title'], q['options']) for q in parsed]}")
    record("Форма содержит вопрос про доступность (Monday..Friday)", avail_ok,
           f"questions={[(q['title'], q['options']) for q in parsed]}")
    record(crit_aspects, aspects_ok and avail_ok,
           f"aspects={aspects_ok} availability={avail_ok}", critical=True)


# ---------------------------------------------------------------------------
# Calendar (kept foreign infra: google_calendar; gcal.* data layer)
# ---------------------------------------------------------------------------
def check_gcal():
    print("\n=== Check 3: Google Calendar — Reading Group Session 1..4 ===")

    crit_gcal = ("Календарь: 4 сессии 'Reading Group Session' на 07/14/21/28 апреля 2026, "
                 "каждая длительностью 1.5 ч")

    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("At least 4 reading group events in April 2026", False, str(e))
        record(crit_gcal, False, "no db", critical=True)
        return

    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE start_datetime >= '2026-04-01' AND start_datetime < '2026-05-01'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    reading_events = [
        e for e in events
        if "reading group" in (e[0] or "").lower()
        or "reading session" in (e[0] or "").lower()
    ]

    record("At least 4 reading group events in April 2026", len(reading_events) >= 4,
           f"Found {len(reading_events)} reading events: "
           f"{[(e[0], str(e[1])) for e in reading_events]}")

    # Duration ~1.5h sanity (non-critical) on the first event.
    if reading_events:
        _s, sdt, edt = reading_events[0]
        if sdt and edt:
            dur = (edt - sdt).total_seconds() / 3600
            record("Reading sessions are ~1.5 hours", 1.0 <= dur <= 2.0,
                   f"Duration: {dur:.2f} hours")

    # --- CRITICAL: the four exact dates, each exactly 1.5h ---
    # start_datetime is timestamptz; compare the calendar day in the value's own
    # offset by matching the date portion of the localized-as-stored timestamp.
    # We bucket events by their UTC-naive date as returned by psycopg2 and also by
    # +/-1 day to tolerate timezone offsets, but require exact 1.5h duration.
    expected_days = {7, 14, 21, 28}
    by_day = {}
    for summ, sdt, edt in reading_events:
        if sdt is None or edt is None:
            continue
        dur_min = round((edt - sdt).total_seconds() / 60.0)
        # The stored day as returned by psycopg2 (tz-aware). The agent schedules
        # these wall-clock April dates via google_calendar; duration is checked
        # exactly (TZ-independent) and the day must match one of the four targets.
        day = sdt.day
        if day in expected_days:
            by_day.setdefault(day, []).append(dur_min)

    days_covered = set(by_day.keys())
    all_days = (days_covered == expected_days)
    # every covered expected day has at least one event of exactly 90 minutes
    durations_ok = all(
        any(abs(dm - 90) <= 1 for dm in by_day.get(d, []))
        for d in expected_days
    )
    record(crit_gcal, all_days and durations_ok,
           f"days_covered={sorted(days_covered)} expected={sorted(expected_days)} "
           f"durations(min)_by_day={by_day}",
           critical=True)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gform()
    check_gcal()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILS,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILS}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
