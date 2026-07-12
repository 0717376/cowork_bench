"""
Evaluation для insales-competitor-pw-pricing-excel-gcal (InSales).

Сценарий: агент извлекает минимальные цены конкурента со страницы-монитора
(http://localhost:30341, 12 позиций в 3 категориях: Электроника / Наушники / Колонки),
выгружает каталог собственного магазина InSales по этим категориям, через терминал
запускает Python-скрипт pricing_analyzer.py (создаёт pricing_analysis.json), строит
Excel Competitive_Pricing_Analysis.xlsx (3 листа) и планирует одно событие календаря
"Pricing Strategy Review".

Что проверяем:
1. Competitive_Pricing_Analysis.xlsx с 3 листами и корректными данными конкурента.
2. Python-скрипт анализа в рабочей директории (использование терминала).
3. Событие календаря "Pricing Strategy Review" (шумовые события не удалены).

Якоря (не волатильные):
- Средние цены конкурента по категориям COMPETITOR_AVGS и счётчики COMPETITOR_COUNTS
  вычислены из фиксированной страницы files/mock_pages/index.html (USD-цены), их агент
  обязан воспроизвести. Это и есть содержательный, негеймимый якорь.
- Цены собственного магазина (Our_*) зависят от живой БД и НЕ проверяются на точное
  значение, чтобы не вводить волатильные хардкоды.

CRITICAL_CHECKS: любой провал => итог FAIL (sys.exit(1)) ДО порога accuracy.
Порог: accuracy >= 70 И нет критичных провалов => PASS.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Содержательные проверки. Провал любой => вся задача FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "Competitor avg per category matches mock page",
    "Summary: Total_Competitor_Products = 12",
    "Calendar: Pricing Strategy Review event exists",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        marker = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{marker} {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_contains_any(haystack, needles):
    if haystack is None:
        return False
    h = str(haystack).strip().lower()
    return any(n and n.strip().lower() in h for n in needles)


# Средние цены конкурента по категориям (из USD-цен files/mock_pages/index.html). Ключи — RU.
#   Электроника: mean(129.99, 79.50, 42.00, 199.99, 64.95) = 103.286
#   Наушники:    mean(59.99, 119.99, 24.99, 89.95)         = 73.73
#   Колонки:     mean(149.99, 39.99, 219.00)               = 136.327
COMPETITOR_AVGS = {"электроника": 103.29, "наушники": 73.73, "колонки": 136.33}
COMPETITOR_COUNTS = {"электроника": 5, "наушники": 4, "колонки": 3}

CAT_SYNONYMS = {
    "электроника": ["электроника", "electronics"],
    "наушники": ["наушники", "headphone"],
    "колонки": ["колонки", "speaker"],
}


def cat_key(label):
    if label is None:
        return None
    s = str(label).strip().lower()
    for key, syns in CAT_SYNONYMS.items():
        for syn in syns:
            if syn in s:
                return key
    return None


def check_excel(agent_workspace):
    print("\n=== Checking Excel Output ===")

    fpath = os.path.join(agent_workspace, "Competitive_Pricing_Analysis.xlsx")
    if not os.path.isfile(fpath):
        record("Excel file exists", False, f"Not found: {fpath}")
        # Критичные якоря зависят от файла — отметим их проваленными.
        record("Competitor avg per category matches mock page", False, "no Excel")
        record("Summary: Total_Competitor_Products = 12", False, "no Excel")
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(fpath, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        record("Competitor avg per category matches mock page", False, "unreadable")
        record("Summary: Total_Competitor_Products = 12", False, "unreadable")
        return

    # --- Лист 1: Price_Comparison (по категориям) ---
    pc_sheet = None
    for name in wb.sheetnames:
        if "price" in name.lower() and "comparison" in name.lower():
            pc_sheet = name
            break
    if pc_sheet is None:
        # fallback: первый лист, где встречается слово сравнения
        for name in wb.sheetnames:
            if "comparison" in name.lower() or "price" in name.lower():
                pc_sheet = name
                break

    if not pc_sheet:
        record("Price_Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Competitor avg per category matches mock page", False, "no sheet")
    else:
        record("Price_Comparison sheet exists", True)
        ws = wb[pc_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []
        record("Price_Comparison has >= 3 category rows", len(data_rows) >= 3,
               f"Found {len(data_rows)}")

        # CRITICAL: для каждой категории есть строка, где среди числовых ячеек
        # присутствует ожидаемая средняя цена конкурента.
        matched = {}
        for row in data_rows:
            if not row or row[0] is None:
                continue
            key = cat_key(row[0])
            if key is None or key not in COMPETITOR_AVGS:
                continue
            expected = COMPETITOR_AVGS[key]
            found = any(num_close(cell, expected, tol=1.0) for cell in row[1:] if cell is not None)
            # учитываем только если ещё не нашли совпадение для этой категории
            matched[key] = matched.get(key, False) or found

        comp_ok = (set(matched.keys()) == set(COMPETITOR_AVGS.keys())) and all(matched.values())
        record("Competitor avg per category matches mock page", comp_ok,
               f"Per-category match: {matched}")

    # --- Лист 2: Market_Position (summary) ---
    mp_sheet = None
    for name in wb.sheetnames:
        if "market" in name.lower() or "position" in name.lower() or "summary" in name.lower():
            mp_sheet = name
            break
    if not mp_sheet:
        record("Market_Position sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Summary: Total_Competitor_Products = 12", False, "no sheet")
    else:
        record("Market_Position sheet exists", True)
        ws = wb[mp_sheet]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:] if len(rows) > 1 else []

        seen_total = False
        for row in data_rows:
            if not row or row[0] is None:
                continue
            metric = str(row[0]).strip().lower()
            val = row[1] if len(row) > 1 else None
            if "total_competitor" in metric or ("competitor" in metric and "product" in metric):
                seen_total = True
                record("Summary: Total_Competitor_Products = 12", num_close(val, 12, tol=0),
                       f"Got {val}")
        if not seen_total:
            record("Summary: Total_Competitor_Products = 12", False,
                   "Total_Competitor_Products row missing")

        # Структурные (НЕ критичные) проверки направлений по категориям:
        # значение должно быть одной из трёх валидных RU-категорий, но конкретную
        # категорию не хардкодим (зависит от живых цен магазина).
        cheapest_val = None
        expensive_val = None
        for row in data_rows:
            if not row or row[0] is None:
                continue
            m = str(row[0]).strip().lower()
            v = row[1] if len(row) > 1 else None
            if "cheapest" in m:
                cheapest_val = v
            elif "expensive" in m or "most_expensive" in m:
                expensive_val = v
        record("Summary: Cheapest category is a valid category",
               cat_key(cheapest_val) is not None, f"Got {cheapest_val}")
        record("Summary: Most expensive category is a valid category",
               cat_key(expensive_val) is not None, f"Got {expensive_val}")

    # --- Лист 3: Action_Items (структурно) ---
    ai_sheet = None
    for name in wb.sheetnames:
        if "action" in name.lower() or "item" in name.lower():
            ai_sheet = name
            break
    if not ai_sheet:
        record("Action_Items sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        record("Action_Items sheet exists", True)
        ws = wb[ai_sheet]
        rows = list(ws.iter_rows(values_only=True))
        record("Action_Items has >= 1 data row", len(rows) > 1, f"Rows: {len(rows)}")

    wb.close()


def check_script(agent_workspace):
    print("\n=== Checking Terminal/Python Script ===")
    try:
        py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    except Exception as e:
        record("Python analysis script exists", False, str(e))
        return
    record("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")
    # pricing_analysis.json — продукт работы скрипта (структурная проверка).
    json_path = os.path.join(agent_workspace, "pricing_analysis.json")
    record("pricing_analysis.json exists", os.path.isfile(json_path), json_path)


def check_calendar():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT summary, start_datetime FROM gcal.events")
        events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Calendar DB accessible", False, str(e))
        record("Calendar: Pricing Strategy Review event exists", False, str(e))
        return

    review_found = False
    for summary, _start in events:
        s = (summary or "").lower()
        if "pricing strategy review" in s or ("pricing" in s and "review" in s):
            review_found = True
            break
    record("Calendar: Pricing Strategy Review event exists", review_found,
           f"events: {[e[0] for e in events]}")

    # Обратная проверка: шумовые события не удалены агентом.
    noise = sum(1 for summary, _ in events
                if str_contains_any(summary, ["standup", "стендап", "lunch", "обед"]))
    record("Noise events preserved (not deleted)", noise >= 1, f"noise events: {noise}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_script(args.agent_workspace)
    check_calendar()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100

    print("\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")
    if CRITICAL_FAILS:
        print(f"  Critical fails: {CRITICAL_FAILS}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_fails": CRITICAL_FAILS,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"  Overall: FAIL (критичные чеки провалены: {len(CRITICAL_FAILS)})")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL (accuracy < 70)")
    sys.exit(1)


if __name__ == "__main__":
    main()
