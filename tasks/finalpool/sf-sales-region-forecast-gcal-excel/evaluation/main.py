"""Evaluation for sf-sales-region-forecast-gcal-excel (ClickHouse Sales DW, russified regions)."""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    agent_ws = args.agent_workspace or task_root

    all_errors = []
    critical_errors = []
    # checks_total / checks_passed feed the accuracy gate (>=70)
    checks = {"total": 0, "passed": 0}

    def check(cond, err, critical=False):
        checks["total"] += 1
        if cond:
            checks["passed"] += 1
        else:
            all_errors.append(err)
            if critical:
                critical_errors.append(err)

    # --- Check 1: Excel file ---
    import openpyxl

    print("Checking Excel file...")
    agent_file = os.path.join(agent_ws, "Regional_Forecast.xlsx")
    gt_file = os.path.join(gt_dir, "Regional_Forecast.xlsx")

    if not os.path.exists(agent_file):
        critical_errors.append("Regional_Forecast.xlsx not found in agent workspace")
        all_errors.append("Regional_Forecast.xlsx not found in agent workspace")
        checks["total"] += 1
    else:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

        # Check Region Performance sheet
        print("  Checking Region Performance sheet...")
        a_rows = load_sheet_rows(agent_wb, "Region Performance")
        g_rows = load_sheet_rows(gt_wb, "Region Performance")

        if a_rows is None:
            critical_errors.append("Sheet 'Region Performance' not found in agent output")
            all_errors.append("Sheet 'Region Performance' not found in agent output")
            checks["total"] += 1
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []

            check(len([r for r in a_data if r and r[0]]) == 5,
                  f"Region Performance row count: {len(a_data)}, expected 5")

            a_lookup = {}
            for row in a_data:
                if row and row[0]:
                    a_lookup[str(row[0]).strip().lower()] = row

            for g_row in g_data:
                if not g_row or not g_row[0]:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup.get(key)
                if a_row is None:
                    # CRITICAL: region key (russified) must be present
                    critical_errors.append(f"Missing region: {g_row[0]}")
                    all_errors.append(f"Missing region: {g_row[0]}")
                    checks["total"] += 4  # target, actual, variance, var_pct
                    continue
                # Target (col 1) -- CRITICAL: must match PDF exactly (tol=1)
                check(num_close(a_row[1], g_row[1], 1),
                      f"{g_row[0]} Target: {a_row[1]} vs {g_row[1]}", critical=True)
                # Actual (col 2) -- CRITICAL: DB-derived, tol=1
                check(num_close(a_row[2], g_row[2], 1),
                      f"{g_row[0]} Actual: {a_row[2]} vs {g_row[2]}", critical=True)
                # Variance (col 3) -- CRITICAL: core derived deliverable, tol=1
                check(len(a_row) > 3 and num_close(a_row[3], g_row[3], 1),
                      f"{g_row[0]} Variance: {a_row[3] if len(a_row) > 3 else None} vs {g_row[3]}",
                      critical=True)
                # Variance_Pct (col 4) -- non-critical, tol=0.5
                check(len(a_row) > 4 and num_close(a_row[4], g_row[4], 0.5),
                      f"{g_row[0]} Variance_Pct: {a_row[4] if len(a_row) > 4 else None} vs {g_row[4]}")
            print("    Done.")

        # Check Summary sheet
        print("  Checking Summary sheet...")
        a_rows2 = load_sheet_rows(agent_wb, "Summary")
        g_rows2 = load_sheet_rows(gt_wb, "Summary")
        if a_rows2 is None:
            critical_errors.append("Sheet 'Summary' not found in agent output")
            all_errors.append("Sheet 'Summary' not found in agent output")
            checks["total"] += 1
        else:
            a_data2 = a_rows2[1:] if len(a_rows2) > 1 else []
            g_data2 = g_rows2[1:] if len(g_rows2) > 1 else []

            a_lookup2 = {}
            for row in a_data2:
                if row and row[0]:
                    a_lookup2[str(row[0]).strip().lower()] = row

            # Metrics whose exact value encodes the core conclusion -> CRITICAL
            critical_metrics = {"total_target", "met_target_count", "missed_target_count"}
            for g_row in g_data2:
                if not g_row or not g_row[0]:
                    continue
                key = str(g_row[0]).strip().lower()
                a_row = a_lookup2.get(key)
                if a_row is None:
                    is_crit = key in critical_metrics
                    if is_crit:
                        critical_errors.append(f"Summary missing row: {g_row[0]}")
                    all_errors.append(f"Summary missing row: {g_row[0]}")
                    checks["total"] += 1
                    continue
                # counts must be exact; money totals tol=100
                tol = 0.5 if key in {"met_target_count", "missed_target_count"} else 100
                check(num_close(a_row[1], g_row[1], tol),
                      f"Summary {g_row[0]}: {a_row[1]} vs {g_row[1]}",
                      critical=(key in critical_metrics))
            print("    Done.")

    # --- Check 2: GCal event (date-specific, CRITICAL) ---
    print("Checking Google Calendar event...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE (LOWER(summary) LIKE '%forecast%' OR LOWER(summary) LIKE '%quarterly%'
                   OR LOWER(summary) LIKE '%прогноз%' OR LOWER(summary) LIKE '%квартал%')
            AND start_datetime::date = '2026-03-31'
        """)
        count = cur.fetchone()[0]
        check(count > 0,
              "No GCal event for Quarterly Forecast Review on 2026-03-31",
              critical=True)
        if count > 0:
            print(f"    GCal event on 2026-03-31 found ({count} events)")
        cur.close()
        conn.close()
    except Exception as e:
        critical_errors.append(f"Error checking GCal: {e}")
        all_errors.append(f"Error checking GCal: {e}")
        checks["total"] += 1

    # --- Check 3: Email sent + body content (CRITICAL) ---
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COALESCE(subject,'') || ' ' || COALESCE(body_text,'') || ' ' || COALESCE(body_html,'')
            FROM email.messages
            WHERE to_addr::text ILIKE '%vp_sales@company.com%'
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()

        check(len(rows) > 0, "No email sent to vp_sales@company.com", critical=True)
        if rows:
            print(f"    Email found ({len(rows)} messages)")
            blob = " ".join(r[0] for r in rows).lower()
            # Body must reference the total variance figure (313998 / 313 998 / 314000 region)
            variance_ok = any(t in blob for t in
                              ["313998", "313 998", "313,998", "313998.33", "314000", "314 000"])
            check(variance_ok,
                  "Email body does not reference total variance figure (~313998)",
                  critical=True)
            # Body must confirm a quarterly review meeting was scheduled (RU+EN keywords)
            meeting_ok = (any(t in blob for t in ["meeting", "review", "встреч", "обзор", "совещан"])
                          and any(t in blob for t in ["quarterly", "scheduled", "квартал",
                                                      "запланир", "заплан"]))
            check(meeting_ok,
                  "Email body does not confirm a scheduled quarterly review meeting (RU+EN)",
                  critical=True)
    except Exception as e:
        critical_errors.append(f"Error checking email: {e}")
        all_errors.append(f"Error checking email: {e}")
        checks["total"] += 1

    # --- Critical gate (before accuracy) ---
    if critical_errors:
        print(f"\n=== CRITICAL FAILURE ({len(critical_errors)}) ===")
        for e in critical_errors[:15]:
            print(f"  [CRITICAL] {e}")
        sys.exit(1)

    # --- Accuracy gate (>=70) ---
    total = max(checks["total"], 1)
    accuracy = 100.0 * checks["passed"] / total
    print(f"\nAccuracy: {accuracy:.1f}% ({checks['passed']}/{total})")
    if all_errors:
        print(f"Non-critical errors ({len(all_errors)}):")
        for e in all_errors[:15]:
            print(f"  {e}")

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL (accuracy below 70) ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
