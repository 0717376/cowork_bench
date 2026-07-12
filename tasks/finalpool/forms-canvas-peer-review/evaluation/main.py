"""
Evaluation для gform-canvas-peer-review (RU-стек: forms/teamly + canvas/gsheet/excel).

Проверки:
1. Excel (Peer_Review_Analysis.xlsx) с 3 листами и корректными данными.
2. Google Sheet "Peer Review Results" с данными Individual Summary.
3. Страница Teamly с обзором, flagged-студентами и рекомендациями.

Модель успеха: accuracy >= 70% И нет проваленных CRITICAL_CHECKS.
Любой fail из CRITICAL_CHECKS => немедленный FAIL (sys.exit(1)) до порога accuracy.
Структурные проверки (лист есть, столбец есть, счётчики ~N) — НЕ критические.
"""

import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

# Критические чеки по имени check(): любой fail => задача FAIL.
CRITICAL_CHECKS = {
    "CRITICAL: Individual Summary содержит ровно 6 студентов",
    "CRITICAL: Frank Liu Overall_Avg вычислен корректно (~2.33)",
    "CRITICAL: Flagged содержит ровно flagged-студентов (Frank Liu), без ложных",
    "CRITICAL: Raw Scores содержит 20 строк (по числу ответов формы)",
    "CRITICAL: Google Sheet Individual Summary совпадает с Excel (6 студентов, Overall_Avg)",
    "CRITICAL: Teamly страница: 20 оценок, 6 студентов, упомянут Frank Liu",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=0.5):
    """Сравнить два числа с допуском."""
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    """Регистронезависимое сравнение строк с нормализацией пробелов."""
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def compute_expected():
    """
    Вычислить ожидаемые значения из gform-ответов в БД.
    Возвращает raw_scores list, individual_summary dict и flagged list.
    """
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute(
        "SELECT id FROM gform.forms WHERE LOWER(title) LIKE '%peer review%' LIMIT 1"
    )
    form_row = cur.fetchone()
    if not form_row:
        cur.close()
        conn.close()
        return None, None, None

    form_id = form_row[0]

    cur.execute(
        "SELECT id, title FROM gform.questions WHERE form_id = %s ORDER BY position",
        (form_id,),
    )
    questions = cur.fetchall()
    q_map = {}
    for qid, qtitle in questions:
        title_lower = qtitle.lower()
        if "your name" in title_lower:
            q_map["reviewer"] = qid
        elif "person" in title_lower or "reviewed" in title_lower:
            q_map["reviewee"] = qid
        elif "contribution" in title_lower:
            q_map["contribution"] = qid
        elif "communication" in title_lower:
            q_map["communication"] = qid
        elif "quality" in title_lower:
            q_map["quality"] = qid
        elif "comment" in title_lower:
            q_map["comments"] = qid

    cur.execute(
        "SELECT answers FROM gform.responses WHERE form_id = %s",
        (form_id,),
    )
    response_rows = cur.fetchall()

    raw_scores = []
    for (answers_json,) in response_rows:
        answers = answers_json if isinstance(answers_json, dict) else json.loads(answers_json)
        reviewer = answers.get(q_map.get("reviewer", ""), "")
        reviewee = answers.get(q_map.get("reviewee", ""), "")
        contrib = int(answers.get(q_map.get("contribution", ""), 0))
        comm = int(answers.get(q_map.get("communication", ""), 0))
        quality = int(answers.get(q_map.get("quality", ""), 0))
        avg_score = round((contrib + comm + quality) / 3.0, 2)
        raw_scores.append({
            "Reviewer": reviewer,
            "Reviewee": reviewee,
            "Contribution": contrib,
            "Communication": comm,
            "Quality": quality,
            "Average_Score": avg_score,
        })

    from collections import defaultdict
    student_data = defaultdict(lambda: {"contrib": [], "comm": [], "quality": []})
    for r in raw_scores:
        name = r["Reviewee"]
        student_data[name]["contrib"].append(r["Contribution"])
        student_data[name]["comm"].append(r["Communication"])
        student_data[name]["quality"].append(r["Quality"])

    individual_summary = {}
    for name, data in student_data.items():
        avg_c = round(sum(data["contrib"]) / len(data["contrib"]), 2)
        avg_m = round(sum(data["comm"]) / len(data["comm"]), 2)
        avg_q = round(sum(data["quality"]) / len(data["quality"]), 2)
        overall = round((avg_c + avg_m + avg_q) / 3.0, 2)
        individual_summary[name] = {
            "Avg_Contribution": avg_c,
            "Avg_Communication": avg_m,
            "Avg_Quality": avg_q,
            "Overall_Avg": overall,
            "Review_Count": len(data["contrib"]),
        }

    flagged = []
    for name, stats in individual_summary.items():
        if stats["Overall_Avg"] < 3.0:
            flagged.append(name)

    cur.close()
    conn.close()

    return raw_scores, individual_summary, flagged


def get_sheet(wb, name):
    """Найти лист регистронезависимо."""
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def _find_student_row(data_rows, name_col, target_name):
    """Найти строку для студента по имени (подстрочное совпадение)."""
    for row in data_rows:
        sn = str(row[name_col]).strip().lower() if row[name_col] else ""
        if target_name.lower() in sn or sn in target_name.lower():
            return row
    return None


def check_excel(agent_workspace, raw_scores, individual_summary, flagged):
    """Проверить выходной файл Excel. Возвращает dict для кросс-проверки с Google Sheet."""
    print("\n=== Проверка Excel ===")
    excel_summary = {}  # name -> Overall_Avg (для сравнения с Google Sheet)

    excel_path = os.path.join(agent_workspace, "Peer_Review_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Ожидался {excel_path}")
    if not os.path.isfile(excel_path):
        # Помечаем зависящие критические чеки как проваленные
        check("CRITICAL: Individual Summary содержит ровно 6 студентов", False, "нет файла")
        check("CRITICAL: Frank Liu Overall_Avg вычислен корректно (~2.33)", False, "нет файла")
        check("CRITICAL: Flagged содержит ровно flagged-студентов (Frank Liu), без ложных", False, "нет файла")
        check("CRITICAL: Raw Scores содержит 20 строк (по числу ответов формы)", False, "нет файла")
        return excel_summary

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        return excel_summary

    check("Excel file readable", True)

    check("Has 'Raw Scores' sheet",
          any(str_match(s, "Raw Scores") for s in wb.sheetnames),
          f"Листы: {wb.sheetnames}")
    check("Has 'Individual Summary' sheet",
          any(str_match(s, "Individual Summary") for s in wb.sheetnames),
          f"Листы: {wb.sheetnames}")
    check("Has 'Flagged' sheet",
          any(str_match(s, "Flagged") for s in wb.sheetnames),
          f"Листы: {wb.sheetnames}")

    # --- Raw Scores ---
    print("\n--- Raw Scores ---")
    ws = get_sheet(wb, "Raw Scores")
    raw_ok_count = False
    if ws:
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]
        expected_count = len(raw_scores)  # 20

        headers = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower().replace("_", "").replace(" ", "") if h else "" for h in headers]
        check("Raw Scores has Reviewer column",
              any("reviewer" in h for h in header_lower), f"Headers: {headers}")
        check("Raw Scores has Reviewee column",
              any("reviewee" in h for h in header_lower), f"Headers: {headers}")
        check("Raw Scores has Contribution column",
              any("contribution" in h for h in header_lower), f"Headers: {headers}")
        check("Raw Scores has Average_Score column",
              any("average" in h or "avg" in h for h in header_lower), f"Headers: {headers}")

        # CRITICAL: ровно 20 строк (по числу инжектированных ответов)
        raw_ok_count = (len(data_rows) == expected_count)
        # Дополнительно: проверить, что Average_Score реально вычислен (не выдуман),
        # на нескольких выборочных строках.
        avg_idx = None
        c_idx = m_idx = q_idx = None
        for i, h in enumerate(header_lower):
            if ("average" in h or "avg" in h) and avg_idx is None:
                avg_idx = i
            elif "contribution" in h:
                c_idx = i
            elif "communication" in h:
                m_idx = i
            elif "quality" in h:
                q_idx = i
        avg_correct = True
        if None not in (avg_idx, c_idx, m_idx, q_idx):
            sampled = 0
            for row in data_rows:
                try:
                    c, m, q = float(row[c_idx]), float(row[m_idx]), float(row[q_idx])
                    exp_avg = round((c + m + q) / 3.0, 2)
                    if not num_close(row[avg_idx], exp_avg, 0.05):
                        avg_correct = False
                        break
                    sampled += 1
                    if sampled >= 5:
                        break
                except (TypeError, ValueError):
                    continue
            avg_correct = avg_correct and sampled >= 1
        else:
            avg_correct = False
        check("CRITICAL: Raw Scores содержит 20 строк (по числу ответов формы)",
              raw_ok_count and avg_correct,
              f"строк={len(data_rows)} (ожид. {expected_count}), avg_correct={avg_correct}")
    else:
        check("CRITICAL: Raw Scores содержит 20 строк (по числу ответов формы)", False, "лист не найден")

    # --- Individual Summary ---
    print("\n--- Individual Summary ---")
    ws = get_sheet(wb, "Individual Summary")
    if ws:
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]

        # CRITICAL: ровно 6 студентов
        check("CRITICAL: Individual Summary содержит ровно 6 студентов",
              len(data_rows) == 6, f"Ожидалось 6, получено {len(data_rows)}")

        headers = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower().replace("_", "").replace(" ", "") if h else "" for h in headers]

        name_col = None
        overall_col = None
        count_col = None
        for i, h in enumerate(header_lower):
            if name_col is None and ("student" in h or "name" in h):
                name_col = i
            if "overall" in h:
                overall_col = i
            if "count" in h or "reviewcount" in h:
                count_col = i

        if name_col is not None and overall_col is not None:
            # Собрать excel_summary для кросс-проверки
            for row in data_rows:
                sn = str(row[name_col]).strip() if row[name_col] else ""
                if sn and row[overall_col] is not None:
                    try:
                        excel_summary[sn.lower()] = float(row[overall_col])
                    except (TypeError, ValueError):
                        pass

            # Структурные (НЕ критические) проверки каждого студента
            for exp_name, exp_stats in individual_summary.items():
                row = _find_student_row(data_rows, name_col, exp_name)
                if row is not None:
                    check(f"Student '{exp_name}' Overall_Avg (struct)",
                          num_close(row[overall_col], exp_stats["Overall_Avg"], 0.5),
                          f"Ожид. ~{exp_stats['Overall_Avg']}, получено {row[overall_col]}")
                    if count_col is not None and row[count_col] is not None:
                        check(f"Student '{exp_name}' Review_Count (struct)",
                              int(row[count_col]) == exp_stats["Review_Count"],
                              f"Ожид. {exp_stats['Review_Count']}, получено {row[count_col]}")

            # CRITICAL: Frank Liu Overall_Avg вычислен корректно с жёстким допуском
            frank_exp = None
            for exp_name, exp_stats in individual_summary.items():
                if "frank" in exp_name.lower():
                    frank_exp = (exp_name, exp_stats)
                    break
            if frank_exp:
                fname, fstats = frank_exp
                frow = _find_student_row(data_rows, name_col, fname)
                check("CRITICAL: Frank Liu Overall_Avg вычислен корректно (~2.33)",
                      frow is not None and frow[overall_col] is not None
                      and num_close(frow[overall_col], fstats["Overall_Avg"], 0.05),
                      f"Ожид. {fstats['Overall_Avg']}, получено "
                      f"{frow[overall_col] if frow else 'нет строки'}")
            else:
                check("CRITICAL: Frank Liu Overall_Avg вычислен корректно (~2.33)",
                      False, "Frank Liu отсутствует среди ожидаемых")
        else:
            check("CRITICAL: Individual Summary содержит ровно 6 студентов", False,
                  f"Не удалось найти столбцы name/overall: {headers}")
            check("CRITICAL: Frank Liu Overall_Avg вычислен корректно (~2.33)", False,
                  "нет столбцов name/overall")
    else:
        check("CRITICAL: Individual Summary содержит ровно 6 студентов", False, "лист не найден")
        check("CRITICAL: Frank Liu Overall_Avg вычислен корректно (~2.33)", False, "лист не найден")

    # --- Flagged ---
    print("\n--- Flagged ---")
    ws = get_sheet(wb, "Flagged")
    if ws:
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and r[0] is not None]

        headers = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower().replace("_", "").replace(" ", "") if h else "" for h in headers]
        overall_col = None
        for i, h in enumerate(header_lower):
            if "overall" in h or "avg" in h:
                overall_col = i
                break

        listed_names = [str(r[0]).strip().lower() for r in data_rows if r[0]]
        all_names = " ".join(listed_names)

        # Структурно: каждый flagged присутствует
        for fname in flagged:
            check(f"Flagged: '{fname}' is listed (struct)",
                  fname.lower() in all_names,
                  f"Имена в Flagged: {all_names}")

        # CRITICAL: множество flagged совпадает (все < 3.0, нет ложных >= 3.0)
        expected_flagged = {n.lower() for n in flagged}
        # Считаем фактически перечисленные имена сопоставимыми с ожидаемыми студентами
        matched_listed = set()
        false_positive = False
        for ln in listed_names:
            hit = None
            for exp_name in individual_summary:
                en = exp_name.lower()
                if en in ln or ln in en:
                    hit = en
                    break
            if hit is None:
                continue
            matched_listed.add(hit)
            # ложный flagged: студент с Overall_Avg >= 3.0
            if individual_summary:
                for exp_name, st in individual_summary.items():
                    if exp_name.lower() == hit and st["Overall_Avg"] >= 3.0:
                        false_positive = True
        flagged_set_ok = (matched_listed == expected_flagged) and not false_positive
        check("CRITICAL: Flagged содержит ровно flagged-студентов (Frank Liu), без ложных",
              flagged_set_ok,
              f"ожид={sorted(expected_flagged)}, получено={sorted(matched_listed)}, "
              f"false_positive={false_positive}")

        if overall_col is not None and data_rows:
            for row in data_rows:
                if row[overall_col] is not None:
                    check(f"Flagged student avg < 3.0 ({row[0]}) (struct)",
                          float(row[overall_col]) < 3.0,
                          f"Overall_Avg = {row[overall_col]}")
                    break
    else:
        check("CRITICAL: Flagged содержит ровно flagged-студентов (Frank Liu), без ложных",
              False, "лист не найден")

    return excel_summary


def check_google_sheet(individual_summary, excel_summary):
    """Проверить, что Google Sheet создан с данными Individual Summary."""
    print("\n=== Проверка Google Sheet ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute(
        """SELECT id, title FROM gsheet.spreadsheets
           WHERE LOWER(title) LIKE '%peer review%'
           LIMIT 1"""
    )
    ss_row = cur.fetchone()
    check("Google Sheet with 'peer review' in title exists (struct)",
          ss_row is not None,
          "Таблица с 'peer review' в названии не найдена")

    if not ss_row:
        check("CRITICAL: Google Sheet Individual Summary совпадает с Excel (6 студентов, Overall_Avg)",
              False, "нет таблицы")
        cur.close()
        conn.close()
        return

    ss_id = ss_row[0]
    ss_title = ss_row[1]
    print(f"  Найдена таблица: '{ss_title}' (id={ss_id})")

    cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id = %s", (ss_id,))
    cell_count = cur.fetchone()[0]
    check("Google Sheet has data (cells > 0) (struct)",
          cell_count > 0, f"Найдено {cell_count} ячеек")

    cur.execute(
        "SELECT COUNT(DISTINCT row_index) FROM gsheet.cells WHERE spreadsheet_id = %s",
        (ss_id,),
    )
    row_count = cur.fetchone()[0]
    check("Google Sheet has at least 7 rows (struct)",
          row_count >= 7, f"Найдено {row_count} различных строк")

    # CRITICAL: значения Individual Summary в Google Sheet совпадают с эталоном.
    # Собираем все строковые значения ячеек и сопоставляем имена студентов с Overall_Avg.
    cur.execute(
        """SELECT row_index, col_index, value FROM gsheet.cells
           WHERE spreadsheet_id = %s""",
        (ss_id,),
    )
    cells = cur.fetchall()
    cur.close()
    conn.close()

    # Построить таблицу: row -> {col: value}
    grid = {}
    for ri, ci, val in cells:
        grid.setdefault(ri, {})[ci] = val

    # Для каждого ожидаемого студента: найти строку, где встречается его имя,
    # и проверить, что где-то в этой строке есть число ~ Overall_Avg.
    matched_students = 0
    for exp_name, st in individual_summary.items():
        target = exp_name.lower()
        for ri, cols in grid.items():
            row_vals = [str(v).strip().lower() for v in cols.values() if v is not None]
            name_hit = any((target in rv or rv in target) and len(rv) >= 3 for rv in row_vals)
            if not name_hit:
                continue
            # Есть число близкое к Overall_Avg в этой строке
            for v in cols.values():
                if num_close(v, st["Overall_Avg"], 0.05):
                    matched_students += 1
                    break
            break

    check("CRITICAL: Google Sheet Individual Summary совпадает с Excel (6 студентов, Overall_Avg)",
          matched_students == len(individual_summary) and len(individual_summary) == 6,
          f"совпало студентов={matched_students} из {len(individual_summary)}")


def check_teamly(individual_summary, flagged, raw_count):
    """Проверить страницу Teamly с обзором, flagged-студентами и рекомендациями."""
    print("\n=== Проверка Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        check("CRITICAL: Teamly страница: 20 оценок, 6 студентов, упомянут Frank Liu", False, str(e))
        return
    cur = conn.cursor()
    cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
    pages = cur.fetchall()
    cur.close()
    conn.close()

    def title_matches(title):
        t = (title or "").lower()
        return "peer review" in t or ("оценив" in t and ("проект" in t or "project" in t))

    candidates = [(t, b) for t, b in pages if title_matches(t)]
    check("Teamly page with 'peer review' in title exists (struct)",
          len(candidates) >= 1, f"Заголовки: {[t for t, _ in pages]}")

    body = "\n\n".join(b for _t, b in candidates)
    bl = body.lower()

    # Обзор: число собранных оценок (20) и число оценённых студентов (6)
    has_count = str(raw_count) in body  # 20
    n_students = len(individual_summary)  # 6
    has_students = str(n_students) in body

    check("Teamly: упомянуто число собранных оценок (struct)",
          bool(candidates) and has_count, f"raw_count={raw_count} present={has_count}")
    check("Teamly: упомянуто число оценённых студентов (struct)",
          bool(candidates) and has_students, f"n_students={n_students} present={has_students}")

    # Flagged-студенты упомянуты в .lower() оригинального текста (RU+EN имена — английские)
    flagged_ok = all(fname.lower() in bl for fname in flagged) if flagged else True
    for fname in flagged:
        check(f"Teamly: упомянут flagged-студент '{fname}' (struct)",
              fname.lower() in bl, "имя не найдено в тексте страницы")

    # Рекомендации / follow-up: RU + EN ключевые слова
    rec_keywords = ("recommend", "follow", "action", "flag",
                    "рекоменд", "действи", "отмеч", "дальнейш")
    has_rec = any(k in bl for k in rec_keywords)
    check("Teamly: есть рекомендации / follow-up (struct)",
          has_rec, "не найдено ключевых слов рекомендаций")

    # CRITICAL: страница содержит корректные счётчики обзора и называет Frank Liu
    check("CRITICAL: Teamly страница: 20 оценок, 6 студентов, упомянут Frank Liu",
          bool(candidates) and has_count and has_students and flagged_ok,
          f"candidates={len(candidates)}, count={has_count}, students={has_students}, flagged_ok={flagged_ok}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Запустить все проверки."""
    print("=== Вычисление ожидаемых значений из БД ===")
    try:
        raw_scores, individual_summary, flagged = compute_expected()
        if raw_scores is None:
            print("  ОШИБКА: форма взаимного оценивания не найдена в БД.")
            return False, "Form not found in database"
        print(f"  raw_scores: {len(raw_scores)}")
        print(f"  студентов в summary: {len(individual_summary)}")
        print(f"  flagged: {flagged}")
    except Exception as e:
        print(f"  ОШИБКА при вычислении ожидаемых значений: {e}")
        import traceback
        traceback.print_exc()
        return False, f"Failed to compute expected values: {e}"

    excel_summary = check_excel(agent_workspace, raw_scores, individual_summary, flagged)
    check_google_sheet(individual_summary, excel_summary)
    check_teamly(individual_summary, flagged, len(raw_scores))

    total = PASS_COUNT + FAIL_COUNT
    pass_rate = PASS_COUNT / total if total > 0 else 0
    pct = 100.0 * pass_rate

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {pct:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL FAIL: {CRITICAL_FAILED}")
    success = (not CRITICAL_FAILED) and pct >= 70.0
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "pass_rate": round(pass_rate, 3),
            "accuracy": round(pct, 1),
            "critical_failed": CRITICAL_FAILED,
            "success": success,
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return success, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {pct:.1f}%"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)

    # Критический гейт раньше порога accuracy
    if CRITICAL_FAILED:
        sys.exit(1)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
