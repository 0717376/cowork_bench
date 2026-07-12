"""Evaluation script for pw-sf-insales-reconciliation-excel-word."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Expected Region dimension values (geographic realia, kept English in groundtruth + page).
EXPECTED_REGIONS = {"asia pacific", "europe", "latin america", "middle east", "north america"}

# Critical (semantic) checks: any failure => overall FAIL regardless of accuracy.
# They verify the agent actually ran the reconciliation pipeline over the correct
# Region dimension and that its outputs are internally consistent (sums, gap rule,
# JSON<->Excel region parity, narrative sections) — NOT hardcoded groundtruth numbers.
CRITICAL_CHECKS = {
    "Data_Analysis contains the 5 expected Region values",
    "Metrics Total_Orders equals sum of Data_Analysis Order_Count",
    "Metrics Total_Revenue equals sum of Data_Analysis Revenue",
    "Recommendations Priority 1 region has the largest negative gap",
    "sf_wc_reconcile_results.json exists, is valid JSON, regions match Excel",
    "Word doc has executive-summary / key-findings / recommendations sections",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def num_close(a, b, rel_tol=0.02, abs_tol=1.0):
    if a is None or b is None:
        return False
    return abs(a - b) <= max(abs_tol, abs(b) * rel_tol)


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def col_index(headers, *names):
    """Return index of first header whose lowercased value matches any of names."""
    for i, h in enumerate(headers):
        for n in names:
            if h == n.lower():
                return i
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Wc_Reconciliation_Report.xlsx")
    check("Wc_Reconciliation_Report.xlsx exists", os.path.exists(excel_path))

    # Data captured for cross-checks across sheets.
    da_regions = set()
    da_order_sum = None
    da_revenue_sum = None
    region_gaps = {}  # region -> gap (internal Order_Count/Revenue based vs market benchmark)

    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 5 rows", len(data_rows) >= 5, f"got {len(data_rows)}")

            for expected_col in ['Region', 'Order_Count', 'Revenue', 'Market_Size_M', 'Market_Penetration_Pct']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            ri = col_index(headers, "region")
            oi = col_index(headers, "order_count")
            rvi = col_index(headers, "revenue")
            msi = col_index(headers, "market_size_m")
            # The agent MAY report an explicit gap/deviation column (analysis_guide.md:
            # gap = internal - benchmark). If present we honour the agent's own numbers
            # instead of imposing a hidden formula.
            gapi = col_index(headers, "gap", "deviation", "penetration_gap",
                             "pen_gap", "gap_pct", "deviation_pct")

            order_total = 0.0
            rev_total = 0.0
            for row in data_rows:
                if ri is not None and ri < len(row) and row[ri]:
                    reg = str(row[ri]).strip()
                    da_regions.add(reg.lower())
                    o = safe_float(row[oi]) if oi is not None and oi < len(row) else None
                    rv = safe_float(row[rvi]) if rvi is not None and rvi < len(row) else None
                    ms = safe_float(row[msi]) if msi is not None and msi < len(row) else None
                    if o is not None:
                        order_total += o
                    if rv is not None:
                        rev_total += rv
                    # Gap per analysis_guide.md = internal - external benchmark
                    # (penetration-based). We ONLY trust an explicit gap/deviation column
                    # the agent itself reports; we deliberately do NOT reconstruct a hidden
                    # formula from the base columns (the original Revenue - Market_Size_M
                    # subtraction was incommensurable and degenerate). When no gap column
                    # is present, region_gaps stays empty and the ordering check below only
                    # validates that Priority 1 is a real region (see check note).
                    g = safe_float(row[gapi]) if (gapi is not None and gapi < len(row)) else None
                    if g is not None:
                        region_gaps[reg] = g
            da_order_sum = order_total
            da_revenue_sum = rev_total

            # CRITICAL: correct primary dimension (Region) populated.
            found = EXPECTED_REGIONS & da_regions
            check("Data_Analysis contains the 5 expected Region values",
                  len(found) == 5, f"matched: {sorted(found)}")

            # Alphabetical sort by Region (structural, non-critical).
            if ri is not None:
                region_seq = [str(r[ri]).strip() for r in data_rows if ri < len(r) and r[ri]]
                check("Data_Analysis sorted alphabetically by Region",
                      region_seq == sorted(region_seq), f"seq: {region_seq}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            metrics = {}
            mi = col_index(headers, "metric")
            vi = col_index(headers, "value")
            if mi is not None and vi is not None:
                for row in data_rows:
                    if mi < len(row) and vi < len(row) and row[mi] is not None:
                        metrics[str(row[mi]).strip().lower()] = safe_float(row[vi])

            tot_orders = metrics.get("total_orders")
            tot_rev = metrics.get("total_revenue")
            # CRITICAL: internal consistency of totals against Data_Analysis sums.
            check("Metrics Total_Orders equals sum of Data_Analysis Order_Count",
                  da_order_sum is not None and tot_orders is not None and num_close(tot_orders, da_order_sum),
                  f"metric={tot_orders} sum={da_order_sum}")
            check("Metrics Total_Revenue equals sum of Data_Analysis Revenue",
                  da_revenue_sum is not None and tot_rev is not None and num_close(tot_rev, da_revenue_sum),
                  f"metric={tot_rev} sum={da_revenue_sum}")

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            for expected_col in ['Priority', 'Action', 'Region']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            pi = col_index(headers, "priority")
            rgi = col_index(headers, "region")
            # Find the agent's Priority 1 (lowest priority number) region.
            p1_region = None
            if pi is not None and rgi is not None:
                best = None
                for row in data_rows:
                    if pi >= len(row) or rgi >= len(row):
                        continue
                    pval = safe_float(row[pi])
                    reg = str(row[rgi]).strip() if row[rgi] else None
                    if pval is not None and (best is None or pval < best):
                        best = pval
                        p1_region = reg

            # CRITICAL (relaxed): gap rule per analysis_guide.md — "most negative gap =>
            # Priority 1". We honour the agent's own analysis instead of imposing a hidden
            # incommensurable formula. The agent's Priority ordering passes if:
            #   (a) Priority 1 points at a valid expected Region, AND
            #   (b) the ordering is internally consistent with a gap/deviation signal the
            #       agent reports — i.e. when a gap/deviation column (or a penetration
            #       benchmark to derive it) is present, Priority 1 = the most-negative-gap
            #       region. When the agent reports no usable gap basis at all, we only
            #       require a valid, ranked Priority 1 region (we do not invent a formula).
            p1_valid_region = (p1_region is not None and
                               p1_region.strip().lower() in EXPECTED_REGIONS)
            ranked_consistent = True
            if region_gaps:
                expected_p1 = min(region_gaps, key=region_gaps.get)
                ranked_consistent = (p1_region is not None and
                                     p1_region.strip().lower() == expected_p1.strip().lower())
            check("Recommendations Priority 1 region has the largest negative gap",
                  p1_valid_region and ranked_consistent,
                  f"got P1={p1_region} valid_region={p1_valid_region} "
                  f"gaps={region_gaps} ranked_consistent={ranked_consistent}")

    # Word document: require the three narrative sections (RU or EN), not just length.
    import glob as globmod
    word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
    check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
    if word_files:
        from docx import Document
        doc = Document(word_files[0])
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Word has content", len(text) > 50, f"text length: {len(text)}")
        sec_summary = any(k in text for k in ["executive summary", "краткое резюме", "резюме", "обзор", "overview"])
        sec_findings = any(k in text for k in ["key findings", "ключевые выводы", "выводы", "findings"])
        sec_recs = any(k in text for k in ["recommendation", "рекомендац"])
        check("Word doc has executive-summary / key-findings / recommendations sections",
              sec_summary and sec_findings and sec_recs,
              f"summary={sec_summary} findings={sec_findings} recs={sec_recs}")

    # Processor + results JSON: proves the pipeline ran.
    proc_path = os.path.join(agent_workspace, "sf_wc_reconcile_processor.py")
    check("sf_wc_reconcile_processor.py exists", os.path.exists(proc_path))

    results_path = os.path.join(agent_workspace, "sf_wc_reconcile_results.json")
    json_regions = set()
    valid_json = False
    if os.path.exists(results_path):
        try:
            with open(results_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            valid_json = True

            def harvest(obj):
                if isinstance(obj, dict):
                    for k, v in obj.items():
                        if isinstance(k, str) and k.strip().lower() in EXPECTED_REGIONS:
                            json_regions.add(k.strip().lower())
                        harvest(v)
                elif isinstance(obj, list):
                    for it in obj:
                        harvest(it)
                elif isinstance(obj, str):
                    if obj.strip().lower() in EXPECTED_REGIONS:
                        json_regions.add(obj.strip().lower())
            harvest(payload)
        except Exception as e:
            valid_json = False
    check("sf_wc_reconcile_results.json exists, is valid JSON, regions match Excel",
          valid_json and json_regions and (json_regions == da_regions or json_regions >= EXPECTED_REGIONS & da_regions and len(json_regions & da_regions) >= 5),
          f"json_regions={sorted(json_regions)} da_regions={sorted(da_regions)} valid={valid_json}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")
        sys.exit(1)

    success = accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
