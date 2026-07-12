"""Evaluation for canvas-at-risk-intervention.

Критические проверки (CRITICAL_CHECKS): провал любой из них => общий FAIL
независимо от accuracy. В остальном PASS требует accuracy >= 70%.

Ожидаемые значения вычисляются «вживую» из БД canvas (read-only, глобальный
сид). Никаких захардкоженных счётчиков для волатильных данных.
"""
import argparse
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Критические (семантические) проверки: провал любой => общий FAIL.
CRITICAL_CHECKS = {
    "At Risk Students row count matches DB at-risk count",
    "Risk_Level thresholds correct (Critical<40, 40<=Warning<50)",
    "Course Summary Avg_Score for Creative Computing (Spring) matches DB avg over ALL enrolled",
    "Every Critical row has counseling referral in Recommended_Support",
    "Email to academic-support@uni.edu with correct subject and at-risk total in body",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def safe_float(val):
    try:
        if val is None:
            return None
        return float(str(val).replace(",", ".").replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def db_at_risk_total():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) FROM canvas.enrollments e
        WHERE e.type = 'StudentEnrollment'
          AND e.grades->>'current_score' IS NOT NULL
          AND (e.grades->>'current_score')::numeric < 50
    """)
    n = cur.fetchone()[0]
    cur.close()
    conn.close()
    return n


def db_ccc_spring_avg():
    """Средний current_score по ВСЕМ записанным StudentEnrollment курса CCC Spring."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT AVG((e.grades->>'current_score')::numeric)
        FROM canvas.enrollments e
        WHERE e.type = 'StudentEnrollment'
          AND e.grades->>'current_score' IS NOT NULL
          AND e.course_id = (
              SELECT id FROM canvas.courses
              WHERE name LIKE '%%вычислен%%Весна%%' LIMIT 1
          )
    """)
    avg = cur.fetchone()[0]
    cur.close()
    conn.close()
    return float(avg) if avg is not None else None


def db_ccc_spring_atrisk():
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) FROM canvas.enrollments e
        WHERE e.type = 'StudentEnrollment'
          AND (e.grades->>'current_score')::numeric < 50
          AND e.grades->>'current_score' IS NOT NULL
          AND e.course_id = (
              SELECT id FROM canvas.courses
              WHERE name LIKE '%%вычислен%%Весна%%' LIMIT 1
          )
    """)
    n = cur.fetchone()[0]
    cur.close()
    conn.close()
    return n


def check_excel(agent_workspace):
    print("\n=== Проверка Excel ===")
    xlsx_path = os.path.join(agent_workspace, "At_Risk_Report.xlsx")
    if not os.path.isfile(xlsx_path):
        check("At_Risk_Report.xlsx exists", False, f"Not found: {xlsx_path}")
        # Зависимые критические проверки помечаем как провал.
        for n in ["At Risk Students row count matches DB at-risk count",
                  "Risk_Level thresholds correct (Critical<40, 40<=Warning<50)",
                  "Course Summary Avg_Score for Creative Computing (Spring) matches DB avg over ALL enrolled",
                  "Every Critical row has counseling referral in Recommended_Support"]:
            check(n, False, "workbook missing")
        return
    check("At_Risk_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return
    check("Excel readable", True)

    # ---------- At Risk Students ----------
    ar_rows = load_sheet_rows(wb, "At Risk Students")
    ar_data = []
    header_lower = []
    if ar_rows is None:
        check("Sheet 'At Risk Students' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'At Risk Students' exists", True)
        ar_data = ar_rows[1:] if len(ar_rows) > 1 else []
        check("At Risk Students has data (>100 rows)", len(ar_data) > 100,
              f"Found {len(ar_data)}")

        header = ar_rows[0] if ar_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        for col in ["student_name", "course", "current_score", "risk_level", "recommended_support"]:
            check(f"Column '{col}' present", any(col in h for h in header_lower),
                  f"Header: {header}")

    # Индексы колонок по заголовку (с запасным вариантом на фиксированный порядок).
    def col_idx(key, default):
        for i, h in enumerate(header_lower):
            if key in h:
                return i
        return default

    i_score = col_idx("current_score", 2)
    i_risk = col_idx("risk_level", 3)
    i_supp = col_idx("recommended_support", 4)

    # Все баллы < 50 (структурная)
    if ar_data:
        all_below_50 = all(
            (safe_float(r[i_score]) is None) or (safe_float(r[i_score]) < 50)
            for r in ar_data if len(r) > i_score
        )
        check("All scores below 50", all_below_50)

        risk_levels = set(str(r[i_risk]).strip() if len(r) > i_risk and r[i_risk] else ""
                          for r in ar_data)
        check("Risk levels include 'Critical' and 'Warning'",
              "Critical" in risk_levels and "Warning" in risk_levels,
              f"Found: {risk_levels}")

    # CRITICAL: число строк ~ счётчику at-risk из БД
    try:
        db_total = db_at_risk_total()
        # допускаем малое расхождение из-за пограничных округлений баллов
        tol = max(2, int(round(db_total * 0.03)))
        check("At Risk Students row count matches DB at-risk count",
              abs(len(ar_data) - db_total) <= tol,
              f"sheet={len(ar_data)} db={db_total} tol={tol}")
    except Exception as e:
        check("At Risk Students row count matches DB at-risk count", False, str(e))

    # CRITICAL: корректность порогов Risk_Level
    if ar_data:
        bad = []
        for r in ar_data:
            if len(r) <= max(i_score, i_risk):
                continue
            sc = safe_float(r[i_score])
            lvl = str(r[i_risk]).strip() if r[i_risk] else ""
            if sc is None:
                continue
            if sc < 40 and lvl != "Critical":
                bad.append((sc, lvl))
            elif 40 <= sc < 50 and lvl != "Warning":
                bad.append((sc, lvl))
        check("Risk_Level thresholds correct (Critical<40, 40<=Warning<50)",
              len(bad) == 0, f"{len(bad)} mismatched, e.g. {bad[:5]}")
    else:
        check("Risk_Level thresholds correct (Critical<40, 40<=Warning<50)", False,
              "no data rows")

    # CRITICAL: для каждой Critical-строки в Recommended_Support есть упоминание
    # консультирования (counseling / консультир / психолог).
    if ar_data:
        counsel_re = re.compile(r"counsel|консультир|психолог", re.IGNORECASE)
        missing = 0
        crit_seen = 0
        for r in ar_data:
            if len(r) <= max(i_risk, i_supp):
                continue
            lvl = str(r[i_risk]).strip() if r[i_risk] else ""
            if lvl == "Critical":
                crit_seen += 1
                supp = str(r[i_supp]) if r[i_supp] else ""
                if not counsel_re.search(supp):
                    missing += 1
        check("Every Critical row has counseling referral in Recommended_Support",
              crit_seen > 0 and missing == 0,
              f"critical={crit_seen} missing_counseling={missing}")
    else:
        check("Every Critical row has counseling referral in Recommended_Support", False,
              "no data rows")

    # ---------- Course Summary ----------
    cs_rows = load_sheet_rows(wb, "Course Summary")
    if cs_rows is None:
        check("Sheet 'Course Summary' exists", False, f"Available: {wb.sheetnames}")
        check("Course Summary Avg_Score for Creative Computing (Spring) matches DB avg over ALL enrolled",
              False, "sheet missing")
    else:
        check("Sheet 'Course Summary' exists", True)
        cs_data = cs_rows[1:] if len(cs_rows) > 1 else []
        check("Course Summary has 22 rows", len(cs_data) == 22, f"Found {len(cs_data)}")

        # колонки Course Summary
        cs_header = [str(h).lower().replace(" ", "_") if h else "" for h in (cs_rows[0] if cs_rows else [])]

        def cs_col(key, default):
            for i, h in enumerate(cs_header):
                if key in h:
                    return i
            return default

        j_count = cs_col("at_risk_count", 1)
        j_avg = cs_col("avg_score", 2)

        # найти строку CCC Spring
        ccc_row = None
        for row in cs_data:
            if row and row[0] and "вычислен" in str(row[0]).lower() \
                    and "весна" in str(row[0]).lower():
                ccc_row = row
                break
        check("Creative Computing (Spring) row found", ccc_row is not None)

        # at-risk count CCC Spring (структурная, мягкая) — оставляем как было
        try:
            expected_cnt = db_ccc_spring_atrisk()
            if ccc_row is not None:
                check(f"CCC Spring at-risk count ~{expected_cnt}",
                      num_close(ccc_row[j_count], expected_cnt, 5),
                      f"Got {ccc_row[j_count]}")
        except Exception as e:
            check("CCC Spring at-risk count", False, str(e))

        # CRITICAL: Avg_Score CCC Spring = среднее по ВСЕМ записанным
        try:
            expected_avg = db_ccc_spring_avg()
            got_avg = safe_float(ccc_row[j_avg]) if ccc_row is not None and len(ccc_row) > j_avg else None
            check("Course Summary Avg_Score for Creative Computing (Spring) matches DB avg over ALL enrolled",
                  expected_avg is not None and got_avg is not None
                  and num_close(got_avg, expected_avg, 2.0),
                  f"got={got_avg} expected={expected_avg}")
        except Exception as e:
            check("Course Summary Avg_Score for Creative Computing (Spring) matches DB avg over ALL enrolled",
                  False, str(e))

    # ---------- Action Plan ----------
    ap_rows = load_sheet_rows(wb, "Action Plan")
    if ap_rows is None:
        check("Sheet 'Action Plan' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Action Plan' exists", True)
        ap_data = ap_rows[1:] if len(ap_rows) > 1 else []
        check("Action Plan has at least 5 actions", len(ap_data) >= 5,
              f"Found {len(ap_data)}")


def check_email():
    print("\n=== Проверка письма ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, subject, to_addr, body_text
            FROM email.messages
            WHERE to_addr::text ILIKE '%%academic-support%%'
               OR to_addr::text ILIKE '%%academic_support%%'
               OR subject ILIKE '%%at-risk%%'
               OR subject ILIKE '%%at_risk%%'
               OR subject ILIKE '%%intervention%%'
        """)
        emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Email to academic-support@uni.edu with correct subject and at-risk total in body",
              False, str(e))
        return

    # структурная: письмо существует и тело непустое
    has_email = len(emails) >= 1
    check("Email sent about at-risk students", has_email, "No matching email found")
    if has_email:
        body0 = str(emails[0][3]) if emails[0][3] else ""
        check("Email body has content", len(body0) > 20, f"Body length: {len(body0)}")

    # CRITICAL: письмо на academic-support@uni.edu с правильной темой И
    # итоговым числом at-risk (цифрами) в теле, совпадающим с БД (с допуском).
    try:
        db_total = db_at_risk_total()
    except Exception:
        db_total = None

    ok = False
    detail = "no qualifying email"
    if db_total is not None:
        candidates = []
        for _id, subj, to_addr, body in emails:
            subj_s = (subj or "")
            to_s = (str(to_addr) or "")
            to_match = "academic-support@uni.edu" in to_s.lower() \
                or "academic-support" in to_s.lower()
            subj_match = "at-risk" in subj_s.lower() and "intervention" in subj_s.lower()
            if to_match and subj_match:
                candidates.append(body or "")
        for body in candidates:
            nums = [int(n) for n in re.findall(r"\d+", body or "")]
            if any(abs(n - db_total) <= max(2, int(round(db_total * 0.03))) for n in nums):
                ok = True
                detail = ""
                break
        if candidates and not ok:
            detail = f"email found but total {db_total} not in body numbers"
        elif not candidates:
            detail = "no email to academic-support with subject At-Risk...Intervention"

    check("Email to academic-support@uni.edu with correct subject and at-risk total in body",
          ok, detail)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Результат: {PASS_COUNT}/{total} пройдено ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"КРИТИЧЕСКИЕ ПРОВАЛЫ: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")
        sys.exit(1)

    if accuracy >= 70:
        print("Все условия выполнены (нет критических провалов, accuracy >= 70%).")
        sys.exit(0)
    else:
        print(f"accuracy {accuracy:.1f}% < 70%")
        sys.exit(1)


if __name__ == "__main__":
    main()
