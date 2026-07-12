"""
Evaluation for the wellness-tracker task (Excel + Word + Teamly).

Критические проверки (CRITICAL_CHECKS): любой провал => общий FAIL независимо
от accuracy. Иначе PASS требует accuracy >= 70% и отсутствия критических провалов.
"""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Критические проверки: любой провал => общий FAIL независимо от accuracy.
CRITICAL_CHECKS = {
    "All Wellness_Score values numeric and in 1-10 across all rows",
    "Recommended consistent with rule (Yes iff Wellness_Score >= 7)",
    "Progress Metrics targets correct (Veggie=5, Sugar=25, Average=mean of scores)",
    "Weekly Plan uses only Recommended=Yes recipes (>=10 rows)",
    "Teamly 'Client Wellness Dashboard' page exists with >= 5 content blocks",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower().replace(" ", "_") == sheet_name.strip().lower().replace(" ", "_"):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
        if name.strip().lower().replace("_", " ") == sheet_name.strip().lower().replace("_", " "):
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def find_col(header, names):
    if not header:
        return None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        cl = str(cell).strip().lower().replace(" ", "_")
        for n in names:
            if n.lower().replace(" ", "_") == cl:
                return i
    return None


def _to_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def check_excel(workspace):
    print("\n=== Checking Excel ===")
    path = os.path.join(workspace, "Wellness_Tracker.xlsx")
    if not os.path.isfile(path):
        record("Excel exists", False, f"Not found: {path}")
        # Зависимые критические проверки помечаем как проваленные.
        record("All Wellness_Score values numeric and in 1-10 across all rows", False, "no excel")
        record("Recommended consistent with rule (Yes iff Wellness_Score >= 7)", False, "no excel")
        record("Progress Metrics targets correct (Veggie=5, Sugar=25, Average=mean of scores)", False, "no excel")
        record("Weekly Plan uses only Recommended=Yes recipes (>=10 rows)", False, "no excel")
        return False
    record("Excel exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)

    # --- Recipe Wellness Score ---
    rw_rows = load_sheet_rows(wb, "Recipe Wellness Score") or load_sheet_rows(wb, "Recipe_Wellness_Score")
    recommended_yes_names = set()
    scores = []
    if rw_rows is None:
        record("Sheet 'Recipe Wellness Score' exists", False, f"Sheets: {wb.sheetnames}")
        record("All Wellness_Score values numeric and in 1-10 across all rows", False, "no sheet")
        record("Recommended consistent with rule (Yes iff Wellness_Score >= 7)", False, "no sheet")
    else:
        record("Sheet 'Recipe Wellness Score' exists", True)
        header = rw_rows[0]
        data = [r for r in rw_rows[1:] if r and r[0] is not None and str(r[0]).strip()]
        record("Has >= 8 recipes", len(data) >= 8, f"Found {len(data)}")

        name_col = find_col(header, ["Recipe_Name", "Recipe Name", "Name"])
        ws_col = find_col(header, ["Wellness_Score", "Wellness Score", "Score"])
        rec_col = find_col(header, ["Recommended", "recommended"])
        cat_col = find_col(header, ["Category", "category"])

        # CRITICAL: все Wellness_Score числовые и в диапазоне 1-10.
        all_scores_ok = ws_col is not None and len(data) > 0
        if ws_col is not None:
            for r in data:
                s = _to_float(r[ws_col]) if ws_col < len(r) else None
                if s is None or not (1 <= s <= 10):
                    all_scores_ok = False
                    break
                scores.append(s)
        record("All Wellness_Score values numeric and in 1-10 across all rows",
               all_scores_ok, f"scores={scores[:12]}")

        # CRITICAL: Recommended == Yes ровно когда Wellness_Score >= 7.
        rule_ok = ws_col is not None and rec_col is not None and len(data) > 0
        if rule_ok:
            for r in data:
                s = _to_float(r[ws_col]) if ws_col < len(r) else None
                rec = str(r[rec_col]).strip().lower() if rec_col < len(r) and r[rec_col] is not None else ""
                if s is None:
                    rule_ok = False
                    break
                expected = "yes" if s >= 7 else "no"
                if rec not in ("yes", "no") or rec != expected:
                    rule_ok = False
                    break
                if rec == "yes" and name_col is not None and name_col < len(r) and r[name_col]:
                    recommended_yes_names.add(str(r[name_col]).strip().lower())
        record("Recommended consistent with rule (Yes iff Wellness_Score >= 7)",
               rule_ok, "mismatch between score and Recommended")

        if rec_col is not None:
            vals = {str(r[rec_col]).strip().lower() for r in data
                    if rec_col < len(r) and r[rec_col] is not None}
            record("Recommended has both Yes and No", "yes" in vals and "no" in vals, f"Values: {vals}")

        if cat_col is not None:
            cats = {str(r[cat_col]).strip().lower() for r in data if cat_col < len(r) and r[cat_col]}
            record("At least 3 categories", len(cats) >= 3, f"Found {len(cats)}: {cats}")

    # --- Progress Metrics (читаем раньше Weekly Plan для среднего) ---
    pm_rows = load_sheet_rows(wb, "Progress Metrics") or load_sheet_rows(wb, "Progress_Metrics")
    if pm_rows is None:
        record("Sheet 'Progress Metrics' exists", False, f"Sheets: {wb.sheetnames}")
        record("Progress Metrics targets correct (Veggie=5, Sugar=25, Average=mean of scores)", False, "no sheet")
    else:
        record("Sheet 'Progress Metrics' exists", True)
        metrics = {}
        for row in pm_rows[1:]:
            if row and row[0]:
                metrics[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        vt_key = next((k for k in metrics if "veggie" in k and "target" in k), None)
        sl_key = next((k for k in metrics if "sugar" in k and "limit" in k), None)
        avg_key = next((k for k in metrics if "average" in k and ("wellness" in k or "score" in k)), None)

        veggie_ok = vt_key is not None and (_to_float(metrics[vt_key]) is not None) and abs(_to_float(metrics[vt_key]) - 5) < 1
        sugar_ok = sl_key is not None and (_to_float(metrics[sl_key]) is not None) and abs(_to_float(metrics[sl_key]) - 25) < 5
        # Average_Wellness_Score == среднее по всем рецептам (округл. до 1 знака).
        avg_ok = False
        if avg_key is not None and scores:
            got = _to_float(metrics[avg_key])
            expected = round(sum(scores) / len(scores), 1)
            avg_ok = got is not None and abs(got - expected) <= 0.15
        record("Progress Metrics targets correct (Veggie=5, Sugar=25, Average=mean of scores)",
               veggie_ok and sugar_ok and avg_ok,
               f"veggie={veggie_ok} sugar={sugar_ok} avg={avg_ok}")

    # --- Weekly Plan ---
    wp_rows = load_sheet_rows(wb, "Weekly Plan") or load_sheet_rows(wb, "Weekly_Plan")
    if wp_rows is None:
        record("Sheet 'Weekly Plan' exists", False, f"Sheets: {wb.sheetnames}")
        record("Weekly Plan uses only Recommended=Yes recipes (>=10 rows)", False, "no sheet")
    else:
        record("Sheet 'Weekly Plan' exists", True)
        wheader = wp_rows[0]
        wdata = [r for r in wp_rows[1:] if r and r[0] is not None]
        record("Weekly Plan has >= 10 rows (5 days x 2 meals)", len(wdata) >= 10, f"Found {len(wdata)}")

        wp_name_col = find_col(wheader, ["Recipe_Name", "Recipe Name", "Name"])
        # CRITICAL: >=10 строк, и каждый Recipe_Name присутствует среди Recommended=Yes.
        plan_ok = len(wdata) >= 10 and wp_name_col is not None and len(recommended_yes_names) > 0
        if plan_ok:
            for r in wdata:
                nm = str(r[wp_name_col]).strip().lower() if wp_name_col < len(r) and r[wp_name_col] else ""
                if not nm or nm not in recommended_yes_names:
                    plan_ok = False
                    break
        record("Weekly Plan uses only Recommended=Yes recipes (>=10 rows)",
               plan_ok, "plan references a non-recommended recipe or too few rows")

    return True


def check_word(workspace):
    print("\n=== Checking Word ===")
    path = os.path.join(workspace, "Wellness_Guide.docx")
    if not os.path.isfile(path):
        record("Word exists", False, f"Not found: {path}")
        return False
    record("Word exists", True)

    try:
        from docx import Document
        doc = Document(path)
        # .lower() ОРИГИНАЛЬНОГО текста — ищем RU и EN ключевые слова.
        text = "\n".join(p.text for p in doc.paragraphs).lower()
        record("Has substantial content", len(text) > 300, f"Only {len(text)} chars")
        record("Mentions wellness/health",
               any(k in text for k in ("wellness", "health", "велнес", "здоров", "самочувств")))
        record("Mentions vegetable/sugar",
               any(k in text for k in ("vegetable", "sugar", "овощ", "сахар")))
        record("Mentions meal plan",
               any(k in text for k in ("meal", "plan", "питани", "план", "рацион")))
        return True
    except Exception as e:
        record("Word readable", False, str(e))
        return False


def check_teamly():
    print("\n=== Checking Teamly ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        # Seed-страницы имеют id <= 3 — рассматриваем только созданные агентом.
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
    except Exception as e:
        record("Teamly 'Client Wellness Dashboard' page exists with >= 5 content blocks", False, str(e))
        return False

    # Заголовок: английский маркер ИЛИ русский эквивалент (велнес/панель/дашборд клиента).
    def title_matches(t):
        tl = (t or "").lower()
        en = ("wellness" in tl or "dashboard" in tl or "client" in tl)
        ru = ("велнес" in tl or "панель" in tl or "дашборд" in tl or
              ("клиент" in tl and ("здоров" in tl or "велнес" in tl or "панель" in tl)))
        return en or ru

    matched = [(pid, title, body) for pid, title, body in pages if title_matches(title)]

    if not matched:
        record("Teamly 'Client Wellness Dashboard' page exists with >= 5 content blocks", False,
               f"new pages: {[(p[0], p[1]) for p in pages]}")
        cur.close()
        conn.close()
        return False

    # Считаем "блоки" ТОЛЬКО для найденной страницы (без глобального фолбэка).
    # Блок = непустая строка тела; markdown-заголовки/пункты считаем отдельно.
    # Дополнительно засчитываем по длине, если MCP хранит тело единым блобом.
    best = 0
    for _pid, _title, body in matched:
        b = str(body)
        lines = [ln for ln in b.splitlines() if ln.strip()]
        n = len(lines)
        # Эвристика для тела-блоба: каждые ~200 непустых символов = 1 блок.
        n = max(n, len(b.strip()) // 200)
        best = max(best, n)

    record("Teamly 'Client Wellness Dashboard' page exists with >= 5 content blocks",
           best >= 5, f"matched={[(p[0], p[1]) for p in matched]} blocks~={best}")

    cur.close()
    conn.close()
    return True


def run_evaluation(agent_workspace):
    check_excel(agent_workspace)
    check_word(agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and accuracy >= 70
    return success


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success = run_evaluation(args.agent_workspace)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
