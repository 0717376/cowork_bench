"""Evaluation script for fetch-canvas-performance-excel-word-email.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

All semantic checks are computed from the agent's OWN outputs (Excel values it
wrote, files it created, emails it sent). The Canvas course set is live/external
and is NOT hardcoded here.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Email subject markers accepted (RU + EN). Agent is told to use the literal
# English subject "Analysis Report Complete"; RU alternatives broaden matching.
SUBJECT_MARKERS = ["report", "analysis", "анализ", "отчёт", "отчет"]

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Data_Analysis sorted alphabetically by Course",
    "Metrics internal consistency (Total_Enrollment + averages)",
    "Recommendations references lowest Pass_Rate course",
    "Analysis email sent to team-lead@company.com",
    "course_perf_processor.py + valid course_perf_results.json",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def col_index(headers, name):
    """Return index of a header (case-insensitive), or -1."""
    name = name.strip().lower()
    for i, h in enumerate(headers):
        if h == name:
            return i
    return -1


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Performance_Report.xlsx")
    excel_exists = os.path.exists(excel_path)
    check("Performance_Report.xlsx exists", excel_exists)

    # Values harvested from Data_Analysis for cross-sheet consistency checks.
    da_courses = []        # Course values, in row order
    da_enrollment = []
    da_avg_score = []
    da_pass_rate = []

    if excel_exists:
        wb = openpyxl.load_workbook(excel_path)

        # ---- Data_Analysis ----
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(data_rows) >= 6, f"got {len(data_rows)}")

            for expected_col in ['Course', 'Code', 'Enrollment', 'Avg_Score', 'Pass_Rate']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            ci_course = col_index(headers, 'course')
            ci_enr = col_index(headers, 'enrollment')
            ci_score = col_index(headers, 'avg_score')
            ci_pass = col_index(headers, 'pass_rate')

            for r in data_rows:
                if not r or all(v is None for v in r):
                    continue
                if ci_course >= 0 and ci_course < len(r) and r[ci_course] is not None:
                    da_courses.append(str(r[ci_course]).strip())
                if ci_enr >= 0 and ci_enr < len(r):
                    v = safe_float(r[ci_enr])
                    if v is not None:
                        da_enrollment.append(v)
                if ci_score >= 0 and ci_score < len(r):
                    v = safe_float(r[ci_score])
                    if v is not None:
                        da_avg_score.append(v)
                if ci_pass >= 0 and ci_pass < len(r):
                    v = safe_float(r[ci_pass])
                    if v is not None:
                        da_pass_rate.append(v)

            # CRITICAL: alphabetical sort by Course (case-insensitive, ascending).
            sorted_ok = da_courses == sorted(da_courses, key=lambda s: s.lower())
            check("Data_Analysis sorted alphabetically by Course",
                  len(da_courses) >= 6 and sorted_ok,
                  f"order: {da_courses[:8]}")
        else:
            check("Data_Analysis sorted alphabetically by Course", False, "no Data_Analysis sheet")

        # ---- Metrics ----
        metrics = {}
        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(data_rows) >= 4, f"got {len(data_rows)}")

            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            for r in data_rows:
                if r and len(r) >= 2 and r[0] is not None:
                    metrics[str(r[0]).strip().lower()] = r[1]

        # CRITICAL: Metrics internal consistency vs Data_Analysis.
        consistency_ok = True
        details = []
        if da_enrollment:
            sum_enr = sum(da_enrollment)
            tot = safe_float(metrics.get('total_enrollment'))
            if tot is None or abs(tot - sum_enr) > 0.5:
                consistency_ok = False
                details.append(f"Total_Enrollment={tot} vs sum={sum_enr}")
        else:
            consistency_ok = False
            details.append("no enrollment values")
        if da_avg_score:
            mean_score = sum(da_avg_score) / len(da_avg_score)
            avs = safe_float(metrics.get('avg_score'))
            if avs is None or abs(avs - mean_score) > 0.5:
                consistency_ok = False
                details.append(f"Avg_Score={avs} vs mean={mean_score:.2f}")
        if da_pass_rate:
            mean_pass = sum(da_pass_rate) / len(da_pass_rate)
            avp = safe_float(metrics.get('avg_pass_rate'))
            if avp is None or abs(avp - mean_pass) > 0.5:
                consistency_ok = False
                details.append(f"Avg_Pass_Rate={avp} vs mean={mean_pass:.2f}")
        check("Metrics internal consistency (Total_Enrollment + averages)",
              consistency_ok, "; ".join(details))

        # ---- Recommendations ----
        rec_courses = []
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            for expected_col in ['Priority', 'Action', 'Course']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            ci_rcourse = col_index(headers, 'course')
            for r in data_rows:
                if r and ci_rcourse >= 0 and ci_rcourse < len(r) and r[ci_rcourse] is not None:
                    rec_courses.append(str(r[ci_rcourse]).strip())

        # CRITICAL: Recommendations references the course with the lowest Pass_Rate
        # (evidence that gap analysis vs the academic_benchmarks was performed).
        if da_courses and da_pass_rate and len(da_courses) == len(da_pass_rate) and rec_courses:
            worst_course = min(zip(da_courses, da_pass_rate), key=lambda t: t[1])[0]
            referenced = any(worst_course.lower() in rc.lower() or rc.lower() in worst_course.lower()
                             for rc in rec_courses)
            check("Recommendations references lowest Pass_Rate course",
                  referenced, f"worst={worst_course!r} not in {rec_courses}")
        else:
            check("Recommendations references lowest Pass_Rate course", False,
                  "insufficient data to validate gap analysis")

        # ---- Word document ----
        import glob as globmod
        word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
        check("Word document exists", len(word_files) >= 1, f"found {len(word_files)} docx files")
        if word_files:
            from docx import Document
            doc = Document(word_files[0])
            text = " ".join(p.text for p in doc.paragraphs)
            check("Word has content", len(text) > 50, f"text length: {len(text)}")

    # ---- Email (CRITICAL): agent actually sent the report to team-lead ----
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, body_text, from_addr
            FROM email.messages
            WHERE to_addr::text ILIKE %s
        """, ('%team-lead@company.com%',))
        rows = cur.fetchall()
        conn.close()
        email_ok = False
        for subj, body, from_addr in rows:
            subj_l = (subj or "").lower()
            # Distinct from the injected newsletter noise.
            if (from_addr or "") == 'newsletter@company.com':
                continue
            if any(m in subj_l for m in SUBJECT_MARKERS) and (body or "").strip():
                email_ok = True
                break
        check("Analysis email sent to team-lead@company.com",
              email_ok, f"{len(rows)} msgs to team-lead, none matched subject+body")
    except Exception as e:
        check("Analysis email sent to team-lead@company.com", False, str(e))

    # ---- Python pipeline (CRITICAL): processor ran and produced valid JSON ----
    proc_path = os.path.join(agent_workspace, "course_perf_processor.py")
    results_path = os.path.join(agent_workspace, "course_perf_results.json")
    pipeline_ok = os.path.exists(proc_path) and os.path.exists(results_path)
    if pipeline_ok:
        try:
            with open(results_path, "r", encoding="utf-8") as f:
                json.load(f)
        except Exception as e:
            pipeline_ok = False
            print(f"  [info] course_perf_results.json not valid JSON: {e}")
    check("course_perf_processor.py + valid course_perf_results.json",
          pipeline_ok,
          f"processor={os.path.exists(proc_path)} results={os.path.exists(results_path)}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w", encoding="utf-8") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
