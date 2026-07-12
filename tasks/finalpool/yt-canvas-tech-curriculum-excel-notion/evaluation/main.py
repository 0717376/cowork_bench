"""
Evaluation for yt-canvas-tech-curriculum-excel-notion (russified, teamly swap).

The agent honestly reads russified Canvas courses (names contain «вычисления»
or «данных») via the live Canvas MCP, reads Fireship videos from the YouTube
video DB, builds Curriculum_Video_Map.xlsx (Course_Videos + Summary sheets) and
creates a teamly knowledge-base page titled "Tech Course Video Resources".

Checks:
1. Curriculum_Video_Map.xlsx exists with Course_Videos and Summary sheets
2. Course_Videos has required columns and rows for the matched courses
3. Summary sheet has Total_Courses_Mapped, Total_Videos_Recommended,
   Avg_Views_Recommended with correct values
4. Teamly page 'Tech Course Video Resources' exists with course/video content
5. Groundtruth XLSX value comparison (row counts + first rows)

Critical checks (see CRITICAL_CHECKS): any failure => overall FAIL regardless
of accuracy. Otherwise pass threshold: accuracy >= 70%.

RU note: the agent legitimately writes Russian prose around the English literal
markers (video titles, headers). RU keyword checks search .lower() ORIGINAL text
(never normalize()).
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Expected groundtruth aggregate values (Computing + Data courses, 6 courses x 2
# videos = 12 rows). See groundtruth_workspace/Curriculum_Video_Map.xlsx.
EXPECTED_TOTAL_COURSES = 6
EXPECTED_TOTAL_VIDEOS = 12
EXPECTED_AVG_VIEWS = 1252213

# Semantic critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Course_Videos has exactly 12 data rows (6 courses x 2 videos)",
    "Course_Videos covers both Computing and Data course types",
    "Summary Total_Videos_Recommended equals 12",
    "Summary Avg_Views_Recommended approximately 1252213",
    "Teamly 'Tech Course Video Resources' page has course sections with video content",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def find_summary_value(sum_rows, label_keywords):
    """Return the numeric value on a Summary row whose first cell matches a label."""
    for r in sum_rows:
        if not r:
            continue
        label = str(r[0]).strip().lower() if r[0] is not None else ""
        if any(k in label for k in label_keywords):
            for cell in r[1:]:
                if cell is None:
                    continue
                try:
                    return float(cell)
                except (TypeError, ValueError):
                    continue
    return None


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Curriculum_Video_Map.xlsx ===")
    xlsx_path = os.path.join(agent_workspace, "Curriculum_Video_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Curriculum_Video_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        return
    record("Curriculum_Video_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # Check Course_Videos sheet
    cv_idx = next((i for i, s in enumerate(sheet_names_lower)
                   if "course_video" in s or "course video" in s), None)
    if cv_idx is None:
        record("Course_Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Course_Videos sheet exists", True)

    ws = wb[wb.sheetnames[cv_idx]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        record("Has data rows", False, "Sheet is empty")
        return

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    has_course = any("course" in h for h in headers)
    has_video = any("video" in h for h in headers)
    has_view = any("view" in h for h in headers)
    has_duration = any("duration" in h for h in headers)
    record("Has required columns (Course, Video, View, Duration)",
           has_course and has_video and has_view and has_duration,
           f"Headers: {rows[0]}")

    data_rows = [r for r in rows[1:] if any(c for c in r)]

    # ---- CRITICAL: exactly 12 data rows (6 courses x 2 videos) ----
    record("Course_Videos has exactly 12 data rows (6 courses x 2 videos)",
           len(data_rows) == EXPECTED_TOTAL_VIDEOS,
           f"Found {len(data_rows)} data rows")

    all_text = " ".join(str(c) for r in rows for c in r if c).lower()
    # Computing courses: «Креативные вычисления ...»; Data courses:
    # «Проектирование на основе данных ...». RU keyword on ORIGINAL .lower() text.
    has_computing = "вычислен" in all_text or "креативн" in all_text
    has_data = "на основе данных" in all_text or "проектирован" in all_text
    # ---- CRITICAL: both course types present ----
    record("Course_Videos covers both Computing and Data course types",
           has_computing and has_data,
           f"Computing:{has_computing}, Data:{has_data}")

    # Video titles (English literal) should be present in the rows.
    has_video_titles = (
        "some bad code just broke" in all_text
        or "uk demands backdoor" in all_text
        or "silly linux mistake" in all_text
        or "21-year old dev" in all_text
    )
    record("Course_Videos rows contain expected Fireship video titles",
           has_video_titles,
           f"Snippet: {all_text[:200]}")

    # Check Summary sheet
    sum_idx = next((i for i, s in enumerate(sheet_names_lower) if "summary" in s), None)
    if sum_idx is None:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}")
        return
    record("Summary sheet exists", True)

    ws_sum = wb[wb.sheetnames[sum_idx]]
    sum_rows = list(ws_sum.iter_rows(values_only=True))
    sum_text = " ".join(str(c) for r in sum_rows for c in r if c).lower()
    has_total_courses = "total_courses" in sum_text or "total courses" in sum_text
    has_total_videos = "total_videos" in sum_text or "total videos" in sum_text
    has_avg = "avg" in sum_text
    record("Summary has required labels (Total_Courses, Total_Videos, Avg_Views)",
           has_total_courses and has_total_videos and has_avg,
           f"Summary text: {sum_text[:200]}")

    # ---- CRITICAL: Summary numeric values match groundtruth aggregates ----
    tc_val = find_summary_value(sum_rows, ["total_courses", "total courses"])
    record("Summary Total_Courses_Mapped equals 6",
           tc_val is not None and num_close(tc_val, EXPECTED_TOTAL_COURSES, 0.5),
           f"matched={tc_val}")

    tv_val = find_summary_value(sum_rows, ["total_videos", "total videos"])
    record("Summary Total_Videos_Recommended equals 12",
           tv_val is not None and num_close(tv_val, EXPECTED_TOTAL_VIDEOS, 0.5),
           f"matched={tv_val}")

    av_val = find_summary_value(sum_rows, ["avg", "average"])
    record("Summary Avg_Views_Recommended approximately 1252213",
           av_val is not None and num_close(av_val, EXPECTED_AVG_VIEWS,
                                            max(EXPECTED_AVG_VIEWS * 0.02, 1.0)),
           f"matched={av_val}")

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Curriculum_Video_Map.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                       f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                       if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_teamly():
    print("\n=== Check 2: Teamly page 'Tech Course Video Resources' ===")
    crit = "Teamly 'Tech Course Video Resources' page has course sections with video content"
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT title, COALESCE(body, '') FROM teamly.pages")
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        record("Teamly page 'Tech Course Video Resources' exists", False, f"db error: {e}")
        record(crit, False, f"db error: {e}")
        return

    pages = [
        (t, b) for t, b in rows
        if t and ("tech course video" in t.lower()
                  or "course video resource" in t.lower())
    ]
    record("Teamly page 'Tech Course Video Resources' exists", len(pages) >= 1,
           f"Total pages: {len(rows)}")

    page_text = " ".join((str(t) + " " + str(b)) for t, b in pages)
    page_lower = page_text.lower()

    # RU keyword on ORIGINAL .lower() text (never normalize()).
    has_computing = "вычислен" in page_lower or "креативн" in page_lower
    has_data = "на основе данных" in page_lower or "проектирован" in page_lower
    has_video = (
        "video" in page_lower or "fireship" in page_lower
        or "some bad code" in page_lower or "uk demands backdoor" in page_lower
    )
    # ---- CRITICAL: page exists AND has course sections AND video content ----
    record(crit,
           bool(pages) and (has_computing or has_data) and has_video,
           f"computing={has_computing}, data={has_data}, video={has_video}, "
           f"snippet={page_lower[:200]}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("YT CANVAS TECH CURRICULUM EXCEL TEAMLY - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
        "success": (not critical_failed) and accuracy >= 70,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("Overall: PASS")
        sys.exit(0)
    print("Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
