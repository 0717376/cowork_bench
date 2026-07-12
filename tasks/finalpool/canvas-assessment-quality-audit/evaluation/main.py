"""Evaluation for canvas-assessment-quality-audit."""
import argparse
import os
import re
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")


def num_close(a, b, tol=2.0):
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def to_float(v):
    if v is None:
        return None
    try:
        return float(str(v).replace("%", "").strip())
    except (TypeError, ValueError):
        return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def get_expected_quiz_data():
    """Get expected quiz data from DB (canvas.* seeded globally, read live)."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c.name, q.title,
               (SELECT COUNT(*) FROM canvas.quiz_questions qq WHERE qq.quiz_id = q.id) as qcount,
               ROUND(AVG(qs.score), 1) as avg_score,
               q.points_possible
        FROM canvas.quizzes q
        JOIN canvas.courses c ON c.id = q.course_id
        LEFT JOIN canvas.quiz_submissions qs ON qs.quiz_id = q.id
        GROUP BY c.name, q.id, q.title, q.points_possible
        ORDER BY c.name, q.title
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()

    quiz_data = []
    for course, quiz, qcount, avg_score, pts in rows:
        if pts and float(pts) > 0 and avg_score and qcount > 0:
            difficulty = round(float(avg_score) / (float(pts) * qcount), 3)
        else:
            difficulty = 0
        flagged = difficulty < 0.3 or difficulty > 0.8
        issue = "too easy" if difficulty > 0.8 else ("too hard" if difficulty < 0.3 else "")
        quiz_data.append({
            "course": course, "quiz": quiz, "qcount": qcount,
            "avg_score": float(avg_score) if avg_score else 0,
            "difficulty": difficulty, "flagged": flagged, "issue": issue,
        })
    return quiz_data


def _key(course, quiz):
    return (str(course).strip().lower(), str(quiz).strip().lower())


def check_excel(agent_workspace):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Assessment_Quality.xlsx")
    if not os.path.isfile(xlsx_path):
        check("Assessment_Quality.xlsx exists", False, f"Not found: {xlsx_path}", critical=True)
        return
    check("Assessment_Quality.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e), critical=True)
        return
    check("Excel readable", True)

    expected = get_expected_quiz_data()
    expected_by_key = {_key(q["course"], q["quiz"]): q for q in expected}
    total_quizzes = len(expected)
    flagged_expected = [q for q in expected if q["flagged"]]
    flagged_count = len(flagged_expected)
    total_questions = sum(q["qcount"] for q in expected)

    # --- Quiz Overview sheet ---
    qo_rows = load_sheet_rows(wb, "Quiz Overview")
    if qo_rows is None:
        check("Sheet 'Quiz Overview' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Quiz Overview' exists", True)
        data_rows = qo_rows[1:] if len(qo_rows) > 1 else []
        check(f"Quiz Overview has {total_quizzes} rows",
              abs(len(data_rows) - total_quizzes) <= 2,
              f"Found {len(data_rows)}")

        header = qo_rows[0] if qo_rows else []
        header_lower = [str(h).lower().replace(" ", "_") if h else "" for h in header]
        col_idx = {}
        for col in ["course", "quiz", "question_count", "avg_score", "avg_difficulty"]:
            present = any(col in h for h in header_lower)
            check(f"Column '{col}' present", present, f"Header: {header}")
            for i, h in enumerate(header_lower):
                if col in h:
                    col_idx[col] = i
                    break

        # CRITICAL: per-quiz difficulty index correctness against DB groundtruth.
        # Mirror eval rounding exactly: avg_score ROUND 1dp, difficulty round(_,3).
        if {"course", "quiz", "avg_difficulty"} <= set(col_idx) and data_rows:
            matched = 0
            comparable = 0
            for r in data_rows:
                ci, qi, di = col_idx["course"], col_idx["quiz"], col_idx["avg_difficulty"]
                if ci >= len(r) or qi >= len(r) or di >= len(r):
                    continue
                exp = expected_by_key.get(_key(r[ci], r[qi]))
                if exp is None:
                    continue
                comparable += 1
                got = to_float(r[di])
                if got is not None and abs(got - exp["difficulty"]) <= 0.02:
                    matched += 1
            ratio = (matched / comparable) if comparable else 0
            check("Per-quiz Avg_Difficulty matches DB for >=80% of rows (tol 0.02)",
                  comparable > 0 and ratio >= 0.8,
                  f"matched {matched}/{comparable} (ratio {ratio:.2f})", critical=True)

    # --- Flagged Items sheet ---
    fi_rows = load_sheet_rows(wb, "Flagged Items")
    if fi_rows is None:
        check("Sheet 'Flagged Items' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Flagged Items' exists", True)
        data_rows = fi_rows[1:] if len(fi_rows) > 1 else []
        check(f"Flagged Items has ~{flagged_count} rows",
              abs(len(data_rows) - flagged_count) <= 5,
              f"Found {len(data_rows)}, expected ~{flagged_count}")

        # Locate Course / Quiz / Issue columns by header
        fheader = fi_rows[0] if fi_rows else []
        fhl = [str(h).lower().replace(" ", "_") if h else "" for h in fheader]
        fidx = {}
        for col in ["course", "quiz", "issue"]:
            for i, h in enumerate(fhl):
                if col in h:
                    fidx[col] = i
                    break

        # CRITICAL: flagged set correctness + Issue literal correctness across ALL rows.
        if {"course", "quiz", "issue"} <= set(fidx):
            got_set = set()
            issue_ok = True
            issue_detail = ""
            for r in data_rows:
                ci, qi, ii = fidx["course"], fidx["quiz"], fidx["issue"]
                if ci >= len(r) or qi >= len(r):
                    continue
                k = _key(r[ci], r[qi])
                got_set.add(k)
                exp = expected_by_key.get(k)
                issue_val = str(r[ii]).strip().lower() if ii < len(r) and r[ii] else ""
                if exp is not None and exp["issue"]:
                    if issue_val != exp["issue"]:
                        issue_ok = False
                        if not issue_detail:
                            issue_detail = f"{k}: got '{issue_val}' exp '{exp['issue']}'"
            exp_set = {_key(q["course"], q["quiz"]) for q in flagged_expected}
            check("Flagged Items set equals DB flagged set (difficulty<0.3 or >0.8)",
                  got_set == exp_set,
                  f"missing={exp_set - got_set} extra={got_set - exp_set}", critical=True)
            check("Every Issue value correct ('Too Easy'>0.8 / 'Too Hard'<0.3) across ALL rows",
                  issue_ok and len(got_set) > 0, issue_detail or "no rows", critical=True)
        else:
            check("Flagged Items has Course/Quiz/Issue columns", False, f"Header: {fheader}")

    # --- Summary sheet ---
    sum_rows = load_sheet_rows(wb, "Summary")
    if sum_rows is None:
        check("Sheet 'Summary' exists", False, f"Available: {wb.sheetnames}")
    else:
        check("Sheet 'Summary' exists", True)
        data_rows = sum_rows[1:] if len(sum_rows) > 1 else []
        lookup = {}
        for row in data_rows:
            if row and row[0]:
                lookup[str(row[0]).strip().lower().replace(" ", "_")] = row[1] if len(row) > 1 else None

        # CRITICAL: exact total quizzes & total questions.
        check(f"Total_Quizzes exactly = {total_quizzes}",
              num_close(lookup.get("total_quizzes"), total_quizzes, 0),
              f"Got {lookup.get('total_quizzes')}", critical=True)
        check(f"Total_Questions exactly = {total_questions}",
              num_close(lookup.get("total_questions"), total_questions, 0),
              f"Got {lookup.get('total_questions')}", critical=True)
        check(f"Flagged_Quizzes close to {flagged_count}",
              num_close(lookup.get("flagged_quizzes"), flagged_count, 5),
              f"Got {lookup.get('flagged_quizzes')}")


def check_word(agent_workspace):
    print("\n=== Checking Word Document ===")
    docx_path = os.path.join(agent_workspace, "Assessment_Report.docx")
    if not os.path.isfile(docx_path):
        check("Assessment_Report.docx exists", False, f"Not found: {docx_path}")
        return
    check("Assessment_Report.docx exists", True)
    check("Word doc has content (> 1KB)", os.path.getsize(docx_path) > 1000,
          f"Size: {os.path.getsize(docx_path)}")

    try:
        from docx import Document
        doc = Document(docx_path)
        # ORIGINAL lower-cased text (NEVER normalized) so Cyrillic survives.
        all_text = " ".join(p.text for p in doc.paragraphs).lower()

        def any_in(subs):
            return any(s in all_text for s in subs)

        check("Report mentions difficulty (RU/EN)",
              any_in(["difficult", "сложн", "трудн"]), f"Sample: {all_text[:200]}")
        check("Report mentions flagged/quality (RU/EN)",
              any_in(["flag", "quality", "флаг", "помеч", "качеств"]),
              f"Sample: {all_text[:200]}")
        check("Report has recommendations (RU/EN)",
              any_in(["recommend", "рекоменд"]), f"Sample: {all_text[:200]}")

        # CRITICAL: narrative reflects real DB numbers, not boilerplate.
        try:
            expected = get_expected_quiz_data()
            total_quizzes = len(expected)
            flagged_count = sum(1 for q in expected if q["flagged"])
            nums = set(re.findall(r"\d+", all_text))
            # CRITICAL: narrative must cite the real total quiz count from the DB
            # (confirms the report reflects actual data, not boilerplate).
            check("Report contains correct total quiz count (matches DB)",
                  str(total_quizzes) in nums,
                  f"need quizzes={total_quizzes}; nums={sorted(nums)[:30]}",
                  critical=True)
            # Non-critical: flagged count also stated as a number (task requests it).
            check("Report also states flagged count (matches DB)",
                  str(flagged_count) in nums,
                  f"need flagged={flagged_count}; nums={sorted(nums)[:30]}")
        except Exception as e:
            check("Report numeric verification (DB)", False, str(e), critical=True)
    except ImportError:
        check("python-docx available", False)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    # CRITICAL gate: any critical failure => FAIL regardless of accuracy.
    if CRITICAL_FAILED:
        print(f"CRITICAL checks failed ({len(CRITICAL_FAILED)}): {CRITICAL_FAILED}")
        sys.exit(1)

    # Accuracy gate.
    if accuracy >= 70:
        print("All critical checks passed and accuracy >= 70%. PASS")
        sys.exit(0)
    else:
        print(f"Accuracy {accuracy:.1f}% < 70%. FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
