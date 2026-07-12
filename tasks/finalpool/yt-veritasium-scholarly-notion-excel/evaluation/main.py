"""
Evaluation for yt-veritasium-scholarly-teamly-excel task.

Checks:
1. Science_Resource_Map.xlsx exists with Videos and Papers sheets
2. Videos sheet has 5 rows with the top Veritasium videos in view-desc order,
   correct Video_IDs / titles / view counts, and a Main_Topic column
3. Papers sheet lists the topical papers with correct citation counts
4. Teamly page 'Science Video-Paper Resource Map' exists with substantive
   content (bridges several real video topics to paper counts)

CRITICAL_CHECKS (semantic): any failure => overall FAIL regardless of accuracy.
Pass threshold otherwise: accuracy >= 70%.
"""
import json
import os
import sys
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks reflect the task's substance (correct ranking by views,
# correct citation counts pulled from the paper DB, a real bridge page), not
# mere structure.
CRITICAL_CHECKS = {
    "Rank 1 video is 'How One Company Secretly Poisoned The Planet'",
    "Videos sheet has top-5 Veritasium videos in view-desc order",
    "Papers sheet citation counts correct for key papers",
    "Teamly page has substantive video/topic content",
}

# Top-5 most-viewed Veritasium videos (from youtube.videos, read-only source),
# ordered by view_count desc. (video_id, title, view_count)
EXPECTED_VIDEOS = [
    ("SC2eSujzrUY", "How One Company Secretly Poisoned The Planet", 32716281),
    ("88bMVbx1dzM", "Can you keep zooming in forever?", 24771026),
    ("Q56PMJbCFXQ", "How a Student's Question Saved This NYC Skyscraper", 23724425),
    ("qJZ1Ez28C-A", "Something Strange Happens When You Trust Quantum Mechanics", 17545470),
    ("Q10_srZ-pbs", "The Closest We've Come to a Theory of Everything", 16347242),
]

# Per-paper citation source-of-truth (title-keyword -> citation_count).
PAPER_CITATIONS = {
    "environmental impact of industrial pollutants": 210,
    "atmospheric dispersion of toxic chemicals": 145,
    "long-term health effects of chemical pollution": 189,
    "fractal geometry in natural structures": 98,
    "infinite series and convergence in optical microscopy": 67,
    "mandelbrot set properties": 54,
    "resonance failures in tall building design": 134,
    "wind load analysis for high-rise": 112,
    "tuned mass dampers": 88,
    "quantum superposition and measurement": 276,
    "quantum entanglement as a resource": 198,
    "quantum randomness and its applications": 143,
    "string theory and the quest for unification": 321,
    "loop quantum gravity": 245,
    "the standard model and beyond": 287,
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1: Science_Resource_Map.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Science_Resource_Map.xlsx")
    if not os.path.exists(xlsx_path):
        record("Science_Resource_Map.xlsx exists", False, f"Not found at {xlsx_path}")
        # Critical checks that depend on the file fail too.
        record("Rank 1 video is 'How One Company Secretly Poisoned The Planet'", False, "no xlsx")
        record("Videos sheet has top-5 Veritasium videos in view-desc order", False, "no xlsx")
        record("Papers sheet citation counts correct for key papers", False, "no xlsx")
        return
    record("Science_Resource_Map.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        record("Rank 1 video is 'How One Company Secretly Poisoned The Planet'", False, "xlsx unreadable")
        record("Videos sheet has top-5 Veritasium videos in view-desc order", False, "xlsx unreadable")
        record("Papers sheet citation counts correct for key papers", False, "xlsx unreadable")
        return
    record("Excel readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # --- Videos sheet ---
    if "videos" not in sheet_names_lower:
        record("Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Rank 1 video is 'How One Company Secretly Poisoned The Planet'", False, "no Videos sheet")
        record("Videos sheet has top-5 Veritasium videos in view-desc order", False, "no Videos sheet")
    else:
        record("Videos sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("videos")]]
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0] if rows else ()
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
        record("Videos sheet has 5 rows (top Veritasium videos)", len(data_rows) == 5,
               f"Found {len(data_rows)} rows")

        # Topic column
        has_topic_col = any("topic" in norm(c) for c in header)
        record("Videos sheet has Main_Topic column", has_topic_col, f"Headers: {header}")

        # Locate columns
        hidx = {norm(c): i for i, c in enumerate(header) if c is not None}
        def col(*cands):
            for cand in cands:
                for h, i in hidx.items():
                    if cand in h:
                        return i
            return None
        rank_i = col("rank")
        vid_i = col("video_id")
        title_i = col("title")
        views_i = col("view")

        # CRITICAL: rank 1 row is the "poisoned the planet" video
        rank1_ok = False
        if data_rows:
            row1 = None
            if rank_i is not None:
                for r in data_rows:
                    if rank_i < len(r) and num_close(r[rank_i], 1, 0.5):
                        row1 = r
                        break
            if row1 is None:
                row1 = data_rows[0]
            t = norm(row1[title_i]) if (title_i is not None and title_i < len(row1)) else ""
            vid = norm(row1[vid_i]) if (vid_i is not None and vid_i < len(row1)) else ""
            rank1_ok = ("poisoned the planet" in t) or (norm(EXPECTED_VIDEOS[0][0]) in vid)
        record("Rank 1 video is 'How One Company Secretly Poisoned The Planet'", rank1_ok,
               "Rank 1 row is not the expected top video")

        # CRITICAL: all 5 expected videos present in view-desc order (by id or title)
        order_ok = False
        if len(data_rows) >= 5:
            order_ok = True
            for idx, (exp_id, exp_title, _) in enumerate(EXPECTED_VIDEOS):
                r = data_rows[idx]
                t = norm(r[title_i]) if (title_i is not None and title_i < len(r)) else ""
                vid = norm(r[vid_i]) if (vid_i is not None and vid_i < len(r)) else ""
                if norm(exp_id) not in vid and norm(exp_title) not in t:
                    order_ok = False
                    break
        record("Videos sheet has top-5 Veritasium videos in view-desc order", order_ok,
               "Videos missing or not ordered by views desc")

        # Views match for top 3 (structural sanity)
        if views_i is not None and len(data_rows) >= 3:
            views_ok = all(
                num_close(data_rows[i][views_i] if views_i < len(data_rows[i]) else None,
                          EXPECTED_VIDEOS[i][2], max(EXPECTED_VIDEOS[i][2] * 0.02, 1))
                for i in range(3)
            )
            record("Video view counts match source for top 3", views_ok,
                   "View counts deviate from source")

    # --- Papers sheet ---
    cites_ok = False
    if "papers" not in sheet_names_lower:
        record("Papers sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Papers sheet citation counts correct for key papers", False, "no Papers sheet")
    else:
        record("Papers sheet exists", True)
        ws2 = wb[wb.sheetnames[sheet_names_lower.index("papers")]]
        rows2 = list(ws2.iter_rows(values_only=True))
        header2 = rows2[0] if rows2 else ()
        data_rows2 = [r for r in rows2[1:] if any(c is not None and str(c).strip() for c in r)]
        record("Papers sheet has at least 12 rows",
               len(data_rows2) >= 12, f"Found {len(data_rows2)} rows")

        headers2 = [norm(c) for c in header2]
        has_topic = any("topic" in h for h in headers2)
        has_title = any("title" in h for h in headers2)
        has_year = any("year" in h for h in headers2)
        record("Papers sheet has Topic, Paper_Title, Year columns",
               has_topic and has_title and has_year, f"Headers: {header2}")

        hidx2 = {norm(c): i for i, c in enumerate(header2) if c is not None}
        title_i = next((i for h, i in hidx2.items() if "paper_title" in h or h == "title" or "title" in h), None)
        cite_i = next((i for h, i in hidx2.items() if "citation" in h), None)

        # CRITICAL: citation counts correct for the key papers present
        if title_i is not None and cite_i is not None:
            matched = 0
            wrong = []
            for r in data_rows2:
                t = norm(r[title_i]) if title_i < len(r) else ""
                for kw, exp_c in PAPER_CITATIONS.items():
                    if kw in t:
                        matched += 1
                        actual = r[cite_i] if cite_i < len(r) else None
                        if not num_close(actual, exp_c, 1):
                            wrong.append((kw, actual, exp_c))
                        break
            cites_ok = (matched >= 12 and not wrong)
            record("Papers sheet citation counts correct for key papers", cites_ok,
                   f"matched={matched} (need >=12), wrong={wrong[:3]}")
        else:
            record("Papers sheet citation counts correct for key papers", False,
                   f"Missing Paper_Title/Citations column; header={header2}")

    # --- Groundtruth XLSX value comparison (structural sanity, non-critical) ---
    gt_path = os.path.join(groundtruth_workspace, "Science_Resource_Map.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False, f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
        gt_wb.close()


def check_teamly():
    print("\n=== Check 2: Teamly page 'Science Video-Paper Resource Map' ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Teamly database accessible", False, str(e))
        record("Teamly page has substantive video/topic content", False, str(e))
        return
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, title, COALESCE(body, '')
            FROM teamly.pages
            WHERE title ILIKE '%%science video%%paper%%'
               OR title ILIKE '%%resource map%%'
        """)
        pages = cur.fetchall()
        record("Teamly page 'Science Video-Paper Resource Map' exists", len(pages) >= 1,
               f"Found {len(pages)} matching pages")

        if not pages:
            record("Teamly page has substantive video/topic content", False, "No matching page")
            return

        combined = " ".join((str(t) + " " + str(b)).lower() for _, t, b in pages)
        max_len = max(len(str(b)) for _, _, b in pages)
        record("Teamly page has non-trivial body", max_len >= 100,
               f"Longest matching body is {max_len} chars")

        # CRITICAL semantic: page references the real videos / topics.
        video_keywords = ["poisoned", "zooming", "skyscraper", "quantum", "theory of everything"]
        topic_keywords = ["pollution", "fractal", "structural", "quantum", "theoretical"]
        video_hits = sum(1 for kw in video_keywords if kw in combined)
        topic_hits = sum(1 for kw in topic_keywords if kw in combined)
        substantive = video_hits >= 3 and topic_hits >= 2
        record("Teamly page has substantive video/topic content", substantive,
               f"video_hits={video_hits} (need >=3), topic_hits={topic_hits} (need >=2)")
    except Exception as e:
        record("Teamly page check", False, str(e))
        record("Teamly page has substantive video/topic content", False, str(e))
    finally:
        cur.close()
        conn.close()


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace, args.groundtruth_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
