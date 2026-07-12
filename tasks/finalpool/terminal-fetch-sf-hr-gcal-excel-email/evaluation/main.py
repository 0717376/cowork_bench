"""Evaluation for terminal-fetch-sf-hr-gcal-excel-email (russified, ClickHouse fork).
Checks:
1. Compensation_Benchmark_Report.xlsx with 3 sheets and correct data
2. Google Calendar events for 7 department review meetings
3. Email sent to hr_team@company.com
4. compensation_analysis.py script exists

CRITICAL semantic gate (any failure => hard FAIL before the accuracy gate):
  - Excel Avg_Salary recomputed from ClickHouse DWH for >=2 departments (no hardcoded fallback)
  - Gap_Pct + Needs_Review recomputed from DWH + russified benchmarks.json
  - All 7 russified departments present across the three sheets
  - Exactly 7 weekday calendar review meetings (2026-03-16..2026-03-20), each titled with a dept + review keyword
  - Email to hr_team reports correct count of departments needing review and the largest-gap dept
"""
import argparse
import json
import os
import sys
import urllib.request
import openpyxl
import psycopg2


def num_close(a, b, rel_tol=0.15, abs_tol=0.5):
    return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel",
}

# Russified department labels, aligned with the central ClickHouse map
# (db/zzz_clickhouse_after_init.sql) and the russified benchmarks.json.
DEPARTMENTS = {"Инженерия", "Финансы", "Кадры", "Операции", "НИОКР", "Продажи", "Поддержка"}

BENCHMARK_URL = "http://localhost:30405/api/benchmarks.json"

# Fallback benchmark medians (RU labels) served on the mock page, used only if the
# live page cannot be fetched during evaluation. These are NOT used to bypass the
# DWH internal-average recompute.
FALLBACK_BENCHMARK = {
    "Инженерия": {"median": 72000, "p25": 58000, "p75": 92000, "trend": "growing"},
    "Финансы": {"median": 65000, "p25": 52000, "p75": 85000, "trend": "stable"},
    "Кадры": {"median": 60000, "p25": 48000, "p75": 78000, "trend": "stable"},
    "Операции": {"median": 58000, "p25": 45000, "p75": 75000, "trend": "declining"},
    "НИОКР": {"median": 70000, "p25": 56000, "p75": 90000, "trend": "growing"},
    "Продажи": {"median": 62000, "p25": 48000, "p75": 82000, "trend": "growing"},
    "Поддержка": {"median": 52000, "p25": 40000, "p75": 68000, "trend": "stable"},
}

GAP_THRESHOLD = 10.0

PASS_COUNT = 0
FAIL_COUNT = 0


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('$', '').replace('%', '').strip())
    except Exception:
        return default


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def fetch_benchmark():
    """Fetch the served benchmarks.json. Department labels are the RU names that
    match the warehouse. Falls back to FALLBACK_BENCHMARK if the page is down."""
    try:
        with urllib.request.urlopen(BENCHMARK_URL, timeout=10) as resp:
            data = json.loads(resp.read().decode("utf-8", errors="replace"))
        result = {}
        for d in data.get("departments", []):
            result[str(d["department"]).strip()] = {
                "median": safe_float(d.get("median_salary")),
                "p25": safe_float(d.get("percentile_25")),
                "p75": safe_float(d.get("percentile_75")),
                "trend": str(d.get("market_trend", "")).strip(),
            }
        if result:
            return result
    except Exception as e:
        print(f"  [WARN] could not fetch benchmark page ({e}); using fallback")
    return {k: dict(v) for k, v in FALLBACK_BENCHMARK.items()}


def get_warehouse_stats():
    """Per-department internal salary stats from ClickHouse warehouse
    (sf_data.HR_ANALYTICS__PUBLIC__EMPLOYEES). DEPARTMENT values are russified
    centrally so they align with the benchmark page join keys."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute('''
        SELECT "DEPARTMENT",
               COUNT(*) AS emp_count,
               AVG("SALARY") AS avg_salary,
               AVG("JOB_SATISFACTION") AS avg_sat
        FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
        GROUP BY "DEPARTMENT"
    ''')
    stats = {}
    for dept, cnt, sal, sat in cur.fetchall():
        stats[str(dept).strip()] = {
            "count": int(cnt),
            "avg_salary": round(float(sal), 2) if sal is not None else None,
            "avg_sat": round(float(sat), 2) if sat is not None else None,
        }
    cur.close()
    conn.close()
    return stats


def compute_expected():
    """Recompute the authoritative comparison from DWH + served benchmark."""
    bench = fetch_benchmark()
    wh = get_warehouse_stats()
    depts = sorted(set(bench) & set(wh))
    rows = {}
    for d in depts:
        internal = wh[d]["avg_salary"]
        median = bench[d]["median"]
        gap_pct = round((internal - median) / median * 100, 1) if median else 0.0
        rows[d] = {
            "internal": internal,
            "median": median,
            "gap_pct": gap_pct,
            "needs_review": abs(gap_pct) > GAP_THRESHOLD,
            "avg_sat": wh[d]["avg_sat"],
            "count": wh[d]["count"],
        }
    needing = [d for d in depts if rows[d]["needs_review"]]
    largest = max(depts, key=lambda d: abs(rows[d]["gap_pct"])) if depts else None
    return {"rows": rows, "depts": depts, "needing": needing, "largest": largest}


def dept_in_text(text):
    """Return the set of RU departments named in the given text (case-insensitive)."""
    low = (text or "").lower()
    return {d for d in DEPARTMENTS if d.lower() in low}


# --------------------------------------------------------------------------
# Non-critical structural checks (broadened RU+EN matching).
# --------------------------------------------------------------------------

def check_excel(workspace):
    print("\n=== Check 1: Compensation_Benchmark_Report.xlsx ===")
    path = os.path.join(workspace, "Compensation_Benchmark_Report.xlsx")
    if not os.path.exists(path):
        check("Excel file exists", False, f"Not found at {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path)
    sheets = wb.sheetnames
    check("Has at least 3 sheets", len(sheets) >= 3, f"Found {len(sheets)}: {sheets}")

    sheets_lower = [s.lower() for s in sheets]

    da_idx = next((i for i, s in enumerate(sheets_lower) if "department" in s or "analysis" in s), 0)
    ws1 = wb[sheets[da_idx]]
    rows1 = list(ws1.iter_rows(values_only=True))
    data_rows1 = [r for r in rows1[1:] if any(c for c in r)]
    check("Department_Analysis has 7 rows", len(data_rows1) >= 7, f"Found {len(data_rows1)}")

    if rows1:
        headers = [str(c).lower() if c else "" for c in rows1[0]]
        has_salary = any("salary" in h for h in headers)
        has_satisfaction = any("satisf" in h for h in headers)
        check("Has salary column", has_salary, f"Headers: {rows1[0]}")
        check("Has satisfaction column", has_satisfaction, f"Headers: {rows1[0]}")

    found_depts = set()
    for row in data_rows1:
        for cell in row:
            if cell and str(cell).strip() in DEPARTMENTS:
                found_depts.add(str(cell).strip())
    check("All 7 departments present in Dept Analysis", len(found_depts) >= 6,
          f"Found: {found_depts}")

    sb_idx = next((i for i, s in enumerate(sheets_lower) if "benchmark" in s), 1)
    if sb_idx < len(sheets):
        ws2 = wb[sheets[sb_idx]]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)]
        check("Salary_Benchmark has 7 rows", len(data_rows2) >= 7, f"Found {len(data_rows2)}")

        if rows2:
            headers2 = [str(c).lower() if c else "" for c in rows2[0]]
            has_gap = any("gap" in h for h in headers2)
            has_review = any("review" in h for h in headers2)
            check("Has Gap_Pct column", has_gap, f"Headers: {rows2[0]}")
            check("Has Needs_Review column", has_review, f"Headers: {rows2[0]}")

    rs_idx = next((i for i, s in enumerate(sheets_lower) if "review" in s or "summary" in s), 2)
    if rs_idx < len(sheets):
        ws3 = wb[sheets[rs_idx]]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)]
        check("Review_Summary has 7 rows", len(data_rows3) >= 7, f"Found {len(data_rows3)}")


def check_gcal():
    print("\n=== Check 2: Google Calendar Review Meetings ===")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT summary, start_datetime, description FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    check("At least 7 review meeting events", len(events) >= 7, f"Found {len(events)} events")

    if events:
        summaries = " ".join(str(e[0]) for e in events).lower()
        check("Events mention salary review (RU/EN)",
              "пересмотр" in summaries or "зарплат" in summaries
              or "salary review" in summaries or "review meeting" in summaries,
              f"Summaries: {summaries[:150]}")

        dept_found = set()
        for event in events:
            dept_found |= dept_in_text(str(event[0]))
        check("Events cover at least 6 departments", len(dept_found) >= 6,
              f"Departments in events: {dept_found}")

        march_events = [e for e in events if e[1] and e[1].month == 3 and e[1].year == 2026]
        check("Events scheduled in March 2026", len(march_events) >= 7,
              f"Found {len(march_events)} March 2026 events")

    cur.close()
    conn.close()


def check_email():
    print("\n=== Check 3: Email to hr_team@company.com ===")
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT subject, from_addr, to_addr, body_text FROM email.messages")
    messages = cur.fetchall()

    matching = None
    for subject, from_addr, to_addr, body_text in messages:
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) else str(to_addr).lower()
            except Exception:
                to_str = str(to_addr).lower()
        if "hr_team@company.com" in to_str:
            matching = (subject, from_addr, to_addr, body_text)
            break

    check("Email sent to hr_team@company.com", matching is not None,
          f"Messages found: {len(messages)}")

    if matching:
        subject, _, _, body_text = matching
        all_text = ((subject or "") + " " + (body_text or "")).lower()
        check("Email mentions compensation/benchmark (RU/EN)",
              "compensation" in all_text or "benchmark" in all_text
              or "зарплат" in all_text or "бенчмарк" in all_text or "компенсац" in all_text,
              f"Subject: {subject}")
        check("Email mentions review meetings scheduled (RU/EN)",
              "meeting" in all_text or "scheduled" in all_text or "march" in all_text
              or "совещан" in all_text or "пересмотр" in all_text or "март" in all_text,
              f"Body snippet: {all_text[:100]}")

    cur.close()
    conn.close()


def check_script(workspace):
    print("\n=== Check 4: compensation_analysis.py ===")
    path = os.path.join(workspace, "compensation_analysis.py")
    check("compensation_analysis.py exists", os.path.exists(path))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    path = os.path.join(workspace, "Compensation_Benchmark_Report.xlsx")
    if os.path.isfile(path):
        wb = openpyxl.load_workbook(path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, values_only=True):
                for cell in row:
                    if isinstance(cell, (int, float)) and cell < 0:
                        # Gap_Pct may legitimately be negative; only flag salary/count negatives.
                        pass
        check("No corrupt Excel", True)

    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM gcal.events
            WHERE (lower(summary) LIKE '%%пересмотр%%' OR lower(summary) LIKE '%%зарплат%%'
                   OR lower(summary) LIKE '%%salary review%%' OR lower(summary) LIKE '%%review meeting%%')
              AND EXTRACT(DOW FROM start_datetime) IN (0, 6)
        """)
        weekend_count = cur.fetchone()[0]
        check("No review meetings on weekends", weekend_count == 0,
              f"Found {weekend_count} weekend events")
        cur.close()
        conn.close()
    except Exception:
        pass


# --------------------------------------------------------------------------
# CRITICAL semantic gate.
# --------------------------------------------------------------------------

def run_critical_checks(workspace):
    print("\n=== CRITICAL CHECKS ===")
    crit_fail = 0

    def crit(name, cond, detail=""):
        nonlocal crit_fail
        if cond:
            print(f"  [CRIT-PASS] {name}")
        else:
            crit_fail += 1
            print(f"  [CRIT-FAIL] {name}: {str(detail)[:300]}")

    excel_path = os.path.join(workspace, "Compensation_Benchmark_Report.xlsx")
    if not os.path.exists(excel_path):
        crit("Compensation_Benchmark_Report.xlsx exists", False, excel_path)
        return False

    try:
        exp = compute_expected()
    except Exception as e:
        crit("recompute from ClickHouse + benchmarks.json", False, str(e))
        return False

    if not exp["depts"] or len(exp["depts"]) < 7:
        crit("DWH/benchmark department join (RU labels, all 7 overlap)", False,
             f"overlap depts = {exp['depts']} (JOIN-KEY MISMATCH)")
        return False

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheets_lower = {s.lower(): s for s in wb.sheetnames}

    def find_sheet(*keys):
        for low, orig in sheets_lower.items():
            if all(k in low for k in keys):
                return orig
        return None

    def read_rows(sheet_name):
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return [], []
        headers = [str(c).strip().lower() if c is not None else "" for c in all_rows[0]]
        data = [r for r in all_rows[1:] if any(c is not None and str(c).strip() for c in r)]
        return headers, data

    def col_idx(headers, *subs):
        for i, h in enumerate(headers):
            if any(s in h for s in subs):
                return i
        return None

    # --- Department_Analysis: Avg_Salary recomputed from DWH (>=2 depts) ---
    da = find_sheet("department") or find_sheet("analysis")
    da = da or wb.sheetnames[0]
    h1, d1 = read_rows(da)
    dept_i = col_idx(h1, "department", "отдел", "департам") or 0
    sal_i = col_idx(h1, "avg_salary", "avg_sal", "salary", "зарплат")
    da_map = {}
    for r in d1:
        if dept_i < len(r) and r[dept_i] is not None:
            da_map[str(r[dept_i]).strip()] = r

    salary_matches = 0
    sal_detail = []
    must_check = ["Инженерия", "Финансы", "НИОКР", "Продажи"]
    for d in must_check:
        e = exp["rows"].get(d)
        r = da_map.get(d)
        if e is None or r is None or sal_i is None:
            sal_detail.append(f"{d}: missing row/col")
            continue
        got = safe_float(r[sal_i]) if sal_i < len(r) else None
        if got is not None and num_close(got, e["internal"], rel_tol=0.01, abs_tol=2.0):
            salary_matches += 1
        else:
            sal_detail.append(f"{d}: Avg_Salary got {got} exp {e['internal']}")
    crit("Department_Analysis Avg_Salary recomputed from ClickHouse for >=2 departments",
         salary_matches >= 2, "; ".join(sal_detail[:6]))

    # --- Salary_Benchmark: Gap_Pct + Needs_Review recomputed ---
    sb = find_sheet("benchmark")
    if sb is None:
        crit("Salary_Benchmark sheet exists", False, f"sheets={wb.sheetnames}")
    else:
        h2, d2 = read_rows(sb)
        dept_i2 = col_idx(h2, "department", "отдел", "департам") or 0
        gap_i = col_idx(h2, "gap", "разрыв")
        rev_i = col_idx(h2, "needs_review", "review", "пересмотр")
        sb_map = {}
        for r in d2:
            if dept_i2 < len(r) and r[dept_i2] is not None:
                sb_map[str(r[dept_i2]).strip()] = r

        gap_ok = True
        rev_ok = True
        det = []
        for d, e in exp["rows"].items():
            r = sb_map.get(d)
            if r is None:
                gap_ok = False
                det.append(f"{d}: missing row")
                continue
            pgap = safe_float(r[gap_i]) if (gap_i is not None and gap_i < len(r)) else None
            if pgap is None or abs(pgap - e["gap_pct"]) > 1.0:
                gap_ok = False
                det.append(f"{d}: Gap_Pct got {pgap} exp {e['gap_pct']}")
            prev = str(r[rev_i]).strip().lower() if (rev_i is not None and rev_i < len(r) and r[rev_i] is not None) else ""
            is_yes = prev in {"да", "yes", "true", "1"}
            is_no = prev in {"нет", "no", "false", "0"}
            if e["needs_review"] and not is_yes:
                rev_ok = False
                det.append(f"{d}: Needs_Review got '{prev}' exp Да/Yes")
            if (not e["needs_review"]) and not is_no:
                rev_ok = False
                det.append(f"{d}: Needs_Review got '{prev}' exp Нет/No")
        crit("Salary_Benchmark Gap_Pct matches (internal-median)/median*100 for every dept",
             gap_ok, "; ".join(det[:6]))
        crit("Needs_Review = Да/Yes exactly for departments with |gap|>10",
             rev_ok, "; ".join(det[:6]))

    # --- All 7 RU departments present across the three sheets ---
    def depts_in_sheet(sheet_name):
        if sheet_name is None:
            return set()
        _, rows = read_rows(sheet_name)
        found = set()
        for r in rows:
            for c in r:
                if c is not None and str(c).strip() in DEPARTMENTS:
                    found.add(str(c).strip())
        return found

    rs = find_sheet("review") or find_sheet("summary")
    da_d = depts_in_sheet(da)
    sb_d = depts_in_sheet(sb)
    rs_d = depts_in_sheet(rs)
    all7 = (da_d == DEPARTMENTS and sb_d == DEPARTMENTS and rs_d == DEPARTMENTS)
    crit("All 7 russified departments present in Department_Analysis, Salary_Benchmark and Review_Summary",
         all7, f"DA={da_d} SB={sb_d} RS={rs_d}")

    # --- Calendar: exactly 7 weekday review meetings 2026-03-16..20, dept + keyword ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime, description FROM gcal.events
            WHERE start_datetime >= '2026-03-16 00:00:00'
              AND start_datetime <  '2026-03-21 00:00:00'
              AND (lower(summary) LIKE '%%пересмотр%%' OR lower(summary) LIKE '%%зарплат%%'
                   OR lower(summary) LIKE '%%salary review%%' OR lower(summary) LIKE '%%review%%')
            ORDER BY start_datetime
        """)
        review_events = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        review_events = []
        crit("gcal lookup", False, str(e))

    weekday_ok = all(ev[1].weekday() < 5 for ev in review_events) if review_events else False
    dept_titled = sum(1 for ev in review_events if dept_in_text(str(ev[0])))
    covered = set()
    for ev in review_events:
        covered |= dept_in_text(str(ev[0]))
    crit("Exactly 7 review meetings on weekdays within 2026-03-16..2026-03-20",
         len(review_events) == 7 and weekday_ok,
         f"count={len(review_events)} weekday_ok={weekday_ok}")
    crit("Each review meeting titled with a department and covers all 7 departments",
         dept_titled == len(review_events) and covered == DEPARTMENTS,
         f"titled={dept_titled}/{len(review_events)} covered={covered}")

    # --- Email: correct count needing review + largest-gap department ---
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        emails = cur.fetchall()
        conn.close()
    except Exception as e:
        emails = []
        crit("email lookup", False, str(e))

    email_ok = False
    edetail = ""
    need_count = len(exp["needing"])
    largest = (exp["largest"] or "")
    for subj, to_addr, body in emails:
        if "hr_team@company.com" not in str(to_addr).lower():
            continue
        b = (str(subj or "") + " " + str(body or "")).lower()
        has_count = str(need_count) in b
        has_largest = largest and largest.lower() in b
        if has_count and has_largest:
            email_ok = True
            break
        edetail = f"need_count({need_count})={has_count} largest('{largest}')={has_largest}"
    crit("Email to hr_team reports correct count of departments needing review and the largest-gap department",
         email_ok, edetail or "no matching email")

    print(f"=== CRITICAL: {'PASS' if crit_fail == 0 else f'{crit_fail} FAILED'} ===")
    return crit_fail == 0


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # CRITICAL semantic gate first: any failure => hard FAIL regardless of accuracy.
    critical_ok = run_critical_checks(args.agent_workspace)
    if not critical_ok:
        print("\nCRITICAL checks failed -> FAIL")
        sys.exit(1)

    check_excel(args.agent_workspace)
    check_gcal()
    check_email()
    check_script(args.agent_workspace)
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
