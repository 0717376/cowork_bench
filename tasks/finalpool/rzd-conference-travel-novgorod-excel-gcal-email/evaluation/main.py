"""
Evaluation for train-conference-travel-excel-gcal-email task (RZD / RU).

Checks:
1. Conference_Travel_Plan.xlsx exists with Outbound, Return, Summary sheets.
2. Outbound: 3 rows; Train_Code cells contain only valid rzd economy trains
   (Москва→Новгород: 818А/820А, СПб→Новгород: 822А/824А).
3. Return: 3 rows; valid return trains (Новгород→Москва: 819А/821А,
   Новгород→СПб: 823А/825А).
4. Summary: Total_Travel_Cost_RUB == 13000 exactly; Avg_Cost_Per_Person == 4333.33.
5. GCal: 3 outbound events (2026-03-12) + 3 return events (2026-03-15).
6. Email with RU subject delivered to each of the 3 attendee addresses.

CRITICAL_CHECKS gate: any critical failure => sys.exit(1) before accuracy gate.
"""
import json
import os
import sys
import unicodedata
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Valid economy (Эконом) trains per route/date in the rzd seed.
OUTBOUND_MOW = {"818А", "820А"}   # Москва → Великий Новгород, 2026-03-12
OUTBOUND_SPB = {"822А", "824А"}   # Санкт-Петербург → Великий Новгород, 2026-03-12
RETURN_MOW = {"819А", "821А"}     # Великий Новгород → Москва, 2026-03-15
RETURN_SPB = {"823А", "825А"}     # Великий Новгород → Санкт-Петербург, 2026-03-15


def normalize(s):
    """NFKD + cyrillic->latin translit. ONLY for mixed cyr/lat train-code matching."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    table = str.maketrans({"А": "A", "а": "a", "В": "B", "Е": "E", "е": "e",
                           "К": "K", "М": "M", "Н": "H", "О": "O", "о": "o",
                           "Р": "P", "С": "C", "с": "c", "Т": "T", "Х": "X"})
    return s.translate(table).strip().upper()


VALID_OUT = {normalize(x) for x in (OUTBOUND_MOW | OUTBOUND_SPB)}
VALID_RET = {normalize(x) for x in (RETURN_MOW | RETURN_SPB)}


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")


def _find_train_codes(rows, valid_set):
    """Return list of normalized train codes found in any cell that match valid_set."""
    found = []
    invalid = []
    for r in rows:
        for c in r:
            if c is None:
                continue
            nc = normalize(c)
            # a cell may be exactly a train code like '818А'
            for code in (VALID_OUT | VALID_RET):
                if code in nc and len(nc) <= 8:
                    if code in valid_set:
                        found.append(code)
                    else:
                        invalid.append(code)
    return found, invalid


def check_excel(agent_workspace):
    print("\n=== Check 1: Conference_Travel_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Conference_Travel_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Conference_Travel_Plan.xlsx exists", False, f"Not found at {xlsx_path}",
               critical=True)
        return
    record("Conference_Travel_Plan.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e), critical=True)
        return
    record("Excel file readable", True)

    sheet_names_lower = [s.lower() for s in wb.sheetnames]

    # ---- Outbound ----
    if "outbound" not in sheet_names_lower:
        record("Outbound sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Outbound sheet exists", True)
        ws = wb[wb.sheetnames[sheet_names_lower.index("outbound")]]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)]
        record("Outbound has 3 attendee rows", len(data_rows) == 3, f"Found {len(data_rows)}")

        found, invalid = _find_train_codes(data_rows, VALID_OUT)
        n_mow = sum(1 for c in found if c in {normalize(x) for x in OUTBOUND_MOW})
        n_spb = sum(1 for c in found if c in {normalize(x) for x in OUTBOUND_SPB})
        ok_routes = (n_mow == 2 and n_spb == 1 and not invalid)
        record("Outbound: 2 Москва→Новгород (818А/820А) + 1 СПб→Новгород (822А/824А), no invalid",
               ok_routes, f"found={found} invalid={invalid} mow={n_mow} spb={n_spb}",
               critical=True)

    # ---- Return ----
    if "return" not in sheet_names_lower:
        record("Return sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Return sheet exists", True)
        ws_r = wb[wb.sheetnames[sheet_names_lower.index("return")]]
        rows_r = list(ws_r.iter_rows(values_only=True))
        data_rows_r = [r for r in rows_r[1:] if any(c for c in r)]
        record("Return has 3 attendee rows", len(data_rows_r) == 3, f"Found {len(data_rows_r)}")

        found_r, invalid_r = _find_train_codes(data_rows_r, VALID_RET)
        n_mow_r = sum(1 for c in found_r if c in {normalize(x) for x in RETURN_MOW})
        n_spb_r = sum(1 for c in found_r if c in {normalize(x) for x in RETURN_SPB})
        ok_routes_r = (n_mow_r == 2 and n_spb_r == 1 and not invalid_r)
        record("Return: 2 Новгород→Москва (819А/821А) + 1 Новгород→СПб (823А/825А), no invalid",
               ok_routes_r, f"found={found_r} invalid={invalid_r} mow={n_mow_r} spb={n_spb_r}",
               critical=True)

    # ---- Summary ----
    if "summary" not in sheet_names_lower:
        record("Summary sheet exists", False, f"Sheets: {wb.sheetnames}", critical=True)
    else:
        record("Summary sheet exists", True)
        ws_s = wb[wb.sheetnames[sheet_names_lower.index("summary")]]
        rows_s = list(ws_s.iter_rows(values_only=True))

        numeric_vals = []
        for r in rows_s:
            for c in r:
                try:
                    numeric_vals.append(round(float(c), 2))
                except (TypeError, ValueError):
                    pass
        # Deterministic: 2 Москва × (2500+2500) + 1 СПб × (1500+1500) = 13000
        has_total = any(abs(v - 13000.0) < 0.5 for v in numeric_vals)
        record("Summary Total_Travel_Cost_RUB == 13000", has_total,
               f"Numeric values: {numeric_vals}", critical=True)

        # 13000 / 3 = 4333.33 (rounded 2dp)
        has_avg = any(abs(v - 4333.33) < 0.5 for v in numeric_vals)
        record("Summary Avg_Cost_Per_Person == 4333.33", has_avg,
               f"Numeric values: {numeric_vals}", critical=True)


def check_gcal():
    print("\n=== Check 2: Calendar Events ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime, end_datetime
        FROM gcal.events
        WHERE start_datetime >= '2026-03-12' AND start_datetime < '2026-03-16'
        ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    def is_travel(title):
        if not title:
            return False
        t = title.lower()
        return any(kw in t for kw in [
            "conference", "travel", "return", "qufu",
            "новгород", "поездк", "конференц", "возвращ",
        ])

    travel_events = [e for e in events if is_travel(e[0])]
    record("At least 6 travel calendar events created", len(travel_events) >= 6,
           f"Found {len(travel_events)}. All: {[e[0] for e in events]}")

    outbound_events = [e for e in travel_events if "2026-03-12" in str(e[1])]
    return_events = [e for e in travel_events if "2026-03-15" in str(e[1])]
    record("3 outbound events on 2026-03-12", len(outbound_events) >= 3,
           f"Found {len(outbound_events)}", critical=True)
    record("3 return events on 2026-03-15", len(return_events) >= 3,
           f"Found {len(return_events)}", critical=True)


def check_emails():
    print("\n=== Check 3: Emails to Attendees ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    attendee_emails = [
        "zhang.wei@uni.edu",
        "liu.mei@institute.org",
        "wang.fang@college.cn",
    ]

    cur.execute("""
        SELECT to_addr, subject FROM email.messages
        WHERE subject ILIKE '%поездк%' OR subject ILIKE '%конференц%'
           OR subject ILIKE '%conference travel%' OR subject ILIKE '%travel plan%'
    """)
    messages = cur.fetchall()
    cur.close()
    conn.close()

    all_msgs = list(messages)
    record("At least 3 travel plan emails sent", len(all_msgs) >= 3,
           f"Found {len(all_msgs)} matching emails")

    all_recipients = []
    for row in all_msgs:
        to_raw = row[0]
        if isinstance(to_raw, list):
            all_recipients.extend([str(r).lower() for r in to_raw])
        elif isinstance(to_raw, str):
            all_recipients.append(to_raw.lower())

    all_recipients_str = " ".join(all_recipients)
    covered = all(em.lower() in all_recipients_str for em in attendee_emails)
    for em in attendee_emails:
        record(f"Email sent to {em}",
               em.lower() in all_recipients_str,
               f"Recipients: {all_recipients_str[:200]}")
    record("Travel-plan email covers all 3 attendee addresses", covered,
           f"Recipients: {all_recipients_str[:200]}", critical=True)


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": CRITICAL_FAILED,
    }

    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILED}")
        sys.exit(1)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
