"""Evaluation для yf-sector-rotation-excel-notion-email (RU: moex + teamly).

Проверки:
1. Excel Sector_Rotation.xlsx — 3 листа, структура и значения совпадают с
   groundtruth (regenerated из moex.* через build_groundtruth.py).
2. Teamly: пространство «Sector Research» с 5 страницами по акциям, каждая со
   своим сектором / сигналом / композитным импульсом.
3. Два письма с точными темами и получателями.

CRITICAL_CHECKS: любой провал => общий FAIL независимо от accuracy.
Иначе PASS требует accuracy >= 70%.
"""
import argparse
import os
import sys
import psycopg2
import openpyxl

# Семантический эталон считаем из того же сида moex.*, что и groundtruth.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_groundtruth import compute, SYMS, SECTOR

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

SYMBOLS = sorted(SYMS)  # ["GAZP.ME","LKOH.ME","MGNT.ME","MTSS.ME","SBER.ME"]

# Tolerances
PRICE_TOL = 0.5
RETURN_TOL = 0.5
RS_TOL = 0.05

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Семантические критические проверки: их провал = общий FAIL.
CRITICAL_CHECKS = {
    "Sector_Rotation.xlsx exists with 5 stock rows",
    "Composite_Momentum matches 0.2*R1M+0.3*R3M+0.5*R6M from moex data",
    "Signal per stock follows composite-vs-benchmark deadband (+/-2)",
    "Portfolio_Signal follows majority rule",
    "Teamly 'Sector Research' has 5 stock pages with sector/signal/momentum",
    "Two emails with exact subjects to investment_team and trading_desk",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def num_close(a, b, tol):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    try:
        return str(a).strip().lower() == str(b).strip().lower()
    except (TypeError, AttributeError):
        return False


def get_conn():
    return psycopg2.connect(**DB)


# --------------------------------------------------------------------------- #
# Excel
# --------------------------------------------------------------------------- #
def check_excel(agent_ws, gt_ws):
    agent_path = os.path.join(agent_ws, "Sector_Rotation.xlsx")
    gt_path = os.path.join(gt_ws, "Sector_Rotation.xlsx")

    if not os.path.exists(agent_path):
        check("Sector_Rotation.xlsx exists with 5 stock rows", False, "agent file missing")
        return
    if not os.path.exists(gt_path):
        check("Sector_Rotation.xlsx exists with 5 stock rows", False, "groundtruth file missing")
        return

    try:
        wb_agent = openpyxl.load_workbook(agent_path, data_only=True)
        wb_gt = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        check("Sector_Rotation.xlsx exists with 5 stock rows", False, f"load error: {e}")
        return

    # Семантический эталон, пересчитанный из moex.* (не из стейл-данных).
    R, B = compute()

    # --- Sheet 1: Momentum Analysis ---
    agent_rows = load_sheet_rows(wb_agent, "Momentum Analysis")
    gt_rows = load_sheet_rows(wb_gt, "Momentum Analysis")

    agent_by_sym = {}
    if agent_rows is None:
        check("Sector_Rotation.xlsx exists with 5 stock rows", False, "Momentum Analysis sheet missing")
    else:
        agent_data = [r for r in agent_rows[1:] if r and r[0] is not None]
        agent_by_sym = {str(r[0]).strip().upper(): r for r in agent_data}
        check("Sector_Rotation.xlsx exists with 5 stock rows",
              len(agent_data) >= 5, f"{len(agent_data)} rows")
        gt_data = [r for r in gt_rows[1:] if r and r[0] is not None] if gt_rows else []
        gt_by_sym = {str(r[0]).strip().upper(): r for r in gt_data}

        # Sorted-by-Symbol (non-critical, structural)
        syms_in_order = [str(r[0]).strip().upper() for r in agent_data[:5]]
        check("Momentum Analysis sorted by Symbol ascending",
              syms_in_order == sorted(syms_in_order), syms_in_order)

        # ---- Per-stock numeric/structural checks vs groundtruth (non-critical) ----
        for sym in SYMBOLS:
            a = agent_by_sym.get(sym)
            g = gt_by_sym.get(sym)
            if a is None:
                check(f"{sym} present in Momentum Analysis", False, "row missing")
                continue
            check(f"{sym} present in Momentum Analysis", True)
            if g is None:
                continue
            check(f"{sym} Latest_Price", num_close(a[3], g[3], PRICE_TOL), f"{a[3]} != {g[3]}")
            for idx, nm in [(4, "Return_1M_Pct"), (5, "Return_3M_Pct"), (6, "Return_6M_Pct")]:
                check(f"{sym} {nm}", num_close(a[idx], g[idx], RETURN_TOL), f"{a[idx]} != {g[idx]}")
            check(f"{sym} Benchmark_Momentum", num_close(a[8], g[8], RETURN_TOL), f"{a[8]} != {g[8]}")

        # ---- CRITICAL: Composite formula recomputed from moex prices ----
        comp_ok = True
        comp_detail = ""
        for sym in SYMBOLS:
            a = agent_by_sym.get(sym)
            if a is None or a[7] is None:
                comp_ok = False
                comp_detail = f"{sym} composite missing"
                break
            expected_comp = R[sym]["comp"]
            if not num_close(a[7], expected_comp, RETURN_TOL):
                comp_ok = False
                comp_detail = f"{sym}: {a[7]} != {round(expected_comp,2)}"
                break
        check("Composite_Momentum matches 0.2*R1M+0.3*R3M+0.5*R6M from moex data",
              comp_ok, comp_detail)

        # ---- CRITICAL: Signal follows composite-vs-benchmark deadband ----
        sig_ok = True
        sig_detail = ""
        for sym in SYMBOLS:
            a = agent_by_sym.get(sym)
            if a is None:
                sig_ok = False
                sig_detail = f"{sym} missing"
                break
            expected_sig = R[sym]["signal"]
            if not str_match(a[9], expected_sig):
                sig_ok = False
                sig_detail = f"{sym}: '{a[9]}' != '{expected_sig}'"
                break
        check("Signal per stock follows composite-vs-benchmark deadband (+/-2)",
              sig_ok, sig_detail)

    # --- Sheet 2: Relative Strength (non-critical) ---
    agent_rs = load_sheet_rows(wb_agent, "Relative Strength")
    gt_rs = load_sheet_rows(wb_gt, "Relative Strength")
    if agent_rs is None or gt_rs is None:
        check("Relative Strength sheet present", agent_rs is not None, "missing")
    else:
        a_rs = {str(r[0]).strip().upper(): r for r in agent_rs[1:] if r and r[0] is not None}
        g_rs = {str(r[0]).strip().upper(): r for r in gt_rs[1:] if r and r[0] is not None}
        check("Relative Strength sheet present", True)
        for sym in SYMBOLS:
            a = a_rs.get(sym)
            g = g_rs.get(sym)
            if a is None or g is None:
                check(f"{sym} in Relative Strength", a is not None, "missing")
                continue
            for idx, nm in [(1, "RS_1M"), (2, "RS_3M"), (3, "RS_6M"), (4, "Avg_RS")]:
                check(f"{sym} {nm}", num_close(a[idx], g[idx], RS_TOL), f"{a[idx]} != {g[idx]}")
            if a[5] is not None and g[5] is not None:
                check(f"{sym} RS_Rank", int(a[5]) == int(g[5]), f"{a[5]} != {g[5]}")

    # --- Sheet 3: Strategy Summary ---
    agent_sum = load_sheet_rows(wb_agent, "Strategy Summary")
    gt_sum = load_sheet_rows(wb_gt, "Strategy Summary")

    def build_summary_dict(rows):
        d = {}
        for r in rows or []:
            if r and r[0] is not None:
                d[str(r[0]).strip()] = r[1]
        return d

    a_sum = build_summary_dict(agent_sum[1:] if agent_sum else [])
    g_sum = build_summary_dict(gt_sum[1:] if gt_sum else [])

    # Counts / top-bottom / avg (non-critical, vs groundtruth)
    for label in ["Overweight_Count", "Neutral_Count", "Underweight_Count"]:
        if label in g_sum:
            av = a_sum.get(label)
            check(f"Summary {label}", av is not None and int(av) == int(g_sum[label]),
                  f"{av} != {g_sum.get(label)}")
    for label in ["Top_Momentum_Stock", "Bottom_Momentum_Stock"]:
        if label in g_sum:
            check(f"Summary {label}", str_match(a_sum.get(label, ""), g_sum[label]),
                  f"{a_sum.get(label)} != {g_sum[label]}")
    if "Avg_Composite_Momentum" in g_sum:
        check("Summary Avg_Composite_Momentum",
              num_close(a_sum.get("Avg_Composite_Momentum", 0), g_sum["Avg_Composite_Momentum"], RETURN_TOL),
              f"{a_sum.get('Avg_Composite_Momentum')} != {g_sum['Avg_Composite_Momentum']}")
    if "Benchmark_Return_6M" in g_sum:
        check("Summary Benchmark_Return_6M",
              num_close(a_sum.get("Benchmark_Return_6M", 0), g_sum["Benchmark_Return_6M"], RETURN_TOL),
              f"{a_sum.get('Benchmark_Return_6M')} != {g_sum['Benchmark_Return_6M']}")

    # ---- CRITICAL: Portfolio_Signal majority rule ----
    ow = sum(1 for s in SYMBOLS if R[s]["signal"] == "Overweight")
    uw = sum(1 for s in SYMBOLS if R[s]["signal"] == "Underweight")
    expected_ps = "Bullish" if ow > 2.5 else ("Bearish" if uw > 2.5 else "Mixed")
    check("Portfolio_Signal follows majority rule",
          str_match(a_sum.get("Portfolio_Signal", ""), expected_ps),
          f"got '{a_sum.get('Portfolio_Signal')}', expected '{expected_ps}'")


# --------------------------------------------------------------------------- #
# Teamly
# --------------------------------------------------------------------------- #
def check_teamly():
    """Раздел 'Sector Research' с 5 страницами по акциям.

    Сид teamly содержит страницы id <= 3 и шумовое пространство OPS; засчитываем
    только страницы агента (id > 3), которые относятся к тикерам.
    """
    R, B = compute()
    try:
        conn = get_conn()
        cur = conn.cursor()
        # Все пользовательские страницы (id>3) с привязкой к пространству.
        cur.execute("""
            SELECT p.id, p.title, COALESCE(p.body,''), COALESCE(s.name,''), COALESCE(s.key,'')
            FROM teamly.pages p
            LEFT JOIN teamly.spaces s ON s.id = p.space_id
            WHERE p.id > 3
        """)
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Sector Research' has 5 stock pages with sector/signal/momentum", False, str(e))
        return

    # Кандидаты: исключаем шумовое пространство OPS.
    candidate = [p for p in pages if p[4].upper() != "OPS"]

    matched = 0
    details = []
    for sym in SYMBOLS:
        ticker = sym  # e.g. SBER.ME
        base = sym.split(".")[0]  # SBER
        page = None
        for pid, title, body, sname, skey in candidate:
            t = (title or "").upper()
            if ticker.upper() in t or base in t:
                page = (title, body)
                break
        if page is None:
            details.append(f"{sym}: no page")
            continue
        text = ((page[0] or "") + " " + (page[1] or "")).lower()
        sector_ok = SECTOR[sym].lower() in text
        signal_ok = R[sym]["signal"].lower() in text
        # Композит как число с точностью до целого/одного знака.
        comp = R[sym]["comp"]
        mom_ok = (f"{comp:.2f}" in text) or (f"{comp:.1f}" in text) or (f"{round(comp):d}" in text) \
            or (f"{abs(comp):.2f}" in text) or (f"{abs(comp):.1f}" in text)
        if sector_ok and signal_ok and mom_ok:
            matched += 1
        else:
            details.append(f"{sym}: sector={sector_ok} signal={signal_ok} momentum={mom_ok}")

    check("Teamly 'Sector Research' has 5 stock pages with sector/signal/momentum",
          matched >= 5, "; ".join(details) or f"matched={matched}")

    # Структурная (non-critical): существует пространство/страница с маркером Sector Research.
    sr_marker = any("sector research" in (p[1] or "").lower()
                    or "sector research" in (p[3] or "").lower()
                    or "секторн" in (p[1] or "").lower()
                    for p in candidate)
    check("Teamly hub named 'Sector Research' present", sr_marker or matched >= 5,
          "no 'Sector Research' marker")


# --------------------------------------------------------------------------- #
# Emails
# --------------------------------------------------------------------------- #
def check_emails():
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("""
            SELECT subject FROM email.messages
            WHERE to_addr::text ILIKE '%investment_team@firm.com%'
        """)
        inv = [r[0] for r in cur.fetchall()]
        cur.execute("""
            SELECT subject FROM email.messages
            WHERE to_addr::text ILIKE '%trading_desk@firm.com%'
        """)
        trd = [r[0] for r in cur.fetchall()]
        conn.close()
    except Exception as e:
        check("Two emails with exact subjects to investment_team and trading_desk", False, str(e))
        return

    inv_ok = any(str_match(s, "Sector Rotation Analysis Update") for s in inv)
    trd_ok = any(str_match(s, "Actionable Sector Signals") for s in trd)
    check("Two emails with exact subjects to investment_team and trading_desk",
          inv_ok and trd_ok, f"inv={inv}, trd={trd}")

    # Non-critical: recipients exist at all.
    check("Email to investment_team@firm.com sent", bool(inv), "none")
    check("Email to trading_desk@firm.com sent", bool(trd), "none")


# --------------------------------------------------------------------------- #
def run_evaluation(agent_ws, gt_ws, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    print("  Checking Excel...")
    check_excel(agent_ws, gt_ws)
    print("  Checking Teamly...")
    check_teamly()
    print("  Checking emails...")
    check_emails()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            import json
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    # Критический гейт ДО порога точности.
    if critical_failed:
        print("\n=== RESULT: FAIL (critical) ===")
        return False
    success = accuracy >= 70
    print(f"\n=== RESULT: {'PASS' if success else 'FAIL'} ===")
    return success


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    agent_ws = args.agent_workspace or os.path.join(task_root, "groundtruth_workspace")
    gt_ws = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    success = run_evaluation(agent_ws, gt_ws, args.launch_time, args.res_log_file)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
