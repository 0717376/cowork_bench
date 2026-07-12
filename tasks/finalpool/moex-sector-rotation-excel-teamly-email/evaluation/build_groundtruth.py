"""Регенерация groundtruth_workspace/Sector_Rotation.xlsx из сида moex.stock_prices.

Источник цен — db/zzz_moex_after_init.sql (глобальный read-only сид moex.*).
Методология полностью соответствует initial_workspace/guide.md и Rotation_Strategy.pdf:
  - лукбэки 30/60/90 календарных дней (последняя торговая цена на дату или раньше);
  - Composite = 0.2*R1M + 0.3*R3M + 0.5*R6M;
  - бенчмарк = равновзвешенное среднее доходностей 5 акций;
  - сигнал Overweight/Underweight/Neutral по дедбенду +/-2 п.п.

Запуск:
  uv run --with openpyxl python3 evaluation/build_groundtruth.py
"""
import os
import re
from collections import defaultdict
from datetime import date, timedelta

import openpyxl

TASK_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROJECT_ROOT = os.path.abspath(os.path.join(TASK_ROOT, "..", "..", ".."))
SEED = os.path.join(PROJECT_ROOT, "db", "zzz_moex_after_init.sql")

SYMS = ["GAZP.ME", "LKOH.ME", "MGNT.ME", "MTSS.ME", "SBER.ME"]
SECTOR = {
    "SBER.ME": "Financial Services",
    "GAZP.ME": "Energy",
    "LKOH.ME": "Energy",
    "MGNT.ME": "Consumer Defensive",
    "MTSS.ME": "Communication Services",
}
NAME = {
    "SBER.ME": "ПАО Сбербанк",
    "GAZP.ME": "ПАО Газпром",
    "LKOH.ME": "ПАО ЛУКОЙЛ",
    "MGNT.ME": "ПАО Магнит",
    "MTSS.ME": "ПАО МТС",
}
LOOKBACKS = {"r1": 30, "r3": 60, "r6": 90}


def _load_prices_from_db():
    """Авторитетный источник — живая БД moex.* (то же, что читает агент через MCP)."""
    import os as _os
    import psycopg2
    db = dict(host=_os.environ.get("PGHOST", "localhost"), port=5432,
              dbname="cowork_gym", user="eigent", password="camel")
    conn = psycopg2.connect(**db)
    cur = conn.cursor()
    cur.execute(
        "SELECT symbol, date::text, close FROM moex.stock_prices "
        "WHERE symbol = ANY(%s) ORDER BY symbol, date", (SYMS,))
    d = defaultdict(dict)
    for sym, dt, close in cur.fetchall():
        d[sym][dt] = float(close)
    cur.close()
    conn.close()
    if not all(d.get(s) for s in SYMS):
        raise RuntimeError("moex.stock_prices incomplete")
    return d


def _load_prices_from_seed():
    txt = open(SEED, encoding="utf-8").read()
    pat = re.compile(
        r"stock_prices \(symbol, date, open, high, low, close, volume\) VALUES "
        r"\('([^']+)', '([0-9-]+)', [0-9.]+, [0-9.]+, [0-9.]+, ([0-9.]+), \d+\);"
    )
    d = defaultdict(dict)
    for sym, dt, close in pat.findall(txt):
        d[sym][dt] = float(close)
    return d


def load_prices():
    """Сначала живая БД (как у агента), иначе разбор глобального сида."""
    try:
        return _load_prices_from_db()
    except Exception:
        return _load_prices_from_seed()


def price_on_or_before(series, target):
    cands = [x for x in series if date.fromisoformat(x) <= target]
    return series[max(cands)] if cands else series[min(series)]


def compute():
    prices = load_prices()
    R = {}
    for s in SYMS:
        series = prices[s]
        latest_dt = max(series)
        latest = series[latest_dt]
        out = {"latest": latest, "latest_dt": latest_dt}
        for lbl, days in LOOKBACKS.items():
            past = price_on_or_before(series, date.fromisoformat(latest_dt) - timedelta(days=days))
            out[lbl] = (latest / past - 1) * 100
        out["comp"] = out["r1"] * 0.2 + out["r3"] * 0.3 + out["r6"] * 0.5
        R[s] = out

    B = {k: sum(R[s][k] for s in SYMS) / len(SYMS) for k in ("r1", "r3", "r6")}
    B["comp"] = B["r1"] * 0.2 + B["r3"] * 0.3 + B["r6"] * 0.5

    for s in SYMS:
        diff = R[s]["comp"] - B["comp"]
        R[s]["signal"] = "Overweight" if diff > 2 else ("Underweight" if diff < -2 else "Neutral")
        R[s]["rs1"] = R[s]["r1"] - B["r1"]
        R[s]["rs3"] = R[s]["r3"] - B["r3"]
        R[s]["rs6"] = R[s]["r6"] - B["r6"]
        R[s]["avg_rs"] = (R[s]["rs1"] + R[s]["rs3"] + R[s]["rs6"]) / 3

    for i, s in enumerate(sorted(SYMS, key=lambda x: -R[x]["avg_rs"])):
        R[s]["rank"] = i + 1

    return R, B


def main():
    R, B = compute()
    analysis_date = R["SBER.ME"]["latest_dt"]
    ow = sum(1 for s in SYMS if R[s]["signal"] == "Overweight")
    nw = sum(1 for s in SYMS if R[s]["signal"] == "Neutral")
    uw = sum(1 for s in SYMS if R[s]["signal"] == "Underweight")
    portfolio_signal = "Bullish" if ow > 2.5 else ("Bearish" if uw > 2.5 else "Mixed")
    top = max(SYMS, key=lambda s: R[s]["comp"])
    bot = min(SYMS, key=lambda s: R[s]["comp"])
    avg_comp = sum(R[s]["comp"] for s in SYMS) / len(SYMS)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Momentum Analysis"
    ws1.append(["Symbol", "Company_Name", "Sector", "Latest_Price", "Return_1M_Pct",
                "Return_3M_Pct", "Return_6M_Pct", "Composite_Momentum",
                "Benchmark_Momentum", "Signal"])
    for s in sorted(SYMS):
        r = R[s]
        ws1.append([s, NAME[s], SECTOR[s], round(r["latest"], 2), round(r["r1"], 2),
                    round(r["r3"], 2), round(r["r6"], 2), round(r["comp"], 2),
                    round(B["comp"], 2), r["signal"]])

    ws2 = wb.create_sheet("Relative Strength")
    ws2.append(["Symbol", "RS_1M", "RS_3M", "RS_6M", "Avg_RS", "RS_Rank"])
    for s in sorted(SYMS):
        r = R[s]
        ws2.append([s, round(r["rs1"], 3), round(r["rs3"], 3), round(r["rs6"], 3),
                    round(r["avg_rs"], 3), r["rank"]])

    ws3 = wb.create_sheet("Strategy Summary")
    ws3.append(["Label", "Value"])
    for label, value in [
        ("Analysis_Date", analysis_date),
        ("Benchmark_Return_6M", round(B["r6"], 2)),
        ("Overweight_Count", ow),
        ("Neutral_Count", nw),
        ("Underweight_Count", uw),
        ("Top_Momentum_Stock", top),
        ("Bottom_Momentum_Stock", bot),
        ("Avg_Composite_Momentum", round(avg_comp, 2)),
        ("Portfolio_Signal", portfolio_signal),
    ]:
        ws3.append([label, value])

    out_path = os.path.join(TASK_ROOT, "groundtruth_workspace", "Sector_Rotation.xlsx")
    wb.save(out_path)
    print(f"[build_groundtruth] wrote {out_path}")
    print(f"  Analysis_Date={analysis_date} OW={ow} N={nw} UW={uw} "
          f"Portfolio_Signal={portfolio_signal} Top={top} Bottom={bot} "
          f"Avg_Composite={round(avg_comp,2)} Benchmark_Composite={round(B['comp'],2)}")


if __name__ == "__main__":
    main()
