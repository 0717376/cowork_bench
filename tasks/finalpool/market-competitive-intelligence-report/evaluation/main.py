#!/usr/bin/env python3
"""Evaluation for market-competitive-intelligence-report.

The agent reads seeded RU competitor source data (competitors.csv), consolidates
it into a ClickHouse table sf_data."MARKET_INTEL__PUBLIC__COMPETITORS", builds an
Excel comparison, drafts a Word report, emails leadership and schedules a Google
Calendar briefing.

CRITICAL_CHECKS reflect task substance: any failure there => overall FAIL
regardless of accuracy. Structural checks (sheet/column/file existence) are
non-critical. Accuracy gate: >= 70 AND no critical fail => PASS.
"""

from argparse import ArgumentParser
import os
import sys
import json
from datetime import datetime

import psycopg2

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    from docx import Document as DocxDocument
except Exception:
    DocxDocument = None

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# Deterministic groundtruth derived from competitors.csv.
# PRICE_PER_FEATURE_RUB = MONTHLY_PRICE_RUB / FEATURE_COUNT.
COMPETITORS = {
    "ОблакоПро":  {"price": 24000, "features": 40, "ppf": 24000 / 40},
    "ВегаСофт":   {"price": 8900,  "features": 22, "ppf": 8900 / 22},
    "ТехноЛайн":  {"price": 18500, "features": 35, "ppf": 18500 / 35},
    "СтартПлан":  {"price": 4500,  "features": 14, "ppf": 4500 / 14},
    "ГранитДата": {"price": 31000, "features": 48, "ppf": 31000 / 48},
}
CHEAPEST_PER_FEATURE = "СтартПлан"   # min ppf = 321.43
MOST_FEATURES = "ГранитДата"         # 48 features
AVG_PPF = sum(c["ppf"] for c in COMPETITORS.values()) / len(COMPETITORS)

AGENT_TABLE = 'sf_data."MARKET_INTEL__PUBLIC__COMPETITORS"'

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "ClickHouse table has all 5 competitors",
    "PRICE_PER_FEATURE_RUB values correct in ClickHouse",
    "Excel Cheapest_Per_Feature correct",
    "Word report has all required RU headings",
    "Email to leadership with [Competitive Intelligence] marker",
    "Future Google Calendar briefing event",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def num_close(a, b, rel_tol=0.02, abs_tol=1.0):
    try:
        return abs(float(a) - float(b)) <= max(abs_tol, abs(float(b)) * rel_tol)
    except (TypeError, ValueError):
        return False


def safe_float(v):
    try:
        if v is None:
            return None
        return float(str(v).replace(",", "").replace("₽", "").replace("руб", "").strip())
    except (ValueError, TypeError):
        return None


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


# --- Check 1: ClickHouse consolidation table ---
def check_clickhouse():
    print("\n=== Check 1: ClickHouse sf_data consolidation ===")
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute(f"SELECT COUNT(*) FROM {AGENT_TABLE}")
        cnt = cur.fetchone()[0]
        record("ClickHouse table has all 5 competitors", cnt >= 5, f"got {cnt} rows")

        cur.execute(
            f'SELECT "COMPETITOR", "MONTHLY_PRICE_RUB", "FEATURE_COUNT", '
            f'"PRICE_PER_FEATURE_RUB" FROM {AGENT_TABLE}'
        )
        rows = cur.fetchall()
        by_name = {str(r[0]).strip(): r for r in rows}

        # column presence already implied by successful select
        record("ClickHouse has PRICE_PER_FEATURE_RUB column", True)

        ppf_ok = True
        for name, gt in COMPETITORS.items():
            r = by_name.get(name)
            if r is None:
                ppf_ok = False
                continue
            if not num_close(r[3], gt["ppf"], rel_tol=0.02, abs_tol=1.0):
                ppf_ok = False
        record("PRICE_PER_FEATURE_RUB values correct in ClickHouse", ppf_ok,
               f"expected ppf e.g. {CHEAPEST_PER_FEATURE}={COMPETITORS[CHEAPEST_PER_FEATURE]['ppf']:.2f}")
    except Exception as e:
        record("ClickHouse table has all 5 competitors", False, str(e))
        record("ClickHouse has PRICE_PER_FEATURE_RUB column", False, str(e))
        record("PRICE_PER_FEATURE_RUB values correct in ClickHouse", False, str(e))
    finally:
        cur.close()
        conn.close()


# --- Check 2: Excel comparison ---
def load_sheet(wb, target):
    for nm in wb.sheetnames:
        if nm.strip().lower() == target.strip().lower():
            return wb[nm]
    return None


def check_excel(agent_ws):
    print("\n=== Check 2: Excel Competitive_Analysis.xlsx ===")
    path = os.path.join(agent_ws, "Competitive_Analysis.xlsx")
    if not os.path.exists(path):
        record("Competitive_Analysis.xlsx exists", False, path)
        return
    record("Competitive_Analysis.xlsx exists", True)
    if openpyxl is None:
        record("Excel readable", False, "openpyxl not installed")
        return
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        return

    comp = load_sheet(wb, "Competitors")
    record("Competitors sheet exists", comp is not None)
    if comp is not None:
        rows = list(comp.iter_rows(values_only=True))
        data = rows[1:] if rows else []
        record("Competitors has >= 5 data rows", len(data) >= 5, f"got {len(data)}")
        headers = [str(c).strip().lower() if c else "" for c in (rows[0] if rows else [])]
        for col in ["COMPETITOR", "MONTHLY_PRICE_RUB", "FEATURE_COUNT", "PRICE_PER_FEATURE_RUB"]:
            record(f"Competitors has {col} column", col.lower() in headers, f"headers: {headers[:8]}")

    pc = load_sheet(wb, "Pricing_Comparison")
    record("Pricing_Comparison sheet exists", pc is not None)
    if pc is not None:
        kv = {}
        for row in pc.iter_rows(values_only=True):
            if row and row[0] is not None and len(row) >= 2:
                kv[str(row[0]).strip().lower()] = row[1]
        cheapest = kv.get("cheapest_per_feature")
        record("Excel Cheapest_Per_Feature correct",
               cheapest is not None and CHEAPEST_PER_FEATURE.lower() in str(cheapest).strip().lower(),
               f"got {cheapest}, expected {CHEAPEST_PER_FEATURE}")
        most = kv.get("most_features")
        record("Excel Most_Features correct",
               most is not None and MOST_FEATURES.lower() in str(most).strip().lower(),
               f"got {most}, expected {MOST_FEATURES}")
        avg = safe_float(kv.get("avg_price_per_feature"))
        record("Excel Avg_Price_Per_Feature correct",
               avg is not None and num_close(avg, AVG_PPF, rel_tol=0.03, abs_tol=2.0),
               f"got {avg}, expected {AVG_PPF:.2f}")


# --- Check 3: Word report ---
def check_word(agent_ws):
    print("\n=== Check 3: Word Competitive_Report.docx ===")
    path = os.path.join(agent_ws, "Competitive_Report.docx")
    if not os.path.exists(path):
        record("Competitive_Report.docx exists", False, path)
        return
    record("Competitive_Report.docx exists", True)
    if DocxDocument is None:
        record("Word readable", False, "python-docx not installed")
        return
    try:
        doc = DocxDocument(path)
    except Exception as e:
        record("Word readable", False, str(e))
        return

    text = "\n".join(p.text for p in doc.paragraphs)
    low = text.lower()

    # Required headings accepted in RU or EN.
    required = [
        ["обзор конкурентной среды", "competitive landscape"],
        ["анализ ценообразования", "pricing analysis"],
        ["сравнение функций", "feature comparison"],
        ["рекомендации", "recommendations"],
    ]
    all_headings = all(any(v in low for v in variants) for variants in required)
    record("Word report has all required RU headings", all_headings,
           f"text head: {low[:120]}")

    has_rub = any(k in low for k in ["руб", "₽", "rub"])
    record("Word report mentions RUB currency", has_rub)

    names_found = sum(1 for n in COMPETITORS if n.lower() in low)
    record("Word report names at least one competitor", names_found >= 1,
           f"found {names_found} names")


# --- Check 4: Email ---
def check_email():
    print("\n=== Check 4: Email to leadership ===")
    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        msgs = cur.fetchall()
    except Exception as e:
        record("Email to leadership with [Competitive Intelligence] marker", False, str(e))
        cur.close()
        conn.close()
        return
    cur.close()
    conn.close()

    match = None
    for subject, to_addr, body in msgs:
        subj = subject or ""
        to_str = ""
        if isinstance(to_addr, list):
            to_str = " ".join(str(r) for r in to_addr)
        else:
            try:
                parsed = json.loads(to_addr) if to_addr else None
                to_str = " ".join(str(r) for r in parsed) if isinstance(parsed, list) else str(to_addr or "")
            except (json.JSONDecodeError, TypeError):
                to_str = str(to_addr or "")
        if "[competitive intelligence]" in subj.lower() and "leadership@company.com" in to_str.lower():
            match = (subject, to_str, body or "")
            break

    record("Email to leadership with [Competitive Intelligence] marker", match is not None,
           "No email to leadership@company.com with [Competitive Intelligence] subject marker")
    if match:
        body_low = (match[2] or "").lower()
        record("Email body mentions cheapest competitor",
               CHEAPEST_PER_FEATURE.lower() in body_low,
               f"expected {CHEAPEST_PER_FEATURE} in body")


# --- Check 5: Google Calendar ---
def check_calendar(launch_time):
    print("\n=== Check 5: Google Calendar briefing ===")
    try:
        now = datetime.strptime(launch_time, "%Y-%m-%d %H:%M:%S") if launch_time else datetime.now()
    except (ValueError, TypeError):
        now = datetime.now()

    conn = get_conn()
    cur = conn.cursor()
    try:
        cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
    except Exception as e:
        record("Future Google Calendar briefing event", False, str(e))
        cur.close()
        conn.close()
        return
    cur.close()
    conn.close()

    kws = ["конкурент", "competit", "брифинг", "briefing", "разведк", "intelligence"]
    found = None
    for summary, description, start_dt in events:
        text = ((summary or "") + " " + (description or "")).lower()
        if any(k in text for k in kws):
            future = start_dt is not None and start_dt.replace(tzinfo=None) > now.replace(tzinfo=None)
            if future:
                found = (summary, start_dt)
                break
    record("Future Google Calendar briefing event", found is not None,
           f"checked {len(events)} events, launch_time={launch_time}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    agent_ws = args.agent_workspace or "."

    check_clickhouse()
    check_excel(agent_ws)
    check_word(agent_ws)
    check_email()
    check_calendar(args.launch_time)

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks performed.")
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\nOverall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT,
                    "total_checks": total,
                    "accuracy": accuracy,
                    "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    if critical_failed:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
