"""
Evaluation script for insales-refund-analysis-notion (InSales store + Teamly).

PASS requires accuracy >= 70% AND no CRITICAL check failed.
Any critical failure => sys.exit(1) regardless of accuracy.

All expected values (refund count, total/avg amount, refund rate, the
most-common refund reason) are computed from the live DB at eval time —
nothing about the data values is hardcoded, so RU/EN realia stay in sync.
"""

import argparse
import json
import os
import sys
from collections import Counter

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Refund Details: per-Refund_ID Amount matches DB",
    "Summary: Total_Refunds / Total_Refund_Amount / Avg_Refund_Amount / Refund_Rate match DB",
    "Teamly 'Refund Analysis Dashboard' page reflects count, amount and most-common reason",
    "Email to cfo@company.com body contains all key figures and the most-common reason",
}


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    return psycopg2.connect(**DB_CONFIG)


def get_expected():
    """Compute all expected aggregates from the live DB."""
    conn = get_conn()
    cur = conn.cursor()
    cur.execute("SELECT id, order_id, amount, reason, date_created FROM wc.refunds ORDER BY id")
    refunds = cur.fetchall()
    cur.execute("SELECT COUNT(*) FROM wc.orders")
    order_count = cur.fetchone()[0]
    cur.close()
    conn.close()

    count = len(refunds)
    total = round(sum(float(r[2]) for r in refunds), 2)
    avg = round(total / count, 2) if count else 0.0
    rate = round(count / order_count * 100, 2) if order_count else 0.0
    reasons = [(r[3] or "").strip() for r in refunds if r[3]]
    mode_reason = Counter(reasons).most_common(1)[0][0] if reasons else ""
    # id -> amount map for per-row verification.
    amount_by_id = {int(r[0]): round(float(r[2]), 2) for r in refunds}
    return {
        "refunds": refunds,
        "order_count": order_count,
        "count": count,
        "total": total,
        "avg": avg,
        "rate": rate,
        "mode_reason": mode_reason,
        "amount_by_id": amount_by_id,
    }


def check_excel(agent_workspace, exp):
    print("\n=== Checking Excel Output ===")
    excel_path = os.path.join(agent_workspace, "Refund_Analysis.xlsx")
    check("Excel file exists", os.path.isfile(excel_path), f"Expected {excel_path}")
    if not os.path.isfile(excel_path):
        check("Refund Details: per-Refund_ID Amount matches DB", False, "no excel")
        check("Summary: Total_Refunds / Total_Refund_Amount / Avg_Refund_Amount / Refund_Rate match DB",
              False, "no excel")
        return

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e))
        check("Refund Details: per-Refund_ID Amount matches DB", False, "unreadable")
        check("Summary: Total_Refunds / Total_Refund_Amount / Avg_Refund_Amount / Refund_Rate match DB",
              False, "unreadable")
        return

    # --- Refund Details sheet ---
    ws = None
    for s in wb.sheetnames:
        if "refund" in s.lower() and "detail" in s.lower():
            ws = wb[s]
            break
    if ws is None:
        for s in wb.sheetnames:
            if "refund" in s.lower():
                ws = wb[s]
                break
    check("Sheet with refund details exists", ws is not None, f"Sheets: {wb.sheetnames}")

    if ws is not None:
        header = [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]
        hmap = {h: i for i, h in enumerate(header)}
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if any(c is not None for c in r)]
        check(f"Refund count matches DB ({exp['count']})",
              len(rows) == exp["count"],
              f"Got {len(rows)} rows, expected {exp['count']}")

        # Structural column checks (non-critical).
        for col in ("refund_id", "order_id", "amount", "reason", "date"):
            check(f"Refund Details has {col} column", col in hmap, f"headers: {header}")

        # CRITICAL: per-Refund_ID Amount matches DB exactly (tol 0.01).
        id_i = hmap.get("refund_id")
        amt_i = hmap.get("amount")
        amounts_ok = id_i is not None and amt_i is not None and len(rows) == exp["count"]
        if amounts_ok:
            seen = {}
            for r in rows:
                rid = r[id_i] if id_i < len(r) else None
                amt = safe_float(r[amt_i]) if amt_i < len(r) else None
                try:
                    rid = int(rid)
                except (ValueError, TypeError):
                    amounts_ok = False
                    break
                seen[rid] = amt
            if amounts_ok:
                if set(seen.keys()) != set(exp["amount_by_id"].keys()):
                    amounts_ok = False
                else:
                    for rid, amt in seen.items():
                        if amt is None or abs(amt - exp["amount_by_id"][rid]) > 0.01:
                            amounts_ok = False
                            break
        check("Refund Details: per-Refund_ID Amount matches DB", amounts_ok,
              f"expected ids/amounts: {exp['amount_by_id']}")
    else:
        check("Refund Details: per-Refund_ID Amount matches DB", False, "no detail sheet")

    # --- Summary sheet ---
    ws2 = None
    for s in wb.sheetnames:
        if "summary" in s.lower():
            ws2 = wb[s]
            break
    check("Summary sheet exists", ws2 is not None, f"Sheets: {wb.sheetnames}")

    summary = {}
    if ws2:
        for row in ws2.iter_rows(min_row=1, values_only=True):
            if row and row[0] is not None:
                key = str(row[0]).strip().lower().replace(" ", "_")
                summary[key] = row[1] if len(row) > 1 else None

    def sget(*keys):
        for k in keys:
            if k in summary:
                return safe_float(summary[k])
        return None

    total_refunds = sget("total_refunds")
    total_amount = sget("total_refund_amount")
    avg_amount = sget("avg_refund_amount", "average_refund_amount")
    refund_rate = sget("refund_rate")

    # Structural presence (non-critical).
    check("Summary has Total_Refunds entry", total_refunds is not None, f"keys: {list(summary.keys())}")
    check("Summary has Total_Refund_Amount entry", total_amount is not None, f"keys: {list(summary.keys())}")
    check("Summary has Avg_Refund_Amount entry", avg_amount is not None, f"keys: {list(summary.keys())}")
    check("Summary has Refund_Rate entry", refund_rate is not None, f"keys: {list(summary.keys())}")

    # CRITICAL: all four computed values correct.
    summary_ok = (
        total_refunds is not None and abs(total_refunds - exp["count"]) < 0.5 and
        total_amount is not None and abs(total_amount - exp["total"]) <= 0.01 and
        avg_amount is not None and abs(avg_amount - exp["avg"]) <= 0.01 and
        refund_rate is not None and abs(refund_rate - exp["rate"]) <= 0.05
    )
    check("Summary: Total_Refunds / Total_Refund_Amount / Avg_Refund_Amount / Refund_Rate match DB",
          summary_ok,
          f"got {total_refunds}/{total_amount}/{avg_amount}/{refund_rate}, "
          f"expected {exp['count']}/{exp['total']}/{exp['avg']}/{exp['rate']}")


def check_teamly(exp):
    print("\n=== Checking Teamly ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly 'Refund Analysis Dashboard' page exists", False, str(e))
        check("Teamly 'Refund Analysis Dashboard' page reflects count, amount and most-common reason",
              False, str(e))
        return

    dash = None
    for pid, title, body in pages:
        if "refund analysis dashboard" in (title or "").lower():
            dash = (pid, title, body)
            break
    check("Teamly 'Refund Analysis Dashboard' page exists", dash is not None,
          f"new pages: {[(p[0], p[1]) for p in pages]}")

    if dash is None:
        check("Teamly 'Refund Analysis Dashboard' page reflects count, amount and most-common reason",
              False, "no dashboard page")
        return

    text = ((dash[1] or "") + " " + (dash[2] or ""))
    low = text.lower()

    # Structural: 'Refund Trends' marker present (non-critical).
    check("Teamly page contains 'Refund Trends' structure",
          "refund trends" in low, "marker absent")

    # CRITICAL: page reflects total count, total amount, most-common reason.
    has_count = str(exp["count"]) in text
    total_str = f"{exp['total']:.2f}"
    has_total = total_str in text or str(exp["total"]) in text or f"{exp['total']:.1f}" in text
    has_reason = exp["mode_reason"].lower() in low if exp["mode_reason"] else False
    check("Teamly 'Refund Analysis Dashboard' page reflects count, amount and most-common reason",
          has_count and has_total and has_reason,
          f"count={has_count} total={has_total} reason='{exp['mode_reason']}'={has_reason}")


def check_email(exp):
    print("\n=== Checking Email ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        all_emails = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Email sent to cfo@company.com", False, str(e))
        check("Email to cfo@company.com body contains all key figures and the most-common reason",
              False, str(e))
        return

    print(f"[check_email] Found {len(all_emails)} emails.")

    target = None
    for subject, to_addr, body_text in all_emails:
        if isinstance(to_addr, list):
            to_str = " ".join(str(r).lower() for r in to_addr)
        elif isinstance(to_addr, str):
            try:
                parsed = json.loads(to_addr)
                to_str = " ".join(str(r).lower() for r in parsed) if isinstance(parsed, list) \
                    else str(to_addr).lower()
            except (json.JSONDecodeError, TypeError):
                to_str = str(to_addr).lower()
        else:
            to_str = str(to_addr).lower()
        if "cfo@company.com" in to_str:
            target = (subject, body_text)
            break

    check("Email sent to cfo@company.com", target is not None)
    if target is None:
        check("Email to cfo@company.com body contains all key figures and the most-common reason",
              False, "no email to cfo")
        return

    subject, body_text = target
    check("Email subject contains 'refund'",
          "refund" in (subject or "").lower(), f"Subject: {subject}")

    body = body_text or ""
    low = body.lower()
    has_count = str(exp["count"]) in body
    total_str = f"{exp['total']:.2f}"
    has_total = total_str in body or str(exp["total"]) in body or f"{exp['total']:.1f}" in body
    avg_str = f"{exp['avg']:.2f}"
    has_avg = avg_str in body or str(exp["avg"]) in body or f"{exp['avg']:.1f}" in body
    has_reason = exp["mode_reason"].lower() in low if exp["mode_reason"] else False
    check("Email to cfo@company.com body contains all key figures and the most-common reason",
          has_count and has_total and has_avg and has_reason,
          f"count={has_count} total={has_total} avg={has_avg} reason='{exp['mode_reason']}'={has_reason}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    exp = get_expected()
    print(f"Expected {exp['count']} refunds, {exp['order_count']} orders, "
          f"total {exp['total']}, avg {exp['avg']}, rate {exp['rate']}, "
          f"most-common reason '{exp['mode_reason']}'")

    check_excel(args.agent_workspace, exp)
    check_teamly(exp)
    check_email(exp)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    success = (not critical_failed) and accuracy >= 70
    print(f"  Overall: {'PASS' if success else 'FAIL'}")

    if args.res_log_file:
        try:
            with open(args.res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                    "success": success,
                }, f, indent=2)
        except Exception:
            pass

    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
