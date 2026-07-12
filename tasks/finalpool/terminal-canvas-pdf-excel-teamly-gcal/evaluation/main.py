"""Evaluation script for terminal-canvas-pdf-excel-teamly-gcal.

Pass rule: accuracy >= 70%  AND  no CRITICAL check failed.
CRITICAL checks reflect the task's substance (correct computed metrics +
correctly derived compliance statuses + the 8-criterion Teamly page + the
three accreditation calendar events), not mere structure. Any critical
failure => sys.exit(1) regardless of accuracy.

All expected numeric values are recomputed LIVE from canvas.* per the task.md
semantics (avg_gpa = mean of per-course average submission grades). Volatile
canvas data is never hardcoded.
"""
import os
import argparse, json, os, sys

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Thresholds legitimately come from accreditation_standards.pdf.
C1_THRESHOLD = 78.0   # Student Learning Outcomes (avg_gpa, 0-100 scale)
C2_THRESHOLD = 0.65   # Assessment Diversity (fraction)

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "program_metrics.json avg_gpa matches per-course mean (task.md formula)",
    "C1 Status correctly derived from avg_gpa vs 78.0 threshold",
    "C2 Assessment Diversity Actual_Value correct and Status derived vs 0.65",
    "Teamly 'Accreditation Action Items' page has 8 criterion entries with statuses",
    "Three accreditation calendar events exist (Evidence/Draft/Final)",
}


def get_expected_from_db():
    """Recompute expected values LIVE from canvas.* per task.md semantics.

    - per-course Avg_Grade = AVG(canvas.submissions.score) for that course
    - avg_gpa = mean of those per-course averages (NOT a flat AVG over all
      submissions) — matches task.md.
    - assessment_diversity = (#courses with BOTH assignments AND quizzes) /
      (#courses).
    """
    defaults = {
        "avg_gpa": 75.4,
        "assessment_diversity": 0.5,
        "course_count": 22,
    }
    try:
        import psycopg2
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("SELECT id FROM canvas.courses")
        course_ids = [r[0] for r in cur.fetchall()]
        defaults["course_count"] = len(course_ids)

        per_course_avgs = []
        diversity_hits = 0
        for cid in course_ids:
            cur.execute("""
                SELECT AVG(s.score)
                FROM canvas.submissions s
                JOIN canvas.assignments a ON s.assignment_id = a.id
                WHERE a.course_id = %s AND s.score IS NOT NULL
            """, (cid,))
            avg = cur.fetchone()[0]
            if avg is not None:
                per_course_avgs.append(float(avg))

            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = %s", (cid,))
            n_asgn = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM canvas.quizzes WHERE course_id = %s", (cid,))
            n_quiz = cur.fetchone()[0]
            if n_asgn > 0 and n_quiz > 0:
                diversity_hits += 1

        if per_course_avgs:
            defaults["avg_gpa"] = sum(per_course_avgs) / len(per_course_avgs)
        if course_ids:
            defaults["assessment_diversity"] = diversity_hits / len(course_ids)

        cur.close()
        conn.close()
    except Exception as e:
        print(f"  [WARN] DB query for expected values failed, using defaults: {e}")
    return defaults


EXPECTED = get_expected_from_db()


def derive_status(actual, threshold):
    """Compliant if >= threshold; Partial if within 10% below; else Non-Compliant."""
    try:
        a = float(actual)
        t = float(threshold)
    except (TypeError, ValueError):
        return None
    if a >= t:
        return "Compliant"
    if a >= t * 0.9:
        return "Partial"
    return "Non-Compliant"


EXPECTED_C1_STATUS = derive_status(EXPECTED["avg_gpa"], C1_THRESHOLD)
EXPECTED_C2_STATUS = derive_status(EXPECTED["assessment_diversity"], C2_THRESHOLD)


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def num_close(a, b, tol=2.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except:
        return False


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def check_excel(agent_workspace):
    print("\n=== Checking Excel Workbook ===")
    import openpyxl

    path = os.path.join(agent_workspace, "Accreditation_Self_Study.xlsx")
    check("Accreditation_Self_Study.xlsx exists", os.path.exists(path))
    if not os.path.exists(path):
        return

    wb = openpyxl.load_workbook(path)
    sheet_names = [s.lower() for s in wb.sheetnames]

    # Check 4 sheets exist
    has_course = any("course" in s and "data" in s for s in sheet_names)
    has_matrix = any("accreditation" in s or "matrix" in s for s in sheet_names)
    has_gap = any("gap" in s for s in sheet_names)
    has_resource = any("resource" in s for s in sheet_names)
    check("Course_Data sheet exists", has_course, f"Sheets: {wb.sheetnames}")
    check("Accreditation_Matrix sheet exists", has_matrix, f"Sheets: {wb.sheetnames}")
    check("Gap_Analysis sheet exists", has_gap, f"Sheets: {wb.sheetnames}")
    check("Resource_Needs sheet exists", has_resource, f"Sheets: {wb.sheetnames}")

    # Check Course_Data content
    for sn in wb.sheetnames:
        if "course" in sn.lower() and "data" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            # Row count derived live (not hardcoded) — allow small slack.
            check("Course_Data has one row per course",
                  len(rows) >= max(1, EXPECTED["course_count"] - 1),
                  f"Got {len(rows)}, expected ~{EXPECTED['course_count']}")
            if rows:
                # Check some known course names
                names = [str(r[0]).lower() if r[0] else "" for r in rows]
                check("Course_Data includes Creative Computing", any("креативн" in n or "вычислен" in n for n in names))
                check("Course_Data includes Основы финансов", any("финанс" in n for n in names))
            # Check headers
            headers = [str(c.value).lower() if c.value else "" for c in ws[1]]
            check("Course_Data has enrollment column", any("enroll" in h for h in headers), f"Headers: {headers}")
            check("Course_Data has grade column", any("grade" in h or "avg" in h for h in headers), f"Headers: {headers}")
            break

    # Check Accreditation_Matrix content
    matrix_c1_status = None
    matrix_c2_actual = None
    matrix_c2_status = None
    for sn in wb.sheetnames:
        if "accreditation" in sn.lower() or "matrix" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Accreditation_Matrix has 8 criteria rows", len(rows) >= 8, f"Got {len(rows)}")
            if rows:
                all_text = " ".join(str(c).lower() for r in rows for c in r if c)
                check("Matrix has Compliant status", "compliant" in all_text)
                check("Matrix has Partial status", "partial" in all_text)
                # Check specific criteria
                check("Matrix mentions Student Learning Outcomes",
                      "student learning" in all_text or "learning outcome" in all_text,
                      f"Text sample: {all_text[:200]}")
                # Check threshold values
                has_78 = any(num_close(r[1], C1_THRESHOLD, tol=1.0) for r in rows if len(r) > 1 and r[1] is not None)
                has_065 = any(num_close(r[1], C2_THRESHOLD, tol=0.05) for r in rows if len(r) > 1 and r[1] is not None)
                check("Matrix has C1 threshold ~78", has_78, f"Thresholds: {[r[1] for r in rows if len(r) > 1]}")
                check("Matrix has C2 threshold ~0.65", has_065, f"Thresholds: {[r[1] for r in rows if len(r) > 1]}")

                # Locate C1 (Student Learning Outcomes) row -> Status (col 4)
                for r in rows:
                    crit = str(r[0]).lower() if r and r[0] else ""
                    if "student learning" in crit or "learning outcome" in crit:
                        matrix_c1_status = str(r[3]).strip() if len(r) > 3 and r[3] else ""
                    if "assessment diversity" in crit or ("diversity" in crit and "assess" in crit):
                        matrix_c2_actual = r[2] if len(r) > 2 else None
                        matrix_c2_status = str(r[3]).strip() if len(r) > 3 and r[3] else ""
            break

    # CRITICAL: C1 status correctly derived from live avg_gpa vs 78.0
    check("C1 Status correctly derived from avg_gpa vs 78.0 threshold",
          matrix_c1_status is not None and EXPECTED_C1_STATUS is not None
          and matrix_c1_status.lower() == EXPECTED_C1_STATUS.lower(),
          f"Got {matrix_c1_status!r}, expected {EXPECTED_C1_STATUS!r} (avg_gpa={EXPECTED['avg_gpa']:.2f})")

    # CRITICAL: C2 actual value correct and status derived vs 0.65
    c2_actual_ok = matrix_c2_actual is not None and num_close(matrix_c2_actual, EXPECTED["assessment_diversity"], tol=0.06)
    c2_status_ok = (matrix_c2_status is not None and EXPECTED_C2_STATUS is not None
                    and matrix_c2_status.lower() == EXPECTED_C2_STATUS.lower())
    check("C2 Assessment Diversity Actual_Value correct and Status derived vs 0.65",
          c2_actual_ok and c2_status_ok,
          f"actual={matrix_c2_actual!r} (exp {EXPECTED['assessment_diversity']:.3f}), "
          f"status={matrix_c2_status!r} (exp {EXPECTED_C2_STATUS!r})")

    # Check Gap_Analysis content
    for sn in wb.sheetnames:
        if "gap" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            non_empty = [r for r in rows if any(c for c in r)]
            check("Gap_Analysis has rows for partial/non-compliant criteria", len(non_empty) >= 1, f"Got {len(non_empty)}")
            if non_empty:
                all_text = " ".join(str(c).lower() for r in non_empty for c in r if c)
                check("Gap_Analysis mentions remediation",
                      any(kw in all_text for kw in
                          ["remed", "action", "improve", "устран", "меропри", "действ", "улучш", "план"]))
            break

    # Check Resource_Needs content
    for sn in wb.sheetnames:
        if "resource" in sn.lower():
            ws = wb[sn]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            non_empty = [r for r in rows if any(c for c in r)]
            check("Resource_Needs has at least 3 rows", len(non_empty) >= 3, f"Got {len(non_empty)}")
            if non_empty:
                headers = [str(c.value).lower() if c.value else "" for c in ws[1]]
                check("Resource_Needs has cost column", any("cost" in h or "estimated" in h for h in headers), f"Headers: {headers}")
            break


def _teamly_accred_pages(cur):
    """Return list of (id, title, body) for the agent's accreditation page(s)."""
    cur.execute("""
        SELECT id, title, COALESCE(body, '')
        FROM teamly.pages
        WHERE title ILIKE '%%accreditation action items%%'
    """)
    return cur.fetchall()


def check_teamly():
    print("\n=== Checking Teamly Page ===")
    try:
        conn = get_conn()
        cur = conn.cursor()

        pages = _teamly_accred_pages(cur)
        check("Accreditation Action Items page exists", len(pages) > 0,
              f"Found {len(pages)} matching pages")

        if pages:
            combined = " ".join((str(t) + " " + str(b)) for _, t, b in pages)
            low = combined.lower()

            # Status presence
            check("Teamly page has Compliant status", "compliant" in low)
            check("Teamly page has Partial status", "partial" in low)

            # Count Compliant/Partial/Non-Compliant occurrences as a proxy for
            # the 8 per-criterion rows (the page is a table, one row per crit).
            n_status = (low.count("compliant") + low.count("partial"))
            # 'non-compliant' also contains 'compliant'; this is a permissive
            # structural proxy. The CRITICAL count check below is stricter.
            check("Teamly page has Owner column", "owner" in low, f"sample: {low[:150]}")
            check("Teamly page has Evidence column", "evidence" in low, f"sample: {low[:150]}")

            # CRITICAL: page must enumerate all 8 criteria with statuses.
            # Count distinct criterion mentions (C1-C8 names) present.
            criteria_keywords = [
                ("student learning", "learning outcome"),
                ("assessment diversity",),
                ("student retention", "retention"),
                ("faculty qualif",),
                ("curriculum coverage", "curriculum"),
                ("course enrollment adequacy", "enrollment adequacy"),
                ("quiz-based", "quiz based", "quiz-based assessment"),
                ("program breadth", "breadth"),
            ]
            crit_hits = sum(1 for kws in criteria_keywords if any(k in low for k in kws))
            # At least 6 explicit status tokens across the table (8 rows expected,
            # allow some leniency), and >=6 of the 8 criteria named.
            status_tokens = low.count("compliant") + low.count("partial")
            check("Teamly 'Accreditation Action Items' page has 8 criterion entries with statuses",
                  crit_hits >= 6 and status_tokens >= 6,
                  f"criterion-name hits={crit_hits} (need >=6), status tokens={status_tokens} (need >=6)")

        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly accessible", False, str(e))
        check("Teamly 'Accreditation Action Items' page has 8 criterion entries with statuses",
              False, str(e))


def check_gcal(launch_time):
    print("\n=== Checking Google Calendar Events ===")
    try:
        conn = get_conn()
        cur = conn.cursor()
        launch_dt = None
        if launch_time:
            from datetime import datetime
            launch_dt = datetime.strptime(launch_time, "%Y-%m-%d %H:%M:%S")

        cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
        events = cur.fetchall()
        summaries = [str(e[0]).lower() for e in events]

        has_evidence = any("evidence" in s and "deadline" in s for s in summaries)
        has_draft = any("draft" in s and ("report" in s or "due" in s) for s in summaries)
        has_final = any("final" in s and "submission" in s for s in summaries)
        check("Evidence Collection Deadline event exists", has_evidence, f"Events: {summaries}")
        check("Draft Report Due event exists", has_draft, f"Events: {summaries}")
        check("Final Submission event exists", has_final, f"Events: {summaries}")

        # CRITICAL: all three accreditation events present.
        check("Three accreditation calendar events exist (Evidence/Draft/Final)",
              has_evidence and has_draft and has_final, f"Events: {summaries}")

        # Check descriptions mention accreditation (RU or EN)
        accred_events = [e for e in events if any(kw in str(e[0]).lower() for kw in ["evidence", "draft report", "final submission"])]
        if accred_events:
            descs = " ".join(str(e[1]).lower() for e in accred_events if e[1])
            check("Calendar events mention accreditation",
                  any(kw in descs for kw in ["accreditation", "self-study", "self study",
                                             "аккредит", "самообслед"]),
                  f"Descriptions: {descs[:200]}")

        # Check timing (roughly 30/60/90 days from launch)
        if launch_dt and accred_events:
            for e in events:
                s = str(e[0]).lower()
                if e[2] and "evidence" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Evidence deadline ~30 days from launch", 25 <= days_diff <= 35, f"Days: {days_diff}")
                elif e[2] and "draft" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Draft due ~60 days from launch", 55 <= days_diff <= 65, f"Days: {days_diff}")
                elif e[2] and "final" in s and "submission" in s:
                    days_diff = (e[2].replace(tzinfo=None) - launch_dt).days
                    check("Final submission ~90 days from launch", 85 <= days_diff <= 95, f"Days: {days_diff}")

        cur.close()
        conn.close()
    except Exception as e:
        check("GCal accessible", False, str(e))
        check("Three accreditation calendar events exist (Evidence/Draft/Final)", False, str(e))


def check_scripts(agent_workspace):
    print("\n=== Checking Scripts and Outputs ===")
    check("compute_metrics.py exists", os.path.exists(os.path.join(agent_workspace, "compute_metrics.py")))
    check("evaluate_compliance.py exists", os.path.exists(os.path.join(agent_workspace, "evaluate_compliance.py")))
    check("generate_summary.py exists", os.path.exists(os.path.join(agent_workspace, "generate_summary.py")))

    # Check program_metrics.json
    metrics_path = os.path.join(agent_workspace, "program_metrics.json")
    check("program_metrics.json exists", os.path.exists(metrics_path))
    if os.path.exists(metrics_path):
        with open(metrics_path) as f:
            metrics = json.load(f)
        check("Metrics has avg_gpa", "avg_gpa" in metrics, f"Keys: {list(metrics.keys())}")
        # CRITICAL: avg_gpa must match the task.md per-course-mean formula.
        check("program_metrics.json avg_gpa matches per-course mean (task.md formula)",
              "avg_gpa" in metrics and num_close(metrics["avg_gpa"], EXPECTED["avg_gpa"], tol=2.0),
              f"Got {metrics.get('avg_gpa')}, expected ~{EXPECTED['avg_gpa']:.2f}")

    # Check compliance_assessment.json
    compliance_path = os.path.join(agent_workspace, "compliance_assessment.json")
    check("compliance_assessment.json exists", os.path.exists(compliance_path))
    if os.path.exists(compliance_path):
        with open(compliance_path) as f:
            compliance = json.load(f)
        if isinstance(compliance, list):
            check("Compliance has 8 entries", len(compliance) >= 8, f"Got {len(compliance)}")
        elif isinstance(compliance, dict):
            check("Compliance has criteria data", len(compliance) >= 3, f"Keys: {list(compliance.keys())[:10]}")

    # Check accreditation_summary.txt
    summary_path = os.path.join(agent_workspace, "accreditation_summary.txt")
    check("accreditation_summary.txt exists", os.path.exists(summary_path))
    if os.path.exists(summary_path):
        with open(summary_path) as f:
            text = f.read().lower()
        check("Summary mentions compliant", "compliant" in text)
        check("Summary mentions partial", "partial" in text)
        check("Summary has compliance percentage",
              "%" in text or "percent" in text or "процент" in text, f"Length: {len(text)}")


def check_reverse_validation():
    """Check that RU noise Teamly pages are NOT in the accreditation page."""
    print("\n=== Reverse Validation ===")
    try:
        conn = get_conn()
        cur = conn.cursor()

        # RU noise teamly page titles that should NOT appear in the accreditation
        # tracker. Matched against ORIGINAL-case text lowered (NOT normalized).
        noise_titles = ["обзор бюджета q1", "миграция серверов", "запуск маркетинговой кампании"]

        pages = _teamly_accred_pages(cur)
        if pages:
            combined = " ".join((str(t) + " " + str(b)).lower() for _, t, b in pages)
            no_noise = not any(nt in combined for nt in noise_titles)
            check("No noise Teamly pages in accreditation tracker (budget, migration, marketing)",
                  no_noise, f"Accreditation page text sample: {combined[:200]}")
        else:
            check("No noise Teamly pages in accreditation tracker", True,
                  "No accreditation page found to check")

        cur.close()
        conn.close()
    except Exception as e:
        check("Reverse validation (teamly noise)", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print(f"[eval] Expected (live from canvas): avg_gpa={EXPECTED['avg_gpa']:.2f}, "
          f"diversity={EXPECTED['assessment_diversity']:.3f}, "
          f"C1={EXPECTED_C1_STATUS}, C2={EXPECTED_C2_STATUS}, "
          f"courses={EXPECTED['course_count']}")

    check_excel(args.agent_workspace)
    check_teamly()
    check_gcal(args.launch_time)
    check_scripts(args.agent_workspace)
    check_reverse_validation()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\nOverall: {PASS_COUNT}/{total} ({accuracy:.1f}%)")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        sys.exit(1)
    sys.exit(0 if accuracy >= 70 else 1)


if __name__ == "__main__":
    main()
