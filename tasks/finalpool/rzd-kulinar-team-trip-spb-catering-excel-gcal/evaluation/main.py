"""Evaluation для rzd-kulinar-team-trip-spb-catering-excel-gcal.

Строгий evaluator, 11 проверок:
  Excel (7):
    1.  Team_Trip_Plan.xlsx существует
    2.  Документ читается
    3.  Все три листа Travel/Menu/Timeline присутствуют
    4.  Travel: номер поезда содержит '752' (самый ранний Сапсан Мск→СПб)
    5.  Travel: указано время '06:50'
    6.  Menu: >= 5 строк данных, есть колонки Course_Type и Dish_Name
    7.  Timeline: >= 5 строк данных
  Calendar (2):
    8.  >= 2 события на 2026-03-10 НЕ из preprocess
    9.  Среди них есть отправление/прибытие/ужин
  Email (2):
   10.  >= 1 отправленное письмо на events@company.ru
   11.  В теле письма упоминание поезда, плана и Санкт-Петербурга
"""
import json
import os
import sys
import unicodedata
from argparse import ArgumentParser

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": int(os.environ.get("PGPORT", "5432")),
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")


def normalize(s: str) -> str:
    """Lowercase + collapse cyrillic/latin lookalikes (А/A, С/C ...)."""
    s = s.lower()
    s = unicodedata.normalize("NFKD", s)
    table = str.maketrans({"а": "a", "о": "o", "е": "e", "р": "p",
                           "с": "c", "у": "y", "к": "k", "х": "x"})
    return s.translate(table)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def check_excel(agent_workspace):
    print("\n=== Excel: Team_Trip_Plan.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Team_Trip_Plan.xlsx")
    if not os.path.exists(xlsx_path):
        record("Team_Trip_Plan.xlsx существует", False, f"не найден в {agent_workspace}")
        return
    record("Team_Trip_Plan.xlsx существует", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Документ читается", False, str(e))
        return
    record("Документ читается", True)

    sheet_lower = [s.lower() for s in wb.sheetnames]
    has_travel = any("travel" in s for s in sheet_lower)
    has_menu = any("menu" in s for s in sheet_lower)
    has_timeline = any("timeline" in s for s in sheet_lower)
    record("Все три листа Travel/Menu/Timeline присутствуют",
           has_travel and has_menu and has_timeline,
           f"листы: {wb.sheetnames}")

    if has_travel:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_lower) if "travel" in s)]
        ws = wb[ws_name]
        all_text = " ".join(str(c) for row in ws.iter_rows(values_only=True) for c in row if c is not None)
        norm = normalize(all_text)
        record("Travel: номер поезда '752' (самый ранний Сапсан)",
               "752" in norm, f"sample: {all_text[:200]!r}")
        has_time = "06:50" in all_text or "6:50" in all_text
        record("Travel: время отправления 06:50", has_time,
               f"sample: {all_text[:200]!r}")
    else:
        record("Travel: номер поезда '752'", False, "лист Travel отсутствует")
        record("Travel: время отправления 06:50", False, "лист Travel отсутствует")

    if has_menu:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_lower) if "menu" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
        headers = [str(c).lower() if c else "" for c in (rows[0] if rows else [])]
        has_course = any("course" in h or "type" in h or "категори" in h or "блюд" in h for h in headers)
        has_dish = any("dish" in h or "name" in h or "назван" in h for h in headers)
        record("Menu: >=5 строк, есть Course_Type и Dish_Name",
               len(data_rows) >= 5 and has_course and has_dish,
               f"rows={len(data_rows)}, headers={rows[0] if rows else []}")
    else:
        record("Menu: >=5 строк, есть Course_Type и Dish_Name", False, "лист Menu отсутствует")

    if has_timeline:
        ws_name = wb.sheetnames[next(i for i, s in enumerate(sheet_lower) if "timeline" in s)]
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c is not None and str(c).strip() for c in r)]
        record("Timeline: >=5 строк данных", len(data_rows) >= 5, f"rows={len(data_rows)}")
    else:
        record("Timeline: >=5 строк данных", False, "лист Timeline отсутствует")


# ---------------------------------------------------------------------------
# Calendar
# ---------------------------------------------------------------------------

def check_gcal():
    print("\n=== Calendar: новые события 10.03.2026 ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT summary, start_datetime FROM gcal.events
         WHERE start_datetime::date = '2026-03-10'
           AND summary NOT ILIKE '%общее собрание%'
           AND summary NOT ILIKE '%партн%'
         ORDER BY start_datetime
    """)
    events = cur.fetchall()
    cur.close()
    conn.close()

    record("Не менее 2 новых событий на 10.03 (без preprocess)",
           len(events) >= 2,
           f"events: {[(e[0], str(e[1])) for e in events]}")

    summaries_low = " ".join(str(e[0]).lower() for e in events)
    has_dep = any(kw in summaries_low for kw in
                  ["отправ", "выезд", "departure", "trip", "вокзал", "поездк"])
    has_arrive = any(kw in summaries_low for kw in
                     ["прибыт", "заселен", "чек-ин", "check", "hotel", "arriv"])
    has_dinner = any(kw in summaries_low for kw in
                     ["ужин", "dinner"])
    record("События содержат отправление/прибытие/ужин",
           sum([has_dep, has_arrive, has_dinner]) >= 2,
           f"dep={has_dep}, arr={has_arrive}, dinner={has_dinner}")


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

def check_email():
    print("\n=== Email: events@company.ru ===")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute(
        """
        SELECT m.subject, m.body_text
          FROM email.messages m
          JOIN email.sent_log s ON s.message_id = m.id
         WHERE m.to_addr::text ILIKE %s
           AND m.from_addr   NOT ILIKE %s
        """,
        ("%events@company.ru%", "%events@company.ru%"),
    )
    rows = cur.fetchall()
    if not rows:
        cur.execute(
            """
            SELECT subject, body_text FROM email.messages
             WHERE to_addr::text ILIKE %s
               AND from_addr  NOT ILIKE %s
            """,
            ("%events@company.ru%", "%events@company.ru%"),
        )
        rows = cur.fetchall()
    cur.close()
    conn.close()

    record("Письмо отправлено на events@company.ru", len(rows) >= 1,
           f"matched: {len(rows)}")
    if not rows:
        record("В теле упоминание поезда/плана/СПб", False)
        return

    body = " ".join(((s or "") + " " + (b or "")) for s, b in rows)
    body_low = body.lower()
    body_norm = normalize(body)
    has_train = "752" in body_norm or "сапсан" in body_low or "sapsan" in body_norm
    has_plan = "план" in body_low or "plan" in body_norm or "excel" in body_norm or "notion" in body_norm
    has_city = "петербург" in body_low or "spb" in body_norm or "сп6" in body_low
    record("В теле упоминание поезда / плана / СПб",
           has_train and has_plan and has_city,
           f"train={has_train}, plan={has_plan}, city={has_city}")


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    check_excel(args.agent_workspace)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")

    result = {"total_passed": PASS_COUNT, "total_checks": total, "accuracy": accuracy}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    else:
        print("FAIL")
        sys.exit(1)


if __name__ == "__main__":
    main()
