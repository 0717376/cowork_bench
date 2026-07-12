"""Evaluation for sf-hr-rating-distribution (ClickHouse, RU dept names)."""
import argparse
import os
import sys
import openpyxl
import psycopg2

DB = {"host": os.environ.get("PGHOST", "localhost"), "port": 5432, "dbname": "cowork_gym", "user": "eigent", "password": "camel"}


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def rows_to_lookup(rows):
    data = rows[1:] if len(rows) > 1 else []
    lookup = {}
    for row in data:
        if row and row[0] is not None:
            lookup[str(row[0]).strip().lower()] = row
    return lookup


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "HR_Rating_Distribution.xlsx")
    gt_file = os.path.join(gt_dir, "HR_Rating_Distribution.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_errors = []

    # ---------------- Rating Distribution sheet ----------------
    print("  Checking Rating Distribution...")
    a_rows = load_sheet_rows(agent_wb, "Rating Distribution")
    g_rows = load_sheet_rows(gt_wb, "Rating Distribution")
    if a_rows is None:
        all_errors.append("Sheet 'Rating Distribution' not found in agent output")
        critical_errors.append("Sheet 'Rating Distribution' missing")
    elif g_rows is None:
        all_errors.append("Sheet 'Rating Distribution' not found in groundtruth")
    else:
        errors = []
        a_lookup = rows_to_lookup(a_rows)
        g_lookup = rows_to_lookup(g_rows)
        # CRITICAL: every groundtruth Dept_Rating row present with Count within tight tol (<=2)
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                critical_errors.append(f"Rating Distribution missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 2):
                    errors.append(f"{key}.Count: {a_row[1]} vs {g_row[1]} (tol=2)")
                    critical_errors.append(f"Rating Distribution {key}.Count {a_row[1]} != {g_row[1]}")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # ---------------- Summary sheet ----------------
    print("  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
        critical_errors.append("Sheet 'Summary' missing")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        errors = []
        a_lookup = rows_to_lookup(a_rows)
        g_lookup = rows_to_lookup(g_rows)
        # Tight tolerances per metric. Total_Employees exact; counts <=2; pct <=0.2.
        metric_tol = {
            "total_employees": 0,
            "rating_5_count": 2,
            "rating_4_count": 2,
            "rating_1_count": 2,
            "high_performers_pct": 0.2,
        }
        critical_metrics = set(metric_tol.keys())
        for key, g_row in g_lookup.items():
            a_row = a_lookup.get(key)
            tol = metric_tol.get(key, 2)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                if key in critical_metrics:
                    critical_errors.append(f"Summary missing metric: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol={tol})")
                    if key in critical_metrics:
                        critical_errors.append(f"Summary {key} {a_row[1]} != {g_row[1]} (tol={tol})")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    # ---------------- Email check ----------------
    print("  Checking Email...")
    email_ok = False
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, to_addr, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%hr-analytics@company.com%'
               OR subject ILIKE '%Performance Rating Distribution Report%'
            LIMIT 20
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        to_addrs = [str(r[1] or "") for r in rows]
        subjects = [str(r[0] or "") for r in rows]
        recip_ok = any("hr-analytics@company.com" in a.lower() for a in to_addrs)
        # subject must contain the literal English identifier (kept English)
        subj_ok = any("performance rating distribution report" in s.lower() for s in subjects)
        email_ok = recip_ok and subj_ok
        if not recip_ok:
            all_errors.append("Email not sent to hr-analytics@company.com")
            critical_errors.append("Email recipient hr-analytics@company.com missing")
        if not subj_ok:
            all_errors.append("Email subject 'Performance Rating Distribution Report' missing")
            critical_errors.append("Email subject 'Performance Rating Distribution Report' missing")
    except Exception as e:
        all_errors.append(f"Email check error: {e}")
        critical_errors.append(f"Email check failed: {e}")
    print("    PASS" if email_ok else "    ERRORS")

    # ---------------- CRITICAL gate ----------------
    if critical_errors:
        print(f"\n=== CRITICAL FAIL ({len(critical_errors)} critical errors) ===")
        for e in critical_errors[:10]:
            print(f"  CRITICAL: {e}")
        sys.exit(1)

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
