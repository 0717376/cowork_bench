"""Evaluation for sf-hr-performance-excel-gcal-email.

Checks:
1. Performance_Review_Summary.xlsx with 3 sheets
2. Google Calendar event "Annual Performance Review Board Meeting" 21 days from launch
3. Email to executives@company.example.com
"""
import argparse
import os
import re
import sys
from datetime import datetime, timedelta

import openpyxl
import psycopg2


def normalize_ru_numbers(text):
    """RU number normalization for substring/regex checks: collapse digit-group
    separators (space/NBSP/NNBSP/dot/comma before a 3-digit group) and turn
    decimal commas into dots ("31 588" -> "31588", "4 586,91" -> "4586.91")."""
    t = str(text or "")
    t = re.sub(r"(?<=\d)[ \xa0\u202f\u2009.,](?=\d{3}\b)", "", t)
    return re.sub(r"(?<=\d),(?=\d)", ".", t)

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432, dbname="cowork_gym", user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0

# Department names are russified centrally by db/zzz_clickhouse_after_init.sql:
#   Engineering->Инженерия, Finance->Финансы, HR->Кадры, Operations->Операции,
#   R&D->НИОКР, Sales->Продажи, Support->Поддержка
# The agent reads these Russian names live and writes them to the Department column,
# so eval keys + groundtruth xlsx use the Russian forms. Numeric aggregates are unchanged.
TOP_PERFORMERS = [
    ("Инженерия", 721, 59150.56, 8.35),
    ("Финансы",   721, 58641.91, 8.29),
    ("Кадры",     740, 57240.73, 8.79),
    ("Операции",  687, 57575.52, 8.60),
    ("НИОКР",     693, 56086.53, 8.38),
    ("Продажи",   694, 59193.14, 7.77),
    ("Поддержка", 752, 59869.03, 8.09),
]

UNDERPERFORMERS = [
    ("Инженерия", 1381, 59347.28, 8.35),
    ("Финансы",   1387, 59698.81, 8.29),
    ("Кадры",     1425, 59741.41, 8.15),
    ("Операции",  1431, 57394.68, 8.49),
    ("НИОКР",     1415, 57839.57, 8.06),
    ("Продажи",   1494, 58111.22, 8.54),
    ("Поддержка", 1479, 58491.37, 8.04),
]


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        print(f"  [FAIL] {name}: {str(detail)[:200]}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Performance_Review_Summary.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Performance_Review_Summary.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Performance_Review_Summary.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return False

    all_ok = True

    # Check Top_Performers sheet
    agent_top = get_sheet(agent_wb, "Top_Performers")
    gt_top = get_sheet(gt_wb, "Top_Performers")
    check("Sheet 'Top_Performers' exists", agent_top is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_top is None:
        all_ok = False
    else:
        a_rows = list(agent_top.iter_rows(min_row=2, values_only=True))
        check("Top_Performers has 7 rows", len(a_rows) == 7, f"Got {len(a_rows)}")
        if len(a_rows) != 7:
            all_ok = False

        a_lookup = {str(r[0]).strip().lower(): r for r in a_rows if r and r[0]}
        for gt_dept, gt_cnt, gt_sal, gt_exp in TOP_PERFORMERS:
            key = gt_dept.lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Top '{gt_dept}' present", False, "Missing")
                all_ok = False
                continue
            ok_cnt = num_close(a_row[1], gt_cnt, 5)
            check(f"Top '{gt_dept}' Count_Rating_5", ok_cnt, f"Expected {gt_cnt}, got {a_row[1]}")
            if not ok_cnt:
                all_ok = False
            ok_sal = num_close(a_row[2], gt_sal, 500)
            check(f"Top '{gt_dept}' Avg_Salary_Top", ok_sal, f"Expected {gt_sal}, got {a_row[2]}")
            if not ok_sal:
                all_ok = False
            ok_exp = num_close(a_row[3], gt_exp, 0.5)
            check(f"Top '{gt_dept}' Avg_Experience_Top", ok_exp, f"Expected {gt_exp}, got {a_row[3]}")
            if not ok_exp:
                all_ok = False

    # Check Underperformers sheet
    agent_low = get_sheet(agent_wb, "Underperformers")
    check("Sheet 'Underperformers' exists", agent_low is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_low is None:
        all_ok = False
    else:
        a_rows = list(agent_low.iter_rows(min_row=2, values_only=True))
        check("Underperformers has 7 rows", len(a_rows) == 7, f"Got {len(a_rows)}")
        if len(a_rows) != 7:
            all_ok = False

        a_lookup = {str(r[0]).strip().lower(): r for r in a_rows if r and r[0]}
        for gt_dept, gt_cnt, gt_sal, gt_exp in UNDERPERFORMERS:
            key = gt_dept.lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Low '{gt_dept}' present", False, "Missing")
                all_ok = False
                continue
            ok_cnt = num_close(a_row[1], gt_cnt, 10)
            check(f"Low '{gt_dept}' Count_Low_Rating", ok_cnt, f"Expected {gt_cnt}, got {a_row[1]}")
            if not ok_cnt:
                all_ok = False
            ok_sal = num_close(a_row[2], gt_sal, 500)
            check(f"Low '{gt_dept}' Avg_Salary_Low", ok_sal, f"Expected {gt_sal}, got {a_row[2]}")
            if not ok_sal:
                all_ok = False
            ok_exp = num_close(a_row[3], gt_exp, 0.5)
            check(f"Low '{gt_dept}' Avg_Experience_Low", ok_exp, f"Expected {gt_exp}, got {a_row[3]}")
            if not ok_exp:
                all_ok = False

    # Check Summary sheet
    agent_sum = get_sheet(agent_wb, "Summary")
    check("Sheet 'Summary' exists", agent_sum is not None, f"Sheets: {agent_wb.sheetnames}")
    if agent_sum is None:
        all_ok = False
    else:
        a_summary = {}
        for row in agent_sum.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_summary[str(row[0]).strip().lower()] = row[1]

        ttp = a_summary.get("total_top_performers")
        check("Total_Top_Performers = 5008", num_close(ttp, 5008, 20), f"Got {ttp}")
        tup = a_summary.get("total_underperformers")
        check("Total_Underperformers = 10012", num_close(tup, 10012, 40), f"Got {tup}")
        oar = a_summary.get("overall_avg_rating")
        check("Overall_Avg_Rating close to 3.20", num_close(oar, 3.20, 0.1), f"Got {oar}")

    return all_ok


def check_gcal(launch_time_str):
    print("\n=== Checking Google Calendar ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("SELECT summary, description, start_datetime FROM gcal.events ORDER BY start_datetime")
    events = cur.fetchall()
    cur.close()
    conn.close()

    print(f"  Found {len(events)} calendar events")
    check("At least 1 calendar event", len(events) >= 1, f"Found {len(events)}")

    date_ok = False
    board_events = [e for e in events
                    if "annual" in (e[0] or "").lower() and "performance" in (e[0] or "").lower()]
    check("Annual Performance Review Board Meeting event exists", len(board_events) >= 1,
          f"Events: {[e[0] for e in events]}")

    if launch_time_str and board_events:
        try:
            # launch_time arrives as strftime("%Y-%m-%d %H:%M:%S %A") with a
            # trailing weekday (e.g. '2026-06-08 03:21:41 Monday'), which
            # fromisoformat cannot parse. Strip the weekday before parsing,
            # mirroring utils/roles/task_agent.py:179-182.
            lt = " ".join(launch_time_str.split()[:2])
            launch_dt = datetime.fromisoformat(lt)
            expected_dt = launch_dt + timedelta(days=21)
            for ev in board_events:
                if ev[2]:
                    ev_dt = ev[2]
                    diff = abs((ev_dt.replace(tzinfo=None) - expected_dt).total_seconds())
                    # Tightened from +/-2 days to ~12h so the 21-day rule is actually enforced.
                    date_ok = diff <= 3600 * 12
                    check("Board meeting 21 days from launch", date_ok,
                          f"Expected around {expected_dt}, got {ev_dt}")
                    break
        except Exception as e:
            print(f"  [INFO] Could not verify date: {e}")

    return {"event_ok": len(board_events) >= 1, "date_ok": date_ok}


def check_email():
    print("\n=== Checking Email ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()
    cur.execute("""
        SELECT subject, from_addr, to_addr, body_text
        FROM email.messages
        WHERE subject ILIKE '%performance%review%'
          AND (subject ILIKE '%annual%' OR from_addr ILIKE '%hr@company%' OR to_addr::text ILIKE '%executives%')
        ORDER BY date DESC
    """)
    emails = cur.fetchall()
    cur.close()
    conn.close()

    check("Performance review email exists", len(emails) >= 1, f"Found {len(emails)}")
    result = {"email_ok": len(emails) >= 1, "from_ok": False, "to_ok": False, "body_ok": False}
    if emails:
        e = emails[0]
        to_str = str(e[2])
        to_ok = "executives@company.example.com" in to_str.lower()
        check("Email to executives@company.example.com", to_ok, f"to: {to_str}")
        from_ok = "hr@company.example.com" in (e[1] or "").lower()
        check("Email from hr@company.example.com", from_ok, f"from: {e[1]}")
        body = normalize_ru_numbers((e[3] or "").lower())
        # Require the ACTUAL numeric totals (tolerate light rounding/formatting of the totals),
        # plus the overall average rating. Not just the word "top".
        def has_total(body, val):
            return any(str(v) in body for v in (val, val - 1, val + 1))
        top_ok = has_total(body, 5008)
        low_ok = has_total(body, 10012)
        rating_ok = ("3.20" in body) or ("3.2" in body) or ("3,20" in body) or ("3,2" in body)
        body_ok = top_ok and low_ok and rating_ok
        check("Email body contains real totals (5008, 10012) and avg rating 3.20",
              body_ok,
              f"top={top_ok} low={low_ok} rating={rating_ok}")
        result.update(from_ok=from_ok, to_ok=to_ok, body_ok=body_ok)
    return result


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    excel_ok = check_excel(args.agent_workspace, gt_dir)
    gcal_res = check_gcal(args.launch_time)
    email_res = check_email()

    # --- CRITICAL CHECKS (semantic substance). Any failure => hard FAIL (sys.exit(1)). ---
    print("\n=== CRITICAL CHECKS ===")
    crit = critical_checks(args.agent_workspace, gcal_res, email_res)
    critical_failed = []
    for name, ok in crit:
        status = "PASS" if ok else "FAIL"
        print(f"  [CRITICAL {status}] {name}")
        if not ok:
            critical_failed.append(name)

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")

    if critical_failed:
        print(f"  Overall: FAIL (critical checks failed: {critical_failed})")
        sys.exit(1)

    overall = excel_ok and FAIL_COUNT == 0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")

    sys.exit(0 if overall else 1)


def _load_excel_lookup(agent_workspace):
    """Return (top_lookup, low_lookup, summary) dicts keyed on Russian dept name, or None."""
    agent_file = os.path.join(agent_workspace, "Performance_Review_Summary.xlsx")
    if not os.path.isfile(agent_file):
        return None, None, None
    try:
        wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception:
        return None, None, None
    top = get_sheet(wb, "Top_Performers")
    low = get_sheet(wb, "Underperformers")
    summ = get_sheet(wb, "Summary")
    top_lk = {}
    if top is not None:
        for r in top.iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                top_lk[str(r[0]).strip().lower()] = r
    low_lk = {}
    if low is not None:
        for r in low.iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                low_lk[str(r[0]).strip().lower()] = r
    summary = {}
    if summ is not None:
        for r in summ.iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                summary[str(r[0]).strip().lower()] = r[1]
    return top_lk, low_lk, summary


def critical_checks(agent_workspace, gcal_res, email_res):
    """List of (name, bool) — semantic core deliverables keyed on Russian dept names."""
    checks = []
    top_lk, low_lk, summary = _load_excel_lookup(agent_workspace)

    # 1. Top_Performers: all 7 Russian-keyed depts present with correct Count_Rating_5.
    if top_lk is None:
        checks.append(("Top_Performers 7 RU-dept Count_Rating_5 match groundtruth", False))
    else:
        ok = True
        for dept, cnt, _sal, _exp in TOP_PERFORMERS:
            row = top_lk.get(dept.lower())
            if row is None or not num_close(row[1], cnt, 5):
                ok = False
                break
        checks.append(("Top_Performers 7 RU-dept Count_Rating_5 match groundtruth", ok))

    # 2. Underperformers: all 7 Russian-keyed depts present with correct Count_Low_Rating.
    if low_lk is None:
        checks.append(("Underperformers 7 RU-dept Count_Low_Rating match groundtruth", False))
    else:
        ok = True
        for dept, cnt, _sal, _exp in UNDERPERFORMERS:
            row = low_lk.get(dept.lower())
            if row is None or not num_close(row[1], cnt, 10):
                ok = False
                break
        checks.append(("Underperformers 7 RU-dept Count_Low_Rating match groundtruth", ok))

    # 3. Summary core totals: 5008 / 10012 / 3.20.
    if summary is None:
        checks.append(("Summary totals (5008/10012/3.20) correct", False))
    else:
        ok = (num_close(summary.get("total_top_performers"), 5008, 20)
              and num_close(summary.get("total_underperformers"), 10012, 40)
              and num_close(summary.get("overall_avg_rating"), 3.20, 0.1))
        checks.append(("Summary totals (5008/10012/3.20) correct", ok))

    # 4. Board meeting event exists AND is ~21 days from launch.
    checks.append(("Board meeting event exists & ~21 days from launch",
                   bool(gcal_res.get("event_ok")) and bool(gcal_res.get("date_ok"))))

    # 5. Email hr->executives with correct subject and real numeric totals in body.
    checks.append(("Email hr->executives with real totals (5008/10012) and avg 3.20 in body",
                   bool(email_res.get("email_ok")) and bool(email_res.get("from_ok"))
                   and bool(email_res.get("to_ok")) and bool(email_res.get("body_ok"))))

    return checks


if __name__ == "__main__":
    main()
