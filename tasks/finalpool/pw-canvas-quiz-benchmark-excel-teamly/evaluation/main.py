"""Evaluation script for pw-canvas-quiz-benchmark-excel-teamly.

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

Canvas course data is LIVE, so the eval never hardcodes course values. Instead
it verifies internal consistency (Metrics recomputed from Data_Analysis rows)
and the analysis-guide rule (top recommendation targets the lowest pass rate).
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Quiz_Benchmark_Report.xlsx exists",
    "Data_Analysis has required columns and >= 6 course rows",
    "Metrics consistent with Data_Analysis (counts/sums/averages)",
    "Recommendations top priority targets lowest pass-rate course",
    "Teamly 'Canvas Quiz Dashboard' page exists (not the noise page)",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def get_sheet(wb, name):
    target = name.strip().lower().replace(" ", "_")
    for n in wb.sheetnames:
        if n.strip().lower().replace(" ", "_") == target:
            return wb[n]
    return None


def header_map(ws):
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return headers, {h: i for i, h in enumerate(headers)}


def data_rows_of(ws):
    return [r for r in ws.iter_rows(min_row=2, values_only=True)
            if any(c is not None for c in r)]


def check_data_analysis(wb):
    """Non-critical structural checks + critical column/row-count check."""
    ws = get_sheet(wb, "Data_Analysis")
    check("Data_Analysis sheet exists", ws is not None)
    if ws is None:
        check("Data_Analysis has required columns and >= 6 course rows", False, "no sheet")
        return None
    headers, hmap = header_map(ws)
    rows = data_rows_of(ws)
    check("Data_Analysis has >= 6 rows", len(rows) >= 6, f"got {len(rows)}")
    required = ['course', 'code', 'enrollment', 'avg_score', 'pass_rate']
    for col in required:
        check(f"Data_Analysis has {col} column", col in headers, f"headers: {headers[:8]}")
    cols_ok = all(c in headers for c in required)
    check("Data_Analysis has required columns and >= 6 course rows",
          cols_ok and len(rows) >= 6, f"cols_ok={cols_ok}, rows={len(rows)}")
    if not cols_ok:
        return None
    # Sorted alphabetically by Course (non-critical).
    ci = hmap['course']
    names = [str(r[ci]).strip() for r in rows if ci < len(r) and r[ci] is not None]
    sorted_ok = names == sorted(names, key=lambda s: s.lower())
    check("Data_Analysis sorted alphabetically by Course", sorted_ok, f"{names[:4]}")
    return ws


def check_metrics_consistency(wb, da_ws):
    """Critical: Metrics values recomputed from Data_Analysis rows."""
    ws = get_sheet(wb, "Metrics")
    check("Metrics sheet exists", ws is not None)
    if ws is None:
        check("Metrics consistent with Data_Analysis (counts/sums/averages)", False, "no sheet")
        return
    headers, _ = header_map(ws)
    rows = data_rows_of(ws)
    check("Metrics has >= 4 rows", len(rows) >= 4, f"got {len(rows)}")
    for col in ['metric', 'value']:
        check(f"Metrics has {col} column", col in headers, f"headers: {headers[:8]}")

    if da_ws is None or 'metric' not in headers or 'value' not in headers:
        check("Metrics consistent with Data_Analysis (counts/sums/averages)", False,
              "missing prerequisites")
        return

    # Build metric -> value map (key normalised).
    mvals = {}
    for r in rows:
        if len(r) >= 2 and r[0] is not None:
            mvals[str(r[0]).strip().lower()] = safe_float(r[1])

    # Recompute from Data_Analysis.
    _, dh = header_map(da_ws)
    da_rows = data_rows_of(da_ws)
    enr_i, pr_i = dh.get('enrollment'), dh.get('pass_rate')
    enrolls = [safe_float(r[enr_i]) for r in da_rows if enr_i is not None and enr_i < len(r)]
    enrolls = [e for e in enrolls if e is not None]
    prates = [safe_float(r[pr_i]) for r in da_rows if pr_i is not None and pr_i < len(r)]
    prates = [p for p in prates if p is not None]

    exp_courses = len(da_rows)
    exp_enroll = sum(enrolls) if enrolls else None
    exp_avg_pr = (sum(prates) / len(prates)) if prates else None

    def close(a, b, tol):
        return a is not None and b is not None and abs(a - b) <= tol

    courses_ok = close(mvals.get('total_courses'), exp_courses, 0.5)
    enroll_ok = close(mvals.get('total_enrollment'), exp_enroll, max(1.0, 0.01 * (exp_enroll or 0)))
    avgpr_ok = close(mvals.get('avg_pass_rate'), exp_avg_pr, 1.5)

    check("Metrics Total_Courses == Data_Analysis row count", courses_ok,
          f"metrics={mvals.get('total_courses')} expected={exp_courses}")
    check("Metrics Total_Enrollment == sum(Enrollment)", enroll_ok,
          f"metrics={mvals.get('total_enrollment')} expected={exp_enroll}")
    check("Metrics Avg_Pass_Rate == mean(Pass_Rate)", avgpr_ok,
          f"metrics={mvals.get('avg_pass_rate')} expected={exp_avg_pr}")

    check("Metrics consistent with Data_Analysis (counts/sums/averages)",
          courses_ok and enroll_ok and avgpr_ok,
          f"courses={courses_ok} enroll={enroll_ok} avgpr={avgpr_ok}")


def check_recommendations(wb, da_ws):
    """Critical: top-priority recommendation targets the lowest pass-rate course."""
    ws = get_sheet(wb, "Recommendations")
    check("Recommendations sheet exists", ws is not None)
    if ws is None:
        check("Recommendations top priority targets lowest pass-rate course", False, "no sheet")
        return
    headers, hmap = header_map(ws)
    rows = data_rows_of(ws)
    check("Recommendations has >= 2 rows", len(rows) >= 2, f"got {len(rows)}")
    for col in ['priority', 'action', 'course']:
        check(f"Recommendations has {col} column", col in headers, f"headers: {headers[:8]}")

    if da_ws is None or 'priority' not in headers or 'course' not in headers:
        check("Recommendations top priority targets lowest pass-rate course", False,
              "missing prerequisites")
        return

    # Determine lowest pass-rate course from Data_Analysis.
    _, dh = header_map(da_ws)
    da_rows = data_rows_of(da_ws)
    c_i, pr_i = dh.get('course'), dh.get('pass_rate')
    worst_course = None
    worst_pr = None
    if c_i is not None and pr_i is not None:
        for r in da_rows:
            pr = safe_float(r[pr_i]) if pr_i < len(r) else None
            if pr is None:
                continue
            if worst_pr is None or pr < worst_pr:
                worst_pr = pr
                worst_course = str(r[c_i]).strip() if c_i < len(r) and r[c_i] is not None else None

    # Find the top-priority recommendation row (smallest numeric Priority).
    p_i, rc_i = hmap['priority'], hmap['course']
    top_course = None
    top_prio = None
    for r in rows:
        prio = safe_float(r[p_i]) if p_i < len(r) else None
        if prio is None:
            continue
        if top_prio is None or prio < top_prio:
            top_prio = prio
            top_course = str(r[rc_i]).strip() if rc_i < len(r) and r[rc_i] is not None else None

    def norm(s):
        return (s or "").strip().lower()

    ok = (worst_course is not None and top_course is not None
          and (norm(worst_course) in norm(top_course) or norm(top_course) in norm(worst_course)))
    check("Recommendations top priority targets lowest pass-rate course", ok,
          f"worst={worst_course}({worst_pr}) top_rec={top_course}")


def check_excel(agent_workspace):
    excel_path = os.path.join(agent_workspace, "Quiz_Benchmark_Report.xlsx")
    check("Quiz_Benchmark_Report.xlsx exists", os.path.exists(excel_path))
    if not os.path.exists(excel_path):
        check("Data_Analysis has required columns and >= 6 course rows", False, "no excel")
        check("Metrics consistent with Data_Analysis (counts/sums/averages)", False, "no excel")
        check("Recommendations top priority targets lowest pass-rate course", False, "no excel")
        return
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    da_ws = check_data_analysis(wb)
    check_metrics_consistency(wb, da_ws)
    check_recommendations(wb, da_ws)


def check_script(agent_workspace):
    check("course_quiz_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "course_quiz_processor.py")))


def check_teamly():
    """Critical: a 'Canvas Quiz Dashboard' page exists with a non-empty summary.

    Seed pages have id <= 3; the noise page ('Старые заметки по проекту') is a
    user-leftover and must NOT satisfy the dashboard check.
    """
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, '') FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        conn.close()
    except Exception as e:
        check("Teamly 'Canvas Quiz Dashboard' page exists (not the noise page)", False, str(e))
        return

    dashboard = None
    for pid, title, body in pages:
        tl = (title or "").lower()
        if "старые заметки" in tl:
            continue
        if "dashboard" in tl or "канвас" in tl or ("quiz" in tl and "canvas" in tl) \
                or "дашборд" in tl or "панель" in tl:
            dashboard = (pid, title, body)
            break
    check("Teamly 'Canvas Quiz Dashboard' page exists (not the noise page)",
          dashboard is not None, f"new pages: {[(p[0], p[1]) for p in pages]}")

    # Non-empty summary body (non-critical).
    if dashboard is not None:
        body = dashboard[2] or ""
        check("Teamly dashboard page has a non-empty summary body",
              len(body.strip()) >= 20, f"body len: {len(body.strip())}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    check_excel(agent_workspace)
    check_script(agent_workspace)
    check_teamly()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT, "total_checks": total,
                    "accuracy": accuracy, "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
