"""Evaluation for terminal-arxiv-scholarly-teamly-word-excel.

Structural checks (sheet exists, row counts, substring presence) are
NON-critical. A small set of SEMANTIC checks (CRITICAL_CHECKS) validates the
task's substance: the 6 transformer papers are present and marked as overlap,
the keyword-anchored categorization is correct, the Word doc is a real review,
and the Teamly tracker carries per-paper Category/Source. Any critical failure
=> overall FAIL regardless of accuracy. Otherwise PASS requires accuracy >= 70%.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

TRANSFORMER_IDS = {"1706.03762", "1810.04805", "2005.14165",
                   "1409.0473", "1910.10683", "2009.06732"}
NOISE_IDS = {"1207.00580", "1502.03167", "1312.06199"}
CATEGORIES = ["Architecture Design", "Training Methods", "Applications", "Survey"]

# Semantic checks. Any failure here => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "All 6 transformer papers in Paper_Catalog",
    "6 transformer papers marked as overlap in Citation_Matrix",
    "Attention Is All You Need categorized as Architecture Design",
    "Efficient Transformers categorized as Survey",
    "Word document is a substantive review (>=500 words, >=4 paper titles)",
    "Teamly tracker has >=6 paper pages with correct Category/Source",
    "Method_Comparison Paper_Count sums to catalog size",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        d = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{d}")


def _norm_cat(v):
    return str(v).strip() if v else ""


def check_excel(ws_path):
    """Check Research_Paper_Analysis.xlsx."""
    print("\n=== Checking Excel ===")
    path = os.path.join(ws_path, "Research_Paper_Analysis.xlsx")
    if not os.path.isfile(path):
        check("Excel file exists", False, f"Not found: {path}")
        return
    check("Excel file exists", True)

    wb = openpyxl.load_workbook(path, data_only=True)
    sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}

    catalog_id_to_cat = {}
    catalog_size = 0

    # Paper_Catalog
    pc_name = sn.get("paper_catalog")
    if pc_name is None:
        check("Paper_Catalog sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Paper_Catalog sheet exists", True)
        ws = wb[pc_name]
        rows = list(ws.iter_rows(values_only=True))
        data = [r for r in rows[1:] if r and r[0] is not None]
        catalog_size = len(data)
        check("Paper_Catalog has 9 rows", len(data) == 9, f"Found {len(data)}")

        # Map Paper_ID -> Category (col0 = Paper_ID, col5 = Category)
        for r in data:
            pid = str(r[0]).strip()
            cat = _norm_cat(r[5]) if len(r) > 5 else ""
            catalog_id_to_cat[pid] = cat

        ids_found = set(catalog_id_to_cat.keys())
        transformer_found = len(TRANSFORMER_IDS & ids_found)
        # CRITICAL: all 6 transformer papers present.
        check("All 6 transformer papers in Paper_Catalog",
              transformer_found == 6, f"Found {transformer_found}/6")

        cats_found = {c for c in catalog_id_to_cat.values()}
        valid_cats = sum(1 for c in cats_found if c in CATEGORIES)
        check("Valid categories assigned", valid_cats >= 2,
              f"Categories found: {cats_found}")

        # CRITICAL: keyword-anchored categorization is correct.
        check("Attention Is All You Need categorized as Architecture Design",
              catalog_id_to_cat.get("1706.03762") == "Architecture Design",
              f"Got {catalog_id_to_cat.get('1706.03762')!r}")
        check("Efficient Transformers categorized as Survey",
              catalog_id_to_cat.get("2009.06732") == "Survey",
              f"Got {catalog_id_to_cat.get('2009.06732')!r}")

    # Method_Comparison
    mc_name = sn.get("method_comparison")
    if mc_name is None:
        check("Method_Comparison sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Method_Comparison sheet exists", True)
        ws2 = wb[mc_name]
        rows2 = list(ws2.iter_rows(values_only=True))
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        check("Method_Comparison has 4 category rows", len(data2) == 4,
              f"Found {len(data2)}")

        # CRITICAL: aggregation reflects the catalog, not hardcoded numbers.
        # Paper_Count is col index 1. Sum must equal the catalog size (9 if all
        # papers catalogued, or 6 if scoped to transformer-only).
        total = 0
        ok_counts = True
        for r in data2:
            try:
                total += int(r[1])
            except (TypeError, ValueError, IndexError):
                ok_counts = False
        check("Method_Comparison Paper_Count sums to catalog size",
              ok_counts and catalog_size > 0 and total in (catalog_size, 6, 9),
              f"Sum={total}, catalog_size={catalog_size}")

    # Citation_Matrix
    cm_name = sn.get("citation_matrix")
    if cm_name is None:
        check("Citation_Matrix sheet exists", False, f"Sheets: {wb.sheetnames}")
    else:
        check("Citation_Matrix sheet exists", True)
        ws3 = wb[cm_name]
        rows3 = list(ws3.iter_rows(values_only=True))
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        check("Citation_Matrix has 9 rows", len(data3) == 9, f"Found {len(data3)}")

        # CRITICAL: the 6 transformer papers (and only those) are overlap=Yes.
        overlap_yes = 0
        noise_overlap_wrong = 0
        for r in data3:
            pid = str(r[0]).strip()
            overlap_val = str(r[4]).strip().lower() if len(r) > 4 and r[4] else ""
            is_yes = "yes" in overlap_val or overlap_val in ("да", "true", "1")
            if pid in TRANSFORMER_IDS and is_yes:
                overlap_yes += 1
            if pid in NOISE_IDS and is_yes:
                noise_overlap_wrong += 1
        check("6 transformer papers marked as overlap in Citation_Matrix",
              overlap_yes == 6 and noise_overlap_wrong == 0,
              f"transformer overlap={overlap_yes}/6, noise wrongly overlapped={noise_overlap_wrong}")

    wb.close()


def check_word(ws_path):
    """Check Transformer_Literature_Review.docx — substance, not just length."""
    print("\n=== Checking Word Document ===")
    path = os.path.join(ws_path, "Transformer_Literature_Review.docx")
    if not os.path.isfile(path):
        check("Word document exists", False, f"Not found: {path}")
        return
    check("Word document exists", True)

    from docx import Document
    doc = Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs)
    low = full_text.lower()

    check("Document mentions transformer", "transformer" in low)
    check("Document mentions attention", "attention" in low)
    check("Document mentions BERT or pre-training",
          "bert" in low or "pre-train" in low or "pretrain" in low)
    # conclusion/summary OR Russian equivalents (вывод/заключение/итог)
    check("Document has conclusion section",
          any(k in low for k in ["conclusion", "summary", "вывод", "заключени", "итог"]))

    # Exact review title present.
    has_title = "literature review: transformer architectures in nlp" in low
    check("Document has the review title", has_title)

    # CRITICAL: substantive review — >=500 words AND names >=4 of 6 paper titles.
    word_count = len(full_text.split())
    paper_titles = [
        "attention is all you need",
        "bert",
        "language models are few-shot learners",
        "neural machine translation by jointly learning to align and translate",
        "exploring the limits of transfer learning",
        "efficient transformers",
    ]
    titles_found = sum(1 for t in paper_titles if t in low)
    has_gaps = any(k in low for k in ["gap", "future", "пробел", "направлен", "будущ"])
    check("Word document is a substantive review (>=500 words, >=4 paper titles)",
          word_count >= 500 and titles_found >= 4 and has_title and has_gaps,
          f"words={word_count}, titles_found={titles_found}/6, title={has_title}, gaps={has_gaps}")


def check_teamly():
    """Check the Teamly 'Research Paper Tracker' space and its paper pages."""
    print("\n=== Checking Teamly 'Research Paper Tracker' ===")
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # Locate the tracker space.
    cur.execute("SELECT id, key, name FROM teamly.spaces")
    spaces = cur.fetchall()
    tracker_space_ids = [
        sid for sid, key, name in spaces
        if (("research" in (name or "").lower() and "paper" in (name or "").lower())
            or "tracker" in (name or "").lower()
            or "tracker" in (key or "").lower())
    ]

    # Gather candidate paper pages: prefer pages in the tracker space; otherwise
    # fall back to any page whose title/body references one of the paper ids.
    if tracker_space_ids:
        cur.execute(
            "SELECT title, COALESCE(body,'') FROM teamly.pages WHERE space_id = ANY(%s)",
            (tracker_space_ids,),
        )
    else:
        cur.execute("SELECT title, COALESCE(body,'') FROM teamly.pages")
    pages = cur.fetchall()
    conn.close()

    check("Research Paper Tracker space exists", bool(tracker_space_ids),
          f"Spaces: {[(k, n) for _, k, n in spaces]}")

    # A page is a "paper page" if it references a paper id OR a known title.
    title_anchors = {
        "1706.03762": "attention is all you need",
        "1810.04805": "bert",
        "2005.14165": "few-shot",
        "1409.0473": "align and translate",
        "1910.10683": "text-to-text",
        "2009.06732": "efficient transformers",
    }
    expected_cat = {
        "1706.03762": "architecture design",
        "1810.04805": "training methods",
        "2005.14165": "training methods",
        "1409.0473": "applications",
        "1910.10683": "architecture design",
        "2009.06732": "survey",
    }

    # Category is strictly required only for the keyword-anchored papers
    # (same anchors as the Excel criticals); other categories stay non-critical
    # because keyword scoring on the remaining abstracts is not fully decisive.
    anchored_ids = {"1706.03762", "2009.06732"}

    matched_ids = set()
    src_ok_ids = set()
    cat_ok_ids = set()
    for title, body in pages:
        blob = (str(title) + " " + str(body)).lower()
        for pid, anchor in title_anchors.items():
            if pid in blob or anchor in blob:
                matched_ids.add(pid)
                if expected_cat[pid] in blob:
                    cat_ok_ids.add(pid)
                if "both" in blob:
                    src_ok_ids.add(pid)
                break

    check("Teamly has >= 6 paper pages", len(matched_ids) >= 6,
          f"Matched paper ids: {sorted(matched_ids)} across {len(pages)} pages")

    # CRITICAL: all 6 paper pages carry Source=Both, and the two
    # keyword-anchored papers carry the correct Category.
    check("Teamly tracker has >=6 paper pages with correct Category/Source",
          len(src_ok_ids) >= 6 and anchored_ids <= cat_ok_ids,
          f"Source=Both on {len(src_ok_ids)}/6 pages, "
          f"anchored Category ok: {sorted(anchored_ids & cat_ok_ids)}")

    # Non-critical: full expected Category mapping across all 6 pages.
    check("Teamly pages carry expected Category on all 6 papers",
          len(cat_ok_ids) >= 6,
          f"Correct Category on {sorted(cat_ok_ids)}")


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    excel_path = os.path.join(workspace, "Research_Paper_Analysis.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            valid_keywords = ["paper", "catalog", "method", "comparison", "citation", "matrix"]
            unexpected = [s for s in wb.sheetnames
                          if not any(k in s.lower() for k in valid_keywords) and s.lower() != "sheet1"]
            check("No unexpected sheets in Excel", len(unexpected) == 0,
                  f"Unexpected sheets: {unexpected}")
            sn = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}
            pc_name = sn.get("paper_catalog")
            if pc_name:
                ws = wb[pc_name]
                rows = list(ws.iter_rows(values_only=True))
                ids = [str(r[0]).strip() for r in rows[1:] if r and r[0]]
                check("No duplicate paper IDs in Paper_Catalog",
                      len(ids) == len(set(ids)),
                      f"Found {len(ids)} IDs but {len(set(ids))} unique")
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    print("=" * 70)
    print("TERMINAL-ARXIV-SCHOLARLY-TEAMLY-WORD-EXCEL - EVALUATION")
    print("=" * 70)

    check_excel(args.agent_workspace)
    check_word(args.agent_workspace)
    check_teamly()
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"  CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"    - {n}")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failed": critical_failed,
    }
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if critical_failed:
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("  Overall: PASS")
        sys.exit(0)
    print("  Overall: FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
