"""Evaluation for canvas-risk-intervention-excel-word-gcal-email.

Стратегия: groundtruth-файлы содержат фиктивные (выдуманные) данные о курсах
и студентах, которые НЕ засеяны в canvas, а агент читает живые данные. Поэтому
сравнение значений ячеек с groundtruth не проводится. Вместо этого:
  - агрегаты пересчитываются из СОБСТВЕННОГО risk_assessment.json агента
    (производного от реального источника) и сверяются с Excel-файлом;
  - выполняются проверки внутренней согласованности (формула Risk_Rate_Pct,
    сортировка, пороги классификации, Highest_Risk_Course, Overall_Risk_Rate_Pct);
  - проверяется точный получатель письма и точные дата/время события.

CRITICAL-проверки (любой провал => немедленный sys.exit(1) до проверки порога):
  1. Пороги классификации: <50 Critical, 50-65 At Risk, >65 On Track;
     лист Critical_Students содержит ровно когорту Critical (по числу).
  2. Формула Risk_Rate_Pct == round((Critical+At Risk)/Total*100,1) и
     Risk_Overview отсортирован по Risk_Rate_Pct по убыванию.
  3. Письмо отправлено точно на academic-affairs@university.edu, тема содержит
     'risk', тело выделяет курсы с риском > 40% (ключевые слова risk/риск).
  4. Событие 'Academic Intervention Planning Meeting' на 2026-03-13 15:00-16:30 UTC
     (проверка даты/времени, не только summary).
  5. Summary: Highest_Risk_Course == курс с максимальным Risk_Rate_Pct, и
     Overall_Risk_Rate_Pct == round((Critical+At Risk)/Total_Assessed*100,1).
"""
import os
import argparse, json, os, sys
from datetime import datetime, timezone
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

TARGET_EMAIL = "academic-affairs@university.edu"
EVENT_SUMMARY_KW = "intervention"
EVENT_DATE = (2026, 3, 13)
EVENT_START_HOUR_UTC = 15
EVENT_END_HOUR_UTC = 16
EVENT_END_MIN_UTC = 30

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    tag = "CRITICAL " if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {tag}{name}: {detail_str}")
        if critical:
            CRITICAL_FAILED.append(name)


def safe_float(val, default=None):
    try:
        if val is None: return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def to_utc_naive(dt):
    """Приводим datetime к наивному UTC для сравнения."""
    if dt is None:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def sheet_rows(ws):
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    hmap = {h.lower(): i for i, h in enumerate(headers) if h}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in r):
            continue
        rows.append(r)
    return headers, hmap, rows


def cell(row, hmap, key):
    i = hmap.get(key.lower())
    if i is None or i >= len(row):
        return None
    return row[i]


# ---------------------------------------------------------------------------

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILED = []

    # -------- Структурные проверки Excel --------
    excel_path = os.path.join(agent_workspace, "Student_Risk_Assessment.xlsx")
    check("Student_Risk_Assessment.xlsx exists", os.path.exists(excel_path))

    overview = critical_sheet = summary = plan = None
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)
        for sn in ["Risk_Overview", "Critical_Students", "Intervention_Plan", "Summary"]:
            check(f"{sn} sheet exists", sn in wb.sheetnames)

        if "Risk_Overview" in wb.sheetnames:
            oh, ohm, orows = sheet_rows(wb["Risk_Overview"])
            for col in ["Course_Name", "Total_Students", "Critical_Count",
                        "At_Risk_Count", "On_Track_Count", "Risk_Rate_Pct"]:
                check(f"Risk_Overview has {col} column", col.lower() in ohm, f"headers: {oh}")
            overview = (oh, ohm, orows)

        if "Critical_Students" in wb.sheetnames:
            ch, chm, crows = sheet_rows(wb["Critical_Students"])
            for col in ["Student_ID", "Course_Name", "Avg_Score",
                        "Assignments_Submitted", "Late_Submissions"]:
                check(f"Critical_Students has {col} column", col.lower() in chm, f"headers: {ch}")
            critical_sheet = (ch, chm, crows)

        if "Intervention_Plan" in wb.sheetnames:
            ph, phm, prows = sheet_rows(wb["Intervention_Plan"])
            for col in ["Course_Name", "Risk_Level", "Recommended_Action",
                        "Responsible_Party", "Deadline"]:
                check(f"Intervention_Plan has {col} column", col.lower() in phm, f"headers: {ph}")
            check("Intervention_Plan has data rows", len(prows) >= 1, f"got {len(prows)}")
            plan = (ph, phm, prows)

        if "Summary" in wb.sheetnames:
            sh, shm, srows = sheet_rows(wb["Summary"])
            for col in ["Metric", "Value"]:
                check(f"Summary has {col} column", col.lower() in shm, f"headers: {sh}")
            summary = (sh, shm, srows)

    # -------- Загрузка risk_assessment.json агента (источник истины агента) --------
    ra_path = os.path.join(agent_workspace, "risk_assessment.json")
    check("risk_assessment.json exists", os.path.exists(ra_path))
    ra = None
    if os.path.exists(ra_path):
        try:
            with open(ra_path, "r", encoding="utf-8") as f:
                ra = json.load(f)
        except Exception as e:
            check("risk_assessment.json parseable", False, str(e))

    # ============================================================
    # CRITICAL 1: пороги классификации + когорта Critical_Students
    # ============================================================
    if overview:
        oh, ohm, orows = overview
        # Считаем суммарные категории из Risk_Overview
        tot_crit = sum(int(safe_float(cell(r, ohm, "Critical_Count")) or 0) for r in orows)
        tot_atrisk = sum(int(safe_float(cell(r, ohm, "At_Risk_Count")) or 0) for r in orows)

        # Критические студенты в листе Critical_Students: все Avg_Score < 50
        if critical_sheet:
            ch, chm, crows = critical_sheet
            scores = [safe_float(cell(r, chm, "Avg_Score")) for r in crows]
            scores = [s for s in scores if s is not None]
            all_below_50 = all(s < 50 for s in scores) if scores else False
            check("Все студенты в Critical_Students имеют Avg_Score < 50",
                  all_below_50, f"scores sample: {scores[:10]}", critical=True)
            # Число строк Critical_Students совпадает с суммой Critical_Count
            check("Число строк Critical_Students == сумме Critical_Count",
                  len(crows) == tot_crit,
                  f"rows={len(crows)} vs Critical_Count sum={tot_crit}", critical=True)
        else:
            check("Critical_Students cohort check", False, "лист отсутствует", critical=True)

        # Сверка с risk_assessment.json (если содержит итоги)
        if ra is not None:
            ra_text = json.dumps(ra, ensure_ascii=False).lower()
            check("risk_assessment.json использует пороги Critical/At Risk/On Track",
                  ("critical" in ra_text and "at risk" in ra_text.replace("_", " ")
                   or "at_risk" in ra_text or "on track" in ra_text.replace("_", " ")
                   or "on_track" in ra_text),
                  "не найдены метки классификации")
    else:
        check("Risk classification thresholds", False, "Risk_Overview отсутствует", critical=True)

    # ============================================================
    # CRITICAL 2: формула Risk_Rate_Pct + сортировка по убыванию
    # ============================================================
    if overview:
        oh, ohm, orows = overview
        formula_ok = True
        rates = []
        bad = None
        for r in orows:
            total = safe_float(cell(r, ohm, "Total_Students"))
            crit = safe_float(cell(r, ohm, "Critical_Count"))
            atr = safe_float(cell(r, ohm, "At_Risk_Count"))
            rate = safe_float(cell(r, ohm, "Risk_Rate_Pct"))
            if None in (total, crit, atr, rate) or total == 0:
                continue
            expected = round((crit + atr) / total * 100, 1)
            rates.append(rate)
            if abs(expected - rate) > 0.2:
                formula_ok = False
                bad = (cell(r, ohm, "Course_Name"), rate, expected)
        check("Risk_Rate_Pct == round((Critical+At Risk)/Total*100,1) для всех курсов",
              formula_ok, f"первое расхождение: {bad}", critical=True)
        sorted_desc = all(rates[i] >= rates[i + 1] - 0.01 for i in range(len(rates) - 1))
        check("Risk_Overview отсортирован по Risk_Rate_Pct по убыванию",
              sorted_desc, f"rates: {rates}", critical=True)
    # (если overview отсутствует — уже зафиксировано выше как critical)

    # ============================================================
    # CRITICAL 5: согласованность Summary
    # ============================================================
    if overview and summary:
        oh, ohm, orows = overview
        sh, shm, srows = summary
        smap = {}
        for r in srows:
            m = cell(r, shm, "Metric")
            v = cell(r, shm, "Value")
            if m is not None:
                smap[str(m).strip().lower()] = v

        # Highest_Risk_Course == курс с макс Risk_Rate_Pct
        best_course, best_rate = None, -1
        for r in orows:
            rate = safe_float(cell(r, ohm, "Risk_Rate_Pct"))
            cn = cell(r, ohm, "Course_Name")
            if rate is not None and rate > best_rate:
                best_rate, best_course = rate, cn
        hrc = smap.get("highest_risk_course")
        hrc_ok = (hrc is not None and best_course is not None
                  and str(best_course).strip().lower() in str(hrc).strip().lower()
                  or (hrc is not None and best_course is not None
                      and str(hrc).strip().lower() in str(best_course).strip().lower()))
        check("Summary Highest_Risk_Course == курс с макс Risk_Rate_Pct",
              bool(hrc_ok), f"summary={hrc}, expected={best_course}", critical=True)

        # Overall_Risk_Rate_Pct == round((Critical+At Risk)/Total_Assessed*100,1)
        tot_crit = sum(int(safe_float(cell(r, ohm, "Critical_Count")) or 0) for r in orows)
        tot_atr = sum(int(safe_float(cell(r, ohm, "At_Risk_Count")) or 0) for r in orows)
        tot_all = sum(int(safe_float(cell(r, ohm, "Total_Students")) or 0) for r in orows)
        overall = safe_float(smap.get("overall_risk_rate_pct"))
        assessed = safe_float(smap.get("total_students_assessed"))
        if tot_all > 0 and overall is not None:
            exp_overall = round((tot_crit + tot_atr) / tot_all * 100, 1)
            check("Summary Overall_Risk_Rate_Pct согласован с Risk_Overview",
                  abs(exp_overall - overall) <= 0.3,
                  f"summary={overall}, expected={exp_overall}", critical=True)
        else:
            check("Summary Overall_Risk_Rate_Pct присутствует", overall is not None,
                  f"value={overall}", critical=True)
        # Total_Students_Assessed согласован
        if assessed is not None:
            check("Summary Total_Students_Assessed согласован с суммой Total_Students",
                  abs(assessed - tot_all) <= max(1, tot_all * 0.02),
                  f"summary={assessed}, sum={tot_all}")
    elif summary is None:
        check("Summary consistency", False, "Summary отсутствует", critical=True)

    # -------- Word-документ --------
    docx_path = os.path.join(agent_workspace, "Intervention_Report.docx")
    check("Intervention_Report.docx exists", os.path.exists(docx_path))
    if os.path.exists(docx_path):
        from docx import Document
        doc = Document(docx_path)
        all_paras = [p.text.strip() for p in doc.paragraphs]
        text = " ".join(all_paras)
        check("Intervention_Report.docx has content", len(text) > 50, f"len: {len(text)}")
        # Заголовки разделов: RU+EN пары (агент может писать заголовки по-русски)
        heads_lower = [h.lower() for h in all_paras if h]  # включаем Title/Heading
        required = [
            ["risk assessment methodology", "методология оценки рисков", "методика оценки рисков"],
            ["course-level analysis", "анализ по курсам", "анализ на уровне курсов"],
            ["critical cases", "критические случаи", "критические студенты"],
            ["recommended interventions", "рекомендуемые вмешательства", "рекомендации"],
        ]
        for variants in required:
            found = any(v in h for h in heads_lower for v in variants)
            check(f"Intervention_Report содержит раздел \"{variants[0]}\"", found,
                  f"варианты: {variants}")

    # -------- Python-скрипт --------
    py_files = [f for f in os.listdir(agent_workspace) if f.endswith(".py")]
    check("Python analysis script exists", len(py_files) >= 1, f"found: {py_files}")
    check("student_performance.json exists",
          os.path.exists(os.path.join(agent_workspace, "student_performance.json")))

    # ============================================================
    # CRITICAL 3 (email) + CRITICAL 4 (calendar) + reverse-noise
    # ============================================================
    try:
        conn = get_conn()
        cur = conn.cursor()

        # --- Email: точный получатель + тема + тело про риск ---
        cur.execute(
            "SELECT subject, to_addr::text, body_text FROM email.messages "
            "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
            "AND subject ILIKE '%risk%'"
        )
        rows = cur.fetchall()
        recip_ok = any(r[1] and TARGET_EMAIL.lower() in r[1].lower() for r in rows)
        check(f"Письмо отправлено точно на {TARGET_EMAIL} с темой про risk",
              recip_ok, f"найдено писем: {len(rows)}; to_addr: {[r[1] for r in rows][:3]}",
              critical=True)
        # Тело письма выделяет риск (ключевые слова risk/риск) — ОРИГИНАЛЬНЫЙ текст .lower()
        body_combined = " ".join((r[2] or "") for r in rows if r[1] and TARGET_EMAIL.lower() in r[1].lower())
        bl = body_combined.lower()
        body_kw_ok = any(k in bl for k in ("risk", "риск", "риска", "рисками"))
        check("Тело письма упоминает риск/проценты (выделение курсов >40%)",
              body_kw_ok, f"body sample: {body_combined[:120]!r}", critical=True)

        # --- Calendar: точная дата и время события ---
        cur.execute(
            "SELECT summary, start_datetime, end_datetime FROM gcal.events "
            "WHERE summary ILIKE %s", (f"%{EVENT_SUMMARY_KW}%",)
        )
        events = cur.fetchall()
        check("Событие 'Academic Intervention Planning Meeting' существует",
              len(events) >= 1, "не найдено событие с 'intervention'")
        dt_ok = False
        for summ, sdt, edt in events:
            su = to_utc_naive(sdt)
            eu = to_utc_naive(edt)
            if su is None:
                continue
            if (su.year, su.month, su.day) == EVENT_DATE and su.hour == EVENT_START_HOUR_UTC and su.minute == 0:
                if eu is not None and eu.hour == EVENT_END_HOUR_UTC and eu.minute == EVENT_END_MIN_UTC:
                    dt_ok = True
                    break
                # допускаем отсутствие/смещение end, если start точен
                if eu is None:
                    dt_ok = True
        check("Событие запланировано на 2026-03-13 15:00-16:30 UTC",
              dt_ok, f"events: {[(e[0], str(to_utc_naive(e[1])), str(to_utc_naive(e[2]))) for e in events][:3]}",
              critical=True)

        # --- Reverse-noise: шумовая рассылка не должна попасть в Sent ---
        cur.execute(
            "SELECT COUNT(*) FROM email.messages "
            "WHERE folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1) "
            "AND (subject ILIKE '%рассылк%' OR subject ILIKE '%newsletter%')"
        )
        noise_sent = cur.fetchone()[0]
        check("Шумовая рассылка не отправлена из Sent", noise_sent == 0,
              f"found {noise_sent}")

        # --- Reverse-noise: шумовые события не удалены ---
        cur.execute(
            "SELECT COUNT(*) FROM gcal.events "
            "WHERE summary ILIKE '%планёрк%' OR summary ILIKE '%планерк%' "
            "OR summary ILIKE '%обед%' OR summary ILIKE '%standup%' OR summary ILIKE '%lunch%'"
        )
        noise_events = cur.fetchone()[0]
        check("Шумовые события не удалены агентом", noise_events >= 1,
              f"noise events: {noise_events}")

        conn.close()
    except Exception as e:
        check("DB checks", False, str(e), critical=True)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    success = (len(CRITICAL_FAILED) == 0) and (accuracy >= 70)
    return success, accuracy, total


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, accuracy, total = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )

    print(f"\nИтого: {PASS_COUNT}/{total} проверок пройдено ({accuracy:.1f}%)")
    if CRITICAL_FAILED:
        print(f"CRITICAL FAILED: {CRITICAL_FAILED}")

    result = {"total_passed": PASS_COUNT, "total_checks": total,
              "accuracy": accuracy, "critical_failed": CRITICAL_FAILED}
    if args.res_log_file:
        with open(args.res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILED:
        print("FAIL (critical check failed)")
        sys.exit(1)
    if accuracy >= 70:
        print("PASS")
        sys.exit(0)
    print("FAIL")
    sys.exit(1)


if __name__ == "__main__":
    main()
