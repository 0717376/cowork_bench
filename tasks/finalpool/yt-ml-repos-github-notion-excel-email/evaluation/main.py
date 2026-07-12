"""
Evaluation для yt-ml-repos-github-notion-excel-email (русифицировано, notion -> teamly).

Проверки:
1. ML_Research_Tracker.xlsx существует и читается.
2. Лист Videos: >=5 строк, есть столбцы Video_ID/Title, присутствуют реальные ID видео.
3. Лист Papers: >=3 строк, есть столбцы ArXiv_ID/Title, присутствуют реальные arXiv ID.
4. Лист Summary: >=3 строк.
5. Сверка значений с groundtruth XLSX (число строк + значения первых строк).
6. Teamly: страница «ML Tech Research Hub» создана, тело содержит реальные ссылки
   на GitHub-репозитории и ID видео (доказательство фактической каталогизации,
   а не предзаготовленного ответа).
7. Email: письмо на research@lab.edu отправлено (не исходное письмо PI из INBOX).

КРИТИЧЕСКИЕ чеки (CRITICAL_CHECKS): любой их провал => FAIL независимо от accuracy.
Порог: accuracy >= 70% И нет критических провалов => PASS.

ID видео и arXiv — латиница/цифры, normalize не требуется.
"""
import os
import sys
import json
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

# Реальные данные плейлиста ML Tech Reviews (из preprocess) — для семантической сверки.
REAL_VIDEO_IDS = [
    "mlu7idceolY", "dA-NhSBt4To", "J87hQFtSmas",
    "9vM4p9NN0Ts", "izrG86jG1Xk", "wh3uuJTK9O0", "klTvEwg3oJ4",
]
REAL_ARXIV_IDS = ["2307.08691", "2106.09685", "2310.06825", "2204.05149", "2208.07339"]
REAL_REPO_FRAGMENTS = [
    "Dao-AILab/flash-attention", "microsoft/LoRA", "CompVis/stable-diffusion",
    "openai/lm-human-preferences", "mistralai/mistral-src",
    "TimDettmers/bitsandbytes", "pgvector/pgvector",
]

CRITICAL_CHECKS = {
    "CRITICAL Videos: >=5 строк и >=5 реальных ID видео в листе",
    "CRITICAL Papers: >=3 строк и >=3 реальных arXiv ID в листе",
    "CRITICAL Teamly: страница-хаб с >=4 реальными ссылками на репозитории",
    "CRITICAL Email: письмо на research@lab.edu отправлено (не от pi@lab.edu)",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILED
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {name}{msg}")
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILED.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _cells_blob(ws):
    """Все непустые значения листа в одну строку (для подсчёта реальных ID)."""
    parts = []
    for row in ws.iter_rows(values_only=True):
        for c in row:
            if c is not None:
                parts.append(str(c))
    return "\n".join(parts)


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-5: ML_Research_Tracker.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "ML_Research_Tracker.xlsx")
    if not os.path.exists(xlsx_path):
        record("ML_Research_Tracker.xlsx exists", False, f"Not found at {xlsx_path}")
        record("CRITICAL Videos: >=5 строк и >=5 реальных ID видео в листе", False, "File missing")
        record("CRITICAL Papers: >=3 строк и >=3 реальных arXiv ID в листе", False, "File missing")
        record("Summary sheet has >= 3 rows", False, "File missing")
        return
    record("ML_Research_Tracker.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel readable", False, str(e))
        record("CRITICAL Videos: >=5 строк и >=5 реальных ID видео в листе", False, str(e))
        record("CRITICAL Papers: >=3 строк и >=3 реальных arXiv ID в листе", False, str(e))
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # --- Videos sheet (CRITICAL) ---
    videos_key = next((sheet_names_lower[k] for k in sheet_names_lower if "video" in k), None)
    if not videos_key:
        record("CRITICAL Videos: >=5 строк и >=5 реальных ID видео в листе", False,
               f"No Videos sheet. Sheets: {wb.sheetnames}")
    else:
        ws = wb[videos_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        headers = [str(c).strip().lower() if c else "" for c in rows[0]] if rows else []
        has_video_id = any("video_id" in h or "videoid" in h or "video id" in h for h in headers)
        has_title = any("title" in h for h in headers)
        blob = _cells_blob(ws)
        real_vid = sum(1 for vid in REAL_VIDEO_IDS if vid in blob)
        record("CRITICAL Videos: >=5 строк и >=5 реальных ID видео в листе",
               len(data_rows) >= 5 and has_video_id and has_title and real_vid >= 5,
               f"Rows: {len(data_rows)}, headers={rows[0] if rows else []}, "
               f"реальных ID видео: {real_vid}/7")

    # --- Papers sheet (CRITICAL) ---
    papers_key = next((sheet_names_lower[k] for k in sheet_names_lower if "paper" in k), None)
    if not papers_key:
        record("CRITICAL Papers: >=3 строк и >=3 реальных arXiv ID в листе", False,
               f"No Papers sheet. Sheets: {wb.sheetnames}")
    else:
        ws2 = wb[papers_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        headers2 = [str(c).strip().lower() if c else "" for c in rows2[0]] if rows2 else []
        has_arxiv_id = any("arxiv" in h or "id" in h for h in headers2)
        has_title2 = any("title" in h for h in headers2)
        blob2 = _cells_blob(ws2)
        real_pap = sum(1 for aid in REAL_ARXIV_IDS if aid in blob2)
        record("CRITICAL Papers: >=3 строк и >=3 реальных arXiv ID в листе",
               len(data_rows2) >= 3 and has_arxiv_id and has_title2 and real_pap >= 3,
               f"Rows: {len(data_rows2)}, headers={rows2[0] if rows2 else []}, "
               f"реальных arXiv ID: {real_pap}/5")

    # --- Summary sheet (non-critical) ---
    summary_key = next((sheet_names_lower[k] for k in sheet_names_lower if "summar" in k), None)
    if not summary_key:
        record("Summary sheet has >= 3 rows", False, f"No Summary sheet. Sheets: {wb.sheetnames}")
    else:
        ws3 = wb[summary_key]
        rows3 = list(ws3.iter_rows(values_only=True))
        data_rows3 = [r for r in rows3[1:] if any(c for c in r)] if rows3 else []
        record("Summary sheet has >= 3 rows", len(data_rows3) >= 3,
               f"Found {len(data_rows3)} data rows")

    # --- Groundtruth XLSX value comparison (non-critical) ---
    gt_path = os.path.join(groundtruth_workspace, "ML_Research_Tracker.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        for gt_sname in gt_wb.sheetnames:
            gt_ws = gt_wb[gt_sname]
            a_ws = None
            for asn in wb.sheetnames:
                if asn.strip().lower() == gt_sname.strip().lower():
                    a_ws = wb[asn]
                    break
            if a_ws is None:
                record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                       f"Available: {wb.sheetnames}")
                continue
            gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                       if any(c is not None for c in r)]
            a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                      if any(c is not None for c in r)]
            record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                   f"Expected {len(gt_rows)}, got {len(a_rows)}")
            for ri in range(min(3, len(gt_rows))):
                if ri >= len(a_rows):
                    break
                ok = True
                for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                    gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                    if gv is None:
                        continue
                    if isinstance(gv, (int, float)):
                        if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                            ok = False
                            break
                    else:
                        if not str_match(av, gv):
                            ok = False
                            break
                record(f"GT '{gt_sname}' row {ri+1} values", ok,
                       f"gt={gt_rows[ri][:4]}, agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
        gt_wb.close()


def check_teamly():
    print("\n=== Check 6: Teamly — страница ML Tech Research Hub ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            # Пользовательские страницы (сидовые id<=3 пропускаем).
            cur.execute("""
                SELECT title, body FROM teamly.pages
                WHERE id > 3
                  AND (title ILIKE '%ML Tech%'
                       OR title ILIKE '%Research Hub%'
                       OR title ILIKE '%Research Items%'
                       OR title ILIKE '%ML Research%')
            """)
            rows = cur.fetchall()
        conn.close()
    except Exception as e:
        record("CRITICAL Teamly: страница-хаб с >=4 реальными ссылками на репозитории",
               False, str(e))
        return

    record("Teamly: страница ML Tech Research Hub найдена", len(rows) >= 1,
           f"найдено страниц: {len(rows)}")

    body = "\n\n".join((b or "") for _t, b in rows)
    repo_hits = sum(1 for frag in REAL_REPO_FRAGMENTS if frag in body)
    vid_hits = sum(1 for vid in REAL_VIDEO_IDS if vid in body)

    # CRITICAL: тело хаба содержит реальные ссылки на репозитории (доказательство
    # фактической каталогизации, а не предзаготовки ответа).
    record("CRITICAL Teamly: страница-хаб с >=4 реальными ссылками на репозитории",
           len(rows) >= 1 and (vid_hits >= 4 or repo_hits >= 4),
           f"реальных репозиториев в теле: {repo_hits}/7, ID видео: {vid_hits}/7")

    # NON-critical: наличие меток полей Type/Topic/Status в теле.
    bl = body.lower()
    label_hits = sum(1 for lbl in ("type", "topic", "status", "video", "paper") if lbl in bl)
    record("Teamly: тело хаба содержит поля Type/Topic/Status и записи Video/Paper",
           label_hits >= 4, f"меток найдено: {label_hits}/5")


def check_email():
    print("\n=== Check 7: Email на research@lab.edu ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        count = 0
        with conn.cursor() as cur:
            # Письмо в Sent на research@lab.edu (не исходное письмо PI).
            cur.execute("""
                SELECT COUNT(*) FROM email.messages
                WHERE to_addr::text ILIKE '%research@lab.edu%'
                  AND COALESCE(from_addr, '') <> 'pi@lab.edu'
                  AND folder_id = (SELECT id FROM email.folders WHERE name = 'Sent' LIMIT 1)
            """)
            count = cur.fetchone()[0]
            if count == 0:
                cur.execute("""
                    SELECT COUNT(*) FROM email.messages
                    WHERE to_addr::text ILIKE '%research@lab.edu%'
                      AND COALESCE(from_addr, '') <> 'pi@lab.edu'
                """)
                count = cur.fetchone()[0]
            if count == 0:
                try:
                    cur.execute("""
                        SELECT COUNT(*) FROM email.sent_log
                        WHERE to_addr ILIKE '%research@lab.edu%'
                    """)
                    count = cur.fetchone()[0]
                except Exception:
                    conn.rollback()
        conn.close()
    except Exception as e:
        record("CRITICAL Email: письмо на research@lab.edu отправлено (не от pi@lab.edu)",
               False, str(e))
        return

    record("CRITICAL Email: письмо на research@lab.edu отправлено (не от pi@lab.edu)",
           count > 0, f"найдено писем: {count}")


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print("Running evaluation for yt-ml-repos-github-notion-excel-email")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_teamly()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    pct = 100.0 * PASS_COUNT / total if total else 0.0
    print(f"\n{'='*40}")
    print(f"=== SUMMARY: {PASS_COUNT}/{total} проверок пройдено ({pct:.1f}%) ===")

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({"passed": PASS_COUNT, "failed": FAIL_COUNT, "pct": pct,
                       "critical_failed": CRITICAL_FAILED}, f)

    if CRITICAL_FAILED:
        print(f"CRITICAL FAIL: {CRITICAL_FAILED}")
        print("Result: FAIL")
        return False, f"CRITICAL FAIL: {CRITICAL_FAILED}"
    if pct < 70.0:
        print("Result: FAIL (accuracy < 70%)")
        return False, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT} (<70%)"
    print("Result: PASS")
    return True, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
