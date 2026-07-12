"""Evaluation for canvas-assignment-stats."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Assignment_Stats.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Assignment_Stats.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []

    # ------------------------------------------------------------------
    # CRITICAL CHECKS — любой провал => немедленный FAIL (sys.exit(1)).
    # Значения детерминированы из seed-данных Canvas (OU-Anonymised dataset),
    # поэтому жёсткое сравнение корректно и не "хрупко".
    # ------------------------------------------------------------------
    critical_errors = []

    def _is_summary_header(row):
        # Распознаём строку-заголовок Metric/Value, не отбрасывая безусловно
        # первую строку. task.md НЕ предписывает строку-заголовок для листа
        # Summary, поэтому первая строка может быть реальной метрикой.
        if not row or row[0] is None:
            return False
        return str(row[0]).strip().lower() in ("metric", "показатель", "metric_name")

    def _summary_lookup(wb):
        rows = load_sheet_rows(wb, "Summary")
        out = {}
        if rows:
            for row in rows:
                if not row or row[0] is None:
                    continue
                if _is_summary_header(row):
                    continue
                out[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None
        return out

    def _stats_lookup(wb):
        rows = load_sheet_rows(wb, "Assignment Stats")
        out = {}
        if rows:
            for row in rows[1:]:
                if row and row[0] is not None:
                    out[str(row[0]).strip().lower()] = row
        return out

    a_summary = _summary_lookup(agent_wb)
    a_stats = _stats_lookup(agent_wb)

    # 1. Summary.Total_Assignments == 206 (точное, tol<=1)
    if not num_close(a_summary.get("total_assignments"), 206, 1):
        critical_errors.append(
            f"CRITICAL: Summary.Total_Assignments={a_summary.get('total_assignments')} != 206 (tol<=1)")

    # 2. Summary.Total_TMAs == 106 и Summary.Total_CMAs == 76 (точное, tol<=1)
    if not num_close(a_summary.get("total_tmas"), 106, 1):
        critical_errors.append(
            f"CRITICAL: Summary.Total_TMAs={a_summary.get('total_tmas')} != 106 (tol<=1)")
    if not num_close(a_summary.get("total_cmas"), 76, 1):
        critical_errors.append(
            f"CRITICAL: Summary.Total_CMAs={a_summary.get('total_cmas')} != 76 (tol<=1)")

    # 3. Summary.Course_Most_Assignments == 'DDD-2013B' (без учёта регистра)
    if not str_match(a_summary.get("course_most_assignments"), "DDD-2013B"):
        critical_errors.append(
            f"CRITICAL: Summary.Course_Most_Assignments={a_summary.get('course_most_assignments')} != DDD-2013B")

    # 4. Покурсовые счётчики для курсов с наибольшим объёмом заданий
    #    (Total_Assignments / TMA_Count / CMA_Count), tol<=1.
    high_volume = {
        "ddd-2013b": (14, 6, 7),
        "fff-2013b": (13, 5, 7),
        "bbb-2013b": (12, 6, 5),
    }
    for code, (tot, tma, cma) in high_volume.items():
        row = a_stats.get(code)
        if row is None:
            critical_errors.append(f"CRITICAL: Assignment Stats missing high-volume course '{code}'")
            continue
        if not num_close(row[1] if len(row) > 1 else None, tot, 1):
            critical_errors.append(
                f"CRITICAL: {code}.Total_Assignments={row[1] if len(row)>1 else None} != {tot} (tol<=1)")
        if not num_close(row[3] if len(row) > 3 else None, tma, 1):
            critical_errors.append(
                f"CRITICAL: {code}.TMA_Count={row[3] if len(row)>3 else None} != {tma} (tol<=1)")
        if not num_close(row[4] if len(row) > 4 else None, cma, 1):
            critical_errors.append(
                f"CRITICAL: {code}.CMA_Count={row[4] if len(row)>4 else None} != {cma} (tol<=1)")

    # 5. Оба выходных файла существуют с обоими обязательными листами.
    for req_sheet in ("Assignment Stats", "Summary"):
        if load_sheet_rows(agent_wb, req_sheet) is None:
            critical_errors.append(f"CRITICAL: required sheet '{req_sheet}' missing in agent output")

    if critical_errors:
        print("\n=== RESULT: FAIL (critical checks) ===")
        for e in critical_errors:
            print(f"  {e}")
        sys.exit(1)


    # Check sheet: Assignment Stats
    print(f"  Checking Assignment Stats...")
    a_rows = load_sheet_rows(agent_wb, "Assignment Stats")
    g_rows = load_sheet_rows(gt_wb, "Assignment Stats")
    if a_rows is None:
        all_errors.append("Sheet 'Assignment Stats' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Assignment Stats' not found in groundtruth")
    else:
        sheet_name = "Assignment Stats"
        errors = []
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        
        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    errors.append(f"{key}.Total_Assignments: {a_row[1]} vs {g_row[1]} (tol=1)")

            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1.0):
                    errors.append(f"{key}.Avg_Points: {a_row[2]} vs {g_row[2]} (tol=1.0)")

            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1):
                    errors.append(f"{key}.TMA_Count: {a_row[3]} vs {g_row[3]} (tol=1)")

            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1):
                    errors.append(f"{key}.CMA_Count: {a_row[4]} vs {g_row[4]} (tol=1)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")


    # Check sheet: Summary
    print(f"  Checking Summary...")
    a_rows = load_sheet_rows(agent_wb, "Summary")
    g_rows = load_sheet_rows(gt_wb, "Summary")
    if a_rows is None:
        all_errors.append("Sheet 'Summary' not found in agent output")
    elif g_rows is None:
        all_errors.append("Sheet 'Summary' not found in groundtruth")
    else:
        sheet_name = "Summary"
        errors = []
        # Не отбрасываем безусловно первую строку: строку-заголовок Metric/Value
        # распознаём явно, иначе первая реальная метрика агента терялась бы.
        a_data = [r for r in a_rows if not _is_summary_header(r)]
        g_data = [r for r in g_rows if not _is_summary_header(r)]

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]}")
                continue
            
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1.0):
                    errors.append(f"{key}.Value: {a_row[1]} vs {g_row[1]} (tol=1.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print(f"    PASS")

    
    docx_path = os.path.join(args.agent_workspace, "Assignment_Overview.docx")
    if not os.path.exists(docx_path):
        all_errors.append("Assignment_Overview.docx not found")
    else:
        try:
            from docx import Document as _DocCheck
            _doc = _DocCheck(docx_path)
            _text = " ".join(p.text for p in _doc.paragraphs).lower()
            _headings = " ".join(p.text for p in _doc.paragraphs if p.style.name.startswith("Heading")).lower()
            if len(_text.strip()) < 50:
                all_errors.append("Assignment_Overview.docx has too little text content (< 50 chars)")
            # RU+EN: агент пишет сводку на русском, поэтому принимаем как
            # английские, так и русские ключевые слова. Достаточно совпадения
            # хотя бы одного из вариантов.
            _kws = ["assignment", "overview", "задани", "обзор", "сводк"]
            _matched = [k for k in _kws if (k in _text or k in _headings)]
            if not _matched:
                all_errors.append(f"Assignment_Overview.docx missing expected keywords (RU+EN): {_kws}")
        except ImportError:
            if os.path.getsize(docx_path) < 100:
                all_errors.append("Assignment_Overview.docx too small")
        except Exception as _e:
            all_errors.append(f"Error reading Assignment_Overview.docx: {_e}")

    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
