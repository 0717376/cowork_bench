"""Evaluation for terminal-canvas-kulinar-excel-word-gcal.

Checks:
1. Nutrition_Course_Assessment.xlsx with 4 sheets
2. Assessment_Feedback.docx
3. Google Calendar events for grading sessions
"""
import argparse
import os
import sys

import openpyxl
import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

# Ground truth derived offline from the kulinar MCP recipe database
# (local_servers/kulinar-mcp/src/data/all_recipes.json): 50 recipes across
# 8 RU categories. Avg_Difficulty = mean of per-recipe `difficulty` field,
# rounded to 1 decimal. These replace the legacy Chinese recipe categories.
KULINAR_TOTAL_RECIPES = 50
KULINAR_TOTAL_CATEGORIES = 8
KULINAR_CATEGORIES = {
    "выпечка": {"count": 5, "avg_difficulty": 3.0},
    "гарнир": {"count": 5, "avg_difficulty": 1.0},
    "горячее": {"count": 10, "avg_difficulty": 2.3},
    "десерт": {"count": 6, "avg_difficulty": 2.3},
    "закуска": {"count": 5, "avg_difficulty": 2.2},
    "напиток": {"count": 4, "avg_difficulty": 1.2},
    "салат": {"count": 7, "avg_difficulty": 1.6},
    "суп": {"count": 8, "avg_difficulty": 1.9},
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        tag = "CRITICAL FAIL" if critical else "FAIL"
        print(f"  [{tag}] {name}: {str(detail)[:300]}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if s.strip().lower() == name.strip().lower():
            return wb[s]
    return None


def check_excel(agent_workspace, groundtruth_workspace):
    print("\n=== Checking Nutrition_Course_Assessment.xlsx ===")
    agent_file = os.path.join(agent_workspace, "Nutrition_Course_Assessment.xlsx")
    gt_file = os.path.join(groundtruth_workspace, "Nutrition_Course_Assessment.xlsx")

    check("Excel file exists", os.path.isfile(agent_file), agent_file)
    if not os.path.isfile(agent_file):
        return

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e))
        return

    # Sheet 1: Student_Assignments
    print("  Checking Student_Assignments...")
    a_sheet = get_sheet(agent_wb, "Student_Assignments")
    g_sheet = get_sheet(gt_wb, "Student_Assignments")
    check("Sheet 'Student_Assignments' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        g_rows = list(g_sheet.iter_rows(min_row=2, values_only=True))
        # Query dynamic assignment count from Canvas DB
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id IN (3, 4)")
            expected_assign_count = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_assign_count = 18
        check(f"Student_Assignments has {expected_assign_count} rows",
              len(a_rows) == expected_assign_count, f"Got {len(a_rows)}")

        # Check a few key assignments by name
        a_lookup = {}
        for r in a_rows:
            if r and r[1]:
                a_lookup[str(r[1]).strip().lower()] = r
        for g_row in g_rows:
            if not g_row or not g_row[1]:
                continue
            key = str(g_row[1]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                check(f"Assignment '{g_row[1]}' present", False, "Missing")
                continue
            if len(a_row) > 2 and len(g_row) > 2:
                check(f"'{key}' Points",
                      num_close(a_row[2], g_row[2], 1.0),
                      f"Expected {g_row[2]}, got {a_row[2]}")

    # Sheet 2: Recipe_Analysis (kulinar: 8 RU categories)
    print("  Checking Recipe_Analysis...")
    a_sheet = get_sheet(agent_wb, "Recipe_Analysis")
    check("Sheet 'Recipe_Analysis' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        # CRITICAL: exactly the 8 kulinar categories, not the legacy 10 Chinese ones.
        check("Recipe_Analysis has 8 rows", len(a_rows) == 8, f"Got {len(a_rows)}",
              critical=True)

        a_lookup = {str(r[0]).strip(): r for r in a_rows if r and r[0]}
        # CRITICAL: the agent must list exactly the kulinar RU categories with
        # correct per-category count and avg difficulty (proves it queried kulinar).
        for cat, gt in KULINAR_CATEGORIES.items():
            a_row = a_lookup.get(cat)
            if a_row is None:
                check(f"Category '{cat}' present", False, "Missing", critical=True)
                continue
            if len(a_row) > 1:
                check(f"'{cat}' Recipe_Count",
                      num_close(a_row[1], gt["count"], 1),
                      f"Expected {gt['count']}, got {a_row[1]}", critical=True)
            if len(a_row) > 2:
                check(f"'{cat}' Avg_Difficulty",
                      num_close(a_row[2], gt["avg_difficulty"], 0.3),
                      f"Expected {gt['avg_difficulty']}, got {a_row[2]}")

    # Sheet 3: Course_Summary
    print("  Checking Course_Summary...")
    a_sheet = get_sheet(agent_wb, "Course_Summary")
    g_sheet = get_sheet(gt_wb, "Course_Summary")
    check("Sheet 'Course_Summary' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet and g_sheet:
        a_data = {}
        for row in a_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                a_data[str(row[0]).strip().lower()] = row[1]
        g_data = {}
        for row in g_sheet.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                g_data[str(row[0]).strip().lower()] = row[1]

        # Query dynamic assignment counts + submission totals from Canvas DB.
        expected_c3 = expected_c4 = None
        expected_c3_sub = expected_c4_sub = None
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = 3")
            expected_c3 = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = 4")
            expected_c4 = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM canvas.submissions s "
                        "JOIN canvas.assignments a ON s.assignment_id = a.id "
                        "WHERE a.course_id = 3")
            expected_c3_sub = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM canvas.submissions s "
                        "JOIN canvas.assignments a ON s.assignment_id = a.id "
                        "WHERE a.course_id = 4")
            expected_c4_sub = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_c3, expected_c4 = 12, 6

        check("Total_Courses = 2",
              num_close(a_data.get("total_courses"), 2, 0),
              f"Got {a_data.get('total_courses')}")
        # CRITICAL: Canvas aggregates must match live DB for courses 3 and 4.
        check(f"Course_3_Assignments = {expected_c3}",
              num_close(a_data.get("course_3_assignments"), expected_c3, 0),
              f"Got {a_data.get('course_3_assignments')}", critical=True)
        check(f"Course_4_Assignments = {expected_c4}",
              num_close(a_data.get("course_4_assignments"), expected_c4, 0),
              f"Got {a_data.get('course_4_assignments')}", critical=True)
        if expected_c3_sub is not None:
            check(f"Course_3_Total_Submissions = {expected_c3_sub}",
                  num_close(a_data.get("course_3_total_submissions"), expected_c3_sub, 1),
                  f"Got {a_data.get('course_3_total_submissions')}", critical=True)
        if expected_c4_sub is not None:
            check(f"Course_4_Total_Submissions = {expected_c4_sub}",
                  num_close(a_data.get("course_4_total_submissions"), expected_c4_sub, 1),
                  f"Got {a_data.get('course_4_total_submissions')}", critical=True)
        # CRITICAL: recipe totals must match kulinar (8 categories / 50 recipes),
        # not the retired legacy 10/322.
        check(f"Total_Recipe_Categories = {KULINAR_TOTAL_CATEGORIES}",
              num_close(a_data.get("total_recipe_categories"), KULINAR_TOTAL_CATEGORIES, 0),
              f"Got {a_data.get('total_recipe_categories')}", critical=True)
        check(f"Total_Recipes = {KULINAR_TOTAL_RECIPES}",
              num_close(a_data.get("total_recipes"), KULINAR_TOTAL_RECIPES, 2),
              f"Got {a_data.get('total_recipes')}", critical=True)

    # Sheet 4: Grading_Schedule
    print("  Checking Grading_Schedule...")
    a_sheet = get_sheet(agent_wb, "Grading_Schedule")
    check("Sheet 'Grading_Schedule' exists", a_sheet is not None,
          f"Sheets: {agent_wb.sheetnames}")
    if a_sheet:
        a_rows = list(a_sheet.iter_rows(min_row=2, values_only=True))
        # Dynamic: grading schedule should match course 3 assignment count
        try:
            conn = psycopg2.connect(**DB)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = 3")
            expected_sched_rows = cur.fetchone()[0]
            cur.close(); conn.close()
        except Exception:
            expected_sched_rows = 12
        # CRITICAL: one grading session per course-3 assignment.
        check(f"Grading_Schedule has {expected_sched_rows} rows",
              len(a_rows) == expected_sched_rows, f"Got {len(a_rows)}", critical=True)
        # CRITICAL: first session starts on 2026-03-10.
        if a_rows:
            first_date = str(a_rows[0][0]).strip()
            check("First grading session on 2026-03-10",
                  "2026-03-10" in first_date,
                  f"Got {first_date}", critical=True)
        # CRITICAL: all sessions fall on weekdays only.
        import datetime as _dt
        weekend_rows = []
        for r in a_rows:
            if not r or not r[0]:
                continue
            ds = str(r[0]).strip()[:10]
            try:
                d = _dt.date.fromisoformat(ds)
                if d.weekday() >= 5:
                    weekend_rows.append(ds)
            except ValueError:
                pass
        check("All grading sessions on weekdays",
              len(weekend_rows) == 0, f"Weekend dates: {weekend_rows}",
              critical=True)


def check_word(agent_workspace):
    print("\n=== Checking Assessment_Feedback.docx ===")
    docx_path = os.path.join(agent_workspace, "Assessment_Feedback.docx")
    check("Assessment_Feedback.docx exists", os.path.isfile(docx_path))
    if not os.path.isfile(docx_path):
        return
    try:
        from docx import Document
        doc = Document(docx_path)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Document has substantial content", len(text) > 200, f"Length: {len(text)}")
        # The agent may write Russian prose: accept RU+EN alternatives.
        has_course = any(k in text for k in ("course", "курс", "программ"))
        has_assign = any(k in text for k in ("assignment", "assessment", "задани", "оцен", "проверк"))
        check("Contains course/assignment reference", has_course and has_assign,
              text[:200])
        check("Contains recipe reference",
              any(k in text for k in ("recipe", "cooking", "category",
                                      "рецеп", "кулинар", "категор", "блюд")),
              text[:200])
        check("Contains recommendation",
              any(k in text for k in ("recommend", "suggest", "curriculum",
                                      "рекоменд", "предлаг", "учебн", "дополн")),
              text[:200])
    except ImportError:
        check("python-docx available", False)
    except Exception as e:
        check("Word document readable", False, str(e))


GRADING_EVENT_FILTER = """
    lower(summary) LIKE '%%grading%%'
       OR lower(summary) LIKE '%%cma%%'
       OR lower(summary) LIKE '%%tma%%'
       OR lower(summary) LIKE '%%final exam%%'
       OR lower(summary) LIKE '%%biochem%%'
       OR lower(summary) LIKE '%%проверк%%'
       OR lower(summary) LIKE '%%оцен%%'
       OR lower(summary) LIKE '%%биохим%%'
"""


def check_gcal():
    print("\n=== Checking Google Calendar ===")
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        # Expected grading events = number of course-3 assignments.
        try:
            cur.execute("SELECT COUNT(*) FROM canvas.assignments WHERE course_id = 3")
            expected_events = cur.fetchone()[0]
        except Exception:
            expected_events = 12

        cur.execute(f"SELECT COUNT(*) FROM gcal.events WHERE {GRADING_EVENT_FILTER}")
        cnt = cur.fetchone()[0]
        # CRITICAL: one grading event per course-3 assignment.
        check(f"Calendar has grading events (>={expected_events})",
              cnt >= expected_events, f"Found {cnt} grading events", critical=True)

        # CRITICAL: events on weekdays only.
        cur.execute(f"""
            SELECT start_datetime, EXTRACT(DOW FROM start_datetime) as dow
            FROM gcal.events WHERE {GRADING_EVENT_FILTER}
        """)
        rows = cur.fetchall()
        if rows:
            weekend_events = [r for r in rows if r[1] in (0, 6)]
            check("No weekend grading events", len(weekend_events) == 0,
                  f"Found {len(weekend_events)} weekend events", critical=True)

        cur.close()
        conn.close()
    except Exception as e:
        check("Calendar check", False, str(e))


def check_reverse_validation(workspace):
    print("\n=== Reverse Validation ===")
    # Check that RU noise calendar events (Совещание кафедры / Часы консультаций /
    # Обед преподавателей) are not included in the grading schedule sheet
    excel_path = os.path.join(workspace, "Nutrition_Course_Assessment.xlsx")
    if os.path.isfile(excel_path):
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            ws = get_sheet(wb, "Grading_Schedule")
            if ws:
                all_text = " ".join(
                    str(c).lower() for r in ws.iter_rows(values_only=True) for c in r if c
                )
                noise_terms = ["совещание кафедры", "часы консультаций",
                               "обед преподавателей"]
                noise_found = [t for t in noise_terms if t in all_text]
                check("No noise events in Grading_Schedule",
                      len(noise_found) == 0,
                      f"Found noise: {noise_found}")
            wb.close()
        except Exception as e:
            check("Reverse validation readable", False, str(e))


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_word(args.agent_workspace)
    check_gcal()
    check_reverse_validation(args.agent_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (100.0 * PASS_COUNT / total) if total else 0.0
    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")

    # Any CRITICAL failure => hard FAIL regardless of accuracy.
    if CRITICAL_FAILS:
        print(f"  Critical failures: {CRITICAL_FAILS}")
        print("  Overall: FAIL (critical check failed)")
        sys.exit(1)

    overall = accuracy >= 70.0
    print(f"  Overall: {'PASS' if overall else 'FAIL'}")
    sys.exit(0 if overall else 1)


if __name__ == "__main__":
    main()
