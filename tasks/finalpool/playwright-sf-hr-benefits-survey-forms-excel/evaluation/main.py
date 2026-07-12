"""Evaluation for playwright-sf-hr-benefits-survey-gform-excel (russified).

ClickHouse swap russifies sf_data department values centrally
(Engineering->Инженерия, Finance->Финансы, HR->Кадры, Operations->Операции,
R&D->НИОКР, Sales->Продажи, Support->Поддержка). The agent reads RU dept names
from the DWH and writes them into the 'Department Satisfaction' sheet AND the
forms (gform) dropdown, so this evaluator matches departments by RU name with EN
fallback. Market/website data (competitors, %, PTO, match) stays English.
"""
import argparse
import json
import os
import sys

import psycopg2

DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

# RU-primary + EN-fallback department aliases (post-swap DWH values).
DEPT_ALIASES = [
    ["инженерия", "engineering"],
    ["финансы", "finance"],
    ["кадры", "hr"],
    ["операции", "operations"],
    ["ниокр", "r&d"],
    ["продажи", "sales"],
    ["поддержка", "support"],
]


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def _row_for_dept(rows, aliases):
    """Find a data row whose dept cell matches any alias (RU or EN)."""
    al = {a for a in aliases}
    for r in rows:
        v = str(r[0]).strip().lower() if r and r[0] is not None else ""
        if v in al:
            return r
    return None


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_workspace):
    """Returns (errors, criticals) where criticals are (label, ok) tuples."""
    errors = []
    criticals = []
    import openpyxl
    path = os.path.join(agent_workspace, "Benefits_Analysis.xlsx")
    if not os.path.exists(path):
        errors.append("Benefits_Analysis.xlsx not found")
        criticals.append(("Excel Benefits_Analysis.xlsx существует", False))
        return errors, criticals
    try:
        wb = openpyxl.load_workbook(path, data_only=True)

        # ---- Competitor Comparison ----
        comp_ok = True
        rows = load_sheet_rows(wb, "Competitor Comparison")
        if rows is None:
            errors.append("Sheet 'Competitor Comparison' not found")
            comp_ok = False
        else:
            data_rows = [r for r in rows[1:] if r and r[0] is not None]
            if len(data_rows) < 7:
                errors.append(f"Competitor Comparison has {len(data_rows)} rows, expected 7")
                comp_ok = False
            companies = {str(r[0]).strip().lower() for r in data_rows if r[0]}
            required_companies = [
                "acme technologies", "beta financial group", "cypress industries",
                "delta software corp", "echo systems inc", "frontier analytics",
                "our company",
            ]
            for c in required_companies:
                if c not in companies:
                    errors.append(f"Company '{c}' missing from Competitor Comparison")
                    comp_ok = False
            ours = [r for r in data_rows if r[0] and "our company" in str(r[0]).lower()]
            if ours:
                if not num_close(ours[0][1], 80, 1):
                    errors.append(f"Our Health Insurance={ours[0][1]}, expected 80")
                    comp_ok = False
                if not num_close(ours[0][2], 20, 1):
                    errors.append(f"Our PTO Days={ours[0][2]}, expected 20")
                    comp_ok = False
                if len(ours[0]) >= 4 and not num_close(ours[0][3], 4.0, 0.2):
                    errors.append(f"Our Retirement Match={ours[0][3]}, expected 4.0")
                    comp_ok = False
            else:
                comp_ok = False
        criticals.append((
            "Competitor Comparison: 6 конкурентов с сайта + 'Our Company' (Health=80, PTO=20, Match=4.0)",
            comp_ok))

        # ---- Department Satisfaction ----
        dept_ok = True
        rows2 = load_sheet_rows(wb, "Department Satisfaction")
        if rows2 is None:
            errors.append("Sheet 'Department Satisfaction' not found")
            dept_ok = False
        else:
            data_rows2 = [r for r in rows2[1:] if r and r[0] is not None]
            if len(data_rows2) < 7:
                errors.append(f"Department Satisfaction has {len(data_rows2)} rows, expected 7")
                dept_ok = False
            for aliases in DEPT_ALIASES:
                if _row_for_dept(data_rows2, aliases) is None:
                    errors.append(f"Department {aliases[0]}/{aliases[1]} missing")
                    dept_ok = False
            # Engineering -> ~6.58 avg, rating High
            eng = _row_for_dept(data_rows2, ["инженерия", "engineering"])
            if eng:
                if not num_close(eng[1], 6.58, 0.15):
                    errors.append(f"Engineering satisfaction={eng[1]}, expected ~6.58")
                    dept_ok = False
                if len(eng) >= 5 and str(eng[4]).strip().lower() != "high":
                    errors.append(f"Engineering rating={eng[4]}, expected High")
                    dept_ok = False
            # Operations -> Moderate (below 6.55 cutoff)
            ops = _row_for_dept(data_rows2, ["операции", "operations"])
            if ops and len(ops) >= 5 and str(ops[4]).strip().lower() != "moderate":
                errors.append(f"Operations rating={ops[4]}, expected Moderate")
                dept_ok = False
        criticals.append((
            "Department Satisfaction: все 7 подразделений (RU/EN) + Инженерия=High(~6.58), Операции=Moderate (порог 6.55)",
            dept_ok))

        # ---- Gap Analysis ----
        gap_ok = True
        rows3 = load_sheet_rows(wb, "Gap Analysis")
        if rows3 is None:
            errors.append("Sheet 'Gap Analysis' not found")
            gap_ok = False
        else:
            data_rows3 = [r for r in rows3[1:] if r and r[0] is not None]
            if len(data_rows3) < 3:
                errors.append(f"Gap Analysis has {len(data_rows3)} rows, expected 3")
                gap_ok = False
            # Health Insurance: market avg ~85.8 (mean of 90,85,82,88,78,92), Priority High (80<85.8)
            health = [r for r in data_rows3 if r[0] and "health" in str(r[0]).lower()]
            if health:
                if not num_close(health[0][2], 85.8, 1.0):
                    errors.append(f"Health Market Avg={health[0][2]}, expected ~85.8")
                    gap_ok = False
                if len(health[0]) >= 5 and str(health[0][4]).strip().lower() != "high":
                    errors.append(f"Health Priority={health[0][4]}, expected High")
                    gap_ok = False
            else:
                gap_ok = False
            # PTO: market avg ~22.8 (25,22,20,24,18,28), our 20 -> negative gap -> High
            pto = [r for r in data_rows3 if r[0] and ("pto" in str(r[0]).lower() or "отпуск" in str(r[0]).lower())]
            if pto:
                if not num_close(pto[0][2], 22.8, 1.0):
                    errors.append(f"PTO Market Avg={pto[0][2]}, expected ~22.8")
                    gap_ok = False
            # Retirement: market avg ~5.3 (6.0,5.5,4.5,5.0,4.0,6.5), our 4.0 -> negative gap -> High
            ret = [r for r in data_rows3 if r[0] and ("retirement" in str(r[0]).lower() or "пенси" in str(r[0]).lower())]
            if ret:
                if not num_close(ret[0][2], 5.3, 0.6):
                    errors.append(f"Retirement Market Avg={ret[0][2]}, expected ~5.3")
                    gap_ok = False
        criticals.append((
            "Gap Analysis: Health Market_Avg~85.8 & Priority=High; PTO~22.8; Retirement~5.3 (рынок из данных конкурентов)",
            gap_ok))

    except Exception as e:
        errors.append(f"Error reading Excel: {e}")
        criticals.append(("Excel читается без ошибок", False))
    return errors, criticals


def _option_values(config):
    """Extract option text values from a question config (RU forms-mcp shape)."""
    vals = []
    if not config:
        return vals
    cfg = config if isinstance(config, dict) else (
        json.loads(config) if config else {})
    opts = cfg.get("options") if isinstance(cfg, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                v = o.get("value")
                if v is not None:
                    vals.append(str(v))
            else:
                vals.append(str(o))
    return vals


def check_gform():
    """Returns (errors, criticals)."""
    errors = []
    criticals = []
    dropdown_ok = False
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("SELECT id, title FROM gform.forms ORDER BY id DESC LIMIT 10")
        forms = cur.fetchall()
        if not forms:
            errors.append("No forms found")
            criticals.append((
                "Форма 'Employee Benefits Improvement Survey' с выпадающим списком из 7 подразделений (RU/EN)",
                False))
            cur.close()
            conn.close()
            return errors, criticals

        form_id = None
        for f in forms:
            t = (f[1] or "").lower()
            if "benefit" in t or "survey" in t or "опрос" in t or "льгот" in t:
                form_id = f[0]
                break
        if form_id is None:
            errors.append(f"No benefits survey form found (forms: {[f[1] for f in forms]})")
        else:
            cur.execute("SELECT title, question_type, config FROM gform.questions "
                        "WHERE form_id = %s", (form_id,))
            questions = cur.fetchall()
            if len(questions) < 5:
                errors.append(f"Form has {len(questions)} questions, expected at least 5")
            # Dropdown question must include all 7 dept names (RU or EN).
            for q_title, q_type, q_config in questions:
                opts = [v.strip().lower() for v in _option_values(q_config)]
                if not opts:
                    continue
                matched = all(
                    any(a in opts for a in aliases) for aliases in DEPT_ALIASES
                )
                if matched and len(opts) >= 7:
                    dropdown_ok = True
                    break
            if not dropdown_ok:
                errors.append("Department dropdown does not contain all 7 dept names (RU/EN)")

        criticals.append((
            "Форма 'Employee Benefits Improvement Survey' с выпадающим списком из 7 подразделений (RU/EN)",
            form_id is not None and dropdown_ok))
        cur.close()
        conn.close()
    except Exception as e:
        errors.append(f"Error checking forms: {e}")
        criticals.append((
            "Форма 'Employee Benefits Improvement Survey' с выпадающим списком из 7 подразделений (RU/EN)",
            False))
    return errors, criticals


def check_email():
    """Returns (errors, criticals)."""
    errors = []
    criticals = []
    email_ok = False
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%hr_leadership@company.com%'
            ORDER BY id DESC LIMIT 10
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
        if not rows:
            errors.append("No email found to hr_leadership@company.com")
        else:
            subj_ok = any("benefits competitiveness analysis" in (r[0] or "").lower()
                          and "survey launch" in (r[0] or "").lower()
                          for r in rows)
            if not subj_ok:
                errors.append(
                    f"Email subject mismatch (got: {[r[0] for r in rows][:3]})")
            # Body must be non-empty and mention gaps/survey (RU or EN).
            body_ok = False
            for r in rows:
                b = (r[1] or "").lower()
                if len(b.strip()) >= 20 and (
                    "gap" in b or "разрыв" in b or "льгот" in b or "benefit" in b
                ):
                    body_ok = True
                    break
            if not body_ok:
                errors.append("Email body empty or does not mention gaps/benefits")
            email_ok = subj_ok and body_ok
    except Exception as e:
        errors.append(f"Error checking email: {e}")
    criticals.append((
        "Письмо на hr_leadership@company.com с темой 'Benefits Competitiveness Analysis & Survey Launch' и непустым телом про разрывы",
        email_ok))
    return errors, criticals


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace")

    all_errors = []
    all_criticals = []

    print("  Checking Excel file...")
    errs, crits = check_excel(agent_ws)
    all_errors.extend(errs)
    all_criticals.extend(crits)
    for e in errs[:5]:
        print(f"    ERROR: {e}")
    if not errs:
        print("    PASS")

    print("  Checking forms (Google Form)...")
    errs, crits = check_gform()
    all_errors.extend(errs)
    all_criticals.extend(crits)
    for e in errs[:3]:
        print(f"    ERROR: {e}")
    if not errs:
        print("    PASS")

    print("  Checking email...")
    errs, crits = check_email()
    all_errors.extend(errs)
    all_criticals.extend(crits)
    for e in errs[:3]:
        print(f"    ERROR: {e}")
    if not errs:
        print("    PASS")

    # --- CRITICAL CHECKS: any failure => immediate FAIL before accuracy gate ---
    print("\n=== CRITICAL CHECKS ===")
    failed_crit = []
    for label, ok in all_criticals:
        print(f"  [{'OK' if ok else 'FAIL'}] {label}")
        if not ok:
            failed_crit.append(label)
    if failed_crit:
        print(f"\n=== RESULT: FAIL (critical checks failed: {len(failed_crit)}) ===")
        for label in failed_crit:
            print(f"  CRITICAL FAIL: {label}")
        sys.exit(1)

    # --- accuracy gate (>=70%) over all granular checks ---
    if all_errors:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)
    else:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)


if __name__ == "__main__":
    main()
