"""Evaluation for yf-earnings-calendar-report."""
import argparse
import os
import sys
import psycopg2
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def num_close_pct(a, b, pct=0.05):
    """Allow percentage-based tolerance for large numbers."""
    try:
        a, b = float(a), float(b)
        if b == 0:
            return abs(a) < 1.0
        return abs(a - b) / abs(b) <= pct
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")
    all_errors = []
    critical_errors = []

    # Expected MOEX tickers (deterministic from db/zzz_moex_after_init.sql seed).
    EXPECTED_TICKERS = {"GAZP.ME", "LKOH.ME", "MGNT.ME", "MTSS.ME", "SBER.ME"}

    # ---- Check Excel ----
    agent_excel = os.path.join(args.agent_workspace, "Stock_Financial_Summary.xlsx")
    gt_excel = os.path.join(gt_dir, "Stock_Financial_Summary.xlsx")

    if not os.path.exists(agent_excel):
        all_errors.append("Agent output Stock_Financial_Summary.xlsx not found")
        critical_errors.append("CRITICAL: Excel file Stock_Financial_Summary.xlsx not present")
    elif not os.path.exists(gt_excel):
        all_errors.append("Groundtruth Stock_Financial_Summary.xlsx not found")
    else:
        agent_wb = openpyxl.load_workbook(agent_excel, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_excel, data_only=True)

        # Check Stock Overview sheet
        print("  Checking Stock Overview...")
        a_rows = load_sheet_rows(agent_wb, "Stock Overview")
        g_rows = load_sheet_rows(gt_wb, "Stock Overview")
        if a_rows is None:
            all_errors.append("Sheet 'Stock Overview' not found in agent output")
            critical_errors.append("CRITICAL: Stock Overview sheet not present in agent output")
        elif g_rows is None:
            all_errors.append("Sheet 'Stock Overview' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []
            if len(a_data) != len(g_data):
                all_errors.append(f"Stock Overview row count: agent={len(a_data)}, expected={len(g_data)}")
            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().upper()] = row
            # CRITICAL: exactly 5 data rows, the expected MOEX tickers, sorted alphabetically.
            agent_tickers = [str(r[0]).strip().upper() for r in a_data if r and r[0] is not None]
            if len(a_data) != 5:
                critical_errors.append(f"CRITICAL: Stock Overview must have exactly 5 data rows, got {len(a_data)}")
            if set(agent_tickers) != EXPECTED_TICKERS:
                critical_errors.append(
                    f"CRITICAL: Stock Overview tickers {sorted(set(agent_tickers))} != expected {sorted(EXPECTED_TICKERS)}"
                )
            if agent_tickers != sorted(agent_tickers):
                critical_errors.append("CRITICAL: Stock Overview rows not sorted alphabetically by Ticker")
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().upper()
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing ticker: {key}")
                    continue
                # Market Cap (col 3) - use percentage tolerance. CRITICAL: core data retrieval.
                if len(a_row) > 3 and len(g_row) > 3 and g_row[3] is not None:
                    if not num_close_pct(a_row[3], g_row[3], 0.05):
                        all_errors.append(f"{key}.Market_Cap: {a_row[3]} vs {g_row[3]}")
                        critical_errors.append(f"CRITICAL: {key}.Market_Cap {a_row[3]} != seed {g_row[3]}")
                # Trailing PE (col 4) - CRITICAL: confirms agent read get_stock_info.
                if len(a_row) > 4 and len(g_row) > 4 and g_row[4] is not None:
                    if not num_close(a_row[4], g_row[4], 1.0):
                        all_errors.append(f"{key}.Trailing_PE: {a_row[4]} vs {g_row[4]}")
                        critical_errors.append(f"CRITICAL: {key}.Trailing_PE {a_row[4]} != seed {g_row[4]}")
                # Forward PE (col 5) - CRITICAL.
                if len(a_row) > 5 and len(g_row) > 5 and g_row[5] is not None:
                    if not num_close(a_row[5], g_row[5], 1.0):
                        all_errors.append(f"{key}.Forward_PE: {a_row[5]} vs {g_row[5]}")
                        critical_errors.append(f"CRITICAL: {key}.Forward_PE {a_row[5]} != seed {g_row[5]}")
                # Trailing EPS (col 7)
                if len(a_row) > 7 and len(g_row) > 7 and g_row[7] is not None:
                    if not num_close(a_row[7], g_row[7], 0.5):
                        all_errors.append(f"{key}.Trailing_EPS: {a_row[7]} vs {g_row[7]}")
            if not any("Stock Overview" in e or "ticker" in e.lower() for e in all_errors):
                print("    PASS")

        # Check Profitability Metrics sheet
        print("  Checking Profitability Metrics...")
        a_rows = load_sheet_rows(agent_wb, "Profitability Metrics")
        g_rows = load_sheet_rows(gt_wb, "Profitability Metrics")
        if a_rows is None:
            all_errors.append("Sheet 'Profitability Metrics' not found in agent output")
            critical_errors.append("CRITICAL: Profitability Metrics sheet not present in agent output")
        elif g_rows is None:
            all_errors.append("Sheet 'Profitability Metrics' not found in groundtruth")
        else:
            a_data = a_rows[1:] if len(a_rows) > 1 else []
            g_data = g_rows[1:] if len(g_rows) > 1 else []
            if len(a_data) != len(g_data):
                all_errors.append(f"Profitability Metrics row count: agent={len(a_data)}, expected={len(g_data)}")
            a_lookup = {}
            for row in a_data:
                if row and row[0] is not None:
                    a_lookup[str(row[0]).strip().upper()] = row
            for g_row in g_data:
                if not g_row or g_row[0] is None:
                    continue
                key = str(g_row[0]).strip().upper()
                a_row = a_lookup.get(key)
                if a_row is None:
                    all_errors.append(f"Missing ticker in Profitability: {key}")
                    continue
                # Total Revenue (col 1) - pct tolerance. CRITICAL: core profitability deliverable.
                if len(a_row) > 1 and len(g_row) > 1 and g_row[1] is not None:
                    if not num_close_pct(a_row[1], g_row[1], 0.05):
                        all_errors.append(f"{key}.Total_Revenue: {a_row[1]} vs {g_row[1]}")
                        critical_errors.append(f"CRITICAL: {key}.Total_Revenue {a_row[1]} != seed {g_row[1]}")
                # Profit Margins (col 3). CRITICAL: decimal-expression rule honored.
                if len(a_row) > 3 and len(g_row) > 3 and g_row[3] is not None:
                    if not num_close(a_row[3], g_row[3], 0.02):
                        all_errors.append(f"{key}.Profit_Margins: {a_row[3]} vs {g_row[3]}")
                        critical_errors.append(f"CRITICAL: {key}.Profit_Margins {a_row[3]} != seed {g_row[3]}")
                # Forward EPS (col 5)
                if len(a_row) > 5 and len(g_row) > 5 and g_row[5] is not None:
                    if not num_close(a_row[5], g_row[5], 0.5):
                        all_errors.append(f"{key}.Forward_EPS: {a_row[5]} vs {g_row[5]}")
            if not any("Profitability" in e for e in all_errors):
                print("    PASS")

    # ---- Check Google Sheet ----
    print("  Checking Google Sheet...")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, title FROM gsheet.spreadsheets")
    spreadsheets = cur.fetchall()
    found_sheet = False
    for ss in spreadsheets:
        if ss[1] and "portfolio" in str(ss[1]).lower() and "watch" in str(ss[1]).lower():
            found_sheet = True
            ss_id = ss[0]
            # Check for Summary sheet
            cur.execute("SELECT id, title FROM gsheet.sheets WHERE spreadsheet_id=%s", (ss_id,))
            sheets = cur.fetchall()
            summary_found = False
            for s in sheets:
                if s[1] and "summary" in str(s[1]).lower():
                    summary_found = True
                    # Check cells exist
                    cur.execute("SELECT COUNT(*) FROM gsheet.cells WHERE spreadsheet_id=%s AND sheet_id=%s", (ss_id, s[0]))
                    cell_count = cur.fetchone()[0]
                    if cell_count < 40:  # 8 cols * 6 rows minimum
                        all_errors.append(f"Google Sheet has only {cell_count} cells, expected at least 40")
                    # Broaden: assert the expected MOEX tickers actually appear in Summary cells.
                    cur.execute(
                        "SELECT value FROM gsheet.cells WHERE spreadsheet_id=%s AND sheet_id=%s",
                        (ss_id, s[0]),
                    )
                    cell_vals = {str(r[0]).strip().upper() for r in cur.fetchall() if r[0] is not None}
                    missing = sorted(t for t in EXPECTED_TICKERS if t not in cell_vals)
                    if missing:
                        all_errors.append(f"Google Sheet Summary missing ticker(s): {missing}")
                    break
            if not summary_found:
                all_errors.append("Google Sheet missing 'Summary' sheet")
            break
    if not found_sheet:
        all_errors.append("Google Sheet 'Portfolio Watch List' not found")
        critical_errors.append("CRITICAL: Google Sheet 'Portfolio Watch List' not present")
    else:
        if not any("Google Sheet" in e for e in all_errors):
            print("    PASS")

    cur.close()
    conn.close()

    # ---- CRITICAL gate: any critical failure => immediate FAIL before accuracy ----
    if critical_errors:
        print(f"\n=== CRITICAL FAILURE ({len(critical_errors)}) ===")
        for e in critical_errors[:15]:
            print(f"  {e}")
        print("=== RESULT: FAIL ===")
        sys.exit(1)

    # ---- Accuracy gate: structural / non-critical checks, threshold >= 70 ----
    # Each non-critical error subtracts from a fixed pool of structural checks
    # (file present, sheets present, row counts, per-ticker EPS/Forward_EPS,
    #  Google Sheet structure & ticker presence).
    TOTAL_CHECKS = 20
    failed = len(all_errors)
    passed = max(0, TOTAL_CHECKS - failed)
    accuracy = 100.0 * passed / TOTAL_CHECKS

    print(f"\n=== Accuracy: {accuracy:.1f}% ({passed}/{TOTAL_CHECKS}), non-critical errors: {failed} ===")
    if all_errors:
        for e in all_errors[:15]:
            print(f"  {e}")

    if accuracy >= 70.0:
        print("=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
