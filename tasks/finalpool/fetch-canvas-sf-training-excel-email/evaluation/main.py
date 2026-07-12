"""Evaluation script for fetch-canvas-sf-training-excel-email.

Critical checks (see CRITICAL_CHECKS): any failure there => overall FAIL
regardless of accuracy. Otherwise pass threshold: accuracy >= 70%.
"""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Semantic floor: any failure here => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Sf_Training_Report.xlsx exists",
    "Data_Analysis has course-comparison columns",
    "Metrics values arithmetically consistent with Data_Analysis",
    "Analysis email sent to team-lead with correct subject and substantive body",
    "course_sf_training_processor.py ran and produced valid results JSON",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")

def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default

def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False

def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)

def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    excel_path = os.path.join(agent_workspace, "Sf_Training_Report.xlsx")
    check("Sf_Training_Report.xlsx exists", os.path.exists(excel_path))
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames)
        da_rows = []
        da_col_idx = {}
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            da_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 6 rows", len(da_rows) >= 6, f"got {len(da_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for i, h in enumerate(headers):
                da_col_idx[h] = i
            # Course-comparison framing (matches groundtruth Canvas course data).
            expected_cols = ['course', 'code', 'enrollment', 'avg_score', 'pass_rate']
            cols_present = all(ec in headers for ec in expected_cols)
            check("Data_Analysis has course-comparison columns", cols_present,
                  f"headers: {headers[:8]}")

        check("Metrics sheet exists", "Metrics" in wb.sheetnames)
        metrics_map = {}
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            m_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 4 rows", len(m_rows) >= 4, f"got {len(m_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")
            for row in m_rows:
                if row and row[0] is not None:
                    metrics_map[str(row[0]).strip().lower()] = row[1] if len(row) > 1 else None

        # ---- CRITICAL: Metrics must be arithmetically consistent with Data_Analysis ----
        consistent = False
        detail = "could not recompute"
        try:
            if da_rows and all(c in da_col_idx for c in ['enrollment', 'avg_score', 'pass_rate']):
                ei, si, pi = da_col_idx['enrollment'], da_col_idx['avg_score'], da_col_idx['pass_rate']
                enrolls = [safe_float(r[ei]) for r in da_rows if len(r) > ei and r[ei] is not None]
                scores = [safe_float(r[si]) for r in da_rows if len(r) > si and r[si] is not None]
                prates = [safe_float(r[pi]) for r in da_rows if len(r) > pi and r[pi] is not None]
                enrolls = [x for x in enrolls if x is not None]
                scores = [x for x in scores if x is not None]
                prates = [x for x in prates if x is not None]

                exp_total_courses = len(da_rows)
                exp_total_enroll = sum(enrolls) if enrolls else None
                exp_avg_score = sum(scores) / len(scores) if scores else None
                exp_avg_pr = sum(prates) / len(prates) if prates else None

                got_tc = safe_float(metrics_map.get('total_courses'))
                got_te = safe_float(metrics_map.get('total_enrollment'))
                got_as = safe_float(metrics_map.get('avg_score'))
                got_ap = safe_float(metrics_map.get('avg_pass_rate'))

                checks = []
                if got_tc is not None:
                    checks.append(num_close(got_tc, exp_total_courses, tol=0.5))
                if got_te is not None and exp_total_enroll is not None:
                    checks.append(num_close(got_te, exp_total_enroll, tol=1.0))
                if got_as is not None and exp_avg_score is not None:
                    checks.append(num_close(got_as, exp_avg_score, tol=0.5))
                if got_ap is not None and exp_avg_pr is not None:
                    checks.append(num_close(got_ap, exp_avg_pr, tol=0.5))

                # Require at least 3 of the 4 aggregate metrics present AND all present ones correct.
                consistent = len(checks) >= 3 and all(checks)
                detail = (f"recomputed: courses={exp_total_courses}, enroll={exp_total_enroll}, "
                          f"avg_score={round(exp_avg_score,2) if exp_avg_score else None}, "
                          f"avg_pr={round(exp_avg_pr,2) if exp_avg_pr else None}; "
                          f"reported: {metrics_map}")
        except Exception as e:
            detail = f"error: {e}"
        check("Metrics values arithmetically consistent with Data_Analysis", consistent, detail)

        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            r_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 2 rows", len(r_rows) >= 2, f"got {len(r_rows)}")

            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            for expected_col in ['Priority', 'Action', 'Course']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

        # ---- CRITICAL: email to team-lead with correct subject + substantive body ----
        try:
            conn = get_conn()
            cur = conn.cursor()
            cur.execute(
                "SELECT subject, to_addr, body_text FROM email.messages "
                "WHERE subject = %s",
                ('Analysis Report Complete',))
            rows = cur.fetchall()
            conn.close()

            ok = False
            detail = f"found {len(rows)} message(s) with exact subject"
            for subject, to_addr, body in rows:
                to_str = ",".join(to_addr).lower() if isinstance(to_addr, list) else (to_addr or "").lower()
                if 'team-lead@company.com' not in to_str:
                    continue
                body_l = (body or "").lower()
                # RU/EN substantive-body keywords (original .lower() text, NOT normalized).
                kw = ['отчёт', 'отчет', 'анализ', 'разрыв', 'курс', 'балл',
                      'report', 'analysis', 'gap', 'course', 'score', 'enrollment']
                if len(body_l) >= 20 and any(k in body_l for k in kw):
                    ok = True
                    break
            check("Analysis email sent to team-lead with correct subject and substantive body",
                  ok, detail)
        except Exception as e:
            check("Analysis email sent to team-lead with correct subject and substantive body",
                  False, str(e))

        # ---- CRITICAL: processor script ran and emitted valid results JSON ----
        proc_path = os.path.join(agent_workspace, "course_sf_training_processor.py")
        results_path = os.path.join(agent_workspace, "course_sf_training_results.json")
        proc_exists = os.path.exists(proc_path)
        results_ok = False
        rdetail = ""
        if os.path.exists(results_path):
            try:
                with open(results_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                results_ok = isinstance(data, (dict, list)) and len(data) > 0
                rdetail = f"results JSON loaded ({type(data).__name__})"
            except Exception as e:
                rdetail = f"results JSON invalid: {e}"
        else:
            rdetail = "course_sf_training_results.json missing"
        check("course_sf_training_processor.py ran and produced valid results JSON",
              proc_exists and results_ok, f"script_exists={proc_exists}; {rdetail}")

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    if res_log_file:
        try:
            with open(res_log_file, "w") as f:
                json.dump({
                    "total_passed": PASS_COUNT,
                    "total_checks": total,
                    "accuracy": accuracy,
                    "critical_failed": critical_failed,
                }, f, indent=2)
        except Exception:
            pass

    success = (not critical_failed) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks ({accuracy:.1f}%)"

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)

if __name__ == "__main__":
    main()
