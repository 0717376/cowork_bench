"""Evaluation for fetch-yf-macro-portfolio-excel-gcal-email (MOEX variant).

Russified swap: yahoo-finance -> moex-finance. The five holdings are the MOEX
tickers SBER.ME, GAZP.ME, LKOH.ME, MGNT.ME, MTSS.ME (schema moex.*, globally
seeded). Rate_Sensitivity is keyed off trailingPE (NOT beta, which does not exist
for RU issuers): High if a sector's average trailingPE > 5.0, else Low.

Grading model:
  * CRITICAL_CHECKS encode the substance of the deliverable. Any critical failure
    => immediate FAIL (sys.exit(1)) regardless of accuracy.
  * Remaining (structural) checks contribute to an accuracy score. PASS requires
    no critical failure AND accuracy >= 70%.

All groundtruth values below are the REAL MOEX seed values (db/zzz_moex_after_init.sql)
and the Russian macro_forecast.json served on :30211, not invented numbers.
"""
import argparse
import json
import os
import sys

import psycopg2


DB = dict(host=os.environ.get("PGHOST", "localhost"), port=5432,
          dbname=os.environ.get("PGDATABASE", "cowork_gym"),
          user="eigent", password="camel")

# --- Groundtruth from MOEX seed --------------------------------------------
# symbol -> (sector, currentPrice, previousClose, marketCap, trailingPE)
# currentPrice/previousClose are the latest two close rows from moex.stock_prices
# (ORDER BY date DESC LIMIT 2), since the MCP server's pg_adapter.info() overrides
# those fields from stock_prices. sector/marketCap/trailingPE come from
# stock_info.data, which the server does NOT override.
STOCKS = {
    "SBER.ME": ("Financial Services", 120.8171, 122.9152, 2877540270080, 2.6448412),
    "GAZP.ME": ("Energy", 208.6306, 210.7187, 4681709912064, 2.2370353),
    "LKOH.ME": ("Energy", 3727.0965, 3655.1261, 2543495413760, 3.4640048),
    "MGNT.ME": ("Consumer Defensive", 4182.1923, 4110.1816, 435439271936, 9.077895),
    "MTSS.ME": ("Communication Services", 255.0623, 259.9887, 465162960896, 8.8383665),
}
SYMBOLS = sorted(STOCKS)

# Macro forecast served from the RU research API (macro_forecast.json)
MACRO = [("Q1 2026", 2.3, 6.5, 21.0),
         ("Q2 2026", 2.2, 6.0, 20.0),
         ("Q3 2026", 1.9, 5.5, 19.0),
         ("Q4 2026", 2.0, 5.0, 18.0)]
AVG_GDP = round(sum(m[1] for m in MACRO) / len(MACRO), 1)  # 2.1

# Sector Sensitivity groundtruth (PE rule, AVG_GDP=2.1 -> High sectors Overweight)
#   Communication Services: avgPE 8.84 -> High -> Overweight
#   Consumer Defensive:     avgPE 9.08 -> High -> Overweight
#   Energy:                 avgPE 2.85 -> Low  -> Hold
#   Financial Services:     avgPE 2.65 -> Low  -> Hold
SECTOR_GT = {
    "communication services": ("high", "overweight"),
    "consumer defensive": ("high", "overweight"),
    "energy": ("low", "hold"),
    "financial services": ("low", "hold"),
}


def num_close(a, b, tol=0.5):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


# ---------------------------------------------------------------------------
# Excel checks. Returns (critical_errors, errors).
# ---------------------------------------------------------------------------
def check_excel(agent_workspace):
    crit = []
    errs = []
    import openpyxl
    path = os.path.join(agent_workspace, "Macro_Portfolio_Analysis.xlsx")
    if not os.path.exists(path):
        return (["Macro_Portfolio_Analysis.xlsx not found"], [])
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        return ([f"Cannot open Macro_Portfolio_Analysis.xlsx: {e}"], [])

    # ---- Macro Forecast ----
    rows = load_sheet_rows(wb, "Macro Forecast")
    if rows is None:
        errs.append("Sheet 'Macro Forecast' not found")
    else:
        data = [r for r in rows[1:] if r and r[0] is not None]
        if len(data) < 4:
            errs.append(f"Macro Forecast has {len(data)} rows, expected 4")
        # CRITICAL: each of the 4 quarters present with correct values
        by_q = {}
        for r in data:
            key = str(r[0]).strip()
            for q, *_ in MACRO:
                if q.replace(" ", "").lower() in key.replace(" ", "").lower():
                    by_q[q] = r
        for q, gdp, infl, rate in MACRO:
            r = by_q.get(q)
            if r is None:
                crit.append(f"Macro Forecast: row {q} missing")
                continue
            if not num_close(r[1], gdp, 0.15):
                crit.append(f"Macro Forecast {q}: GDP={r[1]}, expected {gdp}")
            if len(r) > 2 and not num_close(r[2], infl, 0.15):
                errs.append(f"Macro Forecast {q}: Inflation={r[2]}, expected {infl}")
            if len(r) > 3 and not num_close(r[3], rate, 0.15):
                crit.append(f"Macro Forecast {q}: Interest={r[3]}, expected {rate}")

    # ---- Stock Holdings ----
    rows2 = load_sheet_rows(wb, "Stock Holdings")
    if rows2 is None:
        errs.append("Sheet 'Stock Holdings' not found")
    else:
        data2 = [r for r in rows2[1:] if r and r[0] is not None]
        if len(data2) < 5:
            errs.append(f"Stock Holdings has {len(data2)} rows, expected 5")
        by_sym = {}
        for r in data2:
            by_sym[str(r[0]).strip().upper()] = r
        # CRITICAL: exactly the 5 MOEX tickers present
        for sym in SYMBOLS:
            if sym.upper() not in by_sym:
                crit.append(f"Stock Holdings: ticker {sym} missing")
        extra = set(by_sym) - {s.upper() for s in SYMBOLS}
        if extra:
            errs.append(f"Stock Holdings: unexpected tickers {sorted(extra)}")
        # CRITICAL: sector, price, PE per ticker correct vs MOEX seed
        for sym in SYMBOLS:
            r = by_sym.get(sym.upper())
            if r is None:
                continue
            sector, cprice, pclose, mcap, pe = STOCKS[sym]
            if len(r) < 2 or str(r[1]).strip().lower() != sector.lower():
                crit.append(f"{sym}: Sector={r[1] if len(r) > 1 else None}, "
                            f"expected {sector}")
            # Price: accept currentPrice or previousClose (both seed-real)
            if len(r) < 3 or not (num_close(r[2], cprice, max(2.0, cprice * 0.03))
                                  or num_close(r[2], pclose, max(2.0, pclose * 0.03))):
                crit.append(f"{sym}: Price={r[2] if len(r) > 2 else None}, "
                            f"expected ~{cprice} (or {pclose})")
            # Market_Cap_B in billions, rounded to 1 decimal
            if len(r) > 3 and not num_close(r[3], round(mcap / 1e9, 1),
                                            max(0.2, mcap / 1e9 * 0.02)):
                errs.append(f"{sym}: Market_Cap_B={r[3]}, "
                            f"expected ~{round(mcap / 1e9, 1)}")
            # PE_Ratio
            if len(r) > 4 and not num_close(r[4], round(pe, 1), 0.2):
                errs.append(f"{sym}: PE_Ratio={r[4]}, expected ~{round(pe, 1)}")

    # ---- Sector Sensitivity ----
    rows3 = load_sheet_rows(wb, "Sector Sensitivity")
    if rows3 is None:
        errs.append("Sheet 'Sector Sensitivity' not found")
    else:
        data3 = [r for r in rows3[1:] if r and r[0] is not None]
        if len(data3) < 4:
            errs.append(f"Sector Sensitivity has {len(data3)} rows, expected 4")
        by_sec = {}
        for r in data3:
            by_sec[str(r[0]).strip().lower()] = r
        # CRITICAL: every unique sector has correct Rate_Sensitivity + Action
        for sec, (exp_rs, exp_act) in SECTOR_GT.items():
            r = by_sec.get(sec)
            if r is None:
                crit.append(f"Sector Sensitivity: sector '{sec}' missing")
                continue
            if len(r) > 1 and not num_close(r[1], AVG_GDP, 0.15):
                errs.append(f"Sector Sensitivity {sec}: Avg_GDP_Growth={r[1]}, "
                            f"expected {AVG_GDP}")
            if len(r) < 3 or str(r[2]).strip().lower() != exp_rs:
                crit.append(f"Sector Sensitivity {sec}: Rate_Sensitivity="
                            f"{r[2] if len(r) > 2 else None}, expected {exp_rs}")
            if len(r) < 4 or str(r[3]).strip().lower() != exp_act:
                crit.append(f"Sector Sensitivity {sec}: Recommended_Action="
                            f"{r[3] if len(r) > 3 else None}, expected {exp_act}")

    return (crit, errs)


# ---------------------------------------------------------------------------
def check_gcal():
    crit = []
    errs = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, description, start_datetime, end_datetime
            FROM gcal.events
            WHERE start_datetime::date = '2026-03-30'
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        return ([f"Error checking GCal: {e}"], [])

    if not rows:
        return (["No GCal event found on 2026-03-30"], [])

    # pick the rebalancing event
    target = None
    for summ, desc, st, et in rows:
        s = (summ or "").lower()
        if ("portfolio" in s or "rebalanc" in s or "ребаланс" in s or "портфел" in s):
            target = (summ, desc, st, et)
            break
    if target is None:
        errs.append(f"No portfolio rebalancing event "
                    f"(found summaries: {[r[0] for r in rows]})")
        target = (rows[0][0], rows[0][1], rows[0][2], rows[0][3])

    summ, desc, st, et = target
    # structural: time window 14:00-15:30
    try:
        if st is not None and st.strftime("%H:%M") != "14:00":
            errs.append(f"GCal event start time {st}, expected 14:00")
        if et is not None and et.strftime("%H:%M") != "15:30":
            errs.append(f"GCal event end time {et}, expected 15:30")
    except Exception:
        pass

    # CRITICAL: description contains all 5 tickers + avg GDP + a stance keyword
    d = (desc or "").lower()
    missing = [sym for sym in SYMBOLS if sym.lower() not in d]
    if missing:
        crit.append(f"GCal description missing tickers: {missing}")
    if str(AVG_GDP) not in d and f"{AVG_GDP:.1f}" not in d:
        crit.append(f"GCal description missing average GDP growth {AVG_GDP}")
    if not any(k in d for k in ("overweight", "underweight", "hold",
                                "увеличить", "уменьшить", "сохранить",
                                "нараст", "сократ")):
        crit.append("GCal description missing overall portfolio stance")
    return (crit, errs)


# ---------------------------------------------------------------------------
def check_email():
    crit = []
    errs = []
    try:
        conn = psycopg2.connect(**DB)
        cur = conn.cursor()
        cur.execute("""
            SELECT subject, body_text FROM email.messages
            WHERE to_addr::text ILIKE '%investment_committee@firm.com%'
            ORDER BY id DESC LIMIT 5
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        return ([f"Error checking email: {e}"], [])

    if not rows:
        return (["No email found to investment_committee@firm.com"], [])

    # CRITICAL: a message with the right subject exists
    subj_ok = None
    for subj, body in rows:
        s = (subj or "").lower()
        if "macro outlook" in s and "portfolio impact" in s:
            subj_ok = (subj, body)
            break
    if subj_ok is None:
        # looser fallback for non-critical credit
        for subj, body in rows:
            s = (subj or "").lower()
            if "macro" in s or "portfolio" in s or "outlook" in s:
                subj_ok = (subj, body)
                break
        crit.append(f"Email subject != 'Macro Outlook & Portfolio Impact - Q2 2026' "
                    f"(found: {[r[0] for r in rows]})")
    if subj_ok is None:
        subj_ok = rows[0]

    body = (subj_ok[1] or "").lower()
    # CRITICAL: body summarizes macro highlights + names an Overweight/Underweight sector
    if str(AVG_GDP) not in body and f"{AVG_GDP:.1f}" not in body:
        crit.append(f"Email body missing average GDP growth {AVG_GDP}")
    ow_sectors = [s for s, (rs, act) in SECTOR_GT.items() if act in ("overweight", "underweight")]
    if not any(s in body for s in ow_sectors):
        crit.append(f"Email body missing any Overweight/Underweight sector "
                    f"({ow_sectors})")
    # structural: mentions inflation and rate direction
    if not any(k in body for k in ("inflation", "инфляц")):
        errs.append("Email body does not mention inflation trend")
    if not any(k in body for k in ("rate", "ставк")):
        errs.append("Email body does not mention interest rate direction")
    return (crit, errs)


# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    agent_ws = args.agent_workspace or os.path.join(
        os.path.dirname(__file__), "..", "groundtruth_workspace")

    critical_errors = []
    errors = []

    print("  Checking Excel file...")
    c, e = check_excel(agent_ws)
    critical_errors += c
    errors += e
    for x in (c + e)[:8]:
        print(f"    ERROR: {x}")
    if not (c or e):
        print("    PASS")

    print("  Checking GCal event...")
    c, e = check_gcal()
    critical_errors += c
    errors += e
    for x in (c + e)[:6]:
        print(f"    ERROR: {x}")
    if not (c or e):
        print("    PASS")

    print("  Checking email...")
    c, e = check_email()
    critical_errors += c
    errors += e
    for x in (c + e)[:6]:
        print(f"    ERROR: {x}")
    if not (c or e):
        print("    PASS")

    # ---- Critical gate ----
    if critical_errors:
        print(f"\n=== CRITICAL FAILURE ({len(critical_errors)}) ===")
        for x in critical_errors[:15]:
            print(f"  CRITICAL: {x}")
        print("\n=== RESULT: FAIL ===")
        sys.exit(1)

    # ---- Accuracy gate (non-critical) ----
    # Total grading points: critical checks (all passed here) + structural checks.
    # We approximate weight: every critical check is 1 pt (all passed), each
    # remaining error is a structural miss.
    total_structural = 12  # number of structural (non-critical) checks above
    failed = len(errors)
    passed = max(0, total_structural - failed)
    accuracy = passed / total_structural * 100
    print(f"\n  Structural accuracy: {passed}/{total_structural} = {accuracy:.0f}%")

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL (accuracy {accuracy:.0f}% < 70%) ===")
        for x in errors[:10]:
            print(f"  {x}")
        sys.exit(1)


if __name__ == "__main__":
    main()
