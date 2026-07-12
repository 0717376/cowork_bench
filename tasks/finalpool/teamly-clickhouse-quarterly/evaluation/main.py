"""
Evaluation script for the Q4 sales review task (Teamly + ClickHouse fork).

Checks:
1. Excel Q4_Performance.xlsx - sheet "Regional Comparison" with correct data
   (Region/Target/Actual/Variance/Achievement_Pct), derived from ClickHouse.
2. Google Sheets - "Q4 Dashboard" spreadsheet with "Summary" sheet matching
   the Excel data.

Scoring:
- A tier of CRITICAL semantic checks. Any critical failure => sys.exit(1).
- Otherwise PASS requires accuracy (passed / total) >= 70%.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth
"""

import argparse
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []

# Known targets. Region keys are the russified strings produced centrally by
# db/zzz_clickhouse_after_init.sql (CUSTOMERS.REGION). They MUST match the
# values returned by get_actuals_from_clickhouse() and the page seeded by
# preprocess, or every Actual would silently resolve to 0.0.
TARGETS = {
    "Азиатско-Тихоокеанский регион": 80000,
    "Европа": 85000,
    "Латинская Америка": 75000,
    "Ближний Восток": 85000,
    "Северная Америка": 80000,
}

EXPECTED_HEADER = ["region", "target", "actual", "variance", "achievement_pct"]


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {detail[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def get_actuals_from_clickhouse():
    """Query actual Q4 2024 revenue by region from the ClickHouse proxy tables."""
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()
    cur.execute("""
        SELECT c."REGION",
               ROUND(SUM(o."TOTAL_AMOUNT")::numeric, 2) as total_revenue
        FROM sf_data."SALES_DW__PUBLIC__ORDERS" o
        JOIN sf_data."SALES_DW__PUBLIC__CUSTOMERS" c
          ON o."CUSTOMER_ID" = c."CUSTOMER_ID"
        WHERE o."ORDER_DATE" >= '2024-10-01'
          AND o."ORDER_DATE" <= '2024-12-31'
        GROUP BY c."REGION"
        ORDER BY c."REGION"
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {row[0]: float(row[1]) for row in rows}


def compute_expected_data():
    """Compute expected rows: Region, Target, Actual, Variance, Achievement_Pct."""
    actuals = get_actuals_from_clickhouse()
    rows = []
    for region in sorted(TARGETS.keys()):
        target = TARGETS[region]
        actual = actuals.get(region, 0.0)
        variance = round(actual - target, 2)
        achievement = round(actual / target * 100, 1)
        rows.append((region, target, actual, variance, achievement))
    return rows


def load_sheet_rows(wb, sheet_name):
    """Load all rows from a sheet (case-insensitive name lookup)."""
    matched = None
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            matched = name
            break
    if matched is None:
        return None
    ws = wb[matched]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def check_excel(agent_workspace, expected_data):
    """Check Q4_Performance.xlsx content."""
    print("\n=== Checking Excel Output ===")

    xlsx_path = os.path.join(agent_workspace, "Q4_Performance.xlsx")

    if not os.path.isfile(xlsx_path):
        record("Excel file exists", False, f"Not found: {xlsx_path}", critical=True)
        return

    record("Excel file exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e), critical=True)
        return

    agent_rows = load_sheet_rows(wb, "Regional Comparison")
    if agent_rows is None:
        record("Sheet 'Regional Comparison' exists", False, "Not found", critical=True)
        return

    record("Sheet 'Regional Comparison' exists", True)

    # Header order (non-critical structural)
    header = [str(c).strip().lower() if c is not None else "" for c in (agent_rows[0] if agent_rows else [])]
    record("Header order Region/Target/Actual/Variance/Achievement_Pct",
           header[:5] == EXPECTED_HEADER, f"Got {header[:5]}")

    # Skip header row
    data_rows = agent_rows[1:] if len(agent_rows) > 1 else []
    # Drop fully-empty trailing rows
    data_rows = [r for r in data_rows if any(c is not None and str(c).strip() != "" for c in r)]

    record("Row count is exactly 5", len(data_rows) == 5, f"Got {len(data_rows)}", critical=True)

    # Alphabetical sort by Region (structural)
    region_seq = [str(r[0]).strip() for r in data_rows if r and r[0] is not None]
    record("Rows sorted alphabetically by Region",
           region_seq == sorted(region_seq), f"Got {region_seq}")

    # Build lookup by region name
    agent_by_region = {}
    for row in data_rows:
        if row and row[0]:
            agent_by_region[str(row[0]).strip().lower()] = row

    for region, target, actual, variance, achievement in expected_data:
        region_key = region.lower()
        if region_key not in agent_by_region:
            record(f"Region '{region}' present", False, "Missing", critical=True)
            continue

        a_row = agent_by_region[region_key]
        record(f"Region '{region}' present", True)

        # Target must equal the value seeded on the Teamly page (proves KB read).
        record(f"{region}.Target", num_close(a_row[1], target, 1.0),
               f"{a_row[1]} vs expected {target}", critical=True)

        # Actual from ClickHouse (core data-warehouse deliverable).
        record(f"{region}.Actual", num_close(a_row[2], actual, 1.0),
               f"{a_row[2]} vs expected {actual}", critical=True)

        # Variance == Actual - Target (derived formula).
        record(f"{region}.Variance", num_close(a_row[3], variance, 1.0),
               f"{a_row[3]} vs expected {variance}", critical=True)

        # Achievement_Pct == Actual/Target*100 (derived formula).
        record(f"{region}.Achievement_Pct", num_close(a_row[4], achievement, 0.5),
               f"{a_row[4]} vs expected {achievement}", critical=True)


def check_gsheet(expected_data):
    """Check Google Sheets Q4 Dashboard with Summary sheet."""
    print("\n=== Checking Google Sheet ===")

    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Find spreadsheet titled "Q4 Dashboard"
    cur.execute("""
        SELECT id, title FROM gsheet.spreadsheets
        WHERE LOWER(title) LIKE '%%q4 dashboard%%'
        LIMIT 1
    """)
    row = cur.fetchone()

    if not row:
        record("Spreadsheet 'Q4 Dashboard' exists", False, "Not found in gsheet.spreadsheets")
        cur.close()
        conn.close()
        return

    record("Spreadsheet 'Q4 Dashboard' exists", True)
    spreadsheet_id = row[0]

    # Find "Summary" sheet
    cur.execute("""
        SELECT id, title FROM gsheet.sheets
        WHERE spreadsheet_id = %s AND LOWER(title) LIKE '%%summary%%'
        LIMIT 1
    """, (spreadsheet_id,))
    sheet_row = cur.fetchone()

    if not sheet_row:
        record("Sheet 'Summary' exists", False, "Not found")
        cur.close()
        conn.close()
        return

    record("Sheet 'Summary' exists", True)
    sheet_id = sheet_row[0]

    # Read all cells
    cur.execute("""
        SELECT row_index, col_index, value
        FROM gsheet.cells
        WHERE spreadsheet_id = %s AND sheet_id = %s
        ORDER BY row_index, col_index
    """, (spreadsheet_id, sheet_id))
    cells = cur.fetchall()
    cur.close()
    conn.close()

    if not cells:
        record("Summary sheet has data", False, "No cells found")
        return

    record("Summary sheet has data", True)

    # Build grid from cells
    grid = {}
    for row_idx, col_idx, value in cells:
        grid[(row_idx, col_idx)] = value

    for region, target, actual, variance, achievement in expected_data:
        # Find the row for this region
        matched_row = None
        for (r, c), v in grid.items():
            if v and str(v).strip().lower() == region.lower():
                matched_row = r
                break

        if matched_row is None:
            record(f"GSheet: Region '{region}' present", False, "Not found")
            continue

        record(f"GSheet: Region '{region}' present", True)

        # Collect numeric values in this row
        numeric_vals = []
        for ci in range(12):
            v = grid.get((matched_row, ci))
            if v is not None:
                try:
                    numeric_vals.append(float(v))
                except (ValueError, TypeError):
                    pass

        # Cross-surface consistency: Target AND Actual must match the Excel output.
        record(f"GSheet: {region}.Target",
               any(num_close(nv, target, 1.0) for nv in numeric_vals),
               f"Expected ~{target} in {numeric_vals}")
        record(f"GSheet: {region}.Actual",
               any(num_close(nv, actual, 1.0) for nv in numeric_vals),
               f"Expected ~{actual} in {numeric_vals}")
        # Derived columns must also be present (tighter than original).
        record(f"GSheet: {region}.Variance",
               any(num_close(nv, variance, 1.0) for nv in numeric_vals),
               f"Expected ~{variance} in {numeric_vals}")
        record(f"GSheet: {region}.Achievement_Pct",
               any(num_close(nv, achievement, 0.5) for nv in numeric_vals),
               f"Expected ~{achievement} in {numeric_vals}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    # Compute expected data dynamically from ClickHouse
    expected_data = compute_expected_data()
    print("Expected data:")
    for row in expected_data:
        print(f"  {row}")

    check_excel(args.agent_workspace, expected_data)
    check_gsheet(expected_data)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%")

    if CRITICAL_FAILED:
        print(f"  CRITICAL checks failed: {CRITICAL_FAILED}")
        print("  Overall: FAIL (critical)")
        sys.exit(1)

    if accuracy >= 70.0:
        print("  Overall: PASS")
        sys.exit(0)
    else:
        print("  Overall: FAIL (accuracy < 70%)")
        sys.exit(1)


if __name__ == "__main__":
    main()
