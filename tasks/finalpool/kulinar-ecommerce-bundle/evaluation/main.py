"""
Evaluation script for kulinar-ecommerce-bundle task (InSales + MOEX + Kulinar).

Dynamically computes expected values from PostgreSQL:
  - wc.products: 'Home Appliances' category products (RU label 'Бытовая техника')
  - moex.stock_prices: LKOH.ME last 5 trading days close prices

Recipe Bundles sheet must cover the three required kulinar categories
{напиток, гарнир, горячее} (one each), each with a valid kulinar category,
a non-empty (Cyrillic) recipe name, an integer difficulty 1-4, a real
Бытовая техника paired product, and Bundle_Price == product price * 0.85.

Critical (semantic) checks abort with sys.exit(1) before the accuracy gate.
PASS requires: no critical failure AND accuracy >= 70%.

Falls back to static groundtruth Excel only for row-count parity if PostgreSQL
is unavailable.

Usage:
    python evaluation/main.py \
        --agent_workspace /path/to/workspace \
        --groundtruth_workspace /path/to/groundtruth \
        --launch_time "2026-03-06 10:00:00" \
        --res_log_file /path/to/result.json
"""

import argparse
import json
import os
import sys

import openpyxl

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILURES = []

DB_CONFIG = dict(host=os.environ.get('PGHOST', 'localhost'), port=5432, database='cowork_gym',
                 user=os.environ.get('PGUSER', 'eigent'), password=os.environ.get('PGPASSWORD', 'camel'))

# Valid kulinar (Кулинар) recipe categories.
VALID_RECIPE_CATEGORIES = {
    'выпечка', 'гарнир', 'горячее', 'десерт',
    'закуска', 'напиток', 'салат', 'суп',
}

# Categories required by the task: drinks / staple-side / hot main.
REQUIRED_CATEGORIES = {'напиток', 'гарнир', 'горячее'}

# Chosen moex commodity-emitter ticker replacing the retired gold future GC=F.
GOLD_SYMBOL = 'LKOH.ME'


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{' [CRITICAL]' if critical else ''} {name}")
    else:
        FAIL_COUNT += 1
        detail_str = f": {detail[:200]}" if detail else ""
        print(f"  [FAIL]{' [CRITICAL]' if critical else ''} {name}{detail_str}")
        if critical:
            CRITICAL_FAILURES.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def has_cyrillic(s):
    return any('Ѐ' <= ch <= 'ӿ' for ch in str(s))


def get_sheet(wb, name):
    for s in wb.sheetnames:
        if str_match(s, name):
            return wb[s]
    return None


def compute_expected_from_db():
    """Compute expected values from PostgreSQL for Store Products and Gold Trend."""
    try:
        import psycopg2
    except ImportError:
        return None

    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"  WARNING: Could not connect to PostgreSQL: {e}")
        return None

    # --- Home Appliances products (RU label 'Бытовая техника') ---
    # The category NAME shown to the agent in InSales is russified to
    # 'Бытовая техника', but the products.categories jsonb keeps the
    # English literal 'Home Appliances', which is what we grep here.
    cur.execute("""
        SELECT name, price, stock_quantity
        FROM wc.products
        WHERE categories::text LIKE '%%Home Appliances%%'
        ORDER BY price::numeric ASC
    """)
    products = cur.fetchall()
    if not products:
        conn.close()
        return None

    store_products = []
    product_prices_by_name = {}
    for row in products:
        name = str(row[0]).strip()
        price = round(float(row[1]), 2)
        stock = int(row[2]) if row[2] is not None else 0
        store_products.append((name, price, stock))
        product_prices_by_name[name.lower()] = price

    # --- LKOH.ME close prices (last 5 trading days) ---
    cur.execute("""
        SELECT date, close
        FROM moex.stock_prices
        WHERE symbol = %s
        ORDER BY date DESC
        LIMIT 5
    """, (GOLD_SYMBOL,))
    gold_rows = cur.fetchall()
    if len(gold_rows) < 5:
        conn.close()
        return None

    gold_trend = []
    for row in gold_rows:
        date_str = str(row[0])
        close_price = round(float(row[1]), 2)
        gold_trend.append((date_str, close_price))

    conn.close()

    return {
        'store_products': store_products,
        'product_prices_by_name': product_prices_by_name,
        'gold_trend': gold_trend,
    }


def check_excel(agent_workspace, expected, groundtruth_workspace=None):
    """Check the agent's Excel output."""
    print("\n=== Checking Excel Output ===")

    agent_file = os.path.join(agent_workspace, "Bundle_Pricing.xlsx")
    check("Excel file exists", os.path.isfile(agent_file),
          f"Expected {agent_file}", critical=True)
    if not os.path.isfile(agent_file):
        return False

    try:
        agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    except Exception as e:
        check("Excel file readable", False, str(e), critical=True)
        return False

    expected_sheets = ["Store Products", "Gold Trend", "Recipe Bundles"]
    for sheet_name in expected_sheets:
        found = get_sheet(agent_wb, sheet_name) is not None
        check(f"Sheet '{sheet_name}' exists", found,
              f"Found sheets: {agent_wb.sheetnames}")

    use_db = expected is not None

    # ── Sheet 1: Store Products ──
    print("\n--- Store Products ---")
    agent_ws = get_sheet(agent_wb, "Store Products")
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        if use_db:
            exp_products = expected['store_products']
            check("Store Products row count",
                  len(agent_rows) == len(exp_products),
                  f"Expected {len(exp_products)}, got {len(agent_rows)}",
                  critical=True)

            # Check headers
            agent_headers = [c.value for c in agent_ws[1]]
            check("Store Products has 3+ columns",
                  len(agent_headers) >= 3,
                  f"Got {len(agent_headers)} columns: {agent_headers}")

            # Check sort order (by Price ascending)
            prices = []
            for row in agent_rows:
                try:
                    prices.append(float(row[1]))
                except (TypeError, ValueError, IndexError):
                    prices.append(0)
            if len(agent_rows) >= 2:
                check("Store Products sorted by Price ascending",
                      prices == sorted(prices),
                      f"Prices: {prices}", critical=True)

            # CRITICAL: every expected DB row must be matched in the agent
            # sheet by (price, stock) within tolerance (not loose substring).
            unmatched = []
            agent_pairs = []
            for row in agent_rows:
                if row and len(row) >= 3:
                    agent_pairs.append((row[1], row[2]))
            for exp_name, exp_price, exp_stock in exp_products:
                hit = any(
                    num_close(ap, exp_price, 0.05) and num_close(astk, exp_stock, 1)
                    for ap, astk in agent_pairs
                )
                if not hit:
                    unmatched.append(f"{exp_name} (price={exp_price}, stock={exp_stock})")
            check("Store Products: every Бытовая техника DB product present with correct Price+Stock",
                  len(unmatched) == 0,
                  f"Unmatched: {unmatched[:5]}", critical=True)
        else:
            # Fallback: use groundtruth (row-count parity only)
            gt_file = os.path.join(groundtruth_workspace, "Bundle_Pricing.xlsx")
            if os.path.isfile(gt_file):
                gt_wb = openpyxl.load_workbook(gt_file, data_only=True)
                gt_ws = get_sheet(gt_wb, "Store Products")
                if gt_ws:
                    gt_rows = list(gt_ws.iter_rows(min_row=2, values_only=True))
                    check("Store Products row count",
                          len(agent_rows) == len(gt_rows),
                          f"Expected {len(gt_rows)}, got {len(agent_rows)}")

    # ── Sheet 2: Gold Trend (LKOH.ME close prices) ──
    print("\n--- Gold Trend ---")
    agent_ws = get_sheet(agent_wb, "Gold Trend")
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Gold Trend has exactly 5 rows",
              len(agent_rows) == 5,
              f"Got {len(agent_rows)} rows", critical=True)

        agent_headers = [c.value for c in agent_ws[1]]
        check("Gold Trend has 2+ columns",
              len(agent_headers) >= 2,
              f"Got {len(agent_headers)} columns: {agent_headers}")

        if use_db:
            exp_gold = expected['gold_trend']

            # CRITICAL: most-recent close matches DB within tight tolerance.
            if len(agent_rows) >= 1 and len(exp_gold) >= 1:
                agent_first_close = agent_rows[0][1] if len(agent_rows[0]) >= 2 else None
                exp_first_close = exp_gold[0][1]
                check("Gold Trend most-recent close matches LKOH.ME DB",
                      num_close(agent_first_close, exp_first_close, 2.0),
                      f"Expected ~{exp_first_close}, got {agent_first_close}",
                      critical=True)

            # Check descending date order
            if len(agent_rows) >= 2:
                first_date = str(agent_rows[0][0])
                second_date = str(agent_rows[1][0])
                check("Gold Trend sorted by date descending",
                      first_date >= second_date,
                      f"First date={first_date}, second={second_date}",
                      critical=True)

            # CRITICAL: oldest close matches DB within tight tolerance.
            if len(agent_rows) >= 5 and len(exp_gold) >= 5:
                agent_last_close = agent_rows[4][1] if len(agent_rows[4]) >= 2 else None
                exp_last_close = exp_gold[4][1]
                check("Gold Trend oldest close matches LKOH.ME DB",
                      num_close(agent_last_close, exp_last_close, 2.0),
                      f"Expected ~{exp_last_close}, got {agent_last_close}",
                      critical=True)

    # ── Sheet 3: Recipe Bundles ──
    print("\n--- Recipe Bundles ---")
    agent_ws = get_sheet(agent_wb, "Recipe Bundles")
    if agent_ws:
        agent_rows = list(agent_ws.iter_rows(min_row=2, values_only=True))

        check("Recipe Bundles has exactly 3 rows",
              len(agent_rows) == 3,
              f"Got {len(agent_rows)} rows", critical=True)

        agent_headers = [c.value for c in agent_ws[1]]
        check("Recipe Bundles has 5+ columns",
              len(agent_headers) >= 5,
              f"Got {len(agent_headers)} columns: {agent_headers}")

        # Per-row checks; accumulate flags for the critical aggregate checks.
        seen_categories = set()
        all_bundle_prices_ok = True
        all_recipe_names_ok = True
        all_difficulty_ok = True
        for i, row in enumerate(agent_rows):
            if not row or len(row) < 5:
                check(f"Bundle row {i+1} has enough columns", False,
                      f"Row has {len(row) if row else 0} columns")
                all_bundle_prices_ok = False
                all_recipe_names_ok = False
                all_difficulty_ok = False
                continue

            recipe_name, recipe_cat, difficulty, paired_product, bundle_price = (
                row[0], row[1], row[2], row[3], row[4]
            )

            # Recipe name should be a non-empty Cyrillic (kulinar) name.
            name_ok = (recipe_name is not None
                       and len(str(recipe_name).strip()) > 0
                       and has_cyrillic(recipe_name))
            check(f"Bundle {i+1} has a Cyrillic recipe name",
                  name_ok, f"Got: {recipe_name}")
            if not name_ok:
                all_recipe_names_ok = False

            # Recipe category should be a valid kulinar category.
            cat_str = str(recipe_cat).strip().lower() if recipe_cat else ""
            check(f"Bundle {i+1} category is a valid kulinar category",
                  cat_str in VALID_RECIPE_CATEGORIES,
                  f"Got '{cat_str}', expected one of {sorted(VALID_RECIPE_CATEGORIES)}")
            seen_categories.add(cat_str)

            # Difficulty should be an integer 1-4.
            diff_ok = False
            try:
                dv = int(float(str(difficulty).strip()))
                diff_ok = 1 <= dv <= 4
            except (TypeError, ValueError):
                diff_ok = False
            check(f"Bundle {i+1} difficulty is integer 1-4",
                  diff_ok, f"Got: {difficulty}")
            if not diff_ok:
                all_difficulty_ok = False

            # Paired product should be non-empty.
            check(f"Bundle {i+1} has paired product name",
                  paired_product is not None and len(str(paired_product).strip()) > 0,
                  f"Got: {paired_product}")

            # Bundle price should be a real product price * 0.85.
            if use_db and paired_product:
                product_prices = expected['product_prices_by_name']
                product_name_lower = str(paired_product).strip().lower()
                matched_price = None
                for pname, pprice in product_prices.items():
                    if pname == product_name_lower or pname[:30] in product_name_lower or product_name_lower[:30] in pname:
                        matched_price = pprice
                        break

                if matched_price is not None:
                    expected_bundle = round(matched_price * 0.85, 2)
                    row_ok = num_close(bundle_price, expected_bundle, 0.5)
                    check(f"Bundle {i+1} price = real product price * 0.85",
                          row_ok,
                          f"Expected ~{expected_bundle} (product price {matched_price} * 0.85), got {bundle_price}")
                    if not row_ok:
                        all_bundle_prices_ok = False
                else:
                    # Paired product is not a real Бытовая техника product.
                    check(f"Bundle {i+1} paired product is a real Бытовая техника product",
                          False,
                          f"'{paired_product}' not found among Home Appliances products")
                    all_bundle_prices_ok = False
            else:
                price_ok = bundle_price is not None
                try:
                    price_ok = price_ok and float(bundle_price) > 0
                except (TypeError, ValueError):
                    price_ok = False
                check(f"Bundle {i+1} has numeric bundle price",
                      price_ok, f"Got: {bundle_price}")
                if not price_ok:
                    all_bundle_prices_ok = False

        # CRITICAL: all three required kulinar categories represented (one each).
        check("Recipe Bundles covers напиток (drinks) category",
              'напиток' in seen_categories,
              f"Categories found: {seen_categories}", critical=True)
        check("Recipe Bundles covers гарнир (side/staple) category",
              'гарнир' in seen_categories,
              f"Categories found: {seen_categories}", critical=True)
        check("Recipe Bundles covers горячее (hot main) category",
              'горячее' in seen_categories,
              f"Categories found: {seen_categories}", critical=True)

        # CRITICAL: every bundle row is a valid recipe+product pairing.
        if len(agent_rows) == 3:
            check("All 3 recipe names are non-empty Cyrillic kulinar names",
                  all_recipe_names_ok, "", critical=True)
            check("All 3 difficulties are integers 1-4",
                  all_difficulty_ok, "", critical=True)
            if use_db:
                check("All 3 Bundle_Price == real Бытовая техника product price * 0.85",
                      all_bundle_prices_ok, "", critical=True)

    return True


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    """Run all evaluation checks."""
    db_expected = compute_expected_from_db()
    use_db = db_expected is not None

    if use_db:
        print("INFO: Using dynamically computed expected values from PostgreSQL")
    else:
        print("INFO: Falling back to static groundtruth Excel file")

    check_excel(agent_workspace, db_expected, groundtruth_workspace)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100.0) if total else 0.0

    print(f"\n=== SUMMARY ===")
    print(f"  Passed: {PASS_COUNT}")
    print(f"  Failed: {FAIL_COUNT}")
    print(f"  Accuracy: {accuracy:.1f}%")
    print(f"  Source: {'postgresql' if use_db else 'groundtruth_excel'}")
    print(f"  Critical failures: {CRITICAL_FAILURES}")

    critical_ok = len(CRITICAL_FAILURES) == 0
    success = critical_ok and accuracy >= 70.0

    if not critical_ok:
        print(f"=== RESULT: FAIL (critical checks failed: {CRITICAL_FAILURES}) ===")
    elif success:
        print(f"=== RESULT: PASS (accuracy {accuracy:.1f}%) ===")
    else:
        print(f"=== RESULT: FAIL (accuracy {accuracy:.1f}% < 70%) ===")

    if res_log_file:
        result = {
            "passed": PASS_COUNT,
            "failed": FAIL_COUNT,
            "accuracy": accuracy,
            "success": success,
            "critical_failures": CRITICAL_FAILURES,
            "source": "postgresql" if use_db else "groundtruth_excel",
        }
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    return success, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, Accuracy: {accuracy:.1f}%"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace,
        args.groundtruth_workspace,
        args.launch_time,
        args.res_log_file,
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
