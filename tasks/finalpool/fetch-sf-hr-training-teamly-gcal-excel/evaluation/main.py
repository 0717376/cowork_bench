"""Evaluation for fetch-sf-hr-training (ClickHouse warehouse + Teamly + GCal + Excel + email).

Critical checks (CRITICAL_CHECKS): any failure => overall FAIL regardless of
accuracy. Otherwise PASS requires accuracy >= 70%.

The warehouse (sf_data, logical DB HR_ANALYTICS) is russified centrally: department
names and education levels are stored in Russian (Операции/Продажи/...,
Среднее образование/Диплом/...). Expected per-department metrics are recomputed
LIVE from the warehouse at evaluation time, then matched against the agent output
through a RU<->EN canonical department map (the agent legitimately writes Russian
names, the groundtruth xlsx carries the English ones).
"""
import argparse
import os
import sys

import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
FAILED_NAMES = []

# Canonical department key <- accepted aliases (RU realia + English groundtruth).
DEPT_ALIASES = {
    "operations": ["operations", "операции"],
    "sales": ["sales", "продажи"],
    "rnd": ["r&d", "rnd", "r&d", "research & development", "research and development", "ниокр"],
    "finance": ["finance", "финансы"],
    "hr": ["hr", "human resources", "кадры"],
    "support": ["support", "поддержка"],
    "engineering": ["engineering", "инженерия"],
}
ALIAS_TO_CANON = {a.strip().lower(): canon for canon, al in DEPT_ALIASES.items() for a in al}

# Education levels counted as "low education" (RU realia + English).
LOW_EDU_VALUES = {"high school", "diploma", "среднее образование", "диплом"}

# Expected assigned course per priority department (skill-area match), with the
# cost_per_attendee from the catalog. Accept either RU or EN dept naming.
EXPECTED_COURSE = {
    "operations": ("Advanced Operations Management", 850),
    "sales": ("Sales Excellence Program", 750),
    "rnd": ("R&D Innovation Workshop", 1200),
}

# Critical checks: any failure => overall FAIL regardless of accuracy.
CRITICAL_CHECKS = {
    "Budget Allocation lists the 3 lowest-avg-performance departments (RU or EN)",
    "Department Analysis Avg_Performance matches the live warehouse per department",
    "Department Analysis Low_Education_Pct matches the live warehouse share for priority depts",
    "Estimated_Cost/Total_Cost equals cost_per_attendee x low-education headcount for priority depts",
    "Three GCal training events exist on 2026-05-04, 2026-05-11, 2026-05-18",
}


def check(name, condition, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        FAILED_NAMES.append(name)
        detail_str = str(detail)[:240] if detail else ""
        print(f"  [FAIL] {name}: {detail_str}")


def canon_dept(name):
    return ALIAS_TO_CANON.get(str(name or "").strip().lower())


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(",", "").replace("%", "").replace("$", "").replace("₽", "").strip())
    except (TypeError, ValueError):
        return default


def num_close(a, b, tol=1.0):
    fa, fb = safe_float(a), safe_float(b)
    if fa is None or fb is None:
        return False
    return abs(fa - fb) <= tol


def load_expected_depts():
    """Recompute per-department metrics LIVE from the warehouse, keyed by canonical
    dept. Returns {canon: {avg_perf, total, low_edu, low_pct}}."""
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT
                "DEPARTMENT",
                COUNT(*) AS total,
                AVG("PERFORMANCE_RATING")::float AS avg_perf,
                SUM(CASE WHEN lower("EDUCATION_LEVEL") IN
                    ('high school','diploma','среднее образование','диплом')
                    THEN 1 ELSE 0 END) AS low_edu
            FROM sf_data."HR_ANALYTICS__PUBLIC__EMPLOYEES"
            GROUP BY "DEPARTMENT"
        """)
        rows = cur.fetchall()
        cur.close()
    finally:
        conn.close()
    out = {}
    for dept, total, avg_perf, low_edu in rows:
        canon = canon_dept(dept)
        if canon is None:
            continue
        total = int(total)
        low_edu = int(low_edu)
        out[canon] = {
            "avg_perf": round(float(avg_perf), 4),
            "total": total,
            "low_edu": low_edu,
            "low_pct": round(low_edu / total * 100, 2) if total else 0.0,
        }
    return out


def lowest_three(expected):
    """Three canonical depts with the lowest avg performance (stable tie-break by
    canon name). Returns (priority, candidates) where:
      - priority: the deterministic top-3 list (lowest first).
      - candidates: the set of depts that could legitimately fill the 3rd slot
        because their avg perf is within a tiny window of the 3rd-place value
        (the seven departments are near-tied at ~3.2, so the 3rd slot is fuzzy).
    The two strictly-lowest depts (Operations, Sales) are unambiguous; only the
    3rd-place membership is tolerant."""
    ordered = sorted(expected.items(), key=lambda kv: (kv[1]["avg_perf"], kv[0]))
    priority = [c for c, _ in ordered[:3]]
    if len(ordered) < 3:
        return priority, set(priority)
    third_val = ordered[2][1]["avg_perf"]
    candidates = {c for c, v in ordered if v["avg_perf"] <= third_val + 0.02}
    return priority, candidates


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def check_excel(agent_ws, expected, priority, candidates):
    import openpyxl

    agent_file = os.path.join(agent_ws, "Training_Budget_Plan.xlsx")
    print("Checking Excel file...")
    if not os.path.exists(agent_file):
        check("Training_Budget_Plan.xlsx exists", False, agent_file)
        # Cascade-fail the dependent critical checks so absence => FAIL.
        check("Department Analysis Avg_Performance matches the live warehouse per department", False, "no file")
        check("Department Analysis Low_Education_Pct matches the live warehouse share for priority depts", False, "no file")
        check("Estimated_Cost/Total_Cost equals cost_per_attendee x low-education headcount for priority depts", False, "no file")
        check("Budget Allocation lists the 3 lowest-avg-performance departments (RU or EN)", False, "no file")
        return
    check("Training_Budget_Plan.xlsx exists", True)
    wb = openpyxl.load_workbook(agent_file, data_only=True)

    # --- Department Analysis ---
    da = load_sheet_rows(wb, "Department Analysis")
    check("Sheet 'Department Analysis' present", da is not None)
    da_lookup = {}
    if da:
        header = [str(c).strip() if c is not None else "" for c in da[0]]

        def col(name):
            for i, h in enumerate(header):
                if h.strip().lower() == name.strip().lower():
                    return i
            return None

        c_dept = col("Department")
        c_avg = col("Avg_Performance")
        c_low = col("Low_Education_Pct")
        c_hs = col("High_School_Count")
        c_dip = col("Diploma_Count")
        for row in da[1:]:
            if not row or c_dept is None or row[c_dept] is None:
                continue
            canon = canon_dept(row[c_dept])
            if canon:
                da_lookup[canon] = (row, c_avg, c_low, c_hs, c_dip)
        check("Department Analysis has 7 department rows", len(da_lookup) >= 7,
              f"found {len(da_lookup)}")

        # Critical: Avg_Performance per department vs live warehouse.
        avg_ok = True
        for canon, exp in expected.items():
            entry = da_lookup.get(canon)
            if entry is None or entry[1] is None:
                avg_ok = False
                break
            row, c_avg, *_ = entry
            if not num_close(row[c_avg], exp["avg_perf"], 0.05):
                avg_ok = False
                break
        check("Department Analysis Avg_Performance matches the live warehouse per department",
              avg_ok, "per-dept avg perf mismatch (tol 0.05)")

        # Critical: Low_Education_Pct (and HS+Diploma counts) per department vs the
        # live warehouse. Validated for ALL departments (unambiguous, avoids the
        # near-tie at the 3rd priority slot), since this share drives Estimated_Cost.
        lowpct_ok = True
        detail = ""
        for canon, exp in expected.items():
            entry = da_lookup.get(canon)
            if entry is None or entry[2] is None:
                lowpct_ok = False
                detail = f"{canon}: missing Low_Education_Pct"
                break
            row, _, c_low, c_hs, c_dip = entry
            if not num_close(row[c_low], exp.get("low_pct"), 1.5):
                lowpct_ok = False
                detail = f"{canon}: pct {row[c_low]} vs {exp.get('low_pct')}"
                break
            # HS+Diploma counts must sum to the low-education headcount.
            if c_hs is not None and c_dip is not None:
                s = (safe_float(row[c_hs], 0) or 0) + (safe_float(row[c_dip], 0) or 0)
                if not num_close(s, exp.get("low_edu"), 10):
                    lowpct_ok = False
                    detail = f"{canon}: HS+Dip {s} vs {exp.get('low_edu')}"
                    break
        check("Department Analysis Low_Education_Pct matches the live warehouse share for priority depts",
              lowpct_ok, detail)
    else:
        check("Department Analysis has 7 department rows", False, "no sheet")
        check("Department Analysis Avg_Performance matches the live warehouse per department", False, "no sheet")
        check("Department Analysis Low_Education_Pct matches the live warehouse share for priority depts", False, "no sheet")

    # --- Training Catalog ---
    tc = load_sheet_rows(wb, "Training Catalog")
    if tc is None:
        check("Training Catalog sheet present with >=12 courses", False, "no sheet")
    else:
        body = [r for r in tc[1:] if r and r[0] and str(r[0]).strip()]
        check("Training Catalog sheet present with >=12 courses", len(body) >= 12,
              f"{len(body)} rows")

    # --- Budget Allocation ---
    ba = load_sheet_rows(wb, "Budget Allocation")
    if ba is None:
        check("Budget Allocation sheet present with 3 department rows", False, "no sheet")
        check("Budget Allocation lists the 3 lowest-avg-performance departments (RU or EN)", False, "no sheet")
        check("Estimated_Cost/Total_Cost equals cost_per_attendee x low-education headcount for priority depts", False, "no sheet")
        return

    header = [str(c).strip() if c is not None else "" for c in ba[0]]

    def bcol(name):
        for i, h in enumerate(header):
            if h.strip().lower() == name.strip().lower():
                return i
        return None

    c_dept = bcol("Department")
    c_total = bcol("Total_Cost")
    c_trainees = bcol("Trainees")
    body = [r for r in ba[1:] if r and c_dept is not None and r[c_dept] and str(r[c_dept]).strip()]
    check("Budget Allocation sheet present with 3 department rows", len(body) >= 3,
          f"{len(body)} rows")

    present = {canon_dept(r[c_dept]) for r in body}
    present.discard(None)
    # Critical: the 2 unambiguously-lowest depts must be allocated, and every
    # allocated dept must be a legitimate low-performer (within the candidate
    # window). The 3rd slot is fuzzy because the 7 depts are near-tied at ~3.2.
    strict = set(priority[:2])
    missing_strict = [p for p in strict if p not in present]
    all_in_window = present.issubset(candidates) and len(present) >= 3
    check("Budget Allocation lists the 3 lowest-avg-performance departments (RU or EN)",
          not missing_strict and all_in_window,
          f"missing_strict {missing_strict}; present {present}; candidates {candidates}")

    # Critical: Total_Cost == cost_per_attendee(assigned course) x low-edu headcount,
    # for at least the two lowest departments (Operations, Sales by construction).
    cost_ok = True
    detail = ""
    checked = 0
    by_canon = {}
    for r in body:
        c = canon_dept(r[c_dept])
        if c:
            by_canon[c] = r
    for canon in priority[:2]:
        exp = expected.get(canon, {})
        course = EXPECTED_COURSE.get(canon)
        r = by_canon.get(canon)
        if r is None or course is None or c_total is None:
            cost_ok = False
            detail = f"{canon}: row/course/Total_Cost missing"
            break
        expected_cost = course[1] * exp.get("low_edu", 0)
        # tolerance: 2% of expected to absorb sessions rounding interpretations
        tol = max(50.0, expected_cost * 0.02)
        if not num_close(r[c_total], expected_cost, tol):
            cost_ok = False
            detail = f"{canon}: Total_Cost {r[c_total]} vs {expected_cost}"
            break
        checked += 1
    check("Estimated_Cost/Total_Cost equals cost_per_attendee x low-education headcount for priority depts",
          cost_ok and checked >= 2, detail)


def check_teamly(priority, candidates):
    print("Checking Teamly training program...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("SELECT id, title, COALESCE(body, ''), space_id FROM teamly.pages WHERE id > 3")
        pages = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Teamly 'Training Program 2026' tracker exists", False, str(e))
        check("Teamly has >=3 department training pages", False, str(e))
        return

    # The tracker: a page or space carrying the 'training program' marker (English
    # identifier preserved); the noise page ('справочник сотрудников') must not count.
    real = [p for p in pages if "справочник сотрудников" not in (p[1] or "").lower()]

    def looks_tracker(t):
        tl = (t or "").lower()
        return ("training" in tl and "program" in tl) or "training program 2026" in tl

    tracker_pages = [p for p in real if looks_tracker(p[1])]
    tracker_present = bool(tracker_pages)
    if not tracker_present:
        # fallback: a training space exists and contains >=3 dept pages
        try:
            conn = psycopg2.connect(**DB_CONFIG)
            cur = conn.cursor()
            cur.execute("SELECT id FROM teamly.spaces WHERE lower(name) LIKE '%%обучен%%' OR lower(key) LIKE '%%train%%'")
            tracker_present = cur.fetchone() is not None
            cur.close(); conn.close()
        except Exception:
            pass
    check("Teamly 'Training Program 2026' tracker exists", tracker_present,
          f"new pages: {[(p[0], p[1]) for p in real]}")

    # >=3 department pages: pages (other than the tracker/noise) whose title or body
    # references one of the priority departments (RU or EN).
    dept_pages = 0
    for pid, title, body, sid in real:
        text = ((title or "") + " " + (body or "")).lower()
        if canon_dept(title) in candidates or any(
            alias in text for canon in candidates for alias in DEPT_ALIASES[canon]
        ):
            dept_pages += 1
    check("Teamly has >=3 department training pages", dept_pages >= 3,
          f"dept-referencing pages: {dept_pages}")


def check_gcal():
    print("Checking GCal events...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT summary, start_datetime::date
            FROM gcal.events
            WHERE summary ILIKE '%training%'
            ORDER BY start_datetime
        """)
        rows = cur.fetchall()
        cur.close()
        conn.close()
    except Exception as e:
        check("Three GCal training events exist on 2026-05-04, 2026-05-11, 2026-05-18", False, str(e))
        return
    dates = {str(r[1]) for r in rows}
    needed = ["2026-05-04", "2026-05-11", "2026-05-18"]
    missing = [d for d in needed if d not in dates]
    check("Three GCal training events exist on 2026-05-04, 2026-05-11, 2026-05-18",
          len(rows) >= 3 and not missing,
          f"events={len(rows)}, dates={sorted(dates)}, missing={missing}")


def check_email():
    print("Checking email...")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
        cur.execute("""
            SELECT COUNT(*) FROM email.messages
            WHERE to_addr::text ILIKE '%department-managers@company.com%'
            AND subject ILIKE '%training%'
        """)
        count = cur.fetchone()[0]
        cur.close()
        conn.close()
    except Exception as e:
        check("Email to department-managers with training subject sent", False, str(e))
        return
    check("Email to department-managers with training subject sent", count > 0, f"count={count}")


def run_evaluation(agent_ws):
    global PASS_COUNT, FAIL_COUNT, FAILED_NAMES
    PASS_COUNT = 0
    FAIL_COUNT = 0
    FAILED_NAMES = []

    expected = load_expected_depts()
    priority, candidates = lowest_three(expected)
    print(f"Live priority departments (lowest avg perf): {priority}; candidates: {candidates}")

    check_excel(agent_ws, expected, priority, candidates)
    check_teamly(priority, candidates)
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total else 0
    critical_failed = [n for n in FAILED_NAMES if n in CRITICAL_CHECKS]

    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")
    if critical_failed:
        print(f"CRITICAL FAILURES: {len(critical_failed)}")
        for n in critical_failed:
            print(f"  - {n}")

    success = (not critical_failed) and accuracy >= 70
    return success


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    agent_ws = args.agent_workspace or task_root

    success = run_evaluation(agent_ws)

    if success:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print("\n=== RESULT: FAIL ===")
        sys.exit(1)


if __name__ == "__main__":
    main()
