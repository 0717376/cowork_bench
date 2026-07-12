"""Evaluation for canvas-grade-summary."""
import argparse
import os
import sys
import openpyxl


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def course_match(agent_val, gt_val):
    """Match course names tolerantly: gt may be truncated, agent may be full.
    Pass if either is a prefix/substring of the other (case-insensitive)."""
    if agent_val is None or gt_val is None:
        return False
    a = str(agent_val).strip().lower()
    g = str(gt_val).strip().lower()
    if not a or not g:
        return False
    return a == g or a.startswith(g) or g.startswith(a) or a in g or g in a


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    agent_file = os.path.join(args.agent_workspace, "Canvas_Grade_Summary.xlsx")
    gt_file = os.path.join(gt_dir, "Canvas_Grade_Summary.xlsx")

    if not os.path.exists(agent_file):
        print(f"FAIL: Agent output not found: {agent_file}")
        sys.exit(1)
    if not os.path.exists(gt_file):
        print(f"FAIL: Groundtruth not found: {gt_file}")
        sys.exit(1)

    agent_wb = openpyxl.load_workbook(agent_file, data_only=True)
    gt_wb = openpyxl.load_workbook(gt_file, data_only=True)

    all_errors = []
    critical_failures = []

    # ---- Load Grade Summary rows ----
    a_gs = load_sheet_rows(agent_wb, "Grade Summary")
    g_gs = load_sheet_rows(gt_wb, "Grade Summary")

    a_gs_lookup = {}
    if a_gs:
        for row in a_gs[1:]:
            if row and row[0] is not None:
                a_gs_lookup[str(row[0]).strip().lower()] = row
    g_gs_data = g_gs[1:] if g_gs and len(g_gs) > 1 else []

    # ---- Load Summary rows ----
    a_sum = load_sheet_rows(agent_wb, "Summary")
    g_sum = load_sheet_rows(gt_wb, "Summary")

    a_sum_lookup = {}
    if a_sum:
        for row in a_sum[1:]:
            if row and row[0] is not None:
                a_sum_lookup[str(row[0]).strip().lower()] = row
    g_sum_lookup = {}
    if g_sum:
        for row in g_sum[1:]:
            if row and row[0] is not None:
                g_sum_lookup[str(row[0]).strip().lower()] = row

    # =====================================================================
    # CRITICAL CHECKS — any failure => immediate FAIL before accuracy gate.
    # These reflect the SUBSTANCE of the deliverable: completeness of the
    # course breakdown, correctness of the distinct-student aggregation,
    # the exact course count, the correct top course, and the overall avg.
    # =====================================================================
    print("=== CRITICAL CHECKS ===")

    # C1: every groundtruth course row present in agent output (completeness).
    if a_gs is None:
        critical_failures.append("CRITICAL: Sheet 'Grade Summary' missing in agent output")
    elif not g_gs_data:
        critical_failures.append("CRITICAL: Sheet 'Grade Summary' missing in groundtruth")
    else:
        missing = []
        for g_row in g_gs_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            if key not in a_gs_lookup:
                missing.append(g_row[0])
        if missing:
            critical_failures.append(
                f"CRITICAL: missing {len(missing)} course rows, e.g. {missing[:3]}")
        else:
            print(f"  [OK] all {len(g_gs_data)} course rows present")

    # C2: Avg_Score per course matches groundtruth within tol<=1.0
    #     (proves correct distinct-student aggregation, not a near-miss).
    if a_gs is not None and g_gs_data:
        avg_errs = []
        for g_row in g_gs_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_gs_lookup.get(key)
            if a_row is None:
                continue  # already counted in C1
            if len(a_row) > 2 and len(g_row) > 2:
                if not num_close(a_row[2], g_row[2], 1.0):
                    avg_errs.append(f"{g_row[0]}: {a_row[2]} vs {g_row[2]}")
        if avg_errs:
            critical_failures.append(
                f"CRITICAL: {len(avg_errs)} Avg_Score mismatches (tol=1.0), e.g. {avg_errs[:3]}")
        else:
            print("  [OK] Avg_Score matches groundtruth (tol=1.0)")

    # C3: Summary.Total_Courses == 22 EXACTLY (tol=0).
    g_tc = g_sum_lookup.get("total_courses")
    a_tc = a_sum_lookup.get("total_courses")
    if g_tc is None:
        critical_failures.append("CRITICAL: Total_Courses missing in groundtruth")
    elif a_tc is None:
        critical_failures.append("CRITICAL: Total_Courses missing in agent Summary")
    else:
        if num_close(a_tc[1] if len(a_tc) > 1 else None,
                     g_tc[1] if len(g_tc) > 1 else None, 0):
            print(f"  [OK] Total_Courses == {g_tc[1]} (exact)")
        else:
            critical_failures.append(
                f"CRITICAL: Total_Courses {a_tc[1] if len(a_tc)>1 else None} != {g_tc[1]} (exact)")

    # C4: Summary.Highest_Avg_Course matches the top course (substring-tolerant).
    g_hac = g_sum_lookup.get("highest_avg_course")
    a_hac = a_sum_lookup.get("highest_avg_course")
    if g_hac is None:
        critical_failures.append("CRITICAL: Highest_Avg_Course missing in groundtruth")
    elif a_hac is None:
        critical_failures.append("CRITICAL: Highest_Avg_Course missing in agent Summary")
    else:
        g_name = g_hac[1] if len(g_hac) > 1 else None
        a_name = a_hac[1] if len(a_hac) > 1 else None
        if course_match(a_name, g_name):
            print(f"  [OK] Highest_Avg_Course matches ({a_name})")
        else:
            critical_failures.append(
                f"CRITICAL: Highest_Avg_Course '{a_name}' != '{g_name}'")

    # C5: Summary.Overall_Avg_Score matches groundtruth within tol<=0.5.
    g_oa = g_sum_lookup.get("overall_avg_score")
    a_oa = a_sum_lookup.get("overall_avg_score")
    if g_oa is None:
        critical_failures.append("CRITICAL: Overall_Avg_Score missing in groundtruth")
    elif a_oa is None:
        critical_failures.append("CRITICAL: Overall_Avg_Score missing in agent Summary")
    else:
        if num_close(a_oa[1] if len(a_oa) > 1 else None,
                     g_oa[1] if len(g_oa) > 1 else None, 0.5):
            print(f"  [OK] Overall_Avg_Score matches (tol=0.5)")
        else:
            critical_failures.append(
                f"CRITICAL: Overall_Avg_Score {a_oa[1] if len(a_oa)>1 else None} "
                f"!= {g_oa[1] if len(g_oa)>1 else None} (tol=0.5)")

    if critical_failures:
        print("\n=== CRITICAL FAILURES ===")
        for c in critical_failures:
            print(f"  {c}")
        print("\n=== RESULT: FAIL (critical) ===")
        sys.exit(1)
    print("  All critical checks passed.")

    # =====================================================================
    # NON-CRITICAL accuracy checks (structural / loose-tolerance cells).
    # =====================================================================
    total_checks = 0
    passed_checks = 0

    print("\n  Checking Grade Summary (non-critical)...")
    if a_gs is None:
        all_errors.append("Sheet 'Grade Summary' not found in agent output")
    else:
        errors = []
        for g_row in g_gs_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_gs_lookup.get(key)
            if a_row is None:
                continue

            if len(a_row) > 1 and len(g_row) > 1:
                total_checks += 1
                if num_close(a_row[1], g_row[1], 2):
                    passed_checks += 1
                else:
                    errors.append(f"{key}.Students_Submitted: {a_row[1]} vs {g_row[1]} (tol=2)")

            if len(a_row) > 3 and len(g_row) > 3:
                total_checks += 1
                if num_close(a_row[3], g_row[3], 2.0):
                    passed_checks += 1
                else:
                    errors.append(f"{key}.Max_Score: {a_row[3]} vs {g_row[3]} (tol=2.0)")

            if len(a_row) > 4 and len(g_row) > 4:
                total_checks += 1
                if num_close(a_row[4], g_row[4], 2.0):
                    passed_checks += 1
                else:
                    errors.append(f"{key}.Min_Score: {a_row[4]} vs {g_row[4]} (tol=2.0)")
        if errors:
            all_errors.extend(errors)
            print(f"    ERRORS: {len(errors)}")
            for e in errors[:5]:
                print(f"      {e}")
        else:
            print("    PASS")

    accuracy = (passed_checks / total_checks * 100.0) if total_checks else 100.0
    print(f"\n=== Accuracy: {passed_checks}/{total_checks} = {accuracy:.1f}% ===")

    if accuracy >= 70:
        print("\n=== RESULT: PASS ===")
        sys.exit(0)
    else:
        print(f"\n=== RESULT: FAIL ({len(all_errors)} non-critical errors) ===")
        for e in all_errors[:10]:
            print(f"  {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
