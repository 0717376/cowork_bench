"""
Evaluation for yt-fireship-tech-report-excel-notion task.

Checks:
1. Tech_Trend_Report.xlsx exists with "Videos" sheet having >= 8 data rows
2. "Videos" sheet has Video_ID, Title, View_Count columns (case-insensitive)
3. "Topic_Summary" sheet exists with >= 3 topic rows
4. Topic_Summary has Topic and Video_Count columns
5. Topic_Summary aggregation matches groundtruth: per-topic Total_Views are
   compared label-independently (the GT is derived from the seeded top-10
   Fireship 2024-2025 videos). Two conditions: (A) the agent's per-topic
   Total_Views sum equals the GT grand total (= real top-10 total views), and
   (B) each GT top topic total is matched by value (+/-10%) by some agent topic
   total. Primary_Topic labels are agent-inferred, so they are NOT matched as
   strings.
6. Teamly page exists in space TEAM with title containing "Fireship" or "Tech Trends"
7. GCal has a new event in March 2026 with "Tech Review" in summary (not the noise Team Sync)
8. Email was sent to techteam@company.com

CRITICAL_CHECKS gate the result: if any critical check fails, the task FAILS
regardless of overall accuracy. They capture the substance of the task
(real video data pulled, correct topic aggregation, page published, meeting
scheduled, team notified) — not merely structural shape.
"""
import os
import sys
import json
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Semantic checks that gate the result. Any failure => overall FAIL.
CRITICAL_CHECKS = {
    "Videos sheet has >= 8 data rows",
    "Topic_Summary Total_Views per topic matches groundtruth",
    "Teamly page (TEAM) with 'Fireship'/'Tech Trends' in title exists",
    "GCal has 'Tech Review' event in March 2026",
    "Email sent to techteam@company.com",
}


def record(name, passed, detail=""):
    global PASS_COUNT, FAIL_COUNT
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {name}")
    else:
        FAIL_COUNT += 1
        if name in CRITICAL_CHECKS:
            CRITICAL_FAILS.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        tag = " [CRITICAL]" if name in CRITICAL_CHECKS else ""
        print(f"  [FAIL]{tag} {name}{msg}")


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def _seed_view_counts():
    """view_count реальных топ-10 видео Fireship 2024-2025 из сида БД."""
    conn = psycopg2.connect(**DB_CONFIG)
    with conn.cursor() as cur:
        cur.execute("""
            SELECT view_count FROM youtube.videos
            WHERE channel_title ILIKE '%fireship%'
              AND published_at >= '2024-01-01' AND published_at < '2026-01-01'
            ORDER BY view_count DESC LIMIT 10
        """)
        counts = [int(r[0]) for r in cur.fetchall()]
    conn.close()
    return counts


def _partition_valid(targets, counts):
    """Каждый target — сумма непересекающегося подмножества counts,
    и все counts использованы (exact cover)."""
    from itertools import combinations
    targets = sorted((int(round(float(t))) for t in targets), reverse=True)
    counts = sorted((int(c) for c in counts), reverse=True)
    if sum(targets) != sum(counts):
        return False

    def fit(ti, remaining):
        if ti == len(targets):
            return not remaining
        target = targets[ti]
        n = len(remaining)
        for k in range(1, n + 1):
            for combo in combinations(range(n), k):
                if sum(remaining[i] for i in combo) == target:
                    rest = [remaining[i] for i in range(n) if i not in combo]
                    if fit(ti + 1, rest):
                        return True
        return False

    return fit(0, counts)


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-5: Tech_Trend_Report.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Tech_Trend_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Tech_Trend_Report.xlsx exists", False, f"Not found at {xlsx_path}")
        record("Videos sheet has >= 8 data rows", False, "File missing")
        record("Videos sheet has required columns", False, "File missing")
        record("Topic_Summary sheet exists with >= 3 rows", False, "File missing")
        record("Topic_Summary has Topic and Video_Count columns", False, "File missing")
        record("Topic_Summary Total_Views per topic matches groundtruth", False, "File missing")
        return
    record("Tech_Trend_Report.xlsx exists", True)

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        record("Excel file readable", False, str(e))
        record("Videos sheet has >= 8 data rows", False, "File unreadable")
        record("Topic_Summary Total_Views per topic matches groundtruth", False, "File unreadable")
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Check Videos sheet
    videos_key = None
    videos_topic_views = []  # (primary_topic, view_count) для fallback-агрегации
    for k in sheet_names_lower:
        if "video" in k:
            videos_key = sheet_names_lower[k]
            break
    if not videos_key:
        record("Videos sheet exists", False, f"Sheets: {wb.sheetnames}")
        record("Videos sheet has >= 8 data rows", False, "Sheet missing")
        record("Videos sheet has required columns", False, "Sheet missing")
    else:
        ws = wb[videos_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        record("Videos sheet has >= 8 data rows", len(data_rows) >= 8,
               f"Found {len(data_rows)} data rows")
        if rows:
            headers = [str(c).strip().lower() if c else "" for c in rows[0]]
            has_video_id = any("video_id" in h or "videoid" in h or "video id" in h for h in headers)
            has_title = any("title" in h for h in headers)
            has_views = any("view" in h for h in headers)
            record("Videos sheet has required columns (Video_ID, Title, View_Count)",
                   has_video_id and has_title and has_views,
                   f"Headers: {rows[0]}")
            v_idx = next((i for i, h in enumerate(headers) if "view" in h), None)
            t_idx = next((i for i, h in enumerate(headers) if "topic" in h), None)
            if v_idx is not None and t_idx is not None:
                for r in data_rows:
                    try:
                        if r[t_idx] is not None and r[v_idx] is not None:
                            videos_topic_views.append(
                                (str(r[t_idx]).strip().lower(), float(r[v_idx])))
                    except (TypeError, ValueError, IndexError):
                        pass
        else:
            record("Videos sheet has required columns", False, "Sheet is empty")

    # Check Topic_Summary sheet
    topic_key = None
    for k in sheet_names_lower:
        if "topic" in k or "summary" in k:
            topic_key = sheet_names_lower[k]
            break
    agent_topic_totals = {}
    agent_total_views_list = []  # all per-topic Total_Views values (label-independent)
    if not topic_key:
        record("Topic_Summary sheet exists with >= 3 rows", False, f"Sheets: {wb.sheetnames}")
        record("Topic_Summary has Topic and Video_Count columns", False, "Sheet missing")
    else:
        ws2 = wb[topic_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        record("Topic_Summary sheet exists with >= 3 rows", len(data_rows2) >= 3,
               f"Found {len(data_rows2)} data rows")
        if rows2:
            headers2 = [str(c).strip().lower() if c else "" for c in rows2[0]]
            has_topic = any("topic" in h for h in headers2)
            has_count = any("count" in h or "video" in h for h in headers2)
            record("Topic_Summary has Topic and Video_Count columns",
                   has_topic and has_count,
                   f"Headers: {rows2[0]}")
            # Build topic -> total_views map from agent sheet
            topic_idx = next((i for i, h in enumerate(headers2) if "topic" in h), 0)
            total_idx = next((i for i, h in enumerate(headers2)
                              if "total" in h and "view" in h), None)
            if total_idx is None:
                total_idx = next((i for i, h in enumerate(headers2) if "view" in h), None)
            if total_idx is not None:
                for r in data_rows2:
                    try:
                        topic = str(r[topic_idx]).strip().lower()
                        agent_topic_totals[topic] = r[total_idx]
                        if r[total_idx] is not None:
                            agent_total_views_list.append(float(r[total_idx]))
                    except Exception:
                        pass
        else:
            record("Topic_Summary has Topic and Video_Count columns", False, "Sheet is empty")

    # --- Semantic: Topic_Summary totals образуют валидное разбиение топ-10 ---
    # Primary_Topic — свободная метка агента, поэтому вместо сравнения с одной
    # произвольной GT-разбивкой проверяем: (A) сумма per-topic Total_Views равна
    # сумме просмотров реальных топ-10 видео; (B) каждый per-topic Total_Views —
    # сумма непересекающегося подмножества реальных view_count (exact cover).
    real_counts = []
    try:
        real_counts = _seed_view_counts()
    except Exception as e:
        print(f"  WARNING: seed view counts unavailable from DB: {e}")
    if len(real_counts) < 10:
        # Fallback: лист Videos groundtruth-файла содержит те же топ-10.
        gt_path = os.path.join(groundtruth_workspace, "Tech_Trend_Report.xlsx")
        if os.path.isfile(gt_path):
            gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
            for sn in gt_wb.sheetnames:
                if "video" in sn.lower():
                    g_rows = list(gt_wb[sn].iter_rows(values_only=True))
                    gh = [str(c).strip().lower() if c else "" for c in g_rows[0]]
                    gvi = next((i for i, h in enumerate(gh) if "view" in h), None)
                    if gvi is not None:
                        real_counts = [int(r[gvi]) for r in g_rows[1:]
                                       if r and r[gvi] is not None]
                    break
            gt_wb.close()

    # Fallback: Total_Views в Topic_Summary записаны формулами без кэша
    # (openpyxl с data_only=True читает None) — пересобираем суммы по
    # Primary_Topic из листа Videos агента.
    if not agent_total_views_list and videos_topic_views:
        regroup = {}
        for t, v in videos_topic_views:
            regroup[t] = regroup.get(t, 0.0) + v
        agent_total_views_list = list(regroup.values())

    if not real_counts:
        record("Topic_Summary Total_Views per topic matches groundtruth", False,
               "No reference view counts (DB and groundtruth unavailable)")
    else:
        real_grand = sum(real_counts)
        agent_grand = sum(agent_total_views_list)
        cond_a = num_close(agent_grand, real_grand, max(abs(real_grand) * 0.01, 1.0))
        cond_b = (len(agent_total_views_list) >= 2
                  and _partition_valid(agent_total_views_list, real_counts))
        record("Topic_Summary Total_Views per topic matches groundtruth",
               cond_a and cond_b,
               f"sum agent={agent_grand:.0f} real={real_grand:.0f} (A={cond_a}); "
               f"partition_valid={cond_b} over {real_counts}")


def check_teamly():
    print("\n=== Check 6: Teamly page (TEAM) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT p.id, p.title FROM teamly.pages p
                JOIN teamly.spaces s ON s.id = p.space_id
                WHERE s.key = 'TEAM'
                  AND (p.title ILIKE '%Fireship%' OR p.title ILIKE '%Tech Trends%')
            """)
            rows = cur.fetchall()
        conn.close()
        record("Teamly page (TEAM) with 'Fireship'/'Tech Trends' in title exists",
               len(rows) > 0,
               f"Found {len(rows)} matching pages: {[r[1] for r in rows]}")
    except Exception as e:
        record("Teamly page (TEAM) with 'Fireship'/'Tech Trends' in title exists", False, str(e))


def check_gcal():
    print("\n=== Check 7: GCal Tech Review event ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, summary, start_datetime FROM gcal.events
                WHERE summary ILIKE '%Tech Review%'
                  AND start_datetime >= '2026-03-01'
                  AND start_datetime < '2026-04-01'
            """)
            rows = cur.fetchall()
        conn.close()
        record("GCal has 'Tech Review' event in March 2026",
               len(rows) > 0,
               f"Found {len(rows)} matching events")
    except Exception as e:
        record("GCal has 'Tech Review' event in March 2026", False, str(e))


def check_email():
    print("\n=== Check 8: Email sent to techteam@company.com ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, subject, to_addr FROM email.messages
                WHERE to_addr::text ILIKE '%techteam@company.com%'
            """)
            rows = cur.fetchall()
            if not rows:
                try:
                    cur.execute("""
                        SELECT id FROM email.sent_log
                        WHERE to_addr ILIKE '%techteam@company.com%'
                    """)
                    rows = cur.fetchall()
                except Exception:
                    pass
        conn.close()
        record("Email sent to techteam@company.com",
               len(rows) > 0,
               f"Found {len(rows)} matching emails")
    except Exception as e:
        record("Email sent to techteam@company.com", False, str(e))


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print(f"Running evaluation for yt-fireship-tech-report-excel-notion")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_teamly()
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    accuracy = PASS_COUNT / total * 100 if total > 0 else 0
    print(f"\n{'='*40}")
    print(f"Passed: {PASS_COUNT}/{total} ({accuracy:.1f}%)")

    if CRITICAL_FAILS:
        print(f"CRITICAL failures: {CRITICAL_FAILS}")

    summary = f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}, accuracy={accuracy:.1f}%"

    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump({
                "total_passed": PASS_COUNT,
                "total_checks": total,
                "accuracy": accuracy,
                "critical_fails": CRITICAL_FAILS,
            }, f, indent=2)

    # Critical gate first, then accuracy gate.
    if CRITICAL_FAILS:
        print(f"Result: FAIL (critical) - {summary}")
        return False, summary
    if accuracy >= 70:
        print(f"Result: PASS - {summary}")
        return True, summary
    print(f"Result: FAIL - {summary}")
    return False, summary


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
