"""
Evaluation for yt-fireship-gform-survey-excel-gcal task (RU stack: forms-mcp).

Service identity stays foreign (keep-foreign): YouTube content, channel/video
titles, video IDs and the survey form remain in English. The survey form is now
backed by the RU forms MCP (local_servers/forms-mcp) which stores data in the same
gform.* schema (forms / questions / responses).

CRITICAL_CHECKS gate the substance of the task: a single critical failure => overall
FAIL (sys.exit(1)) regardless of accuracy. The accuracy>=70 gate applies afterward.
These critical checks ensure a correct RU agent passes while a non-doer (missing
outputs, preseeded/empty artefacts, wrong topics) fails.

Checks:
  Excel  (CRITICAL): Community_Report.xlsx with Top_Videos sheet (8 ranked videos,
          Engagement_Rate column) and Engagement_Analysis sheet (>= 4 topic rows),
          plus value match against groundtruth.
  Form   (CRITICAL): survey form exists (title references Fireship/Survey),
          has >= 5 questions, and the topic question offers the real Fireship topics.
  GCal   (CRITICAL): a new April-2026 "Community Standup" event distinct from the
          preseeded "Community Q&A" noise event at 16:00.
  Email  (CRITICAL): a report email actually sent to community@devclub.io.
"""
import os
import sys
import json
from argparse import ArgumentParser

import psycopg2
import openpyxl

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Topic labels the survey's first (topic) question must offer. Mixed cyr/lat is not
# expected here (English service content), but normalize keeps comparison robust.
TOPIC_OPTION_KEYS = ["javascript", "typescript", "react"]


def normalize(s):
    """Lowercase + strip; only meant to neutralise stray casing/whitespace in
    mixed cyr/lat identifiers, never used for RU-keyword greps."""
    return str(s).strip().lower() if s is not None else ""


def record(name, passed, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "CRITICAL " if critical else ""
    if passed:
        PASS_COUNT += 1
        print(f"  [PASS] {tag}{name}")
    else:
        FAIL_COUNT += 1
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL] {tag}{name}{msg}")
        if critical:
            CRITICAL_FAILS.append(name)


def num_close(a, b, tol=1.0):
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return False


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return normalize(a) == normalize(b)


def _option_values(config):
    """Extract option text values from a forms-mcp choice question config."""
    if not config:
        return []
    if isinstance(config, str):
        try:
            config = json.loads(config)
        except Exception:
            return []
    out = []
    opts = config.get("options") if isinstance(config, dict) else None
    if isinstance(opts, list):
        for o in opts:
            if isinstance(o, dict):
                out.append(normalize(o.get("value") or o.get("label") or ""))
            else:
                out.append(normalize(o))
    return [o for o in out if o]


def check_excel(agent_workspace, groundtruth_workspace="."):
    print("\n=== Check 1-3: Community_Report.xlsx ===")

    xlsx_path = os.path.join(agent_workspace, "Community_Report.xlsx")
    if not os.path.exists(xlsx_path):
        record("Community_Report.xlsx exists", False, f"Not found at {xlsx_path}", critical=True)
        record("Top_Videos sheet has >= 6 data rows", False, "File missing", critical=True)
        record("Engagement_Rate column or analysis exists", False, "File missing")
        record("Engagement_Analysis sheet has >= 4 rows", False, "File missing", critical=True)
        return
    record("Community_Report.xlsx exists", True, critical=True)

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        record("Excel readable", False, str(e), critical=True)
        return

    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}

    # Top_Videos sheet
    top_key = next((sheet_names_lower[k] for k in sheet_names_lower
                    if "video" in k or "top" in k), None)
    if not top_key:
        record("Top_Videos sheet has >= 6 data rows", False,
               f"No Top_Videos sheet. Sheets: {wb.sheetnames}", critical=True)
        record("Engagement_Rate column exists in Top_Videos", False, "Sheet missing")
    else:
        ws = wb[top_key]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = [r for r in rows[1:] if any(c for c in r)] if rows else []
        record("Top_Videos sheet has >= 6 data rows", len(data_rows) >= 6,
               f"Found {len(data_rows)} data rows", critical=True)
        if rows:
            headers = [normalize(c) for c in rows[0]]
            has_engagement = any("engagement" in h or "rate" in h for h in headers)
            record("Engagement_Rate column exists in Top_Videos", has_engagement,
                   f"Headers: {rows[0]}")
        else:
            record("Engagement_Rate column exists in Top_Videos", False, "Sheet empty")

    # Engagement_Analysis sheet
    eng_key = next((sheet_names_lower[k] for k in sheet_names_lower
                    if "engagement" in k or "analysis" in k), None)
    if not eng_key:
        record("Engagement_Analysis sheet has >= 4 rows", False,
               f"No Engagement_Analysis sheet. Sheets: {wb.sheetnames}", critical=True)
    else:
        ws2 = wb[eng_key]
        rows2 = list(ws2.iter_rows(values_only=True))
        data_rows2 = [r for r in rows2[1:] if any(c for c in r)] if rows2 else []
        record("Engagement_Analysis sheet has >= 4 rows", len(data_rows2) >= 4,
               f"Found {len(data_rows2)} data rows", critical=True)

    # --- Groundtruth XLSX value comparison ---
    gt_path = os.path.join(groundtruth_workspace, "Community_Report.xlsx")
    if os.path.isfile(gt_path):
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
        try:
            a_wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        except Exception:
            a_wb = None
        if a_wb:
            for gt_sname in gt_wb.sheetnames:
                gt_ws = gt_wb[gt_sname]
                a_ws = None
                for asn in a_wb.sheetnames:
                    if normalize(asn) == normalize(gt_sname):
                        a_ws = a_wb[asn]
                        break
                if a_ws is None:
                    record(f"GT sheet '{gt_sname}' exists in agent xlsx", False,
                           f"Available: {a_wb.sheetnames}", critical=True)
                    continue
                gt_rows = [r for r in gt_ws.iter_rows(min_row=2, values_only=True)
                           if any(c is not None for c in r)]
                a_rows = [r for r in a_ws.iter_rows(min_row=2, values_only=True)
                          if any(c is not None for c in r)]
                record(f"GT '{gt_sname}' row count", len(a_rows) == len(gt_rows),
                       f"Expected {len(gt_rows)}, got {len(a_rows)}")
                for ri in range(min(3, len(gt_rows))):
                    if ri >= len(a_rows):
                        break
                    ok = True
                    for ci in range(min(len(gt_rows[ri]), len(a_rows[ri]))):
                        gv, av = gt_rows[ri][ci], a_rows[ri][ci]
                        if gv is None:
                            continue
                        if isinstance(gv, (int, float)):
                            if not num_close(av, gv, max(abs(gv) * 0.1, 1.0)):
                                ok = False
                                break
                        else:
                            if not str_match(av, gv):
                                ok = False
                                break
                    record(f"GT '{gt_sname}' row {ri+1} values", ok,
                           f"gt={gt_rows[ri][:4]}, "
                           f"agent={a_rows[ri][:4] if ri < len(a_rows) else 'missing'}")
            a_wb.close()
        gt_wb.close()


def check_form():
    print("\n=== Check 4-5: Survey form (forms-mcp / gform.*) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
    except Exception as e:
        record("Survey form exists", False, str(e), critical=True)
        record("Survey form has >= 5 questions", False, "no db", critical=True)
        record("Topic question offers Fireship topics", False, "no db", critical=True)
        return
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT f.id, f.title FROM gform.forms f
                WHERE f.title ILIKE '%Survey%'
                   OR f.title ILIKE '%Fireship%'
            """)
            forms = cur.fetchall()
            if not forms:
                record("Survey form exists (title references Fireship/Survey)", False,
                       "No matching form found", critical=True)
                record("Survey form has >= 5 questions", False, "Form missing", critical=True)
                record("Topic question offers Fireship topics", False, "Form missing", critical=True)
                return
            record("Survey form exists (title references Fireship/Survey)", True,
                   f"Found: {[f[1] for f in forms]}", critical=True)

            form_id = forms[0][0]
            cur.execute("""
                SELECT title, question_type, config
                FROM gform.questions WHERE form_id = %s ORDER BY position
            """, (form_id,))
            questions = cur.fetchall()
            record("Survey form has >= 5 questions", len(questions) >= 5,
                   f"Found {len(questions)} questions", critical=True)

            # Locate the topic question (choice question that offers JS/TS/React).
            topic_ok = False
            best_detail = "no choice question listed the core topics"
            for q_title, q_type, q_config in questions:
                opts = _option_values(q_config)
                if not opts:
                    continue
                joined = " | ".join(opts)
                hits = sum(1 for k in TOPIC_OPTION_KEYS if k in joined)
                if hits >= 2:
                    topic_ok = True
                    best_detail = f"'{q_title}' options={opts}"
                    break
            record("Topic question offers Fireship topics (JS/TS/React)", topic_ok,
                   best_detail, critical=True)
    finally:
        conn.close()


def check_gcal():
    print("\n=== Check 6: GCal Community Standup in April 2026 ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM gcal.events
                WHERE (summary ILIKE '%Community%' OR summary ILIKE '%Standup%')
                  AND start_datetime >= '2026-04-01'
                  AND start_datetime < '2026-05-01'
                  AND summary NOT ILIKE '%Q&A%'
            """)
            count = cur.fetchone()[0]
        conn.close()
        record("GCal has new Community/Standup event in April 2026 (not Q&A noise)",
               count > 0, f"Found {count} events", critical=True)
    except Exception as e:
        record("GCal check", False, str(e), critical=True)


def check_email():
    print("\n=== Check 7: Email sent to community@devclub.io ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        with conn.cursor() as cur:
            cur.execute("""
                SELECT COUNT(*) FROM email.messages
                WHERE to_addr::text ILIKE '%community@devclub.io%'
                  AND from_addr != 'community@devclub.io'
            """)
            count = cur.fetchone()[0]
            if count == 0:
                try:
                    cur.execute("""
                        SELECT COUNT(*) FROM email.sent_log
                        WHERE to_addr ILIKE '%community@devclub.io%'
                    """)
                    count = cur.fetchone()[0]
                except Exception:
                    pass
        conn.close()
        record("Email sent to community@devclub.io", count > 0, f"Found {count}",
               critical=True)
    except Exception as e:
        record("Email check", False, str(e), critical=True)


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    print("Running evaluation for yt-fireship-gform-survey-excel-gcal")
    print(f"Agent workspace: {agent_workspace}")

    check_excel(agent_workspace, groundtruth_workspace)
    check_form()
    check_gcal()
    check_email()

    total = PASS_COUNT + FAIL_COUNT
    if total == 0:
        print("\nFAIL: No checks were performed.")
        if res_log_file:
            with open(res_log_file, "w") as f:
                json.dump({"total_passed": 0, "total_checks": 0,
                           "accuracy": 0.0, "critical_failures": ["no checks"]}, f)
        sys.exit(1)

    accuracy = PASS_COUNT / total * 100
    print(f"\n{'='*40}")
    print(f"Overall: {PASS_COUNT}/{total} checks passed ({accuracy:.1f}%)")

    result = {
        "total_passed": PASS_COUNT,
        "total_checks": total,
        "accuracy": accuracy,
        "critical_failures": CRITICAL_FAILS,
    }
    if res_log_file:
        with open(res_log_file, "w") as f:
            json.dump(result, f, indent=2)

    if CRITICAL_FAILS:
        print(f"\nFAIL: critical checks failed: {CRITICAL_FAILS}")
        return False, f"Critical failures: {CRITICAL_FAILS}"

    if accuracy >= 70:
        print("PASS")
        return True, f"Passed: {PASS_COUNT}, Failed: {FAIL_COUNT}"
    print("FAIL")
    return False, f"Accuracy {accuracy:.1f}% < 70%"


def main():
    parser = ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()
    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
