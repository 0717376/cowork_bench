"""Evaluation script for fetch-email-gcal-scheduler-excel-word (RU)."""
import os
import argparse, json, os, sys
import openpyxl


DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"), "port": 5432,
    "dbname": os.environ.get("PGDATABASE", "cowork_gym"),
    "user": "eigent", "password": "camel"
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILS = []

# Source-of-truth derived from the fetched mock JSON
# data.json:  Отдел продаж=100, Отдел маркетинга=80  (Internal_Value)
# benchmarks.json: Отдел продаж=95, Отдел маркетинга=90  (External_Benchmark)
# Gap = Internal - Benchmark => Отдел продаж=5, Отдел маркетинга=-10
EXPECTED = {
    "отдел продаж": {"internal": 100, "benchmark": 95, "gap": 5},
    "отдел маркетинга": {"internal": 80, "benchmark": 90, "gap": -10},
}
EXPECTED_TOTAL = 2
EXPECTED_AVG_GAP = -2.5
# item with the largest negative gap (priority focus per analysis_guide.md)
LARGEST_NEG_ITEM = "отдел маркетинга"


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    marker = " [CRITICAL]" if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{marker} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILS.append(name)
        detail_str = str(detail)[:200] if detail else ""
        print(f"  [FAIL]{marker} {name}: {detail_str}")


def safe_float(val, default=None):
    try:
        if val is None:
            return default
        return float(str(val).replace(',', '').replace('%', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return default


def get_conn():
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def col_index(headers, name):
    name = name.lower()
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None


def run_evaluation(agent_workspace, groundtruth_workspace, launch_time, res_log_file):
    global PASS_COUNT, FAIL_COUNT, CRITICAL_FAILS
    PASS_COUNT = 0
    FAIL_COUNT = 0
    CRITICAL_FAILS = []

    excel_path = os.path.join(agent_workspace, "Scheduler_Report.xlsx")
    # File existence is CRITICAL: the whole report is the core deliverable.
    check("Scheduler_Report.xlsx exists", os.path.exists(excel_path),
          excel_path, critical=True)

    parsed = {}  # item(lower) -> {internal, benchmark, gap}
    if os.path.exists(excel_path):
        wb = openpyxl.load_workbook(excel_path)

        # ---------- Data_Analysis ----------
        check("Data_Analysis sheet exists", "Data_Analysis" in wb.sheetnames, critical=True)
        if "Data_Analysis" in wb.sheetnames:
            ws = wb["Data_Analysis"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Data_Analysis has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")

            for expected_col in ['Item', 'Internal_Value', 'External_Benchmark', 'Gap']:
                check(f"Data_Analysis has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            ci_item = col_index(headers, 'item')
            ci_int = col_index(headers, 'internal_value')
            ci_bm = col_index(headers, 'external_benchmark')
            ci_gap = col_index(headers, 'gap')

            if None not in (ci_item, ci_int, ci_bm, ci_gap):
                for r in data_rows:
                    if ci_item >= len(r) or r[ci_item] is None:
                        continue
                    item = str(r[ci_item]).strip().lower()
                    parsed[item] = {
                        "internal": safe_float(r[ci_int]),
                        "benchmark": safe_float(r[ci_bm]),
                        "gap": safe_float(r[ci_gap]),
                    }

                # CRITICAL: values must match the fetched source AND gap arithmetic must hold
                values_ok = True
                gap_ok = True
                detail_v = []
                detail_g = []
                for key, exp in EXPECTED.items():
                    got = parsed.get(key)
                    if not got:
                        values_ok = False
                        detail_v.append(f"missing row '{key}'")
                        continue
                    if got["internal"] != exp["internal"] or got["benchmark"] != exp["benchmark"]:
                        values_ok = False
                        detail_v.append(f"{key}: int={got['internal']} bm={got['benchmark']} exp int={exp['internal']} bm={exp['benchmark']}")
                    # gap == internal - benchmark
                    if got["internal"] is not None and got["benchmark"] is not None and got["gap"] is not None:
                        if abs(got["gap"] - (got["internal"] - got["benchmark"])) > 0.01:
                            gap_ok = False
                            detail_g.append(f"{key}: gap={got['gap']} != {got['internal']}-{got['benchmark']}")
                    else:
                        gap_ok = False
                        detail_g.append(f"{key}: missing numeric value")

                check("Data_Analysis values match fetched source (Internal_Value & External_Benchmark)",
                      values_ok, "; ".join(detail_v), critical=True)
                check("Data_Analysis Gap == Internal_Value - External_Benchmark for every row",
                      gap_ok, "; ".join(detail_g), critical=True)

        # ---------- Metrics ----------
        check("Metrics sheet exists", "Metrics" in wb.sheetnames, critical=True)
        if "Metrics" in wb.sheetnames:
            ws = wb["Metrics"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Metrics has >= 2 rows", len(data_rows) >= 2, f"got {len(data_rows)}")
            for expected_col in ['Metric', 'Value']:
                check(f"Metrics has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # Build metric->value map
            metrics_map = {}
            for r in data_rows:
                if r and r[0] is not None:
                    metrics_map[str(r[0]).strip().lower()] = r[1] if len(r) > 1 else None

            total_v = safe_float(metrics_map.get("total_items"))
            avg_v = safe_float(metrics_map.get("avg_gap"))
            # CRITICAL: semantically correct aggregates derived from the actual data
            total_ok = (total_v is not None and abs(total_v - EXPECTED_TOTAL) < 0.01)
            avg_ok = (avg_v is not None and abs(avg_v - EXPECTED_AVG_GAP) < 0.05)
            check("Metrics: Total_Items equals number of Data_Analysis rows",
                  total_ok, f"Total_Items={total_v} expected {EXPECTED_TOTAL}", critical=True)
            check("Metrics: Avg_Gap equals mean of Gap column",
                  avg_ok, f"Avg_Gap={avg_v} expected {EXPECTED_AVG_GAP}", critical=True)

        # ---------- Recommendations ----------
        check("Recommendations sheet exists", "Recommendations" in wb.sheetnames, critical=True)
        if "Recommendations" in wb.sheetnames:
            ws = wb["Recommendations"]
            headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            check("Recommendations has >= 1 rows", len(data_rows) >= 1, f"got {len(data_rows)}")
            for expected_col in ['Priority', 'Action']:
                check(f"Recommendations has {expected_col} column",
                      expected_col.lower() in headers, f"headers: {headers[:8]}")

            # CRITICAL: recommendations derive from gap analysis -> reference largest-neg-gap item
            all_text = " ".join(
                str(c).lower() for r in data_rows for c in r if c is not None
            )
            refs_largest_neg = (LARGEST_NEG_ITEM in all_text) or ("маркетинг" in all_text)
            check("Recommendations reference the item with the largest negative gap",
                  refs_largest_neg, f"text: {all_text[:200]}", critical=True)
    else:
        # Excel missing -> mark dependent deliverables as failed (do not silently skip)
        check("Metrics sheet exists", False, "Excel missing", critical=True)
        check("Recommendations sheet exists", False, "Excel missing", critical=True)

    # ---------- Email (independent of Excel) ----------
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT subject, to_addr, body_text FROM email.messages")
        messages = cur.fetchall()
        conn.close()

        def to_addresses(to_addr):
            if isinstance(to_addr, list):
                return " ".join(str(r).lower() for r in to_addr)
            if to_addr:
                try:
                    parsed_addr = json.loads(str(to_addr))
                    if isinstance(parsed_addr, list):
                        return " ".join(str(r).lower() for r in parsed_addr)
                    return str(to_addr).lower()
                except Exception:
                    return str(to_addr).lower()
            return ""

        # broad existence (non-critical): any analysis/report-ish outgoing email
        analysisish = [m for m in messages if m[0] and (
            "report" in str(m[0]).lower() or "analysis" in str(m[0]).lower()
            or "отчёт" in str(m[0]).lower() or "анализ" in str(m[0]).lower()
        )]
        check("Analysis email sent", len(analysisish) >= 1,
              f"found {len(analysisish)} matching emails")

        # CRITICAL: correct recipient AND exact subject
        target = [m for m in messages
                  if "team-lead@company.com" in to_addresses(m[1])
                  and m[0] is not None and str(m[0]).strip() == "Analysis Report Complete"]
        check("Email to team-lead@company.com with subject 'Analysis Report Complete'",
              len(target) >= 1,
              f"subjects/to: {[(str(m[0]), to_addresses(m[1])) for m in messages][:5]}",
              critical=True)
    except Exception as e:
        check("Email check", False, str(e), critical=True)

    # ---------- Calendar (independent of Excel) ----------
    try:
        conn = get_conn()
        cur = conn.cursor()
        cur.execute("SELECT summary, start_datetime, end_datetime FROM gcal.events")
        events = cur.fetchall()
        conn.close()

        def parse_dt(v):
            """Return (date_iso, hour) tolerant to str/datetime and tz suffix."""
            if v is None:
                return None, None
            try:
                # datetime object
                return v.date().isoformat(), v.hour
            except AttributeError:
                pass
            s = str(v)
            date_part = s[:10]
            hour = None
            # find a HH after a space or 'T'
            import re
            m = re.search(r"[ T](\d{2}):", s)
            if m:
                hour = int(m.group(1))
            return date_part, hour

        # broad existence (non-critical)
        review_ish = [e for e in events if e[0] and (
            "review" in str(e[0]).lower() or "meeting" in str(e[0]).lower()
            or "обзор" in str(e[0]).lower() or "анализ" in str(e[0]).lower()
        )]
        check("Review event created", len(review_ish) >= 1,
              f"found {len(review_ish)} events")

        # CRITICAL: correct title-ish AND correct schedule (2026-03-14, start 14:00 UTC)
        schedule_ok = False
        detail_sched = []
        for e in events:
            summ = str(e[0]).lower() if e[0] else ""
            title_match = ("analysis review" in summ or "обзор анализа" in summ
                           or ("review" in summ or "обзор" in summ))
            sd, sh = parse_dt(e[1])
            ed, eh = parse_dt(e[2])
            detail_sched.append((summ, sd, sh, ed, eh))
            if title_match and sd == "2026-03-14" and sh == 14 and (eh is None or eh == 15):
                schedule_ok = True
                break
        check("Calendar 'Analysis Review' event on 2026-03-14 14:00-15:00 UTC",
              schedule_ok, f"events: {detail_sched[:5]}", critical=True)
    except Exception as e:
        check("Calendar check", False, str(e), critical=True)

    # ---------- Word document (independent of Excel) ----------
    import glob as globmod
    word_files = globmod.glob(os.path.join(agent_workspace, "*.docx"))
    sched_doc = os.path.join(agent_workspace, "Scheduler_Analysis.docx")
    check("Scheduler_Analysis.docx exists", os.path.exists(sched_doc),
          f"found {word_files}", critical=True)
    if os.path.exists(sched_doc):
        from docx import Document
        doc = Document(sched_doc)
        text = " ".join(p.text for p in doc.paragraphs).lower()
        check("Word has content", len(text) > 50, f"text length: {len(text)}")

    # ---------- processor script ----------
    check("email_scheduler_processor.py exists",
          os.path.exists(os.path.join(agent_workspace, "email_scheduler_processor.py")))

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\nAccuracy: {accuracy:.1f}% ({PASS_COUNT}/{total})")
    if CRITICAL_FAILS:
        print(f"Critical fails: {CRITICAL_FAILS}")

    success = (not CRITICAL_FAILS) and accuracy >= 70
    return success, f"Passed {PASS_COUNT}/{total} checks; accuracy={accuracy:.1f}%; critical_fails={CRITICAL_FAILS}"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=False, default=".")
    parser.add_argument("--groundtruth_workspace", required=False, default=".")
    parser.add_argument("--launch_time", required=False, default="2026-03-07 10:00:00")
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    success, message = run_evaluation(
        args.agent_workspace, args.groundtruth_workspace,
        args.launch_time, args.res_log_file
    )
    print(message)
    # Critical gate: any critical fail => hard fail regardless of accuracy
    if CRITICAL_FAILS:
        sys.exit(1)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
