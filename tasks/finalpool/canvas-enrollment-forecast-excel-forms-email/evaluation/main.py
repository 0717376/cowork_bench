"""Evaluation for canvas-enrollment-forecast-excel-gform-email (RU).

Gate: accuracy >= 70% AND no CRITICAL check failed => PASS.
Any CRITICAL failure => immediate FAIL (sys.exit(1)) regardless of accuracy.
"""
import argparse
import json
import os
import sys

import openpyxl
import psycopg2

DB_CONFIG = {
    "host": os.environ.get("PGHOST", "localhost"),
    "port": 5432,
    "dbname": "cowork_gym",
    "user": "eigent",
    "password": "camel",
}

PASS_COUNT = 0
FAIL_COUNT = 0
CRITICAL_FAILED = []


def check(name, condition, detail="", critical=False):
    global PASS_COUNT, FAIL_COUNT
    tag = "[CRIT]" if critical else ""
    if condition:
        PASS_COUNT += 1
        print(f"  [PASS]{tag} {name}")
    else:
        FAIL_COUNT += 1
        if critical:
            CRITICAL_FAILED.append(name)
        msg = f": {str(detail)[:300]}" if detail else ""
        print(f"  [FAIL]{tag} {name}{msg}")


def num_close(a, b, tol=1.0):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= tol
    except (TypeError, ValueError):
        return str(a).strip().lower() == str(b).strip().lower()


def str_match(a, b):
    if a is None or b is None:
        return a is None and b is None
    return str(a).strip().lower() == str(b).strip().lower()


def load_sheet_rows(wb, sheet_name):
    for name in wb.sheetnames:
        if name.strip().lower() == sheet_name.strip().lower():
            return [[cell.value for cell in row] for row in wb[name].iter_rows()]
    return None


# Maps chair email prefix -> GT base course name (для проверки писем).
CHAIR_TO_COURSE = {
    "analytics": "Прикладная аналитика и алгоритмы",
    "biochem": "Биохимия и биоинформатика",
    "computing": "Креативные вычисления и культура",
    "design": "Проектирование на основе данных",
    "economics": "Экологическая экономика и этика",
    "finance": "Основы финансов",
    "governance": "Глобальное управление и геополитика",
}


def load_gt_capacity(gt_dir):
    """Возвращает {course_base_name_lower: row} из листа Course Capacity эталона."""
    gt_path = os.path.join(gt_dir, "Enrollment_Analysis.xlsx")
    lookup = {}
    if not os.path.isfile(gt_path):
        return lookup
    try:
        wb = openpyxl.load_workbook(gt_path, data_only=True)
        rows = load_sheet_rows(wb, "Course Capacity")
        if rows:
            for r in rows[1:]:
                if r and r[0] is not None:
                    lookup[str(r[0]).strip().lower()] = r
    except Exception:
        pass
    return lookup


def check_excel(agent_workspace, gt_dir):
    print("\n=== Checking Excel ===")
    xlsx_path = os.path.join(agent_workspace, "Enrollment_Analysis.xlsx")
    gt_path = os.path.join(gt_dir, "Enrollment_Analysis.xlsx")

    if not os.path.isfile(xlsx_path):
        check("Enrollment_Analysis.xlsx exists", False, f"Not found: {xlsx_path}", critical=True)
        return
    check("Enrollment_Analysis.xlsx exists", True)

    if not os.path.isfile(gt_path):
        check("Groundtruth Excel exists", False, f"Not found: {gt_path}")
        return

    try:
        agent_wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        gt_wb = openpyxl.load_workbook(gt_path, data_only=True)
    except Exception as e:
        check("Excel readable", False, str(e), critical=True)
        return
    check("Excel readable", True)

    # Sheet 1: Enrollment Trends (структурный + значения; значения CRITICAL)
    print("  --- Enrollment Trends ---")
    a_rows = load_sheet_rows(agent_wb, "Enrollment Trends")
    g_rows = load_sheet_rows(gt_wb, "Enrollment Trends")
    if a_rows is None:
        check("Sheet 'Enrollment Trends' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Enrollment Trends' exists in GT", False)
    else:
        check("Sheet 'Enrollment Trends' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Enrollment Trends has 22 data rows", len(a_data) == 22, f"Found {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                key = (str(row[0]).strip().lower(), str(row[1]).strip().lower(), int(row[2]) if row[2] else 0)
                a_lookup[key] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = (str(g_row[0]).strip().lower(), str(g_row[1]).strip().lower(), int(g_row[2]) if g_row[2] else 0)
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing row: {g_row[0]} {g_row[1]} {g_row[2]}")
                continue
            if len(a_row) > 3 and len(g_row) > 3:
                if not num_close(a_row[3], g_row[3], 1):
                    errors.append(f"{key}: Student_Count {a_row[3]} vs {g_row[3]}")
            if len(a_row) > 4 and len(g_row) > 4:
                if not num_close(a_row[4], g_row[4], 1):
                    errors.append(f"{key}: Teacher_Count {a_row[4]} vs {g_row[4]}")
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 1):
                    errors.append(f"{key}: TA_Count {a_row[5]} vs {g_row[5]}")
        # CRITICAL: данные по набору должны совпадать с эталоном
        check("Enrollment Trends data matches GT", not errors,
              "; ".join(errors[:5]), critical=True)

    # Sheet 2: Course Capacity (ядро прогноза — CRITICAL по тренду/split/consolidation/projection)
    print("  --- Course Capacity ---")
    a_rows = load_sheet_rows(agent_wb, "Course Capacity")
    g_rows = load_sheet_rows(gt_wb, "Course Capacity")
    if a_rows is None:
        check("Sheet 'Course Capacity' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Course Capacity' exists in GT", False)
    else:
        check("Sheet 'Course Capacity' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []
        check("Course Capacity has 7 data rows", len(a_data) == 7, f"Found {len(a_data)}")

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        forecast_errors = []   # Growth_Trend / Needs_Split / Consider_Consolidation / Projected_Next
        faculty_errors = []
        latest_errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                forecast_errors.append(f"Missing row: {g_row[0]}")
                latest_errors.append(f"Missing row: {g_row[0]}")
                faculty_errors.append(f"Missing row: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                if not num_close(a_row[1], g_row[1], 1):
                    latest_errors.append(f"{key}: Latest_Enrollment {a_row[1]} vs {g_row[1]}")
            if len(a_row) > 2 and len(g_row) > 2:
                if not str_match(a_row[2], g_row[2]):
                    forecast_errors.append(f"{key}: Growth_Trend '{a_row[2]}' vs '{g_row[2]}'")
            if len(a_row) > 3 and len(g_row) > 3:
                if not str_match(a_row[3], g_row[3]):
                    forecast_errors.append(f"{key}: Needs_Split '{a_row[3]}' vs '{g_row[3]}'")
            if len(a_row) > 4 and len(g_row) > 4:
                if not str_match(a_row[4], g_row[4]):
                    forecast_errors.append(f"{key}: Consider_Consolidation '{a_row[4]}' vs '{g_row[4]}'")
            if len(a_row) > 5 and len(g_row) > 5:
                if not num_close(a_row[5], g_row[5], 5):
                    forecast_errors.append(f"{key}: Projected_Next {a_row[5]} vs {g_row[5]}")
            if len(a_row) > 6 and len(g_row) > 6:
                if not num_close(a_row[6], g_row[6], 1):
                    faculty_errors.append(f"{key}: Faculty_Needed {a_row[6]} vs {g_row[6]}")
        check("Course Capacity Latest_Enrollment matches GT", not latest_errors,
              "; ".join(latest_errors[:5]))
        # CRITICAL: основной прогнозный результат
        check("Course Capacity forecast (Trend/Split/Consolidation/Projected) matches GT",
              not forecast_errors, "; ".join(forecast_errors[:5]), critical=True)
        # CRITICAL: кадровые нормативы из PDF (1/30 если <500, иначе 1/50)
        check("Course Capacity Faculty_Needed respects staffing ratio (GT)",
              not faculty_errors, "; ".join(faculty_errors[:5]), critical=True)

    # Sheet 3: Department Summary (агрегаты всей задачи — CRITICAL)
    print("  --- Department Summary ---")
    a_rows = load_sheet_rows(agent_wb, "Department Summary")
    g_rows = load_sheet_rows(gt_wb, "Department Summary")
    if a_rows is None:
        check("Sheet 'Department Summary' exists", False, f"Available: {agent_wb.sheetnames}")
    elif g_rows is None:
        check("Sheet 'Department Summary' exists in GT", False)
    else:
        check("Sheet 'Department Summary' exists", True)
        a_data = a_rows[1:] if len(a_rows) > 1 else []
        g_data = g_rows[1:] if len(g_rows) > 1 else []

        a_lookup = {}
        for row in a_data:
            if row and row[0] is not None:
                a_lookup[str(row[0]).strip().lower()] = row
        errors = []
        for g_row in g_data:
            if not g_row or g_row[0] is None:
                continue
            key = str(g_row[0]).strip().lower()
            a_row = a_lookup.get(key)
            if a_row is None:
                errors.append(f"Missing metric: {g_row[0]}")
                continue
            if len(a_row) > 1 and len(g_row) > 1:
                tol = 5 if "projected" in key else 1
                if not num_close(a_row[1], g_row[1], tol):
                    errors.append(f"{key}: {a_row[1]} vs {g_row[1]} (tol={tol})")
        # CRITICAL: агрегаты всей задачи (Total_Student_Enrollments=32593 и т.д.)
        check("Department Summary aggregates match GT", not errors,
              "; ".join(errors[:5]), critical=True)


def check_gform():
    print("\n=== Checking Google Form (forms RU fork schema) ===")
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        cur.execute("""
            SELECT id, title FROM gform.forms
            WHERE title ILIKE '%%course preference%%'
               OR title ILIKE '%%preference survey%%'
               OR title ILIKE '%%предпочтени%%'
               OR title ILIKE '%%опрос%%курс%%'
        """)
        forms = cur.fetchall()
        check("Course Preference Survey form exists", len(forms) >= 1,
              f"Found {len(forms)} matching forms")

        if forms:
            form_id = forms[0][0]
            cur.execute("""
                SELECT id, title, question_type FROM gform.questions
                WHERE form_id = %s ORDER BY position
            """, (form_id,))
            questions = cur.fetchall()
            check("Form has 4 questions", len(questions) == 4,
                  f"Found {len(questions)} questions: {[q[1] for q in questions]}")

            # Нормализуем тип к фрагментам форка: textQuestion / choiceQuestion (+EN-совместимость)
            def is_text(t):
                t = (t or "")
                return t in ("textQuestion", "TEXT", "SHORT_ANSWER", "PARAGRAPH")

            def is_choice(t):
                t = (t or "")
                return t in ("choiceQuestion", "RADIO", "MULTIPLE_CHOICE", "CHOICE", "CHECKBOX")

            titles = [(q[1] or "").lower() for q in questions]
            types = [q[2] for q in questions]

            text_count = sum(1 for t in types if is_text(t))
            choice_count = sum(1 for t in types if is_choice(t))

            # Вопрос про имя (текстовый)
            has_name_text = any(("name" in titles[i] or "имя" in titles[i] or "фио" in titles[i]
                                 or "имени" in titles[i]) and is_text(types[i])
                                for i in range(len(titles)))
            check("Has student name text question", has_name_text,
                  f"Questions: {list(zip(titles, types))}")

            # Два текстовых вопроса: имя + accessibility (forms-fork пишет только textQuestion)
            check("Has exactly 2 text questions (name + accessibility)", text_count == 2,
                  f"text_count={text_count}, types={types}")

            # Курсы (множественный выбор) + расписание -> два choiceQuestion в форке
            # (форк не умеет CHECKBOX, поэтому курсы оформляются как choiceQuestion)
            check("Has exactly 2 choice questions (courses + schedule)", choice_count == 2,
                  f"choice_count={choice_count}, types={types}")

            # Вопрос о расписании: ищем по опциям (если таблица опций есть) ИЛИ по заголовку.
            # Схема форка варьируется, поэтому пробуем варианты и не падаем при их отсутствии.
            all_opts = ""
            for tbl, col in (("gform.options", "value"),
                             ("gform.question_options", "value"),
                             ("gform.options", "label"),
                             ("gform.question_options", "label")):
                try:
                    cur.execute(
                        f"SELECT o.{col} FROM {tbl} o "
                        f"JOIN gform.questions q ON q.id = o.question_id "
                        f"WHERE q.form_id = %s", (form_id,))
                    rows = cur.fetchall()
                    all_opts = " ".join((r[0] or "").lower() for r in rows)
                    if all_opts.strip():
                        break
                except Exception:
                    conn.rollback()
                    continue

            sched_by_opts = (
                ("morning" in all_opts and "afternoon" in all_opts and "evening" in all_opts)
                or ("утро" in all_opts and "вечер" in all_opts)
            )
            # Запасной вариант: вопрос о расписании по заголовку среди choice-вопросов.
            sched_by_title = any(
                is_choice(types[i]) and any(
                    kw in titles[i] for kw in
                    ("schedule", "расписан", "время", "morning", "утро"))
                for i in range(len(titles))
            )
            sched_ok = sched_by_opts or sched_by_title
            check("Schedule choice question present (options Morning/Afternoon/Evening or RU)",
                  sched_ok, f"opts='{all_opts[:120]}' by_title={sched_by_title}")

            # CRITICAL: ровно 4 вопроса, 2 текстовых + 2 choice, с именем и расписанием
            check("Course Preference Survey: 4 questions (2 text + 2 choice) with name & schedule",
                  (len(questions) == 4 and text_count == 2 and choice_count == 2
                   and has_name_text and sched_ok),
                  f"q={len(questions)} text={text_count} choice={choice_count} "
                  f"name={has_name_text} sched={sched_ok}",
                  critical=True)

        cur.close()
        conn.close()
    except Exception as e:
        check("Google Form check", False, str(e))


def check_emails(gt_dir):
    print("\n=== Checking Emails ===")
    gt_cap = load_gt_capacity(gt_dir)  # {course_lower: row}, Projected_Next at index 5
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()

        chair_emails = [f"{p}_chair@university.edu" for p in CHAIR_TO_COURSE]

        # Тема: EN "enrollment/forecast/projection" или RU "прогноз/набор"
        cur.execute("""
            SELECT id, subject, to_addr, body_text FROM email.messages
            WHERE subject ILIKE '%%enrollment%%'
               OR subject ILIKE '%%forecast%%'
               OR subject ILIKE '%%projection%%'
               OR subject ILIKE '%%прогноз%%'
               OR subject ILIKE '%%набор%%'
        """)
        emails = cur.fetchall()
        check("At least 7 enrollment-related emails sent", len(emails) >= 7,
              f"Found {len(emails)} matching emails")

        # Сопоставление адрес -> письмо (берём последнее по каждому адресату)
        addr_to_email = {}
        for email_row in emails:
            to_addr = email_row[2]
            if isinstance(to_addr, str):
                try:
                    to_addr = json.loads(to_addr)
                except json.JSONDecodeError:
                    to_addr = [to_addr]
            if not isinstance(to_addr, list):
                to_addr = [to_addr]
            for addr in to_addr:
                addr_lower = str(addr).lower().strip()
                for chair in chair_emails:
                    if chair in addr_lower:
                        addr_to_email[chair] = email_row

        found_chairs = set(addr_to_email.keys())
        for chair in chair_emails:
            check(f"Email sent to {chair}", chair in found_chairs,
                  f"Found emails to: {found_chairs}")

        # CRITICAL: каждому из 7 завкафедрам отправлено письмо И тело содержит
        # число Projected_Next его кафедры (не просто length>20).
        body_ok = []
        for prefix, course in CHAIR_TO_COURSE.items():
            chair = f"{prefix}_chair@university.edu"
            er = addr_to_email.get(chair)
            if er is None:
                body_ok.append(f"{chair}: no email")
                continue
            body = str(er[3] or "")
            gt_row = gt_cap.get(course.strip().lower())
            proj = None
            if gt_row is not None and len(gt_row) > 5:
                proj = gt_row[5]
            if proj is None:
                body_ok.append(f"{chair}: GT projection missing")
                continue
            # Нормализуем тело от разделителей разрядов (запятая/пробел/NBSP),
            # чтобы "2,215" и "2 215" совпадали с "2215".
            body_norm = (body.replace(",", "").replace(" ", "")
                             .replace(" ", "").replace(" ", ""))
            # Допускаем небольшое расхождение округления: проверяем точное и +/-1.
            candidates = {str(int(round(float(proj)))),
                          str(int(round(float(proj))) + 1),
                          str(int(round(float(proj))) - 1)}
            if not any(c in body or c in body_norm for c in candidates):
                body_ok.append(f"{chair}: projection {proj} not in body")

        check("All 7 chair emails contain their dept Projected_Next number in body",
              not body_ok, "; ".join(body_ok[:7]), critical=True)

        cur.close()
        conn.close()
    except Exception as e:
        check("Email check", False, str(e))
        check("All 7 chair emails contain their dept Projected_Next number in body",
              False, str(e), critical=True)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--agent_workspace", required=True)
    parser.add_argument("--groundtruth_workspace", required=False)
    parser.add_argument("--launch_time", required=False)
    parser.add_argument("--res_log_file", required=False)
    args = parser.parse_args()

    task_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    gt_dir = args.groundtruth_workspace or os.path.join(task_root, "groundtruth_workspace")

    check_excel(args.agent_workspace, gt_dir)
    check_gform()
    check_emails(gt_dir)

    total = PASS_COUNT + FAIL_COUNT
    accuracy = (PASS_COUNT / total * 100) if total else 0.0
    print(f"\n=== Results: {PASS_COUNT}/{total} passed ({accuracy:.1f}%) ===")

    if CRITICAL_FAILED:
        print("CRITICAL checks failed:")
        for c in CRITICAL_FAILED:
            print(f"  - {c}")
        print("=> FAIL (critical)")
        sys.exit(1)

    if accuracy >= 70:
        print("=> PASS (accuracy >= 70% and no critical failure)")
        sys.exit(0)
    else:
        print(f"=> FAIL (accuracy {accuracy:.1f}% < 70%)")
        sys.exit(1)


if __name__ == "__main__":
    main()
